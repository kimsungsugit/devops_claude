import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'D:\Project\devops\260105\reports\sts\sts_local_20260310_164052.xlsm',
    read_only=True, data_only=True
)
ws = wb['3.SW Integration Test Spec']
count = 0
srs_linked = 0
safety_count = 0
sample_rows = []
for r in range(7, min((ws.max_row or 7) + 1, 5000)):
    tc_id = ws.cell(row=r, column=2).value
    if tc_id:
        count += 1
        srs_id = str(ws.cell(row=r, column=13).value or '').strip()
        safety = str(ws.cell(row=r, column=4).value or '').strip()
        method = str(ws.cell(row=r, column=6).value or '').strip()
        if srs_id and srs_id != 'None':
            srs_linked += 1
        if safety and safety.upper() == 'X':
            safety_count += 1
        if len(sample_rows) < 5:
            sample_rows.append((str(tc_id), srs_id, safety, method))
wb.close()

with open(r'D:\Project\devops\260105\_sts_result.txt', 'w', encoding='utf-8') as f:
    f.write(f'Total STS TCs: {count}\n')
    f.write(f'With SRS link (col13): {srs_linked} ({100*srs_linked/max(count,1):.1f}%)\n')
    f.write(f'Safety Related (col4): {safety_count} ({100*safety_count/max(count,1):.1f}%)\n')
    f.write(f'Sample TCs:\n')
    for tc_id, srs_id, safety, method in sample_rows:
        f.write(f'  {tc_id} -> srs_id={srs_id}, safety={safety}, method={method}\n')
print('Done')
