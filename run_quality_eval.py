"""
UDS / STS / SUTS 전체 생성 + 품질 평가 스크립트
모든 문서(SRS, SDS, HSIS, 소스코드) 를 사용하여 생성 후 품질 지표를 출력한다.
"""
import sys
import os
import time
import json
from pathlib import Path
from datetime import datetime

# 프로젝트 루트를 sys.path에 추가
ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

# ── 경로 설정 ────────────────────────────────────────────────────────────────
DOCS_DIR   = ROOT / "docs"
SOURCE_ROOT = Path("D:/Project/Ados/PDS_64_RD")
REPORTS_DIR = ROOT / "reports"

SRS_PATH  = DOCS_DIR / "(HDPDM01_SRS) Software Requirements Specification_v1.05_20230510.docx"
SDS_PATH  = DOCS_DIR / "(HDPDM01_SDS) Software Architecture Design Specification_v1.04_20230512.docx"
HSIS_PATH = DOCS_DIR / "(HDPDM01_HSIS) Hardware Software Interface Specification_v5.00_20221115_B.xlsx"

UDS_OUT_DIR  = REPORTS_DIR / "uds_local"
STS_OUT_DIR  = REPORTS_DIR / "sts"
SUTS_OUT_DIR = REPORTS_DIR / "suts"

for d in [UDS_OUT_DIR, STS_OUT_DIR, SUTS_OUT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

TS = datetime.now().strftime("%Y%m%d_%H%M%S")

def banner(title):
    print("\n" + "="*70)
    print(f"  {title}")
    print("="*70)

# ── UDS 생성 ─────────────────────────────────────────────────────────────────
banner("1/3  UDS 생성 시작")
t0 = time.time()

from generators.uds import generate_uds_payload
from helpers.uds import enrich_function_details_with_docs
from report_gen.docx_builder import generate_uds_docx

uds_out_path = UDS_OUT_DIR / f"uds_eval_{TS}.docx"
uds_payload_path = UDS_OUT_DIR / f"uds_eval_{TS}.payload.json"

print(f"  Source  : {SOURCE_ROOT}")
print(f"  SRS     : {SRS_PATH.name}")
print(f"  SDS     : {SDS_PATH.name}")
print(f"  HSIS    : {HSIS_PATH.name}")
print(f"  Output  : {uds_out_path.name}")

try:
    # UDS payload 생성
    uds_payload = generate_uds_payload(
        source_root=str(SOURCE_ROOT),
        req_doc_paths=[str(SRS_PATH), str(SDS_PATH)],
        sds_doc_paths=[str(SDS_PATH)],
    )

    # HSIS enrichment (backend/helpers/uds.py 또는 inline)
    try:
        from generators.sts import _load_hsis_signals
        import re
        hsis_data = _load_hsis_signals(str(HSIS_PATH))
        hsis_sigs = hsis_data.get("signals", [])
        hvar = {}
        for s in hsis_sigs:
            sw_raw = str(s.get("sw_var_name") or "")
            for tok in re.split(r"[\n,\s]+", sw_raw):
                tok = tok.strip()
                if tok and re.match(r"^[A-Za-z_]\w+$", tok):
                    hvar[tok] = s
        enriched = 0
        fn_map = uds_payload.get("function_details") or {}
        for fn_info in fn_map.values():
            if not isinstance(fn_info, dict): continue
            fvars = set()
            for x in (fn_info.get("inputs") or []):  fvars.add(str(x.get("name") or ""))
            for x in (fn_info.get("outputs") or []): fvars.add(str(x.get("name") or ""))
            fvars.update((fn_info.get("globals_write") or {}).keys())
            fvars.update((fn_info.get("globals_read")  or {}).keys())
            matched = [hvar[v] for v in fvars if v in hvar]
            if not matched: continue
            if fn_info.get("description_source", "inference") in {"inference", ""}:
                fn_info["description_source"] = "hsis"
            cur_rel = str(fn_info.get("related") or "").strip()
            if not cur_rel or cur_rel.upper() in {"TBD", "N/A", "-"}:
                rel_ids = [str(s.get("related_id") or "").strip() for s in matched if str(s.get("related_id") or "").strip()]
                if rel_ids:
                    fn_info["related"] = rel_ids[0]
                    fn_info["related_source"] = "hsis"
            enriched += 1
        print(f"  HSIS enrichment: {enriched}개 함수 보강 ({len(hsis_sigs)} 신호)")
    except Exception as he:
        print(f"  HSIS enrichment skipped: {he}")

    # DOCX 생성
    generate_uds_docx(
        template_path=None,
        uds_payload=uds_payload,
        output_path=str(uds_out_path),
    )

    # payload 저장
    with open(uds_payload_path, "w", encoding="utf-8") as f:
        json.dump(uds_payload, f, ensure_ascii=False, indent=2, default=str)

    uds_elapsed = time.time() - t0
    print(f"  완료: {uds_elapsed:.1f}s → {uds_out_path.name}")
    uds_ok = True
    uds_fn_count = len(fn_map)

except Exception as e:
    uds_elapsed = time.time() - t0
    print(f"  UDS 생성 실패: {e}")
    uds_ok = False
    uds_fn_count = 0
    uds_payload = {}

# ── STS 생성 ─────────────────────────────────────────────────────────────────
banner("2/3  STS 생성 시작")
t0 = time.time()

sts_out_path = STS_OUT_DIR / f"sts_eval_{TS}.xlsx"
sts_val_path = STS_OUT_DIR / f"sts_eval_{TS}.validation.md"

print(f"  Source  : {SOURCE_ROOT}")
print(f"  SRS     : {SRS_PATH.name}")
print(f"  SDS     : {SDS_PATH.name}")
print(f"  HSIS    : {HSIS_PATH.name}")
print(f"  Output  : {sts_out_path.name}")

try:
    from generators.sts import generate_sts

    sts_result = generate_sts(
        source_root=str(SOURCE_ROOT),
        srs_docx_path=str(SRS_PATH),
        sds_docx_path=str(SDS_PATH),
        hsis_path=str(HSIS_PATH),
        output_path=str(sts_out_path),
        project_config={
            "project_id": "HDPDM01",
            "doc_id": "HDPDM01-STS",
            "version": "v1.00",
            "asil_level": "ASIL-B",
        },
        max_tc_per_req=5,
        on_progress=lambda p, m: print(f"    [{p:3d}%] {m}"),
    )

    sts_elapsed = time.time() - t0
    sts_qr = sts_result.get("quality_report", {})
    print(f"  완료: {sts_elapsed:.1f}s")
    sts_ok = True

except Exception as e:
    sts_elapsed = time.time() - t0
    print(f"  STS 생성 실패: {e}")
    import traceback; traceback.print_exc()
    sts_ok = False
    sts_result = {}
    sts_qr = {}

# ── SUTS 생성 ─────────────────────────────────────────────────────────────────
banner("3/3  SUTS 생성 시작")
t0 = time.time()

suts_out_path = SUTS_OUT_DIR / f"suts_eval_{TS}.xlsx"

print(f"  Source  : {SOURCE_ROOT}")
print(f"  SRS     : {SRS_PATH.name}")
print(f"  SDS     : {SDS_PATH.name}")
print(f"  HSIS    : {HSIS_PATH.name}")
print(f"  Output  : {suts_out_path.name}")

try:
    from generators.suts import generate_suts

    suts_result = generate_suts(
        source_root=str(SOURCE_ROOT),
        srs_docx_path=str(SRS_PATH),
        sds_docx_path=str(SDS_PATH),
        hsis_path=str(HSIS_PATH),
        output_path=str(suts_out_path),
        project_config={
            "project_id": "HDPDM01",
            "doc_id": "HDPDM01-SUTS",
            "version": "v1.00",
            "asil_level": "ASIL-B",
        },
        max_sequences=6,
        on_progress=lambda p, m: print(f"    [{p:3d}%] {m}"),
    )

    suts_elapsed = time.time() - t0
    suts_qr = suts_result.get("quality_report", {})
    print(f"  완료: {suts_elapsed:.1f}s")
    suts_ok = True

except Exception as e:
    suts_elapsed = time.time() - t0
    print(f"  SUTS 생성 실패: {e}")
    import traceback; traceback.print_exc()
    suts_ok = False
    suts_result = {}
    suts_qr = {}

# ── 품질 평가 리포트 ───────────────────────────────────────────────────────────
banner("품질 평가 리포트")

# UDS 품질
fn_map = uds_payload.get("function_details") or {}
uds_desc_sources = {}
uds_asil_tbd = 0
uds_related_tbd = 0
for fi in fn_map.values():
    if not isinstance(fi, dict): continue
    src = fi.get("description_source", "inference")
    uds_desc_sources[src] = uds_desc_sources.get(src, 0) + 1
    if str(fi.get("asil") or "TBD").upper() in {"TBD", "", "N/A"}:
        uds_asil_tbd += 1
    if str(fi.get("related") or "TBD").upper() in {"TBD", "", "N/A", "-"}:
        uds_related_tbd += 1

SOURCE_SCORES = {"comment":1.0,"sds":0.95,"srs":0.9,"rule":0.8,"hsis":0.75,"reference":0.7,"inference":0.6}
total_fns = len(fn_map)
avg_src_score = 0.0
if total_fns > 0:
    avg_src_score = sum(SOURCE_SCORES.get(fi.get("description_source","inference"),0.6)
                        for fi in fn_map.values() if isinstance(fi, dict)) / total_fns

print(f"""
┌─────────────────────────────────────────────────────────────────────┐
│                         UDS 품질 평가                                │
├─────────────────────────────────────────────────────────────────────┤
│  총 함수 수     : {total_fns:>4}                                          │
│  ASIL TBD      : {uds_asil_tbd:>4} / {total_fns}  ({100*uds_asil_tbd/max(total_fns,1):.1f}%)                    │
│  Related TBD   : {uds_related_tbd:>4} / {total_fns}  ({100*uds_related_tbd/max(total_fns,1):.1f}%)                    │
│  설명 소스 평균 점수 : {avg_src_score:.3f} / 1.000                           │
├──────────────────────────────────────────────────────────────────── │
│  설명 소스 분포:""")
for src, cnt in sorted(uds_desc_sources.items(), key=lambda x:-x[1]):
    score = SOURCE_SCORES.get(src, 0.6)
    pct = 100*cnt/max(total_fns,1)
    print(f"│    {src:<12}: {cnt:>4}개 ({pct:5.1f}%)  점수={score:.2f}               │")
print(f"""\
└─────────────────────────────────────────────────────────────────────┘""")

# STS 품질
print(f"""
┌─────────────────────────────────────────────────────────────────────┐
│                         STS 품질 평가                                │
├─────────────────────────────────────────────────────────────────────┤""")
if sts_ok:
    total_tc = sts_qr.get("total_test_cases", 0)
    complete_tc = sts_qr.get("complete_test_cases", 0)
    complete_pct = sts_qr.get("completeness_pct", 0)
    coverage = sts_qr.get("requirement_coverage", {})
    cov_pct = coverage.get("pct", 0)
    methods = sts_qr.get("test_method_distribution", {})
    gen_methods = sts_qr.get("gen_method_distribution", {})
    print(f"""│  총 TC 수       : {total_tc:>4}                                          │
│  완전한 TC      : {complete_tc:>4} ({complete_pct:.1f}%)                              │
│  요구사항 커버리지: {cov_pct:.1f}%                                         │
│  경과 시간      : {sts_elapsed:.1f}s                                        │
├─────────────────────────────────────────────────────────────────────┤
│  테스트 메서드 분포:""")
    for m, c in sorted(methods.items(), key=lambda x:-x[1]):
        pct = 100*c/max(total_tc,1)
        print(f"│    {m:<8}: {c:>4}개 ({pct:5.1f}%)                                   │")
    print(f"│  생성 방법 분포:")
    for m, c in sorted(gen_methods.items(), key=lambda x:-x[1]):
        pct = 100*c/max(total_tc,1)
        print(f"│    {m:<8}: {c:>4}개 ({pct:5.1f}%)                                   │")

    # 단일 스텝 TC 직접 분석
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(sts_out_path), read_only=True, data_only=True)
        spec_sheet = None
        for sn in wb.sheetnames:
            if "Test Spec" in sn:
                spec_sheet = wb[sn]; break
        single_step_tcs = 0
        if spec_sheet:
            step_counts = {}
            cur_tc = None
            for row in spec_sheet.iter_rows(values_only=True):
                tc_id = str(row[0] or "").strip() if row else ""
                if tc_id.startswith("TC-"):
                    cur_tc = tc_id
                    step_counts[cur_tc] = 0
                elif cur_tc and tc_id:
                    step_counts[cur_tc] = step_counts.get(cur_tc, 0) + 1
            single_step_tcs = sum(1 for c in step_counts.values() if c <= 1)
        print(f"│  단일 스텝(≤1) TC  : {single_step_tcs:>4}개                                  │")
    except Exception:
        pass
else:
    print(f"│  생성 실패                                                          │")
print(f"└─────────────────────────────────────────────────────────────────────┘")

# SUTS 품질
print(f"""
┌─────────────────────────────────────────────────────────────────────┐
│                         SUTS 품질 평가                               │
├─────────────────────────────────────────────────────────────────────┤""")
if suts_ok:
    total_suts_tc = suts_qr.get("total_test_cases", 0)
    total_seq = suts_qr.get("total_sequences", 0)
    avg_seq = suts_qr.get("avg_sequences_per_tc", 0)
    with_io = suts_qr.get("with_io_count", 0)
    io_pct = suts_qr.get("io_coverage_pct", 0)
    fn_cov = suts_qr.get("function_coverage_pct", 0)
    gen_dist = suts_qr.get("gen_method_distribution", {})
    comp_dist = suts_qr.get("component_distribution", {})
    print(f"""│  총 TC 수       : {total_suts_tc:>4}  (소스 함수 {suts_qr.get('total_source_functions',0)}개)      │
│  총 시퀀스 수   : {total_seq:>4}                                          │
│  TC당 평균 시퀀스: {avg_seq:.1f}                                          │
│  I/O 보유 TC   : {with_io:>4} ({io_pct:.1f}%)                              │
│  함수 커버리지  : {fn_cov:.1f}%                                          │
│  경과 시간      : {suts_elapsed:.1f}s                                        │
├─────────────────────────────────────────────────────────────────────┤
│  생성 방법 분포:""")
    for m, c in sorted(gen_dist.items(), key=lambda x:-x[1]):
        pct = 100*c/max(total_suts_tc,1)
        print(f"│    {m:<14}: {c:>4}개 ({pct:5.1f}%)                           │")
    print(f"│  컴포넌트 분포:")
    for comp, c in sorted(comp_dist.items(), key=lambda x:-x[1]):
        print(f"│    {comp:<12}: {c:>4}개                                        │")

    # SRS req_id 채워진 TC 비율 확인
    srs_linked = 0
    try:
        import openpyxl
        wb2 = openpyxl.load_workbook(str(suts_out_path), read_only=True, data_only=True)
        spec_sh = None
        for sn in wb2.sheetnames:
            if "Unit Test Spec" in sn:
                spec_sh = wb2[sn]; break
        if spec_sh:
            headers = None
            srs_col = None
            for row in spec_sh.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c or "").strip() for c in row]
                    for i, h in enumerate(headers):
                        if "SRS" in h.upper() or "REQ" in h.upper():
                            srs_col = i; break
                    continue
                if srs_col is not None and row[srs_col] and str(row[srs_col]).strip() not in {"", "-", "N/A"}:
                    srs_linked += 1
    except Exception:
        pass
    if srs_linked > 0:
        print(f"│  SRS req_id 연결: {srs_linked:>4}개                                      │")
else:
    print(f"│  생성 실패                                                          │")
print(f"└─────────────────────────────────────────────────────────────────────┘")

# ── 종합 점수 계산 ────────────────────────────────────────────────────────────
banner("종합 품질 점수")

def score_uds(fn_count, asil_tbd, related_tbd, avg_src):
    if fn_count == 0: return 0.0
    asil_ok  = 1 - asil_tbd    / fn_count
    rel_ok   = 1 - related_tbd / fn_count
    # 가중치: 설명소스 40%, ASIL 30%, 관련 30%
    return round((avg_src * 0.4 + asil_ok * 0.3 + rel_ok * 0.3) * 10, 2)

def score_sts(complete_pct, cov_pct, method_cnt):
    comp = complete_pct / 100
    cov  = cov_pct / 100
    variety = min(method_cnt / 5, 1.0)   # 5종 이상이면 만점
    return round((comp * 0.5 + cov * 0.3 + variety * 0.2) * 10, 2)

def score_suts(io_pct, fn_cov, avg_seq):
    io   = io_pct / 100
    fc   = fn_cov / 100
    seq  = min(avg_seq / 6, 1.0)         # 6 이상이면 만점
    return round((io * 0.4 + fc * 0.3 + seq * 0.3) * 10, 2)

uds_score  = score_uds(total_fns, uds_asil_tbd, uds_related_tbd, avg_src_score) if uds_ok else 0.0
sts_qr_cp  = sts_qr.get("completeness_pct", 0)
sts_cov    = sts_qr.get("requirement_coverage", {}).get("pct", 0)
sts_methods= len(sts_qr.get("test_method_distribution", {}))
sts_score  = score_sts(sts_qr_cp, sts_cov, sts_methods) if sts_ok else 0.0
suts_io    = suts_qr.get("io_coverage_pct", 0)
suts_fc    = suts_qr.get("function_coverage_pct", 0)
suts_avg   = suts_qr.get("avg_sequences_per_tc", 0)
suts_score = score_suts(suts_io, suts_fc, suts_avg) if suts_ok else 0.0
total_score= round((uds_score + sts_score + suts_score) / 3, 2)

print(f"""
  ┌────────────────────────────────┐
  │  UDS  점수 :  {uds_score:5.2f} / 10.00  │
  │  STS  점수 :  {sts_score:5.2f} / 10.00  │
  │  SUTS 점수 :  {suts_score:5.2f} / 10.00  │
  │────────────────────────────────│
  │  종합 평균 :  {total_score:5.2f} / 10.00  │
  └────────────────────────────────┘

  출력 파일:
    UDS  → {uds_out_path.name if uds_ok else "FAILED"}
    STS  → {sts_out_path.name if sts_ok else "FAILED"}
    SUTS → {suts_out_path.name if suts_ok else "FAILED"}
""")
