"""
STS + SUTS 생성 + 품질 평가 스크립트
Python: C:/Users/kss11/AppData/Local/Programs/Python/Python312/python.exe
"""
import sys
import time
import json
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

DOCS_DIR    = ROOT / "docs"
SOURCE_ROOT = Path("D:/Project/Ados/PDS_64_RD")
REPORTS_DIR = ROOT / "reports"

SRS_PATH  = DOCS_DIR / "(HDPDM01_SRS) Software Requirements Specification_v1.05_20230510.docx"
SDS_PATH  = DOCS_DIR / "(HDPDM01_SDS) Software Architecture Design Specification_v1.04_20230512.docx"
HSIS_PATH = DOCS_DIR / "(HDPDM01_HSIS) Hardware Software Interface Specification_v5.00_20221115_B.xlsx"

STS_OUT_DIR  = REPORTS_DIR / "sts"
SUTS_OUT_DIR = REPORTS_DIR / "suts"
STS_OUT_DIR.mkdir(parents=True, exist_ok=True)
SUTS_OUT_DIR.mkdir(parents=True, exist_ok=True)

TS = datetime.now().strftime("%Y%m%d_%H%M%S")

def banner(t):
    print("\n" + "="*70)
    print(f"  {t}")
    print("="*70)

# ─── 소스 코드 파싱 ──────────────────────────────────────────────────────────
banner("소스 코드 파싱")
t0 = time.time()
print(f"  Source: {SOURCE_ROOT}")
try:
    from report_generator import generate_uds_source_sections, enrich_function_details_with_docs
    sections = generate_uds_source_sections(str(SOURCE_ROOT))
    function_details = sections.get("function_details", {})
    print(f"  파싱 완료: {len(function_details)}개 함수 ({time.time()-t0:.1f}s)")

    # 문서 보강
    t1 = time.time()
    from report_generator import _build_req_map_from_doc_paths
    req_doc_paths = [str(SRS_PATH), str(SDS_PATH)]
    sds_doc_paths = [str(SDS_PATH)]
    function_details = enrich_function_details_with_docs(
        function_details,
        req_doc_paths=req_doc_paths,
        sds_doc_paths=sds_doc_paths,
        function_table_rows=sections.get("function_table_rows", []),
    )
    print(f"  SRS/SDS 보강 완료 ({time.time()-t1:.1f}s)")

    # HSIS 보강
    t2 = time.time()
    try:
        import re
        from generators.sts import _load_hsis_signals
        hsis_data = _load_hsis_signals(str(HSIS_PATH))
        hsis_sigs = hsis_data.get("signals", [])
        hvar = {}
        for s in hsis_sigs:
            sw_raw = str(s.get("sw_var_name") or "")
            for tok in re.split(r"[\n,\s]+", sw_raw):
                tok = tok.strip()
                if tok and re.match(r"^[A-Za-z_]\w+$", tok):
                    hvar[tok] = s
        enriched_hsis = 0
        for fi in function_details.values():
            if not isinstance(fi, dict): continue
            fvars = set()
            for x in (fi.get("inputs") or []):  fvars.add(str(x.get("name") or ""))
            for x in (fi.get("outputs") or []): fvars.add(str(x.get("name") or ""))
            fvars.update((fi.get("globals_write") or {}).keys())
            fvars.update((fi.get("globals_read")  or {}).keys())
            matched = [hvar[v] for v in fvars if v in hvar]
            if not matched: continue
            if fi.get("description_source", "inference") in {"inference", ""}:
                fi["description_source"] = "hsis"
            cur_rel = str(fi.get("related") or "").strip()
            if not cur_rel or cur_rel.upper() in {"TBD", "N/A", "-"}:
                rel_ids = [str(s.get("related_id") or "").strip() for s in matched if str(s.get("related_id") or "").strip()]
                if rel_ids:
                    fi["related"] = rel_ids[0]
                    fi["related_source"] = "hsis"
            enriched_hsis += 1
        print(f"  HSIS 보강: {enriched_hsis}개 함수 ({len(hsis_sigs)} 신호, {time.time()-t2:.1f}s)")
    except Exception as he:
        print(f"  HSIS 보강 실패: {he}")

    parse_ok = True
except Exception as e:
    import traceback; traceback.print_exc()
    function_details = {}
    parse_ok = False

# ─── SRS 텍스트 로드 ─────────────────────────────────────────────────────────
req_texts = []
try:
    from generators.sts import parse_srs_docx_tables
    srs_reqs = parse_srs_docx_tables(str(SRS_PATH))
    req_texts = [f"{r.get('id','')} {r.get('description','')}" for r in srs_reqs if r.get('description')]
    print(f"  SRS 요구사항 로드: {len(req_texts)}개")
except Exception as e:
    print(f"  SRS 텍스트 로드 실패: {e}")

# ─── STS 생성 ────────────────────────────────────────────────────────────────
banner("1/2  STS 생성 (SRS + SDS + HSIS + 소스코드)")
t0 = time.time()

sts_out = STS_OUT_DIR / f"sts_eval_{TS}.xlsx"
print(f"  OUT : {sts_out.name}")

try:
    from generators.sts import generate_sts

    sts_result = generate_sts(
        requirements_text=req_texts,
        function_details=function_details,
        output_path=str(sts_out),
        template_path=None,
        project_config={
            "project_id": "HDPDM01",
            "doc_id": "HDPDM01-STS",
            "version": "v1.00",
            "asil_level": "ASIL-B",
            "max_tc_per_req": 5,
            "default_test_env": "SwTE_01",
        },
        srs_docx_path=str(SRS_PATH),
        sds_docx_path=str(SDS_PATH),
        hsis_path=str(HSIS_PATH),
        on_progress=lambda p, m: print(f"    [{p:3d}%] {m}"),
    )
    sts_elapsed = time.time() - t0
    sts_ok = True
    sts_qr = sts_result.get("quality_report", {})
    print(f"\n  완료: {sts_elapsed:.1f}s  TC={sts_qr.get('total_test_cases',0)}")
    with open(STS_OUT_DIR / f"sts_eval_{TS}.payload.json", "w", encoding="utf-8") as f:
        json.dump(sts_result, f, ensure_ascii=False, indent=2, default=str)
except Exception as e:
    sts_elapsed = time.time() - t0
    sts_ok = False
    sts_qr = {}
    import traceback; traceback.print_exc()

# ─── SUTS 생성 ───────────────────────────────────────────────────────────────
banner("2/2  SUTS 생성 (SRS + SDS + HSIS + 소스코드)")
t0 = time.time()

suts_out = SUTS_OUT_DIR / f"suts_eval_{TS}.xlsx"
print(f"  OUT : {suts_out.name}")

try:
    from generators.suts import generate_suts

    # Load AI config for void-function enhancement
    _suts_ai_cfg = None
    try:
        from workflow.ai import load_oai_config
        _suts_ai_cfg = load_oai_config(str(ROOT / "OAI_CONFIG_LIST"))
        if _suts_ai_cfg:
            print(f"  AI 설정 로드: {_suts_ai_cfg.get('model','?')}")
    except Exception as _ae:
        print(f"  AI 설정 로드 실패 (스킵): {_ae}")

    suts_result = generate_suts(
        source_root=str(SOURCE_ROOT),
        srs_docx_path=str(SRS_PATH),
        sds_docx_path=str(SDS_PATH),
        hsis_path=str(HSIS_PATH),
        output_path=str(suts_out),
        project_config={
            "project_id": "HDPDM01",
            "doc_id": "HDPDM01-SUTS",
            "version": "v1.00",
            "asil_level": "ASIL-B",
        },
        max_sequences=6,
        ai_config=_suts_ai_cfg,
        on_progress=lambda p, m: print(f"    [{p:3d}%] {m}"),
    )
    suts_elapsed = time.time() - t0
    suts_ok = True
    suts_qr = suts_result.get("quality_report", {})
    print(f"\n  완료: {suts_elapsed:.1f}s  TC={suts_qr.get('total_test_cases',0)}")
    with open(SUTS_OUT_DIR / f"suts_eval_{TS}.payload.json", "w", encoding="utf-8") as f:
        json.dump(suts_result, f, ensure_ascii=False, indent=2, default=str)
except Exception as e:
    suts_elapsed = time.time() - t0
    suts_ok = False
    suts_qr = {}
    import traceback; traceback.print_exc()

# ─── 품질 평가 ────────────────────────────────────────────────────────────────
banner("품질 평가 결과")

# UDS (기존 최신 결과 재사용)
print("\n>> UDS  [2026-03-18 기존 결과 - 전체 문서 적용]")
print("  총 함수       : 368")
print("  Description   : 100% High (comment/sds/reference)")
print("  ASIL TBD      : 0 / 368 (0.0%)")
print("  Related TBD   : 0 / 368 (0.0%)")
print("  Quality Gate  : 10/10 PASS")

# STS
print("\n▶ STS")
if sts_ok:
    total_tc    = sts_qr.get("total_test_cases", 0)
    complete_tc = sts_qr.get("complete_test_cases", 0)
    complete_pct= sts_qr.get("completeness_pct", 0.0)
    cov         = sts_qr.get("requirement_coverage", {})
    cov_pct     = cov.get("pct", 0.0)
    covered     = cov.get("covered_reqs", 0)
    total_req   = cov.get("total_reqs", 0)
    methods     = sts_qr.get("test_method_distribution", {})
    gen_methods = sts_qr.get("gen_method_distribution", {})

    print(f"  TC 수          : {total_tc}")
    print(f"  완전한 TC      : {complete_tc} ({complete_pct:.1f}%)")
    print(f"  요구사항 커버리지: {covered}/{total_req} ({cov_pct:.1f}%)")
    print(f"  테스트 메서드   : {dict(sorted(methods.items(), key=lambda x:-x[1]))}")
    print(f"  생성 방법       : {dict(sorted(gen_methods.items(), key=lambda x:-x[1]))}")
    print(f"  경과 시간       : {sts_elapsed:.1f}s")

    # 단일 스텝 TC
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(sts_out), read_only=True, data_only=True)
        spec_sh = next((wb[sn] for sn in wb.sheetnames if "Test Spec" in sn), None)
        if spec_sh:
            tc_steps = {}; cur = None
            for row in spec_sh.iter_rows(values_only=True):
                cell0 = str(row[0] or "").strip()
                if cell0.startswith("TC-"):
                    cur = cell0; tc_steps[cur] = 0
                elif cur and len(row) > 1 and row[1]:
                    tc_steps[cur] += 1
            single = sum(1 for s in tc_steps.values() if s <= 1)
            print(f"  단일 스텝(≤1) TC : {single}개 / {len(tc_steps)}개 ({100*single/max(len(tc_steps),1):.1f}%)")
    except Exception as e2:
        print(f"  (단일 스텝 분석 실패: {e2})")
else:
    print("  생성 실패")

# SUTS
print("\n▶ SUTS")
if suts_ok:
    total_suts = suts_qr.get("total_test_cases", 0)
    total_seq  = suts_qr.get("total_sequences", 0)
    avg_seq    = suts_qr.get("avg_sequences_per_tc", 0)
    with_io    = suts_qr.get("with_io_count", 0)
    io_pct     = suts_qr.get("io_coverage_pct", 0.0)
    fn_cov     = suts_qr.get("function_coverage_pct", 0.0)
    gen_dist   = suts_qr.get("gen_method_distribution", {})
    comp_dist  = suts_qr.get("component_distribution", {})

    print(f"  TC 수           : {total_suts}")
    print(f"  시퀀스 수       : {total_seq}")
    print(f"  TC당 평균 시퀀스 : {avg_seq:.1f}")
    print(f"  I/O 보유 TC     : {with_io} ({io_pct:.1f}%)")
    print(f"  함수 커버리지   : {fn_cov:.1f}%")
    print(f"  생성 방법       : {dict(sorted(gen_dist.items(), key=lambda x:-x[1]))}")
    print(f"  컴포넌트 분포   : {comp_dist}")
    print(f"  경과 시간       : {suts_elapsed:.1f}s")
else:
    print("  생성 실패")

# ─── 종합 점수 ────────────────────────────────────────────────────────────────
banner("종합 품질 점수")

# UDS 고정
uds_score = 9.3

# STS 점수
if sts_ok:
    cp  = sts_qr.get("completeness_pct", 0) / 100
    cv  = sts_qr.get("requirement_coverage", {}).get("pct", 0) / 100
    nm  = len(sts_qr.get("test_method_distribution", {}))
    sts_score = round((cp * 0.5 + cv * 0.3 + min(nm / 5, 1.0) * 0.2) * 10, 2)
else:
    sts_score = 0.0

# SUTS 점수
if suts_ok:
    ip  = suts_qr.get("io_coverage_pct", 0) / 100
    fc  = suts_qr.get("function_coverage_pct", 0) / 100
    av  = suts_qr.get("avg_sequences_per_tc", 0)
    suts_score = round((ip * 0.4 + fc * 0.3 + min(av / 6, 1.0) * 0.3) * 10, 2)
else:
    suts_score = 0.0

total = round((uds_score + sts_score + suts_score) / 3, 2)

print(f"""
  ┌───────────────────────────────────────┐
  │  UDS  : {uds_score:5.2f} / 10  (기존 결과 재사용)  │
  │  STS  : {sts_score:5.2f} / 10                   │
  │  SUTS : {suts_score:5.2f} / 10                   │
  │───────────────────────────────────────│
  │  평균 : {total:5.2f} / 10                   │
  └───────────────────────────────────────┘
""")
