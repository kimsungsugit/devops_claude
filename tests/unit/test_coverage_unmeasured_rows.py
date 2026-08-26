"""미측정(Total=0) 행이 커버리지를 오염시키지 않는가 — §6 후보 8 / UTCV-001.

## 실측한 결함 (2026-08-04)

정상 1함수(10/10) + `total=0` 2함수를 넣고 `build_coverage_report` 를 돌리면:

    r5(Statement)  N=3  Fail=0  Exc=2  Coverage=**1.6666…**   ← 166.67%

원인은 **fail 판정이 두 곳에서 다르게** 되기 때문이다:

    데이터행 Exception 스탬프  `not fc.statement.passed`   ← total=0 이면 True(미측정)
    요약 Fail 카운트           `f != g`                    ← 0 == 0 이라 False

미측정 행이 Fail 로는 안 세어지는데 Exception 으로는 세어져 **분자에만** 더해진다.
요약식이 `(Total - Fail + Exception) / Total` 이라 100% 를 넘는다.

## 왜 "100% 로 되돌리기" 가 오답인가

Exception 스탬프만 막으면 `(3 - 0 + 0)/3 = 100%` 가 된다. 3함수 중 1함수만 측정했는데
100% 라고 말하는 것이고, 166.67% 라는 **불가능한 값이 유일하게 시끄러운 신호**였는데
그걸 조용한 정상값으로 바꾸는 셈이다. 이 저장소가 반복해 겪은 fake-green 이다.

그래서 둘 다 한다:
  1. 미측정 행에는 Exception 'O' 를 찍지 않는다 (면제 결정이 아니라 판단 이전 상태다)
  2. 미측정 행을 **분모에서 뺀다** + 뺐다는 사실을 `out_warnings` 로 보고한다
  3. 실측 행이 0이면 숫자가 아니라 `"미측정"` — 0% 도 100% 도 아니다

## 안 건드린 것

PV 자동 Exception 정책(`_pv_auto_exception`, `:994` 사용자 결정 "PV 기준 유지")은
그대로다. 이 라운드가 가른 것은 **미측정 vs 미달**이지 면제 정책이 아니다.
`applicable` 정책축 신설은 **미착수 확정** — 저장소에 해당 축이 없고, 있는 것은
`exception`(면제)·`not_measured`·ASIL threshold 분리다. disposition 워크플로는
신규 기능이라 결함 수정과 분리한다.
"""
from __future__ import annotations

import pytest

from tests.unit._source_probe import source_of

pytest.importorskip("openpyxl")

from backend.services.swut_coverage_aggregator import _write_spec_totals  # noqa: E402


def _new_sheet():
    """**실제 openpyxl 워크시트**를 쓴다.

    ⚠ 처음엔 최소 흉내 클래스로 짰는데 `_write_spec_totals` 가 셀 `font`·`border` 를
    읽어 전부 AttributeError 로 죽었다. 흉내를 늘리면 그만큼 라이브와 벌어지므로
    ([[reference_sim_harness_live_parity]]) 진짜 시트를 쓴다. 판정 로직은 복제하지
    않는다 — 실제 함수를 태우고 셀 값만 관찰한다.
    """
    from openpyxl import Workbook

    return Workbook().active


def _set(ws, row, col, value):
    ws.cell(row=row, column=col).value = value


def _get(ws, row, col):
    return ws.cell(row=row, column=col).value


# 컬럼 배치 — PV 10열 기준(component 열 없음)
UNIT_ID_COL = 4     # D
NAME_COL = 5        # E
STMT_COUNT_COL = 6  # F  (G=total, H=pass, I=exception)
BRANCH_COUNT_COL = 10  # J (K=total, L=pass, M=exception)
DATA_START = 13


def _sheet_with(rows):
    """rows = [(name, stmt_cov, stmt_tot, stmt_exc, br_cov, br_tot, br_exc), …]"""
    ws = _new_sheet()
    for i, (name, sc, st, se, bc, bt, be) in enumerate(rows):
        r = DATA_START + i
        _set(ws, r, UNIT_ID_COL, f"SwUFn_{i:04d}")
        _set(ws, r, NAME_COL, name)
        _set(ws, r, STMT_COUNT_COL, sc)
        _set(ws, r, STMT_COUNT_COL + 1, st)
        if se:
            _set(ws, r, STMT_COUNT_COL + 3, "O")
        _set(ws, r, BRANCH_COUNT_COL, bc)
        _set(ws, r, BRANCH_COUNT_COL + 1, bt)
        if be:
            _set(ws, r, BRANCH_COUNT_COL + 3, "O")
    return ws, DATA_START + len(rows) - 1


def _run(rows, *, stats=None):
    """기존 호출부 호환을 위해 반환은 `(ws, warnings)` 2-tuple 그대로 둔다.

    `stats` 를 주면 `out_stats` 로 전달돼 그 dict 가 채워진다(2026-08-26 신설).
    """
    ws, last = _sheet_with(rows)
    warnings: list[str] = []
    _write_spec_totals(
        ws, data_start=DATA_START, last_data_row=last,
        unit_id_col=UNIT_ID_COL, no_col=UNIT_ID_COL - 1,
        stmt_count_col=STMT_COUNT_COL, branch_count_col=BRANCH_COUNT_COL,
        out_warnings=warnings, out_stats=stats,
    )
    return ws, warnings


class TestImpossibleCoverageIsGone:
    def test_unmeasured_rows_no_longer_exceed_100_percent(self):
        """재현 케이스 — 정상 1 + 미측정 2. 예전엔 166.67% 였다."""
        ws, _ = _run([
            ("alpha", 10, 10, False, 10, 10, False),
            ("beta", 0, 0, False, 0, 0, False),
            ("gamma", 0, 0, False, 0, 0, False),
        ])
        cov = _get(ws, 5, 8)
        assert isinstance(cov, (int, float)), f"커버리지가 숫자가 아니다: {cov!r}"
        assert cov <= 1.0, f"커버리지가 100% 를 넘는다: {cov}"

    def test_unmeasured_rows_are_excluded_from_the_denominator(self):
        """분모는 실측 행 수 — 1함수만 측정했으면 그 1함수 기준이다."""
        ws, _ = _run([
            ("alpha", 10, 10, False, 10, 10, False),
            ("beta", 0, 0, False, 0, 0, False),
            ("gamma", 0, 0, False, 0, 0, False),
        ])
        assert _get(ws, 5, 8) == pytest.approx(1.0)
        # 전체 행 수(E)는 감사본 정합을 위해 그대로 3
        assert _get(ws, 5, 5) == 3

    def test_exclusion_is_reported_not_silent(self):
        """뺐다는 사실이 경고로 나가야 한다 — 이게 없으면 조용한 100% 다."""
        _ws, warnings = _run([
            ("alpha", 10, 10, False, 10, 10, False),
            ("beta", 0, 0, False, 0, 0, False),
            ("gamma", 0, 0, False, 0, 0, False),
        ])
        hits = [w for w in warnings if "미측정" in w]
        assert hits, f"미측정 제외가 보고되지 않았다: {warnings}"
        assert "Statement 2" in hits[0], f"미측정 건수가 경고에 없다: {hits[0]}"

    def test_all_unmeasured_is_not_100_percent(self):
        """전부 미측정이면 0% 도 100% 도 아니라 **판정 불가**다.

        ⚠ 이 대조군이 핵심이다. 산술만 정규화하면 이 케이스가 조용히 `1`(=100%)이 되고,
          그게 바로 '측정 실패를 정상으로 오독' 이다.
        """
        ws, warnings = _run([
            ("alpha", 0, 0, False, 0, 0, False),
            ("beta", 0, 0, False, 0, 0, False),
        ])
        assert _get(ws, 5, 8) == "미측정", f"전부 미측정인데 {_get(ws, 5, 8)!r} 을 썼다"
        assert _get(ws, 6, 8) == "미측정"
        assert any("미측정" in w for w in warnings)


class TestMeasuredRowsBehaveAsBefore:
    """대조군 — 미측정이 없으면 기존 동작과 같아야 한다(감사본 정합)."""

    def test_all_pass_is_100_percent(self):
        ws, warnings = _run([
            ("alpha", 10, 10, False, 8, 8, False),
            ("beta", 5, 5, False, 3, 3, False),
        ])
        assert _get(ws, 5, 8) == pytest.approx(1.0)
        assert _get(ws, 6, 8) == pytest.approx(1.0)
        assert not [w for w in warnings if "미측정" in w]

    def test_pv_exception_offset_still_reaches_100(self):
        """PV 정책(미달행 Exception 'O' 로 상쇄 → 100%)은 **그대로다**.

        `:994` 의 사용자 결정("PV 기준 유지")을 건드리지 않았음을 값으로 고정한다.
        """
        ws, _ = _run([
            ("alpha", 10, 10, False, 10, 10, False),
            ("beta", 3, 5, True, 2, 4, True),      # 미달 + 면제
        ])
        assert _get(ws, 5, 6) == 1, "Fail 카운트가 바뀌었다"
        assert _get(ws, 5, 7) == 1, "Exception 카운트가 바뀌었다"
        assert _get(ws, 5, 8) == pytest.approx(1.0), "PV 상쇄가 깨졌다"

    def test_real_shortfall_is_not_hidden(self):
        """면제 없는 미달은 100% 가 아니어야 한다."""
        ws, _ = _run([
            ("alpha", 10, 10, False, 10, 10, False),
            ("beta", 3, 5, False, 2, 4, False),
        ])
        assert _get(ws, 5, 8) == pytest.approx(0.5)


class TestUnmeasuredIsNotStampedAsException:
    """데이터행 쪽 — 미측정을 '면제 결정' 으로 기록하지 않는다."""

    def test_stamp_site_distinguishes_unmeasured(self):
        from backend.services import swut_coverage_aggregator as mod

        source = source_of(mod)
        assert "_stmt_unmeasured = int(fc.statement.total or 0) <= 0" in source, (
            "미측정 판정이 사라졌다 — `not passed` 만으로 Exception 을 찍으면 "
            "한 번도 재지 않은 행이 '면제 결정된 행' 으로 기록된다"
        )
        assert "if not fc.statement.passed and not _stmt_unmeasured:" in source
        assert "if not fc.branch.passed and not _branch_unmeasured:" in source

    def test_pv_auto_exception_policy_is_untouched(self):
        """면제 **정책** 자체는 이 라운드 범위가 아니다 — 사용자 결정."""
        from backend.services import swut_coverage_aggregator as mod

        source = source_of(mod)
        assert "_pv_auto_exception = not has_component_col" in source, (
            "PV 자동 예외 정책이 바뀌었다 — `:994` 사용자 결정('PV 기준 유지')을 "
            "건드리려면 별도 판단이 필요하다"
        )


class TestDocumentValueReachesTheGate:
    """문서에 찍은 값이 **호출부로 돌아오는가** — 2026-08-26.

    이 양식은 미달 행의 Exception 을 'O' 로 상쇄해 요약을 100% 로 만든다(PV 정책,
    `:994` 사용자 결정). 게이트는 raw `covered/total` 로 채점하므로 두 숫자가 갈린다.
    실측(KJPDS02 PV): 게이트 99.45% ↔ 문서 100%. 예전엔 그 격차가 summary 에 없어
    **어디에도 안 보였다** — 감사자가 문서만 보면 "달성", 화면만 보면 "미달"이다.

    ⚠ 숫자를 게이트 쪽에서 다시 세면 판정이 두 벌이 되어 드리프트한다(이 저장소가
      `_load_workbook_summary` 에서 겪은 그 결함). 라이터의 스캔 결과를 그대로 넘긴다.
    """

    def test_gap_between_document_and_raw_is_visible(self):
        stats: dict = {}
        ws, _ = _run([
            ("ok", 10, 10, False, 5, 5, False),
            ("ng", 8, 10, True, 4, 5, True),      # 미달 + 면제
        ], stats=stats)
        assert stats["coverage_fail_statement_functions"] == 1
        assert stats["coverage_exception_statement_functions"] == 1
        assert stats["doc_reported_statement_pct"] == 100.0     # (2-1+1)/2
        # 문서 셀과 **같은 값**이어야 한다 — 두 벌이면 여기서 갈린다.
        assert _get(ws, 5, 6) == 1 and _get(ws, 5, 7) == 1

    def test_unexcused_shortfall_is_not_offset(self):
        """면제가 없으면 문서도 100% 라고 쓰지 않는다(과잉 상쇄 방지)."""
        stats: dict = {}
        _run([
            ("ok", 10, 10, False, 5, 5, False),
            ("ng", 8, 10, False, 4, 5, False),
        ], stats=stats)
        assert stats["coverage_exception_statement_functions"] == 0
        assert stats["doc_reported_statement_pct"] == 50.0       # (2-1+0)/2

    def test_unmeasured_rows_reported_and_excluded(self):
        stats: dict = {}
        _run([
            ("ok", 10, 10, False, 5, 5, False),
            ("none", 0, 0, False, 0, 0, False),
        ], stats=stats)
        assert stats["coverage_unmeasured_statement_rows"] == 1
        assert stats["doc_reported_statement_pct"] == 100.0      # 미측정은 분모에서 뺀다

    def test_all_unmeasured_writes_no_number(self):
        """실측 0행이면 문서도 숫자를 안 쓴다 — stats 도 0 을 지어내지 않는다."""
        stats: dict = {}
        ws, _ = _run([("none", 0, 0, False, 0, 0, False)], stats=stats)
        assert "doc_reported_statement_pct" not in stats
        assert _get(ws, 5, 8) == "미측정"

    def test_out_stats_is_optional(self):
        """기존 호출부(미전달)가 깨지지 않는다."""
        ws, _ = _run([("ok", 10, 10, False, 5, 5, False)])
        assert _get(ws, 5, 5) == 1


class TestStatsWiringSurvivesTheChain:
    """`out_stats` 가 **빌더 → 시트 라이터 → TOTALS** 전 구간을 통과하는가.

    ⚠ 2026-08-26 실측 사고: 위 `TestDocumentValueReachesTheGate` 는 `_write_spec_totals`
      를 **직접** 불러서 통과했지만, 라이브에서는 값이 quality DB 에 하나도 안 남았다.
      링크 중 하나만 끊겨도 말단 테스트는 전부 초록이다(뮤테이션 E1 이 실제로 생존했다).
      그래서 여기서는 함수 안이 아니라 **함수 사이**를 잰다.
    """

    def test_sheet_writer_forwards_out_stats(self, monkeypatch):
        """`_write_coverage_sheet` → `_write_spec_totals` 링크."""
        import backend.services.swut_coverage_aggregator as mod

        seen: dict = {}

        def _spy(ws, **kw):
            seen.update(kw)
            if kw.get("out_stats") is not None:
                kw["out_stats"]["_probe"] = 1

        monkeypatch.setattr(mod, "_write_spec_totals", _spy)
        ws, agg = _spec_based_sheet_and_agg()
        stats: dict = {}
        mod._write_coverage_sheet(ws, agg, out_stats=stats)
        assert seen, "_write_spec_totals 가 호출되지 않았다 — spec_based 분기 확인"
        assert stats.get("_probe") == 1, "out_stats 가 TOTALS 로 전달되지 않는다"

    def test_builder_merges_stats_into_summary(self, monkeypatch):
        """`build_coverage_report` → 시트 라이터 → `summary` 링크.

        빌더가 dict 를 만들어 넘기고 **그 결과를 summary 에 합치는지**만 본다.
        (라이터 내부는 위 테스트가 따로 잰다.)
        """
        import backend.services.swut_coverage_aggregator as mod

        def _spy(ws, agg, **kw):
            assert "out_stats" in kw, "빌더가 out_stats 를 넘기지 않는다"
            assert isinstance(kw["out_stats"], dict)
            kw["out_stats"]["doc_reported_statement_pct"] = 42.0
            return 3

        monkeypatch.setattr(mod, "_write_coverage_sheet", _spy)
        session, meta, template = _minimal_build_inputs()
        result = mod.build_coverage_report(session, meta, template)
        assert result.ok
        assert result.summary.get("doc_reported_statement_pct") == 42.0, (
            "라이터가 채운 값이 summary 에 합쳐지지 않는다 — quality DB 에 안 남는다"
        )


def _spec_based_sheet_and_agg():
    """`spec_based` 가 켜지는 최소 조건 — Component 헤더 + SwUDS 이름→ID 맵."""
    from backend.services.swut_input_adapter import CoverageStats, FunctionCoverage

    ws = _new_sheet()
    # header_row 탐지용: 'No' 와 'Component' 가 같은 행에 있어야 한다.
    ws.cell(row=1, column=2).value = "No"
    ws.cell(row=1, column=3).value = "Component"
    ws.cell(row=1, column=4).value = "Unit ID"
    ws.cell(row=1, column=5).value = "Name"
    fc = FunctionCoverage(
        unit_id="SwUFn_0001", name="fn_a",
        statement=CoverageStats(covered=10, total=10),
        branch=CoverageStats(covered=5, total=5),
    )
    agg = {
        "function_rows": [fc],
        "function_name_to_swufn_from_suds": {"fn_a": "SwUFn_0001"},
    }
    return ws, agg


def _minimal_build_inputs():
    """빌더를 태울 최소 입력 — 기존 스위트의 픽스처를 재사용한다."""
    from tests.unit.test_swut_aggregators import (
        CoverageBuildMeta,
        _build_coverage_template,
        _make_session,
    )

    return (
        _make_session(),
        CoverageBuildMeta(release_sw_version="1.0.0", test_date="2024-02-19"),
        _build_coverage_template(),
    )
