"""SwSA aggregator 단위테스트 — 실 템플릿 빌드 + graceful degradation."""
from __future__ import annotations

import io
import os

import pytest
from openpyxl import Workbook, load_workbook

from backend.services.swsa_aggregator import build_swsa_report
from backend.services.swsa_meta import SwsaBuildMeta
from backend.services.swsa_qac_xml_parser import parse_qac_results_xml

_S = os.path.join(os.path.dirname(__file__), "..", "..", ".codex_tmp", "swsa_samples")
_TEMPLATE = os.path.join(_S, "TEMPLATE_(XXXX_SwSA) Software Static Analysis Report_v0.10_2XXXXX.xlsm")
_XML = os.path.join(_S, "XML_APP_NE1aW.xml")
_HMR = os.path.join(_S, "LOG_QAC_NE1aW_01_HMR_27052026_183745.html")
_PMD = os.path.join(_S, "LOG_NE1AW_PORTING_2631_PMD_Report_20260323.txt")
_DV = os.path.join(_S, "REFERENCE_(KJPDS02_DV_SwSA) Software Static Analysis Report_v1.02_260324_R.xlsm")


def _meta():
    return SwsaBuildMeta(
        project_id="KJPDS02", asil_level="ASIL A",
        doc_id_base="HKY-KJPDS02_PV-SwSA", doc_id_sequence="2884",
        doc_version="v0.11", doc_status="In Review", test_date="2026.04.24",
        test_engineer="김진경", release_sw_version="2631.00",
        phase="PV", platform_version="(APP) 2631.00 / (BOOT) 1.13",
        product="PDS", verification_target="MCU",
        compiler="CodeWarrior HC12Z", mcu="MC9S12ZVLA128MLF",
        analysis_round="1", debugger="이재원/유영규",
    )


def _is_yellow(cell) -> bool:
    """USER_INPUT_FILL_RGB(노란) 배경 여부."""
    try:
        return bool(cell.fill and cell.fill.fill_type) and cell.fill.fgColor.rgb == "FFFFEB9C"
    except (AttributeError, TypeError):
        return False


def _synth_template_bytes() -> bytes:
    """라벨/병합을 갖춘 최소 합성 템플릿(.xlsx, vba 없음) — graceful 테스트용."""
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["C26"] = "Document ID"
    cover["C30"] = "Author"
    st = wb.create_sheet("1.ST101")
    for r, lab in [(4, "분석차수"), (5, "SW Ver."), (6, "Tester"), (7, "Debugger")]:
        st.cell(r, 2).value = lab
    # Test Summary 표 헤더(76) — v0.10 형
    st.cell(76, 2).value = "위반 룰 개수"
    st.cell(76, 4).value = "총 위반 건수"
    st.cell(76, 6).value = "예외 처리 항목 수"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class TestSyntheticGraceful:
    def test_build_fills_and_marks(self):
        tpl = _synth_template_bytes()
        qac = parse_qac_results_xml(_XML)
        res = build_swsa_report(tpl, _meta(), qac_xml=qac)
        out = load_workbook(io.BytesIO(res.xlsm_io.getvalue()))
        st = out["1.ST101"]
        # Test-Info 기입 (LAYOUT-A: label B → value C)
        assert st["C4"].value == "1"
        assert st["C6"].value == "김진경"
        # 총 위반 건수(D77) = MISRA active=286, 위반 룰 개수(B77)=15
        assert st["D77"].value == 286
        assert st["B77"].value == 15
        # 예외 처리(F77): 수식 operand 보존 — 텍스트 없이 노란 배경만 (C1 fix)
        assert st["F77"].value is None
        assert _is_yellow(st["F77"])
        assert res.filled_cells > 0

    def test_extraction_failed_marks_yellow(self):
        # qac_xml=None → ST101 위반 셀 노란 표시 (0 stamp 금지)
        tpl = _synth_template_bytes()
        res = build_swsa_report(tpl, _meta(), qac_xml=None)
        out = load_workbook(io.BytesIO(res.xlsm_io.getvalue()))
        st = out["1.ST101"]
        # 추출 실패: D77(수식 operand) 텍스트 없이 노란 배경만 (#VALUE! 방지)
        assert st["D77"].value is None
        assert _is_yellow(st["D77"])
        assert res.user_input_cells >= 1

    def test_corrupt_xml_extraction_failed_marks(self):
        tpl = _synth_template_bytes()
        bad = parse_qac_results_xml("<AnalysisData><broken")
        assert bad.extraction_failed is True
        res = build_swsa_report(tpl, _meta(), qac_xml=bad)
        out = load_workbook(io.BytesIO(res.xlsm_io.getvalue()))
        assert _is_yellow(out["1.ST101"]["D77"])

    def test_formula_cell_preserved(self):
        # C2: v0.11 처럼 D77 이 수식이면 literal 로 덮지 않고 보존 + 경고
        wb = Workbook()
        st = wb.create_sheet("1.ST101")
        for r, lab in [(4, "분석차수"), (5, "SW Ver."), (6, "Tester"), (7, "Debugger")]:
            st.cell(r, 2).value = lab
        st.cell(76, 4).value = "총 위반 건수"
        st.cell(77, 4).value = "=COUNTIF(L86:L102,\">0\")"  # 수식
        bio = io.BytesIO(); wb.save(bio)
        qac = parse_qac_results_xml(_XML)
        res = build_swsa_report(bio.getvalue(), _meta(), qac_xml=qac)
        out = load_workbook(io.BytesIO(res.xlsm_io.getvalue()))
        # 수식 보존 (286 literal 로 덮이지 않음)
        assert out["1.ST101"]["D77"].value == "=COUNTIF(L86:L102,\">0\")"
        assert any("수식 셀 보존" in w for w in res.warnings)


@pytest.mark.skipif(not os.path.exists(_TEMPLATE) or not os.path.exists(_XML),
                    reason="실 템플릿/XML 샘플 없음")
class TestRealTemplateBuild:
    def setup_method(self):
        with open(_TEMPLATE, "rb") as f:
            self.tpl = f.read()
        self.qac = parse_qac_results_xml(_XML)
        self.res = build_swsa_report(self.tpl, _meta(), qac_xml=self.qac)
        self.out = load_workbook(io.BytesIO(self.res.xlsm_io.getvalue()), keep_vba=True)

    def test_cover_meta_matches_reference(self):
        cov = self.out["Cover"]
        assert cov["G26"].value == "HKY-KJPDS02_PV-SwSA-2884"
        assert cov["G27"].value == "v0.11"
        assert cov["G29"].value == "2026.04.24"
        assert cov["G30"].value == "김진경"  # 사인오프 I2 충돌 회피 검증

    def test_summary_header(self):
        s = self.out["Summary"]
        assert s["E3"].value == "KJPDS02"
        assert s["E8"].value == "A"            # 'ASIL A' → 'A'
        assert s["E10"].value == "MC9S12ZVLA128MLF"  # MCU 라벨 충돌 회피 검증

    def test_st101_violations(self):
        st = self.out["1.ST101"]
        assert st["C4"].value == "1"
        assert st["D77"].value == 286   # MISRA active
        assert st["B77"].value == 15    # distinct rules
        assert st["F77"].value is None and _is_yellow(st["F77"])  # 예외처리 노란(텍스트X)

    def test_integrity_preserved(self):
        tpl_wb = load_workbook(_TEMPLATE, keep_vba=True)
        # 시트 수 / 머지 / vba 보존
        assert tpl_wb.sheetnames == self.out.sheetnames
        for sn in ("Cover", "1.ST101", "2.ST201"):
            assert len(tpl_wb[sn].merged_cells.ranges) == len(self.out[sn].merged_cells.ranges)
        assert bool(self.out.vba_archive) is True
        assert self.res.vba_preserved is True

    def test_st1101_graceful_skip(self):
        # v0.10 템플릿엔 ST1101 없음 → graceful 경고
        assert any("11.ST1101 시트 없음" in w for w in self.res.warnings)

    def test_history_filled(self):
        # History 최신 행 기입 (이전 미채움 갭)
        h = self.out["History"]
        assert "History" in self.res.sheets_filled
        assert h["B5"].value == "v0.11"       # doc_version
        assert h["C5"].value == "2026.04.24"  # test_date
        assert h["E5"].value == "김진경"        # author
        assert _is_yellow(h["F5"])            # reviewer 빈 → 노란


@pytest.mark.skipif(
    not (os.path.exists(_TEMPLATE) and os.path.exists(_HMR) and os.path.exists(_PMD)),
    reason="실 템플릿/HMR/PMD 샘플 없음",
)
class TestST201MetricTable:
    """ST201 Test Summary Report 템플릿 주도 밴드 채우기."""

    def setup_method(self):
        from backend.services.swsa_pmd_parser import parse_pmd_cpd
        from backend.services.swsa_st201_binner import parse_st201_from_hmr
        with open(_TEMPLATE, "rb") as f:
            tpl = f.read()
        st201 = parse_st201_from_hmr(_HMR)
        pmd = parse_pmd_cpd(_PMD)
        res = build_swsa_report(tpl, _meta(), st201=st201, pmd=pmd)
        self.ws = load_workbook(io.BytesIO(res.xlsm_io.getvalue()), keep_vba=True)["2.ST201"]
        self.res = res

    def test_cyclomatic_pass_band(self):
        # 878 함수 전부 v(G) 1~10 → F63=878, 초과밴드 0
        assert self.ws["F63"].value == 878
        assert self.ws["F64"].value == 0

    def test_calling_called_pass_band(self):
        assert self.ws["F70"].value == 878   # Calling 0~5
        assert self.ws["F73"].value == 878   # Called 0~7

    def test_duplicated_from_pmd(self):
        # PMD 21블록 → 10~49=18, >=50=3
        assert self.ws["F79"].value == 18
        assert self.ws["F80"].value == 3

    def test_recursion_no_source_skipped(self):
        # Recursion(STNRA) HMR 부재 → F 미기입 + 경고
        assert self.ws["F77"].value is None
        assert any("무소스" in w and "Recursion" in w for w in self.res.warnings)

    def test_nesting_zero_absorbed_in_band1(self):
        # rank1: nesting=0 함수도 첫 밴드(1~10)에 흡수 → F67=전체(878), '밴드 밖' 경고 없음
        assert self.ws["F67"].value == 878
        assert not any("밴드 밖" in w for w in self.res.warnings)


@pytest.mark.skipif(not (os.path.exists(_TEMPLATE) and os.path.exists(_DV) and os.path.exists(_XML)),
                    reason="실 템플릿/DV/XML 샘플 없음")
class TestVariantSignoffDeviation:
    def test_dv_variant_history_append(self):
        # rank7: DV(기존 v0.10/v1.01/v1.02 3행)에 덮어쓰지 않고 append
        with open(_DV, "rb") as f:
            dv = f.read()
        meta = SwsaBuildMeta(project_id="KJPDS02", release_sw_version="1.03",
                             test_date="2026.06.05", doc_version="v1.03", test_engineer="김진경")
        res = build_swsa_report(dv, meta, qac_xml=parse_qac_results_xml(_XML))
        h = load_workbook(io.BytesIO(res.xlsm_io.getvalue()), keep_vba=True)["History"]
        versions = [h.cell(r, 2).value for r in range(5, 10)]
        assert "v1.02" in versions   # 기존 이력 보존
        assert "v1.03" in versions   # 신규 append

    def test_cover_signoff_and_deviation(self):
        with open(_TEMPLATE, "rb") as f:
            tpl = f.read()
        meta = SwsaBuildMeta(project_id="KJPDS02", release_sw_version="2631.00",
                             test_date="2026.04.24", reviewer_override="박검토",
                             approver_override="이승인")
        res = build_swsa_report(tpl, meta, qac_xml=parse_qac_results_xml(_XML))
        cov = load_workbook(io.BytesIO(res.xlsm_io.getvalue()), keep_vba=True)["Cover"]
        # rank8 사인오프
        assert cov["J3"].value == "박검토"
        assert cov["K3"].value == "이승인"
        # rank5 deviation 경고 (제외 997/1283 = 78%)
        assert any("deviation" in w and "78%" in w for w in res.warnings)
