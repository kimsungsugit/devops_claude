"""34차 SwIT SITR (Software Integration Test Result) v2.02 xlsm 빌더 회귀.

SwUT SUTR 회귀 (17차 + 31차 W27)에서 검증된 시트 writer 5개를 import 재활용.
본 회귀는 SwIT 도구별 차이 (파일명 / doc_id_base / Result dataclass) + smoke
+ keep_vba 보존 + ASIL Test Log col+4/5 대칭 검증.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.services.swit_meta import SwitSitrBuildMeta  # noqa: E402
from backend.services.swit_sitr_aggregator import (  # noqa: E402
    SwitSitrBuildResult,
    build_swit_sitr_report,
)
from backend.services.swut_input_adapter import (  # noqa: E402
    CoverageStats,
    EnvironmentData,
    ExecutionRow,
    FunctionCoverage,
    SwUTSession,
)


def _build_swit_sitr_template() -> bytes:
    """SwIT SITR v2.02 빈 양식 — SwUT SUTR template과 동일 구조 가정.

    34-fix 라운드에서 실 양식과 다르면 incomplete_sheets로 표시.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cover = wb.create_sheet("Cover")
    cover["B1"] = "Project"
    cover["B2"] = "ASIL Level"
    cover["B3"] = "Author"
    cover["B4"] = "Version"

    ts = wb.create_sheet("Test Summary")
    ts["B1"] = "Project Name"
    ts["B2"] = "Release Name(SW)"
    ts["B3"] = "Test Target Version(HW)"
    ts["B4"] = "Test Date"
    ts["B5"] = "Test Engineer"
    ts["B6"] = "Target Coverage"
    ts["B7"] = "Actual Coverage"
    ts["B8"] = "Final Test Result"

    dev = wb.create_sheet("Deviation")
    dev["B1"] = "Test Case ID"
    dev["C1"] = "Issue"
    dev["D1"] = "Deviation"
    dev["E1"] = "Status"

    log = wb.create_sheet("Test Log")
    log["B1"] = "Test Case ID"
    log["C1"] = "Component"
    log["D1"] = "Method"
    log["E1"] = "Pass/Fail"

    wb.create_sheet("2.Consistency")
    hist = wb.create_sheet("History")
    hist["A1"] = "■ Revision History"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_kjpds_swit_sitr_template() -> bytes:
    """KJPDS02 SwITR 4-sheet layout with fixed Test Log columns."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("History")
    wb.create_sheet("1.Test Summary")
    log = wb.create_sheet("2.Test Log")
    log.cell(5, 2).value = "Test Case "
    log.cell(5, 8).value = "Input"
    log.cell(5, 18).value = "Expected Result"
    log.cell(5, 28).value = "Actual Result"
    log.cell(5, 38).value = "Pass/Fail"
    log.cell(5, 40).value = "Log Data"
    log.cell(6, 2).value = "TC ID"
    log.cell(6, 3).value = "Description"
    log.cell(6, 5).value = "Test Case Generation Method"
    log.cell(6, 6).value = "Precondition"
    log.cell(6, 8).value = "Param 1"
    log.cell(6, 18).value = "Param 1"
    log.cell(6, 28).value = "Param 1"
    log.cell(6, 38).value = "Unit"
    log.cell(6, 39).value = "Total"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_reference_swit_sitr_template(slot_count: int = 2) -> bytes:
    """Wide KJPDS02 SwITR reference-like Test Log with merged B slots."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("History")
    wb.create_sheet("1.Test Summary")
    log = wb.create_sheet("2.Test Log")
    log.cell(4, 1).value = "TC ID"
    log.cell(4, 2).value = "No"
    log.cell(4, 3).value = "Test Env."
    log.cell(4, 4).value = "Test Method"
    log.cell(4, 5).value = "Test Case Generation Method"
    log.cell(4, 6).value = "tc_id"
    log.cell(4, 7).value = "It."
    log.cell(4, 8).value = "Description"
    log.cell(4, 9).value = "Precondition"
    log.cell(4, 10).value = "inpt[0]"
    log.cell(4, 108).value = "expected[0]"
    log.cell(4, 242).value = "actual[0]"
    log.cell(4, 376).value = "Unit"
    log.cell(4, 377).value = "Total"
    log.cell(4, 378).value = "Tail"

    for idx in range(slot_count):
        start = 5 + (idx * 2)
        log.merge_cells(start_row=start, start_column=2, end_row=start + 1, end_column=2)
        log.merge_cells(start_row=start, start_column=3, end_row=start + 1, end_column=3)
        log.merge_cells(start_row=start, start_column=4, end_row=start + 1, end_column=4)
        log.merge_cells(start_row=start, start_column=5, end_row=start + 1, end_column=5)
        log.merge_cells(start_row=start, start_column=6, end_row=start + 1, end_column=6)

    log.cell(5 + (slot_count * 2), 2).value = "End of Document"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_kjpds_swit_sitr_session() -> SwUTSession:
    env = EnvironmentData(
        env_name="SwIT_SwUFn_0101_depth4_file12",
        component_name="SysOs_Main",
        test_cases={
            "main.001": [
                SimpleNamespace(
                    input_data={"input_a": 1},
                    expected_result={"output_a": 2},
                )
            ],
        },
        test_results={
            "main.001": ExecutionRow(
                tc_name="main.001",
                component="SysOs_Main",
                passed=True,
                actual_result={"output_a": 2},
            ),
        },
    )
    return SwUTSession(
        project_id="KJPDS02",
        version="v1.01_251205",
        source_kind="log_folder",
        source_path="/tmp/fake/251106_IT_Report",
        environments=[env],
    )


def _make_reference_swit_sitr_session() -> SwUTSession:
    env = EnvironmentData(
        env_name="SwIT_SwUFn_0101_depth7",
        component_name="Comp",
        test_cases={
            "a.001": [
                SimpleNamespace(
                    description="Beta path",
                    input_data={"in": "beta"},
                    expected_result={"out": "0x0A"},
                )
            ],
            "b.001": [
                SimpleNamespace(
                    description="Alpha path",
                    input_data={"in": "alpha"},
                    expected_result={"out": "0x09"},
                )
            ],
        },
        test_results={
            "a.001": ExecutionRow(
                tc_name="a.001",
                passed=True,
                actual_result={"out": "10"},
            ),
            "b.001": ExecutionRow(
                tc_name="b.001",
                passed=True,
                actual_result={"out": "9"},
            ),
        },
    )
    return SwUTSession(
        project_id="KJPDS02",
        version="v1.01_251205",
        source_kind="log_folder",
        source_path="/tmp/fake/251106_IT_Report",
        environments=[env],
    )


def _make_swit_sitr_session() -> SwUTSession:
    """SwIT SITR session — SwUT SUTR과 동일 구조 (SwIT TC name SwITC_SwUFn_NNNN)."""
    env = EnvironmentData(
        env_name="SWTE_01",
        component_name="SysOs_Main",
        test_cases={
            "SwITC_SwUFn_0101.001": [object()],
            "SwITC_SwUFn_0103.001": [object()],
        },
        test_results={
            "SwITC_SwUFn_0101.001": ExecutionRow(
                tc_name="SwITC_SwUFn_0101.001", passed=True,
            ),
            "SwITC_SwUFn_0103.001": ExecutionRow(
                tc_name="SwITC_SwUFn_0103.001", passed=False,
            ),
        },
        function_coverage=[
            FunctionCoverage(
                unit_id="SwUFn_0101", name="main",
                statement=CoverageStats(8, 8, 1.0),
                branch=CoverageStats(2, 2, 1.0),
                complexity=3,
            ),
            FunctionCoverage(
                unit_id="SwUFn_0103", name="s_SystemOperation",
                statement=CoverageStats(8, 8, 1.0),
                branch=CoverageStats(3, 3, 1.0),
                complexity=2,
            ),
        ],
    )
    return SwUTSession(
        project_id="HDPDM01",
        version="v2.02_240219",
        source_kind="log_folder",
        source_path="/tmp/fake/v2.02_240219",
        environments=[env],
    )


def _make_swit_sitr_meta() -> SwitSitrBuildMeta:
    return SwitSitrBuildMeta(
        project_id="HDPDM01",
        release_sw_version="2.02",
        test_date="2024-02-19",
        test_engineer="JK Kim",
        doc_id_sequence="042",
        asil_level="ASIL B",
    )


# ---------------------------------------------------------------------------
# 1) Smoke + basic structure
# ---------------------------------------------------------------------------

class TestBuildSwitSitr:
    """SwIT SITR builder smoke + filename + summary."""

    def test_smoke_minimal(self):
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        assert isinstance(result, SwitSitrBuildResult)
        assert result.ok
        assert result.xlsm_io.tell() == 0
        assert result.result_size_bytes > 0

    def test_filename_has_sitr_and_version(self):
        """`(HDPDM01_SITR) Software Integration Test Result_v2.02_240219_R.xlsm`."""
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        assert "SITR" in result.filename
        assert "(HDPDM01_SITR)" in result.filename
        assert "Software Integration Test Result" in result.filename
        assert "v2.02" in result.filename
        assert "240219" in result.filename
        assert result.filename.endswith("_R.xlsm")

    def test_filename_uses_config_pattern_when_provided(self):
        """KJPDS02 SwITR 파일명 패턴을 최종 산출물명에 반영."""
        meta = _make_swit_sitr_meta()
        meta.project_id = "KJPDS02"
        meta.doc_filename_pattern = (
            "(KJPDS02_DV_SwITR) Software Integration Test Result_"
            "v{version}_{date}_R.xlsm"
        )
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            meta,
            _build_swit_sitr_template(),
        )
        assert result.filename == (
            "(KJPDS02_DV_SwITR) Software Integration Test Result_"
            "v2.02_240219_R.xlsm"
        )

    def test_summary_contains_asil_keys(self):
        """30차 W21 + 31차 W29 ASIL keys (SUTR과 대칭)."""
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        for key in (
            "asil_distribution", "asil_b_function_ids",
            "asil_c_function_ids", "asil_d_function_ids",
            "asil_highlight_policy",
        ):
            assert key in result.summary, f"summary 키 '{key}' 누락"

    def test_summary_contains_basic_metrics(self):
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        s = result.summary
        assert s["environments"] == 1
        # 2 TC — 1 pass / 1 fail
        assert s["total"] == 2
        assert s["passed"] == 1
        assert s["failed"] == 1
        assert "template_sha256_12" in s
        assert "build_timestamp" in s

    def test_tool_qualification_present(self):
        """ISO 26262 ASIL audit evidence_class draft 정책."""
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        tq = result.tool_qualification
        assert tq["evidence_class"] == "auto-generated draft"
        assert "asil_a_usage" in tq
        assert "asil_b_c_d_usage" in tq


# ---------------------------------------------------------------------------
# 2) Test Log v2.02 매핑 (58차 F3 — 31차 W27 col+4/+5 매핑 폐기)
# ---------------------------------------------------------------------------

class TestSitrTestLogAsil:
    """58차 F3 — col+4/+5 Function ID/ASIL 매핑 폐기.

    회사 v2.02 SITR 양식 정확한 매핑: B=TC ID, H=Input, R=Expected, AB=Actual,
    AL=Pass/Fail, AN=Log Data. Function ID + ASIL 컬럼은 양식에 없음.
    이전 31차 W27 col+4 (F) Function ID + col+5 (G) ASIL stamp는 v2.02 양식의
    Input Params 영역 (F~) 침범 — fix로 stamp 제거. ASIL 시각 강조는 Pass/Fail
    row 영역에 적용.
    """

    def test_no_function_id_stamp_at_col_f_v202(self):
        """58차 F3: col 6 (F)는 Input Param 자리 — Function ID stamp 금지."""
        session = _make_swit_sitr_session()
        session.environments[0].function_asil_map = {
            "SwUFn_0101": "D", "SwUFn_0103": "B",
        }
        result = build_swit_sitr_report(
            session, _make_swit_sitr_meta(), _build_swit_sitr_template(),
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["Test Log"]
        # col 6 (F) Input Param 자리 — 'SwUFn_' prefix 절대 stamp 금지
        f_values = [str(log.cell(r, 6).value or "") for r in range(2, 4)]
        for v in f_values:
            assert "SwUFn_" not in v, f"col 6 (F)에 Function ID stamp 발견 (v2.02 양식 위반): {v}"

    def test_no_asil_label_stamp_at_col_g_v202(self):
        """58차 F3: col 7 (G)는 Input Param 자리 — 'ASIL D' 텍스트 stamp 금지."""
        session = _make_swit_sitr_session()
        session.environments[0].function_asil_map = {"SwUFn_0101": "D"}
        result = build_swit_sitr_report(
            session, _make_swit_sitr_meta(), _build_swit_sitr_template(),
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["Test Log"]
        # col 7 (G) Input Param 자리 — 'ASIL' 텍스트 stamp 금지
        for r in range(2, 4):
            v = str(log.cell(r, 7).value or "")
            assert "ASIL" not in v, f"col 7 (G)에 ASIL 라벨 stamp 발견 (v2.02 양식 위반): {v}"


    def test_kjpds_swit_test_log_uses_swits_env_match(self):
        """SwITS unit_name -> VectorCAST env matching fills the real Test Log."""
        swits_map = {
            "SwITC_0101_01": SimpleNamespace(
                tc_id="SwITC_0101_01",
                unit_name="SWIT_SWUFN_0101_DEPTH4_FILE12",
                description="Interface : main -> s_System_InitSequence",
                test_method="REQ,IFT,AOR,ABV",
                precondition="Power on",
            ),
            "SwITC_0101_02": SimpleNamespace(
                tc_id="SwITC_0101_02",
                unit_name="SWIT_SWUFN_0101_DEPTH4_FILE12",
                description="Interface : main -> s_System_InitSequence duplicate",
                test_method="REQ,IFT,AOR,ABV",
                precondition="Power on",
            ),
        }
        result = build_swit_sitr_report(
            _make_kjpds_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_kjpds_swit_sitr_template(),
            swuts_map=swits_map,
        )

        assert result.ok
        assert result.summary["test_log_rows_written"] == 1
        assert result.summary["test_log_iteration_rows_written"] == 1
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["2.Test Log"]

        assert log.cell(7, 2).value == "SwITC_0101_01, SwITC_0101_02"
        assert str(log.cell(7, 3).value).startswith(
            "Interface : main -> s_System_InitSequence"
        )
        assert "Grouped SwITS: SwITC_0101_01, SwITC_0101_02" in log.cell(7, 3).value
        assert log.cell(7, 5).value == "REQ, IFT, AOR, ABV"
        assert log.cell(7, 6).value == "Power on"
        assert log.cell(7, 8).value == "input_a"
        assert log.cell(7, 18).value == "output_a"
        assert log.cell(7, 28).value == "output_a"
        assert log.cell(7, 39).value == "Pass"
        assert log.cell(7, 40).value == "SwIT_SwUFn_0101_depth4_file12"
        assert log.cell(8, 4).value == "main.001"
        assert log.cell(8, 8).value == "1"
        assert log.cell(8, 18).value == "2"
        assert log.cell(8, 28).value == "2"
        assert log.cell(8, 38).value == "Pass"
        assert "B7:B8" in {str(r) for r in log.merged_cells.ranges}
        assert "E7:E8" in {str(r) for r in log.merged_cells.ranges}
        assert "F7:F8" in {str(r) for r in log.merged_cells.ranges}

    def test_swit_test_log_matches_two_digit_spec_suffix_to_base_env(self):
        """SWIT_SWUFN_1109_01 can match log env SwIT_SwUFn_1109."""
        env = EnvironmentData(
            env_name="SwIT_SwUFn_1109",
            component_name="Comp",
            test_cases={"case.001": [SimpleNamespace(input_data={}, expected_result={})]},
            test_results={"case.001": ExecutionRow(tc_name="case.001", passed=True)},
        )
        session = SwUTSession(
            project_id="KJPDS02",
            version="v1.01_251205",
            source_kind="log_folder",
            source_path="/tmp/fake",
            environments=[env],
        )
        swits_map = {
            "SwITC_1109_01": SimpleNamespace(
                tc_id="SwITC_1109_01",
                unit_name="SWIT_SWUFN_1109_01",
                description="1109 case 1",
            ),
            "SwITC_1109_02": SimpleNamespace(
                tc_id="SwITC_1109_02",
                unit_name="SWIT_SWUFN_1109_02",
                description="1109 case 2",
            ),
        }

        result = build_swit_sitr_report(
            session,
            _make_swit_sitr_meta(),
            _build_kjpds_swit_sitr_template(),
            swuts_map=swits_map,
        )
        assert result.summary["test_log_rows_written"] == 1
        assert result.summary["test_log_iteration_rows_written"] == 1
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["2.Test Log"]
        assert log.cell(7, 2).value == "SwITC_1109_01, SwITC_1109_02"
        assert log.cell(7, 40).value == "SwIT_SwUFn_1109"
        assert log.cell(8, 4).value == "case.001"

    def test_reference_layout_uses_description_before_sorted_tc_order(self):
        """Wide reference layout slots keep merges and map TC rows by SwITS description."""
        swits_map = {
            "SwITC_0101_01": SimpleNamespace(
                tc_id="SwITC_0101_01",
                unit_name="SWIT_SWUFN_0101_DEPTH7",
                description="Alpha path",
                test_method="REQ, IFT",
                generation_method="AOR, ABV",
            ),
            "SwITC_0101_02": SimpleNamespace(
                tc_id="SwITC_0101_02",
                unit_name="SWIT_SWUFN_0101_DEPTH7",
                description="Beta path",
                test_method="REQ, IFT",
                generation_method="AOR, ABV",
            ),
        }
        result = build_swit_sitr_report(
            _make_reference_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_reference_swit_sitr_template(slot_count=2),
            swuts_map=swits_map,
        )

        assert result.summary["test_log_rows_written"] == 2
        assert result.summary["test_log_iteration_rows_written"] == 2
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["2.Test Log"]
        assert "B5:B6" in {str(r) for r in log.merged_cells.ranges}
        assert "B7:B8" in {str(r) for r in log.merged_cells.ranges}
        assert log.cell(5, 6).value == "SwITC_0101_01"
        assert log.cell(6, 8).value == "Alpha path"
        assert log.cell(6, 10).value == "alpha"
        assert log.cell(6, 242).value == "0x09"
        assert log.cell(7, 6).value == "SwITC_0101_02"
        assert log.cell(8, 8).value == "Beta path"
        assert log.cell(8, 10).value == "beta"
        assert log.cell(8, 242).value == "0x0A"

    def test_reference_layout_slot_overflow_warns_and_preserves_footer(self):
        """If SwITS has more blocks than template slots, do not overwrite the footer."""
        swits_map = {
            "SwITC_0101_01": SimpleNamespace(
                tc_id="SwITC_0101_01",
                unit_name="SWIT_SWUFN_0101_DEPTH7",
                description="Alpha path",
            ),
            "SwITC_0101_02": SimpleNamespace(
                tc_id="SwITC_0101_02",
                unit_name="SWIT_SWUFN_0101_DEPTH7",
                description="Beta path",
            ),
        }
        result = build_swit_sitr_report(
            _make_reference_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_reference_swit_sitr_template(slot_count=1),
            swuts_map=swits_map,
        )

        assert result.summary["test_log_rows_written"] == 1
        assert any("reference template slot overflow" in w for w in result.warnings)
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["2.Test Log"]
        assert log.cell(7, 2).value == "End of Document"
        assert log.cell(7, 6).value is None

    def test_reference_layout_preserves_existing_parameter_value_on_parser_conflict(self):
        """Duplicate variable names in VectorCAST input data must not overwrite a reference slot."""
        template_wb = openpyxl.load_workbook(
            io.BytesIO(_build_reference_swit_sitr_template(slot_count=1)),
            keep_vba=True,
        )
        template_log = template_wb["2.Test Log"]
        template_log.cell(5, 10).value = "u16g_SysDiag_SystemStatus"
        template_log.cell(6, 10).value = "1"
        template_log.cell(5, 11).value = "u8s_TemplateOnly"
        template_log.cell(6, 11).value = "0"
        template_buf = io.BytesIO()
        template_wb.save(template_buf)

        env = EnvironmentData(
            env_name="SwIT_SwUFn_0101_depth7",
            component_name="Comp",
            test_cases={
                "a.001": [
                    SimpleNamespace(
                        description="Alpha path",
                        input_data={"u16g_SysDiag_SystemStatus": "3000"},
                        expected_result={},
                    )
                ],
            },
            test_results={
                "a.001": ExecutionRow(tc_name="a.001", passed=True, actual_result={}),
            },
        )
        session = SwUTSession(
            project_id="KJPDS02",
            version="v1.01_251205",
            source_kind="log_folder",
            source_path="/tmp/fake",
            environments=[env],
        )
        swits_map = {
            "SwITC_0101_01": SimpleNamespace(
                tc_id="SwITC_0101_01",
                unit_name="SWIT_SWUFN_0101_DEPTH7",
                description="Alpha path",
            ),
        }

        result = build_swit_sitr_report(
            session,
            _make_swit_sitr_meta(),
            template_buf.getvalue(),
            swuts_map=swits_map,
        )

        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["2.Test Log"]
        assert log.cell(5, 10).value == "u16g_SysDiag_SystemStatus"
        assert log.cell(6, 10).value == "1"
        assert log.cell(5, 11).value == "u8s_TemplateOnly"
        assert log.cell(6, 11).value == "0"
        assert any("reference template parameter cells preserved" in w for w in result.warnings)
        assert any("reference template parameter cells restored" in w for w in result.warnings)

    def test_non_reference_grouped_block_keeps_all_env_test_cases(self):
        """Grouped non-reference output must not drop env TCs by SwITS description."""
        env = EnvironmentData(
            env_name="SwIT_SwUFn_0101_depth7",
            component_name="Comp",
            test_cases={
                "a.001": [SimpleNamespace(description="Alpha path", input_data={}, expected_result={})],
                "b.001": [SimpleNamespace(description="Beta path", input_data={}, expected_result={})],
            },
            test_results={
                "a.001": ExecutionRow(tc_name="a.001", passed=True),
                "b.001": ExecutionRow(tc_name="b.001", passed=True),
            },
        )
        session = SwUTSession(
            project_id="KJPDS02",
            version="v1.01_251205",
            source_kind="log_folder",
            source_path="/tmp/fake",
            environments=[env],
        )
        swits_map = {
            "SwITC_0101_01": SimpleNamespace(
                tc_id="SwITC_0101_01",
                unit_name="SWIT_SWUFN_0101_DEPTH7",
                description="Alpha path",
            ),
            "SwITC_0101_02": SimpleNamespace(
                tc_id="SwITC_0101_02",
                unit_name="SWIT_SWUFN_0101_DEPTH7",
                description="Beta path",
            ),
        }

        result = build_swit_sitr_report(
            session,
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
            swuts_map=swits_map,
        )

        assert result.summary["test_log_rows_written"] == 1
        assert result.summary["test_log_iteration_rows_written"] == 2
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["Test Log"]
        assert log.cell(3, 4).value == "a.001"
        assert log.cell(4, 4).value == "b.001"
        assert "B2:B4" in {str(r) for r in log.merged_cells.ranges}


# ---------------------------------------------------------------------------
# 3) Meta + template validation
# ---------------------------------------------------------------------------

class TestSwitSitrMetaValidation:
    def test_invalid_release_sw_version_raises(self):
        bad_meta = SwitSitrBuildMeta(
            project_id="HDPDM01",
            release_sw_version="bad-version",
            test_date="2024-02-19",
        )
        with pytest.raises(Exception):  # noqa: B017
            build_swit_sitr_report(
                _make_swit_sitr_session(), bad_meta, _build_swit_sitr_template(),
            )

    def test_invalid_template_bytes_raises(self):
        """ZIP bomb / magic byte 검증 — non-xlsm bytes 거부."""
        with pytest.raises(Exception):  # noqa: B017
            build_swit_sitr_report(
                _make_swit_sitr_session(),
                _make_swit_sitr_meta(),
                template_bytes=b"not an xlsm",
            )


# ---------------------------------------------------------------------------
# 4) SwitSitrBuildResult dataclass API
# ---------------------------------------------------------------------------

class TestSwitSitrResultDataclass:
    def test_xlsm_bytes_property_returns_full_content(self):
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        assert isinstance(result.xlsm_bytes, bytes)
        assert len(result.xlsm_bytes) == result.result_size_bytes

    def test_to_dict_includes_required_keys(self):
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        d = result.to_dict()
        for k in (
            "ok", "filename", "result_size_bytes", "warnings",
            "incomplete_sheets", "summary", "tool_qualification",
            "vba_macros_preserved",
        ):
            assert k in d, f"to_dict key '{k}' 누락"


# ---------------------------------------------------------------------------
# 5) Missing sheet fallback (graceful)
# ---------------------------------------------------------------------------

class TestSwitSitrMissingSheet:
    def test_cover_sheet_missing(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Test Summary")
        wb.create_sheet("Deviation")
        wb.create_sheet("Test Log")
        wb.create_sheet("History")
        buf = io.BytesIO()
        wb.save(buf)
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            buf.getvalue(),
        )
        assert result.ok
        assert any("Cover" in w for w in result.warnings)

    def test_deviation_sheet_missing(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        wb.create_sheet("Test Summary")
        wb.create_sheet("Test Log")
        wb.create_sheet("History")
        buf = io.BytesIO()
        wb.save(buf)
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            buf.getvalue(),
        )
        assert result.ok
        assert any("Deviation" in w for w in result.warnings)


# ---------------------------------------------------------------------------
# 6) keep_vba=True preservation (xlsx fixture — VBA 없음 → False)
# ---------------------------------------------------------------------------

class TestKeepVbaPreserved:
    def test_vba_macros_flag_false_for_xlsx_template(self):
        """deep-reviewer W2 — plain xlsx template (VBA 없음) → vba_macros_preserved=False.

        실 v2.02 xlsm template에 VBA가 있을 경우 True 반환은 실 환경에서만 검증
        (poc_swit_round34.py 라이브 검증 의무).
        """
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        assert result.vba_macros_preserved is False
        d = result.to_dict()
        assert d["vba_macros_preserved"] is False

    def test_xlsm_reopens_with_keep_vba(self):
        """`load_workbook(BytesIO, keep_vba=True)` 재오픈 — 시트 유지 확인."""
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        wb = openpyxl.load_workbook(
            io.BytesIO(result.xlsm_bytes), keep_vba=True,
        )
        for name in ("Cover", "Test Summary", "Deviation", "Test Log", "History"):
            assert name in wb.sheetnames, f"시트 '{name}' 누락"


# ---------------------------------------------------------------------------
# 7) Filename short_date format
# ---------------------------------------------------------------------------

class TestSwitSitrFilename:
    def test_filename_short_date_format(self):
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        # short_date(2024-02-19) → 240219
        assert "240219" in result.filename
        assert result.filename.endswith("_R.xlsm")


# ---------------------------------------------------------------------------
# 8) 34차 deep-reviewer C1/C2 fix 회귀
# ---------------------------------------------------------------------------

class TestDeepReviewerC1Fix:
    """C1: `_collect_tc_to_function` `.match` → `.search` (SwIT prefix `SwITC_` 호환).

    이전 버그: TC name `SwITC_SwUFn_0101.001`에 대해 `.match`가 None 반환 →
    `_compute_self_consistency` row 2가 항상 FAIL 반환 → 잘못된 audit evidence.
    fix 후: `.search`로 SwIT prefix 시작 매칭 성공.
    """

    def test_swit_tc_matches_function_id_pattern(self):
        from backend.services.swut_coverage_aggregator import _collect_tc_to_function
        session = _make_swit_sitr_session()
        # session TC name = "SwITC_SwUFn_0101.001"
        tc_map = _collect_tc_to_function(session)
        assert "SwITC_SwUFn_0101.001" in tc_map
        assert tc_map["SwITC_SwUFn_0101.001"] == "SwUFn_0101"

    def test_swut_tc_still_matches_after_fix(self):
        """SwUT TC name `SwUFn_0101.001`도 .search로 정상 매칭 (회귀 영향 없음)."""
        from backend.services.swut_coverage_aggregator import _collect_tc_to_function
        from backend.services.swut_input_adapter import EnvironmentData, SwUTSession
        env = EnvironmentData(env_name="E1")
        env.test_cases = {"SwUFn_0101.001": []}
        session = SwUTSession(environments=[env])
        tc_map = _collect_tc_to_function(session)
        assert tc_map["SwUFn_0101.001"] == "SwUFn_0101"


class TestDeepReviewerC2Fix:
    """C2: `_compute_self_consistency` + `_write_consistency_sheet`에 test_kind kwarg.

    SwIT 호출 시 "SwIT" 라벨로 row 5 item + intro 텍스트 치환.
    """

    def test_swit_consistency_uses_swit_label(self):
        """SwIT SITR 빌드 시 2.Consistency intro / row 5에 "SwIT" 라벨 발견."""
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
            swuds_function_ids={"SwUFn_0101"},
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        cons = wb["2.Consistency"]
        # A1 intro 텍스트
        intro = str(cons["A1"].value or "")
        assert "SwIT" in intro, f"SwIT 라벨 미발견 in intro: {intro!r}"
        # row 5 (자체 일관성 4 row 4-7 + SwUDS row 8) item에 SwIT 포함
        items = [str(cons.cell(r, 1).value or "") for r in range(4, 10)]
        swit_rows = [item for item in items if "SwIT" in item]
        assert swit_rows, f"row 5 item에 SwIT 라벨 미발견. all items: {items}"

    def test_swut_consistency_still_uses_swuts_label_default(self):
        """SwUT는 default test_kind="SwUTS" — 회귀 호환."""
        from backend.services.swut_coverage_aggregator import _compute_self_consistency
        session = _make_swit_sitr_session()
        rows = _compute_self_consistency(session, swuds_function_ids={"SwUFn_0101"})
        # default test_kind="SwUTS" — row 5 item에 SwUTS
        last_item = rows[-1]["item"]
        assert "SwUTS" in last_item, f"SwUTS default 라벨 미발견: {last_item}"


# ---------------------------------------------------------------------------
# 54차 T282/T283 — v2.02 양식 호환 회귀 (SITR)
# ---------------------------------------------------------------------------


def _build_v202_sitr_template() -> bytes:
    """SwIT v2.02 SITR mimic — SW Version + TC stats + Test Log AL marker."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cover = wb.create_sheet("Cover")
    cover["B1"] = "Project"
    cover["B2"] = "ASIL Level"
    cover["B3"] = "Author"
    cover["B4"] = "Version"

    ts = wb.create_sheet("1.Test Summary")
    ts["B1"] = "Project Name"
    ts["B2"] = "SW Version"
    ts["B3"] = "HW Version"
    ts["B4"] = "Test Date"
    ts["B5"] = "Test Engineer"
    ts["B6"] = "Target Coverage"
    ts["B7"] = "Actual Coverage"
    ts["B8"] = "Final Test Result"
    ts["A17"] = "Total TC"
    ts["A22"] = "Requirements/Design Coverage"

    dev = wb.create_sheet("Deviation")
    dev["B1"] = "Test Case ID"
    dev["C1"] = "Issue"
    dev["D1"] = "Deviation"
    dev["E1"] = "Status"

    log = wb.create_sheet("Test Log")
    log["B1"] = "Test Case ID"
    log["C1"] = "Component"
    log["D1"] = "Method"
    log["E1"] = "Pass/Fail"
    # AL column = 38, header row에 "Marker" 라벨
    log.cell(row=1, column=38, value="Marker")

    wb.create_sheet("2.Consistency")
    wb.create_sheet("History")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSwitSitrV202LayoutCompat:
    """54차 T282/T283 — SITR v2.02 SW Version label + TC stats + AL marker."""

    def test_sw_version_label_filled(self):
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_v202_sitr_template(),
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        ts = wb["1.Test Summary"]
        assert ts["C2"].value == "2.02"
        assert ts["C3"].value == "1.00"

    def test_tc_stats_row_filled(self):
        """55-fix: 라벨 row=17이 헤더 (가로 배치), data는 row=18에 채움."""
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_v202_sitr_template(),
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        ts = wb["1.Test Summary"]
        # 55-fix: "Total TC" 라벨 A17 → data row=18, col_start=1 (A)
        # 2 TC (one pass, one fail)
        assert ts["A18"].value == 2   # Total
        assert ts["B18"].value == 2   # Tested
        assert ts["C18"].value == 1   # Passed (one passed=True)
        assert ts["D18"].value == 1   # Failed (one passed=False)
        assert result.summary.get("tc_stats_blocked_inferred") is True

    def test_test_log_al_column_marker(self):
        """58차 F3: v2.02 fixture는 col 5(E)='Pass/Fail' 헤더 보유 → layout에서
        test_log_pass_fail_col=5로 인식. col 5에 'Pass'/'Fail' stamp 검증.

        AL(col 38)에는 'Marker' 라벨 fixture가 stamp되어 있어 fallback marker는
        skip (test_log_extra_marker_col=38 == LOG_DATA_COL fallback 38 충돌 — 코드
        가 skip).
        """
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_v202_sitr_template(),
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsm_bytes), keep_vba=True)
        log = wb["Test Log"]
        # session 2 TC (Pass + Fail) — col 5 (E) Pass/Fail Unit stamp
        row2_v = log.cell(row=2, column=5).value
        row3_v = log.cell(row=3, column=5).value
        # 'Pass'/'Fail'/'N/A' 한 row 이상 발견
        values = {str(row2_v or ""), str(row3_v or "")}
        assert any(v in ("Pass", "Fail", "N/A") for v in values), (
            f"col 5 (E='Pass/Fail') stamp 미발견 — values: {values}"
        )

    def test_v301_backward_compat_no_al_fill(self):
        """v3.01 양식 (AL col 없음)에서 marker skip — backward compat."""
        result = build_swit_sitr_report(
            _make_swit_sitr_session(),
            _make_swit_sitr_meta(),
            _build_swit_sitr_template(),
        )
        assert result.ok
        assert "tc_stats_blocked_inferred" not in result.summary
