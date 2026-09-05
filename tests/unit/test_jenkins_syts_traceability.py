"""Regression: /api/jenkins/syts/extract-traceability 전용 파서.

SyTS 레이아웃은 SITS와 달라(TC ID=col B 'Test Case ID', Related ID=상단 병합 그룹헤더
아래 열) 과거 SITS 파서 위임이 Title(col C)을 testcase로, Test Environment(SyTE)를
requirement로 오독해 밴드의 92.9%가 허위 추적이었다. 전용 파서로 교체한 것의 회귀 가드.
TC ID는 SyTC_ 형태, 요구 ID는 col의 Related ID(SyEIF→SwEI 평탄화).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from backend.routers.jenkins import (
    jenkins_syits_extract_traceability,
    jenkins_syts_extract_traceability,
)


@pytest.fixture(autouse=True)
def _local_mode(monkeypatch: pytest.MonkeyPatch):
    """resolver를 local로 고정 — tmp 파일 직접 read(cloudium IPC 회피)."""
    from backend.services import file_resolver as fr
    monkeypatch.setattr(fr, "_resolver", fr.LocalFileResolver())


def _save(wb: openpyxl.Workbook, tmp_path: Path, name: str) -> str:
    fp = tmp_path / name
    wb.save(str(fp))
    return str(fp)


def test_syts_tc_id_and_related_id_columns(tmp_path: Path) -> None:
    """TC ID(col B) + Related ID(row3 병합헤더, col T) 정확 추출, Title은 미유입.

    mutation: TC 열을 col C(Title)로 읽으면 testcase='SBCM LIN Signal' → TC 형태 아님 → 스킵.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "2.System Test Spec"
    ws.cell(3, 20, "Related ID")          # 병합 그룹헤더 — TC 헤더(row4)보다 위
    ws.cell(4, 2, "Test Case ID")
    ws.cell(4, 3, "Title")
    ws.cell(5, 2, "SyTC_SyEIF_01_01")
    ws.cell(5, 3, "SBCM LIN Signal")      # Title — testcase/requirement로 유입되면 안 됨
    ws.cell(5, 20, "SyEIF_01")
    res = jenkins_syts_extract_traceability({"path": _save(wb, tmp_path, "syts.xlsx")})
    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWEI_01", "SyTC_SyEIF_01_01") in pairs     # SyEIF→SwEI 평탄화
    # Title 문자열은 어느 필드에도 유입되지 않는다
    assert all("SBCM" not in str(v) for r in res["vcast_rows"] for v in r.values())
    assert {r["source"] for r in res["vcast_rows"]} == {"SyTS"}


def test_syts_related_id_same_row_as_tc_header(tmp_path: Path) -> None:
    """Related ID가 TC 헤더와 같은 행에 있어도 추출(레이아웃 변형)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Test Specification"
    ws.cell(4, 2, "Test Case ID")
    ws.cell(4, 20, "Related ID")
    ws.cell(5, 2, "SyTC_SyTR_0101_01")
    ws.cell(5, 20, "SyTR_0101")
    res = jenkins_syts_extract_traceability({"path": _save(wb, tmp_path, "s.xlsx")})
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWTR_0101", "SyTC_SyTR_0101_01") in pairs


def test_syts_multi_value_related_id(tmp_path: Path) -> None:
    """한 Related ID 셀의 콤마구분 다중 요구를 모두 추출."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2.System Test Spec"
    ws.cell(4, 2, "Test Case ID")
    ws.cell(4, 20, "Related ID")
    ws.cell(5, 2, "SyTC_SyEIF_01_01")
    ws.cell(5, 20, "SyEIF_01, SyTR_0102")
    res = jenkins_syts_extract_traceability({"path": _save(wb, tmp_path, "s.xlsx")})
    reqs = {r["requirement_id"] for r in res["vcast_rows"]}
    assert {"SWEI_01", "SWTR_0102"} <= reqs


def test_syts_no_related_col_warns(tmp_path: Path) -> None:
    """Related ID 열 미탐지 시 warning + 빈 결과(침묵 오파싱 방지)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Test Spec"
    ws.cell(4, 2, "Test Case ID")
    ws.cell(4, 3, "Title")               # Related ID 열 없음
    ws.cell(5, 2, "SyTC_SyEIF_01_01")
    ws.cell(5, 3, "SBCM LIN Signal")
    res = jenkins_syts_extract_traceability({"path": _save(wb, tmp_path, "s.xlsx")})
    assert res["vcast_rows"] == []
    assert res.get("warning") and "Related ID" in res["warning"]


def test_syts_non_tc_id_row_skipped(tmp_path: Path) -> None:
    """TC 열 값이 TC ID 형태(SyTC_)가 아니면 스킵 — Title/설명 유입 2차 방어."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2.System Test Spec"
    ws.cell(4, 2, "Test Case ID")
    ws.cell(4, 20, "Related ID")
    ws.cell(5, 2, "Not A Test Case")     # TC ID 형태 아님
    ws.cell(5, 20, "SyEIF_01")
    res = jenkins_syts_extract_traceability({"path": _save(wb, tmp_path, "s.xlsx")})
    assert res["vcast_rows"] == []


def test_syits_tc_id_in_col_c(tmp_path: Path) -> None:
    """SyITS: col C 'TC ID'(SyITC_)를 TC로, col T 'Related ID'(SyFN)를 요구로.

    SyITS는 col B='ID'(SyFN_01 별도)·col C='TC ID'(실 TC)라 SyTS(col B=TC)와 위치가
    다르나, 동적 헤더 탐지가 'TC ID' 헤더를 잡아 col C를 tc_col로 쓴다. 엔드포인트가
    SyTS 전용 파서를 공유하는지(SITS 위임 아님) 회귀 가드.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "3.System Integration Test Spec"
    ws.cell(3, 20, "Related ID")
    ws.cell(4, 2, "ID")
    ws.cell(4, 3, "TC ID")
    ws.cell(4, 13, "Sequence")
    ws.cell(5, 2, "SyFN_01")               # col B 'ID' — TC 아님
    ws.cell(5, 3, "SyITC_SyFN_01_01")      # col C 'TC ID' — 실 TC
    ws.cell(5, 20, "SyFN_01")              # col T Related ID
    res = jenkins_syits_extract_traceability({"path": _save(wb, tmp_path, "syits.xlsx")})
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SYFN_01", "SyITC_SyFN_01_01") in pairs
    assert {r["source"] for r in res["vcast_rows"]} == {"SyITS"}


def test_syits_prose_swtc_not_extracted(tmp_path: Path) -> None:
    """Sequence 열의 산문 'SwTC_SwTR_..' 교차참조는 requirement로 안 들어간다.

    과거 SITS 위임이 Sequence/Description 산문에서 'SwTC_SwTR_0501_01'의 substring
    'SwTR_0501'을 요구로 오추출해 허위 밴드를 냈다. Related ID 열 한정으로 차단됨을 고정.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "3.System Integration Test Spec"
    ws.cell(3, 20, "Related ID")
    ws.cell(4, 2, "ID")
    ws.cell(4, 3, "TC ID")
    ws.cell(4, 13, "Sequence")
    ws.cell(5, 3, "SyITC_SyFN_01_01")
    ws.cell(5, 13, "SwTC_SwTR_0501_01: 대체 검증 함")  # 산문 교차참조
    ws.cell(5, 20, "SyFN_01")
    res = jenkins_syits_extract_traceability({"path": _save(wb, tmp_path, "syits2.xlsx")})
    reqs = {r["requirement_id"] for r in res["vcast_rows"]}
    assert "SYFN_01" in reqs
    assert "SWTR_0501" not in reqs         # 산문 SwTC 교차참조 오추출 없음


def _syrs_doc(tmp_path: Path, tokens: str) -> str:
    from docx import Document
    p = tmp_path / "syrs.docx"
    d = Document()
    d.add_paragraph(tokens)
    d.save(str(p))
    return str(p)


def _syts_wb() -> openpyxl.Workbook:
    """SyTS 레이아웃 — 요구는 시스템요구(SyTR_0802 등 SW SRS엔 없음)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "2.System Test Spec"
    ws.cell(3, 20, "Related ID")
    ws.cell(4, 2, "Test Case ID")
    ws.cell(5, 2, "SyTC_SyTR_0802_01")
    ws.cell(5, 20, "SyTR_0802")          # SW SRS엔 없으나 SyRS엔 있는 시스템-only 요구
    ws.cell(6, 2, "SyTC_SyEIF_01_01")
    ws.cell(6, 20, "SyEIF_01")
    return wb


def test_syts_system_basis_joins_syrs(tmp_path: Path) -> None:
    """SyTS: 평탄화 전 원본 Sy* 참조를 SyRS에 조인 — SW 매트릭스가 놓치는 시스템-only 요구 포함.

    mutation: sys_refs 수집(원본 Sy* 보존)을 제거하면 joined=0이 되어 실패.
    band는 SW 평탄화(SwTR_0802 — SW SRS 밖)라 낮으나, system_basis는 SyRS 기준 참 커버리지.
    """
    syrs = _syrs_doc(tmp_path, "SyTR_0802 SyEIF_01 SyEIF_06 SyTR_0101")
    res = jenkins_syts_extract_traceability(
        {"path": _save(_syts_wb(), tmp_path, "s.xlsx"), "syrs_path": syrs})
    assert res["ok"] is True
    sb = res["system_basis"]
    assert sb["refs_total"] == 2
    assert sb["joined"] == 2                       # SyTR_0802 + SyEIF_01 둘 다 SyRS 실재
    assert sb["unmatched"] == []


def test_syts_no_syrs_path_no_system_basis(tmp_path: Path) -> None:
    """하위호환: syrs_path 미제공 시 system_basis 없음, 기존 필드 유지."""
    res = jenkins_syts_extract_traceability(
        {"path": _save(_syts_wb(), tmp_path, "s2.xlsx")})
    assert res["ok"] is True
    assert "system_basis" not in res
    assert "vcast_rows" in res and "requirements_covered" in res


def test_syits_carve_out_no_system_basis(tmp_path: Path) -> None:
    """SyITS carve-out: syrs_path를 줘도 system_basis 산출 안 함.

    SyITS는 SyII/SyDB 등 시스템 설계요소 참조라 SyRS(요구) 미조인이 정상(2뿐) — 오분모 방지.
    source_label='SyITS'로 위임되므로 SyTS 파서가 system_basis를 건너뛴다.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "3.System Integration Test Spec"
    ws.cell(3, 20, "Related ID")
    ws.cell(4, 3, "TC ID")
    ws.cell(5, 3, "SyITC_SyII_01_01")
    ws.cell(5, 20, "SyII_01")
    syrs = _syrs_doc(tmp_path, "SyII_01 SyTR_0101")
    res = jenkins_syits_extract_traceability(
        {"path": _save(wb, tmp_path, "syits3.xlsx"), "syrs_path": syrs})
    assert res["ok"] is True
    assert "system_basis" not in res         # carve-out — SyITS는 SyRS 재기준 안 함


def test_syts_bad_syrs_degrades(tmp_path: Path) -> None:
    """비-docx syrs_path → 크래시 아닌 degrade, vcast_rows(주 산출물) 보존."""
    bad = tmp_path / "bad.docx"
    bad.write_text("not a docx", encoding="utf-8")
    res = jenkins_syts_extract_traceability(
        {"path": _save(_syts_wb(), tmp_path, "s3.xlsx"), "syrs_path": str(bad)})
    assert res["ok"] is True
    assert "vcast_rows" in res                # 주 산출물 보존
    assert "system_basis" not in res
    assert res.get("system_basis_error")     # fail-loud
    assert "/" not in res["system_basis_error"] and "\\" not in res["system_basis_error"]
