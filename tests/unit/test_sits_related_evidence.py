"""SITS Related ID 의 **어휘와 범위**, 그리고 그 근거 시트(R7).

정본(KJPDS02_PV_SwITS v1.02) 50 TC · Related 원소 340개 대조에서 두 가지가 좁았다:

    ① 어휘 — `SwCom_` 만 남기고 SwFn/SwSTR/SwST/SwTK 를 버렸다(문서 토큰의 19%)
    ② 범위 — **진입 함수 한 개**만 조회했다. 정본 340 중 entry 자신으로 설명되는 건
       142(41.8%) 뿐이고 나머지 198 은 호출 트리 아래에서만 온다

    원소 재현율   18.8%(현행) → 41.8%(어휘) → 66.2%(2홉) → 74.4%(3홉, 과잉 2.3배)

정본은 이 계산을 `Related_ID 확인`/`정리` 시트에 남기고 그 합집합을 3번 시트로 옮긴다
(실측: 3번 시트 55건 중 43건이 `정리` D열과 문자 그대로 같다). 그래서 이 시트들은
장식이 아니라 산출물의 **근거표**다 — 없으면 Related 칸의 출처를 문서 안에서 확인할 수 없다.
"""
from __future__ import annotations

import io

from generators.sits import (
    _RELATED_CHAIN_DEPTH,
    _absolute_depth_map,
    _build_call_chain_nodes,
    _build_calls_map,
    _build_file_map,
    _build_strategy_tree,
    collect_integration_flows,
    generate_sits_xlsm,
    validate_sits_xlsm,
)


def _kv_docx(pairs: dict[str, str]) -> bytes:
    """`{함수명: Related ID 문자열}` → SwUDS v3.02 세로 kv 표 docx."""
    from docx import Document

    doc = Document()
    for name, related in pairs.items():
        t = doc.add_table(rows=0, cols=2)
        for label, value in (("Name", name), ("Related ID", related)):
            cells = t.add_row().cells
            cells[0].text = label
            cells[1].text = value
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _fd_related():
    """진입 함수와 **그 아래 세 홉**에 서로 다른 종류의 ID 가 붙어 있는 최소 그래프.

    `Ap_Door_Run`(SwCom_01) → `Drv_Motor_Set`(SwFn_09) → `Hal_Pwm_Set`(SwSTR_04)
    → `Reg_Write`(SwTK_02). 진입 함수만 보면 1개, 2홉이면 3개, 3홉이면 4개다.
    """
    def _f(name, file, calls):
        return {"name": name, "file": file, "calls_list": list(calls),
                "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
                "asil": "B"}
    return {
        "F1": _f("Ap_Door_Run", "Ap_Door.c", ["Drv_Motor_Set"]),
        "F2": _f("Drv_Motor_Set", "Drv_Motor.c", ["Hal_Pwm_Set"]),
        "F3": _f("Hal_Pwm_Set", "Hal_Pwm.c", ["Reg_Write"]),
        "F4": _f("Reg_Write", "Hal_Reg.c", []),
    }


_REL_MAP = {
    "ap_door_run": ["SwCom_01"],
    "drv_motor_set": ["SwFn_09"],
    "hal_pwm_set": ["SwSTR_04"],
    "reg_write": ["SwTK_02"],
}


def _itcs(n_tc: int = 1, n_sub: int = 1):
    return [{
        "tc_id": f"SwITC_{i + 1}", "entry_fn": "Ap_Door_Run", "call_chain": "a -> b",
        "description": "d", "asil": "B", "test_method": "REQ, IFT",
        "gen_method": "AOR, AEC", "related_ids": ["SwCom_01"],
        "sub_cases": [{"case_num": j + 1, "inputs": {"x": "1"}, "expected": {"y": "2"},
                       "precondition": "", "call_chain": ""} for j in range(n_sub)],
    } for i in range(n_tc)]


class TestVocabularyIsNotNarrowedToSwCom:
    def test_extractor_keeps_non_swcom_tokens(self):
        from backend.services.iso26262_doc_asil_extractor import (
            extract_function_related_ids_from_kv_tables,
            extract_function_swcom_from_kv_tables,
        )
        docx = _kv_docx({"g_Foo": "SwCom_01, SwFn_09, SwSTR_04, SwST_07, SwTK_02"})
        # 순서는 문자열 정렬이라 `SwSTR_04` 가 `SwST_07` 앞이다(길이 아닌 사전순).
        assert extract_function_related_ids_from_kv_tables(docx)["g_foo"] == [
            "SwCom_01", "SwFn_09", "SwSTR_04", "SwST_07", "SwTK_02"]
        # 좁은 판(ASIL 컴포넌트 상속용)은 그대로 SwCom 만 — 두 축이 서로를 망가뜨리지 않는다.
        assert extract_function_swcom_from_kv_tables(docx) == {"g_foo": ["SwCom_01"]}

    def test_swstr_is_not_folded_into_swst(self):
        """접두를 부분문자열로 판정하면 `SwSTR_01` 이 `SwST_01` 로 뭉개진다."""
        from backend.services.iso26262_doc_asil_extractor import (
            extract_function_related_ids_from_kv_tables,
        )
        got = extract_function_related_ids_from_kv_tables(_kv_docx({"g_Foo": "SWSTR_01, SWST_01"}))
        assert got["g_foo"] == ["SwSTR_01", "SwST_01"], got

    def test_prose_token_is_not_picked_up(self):
        from backend.services.iso26262_doc_asil_extractor import (
            extract_function_related_ids_from_kv_tables,
        )
        got = extract_function_related_ids_from_kv_tables(
            _kv_docx({"g_Foo": "Software_12 참조, SwCom_03"}))
        assert got["g_foo"] == ["SwCom_03"], got

    def test_function_without_related_id_is_absent(self):
        from backend.services.iso26262_doc_asil_extractor import (
            extract_function_related_ids_from_kv_tables,
        )
        assert extract_function_related_ids_from_kv_tables(_kv_docx({"g_Foo": "미정"})) == {}


class TestRelatedIdSpansTheCallTree:
    def test_ids_below_the_entry_are_collected(self):
        flows = collect_integration_flows(_fd_related(), uds_related_map=_REL_MAP)
        f = next(x for x in flows if x["entry_fn"] == "Ap_Door_Run")
        assert "SwCom_01" in f["related_ids"], "진입 함수 자신의 ID 가 빠졌다"
        assert "SwFn_09" in f["related_ids"], "1홉 아래 ID 를 못 모았다"
        assert "SwSTR_04" in f["related_ids"], "2홉 아래 ID 를 못 모았다"

    def test_entry_id_comes_first(self):
        """상한에 걸릴 때 잘리는 건 항상 **가장 먼 근거**여야 한다."""
        flows = collect_integration_flows(_fd_related(), uds_related_map=_REL_MAP)
        ids = next(x for x in flows if x["entry_fn"] == "Ap_Door_Run")["related_ids"]
        assert ids.index("SwCom_01") < ids.index("SwFn_09") < ids.index("SwSTR_04")

    def test_depth_limit_is_honoured(self):
        """3홉은 담지 않는다 — 재현 +8%p 에 과잉 +168 이라 경계를 2홉에 뒀다."""
        assert _RELATED_CHAIN_DEPTH == 2
        flows = collect_integration_flows(_fd_related(), uds_related_map=_REL_MAP)
        ids = next(x for x in flows if x["entry_fn"] == "Ap_Door_Run")["related_ids"]
        assert "SwTK_02" not in ids, f"3홉 아래까지 끌어왔다: {ids}"

    def test_chain_sourced_ids_are_counted_separately(self):
        """칸만 보면 진입 함수 것과 트리 아래 것을 구별할 수 없다 — 통계로 낸다."""
        stats: dict = {}
        collect_integration_flows(_fd_related(), uds_related_map=_REL_MAP, stats_out=stats)
        assert stats["related_chain_flows"] >= 1
        assert stats["related_chain_ids"] >= 2, stats
        assert stats["related_chain_depth"] == 2

    def test_cell_cap_truncation_is_reported(self):
        """칸 상한에 잘린 사실이 사라지면 읽는 쪽은 그게 전부인 줄 안다."""
        wide = {"ap_door_run": [f"SwCom_{i:02d}" for i in range(60)]}
        stats: dict = {}
        flows = collect_integration_flows(_fd_related(), uds_related_map=wide, stats_out=stats)
        f = next(x for x in flows if x["entry_fn"] == "Ap_Door_Run")
        assert len(f["related_ids"]) == 40
        assert stats["related_truncated_ids"] >= 20, stats

    def test_absent_map_falls_back_to_the_narrow_one(self):
        """맵을 안 주면 예전처럼 SwCom 좁은 판으로 내려간다(조용히 죽지 않는다)."""
        flows = collect_integration_flows(_fd_related(),
                                          uds_swcom_map={"ap_door_run": ["SwCom_77"]})
        assert "SwCom_77" in next(
            x for x in flows if x["entry_fn"] == "Ap_Door_Run")["related_ids"]


class TestBalancingDoesNotEraseDocumentedIds:
    """균형 조정(`_balance_related_ids`)은 순번 합성·주석 유래를 겨눈 것이지 문서 값이 아니다.

    정본 실측(50 TC): `SwFn_42` 가 15건(30%)에 정당하게 쓰인다 — 임계 20% 규칙이면 지워진다.
    어휘를 SwCom 밖으로 넓힌 이번 라운드 전에는 대상이 없어 드러나지 않던 구멍이다.
    """

    def test_documented_id_survives_over_use(self):
        from generators.sits import _balance_related_ids

        flows = [{"related_ids": ["SwFn_42", "SwTR_%02d" % i],
                  "doc_related_ids": ["SwFn_42"]} for i in range(10)]
        _balance_related_ids(flows, max_freq_pct=0.2)
        assert all("SwFn_42" in f["related_ids"] for f in flows), \
            "문서에 적힌 ID 가 과집중으로 오인돼 지워졌다"

    def test_undocumented_over_use_is_still_trimmed(self):
        """반대 축 — 근거 없는 ID 가 전 흐름에 붙는 것은 여전히 걷어낸다."""
        from generators.sits import _balance_related_ids

        flows = [{"related_ids": ["SwTR_99", "SwTR_%02d" % i], "doc_related_ids": []}
                 for i in range(10)]
        _balance_related_ids(flows, max_freq_pct=0.2)
        assert all("SwTR_99" not in f["related_ids"] for f in flows)

    def test_flow_records_which_ids_came_from_the_document(self):
        flows = collect_integration_flows(_fd_related(), uds_related_map=_REL_MAP)
        f = next(x for x in flows if x["entry_fn"] == "Ap_Door_Run")
        assert set(f["doc_related_ids"]) == {"SwCom_01", "SwFn_09", "SwSTR_04"}, f
        assert set(f["doc_related_ids"]) <= set(f["related_ids"])


class TestStrategyTreeShape:
    """전략 시트의 트리는 체인과 **다른 규칙**이다 — 경로별 전개 · 절대 depth."""

    def test_tree_expands_per_path_not_per_function(self):
        """정본 실측: 한 블록에서 `s_HistoryPushDoorState` 가 5회(부모가 다르다)."""
        calls = {"root": ["a", "b"], "a": ["leaf"], "b": ["leaf"], "leaf": []}
        nodes, dropped = _build_strategy_tree("root", calls, 100, 20)
        names = [n for _d, n in nodes]
        assert names.count("leaf") == 2, f"경로별 전개가 아니다: {names}"
        assert dropped == 0

    def test_chain_still_dedupes(self):
        """반대 축 — 체인은 visited 기반이라 같은 함수가 두 번 나오면 안 된다(정본 0/54)."""
        calls = {"root": ["a", "b"], "a": ["leaf"], "b": ["leaf"], "leaf": []}
        nodes, _dropped = _build_call_chain_nodes("root", calls, 100)
        assert nodes.count("leaf") == 1, nodes

    def test_recursion_is_cut_per_path_only(self):
        calls = {"root": ["a"], "a": ["root", "b"], "b": []}
        nodes, _dropped = _build_strategy_tree("root", calls, 100, 20)
        assert [n for _d, n in nodes] == ["root", "a", "b"]

    def test_depth_is_measured_from_the_execution_root(self):
        """정본 `N depth` 는 진입점 기준 상대값이 아니라 main 기준 절대값이다.

        실측: 그래프에서 도달 가능한 정본 블록 21개에서 **불일치 0**.
        """
        calls = {"main": ["x"], "x": ["y"], "y": ["z"], "z": [], "isr": []}
        d = _absolute_depth_map(calls, ("main",))
        assert (d["main"], d["x"], d["y"], d["z"]) == (1, 2, 3, 4)
        assert "isr" not in d, "루트에서 안 닿는 함수를 임의 값으로 채우면 안 된다"

    def test_node_cap_is_reported_not_swallowed(self):
        calls = {"root": [f"f{i}" for i in range(10)], **{f"f{i}": [] for i in range(10)}}
        nodes, dropped = _build_strategy_tree("root", calls, 4, 20)
        assert len(nodes) == 4 and dropped >= 1, (nodes, dropped)

    def test_call_graph_builder_is_shared(self):
        """전략 트리와 규격 체인이 **같은 그래프**를 봐야 한다(외부 심볼 제외 규칙 포함)."""
        got = _build_calls_map(_fd_related())
        assert got["Ap_Door_Run"] == ["Drv_Motor_Set"]
        assert _build_calls_map({"F": {"name": "a", "calls_list": ["memset"]}}) == {"a": []}


class TestEvidenceSheetsAreWritten:
    def _gen(self, tmp_path, **ctx):
        out = tmp_path / "ev.xlsx"
        fd = _fd_related()
        flows = collect_integration_flows(fd, uds_related_map=_REL_MAP)
        base = {"calls_map": _build_calls_map(fd), "file_of": _build_file_map(fd),
                "depth_of": {"Ap_Door_Run": 3}, "uds_related_map": _REL_MAP,
                "stats_out": {}}
        base.update(ctx)
        generate_sits_xlsm(None, _itcs(), str(out), flows=flows, strategy_context=base)
        return out, base["stats_out"]

    def test_sheets_exist_and_carry_the_tree(self, tmp_path):
        import openpyxl

        out, stats = self._gen(tmp_path)
        wb = openpyxl.load_workbook(str(out))
        names = {n.replace(" ", "") for n in wb.sheetnames}
        assert "2.SWIntegrationStrategy" in names, wb.sheetnames
        assert {"Related_ID확인", "Related_ID정리"} <= names, wb.sheetnames
        cells = [c.value for row in wb["2.SW Integration Strategy"].iter_rows()
                 for c in row if c.value]
        assert "Ap_Door_Run" in cells and "Hal_Pwm_Set" in cells, cells
        assert stats["strategy_blocks"] >= 1 and stats["strategy_nodes"] >= 3
        assert "Ap_Door.c" in cells, "트리가 걸치는 파일 목록(B열)이 없다"

    def test_depth_header_uses_absolute_depth(self, tmp_path):
        import openpyxl

        out, _stats = self._gen(tmp_path, depth_of={
            "Ap_Door_Run": 5, "Drv_Motor_Set": 6, "Hal_Pwm_Set": 7, "Reg_Write": 8})
        cells = [str(c.value) for row in
                 openpyxl.load_workbook(str(out))["2.SW Integration Strategy"].iter_rows()
                 for c in row if c.value]
        assert {"5 depth", "6 depth", "7 depth"} <= set(cells), cells
        assert "1 depth" not in cells, "진입점을 무조건 1 로 적으면 정본과 어긋난다"

    def test_unreachable_entry_starts_at_one(self, tmp_path):
        """루트에서 안 닿는 진입점(ISR·`_EntryPoint`)은 그 자체가 실행 시작점이다.

        정본 실측: 그래프 미도달 18블록 중 13이 `1 depth` — 미도달→1 규칙으로 34/39 일치.
        """
        import openpyxl

        out, _stats = self._gen(tmp_path, depth_of={})
        cells = [str(c.value) for row in
                 openpyxl.load_workbook(str(out))["2.SW Integration Strategy"].iter_rows()
                 for c in row if c.value]
        assert "1 depth" in cells, cells

    def test_tidy_sheet_carries_the_same_union_as_the_spec_cell(self, tmp_path):
        """정본 실측: 3번 시트 Related 칸 55건 중 43건이 `정리` D열과 문자 그대로 같다."""
        import openpyxl

        out, _stats = self._gen(tmp_path)
        rows = [r for r in openpyxl.load_workbook(str(out))["Related_ID 정리"]
                .iter_rows(min_row=2, values_only=True) if r[1]]
        assert rows, "정리 시트가 비어 있다"
        got = str(rows[0][3])
        assert "SwCom_01" in got and "SwFn_09" in got, got

    def test_index_side_lists_the_source_document_functions(self, tmp_path):
        """우측 인덱스는 **원문 이름**과 설계 ID 를 적는다.

        맵의 키는 소문자(표기 흔들림 흡수)라 그대로 쓰면 ① `design_ids` 조회가 전부 미스이고
        ② 문서에 없는 `ap_door_run` 같은 이름이 산출물에 실린다 — 실측으로 둘 다 났다.
        """
        import openpyxl

        out, stats = self._gen(tmp_path,
                               design_ids={"by_name": {"Ap_Door_Run": "SwUFn_0101"}})
        ws = openpyxl.load_workbook(str(out))["Related_ID 확인"]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[7]]
        names = {str(r[7]) for r in rows}
        assert "Ap_Door_Run" in names, f"소문자 키가 그대로 실렸다: {sorted(names)}"
        assert "ap_door_run" not in names, sorted(names)
        assert any(str(r[6]) == "SwUFn_0101" for r in rows), \
            "설계 ID 가 대소문자 때문에 전부 비었다"
        assert stats["relid_index_rows"] == len(_REL_MAP)

    def test_integration_id_links_back_to_the_spec_sheet(self, tmp_path):
        """근거 시트의 통합 ID 는 규격 시트 TC ID 에서 파생해야 맞춰 볼 수 있다.

        정본: 규격 `SwITC_SwUFn_0101_01` ↔ 근거 `SwIT_SwUFn_0101_01`(접두만 다름).
        예전엔 내부 `flow_id`(`SwUFn_3596`)를 적어 규격 시트 어디에도 없는 이름이었다.
        """
        import openpyxl

        out, _stats = self._gen(tmp_path)
        wb = openpyxl.load_workbook(str(out))
        tidy_ids = {str(r[1]) for r in wb["Related_ID 정리"].iter_rows(min_row=2, values_only=True)
                    if r[1]}
        assert "SwIT_1" in tidy_ids, tidy_ids     # `_itcs()` 의 `SwITC_1` 에서 파생
        strat = {str(c.value) for row in wb["2.SW Integration Strategy"].iter_rows()
                 for c in row if c.value}
        assert "SwIT_1" in strat, "전략 시트도 같은 ID 를 써야 두 시트가 연결된다"

    def test_integration_id_prefix_conversion_is_single_sourced(self):
        from generators.sits import _integration_id

        assert _integration_id("SwITC_SwUFn_0101_01") == "SwIT_SwUFn_0101_01"
        assert _integration_id("SwUFn_0101") == "SwUFn_0101"   # 접두 없으면 그대로
        assert _integration_id("") == ""

    def test_template_sheet_is_filled_not_duplicated(self, tmp_path):
        """등록 템플릿의 시트 이름은 `2. SW…`(공백), 정본은 `2.SW…` 다.

        정확일치로 찾으면 템플릿 시트를 못 보고 **같은 주제의 시트를 하나 더** 만든다 —
        한 파일에 빈 정본 시트와 채워진 사본이 공존하던 `3.`/`4.` 결함과 같은 형태다.
        """
        import openpyxl

        tpl = tmp_path / "tpl.xlsx"
        twb = openpyxl.Workbook()
        twb.active.title = "2. SW Integration Strategy"
        twb["2. SW Integration Strategy"]["A1"] = "Software Integration Strategy"
        twb.save(str(tpl))

        out = tmp_path / "from_tpl.xlsx"
        fd = _fd_related()
        flows = collect_integration_flows(fd, uds_related_map=_REL_MAP)
        generate_sits_xlsm(str(tpl), _itcs(), str(out), flows=flows, strategy_context={
            "calls_map": _build_calls_map(fd), "file_of": _build_file_map(fd),
            "depth_of": {}, "uds_related_map": _REL_MAP, "stats_out": {}})
        wb = openpyxl.load_workbook(str(out))
        strat = [n for n in wb.sheetnames if "integrationstrategy" in n.replace(" ", "").lower()]
        assert strat == ["2. SW Integration Strategy"], \
            f"템플릿 시트를 못 찾아 사본이 생겼다: {wb.sheetnames}"
        cells = [c.value for row in wb[strat[0]].iter_rows() for c in row if c.value]
        assert "Ap_Door_Run" in cells, "템플릿 시트가 채워지지 않았다"

    def test_no_context_means_no_silent_empty_sheets(self, tmp_path):
        """컨텍스트를 안 주면 시트를 **만들지 않는다** — 빈 시트는 '근거 0' 으로 위장한다."""
        import openpyxl

        out = tmp_path / "plain.xlsx"
        generate_sits_xlsm(None, _itcs(), str(out))
        assert "Related_ID 확인" not in openpyxl.load_workbook(str(out)).sheetnames

    def test_sheet_failure_does_not_kill_the_spec_sheet(self, tmp_path):
        """근거 시트가 깨져도 규격 시트는 저장돼야 한다(그리고 사유가 남아야 한다)."""
        out = tmp_path / "boom.xlsx"
        stats: dict = {}
        generate_sits_xlsm(None, _itcs(), str(out), flows=[{"entry_fn": "x"}],
                           strategy_context={"calls_map": "not-a-dict", "stats_out": stats})
        assert validate_sits_xlsm(str(out))["stats"].get("tc_count") == 1
        assert stats.get("strategy_error"), "실패가 조용히 삼켜졌다"
