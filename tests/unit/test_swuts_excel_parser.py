"""SwUTS / SwITS xlsm parser 단위 테스트 (60차 F6-A).

합성 xlsm으로 양식 시뮬레이션 + parse_swuts_xlsm 검증. 라이브 검증은
``.codex_tmp/round_60_local_build/sanity_parse_swuts.py``에서 별도 수행.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.services.swuts_excel_parser import (  # noqa: E402
    XLSM_MAX_BYTES,
    parse_swuts_xlsm,
)


def _build_swuts_xlsm(
    sheet_name: str = "2.SW Unit Test Spec",
    *,
    headers: list[str] | None = None,
    data_rows: list[list[str | None]] | None = None,
    header_row: int = 4,
) -> bytes:
    """합성 xlsm — Hyundai SwUTS 양식 시뮬레이션.

    Args:
        sheet_name: TC 시트명 (정규식 매칭 대상). default = KJPDS02 형식.
        headers: header_row에 stamp할 라벨 list (1-based col 순서).
        data_rows: header 다음 row부터 stamp할 data row list.
        header_row: header가 위치할 row (1-based).
    """
    from openpyxl import Workbook  # type: ignore

    wb = Workbook()
    # 기본 'Sheet' 삭제 + TC 시트 추가 (시트명 정확 매칭 위해)
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
    ws = wb.create_sheet(sheet_name)

    if headers:
        for col_idx, label in enumerate(headers, 1):
            ws.cell(header_row, col_idx).value = label

    if data_rows:
        for r_idx, row in enumerate(data_rows, header_row + 1):
            for col_idx, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(r_idx, col_idx).value = val

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSwUTSExcelParser:
    def test_empty_bytes_returns_ok_false(self):
        result = parse_swuts_xlsm(b"")
        assert not result.ok
        assert any("비어있음" in w for w in result.parse_warnings)

    def test_oversize_rejected(self):
        oversized = b"x" * (XLSM_MAX_BYTES + 1)
        result = parse_swuts_xlsm(oversized)
        assert not result.ok
        assert any("DoS 방지" in w for w in result.parse_warnings)

    def test_kjpds02_swuts_format_basic(self):
        """KJPDS02 SwUTS v1.01 양식 — TC_ID 4자리 (SwUTC_0121) + 헤더 row 4."""
        xlsm = _build_swuts_xlsm(
            sheet_name="2.SW Unit Test Spec",
            headers=[
                "Index", "TC_ID", "Unit", "Test Method",
                "Test Case Generation Method",
            ],
            data_rows=[
                # TC 메타 row (sub-TC index col 비움)
                ["1", "SwUTC_0121", "s_safe_rotr", None, None],
                # sub-TC row (TC_ID 비우고 method 채움)
                ["1", None, "s_safe_rotr", "REQ", "ABV"],
                ["1", None, "s_safe_rotr", "REQ", "ABV"],
                ["2", "SwUTC_0122", "s_sha256_transform", None, None],
                ["2", None, "s_sha256_transform", "REQ", "ABV"],
            ],
        )
        result = parse_swuts_xlsm(xlsm)
        assert result.ok, result.parse_warnings
        assert len(result.entries) == 2  # sub-TC는 별도 entry 아님 (merge)
        e = result.by_tc_id["SwUTC_0121"]
        assert e.unit_name == "s_safe_rotr"
        assert e.test_method == "REQ"  # sub-TC row에서 merge
        assert e.generation_method == "ABV"
        # function_id KJPDS02 패턴 fallback (SwUTC_0121 → SwUFn_0121)
        assert e.function_id == "SwUFn_0121"

    def test_hdpdm01_suts_format_long_tcid(self):
        """HDPDM01 SUTS v3.01 양식 — TC_ID = SwUTC_SwUFn_0101 (long form).

        header에 'Test Case Generation Method' 라벨이 2 col (multiple col) 등장.
        실제 spec 데이터는 두 번째 col에만 있는 양식 (col 9 empty, col 12 채움).
        """
        xlsm = _build_swuts_xlsm(
            sheet_name="2.SW Unit Test Spec",
            header_row=6,
            headers=[
                "Component", "TC ID", "Name", "Description", "Safety Related",
                "Test Environment", "Test Method",
                "Test Case Generation Method",  # col 8 (첫 매칭)
                "Precondition",  # col 9
                "Sequence",      # col 10
                "Test Case Generation Method",  # col 11 (두 번째 매칭)
            ],
            data_rows=[
                # 메타 row + first sub-TC 데이터 일부
                [
                    "SwCom_01", "SwUTC_SwUFn_0101", "void main( void )",
                    "Main entry", None, None, None, None, None, "1",
                    "AEC, ABV",  # col 11 (실제 값) — col 8 (첫 매칭)은 비움
                ],
            ],
        )
        result = parse_swuts_xlsm(xlsm)
        assert result.ok, result.parse_warnings
        e = result.by_tc_id["SwUTC_SwUFn_0101"]
        # function_id 추출 — TC_ID에 SwUFn_ substring 있음
        assert e.function_id == "SwUFn_0101"
        # multiple col 매칭 — col 11이 nonempty라 채택
        assert e.generation_method == "AEC, ABV"

    def test_no_tc_sheets_returns_ok_false(self):
        """시트명에 'Unit Test Spec' / 'Integration Test Spec' 패턴 없으면 skip."""
        xlsm = _build_swuts_xlsm(
            sheet_name="Cover",  # 매칭 안 되는 시트명
            headers=["TC_ID", "Description"],
            data_rows=[["SwUTC_0101", "test"]],
        )
        result = parse_swuts_xlsm(xlsm)
        assert not result.ok
        assert any("TC 시트 미발견" in w for w in result.parse_warnings)

    def test_customer_sts_sheet_now_parses(self):
        """고객 STS 시트 '3.SW Test Spec'(Unit/Integration 없음)도 파싱된다 — _TC_SHEET_RE에 SW 추가."""
        xlsm = _build_swuts_xlsm(
            sheet_name="3.SW Test Spec",
            header_row=6,
            headers=[
                "Index", "Test Case ID", "Title", "Safety Related", "Test Environment",
                "Test Method", "Test Case Generation Method", "FS_REQ",
                "Description", "Pre-condition", "Test Action(Sequence)",
                "Expected Result", "SRS",
            ],
            data_rows=[
                [None, "SwTC_SwEI_01_01", "Battery 전압 모니터", None, None,
                 "FNCT", None, None, "Battery 전압을 주기적으로 모니터한다", "TRACE 연결", None, None, "SwEI_01"],
            ],
        )
        result = parse_swuts_xlsm(xlsm)
        assert result.ok, result.parse_warnings
        e = result.by_tc_id["SwTC_SwEI_01_01"]
        assert e.description.startswith("Battery 전압")
        assert e.precondition == "TRACE 연결"
        assert e.test_method == "FNCT"

    def test_sts_test_action_and_expected_extracted(self):
        """STS 'Test Action(Sequence)'/'Expected Result' → entry.test_action/expected 채움(라운드 후속).

        기존엔 _LABEL_MAP에 없어 raw_inputs(디버그)로만 흘러 콘텐츠 카드에 미표시였다."""
        xlsm = _build_swuts_xlsm(
            sheet_name="3.SW Test Spec",
            header_row=6,
            headers=[
                "Index", "Test Case ID", "Title", "Safety Related", "Test Environment",
                "Test Method", "Test Case Generation Method", "FS_REQ",
                "Description", "Pre-condition", "Test Action(Sequence)",
                "Expected Result", "SRS",
            ],
            data_rows=[
                [None, "SwTC_SwEI_02_01", "과전압 차단", None, None,
                 "FNCT", None, None, "과전압 시 릴레이 차단", "정상 전압 상태",
                 "1) 전압을 5.5V로 인가한다 2) 100ms 대기", "릴레이 OFF, DTC 0xC101 set", "SwEI_02"],
            ],
        )
        result = parse_swuts_xlsm(xlsm)
        assert result.ok, result.parse_warnings
        e = result.by_tc_id["SwTC_SwEI_02_01"]
        assert e.test_action.startswith("1) 전압을 5.5V")
        assert e.expected == "릴레이 OFF, DTC 0xC101 set"
        # 기존 필드 회귀 없음
        assert e.description.startswith("과전압")
        assert e.precondition == "정상 전압 상태"

    def test_sequence_still_maps_sub_index_not_test_action(self):
        """plain 'Sequence' 헤더는 여전히 sub_index로 매핑 — test_action이 탈취하지 않음(순회순서+break)."""
        xlsm = _build_swuts_xlsm(
            sheet_name="2.SW Unit Test Spec",
            headers=["TC_ID", "Description", "Test Method", "Sequence"],
            data_rows=[["SwUTC_0009", "desc here now", "REQ", "2"]],
        )
        result = parse_swuts_xlsm(xlsm)
        assert result.ok, result.parse_warnings
        e = result.by_tc_id["SwUTC_0009"]
        assert e.sub_index == "2"
        assert e.test_action == ""   # 'Sequence'가 test_action으로 새지 않음

    def test_sheet_regex_widen_no_false_positive(self):
        """SW 추가로 STS는 잡되 SITS/SUTS 매칭 불변 + Strategy/Environment 오탐 없음."""
        from backend.services.swuts_excel_parser import _TC_SHEET_RE
        assert _TC_SHEET_RE.search("3.SW Test Spec")               # 신규(STS)
        assert _TC_SHEET_RE.search("3.SW Integration Test Spec")   # SITS(Integration)
        assert _TC_SHEET_RE.search("2.SW Unit Test Spec")          # SUTS(Unit)
        assert not _TC_SHEET_RE.search("2.SW Integration Strategy")  # Test Spec 없음
        assert not _TC_SHEET_RE.search("2.Test Environment")       # Test Spec 없음

    def test_header_label_variants_normalized(self):
        """'TC ID' / 'TC_ID' / 'tc id' 모두 동일 라벨로 매칭."""
        xlsm = _build_swuts_xlsm(
            sheet_name="2.SW Unit Test Spec",
            headers=["tc id", "DESCRIPTION", "Pre-Condition"],
            data_rows=[["SwUTC_0001", "test description", "test pre"]],
        )
        result = parse_swuts_xlsm(xlsm)
        assert result.ok, result.parse_warnings
        e = result.by_tc_id["SwUTC_0001"]
        assert e.description == "test description"
        assert e.precondition == "test pre"

    def test_description_digit_fallback_to_next_col_round_live_nw15(self):
        """F6 라이브 검증 NW15 fix: HDPDM01 SITS v2.02 양식 layout 결함 대응.
        sub-TC row의 description col이 단순 digit (sub-index)이면 인접 col에서
        진짜 description 텍스트 fallback. 사용자 결정 (자동 감지 fallback).
        """
        from backend.services.swuts_excel_parser import parse_swuts_xlsm

        # HDPDM01 SITS 양식 시뮬레이션 — description col(C3) 다음 C4가 진짜 텍스트
        # header: TC ID(C2) / Description(C3) / (C4 unmapped) / Test Method(C5)
        xlsm_bytes = _build_swuts_xlsm(
            headers=["", "TC ID", "Description", "", "Test Method"],
            data_rows=[
                ["", "SwITC_01", "1", "main -> Init -> System", "AEC, ABV"],
                ["", "SwITC_02", "2", "main -> SystemCheck -> Update", "AEC, ABV"],
            ],
        )
        result = parse_swuts_xlsm(xlsm_bytes)
        assert result.ok is True
        # description이 sub-index "1","2"가 아닌 실제 텍스트로 stamp
        entries_by_tc = {e.tc_id: e for e in result.entries}
        assert "SwITC_01" in entries_by_tc
        assert entries_by_tc["SwITC_01"].description == "main -> Init -> System", (
            f"NW15 회귀: description fallback 실패 — "
            f"실제 stamp: {entries_by_tc['SwITC_01'].description!r}"
        )
        assert entries_by_tc["SwITC_01"].sub_index == "1"
        assert entries_by_tc["SwITC_02"].description == "main -> SystemCheck -> Update"
        assert entries_by_tc["SwITC_02"].sub_index == "2"

    def test_description_text_not_affected_by_nw15_fix(self):
        """NW15 fix는 KJPDS02 SwITS처럼 description이 텍스트인 양식에 영향 없음."""
        from backend.services.swuts_excel_parser import parse_swuts_xlsm

        xlsm_bytes = _build_swuts_xlsm(
            headers=["", "TC ID", "Description", "", "Test Method"],
            data_rows=[
                ["", "SwITC_01", "Interface : main -> System -> Init", "", "REQ, IFT"],
            ],
        )
        result = parse_swuts_xlsm(xlsm_bytes)
        assert result.ok is True
        entry = result.entries[0]
        assert entry.description == "Interface : main -> System -> Init"
        assert entry.sub_index == ""  # description은 digit 아니므로 sub-index 분리 안 함

    def test_first_row_sub_tc_emits_warning_round5_nf1(self):
        """F6 Round 5 NF1 fix: 양식 변종에서 첫 data row가 TC_ID 없는 sub-TC면
        직전 메타 row 없어 silent drop. parse_warnings emit으로 audit reviewer
        인지 가능 (silent → audible 전환).
        """
        from backend.services.swuts_excel_parser import parse_swuts_xlsm

        # header row (3+ 라벨 매칭 필요) + 첫 data row TC_ID 없음 + 두번째 TC_ID 있음
        xlsm_bytes = _build_swuts_xlsm(
            headers=["TC ID", "Description", "Test Method", "Generation Method"],
            data_rows=[
                ["", "desc-orphan", "method-orphan", "gen-orphan"],  # TC_ID 없음 — orphan
                ["SwUTC_0001", "desc-1", "method-1", "gen-1"],
            ],
        )
        result = parse_swuts_xlsm(xlsm_bytes)
        assert result.ok is True, f"parse 실패: {result.parse_warnings}"
        # 정상 entry는 1건 (SwUTC_0001)
        assert len(result.entries) == 1
        assert result.entries[0].tc_id == "SwUTC_0001"
        # NF1: orphan sub-TC row drop 사유가 parse_warnings에 누적
        drop_warnings = [
            w for w in result.parse_warnings
            if "TC_ID 없는 sub-TC" in w and "직전 메타 row 없음" in w
        ]
        assert len(drop_warnings) == 1, (
            f"NF1 회귀: orphan sub-TC row silent drop — warning 누락. "
            f"parse_warnings: {result.parse_warnings}"
        )


class TestRound82Kjpds02FormatCompat:
    """라운드 82 — KJPDS02 양식 (v1.01) 호환 검증 회귀.

    라이브 진단 결과 swuts_excel_parser가 KJPDS02 양식 자동 호환:
    - SwUTS sheet '2.SW Unit Test Spec' (HDPDM01과 동일) — 285 entries
    - SwITS sheet '3. SW Integration Test Spec ' (공백 포함) — 47 entries
    - TC ID 'SwUTC_NNNN' (SwUFn_ prefix 부재) → function_id fallback `SwUFn_NNNN`
    - Header row 4 / col 4 'Unit' (HDPDM01은 row 6 / col 4 'Name') 자동 detect
    - Test Method col 5 / Generation Method col 6 / Inpt[N] col 7~ 변종 layout
    """

    def test_kjpds02_swits_sheet_with_trailing_space(self):
        """KJPDS02 SwITS 시트명 '3. SW Integration Test Spec ' (trailing space) 자동 detect."""
        from backend.services.swuts_excel_parser import parse_swuts_xlsm

        xlsm_bytes = _build_swuts_xlsm(
            sheet_name="3. SW Integration Test Spec ",  # 끝에 공백
            header_row=6,
            headers=[
                "Index", "대응 환경명", "Test Method", "TestCase Generation Method",
                "TC_ID", " ", "Description", "Precondition",
            ],
            data_rows=[
                ["1", "SWIT_SWUFN_0101_DEPTH4_FILE12", "REQ,IFT", "AOR,ABV",
                 "SwITC_0101_01", "1", "Interface : main -> Init", None],
            ],
        )
        result = parse_swuts_xlsm(xlsm_bytes)
        assert result.ok, result.parse_warnings
        # 시트명 trailing space에도 TC entry 추출 성공
        assert "SwITC_0101_01" in result.by_tc_id

    def test_kjpds02_unit_col_name_distinct_from_hdpdm01(self):
        """KJPDS02 col 4 = 'Unit' (함수명) vs HDPDM01 col 4 = 'Name' (TC title) 양식 변종.

        둘 다 unit_name 필드로 매핑 가능 — header label 'Unit' 호환 확인.
        """
        from backend.services.swuts_excel_parser import parse_swuts_xlsm

        xlsm_bytes = _build_swuts_xlsm(
            sheet_name="2.SW Unit Test Spec",
            header_row=4,
            headers=["Index", "TC_ID", "Unit", "Test Method", "Test Case Generation Method"],
            data_rows=[
                ["1", "SwUTC_0121", "s_safe_rotr", None, None],
                ["1", None, "s_safe_rotr", "REQ", "ABV"],
            ],
        )
        result = parse_swuts_xlsm(xlsm_bytes)
        assert result.ok, result.parse_warnings
        e = result.by_tc_id["SwUTC_0121"]
        # 'Unit' col이 unit_name으로 매핑 (Name과 등가)
        assert e.unit_name == "s_safe_rotr"
        # function_id fallback (SwUTC_0121 → SwUFn_0121)
        assert e.function_id == "SwUFn_0121"

    def test_kjpds02_tcid_short_form_function_id_fallback(self):
        """KJPDS02 TC_ID 'SwUTC_NNNN' (4자리, SwUFn_ 부재) → function_id 'SwUFn_NNNN' fallback."""
        from backend.services.swuts_excel_parser import parse_swuts_xlsm

        xlsm_bytes = _build_swuts_xlsm(
            sheet_name="2.SW Unit Test Spec",
            header_row=4,
            headers=["Index", "TC_ID", "Unit", "Test Method", "Generation Method"],
            data_rows=[
                ["1", "SwUTC_0121", "s_safe_rotr", "REQ", "ABV"],
                ["2", "SwUTC_0250", "s_other_fn", "REQ", "ABV"],
            ],
        )
        result = parse_swuts_xlsm(xlsm_bytes)
        assert result.ok, result.parse_warnings
        # 짧은 TC ID 형식 (SwUTC_NNNN) → function_id 매핑
        assert result.by_tc_id["SwUTC_0121"].function_id == "SwUFn_0121"
        assert result.by_tc_id["SwUTC_0250"].function_id == "SwUFn_0250"
