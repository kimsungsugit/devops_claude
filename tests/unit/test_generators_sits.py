"""generators/sits.py — 합성 SwCom과 실제 요구 추적성의 분리.

배경: `_infer_swcom_id`는 모듈 **등장 순번**으로 `SwCom_XX`를 만든다(실제 SDS component ID가
아니다). 이 값이 모든 flow의 related_ids에 무조건 삽입되므로 "Related ID 보유율"은 사실상
항상 100%다. 그 수치를 요구 추적성으로 쓰면 요구 링크가 0건이어도 게이트를 통과한다.
"""
from __future__ import annotations

from generators.sits import (
    _DATA_START_ROW,
    _DESC_COL,
    _SPEC_SHEET_NAME,
    _TCID_COL,
    collect_integration_flows,
    generate_sits_quality_report,
    generate_sits_xlsm,
    validate_sits_xlsm,
)


def _fd_layered():
    """계층 진입점 모양 — `main` 은 **같은 모듈** 함수만 직접 호출한다.

    실제 경계 횡단은 두 홉 아래(`s_Init_Core` → `Drv_Spi_Write`)에서 일어난다.
    정본이 시험하는 통합지점 15개가 전부 이 모양이라 1홉 판정에서 통째로 빠졌다.
    """
    def _f(name, file, calls):
        return {"name": name, "file": file, "calls_list": list(calls),
                "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
                "asil": "B"}

    return {
        "F1": _f("main", "SysOs.c", ["s_System_Init"]),
        "F2": _f("s_System_Init", "SysOs.c", ["s_Init_Core"]),
        "F3": _f("s_Init_Core", "SysOs.c", ["Drv_Spi_Write"]),
        "F4": _f("Drv_Spi_Write", "Drv_Spi.c", []),
    }


def _fd(*, related_by_name=None):
    """cross-module 호출 1건을 갖는 최소 function_details."""
    related_by_name = related_by_name or {}
    return {
        "F1": {
            "name": "Ap_Door_Run",
            "file": "Ap_Door.c",
            "calls_list": ["Drv_Motor_Set"],
            "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
            "asil": "B",
            "related": related_by_name.get("Ap_Door_Run", ""),
        },
        "F2": {
            "name": "Drv_Motor_Set",
            "file": "Drv_Motor.c",
            "calls_list": [],
            "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
            "asil": "B",
            "related": related_by_name.get("Drv_Motor_Set", ""),
        },
    }


class TestSyntheticSwComIsMarked:
    def test_synthetic_id_is_recorded_at_insertion(self):
        """합성 여부는 삽입 지점에서 기록된다 — 소비자가 문자열 prefix로 추측하지 않도록."""
        flows = collect_integration_flows(_fd())
        assert flows, "cross-module flow가 생성되지 않았다"
        f = flows[0]
        assert f["related_ids"], "합성 ID가 항상 들어간다는 전제가 깨졌다"
        assert f["synthetic_related_ids"] == [f["swcom_id"]]
        assert f["swcom_id"] in f["related_ids"]


class TestTransitiveCrossModuleQualifies:
    """계층 진입점은 **같은 모듈 안쪽만** 직접 호출한다 — 1홉 판정이면 통째로 빠진다.

    실측(2026-08-14, KJPDS02_PV): 정본이 시험하는 통합지점 15개가 전부 이 경로로
    누락됐다 — `main`(직접 호출 2개가 둘 다 같은 모듈) · `s_SysMain_Init` ·
    `s_System_MainLoop` · `g_DrvIn_Main` · `g_UDS_RDBI_Paser` · `SCI0_ISR` 등.
    """

    def test_layered_entry_qualifies_via_transitive_reach(self):
        flows = collect_integration_flows(_fd_layered())
        entries = {f["entry_fn"] for f in flows}
        assert "main" in entries, f"계층 진입점이 후보에서 빠졌다: {sorted(entries)}"

    def test_transitive_qualification_is_labelled(self):
        """직접 경계와 전이 경계는 근거가 다르다 — 산출물이 구별할 수 있어야 한다."""
        flows = collect_integration_flows(_fd_layered())
        by_entry = {f["entry_fn"]: f for f in flows}
        assert by_entry["main"]["cross_via"] == "transitive"
        # 대조군: 직접 경계를 넘는 흐름은 `direct` 로 남는다
        direct = collect_integration_flows(_fd())
        assert direct[0]["cross_via"] == "direct"

    def test_hop_limit_is_respected(self):
        """무한 전이가 아니다 — 상한 밖의 경계는 자격이 아니다."""
        from generators.sits import _reach_cross_module

        calls = {"a": ["b"], "b": ["c"], "c": ["d"], "d": []}
        mods = {"a": "M1", "b": "M1", "c": "M1", "d": "M2"}
        assert _reach_cross_module("a", calls, mods, max_hop=2) == []
        assert _reach_cross_module("a", calls, mods, max_hop=3) == ["d"]

    def test_transitive_count_is_reported(self):
        """전이분이 0 이 되면 계층 진입점이 다시 빠진 것 — 회귀 신호로 쓰인다."""
        stats: dict = {}
        collect_integration_flows(_fd_layered(), stats_out=stats)
        assert stats["transitive_entries"] >= 1
        assert stats["transitive_entries_emitted"] >= 1
        assert stats["cross_reach_hops"] >= 1


class TestVarNamesFollowReferenceNotation:
    """변수 이름은 정본(VectorCAST) 표기를 따른다 — SUTS 와 **같은 출처**로.

    실측(2026-08-14, KJPDS02_PV): 정본 기대 656칸 중 **일치 0** · 입력 496 중 26(5.2%).
    원인은 수집 범위가 아니라 이름 표기였다 — 옛 `_clean_var_name` 이 `[IN]` 태그와
    함께 배열 첨자를 지우고, 타입을 버리는 대신 이름에 이어붙였다:
    `const UINT8 * data` → `const_UINT8_*_data`, `return U8` → `return_UINT8`.
    """

    def test_type_tokens_are_dropped_not_glued(self):
        from generators.sits import _clean_var_name

        assert _clean_var_name("[IN] const UINT8 * data") == "data"
        assert _clean_var_name("[IN] u8 u8g_Speed") == "u8g_Speed"

    def test_return_slot_is_the_word_return(self):
        """정본은 반환값을 `return` 이라 적는다 — 타입 이름이 아니다."""
        from generators.sits import _clean_var_name

        assert _clean_var_name("[OUT] return U8") == "return"

    def test_pointer_member_uses_reference_notation(self):
        """정본은 `p->m` 을 `p[0].m` 으로 적는다(VectorCAST 가 1원소 배열을 잡아준다)."""
        from generators.sits import _clean_var_name

        assert _clean_var_name("[IN] LinTpMessageType * msg->cfNum") == "msg[0].cfNum"

    def test_unparseable_yields_empty_not_a_fragment(self):
        """못 뽑으면 빈 값이다 — 예전엔 `raw[:40]` 으로 원문 조각을 흘렸다."""
        from generators.sits import _clean_var_name

        assert _clean_var_name("") == ""
        assert _clean_var_name("[IN] ///") == ""

    def test_global_tags_beyond_in_out_are_stripped(self):
        """전역은 `[INDIRECT]`·`[INDIRECT2]` 로도 온다 — 파라미터 정제기는 그걸 못 벗긴다.

        실측(2026-08-14): 두 경로를 한 함수로 합쳤더니 정본과 맞던 입력 **9칸**이
        통째로 사라졌다(`u8s_E2EInitFlag_SBCM0`·`g_DoorState`·`u32s_SecuritySeed` 등
        전부 전역). 형태 검사에서 남은 대괄호 때문에 버려진 것이다.
        """
        from generators.sits import _clean_global_var_name, _clean_var_name

        assert _clean_global_var_name("[INDIRECT] u8s_E2EInitFlag_SBCM0") == "u8s_E2EInitFlag_SBCM0"
        assert _clean_global_var_name("[INDIRECT2] g_DoorState") == "g_DoorState"
        assert _clean_global_var_name("[OUT] u8 u32s_SecuritySeed") == "u32s_SecuritySeed"
        # 대조군 — 파라미터 정제기는 이 태그를 벗기지 못한다(그래서 함수를 나눴다)
        assert _clean_var_name("[INDIRECT] u8s_E2EInitFlag_SBCM0") == ""

    def test_global_path_also_uses_reference_pointer_notation(self):
        """전역 경로에도 `p->m` → `p[0].m` 이 걸린다.

        (뮤테이션 생존으로 드러난 공백 — 파라미터 경로만 검사하면 전역 쪽 교정을
        지워도 테스트가 전부 초록이다.)
        """
        from generators.sits import _clean_global_var_name

        assert _clean_global_var_name("[INDIRECT] g_Ctx->state") == "g_Ctx[0].state"

    def test_unparseable_input_does_not_enter_as_blank(self):
        """이름을 못 뽑은 항목이 **빈 칸**으로 산출물에 들어가면 안 된다.

        (뮤테이션 생존으로 드러난 공백 — 빈 값 가드를 지워도 잡히지 않았다.)
        """
        fd = {
            "F1": {"name": "Entry_Fn", "file": "A.c", "calls_list": ["Callee_Fn"],
                   "inputs": ["[IN] ///", "[IN] u8 u8g_Ok"], "outputs": [],
                   "globals_global": [], "globals_static": [], "asil": "B"},
            "F2": {"name": "Callee_Fn", "file": "B.c", "calls_list": [],
                   "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
                   "asil": "B"},
        }
        flows = collect_integration_flows(fd)
        iv = flows[0]["input_vars"]
        assert "u8g_Ok" in iv
        assert "" not in iv, f"빈 이름이 실렸다: {iv}"

    def test_global_inputs_survive_collection(self):
        """단위 함수 수준이 아니라 **흐름 산출물**에서 살아남는지 본다."""
        fd = {
            "F1": {"name": "Entry_Fn", "file": "A.c", "calls_list": ["Callee_Fn"],
                   "inputs": [], "outputs": [],
                   "globals_global": ["[INDIRECT] u8s_E2EInitFlag_SBCM0"],
                   "globals_static": [], "asil": "B"},
            "F2": {"name": "Callee_Fn", "file": "B.c", "calls_list": [],
                   "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
                   "asil": "B"},
        }
        flows = collect_integration_flows(fd)
        assert "u8s_E2EInitFlag_SBCM0" in flows[0]["input_vars"]

    def test_expected_column_has_no_function_prefix(self):
        """정본 기대 1,172칸 중 `함수명()` 접두는 0건 — 붙이면 한 칸도 안 맞는다."""
        fd = {
            "F1": {"name": "Entry_Fn", "file": "A.c", "calls_list": ["Callee_Fn"],
                   "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
                   "asil": "B"},
            "F2": {"name": "Callee_Fn", "file": "B.c", "calls_list": [],
                   "inputs": [], "outputs": ["[OUT] u8 u8g_Result"],
                   "globals_global": ["[OUT] u8 g_State"], "globals_static": [], "asil": "B"},
        }
        flows = collect_integration_flows(fd)
        expected = flows[0]["expected_vars"]
        assert "u8g_Result" in expected, expected
        assert not any("(" in v for v in expected), f"함수명 접두가 남았다: {expected}"


class TestFlowStatsReachTheReport:
    """생산자가 낸 흐름 통계가 **리포트까지 도달**하는가.

    "보고를 추가했다" 와 "보고가 도달한다" 는 다른 문제다. 리포트가 자기 화이트리스트를
    따로 들고 있어 새 키가 조용히 버려진 일이 두 번 있었다(2026-07-31 `sds_*`,
    2026-08-14 전이/체인 축 — 후자는 캡에 잘린 정본 지점의 원인 규명을 한 라운드
    늦췄다). 목록을 `_FLOW_COV_KEYS` 하나로 묶고 그 정합을 여기서 강제한다.
    """

    def test_no_flow_stat_is_silently_dropped(self):
        from generators.sits import _FLOW_COV_KEYS

        stats: dict = {}
        collect_integration_flows(_fd_layered(), max_flows=1, stats_out=stats)
        # 흐름 축 = SDS/SwUDS 보강 축을 뺀 나머지(그 둘은 별도 dict 로 나간다)
        flow_keys = {k for k in stats if not k.startswith(("sds_", "uds_swcom_"))}
        missing = sorted(flow_keys - set(_FLOW_COV_KEYS))
        assert not missing, f"리포트에 실리지 않는 흐름 통계: {missing}"

    def test_new_axes_are_actually_in_the_payload(self):
        """목록에 있는 것과 실제로 실리는 것은 또 다르다 — 산출 payload 로 확인."""
        stats: dict = {}
        collect_integration_flows(_fd_layered(), max_flows=1, stats_out=stats)
        cov = generate_sits_quality_report(
            [], total_source_functions=4, flow_stats=stats)["integration_flow_coverage"]
        for key in ("transitive_entries", "cross_reach_hops",
                    "chain_truncated_flows", "dropped_in_design_doc_count"):
            assert key in cov, f"{key} 가 리포트에 없다"


class TestCapKeepsDesignDocumentedFlows:
    """캡이 걸릴 때 **설계 문서(SwUDS)에 등재된** 흐름을 먼저 지킨다.

    실측(2026-08-14, KJPDS02_PV): 후보 367 · 캡 200 에서 정본 통합지점 34개 중
    알파벳순으로는 30개만 살아남았다 — 후보가 늘면 캡이 **기존 정답을 밀어낸다**.
    등재 여부를 키에 넣으면 34/34 가 생존한다(정본 지점은 100% 등재, 후보 전체는 83.4%).
    """

    @staticmethod
    def _fd_many(n):
        out = {}
        for i in range(n):
            out[f"F{i}"] = {
                "name": f"fn_{i:03d}", "file": f"M{i}.c", "calls_list": [f"callee_{i:03d}"],
                "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
                "asil": "QM",
            }
            out[f"C{i}"] = {
                "name": f"callee_{i:03d}", "file": f"Other{i}.c", "calls_list": [],
                "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
                "asil": "QM",
            }
        return out

    def test_documented_flow_survives_the_cap_over_alphabetical_order(self):
        """알파벳 꼴찌라도 등재돼 있으면 살아남는다 — 이게 없으면 순서가 곧 운이다."""
        fd = self._fd_many(6)
        # 알파벳 마지막(fn_005)만 등재
        flows = collect_integration_flows(
            fd, max_flows=1, uds_swcom_map={"fn_005": ["SwCom_09"]})
        assert [f["entry_fn"] for f in flows] == ["fn_005"]

    def test_undocumented_are_dropped_first_and_counted(self):
        stats: dict = {}
        fd = self._fd_many(6)
        collect_integration_flows(
            fd, max_flows=2, stats_out=stats,
            uds_swcom_map={"fn_004": ["SwCom_09"], "fn_005": ["SwCom_10"]})
        assert set(stats["dropped_entry_fns"]) == {"fn_000", "fn_001", "fn_002", "fn_003"}
        assert stats["dropped_in_design_doc_count"] == 0

    def test_dropping_a_documented_flow_is_reported(self):
        """등재분까지 잘리면 캡 값을 다시 볼 신호 — 그 사실이 수치로 나와야 한다."""
        stats: dict = {}
        fd = self._fd_many(4)
        collect_integration_flows(
            fd, max_flows=1, stats_out=stats,
            uds_swcom_map={f"fn_{i:03d}": ["SwCom_01"] for i in range(4)})
        assert stats["dropped_in_design_doc_count"] == 3

    def test_dropping_a_documented_flow_warns(self, caplog):
        """수치만 있고 아무도 안 보면 소용없다 — 로그로도 말한다."""
        import logging

        with caplog.at_level(logging.WARNING, logger="generators.sits"):
            collect_integration_flows(
                self._fd_many(4), max_flows=1,
                uds_swcom_map={f"fn_{i:03d}": ["SwCom_01"] for i in range(4)})
        msg = " ".join(r.getMessage() for r in caplog.records)
        assert "SwUDS 에 등재된" in msg

    def test_no_documented_drop_no_extra_warning(self, caplog):
        """음성 대조군 — 미등재분만 잘리면 그 경고는 안 나온다(경고 피로 방지)."""
        import logging

        with caplog.at_level(logging.WARNING, logger="generators.sits"):
            collect_integration_flows(
                self._fd_many(4), max_flows=1, uds_swcom_map={"fn_000": ["SwCom_01"]})
        msgs = [r.getMessage() for r in caplog.records if "SwUDS 에 등재된" in r.getMessage()]
        assert not msgs

    def test_asil_still_outranks_documentation(self):
        """안전 등급이 먼저다 — 등재 여부가 ASIL 을 뒤집으면 안 된다."""
        fd = self._fd_many(2)
        fd["F0"]["asil"] = "B"          # 미등재이지만 안전 관련
        flows = collect_integration_flows(
            fd, max_flows=1, uds_swcom_map={"fn_001": ["SwCom_09"]})
        assert [f["entry_fn"] for f in flows] == ["fn_000"]


class TestCallChainIsWholePath:
    """정본은 경로 **전체**를 적는다(최대 92홉, 함수 재등장 0/54)."""

    def test_chain_spans_the_whole_tree_not_just_direct_callees(self):
        flows = collect_integration_flows(_fd_layered())
        chain = next(f["call_chain"] for f in flows if f["entry_fn"] == "main")
        assert chain.split(" -> ") == [
            "main", "s_System_Init", "s_Init_Core", "Drv_Spi_Write",
        ], chain

    def test_cycles_do_not_repeat_or_hang(self):
        from generators.sits import _build_call_chain_nodes

        calls = {"a": ["b"], "b": ["c"], "c": ["a", "b"]}
        nodes, dropped = _build_call_chain_nodes("a", calls, max_nodes=50)
        assert nodes == ["a", "b", "c"] and dropped == 0

    def test_truncation_is_counted_not_silent(self):
        """자르는 건 괜찮다 — 자른 걸 **말하지 않는** 게 문제다."""
        from generators.sits import _build_call_chain_nodes

        calls = {"a": ["b", "c", "d"], "b": [], "c": [], "d": []}
        nodes, dropped = _build_call_chain_nodes("a", calls, max_nodes=2)
        assert nodes == ["a", "b"]
        assert dropped == 2, "잘린 함수 수가 0 이면 소비처가 '경로 전부' 로 읽는다"

    def test_truncated_flow_carries_the_count(self):
        stats: dict = {}
        flows = collect_integration_flows(_fd_layered(), stats_out=stats)
        assert all("chain_dropped" in f for f in flows)
        assert stats["chain_max_nodes"] >= 1


class TestUdsSwComIsTheRealSource:
    """Related 칸의 SwCom 은 **SwUDS** 에서 온다.

    정본 실측(KJPDS02_PV_SwITS v1.02): Related 어휘는 SwCom 170회(33종)·SwFn 69·
    SwSTR 62·SwST 38·SwTK 8 이고 요구 ID 는 0 건이다. 그리고 SwUDS 에서 뽑은 SwCom
    33종은 정본 33종과 **차집합 양쪽 0** — 정본이 근거로 삼는 표가 바로 그것이다.
    (SDS 파티션 맵에는 SwCom 축이 아예 없다: 함수 588개의 `related` 는 전부 요구 ID)
    """

    def test_uds_swcom_lands_in_related_ids(self):
        flows = collect_integration_flows(
            _fd(), uds_swcom_map={"ap_door_run": ["SwCom_13", "SwCom_14"]})
        f = flows[0]
        assert "SwCom_13" in f["related_ids"] and "SwCom_14" in f["related_ids"]

    def test_uds_swcom_is_not_marked_synthetic(self):
        """문서 유래이므로 추적성 분자에 들어가야 한다 — 합성으로 찍히면 0% 로 샌다."""
        flows = collect_integration_flows(
            _fd(), uds_swcom_map={"ap_door_run": ["SwCom_13"]})
        assert flows[0]["synthetic_related_ids"] == []

    def test_synthetic_id_is_not_appended_next_to_a_real_one(self):
        """진짜 SwCom 이 있으면 순번 합성값을 **덧붙이지 않는다**.

        덧붙이면 한 칸에 문서 유래 SwCom 과 다른 컴포넌트를 가리키는 합성값이 나란히
        실리고, 셀만 보는 쪽은 구별할 수 없다(합성 표시는 셀에 없다).
        """
        flows = collect_integration_flows(
            _fd(), uds_swcom_map={"ap_door_run": ["SwCom_13"]})
        f = flows[0]
        assert f["related_ids"] == ["SwCom_13"], f"합성값이 섞였다: {f['related_ids']}"
        assert f["swcom_id"] not in f["related_ids"]

    def test_missing_function_falls_back_to_marked_synthetic(self):
        """맵에 없는 함수는 합성으로 내려가되 **합성임이 표시**된다."""
        flows = collect_integration_flows(
            _fd(), uds_swcom_map={"someone_else": ["SwCom_13"]})
        f = flows[0]
        assert f["synthetic_related_ids"] == [f["swcom_id"]]
        assert "SwCom_13" not in f["related_ids"]

    def test_enrichment_yield_is_reported(self):
        """0 건이면 0 건이라고 말해야 한다 — 침묵하면 '보강이 돈다'로 읽힌다."""
        stats: dict = {}
        collect_integration_flows(_fd(), stats_out=stats,
                                  uds_swcom_map={"ap_door_run": ["SwCom_13"]})
        assert stats["uds_swcom_lookups"] >= 1
        assert stats["uds_swcom_hits"] == 1
        assert stats["uds_swcom_ids"] == 1

        empty: dict = {}
        collect_integration_flows(_fd(), stats_out=empty, uds_swcom_map={})
        assert empty["uds_swcom_hits"] == 0
        assert empty["uds_swcom_map_entries"] == 0


class TestLoadUdsSwComMap:
    def test_blank_and_missing_paths_are_empty_maps(self, tmp_path):
        from generators.sits import load_uds_swcom_map

        assert load_uds_swcom_map(None) == {}
        assert load_uds_swcom_map("") == {}
        assert load_uds_swcom_map(str(tmp_path / "nope.docx")) == {}

    def test_unreadable_docx_does_not_raise(self, tmp_path):
        """추출 실패는 빈 맵이다 — 생성 전체를 세우지 않는다(합성으로 내려간다)."""
        from generators.sits import load_uds_swcom_map

        p = tmp_path / "broken.docx"
        p.write_bytes(b"not a docx")
        assert load_uds_swcom_map(str(p)) == {}


class TestQualityReportSeparatesAxes:
    @staticmethod
    def _itcs(*, real_ids):
        """related_ids = 합성 1개 + real_ids."""
        return [{
            "tc_id": "SwITC_01",
            "related_ids": ["SwCom_01", *real_ids],
            "synthetic_related_ids": ["SwCom_01"],
            "sub_cases": [], "input_vars": [], "expected_vars": [],
            "gen_method": "ABV",
        }]

    def test_synthetic_only_is_not_traceability(self):
        """합성 ID만 있으면 Related 보유율 100%, 요구 추적성 0%."""
        qr = generate_sits_quality_report(self._itcs(real_ids=[]), total_source_functions=2)
        assert qr["related_coverage_pct"] == 100.0
        assert qr["requirement_traceability_pct"] == 0.0
        assert qr["synthetic_only_related_count"] == 1

    def test_real_id_counts_as_traceability(self):
        qr = generate_sits_quality_report(
            self._itcs(real_ids=["SwTR_012"]), total_source_functions=2,
        )
        assert qr["requirement_traceability_pct"] == 100.0
        assert qr["with_requirement_trace_count"] == 1
        assert qr["synthetic_only_related_count"] == 0

    def test_sds_sourced_swcom_is_not_treated_as_synthetic(self):
        """문서(SDS)에서 온 SwCom ID는 합성이 아니다 — prefix로 뭉뚱그리지 않는다."""
        itcs = [{
            "tc_id": "SwITC_01",
            "related_ids": ["SwCom_07"],       # SDS 유래
            "synthetic_related_ids": [],       # 삽입 지점이 합성으로 기록하지 않았다
            "sub_cases": [], "input_vars": [], "expected_vars": [], "gen_method": "ABV",
        }]
        qr = generate_sits_quality_report(itcs, total_source_functions=1)
        assert qr["requirement_traceability_pct"] == 100.0
        assert qr["synthetic_only_related_count"] == 0

    def test_legacy_itc_without_marker_is_not_silently_credited(self):
        """marker 필드가 없는 구 데이터는 related_ids를 그대로 신뢰한다(하위호환).

        구 경로에서 만들어진 ITC는 합성 여부를 알 수 없다. 여기서 임의로 prefix 추측을
        하면 SDS 유래 ID까지 깎아내리므로, 판정은 생산 지점 기록에만 의존한다.
        """
        itcs = [{
            "tc_id": "SwITC_01", "related_ids": ["SwCom_01"],
            "sub_cases": [], "input_vars": [], "expected_vars": [], "gen_method": "ABV",
        }]
        qr = generate_sits_quality_report(itcs, total_source_functions=1)
        assert qr["requirement_traceability_pct"] == 100.0


# ---------------------------------------------------------------------------
# max_flows 캡 — 침묵 절단 + 안전등급 무관 선별
# ---------------------------------------------------------------------------

def _fd_many(n: int, *, asil_of=None):
    """entry n개(각각 cross-module 호출 1건)를 갖는 function_details.

    이름은 `Ap_F000`… 형태라 알파벳순 = 번호순이다(경계 확인이 쉬워진다).
    """
    asil_of = asil_of or (lambda i: "QM")
    fd = {}
    for i in range(n):
        fd[f"E{i:03d}"] = {
            "name": f"Ap_F{i:03d}",
            "file": f"Ap_Mod{i:03d}.c",
            "calls_list": ["Drv_Common"],
            "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
            "asil": asil_of(i),
            "related": "",
        }
    fd["DRV"] = {
        "name": "Drv_Common",
        "file": "Drv_Common.c",
        "calls_list": [],
        "inputs": [], "outputs": [], "globals_global": [], "globals_static": [],
        "asil": "QM", "related": "",
    }
    return fd


class TestFlowCapIsSurfaced:
    """캡이 물면 **몇 개가 잘렸는지** 남아야 한다.

    회귀 대상: 수집 루프가 `len(flows) >= max_flows` 에서 그냥 break 했다. 캡 이후
    후보는 세어지지도 않아 소비처에서 결과 길이로 되짚어도 절단을 알 수 없었다.
    실측(KJPDS02 계열 900함수): 145개 중 25개가 조용히 사라졌고 7개가 ASIL A였다.
    """

    def test_stats_out_reports_pre_cap_total(self):
        stats = {}
        flows = collect_integration_flows(_fd_many(10), max_flows=4, stats_out=stats)
        assert len(flows) == 4
        assert stats["total_flows_found"] == 10, "캡 **전** 총량이 안 남았다"
        assert stats["flows_emitted"] == 4
        assert stats["flows_dropped"] == 6
        assert stats["flow_emit_pct"] == 40.0
        assert len(stats["dropped_entry_fns"]) == 6

    def test_no_truncation_reports_zero(self):
        """대조군: 캡에 안 닿으면 제외 0건이고 경고도 없다."""
        stats = {}
        collect_integration_flows(_fd_many(3), max_flows=100, stats_out=stats)
        assert stats["flows_dropped"] == 0
        assert stats["dropped_entry_fns"] == []
        assert stats["dropped_safety_related_count"] == 0

    def test_truncation_is_logged(self, caplog):
        with caplog.at_level("WARNING", logger="generators.sits"):
            collect_integration_flows(_fd_many(10), max_flows=4)
        assert "max_flows" in caplog.text and "제외" in caplog.text, caplog.text

    def test_stats_out_is_optional(self):
        """기존 호출부(stats_out 없음)는 그대로 동작해야 한다."""
        assert len(collect_integration_flows(_fd_many(5), max_flows=2)) == 2


class TestFlowCapKeepsSafetyFirst:
    """자를 때 ASIL 높은 흐름을 먼저 버리면 안 된다 (ISO 26262).

    회귀 대상: 정렬 키가 함수명 알파벳순뿐이라 어느 흐름이 살아남는지가 안전등급과
    완전히 무관했다. 실측에서 ASIL A 7건이 QM 보다 먼저 잘려나갔다.
    """

    @staticmethod
    def _asil_of(i):
        # 알파벳 뒤쪽(번호 큰 쪽)에 안전등급을 몰아둔다 — 옛 로직이면 전부 잘린다.
        return {7: "D", 8: "C", 9: "A"}.get(i, "QM")

    def test_safety_flows_survive_the_cap(self):
        stats = {}
        flows = collect_integration_flows(
            _fd_many(10, asil_of=self._asil_of), max_flows=4, stats_out=stats)
        kept = {f["entry_fn"]: f["asil"] for f in flows}
        assert kept.get("Ap_F007") == "D", f"ASIL D가 잘렸다: {kept}"
        assert kept.get("Ap_F008") == "C", f"ASIL C가 잘렸다: {kept}"
        assert kept.get("Ap_F009") == "A", f"ASIL A가 잘렸다: {kept}"
        assert stats["dropped_safety_related_count"] == 0
        assert stats["dropped_asil_distribution"] == {"QM": 6}

    def test_higher_asil_wins_when_cap_is_tighter_than_safety_count(self):
        """캡이 안전관련 수보다 작으면 D > C > A 순으로 남는다."""
        flows = collect_integration_flows(
            _fd_many(10, asil_of=self._asil_of), max_flows=2)
        assert sorted(f["asil"] for f in flows) == ["C", "D"]

    def test_output_order_stays_alphabetical(self):
        """선별만 안전우선이고 **출력 순서는 알파벳 그대로** — 문서 행 순서를 흔들지 않는다."""
        flows = collect_integration_flows(
            _fd_many(10, asil_of=self._asil_of), max_flows=5)
        names = [f["entry_fn"] for f in flows]
        assert names == sorted(names), names

    def test_unknown_asil_ranks_after_qm(self):
        """미상 등급을 QM 보다 우대하면 근거 없는 우선순위가 된다."""
        flows = collect_integration_flows(
            _fd_many(4, asil_of=lambda i: "???" if i == 3 else "QM"), max_flows=3)
        assert "Ap_F003" not in {f["entry_fn"] for f in flows}


class TestSwComIdIsCapIndependent:
    """같은 모듈은 캡 값이 달라도 같은 SwCom ID 를 받아야 한다.

    회귀 대상: `_infer_swcom_id` 가 캡 안쪽 루프에서 호출돼 ID 가 캡에 의존했다.
    (실측: 이 저장소 실 프로젝트에서는 모듈 29개·변동 0건이라 무해한 변경)
    """

    def test_same_module_same_id_across_caps(self):
        fd = _fd_many(10)
        tight = {f["entry_fn"]: f["swcom_id"]
                 for f in collect_integration_flows(fd, max_flows=3)}
        loose = {f["entry_fn"]: f["swcom_id"]
                 for f in collect_integration_flows(fd, max_flows=100)}
        for name, sid in tight.items():
            assert loose[name] == sid, f"{name}: 캡에 따라 SwCom 이 바뀐다 {sid} != {loose[name]}"


class TestQualityReportExposesFlowCap:
    @staticmethod
    def _itc():
        return [{
            "tc_id": "SwITC_01", "related_ids": ["SwCom_01"],
            "synthetic_related_ids": ["SwCom_01"],
            "sub_cases": [], "input_vars": [], "expected_vars": [], "gen_method": "ABV",
        }]

    def test_flow_stats_are_carried_into_report(self):
        stats = {}
        collect_integration_flows(_fd_many(10), max_flows=4, stats_out=stats)
        qr = generate_sits_quality_report(self._itc(), 10, flow_stats=stats)
        cov = qr["integration_flow_coverage"]
        assert cov["total_flows_found"] == 10
        assert cov["flows_dropped"] == 6
        assert cov["flow_emit_pct"] == 40.0

    def test_report_without_flow_stats_stays_backward_compatible(self):
        """구 호출부(flow_stats 없음)는 키가 비어 있을 뿐 깨지지 않는다."""
        qr = generate_sits_quality_report(self._itc(), 1)
        assert qr["integration_flow_coverage"] == {}
        assert qr["total_test_cases"] == 1

    def test_tc_count_alone_would_hide_the_loss(self):
        """total_test_cases 는 캡에 잘려도 줄지 않는다 — 그래서 별도 축이 필요하다."""
        stats = {}
        collect_integration_flows(_fd_many(10), max_flows=4, stats_out=stats)
        qr = generate_sits_quality_report(self._itc(), 10, flow_stats=stats)
        assert qr["total_test_cases"] == 1
        assert qr["integration_flow_coverage"]["flows_dropped"] == 6


# ---------------------------------------------------------------------------
# validate_sits_xlsm — sub-case 계수를 desc 프리픽스로 추측하지 않는다
# ---------------------------------------------------------------------------

def _write_min_sits(path, sub_labels):
    """TC 1건 + 주어진 라벨의 sub-case 행을 갖는 최소 SITS 시트.

    ⚠ 시트 이름·시작행은 **라이터와 같은 상수**로 짓는다. 예전엔 여기서 `"4.SW …"` 와
    행 `7` 을 손으로 적었는데, 그건 라이터가 `3.…`/행 `5` 로 옮긴 뒤에도 이 테스트가
    초록으로 남는다는 뜻이었다 — 실제로 그렇게 됐다(검증기가 산출물을 한 줄도 못 읽는
    동안 이 파일의 5개 테스트는 전부 통과했다). 왕복 가드는 아래 별도 클래스.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _SPEC_SHEET_NAME
    ws.cell(row=_DATA_START_ROW, column=_TCID_COL, value="SwITC_01")
    ws.cell(row=_DATA_START_ROW, column=_DESC_COL,
            value="Verify integration: Ap_Door_Run → Drv_Motor_Set")
    for i, label in enumerate(sub_labels, start=_DATA_START_ROW + 1):
        ws.cell(row=i, column=_DESC_COL, value=label)   # sub-case 행엔 TC ID 없음
    wb.save(str(path))
    return str(path)


class TestValidatorCountsLabelledSubCases:
    r"""회귀 대상: 검증기가 `re.match(r"^\d", desc)` 로 sub-case 를 셌다.

    라이터는 `case_label or case_num` 을 쓰고 case_label 은 `COND_1 [...]` 처럼 **문자**로
    시작한다 — 라이터 포맷이 바뀌었는데 리더 휴리스틱이 안 따라갔다. 실측(실 프로젝트
    120 TC): 파일에 1288행이 있는데 840 만 세어 34.8% 과소, avg 7.0(실제 10.7).
    그런데도 valid 는 True 였다.
    """

    def test_letter_prefixed_labels_are_counted(self, tmp_path):
        labels = ["COND_1 [g_u8State=최솟값]", "ERR_PROP_1 [하한 초과]", "GLOBAL_2 [x]"]
        p = _write_min_sits(tmp_path / "s.xlsx", labels)
        st = validate_sits_xlsm(p)["stats"]
        assert st["sub_case_count"] == 3, "문자로 시작하는 sub-case 가 누락됐다"
        assert st["tc_count"] == 1

    def test_digit_prefixed_labels_still_counted(self, tmp_path):
        """대조군: 기존에 세던 숫자 시작 라벨도 그대로 세어야 한다(회귀 방지)."""
        p = _write_min_sits(tmp_path / "s.xlsx", ["1", "2", "3", "4"])
        assert validate_sits_xlsm(p)["stats"]["sub_case_count"] == 4

    def test_mixed_labels_are_all_counted(self, tmp_path):
        p = _write_min_sits(tmp_path / "s.xlsx", ["1", "COND_1 [a]", "2", "ERR_PROP_1 [b]"])
        st = validate_sits_xlsm(p)["stats"]
        assert st["sub_case_count"] == 4
        assert st["avg_sub_per_tc"] == 4.0

    def test_blank_rows_are_not_counted(self, tmp_path):
        """빈 desc 를 세면 반대 방향으로 거짓말한다."""
        p = _write_min_sits(tmp_path / "s.xlsx", ["COND_1 [a]", "", "   ", "COND_2 [b]"])
        assert validate_sits_xlsm(p)["stats"]["sub_case_count"] == 2

    def test_tc_row_is_not_double_counted_as_subcase(self, tmp_path):
        """TC 행도 desc 를 갖는다 — 구조 판정이 TC 를 sub 로 겹쳐 세면 안 된다."""
        p = _write_min_sits(tmp_path / "s.xlsx", ["COND_1 [a]"])
        st = validate_sits_xlsm(p)["stats"]
        assert (st["tc_count"], st["sub_case_count"]) == (1, 1)


# ---------------------------------------------------------------------------
# 라이터 → 검증기 왕복 — 시트 이름·시작행이 갈라지면 여기서만 잡힌다
# ---------------------------------------------------------------------------

def _round_trip_itcs(n_tc: int = 2, n_sub: int = 3):
    return [{
        "tc_id": f"SwITC_{i:02d}",
        "gen_method": "ABV",
        "entry_fn": f"fn_{i}",
        "call_chain": f"fn_{i} -> callee_a -> callee_b",
        "module_name": "M.c",
        "input_vars": ["u8_A"],
        "expected_vars": ["u8_B"],
        "related_ids": ["SwCom_01"],
        "synthetic_related_ids": ["SwCom_01"],
        "asil": "B",
        "sub_cases": [{
            "case_num": j + 1,
            "case_label": f"COND_{j + 1} [경계]",
            "call_chain": "",
            "precondition": "1",
            "inputs": {"u8_A": j},
            "expected": {"u8_B": j},
        } for j in range(n_sub)],
    } for i in range(1, n_tc + 1)]


class TestWriterReaderRoundTrip:
    """검증기가 **라이터가 실제로 쓴 파일**을 읽는지 본다.

    위 `TestValidatorCountsLabelledSubCases` 는 시트를 손으로 지어 검증기 로직만 봤다.
    그래서 시트 이름이 `4.…`(리더) ↔ `3.…`(라이터)로, 시작행이 `7` ↔ `5` 로 갈라진
    채로도 5개 테스트가 전부 초록이었고, 실 프로젝트 게이트에서 TC 200 · sub-case
    1,400 짜리 산출물을 **0 · 0 으로 보고**했다(2026-08-14 실측).

    이 클래스는 두 결함 각각을 따로 겨눈다 — 하나가 나머지를 가리지 않도록.
    """

    def test_validator_reads_back_what_writer_wrote(self, tmp_path):
        out = tmp_path / "rt.xlsx"
        generate_sits_xlsm(None, _round_trip_itcs(n_tc=2, n_sub=3), str(out))
        res = validate_sits_xlsm(str(out))
        assert res["stats"].get("tc_count") == 2, f"TC 를 되읽지 못했다: {res['issues']}"
        assert res["stats"].get("sub_case_count") == 6

    def test_writer_sheet_name_is_the_one_validator_requires(self, tmp_path):
        """시트 이름 축 단독 — 라이터가 만든 시트가 검증기 필수 목록과 같은 이름인가."""
        import openpyxl

        out = tmp_path / "rt.xlsx"
        generate_sits_xlsm(None, _round_trip_itcs(n_tc=1, n_sub=1), str(out))
        assert _SPEC_SHEET_NAME in openpyxl.load_workbook(str(out)).sheetnames
        assert not any("Missing required sheet" in i
                       for i in validate_sits_xlsm(str(out))["issues"])

    def test_first_data_row_is_not_skipped(self, tmp_path):
        """시작행 축 단독 — TC 1건·sub 1건만 있으면 시작행이 밀렸을 때 둘 다 0 이 된다."""
        out = tmp_path / "rt.xlsx"
        generate_sits_xlsm(None, _round_trip_itcs(n_tc=1, n_sub=1), str(out))
        st = validate_sits_xlsm(str(out))["stats"]
        assert (st.get("tc_count"), st.get("sub_case_count")) == (1, 1)
