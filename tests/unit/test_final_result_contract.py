# tests/unit/test_final_result_contract.py
"""`Final Test Result` 는 실측 집계에서 나와야 한다 — 정적 상수가 아니라.

## 회귀 대상 (2026-07-29 실측)

`_write_test_summary` 가 `meta.final_test_result` 기본값을 시트에 그대로 썼고,
그 값을 **라우터·워크플로 어디에서도 계산해 대입하지 않았다**(grep 0건). 결과:

    시나리오                        Actual Pass ratio    Final Test Result
    전부 실패 (passed=0/failed=5)             0.0                   OK   ← 거짓 PASS
    전부 미실행 (tested=0)          (사용자 입력 필요 표시)             OK   ← 증거 0인데 OK

한 문서 안에서 "Pass ratio 0.0" 과 "Final Test Result: OK" 가 나란히 찍혔고
경고도 0건이었다. ISO 26262 산출물에서 이건 라이브 거짓 PASS 다.

## 두 판정을 섞지 않는다

- SUTR/SITR(결과 문서) → `compute_final_result` : pass/fail/not_executed 기준
- Coverage(커버리지 문서) → `compute_coverage_final_result` : **커버리지 달성** 기준

Coverage 템플릿(KJPDS02 v1.01 C10)의 해당 셀은 `=IF(AND(B13=C9,...),"OK","NG")` 라는
커버리지 비교 수식이다. 여기에 pass/fail 판정을 넣으면 의미가 뒤바뀐다.
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from backend.services.swut_input_adapter import (
    CoverageStats,
    FunctionCoverage,
    compute_coverage_final_result,
    compute_final_result,
)
from backend.services.swut_sutr_aggregator import SutrBuildMeta, _write_test_summary


def _agg(**kw):
    base = {"total": 0, "total_tcs": 0, "tested": 0, "passed": 0,
            "failed": 0, "not_executed": 0, "deviated": 0}
    base.update(kw)
    return base


# ---------------------------------------------------------------------------
# compute_final_result — 결과 문서(SUTR/SITR)
# ---------------------------------------------------------------------------

class TestResultVerdict:
    def test_all_failed_is_not_ok(self):
        r = compute_final_result(_agg(total=5, tested=5, failed=5))
        assert r["verdict"] == "ng"
        assert r["display"] == "NG"

    def test_partial_failure_is_not_ok(self):
        r = compute_final_result(_agg(total=5, tested=5, passed=3, failed=2))
        assert r["verdict"] == "ng"

    def test_not_executed_alone_is_not_ok(self):
        """실행분이 전부 통과여도 미실행이 남아 있으면 합격이 아니다."""
        r = compute_final_result(_agg(total=7, tested=5, passed=5, not_executed=2))
        assert r["verdict"] == "ng"
        assert "미실행 2건" in r["reason"]

    def test_nothing_executed_is_na_not_ng(self):
        """아무것도 안 돌렸으면 '시험했는데 실패'(NG)도 거짓말이다."""
        r = compute_final_result(_agg(total=5, not_executed=5))
        assert r["verdict"] == "na"
        assert r["display"] == "N/A"

    def test_empty_session_is_na(self):
        assert compute_final_result(_agg())["verdict"] == "na"

    def test_all_pass_keeps_todays_value(self):
        """하위호환: 전부 통과면 기존 meta 기본값과 **글자 그대로** 같아야 한다."""
        assert compute_final_result(
            _agg(total=5, tested=5, passed=5), positive_token="OK")["display"] == "OK"
        assert compute_final_result(
            _agg(total=5, tested=5, passed=5), positive_token="PASS")["display"] == "PASS"

    @pytest.mark.parametrize(("pos", "neg"), [("OK", "NG"), ("PASS", "FAIL"), ("Pass", "Fail")])
    def test_vocabulary_per_document(self, pos, neg):
        assert compute_final_result(
            _agg(total=1, tested=1, failed=1), positive_token=pos)["display"] == neg

    def test_unknown_vocabulary_never_reuses_positive_token(self):
        """모르는 어휘여도 부정 판정이 긍정 토큰과 같아지면 안 된다."""
        r = compute_final_result(_agg(total=1, tested=1, failed=1), positive_token="정상")
        assert r["display"] != "정상"

    @pytest.mark.parametrize("bad", [None, "", "많음", [], {}])
    def test_non_numeric_counts_do_not_crash(self, bad):
        r = compute_final_result({"total": bad, "tested": bad, "failed": bad,
                                  "not_executed": bad})
        assert r["verdict"] == "na"      # 셀 수 없으면 합격 근거도 없다

    def test_reason_is_always_populated(self):
        for agg in (_agg(), _agg(total=1, tested=1, passed=1),
                    _agg(total=1, tested=1, failed=1)):
            assert compute_final_result(agg)["reason"]


# ---------------------------------------------------------------------------
# compute_coverage_final_result — 커버리지 문서
# ---------------------------------------------------------------------------

def _fc(name, st_cov, st_tot, br_cov=0, br_tot=0):
    return FunctionCoverage(
        unit_id=name, name=name,
        statement=CoverageStats(covered=st_cov, total=st_tot),
        branch=CoverageStats(covered=br_cov, total=br_tot),
    )


class TestCoverageVerdict:
    def test_nothing_measured_is_na(self):
        r = compute_coverage_final_result([_fc("f1", 0, 0)])
        assert r["verdict"] == "na"

    def test_empty_rows_is_na(self):
        assert compute_coverage_final_result([])["verdict"] == "na"

    def test_incomplete_function_is_ng(self):
        r = compute_coverage_final_result([_fc("f1", 10, 10), _fc("f2", 3, 10)])
        assert r["verdict"] == "ng"
        assert "f2" in r["reason"]

    def test_full_coverage_is_ok_with_positive_token(self):
        r = compute_coverage_final_result([_fc("f1", 10, 10)], positive_token="PASS")
        assert (r["verdict"], r["display"]) == ("ok", "PASS")

    def test_branch_axis_counts_too(self):
        """statement 만 보면 branch 미달을 놓친다."""
        r = compute_coverage_final_result([_fc("f1", 10, 10, br_cov=1, br_tot=4)])
        assert r["verdict"] == "ng"


# ---------------------------------------------------------------------------
# writer 통합 — 시트에 실제로 무엇이 찍히는가
# ---------------------------------------------------------------------------

_LABELS = ["Project Name", "Release Name(SW)", "Test Target Version(HW)", "Test Date",
           "Test Engineer", "Target Coverage", "Actual Coverage",
           "Target Pass ratio", "Actual Pass ratio", "Final Test Result"]


def _sheet():
    wb = Workbook()
    ws = wb.active
    for i, lab in enumerate(_LABELS, start=1):
        ws.cell(row=i, column=1, value=lab)
    return ws


def _cell_after(ws, label):
    for row in ws.iter_rows():
        for c in row:
            if str(c.value or "").strip() == label:
                return ws.cell(row=c.row, column=c.column + 1).value
    return None


class TestSutrWriterUsesComputedValue:
    def test_failing_session_does_not_write_ok(self):
        ws = _sheet()
        warns: list[str] = []
        _write_test_summary(ws, SutrBuildMeta(), _agg(total=5, tested=5, failed=5), warns)
        assert _cell_after(ws, "Final Test Result") == "NG"
        assert _cell_after(ws, "Actual Pass ratio") == 0.0

    def test_failing_session_emits_warning(self):
        ws = _sheet()
        warns: list[str] = []
        _write_test_summary(ws, SutrBuildMeta(), _agg(total=5, tested=5, failed=5), warns)
        assert any("final-result" in w for w in warns), warns

    def test_passing_session_writes_ok_and_no_warning(self):
        """대조군 — 정상 문서의 값과 경고가 오늘과 같아야 한다."""
        ws = _sheet()
        warns: list[str] = []
        _write_test_summary(ws, SutrBuildMeta(), _agg(total=5, tested=5, passed=5), warns)
        assert _cell_after(ws, "Final Test Result") == "OK"
        assert not [w for w in warns if "final-result" in w]

    def test_verdict_is_exposed_in_summary_dict(self):
        """xlsx 를 열지 않고도 API 응답으로 판정을 알 수 있어야 한다."""
        summary: dict = {}
        _write_test_summary(_sheet(), SutrBuildMeta(),
                            _agg(total=5, tested=5, failed=5), [], summary=summary)
        assert summary["final_result_verdict"] == "ng"
        assert summary["final_result"] == "NG"


class TestCoverageWriterPreservesTemplateFormula:
    """수식 양식에서는 템플릿 판정이 우선 — 계산값으로 덮어쓰면 안 된다."""

    def test_formula_cell_is_not_overwritten(self):
        from backend.services.swut_coverage_aggregator import _write_label_keep_formula

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Final Test Result")
        ws.cell(row=1, column=2, value='=IF(AND(B13=C9),"OK","NG")')
        _write_label_keep_formula(ws, "Final Test Result", "NG", [])
        assert ws.cell(row=1, column=2).value == '=IF(AND(B13=C9),"OK","NG")'

    def test_plain_cell_is_written(self):
        """대조군: 수식이 없으면 계산값이 들어가야 한다(예전엔 정적 'PASS')."""
        from backend.services.swut_coverage_aggregator import _write_label_keep_formula

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Final Test Result")
        _write_label_keep_formula(ws, "Final Test Result", "NG", [])
        assert ws.cell(row=1, column=2).value == "NG"


class TestNoBuilderWritesTheStaticDefault:
    """정적 기본값을 셀에 직접 쓰는 잔재가 남아 있으면 안 된다(한쪽만 고쳐짐 방지)."""

    @pytest.mark.parametrize("module_path", [
        "backend/services/swut_sutr_aggregator.py",
        "backend/services/swut_coverage_aggregator.py",
    ])
    def test_meta_final_test_result_is_not_written_directly(self, module_path):
        import ast
        import inspect
        from pathlib import Path

        root = Path(__file__).resolve().parents[2]
        tree = ast.parse((root / module_path).read_text(encoding="utf-8"))
        offenders = []
        for node in ast.walk(tree):
            if not isinstance(node, ast.Call):
                continue
            fname = getattr(node.func, "id", "") or getattr(node.func, "attr", "")
            if "write_label" not in fname:
                continue
            for arg in node.args:
                # meta.final_test_result 를 값 인자로 넘기는 호출
                if (isinstance(arg, ast.Attribute) and arg.attr == "final_test_result"
                        and getattr(arg.value, "id", "") == "meta"):
                    offenders.append(ast.unparse(node)[:100])
        assert offenders == [], (
            f"{module_path} 가 정적 meta 값을 셀에 그대로 쓴다: {offenders}")
        assert inspect  # 사용 표시(도구 경고 억제)
