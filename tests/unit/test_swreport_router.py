# -*- coding: utf-8 -*-
"""backend.routers.swreport endpoint TestClient 검증.

resolver(cloudium)는 _resolve_template_bytes/_resolve_source_workbooks 모킹으로 우회 —
라우팅·schema 검증·응답(헤더/Content-Type) 배선만 검증한다. 집계 정확성은
test_swreport_summary.py가 담당.
"""
import io
import json

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend.main import app

client = TestClient(app)
_HDR = {"X-User": "tester"}  # conftest autouse가 tester를 admin으로 등록


def _es_bytes() -> bytes:
    """최소 ES95411 양식: Summary(헤더+표 1행+Total) + detail 1개."""
    wb = openpyxl.Workbook()
    summ = wb.active
    assert summ is not None
    summ.title = "Summary"
    summ["B11"], summ["E11"] = "Fail", 0
    summ["B12"], summ["E12"] = "Result", "stale"
    for col, txt in [(5, "ID"), (6, "Test Name"), (7, "점검 대상"), (8, "분석 차수"),
                     (10, "Tester"), (11, "Tool"), (12, "Debugger"), (13, "총 분석시간"),
                     (14, "P/F"), (16, "Sheet Name")]:
        summ.cell(row=15, column=col, value=txt)
    summ["B16"], summ["E16"], summ["F16"] = 21, "UT101", "Req Unit"
    summ["G16"], summ["K16"], summ["P16"] = "O", "VectorCAST", "21.UT101"
    summ["B17"] = "Total"
    det = wb.create_sheet("21.UT101")
    det["B4"], det["C4"] = "분석차수", "0.1"
    det["B6"], det["C6"] = "Tester", "주희영"
    det["B7"], det["C7"] = "Debugger", "이재원"
    det["E4"], det["F4"] = "준비", 10
    det["E7"], det["F7"] = "Total", 30
    det["H7"], det["I7"] = "P/F", "Pass"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_VALID_BODY = {
    "project_id": "KJPDS02",
    "release_sw_version": "1.02",
    "test_date": "2026-03-24",
}


@pytest.fixture
def _patch_resolve(monkeypatch):
    data = _es_bytes()
    from backend.routers import swreport as mod
    monkeypatch.setattr(mod, "_resolve_template_bytes", lambda req: data)
    monkeypatch.setattr(mod, "_resolve_source_workbooks", lambda req, t: [("self", data)])
    return data


# ── 응답 배선 ────────────────────────────────────────────────────────────────
def test_summary_build_200(_patch_resolve):
    r = client.post("/api/swreport/summary/build", json=_VALID_BODY, headers=_HDR)
    assert r.status_code == 200, r.text
    assert "spreadsheet" in r.headers["content-type"] or "excel" in r.headers["content-type"]
    summary = json.loads(r.headers["X-SwReport-Summary"])
    assert summary["performed_count"] == 1
    assert summary["matched_rows"] == 1
    assert summary["overall_result"] == "Pass"
    # 산출물이 valid xlsx로 재오픈되는지
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Summary" in wb.sheetnames


def test_summary_preview_200(_patch_resolve):
    r = client.post("/api/swreport/summary/preview", json=_VALID_BODY, headers=_HDR)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["summary"]["matched_rows"] == 1
    ut = next(x for x in body["rows"] if x["id"] == "UT101")
    assert ut["pf"] == "Pass" and ut["tester"] == "주희영"


# ── schema / 인증 ────────────────────────────────────────────────────────────
def test_missing_required_422():
    r = client.post("/api/swreport/summary/build",
                    json={"project_id": "X"}, headers=_HDR)  # release/test_date 누락
    assert r.status_code == 422


def test_unknown_field_422():
    body = {**_VALID_BODY, "template_path": "", "foo_unknown": "x"}
    r = client.post("/api/swreport/summary/preview", json=body, headers=_HDR)
    assert r.status_code == 422  # extra='forbid'


def test_bad_version_422():
    r = client.post("/api/swreport/summary/build",
                    json={**_VALID_BODY, "release_sw_version": "v1"}, headers=_HDR)
    assert r.status_code == 422


def test_no_xuser_rejected():
    r = client.post("/api/swreport/summary/build", json=_VALID_BODY)  # X-User 없음
    assert r.status_code in (400, 401)
