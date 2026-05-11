"""Tests for backend.services.swut_deviation_generator.

매크로 분석, 분모 변수 추출, init 매크로 분리, end-to-end deviation
rationale 생성을 fixture C 소스 + in-memory xlsx로 검증.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.services.swut_deviation_generator import (  # noqa: E402
    CSourceIndex,
    MacroEvidence,
    _build_rationale,
    _extract_denominators,
    _is_init_macro,
    _is_zero_value,
    _parse_tc_label,
    generate_deviation_rationales,
    generate_deviation_rationales_dict,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_sutr_deviation_wb(cases: list[tuple[str, str]]) -> bytes:
    """Deviation 시트만 있는 minimal SUTR xlsx.

    Args:
        cases: [(tc_label, issue_text), ...]
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Deviation")
    ws.cell(row=1, column=2, value="Deviation Report")
    ws.cell(row=2, column=2, value="■ Deviation List")
    ws.cell(row=3, column=2, value="Test Case ID")
    ws.cell(row=3, column=3, value="Issue")
    ws.cell(row=3, column=4, value="Deviation")
    ws.cell(row=3, column=5, value="Status")
    for i, (tc, issue) in enumerate(cases):
        ws.cell(row=4 + i, column=2, value=tc)
        ws.cell(row=4 + i, column=3, value=issue)
        ws.cell(row=4 + i, column=4, value="dev rationale")
        ws.cell(row=4 + i, column=5, value="Closed")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helper function tests
# ---------------------------------------------------------------------------

class TestIsZeroValue:
    def test_integer_zero(self):
        assert _is_zero_value("0")
        assert _is_zero_value("-0")
        assert _is_zero_value("0.0")
        assert _is_zero_value(" 0 ")

    def test_nonzero(self):
        assert not _is_zero_value("68")
        assert not _is_zero_value("-1")
        assert not _is_zero_value("0.001")

    def test_invalid(self):
        assert not _is_zero_value("abc")
        assert not _is_zero_value("")


class TestIsInitMacro:
    @pytest.mark.parametrize("name,expected", [
        # 기본 패턴
        ("s16g_CLR", True),
        ("s32g_CLEAR", True),
        ("u8_RESET", True),
        ("PORT_INIT", True),
        ("S32_ZERO", True),
        ("CFG_DEFAULT", True),
        # ISO2 추가 패턴 (reviewer 권고)
        ("MOTOR_OFF", True),
        ("STATUS_INV", True),
        ("PARAM_INVALID", True),
        ("VAR_NA", True),
        ("CFG_NONE", True),
        ("LIMIT_MIN", True),
        # 정상 값 매크로 (False 기대)
        ("FRONT_OVER_POS", False),
        ("s16s_USER_CTRL_START_SPD", False),
        ("u16s_FRONT_OVER_OPEN_DEG", False),
    ])
    def test_suffix_pattern(self, name, expected):
        m = MacroEvidence(macro_name=name, value="0")
        assert _is_init_macro(m) is expected


class TestParseTcLabel:
    @pytest.mark.parametrize("label,exp_id,exp_no", [
        ("SwUTC_SwUFn_407 (TC2)", "SwUTC_SwUFn_407", "TC2"),
        ("SwUTC_SwUFn_1512 (TC1~5)", "SwUTC_SwUFn_1512", "TC1~5"),
        ("SwUTC_SwUFn_1513 (TC10)", "SwUTC_SwUFn_1513", "TC10"),
        ("SwUTC_SwUFn_0001", "SwUTC_SwUFn_0001", ""),
    ])
    def test_parse(self, label, exp_id, exp_no):
        tc_id, tc_no = _parse_tc_label(label)
        assert tc_id == exp_id
        assert tc_no == exp_no


class TestExtractDenominators:
    def test_circled_digits(self):
        text = (
            "< Divide by zero - 분모자리에 들어가는 변수 >\n"
            "① s32s_StartSpeed - s32s_TargetSpeed_1\n"
            "② s32s_AccelPos\n"
        )
        denoms = _extract_denominators(text)
        # 줄당 1개만 추출 → 첫 등장 식별자
        assert "s32s_StartSpeed" in denoms
        assert "s32s_AccelPos" in denoms

    def test_skips_guidance_lines(self):
        text = (
            "< Divide by zero - 분모자리에 들어가는 변수 >\n"
            "① u16g_SysOptCtrl_OverOpenDeg\n"
        )
        denoms = _extract_denominators(text)
        # 'Divide by zero' / '분모' / '변수' 가이드 라인은 제외
        assert "Divide" not in denoms
        assert "u16g_SysOptCtrl_OverOpenDeg" in denoms

    def test_dedup_preserves_order(self):
        text = "① foo_var\n② foo_var\n③ bar_var\n"
        denoms = _extract_denominators(text)
        assert denoms == ["foo_var", "bar_var"]


# ---------------------------------------------------------------------------
# _build_rationale logic
# ---------------------------------------------------------------------------

class TestBuildRationale:
    def test_no_macros_returns_none(self):
        is_zero, msg = _build_rationale("foo", [])
        assert is_zero is None
        assert "찾지 못함" in msg

    def test_only_init_macros_returns_none(self):
        macros = [MacroEvidence("s16g_CLR", "0"), MacroEvidence("s16g_RESET", "0")]
        is_zero, msg = _build_rationale("foo", macros)
        assert is_zero is None
        assert "초기화 매크로" in msg
        assert "산술 시점 값 enumerate 불가" in msg

    def test_value_macros_nonzero_returns_false(self):
        macros = [
            MacroEvidence("FRONT_OVER_POS", "2472"),
            MacroEvidence("REAR_OVER_POS", "2253"),
        ]
        is_zero, msg = _build_rationale("foo", macros)
        assert is_zero is False
        assert "0이 아니므로" in msg
        assert "FRONT_OVER_POS=2472" in msg

    def test_value_macros_with_zero_returns_true(self):
        macros = [
            MacroEvidence("MAGIC_ZERO_CONST", "0"),  # init suffix 없음
            MacroEvidence("OTHER", "100"),
        ]
        is_zero, msg = _build_rationale("foo", macros)
        assert is_zero is True
        assert "0 가능" in msg

    def test_init_excluded_from_zero_check(self):
        # init 매크로가 0이어도 value 매크로가 0 아니면 False
        macros = [
            MacroEvidence("FRONT_SPEED", "7000"),
            MacroEvidence("s32g_CLR", "0"),  # init → 제외
        ]
        is_zero, msg = _build_rationale("foo", macros)
        assert is_zero is False
        assert "0이 아니므로" in msg
        assert "초기화 매크로 s32g_CLR" in msg


# ---------------------------------------------------------------------------
# CSourceIndex
# ---------------------------------------------------------------------------

class TestCSourceIndex:
    @pytest.fixture
    def src_dir(self, tmp_path):
        # fake .c / .h 파일 생성
        h = tmp_path / "defines.h"
        h.write_text(
            "#define FRONT_OVER_POS    ( ( S16 )( 2472 ) )\n"
            "#define REAR_OVER_POS     ( ( S16 )( 2253 ) )\n"
            "#define s16g_CLR          ( ( S16 )( 0 ) )\n",
            encoding="utf-8",
        )
        c = tmp_path / "ctrl.c"
        c.write_text(
            "s16g_ApiIn_OverPos = FRONT_OVER_POS;\n"
            "s16g_ApiIn_OverPos = s16g_CLR;\n",
            encoding="utf-8",
        )
        return tmp_path

    def test_allowed_roots_rejects_outside_path(self, tmp_path):
        """deep-reviewer 시나리오 4 — allowed_roots 외 path는 ValueError."""
        outside = tmp_path / "outside"
        outside.mkdir()
        allowed = tmp_path / "allowed"
        allowed.mkdir()
        with pytest.raises(ValueError, match="not within allowed_roots"):
            CSourceIndex(str(outside), allowed_roots=[str(allowed)])

    def test_allowed_roots_accepts_subdirectory(self, tmp_path):
        sub = tmp_path / "project" / "src"
        sub.mkdir(parents=True)
        idx = CSourceIndex(str(sub), allowed_roots=[str(tmp_path / "project")])
        assert idx.src_root == str(sub)

    def test_allowed_roots_none_skips_validation(self, tmp_path):
        # None이면 비검증 (CLI/내부 호출 경로 — 현재 동작 유지)
        idx = CSourceIndex(str(tmp_path), allowed_roots=None)
        assert idx.src_root == str(tmp_path)

    def test_scan_indexes_defines(self, src_dir):
        idx = CSourceIndex(str(src_dir))
        idx.scan()
        assert "FRONT_OVER_POS" in idx.macros
        assert idx.macros["FRONT_OVER_POS"][0].value == "2472"
        assert idx.macros["s16g_CLR"][0].value == "0"

    def test_find_macros_via_assignment(self, src_dir):
        idx = CSourceIndex(str(src_dir))
        results = idx.find_macros_for("s16g_ApiIn_OverPos")
        names = {m.macro_name for m in results}
        # FRONT_OVER_POS와 s16g_CLR 모두 발견되어야 함 (assignment를 통해)
        assert "FRONT_OVER_POS" in names
        assert "s16g_CLR" in names

    def test_nonexistent_var_returns_empty(self, src_dir):
        idx = CSourceIndex(str(src_dir))
        assert idx.find_macros_for("never_assigned_var") == []


# ---------------------------------------------------------------------------
# End-to-end generate_deviation_rationales
# ---------------------------------------------------------------------------

class TestGenerateDeviationRationales:
    @pytest.fixture
    def src_dir(self, tmp_path):
        (tmp_path / "macros.h").write_text(
            "#define u16s_FRONT_OVER_OPEN_DEG    ( ( U16 )( 68U ) )\n"
            "#define u16s_REAR_OVER_OPEN_DEG     ( ( U16 )( 70U ) )\n",
            encoding="utf-8",
        )
        (tmp_path / "stuff.c").write_text(
            "u16g_SysOptCtrl_OverOpenDeg = u16s_FRONT_OVER_OPEN_DEG;\n"
            "u16g_SysOptCtrl_OverOpenDeg = u16s_REAR_OVER_OPEN_DEG;\n",
            encoding="utf-8",
        )
        return tmp_path

    def test_detects_divide_by_zero_and_resolves(self, src_dir):
        sutr = _build_sutr_deviation_wb([
            ("SwUTC_SwUFn_407 (TC2)",
             "< Divide by zero - 분모자리에 들어가는 변수 >\n"
             "① u16g_SysOptCtrl_OverOpenDeg\n"),
        ])
        cases = generate_deviation_rationales(sutr, src_dir)
        assert len(cases) == 1
        case = cases[0]
        assert case.tc_id == "SwUTC_SwUFn_407"
        assert case.tc_no == "TC2"
        assert case.pattern == "divide_by_zero"
        assert len(case.denominators) == 1
        denom = case.denominators[0]
        assert denom.name == "u16g_SysOptCtrl_OverOpenDeg"
        assert denom.is_zero_possible is False
        assert "0이 아니므로" in denom.rationale
        assert "u16s_FRONT_OVER_OPEN_DEG" in denom.rationale

    def test_unknown_pattern_when_no_divide_keyword(self, src_dir):
        sutr = _build_sutr_deviation_wb([
            ("SwUTC_SwUFn_999 (TC1)", "Overflow risk on speed var"),
        ])
        cases = generate_deviation_rationales(sutr, src_dir)
        assert len(cases) == 1
        assert cases[0].pattern == "unknown"
        assert cases[0].denominators == []

    def test_dict_serialization(self, src_dir):
        sutr = _build_sutr_deviation_wb([
            ("SwUTC_SwUFn_001 (TC1)", "< Divide by zero >\n① foo_var\n"),
        ])
        result = generate_deviation_rationales_dict(sutr, src_dir)
        assert result["ok"] is True
        assert result["deviation_count"] == 1
        assert result["pattern_counts"] == {"divide_by_zero": 1}
        assert len(result["cases"]) == 1
        assert "tc_id" in result["cases"][0]

    def test_auto_rationale_has_draft_label(self, src_dir):
        """ISO1 — auto_rationale 출력에 [AUTO-GENERATED DRAFT] 라벨 필수."""
        sutr = _build_sutr_deviation_wb([
            ("SwUTC_SwUFn_407 (TC2)",
             "< Divide by zero >\n① u16g_SysOptCtrl_OverOpenDeg\n"),
        ])
        cases = generate_deviation_rationales(sutr, src_dir)
        assert cases[0].auto_rationale.startswith("[AUTO-GENERATED DRAFT")
        assert "검토자 승인 필요" in cases[0].auto_rationale

    def test_macro_paths_relative_in_dict_output(self, src_dir):
        """X8 — to_dict의 매크로 file 경로는 src_root 상대화되어 절대경로 미노출."""
        sutr = _build_sutr_deviation_wb([
            ("SwUTC_SwUFn_407 (TC2)",
             "< Divide by zero >\n① u16g_SysOptCtrl_OverOpenDeg\n"),
        ])
        result = generate_deviation_rationales_dict(sutr, src_dir)
        macros = result["cases"][0]["denominators"][0]["macros"]
        assert macros  # at least one macro found
        for m in macros:
            # 절대 경로(`C:\...`, `/home/...`) 포함 안 됨
            assert ":" not in m["file"] or m["file"].startswith(".")
            # tmp_path 자체 절대경로 prefix 부재
            assert str(src_dir) not in m["file"]

    def test_empty_when_no_deviation_sheet(self):
        # Deviation 시트 없는 워크북
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        buf = io.BytesIO()
        wb.save(buf)
        cases = generate_deviation_rationales(buf.getvalue())
        assert cases == []

    def test_parse_warning_when_no_deviation_sheet(self):
        """deep-reviewer 시나리오 3 — 헤더 미발견 silent empty 막기."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        buf = io.BytesIO()
        wb.save(buf)
        warnings: list[str] = []
        cases = generate_deviation_rationales(buf.getvalue(), out_warnings=warnings)
        assert cases == []
        assert any("Deviation 시트" in w for w in warnings)

    def test_dict_includes_parse_warnings_and_tool_qualification(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        buf = io.BytesIO()
        wb.save(buf)
        result = generate_deviation_rationales_dict(buf.getvalue())
        assert "parse_warnings" in result
        assert result["parse_warnings"]
        assert "tool_qualification" in result
        assert result["tool_qualification"]["asil_b_c_d_usage"].startswith("단독 evidence")

    def test_no_c_source_skips_macro_lookup(self):
        sutr = _build_sutr_deviation_wb([
            ("SwUTC_SwUFn_001 (TC1)", "< Divide by zero >\n① some_var\n"),
        ])
        cases = generate_deviation_rationales(sutr, c_source_root=None)
        assert len(cases) == 1
        case = cases[0]
        # 분모 변수는 텍스트에서 enumerate하되 c_source_root None이면
        # 매크로 evidence는 비고 rationale은 "수동 검토 필요"
        assert case.pattern == "divide_by_zero"
        assert len(case.denominators) == 1
        denom = case.denominators[0]
        assert denom.name == "some_var"
        assert denom.macros == []
        assert denom.is_zero_possible is None
        assert "수동 검토 필요" in denom.rationale
