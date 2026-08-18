"""문서 앞부분(Cover / Introduction / Test Environment) 정합 — R8-5.

⚠ 이 문서는 **1.5 표에서 자기 Test Method 어휘를 정의한다.** 그런데 저장소 템플릿의
  1.5 는 하드웨어 통합시험 판(`FNCT`/`FIT`/`ELCT`)이 그대로 남아 있어, 3번 시트가 내는
  `REQ, IFT` · `FI` 가 **문서 안에 정의되지 않은 코드**였다. 감사자가 1.5 를 펴면 없다.

  같은 결함을 SwUTS 가 먼저 겪었고(그쪽은 3번 시트를 고쳐 맞췄다), SITS 는 반대로
  3번 시트만 고치고 Introduction 을 두었다.

실측(정본 KJPDS02_PV_SwITS v1.02 Introduction):
    1.5  REQ / IFT / RUT / SEP / FI            (FI 는 ASIL B,C,D)
    1.6  AOR / AOI / AEC / ABV / ERG / AFD / ADF / AUC / STA / ASV
"""
from __future__ import annotations

import re
from typing import Any, Dict, List

import openpyxl

from generators.sits import (
    _INTRO_GEN_METHODS,
    _INTRO_TEST_METHODS,
    _SITS_GEN_BOUNDARY,
    _SITS_GEN_DEFAULT,
    _SITS_METHOD_DEFAULT,
    _SITS_METHOD_FAULT,
    _intro_codes,
    _split_method_codes,
    generate_sits_xlsm,
)


class TestEmittedCodesAreDefinedInTheDocument:
    """이 축이 이 파일의 본체다 — 산출물이 쓰는 코드는 1.5/1.6 안에 있어야 한다."""

    def test_test_method_codes_are_declared(self):
        codes = _intro_codes(_INTRO_TEST_METHODS)
        for value in (_SITS_METHOD_DEFAULT, _SITS_METHOD_FAULT):
            used = _split_method_codes(value)
            assert used, value
            assert set(used) <= codes, f"{value!r} 의 {set(used) - codes} 가 1.5 표에 없다"

    def test_gen_method_codes_are_declared(self):
        codes = _intro_codes(_INTRO_GEN_METHODS)
        for value in (_SITS_GEN_DEFAULT, _SITS_GEN_BOUNDARY):
            used = _split_method_codes(value)
            assert used, value
            assert set(used) <= codes, f"{value!r} 의 {set(used) - codes} 가 1.6 표에 없다"

    def test_hardware_vocabulary_is_gone(self):
        """음성 대조군 — 템플릿의 HW 어휘가 살아 있으면 이 표는 고쳐진 게 아니다."""
        codes = _intro_codes(_INTRO_TEST_METHODS)
        assert not ({"FNCT", "FIT", "ELCT"} & codes), codes

    def test_splitter_handles_both_separators(self):
        """SwITS 는 Test Method 가 쉼표, Gen Method 가 슬래시다 — 둘 다 갈라야 한다."""
        assert _split_method_codes("REQ, IFT") == ["REQ", "IFT"]
        assert _split_method_codes("AOR/ABV") == ["AOR", "ABV"]
        assert _split_method_codes("") == []


def _itcs() -> List[Dict[str, Any]]:
    return [{
        "tc_id": "SwITC_01", "entry_fn": "Ap_Door_Run", "call_chain": "a -> b",
        "description": "d", "asil": "B", "test_method": _SITS_METHOD_DEFAULT,
        "gen_method": _SITS_GEN_DEFAULT, "related_ids": ["SwCom_01"],
        "sub_cases": [{"case_num": 1, "inputs": {"x": "1"}, "expected": {"y": "2"},
                       "precondition": "", "call_chain": ""}],
    }]


def _gen(tmp_path, **fm):
    out = tmp_path / "fm.xlsx"
    generate_sits_xlsm(
        None, _itcs(), str(out),
        project_config={"project_id": "KJPDS02", "doc_id": "HKY-KJPDS02-SwITS-2895",
                        "version": "v1.02", "asil_level": "ASIL B"},
        front_matter=fm or None)
    return openpyxl.load_workbook(str(out))


def _intro(wb):
    for sn in wb.sheetnames:
        if re.sub(r"[\s.]+", "", sn).lower() in ("introduction", "1introduction"):
            return wb[sn]
    return None


def _cells(ws) -> List[str]:
    return [str(c.value).strip() for row in ws.iter_rows() for c in row
            if c.value is not None and str(c.value).strip()]


class TestIntroductionIsFilled:
    def test_the_15_table_carries_the_software_vocabulary(self, tmp_path):
        ws = _intro(_gen(tmp_path))
        assert ws is not None, "Introduction 시트를 못 찾았다"
        blob = " | ".join(_cells(ws))
        for _name, abbr, _asil in _INTRO_TEST_METHODS:
            assert abbr in blob, f"1.5 에 {abbr} 가 없다"
        assert "FNCT" not in blob and "ELCT" not in blob, blob[:300]

    def test_reference_rows_name_the_documents_actually_read(self, tmp_path):
        wb = _gen(tmp_path, references=[("(KJPDS02_SwRS) SRS_v3.01.docx", "SW 요구사항 명세서")])
        blob = " | ".join(_cells(_intro(wb)))
        assert "(KJPDS02_SwRS) SRS_v3.01.docx" in blob, blob[:300]

    def test_absent_documents_leave_the_row_empty(self, tmp_path):
        """못 읽은 문서를 '참조했다'고 적으면 문서가 거짓말한다."""
        blob = " | ".join(_cells(_intro(_gen(tmp_path))))
        assert "SW 요구사항 명세서" not in blob or "docx" not in blob, blob[:300]

    def test_glossary_names_this_project_not_another(self, tmp_path):
        """템플릿엔 **다른 프로젝트**(HDPDM01) 용어사전이 박혀 있었다."""
        blob = " | ".join(_cells(_intro(_gen(tmp_path))))
        assert "HDPDM01" not in blob, blob[:300]


class TestCoverIsFilled:
    def test_placeholders_are_replaced(self, tmp_path):
        wb = _gen(tmp_path, author="김진경", date="2026.08.18")
        cover = next((wb[s] for s in wb.sheetnames if s.lower() == "cover"), None)
        if cover is None:
            return  # 무템플릿 경로에 Cover 가 없으면 이 축은 대상 밖
        blob = " | ".join(_cells(cover))
        assert "[P_Name]" not in blob and "202X.XX.XX" not in blob, blob[:300]
