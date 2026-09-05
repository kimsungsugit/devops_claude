"""Regression: /api/jenkins/suts/extract-traceability 전용 파서 + 내용검증(오태깅).

STS/SUTS는 과거 `jenkins_sts_extract_traceability` 한 함수를 doc_type으로 공유했다.
공유 파서는 SyTS식 over-trace 버그는 아니나(레이아웃 적응형·컬럼 한정) 잠복 위험이 있어
SUTS 전용 파서로 분리했다:
- H2: 내부 생성 SUTS의 함수명 컬럼 헤더가 'Name'이라 unit 탐지가 못 잡던 것 → 'name' 폴백.
- H3: 파서가 doc_type만 믿고 source 태그 → 링크 뒤바뀌면 밴드 조용히 오태깅 → 내용검증 경고.
STS 회귀는 test_jenkins_sts_traceability.py, 공유 코어 무변경은 그쪽 5건이 가드한다.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from backend.routers.jenkins import (
    jenkins_sts_extract_traceability,
    jenkins_suts_extract_traceability,
)


@pytest.fixture(autouse=True)
def _local_mode(monkeypatch: pytest.MonkeyPatch):
    """resolver를 local로 고정 — tmp 파일 직접 read(cloudium IPC 회피)."""
    from backend.services import file_resolver as fr
    monkeypatch.setattr(fr, "_resolver", fr.LocalFileResolver())


def _save(wb: openpyxl.Workbook, tmp_path: Path, name: str) -> str:
    fp = tmp_path / name
    wb.save(str(fp))
    return str(fp)


def _suts_wb() -> openpyxl.Workbook:
    """SUTS 레이아웃: 'Related ID'(row3 병합헤더) + 'TC_ID'(row4), 값=SwUTC/SwUFn."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "2.SW Unit Test Spec"
    ws.cell(3, 10, "Related ID")
    ws.cell(4, 3, "TC_ID")
    ws.cell(5, 3, "SwUTC_0121")
    ws.cell(5, 10, "SwUFn_0121")
    ws.cell(6, 3, "SwUTC_0122")
    ws.cell(6, 10, "SwUFn_0122")
    return wb


def _sts_wb() -> openpyxl.Workbook:
    """STS 레이아웃: 'Test Case ID'(col2) / 'SRS'(col13) 같은 행, 값=SwTC/SwEI."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "3.SW Test Spec"
    ws.cell(5, 2, "Test Case ID")
    ws.cell(5, 13, "SRS")
    ws.cell(6, 2, "SwTC_SwEI_01_01")
    ws.cell(6, 13, "SwEI_01, SwTR_0104")
    ws.cell(7, 2, "SwTC_SwEI_02_01")
    ws.cell(7, 13, "SwEI_02")
    return wb


def test_suts_dedicated_endpoint_extracts(tmp_path: Path) -> None:
    """SUTS 전용 함수: SwUFn↔SwUTC 추출, source='SUTS', 오태깅 경고 없음."""
    res = jenkins_suts_extract_traceability({"path": _save(_suts_wb(), tmp_path, "s.xlsx")})
    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWUFN_0121", "SwUTC_0121") in pairs
    assert ("SWUFN_0122", "SwUTC_0122") in pairs
    assert {r["source"] for r in res["vcast_rows"]} == {"SUTS"}
    assert "warning" not in res  # SUTS 파일을 SUTS로 → 정상, 경고 없음


def test_suts_unit_name_column_detected(tmp_path: Path) -> None:
    """H2: 함수명 컬럼 헤더가 'Name'이어도 unit으로 잡힘(SUTS 전용 'name' 폴백).

    mutation: unit_header_extra=('name',)를 빼면 unit이 ''가 되어 이 assert 실패.
    """
    wb = _suts_wb()
    ws = wb.active
    assert ws is not None
    ws.cell(4, 4, "Name")                 # 함수명 컬럼(내부 생성 SUTS)
    ws.cell(5, 4, "EEPROM_SetByte")
    ws.cell(6, 4, "EEPROM_GetByte")
    res = jenkins_suts_extract_traceability({"path": _save(wb, tmp_path, "n.xlsx")})
    unit_by_tc = {r["testcase"]: r.get("unit") for r in res["vcast_rows"]}
    assert unit_by_tc["SwUTC_0121"] == "EEPROM_SetByte"
    assert unit_by_tc["SwUTC_0122"] == "EEPROM_GetByte"


def test_sts_does_not_capture_name_as_unit(tmp_path: Path) -> None:
    """대조군: STS 파서는 'name'을 unit 후보로 안 씀(오탐 방지) — 'Unit' 헤더만 잡는다.

    STS 파일(SwTC/SwEI)에 'Name' 컬럼이 있어도 STS 경로는 unit_header_extra 미전달이라
    unit이 채워지지 않는다(H2 폴백은 SUTS 한정).
    """
    wb = _sts_wb()
    ws = wb.active
    assert ws is not None
    ws.cell(5, 4, "Name")
    ws.cell(6, 4, "some_func")
    res = jenkins_sts_extract_traceability(
        {"path": _save(wb, tmp_path, "sn.xlsx"), "doc_type": "sts"})
    assert all(not r.get("unit") for r in res["vcast_rows"])


def test_suts_mistag_warns_when_sts_structure(tmp_path: Path) -> None:
    """H3: SUTS 엔드포인트에 STS 구조 파일(SwTC/SwEI)이 오면 오태깅 경고(fail-loud)."""
    res = jenkins_suts_extract_traceability({"path": _save(_sts_wb(), tmp_path, "x.xlsx")})
    assert res["ok"] is True
    assert res.get("warning") and "STS" in res["warning"]
    # 추출 자체는 여전히 수행(경고만 추가, 행 불변)
    assert {r["source"] for r in res["vcast_rows"]} == {"SUTS"}


def test_sts_mistag_warns_when_suts_structure(tmp_path: Path) -> None:
    """H3 역방향: STS 엔드포인트에 SUTS 구조 파일(SwUTC/SwUFn)이 오면 경고."""
    res = jenkins_sts_extract_traceability(
        {"path": _save(_suts_wb(), tmp_path, "y.xlsx"), "doc_type": "sts"})
    assert res["ok"] is True
    assert res.get("warning") and "SUTS" in res["warning"]
    assert {r["source"] for r in res["vcast_rows"]} == {"STS"}


def test_sts_shim_delegates_suts_doctype(tmp_path: Path) -> None:
    """하위호환: 구 /sts 엔드포인트에 doc_type='suts'면 SUTS 전용 파서로 위임."""
    res = jenkins_sts_extract_traceability(
        {"path": _save(_suts_wb(), tmp_path, "z.xlsx"), "doc_type": "suts"})
    assert res["ok"] is True
    assert {r["source"] for r in res["vcast_rows"]} == {"SUTS"}
    assert "warning" not in res  # SUTS 파일을 suts로 선언 → 정상


def test_suts_no_false_warning_on_fixed_fallback(tmp_path: Path) -> None:
    """오경보 억제: fallback 고정컬럼(TC='TC_001')은 doc-type 신호가 없어 경고 없음."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Traceability"
    ws.cell(4, 5, "TC_001")      # 고정 레이아웃(헤더 미탐지 → fallback)
    ws.cell(4, 6, "SwRS_001")
    res = jenkins_suts_extract_traceability({"path": _save(wb, tmp_path, "fb.xlsx")})
    assert res["ok"] is True
    assert ("SWRS_001", "TC_001") in {
        (r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert "warning" not in res  # 판별 불가 → 오경보 없음


def test_suts_no_trace_sheet_returns_available(tmp_path: Path) -> None:
    """trace/test 키워드 시트 없으면 error + available_sheets(STS와 동일 계약)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Cover"
    res = jenkins_suts_extract_traceability({"path": _save(wb, tmp_path, "none.xlsx")})
    assert res["ok"] is False
    assert "available_sheets" in res


def test_suts_wide_related_id_column_beyond_200(tmp_path: Path) -> None:
    """회귀(under-trace 방지): 'Related ID' 컬럼이 200열 밖(col 205)이어도 탐지·추출된다.

    실측 SwITS(v1.02)는 Input(82열)+Expected Result(113열)가 붙어 Related ID가 204열로
    밀렸고, 구 200 하드캡(_detect_trace_header_cols)이 이를 통째로 놓쳐 고정컬럼 폴백(col6)이
    엉뚱한 열을 읽어 매핑 0건으로 침묵 손실됐다(SITS 파서는 1024로 이미 고쳤으나 STS/SUTS
    공유 헤더탐지는 200에 방치). SUTS(v0.10)는 189열이라 아슬히 통과하던 잠복 위험을 상한
    통일(_MAX_SCAN_COLS=1024)로 차단. mutation: max_c 캡을 200으로 되돌리면 0건으로 실패한다.
    over-trace 불가: req 컬럼은 'Related ID' 헤더 텍스트 앵커로만 잡혀 없는 매핑을 만들지 못한다.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "2.SW Unit Test Spec"
    ws.cell(3, 205, "Related ID")     # 200열 밖 — 넓은 Input/Expected 뒤로 밀린 요구 컬럼
    ws.cell(4, 3, "TC_ID")
    ws.cell(5, 3, "SwUTC_0121")
    ws.cell(5, 205, "SwUFn_0121")
    ws.cell(6, 3, "SwUTC_0122")
    ws.cell(6, 205, "SwUFn_0122")
    res = jenkins_suts_extract_traceability({"path": _save(wb, tmp_path, "wide.xlsx")})
    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWUFN_0121", "SwUTC_0121") in pairs   # 200열 밖 컬럼이 실제로 읽혔다
    assert ("SWUFN_0122", "SwUTC_0122") in pairs
