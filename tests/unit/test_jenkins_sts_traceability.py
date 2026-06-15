"""Regression: /api/jenkins/sts/extract-traceability 헤더 기반 컬럼 탐지.

STS/SUTS(SwTS/SwUTS) 추출기가 고정 컬럼(TC=5/req=6)을 가정해 KJPDS02
SwTS/SwUTS 레이아웃에서 0건이던 버그를 헤더 기반 탐지(+멀티행 헤더)로 수정한 것의
회귀 가드. SwTS형(TC/SRS 같은 행), SwUTS형(Related ID 상단행 + TC_ID 다음행),
헤더 미탐지 시 기존 고정 컬럼 fallback을 고정한다.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from backend.routers.jenkins import jenkins_sts_extract_traceability


@pytest.fixture(autouse=True)
def _local_mode(monkeypatch: pytest.MonkeyPatch):
    """resolver를 local로 고정 — tmp 파일을 직접 read (cloudium IPC 회피)."""
    from backend.services import file_resolver as fr
    monkeypatch.setattr(fr, "_resolver", fr.LocalFileResolver())


def _save(wb: openpyxl.Workbook, tmp_path: Path, name: str) -> str:
    fp = tmp_path / name
    wb.save(str(fp))
    return str(fp)


def test_sts_same_row_header_srs_column(tmp_path: Path) -> None:
    """SwTS형: 'Test Case ID'(col2) / 'SRS'(col13) 헤더가 같은 행. TC↔SRS 추출."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "3.SW Test Spec"  # "test spec" → list
    ws.cell(5, 2, "Test Case ID")
    ws.cell(5, 13, "SRS")
    ws.cell(6, 2, "SwTC_SwEI_01_01")
    ws.cell(6, 13, "SwEI_01, SwTR_0104")
    ws.cell(7, 2, "SwTC_SwEI_02_01")
    ws.cell(7, 13, "SwEI_02")

    res = jenkins_sts_extract_traceability(
        {"path": _save(wb, tmp_path, "swts.xlsx"), "doc_type": "sts"})

    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWEI_01", "SwTC_SwEI_01_01") in pairs
    assert ("SWTR_0104", "SwTC_SwEI_01_01") in pairs
    assert ("SWEI_02", "SwTC_SwEI_02_01") in pairs
    assert {r["source"] for r in res["vcast_rows"]} == {"STS"}


def test_suts_multirow_header(tmp_path: Path) -> None:
    """SwUTS형: 멀티행 헤더 — 'Related ID'(row3) + 'TC_ID'(row4). band 스캔으로 매칭."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "2.SW Unit Test Spec"  # "test spec"/"unit test" → list
    ws.cell(3, 10, "Related ID")   # 상단 병합 헤더 행
    ws.cell(4, 3, "TC_ID")         # 상세 헤더 행
    ws.cell(5, 3, "SwUTC_0121")
    ws.cell(5, 10, "SwUFn_0121")
    ws.cell(6, 3, "SwUTC_0122")
    ws.cell(6, 10, "SwUFn_0122")

    res = jenkins_sts_extract_traceability(
        {"path": _save(wb, tmp_path, "swuts.xlsx"), "doc_type": "suts"})

    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWUFN_0121", "SwUTC_0121") in pairs
    assert ("SWUFN_0122", "SwUTC_0122") in pairs
    assert {r["source"] for r in res["vcast_rows"]} == {"SUTS"}


def test_suts_merged_cell_carry_forward(tmp_path: Path) -> None:
    """TC가 한 블록 첫 행에만 있고 연속행은 비었을 때, 직전 TC로 carry-forward 매핑."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "2.SW Unit Test Spec"
    ws.cell(3, 10, "Related ID")
    ws.cell(4, 3, "TC_ID")
    ws.cell(5, 3, "SwUTC_0201")
    ws.cell(5, 10, "SwUFn_0201")
    # 연속행: TC 비어있고 req만 — 직전 TC(0201)에 귀속
    ws.cell(6, 10, "SwUFn_0202")

    res = jenkins_sts_extract_traceability(
        {"path": _save(wb, tmp_path, "swuts2.xlsx"), "doc_type": "suts"})

    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWUFN_0201", "SwUTC_0201") in pairs
    assert ("SWUFN_0202", "SwUTC_0201") in pairs  # carry-forward


def test_sts_fallback_fixed_columns(tmp_path: Path) -> None:
    """헤더 미탐지(Test Case ID/SRS 헤더 없음) → 기존 고정 컬럼(TC=5/req=6) fallback."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Traceability"  # "trace" → list
    ws.cell(4, 5, "TC_001")     # 고정 레이아웃: col5=TC
    ws.cell(4, 6, "SwRS_001")   # col6=req
    ws.cell(4, 4, "func_a")     # col4=unit

    res = jenkins_sts_extract_traceability(
        {"path": _save(wb, tmp_path, "fb.xlsx"), "doc_type": "suts"})

    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWRS_001", "TC_001") in pairs


def test_sts_no_trace_sheet_returns_available(tmp_path: Path) -> None:
    """trace/test 키워드 시트 없으면 error + available_sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Cover"

    res = jenkins_sts_extract_traceability(
        {"path": _save(wb, tmp_path, "none.xlsx"), "doc_type": "sts"})

    assert res["ok"] is False
    assert "available_sheets" in res
