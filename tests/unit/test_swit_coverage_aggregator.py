"""33차 SwIT Coverage Report 빌더 회귀 — SwUT 패턴 차용.

SwUT는 30~32차에 17개 시나리오 커버. SwIT는 SwUT 시트 writer 그대로
import 재활용이라 SwUT 회귀가 SwIT에도 적용. 본 회귀는 SwIT 도구별
차이 (파일명 / doc_id_base / 결과 dataclass) + smoke + 시트 미발견
fallback 위주.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.services.swit_coverage_aggregator import (  # noqa: E402
    SwitCoverageBuildResult,
    build_swit_coverage_report,
)
from backend.services.swit_meta import SwitCoverageBuildMeta  # noqa: E402
from backend.services.swut_input_adapter import (  # noqa: E402
    CoverageStats,
    EnvironmentData,
    ExecutionRow,
    FunctionCoverage,
    SwUTSession,
)


def _build_swit_template() -> bytes:
    """SwIT v2.02 빈 양식 — SwUT _build_coverage_template 와 동일 구조 가정."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cover = wb.create_sheet("Cover")
    cover["B1"] = "Project"
    cover["B2"] = "ASIL Level"
    cover["B3"] = "Author"
    cover["B4"] = "Approver"

    ts = wb.create_sheet("Test Summary")
    ts["B1"] = "Project Name"
    ts["B2"] = "Release Name(SW)"
    ts["B3"] = "Test Target Version(HW)"
    ts["B4"] = "Test Date"
    ts["B5"] = "Test Engineer"
    ts["B6"] = "Final Test Result"

    trace = wb.create_sheet("1.Traceability")
    trace["A1"] = "Traceability matrix placeholder"

    cons = wb.create_sheet("2.Consistency")
    cons["A1"] = "Consistency placeholder"

    cov = wb.create_sheet("3. Coverage")
    cov["A1"] = "Statement Coverage"
    cov["A6"] = "Unit ID"
    cov["B6"] = "Name"
    cov["C6"] = "Count"
    cov["D6"] = "Total"
    cov["E6"] = "Pass"

    hist = wb.create_sheet("History")
    hist["A1"] = "■ Revision History"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_swit_session() -> SwUTSession:
    """Integration test session — SwUT와 동일 구조 (재활용)."""
    env = EnvironmentData(env_name="SWTE_01", component_name="SysOs_Main")
    env.test_cases = {"SwITC_SwUFn_0101.001": []}
    env.test_results = {
        "SwITC_SwUFn_0101.001": ExecutionRow(
            tc_name="SwITC_SwUFn_0101.001",
            component="SysOs_Main", passed=True,
        ),
    }
    env.function_coverage = [
        FunctionCoverage(unit_id="SwUFn_0101", name="SysOs_Main.init"),
    ]
    return SwUTSession(
        project_id="HDPDM01", version="v2.02_240219",
        source_kind="log_folder", source_path="/tmp/fake/v2.02_240219",
        environments=[env],
    )


def _make_swit_meta() -> SwitCoverageBuildMeta:
    return SwitCoverageBuildMeta(
        project_id="HDPDM01",
        release_sw_version="2.02",
        test_date="2024-02-19",
        test_engineer="JK Kim",
        doc_id_sequence="001",
        doc_id_base="HDPDM01-SwIT",
        asil_level="ASIL B",
    )


def _build_swit_template_v202_traceability_row_20() -> bytes:
    """58차 F2 — v2.02 SwIT 양식 mock: 1.Traceability 헤더 row 20에 SwUFn_ prefix."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cover = wb.create_sheet("Cover")
    cover["B1"] = "Project"

    ts = wb.create_sheet("1.Test Summary")
    ts["B1"] = "Project Name"
    ts["B2"] = "SW Version"
    ts["B3"] = "HW Version"
    ts["B4"] = "Test Date"
    ts["B5"] = "Test Engineer"
    ts["B6"] = "Target Coverage"
    ts["B7"] = "Actual Coverage"
    ts["B8"] = "Final Test Result"

    trace = wb.create_sheet("1.Traceability")
    # 헤더 row 20에 SwUFn_ prefix 5개 (v2.02 양식 위치 — 자동 탐색 max_row=20 으로는 발견 못함)
    for i in range(5):
        trace.cell(25, 3 + i).value = f"SwUFn_{i:04d}"
    # data row 26에 SwUTC_SwUFn_0001
    trace.cell(26, 2).value = "SwUTC_SwUFn_0000"

    wb.create_sheet("2.Consistency")
    wb.create_sheet("3. Coverage")
    wb.create_sheet("History")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_swit_session_with_fn_0000() -> SwUTSession:
    """traceability stamp 검증용 — TC 1건 + function 1건 mock."""
    env = EnvironmentData(env_name="SWTE_01", component_name="SysOs_Main")
    env.test_cases = {"SwITC_SwUFn_0000.001": []}
    env.test_results = {
        "SwITC_SwUFn_0000.001": ExecutionRow(
            tc_name="SwITC_SwUFn_0000.001",
            component="SysOs_Main", passed=True,
        ),
    }
    env.function_coverage = [
        FunctionCoverage(unit_id="SwUFn_0000", name="SysOs_Main.init"),
    ]
    return SwUTSession(
        project_id="HDPDM01", version="v2.02_240219",
        source_kind="log_folder", source_path="/tmp/fake/v2.02_240219",
        environments=[env],
    )


class TestTraceabilityV202LayoutF2:
    """58차 F2 — SwIT v2.02 양식 Traceability 헤더 row 자동 감지 + layout 강제."""

    def test_traceability_stamps_with_v202_header_at_row_25(self):
        """헤더 row 25 (v2.02 양식 위치) — 자동 탐색 max_row=30 확장 효과로 발견."""
        result = build_swit_coverage_report(
            _make_swit_session_with_fn_0000(),
            _make_swit_meta(),
            _build_swit_template_v202_traceability_row_20(),
        )
        assert result.ok
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        trace = wb["1.Traceability"]
        # row 26 (data) col 3 (header에서 SwUFn_0000 위치) — 'O' stamp 검증
        # SwUFn_0000은 trace.cell(25,3) 위치 → data row=26, col=3
        assert trace.cell(26, 3).value == "O", (
            f"Traceability 'O' stamp 미발견 — row 26 col 3 value: {trace.cell(26, 3).value!r}, "
            f"summary traceability_o_cells: {result.summary.get('traceability_o_cells')}"
        )
        # summary에 stamp 수 1 이상
        assert result.summary.get("traceability_o_cells", 0) >= 1


class TestBuildSwitCoverage:
    """SwIT Coverage Report builder smoke + structure."""

    def test_smoke_minimal(self):
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
        )
        assert isinstance(result, SwitCoverageBuildResult)
        assert result.ok
        assert result.xlsx_io.tell() == 0   # 처음으로 seek됨 (StreamingResponse 준비)
        assert result.result_size_bytes > 0

    def test_filename_has_swit_keyword_and_version(self):
        """파일명: '(HDPDM01)SwIT Coverage Report_v2.02_240219_R.xlsx'."""
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
        )
        assert "SwIT" in result.filename
        assert "HDPDM01" in result.filename
        assert "v2.02" in result.filename
        assert result.filename.endswith("_R.xlsx")

    def test_filename_uses_config_pattern_when_provided(self):
        """KJPDS02 SwITCV 파일명 패턴을 최종 산출물명에 반영."""
        meta = _make_swit_meta()
        meta.project_id = "KJPDS02"
        meta.doc_filename_pattern = (
            "(KJPDS02_DV_SwITCV) Software Integration Test Coverage "
            "Result_v{version}_{date}_R.xlsx"
        )
        result = build_swit_coverage_report(
            _make_swit_session(), meta, _build_swit_template(),
        )
        assert result.filename == (
            "(KJPDS02_DV_SwITCV) Software Integration Test Coverage "
            "Result_v2.02_240219_R.xlsx"
        )

    def test_summary_contains_asil_keys(self):
        """30차 W21 + 31차 W29 ASIL summary keys 노출 (SwUT 동일)."""
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
        )
        for key in (
            "asil_distribution", "asil_b_function_ids",
            "asil_c_function_ids", "asil_d_function_ids",
            "asil_highlight_policy",
        ):
            assert key in result.summary, f"summary 키 '{key}' 누락"

    def test_summary_contains_basic_metrics(self):
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
        )
        s = result.summary
        assert s["environments"] == 1
        assert s["function_rows"] == 1
        assert s["passed"] == 1
        assert "template_sha256_12" in s
        assert "build_timestamp" in s

    def test_tool_qualification_present(self):
        """ISO 26262 ASIL audit evidence_class draft 정책."""
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
        )
        tq = result.tool_qualification
        assert tq["evidence_class"] == "auto-generated draft"
        assert "asil_a_usage" in tq
        assert "asil_b_c_d_usage" in tq


class TestSwitMissingSheetFallback:
    """일부 시트 미발견 시 incomplete_sheets에 표시 + 빌드 진행 (graceful)."""

    def test_cover_sheet_missing(self):
        """Cover 미발견 → warnings 누적, 빌드는 진행."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Test Summary")
        wb.create_sheet("1.Traceability")
        wb.create_sheet("2.Consistency")
        wb.create_sheet("3. Coverage")
        wb.create_sheet("History")
        buf = io.BytesIO()
        wb.save(buf)
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), buf.getvalue(),
        )
        assert result.ok
        assert any("Cover" in w for w in result.warnings)

    def test_traceability_sheet_missing(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        wb.create_sheet("Test Summary")
        wb.create_sheet("2.Consistency")
        wb.create_sheet("3. Coverage")
        wb.create_sheet("History")
        buf = io.BytesIO()
        wb.save(buf)
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), buf.getvalue(),
        )
        assert result.ok
        # 라운드 F7 D1 fix: 'Traceability 시트 미발견' (prefix '1.' 제거 — 회사 표준
        # SwITCV는 '2.Traceability'이지만 본 fixture는 시트 자체 부재)
        assert any("Traceability" in w and "미발견" in w for w in result.warnings)

    def test_consistency_partial_when_swuds_not_provided(self):
        """SwUDS function_ids 미제공 → incomplete_sheets에 partial 표시."""
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
            swuds_function_ids=None,
        )
        assert result.ok
        assert any("Consistency" in s for s in result.incomplete_sheets)
        assert result.summary.get("consistency_swuds_compared") is False


class TestSwitAsilDistribution33:
    """SwUT _compute_asil_distribution 재활용 → SwIT summary에서도 동작."""

    def test_distribution_when_function_asil_map_provided(self):
        session = _make_swit_session()
        # function_asil_map 주입 (router의 _apply_function_asil_map 시뮬레이션)
        session.environments[0].function_asil_map = {"SwUFn_0101": "D"}
        result = build_swit_coverage_report(
            session, _make_swit_meta(), _build_swit_template(),
        )
        assert result.ok
        assert "ASIL_D" in result.summary["asil_distribution"]
        assert "SwUFn_0101" in result.summary["asil_d_function_ids"]


class TestSwitSwudsFunctionIds:
    """SwUDS function_ids 제공 시 2.Consistency 매핑 row 추가 + summary 갱신."""

    def test_consistency_summary_with_swuds_function_ids(self):
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
            swuds_function_ids={"SwUFn_0101"},
        )
        assert result.ok
        assert result.summary.get("consistency_swuds_compared") is True


class TestSwitSwitsConsistency:
    def test_sds_swits_template_exception_annotations_are_preserved(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        wb.create_sheet("1.Test Summary")
        wb.create_sheet("2.Traceability")
        cons = wb.create_sheet("3.Consistency")
        cons.cell(10, 3, "SwDS")
        cons.cell(10, 4, "SwITS")
        cons.cell(11, 2, 1)
        cons.cell(11, 3, "SDS-001")
        cons.cell(11, 6, "X")
        cons.cell(11, 7, "Legacy exception note")
        cons.cell(12, 2, 2)
        cons.cell(12, 3, "SDS-002")
        wb.create_sheet("4.Coverage")
        wb.create_sheet("History")
        buf = io.BytesIO()
        wb.save(buf)

        result = build_swit_coverage_report(
            _make_swit_session(),
            _make_swit_meta(),
            buf.getvalue(),
        )

        assert result.ok
        assert result.summary.get("consistency_sds_swits_compared") is True
        out_wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        out_cons = out_wb["3.Consistency"]
        assert out_cons.cell(11, 5).value == "O"
        # row 11: 템플릿에 미리 기재된 Exception 'X' + note는 preserved (reviewer 수기).
        assert out_cons.cell(11, 6).value == "X"
        assert out_cons.cell(11, 7).value == "Legacy exception note"
        # 라운드 102 — row 12(미주석 default)는 Exception 공백 (회사 감사본 정합).
        # 기존 'X' 강제 기재 제거: F121=COUNTIF(F,"O") 집계 무의미 + REF 공백.
        assert out_cons.cell(12, 6).value in (None, "")
        assert out_cons.cell(12, 7).value in (None, "")
        assert any("exception annotations preserved" in w for w in result.warnings)

    def test_consistency_uses_swits_unit_name_to_log_env_match(self):
        swits_map = {
            "SwITC_0101_01": SimpleNamespace(
                tc_id="SwITC_0101_01",
                unit_name="SWTE_01",
            ),
            "SwITC_0101_02": SimpleNamespace(
                tc_id="SwITC_0101_02",
                unit_name="SWTE_01_01",
            ),
            "SwITC_0999": SimpleNamespace(
                tc_id="SwITC_0999",
                unit_name="SWTE_MISSING",
            ),
        }
        result = build_swit_coverage_report(
            _make_swit_session(),
            _make_swit_meta(),
            _build_swit_template(),
            swuds_function_ids={"SwUFn_9999"},
            swits_map=swits_map,
        )

        assert result.ok
        assert result.summary.get("consistency_swits_log_compared") is True
        assert result.summary.get("consistency_swuds_compared") is not True

        wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        cons = wb["2.Consistency"]
        assert cons.cell(3, 4).value == "SwITS, VectorCAST Log"
        assert cons.cell(4, 4).value == 2
        assert cons.cell(5, 4).value == 1
        assert cons.cell(11, 3).value == "SwITC_0101_01"
        assert cons.cell(11, 4).value == "SWTE_01"
        assert cons.cell(11, 5).value == "O"
        assert cons.cell(12, 3).value == "SwITC_0101_02"
        assert cons.cell(12, 4).value == "SWTE_01"
        assert cons.cell(12, 5).value == "O"
        assert cons.cell(13, 3).value == "SwITC_0999"
        assert cons.cell(13, 5).value == "X"


class TestSwitMetaValidation:
    """validate_build_meta 패턴 — release_sw_version / test_date 형식."""

    def test_invalid_release_sw_version_raises(self):
        bad_meta = SwitCoverageBuildMeta(
            project_id="HDPDM01",
            release_sw_version="bad-version",
            test_date="2024-02-19",
            test_engineer="JK Kim",
        )
        with pytest.raises(Exception):  # noqa: B017
            build_swit_coverage_report(
                _make_swit_session(), bad_meta, _build_swit_template(),
            )

    def test_invalid_template_bytes_raises(self):
        """ZIP bomb / magic byte 검증 — non-xlsx bytes 거부."""
        with pytest.raises(Exception):  # noqa: B017
            build_swit_coverage_report(
                _make_swit_session(), _make_swit_meta(),
                template_bytes=b"not an xlsx",
            )


class TestSwitResultDataclass:
    """SwitCoverageBuildResult API — xlsx_bytes property + to_dict."""

    def test_xlsx_bytes_property_returns_full_content(self):
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
        )
        assert isinstance(result.xlsx_bytes, bytes)
        assert len(result.xlsx_bytes) == result.result_size_bytes

    def test_to_dict_includes_required_keys(self):
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
        )
        d = result.to_dict()
        for k in ("ok", "filename", "result_size_bytes", "warnings",
                  "incomplete_sheets", "summary", "tool_qualification"):
            assert k in d


class TestSwitFilenameSafeDate:
    """short_date 변환 — '2024-02-19' → '240219' 패턴."""

    def test_filename_short_date_format(self):
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), _build_swit_template(),
        )
        # short_date(2024-02-19) → 240219
        assert "240219" in result.filename


# ---------------------------------------------------------------------------
# 54차 T282/T283 — v2.02 양식 호환 회귀
# ---------------------------------------------------------------------------

def _build_v202_template() -> bytes:
    """SwIT v2.02 양식 mimic — "SW Version" / "HW Version" / B17 TC stats / B22 Requirements."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cover = wb.create_sheet("Cover")
    cover["B1"] = "Project"
    cover["B2"] = "ASIL Level"
    cover["B3"] = "Author"
    cover["B4"] = "Approver"

    # v2.02 양식: "1.Test Summary" + SW Version/HW Version + TC stats row + Requirements row
    ts = wb.create_sheet("1.Test Summary")
    ts["B1"] = "Project Name"
    ts["B2"] = "SW Version"      # v2.02 라벨
    ts["B3"] = "HW Version"      # v2.02 라벨
    ts["B4"] = "Test Date"
    ts["B5"] = "Test Engineer"
    ts["B6"] = "Final Test Result"
    # 신규 row — TC stats (Total/Tested/Passed/Failed/Blocked)
    ts["A17"] = "Total TC"
    # 신규 row — Requirements/Design Coverage
    ts["A22"] = "Requirements/Design Coverage"

    wb.create_sheet("1.Traceability")
    wb.create_sheet("2.Consistency")
    cov = wb.create_sheet("3.Coverage")
    cov["A6"] = "Unit ID"
    cov["B6"] = "Name"
    cov["C6"] = "Count"
    cov["D6"] = "Total"
    cov["E6"] = "Pass"
    wb.create_sheet("History")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSwitV202LayoutCompat:
    """54차 T282/T283 — v2.02 양식 label 매칭 + TC stats + B22 fill."""

    def test_sw_version_label_filled(self):
        """v2.02 'SW Version' 라벨 옆 셀에 release_sw_version 채움."""
        template = _build_v202_template()
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), template,
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        ts = wb["1.Test Summary"]
        # "SW Version" label은 B2 → value는 C2 (col+1)
        assert ts["C2"].value == "2.02"

    def test_hw_version_label_filled(self):
        """v2.02 'HW Version' 라벨 옆 셀에 hw_version 채움."""
        template = _build_v202_template()
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), template,
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        ts = wb["1.Test Summary"]
        # HW Version 라벨 B3 → value C3 (col+1). default "1.00"
        assert ts["C3"].value == "1.00"

    def test_tc_stats_row_filled(self):
        """55-fix: 라벨 row=17이 헤더 (가로 배치), data는 row=18에 채움."""
        template = _build_v202_template()
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), template,
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        ts = wb["1.Test Summary"]
        # 55-fix: "Total TC" 라벨 A17에 위치 → tc_stats_row=18 (data), col_start=1 (A)
        assert ts["A18"].value == 1   # Total
        assert ts["B18"].value == 1   # Tested
        assert ts["C18"].value == 1   # Passed
        assert ts["D18"].value == 0   # Failed
        assert ts["E18"].value == 0   # Blocked (inferred)
        assert result.summary.get("tc_stats_blocked_inferred") is True

    def test_requirements_row_filled_with_swits(self):
        """B22 Requirements/Design Coverage row에 SwITS 표기."""
        template = _build_v202_template()
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), template,
        )
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        ts = wb["1.Test Summary"]
        # Requirements/Design Coverage row=22, B22(col 2)에 "SwITS"
        assert ts["B22"].value == "SwITS"

    def test_v301_backward_compat(self):
        """v3.01 양식 (Release Name(SW) 라벨)도 정상 채움 — backward compat."""
        # 기존 _build_swit_template은 Release Name(SW) 라벨 사용 (v3.01 호환)
        template = _build_swit_template()
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), template,
        )
        # v3.01 호환 — fill 성공 + tc_stats_blocked_inferred 없음 (v2.02 row 없음)
        assert result.ok
        assert "tc_stats_blocked_inferred" not in result.summary


class TestSwutBuilderV202InspectFix54:
    """54-fix C1 — SwUT Coverage 빌더가 v2.02 template 잘못 입력 받아도 silent 빈 셀 차단.

    SwUT 라우터에 사용자가 SwIT v2.02 양식 path를 잘못 지정 시:
    - 이전: hardcode "Release Name(SW)" 라벨 미발견 → silent 빈 셀
    - 54-fix: inspect_swit_layout → v2.02 라벨 매핑 → SW Version 옆 자동 채움
    """

    def test_swut_coverage_with_v202_template_fills_sw_version(self):
        """SwUT 빌더에 v2.02 template 입력 → SW Version cell 자동 채움."""
        from backend.services.swut_coverage_aggregator import (
            CoverageBuildMeta,
            build_coverage_report,
        )
        meta = CoverageBuildMeta(
            project_id="HDPDM01",
            release_sw_version="1.0.5",
            test_date="2024-02-19",
            test_engineer="JK Kim",
            doc_id_sequence="001",
        )
        result = build_coverage_report(
            _make_swit_session(), meta, _build_v202_template(),
        )
        assert result.ok
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        # v2.02 "1.Test Summary" 시트의 SW Version 옆 셀
        ts = wb["1.Test Summary"]
        # SW Version 라벨 B2 → value C2 (col+1)
        assert ts["C2"].value == "1.0.5"

    def test_swut_coverage_with_v301_template_backward_compat(self):
        """SwUT 기존 v3.01 양식도 정상 채움 — fallback_to_v301."""
        from backend.services.swut_coverage_aggregator import (
            CoverageBuildMeta,
            build_coverage_report,
        )
        meta = CoverageBuildMeta(
            project_id="HDPDM01",
            release_sw_version="1.0.5",
            test_date="2024-02-19",
            test_engineer="JK Kim",
            doc_id_sequence="001",
        )
        # 기존 v3.01 호환 template (Release Name(SW) 라벨)
        result = build_coverage_report(
            _make_swit_session(), meta, _build_swit_template(),
        )
        assert result.ok
        # v3.01 → fallback_to_v301=True, tc_stats_blocked_inferred 부재
        assert "tc_stats_blocked_inferred" not in result.summary


class TestV202CoverageLabelMissingFallback56:
    """56차 T306 — 회사 Coverage Report v2.02는 row 17 (TC stats) + row 20
    (Requirements) 라벨이 부재한 사용자 수동 입력 영역. 빈 row 감지 → builder가
    라벨+데이터 모두 stamp. SITR과 audit 완성도 대칭.
    """

    def _v202_label_missing_template(self) -> bytes:
        """SW Version/HW Version 라벨은 있지만 TC stats + Requirements row 라벨 부재."""
        import io
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        cover = wb.create_sheet("Cover")
        cover["B1"] = "Project"
        cover["B2"] = "ASIL Level"
        cover["B3"] = "Author"
        cover["B4"] = "Approver"
        ts = wb.create_sheet("1.Test Summary")
        ts["B1"] = "Project Name"
        ts["B2"] = "SW Version"
        ts["B3"] = "HW Version"
        ts["B4"] = "Test Date"
        ts["B5"] = "Test Engineer"
        ts["B6"] = "Final Test Result"
        # row 17 (TC stats) + row 20~22 (Requirements) 모두 빈 row — 라벨 부재
        wb.create_sheet("1.Traceability")
        wb.create_sheet("2.Consistency")
        cov = wb.create_sheet("3.Coverage")
        cov["A6"] = "Unit ID"
        wb.create_sheet("History").cell(1, 1, "■ Revision History")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_builder_stamps_labels_when_template_has_empty_row_17(self):
        """row 17 빈 cell → builder가 label 5개 stamp + data 5개 fill + summary flag."""
        template = self._v202_label_missing_template()
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), template,
        )
        assert result.ok
        # summary에 fallback_used flag set
        assert result.summary.get("tc_stats_fallback_used") is True
        assert result.summary.get("requirements_fallback_used") is True
        # 산출물 검증 — row 17에 라벨 5개 stamp
        wb = openpyxl.load_workbook(io.BytesIO(result.xlsx_bytes))
        ts = wb["1.Test Summary"]
        assert ts.cell(17, 2).value == "Total Number of TCs"
        assert ts.cell(17, 3).value == "Number of TCs Tested"
        assert ts.cell(17, 4).value == "Number of TCs Passed"
        assert ts.cell(17, 5).value == "Number of TCs Failed"
        assert ts.cell(17, 6).value == "Number of TCs not executed"
        # row 18에 데이터 fill (session: TC 5건)
        assert ts.cell(18, 2).value is not None  # total
        # Requirements row 20~22 fill
        assert ts.cell(20, 2).value == "■  Requirements/Design Coverage"
        assert ts.cell(21, 2).value == "Source"
        assert ts.cell(22, 2).value == "SwITS"


class TestTcStatsDataRowGuard55fix2:
    """55-fix-2 W4 — TC stats data row 비어있지 않을 때 silent overwrite 방어."""

    def test_skip_when_data_row_already_has_value(self):
        """data row에 이미 값이 있으면 fill skip + summary에 reason 누적."""
        # v2.02 template + data row 18에 사용자 값 미리 채움 (회사 양식 변형 시뮬레이션)
        import io
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        cover = wb.create_sheet("Cover")
        cover["B1"] = "Project"
        cover["B2"] = "ASIL Level"
        cover["B3"] = "Author"
        cover["B4"] = "Approver"
        ts = wb.create_sheet("1.Test Summary")
        ts["B1"] = "Project Name"
        ts["B2"] = "SW Version"
        ts["B3"] = "HW Version"
        ts["B4"] = "Test Date"
        ts["B5"] = "Test Engineer"
        ts["B6"] = "Final Test Result"
        ts["A17"] = "Total TC"  # 라벨
        ts["A18"] = "ALREADY_FILLED"  # data row에 이미 값 → 우리 코드 skip
        wb.create_sheet("1.Traceability")
        wb.create_sheet("2.Consistency")
        cov = wb.create_sheet("3.Coverage")
        cov["A6"] = "Unit ID"
        wb.create_sheet("History").cell(1, 1, "■ Revision History")
        buf = io.BytesIO()
        wb.save(buf)
        template = buf.getvalue()

        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), template,
        )
        assert result.ok
        # 55-fix-2 W4: data row 이미 값 있으면 fill skip + reason 누적
        assert "tc_stats_skipped_reason" in result.summary
        assert "ALREADY_FILLED" in result.summary["tc_stats_skipped_reason"]
        # blocked_inferred는 set 안 됨 (fill 자체 skip)
        assert "tc_stats_blocked_inferred" not in result.summary


class TestF7StageR3N7IsSwitCallerBranch:
    """F7 Round 3 N7 fix — is_swit_caller kwarg 분기 회귀.

    SwUT 호출 (build_coverage_report) default False → SwUT 분기 (Statement+Branch).
    SwIT 호출 (build_swit_coverage_report) True 명시 → SwIT 분기 (Functions Pass +
    Function Called). 향후 신규 호출처 추가 시 silent SwUT 분기 silent 결함 검출.
    """

    def _build_company_standard_swit_layout_template(self) -> bytes:
        """회사 표준 SwITCV layout — coverage_metric_kind='function_and_calls'
        + has_component_col=True. SwIT 분기 진입 조건 모두 만족."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        ts = wb.create_sheet("1.Test Summary")
        ts["B1"] = "Project Name"
        ts["B2"] = "SW Version"  # v1.01 signature
        # 회사 표준 v1.01 시트 signature
        trace = wb.create_sheet("2.Traceability")
        # SwST header (≥3) → matrix_kind='switc_x_swst' (detected_version='v1.01')
        for i in range(5):
            trace.cell(11, 4 + i).value = f"SwST_{i+1:02d}"
        cons = wb.create_sheet("3.Consistency")
        cons["A1"] = "Item"
        # 회사 표준 4.Coverage header (No/Component/Unit ID/Name) + Function Called layout
        cov = wb.create_sheet("4.Coverage")
        cov["B8"] = "No"
        cov["C8"] = "Component"
        cov["D8"] = "Unit"
        cov["F8"] = "Functions"
        cov["D9"] = "ID"
        cov["E9"] = "Name"
        cov["F9"] = "Count"
        wb.create_sheet("History")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_swit_caller_uses_function_called_layout(self):
        """build_swit_coverage_report → is_swit_caller=True → SwIT 분기.
        Functions Pass (C6='O') + Function Called metric stamp."""
        from openpyxl import load_workbook
        template = self._build_company_standard_swit_layout_template()
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), template,
        )
        assert result.ok
        wb = load_workbook(io.BytesIO(result.xlsx_io.getvalue()))
        cov = wb["4.Coverage"]
        # data_start = header_row(8) + 2 = 10. R10 stamp 검증
        # SwIT layout: C6=Functions Pass 'O', C8/C9/C10 = Function Called
        # synthetic session function_calls_coverage 빈 default → C8~C10 skip
        assert cov.cell(10, 6).value == "O", (
            f"SwIT 분기 Functions Pass 미stamp — C6={cov.cell(10, 6).value!r}"
        )
        # Statement/Branch col (C7) — SwIT 분기는 stamp 안 함 (양식 default 잔존 가능)
        # SwIT 분기 정상 진입 확인

    def test_swit_function_called_totals_row_after_expansion(self):
        """SwITCV v1.01 row expansion must clear old totals and create new totals."""
        from openpyxl import load_workbook

        template = self._build_company_standard_swit_layout_template()
        wb = load_workbook(io.BytesIO(template))
        cov = wb["4.Coverage"]
        cov["E5"] = "=B12"
        cov["F5"] = "=F12"
        cov["G5"] = "=G12"
        cov["E6"] = "=B12"
        cov["F6"] = "=J12"
        cov["G6"] = "=K12"
        cov.cell(12, 3).value = "Total"
        cov.cell(12, 7).value = '=COUNTIF(G10:G11,"O")'
        buf = io.BytesIO()
        wb.save(buf)

        env = EnvironmentData(env_name="SWTE_01", component_name="Comp")
        env.test_cases = {"SwITC_SwUFn_0101.001": []}
        env.test_results = {
            "SwITC_SwUFn_0101.001": ExecutionRow(
                tc_name="SwITC_SwUFn_0101.001", passed=True,
            )
        }
        env.function_coverage = [
            FunctionCoverage(
                unit_id=f"SwUFn_{i:04d}",
                name=f"fn_{i}",
                component_name="Comp",
                function_calls_coverage=CoverageStats(covered=1 if i % 2 else 0, total=1),
            )
            for i in range(1, 7)
        ]
        session = SwUTSession(
            project_id="KJPDS02",
            version="v1.01_251205",
            source_kind="log_folder",
            source_path="/tmp/fake",
            environments=[env],
        )
        result = build_swit_coverage_report(session, _make_swit_meta(), buf.getvalue())
        assert result.ok

        wb_out = load_workbook(io.BytesIO(result.xlsx_io.getvalue()))
        cov_out = wb_out["4.Coverage"]
        assert "Total" not in [cov_out.cell(r, 3).value for r in range(10, 16)]
        assert cov_out.cell(16, 2).value == 6
        assert cov_out.cell(16, 4).value == "Total"
        assert cov_out.cell(16, 6).value == 0
        assert cov_out.cell(16, 10).value == 3
        assert cov_out.cell(5, 5).value == "=B16"
        assert cov_out.cell(6, 6).value == "=J16"
        assert any("[swit-cov]" in w for w in result.warnings)

    def test_coverage_sheet_clear_form_default_function_rows_round_f8(self):
        """F7 R2 N5 carry-over — C2 Coverage clear 단위 회귀.
        SwIT layout (회사 표준 SwITCV)에서 신규 stamp 후 양식 default 함수 row clear
        + 수식 cell preserve_formula + sentinel 보존."""
        from openpyxl import load_workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        ts = wb.create_sheet("1.Test Summary")
        ts["B1"] = "Project Name"
        ts["B2"] = "SW Version"  # v1.01 signature
        trace = wb.create_sheet("2.Traceability")
        for i in range(5):
            trace.cell(11, 4 + i).value = f"SwST_{i+1:02d}"
        cons = wb.create_sheet("3.Consistency")
        cons["A1"] = "Item"
        cov = wb.create_sheet("4.Coverage")
        cov["B8"] = "No"
        cov["C8"] = "Component"
        cov["D8"] = "Unit"
        cov["D9"] = "ID"
        cov["E9"] = "Name"
        # 양식 default 함수 row R12~R14 (신규 session엔 없는 양식 default)
        cov.cell(12, 4).value = "SwUFn_9999"
        cov.cell(12, 5).value = "form_default_func"
        cov.cell(12, 6).value = "O"
        cov.cell(13, 4).value = "SwUFn_9998"
        cov.cell(13, 5).value = "Fun_B"
        # sentinel — clear 차단
        cov.cell(20, 2).value = "< End of Document >"
        # 수식 — preserve_formula
        cov.cell(14, 8).value = "=IF(F14=G14, \"O\", \"X\")"
        wb.create_sheet("History")

        buf = io.BytesIO()
        wb.save(buf)
        result = build_swit_coverage_report(
            _make_swit_session(), _make_swit_meta(), buf.getvalue(),
        )
        assert result.ok
        wb_out = load_workbook(io.BytesIO(result.xlsx_io.getvalue()))
        cov_out = wb_out["4.Coverage"]
        # 신규 session 1 function → R10 stamp. R12~ default clear, sentinel/formula 보존
        assert cov_out.cell(12, 4).value is None  # default 함수 clear
        assert cov_out.cell(12, 5).value is None
        assert cov_out.cell(12, 6).value is None
        assert cov_out.cell(13, 4).value is None
        # sentinel 보존
        assert cov_out.cell(20, 2).value == "< End of Document >"
        # 수식 preserve_formula
        assert cov_out.cell(14, 8).value == "=IF(F14=G14, \"O\", \"X\")"
        # clear warning emit
        assert any("Coverage 시트" in w and "clear" in w for w in result.warnings)

    def test_swit_caller_default_false_swut_branch_in_unit_call(self):
        """build_coverage_report (SwUT) default is_swit_caller=False → SwUT 분기.
        회사 표준 v1.01 양식 (coverage_metric_kind=function_and_calls + has_component_col)
        에서도 SwUT은 Statement+Branch stamp."""
        from openpyxl import load_workbook

        from backend.services.swut_coverage_aggregator import (
            CoverageBuildMeta,
            build_coverage_report,
        )
        template = self._build_company_standard_swit_layout_template()
        meta = CoverageBuildMeta(
            project_id="HDPDM01", release_sw_version="2.02",
            test_date="2024-02-19", test_engineer="JK Kim",
            doc_id_sequence="001",
        )
        result = build_coverage_report(
            _make_swit_session(), meta, template,
        )
        assert result.ok
        wb = load_workbook(io.BytesIO(result.xlsx_io.getvalue()))
        cov = wb["4.Coverage"]
        # SwUT 분기 stamp: stmt_count_col = no_col(2) + 4 = 6
        # session function_coverage: statement=CoverageStats() (default 0/0/0.0)
        # → C6=0 (total), C7=0 (covered), C8='X' (passed=False, total=0)
        # SwIT 분기와 달리 Statement metric stamp 시도 (값이 0이지만 col 위치 정확)
        # 핵심 — C6에 'O' (SwIT branch) 아닌 정수 0 (SwUT stmt.total)
        c6 = cov.cell(10, 6).value
        assert c6 != "O" or c6 == 0, (
            f"SwUT 분기인데 SwIT 분기 진입 — C6={c6!r} (의도: int/None)"
        )


class TestAlignDroppedFunctionWarning:
    """라운드 96-fix W-D — template universe 밖 실측 함수의 silent drop 경고.

    KJPDS02 PV 실측: BOOT 함수 76개가 SwUDS v2.01에 등재돼 있으나 DV 스코프
    template 571행에 부재 → 4.Coverage에서 조용히 제외되던 결함. 원시 행은
    함수명만 보유하므로 SwUDS name→SwUFn reverse map 교차로 검출.
    """

    @staticmethod
    def _run_align(agg):
        from backend.services.swit_coverage_aggregator import (
            _align_function_rows_to_template,
        )
        warnings: list[str] = []
        template_rows = [("SwUFn_0101", "main", 1)]
        _align_function_rows_to_template(agg, template_rows, out_warnings=warnings)
        return warnings

    def test_designed_dropped_function_warned(self):
        from backend.services.swut_input_adapter import (
            CoverageStats,
            FunctionCoverage,
        )
        matched = FunctionCoverage(
            unit_id="", name="main",
            statement=CoverageStats(1, 1, 1.0), branch=CoverageStats(1, 1, 1.0),
        )
        boot_fn = FunctionCoverage(
            unit_id="", name="FBL_BootMain",
            statement=CoverageStats(1, 1, 1.0), branch=CoverageStats(1, 1, 1.0),
        )
        helper_fn = FunctionCoverage(  # SwUDS 미등재 보조 함수 — 노이즈 차단 확인
            unit_id="", name="memcpy_local",
            statement=CoverageStats(1, 1, 1.0), branch=CoverageStats(1, 1, 1.0),
        )
        agg = {
            "function_rows": [matched, boot_fn, helper_fn],
            "function_name_to_swufn_from_suds": {"FBL_BootMain": "SwUFn_3514"},
        }
        warnings = self._run_align(agg)
        dropped = [w for w in warnings if "제외됨" in w]
        assert len(dropped) == 1
        assert "SwUFn_3514:FBL_BootMain" in dropped[0]
        assert "memcpy_local" not in dropped[0]

    def test_no_dropped_no_warning(self):
        from backend.services.swut_input_adapter import (
            CoverageStats,
            FunctionCoverage,
        )
        matched = FunctionCoverage(
            unit_id="", name="main",
            statement=CoverageStats(1, 1, 1.0), branch=CoverageStats(1, 1, 1.0),
        )
        agg = {
            "function_rows": [matched],
            "function_name_to_swufn_from_suds": {"main": "SwUFn_0101"},
        }
        warnings = self._run_align(agg)
        assert not [w for w in warnings if "제외됨" in w]


class TestNoAutoException96Final:
    """라운드 96-final QA fix (W#5) — 미실행/미달 함수 Exception 자동 'O' 제거.

    이전: 미실행 46함수 전건 Exception 'O' → Coverage 수식 100% 표시 (audit 위험).
    신규: Exception 셀 노란 fill만 (값 비움) + Note(L열) 사유 기재 안내 +
    totals Exception=0 + 정책 warning.
    """

    def _template(self) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Cover")
        ts = wb.create_sheet("1.Test Summary")
        ts["B1"] = "Project Name"
        ts["B2"] = "SW Version"
        trace = wb.create_sheet("2.Traceability")
        for i in range(5):
            trace.cell(11, 4 + i).value = f"SwST_{i+1:02d}"
        cons = wb.create_sheet("3.Consistency")
        cons["A1"] = "Item"
        cov = wb.create_sheet("4.Coverage")
        cov["B8"] = "No"
        cov["C8"] = "Component"
        cov["D8"] = "Unit"
        cov["F8"] = "Functions"
        cov["D9"] = "ID"
        cov["E9"] = "Name"
        cov["F9"] = "Count"
        wb.create_sheet("History")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_missing_function_no_auto_exception(self):
        from openpyxl import load_workbook
        session = _make_swit_session()
        env = session.environments[0]
        missing = FunctionCoverage(
            unit_id="SwUFn_9999", name="never_executed",
            statement=CoverageStats(0, 1, 0.0), branch=CoverageStats(0, 1, 0.0),
            function_calls_coverage=CoverageStats(0, 1, 0.0),
        )
        setattr(missing, "swit_function_present", False)
        env.function_coverage.append(missing)

        result = build_swit_coverage_report(
            session, _make_swit_meta(), self._template(),
        )
        assert result.ok
        wb = load_workbook(io.BytesIO(result.xlsx_io.getvalue()))
        cov = wb["4.Coverage"]
        # data_start=10 — missing 함수 행 탐색 (F='X')
        miss_row = None
        for r in range(10, cov.max_row + 1):
            if cov.cell(r, 6).value == "X":
                miss_row = r
                break
        assert miss_row is not None, "미실행 함수 행(F='X') 미발견"
        # Exception(G) — 자동 'O' 금지, 값 비움 + 노란 fill
        assert cov.cell(miss_row, 7).value in (None, ""), (
            f"Exception 자동 stamp 잔존: {cov.cell(miss_row, 7).value!r}"
        )
        assert cov.cell(miss_row, 7).fill.start_color.rgb == "FFFFEB9C"
        # Note(L=12) — 사유 기재 안내
        assert str(cov.cell(miss_row, 12).value or "").startswith("▶")
        # 정책 warning 노출
        assert any("Exception 자동" in w for w in result.warnings)

    def test_totals_exception_count_zero(self):
        from openpyxl import load_workbook
        session = _make_swit_session()
        env = session.environments[0]
        missing = FunctionCoverage(
            unit_id="SwUFn_9999", name="never_executed",
            statement=CoverageStats(0, 1, 0.0), branch=CoverageStats(0, 1, 0.0),
        )
        setattr(missing, "swit_function_present", False)
        env.function_coverage.append(missing)
        result = build_swit_coverage_report(
            session, _make_swit_meta(), self._template(),
        )
        assert result.ok
        wb = load_workbook(io.BytesIO(result.xlsx_io.getvalue()))
        cov = wb["4.Coverage"]
        # totals 행 — 'Total' 라벨 행에서 G(Exception 합)=0
        total_row = None
        for r in range(10, cov.max_row + 1):
            if any(str(cov.cell(r, c).value or "").strip() == "Total" for c in (3, 4)):
                total_row = r
                break
        assert total_row is not None
        assert cov.cell(total_row, 7).value == 0
