from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from tools.export_suts_vectorcast import build_vectorcast_model, export_suts_to_vectorcast_model


def _make_sample_suts(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    ws.cell(row=6, column=2, value="Component")
    ws.cell(row=6, column=3, value="TC ID")
    ws.cell(row=6, column=4, value="Name")
    ws.cell(row=6, column=5, value="Description")
    ws.cell(row=6, column=8, value="Test Method")
    ws.cell(row=6, column=9, value="Gen.Method")
    ws.cell(row=6, column=10, value="Precondition")
    ws.cell(row=6, column=11, value="Sequence")
    ws.cell(row=6, column=12, value="Test Case Gen.Method")
    ws.cell(row=6, column=13, value="Seq. No.")

    # TC block 1
    ws.cell(row=7, column=2, value="SwCom_01")
    ws.cell(row=7, column=3, value="SwUTC_SwUFn_0001")
    ws.cell(row=7, column=4, value="g_TestFunc")
    ws.cell(row=7, column=5, value="desc\n[SRS: SwTR_0001]")
    ws.cell(row=7, column=8, value="FIT")
    ws.cell(row=7, column=9, value="ABV")
    ws.cell(row=7, column=10, value="precondition")
    ws.cell(row=7, column=14, value="u8g_InA")
    ws.cell(row=7, column=15, value="u8g_InB")
    ws.cell(row=7, column=63, value="u8s_OutA")
    ws.cell(row=7, column=149, value="SwUFn_0001")

    ws.cell(row=8, column=11, value="seq-1")
    ws.cell(row=8, column=13, value=1)
    ws.cell(row=8, column=14, value=1)
    ws.cell(row=8, column=15, value=2)
    ws.cell(row=8, column=63, value=3)

    ws.cell(row=9, column=11, value="seq-2")
    ws.cell(row=9, column=13, value=2)
    ws.cell(row=9, column=14, value=10)
    ws.cell(row=9, column=15, value=20)
    ws.cell(row=9, column=63, value="[검증 필요] 30")

    # TC block 2
    ws.cell(row=10, column=2, value="SwCom_02")
    ws.cell(row=10, column=3, value="SwUTC_SwUFn_0002")
    ws.cell(row=10, column=4, value="s_AnotherFunc")
    ws.cell(row=10, column=5, value="another desc")
    ws.cell(row=10, column=8, value="FNCT")
    ws.cell(row=10, column=9, value="AEC, ABV")
    ws.cell(row=10, column=14, value="g_Input")
    ws.cell(row=10, column=63, value="g_Output")
    ws.cell(row=10, column=149, value="SwUFn_0002")

    ws.cell(row=11, column=11, value="seq-1")
    ws.cell(row=11, column=13, value=1)
    ws.cell(row=11, column=14, value=0)
    ws.cell(row=11, column=63, value=1)

    wb.save(path)
    wb.close()


def test_build_vectorcast_model_parses_tc_blocks(tmp_path: Path) -> None:
    suts_path = tmp_path / "sample.xlsm"
    _make_sample_suts(suts_path)

    model = build_vectorcast_model(str(suts_path), project_id="TEST")

    assert model["project_id"] == "TEST"
    assert len(model["units"]) == 2
    assert model["units"][0]["unit_name"] == "g_TestFunc"
    assert len(model["units"][0]["test_cases"]) == 2
    assert model["units"][0]["test_cases"][0]["inputs"] == {"u8g_InA": 1, "u8g_InB": 2}
    assert model["units"][0]["test_cases"][0]["expected"] == {"u8s_OutA": 3}
    assert model["units"][0]["test_cases"][1]["expected"]["u8s_OutA"]["verification_required"] is True


def test_build_vectorcast_model_filters_target_functions(tmp_path: Path) -> None:
    suts_path = tmp_path / "sample.xlsm"
    _make_sample_suts(suts_path)

    model = build_vectorcast_model(str(suts_path), target_functions=["s_AnotherFunc"])

    assert len(model["units"]) == 1
    assert model["units"][0]["unit_name"] == "s_AnotherFunc"


def test_export_suts_to_vectorcast_model_writes_outputs(tmp_path: Path) -> None:
    suts_path = tmp_path / "sample.xlsm"
    json_path = tmp_path / "out.json"
    warnings_path = tmp_path / "warnings.md"
    _make_sample_suts(suts_path)

    model = export_suts_to_vectorcast_model(
        str(suts_path),
        str(json_path),
        warnings_md=str(warnings_path),
    )

    assert json_path.exists()
    assert warnings_path.exists()
    loaded = json.loads(json_path.read_text(encoding="utf-8"))
    assert loaded["schema_version"] == "1.0"
    assert len(model["export_warnings"]) >= 1


def test_parse_sequence_row_non_numeric_seq_no_no_crash():
    """비숫자 sequence_no(else 분기)에서 NameError로 크래시하지 않는다.

    회귀: L106이 미정의 `sequence_no_text`를 참조해 sequence_no가 비숫자인 SUTS(예 KJPDS02
    SwUTS)에서 build_vectorcast_model 전체가 크래시 → 회귀 TC가 조용히 0이 되던 버그.
    """
    from tools.export_suts_vectorcast import (
        _SEQ_NO_COL,
        _SEQUENCE_TEXT_COL,
        _parse_sequence_row,
    )

    class _Cell:
        def __init__(self, v):
            self.value = v

    class _WS:
        def cell(self, row, column):
            if column == _SEQ_NO_COL:
                return _Cell("A")  # 비숫자 시퀀스 번호 → else 분기 유발
            if column == _SEQUENCE_TEXT_COL:
                return _Cell("desc")
            return _Cell(None)

    _cols = {"seq_no": _SEQ_NO_COL, "seq_text": _SEQUENCE_TEXT_COL}
    seq, _warns = _parse_sequence_row(_WS(), 5, [], [], {"base_tc_id": "TC01", "unit_name": "foo"}, _cols)
    assert seq["name"] == "TC01__SEQ_A"  # seq_no_text 사용(수정 전엔 NameError)
    assert seq["sequence_no"] == "A"


# ── 레이아웃 적응형 컬럼 탐지 (SUTS 템플릿 2종) ───────────────────────────────
# 실측: HDPDM01 v3.01 은 SeqNo=13/Input=14../Expected=63../Related=149(모듈 상수와 일치),
# KJPDS02_PV v1.02 는 SeqNo=8/Inpt[0]=9../ExpR[0]=105../Related=189 로 레이아웃이 다르다.
# 하드코딩 상수는 KJPDS02_PV 에서 입력컬럼(col13)을 SeqNo 게이트로 오용 → 입력 파라미터가
# 4개 이하인 함수의 전 시퀀스행이 '빈 블록'으로 드롭됐다(SUTS 카드 '미파싱' 오표시의 근본).


def _make_kjpds02_style_suts(path: Path) -> None:
    """KJPDS02_PV v1.02 레이아웃(축약): 필드헤더 row4, Inpt[0]=c9, Expected=c20, Related=c30,
    SeqNo=c8(=입력시작-1). Description/Precondition/Sequence-text 컬럼 부재. 병합앵커 'Input'@c8
    이 SeqNo 컬럼에 걸리므로 Inpt[0] 우선 신뢰를 검증한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    # 그룹/병합 헤더 (row3)
    ws.cell(row=3, column=8, value="Input")           # 병합앵커 — SeqNo 컬럼에 걸림(오탐 유발원)
    ws.cell(row=3, column=20, value="Expected Result")
    ws.cell(row=3, column=30, value="Related ID")
    # 필드/개별 헤더 (row4)
    ws.cell(row=4, column=2, value="Index")
    ws.cell(row=4, column=3, value="TC_ID")
    ws.cell(row=4, column=4, value="Unit")
    ws.cell(row=4, column=5, value="Safety Related")
    ws.cell(row=4, column=6, value="Test Method")
    ws.cell(row=4, column=7, value="Test Case Generation Method")
    ws.cell(row=4, column=9, value="Inpt[0]")
    ws.cell(row=4, column=10, value="Inpt[1]")
    ws.cell(row=4, column=20, value="ExpR[0]")
    ws.cell(row=4, column=30, value="SUDS")
    # 블록1 — 입력 2개(핵심 회귀: 구 파서는 col13 비어 전 시퀀스 드롭)
    ws.cell(row=7, column=3, value="SwUTC_SwUFn_0001")
    ws.cell(row=7, column=4, value="s_lowinput")
    ws.cell(row=7, column=9, value="u8_argA")   # 입력 변수명(TC_ID 행)
    ws.cell(row=7, column=10, value="u8_argB")
    ws.cell(row=7, column=20, value="u8_out")   # 출력 변수명
    ws.cell(row=8, column=8, value=1)           # SeqNo (c8)
    ws.cell(row=8, column=9, value=10)
    ws.cell(row=8, column=10, value=20)
    ws.cell(row=8, column=20, value=30)
    ws.cell(row=9, column=8, value=2)
    ws.cell(row=9, column=9, value=11)
    ws.cell(row=9, column=10, value=21)
    ws.cell(row=9, column=20, value=31)
    # 블록2
    ws.cell(row=10, column=3, value="SwUTC_SwUFn_0002")
    ws.cell(row=10, column=4, value="s_other")
    ws.cell(row=10, column=9, value="inX")
    ws.cell(row=10, column=20, value="outY")
    ws.cell(row=11, column=8, value=1)
    ws.cell(row=11, column=9, value=5)
    ws.cell(row=11, column=20, value=6)
    wb.save(path)
    wb.close()


def _make_hdpdm01_style_suts(path: Path) -> None:
    """HDPDM01 v3.01 레이아웃(전역 Input/Expected/Related 헤더 포함): 탐지가 모듈 상수와
    동일한 컬럼(SeqNo=13/Input=14/Expected=63/Related=149)을 산출해야 한다(무회귀)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    ws.cell(row=5, column=14, value="Input")
    ws.cell(row=5, column=63, value="Expected Result")
    ws.cell(row=5, column=149, value="Related ID")
    ws.cell(row=6, column=2, value="Component")
    ws.cell(row=6, column=3, value="TC ID")
    ws.cell(row=6, column=4, value="Name")
    ws.cell(row=6, column=5, value="Description")
    ws.cell(row=6, column=8, value="Test Method")
    ws.cell(row=6, column=9, value="Test Case Generation Method")
    ws.cell(row=6, column=10, value="Precondition")
    ws.cell(row=6, column=11, value="Sequence")
    ws.cell(row=6, column=13, value="Seq. No.")
    ws.cell(row=7, column=3, value="SwUTC_SwUFn_0001")
    ws.cell(row=7, column=4, value="g_hdfunc")
    ws.cell(row=7, column=5, value="desc")
    ws.cell(row=7, column=8, value="FIT")
    ws.cell(row=7, column=10, value="pre")
    ws.cell(row=7, column=14, value="inA")
    ws.cell(row=7, column=63, value="outA")
    ws.cell(row=7, column=149, value="REL_01")
    ws.cell(row=8, column=11, value="seq-1")
    ws.cell(row=8, column=13, value=1)
    ws.cell(row=8, column=14, value=100)
    ws.cell(row=8, column=63, value=200)
    wb.save(path)
    wb.close()


def test_detect_columns_kjpds02_layout(tmp_path: Path) -> None:
    """KJPDS02_PV 레이아웃: SeqNo=8(=Inpt[0]-1), Input=9.., Expected=20.., Related=30.
    부재 컬럼(Description/Precondition/Sequence-text)은 None(입력컬럼 오독 방지)."""
    from openpyxl import load_workbook

    from tools.export_suts_vectorcast import _detect_columns

    p = tmp_path / "kj.xlsm"
    _make_kjpds02_style_suts(p)
    ws = load_workbook(p, keep_vba=True)["2.SW Unit Test Spec"]
    cols = _detect_columns(ws)
    assert cols["input_start"] == 9        # Inpt[0] 우선(병합앵커 'Input'@c8 무시)
    assert cols["seq_no"] == 8             # 불변식: 입력시작-1
    assert cols["input_end"] == 19         # 출력시작-1
    assert cols["output_start"] == 20
    assert cols["output_end"] == 29        # Related-1
    assert cols["related"] == 30
    assert cols["test_method"] == 6
    assert cols["gen_method"] == 7
    # 부재 컬럼 → None (구 파서는 여기서 입력컬럼을 오독해 쓰레기값을 넣었다)
    assert cols["description"] is None
    assert cols["precondition"] is None
    assert cols["seq_text"] is None


def test_kjpds02_low_input_fn_recovered(tmp_path: Path) -> None:
    """핵심 회귀: 입력 2개 함수의 시퀀스행이 col13(구 SeqNo 게이트)에 없어 통째로 드롭되던 것 →
    header-구동 게이트(c8)로 복구. SUTS 카드 '미파싱'의 근본 수정."""
    from openpyxl import load_workbook

    from tools.export_suts_vectorcast import build_vectorcast_model

    p = tmp_path / "kj.xlsm"
    _make_kjpds02_style_suts(p)
    # 전제 증거: 구 게이트 컬럼(13)은 시퀀스행에서 비어 있다(그래서 구 파서가 드롭했다).
    ws = load_workbook(p, keep_vba=True)["2.SW Unit Test Spec"]
    assert ws.cell(row=8, column=13).value in (None, "")

    model = build_vectorcast_model(str(p), project_id="KJ")
    units = {u["unit_name"]: u for u in model["units"]}
    assert "s_lowinput" in units
    tcs = units["s_lowinput"]["test_cases"]
    assert len(tcs) == 2                                   # 복구(구 파서는 0)
    assert tcs[0]["inputs"] == {"u8_argA": 10, "u8_argB": 20}   # 정확 컬럼
    assert tcs[0]["expected"] == {"u8_out": 30}
    assert tcs[0]["description"] == ""                     # seq-text 컬럼 부재 → "" (쓰레기 아님)
    assert units["s_other"]["test_cases"][0]["inputs"] == {"inX": 5}


def test_detect_columns_hdpdm01_layout_matches_constants(tmp_path: Path) -> None:
    """HDPDM01 레이아웃 무회귀: 탐지가 모듈 상수와 동일 컬럼 산출."""
    from openpyxl import load_workbook

    from tools.export_suts_vectorcast import (
        _DESCRIPTION_COL,
        _INPUT_COL_START,
        _OUTPUT_COL_START,
        _PRECONDITION_COL,
        _RELATED_COL,
        _SEQ_NO_COL,
        _SEQUENCE_TEXT_COL,
        _detect_columns,
    )

    p = tmp_path / "hd.xlsm"
    _make_hdpdm01_style_suts(p)
    ws = load_workbook(p, keep_vba=True)["2.SW Unit Test Spec"]
    cols = _detect_columns(ws)
    assert cols["seq_no"] == _SEQ_NO_COL == 13
    assert cols["input_start"] == _INPUT_COL_START == 14
    assert cols["output_start"] == _OUTPUT_COL_START == 63
    assert cols["related"] == _RELATED_COL == 149
    assert cols["description"] == _DESCRIPTION_COL == 5
    assert cols["precondition"] == _PRECONDITION_COL == 10
    assert cols["seq_text"] == _SEQUENCE_TEXT_COL == 11

    model = build_vectorcast_model(str(p), project_id="HD")
    unit = model["units"][0]
    assert unit["unit_name"] == "g_hdfunc"
    assert unit["test_cases"][0]["inputs"] == {"inA": 100}
    assert unit["test_cases"][0]["expected"] == {"outA": 200}
    assert unit["test_cases"][0]["description"] == "seq-1"   # Sequence-text 컬럼(11) 정상


def _make_hdpdm01_newline_headers_suts(path: Path) -> None:
    """generators/suts.py 가 실제로 쓰는 개행 삽입 헤더 재현('Test\\nMethod' 등). 탐지 정규화가
    개행을 접어야 test_method/gen_method 를 찾는다(못 찾으면 header_driven이 None으로 침묵 손실)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    ws.cell(row=5, column=14, value="Input")
    ws.cell(row=5, column=63, value="Expected Result")
    ws.cell(row=5, column=149, value="Related ID")
    ws.cell(row=6, column=3, value="TC ID")
    ws.cell(row=6, column=4, value="Name")
    ws.cell(row=6, column=5, value="Description")
    ws.cell(row=6, column=6, value="Safety\nRelated")
    ws.cell(row=6, column=7, value="Test\nEnvironment")
    ws.cell(row=6, column=8, value="Test\nMethod")           # 개행 삽입 (C1)
    ws.cell(row=6, column=9, value="Gen.\nMethod")            # 개행 + 약어 (C1)
    ws.cell(row=6, column=10, value="Precondition")
    ws.cell(row=6, column=11, value="Sequence")
    ws.cell(row=6, column=12, value="Test Case\nGen.Method")  # 개행 + 약어 (C1)
    ws.cell(row=6, column=13, value="Seq.\nNo.")
    ws.cell(row=7, column=3, value="SwUTC_SwUFn_0001")
    ws.cell(row=7, column=4, value="g_gen_func")
    ws.cell(row=7, column=8, value="FIT")   # test_method 값
    ws.cell(row=7, column=9, value="ABV")   # gen_method 값
    ws.cell(row=7, column=14, value="inA")
    ws.cell(row=7, column=63, value="outA")
    ws.cell(row=8, column=13, value=1)
    ws.cell(row=8, column=14, value=7)
    ws.cell(row=8, column=63, value=8)
    wb.save(path)
    wb.close()


def test_detect_columns_tolerates_newline_headers(tmp_path: Path) -> None:
    """C1(deep-review Critical): generators/suts.py 생성본 헤더는 개행 삽입이라 정규화가 개행을
    접지 않으면 test_method/gen_method 가 header_driven 문서에서 조용히 None→"" 로 손실됐다."""
    from openpyxl import load_workbook

    from tools.export_suts_vectorcast import _detect_columns

    p = tmp_path / "nl.xlsm"
    _make_hdpdm01_newline_headers_suts(p)
    ws = load_workbook(p, keep_vba=True)["2.SW Unit Test Spec"]
    cols = _detect_columns(ws)
    assert cols["test_method"] == 8       # 'Test\nMethod' → 개행 접힘 후 매칭
    assert cols["gen_method"] == 9        # 'Gen.\nMethod' → 약어 매칭
    assert cols["tc_gen_method"] == 12    # 'Test Case\nGen.Method'
    assert cols["description"] == 5

    model = build_vectorcast_model(str(p), project_id="NL")
    unit = model["units"][0]
    # provenance 보존(구 파서 동등) — 침묵 손실 아님
    assert unit["metadata"]["test_method"] == "FIT"
    assert unit["metadata"]["gen_method"] == "ABV"


def _make_partial_header_suts(path: Path) -> None:
    """Inpt[0]는 있는데 'Expected Result' 헤더가 없는 변형(미지 레이아웃) — 부분탐지 유발."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    ws.cell(row=4, column=3, value="TC_ID")
    ws.cell(row=4, column=4, value="Unit")
    ws.cell(row=4, column=9, value="Inpt[0]")   # 입력만 탐지, Expected 헤더 의도적 누락
    wb.save(path)
    wb.close()


def test_detect_columns_partial_detection_falls_back_and_warns(tmp_path: Path) -> None:
    """W1(deep-review): input만 탐지하고 expected 미탐지 시 밴드 혼합(역전 [63..29]·입력이 expected
    열 흡수)을 만들지 않고 전부 상수 폴백 + header_detect_fallback 경고로 표면화(X8)."""
    from openpyxl import load_workbook

    from tools.export_suts_vectorcast import _INPUT_COL_START, _SEQ_NO_COL, _detect_columns

    p = tmp_path / "partial.xlsm"
    _make_partial_header_suts(p)
    ws = load_workbook(p, keep_vba=True)["2.SW Unit Test Spec"]
    cols = _detect_columns(ws)
    # 부분탐지 → 밴드는 상수(탐지된 col9로 넘어가지 않음): 역전/혼합 밴드 방지
    assert cols["input_start"] == _INPUT_COL_START   # 14 (탐지된 9 아님)
    assert cols["seq_no"] == _SEQ_NO_COL             # 13 (8 아님)
    assert cols.get("_detect_warning")               # 침묵 아님
    model = build_vectorcast_model(str(p), project_id="P")
    assert any(w.get("code") == "header_detect_fallback" for w in model["export_warnings"])


def _make_signature_name_suts(path: Path) -> None:
    """HDPDM01 실 doc처럼 'Unit Name' 컬럼이 bare 함수명이 아니라 전체 C 시그니처
    ('void g_SysOs_WdiCtrl( void )')인 레이아웃(header_driven). 영향 target(=SVN diff의 bare
    함수명)과 매칭하려면 시그니처를 bare로 정규화해야 한다 — 안 하면 전 유닛이 침묵 필터링돼
    카드 '미파싱'·회귀 TC 0."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    ws.cell(row=5, column=14, value="Input")
    ws.cell(row=5, column=63, value="Expected Result")
    ws.cell(row=5, column=149, value="Related ID")
    ws.cell(row=6, column=3, value="TC ID")
    ws.cell(row=6, column=4, value="Name")
    ws.cell(row=6, column=13, value="Seq. No.")
    ws.cell(row=7, column=3, value="SwUTC_SwUFn_0001")
    ws.cell(row=7, column=4, value="void g_SysOs_WdiCtrl( void )")   # 전체 시그니처
    ws.cell(row=7, column=14, value="inA")
    ws.cell(row=7, column=63, value="outA")
    ws.cell(row=8, column=13, value=1)
    ws.cell(row=8, column=14, value=42)
    ws.cell(row=8, column=63, value=99)
    wb.save(path)
    wb.close()


def test_bare_fn_name_extracts_identifier() -> None:
    """시그니처↔bare 정규화: 반환타입·파라미터·포인터·한정자를 벗겨 식별자만. bare는 idempotent."""
    from tools.export_suts_vectorcast import bare_fn_name

    assert bare_fn_name("void g_SysOs_WdiCtrl( void )") == "g_SysOs_WdiCtrl"
    assert bare_fn_name("static void s_SystemOperation( void )") == "s_SystemOperation"
    assert bare_fn_name("uint8 * s_GetBuf( uint16 n )") == "s_GetBuf"    # 포인터 반환(분리)
    assert bare_fn_name("uint8* s_GetBuf2(void)") == "s_GetBuf2"          # 포인터 반환(붙음)
    assert bare_fn_name("s_sha256_update") == "s_sha256_update"           # bare → 무변경
    assert bare_fn_name("  g_Lib_Sha256_Nb_Start  ") == "g_Lib_Sha256_Nb_Start"
    assert bare_fn_name("") == ""
    assert bare_fn_name(None) == ""                                       # 방어적


def test_build_vectorcast_model_matches_signature_unit_name(tmp_path: Path) -> None:
    """HDPDM01 회귀: unit_name이 시그니처면 bare target(SVN diff)과 매칭돼 유닛이 포함돼야 한다.
    과거 naive .lower() 매칭이라 전량 필터링→회귀 TC/문서카드 침묵 0이 되던 근본 수정."""
    from tools.export_suts_vectorcast import build_vectorcast_model

    p = tmp_path / "sig.xlsm"
    _make_signature_name_suts(p)
    model = build_vectorcast_model(str(p), target_functions=["g_SysOs_WdiCtrl"])
    assert len(model["units"]) == 1                       # 매칭됨(과거 0)
    # 표시용 unit_name 원본(시그니처) 보존 — 빌더/VectorCAST 산출물 계약 불변
    assert model["units"][0]["unit_name"] == "void g_SysOs_WdiCtrl( void )"
    assert model["units"][0]["test_cases"][0]["inputs"] == {"inA": 42}
    # 대조군: 무관 bare target은 매칭 안 됨(bare 정규화가 과대매칭 아님)
    empty = build_vectorcast_model(str(p), target_functions=["s_unrelated"])
    assert len(empty["units"]) == 0
