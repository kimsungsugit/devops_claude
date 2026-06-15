"""Regression tests for /api/jenkins/sits/extract-traceability.

이 endpoint는 openpyxl `read_only=True`로 워크북을 연 뒤 시트를 스캔한다.
과거에는 `ws.cell(r, c)` 랜덤 접근을 썼는데, read_only 모드에서 random
cell access는 호출마다 시트 상단부터 재파싱돼 O(행²·열)로 폭주했다(1874행
실파일에서 ~75분 측정). iter_rows 순차 패스로 교체(O(행))한 뒤, 두 파싱
경로(Strategy 1: trace 시트 / Strategy 2: spec 시트)의 매핑 정확성을 고정한다.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from backend.routers.jenkins import jenkins_sits_extract_traceability


def _save(wb: openpyxl.Workbook, tmp_path: Path, name: str) -> str:
    fp = tmp_path / name
    wb.save(str(fp))
    return str(fp)


def test_sits_strategy1_trace_sheet(tmp_path: Path) -> None:
    """Strategy 1: trace 키워드 시트 — TC id(col 2/3) + req id(col 4+)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "4.SW Integration Test Spec"  # "test spec" 키워드 → Strategy 1
    # 헤더(1~3행)는 무시되고 본문은 4행부터.
    ws.cell(4, 2, "SwITC_01")              # col 2 = TC id (Sw\w+_\d+ 매칭)
    ws.cell(4, 4, "SwRS_001")              # col 4+ = related req ids
    ws.cell(4, 5, "관련: SyRS_002")
    ws.cell(5, 2, "SwITC_02")
    ws.cell(5, 6, "SwRS_001 SwRS_003")     # 중복 SwRS_001 + 신규
    # tc id가 col 2에 매칭 안 되면 col 3로 폴백
    ws.cell(6, 3, "SwITC_03")
    ws.cell(6, 4, "SwRS_004")

    res = jenkins_sits_extract_traceability({"path": _save(wb, tmp_path, "sits1.xlsx")})

    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWRS_001", "SwITC_01") in pairs
    assert ("SYRS_002", "SwITC_01") in pairs
    assert ("SWRS_001", "SwITC_02") in pairs
    assert ("SWRS_003", "SwITC_02") in pairs
    assert ("SWRS_004", "SwITC_03") in pairs
    # 모든 행이 SITS source
    assert {r["source"] for r in res["vcast_rows"]} == {"SITS"}


def test_sits_strategy1_empty_streak_break(tmp_path: Path) -> None:
    """빈 행 50개 연속이면 조기 종료 — 멀리 떨어진 잔여 데이터는 무시."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Traceability"
    ws.cell(4, 2, "SwITC_01")
    ws.cell(4, 4, "SwRS_001")
    # 4행 이후 60행 공백, 그 뒤 데이터 → break(50) 이후라 미수집
    ws.cell(70, 2, "SwITC_99")
    ws.cell(70, 4, "SwRS_999")

    res = jenkins_sits_extract_traceability({"path": _save(wb, tmp_path, "sits_gap.xlsx")})

    reqs = {r["requirement_id"] for r in res["vcast_rows"]}
    assert "SWRS_001" in reqs
    assert "SWRS_999" not in reqs  # empty_streak break로 도달 못 함


def test_sits_strategy2_spec_sheet(tmp_path: Path) -> None:
    """Strategy 2: trace 키워드 없는 Integration Test 시트 — Related 열 탐지."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "SW Integration Test"  # trace 키워드 미포함 → Strategy 2
    # 헤더 행 5/6에서 Related 열 탐지 (여기서는 10열)
    ws.cell(6, 10, "Related SwRS")
    # 본문은 7행부터: col 2 = TC id, col 3 = desc, related_col(10) = req ids
    ws.cell(7, 2, "SwITC_01")
    ws.cell(7, 3, "Verify integration: MainEntry -> SubA")
    ws.cell(7, 10, "SwRS_001, SwRS_002")
    ws.cell(8, 2, "SwITC_02")
    ws.cell(8, 3, "Verify integration: OtherFn -> SubB")
    ws.cell(8, 10, "SyRS_010")

    res = jenkins_sits_extract_traceability({"path": _save(wb, tmp_path, "sits2.xlsx")})

    assert res["ok"] is True
    by_req = {(r["requirement_id"], r["testcase"]): r for r in res["vcast_rows"]}
    assert ("SWRS_001", "SwITC_01") in by_req
    assert ("SWRS_002", "SwITC_01") in by_req
    assert ("SYRS_010", "SwITC_02") in by_req
    # description에서 entry function 추출
    assert by_req[("SWRS_001", "SwITC_01")]["unit"] == "MainEntry"


def test_sits_no_recognizable_sheet_returns_warning(tmp_path: Path) -> None:
    """trace/spec 시트 모두 없으면 available_sheets와 함께 경고 반환."""
    wb = openpyxl.Workbook()
    cover = wb.active
    assert cover is not None
    cover.title = "Cover"

    res = jenkins_sits_extract_traceability({"path": _save(wb, tmp_path, "sits_none.xlsx")})

    assert res["ok"] is True
    assert res["vcast_rows"] == []
    assert "available_sheets" in res


def test_sits_strategy2_default_related_col_145(tmp_path: Path) -> None:
    """Strategy 2: Related 헤더 미발견 → related_col=145 기본 폴백 (실파일 최빈 경로).

    W1 fix(max_column None → 199 스캔)와 body_max_col가 145까지 확장돼
    col 145의 req id를 실제로 읽는지 회귀 가드.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "SW Integration Test"  # trace 키워드 미포함 → Strategy 2
    # 행 5/6에 "Related"/"swds" 헤더 없음 → related_col 기본 145
    ws.cell(7, 2, "SwITC_05")
    ws.cell(7, 3, "Verify integration: EntryX -> Sub")
    ws.cell(7, 145, "SwRS_500")

    res = jenkins_sits_extract_traceability({"path": _save(wb, tmp_path, "sits_def145.xlsx")})

    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWRS_500", "SwITC_05") in pairs


def test_sits_strategy2_entry_fn_without_req_ids(tmp_path: Path) -> None:
    """Strategy 2: req id 없이 entry_fn만 있으면 requirement_id="" 행을 추가 (역매핑용)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "SW Integration Test"
    ws.cell(6, 10, "Related SwRS")  # related_col=10
    ws.cell(7, 2, "SwITC_06")
    ws.cell(7, 3, "Verify integration: LoneFn -> Sub")
    # col 10(related) 비움 → req_ids 없음, entry_fn=LoneFn만 존재
    res = jenkins_sits_extract_traceability({"path": _save(wb, tmp_path, "sits_lonefn.xlsx")})

    assert res["ok"] is True
    lone = [r for r in res["vcast_rows"] if r["testcase"] == "SwITC_06"]
    assert len(lone) == 1
    assert lone[0]["requirement_id"] == ""
    assert lone[0]["unit"] == "LoneFn"


def test_sits_explicit_sheet_name(tmp_path: Path) -> None:
    """sheet_name 인자로 자동탐색에 안 걸리는 시트를 명시 지정 → Strategy 1 처리."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "MyCustomSheet"  # 키워드 자동탐색 미매칭
    ws.cell(4, 2, "SwITC_07")
    ws.cell(4, 4, "SwRS_700")

    res = jenkins_sits_extract_traceability(
        {"path": _save(wb, tmp_path, "sits_named.xlsx"), "sheet_name": "MyCustomSheet"}
    )

    assert res["ok"] is True
    pairs = {(r["requirement_id"], r["testcase"]) for r in res["vcast_rows"]}
    assert ("SWRS_700", "SwITC_07") in pairs
