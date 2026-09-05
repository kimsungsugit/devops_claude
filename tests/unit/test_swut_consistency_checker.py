"""Tests for backend.services.swut_consistency_checker.

머지셀 패턴(컬럼 A None, 컬럼 B label)을 포함한 in-memory xlsx fixture를
openpyxl로 생성해서 추출 + 일관성 판정 4가지 케이스를 검증.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.services.excel_template_utils import BLANK_MARKUP  # noqa: E402
from backend.services.swut_consistency_checker import (  # noqa: E402
    _compact_row,
    _find_value_after_label,
    check_swut_consistency,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_coverage_wb(
    *,
    tcs: list[str],
    functions: list[str],
    coverage_map: dict[str, list[str]],  # tc_id -> [function_ids covered]
    exception_statement: int = 0,
    exception_branch: int = 0,
    final_result: str = "PASS",
) -> bytes:
    """가짜 Coverage Report 워크북 생성.

    Test Summary, 1.Traceability, 3.Coverage 시트 포함.
    Test Summary는 머지셀 시뮬레이션을 위해 컬럼 A에 빈 셀, 컬럼 B/C에 데이터.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Test Summary (머지셀 시뮬: column A 빈, B 라벨, C 값)
    ws = wb.create_sheet("Test Summary")
    ws.cell(row=1, column=2, value="Test Summary Report")
    ws.cell(row=8, column=2, value="Final Test Result")
    ws.cell(row=8, column=3, value=final_result)

    # 1.Traceability
    trace = wb.create_sheet("1.Traceability")
    # 헤더 행(row 1): 컬럼 A 빈, B 빈, C 'ID', D~ function ids
    trace.cell(row=1, column=3, value="ID")
    for i, fn in enumerate(functions):
        trace.cell(row=1, column=4 + i, value=fn)
    # 헤더에 function id 50개 미만이면 sw_count > 50 검사 통과 못함 → fixture는 충분히 많게.
    # 데이터 행
    for ri, tc in enumerate(tcs):
        trace.cell(row=2 + ri, column=2, value=tc)
        covered = coverage_map.get(tc, [])
        for i, fn in enumerate(functions):
            if fn in covered:
                trace.cell(row=2 + ri, column=4 + i, value="O")

    # 3. Coverage
    cov = wb.create_sheet("3. Coverage")
    cov.cell(row=1, column=2, value="Statement Coverage")
    cov.cell(row=3, column=2, value="Coverage")
    cov.cell(row=3, column=3, value="Total")
    cov.cell(row=3, column=4, value="Fail Count")
    cov.cell(row=3, column=5, value="Exception")
    cov.cell(row=4, column=2, value="Statement")
    cov.cell(row=4, column=3, value=len(functions))
    cov.cell(row=4, column=4, value=exception_statement)
    cov.cell(row=4, column=5, value=exception_statement)
    cov.cell(row=5, column=2, value="Branch")
    cov.cell(row=5, column=3, value=len(functions))
    cov.cell(row=5, column=4, value=exception_branch)
    cov.cell(row=5, column=5, value=exception_branch)

    return _wb_to_bytes(wb)


def _build_sutr_wb(
    *,
    total: int = 420,
    tested: int = 419,
    passed: int = 415,
    failed: int = 4,
    deviated: int = 4,
    not_executed: int = 1,
    not_executed_tcs: list[str] | None = None,
    deviation_tcs: list[str] | None = None,
    final_result: str = "OK",
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Test Summary")
    r = 1
    ws.cell(row=r, column=2, value="Test Summary Report"); r += 1
    ws.cell(row=r, column=2, value="Final Test Result")
    ws.cell(row=r, column=3, value=final_result); r += 1

    # Total Number of TCs 헤더 + 값
    ws.cell(row=r, column=2, value="Total Number of TCs")
    ws.cell(row=r, column=3, value="Number of TCs Tested")
    ws.cell(row=r, column=4, value="Number of TCs Passed")
    ws.cell(row=r, column=5, value="Number of TCs Failed")
    ws.cell(row=r, column=6, value="Number of Deviated TCs")
    ws.cell(row=r, column=7, value="Number of TCs not executed"); r += 1
    ws.cell(row=r, column=2, value=total)
    ws.cell(row=r, column=3, value=tested)
    ws.cell(row=r, column=4, value=passed)
    ws.cell(row=r, column=5, value=failed)
    ws.cell(row=r, column=6, value=deviated)
    ws.cell(row=r, column=7, value=not_executed); r += 1

    # Not Executed section
    ws.cell(row=r, column=2, value="■ List of Test Case not Executed"); r += 1
    ws.cell(row=r, column=2, value="Test Case ID"); r += 1
    for tc in (not_executed_tcs or []):
        ws.cell(row=r, column=2, value=tc); r += 1

    # Deviation section
    ws.cell(row=r, column=2, value="■ Deviation List"); r += 1
    ws.cell(row=r, column=2, value="Test Case ID"); r += 1
    for tc in (deviation_tcs or []):
        ws.cell(row=r, column=2, value=tc); r += 1

    return _wb_to_bytes(wb)


# ---------------------------------------------------------------------------
# Unit tests — helpers
# ---------------------------------------------------------------------------

class TestCompactRow:
    def test_filters_none_and_empty(self):
        row = (None, "  ", "label", None, "value", "")
        assert _compact_row(row) == [(2, "label"), (4, "value")]

    def test_preserves_column_index(self):
        row = (None, None, "X")
        assert _compact_row(row) == [(2, "X")]


class TestFindValueAfterLabel:
    def test_returns_next_pair_value(self):
        pairs = [(1, "Final Test Result"), (2, "PASS")]
        assert _find_value_after_label(pairs, "Final Test Result") == "PASS"

    def test_returns_none_if_label_last(self):
        pairs = [(1, "Final Test Result")]
        assert _find_value_after_label(pairs, "Final Test Result") is None

    def test_returns_none_if_label_absent(self):
        pairs = [(1, "Other"), (2, "Value")]
        assert _find_value_after_label(pairs, "Final Test Result") is None


# ---------------------------------------------------------------------------
# Integration tests — full consistency check
# ---------------------------------------------------------------------------

@pytest.fixture
def hdpdm_like_data():
    """HDPDM01 v3.01과 유사한 fixture (Total 420, NotExec 1, Deviation 4)."""
    # 51개 function이면 sw_count > 50 검사 통과
    functions = [f"SwUFn_{i:04d}" for i in range(1, 52)]
    tcs = [f"SwUTC_SwUFn_{i:04d}" for i in range(1, 51)]  # 50 TC (SwUFn_0051 미커버)
    coverage_map = {tc: [tc.replace("SwUTC_", "")] for tc in tcs}
    return {
        "functions": functions,
        "tcs": tcs,
        "coverage_map": coverage_map,
    }


class TestConsistencyAllOk:
    def test_no_critical_issues_when_aligned(self, hdpdm_like_data):
        cov = _build_coverage_wb(
            tcs=hdpdm_like_data["tcs"],
            functions=hdpdm_like_data["functions"],
            coverage_map=hdpdm_like_data["coverage_map"],
            exception_statement=13,
            exception_branch=17,
            final_result="PASS",
        )
        sutr = _build_sutr_wb(
            total=51, tested=50, passed=46, failed=4, deviated=4, not_executed=1,
            not_executed_tcs=["SwUTC_SwUFn_0051"],
            deviation_tcs=["SwUTC_SwUFn_0001 (TC2)", "SwUTC_SwUFn_0002 (TC1~5)",
                           "SwUTC_SwUFn_0003 (TC10)", "SwUTC_SwUFn_0004 (TC10)"],
            final_result="OK",
        )
        report = check_swut_consistency(cov, sutr)
        rd = report.to_dict()
        assert rd["ok"] is True
        # Total TC 차이(50 vs 51) — 미실행 1과 일치 → INFO
        info_total = [i for i in rd["issues"] if i["category"] == "total_tc"]
        assert info_total
        assert info_total[0]["severity"] == "info"
        # Final result terminology — PASS vs OK → INFO
        info_fr = [i for i in rd["issues"] if i["category"] == "final_result"]
        assert info_fr
        assert info_fr[0]["severity"] == "info"


class TestUncoveredMismatchWarning:
    def test_uncovered_function_not_in_sutr_not_executed(self, hdpdm_like_data):
        # SwUFn_0051은 커버 안 됨. SUTR not_executed 비어 있게 → mismatch warning 기대.
        cov = _build_coverage_wb(
            tcs=hdpdm_like_data["tcs"],
            functions=hdpdm_like_data["functions"],
            coverage_map=hdpdm_like_data["coverage_map"],
            final_result="PASS",
        )
        sutr = _build_sutr_wb(
            total=51, tested=51, passed=51, failed=0, deviated=0, not_executed=0,
            not_executed_tcs=[],  # mismatch!
            final_result="OK",
        )
        report = check_swut_consistency(cov, sutr)
        rd = report.to_dict()
        warns = [i for i in rd["issues"] if i["category"] == "uncovered_mismatch"]
        assert warns
        assert warns[0]["severity"] == "warning"


class TestExceptionDeviationMismatch:
    def test_deviation_exists_but_zero_exception(self, hdpdm_like_data):
        # Deviation 4건, Exception 0 → warning
        cov = _build_coverage_wb(
            tcs=hdpdm_like_data["tcs"],
            functions=hdpdm_like_data["functions"],
            coverage_map=hdpdm_like_data["coverage_map"],
            exception_statement=0,
            exception_branch=0,
            final_result="PASS",
        )
        sutr = _build_sutr_wb(
            total=51, tested=51, passed=47, failed=0, deviated=4, not_executed=0,
            not_executed_tcs=[],
            deviation_tcs=["SwUTC_SwUFn_0001", "SwUTC_SwUFn_0002",
                           "SwUTC_SwUFn_0003", "SwUTC_SwUFn_0004"],
            final_result="OK",
        )
        report = check_swut_consistency(cov, sutr)
        rd = report.to_dict()
        warns = [i for i in rd["issues"] if i["category"] == "exception_deviation"]
        # Coverage map에 SwUFn_0051이 커버 안 됐으므로 uncovered_mismatch warning이 함께 발생할 수 있음.
        # 단, exception_deviation 카테고리 warning은 반드시 1건 이상.
        assert warns
        assert warns[0]["severity"] == "warning"


class TestSmallModuleHeaderDetection:
    """Q1 regression: 51개 미만 function 모듈에서도 Traceability header 탐지."""

    def test_small_module_3_functions(self):
        # 함수 3개만 있는 소규모 모듈 (이전 sw_count>50 임계로 detection 실패하던 케이스)
        functions = ["SwUFn_0001", "SwUFn_0002", "SwUFn_0003"]
        tcs = ["SwUTC_SwUFn_0001", "SwUTC_SwUFn_0002"]  # SwUFn_0003 미커버
        coverage_map = {tc: [tc.replace("SwUTC_", "")] for tc in tcs}
        cov = _build_coverage_wb(
            tcs=tcs, functions=functions, coverage_map=coverage_map,
            final_result="PASS",
        )
        sutr = _build_sutr_wb(
            total=3, tested=2, passed=2, failed=0, deviated=0, not_executed=1,
            not_executed_tcs=["SwUTC_SwUFn_0003"],
            final_result="OK",
        )
        report = check_swut_consistency(cov, sutr)
        rd = report.to_dict()
        # 소규모 모듈에서도 미커버 정확 검출 → false negative 방지 (Critical fix)
        assert rd["coverage_summary"]["total_functions"] == 3
        assert rd["coverage_summary"]["total_tcs"] == 2
        assert rd["coverage_summary"]["uncovered_functions"] == ["SwUFn_0003"]


class TestParseWarnings:
    """deep-reviewer 시나리오 3 — 필수 시트 미발견 시 silent empty 막기."""

    def test_empty_workbook_emits_warnings(self):
        empty = openpyxl.Workbook()
        empty.remove(empty.active)
        empty.create_sheet("RandomSheet")
        buf = io.BytesIO()
        empty.save(buf)
        cov_bytes = buf.getvalue()

        empty2 = openpyxl.Workbook()
        empty2.remove(empty2.active)
        empty2.create_sheet("RandomSheet")
        buf2 = io.BytesIO()
        empty2.save(buf2)
        sutr_bytes = buf2.getvalue()

        report = check_swut_consistency(cov_bytes, sutr_bytes)
        rd = report.to_dict()
        # 시트 미발견 → parse_warnings 다수, ok=False
        assert rd["parse_warnings"]
        assert rd["ok"] is False
        assert any("Traceability" in w for w in rd["parse_warnings"])

    def test_tool_qualification_metadata_in_output(self, hdpdm_like_data):
        cov = _build_coverage_wb(
            tcs=hdpdm_like_data["tcs"],
            functions=hdpdm_like_data["functions"],
            coverage_map=hdpdm_like_data["coverage_map"],
            final_result="PASS",
        )
        sutr = _build_sutr_wb(final_result="OK")
        report = check_swut_consistency(cov, sutr)
        rd = report.to_dict()
        assert "tool_qualification" in rd
        assert rd["tool_qualification"]["evidence_class"] == "auto-generated draft"
        assert "단독 evidence" in rd["tool_qualification"]["asil_b_c_d_usage"]


class TestBlankPlaceholderDetection:
    """시나리오 A (deep-reviewer): 빌더 출력의 placeholder 시트를 self-validation에서 인식."""

    def test_trace_sheet_with_blank_markup_is_skipped(self):
        # placeholder Traceability 시트 (BLANK_MARKUP) + 정상 Test Summary
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ts = wb.create_sheet("Test Summary")
        ts.cell(row=1, column=2, value="Final Test Result")
        ts.cell(row=1, column=3, value="PASS")
        trace = wb.create_sheet("1.Traceability")
        trace.cell(row=1, column=1, value=BLANK_MARKUP)
        cov_bytes = _wb_to_bytes(wb)
        sutr_bytes = _build_sutr_wb(final_result="OK")
        report = check_swut_consistency(cov_bytes, sutr_bytes)
        rd = report.to_dict()
        # parse_warnings에 명시 + cov_summary placeholder flag
        assert any("placeholder" in w for w in rd["parse_warnings"])
        assert rd["coverage_summary"]["trace_sheet_is_placeholder"] is True
        assert rd["coverage_summary"]["total_functions"] == 0
        assert rd["ok"] is False  # parse_warnings 존재로 ok=False


class TestFinalResultMismatch:
    def test_final_result_semantically_different(self, hdpdm_like_data):
        # PASS vs FAIL → warning
        cov = _build_coverage_wb(
            tcs=hdpdm_like_data["tcs"][:50],  # all covered
            functions=hdpdm_like_data["functions"][:50],
            coverage_map={tc: [tc.replace("SwUTC_", "")] for tc in hdpdm_like_data["tcs"][:50]},
            final_result="PASS",
        )
        sutr = _build_sutr_wb(
            total=50, tested=50, passed=50, failed=0, deviated=0, not_executed=0,
            not_executed_tcs=[],
            final_result="FAIL",
        )
        report = check_swut_consistency(cov, sutr)
        rd = report.to_dict()
        warns = [i for i in rd["issues"] if i["category"] == "final_result"]
        assert warns
        assert warns[0]["severity"] == "warning"
