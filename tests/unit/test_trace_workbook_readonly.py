"""STS/SUTS 추적성 로더가 read_only + 값 격자로 도는가 (2026-08-06).

## 사용자 보고

"다른 데는 변경되는데 **SUTS만** 안 바뀐다."

## 실체 — 반영이 안 된 게 아니라 아직 안 온 것이었다

`_load_trace_workbook`(STS/SUTS 공유)만 `read_only` 없이 워크북을 열었다. 정상 모드는
시트의 셀 전부를 `Cell` 객체로 물화한다. 실측(KJPDS02_PV, 2026-08-06):

    SUTS 4.75MB : 로드 28.5초  (read_only 0.50초 — **56배**)
    SITS 1.02MB : 로드  5.6초  ← 이 로더는 처음부터 read_only 였다
    STS  0.38MB : 로드  1.2초

엔드포인트 왕복은 SUTS **36.2초**. 그래서 사용자가 경로를 바꾸고 화면을 보면 SITS 는
몇 초 만에 새 값으로 갱신되고 SUTS 만 한참 옛 상태로 남는다 — "SUTS만 안 바뀐다".

같은 파일의 SITS(:4815)·SyTS(:5094) 로더는 read_only 였고 이 로더만 빠져 있었다.
전형적인 판정 복제다.

## ⚠ 플래그만 뒤집으면 되레 느려진다

이 파서는 `ws.cell(r, c)` 랜덤 접근을 13곳에서 한다. openpyxl read_only 워크시트에서
그 접근은 호출마다 재스캔이라 **O(행²)** 로 폭주한다 — 이 저장소엔 같은 형태로 SITS
파서가 **75분** 걸린 실측 전례가 있다. 그래서 `read_only=True` 와 `_GridSheet`(한 번의
`iter_rows` 로 값 격자를 뜬다)는 **한 세트**이고, 아래 테스트가 둘을 함께 고정한다.

## 결과 불변

실파일로 리팩터 전/후를 대조했다: SUTS `total_mappings`·`requirements_covered`·행 수·
첫 3행·마지막 행 **완전 일치**(1014행), STS 동일(102행). 속도만 36.2초 → 13.1초.
"""
from __future__ import annotations

import ast
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parents[2]
_SRC = REPO / "backend" / "routers" / "jenkins.py"

pytest.importorskip("openpyxl")
import openpyxl  # noqa: E402

from backend.routers.jenkins import _GridSheet  # noqa: E402


# ---------------------------------------------------------------------------
# _GridSheet — openpyxl 워크시트와 **같은 답**을 줘야 한다
# ---------------------------------------------------------------------------
def _sample_ws():
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, val in enumerate(("Test Case ID", "Related ID", "Name"), start=1):
        ws.cell(1, col, val)
    for col, val in enumerate(("TC_001", "SwFn_01", "Foo"), start=1):
        ws.cell(2, col, val)
    for col, val in enumerate(("TC_002", "SwFn_02"), start=1):   # 3열 비움 — 짧은 행
        ws.cell(3, col, val)
    return wb, ws


def test_grid_matches_worksheet_values():
    wb, ws = _sample_ws()
    g = _GridSheet(ws)
    for r in range(1, 4):
        for c in range(1, 4):
            assert g.cell(r, c).value == ws.cell(r, c).value, f"({r},{c}) 불일치"
    wb.close()


def test_grid_out_of_range_is_none_like_openpyxl():
    """범위 밖 접근은 openpyxl 과 같이 `value=None` — 예외를 던지면 파서가 죽는다."""
    wb, ws = _sample_ws()
    g = _GridSheet(ws)
    for r, c in ((0, 1), (1, 0), (999, 1), (1, 999), (-1, -1)):
        assert g.cell(r, c).value is None, f"({r},{c}) 에서 None 이 아니다"
    wb.close()


def test_grid_short_rows_do_not_raise():
    """행마다 길이가 다른 시트(끝 열이 빈 행)에서도 안전해야 한다."""
    wb, ws = _sample_ws()
    g = _GridSheet(ws)
    assert g.cell(3, 3).value is None
    wb.close()


def test_grid_reports_dimensions():
    wb, ws = _sample_ws()
    g = _GridSheet(ws)
    assert g.max_row == 3
    assert g.max_column == 3
    wb.close()


def test_grid_caps_columns():
    """비정상적으로 넓은 시트에서 메모리가 터지지 않게 열을 자른다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 50, "x")
    g = _GridSheet(ws, max_cols=10)
    assert g.max_column <= 10
    assert g.cell(1, 50).value is None
    wb.close()


def test_grid_reads_the_sheet_only_once():
    """`iter_rows` 를 두 번 이상 돌면 read_only 에서 이득이 사라진다."""
    wb, ws = _sample_ws()
    calls = {"n": 0}
    real = ws.iter_rows

    def counting(*a, **k):
        calls["n"] += 1
        return real(*a, **k)

    ws.iter_rows = counting  # type: ignore[method-assign]
    _GridSheet(ws)
    assert calls["n"] == 1, f"시트를 {calls['n']}번 순회했다 — 한 번이어야 한다"
    wb.close()


# ---------------------------------------------------------------------------
# 구조 — read_only 와 _GridSheet 는 한 세트다
# ---------------------------------------------------------------------------
def _src() -> str:
    return _SRC.read_text(encoding="utf-8")


def test_shared_trace_loader_uses_read_only():
    """`_load_trace_workbook`(STS/SUTS 공유)이 read_only 로 연다."""
    tree = ast.parse(_src())
    fn = next(n for n in ast.walk(tree)
              if isinstance(n, ast.FunctionDef) and n.name == "_load_trace_workbook")
    loads = [n for n in ast.walk(fn) if isinstance(n, ast.Call)
             and getattr(n.func, "attr", "") == "load_workbook"]
    assert loads, "_load_trace_workbook 에 load_workbook 호출이 없다"
    for call in loads:
        kw = {k.arg: k for k in call.keywords}
        assert "read_only" in kw, (
            f"line {call.lineno}: read_only 가 빠졌다 — SUTS 로드가 28초가 된다"
        )
        assert getattr(kw["read_only"].value, "value", None) is True


def test_extractor_wraps_sheet_in_grid():
    """read_only 로 열었으면 랜덤 접근 전에 반드시 격자로 감싸야 한다.

    이 단언이 깨지면 O(행²) 폭주다(SITS 75분 전례) — read_only 보다 **더 나쁜** 상태다.
    """
    tree = ast.parse(_src())
    fn = next(n for n in ast.walk(tree)
              if isinstance(n, ast.FunctionDef) and n.name == "_extract_test_spec_traceability")
    wrapped = [n for n in ast.walk(fn) if isinstance(n, ast.Call)
               and getattr(n.func, "id", "") == "_GridSheet"]
    assert wrapped, (
        "_extract_test_spec_traceability 가 시트를 _GridSheet 로 감싸지 않는다 — "
        "read_only 워크시트에 랜덤 .cell() 을 하면 O(행²)로 폭주한다"
    )


def test_all_workbook_loaders_in_this_router_are_read_only():
    """이 라우터의 워크북 로더가 **전부** read_only 인지 — 하나만 빠지는 게 이 결함이었다."""
    tree = ast.parse(_src())
    offenders: list[str] = []
    for node in ast.walk(tree):
        if not (isinstance(node, ast.Call) and getattr(node.func, "attr", "") == "load_workbook"):
            continue
        kw = {k.arg for k in node.keywords}
        if "read_only" not in kw:
            offenders.append(f"jenkins.py:{node.lineno}")
    assert not offenders, (
        "read_only 없이 워크북을 여는 자리가 남아 있다(대용량 문서에서 수십 초): "
        + ", ".join(offenders)
    )
