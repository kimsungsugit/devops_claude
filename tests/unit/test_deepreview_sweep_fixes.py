"""2026-06-19 deep-review 배포후 sweep 확정결함(B control-char / C ci-collision) 회귀 가드.

sweep가 찾은 whack-a-mole sibling:
- B (control-char-crash): XlsxManager.write_data(vcast/qac 공유 sink) + qac add_summary_charts
  직접 대입이 C1 hardening을 우회 → IllegalCharacterError로 generate 전체 크래시.
- C (ci-collision): swut_comprehensive_aggregator._lookup_c_function /
  swit_coverage_aggregator norm_name_to_swufn의 last-wins 소문자 인덱스가 대소문자
  충돌쌍을 오매칭(틀린 C evidence / 틀린 SwUFn ID).
"""
from __future__ import annotations

import io

import openpyxl

from backend.services.swut_comprehensive_aggregator import _lookup_c_function
from backend.services.vcast_excel_generator import XlsxManager


class TestSweepControlCharCrash:
    """B — XlsxManager.write_data 불법 제어문자 sanitize (vcast+qac 공유 sink)."""

    def _mgr(self) -> XlsxManager:
        m = XlsxManager()
        m.workbook = openpyxl.Workbook()
        m.worksheet = m.workbook.active
        return m

    def test_write_data_strips_illegal_chars_no_crash(self):
        m = self._mgr()
        # form-feed(\x0c) + bell(\x07) + EOF(\x1a) — openpyxl이 대입 시 raise하던 문자
        m.write_data(1, 1, "byte x = 0;\x0c\x07 /* dump\x1a */")
        assert m.worksheet.cell(1, 1).value == "byte x = 0; /* dump */"

    def test_write_data_preserves_legal_whitespace(self):
        m = self._mgr()
        m.write_data(1, 1, "a\tb\nc")  # tab/LF는 xlsx 합법
        assert m.worksheet.cell(1, 1).value == "a\tb\nc"

    def test_write_data_numeric_and_bool_unchanged(self):
        m = self._mgr()
        m.write_data(1, 1, 42)
        m.write_data(2, 1, 3.14)
        m.write_data(3, 1, True)  # bool은 int 분기보다 먼저 처리(이전 dead code 수정)
        assert m.worksheet.cell(1, 1).value == 42
        assert m.worksheet.cell(2, 1).value == 3.14
        assert m.worksheet.cell(3, 1).value is True

    def test_workbook_saves_after_illegal_input(self):
        """sanitize 없으면 save 시점이 아니라 대입 시점에 크래시했음 — 라운드트립 확인."""
        m = self._mgr()
        m.write_data(1, 1, "ctrl\x0cchars\x07here")
        bio = io.BytesIO()
        m.workbook.save(bio)  # 예외 없이 저장되어야 함
        assert len(bio.getvalue()) > 0


class TestSweepCiCollision:
    """C — _lookup_c_function 대소문자 충돌쌍 모호 키 제외(틀린 C evidence 차단)."""

    def test_exact_match_wins(self):
        cmap = {"Foo": {"name": "Foo", "body": "A"}, "foo": {"name": "foo", "body": "B"}}
        # exact 매칭은 충돌과 무관하게 정확 반환
        assert _lookup_c_function("Foo", "", cmap)["body"] == "A"
        assert _lookup_c_function("foo", "", cmap)["body"] == "B"

    def test_ambiguous_ci_returns_none_not_wrong_evidence(self):
        # 'Foo'/'foo' 충돌 → 'FOO'(둘 다 아닌 제3 표기)는 CI 폴백 시 모호 → None.
        # (이전 last-wins 구현은 둘 중 하나의 body를 silent 첨부 = 거짓 evidence)
        cmap = {"Foo": {"name": "Foo", "body": "A"}, "foo": {"name": "foo", "body": "B"}}
        assert _lookup_c_function("FOO", "", cmap) is None

    def test_non_ambiguous_ci_fallback_still_works(self):
        # 충돌 없는 단일 키는 CI 폴백 정상 동작(casing만 다른 정상 매칭).
        cmap = {"Calc": {"name": "Calc", "body": "X"}}
        assert _lookup_c_function("calc", "", cmap)["body"] == "X"

    def test_same_value_under_two_keys_not_ambiguous(self):
        # 같은 dict 객체가 두 키에 있으면 모호 아님(동일 evidence) — 폴백 허용.
        shared = {"name": "Bar", "body": "Z"}
        cmap = {"Bar": shared, "bar": shared}
        assert _lookup_c_function("BAR", "", cmap)["body"] == "Z"

    def test_empty_map_returns_none(self):
        assert _lookup_c_function("anything", "x", None) is None
        assert _lookup_c_function("anything", "x", {}) is None
