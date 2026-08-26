from __future__ import annotations

import io

import openpyxl
import pytest

from backend.services.swit_comprehensive_aggregator import (
    SwitcrBuildMeta,
    build_switcr_report,
)
from backend.services.swut_input_adapter import (
    CoverageStats,
    EnvironmentData,
    ExecutionRow,
    FunctionCoverage,
    SwUTSession,
)


def _minimal_switcr_template() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("History")
    wb.create_sheet("Guideline")
    wb.create_sheet("Summary")
    for name in (
        "1.IT101",
        "2.IT201",
        "3.IT301",
        "4.IT401",
        "5.IT501",
        "6.IT601",
        "7.IT701",
        "8.IT801",
    ):
        ws = wb.create_sheet(name)
        ws.cell(1, 2).value = name
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


#: 회사 SwITCV 정본은 두 판이 유통되고 **Component 열 하나 차이로 오른쪽 전부가 밀린다.**
#:   DV(11열): No=B, Component=C, Unit ID=D, Name=E, Functions=F …
#:   PV(10열): No=B,              Unit ID=C, Name=D, Functions=E …
_SWITCV_LAYOUTS = ("DV", "PV")


def _write_switcv_sheet(
    ws,
    layout: str,
    rows: list[tuple],
    summary: tuple[tuple, tuple] | None,
) -> None:
    """SwITCV `4.Coverage` 시트를 **헤더까지** 정본 형태로 쓴다.

    ⚠ 헤더를 빠뜨리지 말 것. 데이터만 두면 `coverage_column_base` 가 'Component'
      라벨을 못 찾아 무조건 PV 로 접히고, 그러면 **픽스처가 코드의 거울**이 된다 —
      2026-08-26 이전 픽스처가 정확히 그 상태(헤더 없이 DV 열에 기입)여서
      `_load_workbook_summary` 의 DV 고정 결함을 한 번도 못 잡았다.

    rows: (unit_id, name, func_result, func_exc, ccount, ctotal, call_result, call_exc, note)
    summary: ((f_total, f_fail, f_exc), (c_total, c_fail, c_exc)) — None 이면 수식(빈 셀) 재현
    """
    base = 4 if layout == "DV" else 3          # Unit ID 열
    if layout == "DV":
        ws.cell(9, 3).value = "Component"      # 이 라벨 하나가 판을 가른다
    ws.cell(9, 2).value = "No"
    ws.cell(9, base).value = "Unit"
    ws.cell(10, base).value = "ID"
    ws.cell(10, base + 1).value = "Name"
    ws.cell(9, base + 2).value = "Functions"
    ws.cell(9, base + 3).value = "Exception"
    ws.cell(9, base + 4).value = "Function Called"
    ws.cell(10, base + 4).value = "Count"
    ws.cell(10, base + 5).value = "Total"
    ws.cell(10, base + 6).value = "Pass"
    ws.cell(9, base + 7).value = "Exception"
    ws.cell(9, base + 8).value = "File"

    # 요약 블록 — 헤더 r4 `Total | Fail Count | Exception | Coverage`, 값 r5/r6.
    total_col = base + 1
    ws.cell(4, 2).value = "Coverage"
    for offset, label in enumerate(("Total", "Fail Count", "Exception", "Coverage")):
        ws.cell(4, total_col + offset).value = label
    ws.cell(5, 2).value = "Functions"
    ws.cell(6, 2).value = "Function Calls"
    if summary is not None:
        (f_total, f_fail, f_exc), (c_total, c_fail, c_exc) = summary
        for row, trio in ((5, (f_total, f_fail, f_exc)), (6, (c_total, c_fail, c_exc))):
            for offset, value in enumerate(trio):
                ws.cell(row, total_col + offset).value = value

    for i, row in enumerate(rows):
        r = 11 + i
        ws.cell(r, 2).value = i + 1
        for offset, value in enumerate(row):
            if value is not None:
                ws.cell(r, base + offset).value = value


def _switcv_workbook(layout: str = "DV") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "4.Coverage"
    _write_switcv_sheet(
        ws, layout,
        [("SwUFn_0101", "FunctionA", "X", "O", None, None, "O", None, "SwITCR 참고")],
        ((570, 1, 1), (570, 0, 0)),
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _switr_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2.Test Log"
    ws.cell(4, 6).value = "SwITC_0101_01"
    ws.cell(4, 20).value = "Pass"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fault_injection_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FI_Test Case"
    ws.cell(9, 2).value = 1
    ws.cell(9, 3).value = "SwIFITC_01"
    ws.cell(9, 15).value = "ExpectedA"
    ws.cell(9, 17).value = "ActualA"
    ws.cell(10, 4).value = 1
    ws.cell(10, 5).value = "FI"
    ws.cell(10, 6).value = "AOR"
    ws.cell(10, 15).value = "0"
    ws.cell(10, 16).value = "1"
    ws.cell(10, 17).value = "0"
    ws.cell(10, 18).value = "1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _session() -> SwUTSession:
    env = EnvironmentData(
        env_name="SWIT_SWUFN_0101_DEPTH4_FILE12",
        component_name="CompA",
        test_cases={"SwITC_0101_01": [object()]},
        test_results={
            "SwITC_0101_01": ExecutionRow(
                tc_name="SwITC_0101_01",
                passed=True,
                actual_result={"Input[0]": ("0x01", "0x01")},
            ),
        },
        function_coverage=[
            FunctionCoverage(
                unit_id="SwUFn_0101",
                name="FunctionA",
                statement=CoverageStats(covered=9, total=10),
                branch=CoverageStats(covered=5, total=5),
            ),
        ],
    )
    return SwUTSession(environments=[env])


def test_build_switcr_preserves_template_and_writes_active_sheets():
    result = build_switcr_report(
        _session(),
        SwitcrBuildMeta(
            project_id="KJPDS02",
            project_full_name="KJPDS02",
            release_sw_version="1.01",
            test_date="2025-12-05",
            test_engineer="JK Kim",
            doc_filename_pattern=(
                "(KJPDS02_DV_SwITCR) Software Integration Test Comprehensive Result_"
                "v{version}_{date}_R.xlsm"
            ),
            project_config={
                "switcr_metadata": {
                    "software_platform_ver": "25A1",
                    "tester": "JK Kim",
                    "qualified_function_total": 570,
                }
            },
        ),
        _minimal_switcr_template(),
        swits_map={"SwITC_0101_01": {"tc_id": "SwITC_0101_01"}},
        switcv_bytes=_switcv_workbook(),
        switr_bytes=_switr_workbook(),
        fault_injection_bytes=_fault_injection_workbook(),
    )

    assert result.ok is True
    assert result.filename == (
        "(KJPDS02_DV_SwITCR) Software Integration Test Comprehensive Result_v1.01_251205_R.xlsm"
    )
    assert result.summary["switcr_function_count"] == 570
    assert result.summary["swits_entries"] == 1
    assert result.result_size_bytes > 0

    wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
    assert "AuditLog" not in wb.sheetnames
    assert wb.sheetnames == [
        "Cover", "History", "Guideline", "Summary",
        "1.IT101", "2.IT201", "3.IT301", "4.IT401",
        "7.IT701", "(해당X)5.IT501", "(해당X)6.IT601", "(해당X)8.IT801",
    ]
    assert "(해당X)5.IT501" in wb.sheetnames
    assert "(해당X)6.IT601" in wb.sheetnames
    assert "(해당X)8.IT801" in wb.sheetnames
    assert wb["1.IT101"]["C5"].value == "25A1"
    assert wb["1.IT101"]["E75"].value == 570
    assert wb["1.IT101"]["C83"].value == "함수커버리지"
    assert wb["1.IT101"]["E83"].value == "SwUFn_0101"
    # IT201: F70(전체 인터페이스=분모)은 agg 실측이라 값 유지. G70(통과 수)은
    # interface_passed 설정이 없어 노란 사용자입력 마킹(A2 — passed=total 100% 위장 제거).
    assert wb["2.IT201"]["F70"].value == 1
    assert str(wb["2.IT201"]["G70"].value or "").startswith("▶ 사용자 입력 필요")
    # IT301: _fault_injection_workbook() 이 FI 실측을 제공 → 증거 경로(값 기입 유지).
    assert wb["3.IT301"]["C85"].value == 1
    assert wb["3.IT301"]["E85"].value == 1
    assert wb["3.IT301"]["F85"].value == 1
    assert wb["3.IT301"]["C102"].value == "식별된 결함"
    assert wb["3.IT301"]["C103"].value == "해당사항 없음"
    assert wb["Summary"]["G20"].value == "X"


class TestSwitcrCoverAndSummary96Final:
    """라운드 96-final QA fix — SwITCR Cover stamp (이전 완전 미스탬프 Critical)
    + Summary O열 시트명 rename 동기 (INDIRECT #REF! 차단)."""

    @staticmethod
    def _meta():
        from backend.services.swit_comprehensive_aggregator import SwitcrBuildMeta
        return SwitcrBuildMeta(
            project_id="KJPDS02", release_sw_version="0.10",
            test_date="2026-06-04", test_engineer="주희영",
            default_author="JK Kim", default_approver="CH In",
            doc_filename_pattern="(KJPDS02_PV_SwITCR) X_v{version}_{date}_R.xlsm",
        )

    def test_switcr_cover_stamps_xxxx_placeholders(self):
        import openpyxl

        from backend.services.swit_comprehensive_aggregator import _write_switcr_cover
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cover"
        ws["I2"] = "Author"
        ws["J2"] = "Reviewer"
        ws["K2"] = "Approver"
        ws["C26"] = "Document ID"
        ws["D26"] = "HKY-[P_Name]-SwITCR-28A1"
        ws["C27"] = "Version"
        ws["D27"] = "v0.10"
        ws["C28"] = "Status"
        ws["D28"] = "Unspecified"
        ws["C29"] = "Date"
        ws["D29"] = "202X.XX.XX"
        ws["C30"] = "Author"
        ws["D30"] = "XXXX"
        warns: list[str] = []
        _write_switcr_cover(ws, self._meta(), {}, out_warnings=warns)
        assert ws["D26"].value == "HKY-KJPDS02_PV-SwITCR-28A1"
        assert ws["D28"].value == "DRAFT — PENDING REVIEW"
        assert ws["D29"].value == "2026.06.04"
        assert ws["D30"].value == "주희영"  # test_engineer 우선
        assert ws["J2"].value == "Reviewer"  # 라벨 보존
        assert ws["I3"].value == "주희영"
        assert ws["K3"].value == "CH In"

    def test_summary_o_col_synced_to_renamed_sheets(self):
        import openpyxl

        from backend.services.swit_comprehensive_aggregator import _write_summary_sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        _write_summary_sheet(ws, self._meta(), {})
        assert ws.cell(20, 15).value == "(해당X)5.IT501"
        assert ws.cell(21, 15).value == "(해당X)6.IT601"
        assert ws.cell(23, 15).value == "(해당X)8.IT801"
        # IT802는 8.IT801 시트 내 섹션 — 존재하지 않는 '8.IT802' stamp 금지
        assert ws.cell(24, 15).value == "(해당X)8.IT801"


class TestSwitcrCoverDeepReviewerW1W2:
    """deep-reviewer 96-final W1/W2 — 비-trio fallback + O열 동적 참조."""

    def test_non_trio_template_fallback_writes_reviewer_approver(self):
        import openpyxl

        from backend.services.swit_comprehensive_aggregator import (
            SwitcrBuildMeta,
            _write_switcr_cover,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cover"
        # 세로 단독 kv 라벨 (가로 trio 없음)
        ws["B5"] = "Reviewer"
        ws["B7"] = "Approver"
        meta = SwitcrBuildMeta(
            project_id="KJPDS02", release_sw_version="0.10",
            test_date="2026-06-04", default_approver="CH In",
        )
        _write_switcr_cover(ws, meta, {}, out_warnings=[])
        # fallback 경로 — Approver 기입 + Reviewer 빈 값 노란
        assert ws["C7"].value == "CH In"
        assert str(ws["C5"].value or "").startswith("▶")

    def test_summary_o_col_resolves_actual_names(self):
        import openpyxl

        from backend.services.swit_comprehensive_aggregator import (
            SwitcrBuildMeta,
            _write_summary_sheet,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        # rename이 일부만 적용된 변형: 5.IT501은 rename됨, 6.IT601은 원명 유지
        wb.create_sheet("(해당X)5.IT501")
        wb.create_sheet("6.IT601")
        meta = SwitcrBuildMeta(
            project_id="KJPDS02", release_sw_version="0.10", test_date="2026-06-04",
        )
        _write_summary_sheet(ws, meta, {})
        assert ws.cell(20, 15).value == "(해당X)5.IT501"
        assert ws.cell(21, 15).value == "6.IT601"  # 원명 유지분은 원명 참조 (desync 차단)


def _switcv_workbook_formula_summary(layout: str = "DV") -> bytes:
    """요약셀을 비워 수식→data_only None 을 모사. 실제 SwITCV는 이 셀들이 회사 템플릿
    수식이라 openpyxl이 cached=None 으로 저장한다. 데이터행은 SwITCV function/call
    경로와 동일하게 O/X·count·exception 모두 리터럴."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "4.Coverage"
    _write_switcv_sheet(
        ws, layout,
        [
            ("SwUFn_0101", "FuncA", "X", "O", 2, 3, "X", "O", None),
            ("SwUFn_0102", "FuncB", "O", None, 3, 3, "O", None, None),
            ("SwUFn_0103", "FuncC", "O", None, None, None, "O", None, None),
        ],
        None,          # 요약셀 미기입 → 수식 캐시 부재 재현
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.parametrize("layout", _SWITCV_LAYOUTS)
def test_load_workbook_summary_computes_from_rows_when_summary_blank(layout):
    """후속 결함 회귀 — 요약셀이 수식(data_only None)이어도 roll-up이 데이터행에서
    함수/호출 fail·exception 을 정확 집계 (거짓 100% 커버리지 차단). 이전엔 None→0
    으로 fail/exception 이 0 처리되어 coverage_fail_details(정상)와 모순이었다.

    2026-08-26 — DV/PV **두 판 모두** 돈다. 예전엔 DV 열에만 돌아서, PV 정본에서
    전 열이 한 칸씩 밀리던 결함을 이 테스트가 통과시켜 줬다."""
    from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
    out = _load_workbook_summary(_switcv_workbook_formula_summary(layout))
    assert out["functions_total"] == 3
    assert out["functions_fail_count"] == 1
    assert out["functions_exception_count"] == 1
    assert out["function_calls_total"] == 3
    assert out["function_calls_fail_count"] == 1
    assert out["function_calls_exception_count"] == 1
    assert len(out["coverage_fail_details"]) == 2  # func X + call X (동일 행)


@pytest.mark.parametrize("layout", _SWITCV_LAYOUTS)
def test_load_workbook_summary_prefers_literal_summary_cells(layout):
    """셀에 리터럴 캐시값이 있으면(외부 재계산본) 데이터행 집계보다 우선."""
    from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
    out = _load_workbook_summary(_switcv_workbook(layout))  # Total=570/Fail=1/Exc=1
    assert out["functions_total"] == 570
    assert out["functions_fail_count"] == 1
    assert out["functions_exception_count"] == 1


@pytest.mark.parametrize("layout", _SWITCV_LAYOUTS)
def test_data_rows_are_read_at_the_right_columns(layout):
    """요약셀이 리터럴이어도 **데이터 행은 따로** 읽는다 — 두 경로가 각각 밀릴 수 있다.

    ⚠ 이 단언이 없으면 요약 라벨 탐지가 정답을 내는 덕에 열 판정이 무력화돼도
      위 테스트가 통과한다(뮤테이션 L6 생존으로 실증). 요약과 데이터는 **다른 축**이다.
    """
    from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
    details = _load_workbook_summary(_switcv_workbook(layout))["coverage_fail_details"]
    fn = next(d for d in details if d["kind"] == "함수커버리지")
    assert fn["unit_id"] == "SwUFn_0101"      # 밀리면 함수명("FunctionA")이 들어온다
    assert fn["function"] == "FunctionA"      # 밀리면 O/X 값("X")이 들어온다
    assert fn["exception"] == "O"
    assert fn["note"] == "SwITCR 참고"        # 마지막 열(File) — 오른쪽 끝까지 정합


class TestSwitcvPvLayoutIsNotReadOneColumnOff:
    """KJPDS02 PV 정본(10열) 회귀 가드 — 2026-08-26 실측 결함.

    `_load_workbook_summary` 가 DV(11열)에 고정돼 있어 PV SwITCV 에서 **전 열이 한 칸씩
    밀렸다.** 증상이 조용했던 이유는 밀린 자리의 값이 우연히 같았기 때문이다:
    `Fail Count` 와 `Exception` 이 둘 다 4 라 "fail 은 맞네" 로 보였고, 정작
    `Total`(1014)이 `Fail Count`(4)로 읽혀 **253배 과소** 보고됐다.

    그래서 이 가드는 **Total / Fail Count / Exception 을 전부 다른 값**으로 쓴다.
    세 값이 같으면 한 칸 밀려도 통과하므로 결함을 못 잡는다
    (guard must change the observable).
    """

    #: 정본 실측 배치(`(KJPDS02_SwITCV) … v2.01_260629_R.xlsx` D5/E5/F5, D6/E6/F6).
    #: 실제 파일은 Exception 이 Fail 과 같은 값이라, 밀림을 관측 가능하게 하려고
    #: Exception 만 다른 수로 바꿨다(Total 은 실측 그대로).
    CANON = ((1014, 4, 40), (1014, 21, 7))

    def _pv_workbook(self) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "4.Coverage"
        _write_switcv_sheet(
            ws, "PV",
            [("SwUFn_1005", "ADC0_stop_current_workaround", "X", "O",
              1, 2, "X", "O", "cov.c")],
            self.CANON,
        )
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_total_is_total_not_fail_count(self):
        from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
        out = _load_workbook_summary(self._pv_workbook())
        assert out["functions_total"] == 1014        # 밀리면 4 가 나온다
        assert out["function_calls_total"] == 1014   # 밀리면 21 이 나온다

    def test_exception_is_exception_not_coverage_ratio(self):
        from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
        out = _load_workbook_summary(self._pv_workbook())
        assert out["functions_fail_count"] == 4
        assert out["functions_exception_count"] == 40   # 밀리면 Coverage 열을 읽는다
        assert out["function_calls_fail_count"] == 21
        assert out["function_calls_exception_count"] == 7

    def test_failed_rows_are_reported_not_swallowed(self):
        """밀려 읽으면 O/X 자리에서 Exception 열을 보게 돼 X 가 없어지고,
        SwITCR 이 실재하는 미달성을 '해당사항 없음' 으로 내보낸다."""
        from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
        details = _load_workbook_summary(self._pv_workbook())["coverage_fail_details"]
        kinds = [d["kind"] for d in details]
        assert "함수커버리지" in kinds
        assert "호출커버리지" in kinds
        fn = next(d for d in details if d["kind"] == "함수커버리지")
        assert fn["unit_id"] == "SwUFn_1005"                      # 밀리면 함수명이 들어온다
        assert fn["function"] == "ADC0_stop_current_workaround"   # 밀리면 "X" 가 들어온다

    def test_trailing_total_row_is_not_counted_as_a_function(self):
        """마감 TOTAL 행을 함수로 세면 합계가 1 늘어난다(정본 실측 1015 vs 1014)."""
        from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "4.Coverage"
        _write_switcv_sheet(
            ws, "PV",
            [("SwUFn_0101", "FuncA", "O", None, 1, 1, "O", None, None),
             ("Total", "Total", None, None, None, None, None, None, None)],
            None,                      # 요약셀 비움 → 행집계 경로를 강제로 밟는다
        )
        buf = io.BytesIO()
        wb.save(buf)
        out = _load_workbook_summary(buf.getvalue())
        assert out["functions_total"] == 1


class TestCoverageSheetNameIsMatchedLoosely:
    """회사 SwUTCV 정본의 시트명은 `4. Coverage`(점 뒤 공백)다 — 2026-08-26 실측.

    정확 매칭이면 통째로 놓치고, 커버리지 증거가 **조용히 0** 이 된다."""

    def test_sheet_with_space_after_dot_is_found(self):
        from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "4. Coverage"           # 공백 있음
        _write_switcv_sheet(
            ws, "PV",
            [("SwUFn_0101", "FuncA", "X", None, 1, 2, "O", None, None)],
            ((7, 1, 0), (7, 0, 0)),
        )
        buf = io.BytesIO()
        wb.save(buf)
        out = _load_workbook_summary(buf.getvalue())
        assert out["functions_total"] == 7
        assert len(out["coverage_fail_details"]) == 1


class TestSummaryColumnPrefersTheHeaderLabel:
    """요약 열은 **헤더 라벨 우선**, 없을 때만 `base + 1` 폴백.

    ⚠ 실물 두 판(SwITCV PV / SwUTCV DV)은 라벨 위치와 폴백이 마침 일치해서, 라벨
      탐지를 통째로 들어내도 아무 테스트가 안 깨졌다(뮤테이션 L7 생존). 계약을
      관측 가능하게 만들려면 **둘이 갈리는 워크북**으로 고정해야 한다.
    """

    def _sheet(self, total_col: int | None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if total_col is not None:
            ws.cell(4, total_col).value = "Total"
        return ws

    def test_label_wins_over_the_fallback(self):
        from backend.services.excel_layout_resolver import coverage_summary_col
        ws = self._sheet(total_col=7)          # 폴백(base+1=4)과 다른 자리
        assert coverage_summary_col(ws, base=3) == 7

    def test_fallback_is_used_when_no_label(self):
        from backend.services.excel_layout_resolver import coverage_summary_col
        ws = self._sheet(total_col=None)
        assert coverage_summary_col(ws, base=3) == 4   # PV
        assert coverage_summary_col(ws, base=4) == 5   # DV

    def test_label_match_is_case_insensitive_and_trimmed(self):
        from backend.services.excel_layout_resolver import coverage_summary_col
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(4, 6).value = "  TOTAL "
        assert coverage_summary_col(ws, base=3) == 6


class TestColumnBaseHasASingleSource:
    """열 판정 복제 금지 — 라운드 102 가 한쪽만 고쳐 이 결함이 났다.

    두 aggregator 가 각자 Component 라벨을 스캔하고 있었고 그중 하나만 DV/PV 적응을
    받았다. 어제 세 라우터의 헤더 절단 로직을 `warnings_header_json` 으로 합친 것과
    같은 이유다."""

    def test_both_aggregators_use_the_shared_detector(self):
        from backend.services import swit_comprehensive_aggregator as comp
        from backend.services import swit_coverage_aggregator as cov
        from tests.unit._source_probe import source_of

        for mod in (comp, cov):
            src = source_of(mod)
            assert "coverage_column_base" in src, f"{mod.__name__}: 공유 판정 미사용"
            assert 'strip() == "Component"' not in src, (
                f"{mod.__name__}: Component 스캔 복제본이 되살아났다"
            )


class TestIt201It301EvidenceGuard:
    """IT201 인터페이스·IT301 결함주입 증거 없을 때 통과 조작 금지 (deep-review A2).

    형제 SwUTCR UT201 과 동일 원칙: 실측/설정 증거가 없으면 passed=total(100%)·5/5·PASS 를
    무측정 stamp 하지 않고 노란 사용자입력 마킹 + 경고로 표면화한다. ISO 26262 인터페이스/
    결함주입은 ASIL 안전기구 검증 증거라 무측정 100% 위장은 audit 무결성을 깬다.
    """
    _PH = "▶ 사용자 입력 필요"

    def _ws(self):
        from backend.services.swit_comprehensive_aggregator import SwitcrBuildMeta
        wb = openpyxl.Workbook()
        meta = SwitcrBuildMeta(
            project_id="P", project_full_name="P", release_sw_version="1.0",
            test_date="2025-12-05", test_engineer="T",
        )
        return wb, wb.active, meta

    def test_it201_without_interface_passed_marks_user_input(self):
        from backend.services.swit_comprehensive_aggregator import _write_it201
        wb, ws, meta = self._ws()
        warns: list[str] = []
        # interface_passed 없음, total 은 agg 실측
        _write_it201(ws, meta, {"total_tcs": 12}, {"switcr_metadata": {}}, [], warns)
        assert ws["F70"].value == 12                     # 분모(전체)는 실측 유지
        assert str(ws["G70"].value or "").startswith(self._PH), "passed=total 100% 위장이 남았다"
        # W1 — 파생 수식 H70(=F70-G70)도 마킹돼야 한다(안 그러면 텍스트 G70 로 Excel #VALUE!).
        assert str(ws["H70"].value or "").startswith(self._PH), "H70 이 텍스트 G70 로 #VALUE! 위험"
        assert any("IT201" in w and "interface_passed" in w for w in warns)

    def test_it201_with_interface_passed_stamps_value(self):
        """대조: interface_passed 실측 있으면 값 기입(마킹 아님)."""
        from backend.services.swit_comprehensive_aggregator import _write_it201
        wb, ws, meta = self._ws()
        _write_it201(ws, meta, {"total_tcs": 12},
                     {"switcr_metadata": {"interface_total": 12, "interface_passed": 9}}, [], [])
        assert ws["F70"].value == 12 and ws["G70"].value == 9

    def test_it301_without_any_fi_evidence_marks_user_input(self):
        from backend.services.swit_comprehensive_aggregator import _write_it301
        wb, ws, meta = self._ws()
        warns: list[str] = []
        # FI 실측(fault_injection_summary=None) + 설정(fault_injection_*) 둘 다 없음
        _write_it301(ws, meta, {}, {"switcr_metadata": {}}, {}, None, warns)
        assert str(ws["E85"].value or "").startswith(self._PH), "5/5 통과 위장이 남았다"
        assert str(ws["F85"].value or "").startswith(self._PH)
        assert str(ws["H85"].value or "").startswith(self._PH)   # PASS 플래그 위장 제거
        assert any("IT301" in w and "결함주입" in w for w in warns)

    def test_it301_with_config_evidence_stamps_value(self):
        """대조: fault_injection 설정 있으면 값 기입(마킹 아님)."""
        from backend.services.swit_comprehensive_aggregator import _write_it301
        wb, ws, meta = self._ws()
        _write_it301(ws, meta, {},
                     {"switcr_metadata": {"fault_injection_count": 8, "fault_injection_passed": 8}},
                     {}, None, [])
        assert ws["E85"].value == 8 and ws["F85"].value == 8
        assert ws["H85"].value == 1   # 8/8 → PASS

    def test_it301_count_only_marks_passed_not_fabricated(self):
        """W2 — count(tc_count)만 있고 passed 부재면 passed=total(100% PASS) 조작 대신 마킹.

        단일 bool 게이트는 이 부분증거를 통과시켜 passed=total 을 부활시켰다(deep-review W2).
        뮤테이션: dual-flag 를 단일 `_has_fi_evidence` 로 되돌리면 F85=10(=total)이 돼 실패.
        """
        from backend.services.swit_comprehensive_aggregator import _write_it301
        wb, ws, meta = self._ws()
        warns: list[str] = []
        _write_it301(ws, meta, {}, {"switcr_metadata": {}}, {}, {"tc_count": 10}, warns)
        assert ws["E85"].value == 10                        # total 실측은 기입
        assert str(ws["F85"].value or "").startswith(self._PH), "count-only 인데 passed=total 조작이 남았다"
        assert str(ws["H85"].value or "").startswith(self._PH)   # PASS 판정도 마킹
        assert any("IT301" in w and ("passed" in w or "count-only" in w) for w in warns)

    def test_it301_unmeasured_fail_report_marks_not_none_applicable(self):
        """W3 — FI 미측정이면 Fail Report(C90)를 '해당사항 없음'(=Fail 0=통과)으로 위장 안 함."""
        from backend.services.swit_comprehensive_aggregator import _write_it301
        wb, ws, meta = self._ws()
        _write_it301(ws, meta, {}, {"switcr_metadata": {}}, {}, None, [])
        assert str(ws["C90"].value or "").startswith(self._PH), \
            "미측정인데 Fail Report 가 '해당사항 없음'(상단 마킹과 모순)"

    def test_it701_without_evidence_marks_safety_mechanism_results(self):
        """W5 — 안전기구(watchdog/RAM/stack) 검증 결과를 실측 없이 'Pass'로 조작 안 함.

        IT701 은 ASIL C/D 안전기구 증거다. system_error_protection 설정이 없으면 G70~ 결과를
        노란 마킹. 뮤테이션: `results.get(str(row), "Pass")` 로 되돌리면 G70="Pass"라 실패.
        """
        from backend.services.swit_comprehensive_aggregator import _write_it701
        wb, ws, meta = self._ws()
        ws.cell(80, 1).value = "x"   # 결과 루프 range(70..min(max_row,76))가 돌도록 max_row 확보
        warns: list[str] = []
        _write_it701(ws, meta, {"switcr_metadata": {}}, warns)
        assert str(ws["G70"].value or "").startswith(self._PH), "안전기구 결과가 무측정 'Pass' 조작"
        assert any("IT701" in w and "안전기구" in w for w in warns)

    def test_it701_with_evidence_stamps_result(self):
        """대조: system_error_protection 실측 있으면 그 값 기입(마킹 아님)."""
        from backend.services.swit_comprehensive_aggregator import _write_it701
        wb, ws, meta = self._ws()
        ws.cell(80, 1).value = "x"   # 결과 루프가 돌도록 max_row 확보
        _write_it701(ws, meta, {"switcr_metadata": {"system_error_protection": {"70": "Pass"}}}, [])
        assert ws["G70"].value == "Pass"   # 실측 제공 → 기입

    def test_it401_without_evidence_marks_resource_pass_rows(self):
        """W5 — 자원사용(RAM/ROM) 실측 없으면 하드코딩 'Pass'+가짜 사용량(=1312/4096) 대신 마킹."""
        from backend.services.swit_comprehensive_aggregator import _write_it401
        wb, ws, meta = self._ws()
        warns: list[str] = []
        _write_it401(ws, meta, {"switcr_metadata": {}}, warns)
        # row 78 = "=1312/4096"/"Pass" 하드코딩 → 값(K78)·판정(L78) 둘 다 마킹
        assert str(ws["K78"].value or "").startswith(self._PH), "가짜 사용량(=1312/4096)이 남았다"
        assert str(ws["L78"].value or "").startswith(self._PH), "하드코딩 Pass 가 남았다"
        # row 80 = 의도적 N/A(동적메모리 제외) → 그대로(마킹 아님)
        assert ws["L80"].value == "N/A"
        assert any("IT401" in w and "자원사용" in w for w in warns)


def _switr_canonical_workbook(
    *,
    sheet_name: str = "3.Test Log",
    summary: tuple[int, int, int] | None = (611, 611, 0),
    total_label: str = "Total",
) -> bytes:
    """회사 정본 SwITR 형태 — 2026-08-26 실측 배치 그대로.

    시트 `Cover / History / 1.Test Summary / 2.Deviation / 3.Test Log`.
    `1.Test Summary` r17 헤더 + r18/r20 부분합 + r21 Total, `3.Test Log` 는 TC ID 가
    **B열**(세로 병합이라 행 수 != TC 수).

    ⚠ 예전 픽스처는 `2.Test Log` 에 TC 를 F열로 넣어 **코드의 거울**이었다. 그래서
      "시트명도 열도 정본과 다르다" 는 사실이 한 번도 드러나지 않았다.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("History")
    ts = wb.create_sheet("1.Test Summary")
    if summary is not None:
        tested, passed, failed = summary
        ts["B17"], ts["C17"] = "Type", "Number of TCs Tested"
        ts["D17"], ts["E17"] = "Number of TCs Passed", "Number of TCs Failed"
        ts["F17"] = "Number of TCs not executed"
        ts["B18"], ts["C18"], ts["D18"], ts["E18"] = "Requirements Based TC", 581, 581, 0
        ts["B20"], ts["C20"], ts["D20"], ts["E20"] = "Fault Injection TC", 30, 30, 0
        ts["B21"] = total_label
        ts["C21"], ts["D21"], ts["E21"] = tested, passed, failed
    wb.create_sheet("2.Deviation")
    log = wb.create_sheet(sheet_name)
    log["B5"] = "Test Case"
    for i, (tc, verdict) in enumerate((
        ("SwITC_SwUFn_0101_01", "Pass"),
        ("SwITC_SwUFn_0102_01", "Pass"),
        ("SwITC_SwUFn_0103_01", "Fail"),
    )):
        r = 7 + i
        log.cell(r, 2).value = tc          # B열 — 정본 배치
        log.cell(r, 38).value = verdict    # AL열
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSwitrEvidenceIsActuallyRead:
    """SwITR 증거 읽기가 정본에서 **동작하는지** — 2026-08-26 실측 결함.

    예전 코드는 시트를 `"2.Test Log"` 로 정확 매칭하고 TC ID 열을 `row[5]`(F) 로 박아
    두었다. 정본은 `3.Test Log` / **B열**이라 둘 다 빗나갔고, `sitr_*` 세 키가 늘 부재였다.
    그 키들은 `or` 폴백 자리에만 쓰여 아무도 안 죽었고, **단언하는 테스트가 0건**이라
    아무도 못 봤다.
    """

    def _load(self, data: bytes) -> dict:
        from backend.services.swit_comprehensive_aggregator import _load_workbook_summary
        return _load_workbook_summary(data, keep_vba=False)

    def test_canonical_switr_yields_the_documents_own_total(self):
        out = self._load(_switr_canonical_workbook())
        assert out["sitr_test_log_tcs"] == 611     # 고치기 전엔 키 자체가 없었다
        assert out["sitr_pass_count"] == 611
        assert out["sitr_fail_count"] == 0

    def test_row_counting_would_have_given_the_wrong_number(self):
        """행을 세면 3 이 나온다 — 문서 Total 은 611. 세는 접근이 이 양식엔 안 맞는다."""
        out = self._load(_switr_canonical_workbook())
        assert out["sitr_test_log_tcs"] != 3

    def test_falls_back_to_test_log_when_no_summary_block(self):
        """`1.Test Summary` 집계가 없는 판(v1.01)은 행을 센다 — 시트명/열은 탐지."""
        out = self._load(_switr_canonical_workbook(summary=None))
        assert out["sitr_test_log_tcs"] == 3       # B열 TC 3건
        assert out["sitr_pass_count"] == 2
        assert out["sitr_fail_count"] == 1

    def test_legacy_two_dot_test_log_still_works(self):
        """예전 판(`2.Test Log`)도 계속 읽힌다 — 정본을 고치며 깨뜨리지 않는다."""
        out = self._load(_switr_canonical_workbook(sheet_name="2.Test Log", summary=None))
        assert out["sitr_test_log_tcs"] == 3

    def test_absent_block_yields_no_keys_not_zeros(self):
        """집계도 로그도 없으면 **키를 안 만든다** — 0 으로 지어내면 '문서가 0건이라 한다'
        는 없는 사실이 된다."""
        wb = openpyxl.Workbook()
        wb.active.title = "Cover"
        buf = io.BytesIO()
        wb.save(buf)
        out = self._load(buf.getvalue())
        assert "sitr_test_log_tcs" not in out
        assert "sitr_pass_count" not in out

    def test_non_numeric_total_is_rejected_wholesale(self):
        """Total 한 칸이라도 수가 아니면 통째로 포기한다(부분 신뢰 금지) → 로그 폴백."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ts = wb.create_sheet("1.Test Summary")
        ts["C17"], ts["D17"], ts["E17"] = (
            "Number of TCs Tested", "Number of TCs Passed", "Number of TCs Failed")
        ts["B21"], ts["C21"], ts["D21"], ts["E21"] = "Total", "-", 611, 0
        buf = io.BytesIO()
        wb.save(buf)
        out = self._load(buf.getvalue())
        # ⚠ **한 키도** 남으면 안 된다. `return {}` 대신 그 칸만 건너뛰면 611/0 이
        #   부분적으로 살아남아 "문서가 611 통과라 한다" 는 절반짜리 근거가 된다.
        assert "sitr_test_log_tcs" not in out
        assert "sitr_pass_count" not in out
        assert "sitr_fail_count" not in out

    def test_swutcv_style_summary_sheet_is_not_mistaken(self):
        """SwITCV 에도 `1.Test Summary` 가 있지만 TC 집계 블록이 없다 — 오염 금지."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ts = wb.create_sheet("1.Test Summary")
        ts["B9"], ts["C9"] = "Target Pass Ratio", 1
        ts["B12"], ts["C12"] = "추적성 커버리지", "정합성 커버리지"
        buf = io.BytesIO()
        wb.save(buf)
        out = self._load(buf.getvalue())
        assert "sitr_pass_count" not in out


class TestSessionVsSwitrDivergenceIsReported:
    """세션 실측 ↔ 승인 문서 대조 — 읽기를 고쳐도 말하지 않으면 폴백에 앉을 뿐이다."""

    def _warn(self, agg, doc):
        from backend.services.swit_comprehensive_aggregator import (
            _switr_divergence_warnings,
        )
        return _switr_divergence_warnings(agg, doc)

    def test_divergence_is_reported_with_both_numbers(self):
        out = self._warn({"total_tcs": 611, "passed": 611, "failed": 0},
                         {"sitr_test_log_tcs": 600, "sitr_pass_count": 611,
                          "sitr_fail_count": 0})
        assert len(out) == 1
        assert "611" in out[0] and "600" in out[0]
        assert out[0].startswith("[evidence]")     # 분류기가 아는 prefix 여야 화면에 닿는다

    def test_agreement_is_silent(self):
        assert self._warn({"total_tcs": 611, "passed": 611, "failed": 0},
                          {"sitr_test_log_tcs": 611, "sitr_pass_count": 611,
                           "sitr_fail_count": 0}) == []

    def test_missing_side_is_not_a_divergence(self):
        """부재 != 불일치. 없는 증거를 0 으로 접으면 없는 사실을 만든다."""
        assert self._warn({"total_tcs": 611}, {}) == []
        assert self._warn({}, {"sitr_test_log_tcs": 611}) == []
        assert self._warn({"total_tcs": 611}, {"sitr_test_log_tcs": None}) == []

    def test_every_axis_is_checked(self):
        out = self._warn({"total_tcs": 1, "passed": 2, "failed": 3},
                         {"sitr_test_log_tcs": 9, "sitr_pass_count": 9,
                          "sitr_fail_count": 9})
        assert len(out) == 3


class TestIt101ReadsSitrKeysFromTheSwitrDict:
    """IT101 이 `sitr_*` 를 **SwITCV** dict 에서 찾고 있었다 — 늘 부재였다."""

    def test_no_sitr_lookup_against_the_coverage_summary(self):
        from backend.services.swit_comprehensive_aggregator import _write_it101
        from tests.unit._source_probe import source_of
        src = source_of(_write_it101)
        assert 'switcv_summary.get("sitr_' not in src, "SwITCV dict 에서 SITR 키를 찾는다"
        assert "_switr.get(" in src, "SwITR dict 를 안 쓴다"

    def test_call_site_passes_the_switr_summary(self):
        import ast

        from backend.services import swit_comprehensive_aggregator as mod
        from tests.unit._source_probe import source_of
        tree = ast.parse(source_of(mod))
        calls = [
            n for n in ast.walk(tree)
            if isinstance(n, ast.Call) and getattr(n.func, "id", "") == "_write_it101"
        ]
        assert len(calls) == 1, f"호출부가 {len(calls)}곳 — 가드 대상 재확인 필요"
        assert any(k.arg == "switr_summary" for k in calls[0].keywords), (
            "호출부가 switr_summary 를 안 넘긴다"
        )
