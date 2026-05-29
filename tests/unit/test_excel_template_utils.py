"""Tests for backend.services.excel_template_utils."""
from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.services.excel_template_utils import (  # noqa: E402
    BLANK_MARKUP,
    BuildMetaValidationError,
    TemplateValidationError,
    auto_expand_row_block,
    clear_data_range,
    find_kv_row,
    push_sentinel_to_last_row,
    resolve_merge_anchor,
    safe_write,
    sheet_is_blank_placeholder,
    short_date,
    update_cross_refs_after_row_expansion,
    validate_build_meta,
    validate_xlsx_template_bytes,
    write_value_after_label,
)


def _minimal_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# validate_xlsx_template_bytes — Critical S (reviewer)
# ---------------------------------------------------------------------------

class TestValidateXlsxBytes:
    def test_valid_xlsx_passes(self):
        validate_xlsx_template_bytes(_minimal_xlsx_bytes())

    def test_empty_bytes_rejected(self):
        with pytest.raises(TemplateValidationError, match="empty"):
            validate_xlsx_template_bytes(b"")

    def test_bad_magic_rejected(self):
        with pytest.raises(TemplateValidationError, match="magic bytes"):
            validate_xlsx_template_bytes(b"NOT_AN_XLSX_FILE_AT_ALL_!!!")

    def test_truncated_zip_rejected(self):
        # PK 헤더만 있고 잘림
        with pytest.raises(TemplateValidationError, match="valid ZIP"):
            validate_xlsx_template_bytes(b"PK\x03\x04\x00\x00\x00\x00")

    def test_zip_without_office_structure_rejected(self):
        # 평범한 ZIP (Office XML 구조 없음)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("readme.txt", "hello")
        with pytest.raises(TemplateValidationError, match="Open Office XML"):
            validate_xlsx_template_bytes(buf.getvalue())


# ---------------------------------------------------------------------------
# Helpers (parity with previous tests)
# ---------------------------------------------------------------------------

class TestShortDate:
    @pytest.mark.parametrize("inp,expected", [
        ("2024-02-19", "240219"),
        ("2024/02/19", "240219"),
        ("24-02-19", "240219"),
        ("", ""),
    ])
    def test_parse(self, inp, expected):
        assert short_date(inp) == expected


class TestFindKvRow:
    def test_finds_label(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B3"] = "Project"
        ws["C3"] = "HDPDM01"
        pos = find_kv_row(ws, "Project")
        assert pos == (3, 2)

    def test_not_found_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert find_kv_row(ws, "NonExistent") is None


class TestMergedCellHandling:
    def test_anchor_resolves_for_merged_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "Project"
        ws.merge_cells("B2:C2")
        assert resolve_merge_anchor(ws, 2, 3) == (2, 2)  # C2 → anchor B2

    def test_anchor_passthrough_for_unmerged(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert resolve_merge_anchor(ws, 5, 5) == (5, 5)

    def test_write_value_after_merged_label(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "Project"
        ws.merge_cells("B2:C2")  # 라벨이 2 컬럼 머지
        assert write_value_after_label(ws, "Project", "HDPDM01")
        # 라벨 머지 영역 다음(D2)에 값 들어가야 함
        assert ws["D2"].value == "HDPDM01"

    def test_safe_write_silent_on_merged_non_anchor(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "anchor"
        ws.merge_cells("A1:B1")
        # B1(머지된 non-anchor)에 쓰기 시도 → anchor로 보정되어 성공
        assert safe_write(ws, 1, 2, "X") is True
        assert ws["A1"].value == "X"  # anchor가 덮어쓰여짐


class TestValidateBuildMeta:
    """deep-reviewer X3: 빌더 입력 메타 검증."""

    def test_valid_meta_passes(self):
        validate_build_meta("1.01.05", "2024-02-19")
        validate_build_meta("2.02", "2024/02/19")

    def test_empty_release_rejected(self):
        with pytest.raises(BuildMetaValidationError, match="release_sw_version is empty"):
            validate_build_meta("", "2024-02-19")
        with pytest.raises(BuildMetaValidationError, match="release_sw_version is empty"):
            validate_build_meta("   ", "2024-02-19")

    def test_invalid_release_format_rejected(self):
        with pytest.raises(BuildMetaValidationError, match="형식 미충족"):
            validate_build_meta("v1.01.05", "2024-02-19")
        with pytest.raises(BuildMetaValidationError, match="형식 미충족"):
            validate_build_meta("1.x.05", "2024-02-19")

    def test_empty_date_rejected(self):
        with pytest.raises(BuildMetaValidationError, match="test_date is empty"):
            validate_build_meta("1.01.05", "")

    def test_invalid_date_format_rejected(self):
        with pytest.raises(BuildMetaValidationError, match="형식 미충족"):
            validate_build_meta("1.01.05", "Feb 19 2024")

    def test_test_engineer_with_newline_rejected(self):
        """5차 H1 Critical: 줄바꿈 포함 이름 거부 (xlsx 셀 깨짐 방지)."""
        with pytest.raises(BuildMetaValidationError, match="줄바꿈"):
            validate_build_meta("1.0.0", "2024-02-19", test_engineer="X\nY")

    def test_test_engineer_too_long_rejected(self):
        with pytest.raises(BuildMetaValidationError, match="길이"):
            validate_build_meta("1.0.0", "2024-02-19", test_engineer="A" * 101)

    def test_doc_id_sequence_non_digit_rejected(self):
        """5차 H2 Warning: doc_id_sequence digit만 허용."""
        with pytest.raises(BuildMetaValidationError, match="doc_id_sequence"):
            validate_build_meta("1.0.0", "2024-02-19", doc_id_sequence="abc")

    def test_doc_id_sequence_empty_ok(self):
        validate_build_meta("1.0.0", "2024-02-19", doc_id_sequence="")
        validate_build_meta("1.0.0", "2024-02-19", doc_id_sequence="851")


class TestTruncateCellText:
    """5차 H3 Critical: xlsx 셀 한도 방어 truncate."""

    def test_short_text_unchanged(self):
        from backend.services.excel_template_utils import truncate_cell_text
        s, truncated = truncate_cell_text("hello")
        assert s == "hello"
        assert truncated is False

    def test_long_text_truncated(self):
        from backend.services.excel_template_utils import truncate_cell_text
        long = "A" * 5000
        s, truncated = truncate_cell_text(long)
        assert truncated is True
        assert len(s) <= 2050  # 2000 + " …(truncated)" 마진
        assert s.endswith("…(truncated)")

    def test_none_returns_empty(self):
        from backend.services.excel_template_utils import truncate_cell_text
        s, truncated = truncate_cell_text(None)
        assert s == ""
        assert truncated is False


class TestSheetIsBlankPlaceholder:
    """deep-reviewer 시나리오 A: placeholder 시트 감지."""

    def test_detects_blank_markup_in_a1(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = BLANK_MARKUP
        assert sheet_is_blank_placeholder(ws) is True

    def test_no_blank_markup_returns_false(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Normal data here"
        assert sheet_is_blank_placeholder(ws) is False

    def test_none_sheet_returns_false(self):
        assert sheet_is_blank_placeholder(None) is False


# ---------------------------------------------------------------------------
# 23차 T192/T194: 시각 강조 헬퍼 회귀
# ---------------------------------------------------------------------------

class TestVisualMarkers23:
    """사용자 입력 필요 (노란) / FAIL (빨강) 시각 강조 회귀."""

    def test_mark_user_input_required_writes_placeholder_and_fill(self):
        from backend.services.excel_template_utils import (
            mark_user_input_required, USER_INPUT_PLACEHOLDER,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        result = mark_user_input_required(ws, 2, 3, hint="Approver 이름")
        assert result is True
        cell = ws.cell(2, 3)
        assert cell.value.startswith(USER_INPUT_PLACEHOLDER)
        assert "Approver 이름" in cell.value
        assert cell.fill.fill_type == "solid"
        assert "FFEB9C" in str(cell.fill.fgColor.rgb).upper()

    def test_mark_user_input_required_no_hint(self):
        from backend.services.excel_template_utils import mark_user_input_required
        wb = openpyxl.Workbook()
        ws = wb.active
        mark_user_input_required(ws, 1, 1)
        assert "▶ 사용자 입력 필요" in ws.cell(1, 1).value
        assert "—" not in ws.cell(1, 1).value  # hint 없을 때 separator 없음

    def test_write_value_or_mark_writes_when_value_present(self):
        from backend.services.excel_template_utils import write_value_or_mark
        wb = openpyxl.Workbook()
        ws = wb.active
        result = write_value_or_mark(ws, 1, 1, "JK Kim", hint="ignored")
        assert result is True
        assert ws.cell(1, 1).value == "JK Kim"

    def test_write_value_or_mark_marks_when_value_empty(self):
        from backend.services.excel_template_utils import write_value_or_mark
        wb = openpyxl.Workbook()
        ws = wb.active
        result = write_value_or_mark(ws, 1, 1, "", hint="hint here")
        assert result is False  # marked, not written
        assert "▶ 사용자 입력 필요" in ws.cell(1, 1).value
        assert "FFEB9C" in str(ws.cell(1, 1).fill.fgColor.rgb).upper()

    def test_mark_fail_cell_applies_red_fill_and_preserves_value(self):
        from backend.services.excel_template_utils import mark_fail_cell
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "FAIL"
        result = mark_fail_cell(ws, 1, 1)
        assert result is True
        assert "FFC7CE" in str(ws.cell(1, 1).fill.fgColor.rgb).upper()
        # value 보존 — 색칠이 텍스트 reset 하지 않음
        assert ws.cell(1, 1).value == "FAIL"

    def test_write_label_or_mark_finds_label_and_marks_empty(self):
        from backend.services.excel_template_utils import write_label_or_mark
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A4"] = "Validation Date"
        ws["B4"] = "(placeholder)"
        out_warnings: list[str] = []
        result = write_label_or_mark(
            ws, "Validation Date", value="", hint="yyyy-mm-dd",
            optional_labels={"Reviewer"}, out_warnings=out_warnings,
        )
        assert result is False  # 빈 value → 노란 mark
        assert "▶ 사용자 입력 필요" in ws.cell(4, 2).value
        assert "yyyy-mm-dd" in ws.cell(4, 2).value
        assert "FFEB9C" in str(ws.cell(4, 2).fill.fgColor.rgb).upper()

    def test_write_label_or_mark_writes_when_value_present(self):
        from backend.services.excel_template_utils import write_label_or_mark
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Author"
        out_warnings: list[str] = []
        result = write_label_or_mark(
            ws, "Author", value="JK Kim", hint="이름",
            optional_labels=None, out_warnings=out_warnings,
        )
        assert result is True  # value 있음
        assert ws.cell(1, 2).value == "JK Kim"
        assert out_warnings == []

    def test_write_label_or_mark_skips_optional_label_warning(self):
        from backend.services.excel_template_utils import write_label_or_mark
        wb = openpyxl.Workbook()
        ws = wb.active
        out_warnings: list[str] = []
        write_label_or_mark(
            ws, "Reviewer", value="someone", hint="",
            optional_labels={"Reviewer"}, out_warnings=out_warnings,
        )
        assert out_warnings == []  # optional 라벨이라 미발견 silent

    def test_write_label_or_mark_warns_non_optional_label_missing(self):
        from backend.services.excel_template_utils import write_label_or_mark
        wb = openpyxl.Workbook()
        ws = wb.active
        out_warnings: list[str] = []
        write_label_or_mark(
            ws, "Validation Date", value="2024-02-19", hint="",
            optional_labels={"Reviewer"}, out_warnings=out_warnings,
        )
        assert any("Validation Date" in w for w in out_warnings)

    def test_mark_handles_merged_cell_anchor(self):
        """머지셀 비-anchor 위치 호출 → anchor 셀에 mark."""
        from backend.services.excel_template_utils import mark_user_input_required
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("B2:D2")
        mark_user_input_required(ws, 2, 4, hint="merged")  # D2 → anchor B2
        assert ws.cell(2, 2).value is not None
        assert "▶ 사용자 입력 필요" in ws.cell(2, 2).value

    def test_design_tokens_single_source_29w17(self):
        """29차 W17: excel_template_utils가 design_tokens에서 RGB를 import."""
        from backend.services import design_tokens
        from backend.services.excel_template_utils import (
            _FAIL_FILL_RGB,
            _USER_INPUT_FILL_RGB,
            USER_INPUT_PLACEHOLDER,
        )

        # 동일 객체(=Final 상수)를 가리키는지 확인 — 단일 출처 보장
        assert _USER_INPUT_FILL_RGB == design_tokens.USER_INPUT_FILL_RGB
        assert _FAIL_FILL_RGB == design_tokens.FAIL_FILL_RGB
        assert USER_INPUT_PLACEHOLDER == design_tokens.USER_INPUT_PLACEHOLDER

        # 값 자체도 회귀 — audit 정책 RGB 변경 시 본 테스트가 실패해야 함
        assert _USER_INPUT_FILL_RGB == "FFFFEB9C"
        assert _FAIL_FILL_RGB == "FFFFC7CE"
        assert USER_INPUT_PLACEHOLDER == "▶ 사용자 입력 필요"


class TestAsilDMarker21:
    """30차 W21: mark_asil_d_function — ASIL D 함수 row 빨간 강조."""

    def test_mark_asil_d_applies_red_fill(self):
        from backend.services.excel_template_utils import mark_asil_d_function
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "SwUFn_0103"
        result = mark_asil_d_function(ws, 1, 1)
        assert result is True
        # RGB는 ASIL_D_FILL_RGB (= FAIL_FILL_RGB와 동일 값)
        assert "FFC7CE" in str(ws.cell(1, 1).fill.fgColor.rgb).upper()
        # value 보존 — 색칠이 함수 ID reset 하지 않음
        assert ws.cell(1, 1).value == "SwUFn_0103"

    def test_asil_d_and_fail_share_same_rgb(self):
        """30차 W21: ASIL D / FAIL은 동일 RGB이나 의미 분리.

        design_tokens에서 두 상수가 같은 값을 가리키는지 검증 — 변경 시
        본 테스트가 실패해야 정책 동기 의무 인지.
        """
        from backend.services import design_tokens
        assert design_tokens.ASIL_D_FILL_RGB == design_tokens.FAIL_FILL_RGB

    def test_mark_asil_d_handles_merged_cell_anchor(self):
        """머지셀 비-anchor 위치 호출 → anchor 셀에 fill."""
        from backend.services.excel_template_utils import mark_asil_d_function
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("C3:E3")
        result = mark_asil_d_function(ws, 3, 5)  # E3 → anchor C3
        assert result is True
        assert "FFC7CE" in str(ws.cell(3, 3).fill.fgColor.rgb).upper()


class TestAsilBCMarker31:
    """31차 W29: mark_asil_b_function / mark_asil_c_function — ASIL B/C 강조."""

    def test_mark_asil_b_applies_blue_fill(self):
        from backend.services.excel_template_utils import mark_asil_b_function
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "SwUFn_0201"
        result = mark_asil_b_function(ws, 1, 1)
        assert result is True
        # ASIL_B_FILL_RGB = "FFE2F0FF" 연한 파랑
        assert "E2F0FF" in str(ws.cell(1, 1).fill.fgColor.rgb).upper()
        assert ws.cell(1, 1).value == "SwUFn_0201"

    def test_mark_asil_c_applies_orange_fill(self):
        from backend.services.excel_template_utils import mark_asil_c_function
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(2, 1).value = "SwUFn_0301"
        result = mark_asil_c_function(ws, 2, 1)
        assert result is True
        # ASIL_C_FILL_RGB = "FFFFE5CC" 연한 주황
        assert "FFE5CC" in str(ws.cell(2, 1).fill.fgColor.rgb).upper()
        assert ws.cell(2, 1).value == "SwUFn_0301"

    def test_asil_b_c_d_rgbs_are_all_distinct(self):
        """ASIL B/C/D RGB가 모두 다른 색 — audit reviewer 등급 구분 가능."""
        from backend.services import design_tokens
        rgbs = {
            design_tokens.ASIL_B_FILL_RGB,
            design_tokens.ASIL_C_FILL_RGB,
            design_tokens.ASIL_D_FILL_RGB,
        }
        assert len(rgbs) == 3, f"중복된 RGB 발견: {rgbs}"

    def test_asil_b_c_distinct_from_user_input_and_fail(self):
        """ASIL B/C가 USER_INPUT(노랑)/FAIL(빨강)와도 다른 색."""
        from backend.services import design_tokens
        assert design_tokens.ASIL_B_FILL_RGB not in (
            design_tokens.USER_INPUT_FILL_RGB,
            design_tokens.FAIL_FILL_RGB,
        )
        assert design_tokens.ASIL_C_FILL_RGB not in (
            design_tokens.USER_INPUT_FILL_RGB,
            design_tokens.FAIL_FILL_RGB,
        )

    def test_mark_asil_b_handles_merged_cell_anchor(self):
        from backend.services.excel_template_utils import mark_asil_b_function
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("B2:D2")
        result = mark_asil_b_function(ws, 2, 4)  # D2 → anchor B2
        assert result is True
        assert "E2F0FF" in str(ws.cell(2, 2).fill.fgColor.rgb).upper()

    def test_mark_asil_c_handles_merged_cell_anchor(self):
        from backend.services.excel_template_utils import mark_asil_c_function
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("E5:G5")
        result = mark_asil_c_function(ws, 5, 7)  # G5 → anchor E5
        assert result is True
        assert "FFE5CC" in str(ws.cell(5, 5).fill.fgColor.rgb).upper()


# ---------------------------------------------------------------------------
# 55-fix-2 W6 — build_release_history_row 빈 입력 warning 누적
# ---------------------------------------------------------------------------


class TestBuildReleaseHistoryRow55fix2:
    """build_release_history_row helper 회귀 (55-fix + 55-fix-2)."""

    def _meta(self, **kwargs):
        """SimpleNamespace mock meta — release_sw_version / test_date / author."""
        from types import SimpleNamespace
        defaults = {
            "release_sw_version": "1.0.0",
            "test_date": "2024-02-19",
            "author": "JK Kim",
        }
        defaults.update(kwargs)
        return SimpleNamespace(**defaults)

    def test_normal_release_row(self):
        """정상 meta → 1 row + doc_kind description."""
        from backend.services.excel_template_utils import build_release_history_row
        rows = build_release_history_row(self._meta(), doc_kind="SwIT Coverage Report")
        assert len(rows) == 1
        assert rows[0]["version"] == "v1.0.0"
        assert rows[0]["date"] == "24.02.19"
        assert rows[0]["author"] == "JK Kim"
        assert "SwIT Coverage Report" in rows[0]["description"]

    def test_empty_release_version_accumulates_warning(self):
        """55-fix-2 W6: release_sw_version 빈 시 warning 누적 + version cell 빈 string."""
        from backend.services.excel_template_utils import build_release_history_row
        warnings = []
        rows = build_release_history_row(
            self._meta(release_sw_version=""),
            doc_kind="SwUT Coverage Report",
            out_warnings=warnings,
        )
        assert rows[0]["version"] == ""
        assert any("release_sw_version 빈" in w for w in warnings)

    def test_empty_test_date_accumulates_warning(self):
        """55-fix-2 W6: test_date 빈 시 warning 누적."""
        from backend.services.excel_template_utils import build_release_history_row
        warnings = []
        rows = build_release_history_row(
            self._meta(test_date=""),
            doc_kind="SwUT SUTR",
            out_warnings=warnings,
        )
        assert rows[0]["date"] == ""
        assert any("test_date 빈" in w for w in warnings)

    def test_out_warnings_none_no_error(self):
        """out_warnings=None (default) — 빈 입력에도 raise 없음."""
        from backend.services.excel_template_utils import build_release_history_row
        rows = build_release_history_row(
            self._meta(release_sw_version="", test_date=""),
        )
        assert len(rows) == 1
        assert rows[0]["version"] == ""
        assert rows[0]["date"] == ""


class TestClearDataRange:
    """라운드 D T611: clear_data_range — partial overwrite 결함 fix 회귀."""

    def _make_ws(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # 양식 default 데이터 시뮬레이션 — R5~R10 × C1~C4
        for r in range(5, 11):
            for c in range(1, 5):
                ws.cell(r, c).value = f"default_R{r}C{c}"
        return ws

    def test_clears_data_rows_within_range(self):
        ws = self._make_ws()
        cleared = clear_data_range(
            ws, start_row=5, end_row=10, start_col=1, end_col=4,
        )
        assert cleared == 24
        for r in range(5, 11):
            for c in range(1, 5):
                assert ws.cell(r, c).value is None

    def test_preserves_formula_cells(self):
        """preserve_formula=True (default)면 = 시작 cell은 clear 안 함."""
        ws = self._make_ws()
        ws.cell(6, 2).value = "='Other Sheet'!A1"
        cleared = clear_data_range(
            ws, start_row=5, end_row=10, start_col=1, end_col=4,
        )
        # 수식 1개 보존 → 24 - 1 = 23
        assert cleared == 23
        assert ws.cell(6, 2).value == "='Other Sheet'!A1"

    def test_preserve_formula_false_clears_formula(self):
        ws = self._make_ws()
        ws.cell(7, 3).value = "=SUM(A1:A10)"
        cleared = clear_data_range(
            ws, start_row=5, end_row=10, start_col=1, end_col=4,
            preserve_formula=False,
        )
        assert cleared == 24  # 수식까지 clear
        assert ws.cell(7, 3).value is None

    def test_skips_none_cells(self):
        ws = openpyxl.Workbook().active
        # 빈 시트 → clear 대상 0
        cleared = clear_data_range(
            ws, start_row=1, end_row=10, start_col=1, end_col=5,
        )
        assert cleared == 0

    def test_preserves_merged_non_anchor(self):
        """머지 영역의 비-anchor cell은 보존 (anchor만 clear)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(5, 1).value = "merged_anchor"
        ws.merge_cells(start_row=5, end_row=5, start_column=1, end_column=3)
        ws.cell(5, 4).value = "non_merged"
        cleared = clear_data_range(
            ws, start_row=5, end_row=5, start_col=1, end_col=4,
        )
        # anchor(5,1) + non_merged(5,4) 2개만 clear (merged 비-anchor는 보존)
        assert cleared == 2
        assert ws.cell(5, 1).value is None
        assert ws.cell(5, 4).value is None

    def test_invalid_range_returns_zero(self):
        ws = self._make_ws()
        # start_row > end_row → 즉시 0
        cleared = clear_data_range(
            ws, start_row=10, end_row=5, start_col=1, end_col=4,
        )
        assert cleared == 0

    def test_sentinel_patterns_preserves_end_marker_round_f7_r1(self):
        """F7 자체평가 R1 C1/C3 fix: sentinel_patterns 발견 시 그 row 직전까지 clear.
        양식 끝 마커 ('< End of Document >', '■ Appendix') 보존."""
        ws = self._make_ws()
        # R8 C2에 sentinel 추가
        ws.cell(8, 2).value = "< End of Document >"
        cleared = clear_data_range(
            ws, start_row=5, end_row=10, start_col=1, end_col=4,
            sentinel_patterns=["End of Document", "Appendix"],
        )
        # R5~R7만 clear (R8~ sentinel 보존)
        # R5/R6/R7 × C1~C4 = 12개 cell
        assert cleared == 12
        # sentinel row R8 보존
        assert ws.cell(8, 2).value == "< End of Document >"
        # R9/R10 보존 (sentinel 뒤)
        assert ws.cell(9, 1).value == "default_R9C1"
        assert ws.cell(10, 4).value == "default_R10C4"

    def test_sentinel_patterns_no_match_clears_all(self):
        """sentinel 미발견 → 전체 clear."""
        ws = self._make_ws()
        cleared = clear_data_range(
            ws, start_row=5, end_row=10, start_col=1, end_col=4,
            sentinel_patterns=["nonexistent_sentinel"],
        )
        # 전체 24 cell clear
        assert cleared == 24

    def test_sentinel_substring_false_positive_strict_match_round_f7_r2_n2(self):
        """F7 R2 N2 fix: 함수명 'AppendixHelper' substring 'Appendix' false positive
        차단. strict 매칭으로 양식 끝 마커 prefix ('<', '■', '※') 또는 exact만 인정."""
        ws = self._make_ws()
        # R7 C2에 함수명 'AppendixHelper' (sentinel false positive 후보)
        ws.cell(7, 2).value = "AppendixHelper"
        cleared = clear_data_range(
            ws, start_row=5, end_row=10, start_col=1, end_col=4,
            sentinel_patterns=["Appendix"],
        )
        # strict 매칭 — '< ' prefix 없으므로 sentinel 인식 안 함 → 전체 clear
        # R5/R6 + R7 (AppendixHelper도 clear됨) + R8~R10 × C1~C4 = 24
        assert cleared == 24

    def test_sentinel_strict_match_prefix_round_f7_r2_n2(self):
        """F7 R2 N2 fix: '■ Appendix' / '< End of Document >' 같은 양식 마커
        prefix만 sentinel로 인식."""
        ws = self._make_ws()
        ws.cell(7, 2).value = "■ Appendix - 발생 가능 값"  # 양식 마커
        cleared = clear_data_range(
            ws, start_row=5, end_row=10, start_col=1, end_col=4,
            sentinel_patterns=["Appendix"],
        )
        # R5/R6만 clear (R7 sentinel + R8~R10 보존)
        assert cleared == 8
        assert ws.cell(7, 2).value == "■ Appendix - 발생 가능 값"
        assert ws.cell(8, 1).value == "default_R8C1"


# ---------------------------------------------------------------------------
# auto_expand_row_block — 라운드 73 T801
# ---------------------------------------------------------------------------

class TestAutoExpandRowBlock:
    def _make_ws(self, rows: int = 10, cols: int = 5):
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).value = f"R{r}C{c}"
        return wb, ws

    def test_basic_insert_shifts_existing_rows(self):
        """row 5에 2 row 삽입 → 기존 R5~R10은 R7~R12로 shift."""
        wb, ws = self._make_ws(rows=10, cols=3)
        inserted = auto_expand_row_block(
            ws,
            insert_at_row=5,
            amount=2,
            template_row_idx=4,
            copy_style=False,
            copy_merge=False,
            copy_dimension=False,
        )
        assert inserted == 2
        # 기존 R5 -> R7로 shift
        assert ws.cell(7, 1).value == "R5C1"
        # 신규 R5/R6는 빈 셀
        assert ws.cell(5, 1).value is None
        assert ws.cell(6, 1).value is None
        # max_row 확장
        assert ws.max_row >= 12

    def test_style_copy_from_template_row(self):
        """template row의 cell._style 복제 — font/fill 모두 보존."""
        from openpyxl.styles import Font, PatternFill
        wb, ws = self._make_ws(rows=10, cols=3)
        # template row 4에 bold + yellow fill
        for c in range(1, 4):
            cell = ws.cell(row=4, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
        inserted = auto_expand_row_block(
            ws,
            insert_at_row=5,
            amount=1,
            template_row_idx=4,
            copy_style=True,
            copy_merge=False,
            copy_dimension=False,
        )
        assert inserted == 1
        # 신규 row 5의 cell이 bold + yellow fill 가져야 함
        new_cell = ws.cell(row=5, column=1)
        assert new_cell.font.bold is True
        assert new_cell.fill.start_color.rgb == "FFFFEB9C"

    def test_merge_copy_single_row(self):
        """template row의 single-row merge가 신규 row마다 동일 col span으로 복제."""
        wb, ws = self._make_ws(rows=10, cols=5)
        # template row 4 col 2~4 merge
        ws.merge_cells("B4:D4")
        inserted = auto_expand_row_block(
            ws,
            insert_at_row=5,
            amount=2,
            template_row_idx=4,
            copy_style=False,
            copy_merge=True,
            copy_dimension=False,
        )
        assert inserted == 2
        merge_strs = {str(mr) for mr in ws.merged_cells.ranges}
        # 신규 R5/R6에 동일 merge가 추가되어야 함
        assert "B5:D5" in merge_strs
        assert "B6:D6" in merge_strs

    def test_zero_or_negative_amount_returns_zero(self):
        """amount<=0 시 silent 0 + ws 변경 없음."""
        wb, ws = self._make_ws(rows=10, cols=3)
        original_max_row = ws.max_row
        assert auto_expand_row_block(ws, insert_at_row=5, amount=0, template_row_idx=4) == 0
        assert auto_expand_row_block(ws, insert_at_row=5, amount=-3, template_row_idx=4) == 0
        assert ws.max_row == original_max_row

    def test_multi_row_merge_block_replicated(self):
        """라운드 73 P3 fix: template row가 multi-row merge의 첫 row일 때 block 단위 복제.

        회사 v3.01 SUTR Test Result 1 TC당 6-row block merge (B17:B22) 패턴. amount=12
        (=block 2개)이면 B17:B22 + B23:B28 (신규 row) 동일 col span 복제 확인.
        """
        wb, ws = self._make_ws(rows=30, cols=5)
        # template_row=10 ~ 15 (6-row block) col 2 merge
        ws.merge_cells("B10:B15")
        # amount=6, template_row_idx=10 → block 1개 복제 (B16:B21)
        inserted = auto_expand_row_block(
            ws,
            insert_at_row=16,
            amount=6,
            template_row_idx=10,
            copy_style=False,
            copy_merge=True,
            copy_dimension=False,
        )
        assert inserted == 6
        merge_strs = {str(mr) for mr in ws.merged_cells.ranges}
        assert "B10:B15" in merge_strs  # 기존 보존
        assert "B16:B21" in merge_strs  # 신규 block 복제


# ---------------------------------------------------------------------------
# push_sentinel_to_last_row — 라운드 73 T804 helper
# ---------------------------------------------------------------------------

class TestPushSentinelToLastRow:
    def test_sentinel_at_end_returns_same_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "data1"
        ws.cell(2, 1).value = "data2"
        ws.cell(3, 1).value = "< End of Document >"
        result = push_sentinel_to_last_row(ws)
        assert result == 3
        assert ws.cell(3, 1).value == "< End of Document >"

    def test_sentinel_in_middle_pushed_to_end(self):
        """sentinel이 R3에 있고 R4~R6에 데이터 → sentinel을 R7로 이동."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "data1"
        ws.cell(2, 1).value = "data2"
        ws.cell(3, 1).value = "< End of Document >"
        ws.cell(4, 1).value = "data4"
        ws.cell(5, 1).value = "data5"
        ws.cell(6, 1).value = "data6"
        result = push_sentinel_to_last_row(ws)
        assert result == 7
        assert ws.cell(7, 1).value == "< End of Document >"
        assert ws.cell(3, 1).value is None

    def test_no_sentinel_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "data1"
        assert push_sentinel_to_last_row(ws) is None


# ---------------------------------------------------------------------------
# update_cross_refs_after_row_expansion — 라운드 76 T1101
# ---------------------------------------------------------------------------

class TestUpdateCrossRefs:
    """양식 cross-ref formula `=E25` → `=E{new}` 동적 갱신."""

    def test_simple_update(self):
        """=E25 → =E396 (단순 갱신)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["E5"] = "=E25"
        ws["F5"] = "=H25"
        n = update_cross_refs_after_row_expansion(
            ws, old_totals_row=25, new_totals_row=396,
        )
        assert n == 2
        assert ws["E5"].value == "=E396"
        assert ws["F5"].value == "=H396"

    def test_multiple_cols_in_row(self):
        """=B25, =E25, =H25, =I25, =L25, =M25 한 row에 다중 col."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B5"] = "=B25"
        ws["E5"] = "=E25"
        ws["H5"] = "=H25"
        ws["I5"] = "=I25"
        ws["L5"] = "=L25"
        ws["M5"] = "=M25"
        n = update_cross_refs_after_row_expansion(
            ws, old_totals_row=25, new_totals_row=100,
        )
        assert n == 6
        assert ws["B5"].value == "=B100"
        assert ws["M5"].value == "=M100"

    def test_cross_sheet_ref_skipped(self):
        """`'2.Traceability'!H9` cross-sheet ref은 변경 안 함."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["E5"] = "='2.Traceability'!H25"
        ws["F5"] = "=H25"  # 같은 시트 cell — 갱신 대상
        n = update_cross_refs_after_row_expansion(
            ws, old_totals_row=25, new_totals_row=396,
        )
        assert n == 1
        assert ws["E5"].value == "='2.Traceability'!H25"  # 변경 없음
        assert ws["F5"].value == "=H396"  # 변경

    def test_calculated_formula_unaffected(self):
        """`=(E5-F5)/E5` 같은 calculated formula는 영향 없음 (R{old} 패턴 미매칭)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["G5"] = "=(E5-F5)/E5"
        ws["E5"] = "=E25"
        n = update_cross_refs_after_row_expansion(
            ws, old_totals_row=25, new_totals_row=396,
        )
        assert n == 1
        assert ws["G5"].value == "=(E5-F5)/E5"  # 영향 없음
        assert ws["E5"].value == "=E396"

    def test_idempotent_re_invocation(self):
        """이미 new_totals_row 참조하는 cell은 2번째 호출 시 변경 0 (idempotent)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["E5"] = "=E25"
        n1 = update_cross_refs_after_row_expansion(
            ws, old_totals_row=25, new_totals_row=396,
        )
        assert n1 == 1
        # 2번째 호출 — old_totals_row=25는 이제 없으므로 변경 0
        n2 = update_cross_refs_after_row_expansion(
            ws, old_totals_row=25, new_totals_row=500,
        )
        assert n2 == 0
        assert ws["E5"].value == "=E396"  # 첫 갱신값 유지
