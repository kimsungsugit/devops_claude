from __future__ import annotations

import io

import openpyxl

from backend.services.swut_comprehensive_aggregator import (
    SwutcrBuildMeta,
    _coverage_failures,
    build_swutcr_report,
)
from backend.services.swut_input_adapter import (
    CoverageStats,
    EnvironmentData,
    ExecutionRow,
    FunctionCoverage,
    SwUTSession,
)


def _minimal_swutcr_template() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Cover")
    for row, label in enumerate((
        "Project", "ASIL Level", "Status", "Validation Date", "Author",
        "Reviewer", "Approver", "Doc. ID", "Version", "Build Timestamp",
    ), start=1):
        cover.cell(row, 1).value = label

    summary = wb.create_sheet("Test Summary")
    for row, label in enumerate((
        "Project Name", "Release Name(SW)", "Test Target Version(HW)",
        "Test Date", "Test Engineer", "Target Coverage", "Actual Coverage",
        "Target Pass ratio", "Actual Pass ratio", "Final Test Result",
    ), start=1):
        summary.cell(row, 1).value = label

    wb.create_sheet("1.Traceability")
    wb.create_sheet("2.Consistency")
    coverage = wb.create_sheet("3.Coverage")
    coverage["A1"] = "Statement Coverage"
    coverage["A6"] = "Unit ID"

    deviation = wb.create_sheet("Deviation")
    deviation["B1"] = "Test Case ID"

    test_log = wb.create_sheet("Test Log")
    test_log["B1"] = "Test Case ID"
    test_log["C1"] = "Description"
    test_log["D1"] = "Generation Method"
    test_log["F1"] = "Input"
    test_log["P1"] = "Expected Result"
    test_log["Z1"] = "Actual Result"
    test_log["AJ1"] = "Pass/Fail Unit"
    test_log["AK1"] = "Pass/Fail Total"
    test_log["AL1"] = "Log Data"

    wb.create_sheet("History")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _minimal_swutcr_specific_template() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cover = wb.create_sheet("Cover")
    cover["A1"] = "Project"
    wb.create_sheet("History")
    wb.create_sheet("Summary")
    wb.create_sheet("1.UT101")
    wb.create_sheet("2.UT201")
    wb.create_sheet("3.UT301")
    wb.create_sheet("21.IT801")
    wb.create_sheet("통합검증_BTB")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _session() -> SwUTSession:
    env = EnvironmentData(
        env_name="SWTE_01",
        component_name="CompA",
        test_cases={"SwUFn_0001.001": [object()]},
        test_results={
            "SwUFn_0001.001": ExecutionRow(
                tc_name="SwUFn_0001.001",
                passed=True,
                actual_result={"Input[0]": ("0x01", "0x01")},
            ),
        },
        function_coverage=[
            FunctionCoverage(unit_id="SwUFn_0001", name="FunctionA"),
        ],
    )
    return SwUTSession(environments=[env])


def test_coverage_failures_draft_reason_from_c_source_default_branch():
    failures = _coverage_failures(
        [
            FunctionCoverage(
                unit_id="SwUFn_0001",
                name="FunctionA",
                branch=CoverageStats(covered=1, total=2),
            )
        ],
        {
            "FunctionA": {
                "name": "FunctionA",
                "file": "SysSafety.c",
                "signature": "void FunctionA(void)",
                "body": "switch (mode) { default: return E_NOT_OK; }",
                "calls": [],
                "used_globals": [],
            }
        },
    )

    assert len(failures) == 1
    assert "switch/default defensive logic" in failures[0]["reason"]
    assert "C code evidence:" in failures[0]["reason"]
    assert "negative/default-path TC" in failures[0]["action"]
    assert "Review basis:" in failures[0]["action"]
    assert "Full C function:" in failures[0]["action"]
    assert "void FunctionA(void)" in failures[0]["action"]
    assert "default:" in failures[0]["c_evidence"]


def test_coverage_failures_classifies_register_self_test_not_range():
    failures = _coverage_failures(
        [
            FunctionCoverage(
                unit_id="SwUFn_0002",
                name="u8s_Register_Test",
                branch=CoverageStats(covered=3, total=5),
            )
        ],
        {
            "u8s_Register_Test": {
                "name": "u8s_Register_Test",
                "file": "SysCtrl_Main_PDS.c",
                "body": "ECCIE &= ~u8g_MAX; if (u8t_Register_InitialValue == u8g_CLR) { return OK; }",
                "calls": [],
                "used_globals": [],
            }
        },
    )

    assert "hardware register self-test" in failures[0]["reason"]
    assert "range or boundary" not in failures[0]["reason"]


def test_coverage_failures_classifies_interpolation_guard():
    failures = _coverage_failures(
        [
            FunctionCoverage(
                unit_id="SwUFn_0003",
                name="s16s_ApiIn_InterpolateTemperature",
                statement=CoverageStats(covered=8, total=9),
            )
        ],
        {
            "s16s_ApiIn_InterpolateTemperature": {
                "name": "s16s_ApiIn_InterpolateTemperature",
                "file": "ApiIn_Main_PDS.c",
                "body": "if ((s32t_Adc2 - s32t_Adc1) == (S32)0) { s32t_TempInterpolated = s32t_Temp1; }",
                "calls": [],
                "used_globals": [],
            }
        },
    )

    assert "interpolation or divide-by-zero guard" in failures[0]["reason"]
    assert "lookup-table boundary/interpolation" in failures[0]["action"]


def test_build_swutcr_preserves_template_and_writes_summary():
    result = build_swutcr_report(
        _session(),
        SwutcrBuildMeta(
            project_id="KJPDS02",
            project_full_name="KJPDS02",
            release_sw_version="1.01",
            test_date="2025-12-05",
            test_engineer="JK Kim",
            doc_filename_pattern=(
                "(KJPDS02_DV_SwUTCR) Software Unit Test Comprehensive Result_"
                "v{version}_{date}_R.xlsm"
            ),
        ),
        _minimal_swutcr_template(),
        swuds_function_ids={"SwUFn_0001"},
    )

    assert result.ok is True
    assert result.filename == (
        "(KJPDS02_DV_SwUTCR) Software Unit Test Comprehensive Result_v1.01_251205_R.xlsm"
    )
    assert result.summary["total_tcs"] == 1
    assert result.summary["passed"] == 1
    assert result.result_size_bytes > 0

    wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
    assert "Cover" in wb.sheetnames
    assert "Test Summary" in wb.sheetnames
    assert "AuditLog" in wb.sheetnames
    assert wb["Test Summary"]["B2"].value == "1.01"


def test_build_swutcr_specific_template_uses_project_config_and_no_generic_warnings():
    result = build_swutcr_report(
        _session(),
        SwutcrBuildMeta(
            project_id="KJPDS02",
            project_full_name="KJPDS02",
            release_sw_version="1.01",
            test_date="2025-12-05",
            test_engineer="",
            project_config={
                "swutcr_metadata": {
                    "project": "KJPDS02",
                    "phase": "DV",
                    "software_platform_ver": "25A1",
                    "product": "PDS",
                    "verification_target": "MCU",
                    "asil": "A",
                    "compiler": "\tCodeWarrior HC12Z ",
                    "mcu": "\tNXP S12ZVMC",
                    "test_iteration": "0.1",
                    "tester": "주희영",
                    "debugger": "이재원, 유영규",
                    "prepare_hours": 10,
                    "execution_hours": 10,
                    "review_hours": 10,
                    "tool_name": "VectorCAST",
                    "tool_version": "2025.sp3",
                    "excluded_scope": "LIN",
                    "excluded_size": 65826,
                    "qualified_function_total": 570,
                    "fault_injection_total": 402,
                    "fault_injection_passed": 402,
                    "reference_document": "SwTP",
                    "reference_id": "SwUTE_01",
                    "ut101_enabled": "O",
                    "ut201_enabled": "O",
                    "ut301_enabled": "X",
                    "heap_memory_leak_count": 0,
                    "heap_access_violation_count": 0,
                },
            },
        ),
        _minimal_swutcr_specific_template(),
    )

    assert result.ok is True
    assert "Traceability" not in result.incomplete_sheets
    assert not any("Traceability sheet not found" in w for w in result.warnings)
    assert not any("Test Log sheet not found" in w for w in result.warnings)
    assert result.summary["swutcr_qualified_function_count"] == 570
    assert result.summary["swutcr_raw_function_count"] == 1

    wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
    assert "AuditLog" not in wb.sheetnames
    assert wb["Summary"]["E5"].value == "25A1"
    assert wb["Summary"]["E6"].value == "PDS"
    assert wb["Summary"]["E9"].value == "\tCodeWarrior HC12Z "
    assert wb["1.UT101"]["C6"].value == "주희영"
    # ⚠ 같은 '총 TC 수' 열이지만 **행마다 세는 대상이 다르다.** 두 축을 함께 고정한다 —
    #   한쪽만 단언하면 둘이 뒤바뀌어도 통과한다(예전엔 F76 만 봤고, 거기 함수 수가
    #   들어가 있었다).
    #   73행 = CVG/Statement → 함수 수(qualified_function_total 오버라이드 = 570)
    #   76행 = Test Case(P/F) → **TC 수** (이 세션의 실행 TC 1건)
    assert wb["1.UT101"]["F73"].value == 570, "CVG 행 분모는 함수 수여야 한다"
    assert wb["1.UT101"]["F76"].value == 1, "Test Case(P/F) 행 분모는 TC 수여야 한다"
    assert wb["1.UT101"]["F77"].value == 570, "요구사항 추적성 행은 단위레벨=함수 수"
    assert wb["2.UT201"]["C85"].value == 570
    assert wb["2.UT201"]["E85"].value == 402
    assert wb["2.UT201"]["F85"].value == 402
    assert wb["3.UT301"]["C6"].value == "주희영"
    assert wb["3.UT301"]["C80"].value == "N/A"
    assert wb["21.IT801"]["C6"].value == "주희영"
    assert wb["21.IT801"]["E45"].value == 0
    assert wb["21.IT801"]["E46"].value == 0
    assert wb["통합검증_BTB"]["D4"].value == "KJPDS02"
    assert wb["통합검증_BTB"]["C37"].value == "N/A"


# ---------------------------------------------------------------------------
# 라운드 96-final QA fix — C-5 (UT201 FI 마킹) + W-6 3건 (라벨행 보존)
# ---------------------------------------------------------------------------

_USER_INPUT_YELLOW = "FFFFEB9C"


def _meta_96final() -> SwutcrBuildMeta:
    return SwutcrBuildMeta(
        project_id="KJPDS02",
        project_full_name="KJPDS02",
        release_sw_version="1.01",
        test_date="2025-12-05",
        test_engineer="주희영",
    )


class TestUt201FaultInjection96Final:
    """C-5 — Fault Injection 실측 config 키 부재 시 노란 사용자입력 마킹.

    이전 fallback(total=함수 수, passed=failed==0이면 total)은 FI 실측 부재 시
    측정한 것처럼 보이는 수치(예: 1014/1014)를 무표식 기입 — 24차 silent N/A
    제거 정책 위반 (fabrication). 키 존재 시에는 기존대로 실측 stamp.
    """

    def test_fi_keys_absent_marks_user_input_and_warns(self):
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        cfg = {"swutcr_metadata": {"tool_name": "VectorCAST"}}
        _write_ut201(ws, _meta_96final(), {"swutcr_qualified_function_count": 570},
                     cfg, warns)
        assert ws.cell(85, 3).value == 570  # 함수 수는 정상 stamp
        # E85/F85/C90 — placeholder 텍스트 + 노란 fill
        for row, col in ((85, 5), (85, 6), (90, 3)):
            assert str(ws.cell(row, col).value or "").startswith("▶"), (
                f"({row},{col}) placeholder 미기입: {ws.cell(row, col).value!r}"
            )
            assert ws.cell(row, col).fill.start_color.rgb == _USER_INPUT_YELLOW
        # 파생 수식(G85/H85)은 미기입 — placeholder operand #VALUE! 차단
        assert ws.cell(85, 7).value is None
        assert ws.cell(85, 8).value is None
        assert any("fault injection 실측 미제공" in w for w in warns)

    def test_fi_keys_present_stamps_measured_values(self):
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        cfg = {"swutcr_metadata": {
            "fault_injection_total": 402,
            "fault_injection_passed": 400,
        }}
        _write_ut201(ws, _meta_96final(), {}, cfg, warns)
        assert ws.cell(85, 5).value == 402
        assert ws.cell(85, 6).value == 400
        assert ws.cell(85, 7).value == "=E85-F85"
        assert ws.cell(90, 3).value == "Fail TC 2건"
        assert not any("fault injection" in w for w in warns)

    def test_fi_all_passed_writes_none_note(self):
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        cfg = {"swutcr_metadata": {
            "fault_injection_total": 402,
            "fault_injection_passed": 402,
        }}
        _write_ut201(ws, _meta_96final(), {}, cfg, [])
        assert str(ws.cell(90, 3).value).strip() == "해당 사항 없음"


class TestUt201FaultInjectionR105Branches:
    """라운드 105 — UT201 FI 분기 가드 (config 우선 / 부재 시 노란 마킹).

    라운드 107(2026-06-12)에서 사용자 규칙 확정으로 spec 산출 분기가 구현됨
    (FI = Test Method 'FI' 포함 TC 블록 수 — DV 405/1,598·PV 808 ground truth
    재현, ``TestUt201SpecFiAutoDerivation`` 참조). DV 감사본 수기 402는 전
    가용 산출물로 재현 불가한 stale 카운트로 판명. 여기서는 기존 두 분기
    (config 실측 stamp / 부재+spec 부재 시 사용자입력 마킹)의 경계를 가드한다
    — agg 카운트 기반 fabrication 재도입 금지는 여전히 유효.
    """

    def test_fi_partial_keys_total_only_marks_passed_and_skips_derived(self):
        """total만 있고 passed 부재 — 있는 키는 stamp, 없는 키만 마킹.

        파생 수식(G85/H85)·C90 비고는 절반 입력 기반 stamp 금지 (placeholder
        operand #VALUE! 차단), warning은 누락 키만 명시.
        """
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        cfg = {"swutcr_metadata": {"fault_injection_total": 402}}
        _write_ut201(ws, _meta_96final(), {}, cfg, warns)
        assert ws.cell(85, 5).value == 402                     # 실측 있는 키 stamp
        assert str(ws.cell(85, 6).value or "").startswith("▶")  # 부재 키 마킹
        assert ws.cell(85, 6).fill.start_color.rgb == _USER_INPUT_YELLOW
        assert ws.cell(85, 7).value is None                    # 파생 수식 미기입
        assert ws.cell(85, 8).value is None
        assert str(ws.cell(90, 3).value or "").startswith("▶")
        assert any("fault_injection_passed" in w for w in warns)
        assert not any("fault_injection_total" in w for w in warns)

    def test_fi_empty_string_values_treated_as_absent(self):
        """config 키가 빈 문자열이면 부재와 동일 — 노란 마킹 + warning.

        빈 값을 stamp하면 측정한 것처럼 보이는 공란이 됨 (silent N/A 정책 위반).
        """
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        cfg = {"swutcr_metadata": {
            "fault_injection_total": "",
            "fault_injection_passed": "",
        }}
        _write_ut201(ws, _meta_96final(), {}, cfg, warns)
        for row, col in ((85, 5), (85, 6), (90, 3)):
            assert str(ws.cell(row, col).value or "").startswith("▶")
            assert ws.cell(row, col).fill.start_color.rgb == _USER_INPUT_YELLOW
        assert ws.cell(85, 7).value is None
        assert ws.cell(85, 8).value is None
        assert any("fault injection 실측 미제공" in w for w in warns)

    def test_fi_no_spec_derivation_even_with_rich_agg(self):
        """fabrication 가드 — agg 함수/패스 카운트 기반 FI 자동 산출 금지.

        함수 수 기반 fallback(과거 fabrication) 재도입 차단: config 키 부재
        + spec 산출값(spec_fi_auto) 미전달이면 무조건 사용자입력 마킹.
        spec 기반 자동 산출은 확정 규칙(2026-06-12) 경로(spec_fi_auto 전달)
        에서만 허용 — ``TestUt201SpecFiAutoDerivation`` 참조.
        """
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        rich_agg = {
            "swutcr_qualified_function_count": 1005,
            "total": 6880, "tested": 6880, "passed": 6880, "failed": 0,
        }
        _write_ut201(ws, _meta_96final(), rich_agg, {"swutcr_metadata": {}}, warns)
        assert ws.cell(85, 3).value == 1005          # 함수 수는 정상 stamp
        for row, col in ((85, 5), (85, 6)):
            v = ws.cell(row, col).value
            assert not isinstance(v, int), f"FI 자동 산출 금지 위반: ({row},{col})={v!r}"
            assert str(v or "").startswith("▶")
        assert any("실측 미제공" in w for w in warns)


class TestUt201SpecFiAutoDerivation:
    """라운드 107 — UT201 FI spec 자동 산출 (확정 규칙 2026-06-12).

    규칙: FI total = spec Test Method 'FI' 포함 TC 블록 수, passed = FI
    iteration 전부 실행+Pass 블록 수. 우선순위: config 키 > spec 산출 >
    노란 마킹. ground truth(DV 405/1,598·PV 808)는 spec 추출기 테스트
    (test_swut_sutr_spec_builder.TestSpecFiExtraction)에서 가드.
    """

    @staticmethod
    def _spec_fi_auto(**over):
        base = {
            "total": 405, "passed": 403,
            "not_executed_blocks": 2, "failed_blocks": 0,
            "iteration_total": 1598,
            "spec_filename": "bk_SwUTS_v0.11_251126.xlsm",
            "rule": "Test Method 'FI' 포함 TC 블록 수",
            # 라운드 108 MINOR-6 — _cross_spec_fi_with_session 출력 미러:
            # 실제 _write_ut201 stamp 분기에서만 True로 갱신된다.
            "stamped": False,
        }
        base.update(over)
        return base

    def test_spec_auto_stamps_when_config_absent(self):
        """config 키 둘 다 부재 + spec 산출값 → E85/F85 + 파생 셀 stamp.

        warning 필수 — 규칙·spec 파일명·total/passed·미실행/fail 명기
        (DV 감사본 수기 402와의 차이 추적 가능성 — 실측 위장 금지).
        라운드 108 MAJOR-1 — C90 비고는 미통과 breakdown 명기 (미실행 블록을
        'Fail TC N건'으로 위장 금지 — 구 단언 동기 수정). INFO-7 — warning의
        passed=M 뒤 공백 1자.
        """
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        spec_fi_auto = self._spec_fi_auto()
        _write_ut201(ws, _meta_96final(), {}, {"swutcr_metadata": {}}, warns,
                     spec_fi_auto)
        assert ws.cell(85, 5).value == 405
        assert ws.cell(85, 6).value == 403
        # 파생 셀 — config-제공 분기와 동일.
        assert ws.cell(85, 7).value == "=E85-F85"
        assert ws.cell(85, 8).value == "=E86"
        # MAJOR-1 — 미통과 2 = 미실행 2 + Fail 0 (Fail 위장 금지 breakdown).
        assert ws.cell(90, 3).value == "미통과 2건(미실행 2, Fail 0)"
        fi_warns = [w for w in warns if "UT201 FI spec 자동 산출" in w]
        assert len(fi_warns) == 1
        assert "규칙: Test Method 'FI' 포함 TC 블록 수" in fi_warns[0]
        assert "spec=bk_SwUTS_v0.11_251126.xlsm" in fi_warns[0]
        # INFO-7 — 'passed=M (미실행 K, fail J)' 공백 포함 형식.
        assert "total=405/passed=403 (미실행 2, fail 0)" in fi_warns[0]
        # 실측 미제공 마킹 warning은 없어야 함.
        assert not any("실측 미제공" in w for w in warns)
        # MINOR-6 — 실제 stamp됐으므로 True.
        assert spec_fi_auto["stamped"] is True

    def test_spec_auto_failed_blocks_in_breakdown_note(self):
        """라운드 108 MAJOR-1 — Fail 블록 포함 시 breakdown에 Fail 수 명기."""
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        _write_ut201(ws, _meta_96final(), {}, {"swutcr_metadata": {}}, [],
                     self._spec_fi_auto(total=405, passed=400,
                                        not_executed_blocks=2,
                                        failed_blocks=3))
        assert ws.cell(90, 3).value == "미통과 5건(미실행 2, Fail 3)"

    def test_spec_auto_all_passed_writes_none_note(self):
        """전 블록 Pass — C90 '해당 사항 없음' (config 분기와 동일 비고)."""
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        _write_ut201(ws, _meta_96final(), {}, {"swutcr_metadata": {}}, [],
                     self._spec_fi_auto(total=808, passed=808,
                                        not_executed_blocks=0))
        assert ws.cell(85, 5).value == 808
        assert ws.cell(85, 6).value == 808
        assert str(ws.cell(90, 3).value).strip() == "해당 사항 없음"

    def test_config_keys_win_over_spec_auto(self):
        """config 키 존재 시 spec 산출값 무시 — 기존 동작 그대로 (단일 진리원)."""
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        cfg = {"swutcr_metadata": {
            "fault_injection_total": 402, "fault_injection_passed": 402,
        }}
        spec_fi_auto = self._spec_fi_auto()
        _write_ut201(ws, _meta_96final(), {}, cfg, warns, spec_fi_auto)
        assert ws.cell(85, 5).value == 402
        assert ws.cell(85, 6).value == 402
        assert not any("UT201 FI spec 자동 산출" in w for w in warns)
        # 라운드 108 MINOR-6 — config 우선이라 spec 산출 stamp 미발생.
        assert spec_fi_auto["stamped"] is False

    def test_config_partial_key_wins_over_spec_auto(self):
        """config 키 하나라도 있으면 기존 partial 동작 — spec 산출 미적용."""
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        cfg = {"swutcr_metadata": {"fault_injection_total": 402}}
        _write_ut201(ws, _meta_96final(), {}, cfg, warns, self._spec_fi_auto())
        assert ws.cell(85, 5).value == 402
        assert str(ws.cell(85, 6).value or "").startswith("▶")
        assert not any("UT201 FI spec 자동 산출" in w for w in warns)
        assert any("fault_injection_passed" in w for w in warns)

    def test_spec_auto_non_numeric_falls_back_to_marking(self):
        """spec 산출값 비정수 — stamp 금지 + 노란 마킹 유지 (정직 폴백)."""
        from backend.services.swut_comprehensive_aggregator import _write_ut201
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        spec_fi_auto = self._spec_fi_auto(total=None, passed=None)
        _write_ut201(ws, _meta_96final(), {}, {"swutcr_metadata": {}}, warns,
                     spec_fi_auto)
        for row, col in ((85, 5), (85, 6), (90, 3)):
            assert str(ws.cell(row, col).value or "").startswith("▶")
        assert any("비정수" in w for w in warns)
        assert any("실측 미제공" in w for w in warns)
        # 라운드 108 MINOR-6 — stamp skip이므로 False 유지.
        assert spec_fi_auto["stamped"] is False

    def test_cross_spec_fi_with_session_classifies_blocks(self):
        """교차 분류 — 전부 Pass / 일부 미실행 / 일부 Fail / 전 블록 미매칭."""
        from backend.services.swut_comprehensive_aggregator import (
            _cross_spec_fi_with_session,
        )
        env = EnvironmentData(env_name="SwUT_01", component_name="T")
        results = {
            "SwUFn_0101.002": True, "SwUFn_0101.003": True,  # 101: 전부 Pass
            "SwUFn_0102.001": True,                          # 102: .002 미실행
            "SwUFn_0103.001": False,                         # 103: Fail
            # 104: 실행 기록 자체 없음 (블록 미매칭)
        }
        for tc_name, passed in results.items():
            env.test_cases[tc_name] = []
            env.test_results[tc_name] = ExecutionRow(tc_name=tc_name, passed=passed)
        session = SwUTSession(environments=[env])
        spec_fi = {
            "fi_block_total": 4,
            "fi_iteration_total": 6,
            "fi_block_keys": {"101", "102", "103", "104"},
            "fi_iter_rows_per_block": {
                "101": [6, 7], "102": [10, 11], "103": [14], "104": [17],
            },
            "fi_iters_per_block": {
                "101": [2, 3], "102": [1, 2], "103": [1], "104": [1],
            },
            "spec_filename": "spec.xlsm",
        }
        out = _cross_spec_fi_with_session(spec_fi, session)
        assert out["total"] == 4
        assert out["passed"] == 1            # 101만
        assert out["not_executed_blocks"] == 2  # 102(.002), 104(전체)
        assert out["failed_blocks"] == 1     # 103
        assert out["iteration_total"] == 6
        assert out["spec_filename"] == "spec.xlsm"

    def test_cross_unparseable_iter_index_counts_as_not_executed(self):
        """iteration index 파싱 불가(rows>iters) — 검증 불가는 미실행 취급."""
        from backend.services.swut_comprehensive_aggregator import (
            _cross_spec_fi_with_session,
        )
        env = EnvironmentData(env_name="SwUT_01", component_name="T")
        env.test_cases["SwUFn_0101.001"] = []
        env.test_results["SwUFn_0101.001"] = ExecutionRow(
            tc_name="SwUFn_0101.001", passed=True,
        )
        session = SwUTSession(environments=[env])
        spec_fi = {
            "fi_block_total": 1,
            "fi_iteration_total": 2,
            "fi_block_keys": {"101"},
            "fi_iter_rows_per_block": {"101": [6, 7]},  # 행 2개
            "fi_iters_per_block": {"101": [1]},         # index 1개만 파싱됨
        }
        out = _cross_spec_fi_with_session(spec_fi, session)
        assert out["passed"] == 0
        assert out["not_executed_blocks"] == 1

    def test_cross_env_conflict_fail_wins_over_pass(self):
        """라운드 108 MAJOR-2 — env1 Fail + env2 Pass 동일 iteration → Fail 고정.

        _build_fn_iteration_map(SUTR 공용, last-wins)은 불변 — 교차 전용
        env 전수 관측으로 'Fail 있으면 미통과' 확정 규칙 위반을 차단 +
        상충 warning emit.
        """
        from backend.services.swut_comprehensive_aggregator import (
            _cross_spec_fi_with_session,
        )
        env1 = EnvironmentData(env_name="SwUT_01", component_name="T")
        env1.test_cases["SwUFn_0101.001"] = []
        env1.test_results["SwUFn_0101.001"] = ExecutionRow(
            tc_name="SwUFn_0101.001", passed=False,  # env1 Fail
        )
        env2 = EnvironmentData(env_name="SwUT_02", component_name="T")
        env2.test_cases["SwUFn_0101.001"] = []
        env2.test_results["SwUFn_0101.001"] = ExecutionRow(
            tc_name="SwUFn_0101.001", passed=True,   # env2 Pass (마지막 env)
        )
        session = SwUTSession(environments=[env1, env2])
        spec_fi = {
            "fi_block_total": 1,
            "fi_iteration_total": 1,
            "fi_block_keys": {"101"},
            "fi_iter_rows_per_block": {"101": [6]},
            "fi_iters_per_block": {"101": [1]},
        }
        warns: list[str] = []
        out = _cross_spec_fi_with_session(spec_fi, session, warns)
        assert out["passed"] == 0          # last-wins였다면 1 (규칙 위반)
        assert out["failed_blocks"] == 1
        assert out["not_executed_blocks"] == 0
        conflict_warns = [w for w in warns if "상충" in w]
        assert len(conflict_warns) == 1
        assert "1건" in conflict_warns[0]
        assert "SwUFn_101.001" in conflict_warns[0]
        assert "Fail" in conflict_warns[0]

    def test_cross_env_conflict_fail_wins_reversed_order(self):
        """라운드 108 MAJOR-2 — env1 Pass + env2 Fail (역순)도 동일 Fail 고정."""
        from backend.services.swut_comprehensive_aggregator import (
            _cross_spec_fi_with_session,
        )
        env1 = EnvironmentData(env_name="SwUT_01", component_name="T")
        env1.test_cases["SwUFn_0101.001"] = []
        env1.test_results["SwUFn_0101.001"] = ExecutionRow(
            tc_name="SwUFn_0101.001", passed=True,
        )
        env2 = EnvironmentData(env_name="SwUT_02", component_name="T")
        env2.test_cases["SwUFn_0101.001"] = []
        env2.test_results["SwUFn_0101.001"] = ExecutionRow(
            tc_name="SwUFn_0101.001", passed=False,
        )
        session = SwUTSession(environments=[env1, env2])
        spec_fi = {
            "fi_block_total": 1,
            "fi_iteration_total": 1,
            "fi_block_keys": {"101"},
            "fi_iter_rows_per_block": {"101": [6]},
            "fi_iters_per_block": {"101": [1]},
        }
        out = _cross_spec_fi_with_session(spec_fi, session)
        assert out["passed"] == 0
        assert out["failed_blocks"] == 1

    def test_cross_multi_env_consistent_pass_no_conflict_warning(self):
        """라운드 108 MAJOR-2 가드 — env 간 결과 일치(Pass/Pass)면 warning 없음."""
        from backend.services.swut_comprehensive_aggregator import (
            _cross_spec_fi_with_session,
        )
        envs = []
        for name in ("SwUT_01", "SwUT_02"):
            env = EnvironmentData(env_name=name, component_name="T")
            env.test_cases["SwUFn_0101.001"] = []
            env.test_results["SwUFn_0101.001"] = ExecutionRow(
                tc_name="SwUFn_0101.001", passed=True,
            )
            envs.append(env)
        session = SwUTSession(environments=envs)
        spec_fi = {
            "fi_block_total": 1,
            "fi_iteration_total": 1,
            "fi_block_keys": {"101"},
            "fi_iter_rows_per_block": {"101": [6]},
            "fi_iters_per_block": {"101": [1]},
        }
        warns: list[str] = []
        out = _cross_spec_fi_with_session(spec_fi, session, warns)
        assert out["passed"] == 1
        assert not any("상충" in w for w in warns)

    def test_cross_zero_verifiable_iterations_not_executed(self):
        """라운드 108 MINOR-3 — 검증 가능 FI iteration 0개 블록 공허 통과 차단.

        FI 세그먼트가 anchor만 덮으면 rows=[]·iters=[] — 구 구현은
        any_missing=False·any_fail=False로 passed 집계(index 파싱불가의
        보수적 미실행 취급과 비대칭). 미실행 분류 + 블록 수 warning.
        """
        from backend.services.swut_comprehensive_aggregator import (
            _cross_spec_fi_with_session,
        )
        env = EnvironmentData(env_name="SwUT_01", component_name="T")
        env.test_cases["SwUFn_0101.001"] = []
        env.test_results["SwUFn_0101.001"] = ExecutionRow(
            tc_name="SwUFn_0101.001", passed=True,
        )
        session = SwUTSession(environments=[env])
        spec_fi = {
            "fi_block_total": 2,
            "fi_iteration_total": 1,
            "fi_block_keys": {"101", "102"},
            "fi_iter_rows_per_block": {"101": [6], "102": []},  # 102: anchor-only
            "fi_iters_per_block": {"101": [1], "102": []},
        }
        warns: list[str] = []
        out = _cross_spec_fi_with_session(spec_fi, session, warns)
        assert out["passed"] == 1            # 101만 — 102 공허 통과 금지
        assert out["not_executed_blocks"] == 1
        zero_warns = [w for w in warns if "검증 가능 FI iteration 0개" in w]
        assert len(zero_warns) == 1
        assert "1건" in zero_warns[0]

    def test_cross_dup_key_blocks_not_executed(self):
        """라운드 108 MINOR-4 — 중복 TC_ID 숫자 키 블록은 보수적 미실행 처리.

        'SwUTC_0101'+'SwUTC_101' → 같은 키 '101'로 병합되어 total은 블록
        단위(2), passed는 키 단위(최대 1)로 어긋났음 — dup 키 블록 전체
        (1 + 추가 entry 수)를 미실행으로 분류해 passed 집계에서 제외.
        """
        from backend.services.swut_comprehensive_aggregator import (
            _cross_spec_fi_with_session,
        )
        env = EnvironmentData(env_name="SwUT_01", component_name="T")
        env.test_cases["SwUFn_0101.001"] = []
        env.test_results["SwUFn_0101.001"] = ExecutionRow(
            tc_name="SwUFn_0101.001", passed=True,
        )
        session = SwUTSession(environments=[env])
        spec_fi = {
            "fi_block_total": 2,             # 블록 단위 (dup 포함)
            "fi_iteration_total": 2,
            "fi_block_keys": {"101"},
            "fi_iter_rows_per_block": {"101": [6, 10]},  # 두 블록 병합 귀속
            "fi_iters_per_block": {"101": [1, 1]},
            "fi_dup_keys": ["101"],          # 추가 블록당 1 entry
        }
        warns: list[str] = []
        out = _cross_spec_fi_with_session(spec_fi, session, warns)
        assert out["total"] == 2
        assert out["passed"] == 0
        assert out["not_executed_blocks"] == 2  # 블록 단위 (1 + dup 1)
        dup_warns = [w for w in warns if "중복 TC_ID" in w]
        assert len(dup_warns) == 1
        assert "2건" in dup_warns[0]

    def test_build_swutcr_report_wires_spec_fi(self):
        """build_swutcr_report(spec_fi=...) — 교차→stamp→summary→warning 배선."""
        env = EnvironmentData(
            env_name="SWTE_01",
            component_name="CompA",
            test_cases={"SwUFn_0101.001": [object()]},
            test_results={
                "SwUFn_0101.001": ExecutionRow(
                    tc_name="SwUFn_0101.001", passed=True,
                ),
            },
            function_coverage=[
                FunctionCoverage(unit_id="SwUFn_0101", name="FunctionA"),
            ],
        )
        session = SwUTSession(environments=[env])
        spec_fi = {
            "rule": "Test Method 'FI' 포함 TC 블록 수",
            "fi_block_total": 1,
            "fi_iteration_total": 1,
            "fi_block_keys": {"101"},
            "fi_iter_rows_per_block": {"101": [6]},
            "fi_iters_per_block": {"101": [1]},
            "spec_filename": "wip_pv_SwUTS_v0.10_260608.xlsm",
        }
        result = build_swutcr_report(
            session, _meta_96final(), _minimal_swutcr_specific_template(),
            spec_fi=spec_fi,
        )
        assert result.ok is True
        assert result.summary["ut201_fi_auto"]["total"] == 1
        assert result.summary["ut201_fi_auto"]["passed"] == 1
        # 라운드 108 MINOR-6 — 실제 2.UT201 stamp 발생 시에만 True.
        assert result.summary["ut201_fi_auto"]["stamped"] is True
        wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
        assert wb["2.UT201"]["E85"].value == 1
        assert wb["2.UT201"]["F85"].value == 1
        assert wb["2.UT201"]["G85"].value == "=E85-F85"
        assert str(wb["2.UT201"]["C90"].value).strip() == "해당 사항 없음"
        fi_warns = [w for w in result.warnings if "UT201 FI spec 자동 산출" in w]
        assert len(fi_warns) == 1
        assert "spec=wip_pv_SwUTS_v0.10_260608.xlsm" in fi_warns[0]

    def test_build_swutcr_report_without_ut201_sheet_not_stamped(self):
        """라운드 108 MINOR-6 — 2.UT201 시트 전무 템플릿: stamped=False.

        구 구현은 summary['ut201_fi_auto']를 시트 존재 게이트 이전에 무조건
        설정 — stamp 없는데 산출된 것처럼 보였다. stamped 플래그는 실제
        _write_ut201 stamp 분기에서만 True.
        """
        wb = openpyxl.load_workbook(
            io.BytesIO(_minimal_swutcr_specific_template()), keep_vba=False,
        )
        del wb["2.UT201"]
        buf = io.BytesIO()
        wb.save(buf)
        env = EnvironmentData(
            env_name="SWTE_01",
            component_name="CompA",
            test_cases={"SwUFn_0101.001": [object()]},
            test_results={
                "SwUFn_0101.001": ExecutionRow(
                    tc_name="SwUFn_0101.001", passed=True,
                ),
            },
            function_coverage=[
                FunctionCoverage(unit_id="SwUFn_0101", name="FunctionA"),
            ],
        )
        session = SwUTSession(environments=[env])
        spec_fi = {
            "fi_block_total": 1,
            "fi_iteration_total": 1,
            "fi_block_keys": {"101"},
            "fi_iter_rows_per_block": {"101": [6]},
            "fi_iters_per_block": {"101": [1]},
            "spec_filename": "spec.xlsm",
        }
        result = build_swutcr_report(
            session, _meta_96final(), buf.getvalue(), spec_fi=spec_fi,
        )
        assert result.ok is True
        # 교차값은 노출되지만 stamped=False — 산출 위장 차단.
        assert result.summary["ut201_fi_auto"]["passed"] == 1
        assert result.summary["ut201_fi_auto"]["stamped"] is False
        assert "ut201" not in result.summary.get("swutcr_specific_written", {})
        assert not any(
            "UT201 FI spec 자동 산출 — 규칙" in w for w in result.warnings
        )

    def test_build_swutcr_report_without_spec_fi_keeps_marking(self):
        """spec_fi 미전달(HDPDM01 경로) — 기존 노란 마킹 유지 (무회귀)."""
        result = build_swutcr_report(
            _session(), _meta_96final(), _minimal_swutcr_specific_template(),
        )
        assert result.ok is True
        assert "ut201_fi_auto" not in result.summary
        wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
        assert str(wb["2.UT201"]["E85"].value or "").startswith("▶")
        assert str(wb["2.UT201"]["F85"].value or "").startswith("▶")
        assert any("fault injection 실측 미제공" in w for w in result.warnings)


class TestW6LabelRowFixes96Final:
    """W-6 ①②③ — 머지 라벨 헤더행 덮어쓰기 결함 fix (값은 헤더 아래 값행에).

    ① 3.UT301 r90:91 세로 머지 라벨 → 값은 92행.
    ② 21.IT801 r50 헤더('파일명' 등) → 값은 51행.
    ③ 통합검증_BTB r10 머지 라벨(C10:D11 'SW 버전'/E10:F11 'Test Period')
       → 값은 12행 (C12:D12/E12:F12 머지 anchor).
    """

    def test_ut301_w6_1_labels_preserved_values_on_row92(self):
        from backend.services.swut_comprehensive_aggregator import _write_ut301
        wb = openpyxl.Workbook()
        ws = wb.active
        labels = {3: "SW Unit(함수)", 5: "미달성 사유", 8: "대책", 12: "문장"}
        for col, text in labels.items():
            ws.cell(90, col).value = text
            ws.merge_cells(start_row=90, end_row=91, start_column=col, end_column=col)
        cfg = {"swutcr_metadata": {"ut301_enabled": "X"}}
        _write_ut301(ws, _meta_96final(), {}, cfg)
        # 라벨 보존 (이전: 91행 기입 → merge anchor redirect로 90행 라벨 덮어씀)
        for col, text in labels.items():
            assert ws.cell(90, col).value == text, f"라벨 col {col} 덮어써짐"
        # 값은 헤더 아래 첫 값행(92행)
        assert ws.cell(92, 3).value == "N/A"
        assert ws.cell(92, 5).value == "-"
        assert "Back-to-back" in str(ws.cell(92, 8).value)
        assert ws.cell(92, 12).value == "추가 B2B evidence source 제공 시 재작성"
        assert ws.cell(80, 3).value == "N/A"  # 기존 80행 stamp 불변

    def test_it801_w6_2_header_row50_preserved_values_on_row51(self):
        from backend.services.swut_comprehensive_aggregator import _write_it801
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(50, 3).value = "파일명"
        ws.cell(50, 5).value = "오류 내용"
        ws.cell(50, 7).value = "수정 여부"
        _write_it801(ws, _meta_96final(), {})
        assert ws.cell(50, 3).value == "파일명"
        assert ws.cell(50, 5).value == "오류 내용"
        assert ws.cell(50, 7).value == "수정 여부"
        assert ws.cell(51, 3).value == "N/A"
        assert ws.cell(51, 5).value == "해당 사항 없음"
        assert ws.cell(51, 7).value == "O"
        assert "mpatrol" in str(ws.cell(51, 10).value)

    def test_btb_w6_3_merged_labels_row10_preserved_values_on_row12(self):
        from backend.services.swut_comprehensive_aggregator import _write_btb_sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(10, 3).value = "SW 버전"
        ws.merge_cells("C10:D11")
        ws.cell(10, 5).value = "Test Period"
        ws.merge_cells("E10:F11")
        ws.merge_cells("C12:D12")
        ws.merge_cells("E12:F12")
        cfg = {"swutcr_metadata": {
            "test_iteration": "0.1",
            "software_platform_ver": "25A1",
            "prepare_hours": 10,
            "execution_hours": 20,
            "review_hours": 30,
        }}
        _write_btb_sheet(ws, _meta_96final(), cfg)
        # 머지 라벨 헤더(10행) 보존
        assert ws.cell(10, 3).value == "SW 버전"
        assert ws.cell(10, 5).value == "Test Period"
        # 값행(12행) — C12/E12 머지 anchor + 시간값 G12:J12
        assert ws.cell(12, 3).value == "0.1"
        assert ws.cell(12, 5).value == "25A1"
        assert ws.cell(12, 7).value == 10
        assert ws.cell(12, 8).value == 20
        assert ws.cell(12, 9).value == 30
        assert ws.cell(12, 10).value == "=SUM(G12:I12)"


# ---------------------------------------------------------------------------
# UT101 4.1 Test Results Summary — 행마다 세는 대상이 다르다
# ---------------------------------------------------------------------------
class TestUt101TestCaseRowCountsTcsNotFunctions:
    """76행('Test Case(P/F)')의 '총 TC 수' 자리에 **함수 수**가 들어가고 있었다.

    회사 양식(`(XXXX_SwUTCR) … v0.10`)의 `1.UT101` 시트 구조를 열어서 확인했다:

        72행 헤더  F='총 TC 수'  G='Pass TC 수'  H='Fail TC 수'  I='테스트 성공율'
        73행  UT / CVG / Statement                → 분모 = 함수 수(SwUFn)
        74행       CVG / Branch                   → 분모 = 함수 수
        75행       CVG / TC                       → '-' (VectorCAST 미산출)
        76행  Test Case(P/F)                      → 분모 = **TC 수**
        77·78행  요구사항 추적성/정합성 커버리지    → 단위레벨이라 분모 = 함수 수

    즉 같은 열이어도 76행만 세는 대상이 다르다. 여기에 함수 수를 쓰면 KJPDS02 실측
    기준 TC 6,882건이 1,014건으로 **5.8배 과소 보고**된다. 비율(=G/F)은 우연히 같아서
    화면상 아무 이상이 없고, ISO 26262 근거로 인용되는 것은 그 절대 수치다.

    ⚠ 이 결함을 기존 테스트가 **고정하고 있었다** — `F76 == 570`(함수 수)만 단언했다.
      그래서 아래는 두 축을 함께 본다: 한쪽만 보면 둘이 뒤바뀌어도 통과한다.
    """

    @staticmethod
    def _sess(n_cases: int, n_results: int) -> SwUTSession:
        # tc_name 하나당 항목 1개 → total_tcs 는 키 수와 같다(집계 방식에 안 흔들린다).
        cases = {f"SwUFn_0001.{i:03d}": [object()] for i in range(1, n_cases + 1)}
        results = {
            k: ExecutionRow(tc_name=k, passed=True, actual_result={})
            for k in list(cases)[:n_results]
        }
        env = EnvironmentData(
            env_name="SWTE_01",
            component_name="CompA",
            test_cases=cases,
            test_results=results,
            function_coverage=[FunctionCoverage(unit_id="SwUFn_0001", name="FunctionA")],
        )
        return SwUTSession(environments=[env])

    @staticmethod
    def _build(session: SwUTSession):
        return build_swutcr_report(
            session,
            SwutcrBuildMeta(
                project_id="KJPDS02",
                project_full_name="KJPDS02",
                release_sw_version="1.01",
                test_date="2025-12-05",
                test_engineer="",
                project_config={
                    "swutcr_metadata": {
                        "project": "KJPDS02",
                        "phase": "DV",
                        # 함수 축을 TC 축과 **다른 값**으로 고정해 둘이 섞이면 즉시 드러나게 한다.
                        "qualified_function_total": 570,
                    },
                },
            ),
            _minimal_swutcr_specific_template(),
        )

    def test_tc_row_uses_tc_count_and_cvg_rows_use_function_count(self):
        result = self._build(self._sess(n_cases=3, n_results=3))
        assert result.ok is True
        wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
        ws = wb["1.UT101"]
        assert ws["F76"].value == 3, "Test Case(P/F) 분모가 TC 수(3)가 아니다"
        assert ws["F73"].value == 570, "CVG/Statement 분모는 함수 수(570)여야 한다"
        assert ws["F74"].value == 570, "CVG/Branch 분모는 함수 수(570)여야 한다"
        assert ws["F77"].value == 570, "요구사항 추적성은 단위레벨=함수 수"

    def test_not_all_executed_warns_that_pass_formula_overstates(self):
        """양식의 Pass 수식은 `=F76-H76` 이라 **미실행을 Pass 로 센다.**

        분모를 TC 수로 고친 뒤에 남는 문제다. 값을 몰래 바꾸지 않고 경고로 드러낸다 —
        수식은 회사 양식 소유물이라 우리가 갈아끼울 대상이 아니다.
        """
        result = self._build(self._sess(n_cases=5, n_results=2))
        wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
        assert wb["1.UT101"]["F76"].value == 5
        hits = [w for w in result.warnings if "미실행 TC" in w and "Test Case(P/F)" in w]
        assert hits, f"미실행 TC 경고가 없다: {result.warnings}"
        assert "3건" in hits[0], f"미실행 건수(5-2=3)가 안 적혔다: {hits[0]}"

    def test_missing_total_stays_zero_and_never_falls_back_to_function_count(self):
        """TC 수를 못 구하면 0 으로 남기고 경고한다 — 함수 수로 메우지 않는다.

        '빈 자리를 그럴듯한 다른 수로 채우는' 것이 바로 이 결함의 정체였다.
        """
        result = self._build(self._sess(n_cases=0, n_results=0))
        wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
        f76 = wb["1.UT101"]["F76"].value
        assert f76 == 0, f"TC 수 부재인데 F76={f76!r} — 다른 수로 메웠다"
        assert f76 != 570, "함수 수로 폴백하면 안 된다"
        assert any("총 TC 수를 못 구했다" in w for w in result.warnings), \
            f"부재를 침묵으로 넘겼다: {result.warnings}"

    def test_ratio_cell_survives_zero_denominator(self):
        """F76 이 0 이어도 #DIV/0! 이 아니라 빈 칸이어야 한다.

        `#DIV/0!` 은 "못 구했다" 가 아니라 "엑셀이 깨졌다" 로 읽힌다 — 형제 행(73/74/
        77/78)이 전부 IFERROR 인 것도 같은 이유다.
        """
        result = self._build(self._sess(n_cases=0, n_results=0))
        wb = openpyxl.load_workbook(result.xlsm_io, data_only=False)
        assert wb["1.UT101"]["I76"].value == '=IFERROR(G76/F76, "")'
