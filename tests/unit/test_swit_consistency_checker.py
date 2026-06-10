"""35차 SwIT Coverage Report ↔ SITR consistency checker 회귀.

SwUT consistency_checker thin wrapper — `tc_prefix="SwITC"` 전달. 회귀는
SwIT-specific (prefix 매칭 + thin wrapper signature) 위주.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.services.swit_consistency_checker import (  # noqa: E402
    ConsistencyIssue,
    ConsistencyReport,
    check_swit_consistency,
)


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_swit_coverage_wb(
    *,
    tcs: list[str],
    functions: list[str],
    coverage_map: dict[str, list[str]],
    final_result: str = "PASS",
) -> bytes:
    """SwIT Coverage Report v2.02 (SwUT v3.01과 시트 구조 동일)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Test Summary")
    ws.cell(row=1, column=2, value="Test Summary Report")
    ws.cell(row=8, column=2, value="Final Test Result")
    ws.cell(row=8, column=3, value=final_result)

    trace = wb.create_sheet("1.Traceability")
    trace.cell(row=1, column=3, value="ID")
    for i, fn in enumerate(functions):
        trace.cell(row=1, column=4 + i, value=fn)
    for ri, tc in enumerate(tcs):
        trace.cell(row=2 + ri, column=2, value=tc)
        covered = coverage_map.get(tc, [])
        for i, fn in enumerate(functions):
            if fn in covered:
                trace.cell(row=2 + ri, column=4 + i, value="O")

    cov = wb.create_sheet("3. Coverage")
    cov.cell(row=1, column=2, value="Statement Coverage")
    cov.cell(row=3, column=2, value="Coverage")
    cov.cell(row=3, column=3, value="Total")
    cov.cell(row=3, column=4, value="Fail Count")
    cov.cell(row=3, column=5, value="Exception")
    cov.cell(row=4, column=2, value="Statement")
    cov.cell(row=4, column=3, value=len(functions))
    cov.cell(row=4, column=4, value=0)
    cov.cell(row=4, column=5, value=0)
    cov.cell(row=5, column=2, value="Branch")
    cov.cell(row=5, column=3, value=len(functions))
    cov.cell(row=5, column=4, value=0)
    cov.cell(row=5, column=5, value=0)

    return _wb_to_bytes(wb)


def _build_swit_sitr_wb(
    *,
    total: int = 50,
    tested: int = 49,
    passed: int = 45,
    failed: int = 4,
    deviated: int = 4,
    not_executed: int = 1,
    not_executed_tcs: list[str] | None = None,
    deviation_tcs: list[str] | None = None,
    final_result: str = "OK",
) -> bytes:
    """SwIT SITR v2.02 (SwUT SUTR과 시트 구조 동일)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Test Summary")
    r = 1
    ws.cell(row=r, column=2, value="Test Summary Report"); r += 1
    ws.cell(row=r, column=2, value="Final Test Result")
    ws.cell(row=r, column=3, value=final_result); r += 1

    ws.cell(row=r, column=2, value="Total Number of TCs")
    ws.cell(row=r, column=3, value="Number of TCs Tested")
    ws.cell(row=r, column=4, value="Number of TCs Passed")
    ws.cell(row=r, column=5, value="Number of TCs Failed")
    ws.cell(row=r, column=6, value="Number of Deviated TCs")
    ws.cell(row=r, column=7, value="Number of TCs not executed"); r += 1
    ws.cell(row=r, column=2, value=total)
    ws.cell(row=r, column=3, value=tested)
    ws.cell(row=r, column=4, value=passed)
    ws.cell(row=r, column=5, value=failed)
    ws.cell(row=r, column=6, value=deviated)
    ws.cell(row=r, column=7, value=not_executed); r += 1

    ws.cell(row=r, column=2, value="■ List of Test Case not Executed"); r += 1
    ws.cell(row=r, column=2, value="Test Case ID"); r += 1
    for tc in (not_executed_tcs or []):
        ws.cell(row=r, column=2, value=tc); r += 1

    ws.cell(row=r, column=2, value="■ Deviation List"); r += 1
    ws.cell(row=r, column=2, value="Test Case ID"); r += 1
    for tc in (deviation_tcs or []):
        ws.cell(row=r, column=2, value=tc); r += 1

    return _wb_to_bytes(wb)


# ---------------------------------------------------------------------------
# Re-export + thin wrapper 검증
# ---------------------------------------------------------------------------

class TestSwitConsistencyApi:
    """SwIT consistency_checker public API — SwUT와 동일 dataclass 재사용."""

    def test_consistency_issue_and_report_imported(self):
        assert ConsistencyIssue is not None
        assert ConsistencyReport is not None

    def test_check_swit_consistency_returns_consistency_report(self):
        functions = [f"SwUFn_{i:04d}" for i in range(1, 52)]
        tcs = [f"SwITC_SwUFn_{i:04d}" for i in range(1, 51)]
        cov_map = {tc: [tc.replace("SwITC_", "")] for tc in tcs}
        cov = _build_swit_coverage_wb(
            tcs=tcs, functions=functions, coverage_map=cov_map, final_result="PASS",
        )
        sitr = _build_swit_sitr_wb(
            total=51, tested=50, passed=46, failed=4, deviated=4, not_executed=1,
            not_executed_tcs=["SwITC_SwUFn_0051"],
            deviation_tcs=["SwITC_SwUFn_0001 (TC1)"],
            final_result="OK",
        )
        report = check_swit_consistency(cov, sitr)
        assert isinstance(report, ConsistencyReport)


# ---------------------------------------------------------------------------
# SwIT TC prefix matching (35차 핵심) — SwITC_SwUFn_X
# ---------------------------------------------------------------------------

class TestSwitPrefixMatching:
    """SwIT TC name prefix `SwITC_` 가 함수 ID 추출에 성공하는지."""

    def test_swit_tc_extracted_in_uncovered_matching(self):
        """미커버 SwUFn_X ↔ SITR `SwITC_SwUFn_X` 미실행 매칭 정상 동작."""
        functions = [f"SwUFn_{i:04d}" for i in range(1, 52)]
        tcs = [f"SwITC_SwUFn_{i:04d}" for i in range(1, 51)]  # SwUFn_0051 미커버
        cov_map = {tc: [tc.replace("SwITC_", "")] for tc in tcs}
        cov = _build_swit_coverage_wb(
            tcs=tcs, functions=functions, coverage_map=cov_map, final_result="PASS",
        )
        # SITR의 not_executed_tcs에 SwITC_SwUFn_0051 — Coverage 미커버와 일치
        sitr = _build_swit_sitr_wb(
            total=51, tested=50, passed=50, failed=0, deviated=0, not_executed=1,
            not_executed_tcs=["SwITC_SwUFn_0051"],
            final_result="OK",
        )
        report = check_swit_consistency(cov, sitr)
        rd = report.to_dict()
        # uncovered_mismatch issue 없어야 함 (일치)
        mismatch_issues = [
            i for i in rd["issues"] if i["category"] == "uncovered_mismatch"
        ]
        assert not mismatch_issues, (
            f"uncovered_mismatch issue 발생: {mismatch_issues} — "
            "SwIT prefix 매칭 실패 의심"
        )

    def test_swit_tc_mismatch_detected_when_sitr_missing_uncovered(self):
        """Coverage 미커버 SwUFn_X가 SITR not_executed에 없으면 warning."""
        functions = [f"SwUFn_{i:04d}" for i in range(1, 52)]
        tcs = [f"SwITC_SwUFn_{i:04d}" for i in range(1, 51)]
        cov_map = {tc: [tc.replace("SwITC_", "")] for tc in tcs}
        cov = _build_swit_coverage_wb(
            tcs=tcs, functions=functions, coverage_map=cov_map, final_result="PASS",
        )
        sitr = _build_swit_sitr_wb(
            total=51, tested=51, passed=51, failed=0, deviated=0, not_executed=0,
            not_executed_tcs=[],  # 미실행 비어있음 → mismatch
            final_result="OK",
        )
        report = check_swit_consistency(cov, sitr)
        rd = report.to_dict()
        warns = [
            i for i in rd["issues"]
            if i["category"] == "uncovered_mismatch" and i["severity"] == "warning"
        ]
        assert warns, f"warning 미발생 (uncovered SwUFn_0051): {rd['issues']}"


# ---------------------------------------------------------------------------
# SwIT report tool_qualification (재활용 확인)
# ---------------------------------------------------------------------------

class TestSwitReportToolQualification:
    """SwUT와 동일 tool_qualification 메타 (재활용 확인)."""

    def test_report_includes_tool_qualification(self):
        functions = [f"SwUFn_{i:04d}" for i in range(1, 52)]
        tcs = [f"SwITC_SwUFn_{i:04d}" for i in range(1, 51)]
        cov_map = {tc: [tc.replace("SwITC_", "")] for tc in tcs}
        cov = _build_swit_coverage_wb(
            tcs=tcs, functions=functions, coverage_map=cov_map,
        )
        sitr = _build_swit_sitr_wb(
            total=51, tested=50, passed=50, failed=0, deviated=0, not_executed=1,
            not_executed_tcs=["SwITC_SwUFn_0051"],
        )
        report = check_swit_consistency(cov, sitr)
        d = report.to_dict()
        assert "tool_qualification" in d
        tq = d["tool_qualification"]
        assert tq["evidence_class"] == "auto-generated draft"
        assert "asil_b_c_d_usage" in tq


# ---------------------------------------------------------------------------
# 라운드 96-fix W-B — KJPDS02 TC 명명 변형 일반화
# ---------------------------------------------------------------------------


class TestTcIdPatternsKJPDS02:
    """_tc_id_patterns — HDPDM01 기존형 + KJPDS02 DV/PV 명명 모두 매칭."""

    @staticmethod
    def _patterns(prefix):
        from backend.services.swut_consistency_checker import _tc_id_patterns
        return _tc_id_patterns(prefix)

    @pytest.mark.parametrize("tc_id", [
        "SwITC_SwUFn_12",          # HDPDM01 기존형
        "SwITC_0201",              # KJPDS02 DV Test Log
        "SwITC_0101_01",           # KJPDS02 DV spec
        "SwIT_SwUFn_0101_01",      # KJPDS02 PV VectorCAST (trailing C 탈락)
        "SwIT_SwUFn_0101_01.001",  # compound iteration suffix
    ])
    def test_swit_match_variants(self, tc_id):
        match_re, _ = self._patterns("SwITC")
        assert match_re.match(tc_id), tc_id

    @pytest.mark.parametrize("bad", [
        "SwUFn_0101",        # 함수 ID 자체 (TC 아님)
        "Range",             # 보조 행
        "SwITCX_0101",       # 오타 prefix
        "SwITC_",            # 숫자 없음
    ])
    def test_swit_reject_non_tc(self, bad):
        match_re, _ = self._patterns("SwITC")
        assert not match_re.match(bad), bad

    @pytest.mark.parametrize("tc_id,expected_fn", [
        ("SwITC_SwUFn_12", "SwUFn_12"),
        ("SwITC_0201", "SwUFn_0201"),
        ("SwIT_SwUFn_0101_01", "SwUFn_0101"),
        ("SwUTC_SwUFn_0034", "SwUFn_0034"),
    ])
    def test_function_id_extraction(self, tc_id, expected_fn):
        from backend.services.swut_consistency_checker import (
            _extract_function_id, _tc_id_patterns,
        )
        prefix = "SwITC" if tc_id.startswith("SwIT") else "SwUTC"
        _, fn_re = _tc_id_patterns(prefix)
        assert _extract_function_id(fn_re, tc_id) == expected_fn

    def test_swut_kjpds02_numeric_form(self):
        """KJPDS02 SwUT TC 'SwUTC_NNNN' (라운드 82 양식)도 매칭."""
        match_re, _ = self._patterns("SwUTC")
        assert match_re.match("SwUTC_0001")
        assert match_re.match("SwUTC_SwUFn_99")


class TestKjpds02TraceabilityMatrix:
    """KJPDS02 SwIT v1.01 — TC×SwST 매트릭스에서 total_tcs/total_functions 추출."""

    def _make_cov_wb_bytes(self):
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "2.Traceability"
        # 헤더: SwST 컬럼 (KJPDS02 SwIT — SwUFn 아님)
        ws.cell(11, 2, "SwDS")
        for i in range(3):
            ws.cell(11, 3 + i, f"SwST_{i+1:02d}")
        # TC rows — KJPDS02 DV 명명
        for r, tc in enumerate(["SwITC_0101_01", "SwITC_0101_02", "SwITC_0201"], start=13):
            ws.cell(r, 1, tc)
            ws.cell(r, 3, "O")
        # 3.Coverage placeholder (Exception 추출 — 없음)
        wb.create_sheet("4.Coverage")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _make_sutr_wb_bytes(self):
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Summary"
        ws.cell(1, 1, "Final Test Result")
        ws.cell(1, 2, "OK")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_total_tcs_extracted_from_swst_matrix(self):
        from backend.services.swit_consistency_checker import check_swit_consistency
        report = check_swit_consistency(
            self._make_cov_wb_bytes(), self._make_sutr_wb_bytes(),
        )
        d = report.to_dict()
        assert d["coverage_summary"]["total_tcs"] == 3
        assert d["coverage_summary"]["total_functions"] == 3
        # SwST_02/03은 O 없음 → uncovered로 식별
        assert "SwST_02" in d["coverage_summary"]["uncovered_functions"]
