# -*- coding: utf-8 -*-
"""swreport_summary_aggregator 단위 테스트.

16MB 실 ES95411 대신 동일 구조(Summary 카탈로그 + detail Result 블록)의 합성
워크북으로 roll-up 규칙을 결정론적으로 검증한다:
  - primary 수행행 stamp / 미수행(X) 공란 / sub행 P/F 승계
  - Total 행 계산 (수행수·시간합·Fail수)
  - Fail 행 빨간 강조
  - 외부링크 정화 후에도 valid xlsm
  - preview JSON 집계
"""
import io

import openpyxl
import pytest

from backend.services.design_tokens import FAIL_FILL_RGB
from backend.services.swreport_summary_aggregator import (
    SwReportBuildMeta,
    build_summary_report,
    normalize_pf,
    preview_summary_report,
    strip_id,
)


def _fill_rgb(cell) -> str:
    return str(getattr(getattr(getattr(cell, "fill", None), "fgColor", None), "rgb", "") or "").upper()


def _is_fail_fill(cell) -> bool:
    return _fill_rgb(cell).endswith(FAIL_FILL_RGB[-6:])


def _detail(wb, title, *, it, swv, tester, dbg, prep, ex, rev, total, pf):
    ws = wb.create_sheet(title)
    ws["B3"] = "■ Test Information"
    ws["E3"] = "■ 분석시간 (HR)"
    ws["H3"] = "■ Result"
    ws["B4"], ws["C4"] = "분석차수", it
    ws["B5"], ws["C5"] = "SW Ver.", swv
    ws["B6"], ws["C6"] = "Tester", tester
    ws["B7"], ws["C7"] = "Debugger", dbg
    ws["E4"], ws["F4"] = "준비", prep
    ws["E5"], ws["F5"] = "수행", ex
    ws["E6"], ws["F6"] = "검토", rev
    ws["E7"], ws["F7"] = "Total", total
    ws["H4"], ws["I4"] = "TC Pass율", 1
    ws["H5"], ws["I5"] = "커버리지", 1
    ws["H7"], ws["I7"] = "P/F", pf
    return ws


# Summary 데이터 행: (B No, E ID, F Name, G 점검, K Tool, P SheetName)
_ROWS = [
    (1, "ST201", "Code Metric", "O", "Helix QAC", "2.ST201"),   # primary 수행
    (2, "ST202", "Nesting", "O", "Helix QAC", "2.ST201"),       # sub (parent 공유)
    (21, "UT101", "Req Unit Test", "O", "VectorCAST", "21.UT101"),
    (22, "UT201", "Fault Injection", "O", "VectorCAST", "22.UT201"),
    (23, "UT301", "Back to Back", "X", "VectorCAST", "23.UT301"),  # 미수행
    (31, "IT101", "Req Integration", "O", "VectorCast", "31.IT101"),  # FAIL
]


@pytest.fixture
def fixture_bytes() -> bytes:
    wb = openpyxl.Workbook()
    summ = wb.active            # 기본 시트를 Summary로 재활용
    assert summ is not None
    summ.title = "Summary"
    summ["A1"] = "Software Test Result Summary"
    # 헤더 블록 (B 라벨 / E 값)
    for r, (lbl, val) in enumerate(
        [("Project", "X"), ("Phase", "X"), ("Software Platform Ver.", "X"),
         ("Product", "X"), ("검증 대상", "X"), ("ASIL 등급", "X"),
         ("Complier", "X"), ("MCU", "X"), ("Fail", 9), ("Result", "stale")],
        start=3,
    ):
        summ.cell(row=r, column=2, value=lbl)
        summ.cell(row=r, column=5, value=val)
    # 표 헤더 (row 15: E='ID' anchor)
    summ["B14"] = "No"
    for col, txt in [(5, "ID"), (6, "Test Name"), (7, "점검 대상"), (8, "분석 차수"),
                     (9, "SW Ver."), (10, "Tester"), (11, "Tool"), (12, "Debugger"),
                     (13, "총 분석시간"), (14, "P/F"), (15, "Note"), (16, "Sheet Name")]:
        summ.cell(row=15, column=col, value=txt)
    # 데이터 행 16~
    r = 16
    for (no, tid, name, perf, tool, sheet) in _ROWS:
        summ.cell(row=r, column=2, value=no)
        summ.cell(row=r, column=5, value=tid)
        summ.cell(row=r, column=6, value=name)
        summ.cell(row=r, column=7, value=perf)
        summ.cell(row=r, column=11, value=tool)
        summ.cell(row=r, column=16, value=sheet)
        # 결과 컬럼에 stale 값을 미리 넣어 — 빌더가 공란화/덮어쓰기 하는지 검증
        summ.cell(row=r, column=8, value="STALE")
        summ.cell(row=r, column=14, value="STALE")
        r += 1
    total_row = r
    summ.cell(row=total_row, column=2, value="Total")

    # ST201(r16)+ST202(r17) 그룹 — 실제 ES95411 구조 재현:
    #   분석차수(H)/SWVer(I)/Tester(J)/시간(M)은 parent와 세로 병합(parent에만 표시),
    #   Debugger(L)는 행별 — sub 행 meta를 빌더가 보존하는지(병합 무손상·행별 값 유지) 검증.
    for col in (8, 9, 10, 13):  # H/I/J/M
        summ.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)
    summ.cell(row=16, column=12, value="원본디버거")   # parent — stamp가 덮어씀
    summ.cell(row=17, column=12, value="유영규-sub")    # sub — 빌더가 보존해야 함

    # detail 시트 (UT301은 미수행 — 시트 없음)
    _detail(wb, "2.ST201", it="0.1", swv="25A1", tester="주희영", dbg="이재원",
            prep=3, ex=2, rev=3, total=8, pf="Pass")
    _detail(wb, "21.UT101", it="0.1", swv="25A1", tester="주희영", dbg="이재원",
            prep=10, ex=10, rev=10, total=30, pf="Pass")
    _detail(wb, "22.UT201", it="0.1", swv="25A1", tester="주희영", dbg="이재원",
            prep=10, ex=10, rev=10, total=30, pf="Pass")
    _detail(wb, "31.IT101", it="0.2", swv="25A1", tester="김진경", dbg="유영규",
            prep=5, ex=10, rev=10, total=25, pf="Fail")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── 단위 헬퍼 ────────────────────────────────────────────────────────────────
@pytest.mark.parametrize("raw,expect", [
    ("21.UT101", "UT101"), ("2. ST201", "ST201"), ("1.ST101", "ST101"),
    ("UT101", "UT101"), ("  38.IT801 ", "IT801"), ("", ""),
])
def test_strip_id(raw, expect):
    assert strip_id(raw) == expect


@pytest.mark.parametrize("raw,expect", [
    ("Pass", "Pass"), ("PASS", "Pass"), ("OK", "Pass"), ("Fail", "Fail"),
    ("FAIL", "Fail"), ("", ""), ("0", "0"),
])
def test_normalize_pf(raw, expect):
    assert normalize_pf(raw) == expect


# ── 빌드 roll-up ─────────────────────────────────────────────────────────────
def test_build_rollup(fixture_bytes):
    res = build_summary_report(fixture_bytes, [("self", fixture_bytes)],
                               SwReportBuildMeta(project_id="KJPDS02"))
    assert res.ok
    s = res.summary
    assert s["performed_count"] == 5      # ST201,ST202,UT101,UT201,IT101 (UT301=X 제외)
    assert s["matched_rows"] == 5
    assert s["fail_count"] == 1
    assert s["fail_ids"] == ["IT101"]
    assert s["total_hours"] == 93         # 8+30+30+25 (sub ST202는 시간 미합산)
    assert s["overall_result"] == "Fail"

    wb = openpyxl.load_workbook(io.BytesIO(res.xlsm_bytes), data_only=True)
    ws = wb["Summary"]

    def cell(r, c):
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    # r16 ST201 primary — stamp (병합 anchor에 기록)
    assert cell(16, 8) == "0.1"        # 분석차수
    assert cell(16, 10) == "주희영"    # Tester
    assert cell(16, 12) == "이재원"    # Debugger (stamp가 '원본디버거' 덮어씀)
    assert cell(16, 13) == "8"         # 시간
    assert cell(16, 14) == "Pass"      # P/F
    # r17 ST202 sub — P/F만 승계. meta는 template에 위임(병합/행별 Debugger 보존).
    assert cell(17, 14) == "Pass"               # PF 기록
    assert cell(17, 12) == "유영규-sub"         # 행별 Debugger 보존 (회귀 방지)
    assert cell(16, 8) == "0.1"                 # parent 병합 anchor 무손상
    # r20 UT301 미수행 — 결과 컬럼 공란 (stale 제거)
    assert cell(20, 8) == ""
    assert cell(20, 14) == ""
    # r21 IT101 Fail
    assert cell(21, 14) == "Fail"
    assert cell(21, 10) == "김진경"
    assert cell(21, 13) == "25"
    # Total 행 (r22)
    assert cell(22, 7) == "5"          # 수행 수
    assert cell(22, 13) == "93"        # 시간 합
    assert cell(22, 14) == "1"         # Fail 수
    # 헤더 블록: Fail 개수 / Result 계산 (stale 덮어씀)
    assert cell(11, 5) == "1"          # Fail count
    assert cell(12, 5) == "Fail"       # 종합 Result


def test_build_fail_cell_marked(fixture_bytes):
    """IT101 P/F(N열) 셀에 FAIL 빨간 배경(design_tokens)이 적용되는지."""
    from backend.services.design_tokens import FAIL_FILL_RGB
    res = build_summary_report(fixture_bytes, [("self", fixture_bytes)])
    wb = openpyxl.load_workbook(io.BytesIO(res.xlsm_bytes))
    # 라운드 107 — ES95411 템플릿 보존 수식 재계산 보장 (fullCalcOnLoad 회귀 가드).
    assert wb.calculation.fullCalcOnLoad is True
    ws = wb["Summary"]
    fill = ws.cell(row=21, column=14).fill
    assert fill is not None and fill.fgColor is not None
    assert str(fill.fgColor.rgb).upper().endswith(FAIL_FILL_RGB[-6:].upper())


def test_build_meta_header_override(fixture_bytes):
    res = build_summary_report(
        fixture_bytes, [("self", fixture_bytes)],
        SwReportBuildMeta(project_id="KJPDS02", phase="DV", product="PDS",
                          asil_level="ASIL A", software_platform_ver="25A1"),
    )
    wb = openpyxl.load_workbook(io.BytesIO(res.xlsm_bytes), data_only=True)
    ws = wb["Summary"]
    # Phase(B4)→E4, Product(B6)→E6
    assert str(ws["E4"].value).strip() == "DV"
    assert str(ws["E6"].value).strip() == "PDS"


def test_output_is_valid_xlsx(fixture_bytes):
    """빌드 산출물이 외부링크 정화 후에도 재오픈 가능 + 시트 보존."""
    res = build_summary_report(fixture_bytes, [("self", fixture_bytes)])
    wb = openpyxl.load_workbook(io.BytesIO(res.xlsm_bytes))
    assert "Summary" in wb.sheetnames
    assert "21.UT101" in wb.sheetnames  # detail 시트 보존


def test_missing_summary_raises():
    wb = openpyxl.Workbook()
    out = io.BytesIO()
    wb.save(out)
    with pytest.raises(ValueError, match="Summary"):
        build_summary_report(out.getvalue(), [])


# ── 수식 캐시 미스 (리뷰 F1/F2 회귀 가드) ────────────────────────────────────
def test_formula_cache_miss_hours_fallback_and_pf_warning():
    """source hours/PF가 수식(캐시 없음)일 때: hours는 'P/F' wrong-pick 없이 준비+수행+
    검토 합산, PF 미캐시는 incomplete로 surface (silent 아님)."""
    wb = openpyxl.Workbook()
    summ = wb.active
    assert summ is not None
    summ.title = "Summary"
    for col, txt in [(5, "ID"), (7, "점검 대상"), (8, "분석 차수"),
                     (13, "총 분석시간"), (14, "P/F"), (16, "Sheet Name")]:
        summ.cell(row=15, column=col, value=txt)
    summ["B16"], summ["E16"], summ["G16"], summ["P16"] = 21, "UT101", "O", "21.UT101"
    summ["B17"] = "Total"
    det = wb.create_sheet("21.UT101")
    det["B4"], det["C4"] = "분석차수", "0.1"
    det["B6"], det["C6"] = "Tester", "주희영"
    det["E4"], det["F4"] = "준비", 10
    det["E5"], det["F5"] = "수행", 10
    det["E6"], det["F6"] = "검토", 10
    det["E7"], det["F7"] = "Total", "=SUM(F4:F6)"          # 수식 (캐시 없음)
    det["H7"], det["I7"] = "P/F", '=IF(1=1,"Pass","Fail")'  # 수식 (캐시 없음)
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    res = build_summary_report(data, [("formula-src", data)])
    ws = openpyxl.load_workbook(io.BytesIO(res.xlsm_bytes), data_only=True)["Summary"]
    # hours: 빈 Total 셀 우측의 'P/F' 라벨을 긁지 않고 컴포넌트 합산 30
    assert str(ws.cell(row=16, column=13).value).strip() == "30"
    # PF 미캐시 → incomplete로 surface
    assert any("P/F 미캐시" in m for m in res.incomplete_rows), res.incomplete_rows
    # Total 시간 합도 30 (오염 없음)
    assert str(ws.cell(row=17, column=13).value).strip() == "30"


# ── 컬럼 자동 감지 (타 프로젝트/공식 템플릿 컬럼 이동 대응) ──────────────────
def test_column_autodetect_shifted_layout():
    """헤더 컬럼이 v1.02 표준과 다른 위치(우측으로 밀림)여도 라벨로 자동 감지해
    올바른(이동된) 셀에 데이터가 기록되는지 — 타 프로젝트/공식 템플릿 일반화 증명."""
    wb = openpyxl.Workbook()
    summ = wb.active
    assert summ is not None
    summ.title = "Summary"
    # 비표준 위치: No=3, ID=7, Name=8, 점검=9, 차수=10, SWVer=11, Tester=12,
    # Tool=13, Debugger=14, hours=15, P/F=16, Note=17, SheetName=18 (표준 대비 +2~)
    hdr = {3: "No", 7: "ID", 8: "Test Name", 9: "점검 대상", 10: "분석 차수",
           11: "SW Ver.", 12: "Tester", 13: "Tool", 14: "Debugger",
           15: "총 분석시간", 16: "P/F", 17: "Note", 18: "Sheet Name"}
    for c, t in hdr.items():
        summ.cell(row=15, column=c, value=t)
    summ.cell(row=16, column=3, value=21)
    summ.cell(row=16, column=7, value="UT101")
    summ.cell(row=16, column=9, value="O")
    summ.cell(row=16, column=18, value="21.UT101")
    summ.cell(row=17, column=3, value="Total")
    _detail(wb, "21.UT101", it="0.1", swv="25A1", tester="주희영", dbg="이재원",
            prep=10, ex=10, rev=10, total=30, pf="Pass")
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    res = build_summary_report(data, [("self", data)])
    assert set(res.summary["detected_columns"]) >= {
        "id", "pf", "tester", "hours", "sheetname", "performed", "name"}
    ws = openpyxl.load_workbook(io.BytesIO(res.xlsm_bytes), data_only=True)["Summary"]
    # 데이터가 '이동된' 컬럼에 정확히 기록 (하드코딩 N/M/J가 아니라 16/15/12)
    assert str(ws.cell(row=16, column=12).value).strip() == "주희영"  # Tester @12
    assert str(ws.cell(row=16, column=14).value).strip() == "이재원"  # Debugger @14
    assert str(ws.cell(row=16, column=15).value).strip() == "30"      # hours @15
    assert str(ws.cell(row=16, column=16).value).strip() == "Pass"    # P/F @16
    # 헤더 없는 미사용 컬럼(5/6)에는 오기록 없음
    assert ws.cell(row=16, column=5).value in (None, "")
    assert ws.cell(row=16, column=6).value in (None, "")
    # Total 행도 이동된 컬럼에
    assert str(ws.cell(row=17, column=9).value).strip() == "1"        # 수행수 @9
    assert str(ws.cell(row=17, column=15).value).strip() == "30"      # 시간합 @15


# ── 데이터 변경 반영 (Fail↔Pass·hours·Total + stale 빨강 제거) ────────────────
def _dc_template() -> "openpyxl.Workbook":
    wb = openpyxl.Workbook()
    s = wb.active
    assert s is not None
    s.title = "Summary"
    for c, t in [(5, "ID"), (7, "점검 대상"), (8, "분석 차수"), (10, "Tester"),
                 (13, "총 분석시간"), (14, "P/F"), (16, "Sheet Name")]:
        s.cell(row=15, column=c, value=t)
    s.cell(row=16, column=5, value="UT101"); s.cell(row=16, column=7, value="O")
    s.cell(row=16, column=16, value="21.UT101")
    s.cell(row=17, column=5, value="UT201"); s.cell(row=17, column=7, value="O")
    s.cell(row=17, column=16, value="22.UT201")
    s.cell(row=18, column=2, value="Total")
    return wb


def _dc_source(ut101_pf, ut201_pf, ut101_hours) -> bytes:
    wb = _dc_template()
    _detail(wb, "21.UT101", it="0.1", swv="25A1", tester="T1", dbg="d",
            prep=ut101_hours, ex=0, rev=0, total=ut101_hours, pf=ut101_pf)
    _detail(wb, "22.UT201", it="0.1", swv="25A1", tester="T2", dbg="d",
            prep=5, ex=0, rev=0, total=5, pf=ut201_pf)
    b = io.BytesIO(); wb.save(b)
    return b.getvalue()


def test_data_change_reflected_and_stale_fail_fill_cleared():
    # 1차: UT101=Fail(빨강), UT201=Pass, hours=10
    tplA = io.BytesIO(); _dc_template().save(tplA)
    o1 = build_summary_report(tplA.getvalue(), [("A", _dc_source("Fail", "Pass", 10))])
    w1 = openpyxl.load_workbook(io.BytesIO(o1.xlsm_bytes))["Summary"]
    assert str(w1.cell(row=16, column=14).value).strip() == "Fail"
    assert _is_fail_fill(w1.cell(row=16, column=14))          # 빨강 적용
    assert o1.summary["fail_count"] == 1

    # 2차: o1(빨강 보유 populated)을 template로, source 변경 — UT101 Fail→Pass,
    #      UT201 Pass→Fail, hours 10→99
    o2 = build_summary_report(o1.xlsm_bytes, [("B", _dc_source("Pass", "Fail", 99))])
    w2 = openpyxl.load_workbook(io.BytesIO(o2.xlsm_bytes))["Summary"]
    # UT101: Pass + stale 빨강 제거
    assert str(w2.cell(row=16, column=14).value).strip() == "Pass"
    assert not _is_fail_fill(w2.cell(row=16, column=14)), "Fail→Pass인데 빨강 잔존"
    assert str(w2.cell(row=16, column=13).value).strip() == "99"   # hours 변경 반영
    # UT201: Fail + 빨강 추가
    assert str(w2.cell(row=17, column=14).value).strip() == "Fail"
    assert _is_fail_fill(w2.cell(row=17, column=14))
    # Total 재계산 (fail 1)
    assert str(w2.cell(row=18, column=14).value).strip() == "1"
    assert o2.summary["fail_ids"] == ["UT201"]


def test_data_change_performed_to_not_performed_clears_row():
    # 1차: UT101 수행(O) Pass — 값 채워짐
    tplA = io.BytesIO(); _dc_template().save(tplA)
    o1 = build_summary_report(tplA.getvalue(), [("A", _dc_source("Pass", "Pass", 10))])
    # 2차: template의 UT101을 미수행(X)으로 바꾸고 재빌드 → 결과 컬럼 공란
    tpl2 = openpyxl.load_workbook(io.BytesIO(o1.xlsm_bytes))
    tpl2["Summary"].cell(row=16, column=7, value="X")    # 점검대상 O→X
    b2 = io.BytesIO(); tpl2.save(b2)
    o2 = build_summary_report(b2.getvalue(), [("B", _dc_source("Pass", "Pass", 10))])
    w2 = openpyxl.load_workbook(io.BytesIO(o2.xlsm_bytes))["Summary"]
    # UT101 결과 컬럼(Tester/hours/PF) 공란화
    assert w2.cell(row=16, column=10).value in (None, "")   # Tester
    assert w2.cell(row=16, column=13).value in (None, "")   # hours
    assert w2.cell(row=16, column=14).value in (None, "")   # P/F


# ── preview ──────────────────────────────────────────────────────────────────
def test_preview(fixture_bytes):
    r = preview_summary_report(fixture_bytes, [("self", fixture_bytes)])
    s = r["summary"]
    assert s["rows_total"] == 6
    assert s["performed_count"] == 5
    assert s["matched_rows"] == 5
    assert s["fail_count"] == 1
    assert s["total_hours"] == 93
    assert s["overall_result"] == "Fail"
    it101 = next(x for x in r["rows"] if x["id"] == "IT101")
    assert it101["pf"] == "Fail" and it101["matched"] is True
    ut301 = next(x for x in r["rows"] if x["id"] == "UT301")
    assert ut301["matched"] is False and ut301["planned"] == "X"
