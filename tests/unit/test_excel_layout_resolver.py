"""54차 T280 — excel_layout_resolver.py 회귀 테스트.

8 시나리오: v2.02 inspect / v3.01 fallback / sha256 caching / graceful missing /
LRU 회전 / kind 분기 (coverage vs sitr).
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import openpyxl  # noqa: E402

from backend.services import excel_layout_resolver as elr  # noqa: E402


def _make_xlsx(sheets: dict[str, list[list[str]]]) -> bytes:
    """sheets dict로 xlsx bytes 생성. {sheet_name: [[cell, cell, ...], ...]}."""
    wb = openpyxl.Workbook()
    # default sheet 제거
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(autouse=True)
def _reset_cache():
    elr.clear_layout_cache()
    yield
    elr.clear_layout_cache()


def _v202_coverage_template() -> bytes:
    """SwIT v2.02 Coverage 양식 mimicry — 'SW Version', 'HW Version' 라벨."""
    return _make_xlsx({
        "Cover": [
            ["", ""],
            ["Project", ""],
            ["ASIL Level", ""],
            ["Author", ""],
            ["Approver", ""],
        ],
        "1.Test Summary": [
            ["", ""],
            ["Project Name", ""],
            ["SW Version", ""],
            ["HW Version", ""],
            ["Test Date", ""],
            ["Test Engineer", ""],
            ["Final Test Result", ""],
            # B17-F17 row 16~17번째 즈음
            [""], [""], [""], [""], [""], [""], [""], [""], [""],
            ["Total TC", ""],  # v2.02 신규 — row 17 (1-based)
            [""], [""], [""], [""],
            ["Requirements/Design Coverage", ""],  # row 22
        ],
        "1.Traceability": [["Function ID"]],
        "2.Consistency": [["Item", "Expected", "Actual", "Result"]],
        "3.Coverage": [["Unit ID"]],
        "History": [["Version", "Date"]],
    })


def _v202_sitr_template() -> bytes:
    """SwIT v2.02 SITR — xlsm은 만들 수 없으나 sheet 구조는 mimicry. 본 fixture는
    xlsx 확장자로 만들지만 kind='sitr'로 호출하여 path 분기만 검증."""
    return _make_xlsx({
        "Cover": [["Project", ""]],
        "1.Test Summary": [
            ["", ""],
            ["SW Version", ""],
            ["HW Version", ""],
            ["Test Engineer", ""],
        ],
        "Deviation": [["Test Case ID", "Issue", "Rationale"]],
        "Test Log": [
            # header row에 marker col 추가
            ["Test Case ID", "Component", "Method", "Result", "Function ID", "ASIL", "", "", "Marker"],
        ],
        "History": [["Version"]],
    })


def _v301_coverage_template() -> bytes:
    """SwUT v3.01 양식 — Release Name(SW) / Test Target Version(HW) 라벨."""
    return _make_xlsx({
        "Cover": [
            ["Project", ""],
            ["ASIL Level", ""],
            ["Author", ""],
            ["Approver", ""],
        ],
        "Test Summary": [
            ["Project Name", ""],
            ["Release Name(SW)", ""],
            ["Test Target Version(HW)", ""],
            ["Test Date", ""],
            ["Test Engineer", ""],
            ["Final Test Result", ""],
        ],
        "1.Traceability": [["Function ID"]],
        "2.Consistency": [["Item"]],
        "3.Coverage": [["Unit ID"]],
        "History": [["Version"]],
    })


def _empty_xlsx() -> bytes:
    """label 없는 빈 xlsx — graceful warnings 확인."""
    return _make_xlsx({"Sheet1": [["dummy"]]})


class TestInspectCoverageV202:
    def test_inspect_coverage_template_v202(self):
        layout = elr.inspect_swit_layout(_v202_coverage_template(), "coverage")
        assert layout.detected_version == "v2.02"
        assert layout.test_summary_labels.get("release_sw_version") == "SW Version"
        assert layout.test_summary_labels.get("hw_version") == "HW Version"
        assert layout.tc_stats_row is not None
        assert layout.tc_stats_col_start is not None
        assert layout.requirements_row is not None
        assert layout.fallback_to_v301 is False

    def test_inspect_kind_coverage_skips_test_log(self):
        """kind='coverage'는 Test Log/Deviation inspect skip."""
        layout = elr.inspect_swit_layout(_v202_coverage_template(), "coverage")
        assert layout.test_log_header_cell is None
        assert layout.deviation_header_cell is None
        assert layout.test_log_extra_marker_col is None


class TestInspectSitrV202:
    def test_inspect_sitr_template_v202_marker_col(self):
        layout = elr.inspect_swit_layout(_v202_sitr_template(), "sitr")
        assert layout.detected_version == "v2.02"
        # Test Log header에 "Marker" 가 col 9에 있음
        assert layout.test_log_extra_marker_col == 9
        # Test Log header cell + Deviation header cell도 잡혀야 함
        assert layout.test_log_header_cell is not None
        assert layout.deviation_header_cell is not None

    def test_inspect_kind_sitr_finds_deviation(self):
        layout = elr.inspect_swit_layout(_v202_sitr_template(), "sitr")
        assert layout.deviation_header_cell is not None
        assert layout.deviation_header_cell.row == 1
        assert layout.deviation_header_cell.col == 1


class TestV301Fallback:
    def test_inspect_v301_fallback(self):
        """v3.01 양식 → fallback_to_v301=True + tc_stats_row=None."""
        layout = elr.inspect_swit_layout(_v301_coverage_template(), "coverage")
        assert layout.detected_version == "v3.01"
        assert layout.fallback_to_v301 is True
        assert layout.tc_stats_row is None
        assert layout.requirements_row is None
        # v3.01 라벨이 검출돼야 함
        assert layout.test_summary_labels.get("release_sw_version") == "Release Name(SW)"

    def test_v301_does_not_trigger_v202_fallback(self):
        """56차 T306 회귀 — v3.01 양식은 v2.02 label-missing fallback 미발동.

        SwUT v3.01 Coverage는 TC stats row 자체가 없는 양식 → fallback path가
        잘못 발동되면 v3.01 회귀 깨짐. detected_version='v3.01'이면 fallback skip 보장.
        """
        layout = elr.inspect_swit_layout(_v301_coverage_template(), "coverage")
        assert layout.tc_stats_label_missing is False
        assert layout.requirements_label_missing is False
        assert layout.tc_stats_row is None
        assert layout.requirements_row is None


class TestV202CoverageLabelMissingFallback:
    """56차 T306 — 회사 Coverage Report v2.02 양식은 row 17 (TC stats) + row 20
    (Requirements) 이 사용자 수동 입력 영역으로 비어있음. label 매칭 실패 시 default
    position fallback + writer가 라벨 stamp.
    """

    def _v202_coverage_label_missing_template(self) -> bytes:
        """SW Version / HW Version 라벨은 있어 v2.02 detect되지만 TC stats +
        Requirements row label은 부재한 회사 Coverage v2.02 양식 mimicry."""
        return _make_xlsx({
            "Cover": [["Project", ""]],
            "1.Test Summary": [
                ["", ""],
                ["Project Name", ""],
                ["SW Version", ""],
                ["HW Version", ""],
                ["Test Date", ""],
                ["Test Engineer", ""],
                ["Final Test Result", ""],
                # row 8~16 빈 row
                [""], [""], [""], [""], [""], [""], [""], [""], [""],
                # row 17 — 라벨 부재 (회사 양식 사용자 수동 입력 영역)
                [""],
                [""],  # row 18
                [""], [""],  # row 19, 20 — Requirements row도 라벨 부재
            ],
            "History": [["Version"]],
        })

    def test_v202_coverage_label_missing_triggers_fallback(self):
        """row 17 B열 빈 cell → tc_stats_label_missing=True + default position 사용."""
        layout = elr.inspect_swit_layout(
            self._v202_coverage_label_missing_template(), "coverage",
        )
        assert layout.detected_version == "v2.02"
        # fallback 적용 — data row=18, col=B(=2), label_missing=True
        assert layout.tc_stats_row == 18
        assert layout.tc_stats_col_start == 2
        assert layout.tc_stats_label_missing is True
        # Requirements도 동일 fallback (row 20 빈 cell)
        assert layout.requirements_row == 20
        assert layout.requirements_label_missing is True
        # warnings에 fallback 명시
        assert any("TC stats row label 미발견" in w and "fallback" in w for w in layout.warnings)
        assert any("Requirements row label 미발견" in w and "fallback" in w for w in layout.warnings)

    def test_v202_sitr_with_labels_no_fallback(self):
        """SITR 양식은 label 있어 fallback 미발동 (회귀 유지)."""
        # _v202_coverage_template에 TC stats 'Total TC' + Requirements 라벨 있음
        layout = elr.inspect_swit_layout(_v202_coverage_template(), "coverage")
        assert layout.detected_version == "v2.02"
        assert layout.tc_stats_label_missing is False
        assert layout.requirements_label_missing is False


class TestTestLogTcRowStep57:
    """57차 T314 — Test Log/Test Result 시트의 1 TC당 row step 동적 감지."""

    def _v202_sutr_template_with_6row_step(self) -> bytes:
        """회사 v2.02 SUTR 양식 mimic — TC ID at B5/B11/B17 (6 row step)."""
        return _make_xlsx({
            "Cover": [["Project", ""]],
            "1.Test Summary": [
                ["", ""],
                ["SW Version", ""],
                ["HW Version", ""],
            ],
            "2.Deviation": [["Test Case ID", "Issue"]],
            "3.Test Result": [
                ["Test Log"],                                       # row 1
                [""],                                               # row 2
                ["", "Test Case ", "", "", "", "", "Input"],        # row 3
                ["", "TC ID", "Title", "Test Case Generation Method", "", "", "Param 1"],  # row 4 header
                ["", "SwUTC_0101"],                                 # row 5 col B
                ["", "", "", "", 1],                                # row 6 sub-row
                ["", "", "", "", 2],
                ["", "", "", "", 3],
                ["", "", "", "", 4],
                ["", "", "", "", 5],
                ["", "SwUTC_0102"],                                 # row 11 col B — step = 6
            ],
            "History": [["Version"]],
        })

    def _v301_sutr_template_with_1row_step(self) -> bytes:
        """v3.01 양식 mimic — TC ID 연속 (1 row step backward compat)."""
        return _make_xlsx({
            "Cover": [["Project", ""]],
            "Test Summary": [
                ["Project Name", ""],
                ["Release Name(SW)", ""],
            ],
            "Deviation": [["Test Case ID"]],
            "Test Log": [
                ["", "Test Case ID", "Component", "Method", "Result"],  # row 1 header (col B label)
                ["", "SwUTC_0001"],  # row 2 col B
                ["", "SwUTC_0002"],  # row 3 col B — step = 1
            ],
            "History": [["Version"]],
        })

    def test_v202_sutr_detects_6_row_step(self):
        """B5='SwUTC_0101', B11='SwUTC_0102' → step = 6 감지."""
        layout = elr.inspect_swit_layout(
            self._v202_sutr_template_with_6row_step(), "sitr",
        )
        assert layout.test_log_tc_row_step == 6

    def test_v301_sutr_default_1_step(self):
        """v3.01 TC ID 연속 → step = 1 (backward compat)."""
        layout = elr.inspect_swit_layout(
            self._v301_sutr_template_with_1row_step(), "sitr",
        )
        assert layout.test_log_tc_row_step == 1

    def test_coverage_kind_does_not_inspect_test_log(self):
        """kind='coverage'는 Test Log 시트 inspect 안 함 → step default 1."""
        layout = elr.inspect_swit_layout(_v202_coverage_template(), "coverage")
        assert layout.test_log_tc_row_step == 1


class TestCaching:
    def test_lru_cache_sha256_keying(self, monkeypatch):
        """같은 bytes 두 번 inspect → _inspect_internal 1회만 호출."""
        template = _v202_coverage_template()
        call_count = {"n": 0}
        original = elr._inspect_internal

        def counting(bytes_, kind):
            call_count["n"] += 1
            return original(bytes_, kind)

        monkeypatch.setattr(elr, "_inspect_internal", counting)
        elr.clear_layout_cache()

        layout1 = elr.inspect_swit_layout(template, "coverage")
        layout2 = elr.inspect_swit_layout(template, "coverage")
        assert call_count["n"] == 1
        assert layout1 is layout2  # 동일 객체 반환

    def test_layout_cache_invalidation_on_bytes_change(self):
        """1 byte라도 다르면 sha256 다름 → 새 inspect."""
        t1 = _v202_coverage_template()
        # 1바이트 변경된 새 template
        t2 = _v202_sitr_template()
        assert t1 != t2
        layout1 = elr.inspect_swit_layout(t1, "coverage")
        layout2 = elr.inspect_swit_layout(t2, "coverage")
        # detected_version은 같을 수 있으나 다른 cache entry
        assert elr.cache_size() == 2
        assert layout1 is not layout2

    def test_lru_evict_when_max_exceeded(self):
        """maxsize 초과 시 oldest entry 제거 (54-fix W3: 4 → 8)."""
        elr.clear_layout_cache()
        # 9개 서로 다른 template
        for i in range(9):
            t = _make_xlsx({"Sheet": [[f"v{i}"]]})
            elr.inspect_swit_layout(t, "coverage")
        # _MAX_CACHE_SIZE = 8 이므로 9번째에 oldest evict됨
        assert elr.cache_size() == 8


class TestGraceful:
    def test_missing_labels_graceful(self):
        """label 없는 xlsx → warnings 누적 + fallback=True."""
        layout = elr.inspect_swit_layout(_empty_xlsx(), "coverage")
        assert layout.fallback_to_v301 is True
        # test_summary_labels 비어 있음 (Test Summary 시트 없음)
        assert layout.test_summary_labels == {}
        assert any("Test Summary" in w for w in layout.warnings)

    def test_inspect_rejects_zip_bomb_magic_bytes(self):
        """54-fix C2 — ZIP bomb / magic byte 위조 거부."""
        # ZIP magic bytes (PK\x03\x04)이지만 후속 ZIP 구조 부재
        fake_zip = b"PK\x03\x04" + b"\x00" * 100
        layout = elr.inspect_swit_layout(fake_zip, "coverage")
        assert layout.fallback_to_v301 is True
        # validate_xlsx_template_bytes가 거부 (Open Office XML 구조 미발견 등)
        assert any(
            "template 입력 검증 실패" in w or "Open Office XML" in w or "not a valid ZIP" in w
            for w in layout.warnings
        )

    def test_corrupted_bytes_graceful(self):
        """invalid bytes → fallback=True + warnings (54-fix C2: ZIP bomb 사전 검증)."""
        layout = elr.inspect_swit_layout(b"not a real xlsx", "coverage")
        assert layout.fallback_to_v301 is True
        # C2 fix 후: validate_xlsx_template_bytes가 거부 → "template 입력 검증 실패" 메시지
        assert any(
            "template 입력 검증 실패" in w or "template load 실패" in w or "load 실패" in w
            for w in layout.warnings
        )


# ---------------------------------------------------------------------------
# 58차 F3 — Test Log column 자동 감지 (v2.02 SITR layout)
# ---------------------------------------------------------------------------


class TestScanTestLogColumnsV202:
    """`_scan_test_log_columns` — 헤더 row에서 'Input', 'Expected Result', 'Actual
    Result', 'Pass/Fail', 'Log Data' 라벨을 매칭하여 column 위치 반환.
    """

    def test_scan_test_log_columns_v202_sitr_mock(self):
        """v2.02 SITR mock — B/H/R/AB/AL/AN 라벨이 col 2/8/18/28/38/40."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 2).value = "Test Case"
        ws.cell(1, 8).value = "Input"
        ws.cell(1, 18).value = "Expected Result"
        ws.cell(1, 28).value = "Actual Result"
        ws.cell(1, 38).value = "Pass/Fail"
        ws.cell(1, 40).value = "Log Data"
        cols = elr._scan_test_log_columns(ws)
        assert cols["input_col"] == 8
        assert cols["expected_col"] == 18
        assert cols["actual_col"] == 28
        assert cols["pass_fail_col"] == 38
        assert cols["log_data_col"] == 40
        assert cols["pass_fail_total_col"] is None

    def test_scan_test_log_columns_v301_sutr_mock(self):
        """v3.01 SUTR mock — F/P/Z/AJ/AK/AL 라벨 col 6/16/26/36/37/38."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 6).value = "Input"
        ws.cell(1, 16).value = "Expected Result"
        ws.cell(1, 26).value = "Actual Result"
        ws.cell(1, 36).value = "Pass/Fail Unit"
        ws.cell(1, 37).value = "Pass/Fail Total"
        ws.cell(1, 38).value = "Log Data"
        cols = elr._scan_test_log_columns(ws)
        assert cols["input_col"] == 6
        assert cols["expected_col"] == 16
        assert cols["actual_col"] == 26
        assert cols["pass_fail_col"] == 36
        assert cols["pass_fail_total_col"] == 37
        assert cols["log_data_col"] == 38

    def test_scan_test_log_columns_no_headers_all_none(self):
        """헤더 라벨 없으면 모두 None (v3.01 hardcode fallback)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 2).value = "Random Header"
        cols = elr._scan_test_log_columns(ws)
        assert all(v is None for v in cols.values())

    def test_scan_test_log_columns_handles_none_ws(self):
        """None ws → 모두 None graceful."""
        cols = elr._scan_test_log_columns(None)
        assert all(v is None for v in cols.values())

    def test_scan_precondition_col_swits_kjpds02_pattern_f6a(self):
        """60차 F6-A — 'Precondition' 라벨 → precondition_col 자동 감지.

        KJPDS02 SwITS v1.01 양식 = col 9, HDPDM01 SUTS v3.01 = col 10,
        SITS v2.02 = col 6.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 9).value = "Precondition"
        ws.cell(1, 10).value = "Sequence"
        cols = elr._scan_test_log_columns(ws)
        assert cols["precondition_col"] == 9

    def test_scan_precondition_col_none_when_label_missing_f6a(self):
        """60차 F6-A — 'Precondition' 라벨 미존재 → precondition_col=None.

        KJPDS02 SwUTS v1.01 양식이 Precondition col 없는 케이스. _write_test_log
        에서 stamp skip하여 backward-compat 유지.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 2).value = "TC_ID"
        ws.cell(1, 5).value = "Test Method"  # Precondition 라벨 없음
        cols = elr._scan_test_log_columns(ws)
        assert cols["precondition_col"] is None


# ---------------------------------------------------------------------------
# 58차 F2 — Traceability 헤더 row 자동 감지
# ---------------------------------------------------------------------------


class TestScanTraceabilityHeader:
    """`_scan_traceability_header` — SwUFn_/SwUTC_/SwITC_ prefix 5개+ 행 row 반환."""

    def test_scan_traceability_header_v202_swit_at_row_20(self):
        """SwIT v2.02 양식 mock — 헤더 row 20에 SwUFn_ prefix 5개+."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(5):
            ws.cell(20, 3 + i).value = f"SwUFn_{i:04d}"
        header_row = elr._scan_traceability_header(ws)
        assert header_row == 20

    def test_scan_traceability_header_v301_at_row_3(self):
        """SwUT v3.01 양식 mock — 헤더 row 3에 SwUFn_ prefix."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(6):
            ws.cell(3, 3 + i).value = f"SwUFn_{i:04d}"
        header_row = elr._scan_traceability_header(ws)
        assert header_row == 3

    def test_scan_traceability_header_returns_none_when_no_swufn(self):
        """SwUFn_/SwUTC_/SwITC_ prefix 5개 미만 → None."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "Header"
        # SwUFn 2개만
        ws.cell(2, 2).value = "SwUFn_0001"
        ws.cell(2, 3).value = "SwUFn_0002"
        header_row = elr._scan_traceability_header(ws)
        assert header_row is None

    def test_scan_traceability_header_handles_none_ws(self):
        """None ws → None graceful."""
        assert elr._scan_traceability_header(None) is None


# ---------------------------------------------------------------------------
# 58차 F3 — SwitLayout inspect SITR 분기에 column 자동 반영
# ---------------------------------------------------------------------------


class TestInspectSitrColumnDetection:
    """inspect_swit_layout(kind='sitr') 호출 시 layout.test_log_*_col 자동 채움."""

    def test_inspect_sitr_v202_layout_has_input_col(self):
        """v2.02 SITR mock template inspect → test_log_input_col=8."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [
                ["Project Name"], ["SW Version"], ["HW Version"],
                ["Test Date"], ["Test Engineer"], ["Target Coverage"],
                ["Actual Coverage"], ["Final Test Result"],
            ],
            "Test Log": [
                # row 1: 헤더 — col 2/8/18/28/38/40
                [None, "Test Case", None, None, None, None, None, "Input"]
                + [None] * 9
                + ["Expected Result"]
                + [None] * 9
                + ["Actual Result"]
                + [None] * 9
                + ["Pass/Fail", None, "Log Data"],
            ],
            "Deviation": [["Test Case ID"]],
        })
        layout = elr.inspect_swit_layout(template, "sitr")
        assert layout.test_log_input_col == 8
        assert layout.test_log_expected_col == 18
        assert layout.test_log_actual_col == 28
        assert layout.test_log_pass_fail_col == 38
        assert layout.test_log_log_data_col == 40

    def test_inspect_coverage_no_test_log_columns_set(self):
        """kind='coverage'에는 test_log_*_col 모두 None."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["Project Name"]],
        })
        layout = elr.inspect_swit_layout(template, "coverage")
        assert layout.test_log_input_col is None
        assert layout.test_log_expected_col is None
        assert layout.test_log_actual_col is None


# ---------------------------------------------------------------------------
# 59차 F4-A — 변수명 헤더 row + Input/Expected/Actual block max counts
# ---------------------------------------------------------------------------


class TestScanTestLogVariableHeaderRow:
    """`_scan_test_log_variable_header_row` — input_col~expected_col 범위에서
    영문 식별자 또는 Inpt[N] 3+ 연속 row 자동 감지 (KJPDS02 v1.01 = row 5).
    """

    def test_scan_variable_header_v101_kjpds02_row_5(self):
        """KJPDS02 v1.01 양식 mimicry — row 5 col 10~ 에 영문 변수 식별자 3+."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(5, 10).value = "u16g_SysDiag_SystemStatus"
        ws.cell(5, 11).value = "u8g_SysEepromCtrl_InLineMod_F"
        ws.cell(5, 12).value = "u16g_SysEepromCtrl_OpLimitVehi"
        ws.cell(5, 13).value = "s_System_I"
        # input_col=10, expected_col=20 → 10~19 scan
        header_row = elr._scan_test_log_variable_header_row(
            ws, input_col=10, expected_col=20,
        )
        assert header_row == 5

    def test_scan_variable_header_inpt_label_pattern(self):
        """`Inpt[0]`, `Inpt[1]`, ... 패턴도 헤더 row로 인식."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(4, 6).value = "Inpt[0]"
        ws.cell(4, 7).value = "Inpt[1]"
        ws.cell(4, 8).value = "Inpt[2]"
        ws.cell(4, 9).value = "Inpt[3]"
        header_row = elr._scan_test_log_variable_header_row(
            ws, input_col=6, expected_col=16,
        )
        assert header_row == 4

    def test_scan_variable_header_returns_none_when_no_match(self):
        """영문 식별자 0개 → None (변수명 헤더 stamp skip)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(3, 6).value = "한글헤더"
        ws.cell(3, 7).value = "공백 들어간 라벨"
        ws.cell(3, 8).value = "12-34"
        header_row = elr._scan_test_log_variable_header_row(
            ws, input_col=6, expected_col=16,
        )
        assert header_row is None

    def test_scan_variable_header_handles_none_input_col(self):
        """input_col=None (column 자동 감지 실패) → None graceful."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(5, 10).value = "u16g_var"
        header_row = elr._scan_test_log_variable_header_row(
            ws, input_col=None, expected_col=None,
        )
        assert header_row is None


class TestScanTestLogMaxCounts:
    """`_scan_test_log_max_counts` — 인접 col 차이로 block 크기 산출."""

    def test_max_counts_v101_kjpds02(self):
        """KJPDS02 v1.01: input=10, expected=20, actual=30, pass_fail=42 →
        max_counts = (10, 10, 12)."""
        cols = {
            "input_col": 10, "expected_col": 20, "actual_col": 30,
            "pass_fail_col": 42, "pass_fail_total_col": None, "log_data_col": None,
        }
        result = elr._scan_test_log_max_counts(cols)
        assert result["input_max_count"] == 10
        assert result["expected_max_count"] == 10
        assert result["actual_max_count"] == 12

    def test_max_counts_all_none_uses_default(self):
        """모든 col None → default 10."""
        cols = {
            "input_col": None, "expected_col": None, "actual_col": None,
            "pass_fail_col": None, "pass_fail_total_col": None, "log_data_col": None,
        }
        result = elr._scan_test_log_max_counts(cols)
        assert result["input_max_count"] == 10
        assert result["expected_max_count"] == 10
        assert result["actual_max_count"] == 10

    def test_max_counts_uses_ws_max_col_for_actual(self):
        """pass_fail_col 부재 시 ws_max_col로 actual_max 산출."""
        cols = {
            "input_col": 6, "expected_col": 16, "actual_col": 26,
            "pass_fail_col": None, "pass_fail_total_col": None, "log_data_col": None,
        }
        result = elr._scan_test_log_max_counts(cols, ws_max_col=50)
        assert result["input_max_count"] == 10
        assert result["expected_max_count"] == 10
        # ws_max_col=50, actual_col=26 → diff = 50+1-26 = 25
        assert result["actual_max_count"] == 25


class TestInspectSitrV4AStepLayout:
    """`_inspect_internal` SITR 분기에서 test_log_step_layout 결정."""

    def test_v202_tc_row_step_gt_1_step_in_rows(self):
        """v2.02 양식 SwUTC_/SwITC_ prefix 2개 row 차이로 tc_row_step=6 →
        test_log_step_layout='step_in_rows'."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["SW Version"]],
            "Test Log": [
                ["Test Case ID", None, None, None, None, "Input"]
                + [None] * 9
                + ["Expected Result"]
                + [None] * 9
                + ["Actual Result"]
                + [None] * 9
                + ["Pass/Fail", None, "Log Data"],
                # row 2: SwUTC_0001 (1번째 TC)
                [None, "SwUTC_0001"],
                # row 3~7: sub-rows (5개)
                [None, None], [None, None], [None, None], [None, None], [None, None],
                # row 8: SwUTC_0002 (2번째 TC) — 차이 6
                [None, "SwUTC_0002"],
            ],
            "Deviation": [["Test Case ID"]],
        })
        layout = elr.inspect_swit_layout(template, "sitr")
        assert layout.test_log_tc_row_step == 6
        assert layout.test_log_step_layout == "step_in_rows"

    def test_v301_tc_row_step_1_single_row(self):
        """v3.01 hardcode default — tc_row_step=1 → 'single_row'."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["Release Name(SW)"]],
            "Test Log": [
                ["Test Case ID"],
                [None, "SwUTC_0001"],
                [None, "SwUTC_0002"],  # 차이 1
            ],
        })
        layout = elr.inspect_swit_layout(template, "sitr")
        assert layout.test_log_tc_row_step == 1
        assert layout.test_log_step_layout == "single_row"


# ---------------------------------------------------------------------------
# 59차 F4-C — KJPDS02 v1.01 양식 시트 구성 자동 감지
# ---------------------------------------------------------------------------


class TestInspectV101SheetConfig:
    """`_inspect_internal` 양식 분류 — 시트 이름 패턴으로 v1.01 / v2.02 / v3.01 구분."""

    def test_v101_coverage_signature_detected(self):
        """2.Traceability + 3.Consistency + 4.Coverage 시트 → detected_version=v1.01.

        라운드 F7 D2: traceability_matrix_kind는 시트 header 내용 inspect로 결정
        (시트명 prefix만이 아닌). SwITCV 양식 (SwST_NN header 다수)이어야 switc_x_swst.
        """
        # SwITCV 회사 표준 시뮬레이션 — 2.Traceability R11에 SwST/SwSTR header 다수
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["Project Name"]],
            "2.Traceability": [
                ["Matrix"], [], [], [], [], [], [], [], [], [],
                ["ID", "101", "SwST_01", "SwST_02", "SwST_03", "SwST_04",
                 "SwST_05", "SwSTR_01", "SwSTR_02"],
            ],
            "3.Consistency": [["Item"]],
            "4.Coverage": [["Function"]],
            "History": [["Version"]],
        })
        layout = elr.inspect_swit_layout(template, "coverage")
        assert layout.detected_version == "v1.01"
        assert layout.traceability_matrix_kind == "switc_x_swst"
        assert layout.coverage_metric_kind == "function_and_calls"
        assert layout.test_summary_coverage_breakdown == 4

    def test_v101_swutcv_swufn_header_not_classified_as_switc_round_f7_d2(self):
        """라운드 F7 D2: 회사 표준 SwUTCV는 v1.01 signature지만 SwUFn matrix.
        2.Traceability header에 SwUFn_NNNN 다수 → swufn_x_env 유지 (switc_x_swst X).
        """
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["Project Name"]],
            "2.Traceability": [
                ["Matrix"], [], [], [], [], [], [], [], [], [], [],
                ["ID", "419", "SwUFn_0101", "SwUFn_0102", "SwUFn_0103",
                 "SwUFn_0104", "SwUFn_0105"],
            ],
            "3.Consistency": [["Item"]],
            "4.Coverage": [["Function"]],
            "History": [["Version"]],
        })
        layout = elr.inspect_swit_layout(template, "coverage")
        assert layout.detected_version == "v1.01"
        assert layout.traceability_matrix_kind == "swufn_x_env", (
            f"D2 회귀: SwUTCV (SwUFn header) → swufn_x_env 유지해야 하나 "
            f"{layout.traceability_matrix_kind}"
        )

    def test_matrix_kind_boundary_3_swst_threshold_round_f8(self):
        """F7 R2 carry-over N4 — matrix_kind 임계값 boundary 회귀.
        Round 5 W4 fix로 임계값 'swst_count >= 3 OR (swst_count > 0 AND swufn==0)'.
        소규모 SwITCV 양식 (3 SwST + 0 SwUFn) → switc_x_swst 정상 분류.
        """
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["Project Name"]],
            "2.Traceability": [
                ["Matrix"], [], [], [], [], [], [], [], [], [], [],
                ["ID", "5", "SwST_01", "SwST_02", "SwST_03"],  # 3 SwST + 0 SwUFn
            ],
            "3.Consistency": [["Item"]],
            "4.Coverage": [["Function"]],
            "History": [["Version"]],
        })
        layout = elr.inspect_swit_layout(template, "coverage")
        assert layout.traceability_matrix_kind == "switc_x_swst", (
            f"N4 boundary: 3 SwST + 0 SwUFn → switc_x_swst (got "
            f"{layout.traceability_matrix_kind})"
        )

    def test_matrix_kind_boundary_1_swst_edge_case_round_f8(self):
        """N4 edge case — 1 SwST + 0 SwUFn → switc_x_swst (extreme MVP)."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["Project Name"]],
            "2.Traceability": [
                ["Matrix"], [], [], [], [], [], [], [], [], [], [],
                ["ID", "1", "SwST_01"],  # 1 SwST 단독
            ],
            "3.Consistency": [["Item"]],
            "4.Coverage": [["Function"]],
            "History": [["Version"]],
        })
        layout = elr.inspect_swit_layout(template, "coverage")
        assert layout.traceability_matrix_kind == "switc_x_swst"

    def test_matrix_kind_boundary_2_swst_1_swufn_falls_to_swufn_round_f8(self):
        """N4 ambiguous — 2 SwST + 1 SwUFn → (3 미만 + SwUFn>0) → swufn_x_env default.
        Round 5 W4 임계값이 의도적으로 보수적: 2 SwST는 1 SwUFn과 함께 있으면 SwUT 양식
        오판 가능성 — default 안전."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["Project Name"]],
            "2.Traceability": [
                ["Matrix"], [], [], [], [], [], [], [], [], [], [],
                ["ID", "3", "SwST_01", "SwST_02", "SwUFn_0001"],
            ],
            "3.Consistency": [["Item"]],
            "4.Coverage": [["Function"]],
            "History": [["Version"]],
        })
        layout = elr.inspect_swit_layout(template, "coverage")
        assert layout.traceability_matrix_kind == "swufn_x_env", (
            f"N4 ambiguous boundary: 2 SwST + 1 SwUFn → default swufn (got "
            f"{layout.traceability_matrix_kind})"
        )

    def test_v202_no_v101_signature(self):
        """1.Traceability + 2.Consistency + 3.Coverage → v1.01 신호 부족 → 기본 유지."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["SW Version"]],
            "1.Traceability": [["Matrix"]],
            "2.Consistency": [["Item"]],
            "3.Coverage": [["Function"]],
        })
        layout = elr.inspect_swit_layout(template, "coverage")
        # v1.01 시그널 < 2 (2.Traceability/3.Consistency/4.Coverage 모두 부재)
        assert layout.detected_version != "v1.01"
        assert layout.traceability_matrix_kind == "swufn_x_env"
        assert layout.coverage_metric_kind == "single"
        assert layout.test_summary_coverage_breakdown == 1

    def test_v101_sitr_no_deviation_sheet(self):
        """v1.01 SITR 시트 (Deviation 없음) → deviation_sheet_present=False."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["Project Name"]],
            "2.Test Log": [["Test Case ID"]],
            "History": [["Version"]],
        })
        layout = elr.inspect_swit_layout(template, "sitr")
        assert layout.deviation_sheet_present is False

    def test_v202_sitr_with_deviation_sheet(self):
        """v2.02 SITR (Deviation 시트 보유) → deviation_sheet_present=True."""
        template = _make_xlsx({
            "Cover": [["Project"]],
            "1.Test Summary": [["SW Version"]],
            "Deviation": [["Test Case ID"]],
            "Test Log": [["Test Case ID"]],
        })
        layout = elr.inspect_swit_layout(template, "sitr")
        assert layout.deviation_sheet_present is True
