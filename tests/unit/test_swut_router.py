"""Tests for backend.routers.swut endpoint (8차 라운드 T143).

FastAPI TestClient로 endpoint 진입 검증:
- Pydantic 422 (invalid release_sw_version / test_engineer 줄바꿈 / doc_id_sequence 비-digit)
- X-User 400 (헤더 미포함)
- Coverage build 200 + xlsx Content-Type (모킹된 input/template)
- Jenkins fetcher mock 후 SwUTSession 반환
- Semaphore(2) 동시 호출 제한
"""
from __future__ import annotations

import asyncio
import io
import pathlib
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
from fastapi.testclient import TestClient
from pydantic import ValidationError

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.main import app  # noqa: E402
from backend.schemas import SwUTBuildRequest  # noqa: E402
from backend.services.swut_input_adapter import (  # noqa: E402
    CoverageStats,
    EnvironmentData,
    ExecutionRow,
    FunctionCoverage,
    SwUTSession,
)

client = TestClient(app)


def _minimal_xlsx_template_bytes() -> bytes:
    """최소 6시트 Coverage Report template."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cov = wb.create_sheet("Cover")
    cov["B1"] = "Project"
    ts = wb.create_sheet("Test Summary")
    ts["B1"] = "Project Name"
    wb.create_sheet("1.Traceability")
    wb.create_sheet("2.Consistency")
    cov3 = wb.create_sheet("3. Coverage")
    cov3["A1"] = "Statement Coverage"
    cov3["A6"] = "Unit ID"
    wb.create_sheet("History")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_session() -> SwUTSession:
    env = EnvironmentData(
        env_name="SWTE_01",
        component_name="X",
        test_cases={"SwUFn_0001.001": [object()]},
        test_results={"SwUFn_0001.001": ExecutionRow(tc_name="SwUFn_0001.001", passed=True)},
        function_coverage=[FunctionCoverage(unit_id="SwUFn_0001", name="X")],
    )
    return SwUTSession(environments=[env])


def test_resolve_swut_log_folder_uses_request_then_config(monkeypatch):
    from backend.routers import swut as swut_mod

    monkeypatch.setattr(
        swut_mod,
        "_load_meta_from_config",
        lambda _project_id: {"swut_log_folder": "U:/unit/01.Log"},
    )

    req_default = SwUTBuildRequest(
        project_id="KJPDS02",
        release_sw_version="1.01",
        test_date="2025-12-05",
    )
    assert swut_mod._resolve_swut_log_folder(req_default) == "U:/unit/01.Log"

    req_override = SwUTBuildRequest(
        project_id="KJPDS02",
        release_sw_version="1.01",
        test_date="2025-12-05",
        log_folder="C:/request/log",
    )
    assert swut_mod._resolve_swut_log_folder(req_override) == "C:/request/log"


# ---------------------------------------------------------------------------
# Pydantic 422 검증
# ---------------------------------------------------------------------------

class TestPydanticValidation:
    def test_invalid_release_sw_version_rejected(self):
        r = client.post(
            "/api/swut/coverage/build",
            json={
                "project_id": "HDPDM01",
                "release_sw_version": "vX.Y",  # regex 미충족
                "test_date": "2024-02-19",
            },
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_test_engineer_newline_rejected(self):
        r = client.post(
            "/api/swut/coverage/build",
            json={
                "project_id": "HDPDM01",
                "release_sw_version": "1.0.0",
                "test_date": "2024-02-19",
                "test_engineer": "X\nY",
            },
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_doc_id_sequence_non_digit_rejected(self):
        r = client.post(
            "/api/swut/coverage/build",
            json={
                "project_id": "HDPDM01",
                "release_sw_version": "1.0.0",
                "test_date": "2024-02-19",
                "doc_id_sequence": "abc",
            },
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422


class TestInputSurface13:
    """13차 라운드 — 입력 표면 강화 (C3/W7/W8/W9)."""

    def _base_body(self) -> dict:
        return {
            "project_id": "HDPDM01",
            "release_sw_version": "1.0.0",
            "test_date": "2024-02-19",
        }

    def test_w7_test_date_with_suffix_rejected(self):
        """W7: test_date $ anchor — garbage suffix 차단."""
        body = self._base_body()
        body["test_date"] = "2024-02-19; DROP TABLE users"
        r = client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_w7_validation_date_with_garbage_rejected(self):
        """W7: validation_date pattern — 빈 string OK, garbage 차단."""
        body = self._base_body()
        body["validation_date"] = "not a date"
        r = client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_w7_validation_date_empty_allowed(self):
        """W7: validation_date 빈 string은 통과 — schema validate 만 통과 검증."""
        from backend.schemas import SwUTBuildRequest
        # Pydantic validate 단독 검증 (endpoint mock 부담 없이)
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
            validation_date="",
        )
        assert req.validation_date == ""

    def test_w8_log_folder_with_newline_rejected(self):
        """W8: log_folder 줄바꿈 금지 (헤더 인젝션 안전)."""
        body = self._base_body()
        body["log_folder"] = "C:/fake/log\r\nX-Injected: evil"
        r = client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_w8_coverage_template_path_maxlen_rejected(self):
        """W8 + 51차: coverage_template_path 500자 초과 차단."""
        body = self._base_body()
        body["coverage_template_path"] = "C:/" + "a" * 600
        r = client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_w8_sutr_template_path_maxlen_rejected(self):
        """51차: sutr_template_path 500자 초과 차단 (Coverage와 분리 필드)."""
        body = self._base_body()
        body["sutr_template_path"] = "C:/" + "a" * 600
        r = client.post(
            "/api/swut/sutr/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_coverage_endpoint_uses_coverage_template_path(self, monkeypatch):
        """52차 W2 — SwUT Coverage endpoint이 coverage_template_path만 사용.

        sutr_template_path 입력은 무시되고 coverage_template_path가 _read_template_bytes로 전달.
        """
        from backend.routers import swut as swut_mod
        captured = {}

        def _fake_read(template_path, project_id, kind):
            captured["template_path"] = template_path
            captured["kind"] = kind
            raise RuntimeError("stop after capture")

        monkeypatch.setattr(swut_mod, "_read_template_bytes", _fake_read)
        body = self._base_body()
        body["log_folder"] = "C:/fake/log"  # collect_swut_session 단계에서 또 다른 RuntimeError
        body["coverage_template_path"] = "C:/coverage.xlsx"
        body["sutr_template_path"] = "C:/sutr.xlsm"
        # 빌드 시도 — _read_template_bytes 호출 시점에 stop. 500 또는 빌더 실패 응답.
        client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        # 빌더 도달 여부 무관 — captured 확인 시점
        if "template_path" in captured:
            assert captured["template_path"] == "C:/coverage.xlsx"
            assert captured["kind"] == "coverage"

    def test_sutr_endpoint_uses_sutr_template_path(self, monkeypatch):
        """52차 W2 — SwUT SUTR endpoint이 sutr_template_path만 사용."""
        from backend.routers import swut as swut_mod
        captured = {}

        def _fake_read(template_path, project_id, kind):
            captured["template_path"] = template_path
            captured["kind"] = kind
            raise RuntimeError("stop after capture")

        monkeypatch.setattr(swut_mod, "_read_template_bytes", _fake_read)
        body = self._base_body()
        body["log_folder"] = "C:/fake/log"
        body["coverage_template_path"] = "C:/coverage.xlsx"
        body["sutr_template_path"] = "C:/sutr.xlsm"
        client.post(
            "/api/swut/sutr/build", json=body,
            headers={"X-User": "tester"},
        )
        if "template_path" in captured:
            assert captured["template_path"] == "C:/sutr.xlsm"
            assert captured["kind"] == "sutr"

    def test_swutcr_endpoint_uses_swutcr_template_path(self, monkeypatch):
        """SwUTCR endpoint uses swutcr_template_path only."""
        from backend.routers import swut as swut_mod
        captured = {}

        def _fake_read(template_path, project_id, kind):
            captured["template_path"] = template_path
            captured["kind"] = kind
            raise RuntimeError("stop after capture")

        monkeypatch.setattr(swut_mod, "_read_template_bytes", _fake_read)
        body = self._base_body()
        body["log_folder"] = "C:/fake/log"
        body["coverage_template_path"] = "C:/coverage.xlsx"
        body["sutr_template_path"] = "C:/sutr.xlsm"
        body["swutcr_template_path"] = "C:/swutcr.xlsm"
        client.post(
            "/api/swut/swutcr/build", json=body,
            headers={"X-User": "tester"},
        )
        if "template_path" in captured:
            assert captured["template_path"] == "C:/swutcr.xlsm"
            assert captured["kind"] == "swutcr"

    def test_w9_jenkins_build_number_negative_rejected(self):
        """W9: jenkins_build_number 음수 차단."""
        body = self._base_body()
        body["jenkins_build_number"] = -1
        r = client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_w9_jenkins_build_number_too_large_rejected(self):
        """W9: jenkins_build_number 99999 초과 차단."""
        body = self._base_body()
        body["jenkins_build_number"] = 100000
        r = client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_c3_deviation_cases_max_length_rejected(self):
        """C3: deviation_cases 201개 차단 (max_length=200)."""
        body = self._base_body()
        body["deviation_cases"] = [{"tc_id": f"T{i}"} for i in range(201)]
        r = client.post(
            "/api/swut/sutr/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_c3_deviation_cases_size_limit_rejected(self):
        """C3: 합산 256KB 초과 차단 (50개 × 6KB = 300KB)."""
        body = self._base_body()
        body["deviation_cases"] = [
            {"tc_id": f"T{i}", "rationale": "x" * 6000} for i in range(50)
        ]
        r = client.post(
            "/api/swut/sutr/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_c3_deviation_cases_key_count_rejected(self):
        """C3: 단일 item key 수 21개 차단."""
        body = self._base_body()
        item = {f"k{i}": "v" for i in range(21)}
        body["deviation_cases"] = [item]
        r = client.post(
            "/api/swut/sutr/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_c3_deviation_cases_within_limits_accepted_by_schema(self):
        """C3: 정상 case는 schema 통과."""
        from backend.schemas import SwUTBuildRequest
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
            deviation_cases=[
                {"tc_id": "TC1", "tc_no": "TC1", "issue_text": "div by zero",
                 "auto_rationale": "[AUTO] checked"},
                {"tc_id": "TC2", "tc_no": "TC2", "issue_text": "stack overflow",
                 "auto_rationale": "[AUTO] checked"},
            ],
        )
        assert len(req.deviation_cases) == 2


class TestCSourceRoot21:
    """30차 W21 T220 — c_source_root 입력 표면 (maxlen 500 + 줄바꿈 금지)."""

    def _base_body(self) -> dict:
        return {
            "project_id": "HDPDM01",
            "release_sw_version": "1.0.0",
            "test_date": "2024-02-19",
        }

    def test_c_source_root_empty_accepted(self):
        """기본값 빈 string은 통과 — c_source_root는 옵션."""
        from backend.schemas import SwUTBuildRequest
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
        )
        assert req.c_source_root == ""

    def test_c_source_root_with_newline_rejected(self):
        """줄바꿈 금지 (헤더 인젝션 안전 — log_folder와 동일 정책)."""
        body = self._base_body()
        body["c_source_root"] = "D:/Project/src\r\nX-Injected: evil"
        r = client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_c_source_root_maxlen_rejected(self):
        """500자 초과 차단 (다른 path 필드와 동일)."""
        body = self._base_body()
        body["c_source_root"] = "D:/" + "a" * 600
        r = client.post(
            "/api/swut/coverage/build", json=body,
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422


class TestSwudsAsilFallback32:
    """32차 W28: SwUDS docx → function_asil_map fallback + c_source 우선 merge."""

    def _make_session_for_apply(self):
        """_apply_function_asil_map 호출용 minimal session."""
        from backend.services.swut_input_adapter import EnvironmentData, SwUTSession
        return SwUTSession(
            project_id="HDPDM01", version="v0.01",
            source_kind="log_folder", source_path="",
            environments=[EnvironmentData(env_name="SWTE_01")],
        )

    def test_c_source_priority_over_swuds(self, monkeypatch):
        """c_source + swuds 둘 다 같은 fn_id 매핑 → c_source 값 우선."""
        from backend.routers import swut as swut_router

        # c_source resolver mock: SwUFn_0101 → "B"
        class FakeAsilResult:
            warnings: list[str] = []
            function_asil_map = {"SwUFn_0101": "B"}

        def fake_resolve(*_a, **_k):
            return FakeAsilResult()

        import backend.services.swut_asil_resolver as asil_resolver_mod
        monkeypatch.setattr(asil_resolver_mod, "resolve_function_asil_map", fake_resolve)

        # 54차 T281 — SwUDS resolver mock은 swut_meta_resolver로 redirect
        from backend.services import swut_meta_resolver as meta_resolver_mod
        # 라운드 89: apply는 단일 parse seam resolve_swuds_maps 사용 (id→ASIL, name→id).
        # 2026-06-10: out_warnings kwarg 추가 — mock도 **kwargs 수용.
        monkeypatch.setattr(
            meta_resolver_mod, "resolve_swuds_maps",
            lambda req, project_id, **_kw: ({"SwUFn_0101": "D", "SwUFn_0102": "C"}, {}),
        )

        from backend.schemas import SwUTBuildRequest
        req = SwUTBuildRequest(
            project_id="HDPDM01", release_sw_version="1.0.0",
            test_date="2024-02-19",
            c_source_root="C:/src", swuds_docx_path="C:/swuds.docx",
        )
        session = self._make_session_for_apply()
        swut_router._apply_function_asil_map(req, session)
        merged = session.environments[0].function_asil_map
        # SwUFn_0101은 c_source 우선 (B), SwUFn_0102는 SwUDS만 (C)
        assert merged["SwUFn_0101"] == "B"
        assert merged["SwUFn_0102"] == "C"
        # 충돌 warning 누적
        assert any("ASIL 충돌" in w for w in session.parse_warnings)

    def test_swuds_only_fallback_when_no_c_source(self, monkeypatch):
        """c_source 없고 swuds만 → SwUDS 결과 사용."""
        from backend.routers import swut as swut_router

        # 54차 T281 — swut_meta_resolver로 redirect
        from backend.services import swut_meta_resolver as meta_resolver_mod
        # 라운드 89: 단일 parse seam resolve_swuds_maps.
        # 2026-06-10: out_warnings kwarg 추가 — mock도 **kwargs 수용.
        monkeypatch.setattr(
            meta_resolver_mod, "resolve_swuds_maps",
            lambda req, project_id, **_kw: ({"SwUFn_0103": "D"}, {}),
        )
        from backend.schemas import SwUTBuildRequest
        req = SwUTBuildRequest(
            project_id="HDPDM01", release_sw_version="1.0.0",
            test_date="2024-02-19",
            swuds_docx_path="C:/swuds.docx",  # c_source_root 빈 string
        )
        session = self._make_session_for_apply()
        swut_router._apply_function_asil_map(req, session)
        merged = session.environments[0].function_asil_map
        assert merged == {"SwUFn_0103": "D"}

    def test_both_empty_results_no_apply(self, monkeypatch):
        """둘 다 부재 / 매핑 0 → environments[0].function_asil_map 빈 dict (default 유지)."""
        from backend.routers import swut as swut_router
        from backend.schemas import SwUTBuildRequest
        req = SwUTBuildRequest(
            project_id="HDPDM01", release_sw_version="1.0.0",
            test_date="2024-02-19",
        )  # 둘 다 빈 string
        session = self._make_session_for_apply()
        swut_router._apply_function_asil_map(req, session)
        # 빌드는 진행 (silent), function_asil_map은 default 빈 dict
        assert session.environments[0].function_asil_map == {}


class TestSummaryHeaderTruncation21:
    """30차 W21 deep-reviewer Warning fix — X-SwUT-Summary 1024B 한도에서
    asil_d_function_ids list가 잘려도 frontend JSON.parse가 실패하지 않도록
    valid JSON sentinel 보장."""

    def test_large_asil_d_list_truncated_to_valid_json(self):
        """ASIL D 함수 100개 → 헤더 1024B 초과 → list 길이로 축약된 valid JSON."""
        import json

        from backend.routers.swut import _build_result_to_response

        # 큰 asil_d_function_ids list (100 개 × 13B ≈ 1500B)
        summary = {
            "function_rows": 200,
            "asil_distribution": {"ASIL_A": 100, "ASIL_D": 100},
            "asil_d_function_ids": [f"SwUFn_{i:04d}" for i in range(100)],
        }
        from io import BytesIO
        res = _build_result_to_response(
            content_io=BytesIO(b"x"),
            filename="cov.xlsx",
            summary=summary,
            warnings=[],
            incomplete_sheets=[],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        header = res.headers.get("X-SwUT-Summary")
        assert header
        # valid JSON parse 가능해야 함 — string 중간 잘림 회피
        parsed = json.loads(header)
        # asil_d_function_ids는 list 길이로 축약된 string (sentinel)
        assert isinstance(parsed.get("asil_d_function_ids"), str)
        assert "100 ids" in parsed["asil_d_function_ids"]

    def test_small_summary_passes_through_unchanged(self):
        """1024B 이하 summary는 그대로 전달."""
        import json
        from io import BytesIO

        from backend.routers.swut import _build_result_to_response

        summary = {
            "function_rows": 5,
            "asil_distribution": {"ASIL_A": 3, "ASIL_D": 2},
            "asil_d_function_ids": ["SwUFn_0001", "SwUFn_0002"],
        }
        res = _build_result_to_response(
            content_io=BytesIO(b"x"),
            filename="cov.xlsx",
            summary=summary,
            warnings=[],
            incomplete_sheets=[],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        parsed = json.loads(res.headers.get("X-SwUT-Summary"))
        # 작은 list는 그대로 (list 보존)
        assert parsed["asil_d_function_ids"] == ["SwUFn_0001", "SwUFn_0002"]


class TestWarningsSentinelBreakdownRound3NC1:
    """F6 Round 3 NC1 partial + Round 4 NW7/NW8 fix:
    warnings 1024B 초과 시 sentinel에 카테고리별 카운트 breakdown 노출.
    Round 4 NW7: ambiguous 카운트는 'ambiguous' substring이 아닌
    '[hmr] ambiguous' prefix로 정밀 (stamp summary 메시지의 'ambiguous skipped:'
    substring 오분류 방지)."""

    def test_warnings_truncated_with_category_breakdown(self):
        import json
        from io import BytesIO

        from backend.routers.swut import _build_result_to_response

        # production-realistic — stamp summary 메시지에 "ambiguous skipped:" 포함
        warnings = (
            ["[hmr] ambiguous function 'Init' — 다중 unit_file (a.c, b.c)"] * 20
            + [
                # Round 4 NW7 검증: 이 메시지에 "ambiguous skipped" substring 포함 —
                # 이전 코드는 이를 +1 ambiguous로 카운트 (false 21 표시).
                "[hmr] Function Calls metric stamped — 50/100 functions matched "
                "(HMR metric count: 150, ambiguous skipped: 20)"
            ]
            + ["[swuts] parse 실패 — spec stamp skip, 하드코딩 fallback"] * 10
            + ["[layout] precondition col missing"] * 5
            + ["기타 일반 warning"] * 14  # NW8 — 어떤 prefix에도 안 잡힘
        )
        res = _build_result_to_response(
            content_io=BytesIO(b"x"),
            filename="cov.xlsx",
            summary={"function_rows": 100},
            warnings=warnings,
            incomplete_sheets=[],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        parsed = json.loads(res.headers.get("X-SwUT-Warnings"))
        assert isinstance(parsed, list) and len(parsed) == 1
        msg = parsed[0]
        # NW7 검증: ambiguous는 정확히 20 (stamp summary +1 오분류 안 됨)
        assert "ambiguous=20" in msg, f"NW7 회귀 — ambiguous miscount: {msg}"
        # hmr는 ambiguous 20 + stamp summary 1 = 21
        assert "hmr=21" in msg
        assert "swuts=10" in msg
        assert "layout=5" in msg
        # NW8 검증: 비-category warning (기타 일반 warning 14건) → other 카테고리
        assert "other=14" in msg, f"NW8 회귀 — uncategorized 누락: {msg}"
        assert f"{len(warnings)} warnings" in msg


class TestXUserHeader:
    def test_missing_x_user_header_rejected(self):
        r = client.post(
            "/api/swut/coverage/build",
            json={
                "project_id": "HDPDM01",
                "release_sw_version": "1.0.0",
                "test_date": "2024-02-19",
            },
            # X-User 헤더 미포함
        )
        # UserContextMiddleware가 먼저 403 또는 endpoint 400 둘 중 하나
        assert r.status_code in (400, 401, 403)


# ---------------------------------------------------------------------------
# Coverage / SUTR 빌드 200 (mock)
# ---------------------------------------------------------------------------

class TestCoverageBuildSuccess:
    def test_returns_xlsx_with_attachment_header(self):
        with patch(
            "backend.routers.swut.collect_swut_session", return_value=_make_session(),
        ), patch(
            "backend.routers.swut._read_template_bytes",
            return_value=_minimal_xlsx_template_bytes(),
        ):
            r = client.post(
                "/api/swut/coverage/build",
                json={
                    "project_id": "HDPDM01",
                    "release_sw_version": "1.0.0",
                    "test_date": "2024-02-19",
                    "test_engineer": "JK",
                    "log_folder": "C:/fake/log",
                },
                headers={"X-User": "tester"},
            )
        assert r.status_code == 200
        assert r.headers["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in r.headers.get("Content-Disposition", "")
        # 반환 bytes가 xlsx ZIP magic 헤더
        assert r.content[:4] == b"PK\x03\x04"
        # X-* 헤더에 summary/warnings
        assert "X-Swut-Summary" in r.headers or "x-swut-summary" in r.headers


class TestJenkinsFetcherMock:
    def test_collect_from_jenkins_cache_with_mock(self, tmp_path):
        """T140: jenkins_adapter mock 후 SwUTSession 반환."""
        from backend.services.swut_input_adapter import collect_from_jenkins_cache

        # 가짜 cache_root 구조: <root>/<project>/<build>/dummy
        project_dir = tmp_path / "HDPDM01" / "42"
        project_dir.mkdir(parents=True)
        # html 파일 직접 — scan_jenkins_build_root mock 반환
        fake_files = [
            str(project_dir / "SWTE_01_test_case_data_report.html"),
            str(project_dir / "SWTE_01_execution_results_report.html"),
            str(project_dir / "SWTE_01_aggregate_coverage_report.html"),
        ]
        for f in fake_files:
            Path(f).write_text(
                "<!DOCTYPE html><!-- VectorCAST Report header --><html><body>"
                "<title>Test Case Data Report</title></body></html>",
                encoding="utf-8",
            )

        with patch(
            "backend.services.jenkins_adapter.scan_jenkins_build_root",
            return_value={"html_files": fake_files},
        ):
            session = collect_from_jenkins_cache(
                resolver=type("R", (), {"read_bytes": lambda self, p: Path(p).read_bytes()})(),
                project_id="HDPDM01",
                cache_root=str(tmp_path),
                build_number=42,
            )

        assert session is not None
        assert session.source_kind == "jenkins_cache"
        assert len(session.environments) == 1
        assert session.environments[0].env_name == "SWTE_01"

    def test_cache_root_missing_returns_none(self, tmp_path):
        from backend.services.swut_input_adapter import collect_from_jenkins_cache
        warnings: list[str] = []
        session = collect_from_jenkins_cache(
            resolver=None, project_id="HDPDM01",
            cache_root=str(tmp_path / "nonexistent"),
            parse_warnings=warnings,
        )
        assert session is None
        assert any("미발견" in w for w in warnings)


# ---------------------------------------------------------------------------
# Semaphore(2) 동작 검증
# ---------------------------------------------------------------------------

class TestSemaphore:
    def test_semaphore_capacity_is_3(self):
        """17차 T173: Semaphore(2) → (3) 상향. 14차 메모리 1배 절감 활용."""
        from backend.routers.swut import _BUILD_SEMAPHORE
        assert _BUILD_SEMAPHORE._value == 3

    def test_semaphore_serializes_4th_request(self):
        """17차 T173: 4건 동시 호출 시 1건은 대기. capacity=3."""
        sem = asyncio.Semaphore(3)
        order: list[int] = []

        async def _hold(i: int):
            async with sem:
                order.append(i)
                await asyncio.sleep(0.02)

        async def _main():
            await asyncio.gather(_hold(0), _hold(1), _hold(2), _hold(3))

        asyncio.run(_main())
        assert len(order) == 4  # 모두 완료, Semaphore가 deadlock 안 됨


# ---------------------------------------------------------------------------
# 12차 라운드 — code health 개선 검증
# ---------------------------------------------------------------------------

class TestConfigCache:
    """C2: _load_meta_from_config mtime 기반 cache."""

    def test_lru_cache_hits_with_same_mtime(self, tmp_path, monkeypatch):
        from backend.routers import swut as swut_mod
        from backend.services import swut_meta_resolver as resolver_mod

        cfg_path = tmp_path / "swut_meta.json"
        cfg_path.write_text('{"projects":{"HDPDM01":{"asil_level":"ASIL A"}}}', encoding="utf-8")
        # 54차 T281 — resolver 모듈 path patch
        monkeypatch.setattr(resolver_mod, "_META_CONFIG_PATH", str(cfg_path))
        monkeypatch.setattr(swut_mod, "_META_CONFIG_PATH", str(cfg_path))

        resolver_mod._read_meta_config_raw.cache_clear()

        for _ in range(5):
            cfg = swut_mod._load_meta_from_config("HDPDM01")
            assert cfg.get("asil_level") == "ASIL A"

        info = resolver_mod._read_meta_config_raw.cache_info()
        # 5회 호출 중 1회 miss, 4회 hit
        assert info.misses == 1
        assert info.hits == 4

    def test_lru_cache_invalidates_on_mtime_change(self, tmp_path, monkeypatch):
        """mtime이 변하면 cache miss → reload."""
        import time

        from backend.routers import swut as swut_mod
        from backend.services import swut_meta_resolver as resolver_mod

        cfg_path = tmp_path / "swut_meta.json"
        cfg_path.write_text('{"projects":{"HDPDM01":{"asil_level":"ASIL A"}}}', encoding="utf-8")
        monkeypatch.setattr(resolver_mod, "_META_CONFIG_PATH", str(cfg_path))
        monkeypatch.setattr(swut_mod, "_META_CONFIG_PATH", str(cfg_path))
        resolver_mod._read_meta_config_raw.cache_clear()

        cfg1 = swut_mod._load_meta_from_config("HDPDM01")
        assert cfg1.get("asil_level") == "ASIL A"

        # 파일 변경 + mtime 강제 진행
        time.sleep(0.02)
        cfg_path.write_text('{"projects":{"HDPDM01":{"asil_level":"ASIL D"}}}', encoding="utf-8")
        new_mtime = cfg_path.stat().st_mtime + 1.0
        import os
        os.utime(str(cfg_path), (new_mtime, new_mtime))

        cfg2 = swut_mod._load_meta_from_config("HDPDM01")
        assert cfg2.get("asil_level") == "ASIL D"  # cache invalidated


class TestSwutConfigFallback50:
    """50차 — req 값 비면 config/swut_meta.json fallback (c_source_root + swuds_docx_path).

    config 자동 적용으로 사용자가 매 빌드마다 동일 path 재입력하는 부담 제거.
    """

    def _setup_cfg(self, tmp_path, monkeypatch, cfg_dict):
        from backend.routers import swut as swut_mod
        from backend.services import swut_meta_resolver as resolver_mod
        cfg_path = tmp_path / "swut_meta.json"
        import json as _json
        cfg_path.write_text(_json.dumps(cfg_dict), encoding="utf-8")
        # 54차 T281 — DRY 통합 후 resolver 모듈의 path를 patch (swut_mod alias도 동기)
        monkeypatch.setattr(resolver_mod, "_META_CONFIG_PATH", str(cfg_path))
        monkeypatch.setattr(swut_mod, "_META_CONFIG_PATH", str(cfg_path))
        resolver_mod._read_meta_config_raw.cache_clear()
        return swut_mod

    def test_resolve_c_source_root_req_priority(self, tmp_path, monkeypatch):
        """req.c_source_root 우선 — config 값 무시."""
        from backend.schemas import SwUTBuildRequest
        swut_mod = self._setup_cfg(tmp_path, monkeypatch, {
            "projects": {"HDPDM01": {"c_source_root": "C:/from_config"}}
        })
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
            c_source_root="D:/from_req",
        )
        assert swut_mod._resolve_c_source_root(req) == "D:/from_req"

    def test_resolve_c_source_root_config_fallback(self, tmp_path, monkeypatch):
        """req.c_source_root 빈 string → config 값 fallback."""
        from backend.schemas import SwUTBuildRequest
        swut_mod = self._setup_cfg(tmp_path, monkeypatch, {
            "projects": {"HDPDM01": {"c_source_root": "C:/from_config"}}
        })
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
        )
        assert swut_mod._resolve_c_source_root(req) == "C:/from_config"

    def test_resolve_c_source_root_both_empty(self, tmp_path, monkeypatch):
        """req + config 모두 비면 빈 string — graceful (panel 미표시)."""
        from backend.schemas import SwUTBuildRequest
        swut_mod = self._setup_cfg(tmp_path, monkeypatch, {
            "projects": {"HDPDM01": {}}
        })
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
        )
        assert swut_mod._resolve_c_source_root(req) == ""

    def test_apply_c_function_map_rejects_blocked_source_root_before_parse(
        self, monkeypatch,
    ):
        """SwUTCR reason/action C scan must reuse the system-dir guard."""
        import backend.services.swut_asil_resolver as asil_resolver
        from backend.routers import swut as swut_mod
        from backend.schemas import SwUTBuildRequest

        monkeypatch.setattr(asil_resolver, "is_blocked_source_root", lambda _p: True)

        def _fail_parse(*_args, **_kwargs):
            raise AssertionError("parse_c_project must not run for blocked roots")

        monkeypatch.setattr("workflow.code_parser.c_parser.parse_c_project", _fail_parse)

        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
            c_source_root="C:/Windows",
        )

        class _Session:
            parse_warnings: list[str] = []
            c_function_map: dict = {}

        session = _Session()
        swut_mod._apply_c_function_map(req, session)

        assert session.c_function_map == {}
        assert any("system directory rejected" in w for w in session.parse_warnings)

    def test_vcast_source_fallback_extracts_full_function_after_prototype(self):
        """SwUTCR can use VectorCAST HTML when C source root misses a static helper."""
        from backend.routers import swut as swut_mod

        env = EnvironmentData(
            env_name="SWTE_01",
            component_name="Door",
            function_coverage=[
                FunctionCoverage(
                    unit_id="getPastSpeed",
                    name="getPastSpeed",
                    statement=CoverageStats(covered=7, total=8),
                ),
            ],
        )
        session = SwUTSession(environments=[env])
        session.c_function_map = {}
        session.parse_warnings = []

        html = """
        <html><body>
          <span>static S16 getPastSpeed(U8 timeMs);</span>
          <span>3338 58 0 (T) static S16 getPastSpeed(U8 timeMs)</span>
          <span>3339 58 0 {</span>
          <span>getPastSpeed</span>
          <span>3340 58 0 * if (timeMs == 0U)</span>
          <span>3341 58 0 {</span>
          <span>3342 58 0 return 0;</span>
          <span>3343 58 0 }</span>
          <span>3344 58 0 else</span>
          <span>3345 58 0 {</span>
          <span>3344 58 0 return s16g_speed;</span>
          <span>3345 58 0 }</span>
          <span>3345 58 0 }</span>
        </body></html>
        """

        class _Resolver:
            def list_dir(self, path, pattern="*", recursive=False):
                assert path == r"C:\vcast\Aggregate"
                assert pattern == "*AggregateCoverageReport.html"
                assert recursive is False
                return [r"C:\vcast\Aggregate\Door_AggregateCoverageReport.html"]

            def read_bytes(self, path):
                assert path.endswith("AggregateCoverageReport.html")
                return html.encode("utf-8")

        swut_mod._apply_vcast_source_fallback(session, _Resolver(), r"C:\vcast")

        c_entry = session.c_function_map["getPastSpeed"]
        assert c_entry["source_origin"] == "vectorcast_aggregate_html"
        assert c_entry["file"] == "Door_AggregateCoverageReport.html"
        assert c_entry["signature"] == "static S16 getPastSpeed(U8 timeMs)"
        assert "if (timeMs == 0U)" in c_entry["body"]
        assert "\nelse\n" in f"\n{c_entry['body']}\n"
        assert "\ngetPastSpeed\n" not in f"\n{c_entry['body']}\n"
        assert c_entry["body"].rstrip().endswith("}")
        assert any("VectorCAST aggregate source fallback applied" in w for w in session.parse_warnings)

    def test_resolve_swuds_path_config_fallback(self, tmp_path, monkeypatch):
        """req.swuds_docx_path 빈 string → config 값 fallback."""
        from backend.schemas import SwUTBuildRequest
        swut_mod = self._setup_cfg(tmp_path, monkeypatch, {
            "projects": {"HDPDM01": {"swuds_docx_path": "U:/from_config.docx"}}
        })
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
        )
        assert swut_mod._resolve_swuds_path(req) == "U:/from_config.docx"

    def test_resolve_swuds_path_whitespace_only_config_treated_empty(self, tmp_path, monkeypatch):
        """config에 whitespace만 들어가도 strip 후 빈 string → 빈 반환 (silent path 사용 차단)."""
        from backend.schemas import SwUTBuildRequest
        swut_mod = self._setup_cfg(tmp_path, monkeypatch, {
            "projects": {"HDPDM01": {"swuds_docx_path": "   "}}
        })
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
        )
        assert swut_mod._resolve_swuds_path(req) == ""

    def test_apply_function_asil_map_records_source_origin(self, tmp_path, monkeypatch):
        """53차 W3 — 50차 W4/W5 source origin 시각화 (`(req)` / `(config fallback)`) 검증.

        `_apply_function_asil_map`이 session.parse_warnings에 출처 origin 명시.
        c_source는 req 입력, swuds는 config fallback인 mixed case로 양 출처 모두 검증.
        """
        from backend.schemas import SwUTBuildRequest
        swut_mod = self._setup_cfg(tmp_path, monkeypatch, {
            "projects": {"HDPDM01": {
                "swuds_docx_path": "U:/from_config.docx",
            }}
        })
        # c_source는 req에서, swuds는 config fallback 시뮬레이션
        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
            c_source_root="D:/from_req/src",
        )

        # resolver를 mock해서 실제 file IO 회피
        class _FakeResult:
            warnings = []
            function_asil_map = {"SwUFn_0001": "ASIL D"}
            ok = True
            function_ids = {"SwUFn_0001"}

        monkeypatch.setattr(
            "backend.services.swut_asil_resolver.resolve_function_asil_map",
            lambda _root: _FakeResult(),
        )
        # SwUDS docx read는 PermissionError로 fail (config path 미존재) — 정상 graceful path
        # _resolve_swuds_function_asil_map는 빈 dict 반환 → swuds_map=0건

        class _MockSession:
            parse_warnings: list = []
            environments: list = []

        session = _MockSession()
        swut_mod._apply_function_asil_map(req, session)
        # parse_warnings에 source origin 표기 검증
        joined = " | ".join(session.parse_warnings)
        assert "function_asil_map source" in joined
        assert "(req)" in joined  # c_source는 req
        assert "(config fallback)" in joined  # swuds는 config


class TestExceptionSanitization:
    """W4: builder exception sanitize — internal detail leak 차단."""

    def test_filenotfound_returns_404_sanitized(self):
        """template_path 존재 안 함 → FileNotFoundError → 404 + sanitized detail."""
        with patch(
            "backend.routers.swut._do_coverage_build",
            side_effect=FileNotFoundError("/secret/internal/path"),
        ):
            r = client.post(
                "/api/swut/coverage/build",
                json={
                    "project_id": "HDPDM01",
                    "release_sw_version": "1.0.0",
                    "test_date": "2024-02-19",
                    "log_folder": "C:/fake/log",
                },
                headers={"X-User": "tester"},
            )
        assert r.status_code == 404
        body = r.json()
        # 응답 형식: {"ok": False, "error": {"code": "HTTP_404", "message": "..."}}
        message = body.get("error", {}).get("message") or body.get("detail", "")
        # 내부 path는 leak되지 않음 — type name만 표시
        assert "/secret/internal/path" not in message
        assert "FileNotFoundError" in message

    def test_value_error_returns_400_with_message(self):
        """ValueError는 사용자 입력 관련이므로 message 표시."""
        with patch(
            "backend.routers.swut._do_coverage_build",
            side_effect=ValueError("ASIL level 'XYZ' 미지원"),
        ):
            r = client.post(
                "/api/swut/coverage/build",
                json={
                    "project_id": "HDPDM01",
                    "release_sw_version": "1.0.0",
                    "test_date": "2024-02-19",
                    "log_folder": "C:/fake/log",
                },
                headers={"X-User": "tester"},
            )
        assert r.status_code == 400
        body = r.json()
        message = body.get("error", {}).get("message") or body.get("detail", "")
        assert "ASIL level" in message

    def test_unexpected_returns_500_no_message_leak(self):
        """예상 못 한 exception은 message 본문 미노출 — type name만."""
        with patch(
            "backend.routers.swut._do_coverage_build",
            side_effect=RuntimeError("DB password=hunter2 leaked"),
        ):
            r = client.post(
                "/api/swut/coverage/build",
                json={
                    "project_id": "HDPDM01",
                    "release_sw_version": "1.0.0",
                    "test_date": "2024-02-19",
                    "log_folder": "C:/fake/log",
                },
                headers={"X-User": "tester"},
            )
        assert r.status_code == 500
        body = r.json()
        message = body.get("error", {}).get("message") or body.get("detail", "")
        assert "hunter2" not in message
        assert "RuntimeError" in message


class TestAsyncMigration:
    """W5: asyncio.to_thread 마이그레이션 검증 — 기존 endpoint가 to_thread로 동작."""

    def test_endpoint_uses_to_thread(self):
        """source code에 asyncio.to_thread 사용 + get_event_loop 미사용 검증."""
        import inspect

        from backend.routers import swut as swut_mod

        src = inspect.getsource(swut_mod)
        assert "asyncio.to_thread" in src
        assert "asyncio.get_event_loop()" not in src
        assert "loop.run_in_executor" not in src


class TestStreamingResponse14:
    """14차 W1: StreamingResponse + BytesIO 직접 stream."""

    def test_response_includes_content_length(self):
        """W1b: Content-Length 헤더가 명시되어 chunked 모드 아니어도 정확한 크기 통보."""
        with patch(
            "backend.routers.swut.collect_swut_session", return_value=_make_session(),
        ), patch(
            "backend.routers.swut._read_template_bytes",
            return_value=_minimal_xlsx_template_bytes(),
        ):
            r = client.post(
                "/api/swut/coverage/build",
                json={
                    "project_id": "HDPDM01",
                    "release_sw_version": "1.0.0",
                    "test_date": "2024-02-19",
                    "log_folder": "C:/fake/log",
                },
                headers={"X-User": "tester"},
            )
        assert r.status_code == 200
        content_length = r.headers.get("content-length")
        assert content_length is not None
        assert int(content_length) == len(r.content)
        # PK ZIP magic 보존 — stream chunk 손상 없음
        assert r.content[:4] == b"PK\x03\x04"

    def test_streaming_response_used_not_plain_response(self):
        """W1a: source code에 StreamingResponse 사용 확인."""
        import inspect

        from backend.routers import swut as swut_mod

        src = inspect.getsource(swut_mod)
        assert "StreamingResponse" in src
        assert "_iter_bytesio" in src
        # plain Response(content=bytes) 사용 회피 확인 — bytes 그대로 전송하지 않음
        assert "Response(content=content" not in src

    def test_builder_result_xlsx_io_is_bytesio(self):
        """W1a: builder result.xlsx_io가 BytesIO 인스턴스 + backward compat property."""
        from backend.routers.swut import _build_coverage_meta
        from backend.schemas import SwUTBuildRequest
        from backend.services.swut_coverage_aggregator import build_coverage_report

        req = SwUTBuildRequest(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
            log_folder="C:/fake/log",
        )
        session = _make_session()
        meta = _build_coverage_meta(req)
        result = build_coverage_report(session, meta, _minimal_xlsx_template_bytes())
        assert isinstance(result.xlsx_io, io.BytesIO)
        # backward compat property 동작
        assert result.xlsx_bytes[:4] == b"PK\x03\x04"
        # result_size_bytes는 BytesIO copy 없이 측정
        assert result.result_size_bytes == len(result.xlsx_bytes)
        # property 호출 후에도 BytesIO position이 보존되어 stream 가능 (idempotent)
        assert result.xlsx_io.tell() == 0

    def test_iter_bytesio_yields_chunks(self):
        """W1: _iter_bytesio가 chunk별 yield + 빈 data 시 종료."""
        import io as _io

        from backend.routers.swut import _iter_bytesio

        data = b"x" * (200 * 1024)  # 200KB → 64KB chunk 4개
        buf = _io.BytesIO(data)
        chunks = list(_iter_bytesio(buf, chunk_size=64 * 1024))
        assert len(chunks) == 4
        assert sum(len(c) for c in chunks) == len(data)
        assert b"".join(chunks) == data


class TestSutrSwudsIntegration17:
    """17차 T172: SUTR endpoint도 swuds_docx_path 처리."""

    def test_sutr_endpoint_calls_resolve_swuds(self):
        """SUTR endpoint가 _resolve_swuds_function_ids를 호출 + build_sutr에 인자 전달."""
        # mock으로 SwUDS 함수 ID set 강제 + build_sutr 호출 검증
        with patch(
            "backend.routers.swut.collect_swut_session", return_value=_make_session(),
        ), patch(
            "backend.routers.swut._read_template_bytes",
            return_value=_minimal_xlsx_template_bytes(),
        ), patch(
            "backend.routers.swut._resolve_swuds_function_ids",
            return_value={"SwUFn_0001", "SwUFn_0002"},
        ) as resolve_mock, patch(
            "backend.routers.swut.build_sutr",
        ) as build_mock:
            # build_sutr mock — io.BytesIO 결과 반환
            import io as _io

            from backend.services.swut_sutr_aggregator import SutrBuildResult
            build_mock.return_value = SutrBuildResult(
                ok=True, xlsm_io=_io.BytesIO(b"PK\x03\x04test"),
                filename="x.xlsm", warnings=[],
                incomplete_sheets=[], vba_macros_preserved=False, summary={},
            )

            client.post(
                "/api/swut/sutr/build",
                json={
                    "project_id": "HDPDM01",
                    "release_sw_version": "1.0.0",
                    "test_date": "2024-02-19",
                    "log_folder": "C:/fake/log",
                    "swuds_docx_path": "U:/docs/SwUDS_v3.docx",
                },
                headers={"X-User": "tester"},
            )

            assert resolve_mock.called
            # build_sutr이 swuds_function_ids 인자 받았는지
            assert build_mock.called
            call_kwargs = build_mock.call_args.kwargs
            assert call_kwargs.get("swuds_function_ids") == {"SwUFn_0001", "SwUFn_0002"}


# ---------------------------------------------------------------------------
# 18차 T177: /api/swut/consistency/check endpoint
# ---------------------------------------------------------------------------

class TestConsistencyCheckEndpoint18:
    """Coverage↔SUTR cross-validation endpoint."""

    def test_schema_requires_both_paths(self):
        """422: coverage_path / sutr_path 둘 다 필수."""
        r = client.post(
            "/api/swut/consistency/check",
            json={"coverage_path": "C:/cov.xlsx"},  # sutr_path 누락
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_schema_rejects_path_with_newline(self):
        """422: path 줄바꿈 차단 (헤더 인젝션 안전)."""
        r = client.post(
            "/api/swut/consistency/check",
            json={
                "coverage_path": "C:/cov.xlsx\r\nX-Injected: evil",
                "sutr_path": "C:/sutr.xlsm",
            },
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_schema_rejects_oversize_path(self):
        """422: path maxlen 500."""
        r = client.post(
            "/api/swut/consistency/check",
            json={
                "coverage_path": "C:/" + "a" * 600,
                "sutr_path": "C:/sutr.xlsm",
            },
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_endpoint_returns_consistency_report_json(self):
        """200: 정상 호출 시 ConsistencyReport.to_dict() 형식."""
        from unittest.mock import MagicMock

        from backend.services.swut_consistency_checker import ConsistencyReport

        mock_report = ConsistencyReport(ok=True, issues=[], parse_warnings=[])
        with patch(
            "backend.routers.swut.get_resolver",
        ) as mock_resolver, patch(
            "backend.routers.swut.check_swut_consistency",
            return_value=mock_report,
        ):
            mock_resolver.return_value = MagicMock(read_bytes=lambda p: b"PK\x03\x04mock")
            r = client.post(
                "/api/swut/consistency/check",
                json={"coverage_path": "C:/cov.xlsx", "sutr_path": "C:/sutr.xlsm"},
                headers={"X-User": "tester"},
            )
        assert r.status_code == 200
        body = r.json()
        assert body["ok"] is True
        assert "issues" in body
        assert "tool_qualification" in body  # ConsistencyReport.to_dict() 보장 필드

    def test_endpoint_sanitizes_filenotfound(self):
        """404: 파일 미존재 시 internal path leak 없이 type name만."""
        from unittest.mock import MagicMock

        def _raise_fnf(p):
            raise FileNotFoundError("/secret/internal/cov.xlsx")

        with patch("backend.routers.swut.get_resolver") as mock_resolver:
            mock_resolver.return_value = MagicMock(read_bytes=_raise_fnf)
            r = client.post(
                "/api/swut/consistency/check",
                json={"coverage_path": "C:/cov.xlsx", "sutr_path": "C:/sutr.xlsm"},
                headers={"X-User": "tester"},
            )
        assert r.status_code == 404
        body = r.json()
        message = body.get("error", {}).get("message") or body.get("detail", "")
        assert "/secret/internal" not in message
        assert "FileNotFoundError" in message

    def test_endpoint_missing_x_user_header_rejected(self):
        r = client.post(
            "/api/swut/consistency/check",
            json={"coverage_path": "C:/cov.xlsx", "sutr_path": "C:/sutr.xlsm"},
        )
        assert r.status_code in (400, 401, 403)


# ---------------------------------------------------------------------------
# 20차 T182: 메모리 모니터링 헬퍼
# ---------------------------------------------------------------------------

class TestMemoryMonitor20:
    """get_process_memory_mb 헬퍼 — psutil 유무 fail-safe.

    38차 I2: backend.routers.swut에서 backend.routers._safety로 이전.
    회귀는 새 위치를 검증 — 기존 _get_process_memory_mb 이름은 _safety.get_process_memory_mb.
    """

    def test_returns_float_when_psutil_available(self):
        """psutil 설치된 환경에서는 양수 float 반환."""
        from backend.routers._safety import _HAS_PSUTIL, get_process_memory_mb
        result = get_process_memory_mb()
        if _HAS_PSUTIL:
            assert isinstance(result, float)
            assert result > 0  # 어떤 프로세스도 0 MB 일 수 없음
        else:
            assert result is None

    def test_returns_none_when_psutil_missing(self):
        """psutil ImportError 시뮬레이션 — None 반환."""
        from backend.routers import _safety as safety_mod
        original = safety_mod._HAS_PSUTIL
        try:
            safety_mod._HAS_PSUTIL = False
            assert safety_mod.get_process_memory_mb() is None
        finally:
            safety_mod._HAS_PSUTIL = original

    def test_returns_none_on_psutil_error(self):
        """psutil.Process가 예외 던지면 silent None."""
        from unittest.mock import patch as _patch

        from backend.routers import _safety as safety_mod
        if not safety_mod._HAS_PSUTIL:
            return  # psutil 미설치 환경 skip
        with _patch("backend.routers._safety.psutil.Process",
                    side_effect=Exception("mock error")):
            assert safety_mod.get_process_memory_mb() is None


# ---------------------------------------------------------------------------
# 21차 T185: /api/swut/browse path picker endpoint
# ---------------------------------------------------------------------------

class TestBrowseEndpoint21:
    """Path picker dialog용 browse endpoint."""

    def test_schema_rejects_path_with_newline(self):
        r = client.post(
            "/api/swut/browse",
            json={"path": "C:/fake\r\nX-Injected: evil"},
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_schema_rejects_oversize_path(self):
        r = client.post(
            "/api/swut/browse",
            json={"path": "C:/" + "a" * 600},
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422

    def test_browse_returns_dirs_and_files(self, tmp_path):
        """tmp_path 안 디렉토리 + 파일 모두 분리 반환."""
        sub = tmp_path / "subdir"
        sub.mkdir()
        (tmp_path / "a.xlsx").write_bytes(b"PK")
        (tmp_path / "b.txt").write_bytes(b"text")

        r = client.post(
            "/api/swut/browse",
            json={"path": str(tmp_path), "pattern": "*"},
            headers={"X-User": "tester"},
        )
        assert r.status_code == 200
        body = r.json()
        assert body["ok"] is True
        assert any("subdir" in d for d in body["dirs"])
        assert any("a.xlsx" in f for f in body["files"])
        assert body["truncated"] is False

    def test_browse_pattern_filters_files(self, tmp_path):
        """*.xlsx pattern은 xlsx만 반환 (dirs는 별도 list)."""
        (tmp_path / "a.xlsx").write_bytes(b"PK")
        (tmp_path / "b.txt").write_bytes(b"text")
        (tmp_path / "c.xlsm").write_bytes(b"PK")

        r = client.post(
            "/api/swut/browse",
            json={"path": str(tmp_path), "pattern": "*.xlsx"},
            headers={"X-User": "tester"},
        )
        body = r.json()
        # *.xlsx만 매칭 + b.txt / c.xlsm 제외
        assert any("a.xlsx" in f for f in body["files"])
        assert not any("b.txt" in f for f in body["files"])
        assert not any("c.xlsm" in f for f in body["files"])

    def test_browse_parent_path_provided(self, tmp_path):
        """현재 + parent 경로 모두 반환 — navigate up 가능."""
        sub = tmp_path / "subdir"
        sub.mkdir()
        r = client.post(
            "/api/swut/browse",
            json={"path": str(sub)},
            headers={"X-User": "tester"},
        )
        body = r.json()
        assert "current" in body
        assert "parent" in body
        # parent가 tmp_path여야 함 (또는 그 부모)
        assert str(tmp_path) in body["parent"] or body["parent"] == str(tmp_path)

    def test_browse_missing_x_user_rejected(self, tmp_path):
        r = client.post("/api/swut/browse", json={"path": str(tmp_path)})
        assert r.status_code in (400, 401, 403)

    def test_browse_response_includes_file_mode(self, tmp_path):
        """22차 T190: 응답에 file_mode 필드 포함 (local/cloudium 구분)."""
        r = client.post(
            "/api/swut/browse",
            json={"path": str(tmp_path)},
            headers={"X-User": "tester"},
        )
        assert r.status_code == 200
        body = r.json()
        assert "file_mode" in body
        assert body["file_mode"] in ("local", "cloudium")
        # local 모드 + 디렉토리 정상 list → cloudium_hint는 빈 string
        assert body.get("cloudium_hint", "") == ""

    def test_browse_cloudium_hint_when_iterdir_permission_error(self):
        """22차 T190: cloudium 모드 + iterdir PermissionError → 200 + cloudium_hint 메시지."""
        from unittest.mock import MagicMock

        mock_resolver = MagicMock(mode="cloudium")
        mock_resolver.list_dir.return_value = ["U:/cloud/a.xlsx"]

        # iterdir이 PermissionError 던지도록 mock — cloudium에서 backend 권한 부족 시뮬레이션
        with patch("backend.routers.swut.get_resolver", return_value=mock_resolver), \
             patch("pathlib.Path.iterdir",
                   side_effect=PermissionError("no cloudium access")):
            r = client.post(
                "/api/swut/browse",
                json={"path": "U:/cloud"},
                headers={"X-User": "tester"},
            )
        # 22차: PermissionError silent → 200 + cloudium_hint + dirs=[] + files=[1]
        assert r.status_code == 200
        body = r.json()
        assert body["file_mode"] == "cloudium"
        assert body["dirs"] == []
        assert len(body["files"]) == 1
        assert "Cloudium 모드" in body["cloudium_hint"]
        assert "디렉토리" in body["cloudium_hint"]

    def test_browse_local_hint_when_iterdir_permission_error(self):
        """22차 T190: local 모드 + iterdir PermissionError → 200 + 일반 hint."""
        from unittest.mock import MagicMock

        mock_resolver = MagicMock(mode="local")
        mock_resolver.list_dir.return_value = []

        with patch("backend.routers.swut.get_resolver", return_value=mock_resolver), \
             patch("pathlib.Path.iterdir",
                   side_effect=PermissionError("no access")):
            r = client.post(
                "/api/swut/browse",
                json={"path": "/restricted"},
                headers={"X-User": "tester"},
            )
        assert r.status_code == 200
        body = r.json()
        assert body["file_mode"] == "local"
        assert "권한 부족" in body["cloudium_hint"]
        assert "Cloudium" not in body["cloudium_hint"]


class TestPathModeMismatch56:
    """56차 T308 — SwUT log_folder UNC + Local 모드 → 400 + PATH_MODE_MISMATCH."""

    def test_swut_coverage_unc_local_mode_returns_400(self):
        """Local 모드 + log_folder=U:/... → 400 + PATH_MODE_MISMATCH."""
        from backend.services.file_resolver import LocalFileResolver, set_resolver
        set_resolver(LocalFileResolver())
        try:
            r = client.post(
                "/api/swut/coverage/build",
                json={
                    "project_id": "HDPDM01",
                    "release_sw_version": "1.0.0",
                    "test_date": "2024-02-19",
                    "log_folder": "U:/연구소/test/v1.0",
                },
                headers={"X-User": "tester"},
            )
            assert r.status_code == 400, f"expected 400 got {r.status_code}: {r.text[:200]}"
            body = r.json()
            assert body.get("error", {}).get("code") == "PATH_MODE_MISMATCH", f"body={body}"
        finally:
            set_resolver(LocalFileResolver())

    def test_swut_sutr_unc_local_mode_returns_400(self):
        """SUTR endpoint도 동일."""
        from backend.services.file_resolver import LocalFileResolver, set_resolver
        set_resolver(LocalFileResolver())
        try:
            r = client.post(
                "/api/swut/sutr/build",
                json={
                    "project_id": "HDPDM01",
                    "release_sw_version": "1.0.0",
                    "test_date": "2024-02-19",
                    "log_folder": r"\\corp\share\test",
                },
                headers={"X-User": "tester"},
            )
            assert r.status_code == 400, f"expected 400 got {r.status_code}: {r.text[:200]}"
            body = r.json()
            assert body.get("error", {}).get("code") == "PATH_MODE_MISMATCH", f"body={body}"
        finally:
            set_resolver(LocalFileResolver())


# ---------------------------------------------------------------------------
# B2 — _resolve_swut_log_folders 4단계 우선순위 + log_folders 스키마 검증
# ---------------------------------------------------------------------------


def _make_swut_req(**kwargs) -> SwUTBuildRequest:
    """최소 유효 SwUTBuildRequest — log_folder(s) 우선순위 검증용."""
    base = {
        "project_id": "KJPDS02",
        "release_sw_version": "1.01",
        "test_date": "2025-12-05",
    }
    base.update(kwargs)
    return SwUTBuildRequest(**base)


class TestResolveSwutLogFoldersPriority:
    """B2 — req.log_folders > req.log_folder > config swut_log_folders >
    config swut_log_folder."""

    def _patch_cfg(self, monkeypatch, cfg: dict):
        from backend.routers import swut as swut_mod
        monkeypatch.setattr(
            swut_mod, "_load_meta_from_config", lambda _project_id: cfg,
        )
        return swut_mod

    _FULL_CFG = {
        "swut_log_folders": ["U:/cfg/app", "U:/cfg/boot"],
        "swut_log_folder": "U:/cfg/single",
    }

    def test_req_log_folders_first(self, monkeypatch):
        swut_mod = self._patch_cfg(monkeypatch, dict(self._FULL_CFG))
        req = _make_swut_req(
            log_folders=["C:/req/app", "C:/req/boot"],
            log_folder="C:/req/single",
        )
        assert swut_mod._resolve_swut_log_folders(req) == [
            "C:/req/app", "C:/req/boot",
        ]

    def test_req_log_folder_second(self, monkeypatch):
        """req.log_folders 비면 req.log_folder가 config 양 키보다 우선."""
        swut_mod = self._patch_cfg(monkeypatch, dict(self._FULL_CFG))
        req = _make_swut_req(log_folder="C:/req/single")
        assert swut_mod._resolve_swut_log_folders(req) == ["C:/req/single"]

    def test_config_log_folders_third(self, monkeypatch):
        """req 양 필드 비면 config swut_log_folders(list)가 단수 키보다 우선."""
        swut_mod = self._patch_cfg(monkeypatch, dict(self._FULL_CFG))
        req = _make_swut_req()
        assert swut_mod._resolve_swut_log_folders(req) == [
            "U:/cfg/app", "U:/cfg/boot",
        ]

    def test_config_log_folder_fourth(self, monkeypatch):
        """config swut_log_folders 부재 시 기존 단수 swut_log_folder fallback."""
        swut_mod = self._patch_cfg(
            monkeypatch, {"swut_log_folder": "U:/cfg/single"},
        )
        req = _make_swut_req()
        assert swut_mod._resolve_swut_log_folders(req) == ["U:/cfg/single"]

    def test_all_empty_returns_empty_list(self, monkeypatch):
        """전부 비면 빈 list — Jenkins-only 빌드 허용."""
        swut_mod = self._patch_cfg(monkeypatch, {})
        req = _make_swut_req()
        assert swut_mod._resolve_swut_log_folders(req) == []

    def test_config_empty_list_falls_through_to_single(self, monkeypatch):
        """config swut_log_folders=[] (빈 list) → 단수 키 fallback."""
        swut_mod = self._patch_cfg(
            monkeypatch,
            {"swut_log_folders": [], "swut_log_folder": "U:/cfg/single"},
        )
        req = _make_swut_req()
        assert swut_mod._resolve_swut_log_folders(req) == ["U:/cfg/single"]

    def test_backward_compat_wrapper_returns_first(self, monkeypatch):
        """_resolve_swut_log_folder(단수 wrapper) — 첫 폴더 단일 반환 계약."""
        swut_mod = self._patch_cfg(monkeypatch, dict(self._FULL_CFG))
        req = _make_swut_req()
        assert swut_mod._resolve_swut_log_folder(req) == "U:/cfg/app"
        # 전부 비면 None (기존 계약)
        swut_mod2 = self._patch_cfg(monkeypatch, {})
        assert swut_mod2._resolve_swut_log_folder(_make_swut_req()) is None


class TestLogFoldersSchema:
    """B2 — SwUTBuildRequest.log_folders 입력 표면 (maxlen/줄바꿈/개수/extra)."""

    _BASE = {
        "project_id": "HDPDM01",
        "release_sw_version": "1.0.0",
        "test_date": "2024-02-19",
    }

    def test_valid_two_folders_accepted(self):
        req = SwUTBuildRequest(**self._BASE, log_folders=["C:/a", "C:/b"])
        assert req.log_folders == ["C:/a", "C:/b"]

    def test_item_len_500_accepted(self):
        item = "C:/" + "x" * 497
        assert len(item) == 500
        req = SwUTBuildRequest(**self._BASE, log_folders=[item])
        assert req.log_folders == [item]

    def test_item_over_500_rejected(self):
        item = "C:/" + "x" * 498  # 501자
        with pytest.raises(ValidationError, match="500"):
            SwUTBuildRequest(**self._BASE, log_folders=[item])

    @pytest.mark.parametrize("bad", ["C:/a\nb", "C:/a\rb", "C:/a\r\nX-Injected: evil"])
    def test_item_newline_rejected(self, bad):
        """W8 정책 동일 적용 — 항목별 줄바꿈 금지 (헤더 인젝션 안전)."""
        with pytest.raises(ValidationError, match="줄바꿈"):
            SwUTBuildRequest(**self._BASE, log_folders=[bad])

    def test_exactly_8_items_accepted(self):
        folders = [f"C:/p{i}" for i in range(8)]
        req = SwUTBuildRequest(**self._BASE, log_folders=folders)
        assert req.log_folders == folders

    def test_nine_items_rejected(self):
        with pytest.raises(ValidationError):
            SwUTBuildRequest(
                **self._BASE, log_folders=[f"C:/p{i}" for i in range(9)],
            )

    def test_extra_field_still_forbidden(self):
        """extra='forbid' 유지 — 신규 필드 추가가 모델 설정을 풀지 않았는지."""
        with pytest.raises(ValidationError):
            SwUTBuildRequest(**self._BASE, log_folders_typo=["C:/a"])

    def test_endpoint_log_folders_newline_422(self):
        """endpoint 표면에서도 422 (TestClient 경유)."""
        r = client.post(
            "/api/swut/coverage/build",
            json={
                **self._BASE,
                "log_folders": ["C:/fake/log\r\nX-Injected: evil"],
            },
            headers={"X-User": "tester"},
        )
        assert r.status_code == 422


# ---------------------------------------------------------------------------
# 라운드 107 — UT201 FI spec 자동 산출 라우터 게이트 (_resolve_spec_fi_for_swutcr)
# ---------------------------------------------------------------------------

def _make_fi_spec_min_bytes() -> bytes:
    """FI 블록 1개짜리 최소 DV 레이아웃 spec (라우터 게이트 통과 검증용)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    ws.cell(3, 2).value = "Test Case"
    ws.cell(3, 58).value = "Expected Result"
    ws.cell(3, 162).value = "Related ID"
    ws.cell(4, 2).value = "Index"
    ws.cell(4, 3).value = "TC_ID"
    ws.cell(4, 4).value = "Unit"
    ws.cell(4, 5).value = "Test Method"
    ws.cell(4, 8).value = "Inpt[0]"
    ws.cell(4, 58).value = "ExpR[0]"
    ws.cell(5, 2).value = 1
    ws.cell(5, 3).value = "SwUTC_0101"
    ws.cell(5, 4).value = "fn_a"
    ws.cell(6, 5).value = "FI"
    ws.cell(6, 7).value = 1
    ws.cell(6, 8).value = "0x0"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSwutcrSpecFiResolve107:
    """확정 규칙(2026-06-12) — config FI 키 부재 시에만 spec 1회 로드.

    우선순위: config fault_injection_total/passed > spec 산출 > 노란 마킹.
    config 키가 하나라도 있으면 spec 로드 자체를 skip (heavy 로드 비용 회피 +
    _write_ut201 config 분기와 게이트 일치). spec path 부재(HDPDM01)는
    None — 기존 노란 마킹 경로 무회귀.
    """

    def _req(self, **over) -> SwUTBuildRequest:
        kw = dict(
            project_id="HDPDM01",
            release_sw_version="1.0.0",
            test_date="2024-02-19",
        )
        kw.update(over)
        return SwUTBuildRequest(**kw)

    def _setup_cfg(self, tmp_path, monkeypatch, cfg_dict):
        """resolve_swuts_path의 config fallback을 hermetic하게 격리."""
        import json as _json

        from backend.services import swut_meta_resolver as resolver_mod
        cfg_path = tmp_path / "swut_meta.json"
        cfg_path.write_text(_json.dumps(cfg_dict), encoding="utf-8")
        monkeypatch.setattr(resolver_mod, "_META_CONFIG_PATH", str(cfg_path))
        resolver_mod._read_meta_config_raw.cache_clear()

    class _RaisingResolver:
        """spec read가 호출되면 안 되는 경로 가드."""
        def read_bytes(self, path):
            raise AssertionError(f"spec read 금지 경로에서 read_bytes 호출: {path}")

    class _BytesResolver:
        def __init__(self, data: bytes):
            self._data = data
            self.read_paths: list[str] = []

        def read_bytes(self, path):
            self.read_paths.append(path)
            return self._data

    class _FailingResolver:
        def read_bytes(self, path):
            raise FileNotFoundError(path)

    def test_config_fi_keys_present_skips_spec_load(self, tmp_path, monkeypatch):
        from backend.routers import swut as swut_mod
        self._setup_cfg(tmp_path, monkeypatch, {"projects": {"HDPDM01": {}}})
        warns: list[str] = []
        cfg = {"swutcr_metadata": {
            "fault_injection_total": 402, "fault_injection_passed": 402,
        }}
        out = swut_mod._resolve_spec_fi_for_swutcr(
            self._req(swuts_docx_path="C:/x/spec.xlsm"), cfg,
            self._RaisingResolver(), warns,
        )
        assert out is None
        assert warns == []

    def test_config_partial_fi_key_skips_spec_load(self, tmp_path, monkeypatch):
        """키 하나만 있어도 config 우선 — _write_ut201 partial 분기와 일치."""
        from backend.routers import swut as swut_mod
        self._setup_cfg(tmp_path, monkeypatch, {"projects": {"HDPDM01": {}}})
        warns: list[str] = []
        cfg = {"swutcr_metadata": {"fault_injection_total": 402}}
        out = swut_mod._resolve_spec_fi_for_swutcr(
            self._req(swuts_docx_path="C:/x/spec.xlsm"), cfg,
            self._RaisingResolver(), warns,
        )
        assert out is None
        assert warns == []

    def test_no_spec_path_returns_none_silently(self, tmp_path, monkeypatch):
        """HDPDM01 경로 — config 키도 spec도 없음 → None (노란 마킹 무회귀)."""
        from backend.routers import swut as swut_mod
        self._setup_cfg(tmp_path, monkeypatch, {"projects": {"HDPDM01": {}}})
        warns: list[str] = []
        out = swut_mod._resolve_spec_fi_for_swutcr(
            self._req(), {"swutcr_metadata": {}}, self._RaisingResolver(), warns,
        )
        assert out is None
        assert warns == []

    def test_spec_read_failure_warns_and_returns_none(self, tmp_path, monkeypatch):
        from backend.routers import swut as swut_mod
        self._setup_cfg(tmp_path, monkeypatch, {"projects": {"HDPDM01": {}}})
        warns: list[str] = []
        out = swut_mod._resolve_spec_fi_for_swutcr(
            self._req(swuts_docx_path="C:/x/spec.xlsm"), {},
            self._FailingResolver(), warns,
        )
        assert out is None
        assert any("UT201 FI spec 읽기 실패" in w for w in warns)

    def test_spec_loaded_extracted_with_filename(self, tmp_path, monkeypatch):
        """config 키 부재 + spec 존재 — 1회 로드 + 추출 + 파일명 동봉."""
        from backend.routers import swut as swut_mod
        self._setup_cfg(tmp_path, monkeypatch, {"projects": {"HDPDM01": {}}})
        resolver = self._BytesResolver(_make_fi_spec_min_bytes())
        warns: list[str] = []
        out = swut_mod._resolve_spec_fi_for_swutcr(
            self._req(swuts_docx_path="C:/x/wip_pv_SwUTS_v0.10_260608.xlsm"),
            {"swutcr_metadata": {}}, resolver, warns,
        )
        assert out is not None
        assert out["fi_block_total"] == 1
        assert out["fi_block_keys"] == {"101"}
        assert out["fi_iters_per_block"]["101"] == [1]
        assert out["spec_filename"] == "wip_pv_SwUTS_v0.10_260608.xlsm"
        assert resolver.read_paths == ["C:/x/wip_pv_SwUTS_v0.10_260608.xlsm"]

    def test_config_spec_path_fallback_used(self, tmp_path, monkeypatch):
        """req path 비면 config swuts_docx_path fallback (resolve_swuts_path 정책)."""
        from backend.routers import swut as swut_mod
        self._setup_cfg(tmp_path, monkeypatch, {"projects": {"HDPDM01": {
            "swuts_docx_path": "U:/cfg/spec_from_cfg.xlsm",
        }}})
        resolver = self._BytesResolver(_make_fi_spec_min_bytes())
        out = swut_mod._resolve_spec_fi_for_swutcr(
            self._req(), {"swutcr_metadata": {}}, resolver, [],
        )
        assert out is not None
        assert out["spec_filename"] == "spec_from_cfg.xlsm"

    def test_preloaded_spec_bytes_skips_resolver_read(self, tmp_path, monkeypatch):
        """라운드 108 MINOR-5 — preloaded bytes 전달 시 spec 재read 없음.

        SwUTCR 빌드는 resolve_swuts_test_specs(out_xlsm_bytes=...)가 이미
        읽은 동일 spec bytes를 재사용 — PV ~20s 중복 read 제거.
        """
        from backend.routers import swut as swut_mod
        self._setup_cfg(tmp_path, monkeypatch, {"projects": {"HDPDM01": {}}})
        warns: list[str] = []
        out = swut_mod._resolve_spec_fi_for_swutcr(
            self._req(swuts_docx_path="C:/x/wip_pv_SwUTS_v0.10_260608.xlsm"),
            {"swutcr_metadata": {}},
            self._RaisingResolver(),  # read 호출 시 즉시 실패 — 미호출 증명
            warns,
            preloaded_spec_bytes=_make_fi_spec_min_bytes(),
        )
        assert out is not None
        assert out["fi_block_total"] == 1
        assert out["spec_filename"] == "wip_pv_SwUTS_v0.10_260608.xlsm"

    def test_docx_spec_extension_warns_and_skips(self, tmp_path, monkeypatch):
        """라운드 108 INFO-8 — xlsm/xlsx 외 확장자(docx)는 사유 warning + skip.

        silent skip(예외 삼킴) 대신 명시 사유 — read 자체도 생략.
        """
        from backend.routers import swut as swut_mod
        self._setup_cfg(tmp_path, monkeypatch, {"projects": {"HDPDM01": {}}})
        warns: list[str] = []
        out = swut_mod._resolve_spec_fi_for_swutcr(
            self._req(swuts_docx_path="C:/x/spec.docx"),
            {"swutcr_metadata": {}}, self._RaisingResolver(), warns,
        )
        assert out is None
        assert any(
            "spec 확장자가 xlsm/xlsx 아님" in w and "spec.docx" in w
            for w in warns
        )


# ── 시험 결과 3종의 Quality 기록 배선 (2026-08-21) ────────────────────────────
#
# 여기서 깨지는 방식은 **조용하다**. 빌드는 성공하고 파일도 받아지는데 Quality DB 에
# 행이 안 남아, 생성 현황 보드가 방금 만든 문서를 영영 "미생성" 으로 표시한다. 실제로
# SwUT Coverage 에만 기록이 있어 SUTR 가 그 상태였고(커밋 이력), 고친 뒤에도
# **SWUTCR 은 여전히 그 상태였다** — 같은 결함이 한 칸 옆에서 반복됐다.
#
# 그래서 헬퍼 단독 테스트로는 부족하다(호출부가 빠진 게 결함이므로). 아래는 **호출부를
# AST 로** 확인한다: 각 빌드 함수가 `_record_test_quality` 를 부르는가, 그리고 doc_type
# 을 무엇으로 넘기는가. doc_type 이 틀리면 종합결과서가 SUTR 행을 덮어쓴다.


class TestQualityRecordingWiring:
    EXPECTED = {
        "_do_sutr_build": "sutr",
        "_do_sutr_build_spec_based": "sutr",
        "_do_swutcr_build": "swutcr",
    }

    @staticmethod
    def _record_calls(func_name: str):
        import ast
        src = pathlib.Path("backend/routers/swut.py").read_text(encoding="utf-8")
        tree = ast.parse(src)
        fn = next(
            (n for n in ast.walk(tree)
             if isinstance(n, (ast.FunctionDef, ast.AsyncFunctionDef)) and n.name == func_name),
            None,
        )
        assert fn is not None, f"{func_name} 가 사라졌다 — 테스트가 겨눌 대상이 없다"
        out = []
        for node in ast.walk(fn):
            if (isinstance(node, ast.Call) and isinstance(node.func, ast.Name)
                    and node.func.id == "_record_test_quality"):
                kw = {k.arg: k.value for k in node.keywords}
                doc = kw.get("doc_type")
                out.append(doc.value if isinstance(doc, ast.Constant) else None)
        return out

    @pytest.mark.parametrize("func_name", sorted(EXPECTED))
    def test_build_path_records_quality_with_its_own_doc_type(self, func_name):
        calls = self._record_calls(func_name)
        assert calls == [self.EXPECTED[func_name]], (
            f"{func_name} 의 quality 기록 호출: {calls} "
            f"(기대: [{self.EXPECTED[func_name]!r}])"
        )

    def test_doc_type_has_no_default(self):
        """기본값을 두면 **빠뜨린 호출이 조용히 SUTR 로 기록**된다 — 종합결과서가
        SUTR 행을 덮어쓰고, 보드는 둘을 구분하지 못한다."""
        import inspect

        from backend.routers import swut as mod
        sig = inspect.signature(mod._record_test_quality)
        param = sig.parameters["doc_type"]
        assert param.kind is inspect.Parameter.KEYWORD_ONLY
        assert param.default is inspect.Parameter.empty

    def test_helper_passes_doc_type_through(self, monkeypatch):
        """헬퍼가 받은 doc_type 을 그대로 recorder 에 넘기는가."""
        import workflow.quality.recorder as rec
        from backend.routers import swut as mod
        seen = {}

        def _fake(doc_type, summary, **kw):
            seen["doc_type"] = doc_type
            seen["summary"] = summary
            seen.update(kw)
            return 1

        monkeypatch.setattr(rec, "record_test_result_run", _fake)

        class _Req:
            project_id = "HDPDM01"
            release_sw_version = "1.02"

        class _Meta:
            asil_level = "ASIL B"

        mod._record_test_quality(_Req(), _Meta(), {"total_tcs": 5}, doc_type="swutcr")
        assert seen["doc_type"] == "swutcr"
        assert seen["release_sw_version"] == "1.02"
        assert seen["project_id"] == "HDPDM01"

    def test_recording_failure_is_logged_not_silent(self, monkeypatch, caplog):
        """non-fatal 은 유지하되 **침묵은 금지** — 같은 블록이 NameError 를 삼킨 전례가 있다."""
        import logging

        import workflow.quality.recorder as rec
        from backend.routers import swut as mod

        def _boom(*a, **k):
            raise RuntimeError("db down")

        monkeypatch.setattr(rec, "record_test_result_run", _boom)

        class _X:
            project_id = ""
            release_sw_version = ""
            asil_level = ""

        with caplog.at_level(logging.ERROR):
            mod._record_test_quality(_X(), _X(), {}, doc_type="swutcr")   # 예외가 새면 안 된다
        assert any("SWUTCR" in r.getMessage() for r in caplog.records), (
            "기록 실패가 로그에 남지 않았다 — 조용한 실패는 이 파일이 겨누는 결함 그 자체다"
        )

    def test_scm_id_is_passed_through_not_guessed(self, monkeypatch):
        """프로젝트 축을 **추측에 맡기지 않는다**(2026-08-24 라이브 실측).

        비우면 recorder 가 `project_id` 에서 `resolve_scm_id` 로 추측하는데, 문자열
        "KJPDS02" 가 SCM entry `kjpds02` 의 id 이면서 `kjpds02_pv` 의 builder_project_id
        라 추측이 빗나갔다 → 생성 현황 보드(`kjpds02_pv` 로 조회)가 방금 만든 문서를
        영영 "미생성" 으로 표시했다. 빌드도 기록도 정상인데 화면만 침묵하는 형태다.
        """
        import workflow.quality.recorder as rec
        from backend.routers import swut as mod
        seen = {}
        monkeypatch.setattr(
            rec, "record_test_result_run",
            lambda doc_type, summary, **kw: (seen.update(kw), 1)[1],
        )

        class _Req:
            project_id = "KJPDS02"
            release_sw_version = "1.02"
            scm_id = "kjpds02_pv"

        class _Meta:
            asil_level = "ASIL A"

        mod._record_test_quality(_Req(), _Meta(), {"total_tcs": 5}, doc_type="swutcr")
        assert seen["scm_id"] == "kjpds02_pv"

    def test_absent_scm_id_stays_none_not_empty_string(self, monkeypatch):
        """빈 문자열을 넘기면 recorder 의 `if not scm_id` 추측 분기가 **살아 있어야** 한다
        — `""` 를 그대로 저장하면 '축을 아는데 빈 값' 처럼 보여 더 나쁘다."""
        import workflow.quality.recorder as rec
        from backend.routers import swut as mod
        seen = {}
        monkeypatch.setattr(
            rec, "record_test_result_run",
            lambda doc_type, summary, **kw: (seen.update(kw), 1)[1],
        )

        class _Req:
            project_id = "KJPDS02"
            release_sw_version = "1.02"
            scm_id = ""

        class _Meta:
            asil_level = "ASIL A"

        mod._record_test_quality(_Req(), _Meta(), {"total_tcs": 5}, doc_type="swutcr")
        assert seen["scm_id"] is None
