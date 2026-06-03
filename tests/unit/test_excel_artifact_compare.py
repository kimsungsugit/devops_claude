"""Tests for strict Excel artifact comparison."""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Border, Font, Side

from backend.services.excel_artifact_compare import compare_excel_artifacts


def _save(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_compare_detects_merge_font_border_and_formula_diffs():
    ref = openpyxl.Workbook()
    ref_ws = ref.active
    ref_ws.title = "Sheet1"
    ref_ws.merge_cells("A1:B1")
    ref_ws["A1"] = "Merged"
    ref_ws["C1"] = "=SUM(A2:A3)"
    ref_ws["D1"] = "Styled"
    ref_ws["D1"].font = Font(name="Arial", size=10, bold=True)
    ref_ws["D1"].border = Border(bottom=Side(style="thin"))

    gen = openpyxl.Workbook()
    gen_ws = gen.active
    gen_ws.title = "Sheet1"
    gen_ws["A1"] = "Merged"
    gen_ws["C1"] = "=SUM(A2:A4)"
    gen_ws["D1"] = "Styled"
    gen_ws["D1"].font = Font(name="Arial", size=10, bold=False)
    gen_ws["D1"].border = Border(bottom=Side(style="thick"))

    report = compare_excel_artifacts(_save(ref), _save(gen), sheets=["Sheet1"])

    assert report.ok is False
    assert report.summary["merge"] == 1
    assert report.summary["formula"] == 1
    assert report.summary["font"] >= 1
    assert report.summary["border"] >= 1


def test_compare_treats_none_and_empty_string_as_blank_equivalent():
    ref = openpyxl.Workbook()
    ref.active.title = "Sheet1"
    ref.active["A1"] = None

    gen = openpyxl.Workbook()
    gen.active.title = "Sheet1"
    gen.active["A1"] = ""

    report = compare_excel_artifacts(
        _save(ref),
        _save(gen),
        sheets=["Sheet1"],
        compare_styles=False,
    )

    assert report.ok is True
    assert report.summary == {}


def test_compare_limits_stored_diffs_but_counts_all():
    ref = openpyxl.Workbook()
    ref.active.title = "Sheet1"
    gen = openpyxl.Workbook()
    gen.active.title = "Sheet1"
    for row in range(1, 6):
        ref.active.cell(row, 1, "A")
        gen.active.cell(row, 1, "B")

    report = compare_excel_artifacts(
        _save(ref),
        _save(gen),
        sheets=["Sheet1"],
        max_diffs_per_category=2,
        compare_styles=False,
    )

    assert report.summary["value"] == 5
    assert len([diff for diff in report.diffs if diff.category == "value"]) == 2
    assert report.warnings


def test_compare_requested_missing_sheet_reports_without_key_error():
    ref = openpyxl.Workbook()
    ref.active.title = "OnlyRef"
    gen = openpyxl.Workbook()
    gen.active.title = "OnlyGen"

    report = compare_excel_artifacts(
        _save(ref),
        _save(gen),
        sheets=["OnlyRef", "OnlyGen"],
    )

    assert report.ok is False
    assert report.summary["sheet"] == 2
