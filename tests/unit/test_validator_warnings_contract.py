"""세 시험규격 검증기의 **`warnings` 계약**이 같은가.

## 왜 이 파일이 필요했나

`validate_sts_xlsm` / `validate_suts_xlsm` / `validate_sits_xlsm` 은 같은 모양의 dict 를
낸다고 공시하지만, 2026-09-02 AST 실측에서 셋이 **전부 달랐다**:

| 검증기 | 선언 | `warnings.append` | 반환 경로 중 포함 | 리포트 절 |
|---|---|---|---|---|
| STS  | ✓ | **5** | **5/5** | ✓ |
| SUTS | ✓(미사용) | **0** | **2/4** | ✗ |
| SITS | ✗ | 0 | **0/4** | ✗ |

그런데 API 는 세 종류 모두에 `warnings` 를 실어 내보낸다
(`backend/helpers/common.py::_build_excel_artifact_payload` 가 `validation` 을
**본문째** 실어 보내고, `summary.validation` 에는 `warning_count` 만 남는다). 즉 SUTS·SITS 는
**항상 `[]`** 이고 그게 "점검했고 깨끗함" 으로 읽혔다 — 점검이 하나도 없다는 사실이
빈 배열에 숨었다. SUTS 의 docstring 은 `warnings` 를 준다고 **명시**하고 있었다.

`issues` 는 `valid` 를 뒤집지만 `warnings` 는 안 뒤집는다. 그래서 경고는 **리포트에
자리가 있어야만** 사람 눈에 닿는다 — 검증기만 고치고 리포트를 놔두면 여전히 아무 데도
안 남는다. 이 파일은 두 면을 같이 잠근다.

⚠ 임계를 새로 발명하지 않는다. 추가한 경고는 전부 **산술적 사실**이다
   ("I/O 없는 TC 가 n개", "시퀀스 수 < TC 수", "sub-case 수 < TC 수").
"""
from __future__ import annotations

import ast
from pathlib import Path
from typing import List, Set

import pytest

REPO = Path(__file__).resolve().parents[2]

VALIDATORS = {
    "sts": ("generators/sts.py", "validate_sts_xlsm"),
    "suts": ("generators/suts.py", "validate_suts_xlsm"),
    "sits": ("generators/sits.py", "validate_sits_xlsm"),
}
REPORTERS = {
    "sts": ("generators/sts.py", "generate_sts_validation_report"),
    "suts": ("generators/suts.py", "generate_suts_validation_report"),
    "sits": ("generators/sits.py", "generate_sits_validation_report"),
}


def _func(rel: str, name: str) -> tuple[ast.FunctionDef, str]:
    src = (REPO / rel).read_text(encoding="utf-8")
    for node in ast.parse(src).body:
        if isinstance(node, ast.FunctionDef) and node.name == name:
            return node, src
    raise AssertionError(f"{rel}: {name} 를 못 찾았다")


def _dict_return_keys(node: ast.FunctionDef) -> List[Set[str]]:
    """이 함수의 **dict 리터럴 return** 들이 각각 담는 키 집합."""
    out: List[Set[str]] = []
    for n in ast.walk(node):
        if isinstance(n, ast.Return) and isinstance(n.value, ast.Dict):
            out.append({k.value for k in n.value.keys if isinstance(k, ast.Constant)})
    return out


# ==============================================================
# 1. 반환 shape — 경로마다 다르면 소비처가 경로에 따라 죽는다
# ==============================================================

class TestEveryReturnPathCarriesWarnings:

    @pytest.mark.parametrize("kind", sorted(VALIDATORS))
    def test_all_dict_returns_include_the_documented_keys(self, kind):
        rel, name = VALIDATORS[kind]
        node, _ = _func(rel, name)
        rets = _dict_return_keys(node)
        assert rets, f"{name}: dict 리턴을 못 찾았다(구조가 바뀌었나)"
        for i, keys in enumerate(rets):
            missing = {"valid", "issues", "warnings", "stats"} - keys
            assert not missing, f"{name} 반환 #{i}: {sorted(missing)} 누락 — {sorted(keys)}"

    @pytest.mark.parametrize("kind", sorted(VALIDATORS))
    def test_the_three_validators_agree_on_the_shape(self, kind):
        """한 곳만 키가 많거나 적으면 소비처가 종류별 분기를 갖게 된다."""
        shapes = {}
        for k, (rel, name) in VALIDATORS.items():
            node, _ = _func(rel, name)
            shapes[k] = [frozenset(s) for s in _dict_return_keys(node)]
        base = {"valid", "issues", "warnings", "stats"}
        for k, rets in shapes.items():
            for keys in rets:
                assert base <= set(keys), f"{k}: {sorted(base - set(keys))} 누락"


# ==============================================================
# 2. 선언만 하고 안 채우면 "점검 없음" 이 "경고 없음" 으로 위장한다
# ==============================================================

class TestWarningsAreActuallyProduced:

    @pytest.mark.parametrize("kind", sorted(VALIDATORS))
    def test_each_validator_can_emit_at_least_one_warning(self, kind):
        """`warnings` 를 내보낸다고 공시하면 **만들 수 있어야** 한다.

        SUTS 는 선언만 하고 append 가 0 이었다 — 빈 배열이 영구 보장이었고,
        docstring 은 준다고 적혀 있었다.
        """
        rel, name = VALIDATORS[kind]
        node, src = _func(rel, name)
        seg = ast.get_source_segment(src, node) or ""
        assert seg.count("warnings.append") >= 1, f"{name}: 경고를 만드는 자리가 없다"


# ==============================================================
# 3. 리포트에 자리가 없으면 경고는 아무 데도 안 남는다
# ==============================================================

class TestTheReportHasSomewhereToPutThem:

    @pytest.mark.parametrize("kind", sorted(REPORTERS))
    def test_the_report_actually_writes_the_warning_text(self, kind, tmp_path):
        """⚠ **행동으로** 잰다. 구조 검사(`for w in warnings` 가 소스에 있나)로는
        `if warnings:` → `if False:` 뮤턴트가 그대로 살아남는다(실측 M235·M236) —
        루프는 남아 있고 도달만 안 하기 때문이다.

        `issues` 는 `valid` 를 뒤집지만 `warnings` 는 안 뒤집는다. 그래서 경고는
        리포트에 실제로 **쓰여야만** 사람 눈에 닿는다.
        """
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        import generators.sits as sits_mod
        import generators.sts as sts_mod
        import generators.suts as suts_mod

        writer = {"sts": sts_mod.generate_sts_validation_report,
                  "suts": suts_mod.generate_suts_validation_report,
                  "sits": sits_mod.generate_sits_validation_report}[kind]

        xlsm = tmp_path / f"{kind}.xlsm"
        Workbook().save(str(xlsm))
        marker = "MARKER-경고가-리포트에-닿았는가"
        # 리포트 라이터는 `validation` 을 주입받는다 — 파싱 가능한 워크북 없이도
        # **렌더 경로 그대로** 태울 수 있다.
        out = writer(str(xlsm), {}, validation={
            "valid": True, "issues": [], "warnings": [marker], "stats": {}})
        text = Path(out).read_text(encoding="utf-8")
        assert marker in text, f"{kind}: 경고가 리포트에 안 적힌다\n{text[:600]}"

    @pytest.mark.parametrize("kind", sorted(REPORTERS))
    def test_the_report_uses_the_verdict_it_was_given(self, kind, tmp_path):
        """호출부가 만든 판정을 리포트가 **버리지 않는가**.

        STS 리포트는 이 인자가 아예 없어 `validate_sts_xlsm` 을 처음부터 다시 돌렸다.
        그래서 사라지던 것(실측):
          1. `apply_write_back_check` — 생성 수 ↔ 파일 기록 수 불일치면 `valid=False`
             인데 재검증은 그걸 모르므로 리포트가 **PASS 로 적었다**
          2. 검증 크래시 시 호출부가 세운 fail-closed(`valid:False` + "미검증")

        빈 워크북이라 재검증하면 반드시 FAIL 이 나온다 — 그러므로 리포트에 **PASS** 가
        적혔다면 주입한 판정을 쓴 것이다(대조가 성립한다).
        """
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        import generators.sits as sits_mod
        import generators.sts as sts_mod
        import generators.suts as suts_mod

        writer = {"sts": sts_mod.generate_sts_validation_report,
                  "suts": suts_mod.generate_suts_validation_report,
                  "sits": sits_mod.generate_sits_validation_report}[kind]
        validator = {"sts": sts_mod.validate_sts_xlsm,
                     "suts": suts_mod.validate_suts_xlsm,
                     "sits": sits_mod.validate_sits_xlsm}[kind]

        xlsm = tmp_path / f"{kind}.xlsm"
        Workbook().save(str(xlsm))
        assert validator(str(xlsm))["valid"] is False, "빈 워크북이 통과하면 이 대조가 무의미"

        out = writer(str(xlsm), {}, validation={
            "valid": True, "issues": [], "warnings": [], "stats": {}})
        text = Path(out).read_text(encoding="utf-8")
        # ⚠ 본문에는 `stats` 에서 파생된 **개별 점검** 행에도 PASS/FAIL 이 있다.
        #   재는 것은 **최상단 종합 판정 줄** 하나다 — 그게 주입값을 따르는지가 계약이다.
        verdict = next((ln for ln in text.splitlines() if ln.startswith("**결과**")), "")
        assert verdict, f"{kind}: 종합 판정 줄을 못 찾았다\n{text[:400]}"
        assert "PASS" in verdict, f"{kind}: 주입한 판정을 버리고 재검증했다 — {verdict!r}"


    @pytest.mark.parametrize("kind", sorted(REPORTERS))
    def test_the_generator_passes_its_verdict_to_the_report(self, kind):
        """고치는 자리가 **둘**이다: 라이터가 받을 수 있는가 + 호출부가 넘기는가.

        라이터만 고치면 뮤테이션이 그대로 산다(실측 M239 생존) — 호출부에서
        판정을 떼면 리포트는 다시 재검증하고 write-back 결과를 잃는다.

        ⚠ 재는 것은 **"판정이 전달되는가"** 이지 키워드로 쓰였는가가 아니다.
          처음엔 `validation=` 키워드만 봤다가 SITS 를 거짓 양성으로 잡았다 —
          거긴 세 번째 **위치 인자**로 정상 전달하고 있었다. 가드가 사실이 아니라
          철자를 재면 멀쩡한 코드를 결함으로 신고한다.
        """
        rel, name = REPORTERS[kind]
        src = (REPO / rel).read_text(encoding="utf-8")
        tree = ast.parse(src)
        calls = [
            n for n in ast.walk(tree)
            if isinstance(n, ast.Call) and isinstance(n.func, ast.Name) and n.func.id == name
        ]
        assert calls, f"{name} 호출부를 못 찾았다 — 생성 흐름이 바뀌었나"
        for c in calls:
            by_keyword = any(k.arg == "validation" for k in c.keywords)
            by_position = len(c.args) >= 3          # (xlsm_path, quality_report, validation)
            assert by_keyword or by_position, (
                f"{rel}:{c.lineno} {name}(...) 이 판정을 안 넘긴다 — 리포트가 재검증하며 "
                f"write-back 대조와 '미검증' 상태를 잃는다")


# ==============================================================
# 4. 행동 — 실제로 만들어 보고 경고가 나오는가
# ==============================================================

class TestBehaviour:

    def test_suts_warns_when_a_minority_of_tcs_lack_io(self, tmp_path):
        """50% **이하**는 예전에 통째로 침묵했다(50% 초과만 `issues`).

        ⚠ 소스에 `warnings.append` 가 몇 개 있나로 재면 안 된다 — 두 개 중 하나를
          지워도 통과한다(실측 M232 생존). 실제 워크북을 만들어 **경고 문장이 나오는지**
          본다. 레이아웃은 모듈 상수에서 파생한다(테스트에 열 번호를 박지 않는다).
        """
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        from generators import suts

        wb = Workbook()
        ws = wb.active
        ws.title = "2.SW Unit Test Spec"
        # TC 3건 중 1건만 I/O 가 비어 있다 → 50% 이하 → `issues` 가 아니라 `warnings`.
        for i in range(3):
            r = suts._DATA_START_ROW + i
            ws.cell(row=r, column=suts._COL_TC_ID, value=f"SwUTC_{i:04d}")
            ws.cell(row=r, column=suts._SEQ_COL, value=1)          # 시퀀스는 전부 채운다
            if i < 2:                                              # 2건만 I/O 를 채운다
                ws.cell(row=r, column=suts._INPUT_COL_START, value="in")
                ws.cell(row=r, column=suts._OUTPUT_COL_START, value="out")
        p = tmp_path / "x.xlsm"
        wb.save(str(p))

        got = suts.validate_suts_xlsm(str(p))
        assert got["stats"]["tc_count"] == 3, got["stats"]
        assert got["stats"]["empty_io_tc_count"] == 1, got["stats"]
        joined = " ".join(got["warnings"])
        assert "1/3" in joined and "lack I/O" in joined, got["warnings"]
        # ⚠ 대조군 — 소수(50% 이하)는 `issues` 가 아니다. 여기서 issues 로 새면
        #   `valid` 판정이 바뀌어 정상 산출물이 결함으로 신고된다.
        assert not any("lack I/O" in i for i in got["issues"]), got["issues"]

    def test_suts_warns_when_a_tc_has_no_sequence(self, tmp_path):
        """시퀀스 수 < TC 수 = 일부 TC 에 시험 절차가 없다(산술적 사실)."""
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        from generators import suts

        wb = Workbook()
        ws = wb.active
        ws.title = "2.SW Unit Test Spec"
        for i in range(2):
            r = suts._DATA_START_ROW + i
            ws.cell(row=r, column=suts._COL_TC_ID, value=f"SwUTC_{i:04d}")
            ws.cell(row=r, column=suts._INPUT_COL_START, value="in")
            if i == 0:                                   # 1건만 시퀀스가 있다
                ws.cell(row=r, column=suts._SEQ_COL, value=1)
        p = tmp_path / "y.xlsm"
        wb.save(str(p))

        got = suts.validate_suts_xlsm(str(p))
        assert got["stats"]["tc_count"] == 2 and got["stats"]["seq_count"] == 1, got["stats"]
        assert any("no test sequence" in w for w in got["warnings"]), got["warnings"]

    def test_sits_returns_warnings_key_even_when_file_is_missing(self, tmp_path):
        from generators import sits

        got = sits.validate_sits_xlsm(str(tmp_path / "nope.xlsm"))
        assert got["valid"] is False
        assert "warnings" in got, got
        assert got["warnings"] == []

    def test_suts_returns_warnings_key_even_when_file_is_missing(self, tmp_path):
        from generators import suts

        got = suts.validate_suts_xlsm(str(tmp_path / "nope.xlsm"))
        assert got["valid"] is False
        assert "warnings" in got, got

    @pytest.mark.parametrize("kind", ["sts", "suts", "sits"])
    def test_the_api_surface_carries_the_warning_through(self, kind):
        """API 응답이 검증기의 경고를 **본문째** 실어 나르는가.

        ⚠ 표면이 둘이고 담는 것이 다르다(실측):
          · `payload["validation"]`        → 검증 dict **그대로**(경고 문장 포함)
          · `payload["summary"]["validation"]` → `issue_count`/`warning_count` **개수만**
        개수만 보는 화면은 "경고 1건" 까지만 알고 무엇인지는 모른다. 두 면을 같이 못박아
        둔다 — 한쪽만 보고 "경고가 사라졌다/남았다" 고 단정하면 틀린다.
        """
        from backend.helpers.common import _build_excel_artifact_payload

        got = _build_excel_artifact_payload(
            kind, {"validation": {"valid": True, "issues": [],
                                  "warnings": ["3/10 TCs lack I/O variables"], "stats": {}}})
        assert "lack I/O variables" in repr(got["validation"]), got["validation"]
        assert got["summary"]["validation"]["warning_count"] == 1, got["summary"]["validation"]
