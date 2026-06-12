"""라운드 91 — spec-based SUTR 빌더 (swut_sutr_spec_builder) 단위 테스트.

회사 감사본 양식(KJPDS02 v1.01 류) SUTR '3.Test Log' 빌드 경로 검증:
- SwUTS spec 시트 통째 복사 (Input/Expected 보존) + Actual/Pass-Fail/Log 추가.
- 함수 anchor 스캔, SwUFn 숫자 매칭, iteration Pass/Fail, JG Total 세로병합.
"""
from __future__ import annotations

import io
import os

import openpyxl
import pytest

from backend.services.swut_input_adapter import (
    EnvironmentData,
    ExecutionRow,
    SwUTSession,
)
from backend.services.swut_sutr_aggregator import SutrBuildMeta
from backend.services.swut_sutr_spec_builder import (
    ACTUAL_MAX,
    COL_ACTUAL_START,
    COL_EXPECTED_START,
    COL_INPUT_START,
    COL_ITER_INDEX,
    COL_LOG_DATA,
    COL_PASS_FAIL,
    COL_PASS_TOTAL,
    COL_RELATED_ID,
    LOG_SHEET_NAME,
    SUBHEADER_ROW,
    SpecLayout,
    _apply_actual_result_style,
    _build_fn_iteration_map,
    _detect_spec_layout,
    _lookup_vcast_actual,
    _scan_spec_blocks,
    build_sutr_from_spec,
)


# ---------------------------------------------------------------------------
# 합성 spec xlsm 생성 (회사 KJPDS02 v1.01 양식 축소판)
# ---------------------------------------------------------------------------

def _make_spec_bytes(blocks: list[tuple[str, str, list[tuple]]]) -> bytes:
    """합성 SwUTS spec xlsm bytes 생성.

    blocks: [(tc_id, unit, [(input_vals, expected_vals), ...]), ...]
        각 iteration tuple = (dict 형태 단순화 — 여기선 변수 1개씩 H/BF에).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    # 헤더
    ws.cell(1, 1).value = "Software Unit Test"
    ws.cell(3, 2).value = "Test Case"
    ws.cell(3, 7).value = "Input"
    ws.cell(3, 58).value = "Expected Result"
    ws.cell(3, 162).value = "Related ID"
    ws.cell(4, 2).value = "Index"
    ws.cell(4, 3).value = "TC_ID"
    ws.cell(4, 4).value = "Unit"
    ws.cell(4, 5).value = "Test Method"
    ws.cell(4, 6).value = "Test Case Generation"
    ws.cell(4, 7).value = " "
    ws.cell(4, 8).value = "Inpt[0]"
    ws.cell(4, 58).value = "ExpR[0]"
    rr = 5
    idx = 1
    for tc_id, unit, iters in blocks:
        anchor = rr
        ws.cell(anchor, 2).value = idx
        ws.cell(anchor, 3).value = tc_id
        ws.cell(anchor, 4).value = unit
        ws.cell(anchor, 8).value = "in_var"   # H 변수명
        ws.cell(anchor, 58).value = "exp_var"  # BF 변수명
        # spec Related ID 세로병합 (실제 양식 재현 — 빌더가 해제해야 함)
        if iters:
            ws.merge_cells(start_row=anchor, end_row=anchor + len(iters),
                           start_column=162, end_column=162)
        for it_no, (inp, exp) in enumerate(iters, start=1):
            ir = anchor + it_no
            ws.cell(ir, 5).value = "REQ"
            ws.cell(ir, 6).value = "ABV"
            ws.cell(ir, 7).value = it_no
            ws.cell(ir, 8).value = inp
            ws.cell(ir, 58).value = exp
        rr = anchor + len(iters) + 1
        idx += 1
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _make_session(fn_results: dict[str, list[bool | None]]) -> SwUTSession:
    """fn_results: {'SwUFn_0101': [True, False, None], ...} → session."""
    env = EnvironmentData(env_name="SwUT_01_Test", component_name="Test")
    for fn_id, results in fn_results.items():
        for i, passed in enumerate(results, start=1):
            key = f"{fn_id}.{i:03d}"
            env.test_cases[key] = []  # 빌더는 키만 사용 (Actual은 Expected 복제)
            if passed is not None:
                env.test_results[key] = ExecutionRow(
                    tc_name=key, passed=passed,
                    actual_result={"exp_var": ("actual_v", "exp_v")},
                )
    sess = SwUTSession(project_id="KJPDS02")
    sess.environments = [env]
    return sess


def _meta() -> SutrBuildMeta:
    return SutrBuildMeta(
        project_id="KJPDS02", project_full_name="KJPDS02", asil_level="ASIL A",
        doc_id_base="KJPDS02-SUTR", doc_id_sequence="1",
        default_author="JK Kim", default_approver="CH In",
        release_sw_version="1.01", test_date="2025-12-05", test_engineer="JK Kim",
    )


# ---------------------------------------------------------------------------
# 테스트
# ---------------------------------------------------------------------------

def test_scan_spec_blocks_basic():
    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("0x0", "0x1"), ("0x2", "0x3")]),
        ("SwUTC_0102", "foo", [("a", "b")]),
    ])
    wb = openpyxl.load_workbook(io.BytesIO(spec))
    ws = wb.active
    blocks = _scan_spec_blocks(ws)
    assert len(blocks) == 2
    assert blocks[0]["tc_id"] == "SwUTC_0101"
    assert blocks[0]["unit"] == "main"
    assert blocks[0]["num"] == "0101"
    assert len(blocks[0]["iter_rows"]) == 2
    assert len(blocks[1]["iter_rows"]) == 1


def test_build_fn_iteration_map():
    sess = _make_session({"SwUFn_0101": [True, False], "SwUFn_0102": [None]})
    m = _build_fn_iteration_map(sess)
    # 4-digit num은 lstrip("0") → "101"
    assert "101" in m
    assert m["101"][1]["passed"] is True
    assert m["101"][2]["passed"] is False
    assert "102" in m


def test_apply_actual_result_style_mirrors_expected():
    """라운드 104 — Actual(162~)이 Expected(58~) 서식을 미러 + 헤더행 미변경."""
    from openpyxl.styles import Border, Font, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    thin = Side(style="thin")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_row = SUBHEADER_ROW + 1
    # Expected 셀에 테두리/폰트 부여 (graft 서식 모사).
    exp = ws.cell(data_row, COL_EXPECTED_START)
    exp.border = bd
    exp.font = Font(name="Arial", size=10)
    # Expected 마지막(161) — Pass/Fail·Total·Log 미러 소스.
    last_exp = ws.cell(data_row, COL_ACTUAL_START - 1)
    last_exp.border = bd
    last_exp.font = Font(name="Arial", size=10)
    # 헤더 행(SUBHEADER_ROW)의 Actual 셀은 변경되면 안 됨 — 기본(무테) 확인용.
    hdr_actual = ws.cell(SUBHEADER_ROW, COL_ACTUAL_START)

    n = _apply_actual_result_style(ws)
    assert n > 0
    # Actual 대응 셀(162)이 Expected 테두리 미러.
    act = ws.cell(data_row, COL_ACTUAL_START)
    assert act.border.left is not None and act.border.left.style == "thin"
    assert act.font.name == "Arial"
    # Pass/Fail(266)·Pass(267) — 라운드 106 DV 감사본 명시 Border + 레퍼런스 폰트.
    # data_row == DATA_START_ROW(첫 데이터행) → Pass/Fail T:medium (헤더 경계).
    pf = ws.cell(data_row, COL_PASS_FAIL)
    assert pf.border.left.style == "double"
    assert pf.border.right.style == "thin"
    assert pf.border.top.style == "medium"
    assert pf.border.bottom.style == "thin"
    pt = ws.cell(data_row, COL_PASS_TOTAL)
    assert pt.border.left.style == "thin"
    assert pt.border.right.style == "thin"
    assert pt.border.top.style == "thin"
    assert pt.border.bottom.style == "thin"
    assert pf.font.name == "맑은 고딕"
    assert pf.font.sz == 10
    assert pt.font.name == "맑은 고딕"
    assert pt.font.sz == 10
    # Log Data(JH)는 값 비움 + L:thin / 첫 데이터행 T:medium / B 무테 (DV 실측).
    log = ws.cell(data_row, COL_LOG_DATA)
    assert log.border.left.style == "thin"
    assert log.border.top.style == "medium"
    assert log.border.bottom is None or log.border.bottom.style is None
    assert log.font.name == "맑은 고딕"
    assert log.font.sz == 10
    # 헤더 행은 미변경 (데이터 행만 대상).
    assert hdr_actual.border.left is None or hdr_actual.border.left.style is None


def test_lookup_vcast_actual_dotted():
    d = {"_LP0DR": ("0x3", "0x3"), "Byte": ("0x4", "0x4")}
    # dotted 변수명 분리 후보 매칭
    assert _lookup_vcast_actual(d, "_LP0DR.Byte") in ("0x3", "0x4")
    assert _lookup_vcast_actual(d, "missing") is None
    assert _lookup_vcast_actual({}, "x") is None


def test_build_sutr_from_spec_layout_matches_reference():
    """생성 '3.Test Log' 시트명 + 헤더 + Actual 채움 검증 (레퍼런스 양식)."""
    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("0x0", "0x1"), ("0x2", "0x3")]),
    ])
    sess = _make_session({"SwUFn_0101": [True, True]})
    res = build_sutr_from_spec(sess, _meta(), spec, function_asil_map={})
    assert res.ok
    wb = openpyxl.load_workbook(res.xlsm_io, data_only=False)
    assert LOG_SHEET_NAME in wb.sheetnames
    ws = wb[LOG_SHEET_NAME]
    # 헤더
    assert ws.cell(1, 1).value == "Software Unit Test Log"
    assert ws.cell(3, COL_ACTUAL_START).value == "Actual Result"
    assert ws.cell(3, COL_PASS_FAIL).value == "Pass/Fail"
    assert ws.cell(3, COL_PASS_TOTAL).value == "Pass"
    assert ws.cell(3, COL_LOG_DATA).value == "Log Data"
    assert ws.cell(4, COL_ACTUAL_START).value == "Param 1"
    assert ws.cell(4, COL_PASS_FAIL).value == "Unit"
    assert ws.cell(4, COL_PASS_TOTAL).value == "Pass"
    assert ws.cell(3, COL_PASS_FAIL).font.name == "맑은 고딕"
    assert ws.cell(3, COL_PASS_FAIL).font.sz == 10
    assert ws.cell(3, COL_PASS_FAIL).font.bold is True
    assert ws.cell(4, COL_PASS_FAIL).font.name == "맑은 고딕"
    assert ws.cell(4, COL_PASS_FAIL).font.sz == 10
    assert ws.cell(4, COL_PASS_FAIL).font.bold is True
    # Input/Expected 보존
    assert ws.cell(6, 8).value == "0x0"     # iteration1 input
    assert ws.cell(6, 58).value == "0x1"    # iteration1 expected
    # anchor Actual 변수명 = Expected 변수명
    assert ws.cell(5, COL_ACTUAL_START).value == "exp_var"
    # iteration Actual = Expected 복제 (Pass)
    assert ws.cell(6, COL_ACTUAL_START).value == "0x1"
    assert ws.cell(7, COL_ACTUAL_START).value == "0x3"
    # Pass/Fail
    assert ws.cell(6, COL_PASS_FAIL).value == "Pass"
    assert ws.cell(7, COL_PASS_FAIL).value == "Pass"
    # Total
    assert ws.cell(5, COL_PASS_TOTAL).value == "Pass"
    # Log Data 데이터 셀에는 로그 경로를 쓰지 않음(레퍼런스 정합).
    assert ws.cell(6, COL_LOG_DATA).value is None


def test_build_sutr_unmatched_function_na():
    """spec에 있지만 vcast 미실행 함수 → Pass/Fail N/A, Total N/A."""
    spec = _make_spec_bytes([
        ("SwUTC_0199", "orphan", [("x", "y")]),
    ])
    sess = _make_session({"SwUFn_0101": [True]})  # 0199 없음
    res = build_sutr_from_spec(sess, _meta(), spec, function_asil_map={})
    assert res.ok
    assert res.summary["unmatched_functions"] == 1
    ws = openpyxl.load_workbook(res.xlsm_io)[LOG_SHEET_NAME]
    assert ws.cell(6, COL_PASS_FAIL).value == "N/A"
    assert ws.cell(5, COL_PASS_TOTAL).value == "N/A"
    merges = [
        str(m) for m in ws.merged_cells.ranges
        if m.min_col == COL_PASS_TOTAL and m.min_row == 5
    ]
    assert merges, "미매칭 함수도 JG Total 세로병합 필요"


def test_build_sutr_fail_iteration():
    """Fail iteration → JF='Fail', Total='Fail'."""
    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("0x0", "0x1"), ("0x2", "0x3")]),
    ])
    sess = _make_session({"SwUFn_0101": [True, False]})
    res = build_sutr_from_spec(sess, _meta(), spec, function_asil_map={})
    ws = openpyxl.load_workbook(res.xlsm_io)[LOG_SHEET_NAME]
    assert ws.cell(6, COL_PASS_FAIL).value == "Pass"
    assert ws.cell(7, COL_PASS_FAIL).value == "Fail"
    assert ws.cell(5, COL_PASS_TOTAL).value == "Fail"  # 하나라도 Fail


def test_build_sutr_uses_spec_iteration_number_when_rows_have_gap():
    """G열 iteration 번호가 비어도 해당 SwUFn suffix 결과를 매칭한다."""
    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("0x0", "0x1"), ("0x2", "0x3")]),
    ])
    wb = openpyxl.load_workbook(io.BytesIO(spec), keep_vba=True)
    ws = wb["2.SW Unit Test Spec"]
    ws.cell(7, 7).value = 3
    buf = io.BytesIO()
    wb.save(buf)
    sess = _make_session({"SwUFn_0101": [True, None, True]})

    res = build_sutr_from_spec(sess, _meta(), buf.getvalue(), function_asil_map={})
    out = openpyxl.load_workbook(res.xlsm_io, keep_vba=True)
    log = out[LOG_SHEET_NAME]

    assert log.cell(6, COL_PASS_FAIL).value == "Pass"
    assert log.cell(7, COL_PASS_FAIL).value == "Pass"
    assert log.cell(7, COL_ACTUAL_START).value == "0x3"
    assert log.cell(5, COL_PASS_TOTAL).value == "Pass"


def test_build_sutr_total_vertical_merge():
    """JG Total 셀이 함수 블록 전체 세로병합."""
    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("a", "b"), ("c", "d"), ("e", "f")]),
    ])
    sess = _make_session({"SwUFn_0101": [True, True, True]})
    res = build_sutr_from_spec(sess, _meta(), spec, function_asil_map={})
    ws = openpyxl.load_workbook(res.xlsm_io)[LOG_SHEET_NAME]
    merges = [str(m) for m in ws.merged_cells.ranges
              if m.min_col == COL_PASS_TOTAL and m.min_row == 5]
    assert merges, "JG Total 세로병합 없음"
    rng = next(m for m in ws.merged_cells.ranges
               if m.min_col == COL_PASS_TOTAL and m.min_row == 5)
    assert rng.max_row == 8  # anchor(5) + 3 iteration


def test_build_sutr_no_spec_sheet_fails():
    """spec 시트 없는 xlsm → ok=False."""
    wb = openpyxl.Workbook()
    wb.active.title = "Random"
    bio = io.BytesIO()
    wb.save(bio)
    res = build_sutr_from_spec(_make_session({}), _meta(), bio.getvalue())
    assert res.ok is False


def test_build_sutr_invalid_date_raises():
    spec = _make_spec_bytes([("SwUTC_0101", "main", [("a", "b")])])
    meta = _meta()
    meta.test_date = "251205"  # 잘못된 형식
    with pytest.raises(Exception):
        build_sutr_from_spec(_make_session({}), meta, spec)


def test_build_sutr_asil_marking_applied():
    """function_asil_map 제공 시 anchor Total 셀에 ASIL 시각 강조 fill 적용."""
    spec = _make_spec_bytes([("SwUTC_0101", "main", [("a", "b")])])
    sess = _make_session({"SwUFn_0101": [True]})
    res = build_sutr_from_spec(
        sess, _meta(), spec, function_asil_map={"SwUFn_0101": "A"},
    )
    ws = openpyxl.load_workbook(res.xlsm_io)[LOG_SHEET_NAME]
    fill = ws.cell(5, COL_PASS_TOTAL).fill
    rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
    assert rgb not in (None, "00000000"), "ASIL A 마킹 fill 미적용"


def test_is_sutr_spec_based_flag():
    """라우터 분기 헬퍼 — config sutr_spec_based 플래그."""
    from backend.routers.swut import _is_sutr_spec_based

    class _Req:
        project_id = "KJPDS02"

    assert _is_sutr_spec_based(_Req(), {"sutr_spec_based": True}) is True
    assert _is_sutr_spec_based(_Req(), {"sutr_spec_based": False}) is False
    assert _is_sutr_spec_based(_Req(), {}) is False  # 미설정 → backward-compat


# ---------------------------------------------------------------------------
# 라운드 92 — 표준 SUTR 템플릿 베이스 + spec 시트 이식
# ---------------------------------------------------------------------------

def _make_standard_template_bytes() -> bytes:
    """합성 표준 SUTR 템플릿 (Cover/History/1.Test Summary/2.Deviation/3.Test Result)."""
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.cell(3, 2).value = "Project"
    cover.cell(4, 2).value = "ASIL Level"
    cover.cell(5, 2).value = "Version"
    cover.cell(6, 2).value = "Test Date"
    cover.cell(7, 2).value = "Author"
    cover.cell(8, 2).value = "Approver"

    hist = wb.create_sheet("History")
    hist.cell(3, 2).value = "No."
    hist.cell(3, 3).value = "Date"
    hist.cell(3, 4).value = "Version"
    hist.cell(3, 5).value = "Description"
    hist.cell(3, 6).value = "Author"

    ts = wb.create_sheet("1.Test Summary")
    ts.cell(4, 2).value = "Project Name"
    ts.cell(5, 2).value = "SW Version"
    ts.cell(6, 2).value = "HW Version"
    ts.cell(7, 2).value = "Test Date"
    ts.cell(8, 2).value = "Test Engineer"
    ts.cell(17, 2).value = "Total Number of TCs"
    ts.cell(17, 3).value = "Number of TCs Tested"
    ts.cell(17, 4).value = "Number of TCs Passed"
    ts.cell(17, 5).value = "Number of TCs Failed"
    ts.cell(17, 6).value = "Number of TCs not executed"
    ts.cell(21, 2).value = "Source"
    ts.cell(22, 2).value = "System Design"

    dev = wb.create_sheet("2.Deviation")
    dev.cell(4, 2).value = "Test Case ID"
    dev.cell(4, 3).value = "Issue"
    dev.cell(4, 4).value = "Deviation"
    dev.cell(4, 5).value = "Status"

    res = wb.create_sheet("3.Test Result")
    res.cell(1, 1).value = "Narrow 38-col result"
    res.cell(5, 38).value = "edge"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_r92_template_base_sheet_composition():
    """R92 — 표준 템플릿 베이스 시 시트 구성이 레퍼런스 정합.

    [Cover, History, 1.Test Summary, 2.Deviation, 3.Test Log] (+AuditLog)
    Introduction / 1.Test Environment / 3.Test Result 없음.
    """
    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("0x0", "0x1"), ("0x2", "0x3")]),
        ("SwUTC_0102", "foo", [("a", "b")]),
    ])
    tmpl = _make_standard_template_bytes()
    sess = _make_session({"SwUFn_0101": [True, True], "SwUFn_0102": [False]})
    res = build_sutr_from_spec(
        sess, _meta(), spec, template_xlsm_bytes=tmpl, function_asil_map={},
    )
    assert res.ok
    wb = openpyxl.load_workbook(res.xlsm_io, data_only=False)
    names = wb.sheetnames
    # 좁은 3.Test Result 제거됨
    assert "3.Test Result" not in names
    # 레퍼런스 4 시트 + 3.Test Log 존재
    for must in ("Cover", "History", "1.Test Summary", "2.Deviation", "3.Test Log"):
        assert must in names, f"{must} 누락 (시트: {names})"
    # spec 전용 시트 미존재
    assert not any("introduction" in n.lower() for n in names)
    assert not any("environment" in n.lower() for n in names)


def test_r92_test_log_wide_sheet_grafted():
    """R92 — 이식된 3.Test Log가 spec 와이드 양식(value/merge) 보존."""
    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("0x0", "0x1"), ("0x2", "0x3")]),
    ])
    tmpl = _make_standard_template_bytes()
    sess = _make_session({"SwUFn_0101": [True, True]})
    res = build_sutr_from_spec(
        sess, _meta(), spec, template_xlsm_bytes=tmpl, function_asil_map={},
    )
    wb = openpyxl.load_workbook(res.xlsm_io, data_only=False)
    ws = wb["3.Test Log"]
    # spec 보존 — Input/Expected
    assert ws.cell(6, 8).value == "0x0"
    assert ws.cell(6, 58).value == "0x1"
    # R91 fill 로직 적용 — Actual/Pass-Fail
    assert ws.cell(3, COL_ACTUAL_START).value == "Actual Result"
    assert ws.cell(6, COL_ACTUAL_START).value == "0x1"  # Pass → Expected 복제
    assert ws.cell(6, COL_PASS_FAIL).value == "Pass"
    assert ws.cell(5, COL_PASS_TOTAL).value == "Pass"


def test_r92_test_summary_function_counts():
    """R92 — 1.Test Summary TC 카운트가 함수 단위(레퍼런스 정합)로 stamp."""
    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("a", "b")]),       # pass fn
        ("SwUTC_0102", "foo", [("c", "d")]),         # fail fn
        ("SwUTC_0103", "bar", [("e", "f")]),         # na fn (vcast 미실행)
    ])
    tmpl = _make_standard_template_bytes()
    sess = _make_session({"SwUFn_0101": [True], "SwUFn_0102": [False]})
    res = build_sutr_from_spec(
        sess, _meta(), spec, template_xlsm_bytes=tmpl, function_asil_map={},
    )
    assert res.summary["test_summary_tc_total"] == 3
    assert res.summary["test_summary_passed"] == 1
    assert res.summary["test_summary_failed"] == 1
    assert res.summary["test_summary_not_executed"] == 1
    assert res.summary["test_summary_tested"] == 2
    wb = openpyxl.load_workbook(res.xlsm_io)
    ts = wb["1.Test Summary"]
    assert ts.cell(18, 2).value == 3   # Total
    assert ts.cell(18, 3).value == 2   # Tested
    assert ts.cell(18, 4).value == 1   # Passed
    assert ts.cell(18, 5).value == 1   # Failed
    assert ts.cell(18, 6).value == 1   # not executed


def test_r92_cover_meta_stamped():
    """R92 — Cover 시트 meta(Project/Version/Date) stamp."""
    spec = _make_spec_bytes([("SwUTC_0101", "main", [("a", "b")])])
    tmpl = _make_standard_template_bytes()
    sess = _make_session({"SwUFn_0101": [True]})
    res = build_sutr_from_spec(
        sess, _meta(), spec, template_xlsm_bytes=tmpl, function_asil_map={},
    )
    wb = openpyxl.load_workbook(res.xlsm_io)
    cover = wb["Cover"]
    # 표준 _write_cover 가 stamp 하는 라벨: Project / ASIL Level / Version.
    assert cover.cell(3, 3).value == "KJPDS02"          # Project value (C열)
    assert cover.cell(4, 3).value == "ASIL A"           # ASIL Level
    assert cover.cell(5, 3).value == "v1.01"            # Version
    # Test Date 는 1.Test Summary 에 stamp.
    ts = wb["1.Test Summary"]
    assert ts.cell(7, 3).value == "2025-12-05"          # Test Date (Summary)


def test_copy_sheet_across_workbooks_fidelity():
    """크로스-워크북 시트 복사 helper — value/merge/width/height 보존."""
    from backend.services.excel_template_utils import copy_sheet_across_workbooks

    src_wb = openpyxl.Workbook()
    src = src_wb.active
    src.title = "Src"
    src.cell(1, 1).value = "hello"
    src.cell(2, 3).value = 42
    src.merge_cells("A5:C5")
    src.column_dimensions["B"].width = 33.0
    src.row_dimensions[4].height = 22.0

    dst_wb = openpyxl.Workbook()
    dst = copy_sheet_across_workbooks(src, dst_wb, new_title="3.Test Log")
    assert "3.Test Log" in dst_wb.sheetnames
    assert dst.cell(1, 1).value == "hello"
    assert dst.cell(2, 3).value == 42
    assert "A5:C5" in [str(m) for m in dst.merged_cells.ranges]
    assert dst.column_dimensions["B"].width == 33.0
    assert dst.row_dimensions[4].height == 22.0


def test_r91_fallback_when_no_template():
    """R92 — template None이면 R91 fallback (spec wb 베이스) + warning."""
    spec = _make_spec_bytes([("SwUTC_0101", "main", [("a", "b")])])
    sess = _make_session({"SwUFn_0101": [True]})
    res = build_sutr_from_spec(sess, _meta(), spec, function_asil_map={})
    assert res.ok
    assert res.summary["builder"] == "spec-based-r91"
    assert any("표준 SUTR 템플릿 미제공" in w for w in res.warnings)


# ---------------------------------------------------------------------------
# 라운드 96-final QA fix — C-4 (coverage 수식 통일) + W-14 (col dims 범위 복제)
# ---------------------------------------------------------------------------

def test_c4_requirements_coverage_formula_unified(monkeypatch):
    """C-4 — Requirements coverage 수식이 DV 감사본 형식(N/A 가산)으로 통일.

    템플릿 잔존 '=IFERROR(Dn/Cn,"")'(N/A 미가산)과 'Actual Coverage' 리터럴이
    모순 수치를 표시하던 결함 — 수식 셀은 '(D+E)/C' 형식으로 교체, Actual
    Coverage는 그 셀 참조 수식('=F22')으로 단일 진리원화. 좌표는 라벨 동적.
    """
    import backend.services.swut_meta_resolver as _meta_resolver
    monkeypatch.setattr(_meta_resolver, "load_meta_from_config", lambda pid: {})

    tmpl_wb = openpyxl.load_workbook(io.BytesIO(_make_standard_template_bytes()))
    ts = tmpl_wb["1.Test Summary"]
    ts.cell(10, 2).value = "Actual Coverage"
    # 동일 블록 잔존 템플릿 수식 (N/A 미가산형) — 통일 대상
    ts.cell(21, 6).value = '=IFERROR(D21/C21,"")'
    bio = io.BytesIO()
    tmpl_wb.save(bio)

    spec = _make_spec_bytes([
        ("SwUTC_0101", "main", [("a", "b")]),
        ("SwUTC_0103", "bar", [("e", "f")]),  # 미실행 fn → not tested 1
    ])
    sess = _make_session({"SwUFn_0101": [True]})
    res = build_sutr_from_spec(
        sess, _meta(), spec, template_xlsm_bytes=bio.getvalue(),
        function_asil_map={},
    )
    assert res.ok
    out = openpyxl.load_workbook(res.xlsm_io, data_only=False)
    ts_out = out["1.Test Summary"]
    # 'System Design' row 22 — coverage 수식 F22가 DV 형식 (tested+N/A 가산)
    assert ts_out.cell(22, 6).value == '=IFERROR((D22+E22)/C22,"")'
    # 동일 블록 잔존 템플릿 수식(F21)도 통일
    assert ts_out.cell(21, 6).value == '=IFERROR((D21+E21)/C21,"")'
    # 'Actual Coverage'(B10) 우측 — 리터럴 대신 F22 참조 수식 (단일 진리원)
    assert ts_out.cell(10, 3).value == "=F22"
    assert res.summary["requirements_coverage_formula"] == '=IFERROR((D22+E22)/C22,"")'
    assert res.summary["actual_coverage_formula"] == "=F22"


def test_w14_copy_sheet_preserves_column_dimension_ranges():
    """W-14 — column_dimensions 범위 정의(min/max)·hidden·width 복제 보존.

    openpyxl은 '<col min="18" max="57" hidden="1"/>' 범위를 첫 열 key 1개에
    min/max로 보존 — min/max 미복제 시 단일 열로 축소돼 DV spec 시트의 hidden
    범위(cols 18-57 등)·폭 설정이 graft에서 소실되던 결함.
    """
    from backend.services.excel_template_utils import copy_sheet_across_workbooks

    src_wb = openpyxl.Workbook()
    src = src_wb.active
    src.title = "Src"
    src.cell(1, 1).value = "x"
    dim = src.column_dimensions["R"]  # col 18
    dim.min = 18
    dim.max = 57
    dim.hidden = True
    dim.width = 9.0

    dst_wb = openpyxl.Workbook()
    dst = copy_sheet_across_workbooks(src, dst_wb, new_title="3.Test Log")
    dd = dst.column_dimensions["R"]
    assert dd.min == 18
    assert dd.max == 57
    assert dd.hidden is True
    assert dd.width == 9.0


def test_not_executed_list_skips_column_subheader():
    """라운드 97 재검증 — W-4 발현 보장: 컬럼 sub-header에서 break 금지.

    KJPDS02 v1.01 실측: '■ List of Test Case not Executed'(R26) 바로 아래
    R27이 컬럼 sub-header('Test Case ID'/'Rationale ...')라 기존 스캔이
    즉시 break → 가용 행 0 → 목록 영영 미기재. sub-header는 skip하고 그
    아래 빈 행(R28~)에 대표 목록 + '외 N건' 요약을 기재해야 한다.
    """
    from backend.services.swut_sutr_spec_builder import _write_not_executed_list

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1.Test Summary"
    ws.cell(row=26, column=2, value="■ List of Test Case not Executed")
    ws.cell(row=27, column=2, value="Test Case ID")
    ws.cell(row=27, column=3, value="Rationale why test case is not executed")
    ws.cell(row=27, column=6, value="When test case executes")
    ws.cell(row=32, column=2, value="■  Test Defects List")

    fill_stats = {
        "fn_na": 44,
        "na_tc_list": [f"SwUT_x_{i:02d} (SwUFn_{i:04d})" for i in range(1, 9)],
    }
    summary: dict = {}
    warnings: list[str] = []
    written = _write_not_executed_list(wb, fill_stats, summary, warnings)

    assert written == 4  # 대표 3 (R28~R30) + '외 41건' 요약 (R31)
    assert ws.cell(row=28, column=2).value == "SwUT_x_01 (SwUFn_0001)"
    assert ws.cell(row=30, column=2).value == "SwUT_x_03 (SwUFn_0003)"
    assert "외 41건" in str(ws.cell(row=31, column=2).value)
    # sub-header(R27)·다음 섹션(R32) 침범 금지.
    assert ws.cell(row=27, column=2).value == "Test Case ID"
    assert ws.cell(row=32, column=2).value == "■  Test Defects List"
    assert summary["not_executed_list_rows"] == 4
    assert summary["not_executed_total"] == 44
    assert not [w for w in warnings if "가용 행 0" in w]


def test_not_executed_list_without_subheader_unchanged():
    """sub-header 없는 양식(헤더 바로 아래 빈 행)은 기존 동작 그대로."""
    from backend.services.swut_sutr_spec_builder import _write_not_executed_list

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1.Test Summary"
    ws.cell(row=10, column=2, value="■ List of Test Case not Executed")
    ws.cell(row=15, column=2, value="■  Test Defects List")

    fill_stats = {"fn_na": 2, "na_tc_list": ["TC_A (SwUFn_0001)", "TC_B (SwUFn_0002)"]}
    summary: dict = {}
    warnings: list[str] = []
    written = _write_not_executed_list(wb, fill_stats, summary, warnings)

    assert written == 2  # 전건 기재 — 요약 행 없음
    assert ws.cell(row=11, column=2).value == "TC_A (SwUFn_0001)"
    assert ws.cell(row=12, column=2).value == "TC_B (SwUFn_0002)"
    assert ws.cell(row=13, column=2).value is None


# ---------------------------------------------------------------------------
# 라운드 105 — spec 컬럼 레이아웃 동적화 (DV 268열 / PV 'Safety Related' 삽입 공존)
#
# PV WIP spec(v0.10_260608) 실측 레이아웃: Safety Related(5) 삽입으로 Test
# Method=6/Generation=7/iter=8/Inpt[0]=9, ExpR[0..83]=105~188, Related ID(r4=
# 'SUDS')=189. DV 고정 상수(162/266/267/268) 그대로 쓰면 Actual stamp가
# Expected(105~188)·SUDS(189)를 덮어쓴다 — _detect_spec_layout이 헤더 스캔으로
# SpecLayout을 동적 산출하고, DV spec에서는 산출값 == 기존 상수(하위 호환 게이트).
# ---------------------------------------------------------------------------

# PV 레이아웃 좌표 (작성중 v0.10_260608 실측 축소판)
_PV_ITER = 8           # iteration index 열 (DV 7)
_PV_INPUT = 9          # Inpt[0] (DV 8)
_PV_EXP0 = 105         # ExpR[0] (DV 58)
_PV_EXP_LAST = 188     # ExpR[83] (DV 161)
_PV_REL = 189          # Related ID / r4 'SUDS' (DV 162)
_PV_PASS_FAIL = 273    # Actual 끝 +1 (DV 266)
_PV_PASS_TOTAL = 274   # (DV 267)
_PV_LOG = 275          # (DV 268)

_YELLOW = "FFFFEB9C"   # design_tokens.USER_INPUT_FILL_RGB (openpyxl ARGB)


def _make_pv_spec_bytes(
    blocks: list[tuple[str, str, list[tuple]]],
    *,
    with_r3_headers: bool = True,
) -> bytes:
    """합성 PV SwUTS spec xlsm (작성중 v0.10_260608 양식 축소판 — 189열).

    blocks: [(tc_id, unit, [(input, exp1, exp2), ...]), ...]
        변수 2개: exp_var@105(ExpR[0]) / exp_var2@188(ExpR[83] — 경계 검증용).
    with_r3_headers=False면 r3 그룹 헤더 생략 — r4 'ExpR[..]' 폴백 검증용.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    ws.cell(1, 1).value = "Software Unit Test"
    if with_r3_headers:
        ws.cell(3, 2).value = "SwUTC_Test Case"
        ws.cell(3, _PV_ITER).value = "Input"
        ws.cell(3, _PV_EXP0).value = "Expected Result"
        # 실물 PV spec의 r3 Expected 병합(DA3:GF3) 재현 — 빌더가 보존해야 함.
        ws.merge_cells(start_row=3, end_row=3,
                       start_column=_PV_EXP0, end_column=_PV_EXP_LAST)
        ws.cell(3, _PV_REL).value = "Related ID"
    ws.cell(4, 2).value = "Index"
    ws.cell(4, 3).value = "TC_ID"
    ws.cell(4, 4).value = "Unit"
    ws.cell(4, 5).value = "Safety Related"
    ws.cell(4, 6).value = "Test Method"
    ws.cell(4, 7).value = "Test Case Generation Method"
    ws.cell(4, _PV_ITER).value = " "
    ws.cell(4, _PV_INPUT).value = "Inpt[0]"
    for i in range(84):  # ExpR[0..83] = 105~188
        ws.cell(4, _PV_EXP0 + i).value = f"ExpR[{i}]"
    ws.cell(4, _PV_REL).value = "SUDS"
    rr = 5
    idx = 1
    for tc_id, unit, iters in blocks:
        anchor = rr
        ws.cell(anchor, 2).value = idx
        ws.cell(anchor, 3).value = tc_id
        ws.cell(anchor, 4).value = unit
        ws.cell(anchor, 5).value = "O"
        ws.cell(anchor, _PV_INPUT).value = "in_var"
        ws.cell(anchor, _PV_EXP0).value = "exp_var"
        ws.cell(anchor, _PV_EXP_LAST).value = "exp_var2"
        ws.cell(anchor, _PV_REL).value = f"SwUDS_{idx:04d}"  # SUDS 추적성 참조
        if iters:
            # 실물 spec의 Related ID 세로병합 재현 — 빌더가 해제해야 함.
            ws.merge_cells(start_row=anchor, end_row=anchor + len(iters),
                           start_column=_PV_REL, end_column=_PV_REL)
        for it_no, (inp, exp1, exp2) in enumerate(iters, start=1):
            ir = anchor + it_no
            ws.cell(ir, 6).value = "REQ"
            if it_no == 1:
                ws.cell(ir, 7).value = "ABV"
            ws.cell(ir, _PV_ITER).value = it_no
            ws.cell(ir, _PV_INPUT).value = inp
            ws.cell(ir, _PV_EXP0).value = exp1
            ws.cell(ir, _PV_EXP_LAST).value = exp2
        if len(iters) > 1:
            # 실물 PV spec의 Generation Method(7) 세그먼트 병합 재현 — 비-anchor
            # 행은 col 7이 None이라 고정 iter=7 사용 시 iteration 행 누락(실측 67%).
            ws.merge_cells(start_row=anchor + 1, end_row=anchor + len(iters),
                           start_column=7, end_column=7)
        rr = anchor + len(iters) + 1
        idx += 1
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class TestR105SpecLayoutDetection:
    """(a) 동적 레이아웃 — DV 상수 재현(하위 호환 게이트) + PV 산출 + fallback."""

    def test_spec_layout_defaults_equal_dv_constants(self):
        """SpecLayout() 기본값(미스캔/레거시 직접 호출)이 기존 DV 상수와 동일.

        기존 268열 테스트가 layout 미전달로도 무수정 통과하는 근거 — 파생
        property 전부 단일 진리원 검증.
        """
        lo = SpecLayout()
        assert lo.detected is False
        assert lo.expected_start == COL_EXPECTED_START == 58
        assert lo.related_id == COL_RELATED_ID == 162
        assert lo.actual_start == COL_ACTUAL_START == 162
        assert lo.actual_max == ACTUAL_MAX == 104
        assert lo.pass_fail == COL_PASS_FAIL == 266
        assert lo.pass_total == COL_PASS_TOTAL == 267
        assert lo.log_data == COL_LOG_DATA == 268
        assert lo.input_start == COL_INPUT_START == 8
        assert lo.iter_index == COL_ITER_INDEX == 7

    def test_detect_dv_layout_reproduces_constants(self):
        """DV 합성 spec 헤더 스캔 산출값 == 기존 상수 (하위 호환 게이트)."""
        spec = _make_spec_bytes([("SwUTC_0101", "main", [("a", "b")])])
        ws = openpyxl.load_workbook(io.BytesIO(spec))["2.SW Unit Test Spec"]
        warns: list[str] = []
        lo = _detect_spec_layout(ws, warns)
        assert lo.detected is True
        assert lo.expected_start == COL_EXPECTED_START
        assert lo.actual_start == COL_ACTUAL_START
        assert lo.actual_max == ACTUAL_MAX
        assert lo.pass_fail == COL_PASS_FAIL
        assert lo.pass_total == COL_PASS_TOTAL
        assert lo.log_data == COL_LOG_DATA
        assert lo.iter_index == COL_ITER_INDEX
        assert lo.input_start == COL_INPUT_START
        assert warns == []

    def test_detect_pv_layout_from_r3_headers(self):
        """PV 합성 spec(189열) — Actual=Related ID(189), Pass/Fail=273/274/275."""
        spec = _make_pv_spec_bytes([("SwUFn_0101", "main", [("a", "b", "c")])])
        ws = openpyxl.load_workbook(io.BytesIO(spec))["2.SW Unit Test Spec"]
        lo = _detect_spec_layout(ws)
        assert lo.detected is True
        assert lo.expected_start == _PV_EXP0
        assert lo.related_id == _PV_REL
        assert lo.actual_start == _PV_REL      # Actual 시작 = Related ID 열 대체
        assert lo.actual_max == 84             # Expected 폭 (ExpR[0..83])
        assert lo.pass_fail == _PV_PASS_FAIL
        assert lo.pass_total == _PV_PASS_TOTAL
        assert lo.log_data == _PV_LOG
        assert lo.input_start == _PV_INPUT
        assert lo.iter_index == _PV_ITER

    def test_detect_pv_layout_r4_expr_fallback_without_r3(self):
        """r3 그룹 헤더 부재 시 r4 'ExpR[..]' 범위 폴백 (Related ID=마지막+1)."""
        spec = _make_pv_spec_bytes(
            [("SwUFn_0101", "main", [("a", "b", "c")])], with_r3_headers=False,
        )
        ws = openpyxl.load_workbook(io.BytesIO(spec))["2.SW Unit Test Spec"]
        lo = _detect_spec_layout(ws)
        assert lo.detected is True
        assert lo.expected_start == _PV_EXP0
        assert lo.related_id == _PV_REL
        assert lo.input_start == _PV_INPUT

    def test_detect_scan_failure_falls_back_dv_constants_with_warning(self):
        """헤더 전무(빈 시트) — DV 상수 fallback + warning (silent 오배치 금지)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        warns: list[str] = []
        lo = _detect_spec_layout(ws, warns)
        assert lo.detected is False
        assert lo.actual_start == COL_ACTUAL_START
        assert lo.pass_fail == COL_PASS_FAIL
        assert any("스캔 실패" in w for w in warns)

    def test_build_sutr_dv_summary_spec_layout_observability(self):
        """DV 빌드 summary.spec_layout 관측성 — 산출값 == 기존 상수 전체."""
        spec = _make_spec_bytes([("SwUTC_0101", "main", [("a", "b")])])
        sess = _make_session({"SwUFn_0101": [True]})
        res = build_sutr_from_spec(sess, _meta(), spec, function_asil_map={})
        assert res.ok
        assert res.summary["spec_layout"] == {
            "detected": True,
            "expected_start": 58,
            "actual_start": 162,
            "actual_max": 104,
            "pass_fail": 266,
            "pass_total": 267,
            "log_data": 268,
            "iter_index": 7,
        }


class TestR105PvLayoutBuild:
    """(a) PV 레이아웃 e2e — Actual=189 시작·Expected(105~188) 비침범·273/274/275."""

    def _build_pv(self):
        spec = _make_pv_spec_bytes([
            ("SwUFn_0101", "main",
             [("0x0", "0x1", "0xA"), ("0x2", "0x3", "0xB"), ("0x4", "0x5", "0xC")]),
        ])
        # it1 Pass / it2 Fail / it3 미실행 — Actual 3경로(Expected 복제/vcast/N-A).
        sess = _make_session({"SwUFn_0101": [True, False, None]})
        res = build_sutr_from_spec(sess, _meta(), spec, function_asil_map={})
        assert res.ok
        return res, openpyxl.load_workbook(res.xlsm_io, data_only=False)[LOG_SHEET_NAME]

    def test_pv_headers_at_dynamic_positions(self):
        res, ws = self._build_pv()
        assert ws.cell(3, _PV_REL).value == "Actual Result"
        assert ws.cell(3, _PV_PASS_FAIL).value == "Pass/Fail"
        assert ws.cell(3, _PV_PASS_TOTAL).value == "Pass"
        assert ws.cell(3, _PV_LOG).value == "Log Data"
        assert ws.cell(4, _PV_REL).value == "Param 1"            # r4 'SUDS' 대체
        assert ws.cell(4, _PV_REL + 83).value == "Param 84"      # Actual 폭 84
        assert ws.cell(4, _PV_PASS_FAIL).value == "Unit"
        assert ws.cell(4, _PV_PASS_TOTAL).value == "Pass"
        # DV 고정 좌표에는 아무 것도 stamp되지 않음 (구 상수 회귀 가드).
        assert ws.cell(3, COL_PASS_FAIL).value is None           # 266
        assert ws.cell(3, COL_LOG_DATA).value is None            # 268
        # r2 COUNTIF 요약 수식이 동적 열 문자 사용.
        from openpyxl.utils import get_column_letter
        jf = get_column_letter(_PV_PASS_FAIL)
        jg = get_column_letter(_PV_PASS_TOTAL)
        assert f"{jf}5:" in str(ws.cell(2, _PV_PASS_FAIL).value)
        assert f"{jg}5:" in str(ws.cell(2, _PV_PASS_TOTAL).value)
        # summary 관측성 — PV 산출 좌표 전체.
        assert res.summary["spec_layout"] == {
            "detected": True,
            "expected_start": _PV_EXP0,
            "actual_start": _PV_REL,
            "actual_max": 84,
            "pass_fail": _PV_PASS_FAIL,
            "pass_total": _PV_PASS_TOTAL,
            "log_data": _PV_LOG,
            "iter_index": _PV_ITER,
        }

    def test_pv_expected_region_not_invaded(self):
        """Expected(105~188) 헤더·데이터·r3 병합 비침범 — DV 상수 사용 시 깨지던 영역."""
        _res, ws = self._build_pv()
        # r3 'Expected Result' anchor + 병합(DA3:GF3) 보존.
        assert ws.cell(3, _PV_EXP0).value == "Expected Result"
        from openpyxl.utils import get_column_letter
        exp_merge = (
            f"{get_column_letter(_PV_EXP0)}3:{get_column_letter(_PV_EXP_LAST)}3"
        )
        assert exp_merge in [str(m) for m in ws.merged_cells.ranges]
        # r4 ExpR 서브헤더 보존 (경계 — DV 상수면 'Param N'으로 덮임).
        assert ws.cell(4, _PV_EXP0).value == "ExpR[0]"
        assert ws.cell(4, 162).value == "ExpR[57]"   # 구 COL_ACTUAL_START 위치
        assert ws.cell(4, _PV_EXP_LAST).value == "ExpR[83]"
        # Expected 데이터 보존 (iteration 행).
        assert ws.cell(6, _PV_EXP0).value == "0x1"
        assert ws.cell(6, _PV_EXP_LAST).value == "0xA"
        assert ws.cell(7, _PV_EXP_LAST).value == "0xB"
        # Input 보존.
        assert ws.cell(6, _PV_INPUT).value == "0x0"

    def test_pv_actual_starts_at_related_id_with_packed_vars(self):
        """Actual 시작=189 (Related ID/SUDS 대체) + 변수 packed stamp."""
        res, ws = self._build_pv()
        # anchor: SUDS 참조가 Actual 변수명으로 대체 (DV 레퍼런스 규칙).
        assert ws.cell(5, _PV_REL).value == "exp_var"
        assert ws.cell(5, _PV_REL + 1).value == "exp_var2"   # packed (188→190)
        # it1 Pass — Actual = Expected 복제.
        assert ws.cell(6, _PV_REL).value == "0x1"
        assert ws.cell(6, _PV_REL + 1).value == "0xA"
        # it2 Fail — exp_var는 vcast actual, exp_var2는 부재 → 노란 마킹.
        assert ws.cell(7, _PV_REL).value == "actual_v"
        assert ws.cell(7, _PV_REL + 1).value is None
        assert ws.cell(7, _PV_REL + 1).fill.start_color.rgb == _YELLOW
        # it3 미실행 — Actual 비움.
        assert ws.cell(8, _PV_REL).value is None
        assert res.summary["actual_from_expected"] == 2
        assert res.summary["actual_from_vcast"] == 1
        assert res.summary["actual_missing"] == 1

    def test_pv_pass_fail_total_positions_and_merge(self):
        """Pass/Fail=273, 함수 Total=274 세로병합, Related ID 병합 해제."""
        _res, ws = self._build_pv()
        assert ws.cell(6, _PV_PASS_FAIL).value == "Pass"
        assert ws.cell(7, _PV_PASS_FAIL).value == "Fail"
        assert ws.cell(8, _PV_PASS_FAIL).value == "N/A"
        assert ws.cell(5, _PV_PASS_FAIL).value == "Fail"     # anchor = 함수 결과
        assert ws.cell(5, _PV_PASS_TOTAL).value == "Fail"
        merges = [
            m for m in ws.merged_cells.ranges
            if m.min_col == _PV_PASS_TOTAL and m.min_row == 5
        ]
        assert merges and merges[0].max_row == 8
        # spec Related ID(189) 세로병합은 해제됨 (겹침 병합/anchor 흘림 차단).
        assert not [
            m for m in ws.merged_cells.ranges
            if m.min_col == _PV_REL and m.max_col == _PV_REL and m.min_row == 5
        ]

    def test_pv_matched_function_counts(self):
        """(b) PV SwUFn_ TC_ID 직접 표기 블록이 VectorCAST와 매칭."""
        res, _ws = self._build_pv()
        assert res.summary["spec_function_blocks"] == 1
        assert res.summary["matched_functions"] == 1
        assert res.summary["unmatched_functions"] == 0
        assert res.summary["fn_fail"] == 1   # it2 Fail 포함 → 함수 Fail


class TestR105SwUFnTcIdBlocks:
    """(b) SwUFn_ TC_ID 블록 매칭 — PV는 SwUTC_가 아닌 SwUFn_NNNN 직접 사용."""

    def test_scan_spec_blocks_swufn_and_lowercase_stub(self):
        """SwUFn_/소문자 'SwUfn_'(WIP 오타 실측) 모두 숫자 추출."""
        spec = _make_spec_bytes([
            ("SwUFn_0101", "main", [("a", "b")]),
            ("SwUfn_1361", "stub", []),   # WIP 실측 stub — iteration 0건
        ])
        ws = openpyxl.load_workbook(io.BytesIO(spec))["2.SW Unit Test Spec"]
        blocks = _scan_spec_blocks(ws)
        assert [b["num"] for b in blocks] == ["0101", "1361"]
        assert blocks[1]["iter_rows"] == []

    def test_build_sutr_swufn_tc_id_matches_dv_layout(self):
        """DV 레이아웃 + SwUFn_ 직접 TC_ID — 매칭·Pass stamp 동일 동작."""
        spec = _make_spec_bytes([("SwUFn_0101", "main", [("0x0", "0x1")])])
        sess = _make_session({"SwUFn_0101": [True]})
        res = build_sutr_from_spec(sess, _meta(), spec, function_asil_map={})
        assert res.ok
        assert res.summary["matched_functions"] == 1
        assert res.summary["unmatched_functions"] == 0
        ws = openpyxl.load_workbook(res.xlsm_io)[LOG_SHEET_NAME]
        assert ws.cell(5, COL_PASS_TOTAL).value == "Pass"
        assert ws.cell(6, COL_ACTUAL_START).value == "0x1"

    def test_scan_spec_blocks_pv_iter_column_8(self):
        """PV iteration index 열=8 — 고정 7 사용 시 iteration 행 누락 회귀 가드."""
        spec = _make_pv_spec_bytes([
            ("SwUFn_0101", "main", [("a", "b", "c"), ("d", "e", "f")]),
        ])
        ws = openpyxl.load_workbook(io.BytesIO(spec))["2.SW Unit Test Spec"]
        lo = _detect_spec_layout(ws)
        blocks = _scan_spec_blocks(ws, layout=lo)
        assert len(blocks) == 1
        assert len(blocks[0]["iter_rows"]) == 2
        assert ws.cell(blocks[0]["iter_rows"][0], lo.iter_index).value == 1
        # 레거시(layout 미전달 → DV iter=7)는 Generation Method(7) 세그먼트 병합
        # 때문에 비-anchor iteration 행을 놓침 — 동적화 필요성의 실측 근거 재현.
        legacy = _scan_spec_blocks(ws)
        assert len(legacy[0]["iter_rows"]) < 2


# ---------------------------------------------------------------------------
# 라운드 107 — UT201 FI spec 자동 산출 추출기 (확정 규칙 2026-06-12)
# ---------------------------------------------------------------------------

def _make_fi_spec_bytes(
    blocks: list[tuple[str, str, list[tuple[str, int]]]],
    *,
    merge_method: bool = True,
    method_header: str = "Test Method",
) -> bytes:
    """합성 spec — Test Method 세그먼트 단위 세로병합 재현 (DV 레이아웃 축소판).

    blocks: [(tc_id, unit, [(method, n_iters), ...]), ...]
        method 세그먼트는 실물처럼 iteration 행만 덮는 세로병합(anchor만 값).
        anchor 행 method는 빈 값 (DV/PV 실측 — '' 570/1,024건).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2.SW Unit Test Spec"
    ws.cell(1, 1).value = "Software Unit Test"
    ws.cell(3, 2).value = "Test Case"
    ws.cell(3, 7).value = "Input"
    ws.cell(3, 58).value = "Expected Result"
    ws.cell(3, 162).value = "Related ID"
    ws.cell(4, 2).value = "Index"
    ws.cell(4, 3).value = "TC_ID"
    ws.cell(4, 4).value = "Unit"
    ws.cell(4, 5).value = method_header
    ws.cell(4, 6).value = "Test Case Generation"
    ws.cell(4, 7).value = " "
    ws.cell(4, 8).value = "Inpt[0]"
    ws.cell(4, 58).value = "ExpR[0]"
    rr = 5
    idx = 1
    for tc_id, unit, segments in blocks:
        anchor = rr
        ws.cell(anchor, 2).value = idx
        ws.cell(anchor, 3).value = tc_id
        ws.cell(anchor, 4).value = unit
        ws.cell(anchor, 8).value = "in_var"
        ws.cell(anchor, 58).value = "exp_var"
        it_no = 0
        seg_start = anchor + 1
        for method, n_iters in segments:
            for _ in range(n_iters):
                it_no += 1
                ir = anchor + it_no
                ws.cell(ir, 7).value = it_no
                ws.cell(ir, 8).value = f"in{it_no}"
                ws.cell(ir, 58).value = f"exp{it_no}"
            seg_end = anchor + it_no
            ws.cell(seg_start, 5).value = method  # 세그먼트 anchor만 값
            if merge_method and seg_end > seg_start:
                ws.merge_cells(start_row=seg_start, end_row=seg_end,
                               start_column=5, end_column=5)
            seg_start = seg_end + 1
        rr = anchor + it_no + 1
        idx += 1
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class TestSpecFiExtraction:
    """확정 규칙(2026-06-12) — FI 수량 = Test Method 'FI' 포함 TC 블록 수.

    ground truth 실측 재현 검증 완료 (extract_spec_fi_stats, 2026-06-12):
    DV bk_SwUTS_v0.11_251126.xlsm → FI 블록 405 / FI iteration 1,598,
    PV wip_pv_SwUTS_v0.10_260608.xlsm → FI 블록 808 / FI iteration 3,229.
    여기서는 합성 spec으로 규칙 요소(병합 전개/정규화/실패 폴백)를 가드.
    """

    def test_fi_blocks_and_iterations_with_merged_segments(self):
        """FI 블록 카운트 + merged 세그먼트 전개로 iteration 블록별 귀속."""
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        spec = _make_fi_spec_bytes([
            ("SwUTC_0101", "fn_a", [("REQ", 2), ("FI", 3)]),
            ("SwUTC_0102", "fn_b", [("REQ", 2)]),          # FI 없음 — 제외
            ("SwUTC_0103", "fn_c", [("fi ", 1)]),          # 소문자+공백 정규화
            ("SwUTC_0104", "fn_d", []),                    # stub — method 없음
        ])
        warns: list[str] = []
        stats = extract_spec_fi_stats(
            spec, spec_filename="t.xlsm", out_warnings=warns,
        )
        assert stats is not None
        assert stats["blocks_total"] == 4
        assert stats["fi_block_total"] == 2
        assert stats["fi_block_keys"] == {"101", "103"}
        assert stats["fi_iteration_total"] == 4
        # 병합 전개 핵심 — FI 세그먼트 3행 중 anchor(첫 행)만 값 보유.
        # 전개 없으면 iteration 1개만 귀속된다 (PV 병합 1,865 range 실측 근거).
        assert stats["fi_iters_per_block"]["101"] == [3, 4, 5]
        assert stats["fi_iters_per_block"]["103"] == [1]
        assert stats["spec_filename"] == "t.xlsm"
        assert stats["method_col"] == 5  # DV
        assert not warns

    def test_unmerged_method_rows_also_counted(self):
        """DV 실측 — method 비병합(행마다 값) spec도 동일 산출."""
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        spec = _make_fi_spec_bytes(
            [("SwUTC_0101", "fn_a", [("FI", 1), ("FI", 1)])],
            merge_method=False,
        )
        stats = extract_spec_fi_stats(spec)
        assert stats is not None
        assert stats["fi_block_total"] == 1
        assert stats["fi_iters_per_block"]["101"] == [1, 2]

    def test_pv_layout_method_col6_detected(self):
        """PV — 'Safety Related' 삽입으로 Test Method=6, iter index=8."""
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        spec = _make_pv_spec_bytes([
            ("SwUFn_0101", "main", [("a", "b", "c"), ("d", "e", "f")]),
        ])
        # _make_pv_spec_bytes는 method(col6)에 'REQ' 기입 — FI로 치환.
        wb = openpyxl.load_workbook(io.BytesIO(spec))
        ws = wb["2.SW Unit Test Spec"]
        ws.cell(6, 6).value = "FI"
        ws.cell(7, 6).value = "FI"
        bio = io.BytesIO()
        wb.save(bio)
        stats = extract_spec_fi_stats(bio.getvalue())
        assert stats is not None
        assert stats["method_col"] == 6
        assert stats["fi_block_total"] == 1
        assert stats["fi_block_keys"] == {"101"}
        assert stats["fi_iters_per_block"]["101"] == [1, 2]

    def test_method_header_missing_returns_none_with_warning(self):
        """'Test Method' 헤더 미발견 — None + warning (노란 마킹 폴백)."""
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        spec = _make_fi_spec_bytes(
            [("SwUTC_0101", "fn_a", [("FI", 1)])],
            method_header="Method종류",  # 비표준 헤더
        )
        warns: list[str] = []
        assert extract_spec_fi_stats(spec, out_warnings=warns) is None
        assert any("'Test Method' 헤더 열 미발견" in w for w in warns)

    def test_no_spec_sheet_returns_none_with_warning(self):
        """spec 시트 미발견 — None + warning."""
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        wb = openpyxl.Workbook()
        wb.active.title = "다른 시트"
        bio = io.BytesIO()
        wb.save(bio)
        warns: list[str] = []
        assert extract_spec_fi_stats(bio.getvalue(), out_warnings=warns) is None
        assert any("spec 시트" in w for w in warns)

    def test_corrupt_bytes_returns_none_with_warning(self):
        """깨진 bytes — 예외 삼키고 None + warning (빌드 전체 실패 차단)."""
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        warns: list[str] = []
        assert extract_spec_fi_stats(b"not a zip", out_warnings=warns) is None
        assert any("자동 산출 실패" in w for w in warns)
        warns2: list[str] = []
        assert extract_spec_fi_stats(b"", out_warnings=warns2) is None
        assert any("비어있음" in w for w in warns2)

    def test_composite_method_segment_not_counted_exact_match_rule(self):
        """확정 규칙 — 'FI' strip+대소문자 무시 '정확 일치'만 집계.

        'REQ/FI' 류 복합 표기는 미집계. DV/PV 실측 method 분포는
        REQ/FI/빈값 3종뿐이라 현재 데이터엔 영향 없음 — 향후 spec 양식이
        복합 표기를 도입하면 규칙 재확정 필요 (concerns 명기). 본 테스트가
        깨지면 매칭 규칙이 부분 일치로 바뀐 것 — 사용자 재확정 없이 금지.
        """
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        spec = _make_fi_spec_bytes([
            ("SwUTC_0101", "fn_a", [("REQ/FI", 2)]),   # 복합 표기 — 미집계
            ("SwUTC_0102", "fn_b", [("FIT", 1)]),      # 부분 문자열 — 미집계
            ("SwUTC_0103", "fn_c", [("FI", 1)]),
        ])
        stats = extract_spec_fi_stats(spec)
        assert stats is not None
        assert stats["blocks_total"] == 3
        assert stats["fi_block_total"] == 1
        assert stats["fi_block_keys"] == {"103"}
        assert stats["fi_iteration_total"] == 1


# ---------------------------------------------------------------------------
# 라운드 107 — 로컬 실파일 ground truth 가드 (.codex_tmp 존재 시에만 — skipif)
# ---------------------------------------------------------------------------

_KJPDS02_REFS = os.path.join(
    os.path.dirname(__file__), "..", "..", ".codex_tmp", "kjpds02_refs",
)
_DV_BK_SPEC = os.path.join(_KJPDS02_REFS, "bk_SwUTS_v0.11_251126.xlsm")
_PV_WIP_SPEC = os.path.join(_KJPDS02_REFS, "wip_pv_SwUTS_v0.10_260608.xlsm")


class TestSpecFiExtractionRealFiles:
    """확정 규칙(2026-06-12) ground truth — 로컬 실파일 재현 가드.

    DV 감사본 수기 402는 전 가용 산출물(SwUTS 백업 2종/SwUDS/코드)로 재현
    불가한 stale 카운트로 판명 — 규칙 기준 실측은 405. 본 가드가 깨지면
    추출 규칙 회귀이므로 산출물 warning의 402 추적성(규칙·spec 파일명
    명기)도 함께 재검토할 것. 실파일은 git 미추적(.codex_tmp) — CI/타
    환경에서는 skip (합성 spec 가드 ``TestSpecFiExtraction`` 이 대체).
    """

    @pytest.mark.skipif(
        not os.path.exists(_DV_BK_SPEC),
        reason="DV 백업 spec 로컬 미존재 (.codex_tmp/kjpds02_refs)",
    )
    def test_dv_backup_spec_fi_405_blocks_1598_iterations(self):
        """DV bk_SwUTS_v0.11_251126.xlsm — FI 블록 405 / iteration 1,598."""
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        with open(_DV_BK_SPEC, "rb") as f:
            data = f.read()
        warns: list[str] = []
        stats = extract_spec_fi_stats(
            data, spec_filename="bk_SwUTS_v0.11_251126.xlsm",
            out_warnings=warns,
        )
        assert stats is not None
        assert stats["blocks_total"] == 570
        assert stats["fi_block_total"] == 405
        assert stats["fi_iteration_total"] == 1598
        assert stats["method_col"] == 5  # DV 레이아웃 — E열
        assert stats["layout_detected"] is True
        assert warns == []

    @pytest.mark.skipif(
        not os.path.exists(_PV_WIP_SPEC),
        reason="PV WIP spec 로컬 미존재 (.codex_tmp/kjpds02_refs)",
    )
    def test_pv_wip_spec_fi_808_blocks_3229_iterations(self):
        """PV wip_pv_SwUTS_v0.10_260608.xlsm — FI 블록 808 / iteration 3,229.

        7,899행×189열 + Test Method 세로병합 1,865 range — merged range
        전개 회귀 시 iteration 귀속이 무너져 본 수치가 깨진다 (~20s 로드).
        """
        from backend.services.swut_sutr_spec_builder import extract_spec_fi_stats
        with open(_PV_WIP_SPEC, "rb") as f:
            data = f.read()
        warns: list[str] = []
        stats = extract_spec_fi_stats(
            data, spec_filename="wip_pv_SwUTS_v0.10_260608.xlsm",
            out_warnings=warns,
        )
        assert stats is not None
        assert stats["blocks_total"] == 1014
        assert stats["fi_block_total"] == 808
        assert stats["fi_iteration_total"] == 3229
        assert stats["method_col"] == 6  # PV — 'Safety Related' 삽입 시프트, F열
        assert stats["layout_detected"] is True
        assert warns == []
