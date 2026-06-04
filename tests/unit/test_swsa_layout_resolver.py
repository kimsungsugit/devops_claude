"""SwSA layout_resolver 단위테스트 — 라벨 앵커 + 병합 라벨 흡수."""
from __future__ import annotations

import os

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as col

from backend.services.swsa_layout_resolver import detect_st_layout, find_value_target


def _coord(rc):
    return f"{col(rc[1])}{rc[0]}"


def _sheet_layout_a():
    wb = Workbook()
    ws = wb.active
    ws.title = "1.ST101"
    labels = {"분석차수": 4, "SW Ver.": 5, "Tester": 6, "Debugger": 7}
    for label, r in labels.items():
        ws.cell(r, 2).value = label  # col B
    return ws


def _sheet_merged_labels():
    """v0.10 ST201 형: 라벨 B4:C4 병합 → 값은 D4."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2.ST201"
    labels = {"분석차수": 4, "SW Ver.": 5, "Tester": 6, "Debugger": 7}
    for label, r in labels.items():
        ws.cell(r, 2).value = label
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)  # B:C
    return ws


def _sheet_layout_b():
    """ST1101 형: 라벨 col C → 값 col D."""
    wb = Workbook()
    ws = wb.active
    ws.title = "11.ST1101"
    labels = {"분석차수": 4, "SW Ver.": 5, "Tester": 6, "Debugger": 7}
    for label, r in labels.items():
        ws.cell(r, 3).value = label  # col C
    return ws


class TestLayoutDetection:
    def test_layout_a_value_in_c(self):
        lay = detect_st_layout(_sheet_layout_a())
        assert lay.missing == []
        assert lay.layout == "A"
        assert _coord(lay.test_info["analysis_round"]) == "C4"
        assert _coord(lay.test_info["debugger"]) == "C7"

    def test_merged_label_value_after_merge(self):
        # 병합 B4:C4 라벨 → 값은 D4 (라벨 덮어쓰기 버그 방지)
        lay = detect_st_layout(_sheet_merged_labels())
        assert lay.missing == []
        assert lay.layout == "A"  # 라벨 col B → A (값이 D여도)
        assert _coord(lay.test_info["analysis_round"]) == "D4"
        assert _coord(lay.test_info["sw_version"]) == "D5"

    def test_layout_b_shifted(self):
        lay = detect_st_layout(_sheet_layout_b())
        assert lay.layout == "B"
        assert _coord(lay.test_info["analysis_round"]) == "D4"

    def test_missing_label_recorded(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(4, 2).value = "분석차수"  # 나머지 라벨 없음
        lay = detect_st_layout(ws)
        assert "SW Ver." in lay.missing
        assert "Tester" in lay.missing
        assert not lay.ok

    def test_find_value_target_none_when_absent(self):
        wb = Workbook()
        assert find_value_target(wb.active, "없는라벨") is None


_TEMPLATE = os.path.join(
    os.path.dirname(__file__), "..", "..", ".codex_tmp", "swsa_samples",
    "TEMPLATE_(XXXX_SwSA) Software Static Analysis Report_v0.10_2XXXXX.xlsm",
)


@pytest.mark.skipif(not os.path.exists(_TEMPLATE), reason="실 템플릿 샘플 없음")
class TestRealTemplate:
    def test_v010_st101_and_st201(self):
        wb = load_workbook(_TEMPLATE, keep_vba=True)
        l101 = detect_st_layout(wb["1.ST101"])
        l201 = detect_st_layout(wb["2.ST201"])
        # ST101 라벨 비병합 → C; ST201 라벨 B:C 병합 → D
        assert _coord(l101.test_info["analysis_round"]) == "C4"
        assert _coord(l201.test_info["analysis_round"]) == "D4"
        assert l101.layout == "A" and l201.layout == "A"
        assert l101.missing == [] and l201.missing == []
