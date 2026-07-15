"""Phase 4 — SITS 콜체인 채굴: sits/extract-traceability가 spec 시트 C4의 'Interface : A -> B -> deep'
콜체인을 파싱해 chain_fns로 반환 → 깊은 callee(entry가 아닌 함수)가 그 SITS TC를 획득(g_drvin 0 해소).
"""
from __future__ import annotations

import io


def _build_sits_xlsm(rows_c2_c4):
    """합성 SITS — '3.SW Integration Test Spec' 시트에 (row, C2=TC, C4=Interface) 스탬프."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "3.SW Integration Test Spec"
    for r, c2, c4 in rows_c2_c4:
        if c2:
            ws.cell(r, 2).value = c2
        if c4:
            ws.cell(r, 4).value = c4
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_sits_chain_mining_attributes_deep_callee(monkeypatch):
    import backend.services.file_resolver as fr
    import backend.services.resolver_helpers as rh
    from backend.routers.jenkins import jenkins_sits_extract_traceability

    data = _build_sits_xlsm([
        (5, "SwITC_SwUFn_0112", None),
        (6, None, "Interface : main -> s_mid_layer -> g_deep_target"),
    ])

    class _R:
        def exists(self, _p):  # noqa: ARG002
            return True

        def read_bytes(self, _p):  # noqa: ARG002
            return data

    monkeypatch.setattr(fr, "get_resolver", lambda: _R())
    monkeypatch.setattr(rh, "enforce_resolver_access", lambda _p: None)

    d = jenkins_sits_extract_traceability({"path": "U:/sits.xlsm"})
    rows = d.get("vcast_rows", [])
    chain_row = next((r for r in rows if r.get("testcase") == "SwITC_SwUFn_0112" and r.get("chain_fns")), None)
    assert chain_row is not None, f"chain 행 없음: {rows}"
    # entry(main)뿐 아니라 중간·깊은 callee 전부 귀속(콜체인 채굴 핵심).
    assert "g_deep_target" in chain_row["chain_fns"]
    assert "s_mid_layer" in chain_row["chain_fns"]
    assert "main" in chain_row["chain_fns"]


def test_sits_chain_mining_ignores_non_identifier_tokens(monkeypatch):
    """체인 파싱은 C 식별자(3자+)만 채택 — 기호/짧은 토큰 오탐 방지(over-attribution 경계)."""
    import backend.services.file_resolver as fr
    import backend.services.resolver_helpers as rh
    from backend.routers.jenkins import jenkins_sits_extract_traceability

    data = _build_sits_xlsm([
        (5, "SwITC_SwUFn_0200", None),
        (6, None, "Interface : g_real_fn -> () -> 12 -> s_ok_fn"),
    ])

    class _R:
        def exists(self, _p):  # noqa: ARG002
            return True

        def read_bytes(self, _p):  # noqa: ARG002
            return data

    monkeypatch.setattr(fr, "get_resolver", lambda: _R())
    monkeypatch.setattr(rh, "enforce_resolver_access", lambda _p: None)

    d = jenkins_sits_extract_traceability({"path": "U:/sits.xlsm"})
    chain_row = next((r for r in d.get("vcast_rows", []) if r.get("chain_fns")), None)
    assert chain_row is not None
    cf = set(chain_row["chain_fns"])
    assert "g_real_fn" in cf and "s_ok_fn" in cf
    assert "()" not in cf and "12" not in cf   # 비식별자 토큰 배제
