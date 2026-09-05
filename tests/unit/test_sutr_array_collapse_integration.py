"""DC-1 통합 — _write_test_log이 다차원 full-box 배열 컬럼을 접어 stamp.

end-to-end: 다차원 배열은 단일 컬럼(헤더 base[d*d] + 값 format_values/_actual)으로,
스칼라 다수는 기존대로 10열 절단(경고)됨을 확인. 일반(배열 없음) 케이스 불변은
test_swut_aggregators.py 전체 통과로 보장.
"""
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.services.swut_input_adapter import (  # noqa: E402
    EnvironmentData,
    ExecutionRow,
    SwUTSession,
)
from backend.services.swut_sutr_aggregator import _write_test_log  # noqa: E402

INPUT_COL = 6   # F
EXPECTED_COL = 16  # P
ACTUAL_COL = 26  # Z


@dataclass
class _MockTCItem:
    input_data: dict
    expected_result: dict


def _ws():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Test Case ID"  # header → start_row=2, col=1(B는 2? — col=pos[1])
    return ws


def _session(input_data, expected_result, actual_result):
    ik = "SwUTC_SwUFn_0103.001"
    env = EnvironmentData(
        env_name="SWTE_01",
        component_name="Comp",
        test_cases={ik: [_MockTCItem(input_data=input_data, expected_result=expected_result)]},
        test_results={ik: ExecutionRow(tc_name=ik, passed=True, actual_result=actual_result)},
    )
    return SwUTSession(
        project_id="X", version="v", source_kind="log_folder", environments=[env]
    )


def _anchor_iter_rows(ws):
    """헤더 row 1 → anchor row 2, iteration row 3."""
    return 2, 3


class TestMergeGuardMerge06:
    """MERGE-06 — tc_row_step 오검출 방어 경고."""

    def test_warns_when_multirow_block_without_template_merges(self):
        # 빈 ws(템플릿 병합 0) + 다중행 블록(2 iteration → height 3) → 경고
        ik1, ik2 = "SwUTC_SwUFn_0103.001", "SwUTC_SwUFn_0103.002"
        env = EnvironmentData(
            env_name="SWTE_01", component_name="Comp",
            test_cases={
                ik1: [_MockTCItem(input_data={"a": "1"}, expected_result={"o": "1"})],
                ik2: [_MockTCItem(input_data={"a": "2"}, expected_result={"o": "2"})],
            },
            test_results={
                ik1: ExecutionRow(tc_name=ik1, passed=True, actual_result={"o": ("1", "1")}),
                ik2: ExecutionRow(tc_name=ik2, passed=True, actual_result={"o": ("2", "2")}),
            },
        )
        session = SwUTSession(
            project_id="X", version="v", source_kind="log_folder", environments=[env]
        )
        ws = _ws()
        warns: list[str] = []
        _write_test_log(ws, session, out_warnings=warns)
        assert any("[merge_guard]" in w for w in warns)


class TestSutrArrayCollapseIntegration:
    def test_multidim_array_collapses_to_single_column(self):
        # buffer[4][3] = 12 요소(>10) → 접기 없으면 12열(2 절단). 접기 → 1열.
        arr = {f"buffer[{i}][{j}]": str(i * 3 + j) for i in range(4) for j in range(3)}
        session = _session(arr, {"out": "1"}, {"out": ("1", "1")})
        ws = _ws()
        warns: list[str] = []
        n = _write_test_log(ws, session, out_warnings=warns)
        assert n >= 1
        anchor, it = _anchor_iter_rows(ws)
        # anchor 행 INPUT 헤더: 단일 접힌 헤더, 다음 열은 비어야(12열로 안 펼쳐짐)
        assert ws.cell(anchor, INPUT_COL).value == "buffer[4*3]"
        assert ws.cell(anchor, INPUT_COL + 1).value in (None, "")
        # iteration 행 값: format_values (비균일 → 바깥 차원별 줄바꿈)
        cell = ws.cell(it, INPUT_COL).value
        assert cell is not None and cell.startswith("[0]: 0, 1, 2")
        assert "[3]: 9, 10, 11" in cell
        # 접혀서 1열 ≤ 10 → 절단 경고 없음 ([diag]의 'truncated_fn'과 구분 위해 '[truncate]')
        assert not any("[truncate]" in w for w in warns)

    def test_scalars_over_limit_still_truncate(self):
        # 스칼라 12개(접기 대상 아님) → 10열 절단 + 경고 (기존 동작 보존)
        scal = {f"var{i:02d}": str(i) for i in range(12)}
        session = _session(scal, {"out": "1"}, {"out": ("1", "1")})
        ws = _ws()
        warns: list[str] = []
        _write_test_log(ws, session, out_warnings=warns)
        anchor, _ = _anchor_iter_rows(ws)
        # input 섹션 F~O(6~15) 10열 채움 — 10번째(O=15) stamp 확인. 11/12번째는 드롭.
        assert ws.cell(anchor, INPUT_COL + 9).value is not None  # O열=10번째 input
        # W2: 절단 경고가 raw count + 누락 컬럼명(var10/var11) 표기
        assert any(
            "[truncate]" in w and "raw 12/" in w and "누락 컬럼" in w
            and ("var10" in w or "var11" in w)
            for w in warns
        )

    def test_actual_collapsed_shows_ng_summary(self):
        # 2x2 배열, actual 일부 불일치 → Actual 셀 NG(k/N)
        arr_in = {f"m[{i}][{j}]": "0" for i in range(2) for j in range(2)}
        arr_exp = dict(arr_in)
        # actual_result: (actual, expected) — m[0][1]만 불일치
        arr_act = {
            "m[0][0]": ("0", "0"), "m[0][1]": ("9", "0"),
            "m[1][0]": ("0", "0"), "m[1][1]": ("0", "0"),
        }
        session = _session(arr_in, arr_exp, arr_act)
        ws = _ws()
        _write_test_log(ws, session)
        anchor, it = _anchor_iter_rows(ws)
        assert ws.cell(anchor, ACTUAL_COL).value == "m[2*2]"
        act_cell = ws.cell(it, ACTUAL_COL).value
        assert act_cell is not None and act_cell.startswith("NG (1/4)")
        assert "[0][1]" in act_cell

    def test_single_dim_array_collapses_collapse_all(self):
        # 실 데이터 케이스: 단일차원 배열 lin_pFrameBuf[0..11] = 12열 → collapse_all로 1열
        arr = {f"lin_pFrameBuf[{i}]": str(i) for i in range(12)}
        session = _session(arr, {"out": "1"}, {"out": ("1", "1")})
        ws = _ws()
        warns: list[str] = []
        n = _write_test_log(ws, session, out_warnings=warns)
        assert n >= 1
        anchor, it = _anchor_iter_rows(ws)
        # 단일 접힌 헤더 'lin_pFrameBuf[12]', 다음 열 비어야(12열로 안 펼쳐짐)
        assert ws.cell(anchor, INPUT_COL).value == "lin_pFrameBuf[12]"
        assert ws.cell(anchor, INPUT_COL + 1).value in (None, "")
        # 값: 콤마 구분 compact
        cell = ws.cell(it, INPUT_COL).value
        assert cell is not None and cell.startswith("[0]: 0, [1]: 1")
        # 12→1열로 절단 경고 없음
        assert not any("[truncate]" in w for w in warns)

    def test_no_array_unchanged_passthrough(self):
        # 스칼라 3개 → 3열 그대로 (접기 no-op)
        session = _session(
            {"a": "1", "b": "2", "c": "3"}, {"out": "1"}, {"out": ("1", "1")}
        )
        ws = _ws()
        warns: list[str] = []
        _write_test_log(ws, session, out_warnings=warns)
        anchor, it = _anchor_iter_rows(ws)
        headers = [ws.cell(anchor, INPUT_COL + i).value for i in range(3)]
        assert headers == ["a", "b", "c"]
        assert ws.cell(it, INPUT_COL).value == "1"  # 단일 값 그대로
        assert not any("[truncate]" in w for w in warns)
