"""vcast_excel_generator 세로 셀 병합 회귀 테스트 (MERGE-01/02).

C# TResultParser(UCTestCaseList.cs:1147-1182) 포팅 — /api/vcast/generate-excel 산출물에서
연속 동일값(TC ID/Unit Name)을 세로 병합한다. 기존 가로 병합(헤더/P-F) 보존, 과병합 방지,
빈/단일 행 무크래시, 마지막 행까지 확장된 run의 하단 테두리 보존을 고정한다.
"""
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from backend.services.vcast_excel_generator import generate_testcase_excel  # noqa: E402
from backend.services.vcast_parser import (  # noqa: E402
    TCBank,
    VCastHeader,
)
from backend.services.vcast_parser import TestCaseItem as TCItem  # noqa: E402
from backend.services.vcast_parser import TestResultItem as TRItem  # noqa: E402


def _h(unit, tc, idx):
    return VCastHeader(component_name="Comp", unit_name=unit, test_case_name=tc,
                       test_case_index=str(idx), filename="f.html")


def _tc(unit, tc, idx):
    it = TCItem(header=_h(unit, tc, idx))
    it.input_data = {"a": "1"}
    it.expected_result = {"r": "1"}
    return it


def _tr(unit, tc, idx, passed):
    it = TRItem(header=_h(unit, tc, idx))
    it.passed = passed
    it.actual_result = {"r": ("1", "1")}
    return it


def _ranges(path):
    wb = load_workbook(path)
    ws = wb.active
    return ws, {str(r) for r in ws.merged_cells.ranges}


class TestVcastVerticalMerge:
    def test_testcase_merges_tcid_and_unit(self, tmp_path):
        bank = TCBank(component_name="Comp")
        bank.input_names = ["a"]
        bank.exp_result_names = ["r"]
        bank.test_cases = {
            "1": [_tc("UnitA", "TC1", 1), _tc("UnitA", "TC1", 2)],  # 행7-8: TC1, UnitA
            "2": [_tc("UnitA", "TC2", 3)],                          # 행9:  TC2, UnitA
            "3": [_tc("UnitB", "TC3", 4)],                          # 행10: TC3, UnitB
        }
        out = tmp_path / "tc.xlsx"
        assert generate_testcase_excel(bank, out, mode="TestCase")
        ws, ranges = _ranges(out)
        assert "C7:C8" in ranges       # TC ID 'TC1' 2행 병합
        assert "D7:D9" in ranges       # Unit 'UnitA' 3행 병합
        assert "D7:D10" not in ranges  # UnitB 흡수(과병합) 없음
        assert ws.cell(7, 3).value == "TC1" and ws.cell(8, 3).value is None

    def test_testresult_merges_tcid_preserves_passfail_hmerge(self, tmp_path):
        bank = TCBank(component_name="Comp")
        bank.act_result_names = ["r"]
        bank.test_results = {
            "1": [_tr("UnitA", "TC1", 1, True), _tr("UnitA", "TC1", 2, True)],  # 행7-8
            "2": [_tr("UnitA", "TC2", 3, False)],                              # 행9
        }
        out = tmp_path / "tr.xlsx"
        assert generate_testcase_excel(bank, out, mode="TestResult")
        _, ranges = _ranges(out)
        assert "C7:C8" in ranges                              # TC ID 세로 병합
        assert {"E7:F7", "E8:F8", "E9:F9"}.issubset(ranges)  # P/F 행별 가로병합 보존

    def test_non_consecutive_duplicates_not_merged(self, tmp_path):
        bank = TCBank(component_name="Comp")
        bank.input_names = ["a"]
        bank.exp_result_names = ["r"]
        bank.test_cases = {  # TC1, TC2, TC1 — 비연속이라 병합 금지
            "1": [_tc("U", "TC1", 1)],
            "2": [_tc("U", "TC2", 2)],
            "3": [_tc("U", "TC1", 3)],
        }
        out = tmp_path / "nc.xlsx"
        assert generate_testcase_excel(bank, out, mode="TestCase")
        _, ranges = _ranges(out)
        assert not any(r.startswith("C") and ":" in r for r in ranges)  # TC ID 병합 없음
        assert "D7:D9" in ranges  # Unit 'U'는 연속 3행 → 병합

    def test_last_row_run_keeps_bottom_border(self, tmp_path):
        """W1 회귀: run이 마지막 데이터 행까지 확장돼도 하단 thick 테두리 보존."""
        bank = TCBank(component_name="Comp")
        bank.input_names = ["a"]
        bank.exp_result_names = ["r"]
        bank.test_cases = {
            "1": [_tc("UnitA", "TC1", 1)],  # 행7
            "2": [_tc("UnitZ", "TC2", 2)],  # 행8
            "3": [_tc("UnitZ", "TC3", 3)],  # 행9 → UnitZ 행8-9 병합, last_data_row(9) 도달
        }
        out = tmp_path / "lr.xlsx"
        assert generate_testcase_excel(bank, out, mode="TestCase")
        ws, ranges = _ranges(out)
        assert "D8:D9" in ranges  # run이 마지막 데이터 행(9)까지 확장
        # 미병합 컬럼(B)과 병합 slave(D9, last_data_row)의 하단 테두리가 동일(thick)해야
        last = 9
        b_bottom = ws.cell(last, 2).border.bottom.style
        d_bottom = ws.cell(last, 4).border.bottom.style
        assert b_bottom == "thick" and d_bottom == "thick"

    @pytest.mark.parametrize("mode", ["TestCase", "TestResult"])
    def test_empty_and_single_no_crash(self, tmp_path, mode):
        bank = TCBank(component_name="Empty")
        bank.input_names = ["a"]
        bank.exp_result_names = ["r"]
        bank.act_result_names = ["r"]
        out = tmp_path / f"empty_{mode}.xlsx"
        assert generate_testcase_excel(bank, out, mode=mode)  # 0행 무크래시
        _, ranges = _ranges(out)
        # 데이터 0행 → 세로 병합 미발행(C/D run 없음)
        assert not any(r.startswith(("C", "D")) and r[1:2].isdigit() and ":" in r
                       and not r.startswith(("C6", "D6")) for r in ranges)
