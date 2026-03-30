# workflow/doc_compare.py
"""Document version comparison - diff UDS/STS/SUTS documents.

Supports comparing:
- Two DOCX files (paragraph-level diff)
- Two XLSM files (cell-level diff)
- UDS JSON payloads (function-detail diff)
"""

from __future__ import annotations

import difflib
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


def diff_text_lines(
    old_lines: List[str],
    new_lines: List[str],
    *,
    context: int = 3,
) -> List[Dict[str, Any]]:
    """Compute unified diff between two text line lists."""
    changes: List[Dict[str, Any]] = []
    matcher = difflib.SequenceMatcher(None, old_lines, new_lines)

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        changes.append({
            "type": tag,
            "old_start": i1 + 1,
            "old_end": i2,
            "new_start": j1 + 1,
            "new_end": j2,
            "old_lines": old_lines[i1:i2],
            "new_lines": new_lines[j1:j2],
        })

    return changes


def diff_docx(old_path: str, new_path: str) -> Dict[str, Any]:
    """Compare two DOCX files paragraph by paragraph."""
    try:
        from docx import Document
    except ImportError:
        return {"error": "python-docx not installed"}

    def _extract_paras(path: str) -> List[str]:
        doc = Document(path)
        return [p.text.strip() for p in doc.paragraphs if p.text.strip()]

    old_paras = _extract_paras(old_path)
    new_paras = _extract_paras(new_path)

    changes = diff_text_lines(old_paras, new_paras)

    return {
        "old_file": Path(old_path).name,
        "new_file": Path(new_path).name,
        "old_paragraphs": len(old_paras),
        "new_paragraphs": len(new_paras),
        "changes": changes,
        "change_count": len(changes),
        "similarity": difflib.SequenceMatcher(None, old_paras, new_paras).ratio(),
    }


def diff_xlsm(
    old_path: str,
    new_path: str,
    *,
    sheet_name: Optional[str] = None,
    max_rows: int = 500,
    max_cols: int = 30,
) -> Dict[str, Any]:
    """Compare two XLSM/XLSX files cell by cell."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed"}

    wb_old = load_workbook(old_path, read_only=True, data_only=True)
    wb_new = load_workbook(new_path, read_only=True, data_only=True)

    sheets_to_compare = [sheet_name] if sheet_name else list(
        set(wb_old.sheetnames) & set(wb_new.sheetnames)
    )

    result_sheets: List[Dict[str, Any]] = []
    total_changes = 0

    for sn in sheets_to_compare:
        if sn not in wb_old.sheetnames or sn not in wb_new.sheetnames:
            continue
        ws_old = wb_old[sn]
        ws_new = wb_new[sn]

        changes: List[Dict[str, Any]] = []
        mr = min(max(ws_old.max_row or 0, ws_new.max_row or 0), max_rows)
        mc = min(max(ws_old.max_column or 0, ws_new.max_column or 0), max_cols)

        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                v_old = ws_old.cell(row=r, column=c).value
                v_new = ws_new.cell(row=r, column=c).value
                if v_old != v_new:
                    changes.append({
                        "row": r,
                        "col": c,
                        "old_value": str(v_old) if v_old is not None else "",
                        "new_value": str(v_new) if v_new is not None else "",
                    })

        total_changes += len(changes)
        result_sheets.append({
            "sheet": sn,
            "changes": changes[:200],
            "change_count": len(changes),
        })

    wb_old.close()
    wb_new.close()

    return {
        "old_file": Path(old_path).name,
        "new_file": Path(new_path).name,
        "sheets": result_sheets,
        "total_changes": total_changes,
    }


def diff_uds_payloads(
    old_payload: Dict[str, Any],
    new_payload: Dict[str, Any],
) -> Dict[str, Any]:
    """Compare two UDS JSON payloads (function_details level)."""
    old_details = old_payload.get("function_details", {})
    new_details = new_payload.get("function_details", {})

    old_ids = set(old_details.keys())
    new_ids = set(new_details.keys())

    added = sorted(new_ids - old_ids)
    removed = sorted(old_ids - new_ids)
    common = old_ids & new_ids

    modified: List[Dict[str, Any]] = []
    compare_fields = [
        "description", "prototype", "inputs", "outputs",
        "called", "calling", "precondition", "asil",
    ]

    for fid in sorted(common):
        old_info = old_details[fid]
        new_info = new_details[fid]
        field_changes = {}
        for field in compare_fields:
            ov = old_info.get(field)
            nv = new_info.get(field)
            if str(ov) != str(nv):
                field_changes[field] = {"old": ov, "new": nv}
        if field_changes:
            modified.append({
                "fid": fid,
                "name": new_info.get("name", old_info.get("name", "")),
                "changes": field_changes,
            })

    return {
        "added_functions": added,
        "removed_functions": removed,
        "modified_functions": modified,
        "added_count": len(added),
        "removed_count": len(removed),
        "modified_count": len(modified),
        "total_old": len(old_ids),
        "total_new": len(new_ids),
    }
