"""QAC router."""
from __future__ import annotations

import logging
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse

from backend.helpers.jenkins import _jenkins_sts_dir, _jenkins_suts_dir, _resolve_cached_build_root
from backend.services.jenkins_helpers import _job_slug
from backend.services.output_paths import drop_empty_reservation, reserve_unique_path
from backend.services.paths import safe_resolve_under
from backend.services.qac_excel_generator import generate_qac_excel
from backend.services.qac_parser import QACDataManager, parse_qac_report
from backend.services.report_parsers import _normalize_prqa_path

repo_root = Path(__file__).resolve().parents[2]
router = APIRouter()
_logger = logging.getLogger("devops_api")


def _artifact_kind(path: Path) -> str:
    name = path.name.upper()
    if "_HMR_" in name or "HMR" in name:
        return "hmr"
    if "_RCR_" in name or "RCR" in name:
        return "rcr"
    if "_CRR_" in name or "CRR" in name:
        return "crr"
    return "other"


def _sniff_old_version(path: Path) -> Optional[bool]:
    if path.suffix.lower() != ".html" or not path.exists():
        return None
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")[:4000]
    except Exception:
        return None
    if "PRQA HIS Metrics Report" in text:
        return True
    if "Helix QAC HIS Metrics Report" in text:
        return False
    return None


def _serialize_qac_manager(
    qac_manager: QACDataManager,
    *,
    project_root: Optional[Path] = None,
    job_slug: Optional[str] = None,
    artifact_path: str = "",
    artifact_rel_path: str = "",
    old_version: Optional[bool] = None,
) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "ok": True,
        "item_count": len(qac_manager.list_result),
        "items": [],
        "artifact_path": artifact_path,
        "artifact_rel_path": artifact_rel_path,
    }
    if old_version is not None:
        result["old_version"] = bool(old_version)
    for item in qac_manager.list_result:
        raw_file_name = str(item.file_name or "").strip()
        normalized_path = _normalize_prqa_path(raw_file_name, project_root, job_slug)
        values: Dict[str, Dict[str, Any]] = {}
        for matrix in QACDataManager.get_matrix_list():
            value = item.get_matrix_value(matrix)
            values[matrix.name] = {
                "value": value,
                "warning_level": qac_manager.check_warning_level(matrix, value),
            }
        result["items"].append(
            {
                "function_name": item.function_name,
                "file_name": raw_file_name,
                "normalized_path": normalized_path,
                "values": values,
            }
        )
    totals: Dict[str, Dict[str, int]] = {}
    for matrix in QACDataManager.get_matrix_list():
        spec = qac_manager.dic_spec_over_count.get(matrix)
        if not spec:
            continue
        totals[matrix.name] = {
            "level_1": spec.list_spec[0] if len(spec.list_spec) > 0 else 0,
            "level_2": spec.list_spec[1] if len(spec.list_spec) > 1 else 0,
            "level_3": spec.list_spec[2] if len(spec.list_spec) > 2 else 0,
        }
    result["totals"] = totals
    return result


def _parse_qac_path(
    path: Path,
    *,
    old_version: bool,
    project_root: Optional[Path] = None,
    job_slug: Optional[str] = None,
    artifact_rel_path: str = "",
) -> Dict[str, Any]:
    qac_manager = parse_qac_report(path, old_version)
    return _serialize_qac_manager(
        qac_manager,
        project_root=project_root,
        job_slug=job_slug,
        artifact_path=str(path),
        artifact_rel_path=artifact_rel_path,
        old_version=old_version,
    )


def _collect_jenkins_qac_artifacts(build_root: Path) -> List[Dict[str, Any]]:
    patterns = (
        "*_HMR_*.html",
        "*HMR*.html",
        "*_CRR_*.html",
        "*CRR*.html",
        "*_RCR_*.html",
        "*RCR*.html",
        "*_HMR_*.xlsx",
        "*HMR*.xlsx",
    )
    seen: set[str] = set()
    items: List[Dict[str, Any]] = []
    for pattern in patterns:
        for path in build_root.rglob(pattern):
            if not path.is_file():
                continue
            rel_path = str(path.relative_to(build_root)).replace("\\", "/")
            key = rel_path.lower()
            if key in seen:
                continue
            seen.add(key)
            kind = _artifact_kind(path)
            old_version = _sniff_old_version(path)
            can_parse = path.suffix.lower() == ".html" and kind == "hmr"
            priority = 0
            rel_upper = rel_path.upper()
            if "/REPORT/PRQA/" in f"/{rel_upper}/":
                priority += 10
            if rel_path.count("/") == 0:
                priority += 20
            if kind == "hmr":
                priority += 5
            items.append(
                {
                    "name": path.name,
                    "rel_path": rel_path,
                    "kind": kind,
                    "extension": path.suffix.lower(),
                    "can_parse": can_parse,
                    "old_version": old_version,
                    "size": path.stat().st_size,
                    "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                    "_priority": priority,
                }
            )
    items.sort(key=lambda row: (-int(row.get("_priority") or 0), row["rel_path"].lower()))
    for row in items:
        row.pop("_priority", None)
    return items


def _resolve_cached_artifact(build_root: Path, rel_path: str) -> Path:
    if not rel_path:
        raise HTTPException(status_code=400, detail="rel_path required")
    target = (build_root / rel_path).resolve()
    try:
        target.relative_to(build_root)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid rel_path") from exc
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="artifact not found")
    return target


def _norm_symbol(value: str) -> str:
    import re
    return re.sub(r"[^a-z0-9_]", "", str(value or "").strip().lower())


def _pick_latest_excel(directory: Path) -> Optional[Path]:
    files = sorted(directory.glob("*.xls*"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def _pick_recent_excels(directory: Path, limit: int = 2) -> List[Path]:
    return sorted(directory.glob("*.xls*"), key=lambda p: p.stat().st_mtime, reverse=True)[: max(1, int(limit or 1))]


def _scan_excel_for_function(file_path: Path, function_name: str, max_hits: int = 20) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed") from exc
    target = _norm_symbol(function_name)
    if not target:
        return {"filename": file_path.name, "match_count": 0, "matches": []}
    wb = load_workbook(str(file_path), read_only=True, data_only=True, keep_vba=False)
    matches: List[Dict[str, Any]] = []
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = ["" if v is None else str(v).strip() for v in row]
                if not any(values):
                    continue
                matched = False
                for value in values:
                    if target and target == _norm_symbol(value):
                        matched = True
                        break
                if not matched:
                    continue
                matches.append(
                    {
                        "sheet": sname,
                        "row_index": row_idx,
                        "cells": values[:12],
                    }
                )
                if len(matches) >= max_hits:
                    break
            if len(matches) >= max_hits:
                break
    finally:
        wb.close()
    return {
        "filename": file_path.name,
        "output_path": str(file_path),
        "match_count": len(matches),
        "matches": matches,
    }


def _write_qac_impact_report(
    *,
    job_url: str,
    function_name: str,
    payload: Dict[str, Any],
) -> Optional[Path]:
    try:
        out_dir = repo_root / "reports" / "qac_impact"
        out_dir.mkdir(parents=True, exist_ok=True)
        stem = _job_slug(job_url)
        safe_fn = "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in function_name).strip("_") or "function"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        # ⚠ `reports/qac_impact` 는 전 사용자 공유다. 키가 job+함수뿐이라 같은 함수를
        #   두 사람이 같은 초에 분석하면 한쪽 리포트가 사라진다.
        from backend.services.output_paths import reserve_unique_path
        out_path = reserve_unique_path(out_dir / f"qac_impact_{stem}_{safe_fn}_{ts}.md")
        summary = payload.get("summary") or {}
        compare = payload.get("compare") or {}
        lines = [
            f"# QAC Impact Report: {function_name}",
            "",
            f"- Job: `{job_url}`",
            f"- Generated: `{datetime.now().isoformat(timespec='seconds')}`",
            *([f"- ⚠ {summary.get('message', 'STS/SUTS 캐시 없음 — 영향도 판정 불가')}"]
              if summary.get("status") in ("insufficient_data", "partial_data") else []),
            f"- STS impacted: `{summary.get('sts_impacted', 0)}`",
            f"- SUTS impacted: `{summary.get('suts_impacted', 0)}`",
            f"- STS delta: `{summary.get('sts_delta', 0)}`",
            f"- SUTS delta: `{summary.get('suts_delta', 0)}`",
            "",
        ]
        for kind in ("sts", "suts"):
            cur = ((compare.get(kind) or {}).get("current") or {})
            prev = ((compare.get(kind) or {}).get("previous") or {})
            delta = (compare.get(kind) or {}).get("delta", 0)
            lines.extend(
                [
                    f"## {kind.upper()}",
                    "",
                    f"- Current: `{cur.get('filename', '')}` / `{cur.get('match_count', 0)}`",
                    f"- Previous: `{prev.get('filename', '')}` / `{prev.get('match_count', 0)}`",
                    f"- Delta: `{delta}`",
                    "",
                ]
            )
            matches = cur.get("matches") or []
            if matches:
                lines.append("| Sheet | Row | Cells |")
                lines.append("| --- | ---: | --- |")
                for row in matches[:8]:
                    cells = " | ".join(str(v) for v in (row.get("cells") or []) if str(v).strip())
                    cells = cells.replace("|", "\\|")
                    lines.append(f"| {row.get('sheet', '')} | {row.get('row_index', '')} | {cells} |")
                lines.append("")
        out_path.write_text("\n".join(lines), encoding="utf-8")
        return out_path
    except Exception as exc:
        _logger.warning("qac impact report write skipped: %s", exc)
        return None


@router.get("/api/qac/jenkins-artifacts")
def qac_jenkins_artifacts(
    job_url: str,
    cache_root: str = "",
    build_selector: str = "lastSuccessfulBuild",
) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(job_url, cache_root, build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    return {"ok": True, "build_root": str(build_root), "items": _collect_jenkins_qac_artifacts(build_root)}


@router.get("/api/qac/jenkins-parse")
def qac_jenkins_parse(
    job_url: str,
    cache_root: str = "",
    build_selector: str = "lastSuccessfulBuild",
    rel_path: str = "",
    source_root: str = "",
    old_version: Optional[bool] = Query(None),
) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(job_url, cache_root, build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    target = _resolve_cached_artifact(build_root, rel_path)
    inferred_old = _sniff_old_version(target)
    use_old = old_version if old_version is not None else (True if inferred_old is None else inferred_old)
    project_root = Path(source_root).expanduser().resolve() if source_root else None
    return _parse_qac_path(
        target,
        old_version=bool(use_old),
        project_root=project_root,
        job_slug=_job_slug(job_url),
        artifact_rel_path=str(rel_path).replace("\\", "/"),
    )


@router.get("/api/qac/jenkins-impact")
def qac_jenkins_impact(
    job_url: str,
    cache_root: str = "",
    build_selector: str = "lastSuccessfulBuild",
    function_name: str = "",
) -> Dict[str, Any]:
    fn_name = str(function_name or "").strip()
    if not fn_name:
        raise HTTPException(status_code=400, detail="function_name required")
    sts_dir = _jenkins_sts_dir(cache_root)
    suts_dir = _jenkins_suts_dir(cache_root)
    sts_files = _pick_recent_excels(sts_dir, 2)
    suts_files = _pick_recent_excels(suts_dir, 2)
    sts_current = _scan_excel_for_function(sts_files[0], fn_name) if len(sts_files) >= 1 else {"filename": "", "match_count": 0, "matches": []}
    sts_previous = _scan_excel_for_function(sts_files[1], fn_name) if len(sts_files) >= 2 else {"filename": "", "match_count": 0, "matches": []}
    suts_current = _scan_excel_for_function(suts_files[0], fn_name) if len(suts_files) >= 1 else {"filename": "", "match_count": 0, "matches": []}
    suts_previous = _scan_excel_for_function(suts_files[1], fn_name) if len(suts_files) >= 2 else {"filename": "", "match_count": 0, "matches": []}
    # B4/W4 — 캐시(STS/SUTS QAC excel) 부재는 "영향 없음"이 아니라 **판정 불가**다. 캐시
    # 미비(cache_root 오설정·미생성)만으로 no-impact 로 흘리면 안전변경이 의존 STS/SUTS
    # 재검증을 조용히 우회한다. 소스별로 판정한다(W4 — 예전엔 OR 라 한 소스만 있어도
    # data_available=True 라, 나머지 소스 캐시 누락이 has_any_impact 확정에서 묻혔다):
    #   · 어느 소스든 영향 확인 → True
    #   · 두 소스 다 스캔했고 영향 0 → False (확정 무영향)
    #   · 일부 소스 캐시 부재 + 나머지 영향 0 → None (그 소스는 미판정이라 무영향 단정 불가)
    sts_ok, suts_ok = bool(sts_files), bool(suts_files)
    sts_impacted = sts_ok and (sts_current.get("match_count") or 0) > 0
    suts_impacted = suts_ok and (suts_current.get("match_count") or 0) > 0
    missing_sources = [s for s, ok in (("STS", sts_ok), ("SUTS", suts_ok)) if not ok]
    if sts_impacted or suts_impacted:
        has_any_impact: bool | None = True
    elif not missing_sources:
        has_any_impact = False
    else:
        has_any_impact = None
    data_available = sts_ok or suts_ok
    payload = {
        "ok": True,
        "function_name": fn_name,
        "build_label": build_selector,
        "sts": sts_current,
        "suts": suts_current,
        "compare": {
            "sts": {
                "current": sts_current,
                "previous": sts_previous,
                "delta": int(sts_current.get("match_count") or 0) - int(sts_previous.get("match_count") or 0),
            },
            "suts": {
                "current": suts_current,
                "previous": suts_previous,
                "delta": int(suts_current.get("match_count") or 0) - int(suts_previous.get("match_count") or 0),
            },
        },
        "summary": {
            "sts_impacted": int(sts_current.get("match_count") or 0),
            "suts_impacted": int(suts_current.get("match_count") or 0),
            "sts_delta": int(sts_current.get("match_count") or 0) - int(sts_previous.get("match_count") or 0),
            "suts_delta": int(suts_current.get("match_count") or 0) - int(suts_previous.get("match_count") or 0),
            # 소스별 판정(W4): None=판정 불가(일부 캐시 부재), True/False=확정.
            "has_any_impact": has_any_impact,
            "data_available": data_available,
            "missing_sources": missing_sources,
            "sts_files_scanned": len(sts_files),
            "suts_files_scanned": len(suts_files),
        },
    }
    if missing_sources:
        # 둘 다 부재=데이터 없음, 하나만 부재=부분 데이터(그 소스 미판정).
        payload["summary"]["status"] = (
            "insufficient_data" if len(missing_sources) == 2 else "partial_data"
        )
        payload["summary"]["message"] = (
            f"{'/'.join(missing_sources)} QAC 캐시가 없어 해당 소스 영향도를 판정할 수 없습니다"
            + (" (cache_root 설정 또는 STS/SUTS 생성 후 재시도)."
               if len(missing_sources) == 2 else " — 나머지 소스만으로는 무영향을 단정할 수 없음.")
        )
    report_path = _write_qac_impact_report(job_url=job_url, function_name=fn_name, payload=payload)
    payload["impact_report_path"] = str(report_path) if report_path and report_path.exists() else ""
    return payload


@router.get("/api/qac/jenkins-excel")
def qac_jenkins_generate_excel(
    job_url: str,
    cache_root: str = "",
    build_selector: str = "lastSuccessfulBuild",
    rel_path: str = "",
    old_version: Optional[bool] = Query(None),
) -> FileResponse:
    build_root = _resolve_cached_build_root(job_url, cache_root, build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    target = _resolve_cached_artifact(build_root, rel_path)
    inferred_old = _sniff_old_version(target)
    use_old = old_version if old_version is not None else (True if inferred_old is None else inferred_old)
    output_dir = repo_root / "reports" / "qac_excel"
    output_dir.mkdir(parents=True, exist_ok=True)
    # ⚠ `reports/qac_excel` 은 전 사용자 공유다 — 키가 job+초뿐이라 같은 job 을 두 사람이
    #   같은 초에 뽑으면 한쪽이 남의 리포트를 받는다.
    output_path = reserve_unique_path(
        output_dir / f"qac_{_job_slug(job_url)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    filename = output_path.name
    qac_manager = parse_qac_report(target, bool(use_old))
    if not generate_qac_excel(qac_manager, output_path):
        # 선점만 남은 0바이트 파일이 `/api/qac/reports` 목록에 뜨지 않게.
        drop_empty_reservation(output_path)
        raise HTTPException(status_code=500, detail="Excel generation failed")
    return FileResponse(
        str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/api/qac/parse")
async def qac_parse(
    file: UploadFile = File(...),
    old_version: bool = Query(False),
) -> Dict[str, Any]:
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".html") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = Path(tmp_file.name)
        try:
            return _parse_qac_path(tmp_path, old_version=old_version)
        finally:
            try:
                tmp_path.unlink()
            except Exception:
                pass
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Parse error: {exc}") from exc


@router.post("/api/qac/generate-excel")
async def qac_generate_excel(
    file: UploadFile = File(...),
    old_version: bool = Query(False),
    output_filename: Optional[str] = None,
) -> FileResponse:
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".html") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = Path(tmp_file.name)
        try:
            qac_manager = parse_qac_report(tmp_path, old_version)
            output_dir = repo_root / "reports" / "qac_excel"
            output_dir.mkdir(parents=True, exist_ok=True)
            # ⚠ `output_filename` 은 **클라이언트가 준 쿼리 파라미터**다. `output_dir / …`
            #   는 봉인이 아니다 — `../` 도, 기준 디렉터리를 통째로 대체하는 절대경로도
            #   통과한다. 같은 파일의 다운로드(`qac_download_report`)는 이미
            #   `safe_resolve_under` 를 쓴다 — **읽기만 봉인된 비대칭**이었다.
            #   (`/api/vcast/generate-excel` 에 같은 결함이 있었다 — 같은 패턴의 다른 입구.)
            if output_filename:
                filename = output_filename if output_filename.endswith(".xlsx") else f"{output_filename}.xlsx"
                try:
                    want = safe_resolve_under(output_dir, filename)
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail=f"invalid output_filename: {exc}")
            else:
                want = output_dir / f"qac_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            want.parent.mkdir(parents=True, exist_ok=True)
            output_path = reserve_unique_path(want)
            filename = output_path.name
            if not generate_qac_excel(qac_manager, output_path):
                drop_empty_reservation(output_path)
                raise HTTPException(status_code=500, detail="Excel generation failed")
            return FileResponse(
                str(output_path),
                filename=filename,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        finally:
            try:
                tmp_path.unlink()
            except Exception:
                pass
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Excel generation error: {exc}") from exc


@router.get("/api/qac/reports")
def qac_list_reports() -> Dict[str, Any]:
    """생성된 QAC Excel 리포트 목록"""
    output_dir = repo_root / "reports" / "qac_excel"
    if not output_dir.exists():
        return {"ok": True, "reports": [], "count": 0}
    files = []
    for f in sorted(output_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
        files.append({
            "name": f.name,
            "size": f.stat().st_size,
            "created": datetime.fromtimestamp(f.stat().st_mtime).isoformat(timespec="seconds"),
        })
    return {"ok": True, "reports": files, "count": len(files)}


@router.get("/api/qac/reports/{filename}")
def qac_download_report(filename: str) -> FileResponse:
    """생성된 QAC Excel 리포트 다운로드"""
    output_dir = repo_root / "reports" / "qac_excel"
    from backend.services.paths import safe_resolve_under
    file_path = safe_resolve_under(output_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="file not found")
    return FileResponse(str(file_path), filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/api/qac/generate-excel-from-path")
def qac_generate_excel_from_path(body: Dict[str, Any]) -> FileResponse:
    """서버 로컬 경로의 QAC HTML → Excel 생성"""
    file_path = str(body.get("path", "")).strip()
    old_version = bool(body.get("old_version", False))
    if not file_path:
        raise HTTPException(status_code=400, detail="path required")
    p = Path(file_path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=400, detail=f"파일을 찾을 수 없습니다: {file_path}")
    try:
        qac_manager = parse_qac_report(p, old_version)
        output_dir = repo_root / "reports" / "qac_excel"
        output_dir.mkdir(parents=True, exist_ok=True)
        # ⚠ 키가 입력 파일 stem 뿐이다 — 같은 HTML 을 두 사람이 같은 초에 변환하면 겹친다.
        output_path = reserve_unique_path(
            output_dir / f"qac_{p.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        filename = output_path.name
        if not generate_qac_excel(qac_manager, output_path):
            drop_empty_reservation(output_path)
            raise HTTPException(status_code=500, detail="Excel generation failed")
        return FileResponse(str(output_path), filename=filename,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/api/qac/scan-folder")
def qac_scan_folder(body: Dict[str, Any]) -> Dict[str, Any]:
    """폴더 경로를 스캔하여 QAC HTML 파일 목록 반환"""
    folder = str(body.get("folder", "")).strip()
    if not folder:
        raise HTTPException(status_code=400, detail="folder path required")
    p = Path(folder).expanduser().resolve()
    if not p.exists() or not p.is_dir():
        raise HTTPException(status_code=400, detail=f"폴더를 찾을 수 없습니다: {folder}")

    files = []
    for f in sorted(p.rglob("*.html")):
        name = f.name
        kind = "hmr" if "_HMR_" in name else "crr" if "_CRR_" in name else "rcr" if "_RCR_" in name else "other"
        can_parse = kind == "hmr"
        files.append({
            "name": f.name,
            "path": str(f),
            "rel_path": str(f.relative_to(p)),
            "kind": kind,
            "can_parse": can_parse,
            "size": f.stat().st_size,
        })
    return {"ok": True, "folder": str(p), "items": files, "count": len(files)}
