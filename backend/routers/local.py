"""Auto-generated router: local"""
import json
import logging
import os
import re
import tempfile
import threading
import time
import traceback
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse

import config
from backend.dependencies.admin import require_admin
from backend.helpers import (
    _apply_uds_view_filters,
    _augment_path,
    _build_excel_artifact_payload,
    _build_excel_artifact_summary,
    _build_preflight,
    _build_quality_evaluation,
    _collect_tool_paths,
    _compute_quick_quality_gate,
    _compute_uds_mapping_summary,
    _enrich_function_quality_fields,
    _generate_docx_with_retry,
    _get_progress,
    _get_source_sections_cached,
    _get_uds_view_payload_cached,
    _is_allowed_req_doc,
    _local_reports_dir,
    _local_sits_dir,
    _local_sts_dir,
    _local_suts_dir,
    _local_uds_dir,
    _open_local_path,
    _parse_component_map_file,
    _parse_path_list,
    _read_excel_artifact_sidecar,
    _resolve_local_report_path,
    _resolve_local_sits_path,
    _resolve_local_sts_path,
    _resolve_local_suts_path,
    _resolve_local_uds_path,
    _resolve_report_dir,
    _resolve_source_root_from_cfg,
    _run_impact_analysis_for_uds,
    _run_report_with_timeout,
    _set_progress,
    _validate_docx_template_bytes,
    _write_excel_artifact_sidecar,
    _write_residual_tbd_report,
    _write_upload_to_temp,
    build_vectorcast_metadata,
    evaluate_vectorcast_readiness,
    load_vectorcast_project_config,
)
from backend.helpers.sds import build_sds_view_model, is_sds_filename, is_srs_filename
from backend.schemas import (
    EditorReadAbsRequest,
    EditorReadRequest,
    EditorReplaceRequest,
    EditorWriteRequest,
    FormatCodeRequest,
    GitRequest,
    KBRequest,
    ListDirRequest,
    LocalImpactTriggerRequest,
    LocalReportGenerateRequest,
    OpenFileRequest,
    OpenFolderRequest,
    PickerRequest,
    PreflightRequest,
    RagIngestRequest,
    RagQueryRequest,
    RagStatusRequest,
    RagStorageRequest,
    ReplaceTextRequest,
    ScmRequest,
    SdsViewRequest,
    SearchRequest,
    TextPreviewRequest,
)
from backend.services.files import read_text_limited
from backend.services.local_report_generator import generate_local_docx, generate_local_xlsx
from backend.services.local_service import (
    delete_kb_entry,
    format_c_code,
    git_branches,
    git_checkout,
    git_commit,
    git_create_branch,
    git_diff,
    git_log,
    git_stage,
    git_status,
    git_unstage,
    list_directory,
    list_kb_entries,
    pick_directory,
    pick_file,
    read_file_text,
    replace_in_file,
    replace_lines,
    run_git,
    run_svn,
    search_in_files,
    write_file_text,
)
from backend.services.paths import is_under_any, safe_resolve_under
from backend.user_context import wrap_with_user
from report_gen.provenance import has_evidence_value, is_weak_source
from report_gen.utils import build_function_details_by_name
from report_generator import (
    _build_req_map_from_doc_paths,
    enrich_function_details_with_docs,
    generate_asil_related_confidence_report,
    generate_called_calling_accuracy_report,
    generate_swcom_context_report,
    generate_uds_constraints_report,
    generate_uds_field_quality_gate_report,
    generate_uds_preview_html,
    generate_uds_requirements_from_docs,
    generate_uds_source_sections,
    generate_uds_validation_report,
)
from workflow.change_trigger import build_registry_trigger
from workflow.impact_jobs import start_impact_job
from workflow.impact_orchestrator import run_impact_update

try:
    from workflow.rag import _read_and_chunk_file, _read_text_from_file, get_kb, ingest_external_sources
except ImportError:
    _read_text_from_file = None
    _read_and_chunk_file = None
    get_kb = None
    ingest_external_sources = None
try:
    from workflow.uds_ai import generate_uds_ai_sections
except ImportError:
    generate_uds_ai_sections = None

repo_root = Path(__file__).resolve().parents[2]


router = APIRouter()
_logger = logging.getLogger("devops_api")


# ---------------------------------------------------------------------------
# 요청자가 지정한 base(`project_root`) 확정 — **단일 출처**
# ---------------------------------------------------------------------------
#
# ## 왜 생겼나 (2026-08-04, 보안 표면 감사)
#
# 이 파일의 endpoint 20곳이 `req.project_root` 를 **그대로 base 로** 써서, 인증만 통과하면
# (당시엔 `X-User` 헤더 한 줄이면 됐다) 디스크 임의 위치를 읽고 쓸 수 있었다. 실측 재현:
#
#   POST /api/local/editor/write  project_root="C:/Users/<me>"                  -> 200, 홈에 파일 생성
#   POST /api/local/editor/write  project_root=…\Start Menu\Programs\Startup    -> 200, **로그인 시 자동실행 지속성**
#   POST /api/local/editor/write  rel_path="backend/routers/__probe.py"         -> 200, **코드 주입 표면**
#   POST /api/local/editor/read   rel_path=".env"                               -> 200, 2,165B (JWT_SECRET 포함)
#   POST /api/local/editor/read   rel_path="reports/quality.sqlite"             -> 200, 'SQLite format 3\x00'
#
# ⚠ **traversal 가드는 이미 있었고 정상 동작했다**(`../` 3종 전부 차단). 결함은 traversal 이
#    아니라 **base 지정**이다 — `rel_path` 만 검사하고 root 는 body 를 그대로 믿었다.
#    같은 파일 `:4155`/`:4177`/`:4236`(open-file/read-abs/open-folder)은 이미 이 확정을
#    하고 있었다 — 읽기전용 3곳은 잠겼고 **쓰기 3곳은 열려 있던** 비대칭이었다.
#
# ⚠ 확정만으로는 부족하다. `.env`·`reports/quality.sqlite` 는 `repo_root` **밑**이라
#    화이트리스트를 통과한다. 그래서 민감 경로 거부를 함께 둔다(`_deny_sensitive_target`).
#
# ⚠ 이 헬퍼는 **라우터 계층 전용**이다. MCP(`write_file`/`replace_in_file`)는 HTTP 를 타지
#    않고 `local_service` 함수를 in-process 로 부르며 자체 가드가 있다 — 건드리지 않는다.

def _allowed_request_roots() -> List[Path]:
    """요청자가 base 로 지정할 수 있는 최상위 경로. `:4157` 과 같은 목록이다."""
    return [(Path.home() / ".devops_pro_cache").resolve(), repo_root.resolve()]


# base 확정을 통과해도 **내용을 내주면 안 되는** 것들. 전부 `repo_root` 밑이라
# 화이트리스트로는 안 걸린다.
_SENSITIVE_NAME_PREFIXES = (".env",)
_SENSITIVE_SUFFIXES = (".sqlite", ".sqlite3", ".db")
_SENSITIVE_RELATIVE = (
    Path("config/admin_users.json"),
    Path("config/users.json"),
    Path("config/allowed_users.json"),
)


def _deny_sensitive_target(target: Path) -> None:
    """자격·신원·감사 저장소는 이 API 로 읽지도 쓰지도 못한다."""
    name = target.name.lower()
    if name.startswith(_SENSITIVE_NAME_PREFIXES):
        raise HTTPException(status_code=403, detail="sensitive file not allowed")
    # `quality.sqlite.bak` 처럼 접미가 더 붙어도 막는다 — 확장자 일치만 보면 새 나간다.
    if any(s in name for s in _SENSITIVE_SUFFIXES):
        raise HTTPException(status_code=403, detail="sensitive file not allowed")
    try:
        rel = target.resolve().relative_to(repo_root.resolve())
    except (ValueError, OSError):
        return
    if any(rel == p or rel.as_posix() == p.as_posix() for p in _SENSITIVE_RELATIVE):
        raise HTTPException(status_code=403, detail="sensitive file not allowed")


def confine_request_root(project_root: Any, *, rel_path: Any = None) -> str:
    """요청 body 의 base 를 허용 루트 안으로 **확정**한다. 밖이면 403.

    Args:
        project_root: 요청자가 준 base. 빈 값이면 `repo_root`.
        rel_path: 있으면 최종 대상까지 민감 경로 검사를 건다.

    Returns:
        확정된 base 문자열(하위 서비스 함수가 그대로 쓸 수 있게).
    """
    raw = str(project_root or "").strip() or str(repo_root)
    try:
        target = Path(raw).expanduser().resolve()
    except (OSError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="invalid project_root") from exc
    if not is_under_any(target, _allowed_request_roots()):
        # ⚠ 어떤 경로가 허용되는지 응답에 적지 않는다 — 실패 자체가 정보다.
        _logger.warning("허용 밖 project_root 요청을 차단했다: %s", target)
        raise HTTPException(status_code=403, detail="project_root not allowed")
    _deny_sensitive_target(target)
    if rel_path is not None and str(rel_path).strip():
        try:
            combined = (target / str(rel_path)).resolve()
        except (OSError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="invalid rel_path") from exc
        _deny_sensitive_target(combined)
    return str(target)


# blocking 오프로딩의 **단일 정의**는 backend/routers/_safety.py 에 있다.
# 여기 사본을 두면 jenkins.py 와 갈라져 한쪽만 고쳐진다(이 저장소 1순위 재발 패턴).
from backend.routers._safety import run_blocking as _run_blocking  # noqa: E402
from backend.services.resolver_helpers import read_requirement_doc  # noqa: E402

_MAX_PREVIEW_COLS = 20


def _pick_excel_suffix(template_path: Optional[str]) -> str:
    if template_path:
        suffix = Path(template_path).suffix.lower()
        if suffix in (".xlsm", ".xlsx"):
            return suffix
    return ".xlsx"


def _build_vectorcast_package_response(
    *,
    package_dir: Path,
    package_name: str,
    manifest: Dict[str, Any],
    project_config: Dict[str, Any],
    units: List[str],
) -> Dict[str, Any]:
    metadata = build_vectorcast_metadata(
        project_config=project_config,
        source_root=str(project_config.get("source_root") or ""),
        units=units,
    )
    readiness = evaluate_vectorcast_readiness(metadata)
    return {
        "ok": True,
        "package_dir": str(package_dir),
        "package_name": package_name,
        "manifest": manifest,
        "files": sorted(str(p.name) for p in package_dir.iterdir() if p.is_file()),
        "project_config": metadata,
        "readiness": readiness,
    }


def _build_local_excel_output(base_dir: Path, category: str, stem: str, template_path: Optional[str]) -> Tuple[str, Path]:
    if category == "sts":
        target_dir = _local_sts_dir(base_dir)
    elif category == "sits":
        target_dir = _local_sits_dir(base_dir)
    else:
        target_dir = _local_suts_dir(base_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S") + f"_{uuid.uuid4().hex[:4]}"
    suffix = _pick_excel_suffix(template_path)
    filename = f"{stem}_{ts}{suffix}"
    return filename, target_dir / filename


def _excel_media_type(file_path: Path) -> str:
    if file_path.suffix.lower() == ".xlsm":
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _write_uds_payload_sidecar(out_path: Path, uds_payload: Dict[str, Any]) -> Optional[Path]:
    try:
        details = uds_payload.get("function_details")
        if not isinstance(details, dict):
            return None
        summary = uds_payload.get("summary")
        if not isinstance(summary, dict):
            summary = {}
        summary["mapping"] = _compute_uds_mapping_summary(details)
        uds_payload["summary"] = summary
        sidecar = out_path.with_suffix(".payload.json")
        payload = {
            "docx_path": str(out_path),
            "summary": summary,
            "function_details": details,
        }
        sidecar.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return sidecar
    except Exception as exc:
        _logger.warning("uds payload sidecar write skipped: %s", exc)
        return None


def _discover_default_req_docs() -> Dict[str, List[str]]:
    docs_dir = repo_root / "docs"
    result: Dict[str, List[str]] = {"req": [], "sds": []}
    if not docs_dir.exists():
        return result
    for path in docs_dir.glob("*.docx"):
        lower = path.name.lower()
        if is_srs_filename(lower) or is_sds_filename(lower):
            result["req"].append(str(path))
        if is_sds_filename(lower):
            result["sds"].append(str(path))
    return result


def _dedupe_paths(paths: Optional[List[str]]) -> List[str]:
    items: List[str] = []
    seen = set()
    for raw in paths or []:
        try:
            p = str(raw or "").strip()
        except Exception:
            p = ""
        if not p or p in seen:
            continue
        seen.add(p)
        items.append(p)
    return items


_GENERIC_SOURCE_DIR_NAMES = {
    "sources",
    "source",
    "src",
    "srcs",
    "code",
    "codes",
    "app",
    "application",
    "firmware",
    "fw",
    "project",
}


def _derive_project_module_names(root: Optional[Path]) -> Tuple[str, str]:
    """Pick sensible (project_name, module_name) from a source root.

    If the leaf directory looks like a generic source folder ("Sources",
    "src" etc.), the parent directory is used as project_name and the leaf
    becomes the module_name. Otherwise the leaf is used for both.
    """
    if not root:
        return "", ""
    try:
        leaf = root.name or ""
        parent = root.parent.name if root.parent else ""
    except Exception:
        return "", ""
    if leaf and leaf.lower() in _GENERIC_SOURCE_DIR_NAMES and parent:
        return parent, leaf
    return leaf, leaf


def _resolve_req_doc_sets(
    req_doc_paths: Optional[List[str]] = None,
    sds_doc_paths: Optional[List[str]] = None,
) -> Tuple[List[str], List[str]]:
    """요구/SDS 문서 경로 확정 — 저장소 `docs/` 글롭은 **아무것도 안 준 경우에만**.

    예전엔 사용자가 준 경로에 저장소 글롭을 **무조건 이어붙였다**(`user + defaults`).
    그러면 어느 프로젝트를 돌리든 저장소 동봉 HDPDM01 SDS 가 항상 섞인다.
    `enrich_function_details_with_docs` 의 병합은 first-wins 라, 사용자 문서가 **빈칸인
    항목을 HDPDM01 의 ASIL·요구ID 가 채우고**, 사용자 문서에 아예 없는 함수는 HDPDM01
    엔트리가 통째로 추가된다.

    실측(KJPDS02 SwDS ↔ 저장소 HDPDM01 SwDS): 함수 엔트리 **473개(KJPDS02 의 80.9%)가
    이름 충돌**하고 그 **전부**가 HDPDM01 쪽에 asil·related 를 갖고 있었다 —
    예: `adc0_stop_current_workaround` ← HDPDM01 의 `SwTR_0107, SwNTR_0103`.

    같은 규율이 이 파일의 `_doc_or_discovered`(SRS)와 `generators/suts.py`
    `load_sds_map_from`(커밋 `1bfdee9` "프로젝트 간 오염 3건")에 이미 있다 —
    여기만 빠져 있었다. 그쪽은 "해석 실패 시 대체 금지"였는데 여기는 조건조차 없었다.
    """
    user_req = _dedupe_paths([p for p in (req_doc_paths or []) if str(p or "").strip()])
    user_sds = _dedupe_paths([p for p in (sds_doc_paths or []) if str(p or "").strip()])
    # 사용자가 준 요구 문서 중 이름에 'sds' 가 든 것도 SDS 로 인정 — Jenkins 경로
    # (`jenkins.py` `sds_doc_paths` 구성)와 동일 규칙이라 모드 간 결과가 갈리지 않는다.
    if not user_sds and user_req:
        user_sds = _dedupe_paths([p for p in user_req if is_sds_filename(p)])
    if user_req or user_sds:
        if not user_sds:
            # 침묵 금지 — SDS 없이 진행하면 ASIL/요구 보강이 비는데, 그게 저장소 문서로
            # 채워지는 것보다는 낫다. 다만 왜 비었는지는 남긴다.
            _logger.warning(
                "SDS 미지정 — 저장소 docs/ 문서로 대체하지 않는다(다른 프로젝트 오염 방지). "
                "함수별 ASIL/요구 보강은 생략된다")
        return user_req, user_sds
    defaults = _discover_default_req_docs()
    req_paths = _dedupe_paths(list(defaults.get("req") or []))
    sds_paths = _dedupe_paths(list(defaults.get("sds") or []))
    if req_paths or sds_paths:
        _logger.info("요구/SDS 문서 미지정 — 저장소 docs/ 에서 자동 탐색(프로젝트 무관): "
                     "req %d · sds %d", len(req_paths), len(sds_paths))
    return req_paths, sds_paths


def _load_sts_ai_config() -> Optional[Dict[str, Any]]:
    """Load AI config for STS enhancement from default OAI config path."""
    try:
        import config as _appconfig
        from workflow.ai import load_oai_config
        cfg_path = getattr(_appconfig, "DEFAULT_OAI_CONFIG_PATH", None)
        cfg = load_oai_config(cfg_path)
        if cfg and isinstance(cfg, dict) and cfg.get("model"):
            return cfg
    except Exception as _e:
        _logger.debug("STS ai_config load skipped: %s", _e)
    return None


def _discover_hsis_path() -> Optional[str]:
    """Auto-discover HSIS xlsx file from docs/ directory."""
    try:
        docs_dir = Path(__file__).resolve().parents[2] / "docs"
        for p in docs_dir.glob("*.xlsx"):
            if "hsis" in p.name.lower():
                return str(p)
    except Exception:
        pass
    return None


def _discover_srs_docx() -> Optional[str]:
    """저장소 `docs/` 에서 SRS docx 하나를 고른다(프로젝트 무관)."""
    for p in _discover_default_req_docs().get("req", []):
        if is_srs_filename(p) and p.endswith(".docx"):
            return p
    return None


def _discover_sds_docx() -> Optional[str]:
    """저장소 `docs/` 에서 SDS docx 하나를 고른다(프로젝트 무관)."""
    for p in _discover_default_req_docs().get("sds", []):
        return p
    return None


def _doc_or_discovered(
    resolved: Optional[str],
    user_supplied: Any,
    discover: Callable[[], Optional[str]],
    *,
    label: str,
    tag: str = "",
) -> Optional[str]:
    """해석된 경로가 없을 때 자동 탐색을 쓸지 결정한다.

    자동 탐색(저장소 `docs/` 글롭)은 사용자가 **아무것도 안 준 경우에만** 쓴다.
    사용자가 경로를 줬는데 해석에 실패했다면(대표 사례: cloudium worker-only `U:\\…` —
    로컬 `Path.exists()` 가 항상 False) **대체하지 않고 경고만 남긴다**.

    과거엔 두 경우를 `if not resolved:` 하나로 묶어, 지정한 문서를 못 읽으면 저장소
    `docs/` 에 들어있는 **다른 프로젝트 문서**(현재 HDPDM01)로 조용히 바꿔치기했다.
    로그가 "auto-discovered" 라 기능처럼 읽혔고, SDS 는 요구-함수 매핑 전체를 좌우하므로
    산출물이 통째로 남의 프로젝트 설계 기준이 됐다.
    """
    if resolved:
        return resolved
    if user_supplied:
        _logger.warning(
            "%s%s: 지정한 입력을 해석하지 못해 건너뛴다 — 저장소 docs/ 문서로 대체하지 "
            "않는다(다른 프로젝트 오염 방지)", tag, label)
        return None
    picked = discover()
    if picked:
        _logger.info("%s%s 미지정 — 저장소 docs/ 에서 자동 탐색(프로젝트 무관): %s",
                     tag, label, picked)
    return picked


def _localize_uds_for_enrich(uds_path: Optional[str]) -> Optional[str]:
    """SwUDS 문서를 로컬 tmp .docx로 복사해 enrich 파서가 직접 open 가능하게 한다.

    enrich/하위 파서는 path.exists()/open() 로컬 fs 직접 접근이라 cloudium U: 경로를
    직접 읽지 못한다. resolver.read_bytes(worker 8765)로 bytes를 받아 tmp .docx에 기록
    후 그 경로를 반환한다. 호출부는 사용 후 반드시 os.unlink로 삭제해야 한다(누수 방지).

    Args:
        uds_path: SwUDS 문서 경로(로컬 또는 cloudium). 빈 값이면 None.

    Returns:
        로컬 tmp .docx 경로. 미입력 또는 접근 실패 시 None(기존 동작 불변).
    """
    if not uds_path or not str(uds_path).strip():
        return None
    try:
        from backend.services.file_resolver import get_resolver
        from backend.services.resolver_helpers import enforce_resolver_access
        enforce_resolver_access(uds_path)
        data = get_resolver().read_bytes(uds_path)
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp.write(data)
            return tmp.name
    except Exception as exc:
        _logger.warning("UDS localize for enrich skipped: %s", exc)
        return None


def _enrich_function_details_map(
    function_details: Optional[Dict[str, Any]],
    *,
    function_table_rows: Optional[List[List[Any]]] = None,
    req_doc_paths: Optional[List[str]] = None,
    sds_doc_paths: Optional[List[str]] = None,
    uds_path: Optional[str] = None,
) -> Tuple[Dict[str, Any], List[str], List[str]]:
    details = function_details if isinstance(function_details, dict) else {}
    req_paths, sds_paths = _resolve_req_doc_sets(req_doc_paths, sds_doc_paths)
    if details:
        _uds_tmp = _localize_uds_for_enrich(uds_path)
        try:
            enrich_function_details_with_docs(
                details,
                function_table_rows,
                req_doc_paths=req_paths,
                sds_doc_paths=sds_paths,
                uds_doc_paths=[_uds_tmp] if _uds_tmp else None,
            )
        except Exception as exc:
            _logger.warning("function detail enrichment skipped: %s", exc)
        finally:
            if _uds_tmp:
                try:
                    os.unlink(_uds_tmp)
                except OSError:
                    pass
    # HSIS enrichment: functions using HSIS signal variables get
    # description_source/related_source upgraded from "inference" to "hsis"
    _hsis_p = _discover_hsis_path()
    if _hsis_p and details:
        try:
            from generators.sts import _load_hsis_signals
            _hsis_d = _load_hsis_signals(_hsis_p)
            _hsis_sigs = _hsis_d.get("signals", [])
            if _hsis_sigs:
                _hvar: Dict[str, Dict] = {}
                for _s in _hsis_sigs:
                    _sw = str(_s.get("sw_var_name") or "")
                    for _tok in re.split(r"[\n,\s]+", _sw):
                        _tok = _tok.strip()
                        if _tok and re.match(r"^[A-Za-z_]\w+$", _tok):
                            _hvar[_tok] = _s
                for _fn_info in details.values():
                    if not isinstance(_fn_info, dict):
                        continue
                    _fvars: set = set()
                    # inputs/outputs 원소는 소스 파서에 따라 dict({name:...}) 또는 str(예 "[OUT]
                    # return U8 …")로 온다. dict 가정으로 .get 호출 시 str에서 AttributeError가
                    # 나 try/except가 전체 함수 루프를 중단 → 이후 함수 HSIS 보강 전멸(실측 877/900).
                    # dict는 name, str은 그대로 수용(HSIS 신호는 주로 globals로 매칭되므로 무해).
                    for _x in (_fn_info.get("inputs") or []):
                        _fvars.add(str(_x.get("name") or "") if isinstance(_x, dict) else str(_x))
                    for _x in (_fn_info.get("outputs") or []):
                        _fvars.add(str(_x.get("name") or "") if isinstance(_x, dict) else str(_x))
                    _fvars.update((_fn_info.get("globals_write") or {}).keys())
                    _fvars.update((_fn_info.get("globals_read") or {}).keys())
                    _matched_sigs = [_hvar[v] for v in _fvars if v in _hvar]
                    if not _matched_sigs:
                        continue
                    # Upgrade description_source from inference → hsis
                    # ⚠ **값이 있을 때만** 올린다. `hsis` 는 별칭이 `sds`(0.95)라, 설명이
                    #    빈 칸인데 라벨만 올리면 "근거는 SDS 급인데 내용이 없다" 는 상태가
                    #    0.95 를 받는다(`_score_for` 는 값 유무를 안 본다).
                    if (
                        has_evidence_value(_fn_info.get("description"))
                        and is_weak_source(_fn_info.get("description_source") or "inference")
                    ):
                        _fn_info["description_source"] = "hsis"
                    # Set related if currently TBD/empty
                    _cur_rel = str(_fn_info.get("related") or "").strip()
                    if not _cur_rel or _cur_rel.upper() in {"TBD", "N/A", "-"}:
                        _rel_ids = [
                            str(s.get("related_id") or "").strip()
                            for s in _matched_sigs
                            if str(s.get("related_id") or "").strip()
                        ]
                        if _rel_ids:
                            _fn_info["related"] = _rel_ids[0]
                            _fn_info["related_source"] = "hsis"
        except Exception as _hsis_exc:
            _logger.warning("HSIS UDS enrichment skipped: %s", _hsis_exc)
    return details, req_paths, sds_paths


def _enrich_source_sections_with_docs(
    source_sections: Optional[Dict[str, Any]],
    *,
    req_doc_paths: Optional[List[str]] = None,
    sds_doc_paths: Optional[List[str]] = None,
    uds_path: Optional[str] = None,
) -> Tuple[Dict[str, Any], List[str], List[str]]:
    sections = source_sections if isinstance(source_sections, dict) else {}
    details = sections.get("function_details", {})
    table_rows = sections.get("function_table_rows", [])
    details, req_paths, sds_paths = _enrich_function_details_map(
        details,
        function_table_rows=table_rows if isinstance(table_rows, list) else None,
        req_doc_paths=req_doc_paths,
        sds_doc_paths=sds_doc_paths,
        uds_path=uds_path,
    )
    sections["function_details"] = details

    # ⚠ 키 규칙은 `report_gen.utils.function_name_key` **단일 출처**를 따른다.
    #    여기 있던 인라인 루프는 `.strip()` 만 해서 **원형 대소문자**를 키로 썼는데,
    #    조회는 전부 소문자다(`docx_builder` 13곳 · `code.py:126` · `test_gen.py:32` ·
    #    `uds_generator` 4곳). 실측 표본 350개 중 267개(76.3%)가 대문자를 포함해
    #    그만큼 **조용히 miss** 했다. jenkins 경로는 처음부터 소문자였다(비대칭).
    sections["function_details_by_name"] = build_function_details_by_name(details)
    return sections, req_paths, sds_paths


def _find_latest_excel_file(directory: Path) -> Optional[Path]:
    files = [
        p for p in directory.iterdir()
        if p.is_file() and p.suffix.lower() in (".xlsm", ".xlsx")
    ]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def _parse_xlsm_preview(file_path: Path, max_rows: int = 30) -> Dict[str, Any]:
    """Parse XLSM/XLSX and return sheet data as JSON for web preview."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True, keep_vba=False)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot open XLSM: {e}")

    # Safety cap: full-viewer sends large max_rows but we cap to avoid timeout
    effective_max_rows = min(max_rows, 5000)

    sheets: List[Dict[str, Any]] = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        mr = ws.max_row or 0
        mc = ws.max_column or 0
        if mr == 0 or mc == 0:
            sheets.append({"name": sname, "headers": [], "rows": [], "total_rows": 0, "total_cols": 0})
            continue

        col_limit = min(mc, _MAX_PREVIEW_COLS)

        headers: List[str] = []
        rows: List[List[Any]] = []
        # Use iter_rows() — much faster than ws.cell() in read_only mode
        for row_idx, row in enumerate(ws.iter_rows(max_col=col_limit, values_only=True)):
            if row_idx == 0:
                headers = [str(v) if v is not None else f"Col{ci + 1}" for ci, v in enumerate(row)]
                continue
            if row_idx >= effective_max_rows:
                break
            row_data: List[Any] = []
            for v in row:
                if v is None:
                    row_data.append("")
                elif isinstance(v, (int, float)):
                    row_data.append(v)
                else:
                    s = str(v).strip()
                    row_data.append(s[:200] if len(s) > 200 else s)
            if any(v != "" for v in row_data):
                rows.append(row_data)

        sheets.append({
            "name": sname,
            "headers": headers,
            "rows": rows,
            "total_rows": mr,
            "total_cols": mc,
        })

    all_sheet_names = list(wb.sheetnames)
    wb.close()
    return {
        "filename": file_path.name,
        "sheets": sheets,
        "sheet_names": all_sheet_names,
    }


def _load_excel_artifact_payload(
    file_path: Path,
    artifact_type: str,
    *,
    download_url: str,
    preview_url: str,
) -> Dict[str, Any]:
    payload = _read_excel_artifact_sidecar(file_path)
    if not payload:
        payload = _repair_excel_artifact_payload(
            file_path,
            artifact_type,
            download_url=download_url,
            preview_url=preview_url,
        )
    payload["filename"] = file_path.name
    payload["output_path"] = str(file_path)
    payload["download_url"] = download_url
    payload["preview_url"] = preview_url
    if not str(payload.get("validation_report_path") or "").strip():
        validation_path = file_path.with_suffix(".validation.md")
        payload["validation_report_path"] = str(validation_path) if validation_path.exists() else ""
    if not str(payload.get("residual_report_path") or "").strip():
        residual_path = file_path.with_suffix(".residual.md")
        if not residual_path.exists():
            residual_path = file_path.with_suffix(".residual_tbd.md")
        payload["residual_report_path"] = str(residual_path) if residual_path.exists() else ""
    if not isinstance(payload.get("summary"), dict):
        payload["summary"] = _build_excel_artifact_summary(artifact_type, payload.get("raw_result") or payload)
    return payload


def _repair_excel_artifact_payload(
    file_path: Path,
    artifact_type: str,
    *,
    download_url: str,
    preview_url: str,
) -> Dict[str, Any]:
    kind = str(artifact_type or "").strip().lower()
    result: Dict[str, Any] = {
        "ok": True,
        "output_path": str(file_path),
        "filename": file_path.name,
        "download_url": download_url,
        "preview_url": preview_url,
    }
    validation_path = file_path.with_suffix(".validation.md")
    if validation_path.exists():
        result["validation_report_path"] = str(validation_path)
    residual_path = file_path.with_suffix(".residual_tbd.md")
    if residual_path.exists():
        result["residual_report_path"] = str(residual_path)
    try:
        if kind == "sts":
            from generators.suts import validate_sts_xlsm
            validation = validate_sts_xlsm(str(file_path))
            stats = validation.get("stats", {}) if isinstance(validation, dict) else {}
            result["validation"] = validation
            result["test_case_count"] = int(stats.get("tc_count") or 0)
        elif kind == "suts":
            from generators.suts import validate_suts_xlsm
            validation = validate_suts_xlsm(str(file_path))
            stats = validation.get("stats", {}) if isinstance(validation, dict) else {}
            result["validation"] = validation
            result["test_case_count"] = int(stats.get("tc_count") or 0)
            result["total_sequences"] = int(stats.get("seq_count") or 0)
            result["quality_report"] = {
                "avg_sequences_per_tc": float(stats.get("avg_seq_per_tc") or 0),
            }
        elif kind == "sits":
            from generators.sits import validate_sits_xlsm
            validation = validate_sits_xlsm(str(file_path))
            stats = validation.get("stats", {}) if isinstance(validation, dict) else {}
            result["validation"] = validation
            result["test_case_count"] = int(stats.get("tc_count") or 0)
            result["total_sub_cases"] = int(stats.get("sub_case_count") or 0)
            result["flow_count"] = int(stats.get("flow_count") or 0)
    except Exception:
        result.setdefault("validation", {})
    payload = _build_excel_artifact_payload(
        kind,
        result,
        output_path=str(file_path),
        filename=file_path.name,
        download_url=download_url,
        preview_url=preview_url,
    )
    _write_excel_artifact_sidecar(file_path, kind, payload)
    return payload

@router.post("/api/local/reports/generate")
def local_reports_generate(req: LocalReportGenerateRequest) -> Dict[str, Any]:
    report_dir = _resolve_report_dir(req.report_dir)
    summary_path = report_dir / "analysis_summary.json"
    if not summary_path.exists():
        raise HTTPException(status_code=404, detail="analysis_summary.json not found")
    try:
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"summary parse error: {e}")

    out_dir = _local_reports_dir(report_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"local_report_{ts}"
    formats = [str(f).lower() for f in (req.formats or [])]

    outputs: List[Dict[str, Any]] = []
    if "docx" in formats:
        out_path = out_dir / f"{base}.docx"
        generate_local_docx(summary, out_path)
        outputs.append({"file": out_path.name, "path": str(out_path)})
    if "xlsx" in formats:
        out_path = out_dir / f"{base}.xlsx"
        generate_local_xlsx(summary, out_path)
        outputs.append({"file": out_path.name, "path": str(out_path)})

    return {"ok": True, "files": outputs}


@router.get("/api/local/reports")
def local_reports_list(report_dir: Optional[str] = None) -> List[Dict[str, Any]]:
    report_path = _resolve_report_dir(report_dir)
    reports_dir = _local_reports_dir(report_path)
    if not reports_dir.exists():
        return []
    rows: List[Dict[str, Any]] = []
    for p in reports_dir.glob("local_report_*.*"):
        if not p.is_file():
            continue
        rows.append(
            {
                "file": p.name,
                "path": str(p),
                "size_mb": round(p.stat().st_size / (1024 * 1024), 2),
                "mtime": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
                "download_url": f"/api/local/reports/download/{p.name}?report_dir={report_path}",
            }
        )
    rows.sort(key=lambda x: x.get("mtime") or "", reverse=True)
    return rows


@router.get("/api/local/reports/download/{filename}")
def local_reports_download(filename: str, report_dir: Optional[str] = None) -> FileResponse:
    report_path = _resolve_report_dir(report_dir)
    file_path = _resolve_local_report_path(report_path, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="report not found")
    media = "application/octet-stream"
    if file_path.suffix.lower() == ".docx":
        media = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif file_path.suffix.lower() == ".xlsx":
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(str(file_path), filename=file_path.name, media_type=media)


@router.post("/api/local/uds/generate")
async def local_uds_generate(
    request: Request,
    source_root: str = Form(""),
    req_files: List[UploadFile] = File(default_factory=list),
    req_paths: str = Form(""),
    template_file: UploadFile = File(default=None),
    template_strict: bool = Form(False),
    component_list: UploadFile = File(default=None),
    logic_max_children: Optional[int] = Form(None),
    logic_max_grandchildren: Optional[int] = Form(None),
    logic_max_depth: Optional[int] = Form(None),
    globals_format_order: str = Form(""),
    globals_format_sep: str = Form(""),
    globals_format_with_labels: bool = Form(True),
    call_relation_mode: str = Form("code"),
    ai_enable: bool = Form(True),
    ai_example_path: str = Form(""),
    ai_detailed: bool = Form(True),
    expand: bool = Form(False),
    doc_only: bool = Form(False),
    test_mode: bool = Form(False),
    rag_top_k: Optional[int] = Form(None),
    rag_categories: str = Form(""),
    report_dir: str = Form(""),
    req_types: str = Form(""),
    show_mapping_evidence: bool = Form(False),
) -> Dict[str, Any]:
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    reject_upload_in_cloudium(*(req_files or []), template_file, component_list)
    req_id = (request.headers.get("x-req-id") or "").strip() or f"uds-gen-{int(time.time() * 1000)}"
    _logger.info("[UDS_GENERATE][%s] start source_root=%s test_mode=%s", req_id, source_root, bool(test_mode))
    template_bytes: Optional[bytes] = None
    template_warning = ""
    if template_file and template_file.filename:
        try:
            template_bytes = await template_file.read()
        except Exception:
            template_bytes = None
        valid_tpl, tpl_err = _validate_docx_template_bytes(template_bytes)
        if not valid_tpl:
            msg = f"template invalid: {tpl_err}"
            if bool(template_strict):
                raise HTTPException(status_code=400, detail=msg)
            template_warning = msg
            template_bytes = None
    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="source_root(코드 루트)가 필요합니다.")
    req_paths_list = _parse_path_list(req_paths)
    has_req_upload = any((f and f.filename) for f in (req_files or []))
    if not has_req_upload and not req_paths_list:
        raise HTTPException(status_code=400, detail="SRS/SDS 요구사항 문서를 최소 1개 이상 제공해주세요.")

    type_list = [t.strip().lower() for t in req_types.split(",") if t.strip()] if req_types else []

    req_texts: List[str] = []
    srs_texts: List[str] = []
    sds_texts: List[str] = []
    req_doc_paths: List[str] = []
    for idx, f in enumerate(req_files):
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            text = _read_text_from_file(tmp_path)
        except Exception:
            text = ""
        if tmp_path.suffix.lower() == ".docx":
            req_doc_paths.append(str(tmp_path))
        if text:
            req_texts.append(text.strip())
            ftype = type_list[idx] if idx < len(type_list) else ""
            if ftype == "srs":
                srs_texts.append(text.strip())
            elif ftype == "sds":
                sds_texts.append(text.strip())
    sds_doc_paths: List[str] = []
    # 탈락 사유를 버리지 않는다 — 예전엔 `except Exception: text = ""` 라
    # 경로 오타·권한 없음·본문 0자가 전부 같은 침묵이었다.
    doc_skips: List[str] = []
    for path_str in req_paths_list:
        p, text, reason = read_requirement_doc(path_str, allow=_is_allowed_req_doc)
        if reason:
            doc_skips.append(reason)
            continue
        if p and text:
            req_texts.append(text)
            if p.suffix.lower() == ".docx":
                req_doc_paths.append(str(p))
            fname_lower = p.name.lower()
            if is_srs_filename(fname_lower):
                srs_texts.append(text.strip())
            elif is_sds_filename(fname_lower):
                sds_texts.append(text.strip())
                if p.suffix.lower() in {".docx", ".doc"}:
                    sds_doc_paths.append(str(p))

    component_map: Dict[str, Dict[str, str]] = {}
    if component_list and component_list.filename:
        tmp = _write_upload_to_temp(component_list, ".json")
        if tmp:
            try:
                component_map = _parse_component_map_file(tmp)
            except Exception:
                component_map = {}
            finally:
                try:
                    tmp.unlink(missing_ok=True)
                except Exception:
                    pass

    # SDS 파티션 맵 로드 (Related ID + ASIL 전파용)
    _sds_pmap: Dict[str, Dict[str, str]] = {}
    if sds_doc_paths:
        from report_gen.requirements import _extract_sds_partition_map
        for sp in sds_doc_paths:
            try:
                _sds_pmap.update(_extract_sds_partition_map(sp))
            except Exception:
                pass

    source_sections: Dict[str, str] = {}
    if source_root_path and source_root_path.exists():
        # 전체 소스트리 파싱 — 수 초~수십 초. 이벤트 루프에서 돌리면 그동안
        # 백엔드 전체가 멈춘다(tests/unit/test_router_event_loop_blocking.py).
        source_sections = await _run_blocking(
            generate_uds_source_sections,
            source_root,  # 콤마 구분 복수 경로 그대로 전달
            component_map=component_map if component_map else None,
            sds_partition_map=_sds_pmap if _sds_pmap else None,
        )
        source_sections, req_doc_paths, sds_doc_paths = _enrich_source_sections_with_docs(
            source_sections,
            req_doc_paths=req_doc_paths,
            sds_doc_paths=sds_doc_paths,
        )

    req_from_docs = generate_uds_requirements_from_docs(req_texts) if req_texts else ""
    req_map = _build_req_map_from_doc_paths(req_doc_paths, req_texts) if req_texts or req_doc_paths else {}
    req_source = source_sections.get("requirements", "")
    if req_from_docs and req_source:
        req_combined = "\n".join([req_from_docs.strip(), req_source.strip()]).strip()
    else:
        req_combined = req_from_docs or req_source

    globals_order_list = [
        x.strip()
        for x in re.split(r"[,\|;]+", globals_format_order or "")
        if x.strip()
    ]
    _project_name_val, _module_name_val = _derive_project_module_names(source_root_path)
    uds_payload = {
        "job_url": "local",
        "build_number": "",
        "project_name": _project_name_val,
        "module_name": _module_name_val,
        "summary": {},
        "overview": source_sections.get("overview", ""),
        "requirements": req_combined,
        "interfaces": source_sections.get("interfaces", ""),
        "uds_frames": source_sections.get("uds_frames", ""),
        "notes": "",
        "logic_diagrams": [],
        "software_unit_design": source_sections.get("software_unit_design", ""),
        "unit_structure": source_sections.get("unit_structure", ""),
        "global_data": source_sections.get("global_data", ""),
        "interface_functions": source_sections.get("interface_functions", ""),
        "internal_functions": source_sections.get("internal_functions", ""),
        "function_table_rows": source_sections.get("function_table_rows", []),
        "global_vars": source_sections.get("global_vars", []),
        "static_vars": source_sections.get("static_vars", []),
        "macro_defs": source_sections.get("macro_defs", []),
        "calibration_params": source_sections.get("calibration_params", []),
        "function_details": source_sections.get("function_details", {}),
        "function_details_by_name": source_sections.get("function_details_by_name", {}),
        "call_map": source_sections.get("call_map", {}),
        "module_map": source_sections.get("module_map", {}),
        "req_map": req_map,
        "globals_info_map": source_sections.get("globals_info_map", {}),
        "common_macros": source_sections.get("common_macros", []),
        "type_defs": source_sections.get("type_defs", []),
        "param_defs": source_sections.get("param_defs", []),
        "version_defs": source_sections.get("version_defs", []),
        "globals_format_order": globals_order_list,
        "globals_format_sep": globals_format_sep,
        "globals_format_with_labels": globals_format_with_labels,
        "call_relation_mode": call_relation_mode,
        "show_mapping_evidence": bool(show_mapping_evidence),
        "logic_max_children": logic_max_children,
        "logic_max_grandchildren": logic_max_grandchildren,
        "logic_max_depth": logic_max_depth,
        "srs_texts": srs_texts,
        "sds_texts": sds_texts,
        "sds_doc_paths": sds_doc_paths,
    }
    impact_path = _run_impact_analysis_for_uds(
        source_root_path,
        os.getenv("UDS_CHANGED_FILES", ""),
    )
    if impact_path:
        notes_text = str(uds_payload.get("notes") or "").strip()
        uds_payload["notes"] = "\n".join([x for x in [notes_text, f"impact:{impact_path.name}"] if x])

    if ai_enable:
        rag_snippets: List[Dict[str, Any]] = []
        try:
            report_path = _resolve_report_dir(report_dir)
            kb = get_kb(report_path)
            rag_query = req_combined.strip()[:2000] or (source_sections.get("overview", "") or "")[:2000]
            if rag_query:
                fn_count = len(source_sections.get("function_details_by_name") or {}) if isinstance(source_sections, dict) else 0
                default_top_k = 12 if fn_count >= 300 else 10 if fn_count >= 120 else 8 if expand else 4
                # ⚠ `rag_top_k` 는 사용자 Form 입력이 그대로 검색 폭 → 프롬프트 크기가 되는
                #    유일한 축이라 상한을 건다(§6 후보 17). clamp 는
                #    `workflow.ai.clamp_rag_top_k` 단일 출처 — 소비처 3곳이 각자 조이면 갈라진다.
                from workflow.ai import clamp_rag_top_k
                use_top_k = clamp_rag_top_k(
                    rag_top_k if rag_top_k and rag_top_k > 0 else int(
                        getattr(config, "AGENT_RAG_TOP_K_DEFAULT", default_top_k)
                    ),
                    default=default_top_k,
                )
                use_categories = [str(c).strip() for c in re.split(r"[,\n;]+", rag_categories or "") if str(c).strip()]
                if not use_categories:
                    use_categories = ["uds", "requirements", "code", "constraints"]
                rag_rows = kb.search(
                    rag_query,
                    top_k=use_top_k,
                    categories=use_categories,
                )
                for row in rag_rows:
                    rag_snippets.append(
                        {
                            "title": row.get("error_raw") or "",
                            "category": row.get("category") or "",
                            "source_type": "rag",
                            "source_file": row.get("source_file") or "",
                            "excerpt": str(row.get("context") or row.get("fix") or "")[:1200],
                            "score": row.get("score"),
                        }
                    )
        except Exception:
            rag_snippets = []
        example_text = ""
        template_text = ""
        if ai_example_path:
            try:
                p = Path(ai_example_path).expanduser().resolve()
                if p.exists() and p.is_file():
                    example_text = _read_text_from_file(p)
            except Exception:
                example_text = ""
        if not example_text and template_file and template_file.filename and template_bytes:
            try:
                suffix = Path(template_file.filename).suffix.lower() or ".docx"
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(template_bytes)
                    template_text = _read_text_from_file(Path(tmp.name))
            except Exception:
                template_text = ""
            example_text = template_text or example_text
        if not example_text:
            try:
                ref_suds_path = Path(config.UDS_REF_SUDS_PATH)
                if ref_suds_path.exists() and ref_suds_path.is_file():
                    example_text = _read_text_from_file(ref_suds_path)
            except Exception:
                pass
        notes_text = ""
        if expand:
            doc_block = "\n\n".join(req_texts)[:40000]
            src_block = "\n\n".join(
                [
                    source_sections.get("overview", ""),
                    source_sections.get("interfaces", ""),
                    source_sections.get("uds_frames", ""),
                ]
            )
            notes_text = "\n\n".join([doc_block, src_block]).strip()
        # Gemini 호출 — 수 분. 이 저장소에서 이벤트 루프를 가장 오래 잡는 축이다.
        ai_sections = await _run_blocking(
            generate_uds_ai_sections,
            requirements_text=req_combined,
            source_sections=source_sections,
            notes_text=notes_text,
            logic_items=[],
            example_text=example_text,
            detailed=bool(True if expand else ai_detailed),
            rag_snippets=rag_snippets,
        )
        if ai_sections:
            uds_payload["ai_sections"] = ai_sections
    _enrich_function_quality_fields(uds_payload)
    quick_quality_gate = _compute_quick_quality_gate(uds_payload)

    out_dir = _local_uds_dir(report_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"uds_local_{ts}.docx"
    tpl_path = None
    template_applied = False
    if template_file and template_file.filename and template_bytes:
        suffix = Path(template_file.filename).suffix.lower() or ".docx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(template_bytes)
            tpl_tmp_path = Path(tmp.name)
            tpl_path = str(tpl_tmp_path)
        try:
            tpl_text = _read_text_from_file(Path(tpl_path))
            template_applied = "{{" in tpl_text and "}}" in tpl_text
        except Exception:
            template_applied = False
    if not tpl_path:
        # Use SUDS reference as default template for 4-level SUDS structure
        try:
            from config import UDS_REF_SUDS_PATH
            _ref_path = Path(UDS_REF_SUDS_PATH)
        except Exception:
            _ref_path = Path(__file__).resolve().parents[2] / "docs" / "(HDPDM01_SUDS) Software Unit Design Specification_v1.07_240213.docx"
        if _ref_path.exists():
            tpl_path = str(_ref_path)
    try:
        # Inject ai_config into payload for subprocess to use in function desc enhancement
        _uds_ai_cfg = _load_sts_ai_config()
        if _uds_ai_cfg:
            uds_payload["_gen_ai_config"] = _uds_ai_cfg
        await _run_blocking(_generate_docx_with_retry, tpl_path, uds_payload, out_path)
    except Exception as docx_exc:
        tb = traceback.format_exc()
        _logger.error("[UDS_GENERATE][%s] DOCX generation error:\n%s", req_id, tb)
        err_detail = str(docx_exc)[:800]
        if "timeout" in err_detail.lower():
            raise HTTPException(
                status_code=504,
                detail=f"UDS DOCX 생성 타임아웃: {err_detail}\n\n재시도하거나 AI를 비활성화하세요.",
            )
        raise HTTPException(
            status_code=500,
            detail=f"UDS DOCX 생성 실패: {err_detail}",
        )
    _write_uds_payload_sidecar(out_path, uds_payload)
    residual_tbd_path = _write_residual_tbd_report(out_path, (uds_payload.get("summary") or {}).get("mapping") or {})
    report_timeout_short = 3600 if bool(test_mode) else 300
    report_timeout_long = 14400 if bool(test_mode) else 600
    if bool(doc_only):
        quality_evaluation = _build_quality_evaluation(
            quick_quality_gate,
            quality_gate_path=None,
            accuracy_path=None,
            template_warning=template_warning,
            doc_only_mode=True,
        )
        # Quality DB recording (non-fatal)
        try:
            from workflow.quality.recorder import record_uds_run
            # project_root 는 sts/suts/sits 와 **같은 어휘**(source_root)로 넘긴다 —
            # recorder 가 이 값으로 scm_id 를 해결한다. 예전엔 UDS 만 아무것도 안
            # 넘겨서 DB 의 uds 행이 3/3 전부 NULL 이었고, 그래서 "이 프로젝트의 UDS
            # 품질" 을 물을 수단이 없었다.
            record_uds_run(
                quality_evaluation,
                project_root=str(source_root or ""),
                output_path=str(out_path),
            )
        except Exception:
            # non-fatal 은 유지하되 침묵은 금지 (608f849 참조).
            _logger.exception("[UDS_GENERATE][%s] quality record skipped (non-fatal)", req_id)
        _logger.info("[UDS_GENERATE][%s] done file=%s (doc_only)", req_id, out_path.name)
        return {
            "ok": True,
            "filename": out_path.name,
            "path": str(out_path),
            "template_applied": template_applied,
            "download_url": f"/api/local/uds/download/{out_path.name}?report_dir={report_dir}",
            "preview_url": "",
            "preview_path": "",
            "validation_path": "",
            "accuracy_path": "",
            "swcom_context_path": "",
            "swcom_diff_path": "",
            "confidence_path": "",
            "constraints_path": "",
            "quality_gate_path": "",
            "impact_path": str(impact_path) if impact_path else "",
            "residual_tbd_report_path": str(residual_tbd_path) if residual_tbd_path else "",
            "quick_quality_gate": quick_quality_gate,
            "quality_evaluation": quality_evaluation,
        }
    # 부가 보고서 (각각 _run_report_with_timeout 내부에서 timeout 관리됨)
    validation_path = out_path.with_suffix(".validation.md")
    ok_validation, _ = await _run_blocking(
        _run_report_with_timeout,
        lambda: generate_uds_validation_report(str(out_path), str(validation_path)),
        timeout_seconds=report_timeout_short,
        report_name="validation report",
    )
    if not ok_validation:
        validation_path = None
    accuracy_path = out_path.with_suffix(".accuracy.md")
    src_root = str(source_root_path) if source_root_path else ""
    ok_accuracy, _ = await _run_blocking(
        _run_report_with_timeout,
        lambda: generate_called_calling_accuracy_report(
            str(out_path),
            src_root,
            str(accuracy_path),
            relation_mode=str(call_relation_mode or "code"),
        ),
        timeout_seconds=report_timeout_long,
        report_name="accuracy report",
    )
    if not ok_accuracy:
        accuracy_path = None
    swcom_context_path = out_path.with_suffix(".swcom_context.md")
    ok_swcom, _ = await _run_blocking(
        _run_report_with_timeout,
        lambda: generate_swcom_context_report(str(out_path), str(swcom_context_path)),
        timeout_seconds=report_timeout_short,
        report_name="swcom context report",
    )
    if not ok_swcom:
        swcom_context_path = None
    swcom_diff_path = None
    confidence_path = out_path.with_suffix(".field_confidence.md")
    ok_confidence, _ = await _run_blocking(
        _run_report_with_timeout,
        lambda: generate_asil_related_confidence_report(
            uds_payload,
            str(confidence_path),
            str(out_path),
        ),
        timeout_seconds=report_timeout_short,
        report_name="ASIL/Related confidence report",
    )
    if not ok_confidence:
        confidence_path = None
    constraints_path = out_path.with_suffix(".constraints.md")
    ok_constraints, _ = await _run_blocking(
        _run_report_with_timeout,
        lambda: generate_uds_constraints_report(uds_payload, str(constraints_path)),
        timeout_seconds=report_timeout_short,
        report_name="constraints report",
    )
    if not ok_constraints:
        constraints_path = None
    quality_gate_path = out_path.with_suffix(".quality_gate.md")
    ok_quality_gate, _ = await _run_blocking(
        _run_report_with_timeout,
        lambda: generate_uds_field_quality_gate_report(str(out_path), str(quality_gate_path)),
        timeout_seconds=report_timeout_short,
        report_name="field quality gate report",
    )
    if not ok_quality_gate:
        quality_gate_path = None
    preview_html = await _run_blocking(generate_uds_preview_html, uds_payload)
    preview_path = out_path.with_suffix(".html")
    preview_path.write_text(preview_html, encoding="utf-8")
    quality_evaluation = _build_quality_evaluation(
        quick_quality_gate,
        quality_gate_path=quality_gate_path,
        accuracy_path=accuracy_path,
        template_warning=template_warning,
        doc_only_mode=False,
    )
    # Quality DB recording (non-fatal)
    try:
        from workflow.quality.recorder import record_uds_run
        # doc_only 경로와 같은 어휘(source_root) — recorder 가 scm_id 를 해결한다.
        record_uds_run(
            quality_evaluation,
            project_root=str(source_root or ""),
            output_path=str(out_path),
        )
    except Exception:
        # non-fatal 은 유지하되 침묵은 금지 (608f849 참조).
        _logger.exception("[UDS_GENERATE][%s] quality record skipped (non-fatal)", req_id)
    _logger.info("[UDS_GENERATE][%s] done file=%s", req_id, out_path.name)

    return {
        "ok": True,
        "filename": out_path.name,
        "path": str(out_path),
        "template_applied": template_applied,
        "download_url": f"/api/local/uds/download/{out_path.name}?report_dir={report_dir}",
        "preview_url": f"/api/local/uds/preview/{preview_path.name}?report_dir={report_dir}",
        "preview_path": str(preview_path),
        "validation_path": str(validation_path) if validation_path else "",
        "accuracy_path": str(accuracy_path) if accuracy_path else "",
        "swcom_context_path": str(swcom_context_path) if swcom_context_path else "",
        "swcom_diff_path": str(swcom_diff_path) if swcom_diff_path else "",
        "confidence_path": str(confidence_path) if confidence_path else "",
        "constraints_path": str(constraints_path) if constraints_path else "",
        "quality_gate_path": str(quality_gate_path) if quality_gate_path else "",
        "impact_path": str(impact_path) if impact_path else "",
        "residual_tbd_report_path": str(residual_tbd_path) if residual_tbd_path else "",
        "quick_quality_gate": quick_quality_gate,
        "quality_evaluation": quality_evaluation,
    }


@router.post("/api/local/uds/generate-async")
async def local_uds_generate_async(
    request: Request,
    source_root: str = Form(""),
    req_files: List[UploadFile] = File(default_factory=list),
    req_paths: str = Form(""),
    template_file: UploadFile = File(default=None),
    template_strict: bool = Form(False),
    component_list: UploadFile = File(default=None),
    logic_max_children: Optional[int] = Form(None),
    logic_max_grandchildren: Optional[int] = Form(None),
    logic_max_depth: Optional[int] = Form(None),
    globals_format_order: str = Form(""),
    globals_format_sep: str = Form(""),
    globals_format_with_labels: bool = Form(True),
    call_relation_mode: str = Form("code"),
    ai_enable: bool = Form(True),
    ai_example_path: str = Form(""),
    ai_detailed: bool = Form(True),
    expand: bool = Form(False),
    doc_only: bool = Form(False),
    test_mode: bool = Form(False),
    rag_top_k: Optional[int] = Form(None),
    rag_categories: str = Form(""),
    report_dir: str = Form(""),
    req_types: str = Form(""),
    show_mapping_evidence: bool = Form(False),
) -> Dict[str, Any]:
    """Non-blocking local UDS generation. Returns job_id for progress polling."""
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    reject_upload_in_cloudium(*(req_files or []), template_file, component_list)
    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="source_root(코드 루트)가 필요합니다.")
    req_paths_list = _parse_path_list(req_paths)
    has_req_upload = any((f and f.filename) for f in (req_files or []))
    if not has_req_upload and not req_paths_list:
        raise HTTPException(status_code=400, detail="SRS/SDS 요구사항 문서를 최소 1개 이상 제공해주세요.")

    job_id = uuid.uuid4().hex
    _set_progress(
        "local_uds", "local", "local",
        {"stage": "start", "percent": 1, "message": "Local UDS 생성 준비 중", "done": False, "error": ""},
        job_id=job_id,
    )

    req_file_paths: List[Path] = []
    for f in req_files:
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            req_file_paths.append(Path(tmp.name))

    template_bytes: Optional[bytes] = None
    if template_file and template_file.filename:
        try:
            template_bytes = await template_file.read()
        except Exception:
            template_bytes = None

    comp_map: Dict[str, Dict[str, str]] = {}
    if component_list and component_list.filename:
        tmp = _write_upload_to_temp(component_list, ".json")
        if tmp:
            try:
                comp_map = _parse_component_map_file(tmp)
            except Exception:
                comp_map = {}
            finally:
                try:
                    tmp.unlink(missing_ok=True)
                except Exception:
                    pass

    type_list = [t.strip().lower() for t in req_types.split(",") if t.strip()] if req_types else []

    def _worker():
        try:
            _set_progress(
                "local_uds", "local", "local",
                {"stage": "source_analysis", "percent": 10, "message": "소스 코드 분석 중"},
                job_id=job_id,
            )
            # SDS 파티션 맵 로드 (Related ID + ASIL 전파)
            _async_sds_pmap: Dict[str, Dict[str, str]] = {}
            for rp in req_paths_list:
                if rp.lower().endswith(".docx") and is_sds_filename(rp):
                    try:
                        from report_gen.requirements import _extract_sds_partition_map
                        _async_sds_pmap.update(_extract_sds_partition_map(rp))
                    except Exception:
                        pass
            source_sections = generate_uds_source_sections(
                str(source_root_path),
                component_map=comp_map if comp_map else None,
                sds_partition_map=_async_sds_pmap if _async_sds_pmap else None,
            ) if source_root_path and source_root_path.exists() else {}

            _set_progress(
                "local_uds", "local", "local",
                {"stage": "requirements", "percent": 30, "message": "요구사항 문서 처리 중"},
                job_id=job_id,
            )
            req_texts: List[str] = []
            srs_texts: List[str] = []
            sds_texts: List[str] = []
            sds_doc_paths: List[str] = []
            req_doc_paths: List[str] = []
            for idx, fp in enumerate(req_file_paths):
                try:
                    text = _read_text_from_file(fp)
                except Exception:
                    text = ""
                if fp.suffix.lower() == ".docx":
                    req_doc_paths.append(str(fp))
                if text:
                    req_texts.append(text.strip())
                    ftype = type_list[idx] if idx < len(type_list) else ""
                    if ftype == "srs":
                        srs_texts.append(text.strip())
                    elif ftype == "sds":
                        sds_texts.append(text.strip())

            _async_doc_skips: List[str] = []
            for path_str in req_paths_list:
                p, text, reason = read_requirement_doc(path_str, allow=_is_allowed_req_doc)
                if reason:
                    _async_doc_skips.append(reason)
                    continue
                if p and text:
                    req_texts.append(text)
                    if p.suffix.lower() == ".docx":
                        req_doc_paths.append(str(p))
                    fname_lower = p.name.lower()
                    if is_srs_filename(fname_lower):
                        srs_texts.append(text)
                    elif is_sds_filename(fname_lower):
                        sds_texts.append(text)
                        if p.suffix.lower() in {".docx", ".doc"}:
                            sds_doc_paths.append(str(p))
            if _async_doc_skips:
                _logger.warning("[UDS-async] 요구사항 문서 %d건 탈락: %s",
                                len(_async_doc_skips), "; ".join(_async_doc_skips[:5]))

            source_sections, req_doc_paths, sds_doc_paths = _enrich_source_sections_with_docs(
                source_sections,
                req_doc_paths=req_doc_paths,
                sds_doc_paths=sds_doc_paths,
            )

            req_from_docs = generate_uds_requirements_from_docs(req_texts) if req_texts else ""
            req_map = _build_req_map_from_doc_paths(req_doc_paths, req_texts) if req_texts or req_doc_paths else {}
            req_source = source_sections.get("requirements", "")
            req_combined = "\n".join([req_from_docs.strip(), req_source.strip()]).strip() if req_from_docs and req_source else (req_from_docs or req_source)

            globals_order_list = [x.strip() for x in re.split(r"[,\|;]+", globals_format_order or "") if x.strip()]
            _project_name_val, _module_name_val = _derive_project_module_names(source_root_path)
            _source_docs: List[str] = []
            for _p in req_doc_paths:
                try:
                    _source_docs.append(Path(_p).name)
                except Exception:
                    continue
            uds_payload = {
                "job_url": "local",
                "build_number": "",
                "project_name": _project_name_val,
                "module_name": _module_name_val,
                "source_docs": _source_docs,
                "summary": {},
                "overview": source_sections.get("overview", ""),
                "requirements": req_combined,
                "interfaces": source_sections.get("interfaces", ""),
                "uds_frames": source_sections.get("uds_frames", ""),
                "notes": "",
                "logic_diagrams": [],
                "software_unit_design": source_sections.get("software_unit_design", ""),
                "unit_structure": source_sections.get("unit_structure", ""),
                "global_data": source_sections.get("global_data", ""),
                "interface_functions": source_sections.get("interface_functions", ""),
                "internal_functions": source_sections.get("internal_functions", ""),
                "function_table_rows": source_sections.get("function_table_rows", []),
                "global_vars": source_sections.get("global_vars", []),
                "static_vars": source_sections.get("static_vars", []),
                "macro_defs": source_sections.get("macro_defs", []),
                "calibration_params": source_sections.get("calibration_params", []),
                "function_details": source_sections.get("function_details", {}),
                "function_details_by_name": source_sections.get("function_details_by_name", {}),
                "call_map": source_sections.get("call_map", {}),
                "module_map": source_sections.get("module_map", {}),
                "req_map": req_map,
                "globals_info_map": source_sections.get("globals_info_map", {}),
                "common_macros": source_sections.get("common_macros", []),
                "type_defs": source_sections.get("type_defs", []),
                "param_defs": source_sections.get("param_defs", []),
                "version_defs": source_sections.get("version_defs", []),
                "globals_format_order": globals_order_list,
                "globals_format_sep": globals_format_sep,
                "globals_format_with_labels": globals_format_with_labels,
                "call_relation_mode": call_relation_mode,
                "show_mapping_evidence": bool(show_mapping_evidence),
                "logic_max_children": logic_max_children,
                "logic_max_grandchildren": logic_max_grandchildren,
                "logic_max_depth": logic_max_depth,
                "srs_texts": srs_texts,
                "sds_texts": sds_texts,
                "sds_doc_paths": sds_doc_paths,
            }

            _set_progress(
                "local_uds", "local", "local",
                {"stage": "docx_generation", "percent": 50, "message": "DOCX 생성 중"},
                job_id=job_id,
            )
            _enrich_function_quality_fields(uds_payload)
            out_dir = _local_uds_dir(report_dir)
            out_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = out_dir / f"uds_local_{ts}.docx"

            tpl_path = None
            if template_bytes:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
                    tmp.write(template_bytes)
                    tpl_path = str(Path(tmp.name))

            _uds_ai_cfg = _load_sts_ai_config()
            if _uds_ai_cfg:
                uds_payload["_gen_ai_config"] = _uds_ai_cfg
            _generate_docx_with_retry(tpl_path, uds_payload, out_path)
            _write_uds_payload_sidecar(out_path, uds_payload)
            residual_tbd_path = _write_residual_tbd_report(out_path, (uds_payload.get("summary") or {}).get("mapping") or {})

            _set_progress(
                "local_uds", "local", "local",
                {"stage": "reports", "percent": 80, "message": "리포트 생성 중"},
                job_id=job_id,
            )
            report_timeout = 3600 if bool(test_mode) else 300
            quick_qg = _compute_quick_quality_gate(uds_payload)
            if not bool(doc_only):
                _run_report_with_timeout(
                    lambda: generate_uds_validation_report(str(out_path), str(out_path.with_suffix(".validation.md"))),
                    timeout_seconds=report_timeout, report_name="validation",
                )
                _run_report_with_timeout(
                    lambda: generate_uds_field_quality_gate_report(str(out_path), str(out_path.with_suffix(".quality_gate.md"))),
                    timeout_seconds=report_timeout, report_name="quality gate",
                )

            # Quality DB recording (non-fatal)
            try:
                from workflow.quality.recorder import record_uds_run
                # 동기 경로와 같은 어휘(source_root) — recorder 가 scm_id 를 해결한다.
                record_uds_run(
                    quick_qg,
                    project_root=str(source_root or ""),
                    output_path=str(out_path),
                )
            except Exception:
                # non-fatal 은 유지하되 침묵은 금지 (608f849 참조).
                _logger.exception("UDS quality record skipped (non-fatal)")

            _set_progress(
                "local_uds", "local", "local",
                {
                    "stage": "done", "percent": 100, "message": "완료",
                    "done": True, "error": "",
                    "result": {
                        "ok": True,
                        "filename": out_path.name,
                        "path": str(out_path),
                        "download_url": f"/api/local/uds/download/{out_path.name}?report_dir={report_dir}",
                        "residual_tbd_report_path": str(residual_tbd_path) if residual_tbd_path else "",
                        "quick_quality_gate": quick_qg,
                    },
                },
                job_id=job_id,
            )
            _logger.info("[UDS_ASYNC_LOCAL][%s] done file=%s", job_id, out_path.name)

        except Exception as exc:
            tb = traceback.format_exc()
            _logger.error("[UDS_ASYNC_LOCAL][%s] FAILED: %s\n%s", job_id, str(exc)[:500], tb)
            _set_progress(
                "local_uds", "local", "local",
                {"stage": "error", "percent": 100, "message": f"실패: {str(exc)[:300]}", "done": True, "error": str(exc)[:500]},
                job_id=job_id,
            )

    threading.Thread(target=wrap_with_user(_worker), daemon=True).start()
    return {"ok": True, "job_id": job_id}


@router.get("/api/local/uds/progress")
def local_uds_progress(job_id: str = "") -> Dict[str, Any]:
    data = _get_progress("local_uds", "local", "local", job_id)
    return {"ok": bool(data), "progress": data}


@router.get("/api/local/uds/download/{filename}")
def local_uds_download(filename: str, report_dir: Optional[str] = None) -> FileResponse:
    file_path = _resolve_local_uds_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="uds report not found")
    media = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    return FileResponse(str(file_path), filename=file_path.name, media_type=media)


@router.get("/api/local/uds/preview/{filename}")
def local_uds_preview(filename: str, report_dir: Optional[str] = None) -> HTMLResponse:
    file_path = _resolve_local_uds_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="uds preview not found")
    return HTMLResponse(file_path.read_text(encoding="utf-8", errors="ignore"))


@router.get("/api/local/uds/files")
def local_uds_files(report_dir: Optional[str] = None) -> List[Dict[str, Any]]:
    base = _resolve_report_dir(report_dir)
    uds_dir = _local_uds_dir(base)
    if not uds_dir.exists():
        return []
    rows: List[Dict[str, Any]] = []
    for p in uds_dir.glob("*.docx"):
        if not p.is_file():
            continue
        rows.append(
            {
                "filename": p.name,
                "path": str(p),
                "size_mb": round(p.stat().st_size / (1024 * 1024), 2),
                "mtime": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
                "download_url": f"/api/local/uds/download/{p.name}?report_dir={report_dir or ''}",
                "preview_url": f"/api/local/uds/preview/{p.with_suffix('.html').name}?report_dir={report_dir or ''}",
            }
        )
    rows.sort(key=lambda x: x.get("mtime") or "", reverse=True)
    return rows


@router.get("/api/local/uds/view/{filename}")
def local_uds_view(
    filename: str,
    report_dir: Optional[str] = None,
    q: str = Query(default=""),
    swcom: str = Query(default="all"),
    asil: str = Query(default="all"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    trace_q: str = Query(default=""),
    trace_page: int = Query(default=1, ge=1),
    trace_page_size: int = Query(default=100, ge=1, le=500),
) -> Dict[str, Any]:
    if Path(filename).suffix.lower() != ".docx":
        raise HTTPException(status_code=400, detail="filename must be .docx")
    docx_path = _resolve_local_uds_path(report_dir, filename)
    if not docx_path.exists():
        raise HTTPException(status_code=404, detail="uds report not found")
    accuracy_path = docx_path.with_suffix(".accuracy.md")
    quality_gate_path = docx_path.with_suffix(".quality_gate.md")
    payload = _get_uds_view_payload_cached(
        docx_path,
        accuracy_path if accuracy_path.exists() else None,
        quality_gate_path if quality_gate_path.exists() else None,
    )
    payload = _apply_uds_view_filters(
        payload,
        q=q,
        swcom=swcom,
        asil=asil,
        page=page,
        page_size=page_size,
        trace_q=trace_q,
        trace_page=trace_page,
        trace_page_size=trace_page_size,
    )
    preview_name = docx_path.with_suffix(".html").name
    payload["download_url"] = f"/api/local/uds/download/{docx_path.name}?report_dir={report_dir or ''}"
    payload["preview_url"] = f"/api/local/uds/preview/{preview_name}?report_dir={report_dir or ''}"
    payload["accuracy_path"] = str(accuracy_path) if accuracy_path.exists() else ""
    payload["quality_gate_path"] = str(quality_gate_path) if quality_gate_path.exists() else ""
    residual_tbd_path = docx_path.with_suffix(".residual_tbd.md")
    payload["residual_tbd_report_path"] = str(residual_tbd_path) if residual_tbd_path.exists() else ""
    return payload


@router.get("/api/local/uds/view-by-path")
def local_uds_view_by_path(
    docx_path: str,
    q: str = Query(default=""),
    swcom: str = Query(default="all"),
    asil: str = Query(default="all"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    trace_q: str = Query(default=""),
    trace_page: int = Query(default=1, ge=1),
    trace_page_size: int = Query(default=100, ge=1, le=500),
) -> Dict[str, Any]:
    target = Path(str(docx_path or "")).expanduser().resolve()
    if target.suffix.lower() != ".docx":
        raise HTTPException(status_code=400, detail="docx_path must be .docx")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="uds report not found")
    accuracy_path = target.with_suffix(".accuracy.md")
    quality_gate_path = target.with_suffix(".quality_gate.md")
    payload = _get_uds_view_payload_cached(
        target,
        accuracy_path if accuracy_path.exists() else None,
        quality_gate_path if quality_gate_path.exists() else None,
    )
    payload = _apply_uds_view_filters(
        payload,
        q=q,
        swcom=swcom,
        asil=asil,
        page=page,
        page_size=page_size,
        trace_q=trace_q,
        trace_page=trace_page,
        trace_page_size=trace_page_size,
    )
    payload["download_url"] = ""
    payload["preview_url"] = ""
    payload["accuracy_path"] = str(accuracy_path) if accuracy_path.exists() else ""
    payload["quality_gate_path"] = str(quality_gate_path) if quality_gate_path.exists() else ""
    residual_tbd_path = target.with_suffix(".residual_tbd.md")
    payload["residual_tbd_report_path"] = str(residual_tbd_path) if residual_tbd_path.exists() else ""
    return payload


@router.post("/api/local/traceability")
def local_traceability(
    request: Request,
    source_root: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    report_dir: str = Form(""),
) -> Dict[str, Any]:
    """Build full traceability matrix: SRS -> Functions -> Test Cases."""
    from report_gen.requirements import (
        _extract_sds_partition_map,
        _normalize_req_id,
        annotate_sds_coarse,
        build_sds_component_maps,
    )
    from sts_generator import (
        map_requirements_to_functions,
        parse_srs_docx_tables,
    )

    srs_docx: Optional[str] = None
    if srs_path:
        p = Path(srs_path).expanduser().resolve()
        if p.exists() and p.is_file():
            srs_docx = str(p)

    # Parse SDS component mapping (V-Model 설계 계층)
    # 추적 정화: sds_req_to_comps=컴포넌트+함수 전체(브리지/함수 표시용), sds_req_to_design_comps=
    # 실 SwCom/모듈만(SDS 밴드 집계용). 판정은 report_gen.requirements.build_sds_component_maps
    # 단일 출처 — Jenkins 경로(jenkins.py sds extract)와 같은 함수를 쓰므로 모드 간 SDS 컴포넌트
    # 수가 갈릴 수 없다(예전엔 두 곳이 복제라 한쪽만 고쳐지면 갈렸다).
    sds_req_to_comps: Dict[str, List[str]] = {}
    sds_req_to_design_comps: Dict[str, List[str]] = {}
    sds_req_to_folded_comps: Dict[str, List[str]] = {}
    sds_req_to_element_comps: Dict[str, List[str]] = {}
    # 아래 map_requirements_to_functions 의 폴백 출처로도 쓴다 — 안 넘기면 저장소 docs/
    # 글롭(프로젝트 무관)이 대신한다.
    sds_partition_map: Optional[Dict[str, Any]] = None
    if sds_path:
        sds_p = Path(sds_path).expanduser().resolve()
        if not is_under_any(sds_p, [repo_root, sds_p.parent.resolve()]):
            raise HTTPException(status_code=403, detail="SDS 경로 접근이 허용되지 않습니다")
        if sds_p.exists() and sds_p.is_file():
            partition_map = _extract_sds_partition_map(str(sds_p))
            sds_partition_map = partition_map or None
            _sds_maps = build_sds_component_maps(partition_map)
            sds_req_to_comps = _sds_maps["req_to_comps"]
            sds_req_to_design_comps = _sds_maps["req_to_design_comps"]
            sds_req_to_folded_comps = _sds_maps["req_to_folded_comps"]
            sds_req_to_element_comps = _sds_maps["req_to_element_comps"]

    # Parse requirements
    reqs: List[Dict[str, Any]] = []
    if srs_docx:
        reqs = parse_srs_docx_tables(srs_docx)

    if not reqs:
        raise HTTPException(status_code=400, detail="SRS 문서에서 요구사항을 추출할 수 없습니다.")

    # Parse source for function details
    function_details: Dict[str, Any] = {}
    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if source_root_path and source_root_path.exists() and source_root_path.is_dir():
        try:
            sections = _get_source_sections_cached(str(source_root_path))
            function_details = sections.get("function_details", {})
        except Exception:
            pass

    # Map requirements to functions
    req_to_fids = map_requirements_to_functions(reqs, function_details,
                                                sds_map=sds_partition_map)

    # Keyword-based fallback mapping if related fields are TBD
    _kw_map = {
        "battery": ["apiin", "apiout", "monitor_adc", "drvin", "vsup"],
        "voltage": ["apiin", "apiout", "monitor_adc", "drvin", "vsup"],
        "buzzer": ["buzzer"],
        "door": ["door", "motor"],
        "motor": ["door", "motor"],
        "latch": ["door"],
        "lin": ["lin", "apiin", "apiout"],
        "signal": ["lin", "apiin", "apiout"],
        "eeprom": ["eeprom"],
        "memory": ["eeprom"],
        "sleep": ["sleep", "wake"],
        "wakeup": ["sleep", "wake"],
        "diag": ["diag", "uds"],
        "diagnostic": ["diag", "uds"],
        "option": ["option"],
        "init": ["init", "main", "sysctrl"],
        "position": ["motor", "speed", "direction"],
        "sensor": ["motor", "speed", "direction"],
        "pwm": ["pwm"],
        "error": ["diag", "error"],
        "close": ["door", "motor"],
        "open": ["door", "motor"],
    }
    mapped_count_before = sum(1 for v in req_to_fids.values() if v)
    if function_details and mapped_count_before < len(reqs) * 0.3:
        for r in reqs:
            rid = r["id"]
            if req_to_fids.get(rid):
                continue
            desc = (r.get("description", "") + " " + r.get("name", "")).lower()
            keywords = set()
            for kw, fns in _kw_map.items():
                if kw in desc:
                    keywords.update(fns)
            if keywords:
                for fid, info in function_details.items():
                    if not isinstance(info, dict):
                        continue
                    fname = str(info.get("name", "")).lower()
                    if any(k in fname for k in keywords):
                        fid_list = req_to_fids.setdefault(rid, [])
                        if fid not in fid_list:
                            fid_list.append(fid)

    # Load STS test cases if available
    sts_test_cases: List[Dict[str, Any]] = []
    base = _resolve_report_dir(report_dir)
    sts_dir = base / "sts"
    sts_file_name = None
    if sts_dir.exists():
        latest_sts = _find_latest_excel_file(sts_dir)
        if latest_sts:
            sts_file_name = latest_sts.name
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(latest_sts), read_only=True, data_only=True)
                if "3.SW Integration Test Spec" in wb.sheetnames:
                    ws = wb["3.SW Integration Test Spec"]
                    for r in range(7, (ws.max_row or 7) + 1):
                        tc_id = ws.cell(row=r, column=2).value
                        if tc_id:
                            sts_test_cases.append({
                                "tc_id": str(tc_id),
                                "title": str(ws.cell(row=r, column=3).value or ""),
                                "method": str(ws.cell(row=r, column=6).value or ""),
                                "srs_id": str(ws.cell(row=r, column=13).value or ""),
                            })
                wb.close()
            except Exception:
                pass

    # Load SUTS test cases if available
    suts_test_cases: List[Dict[str, Any]] = []
    suts_dir = base / "suts"
    suts_file_name = None
    if suts_dir.exists():
        latest_suts = _find_latest_excel_file(suts_dir)
        if latest_suts:
            suts_file_name = latest_suts.name
            try:
                import openpyxl as _xl
                swb = _xl.load_workbook(str(latest_suts), read_only=True, data_only=True)
                if "2.SW Unit Test Spec" in swb.sheetnames:
                    sws = swb["2.SW Unit Test Spec"]
                    for sr in range(7, (sws.max_row or 7) + 1):
                        tc_id = sws.cell(row=sr, column=3).value
                        if tc_id and str(tc_id).startswith("SwUTC"):
                            related = sws.cell(row=sr, column=149).value or ""
                            n_inp = sum(1 for c in range(14, 63) if sws.cell(row=sr, column=c).value is not None)
                            n_out = sum(1 for c in range(63, 149) if sws.cell(row=sr, column=c).value is not None)
                            suts_test_cases.append({
                                "tc_id": str(tc_id),
                                "name": str(sws.cell(row=sr, column=4).value or ""),
                                "related_fid": str(related),
                                "gen_method": str(sws.cell(row=sr, column=12).value or ""),
                                "input_count": n_inp,
                                "output_count": n_out,
                            })
                swb.close()
            except Exception:
                pass

    # Build fid→suts_tc lookup
    fid_to_suts: Dict[str, List[Dict[str, Any]]] = {}
    for stc in suts_test_cases:
        fid = stc.get("related_fid", "")
        if fid:
            fid_to_suts.setdefault(fid, []).append(stc)

    # Build traceability rows
    rows: List[Dict[str, Any]] = []
    for r in reqs:
        rid = r["id"]
        fids = req_to_fids.get(rid, [])
        func_names = []
        # 전체 fid 순회 — 과거 fids[:10] 절단은 UDS 함수를 최대 ~188개 silent 누락시켜
        # source_ids 기반 트리의 단위시험 미연결/orphan SUTS 계산을 거짓으로 만들었다
        # (라운드 재검증 W1: 외곽 슬라이스만 제거됐고 이 내부 루프 절단이 잔존했음).
        # 아래 suts_tcs_for_req는 이미 전체 fids를 쓰므로 비대칭도 해소된다.
        for fid in fids:
            info = function_details.get(fid)
            if isinstance(info, dict):
                func_names.append(info.get("name", fid))

        # STS TC 매칭: 콤마 분리 + ID 정규화 (BUG-1, BUG-6 수정)
        norm_rid = _normalize_req_id(rid)
        sts_tcs = [
            tc for tc in sts_test_cases
            if norm_rid in [_normalize_req_id(s.strip()) for s in (tc.get("srs_id") or "").split(",")]
        ]
        suts_tcs_for_req: List[Dict[str, Any]] = []
        for fid in fids:
            suts_tcs_for_req.extend(fid_to_suts.get(fid, []))

        # SDS 컴포넌트 매핑 (V-Model 아키텍처 설계 계층) — 정화: 컴포넌트만 표시, 함수는 분리.
        sds_comps_all = sds_req_to_comps.get(norm_rid, [])
        sds_comps = sds_req_to_design_comps.get(norm_rid, [])  # 실 SwCom/모듈만(함수 fan-out 제외)
        _scset = set(sds_comps)
        # canonical 접기로 sds_comps 에서 사라진 원 키 — 함께 빼지 않으면 함수로 이중 계상된다.
        _folded = set(sds_req_to_folded_comps.get(norm_rid, []))
        # 설계 요소(설계ID·상태명·표행·heading)는 함수가 아니다 — 함께 빼지 않으면 차집합이
        # "컴포넌트 아닌 것 = 함수" 로 접어 UI 가 상태명·목차 줄을 '멤버 함수'로 표시한다.
        # Jenkins 경로(requirements.py sds_elem_list)와 같은 분리다.
        _elem = set(sds_req_to_element_comps.get(norm_rid, []))
        sds_elems = [c for c in sds_comps_all if c in _elem and c not in _scset]
        sds_funcs = [c for c in sds_comps_all
                     if c not in _scset and c not in _folded and c not in _elem]  # 함수만

        # tests 배열 통합 (Jenkins generate_uds_traceability_matrix 형식)
        tests: List[Dict[str, Any]] = []
        for tc in sts_tcs:
            tests.append({
                "requirement_id": rid,
                "testcase": tc["tc_id"],
                "result": "mapped",
                "source": "STS",
                "confidence": "exact",
                "unit": "",
                "report": "",
            })
        for tc in suts_tcs_for_req:
            tests.append({
                "requirement_id": rid,
                "testcase": tc["tc_id"],
                "result": "mapped",
                "source": "SUTS",
                "confidence": "exact",
                "unit": tc.get("related_fid", ""),
                "report": "",
            })
        test_ids = [t["testcase"] for t in tests]

        # Jenkins 경로(generate_uds_traceability_matrix)와 동일한 행 구조
        # status는 행에 포함하지 않음 — 프론트엔드 deriveStatus()가 단일 판정
        rows.append({
            "requirement_id": rid,
            "sds_components": sds_comps,        # 실 설계 컴포넌트만(함수 fan-out 제외)
            "sds_functions": sds_funcs,         # 인터페이스 함수(분리, 투명성·브리지) — 함수만
            "sds_design_elements": sds_elems,   # 설계ID·상태명·표행 — 함수 아님(Jenkins 경로 lockstep)
            # 전체 함수 유지 — 과거 [:10] 절단은 UDS 함수를 최대 ~188개 silent 누락시켜
            # 트리의 단위시험 미연결/orphan SUTS 계산을 거짓으로 만들었다(deep-analyze).
            # Jenkins 경로(generate_uds_traceability_matrix)는 전량 싣는다 — 표시는 프론트가 스크롤로 제한.
            "source_ids": func_names,
            "tests": tests,
            "test_ids": test_ids,
            "test_count": len(tests),
            "pass_count": 0,
            "fail_count": 0,
            "confidence": "exact" if tests else None,
            # Local 전용 추가 필드 (하위 호환)
            "req_name": r.get("name", ""),
            "req_type": r.get("req_type", ""),
            "asil": r.get("asil", ""),
        })

    # Summary — Jenkins 경로(generate_uds_traceability_matrix)와 동일 키 사용
    # deriveStatus 동일 로직: 설계(SDS or UDS) + 검증(any test) = covered
    def _derive(r):
        # 추적 정화: sds_functions 포함 — 함수로만 추적된 요구사항이 설계 없음으로 회귀하지 않게
        # (프론트 DESIGN_FIELDS·Jenkins _cache_trace_summary와 lockstep). 이게 실제 요약 집계 사이트.
        # jenkins _cache_trace_summary / 프론트 DESIGN_FIELDS와 동일 9필드 디텍터(lockstep 실현).
        # sds_design_elements: 설계ID·상태명 등. sds_functions 를 '함수만'으로 정화하면서
        # 분리됐다 — 빼면 14행(HDPDM01 실측)이 uncovered 로 회귀한다.
        has_d = (bool(r.get("sds_components")) or bool(r.get("sds_functions"))
                 or bool(r.get("sds_design_elements"))
                 or bool(r.get("source_ids")) or bool(r.get("hsis_signals"))
                 or bool(r.get("functions")) or bool(r.get("mapping"))
                 or bool(r.get("sds")) or bool(r.get("source_mapping")))
        has_t = bool(r.get("test_count"))
        # 비기능/안전 요구(SwNTR/SwNTSR)는 설계 없이 시험으로 직접 검증(결정1, jenkins/프론트 lockstep).
        _rid = str(r.get("requirement_id") or "").upper()
        is_nf = _rid.startswith(("SWNTR", "SWNTSR", "SYNTR", "SYNTSR"))  # RAW 철자 — Sy* prefix도 인정
        if has_t and (has_d or is_nf):
            return "covered"
        if has_d or has_t:
            return "partial"
        return "uncovered"

    total = len(rows)
    covered = sum(1 for r in rows if _derive(r) == "covered")
    partial = sum(1 for r in rows if _derive(r) == "partial")
    uncovered = sum(1 for r in rows if _derive(r) == "uncovered")
    safety_total = sum(1 for r in rows if r.get("asil") and r["asil"].upper() not in ("QM", "TBD", ""))
    safety_covered = sum(1 for r in rows if _derive(r) == "covered" and r.get("asil") and r["asil"].upper() not in ("QM", "TBD", ""))
    mapped_sds_count = sum(1 for r in rows if r.get("sds_components"))
    # 거친 입도 — Jenkins 경로(generate_uds_traceability_matrix)와 **같은 헬퍼**로 센다.
    # 각자 세면 같은 문서가 모드에 따라 다른 값을 내고, local 만 빠지면 이 신호가 조용히
    # 사라진다(프론트 칩은 summary.sds_coarse_count 부재를 0으로 읽어 아무것도 안 띄운다).
    _sds_total_comps, sds_coarse_count = annotate_sds_coarse(rows)
    mapped_source_count = sum(1 for r in rows if r.get("source_ids"))
    mapped_test_count = sum(1 for r in rows if r.get("test_count"))
    total_tests = sum(r.get("test_count", 0) for r in rows)

    # source별 테스트 건수 (Jenkins source_stats와 동일)
    source_stats: Dict[str, int] = {}
    for r in rows:
        for t in r.get("tests", []):
            src = t.get("source", "unknown")
            source_stats[src] = source_stats.get(src, 0) + 1

    type_dist: Dict[str, int] = {}
    for r in rows:
        t = r.get("req_type") or "OTHER"
        type_dist[t] = type_dist.get(t, 0) + 1

    # SUTS-specific coverage
    total_suts_fns = len(suts_test_cases)
    fns_with_suts = sum(1 for fid in function_details if fid in fid_to_suts)

    return {
        "ok": True,
        "total_requirements": total,
        "summary": {
            # Jenkins 호환 키
            "requirement_count": total,
            "mapped_sds_count": mapped_sds_count,
            "mapped_source_count": mapped_source_count,
            "mapped_test_count": mapped_test_count,
            "total_tests": total_tests,
            "total_pass": 0,
            "total_fail": 0,
            "source_stats": source_stats,
            # Local 추가 키
            "covered": covered,
            "partial": partial,
            "uncovered": uncovered,
            "coverage_pct": round(covered / max(total, 1) * 100, 1),
            "full_coverage_pct": round((covered + partial) / max(total, 1) * 100, 1),
            "safety_total": safety_total,
            "safety_covered": safety_covered,
            "safety_pct": round(safety_covered / max(safety_total, 1) * 100, 1),
            # 정화: 실 설계 컴포넌트 distinct 수(함수 fan-out 제외) — Jenkins 경로와 동일 의미.
            # **행 기준**으로 센다(annotate_sds_coarse 반환값). 예전엔 sds_req_to_design_comps
            # 전량을 셌는데, 거기엔 매트릭스 행이 없는 요구(SRS 에 없는 ID)의 컴포넌트까지 들어가
            # 거친 입도 임계의 분모(행 기준)와 표시 총수가 어긋났다 — 툴팁이 "33개 중 40%"라
            # 적으면서 실제로는 다른 분모로 판정하는 상태. Jenkins 경로도 행 기준이라 이게 lockstep.
            "total_sds_components": _sds_total_comps,
            # 거친 입도 요구 수 — Jenkins summary 와 같은 키. 없으면 프론트 칩이 침묵한다.
            "sds_coarse_count": sds_coarse_count,
            "total_functions": len(function_details),
            "total_sts_test_cases": len(sts_test_cases),
            "total_suts_test_cases": total_suts_fns,
            "suts_function_coverage": fns_with_suts,
            "suts_function_coverage_pct": round(fns_with_suts / max(len(function_details), 1) * 100, 1),
            "type_distribution": type_dist,
        },
        "has_sds_mapping": any(r.get("sds_components") for r in rows),
        "has_source_mapping": any(r.get("source_ids") for r in rows),
        "has_tests": any(r.get("test_count") for r in rows),
        "rows": rows,
        "sts_file": sts_file_name,
        "suts_file": suts_file_name,
    }


@router.post("/api/local/sts/generate")
async def local_sts_generate(
    request: Request,
    source_root: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    stp_path: str = Form(""),
    hsis_path: str = Form(""),
    req_paths: str = Form(""),
    req_files: List[UploadFile] = File(default_factory=list),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_tc_per_req: int = Form(5),
    report_dir: str = Form(""),
) -> Dict[str, Any]:
    """Generate STS (Software Test Specification) Excel from SRS + source code."""
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    from sts_generator import generate_sts
    reject_upload_in_cloudium(*(req_files or []))

    req_id = (request.headers.get("x-req-id") or "").strip() or f"sts-gen-{int(time.time() * 1000)}"
    _logger.info("[STS_GENERATE][%s] start source_root=%s", req_id, source_root)

    # Resolve SRS path
    srs_docx_path: Optional[str] = None
    if srs_path:
        p = Path(srs_path).expanduser().resolve()
        if p.exists() and p.is_file():
            srs_docx_path = str(p)

    # Collect requirement text from paths/uploads
    req_paths_list = _parse_path_list(req_paths)
    req_texts: List[str] = []
    req_doc_paths: List[str] = []
    sds_doc_paths: List[str] = []

    # 탈락 사유를 버리지 않는다 — 권한 없음/본문 0자가 '문서 미지정'과 구분된다.
    doc_skips: List[str] = []
    for path_str in req_paths_list:
        p, text, reason = read_requirement_doc(path_str, allow=_is_allowed_req_doc)
        if reason:
            doc_skips.append(reason)
            continue
        if not p or not text:
            continue
        req_texts.append(text)
        if p.suffix.lower() == ".docx":
            req_doc_paths.append(str(p))
            if is_sds_filename(p.name):
                sds_doc_paths.append(str(p))
        if not srs_docx_path and is_srs_filename(p.name) and p.suffix.lower() == ".docx":
            srs_docx_path = str(p)
    if doc_skips:
        _logger.warning("[STS] 요구사항 문서 %d건 탈락: %s", len(doc_skips), "; ".join(doc_skips[:5]))

    for f in (req_files or []):
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            text = _read_text_from_file(tmp_path)
            if text:
                req_texts.append(text.strip())
                if tmp_path.suffix.lower() == ".docx":
                    req_doc_paths.append(str(tmp_path))
                    if is_sds_filename(f.filename):
                        sds_doc_paths.append(str(tmp_path))
                if not srs_docx_path and is_srs_filename(f.filename) and suffix == ".docx":
                    srs_docx_path = str(tmp_path)
        except Exception:
            pass

    # Fallback: auto-discover SRS from docs/ if not yet resolved
    srs_docx_path = _doc_or_discovered(
        srs_docx_path, bool(req_paths_list or req_files), _discover_srs_docx,
        label="SRS", tag=f"[STS_GENERATE][{req_id}] ")

    if not req_texts and not srs_docx_path:
        raise HTTPException(status_code=400, detail="SRS 문서를 최소 1개 이상 제공해주세요.")

    # Get function_details from source root
    function_details: Dict[str, Any] = {}
    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if source_root_path and source_root_path.exists() and source_root_path.is_dir():
        try:
            sections = _get_source_sections_cached(str(source_root_path))
            function_details = sections.get("function_details", {})
            function_details, req_doc_paths, sds_doc_paths = _enrich_function_details_map(
                function_details,
                function_table_rows=sections.get("function_table_rows", []),
                req_doc_paths=req_doc_paths,
                sds_doc_paths=sds_doc_paths,
                uds_path=uds_path,
            )
            _logger.info("[STS_GENERATE][%s] parsed %d functions from source", req_id, len(function_details))
        except Exception as e:
            print(f"[STS_GENERATE][{req_id}] source parsing warning: {e}", flush=True)

    # Resolve optional supplementary document paths
    # 선택 입력은 **worker 경유**로 로컬화한다 — 직독은 cloudium `U:` 를 못 읽어
    # 전량 `None` 이 되고, 생성기가 그 문서 **없이** 만든 뒤 "생성 완료" 가 떴다.
    from backend.services.resolver_helpers import resolve_builder_input
    opt_skips: List[str] = []

    def _resolve_opt(val: str) -> Optional[str]:
        return resolve_builder_input(val, reasons=opt_skips)

    sds_docx_path = _resolve_opt(sds_path)
    # Fallback: auto-discover SDS from docs/ if not provided
    sds_docx_path = _doc_or_discovered(sds_docx_path, sds_path, _discover_sds_docx,
                                       label="SDS", tag=f"[STS_GENERATE][{req_id}] ")
    uds_file_path = _resolve_opt(uds_path)
    stp_docx_path = _resolve_opt(stp_path)
    hsis_file_path = _doc_or_discovered(_resolve_opt(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")
    if opt_skips:
        _logger.warning("STS(sync): 선택 입력 %d건이 빠진 채 생성한다 — %s",
                        len(opt_skips), "; ".join(opt_skips)[:400])

    # Resolve template
    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    # Output path
    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "sts", "sts_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}_STS",
        "version": version,
        "asil_level": asil_level,
        "max_tc_per_req": max_tc_per_req,
        "default_test_env": "SwTE_01",
    }

    _sts_ai_cfg = _load_sts_ai_config()

    try:
        # STS 생성 전 구간(문서 파싱 + AI + xlsx 빌드) — 이벤트 루프 밖에서 돈다.
        result = await _run_blocking(
            generate_sts,
            requirements_text=req_texts,
            function_details=function_details,
            output_path=str(out_path),
            template_path=tpl_path,
            project_config=project_config,
            srs_docx_path=srs_docx_path,
            sds_docx_path=sds_docx_path,
            uds_path=uds_file_path,
            stp_path=stp_docx_path,
            hsis_path=hsis_file_path,
            ai_config=_sts_ai_cfg,
            source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달 (품질 DB project_root)
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"STS 생성 실패: {e}")

    download_url = f"/api/local/sts/download/{out_filename}"
    print(f"[STS_GENERATE][{req_id}] done tc={result.get('test_case_count')} file={out_path}", flush=True)

    payload = _build_excel_artifact_payload(
        "sts",
        {
            "ok": True,
            "output_path": str(out_path),
            "filename": out_filename,
            "download_url": download_url,
            "test_case_count": result.get("test_case_count", 0),
            "elapsed_seconds": result.get("elapsed_seconds", 0),
            "quality_report": result.get("quality_report", {}),
            "trace_coverage": result.get("trace_coverage", {}),
            "validation": result.get("validation", {}),
            "validation_report_path": result.get("validation_report_path", ""),
        },
        output_path=str(out_path),
        filename=out_filename,
        download_url=download_url,
        preview_url=f"/api/local/sts/preview/{out_filename}",
    )
    _write_excel_artifact_sidecar(out_path, "sts", payload)
    return payload


@router.post("/api/local/sts/generate-stream")
async def local_sts_generate_stream(
    request: Request,
    source_root: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    stp_path: str = Form(""),
    hsis_path: str = Form(""),
    req_paths: str = Form(""),
    req_files: List[UploadFile] = File(default_factory=list),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_tc_per_req: int = Form(5),
    report_dir: str = Form(""),
):
    """Generate STS with SSE progress streaming."""
    import json as _json
    import queue
    import threading

    from backend.services.resolver_helpers import reject_upload_in_cloudium
    from sts_generator import generate_sts
    reject_upload_in_cloudium(*(req_files or []))

    srs_docx_path: Optional[str] = None
    if srs_path:
        p = Path(srs_path).expanduser().resolve()
        if p.exists() and p.is_file():
            srs_docx_path = str(p)

    req_paths_list = _parse_path_list(req_paths)
    req_texts: List[str] = []
    req_doc_paths: List[str] = []
    sds_doc_paths: List[str] = []
    # 탈락 사유를 버리지 않는다 — 권한 없음/본문 0자가 '문서 미지정'과 구분된다.
    doc_skips: List[str] = []
    for path_str in req_paths_list:
        p, text, reason = read_requirement_doc(path_str, allow=_is_allowed_req_doc)
        if reason:
            doc_skips.append(reason)
            continue
        if not p or not text:
            continue
        req_texts.append(text)
        if p.suffix.lower() == ".docx":
            req_doc_paths.append(str(p))
            if is_sds_filename(p.name):
                sds_doc_paths.append(str(p))
        if not srs_docx_path and is_srs_filename(p.name) and p.suffix.lower() == ".docx":
            srs_docx_path = str(p)
    if doc_skips:
        _logger.warning("[STS] 요구사항 문서 %d건 탈락: %s", len(doc_skips), "; ".join(doc_skips[:5]))

    for f in (req_files or []):
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            text = _read_text_from_file(tmp_path)
            if text:
                req_texts.append(text.strip())
                if tmp_path.suffix.lower() == ".docx":
                    req_doc_paths.append(str(tmp_path))
                    if is_sds_filename(f.filename):
                        sds_doc_paths.append(str(tmp_path))
                if not srs_docx_path and is_srs_filename(f.filename) and suffix == ".docx":
                    srs_docx_path = str(tmp_path)
        except Exception:
            pass

    if not req_texts and not srs_docx_path:
        raise HTTPException(status_code=400, detail="SRS 문서를 최소 1개 이상 제공해주세요.")

    function_details: Dict[str, Any] = {}
    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if source_root_path and source_root_path.exists() and source_root_path.is_dir():
        try:
            sections = _get_source_sections_cached(str(source_root_path))
            function_details = sections.get("function_details", {})
            function_details, req_doc_paths, sds_doc_paths = _enrich_function_details_map(
                function_details,
                function_table_rows=sections.get("function_table_rows", []),
                req_doc_paths=req_doc_paths,
                sds_doc_paths=sds_doc_paths,
                uds_path=uds_path,
            )
        except Exception:
            pass

    # 선택 입력 worker 경유 (위 sync 핸들러와 같은 이유).
    from backend.services.resolver_helpers import resolve_builder_input
    opt_skips2: List[str] = []

    def _resolve_opt2(val: str) -> Optional[str]:
        return resolve_builder_input(val, reasons=opt_skips2)

    sds_docx_path = _resolve_opt2(sds_path)
    uds_file_path = _resolve_opt2(uds_path)
    stp_docx_path = _resolve_opt2(stp_path)
    hsis_file_path2 = _doc_or_discovered(_resolve_opt2(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")
    if opt_skips2:
        _logger.warning("STS(stream): 선택 입력 %d건이 빠진 채 생성한다 — %s",
                        len(opt_skips2), "; ".join(opt_skips2)[:400])

    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "sts", "sts_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}_STS",
        "version": version,
        "asil_level": asil_level,
        "max_tc_per_req": max_tc_per_req,
        "default_test_env": "SwTE_01",
    }

    progress_queue: queue.Queue = queue.Queue()

    def _on_progress(pct: int, msg: str):
        progress_queue.put({"type": "progress", "pct": pct, "message": msg})

    def _run():
        _sts_ai_cfg2 = _load_sts_ai_config()
        try:
            result = generate_sts(
                requirements_text=req_texts,
                function_details=function_details,
                output_path=str(out_path),
                template_path=tpl_path,
                project_config=project_config,
                srs_docx_path=srs_docx_path,
                sds_docx_path=sds_docx_path,
                uds_path=uds_file_path,
                stp_path=stp_docx_path,
                hsis_path=hsis_file_path2,
                ai_config=_sts_ai_cfg2,
                on_progress=_on_progress,
                source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달 (품질 DB project_root)
            )
            download_url = f"/api/local/sts/download/{out_filename}"
            payload = _build_excel_artifact_payload(
                "sts",
                {
                    "ok": True,
                    "output_path": str(out_path),
                    "filename": out_filename,
                    "download_url": download_url,
                    "test_case_count": result.get("test_case_count", 0),
                    "elapsed_seconds": result.get("elapsed_seconds", 0),
                    "quality_report": result.get("quality_report", {}),
                    "trace_coverage": result.get("trace_coverage", {}),
                    "validation": result.get("validation", {}),
                    "validation_report_path": result.get("validation_report_path", ""),
                },
                output_path=str(out_path),
                filename=out_filename,
                download_url=download_url,
                preview_url=f"/api/local/sts/preview/{out_filename}",
            )
            _write_excel_artifact_sidecar(out_path, "sts", payload)
            progress_queue.put({"type": "done", **payload})
        except Exception as e:
            progress_queue.put({"type": "error", "detail": str(e)})

    threading.Thread(target=wrap_with_user(_run), daemon=True).start()

    def _event_stream():
        while True:
            try:
                item = progress_queue.get(timeout=120)
            except queue.Empty:
                yield "data: {\"type\":\"keepalive\"}\n\n"
                continue
            yield f"data: {_json.dumps(item, ensure_ascii=False)}\n\n"
            if item.get("type") in ("done", "error"):
                break

    return StreamingResponse(_event_stream(), media_type="text/event-stream")


@router.post("/api/local/sts/generate-async")
async def local_sts_generate_async(
    request: Request,
    source_root: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    stp_path: str = Form(""),
    hsis_path: str = Form(""),
    req_paths: str = Form(""),
    req_files: List[UploadFile] = File(default_factory=list),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_tc_per_req: int = Form(5),
    report_dir: str = Form(""),
) -> Dict[str, Any]:
    """Non-blocking STS generation. Returns job_id for progress polling."""
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    from sts_generator import generate_sts
    reject_upload_in_cloudium(*(req_files or []))

    srs_docx_path: Optional[str] = None
    if srs_path:
        p = Path(srs_path).expanduser().resolve()
        if p.exists() and p.is_file():
            srs_docx_path = str(p)

    req_paths_list = _parse_path_list(req_paths)
    req_texts: List[str] = []
    req_doc_paths: List[str] = []
    sds_doc_paths: List[str] = []
    # 탈락 사유를 버리지 않는다 — 권한 없음/본문 0자가 '문서 미지정'과 구분된다.
    doc_skips: List[str] = []
    for path_str in req_paths_list:
        p, text, reason = read_requirement_doc(path_str, allow=_is_allowed_req_doc)
        if reason:
            doc_skips.append(reason)
            continue
        if not p or not text:
            continue
        req_texts.append(text)
        if p.suffix.lower() == ".docx":
            req_doc_paths.append(str(p))
            if is_sds_filename(p.name):
                sds_doc_paths.append(str(p))
        if not srs_docx_path and is_srs_filename(p.name) and p.suffix.lower() == ".docx":
            srs_docx_path = str(p)
    if doc_skips:
        _logger.warning("[STS] 요구사항 문서 %d건 탈락: %s", len(doc_skips), "; ".join(doc_skips[:5]))

    for f in (req_files or []):
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            text = _read_text_from_file(tmp_path)
            if text:
                req_texts.append(text.strip())
                if tmp_path.suffix.lower() == ".docx":
                    req_doc_paths.append(str(tmp_path))
                    if is_sds_filename(f.filename):
                        sds_doc_paths.append(str(tmp_path))
                if not srs_docx_path and is_srs_filename(f.filename) and suffix == ".docx":
                    srs_docx_path = str(tmp_path)
        except Exception:
            pass

    # Fallback: auto-discover SRS from docs/ if not yet resolved
    srs_docx_path = _doc_or_discovered(
        srs_docx_path, bool(req_paths_list or req_files), _discover_srs_docx,
        label="SRS", tag="[STS_GENERATE_ASYNC] ")

    if not req_texts and not srs_docx_path:
        raise HTTPException(status_code=400, detail="SRS 문서를 최소 1개 이상 제공해주세요.")

    job_id = uuid.uuid4().hex
    _set_progress(
        "local_sts", "local", "local",
        {"stage": "start", "percent": 1, "message": "STS 생성 준비 중", "done": False, "error": ""},
        job_id=job_id,
    )

    # 선택 입력 worker 경유 (위 sync/stream 핸들러와 같은 이유).
    from backend.services.resolver_helpers import resolve_builder_input
    opt_skips3: List[str] = []

    def _resolve_opt3(val: str) -> Optional[str]:
        return resolve_builder_input(val, reasons=opt_skips3)

    sds_docx_path = _resolve_opt3(sds_path)
    # Fallback: auto-discover SDS from docs/ if not provided
    sds_docx_path = _doc_or_discovered(sds_docx_path, sds_path, _discover_sds_docx,
                                       label="SDS", tag="[STS_GENERATE_ASYNC] ")
    uds_file_path = _resolve_opt3(uds_path)
    stp_docx_path = _resolve_opt3(stp_path)
    hsis_file_path3 = _doc_or_discovered(_resolve_opt3(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")
    if opt_skips3:
        _logger.warning("STS(async): 선택 입력 %d건이 빠진 채 생성한다 — %s",
                        len(opt_skips3), "; ".join(opt_skips3)[:400])

    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "sts", "sts_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}_STS",
        "version": version,
        "asil_level": asil_level,
        "max_tc_per_req": max_tc_per_req,
        "default_test_env": "SwTE_01",
    }

    def _sts_on_progress(pct: int, msg: str):
        _set_progress(
            "local_sts", "local", "local",
            {"stage": "generation", "percent": max(10, min(pct, 95)), "message": msg},
            job_id=job_id,
        )

    def _worker():
        try:
            _set_progress(
                "local_sts", "local", "local",
                {"stage": "source_analysis", "percent": 10, "message": "소스 코드 분석 중"},
                job_id=job_id,
            )
            function_details: Dict[str, Any] = {}
            if source_root_path and source_root_path.exists() and source_root_path.is_dir():
                try:
                    sections = _get_source_sections_cached(str(source_root_path))
                    function_details = sections.get("function_details", {})
                    # req_doc_paths / sds_doc_paths 는 **바깥 스코프**(이 함수를 감싸는
                    # 엔드포인트)의 변수다. 여기서 대입 타깃에 넣으면 _worker 지역변수로
                    # 승격돼, 바로 아래 kwarg 읽기가 UnboundLocalError를 냈다.
                    # 정확히는 **인자 평가 단계**에서 터지므로 _enrich_function_details_map
                    # 은 호출조차 되지 않았고, 예외는 except가 "source parsing warning"
                    # 으로만 찍었다. 이때 function_details 는 바로 위에서 이미 바인딩된
                    # **원본(비보강) 파싱 결과**를 그대로 들고 generate_sts 로 넘어갔다
                    # (빈 dict 아님). 즉 유실된 건 문서/HSIS 기반 **보강분**이다.
                    # 반환된 경로 집합은 generate_sts가 쓰지 않으므로 버린다(동기판도 동일).
                    function_details, _, _ = _enrich_function_details_map(
                        function_details,
                        function_table_rows=sections.get("function_table_rows", []),
                        req_doc_paths=req_doc_paths,
                        sds_doc_paths=sds_doc_paths,
                        uds_path=uds_path,
                    )
                except Exception as e:
                    _logger.warning("[STS_ASYNC][%s] source parsing warning: %s", job_id, e)

            _set_progress(
                "local_sts", "local", "local",
                {"stage": "generation", "percent": 40, "message": "STS 테스트 케이스 생성 중"},
                job_id=job_id,
            )
            result = generate_sts(
                requirements_text=req_texts,
                function_details=function_details,
                output_path=str(out_path),
                template_path=tpl_path,
                project_config=project_config,
                srs_docx_path=srs_docx_path,
                sds_docx_path=sds_docx_path,
                uds_path=uds_file_path,
                stp_path=stp_docx_path,
                hsis_path=hsis_file_path3,
                ai_config=_load_sts_ai_config(),
                on_progress=_sts_on_progress,
                source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달 (품질 DB project_root)
            )

            download_url = f"/api/local/sts/download/{out_filename}"
            result_payload = _build_excel_artifact_payload(
                "sts",
                {
                    "ok": True,
                    "output_path": str(out_path),
                    "filename": out_filename,
                    "download_url": download_url,
                    "test_case_count": result.get("test_case_count", 0),
                    "elapsed_seconds": result.get("elapsed_seconds", 0),
                    "quality_report": result.get("quality_report", {}),
                    "trace_coverage": result.get("trace_coverage", {}),
                    "validation": result.get("validation", {}),
                    "validation_report_path": result.get("validation_report_path", ""),
                },
                output_path=str(out_path),
                filename=out_filename,
                download_url=download_url,
                preview_url=f"/api/local/sts/preview/{out_filename}",
            )
            _write_excel_artifact_sidecar(out_path, "sts", result_payload)
            _set_progress(
                "local_sts", "local", "local",
                {
                    "stage": "done", "percent": 100, "message": "???",
                    "done": True, "error": "",
                    "result": result_payload,
                },
                job_id=job_id,
            )
            _logger.info("[STS_ASYNC][%s] done file=%s tc=%s", job_id, out_filename, result.get("test_case_count"))

        except Exception as exc:
            tb = traceback.format_exc()
            _logger.error("[STS_ASYNC][%s] FAILED: %s\n%s", job_id, str(exc)[:500], tb)
            _set_progress(
                "local_sts", "local", "local",
                {"stage": "error", "percent": 100, "message": f"실패: {str(exc)[:300]}", "done": True, "error": str(exc)[:500]},
                job_id=job_id,
            )

    threading.Thread(target=wrap_with_user(_worker), daemon=True).start()
    return {"ok": True, "job_id": job_id}


@router.get("/api/local/sts/progress")
def local_sts_progress(job_id: str = "") -> Dict[str, Any]:
    data = _get_progress("local_sts", "local", "local", job_id)
    return {"ok": bool(data), "progress": data}


@router.get("/api/local/sts/download/{filename}")
def local_sts_download(filename: str, report_dir: Optional[str] = None) -> FileResponse:
    file_path = _resolve_local_sts_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="STS file not found")
    media = _excel_media_type(file_path)
    return FileResponse(str(file_path), filename=file_path.name, media_type=media)


@router.get("/api/local/sts/files")
def local_sts_files(report_dir: Optional[str] = None) -> List[Dict[str, Any]]:
    base = _resolve_report_dir(report_dir)
    sts_dir = _local_sts_dir(base)
    if not sts_dir.exists():
        return []
    rows: List[Dict[str, Any]] = []
    for f in sorted(sts_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.suffix.lower() in (".xlsm", ".xlsx"):
            payload = _load_excel_artifact_payload(
                f,
                "sts",
                download_url=f"/api/local/sts/download/{f.name}",
                preview_url=f"/api/local/sts/preview/{f.name}",
            )
            rows.append({
                "filename": f.name,
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                "download_url": f"/api/local/sts/download/{f.name}",
                "validation_report_path": payload.get("validation_report_path", ""),
                "residual_report_path": payload.get("residual_report_path", ""),
                "summary": payload.get("summary", {}),
            })
    return rows


@router.get("/api/local/sts/preview/{filename}")
def local_sts_preview(filename: str, report_dir: Optional[str] = None, max_rows: int = 30) -> Dict[str, Any]:
    """Preview STS Excel content as JSON table data."""
    file_path = _resolve_local_sts_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="STS file not found")
    return _parse_xlsm_preview(file_path, max_rows)


@router.get("/api/local/sts/view/{filename}")
def local_sts_view(filename: str, report_dir: Optional[str] = None) -> Dict[str, Any]:
    file_path = _resolve_local_sts_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="STS file not found")
    return _load_excel_artifact_payload(
        file_path,
        "sts",
        download_url=f"/api/local/sts/download/{file_path.name}",
        preview_url=f"/api/local/sts/preview/{file_path.name}",
    )


@router.post("/api/local/suts/generate")
def local_suts_generate(
    request: Request,
    source_root: str = Form(""),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_sequences: int = Form(6),
    report_dir: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    hsis_path: str = Form(""),
) -> Dict[str, Any]:
    """Generate SUTS (Software Unit Test Specification) Excel from source code."""
    from suts_generator import generate_suts

    req_id = (request.headers.get("x-req-id") or "").strip() or f"suts-gen-{int(time.time() * 1000)}"
    print(f"[SUTS_GENERATE][{req_id}] start source_root={source_root}", flush=True)

    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="유효한 소스 코드 루트 경로를 제공해주세요.")

    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    from backend.services.resolver_helpers import resolve_builder_input

    def _resolve_doc_path(val: str) -> Optional[str]:
        # worker 경유 — 직독은 cloudium `U:` 를 못 읽어 선택 문서가 조용히 빠졌다.
        return resolve_builder_input(val)

    srs_docx = _resolve_doc_path(srs_path)
    sds_docx = _resolve_doc_path(sds_path)
    uds_file = _resolve_doc_path(uds_path)
    # Fallback: auto-discover SRS/SDS/HSIS from docs/ if not provided
    srs_docx = _doc_or_discovered(srs_docx, srs_path, _discover_srs_docx,
                                  label="SRS", tag="[SUTS_GENERATE] ")
    sds_docx = _doc_or_discovered(sds_docx, sds_path, _discover_sds_docx,
                                  label="SDS", tag="[SUTS_GENERATE] ")
    hsis_suts = _doc_or_discovered(_resolve_doc_path(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")

    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "suts", "suts_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}-SUTS",
        "version": version,
        "asil_level": asil_level,
    }

    try:
        result = generate_suts(
            source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달
            output_path=str(out_path),
            template_path=tpl_path,
            project_config=project_config,
            max_sequences=max_sequences,
            srs_docx_path=srs_docx,
            sds_docx_path=sds_docx,
            uds_path=uds_file,
            hsis_path=hsis_suts,
            ai_config=_load_sts_ai_config(),
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"SUTS 생성 실패: {e}")

    download_url = f"/api/local/suts/download/{out_filename}"
    print(f"[SUTS_GENERATE][{req_id}] done tc={result.get('test_case_count')} file={out_path}", flush=True)

    payload = _build_excel_artifact_payload(
        "suts",
        {
            "ok": True,
            "output_path": str(out_path),
            "filename": out_filename,
            "download_url": download_url,
            "test_case_count": result.get("test_case_count", 0),
            "total_sequences": result.get("total_sequences", 0),
            "elapsed_seconds": result.get("elapsed_seconds", 0),
            "quality_report": result.get("quality_report", {}),
            "validation": result.get("validation", {}),
            "validation_report_path": result.get("validation_report_path", ""),
        },
        output_path=str(out_path),
        filename=out_filename,
        download_url=download_url,
        preview_url=f"/api/local/suts/preview/{out_filename}",
    )
    _write_excel_artifact_sidecar(out_path, "suts", payload)
    return payload


@router.post("/api/local/suts/generate-stream")
def local_suts_generate_stream(
    request: Request,
    source_root: str = Form(""),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_sequences: int = Form(6),
    report_dir: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    hsis_path: str = Form(""),
):
    """Generate SUTS with SSE progress streaming."""
    import json as _json
    import queue
    import threading

    from suts_generator import generate_suts

    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="유효한 소스 코드 루트 경로를 제공해주세요.")

    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    from backend.services.resolver_helpers import resolve_builder_input

    def _res_doc(val: str) -> Optional[str]:
        return resolve_builder_input(val)

    srs_docx_stream = _res_doc(srs_path)
    sds_docx_stream = _res_doc(sds_path)
    uds_file_stream = _res_doc(uds_path)
    # Fallback: auto-discover SRS/SDS/HSIS from docs/ if not provided
    srs_docx_stream = _doc_or_discovered(srs_docx_stream, srs_path, _discover_srs_docx,
                                         label="SRS", tag="[SUTS_STREAM] ")
    sds_docx_stream = _doc_or_discovered(sds_docx_stream, sds_path, _discover_sds_docx,
                                         label="SDS", tag="[SUTS_STREAM] ")
    hsis_suts_stream = _doc_or_discovered(_res_doc(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")

    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "suts", "suts_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}-SUTS",
        "version": version,
        "asil_level": asil_level,
    }

    progress_queue: queue.Queue = queue.Queue()

    def _on_progress(pct: int, msg: str):
        progress_queue.put({"type": "progress", "pct": pct, "message": msg})

    def _run():
        try:
            result = generate_suts(
                source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달
                output_path=str(out_path),
                template_path=tpl_path,
                project_config=project_config,
                max_sequences=max_sequences,
                on_progress=_on_progress,
                srs_docx_path=srs_docx_stream,
                sds_docx_path=sds_docx_stream,
                uds_path=uds_file_stream,
                hsis_path=hsis_suts_stream,
                ai_config=_load_sts_ai_config(),
            )
            download_url = f"/api/local/suts/download/{out_filename}"
            payload = _build_excel_artifact_payload(
                "suts",
                {
                    "ok": True,
                    "output_path": str(out_path),
                    "filename": out_filename,
                    "download_url": download_url,
                    "test_case_count": result.get("test_case_count", 0),
                    "total_sequences": result.get("total_sequences", 0),
                    "elapsed_seconds": result.get("elapsed_seconds", 0),
                    "quality_report": result.get("quality_report", {}),
                    "validation": result.get("validation", {}),
                    "validation_report_path": result.get("validation_report_path", ""),
                },
                output_path=str(out_path),
                filename=out_filename,
                download_url=download_url,
                preview_url=f"/api/local/suts/preview/{out_filename}",
            )
            _write_excel_artifact_sidecar(out_path, "suts", payload)
            progress_queue.put({"type": "done", **payload})
        except Exception as e:
            progress_queue.put({"type": "error", "detail": str(e)})

    threading.Thread(target=wrap_with_user(_run), daemon=True).start()

    def _event_stream():
        while True:
            try:
                item = progress_queue.get(timeout=120)
            except queue.Empty:
                yield "data: {\"type\":\"keepalive\"}\n\n"
                continue
            yield f"data: {_json.dumps(item, ensure_ascii=False)}\n\n"
            if item.get("type") in ("done", "error"):
                break

    return StreamingResponse(_event_stream(), media_type="text/event-stream")


@router.post("/api/local/suts/generate-async")
def local_suts_generate_async(
    request: Request,
    source_root: str = Form(""),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_sequences: int = Form(6),
    report_dir: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    hsis_path: str = Form(""),
) -> Dict[str, Any]:
    """Non-blocking SUTS generation. Returns job_id for progress polling."""
    from suts_generator import generate_suts

    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="유효한 소스 코드 루트 경로를 제공해주세요.")

    job_id = uuid.uuid4().hex
    _set_progress(
        "local_suts", "local", "local",
        {"stage": "start", "percent": 1, "message": "SUTS 생성 준비 중", "done": False, "error": ""},
        job_id=job_id,
    )

    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    from backend.services.resolver_helpers import resolve_builder_input

    def _res_async(val: str) -> Optional[str]:
        return resolve_builder_input(val)

    srs_docx_async = _res_async(srs_path)
    sds_docx_async = _res_async(sds_path)
    uds_file_async = _res_async(uds_path)
    hsis_suts_async = _doc_or_discovered(_res_async(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")

    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "suts", "suts_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}-SUTS",
        "version": version,
        "asil_level": asil_level,
    }

    def _suts_on_progress(pct: int, msg: str):
        stage = "source_analysis" if pct < 30 else "generation"
        _set_progress(
            "local_suts", "local", "local",
            {"stage": stage, "percent": max(10, min(pct, 95)), "message": msg},
            job_id=job_id,
        )

    def _worker():
        try:
            _set_progress(
                "local_suts", "local", "local",
                {"stage": "source_analysis", "percent": 5, "message": "소스 코드 분석 시작"},
                job_id=job_id,
            )
            _logger.info("[SUTS_ASYNC][%s] calling generate_suts ...", job_id)
            result = generate_suts(
                source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달
                output_path=str(out_path),
                template_path=tpl_path,
                project_config=project_config,
                max_sequences=max_sequences,
                on_progress=_suts_on_progress,
                srs_docx_path=srs_docx_async,
                sds_docx_path=sds_docx_async,
                uds_path=uds_file_async,
                hsis_path=hsis_suts_async,
                ai_config=_load_sts_ai_config(),
            )
            _logger.info("[SUTS_ASYNC][%s] generate_suts returned, setting done", job_id)

            download_url = f"/api/local/suts/download/{out_filename}"
            result_payload = _build_excel_artifact_payload(
                "suts",
                {
                    "ok": True,
                    "output_path": str(out_path),
                    "filename": out_filename,
                    "download_url": download_url,
                    "test_case_count": result.get("test_case_count", 0),
                    "total_sequences": result.get("total_sequences", 0),
                    "elapsed_seconds": result.get("elapsed_seconds", 0),
                    "quality_report": result.get("quality_report", {}),
                    "validation": result.get("validation", {}),
                    "validation_report_path": result.get("validation_report_path", ""),
                },
                output_path=str(out_path),
                filename=out_filename,
                download_url=download_url,
                preview_url=f"/api/local/suts/preview/{out_filename}",
            )
            _write_excel_artifact_sidecar(out_path, "suts", result_payload)
            _set_progress(
                "local_suts", "local", "local",
                {
                    "stage": "done", "percent": 100, "message": "???",
                    "done": True, "error": "",
                    "result": result_payload,
                },
                job_id=job_id,
            )
            _logger.info("[SUTS_ASYNC][%s] done file=%s tc=%s", job_id, out_filename, result.get("test_case_count"))

        except Exception as exc:
            tb = traceback.format_exc()
            _logger.error("[SUTS_ASYNC][%s] FAILED: %s\n%s", job_id, str(exc)[:500], tb)
            _set_progress(
                "local_suts", "local", "local",
                {"stage": "error", "percent": 100, "message": f"실패: {str(exc)[:300]}", "done": True, "error": str(exc)[:500]},
                job_id=job_id,
            )

    threading.Thread(target=wrap_with_user(_worker), daemon=True).start()
    return {"ok": True, "job_id": job_id}


@router.get("/api/local/suts/progress")
def local_suts_progress(job_id: str = "") -> Dict[str, Any]:
    data = _get_progress("local_suts", "local", "local", job_id)
    return {"ok": bool(data), "progress": data}


@router.get("/api/local/suts/download/{filename}")
def local_suts_download(filename: str, report_dir: Optional[str] = None) -> FileResponse:
    file_path = _resolve_local_suts_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="SUTS file not found")
    media = _excel_media_type(file_path)
    return FileResponse(str(file_path), filename=file_path.name, media_type=media)


@router.get("/api/local/suts/files")
def local_suts_files(report_dir: Optional[str] = None) -> List[Dict[str, Any]]:
    base = _resolve_report_dir(report_dir)
    suts_dir = _local_suts_dir(base)
    if not suts_dir.exists():
        return []
    rows: List[Dict[str, Any]] = []
    for f in sorted(suts_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.suffix.lower() in (".xlsm", ".xlsx"):
            payload = _load_excel_artifact_payload(
                f,
                "suts",
                download_url=f"/api/local/suts/download/{f.name}",
                preview_url=f"/api/local/suts/preview/{f.name}",
            )
            rows.append({
                "filename": f.name,
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                "download_url": f"/api/local/suts/download/{f.name}",
                "validation_report_path": payload.get("validation_report_path", ""),
                "residual_report_path": payload.get("residual_report_path", ""),
                "summary": payload.get("summary", {}),
            })
    return rows


@router.get("/api/local/suts/preview/{filename}")
def local_suts_preview(filename: str, report_dir: Optional[str] = None, max_rows: int = 30) -> Dict[str, Any]:
    """Preview SUTS Excel content as JSON table data."""
    file_path = _resolve_local_suts_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="SUTS file not found")
    return _parse_xlsm_preview(file_path, max_rows)


@router.get("/api/local/suts/view/{filename}")
def local_suts_view(filename: str, report_dir: Optional[str] = None) -> Dict[str, Any]:
    file_path = _resolve_local_suts_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="SUTS file not found")
    return _load_excel_artifact_payload(
        file_path,
        "suts",
        download_url=f"/api/local/suts/download/{file_path.name}",
        preview_url=f"/api/local/suts/preview/{file_path.name}",
    )


@router.post("/api/local/sits/generate")
def local_sits_generate(
    request: Request,
    source_root: str = Form(""),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_subcases: int = Form(7),
    report_dir: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    hsis_path: str = Form(""),
    stp_path: str = Form(""),
) -> Dict[str, Any]:
    """Generate SITS (Software Integration Test Specification) Excel from source code."""
    from sits_generator import generate_sits

    req_id = (request.headers.get("x-req-id") or "").strip() or f"sits-gen-{int(time.time() * 1000)}"
    print(f"[SITS_GENERATE][{req_id}] start source_root={source_root}", flush=True)

    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="유효한 소스 코드 루트 경로를 제공해주세요.")

    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    from backend.services.resolver_helpers import resolve_builder_input

    def _resolve_doc_path_sits(val: str) -> Optional[str]:
        return resolve_builder_input(val)

    srs_docx = _resolve_doc_path_sits(srs_path)
    sds_docx = _resolve_doc_path_sits(sds_path)
    uds_file = _resolve_doc_path_sits(uds_path)
    hsis_file = _doc_or_discovered(_resolve_doc_path_sits(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")
    stp_file = _resolve_doc_path_sits(stp_path)

    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "sits", "sits_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}-SITS",
        "version": version,
        "asil_level": asil_level,
    }

    try:
        result = generate_sits(
            source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달
            output_path=str(out_path),
            template_path=tpl_path,
            project_config=project_config,
            max_subcases=max_subcases,
            srs_docx_path=srs_docx,
            sds_docx_path=sds_docx,
            uds_path=uds_file,
            hsis_path=hsis_file,
            stp_path=stp_file,
            ai_config=_load_sts_ai_config(),
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"SITS 생성 실패: {e}")

    download_url = f"/api/local/sits/download/{out_filename}"
    print(f"[SITS_GENERATE][{req_id}] done tc={result.get('test_case_count')} file={out_path}", flush=True)

    payload = _build_excel_artifact_payload(
        "sits",
        {
            "ok": True,
            "output_path": str(out_path),
            "filename": out_filename,
            "download_url": download_url,
            "test_case_count": result.get("test_case_count", 0),
            "elapsed_seconds": result.get("elapsed_seconds", 0),
            "quality_report": result.get("quality_report", {}),
            "validation": result.get("validation", {}),
            "validation_report_path": result.get("validation_report_path", ""),
        },
        output_path=str(out_path),
        filename=out_filename,
        download_url=download_url,
        preview_url=f"/api/local/sits/preview/{out_filename}",
    )
    _write_excel_artifact_sidecar(out_path, "sits", payload)
    return payload


@router.post("/api/local/sits/generate-stream")
def local_sits_generate_stream(
    request: Request,
    source_root: str = Form(""),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_subcases: int = Form(7),
    report_dir: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    hsis_path: str = Form(""),
    stp_path: str = Form(""),
):
    """Generate SITS with SSE progress streaming."""
    import json as _json
    import queue
    import threading

    from sits_generator import generate_sits

    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="유효한 소스 코드 루트 경로를 제공해주세요.")

    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    from backend.services.resolver_helpers import resolve_builder_input

    def _res_doc_sits(val: str) -> Optional[str]:
        return resolve_builder_input(val)

    srs_docx_stream = _res_doc_sits(srs_path)
    sds_docx_stream = _res_doc_sits(sds_path)
    uds_file_stream = _res_doc_sits(uds_path)
    hsis_stream = _doc_or_discovered(_res_doc_sits(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")
    stp_stream = _res_doc_sits(stp_path)

    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "sits", "sits_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}-SITS",
        "version": version,
        "asil_level": asil_level,
    }

    progress_queue: queue.Queue = queue.Queue()

    def _on_progress(pct: int, msg: str):
        progress_queue.put({"type": "progress", "pct": pct, "message": msg})

    def _run():
        try:
            result = generate_sits(
                source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달
                output_path=str(out_path),
                template_path=tpl_path,
                project_config=project_config,
                max_subcases=max_subcases,
                on_progress=_on_progress,
                srs_docx_path=srs_docx_stream,
                sds_docx_path=sds_docx_stream,
                uds_path=uds_file_stream,
                hsis_path=hsis_stream,
                stp_path=stp_stream,
                ai_config=_load_sts_ai_config(),
            )
            download_url = f"/api/local/sits/download/{out_filename}"
            payload = _build_excel_artifact_payload(
                "sits",
                {
                    "ok": True,
                    "output_path": str(out_path),
                    "filename": out_filename,
                    "download_url": download_url,
                    "test_case_count": result.get("test_case_count", 0),
                    "elapsed_seconds": result.get("elapsed_seconds", 0),
                    "quality_report": result.get("quality_report", {}),
                    "validation": result.get("validation", {}),
                    "validation_report_path": result.get("validation_report_path", ""),
                },
                output_path=str(out_path),
                filename=out_filename,
                download_url=download_url,
                preview_url=f"/api/local/sits/preview/{out_filename}",
            )
            _write_excel_artifact_sidecar(out_path, "sits", payload)
            progress_queue.put({"type": "done", **payload})
        except Exception as e:
            progress_queue.put({"type": "error", "detail": str(e)})

    threading.Thread(target=wrap_with_user(_run), daemon=True).start()

    def _event_stream():
        while True:
            try:
                item = progress_queue.get(timeout=120)
            except queue.Empty:
                yield "data: {\"type\":\"keepalive\"}\n\n"
                continue
            yield f"data: {_json.dumps(item, ensure_ascii=False)}\n\n"
            if item.get("type") in ("done", "error"):
                break

    return StreamingResponse(_event_stream(), media_type="text/event-stream")


@router.post("/api/local/sits/generate-async")
def local_sits_generate_async(
    request: Request,
    source_root: str = Form(""),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_subcases: int = Form(7),
    report_dir: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    hsis_path: str = Form(""),
    stp_path: str = Form(""),
) -> Dict[str, Any]:
    """Non-blocking SITS generation. Returns job_id for progress polling."""
    from sits_generator import generate_sits

    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="유효한 소스 코드 루트 경로를 제공해주세요.")

    job_id = uuid.uuid4().hex
    _set_progress(
        "local_sits", "local", "local",
        {"stage": "start", "percent": 1, "message": "SITS 생성 준비 중", "done": False, "error": ""},
        job_id=job_id,
    )

    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)

    # 선택 입력 worker 경유. SITS 는 선택 문서가 5종(SRS·SDS·UDS·HSIS·STP)이라
    # 직독 시절엔 cloudium 에서 **다섯 개가 통째로** 빠진 채 만들어졌다.
    from backend.services.resolver_helpers import resolve_builder_input
    sits_opt_skips: List[str] = []

    def _res_async_sits(val: str) -> Optional[str]:
        return resolve_builder_input(val, reasons=sits_opt_skips)

    srs_docx_async = _res_async_sits(srs_path)
    sds_docx_async = _res_async_sits(sds_path)
    uds_file_async = _res_async_sits(uds_path)
    hsis_async = _doc_or_discovered(_res_async_sits(hsis_path), hsis_path,
                              _discover_hsis_path, label="HSIS")
    stp_async = _res_async_sits(stp_path)
    if sits_opt_skips:
        _logger.warning("SITS: 선택 입력 %d건이 빠진 채 생성한다 — %s",
                        len(sits_opt_skips), "; ".join(sits_opt_skips)[:400])

    base_dir = _resolve_report_dir(report_dir)
    out_filename, out_path = _build_local_excel_output(base_dir, "sits", "sits_local", tpl_path)

    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}-SITS",
        "version": version,
        "asil_level": asil_level,
    }

    def _sits_on_progress(pct: int, msg: str):
        stage = "source_analysis" if pct < 30 else "generation"
        _set_progress(
            "local_sits", "local", "local",
            {"stage": stage, "percent": max(10, min(pct, 95)), "message": msg},
            job_id=job_id,
        )

    def _worker():
        try:
            _set_progress(
                "local_sits", "local", "local",
                {"stage": "source_analysis", "percent": 5, "message": "소스 코드 분석 시작"},
                job_id=job_id,
            )
            _logger.info("[SITS_ASYNC][%s] calling generate_sits ...", job_id)
            result = generate_sits(
                source_root=source_root,  # 콤마 구분 복수 경로 그대로 전달
                output_path=str(out_path),
                template_path=tpl_path,
                project_config=project_config,
                max_subcases=max_subcases,
                on_progress=_sits_on_progress,
                srs_docx_path=srs_docx_async,
                sds_docx_path=sds_docx_async,
                uds_path=uds_file_async,
                hsis_path=hsis_async,
                stp_path=stp_async,
                ai_config=_load_sts_ai_config(),
            )
            _logger.info("[SITS_ASYNC][%s] generate_sits returned, setting done", job_id)

            download_url = f"/api/local/sits/download/{out_filename}"
            result_payload = _build_excel_artifact_payload(
                "sits",
                {
                    "ok": True,
                    "output_path": str(out_path),
                    "filename": out_filename,
                    "download_url": download_url,
                    "test_case_count": result.get("test_case_count", 0),
                    "elapsed_seconds": result.get("elapsed_seconds", 0),
                    "quality_report": result.get("quality_report", {}),
                    "validation": result.get("validation", {}),
                    "validation_report_path": result.get("validation_report_path", ""),
                },
                output_path=str(out_path),
                filename=out_filename,
                download_url=download_url,
                preview_url=f"/api/local/sits/preview/{out_filename}",
            )
            _write_excel_artifact_sidecar(out_path, "sits", result_payload)
            _set_progress(
                "local_sits", "local", "local",
                {
                    "stage": "done", "percent": 100, "message": "완료",
                    "done": True, "error": "",
                    "result": result_payload,
                },
                job_id=job_id,
            )
            _logger.info("[SITS_ASYNC][%s] done file=%s tc=%s", job_id, out_filename, result.get("test_case_count"))

        except Exception as exc:
            tb = traceback.format_exc()
            _logger.error("[SITS_ASYNC][%s] FAILED: %s\n%s", job_id, str(exc)[:500], tb)
            _set_progress(
                "local_sits", "local", "local",
                {"stage": "error", "percent": 100, "message": f"실패: {str(exc)[:300]}", "done": True, "error": str(exc)[:500]},
                job_id=job_id,
            )

    threading.Thread(target=wrap_with_user(_worker), daemon=True).start()
    return {"ok": True, "job_id": job_id}


@router.get("/api/local/sits/progress")
def local_sits_progress(job_id: str = "") -> Dict[str, Any]:
    data = _get_progress("local_sits", "local", "local", job_id)
    return {"ok": bool(data), "progress": data}


@router.get("/api/local/sits/download/{filename}")
def local_sits_download(filename: str, report_dir: Optional[str] = None) -> FileResponse:
    file_path = _resolve_local_sits_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="SITS file not found")
    media = _excel_media_type(file_path)
    return FileResponse(str(file_path), filename=file_path.name, media_type=media)


@router.get("/api/local/sits/files")
def local_sits_files(report_dir: Optional[str] = None) -> List[Dict[str, Any]]:
    base = _resolve_report_dir(report_dir)
    sits_dir = _local_sits_dir(base)
    if not sits_dir.exists():
        return []
    rows: List[Dict[str, Any]] = []
    for f in sorted(sits_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.suffix.lower() in (".xlsm", ".xlsx"):
            payload = _load_excel_artifact_payload(
                f,
                "sits",
                download_url=f"/api/local/sits/download/{f.name}",
                preview_url=f"/api/local/sits/preview/{f.name}",
            )
            rows.append({
                "filename": f.name,
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                "download_url": f"/api/local/sits/download/{f.name}",
                "validation_report_path": payload.get("validation_report_path", ""),
                "residual_report_path": payload.get("residual_report_path", ""),
                "summary": payload.get("summary", {}),
            })
    return rows


@router.get("/api/local/sits/preview/{filename}")
def local_sits_preview(filename: str, report_dir: Optional[str] = None, max_rows: int = 30) -> Dict[str, Any]:
    """Preview SITS Excel content as JSON table data."""
    file_path = _resolve_local_sits_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="SITS file not found")
    return _parse_xlsm_preview(file_path, max_rows)


@router.get("/api/local/sits/view/{filename}")
def local_sits_view(filename: str, report_dir: Optional[str] = None) -> Dict[str, Any]:
    file_path = _resolve_local_sits_path(report_dir, filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="SITS file not found")
    return _load_excel_artifact_payload(
        file_path,
        "sits",
        download_url=f"/api/local/sits/download/{file_path.name}",
        preview_url=f"/api/local/sits/preview/{file_path.name}",
    )


@router.post("/api/local/suts/export-vectorcast")
def local_suts_export_vectorcast(
    filename: str = Form(""),
    report_dir: str = Form(""),
    source_root: str = Form(""),
    project_id: str = Form(""),
    compiler: str = Form("CC"),
) -> Dict[str, Any]:
    """Generate a VectorCAST unit-test package from a SUTS file."""
    from tools.export_suts_vectorcast import export_suts_to_vectorcast_model
    from tools.export_vectorcast_script import export_vectorcast_package

    base_dir = _resolve_report_dir(report_dir)
    suts_dir = _local_suts_dir(base_dir)
    if filename:
        xlsm_path = suts_dir / filename
    else:
        candidates = sorted(suts_dir.glob("*.xlsm"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not candidates:
            raise HTTPException(status_code=404, detail="No SUTS file found")
        xlsm_path = candidates[0]
    if not xlsm_path.exists():
        raise HTTPException(status_code=404, detail="SUTS file not found")

    resolved_source_root = str(source_root or "").strip()
    cfg = load_vectorcast_project_config(project_id=project_id, source_root=resolved_source_root)
    effective_project_id = str(project_id or cfg.get("project_id") or "VECTORCAST").strip()
    effective_source_root = resolved_source_root or str(cfg.get("source_root") or "").strip()

    resolved_source_root = str(source_root or "").strip()
    cfg = load_vectorcast_project_config(project_id=project_id, source_root=resolved_source_root)
    effective_source_root = resolved_source_root or str(cfg.get("source_root") or "").strip()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    # ⚠ 키가 **없는** 이름이다(ts 뿐) — 다른 프로젝트·다른 사용자여도 같은 초면 부딪힌다.
    #   `mkdir(exist_ok=True)` 는 폴더를 **공유**시켜 안의 산출물이 서로 덮어써진다.
    #   원자 선점으로 폴더 자체를 비켜간다.
    from backend.services.output_paths import reserve_unique_dir
    out_dir = reserve_unique_dir(base_dir / "vectorcast" / f"suts_vectorcast_{ts}")
    package_name = out_dir.name
    intermediate_json = out_dir / "suts_vectorcast_model.json"
    warnings_md = out_dir / "suts_vectorcast_warnings.md"

    try:
        model = export_suts_to_vectorcast_model(
            str(xlsm_path),
            str(intermediate_json),
            warnings_md=str(warnings_md),
            project_id=effective_project_id,
        )
        manifest = export_vectorcast_package(
            str(intermediate_json),
            str(out_dir),
            package_name=package_name,
            source_root=effective_source_root,
            compiler=str(cfg.get("compiler") or compiler or "CC"),
            project_config=cfg,
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"VectorCAST package generation failed: {e}")

    unit_names = [str(unit.get("unit_name") or "") for unit in model.get("units") or []]
    return _build_vectorcast_package_response(
        package_dir=out_dir,
        package_name=package_name,
        manifest=manifest,
        project_config=cfg,
        units=unit_names,
    )


@router.post("/api/local/sits/export-vectorcast")
def local_sits_export_vectorcast(
    filename: str = Form(""),
    report_dir: str = Form(""),
    source_root: str = Form(""),
    project_id: str = Form(""),
    compiler: str = Form("CC"),
) -> Dict[str, Any]:
    """Generate a VectorCAST integration test package from a SITS file."""
    from tools.export_sits_vectorcast_package import export_sits_vectorcast_package

    base_dir = _resolve_report_dir(report_dir)
    sits_dir = _local_sits_dir(base_dir)

    # Locate intermediate JSON (generated alongside the XLSM by generate_sits)
    if filename:
        xlsm_path = sits_dir / filename
    else:
        # latest XLSM in sits dir
        candidates = sorted(sits_dir.glob("*.xlsm"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not candidates:
            raise HTTPException(status_code=404, detail="No SITS file found")
        xlsm_path = candidates[0]

    stem = xlsm_path.stem
    intermediate_json = xlsm_path.with_name(f"{stem}_vectorcast.json")
    if not intermediate_json.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Intermediate JSON not found: {intermediate_json.name}. Re-generate the SITS file first.",
        )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_name = f"sits_vectorcast_{ts}"
    out_dir = base_dir / "vectorcast" / package_name

    # source_root / compiler 설정
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    resolved_source_root = str(Path(_first_root).resolve()) if _first_root else ""

    try:
        model = json.loads(intermediate_json.read_text(encoding="utf-8"))
        manifest = export_sits_vectorcast_package(
            str(intermediate_json),
            str(out_dir),
            package_name=package_name,
            source_root=resolved_source_root,
            compiler=compiler or "CC",
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"VectorCAST 패키지 생성 실패: {e}")

    unit_names = sorted(
        {
            str(step.split(".", 1)[0]).strip()
            for item in (model.get("integrations") or [])
            for step in str(item.get("call_chain") or "").split("->")
            if str(step).strip()
        }
    )
    return _build_vectorcast_package_response(
        package_dir=out_dir,
        package_name=package_name,
        manifest=manifest,
        project_config={},
        units=unit_names,
    )


# ── VectorCAST 패키지 목록 / 다운로드 ──

# ── VectorCAST 패키지 루트 ────────────────────────────────────────────────────
# ⚠ 등록(쓰기)이 **두 갈래**다. 읽기가 한쪽만 보면 반대쪽 산출물이 통째로 사라진다:
#     `/api/local/{suts,sits}/export-vectorcast` → `_resolve_report_dir(report_dir)/vectorcast`
#     `/api/jenkins/suts/export-vectorcast`      → `_jenkins_exports_dir(cache_root)/vectorcast`
#   실측(2026-08-07): `reports/vectorcast` 3개 · `.devops_pro_cache/exports/vectorcast` 2개.
#   목록은 전자만 봤고, 심지어 프론트가 cache_root 를 report_dir 로 보내 403 이었다 →
#   화면엔 "등록된 패키지가 없습니다". **403 이 '없음'으로 위장**한 것이다.
#
#   그래서 한쪽 루트만 고르는 수정(어느 쪽이든)은 답이 아니다 — 반대쪽이 그대로 사라진다.
#   목록·다운로드·삭제가 **이 함수 하나**를 공유하게 한다. 따로 세면 곧 어긋나서
#   "목록에 보이는데 못 지운다"(또는 그 반대)가 된다.

def _vcast_roots(report_dir: str, cache_root: str) -> Tuple[List[Tuple[str, Path]], List[str]]:
    """(source, 루트) 목록과 **제외 사유**를 함께 준다. 사유를 버리면 침묵이 된다."""
    from backend.helpers.jenkins import _jenkins_exports_dir

    roots: List[Tuple[str, Path]] = []
    notes: List[str] = []

    try:
        base = _resolve_report_dir(report_dir)
    except HTTPException as exc:
        # 허용 밖 report_dir 이라고 전체를 실패시키지 않는다 — 기본 리포트 루트는 살리고
        # 무시했다는 사실만 올린다(구 프론트가 cache_root 를 여기로 보내던 전례).
        base = _resolve_report_dir("")
        notes.append(f"report_dir 무시됨({exc.detail}) — 기본 리포트 루트로 대체")
    roots.append(("reports", (base / "vectorcast").resolve()))

    raw_cache = str(cache_root or "").strip()
    if raw_cache:
        try:
            # ⚠ create=False — 조회가 디렉터리를 만들면 오타 난 경로도 실재하게 된다.
            roots.append(("jenkins_cache",
                          (_jenkins_exports_dir(raw_cache, create=False) / "vectorcast").resolve()))
            # 캐시는 **사용자별 격리 + legacy 공유 이중구조**다(`.devops_pro_cache/{user}/`
            # 와 `.devops_pro_cache/`). 사용자 세그먼트가 붙기 전에 등록된 패키지는 상위
            # 공유 루트에 남아 있어, 여기를 안 보면 그 등록물이 영영 안 보인다
            # (실측 2026-08-07: 현재 jenkins 경유 등록물 2건이 **전부** 이쪽에 있다).
            # ⚠ **실재할 때만** 추가한다 — 없는 루트를 지어내면 진단이 흐려진다.
            parent = Path(raw_cache).expanduser().resolve().parent
            legacy = (_jenkins_exports_dir(str(parent), create=False) / "vectorcast").resolve()
            if legacy.is_dir():
                roots.append(("jenkins_cache_legacy", legacy))
        except (OSError, ValueError) as exc:
            notes.append(f"cache_root 제외됨 — {type(exc).__name__}: {exc}")

    # 두 루트가 같은 디렉터리를 가리키면 패키지가 두 번 나온다(첫 등장만 남긴다).
    seen: set = set()
    deduped: List[Tuple[str, Path]] = []
    for source, root in roots:
        if root in seen:
            continue
        seen.add(root)
        deduped.append((source, root))
    return deduped, notes


def _scan_vcast_root(source: str, root: Path) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """루트 하나를 훑는다. (패키지, 실패사유). 미존재는 실패가 아니다(아직 등록 전)."""
    if not root.exists():
        return [], None
    packages: List[Dict[str, Any]] = []
    try:
        entries = [d for d in root.iterdir() if d.is_dir()]
    except OSError as exc:
        return [], f"{source} 루트를 읽지 못했다 — {type(exc).__name__}: {exc}"

    for d in entries:
        try:
            mtime = d.stat().st_mtime
            files = sorted(p.name for p in d.iterdir() if p.is_file())
        except OSError as exc:
            # 개별 패키지 실패가 나머지를 가리지 않게 — 행은 남기되 사유를 싣는다.
            packages.append({
                "name": d.name, "doc_type": "suts", "path": str(d), "source": source,
                "files": [], "file_count": 0, "created": None, "summary": {},
                "error": f"{type(exc).__name__}: {exc}",
            })
            continue
        meta: Dict[str, Any] = {}
        manifest_file = d / "manifest.json"
        if manifest_file.exists():
            try:
                meta = json.loads(manifest_file.read_text(encoding="utf-8"))
            except (OSError, ValueError):
                meta = {}          # manifest 파손은 목록 자체를 막지 않는다(summary 만 빈다)
        packages.append({
            "name": d.name,
            "doc_type": "sits" if "sits" in d.name else "suts",
            "path": str(d),
            "source": source,
            "files": files,
            "file_count": len(files),
            "created": datetime.fromtimestamp(mtime).isoformat(),
            "summary": meta.get("summary", {}) if isinstance(meta, dict) else {},
        })
    return packages, None


def _confine_vcast_package(package_path: str, report_dir: str, cache_root: str) -> Path:
    """`package_path` 를 목록이 훑는 루트의 **직계 하위**로 확정한다. 밖이면 403.

    ⚠ 이전엔 검사가 **아예 없었다** — `delete` 는 임의 경로를 `shutil.rmtree` 했고
      `download` 는 임의 파일을 반환했다. 목록이 준 경로를 되받는 설계라 클라이언트를
      믿은 것인데, 클라이언트가 준 값은 클라이언트가 지어낼 수도 있는 값이다.
    ⚠ '직계 하위'인 이유: 단순 하위 검사는 루트 **자기 자신**도 통과시켜
      `vectorcast` 디렉터리째 삭제된다.
    """
    raw = str(package_path or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="package_path required")
    try:
        target = Path(raw).expanduser().resolve()
    except (OSError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="invalid package_path") from exc
    roots, _ = _vcast_roots(report_dir, cache_root)
    if not any(target.parent == root for _, root in roots):
        # ⚠ 어떤 경로가 허용되는지 응답에 적지 않는다 — 실패 자체가 정보다.
        _logger.warning("허용 밖 VectorCAST package_path 를 차단했다: %s", target)
        raise HTTPException(status_code=403, detail="package_path not allowed")
    return target


@router.get("/api/local/vectorcast/list")
def local_vectorcast_list(report_dir: str = "", cache_root: str = "") -> Dict[str, Any]:
    """등록된 VectorCAST 패키지 목록 조회 — **등록 경로 두 갈래를 모두** 훑는다."""
    roots, warnings = _vcast_roots(report_dir, cache_root)
    packages: List[Dict[str, Any]] = []
    scanned: List[Dict[str, Any]] = []
    for source, root in roots:
        found, failure = _scan_vcast_root(source, root)
        if failure:
            warnings.append(failure)
        packages.extend(found)
        scanned.append({
            "source": source, "path": str(root),
            "exists": root.exists(), "count": len(found),
            "error": failure,
        })
    # 루트를 가로질러 최신순. created 가 없는 행(stat 실패)은 뒤로.
    packages.sort(key=lambda p: p.get("created") or "", reverse=True)
    # `scanned_roots` 는 "0건"이 어느 루트에서 온 0건인지 화면이 말할 수 있게 하는 근거다.
    return {"ok": True, "packages": packages, "warnings": warnings, "scanned_roots": scanned}


@router.get("/api/local/vectorcast/download")
def local_vectorcast_download(
    package_path: str = "",
    filename: str = "",
    report_dir: str = "",
    cache_root: str = "",
):
    """VectorCAST 패키지 파일 다운로드 — 허용 루트 하위만."""
    from fastapi.responses import FileResponse
    pkg_dir = _confine_vcast_package(package_path, report_dir, cache_root)
    if not pkg_dir.is_dir():
        raise HTTPException(status_code=404, detail="Package not found")
    if filename:
        # ⚠ `pkg_dir / filename` 만으론 `../` 로 패키지 밖을 짚는다 — 상대경로를 살균한다.
        try:
            target = safe_resolve_under(pkg_dir, filename)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="invalid filename") from exc
        if not target.is_file():
            raise HTTPException(status_code=404, detail="File not found")
        return FileResponse(str(target), filename=target.name)
    # filename 없으면 ZIP으로 전체 패키지 다운로드
    import zipfile
    tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
    with zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in pkg_dir.rglob("*"):
            if f.is_file():
                zf.write(f, f.relative_to(pkg_dir))
    return FileResponse(tmp.name, filename=f"{pkg_dir.name}.zip", media_type="application/zip")


@router.delete("/api/local/vectorcast/delete")
def local_vectorcast_delete(
    package_path: str = "",
    report_dir: str = "",
    cache_root: str = "",
) -> Dict[str, Any]:
    """VectorCAST 패키지 삭제 — 허용 루트의 **직계 하위**만.

    ⚠ 이 함수는 `shutil.rmtree` 다. 이전엔 경로 검사가 없어 인증된 사용자면
      **서버의 아무 디렉터리나** 지울 수 있었다.
    """
    import shutil
    pkg_dir = _confine_vcast_package(package_path, report_dir, cache_root)
    if not pkg_dir.is_dir():
        raise HTTPException(status_code=404, detail="Package not found")
    shutil.rmtree(pkg_dir)
    return {"ok": True, "deleted": str(pkg_dir)}


@router.post("/api/local/scm", dependencies=[Depends(require_admin)])
def local_scm(req: ScmRequest) -> Dict[str, Any]:
    if req.mode.lower() == "git":
        return run_git(
            project_root=confine_request_root(req.project_root, rel_path=req.workdir_rel),
            workdir_rel=req.workdir_rel,
            action=req.action,
            repo_url=req.repo_url,
            branch=req.branch,
            depth=req.depth,
            timeout_sec=req.timeout_sec,
        )
    if req.mode.lower() == "svn":
        return run_svn(
            project_root=confine_request_root(req.project_root, rel_path=req.workdir_rel),
            workdir_rel=req.workdir_rel,
            action=req.action,
            repo_url=req.repo_url,
            revision=req.revision,
            timeout_sec=req.timeout_sec,
        )
    raise HTTPException(status_code=400, detail="unknown scm mode")


@router.post("/api/local/impact/trigger")
def local_impact_trigger(req: LocalImpactTriggerRequest) -> Dict[str, Any]:
    try:
        trigger = build_registry_trigger(
            trigger_type="local",
            scm_id=req.scm_id,
            base_ref=req.base_ref,
            dry_run=req.dry_run,
            auto_generate=req.auto_generate,
            targets=req.targets or None,
            manual_changed_files=req.manual_changed_files or None,
            metadata={"source": "api/local/impact/trigger"},
        )
    except KeyError:
        raise HTTPException(status_code=404, detail="registry entry not found")
    return run_impact_update(trigger)


@router.post("/api/local/impact/trigger-async")
def local_impact_trigger_async(req: LocalImpactTriggerRequest) -> Dict[str, Any]:
    try:
        trigger = build_registry_trigger(
            trigger_type="local",
            scm_id=req.scm_id,
            base_ref=req.base_ref,
            dry_run=req.dry_run,
            auto_generate=req.auto_generate,
            targets=req.targets or None,
            manual_changed_files=req.manual_changed_files or None,
            metadata={"source": "api/local/impact/trigger-async"},
        )
    except KeyError:
        raise HTTPException(status_code=404, detail="registry entry not found")
    return start_impact_job(trigger)


@router.post("/api/local/kb/list", dependencies=[Depends(require_admin)])
def local_kb_list(req: KBRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root)
    return {"entries": list_kb_entries(root, req.report_dir)}


@router.post("/api/local/kb/delete", dependencies=[Depends(require_admin)])
def local_kb_delete(req: KBRequest) -> Dict[str, Any]:
    if not req.entry_key:
        raise HTTPException(status_code=400, detail="entry_key required")
    root = confine_request_root(req.project_root)
    ok, msg = delete_kb_entry(req.entry_key, root, req.report_dir)
    return {"ok": ok, "message": msg}


@router.post("/api/local/editor/read", dependencies=[Depends(require_admin)])
def local_editor_read(req: EditorReadRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.rel_path)
    return read_file_text(root, req.rel_path, req.max_bytes)


@router.post("/api/local/editor/write", dependencies=[Depends(require_admin)])
def local_editor_write(req: EditorWriteRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.rel_path)
    return write_file_text(root, req.rel_path, req.content, req.make_backup)


@router.post("/api/local/editor/replace", dependencies=[Depends(require_admin)])
def local_editor_replace(req: EditorReplaceRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.rel_path)
    return replace_lines(root, req.rel_path, req.start_line, req.end_line, req.content)


@router.post("/api/local/format-c")
def local_format_c(req: FormatCodeRequest) -> Dict[str, Any]:
    return format_c_code(req.text, req.filename)


# ─── Project Setup (component_map / override 자동 생성) ────────────────────

@router.post("/api/local/project-setup/generate-component-map")
def local_generate_component_map(
    sds_path: str = Form(""),
    source_root: str = Form(""),
    output_dir: str = Form(""),
) -> Dict[str, Any]:
    """SDS 문서에서 component_map.json을 자동 생성한다."""
    if not sds_path:
        raise HTTPException(status_code=400, detail="sds_path 필요")
    if not source_root:
        raise HTTPException(status_code=400, detail="source_root 필요")
    from report_gen.project_setup import generate_component_map_from_sds
    out_path = str(Path(output_dir or "docs") / "component_map.json") if output_dir else "docs/component_map.json"
    result = generate_component_map_from_sds(sds_path, source_root, output_path=out_path)
    return {"ok": True, **result}


@router.post("/api/local/project-setup/generate-override")
def local_generate_override(
    uds_path: str = Form(""),
    output_dir: str = Form(""),
) -> Dict[str, Any]:
    """레퍼런스 UDS 문서에서 함수 단위 override 맵을 생성한다."""
    if not uds_path:
        raise HTTPException(status_code=400, detail="uds_path (레퍼런스 UDS) 필요")
    from report_gen.project_setup import generate_override_from_reference_uds
    out_path = str(Path(output_dir or "docs") / "uds_function_swcom_override.json") if output_dir else "docs/uds_function_swcom_override.json"
    result = generate_override_from_reference_uds(uds_path, output_path=out_path)
    return {"ok": True, **result}


@router.get("/api/local/project-setup/status")
def local_project_setup_status() -> Dict[str, Any]:
    """현재 프로젝트의 setup 파일 상태를 확인한다."""
    import json as _json
    docs_dir = repo_root / "docs"
    cm_path = docs_dir / "component_map.json"
    ovr_path = docs_dir / "uds_function_swcom_override.json"

    cm_status = {"exists": False, "entries": 0}
    if cm_path.exists():
        try:
            data = _json.loads(cm_path.read_text(encoding="utf-8"))
            cm_status = {
                "exists": True,
                "entries": len(data),
                "verify_o": sum(1 for r in data if r.get("verify", "").upper() == "O"),
                "verify_x": sum(1 for r in data if r.get("verify", "").upper() == "X"),
                "path": str(cm_path),
            }
        except Exception:
            cm_status = {"exists": True, "entries": 0, "error": "parse failed"}

    ovr_status = {"exists": False, "functions": 0}
    if ovr_path.exists():
        try:
            data = _json.loads(ovr_path.read_text(encoding="utf-8"))
            ovr_status = {
                "exists": True,
                "functions": len(data),
                "with_asil": sum(1 for v in data.values() if v.get("asil")),
                "swcom_count": len({v.get("swcom") for v in data.values()}),
                "path": str(ovr_path),
            }
        except Exception:
            ovr_status = {"exists": True, "functions": 0, "error": "parse failed"}

    return {
        "ok": True,
        "component_map": cm_status,
        "override": ovr_status,
    }


@router.post("/api/local/rag/status")
def local_rag_status(req: RagStatusRequest) -> Dict[str, Any]:
    cfg = req.config or {}
    report_dir = str(req.report_dir or cfg.get("report_dir") or getattr(config, "DEFAULT_REPORT_DIR", "reports"))
    report_path = (repo_root / report_dir).resolve()
    report_path.mkdir(parents=True, exist_ok=True)
    force_pg = bool(getattr(config, "FORCE_PGVECTOR", False))
    force_pg_strict = bool(getattr(config, "FORCE_PGVECTOR_STRICT", False))
    try:
        kb = get_kb(report_path)
        storage = str(getattr(kb, "storage", "sqlite"))
        pg_ok = bool(getattr(kb, "_pg_ok", False))
    except Exception as e:
        return {
            "ok": False,
            "error": str(e),
            "kb_storage": "pgvector" if force_pg else str(getattr(config, "KB_STORAGE", "sqlite")),
            "pgvector_forced": force_pg,
            "pgvector_strict": force_pg_strict,
            "pgvector_ready": False,
            "kb_dir": str(report_path / getattr(config, "KB_DIR_NAME", "kb_store")),
        }
    dsn = str(os.environ.get("PGVECTOR_DSN", "") or getattr(config, "PGVECTOR_DSN", ""))
    url = str(os.environ.get("PGVECTOR_URL", "") or getattr(config, "PGVECTOR_URL", ""))
    stats = {}
    try:
        stats = kb.stats()
    except Exception:
        stats = {}
    return {
        "ok": True,
        "rag_ingest_enable": bool(cfg.get("rag_ingest_enable", getattr(config, "RAG_INGEST_ENABLE", True))),
        "rag_ingest_on_pipeline": bool(cfg.get("rag_ingest_on_pipeline", getattr(config, "RAG_INGEST_ON_PIPELINE", True))),
        "agent_rag": bool(cfg.get("agent_rag", getattr(config, "AGENT_RAG_ENABLED_DEFAULT", True))),
        "kb_storage": storage,
        "pgvector_forced": force_pg,
        "pgvector_strict": force_pg_strict,
        "pgvector_configured": bool(dsn or url),
        "pgvector_ready": pg_ok if storage == "pgvector" else False,
        "kb_dir": str(report_path / getattr(config, "KB_DIR_NAME", "kb_store")),
        "stats": stats,
    }


@router.post("/api/local/rag/ingest")
def local_rag_ingest(req: RagIngestRequest) -> Dict[str, Any]:
    cfg = req.config or {}
    report_dir = str(req.report_dir or cfg.get("report_dir") or getattr(config, "DEFAULT_REPORT_DIR", "reports"))
    report_path = (repo_root / report_dir).resolve()
    report_path.mkdir(parents=True, exist_ok=True)
    kb = get_kb(report_path)
    result = ingest_external_sources(kb, cfg=cfg)
    return {"ok": True, "result": result}


@router.post("/api/local/rag/ingest-files")
async def local_rag_ingest_files(
    files: List[UploadFile] = File(default_factory=list),
    category: str = Form("general"),
    tags: str = Form(""),
    report_dir: str = Form(""),
    chunk_size: Optional[int] = Form(None),
    chunk_overlap: Optional[int] = Form(None),
    max_chunks: Optional[int] = Form(None),
) -> Dict[str, Any]:
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    reject_upload_in_cloudium(*(files or []))
    report_dir = str(report_dir or getattr(config, "DEFAULT_REPORT_DIR", "reports"))
    report_path = (repo_root / report_dir).resolve()
    report_path.mkdir(parents=True, exist_ok=True)
    kb = get_kb(report_path)
    tag_list = [t.strip() for t in re.split(r"[,\n;]+", str(tags or "")) if t.strip()]
    use_chunk_size = int(chunk_size or getattr(config, "RAG_CHUNK_SIZE", 1200))
    use_overlap = int(chunk_overlap or getattr(config, "RAG_CHUNK_OVERLAP", 200))
    use_max_chunks = int(max_chunks or getattr(config, "RAG_INGEST_MAX_CHUNKS_PER_FILE", 12))

    added = 0
    skipped = 0
    for f in files:
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            chunks = _read_and_chunk_file(
                tmp_path,
                chunk_size=use_chunk_size,
                overlap=use_overlap,
                max_chunks=use_max_chunks,
            )
        except Exception:
            chunks = []
        if not chunks:
            skipped += 1
            continue
        for i, ch in enumerate(chunks):
            title = f"{category}:{Path(f.filename).name}#{i+1}"
            kb.add_document(
                title=title,
                content=ch,
                category=str(category or "general"),
                tags=tag_list,
                source_file=str(f.filename),
            )
            added += 1
    return {"ok": True, "added": added, "skipped": skipped, "category": category}


@router.post("/api/local/rag/use-pgvector")
def local_rag_use_pgvector(req: RagStorageRequest) -> Dict[str, Any]:
    dsn = str(req.pgvector_dsn or "").strip()
    url = str(req.pgvector_url or "").strip()
    if not dsn and not url:
        raise HTTPException(status_code=400, detail="pgvector dsn or url required")
    os.environ["KB_STORAGE"] = "pgvector"
    os.environ["PGVECTOR_DSN"] = dsn
    os.environ["PGVECTOR_URL"] = url
    config.KB_STORAGE = "pgvector"
    config.PGVECTOR_DSN = dsn
    config.PGVECTOR_URL = url
    config.FORCE_PGVECTOR = True
    config.FORCE_PGVECTOR_STRICT = True
    # get_kb 캐시는 base_dir 만 key 로 쓰고 storage(KB_STORAGE/FORCE_PGVECTOR) 는
    # key 에 없다 → 캐시를 비우지 않으면 이전에 빌드된 sqlite 인스턴스가 반환되어
    # pgvector 전환이 silent no-op 된다. config 변이 후 무효화.
    from workflow.rag import _clear_kb_cache
    _clear_kb_cache()
    report_dir = str(req.report_dir or getattr(config, "DEFAULT_REPORT_DIR", "reports"))
    report_path = (repo_root / report_dir).resolve()
    report_path.mkdir(parents=True, exist_ok=True)
    try:
        kb = get_kb(report_path)
        pg_ok = bool(getattr(kb, "_pg_ok", False))
        return {
            "ok": pg_ok,
            "kb_storage": str(getattr(kb, "storage", "pgvector")),
            "pgvector_ready": pg_ok,
        }
    except Exception as e:
        return {
            "ok": False,
            "kb_storage": "pgvector",
            "pgvector_ready": False,
            "error": str(e),
            "hint": "pgvector 확장/권한/DSN 설정을 확인하세요 (CREATE EXTENSION vector;)",
        }


@router.post("/api/local/rag/query")
def local_rag_query(req: RagQueryRequest) -> Dict[str, Any]:
    query = str(req.query or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="query required")
    cfg = req.config or {}
    report_dir = str(req.report_dir or cfg.get("report_dir") or getattr(config, "DEFAULT_REPORT_DIR", "reports"))
    report_path = (repo_root / report_dir).resolve()
    report_path.mkdir(parents=True, exist_ok=True)
    kb = get_kb(report_path)
    top_k = max(1, int(req.top_k or 5))
    categories = [str(c).strip() for c in (req.categories or []) if str(c).strip()]
    rows = kb.search(query, top_k=top_k, categories=categories or None)
    items = []
    for row in rows:
        items.append(
            {
                "title": row.get("error_raw") or "",
                "category": row.get("category") or "",
                "source_file": row.get("source_file") or "",
                "score": float(row.get("score") or 0.0),
                "snippet": str(row.get("fix") or "")[:1200],
            }
        )
    return {"ok": True, "items": items}


@router.post("/api/local/pick-directory")
def api_pick_directory(req: PickerRequest) -> Dict[str, Any]:
    path, error = pick_directory(req.title or "폴더 선택")
    return {"ok": bool(path), "path": path, "error": error or None}


@router.post("/api/local/pick-file")
def api_pick_file(req: PickerRequest) -> Dict[str, Any]:
    path, error = pick_file(req.title or "파일 선택")
    return {"ok": bool(path), "path": path, "error": error or None}


@router.post("/api/local/open-file")
def api_open_file(req: OpenFileRequest) -> Dict[str, Any]:
    if not req.path:
        raise HTTPException(status_code=400, detail="path required")
    target = Path(req.path).expanduser().resolve()
    allowed_roots = [
        (Path.home() / ".devops_pro_cache").resolve(),
        repo_root.resolve(),
    ]
    if not is_under_any(target, allowed_roots):
        raise HTTPException(status_code=403, detail="path not allowed")
    if not target.exists() or target.is_dir():
        raise HTTPException(status_code=404, detail="file not found")
    try:
        _open_local_path(target)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return {"ok": True, "path": str(target)}


@router.post("/api/local/editor/read-abs", dependencies=[Depends(require_admin)])
def local_editor_read_abs(req: EditorReadAbsRequest) -> Dict[str, Any]:
    if not req.path:
        raise HTTPException(status_code=400, detail="path required")
    target = Path(req.path).expanduser().resolve()
    allowed_roots = [
        (Path.home() / ".devops_pro_cache").resolve(),
        repo_root.resolve(),
    ]
    if not is_under_any(target, allowed_roots):
        raise HTTPException(status_code=403, detail="path not allowed")
    if not target.exists() or target.is_dir():
        raise HTTPException(status_code=404, detail="file not found")
    text, truncated = read_text_limited(target, req.max_bytes)
    return {"ok": True, "path": str(target), "text": text, "truncated": truncated}


@router.post("/api/local/preview-text")
def local_preview_text(req: TextPreviewRequest) -> Dict[str, Any]:
    if not req.path:
        raise HTTPException(status_code=400, detail="path required")
    target = Path(req.path).expanduser().resolve()
    if not target.exists() or target.is_dir():
        raise HTTPException(status_code=404, detail="file not found")
    if not _is_allowed_req_doc(target):
        raise HTTPException(status_code=400, detail="unsupported file type")
    text = _read_text_from_file(target)
    max_chars = max(1000, int(req.max_chars or 0))
    truncated = False
    if max_chars and len(text) > max_chars:
        text = text[:max_chars]
        truncated = True
    return {"ok": True, "path": str(target), "text": text, "truncated": truncated}


@router.post("/api/local/sds/view")
def local_sds_view(req: SdsViewRequest) -> Dict[str, Any]:
    if not req.path:
        raise HTTPException(status_code=400, detail="path required")
    target = Path(req.path).expanduser().resolve()
    if not target.exists() or target.is_dir():
        raise HTTPException(status_code=404, detail="file not found")
    if target.suffix.lower() != ".docx":
        raise HTTPException(status_code=400, detail="SDS view supports .docx only")
    if not _is_allowed_req_doc(target):
        raise HTTPException(status_code=400, detail="unsupported file type")
    try:
        view = build_sds_view_model(
            str(target),
            max_items=max(1, int(req.max_items or 500)),
            changed_functions=dict(req.changed_functions or {}),
            changed_files=list(req.changed_files or []),
            flagged_modules=list(req.flagged_modules or []),
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return {"ok": True, "view": view}


@router.post("/api/local/open-folder")
def api_open_folder(req: OpenFolderRequest) -> Dict[str, Any]:
    if not req.path:
        raise HTTPException(status_code=400, detail="path required")
    target = Path(req.path).expanduser().resolve()
    allowed_roots = [
        (Path.home() / ".devops_pro_cache").resolve(),
        repo_root.resolve(),
    ]
    if not is_under_any(target, allowed_roots):
        raise HTTPException(status_code=403, detail="path not allowed")
    if not target.exists() or not target.is_dir():
        raise HTTPException(status_code=404, detail="folder not found")
    try:
        _open_local_path(target)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return {"ok": True, "path": str(target)}


@router.post("/api/local/preflight", dependencies=[Depends(require_admin)])
def local_preflight(req: PreflightRequest) -> Dict[str, Any]:
    cfg = dict(req.config or {})
    resolved, root = _resolve_source_root_from_cfg(cfg, confine_request_root(req.project_root))
    extra_paths = _collect_tool_paths()
    original_path = os.environ.get("PATH", "")
    os.environ["PATH"] = _augment_path(original_path, extra_paths)
    try:
        preflight = _build_preflight(cfg)
    finally:
        os.environ["PATH"] = original_path
    root_ok = Path(root).expanduser().resolve().exists()
    if not root_ok:
        preflight["warnings"].append("project_root_not_found")
    ready = root_ok and not preflight.get("missing")
    return {
        "ok": True,
        "ready": ready,
        "resolved": resolved,
        "project_root": root,
        "preflight": preflight,
    }


@router.post("/api/local/list-dir", dependencies=[Depends(require_admin)])
def api_list_dir(req: ListDirRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.rel_path)
    return list_directory(root, req.rel_path)


@router.post("/api/local/search", dependencies=[Depends(require_admin)])
def api_search(req: SearchRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.rel_path)
    return search_in_files(root, req.rel_path, req.query, req.max_results)


@router.post("/api/local/replace-text", dependencies=[Depends(require_admin)])
def api_replace_text(req: ReplaceTextRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.rel_path)
    return replace_in_file(root, req.rel_path, req.search, req.replace)


@router.post("/api/local/git/status", dependencies=[Depends(require_admin)])
def api_git_status(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_status(root, req.workdir_rel)


@router.post("/api/local/git/diff", dependencies=[Depends(require_admin)])
def api_git_diff(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_diff(root, req.workdir_rel, req.staged, req.path)


@router.post("/api/local/git/log", dependencies=[Depends(require_admin)])
def api_git_log(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_log(root, req.workdir_rel, req.max_count)


@router.post("/api/local/git/branches", dependencies=[Depends(require_admin)])
def api_git_branches(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_branches(root, req.workdir_rel)


@router.post("/api/local/git/checkout", dependencies=[Depends(require_admin)])
def api_git_checkout(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_checkout(root, req.workdir_rel, req.branch)


@router.post("/api/local/git/create-branch", dependencies=[Depends(require_admin)])
def api_git_create_branch(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_create_branch(root, req.workdir_rel, req.branch)


@router.post("/api/local/git/stage", dependencies=[Depends(require_admin)])
def api_git_stage(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_stage(root, req.workdir_rel, req.paths)


@router.post("/api/local/git/unstage", dependencies=[Depends(require_admin)])
def api_git_unstage(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_unstage(root, req.workdir_rel, req.paths)


@router.post("/api/local/git/commit", dependencies=[Depends(require_admin)])
def api_git_commit(req: GitRequest) -> Dict[str, Any]:
    root = confine_request_root(req.project_root, rel_path=req.workdir_rel)
    return git_commit(root, req.workdir_rel, req.message)
