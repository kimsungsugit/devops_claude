# /app/backend/routers/health.py
"""Health-check and monitoring endpoints."""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException

from backend.dependencies.admin import require_admin
from pydantic import BaseModel, Field, field_validator

import config
from backend.error_handler import APIError
# C3: 공유 헬퍼로 이동 — health.py / jenkins.py 등이 동일 검증 사용 (방어 비대칭 해소)
from backend.services.resolver_helpers import enforce_resolver_access as _enforce_resolver_access

_GATE_PROCESS_PATTERN = re.compile(r"^[A-Za-z0-9_\-\.]+\.exe$")

try:
    import psutil
except ImportError:
    psutil = None


# ── 문서 미리보기 캐시 ────────────────────────────────────────────────────────
# 페이지 이동/이미지 로드 시 거대 원본(예: 36MB UDS docx)을 resolver로 매번 다시
# 읽고(cloudium은 IPC) python-docx로 전체 재파싱하던 비용을 제거한다.
#  - _preview_bytes  : 원본 바이트(LRU + 바이트 예산). excel/csv 페이지 이동, docx 이미지 공용.
#  - _preview_payload: docx 텍스트 시트(전체 행)(item cap + 바이트 예산). 페이지 이동은 캐시 슬라이스만.
#  - _preview_relmap : docx 이미지 rId→media 경로 맵(작음, item cap만).
# 무효화: ① TTL(_preview_ttl, 기본 120s). ② local 모드는 (mtime,size) 시그니처로 정확 무효화
#   (cloudium은 backend가 U: 경로를 stat 못 함 → sig=None → TTL만). ③ file_mode 전환·
#   /cache/clear 시 clear_preview_cache로 전량 무효화(cross-mode 오염 방지).
import os as _os
import threading as _threading
import time as _time
from collections import OrderedDict as _OrderedDict

_PREVIEW_CACHE_LOCK = _threading.Lock()
_PREVIEW_BYTES_BUDGET = 128 * 1024 * 1024   # 바이트 캐시 총량 상한
_PREVIEW_PAYLOAD_BUDGET = 64 * 1024 * 1024  # docx payload 추정 텍스트 총량 상한
_PREVIEW_PAYLOAD_MAX = 8                     # 파싱된 docx 페이로드 최대 보관 수
_PREVIEW_RELMAP_MAX = 32                     # rId 맵 최대 보관 수(엔트리 작음)


def _preview_ttl() -> float:
    return float(getattr(config, "PREVIEW_CACHE_TTL", 120.0))


def _path_sig(path: str):
    """local 모드: (mtime_ns, size) 시그니처. cloudium(U: 경로 stat 불가)·오류 시 None.

    None이면 get 경로에서 재-stat을 건너뛰어 추가 비용 0 → cloudium은 순수 TTL 동작.
    """
    try:
        st = _os.stat(path)
    except OSError:
        return None
    return (st.st_mtime_ns, st.st_size)


class _PreviewCache:
    """TTL + (mtime,size)시그니처 + LRU(+바이트 예산) 캐시.

    모든 메서드는 _PREVIEW_CACHE_LOCK 보유 상태에서 호출한다(공개 헬퍼가 락 획득).
    엔트리 = [ts, value, sig, size]. self.total은 size 합으로, 만료/시그니처불일치/
    캡/예산 축출 모든 경로에서 일관되게 감산한다(phantom debt 방지).
    """

    def __init__(self, byte_budget: int = 0, item_cap: int = 0):
        self.store: "_OrderedDict" = _OrderedDict()   # key -> [ts, value, sig, size]
        self.byte_budget = byte_budget
        self.item_cap = item_cap
        self.total = 0

    def _drop(self, key: str, ent) -> None:
        self.store.pop(key, None)
        self.total -= ent[3]

    def get(self, key: str):
        ent = self.store.get(key)
        if ent is None:
            return None
        ts, value, sig, _size = ent
        if _time.time() - ts > _preview_ttl():
            self._drop(key, ent)
            return None
        # local 모드: 동일 path가 다른 내용으로 바뀌면(재생성/경로 재사용) 무효화.
        if sig is not None:
            cur = _path_sig(key)
            if cur is not None and cur != sig:
                self._drop(key, ent)
                return None
        self.store.move_to_end(key)
        return value

    def put(self, key: str, value, size: int, sig) -> None:
        old = self.store.pop(key, None)
        if old is not None:
            self.total -= old[3]
        self.store[key] = [_time.time(), value, sig, size]
        self.total += size
        if self.item_cap:
            while len(self.store) > self.item_cap:
                _k, ev = self.store.popitem(last=False)
                self.total -= ev[3]
        if self.byte_budget:
            # 예산 초과 시 LRU 축출. 최근 1개는 보존(거대 단일 파일이 캐시를 무력화하지 않도록).
            while self.total > self.byte_budget and len(self.store) > 1:
                _k, ev = self.store.popitem(last=False)
                self.total -= ev[3]

    def clear(self) -> None:
        self.store.clear()
        self.total = 0


_preview_bytes = _PreviewCache(byte_budget=_PREVIEW_BYTES_BUDGET)
_preview_payload = _PreviewCache(byte_budget=_PREVIEW_PAYLOAD_BUDGET, item_cap=_PREVIEW_PAYLOAD_MAX)
_preview_relmap = _PreviewCache(item_cap=_PREVIEW_RELMAP_MAX)


def _preview_cache_get(cache: "_PreviewCache", key: str):
    with _PREVIEW_CACHE_LOCK:
        return cache.get(key)


def _preview_cache_put(cache: "_PreviewCache", key: str, value, size: int, sig) -> None:
    with _PREVIEW_CACHE_LOCK:
        cache.put(key, value, size, sig)


def _estimate_payload_size(sheets) -> int:
    """docx payload(시트 전체 행) 추정 바이트(문자 수 합). 예산 산정용."""
    total = 0
    for s in sheets:
        for row in s.get("rows", []):
            for cell in row:
                total += len(cell)
    return total


def _read_bytes_cached(resolver, path: str) -> bytes:
    """resolver.read_bytes를 TTL+LRU 캐시. cloudium IPC/디스크 재read를 제거한다."""
    cached = _preview_cache_get(_preview_bytes, path)
    if cached is not None:
        return cached
    # IPC read는 락 밖에서 수행(락 점유 최소화). 동시 miss는 중복 read 후 last-write(무해).
    data = resolver.read_bytes(path)
    _preview_cache_put(_preview_bytes, path, data, len(data), _path_sig(path))
    return data


def clear_preview_cache() -> None:
    """미리보기 캐시 전체 비움(테스트/모드 전환/cache-clear용)."""
    with _PREVIEW_CACHE_LOCK:
        _preview_bytes.clear()
        _preview_payload.clear()
        _preview_relmap.clear()


class FileModeRequest(BaseModel):
    mode: str = "local"
    base_url: Optional[str] = None
    source_root: Optional[str] = None
    allowed_prefixes: Optional[str] = None
    gate_process: Optional[str] = None

    @field_validator("gate_process")
    @classmethod
    def _validate_gate_process(cls, v: Optional[str]) -> Optional[str]:
        if v is None or v == "":
            return v
        if not _GATE_PROCESS_PATTERN.match(v):
            raise ValueError(
                "gate_process must match ^[A-Za-z0-9_\\-\\.]+\\.exe$ "
                "(e.g. 'excel_rename_gui_v2.exe')"
            )
        return v


class PreviewExcelRequest(BaseModel):
    path: str
    # 음수 page는 음수 슬라이스/iter_rows로 전체 노출/오작동 → ge=0. page_size 상한으로
    # 단건 윈도우 비용(메모리/파싱) 제한(DoS 완화).
    page: int = Field(default=0, ge=0)
    page_size: int = Field(default=100, ge=1, le=500)


class CheckAccessRequest(BaseModel):
    path: str = ""


class BrowseFileRequest(BaseModel):
    title: str = ""
    initialdir: str = ""
    kind: str = "file"  # "file" or "directory"

router = APIRouter(prefix="/api", tags=["health"])

# 42차 W2: file-mode 관리 endpoint (add/remove/list extra-prefixes + browse-file) 전용
# admin 라우터 — 라우터 레벨 dependency로 DRY 통합 (endpoint별 _admin 파라미터 제거).
# health/preview-excel 등은 admin 무관이므로 메인 router에 유지.
admin_router = APIRouter(
    prefix="/api",
    tags=["health-admin"],
    dependencies=[Depends(require_admin)],
)


@router.get("/health")
async def health_check():
    from backend.services.file_resolver import get_resolver
    resolver = get_resolver()
    return {
        "status": "ok",
        "engine": getattr(config, "ENGINE_NAME", "DevOps Analyzer"),
        "version": getattr(config, "ENGINE_VERSION", "unknown"),
        "file_mode": resolver.mode,
    }


@router.get("/file-mode")
async def get_file_mode():
    from backend.services.file_resolver import get_resolver
    return get_resolver().get_config()


# NOTE(보안 트레이드오프): 이 endpoint는 의도적으로 admin 비게이트(일반 router)다.
# 이 배포는 비-admin 사용자가 직접 cloudium 모드를 전환해 사용하며, 그 선택을
# 재시작 후에도 유지(config/file_mode.json 영속)하는 것이 요구사항이다. admin_router로
# 옮기면 require_admin(config/admin_users.json 미등록 거부)이 비-admin의 전환을 막아
# 기능이 깨진다. 단, 영속화로 인해 비-admin도 cloudium→local 강등을 durable하게
# 고정할 수 있다는 점(read-only 경계 약화)은 알려진 한계 — 운영상 admin 전용이 필요하면
# 사용자를 admin_users.json에 등록 후 이 데코레이터를 admin_router로 전환할 것.
@router.post("/file-mode")
async def set_file_mode(body: FileModeRequest):
    from backend.services.file_resolver import switch_mode
    kwargs = body.model_dump(exclude={"mode"}, exclude_none=True)
    resolver = switch_mode(body.mode, **kwargs)
    # 모드/프리픽스 전환 시 미리보기 캐시 무효화 — 같은 path가 다른 출처(local FS vs
    # 워커 마운트)를 가리킬 수 있어 cross-mode 캐시 오염을 막는다(리뷰 C2).
    clear_preview_cache()
    # 재시작 간 모드 유지 — 선택을 config/file_mode.json에 영속 (in-memory 소실 fix).
    # allowed_prefixes/gate_process는 UI가 보낸 base 값만 저장한다. SCM 경로와
    # 사용자 추가 prefix는 각각 scm_registry.json / cloudium_extra_prefixes.json에
    # 별도 영속되어 startup(main.py lifespan)에서 재merge되므로 여기서 중복 저장 안 함.
    # 영속 실패가 모드 전환 자체를 막지 않도록 graceful.
    try:
        from backend.services.file_mode_store import save_file_mode
        save_file_mode(
            body.mode,
            allowed_prefixes=body.allowed_prefixes or "",
            gate_process=body.gate_process or "",
        )
    except Exception as _pe:  # noqa: BLE001
        import logging
        logging.getLogger("devops_api").warning("file-mode persist 실패: %s", _pe)
    # cloudium 모드 전환 시 worker 자동 시작 시도 (이미 떠 있으면 skip).
    # 결과를 응답에 포함하여 frontend가 즉시 인지 가능 (W4).
    worker_action: dict = {"action": "skipped_local_mode"}
    scm_merge: dict = {"mode": "skipped_local"}
    if body.mode == "cloudium":
        from backend.services.cloudium_worker_launcher import ensure_cloudium_worker_running
        worker_action = ensure_cloudium_worker_running()
        # N18: 등록된 모든 SCM의 path를 allowed_prefixes에 자동 merge.
        # 사용자가 SCM 수정 저장 안 해도 추적성/분석 통과.
        try:
            from backend.routers.scm import merge_all_scm_paths_to_cloudium
            scm_merge = merge_all_scm_paths_to_cloudium()
        except Exception as e:
            scm_merge = {"merged_entries": 0, "mode": "error", "error": str(e)}
    # 39차: cloudium 전환 시 사용자 추가 prefixes (영구 저장소) 자동 merge
    extra_merge: dict = {"merged_entries": 0, "mode": "skipped_local"}
    if body.mode == "cloudium":
        try:
            from backend.services.cloudium_extra_prefixes import load_extra_prefixes
            _extra = load_extra_prefixes()
            if _extra:
                _apply_extra_prefixes_to_resolver(_extra)
                extra_merge = {"merged_entries": len(_extra), "mode": "cloudium"}
            else:
                extra_merge = {"merged_entries": 0, "mode": "cloudium"}
        except Exception as e:  # noqa: BLE001
            extra_merge = {"merged_entries": 0, "mode": "error", "error": str(e)}

    # 갱신된 resolver 상태로 응답 (merge 결과 반영)
    from backend.services.file_resolver import get_resolver
    return {
        "ok": True,
        "cloudium_worker": worker_action,
        "scm_auto_merge": scm_merge,
        "extra_prefixes_merge": extra_merge,
        **get_resolver().get_config(),
    }


# ─────────────────────────────────────────────────────────────────────
# 39차 — 동적 allowed_prefixes 관리 (Cloudium 전용)
# ─────────────────────────────────────────────────────────────────────

def _apply_extra_prefix_to_resolver(prefix: str) -> dict:
    """현재 cloudium resolver에 prefix 즉시 추가 (switch_mode 재호출).

    Returns: 갱신된 resolver config.
    """
    from backend.services.file_resolver import (
        CloudiumFileResolver, get_resolver, switch_mode,
    )
    resolver = get_resolver()
    if not isinstance(resolver, CloudiumFileResolver):
        return get_resolver().get_config()
    existing = list(resolver.allowed_prefixes or [])
    if prefix not in existing:
        existing.append(prefix)
    merged = ",".join(existing)
    switch_mode("cloudium", allowed_prefixes=merged)
    return get_resolver().get_config()


def _apply_extra_prefixes_to_resolver(prefixes: list[str]) -> dict:
    """여러 prefix 일괄 적용 (startup auto-merge에서 사용)."""
    from backend.services.file_resolver import (
        CloudiumFileResolver, get_resolver, switch_mode,
    )
    resolver = get_resolver()
    if not isinstance(resolver, CloudiumFileResolver):
        return get_resolver().get_config()
    existing = list(resolver.allowed_prefixes or [])
    for p in prefixes:
        if p and p not in existing:
            existing.append(p)
    merged = ",".join(existing)
    switch_mode("cloudium", allowed_prefixes=merged)
    return get_resolver().get_config()


@admin_router.get("/file-mode/extra-prefixes")
async def list_extra_prefixes():
    """39차: 사용자 추가 cloudium prefixes 목록 (영구 저장).

    GET — read-only. Local 모드에서도 조회 가능 (영구 저장 검토용).
    """
    from backend.services.cloudium_extra_prefixes import load_extra_prefixes
    return {"prefixes": load_extra_prefixes()}


class _AddPrefixBody(BaseModel):
    """39차 add-allowed-prefix request — schemas.AddAllowedPrefixRequest 별칭.

    health.py 모듈 안 import 순환 회피용 inline 정의. 검증 로직은 동일.
    """
    prefix: str

    @field_validator("prefix")
    @classmethod
    def _validate(cls, v: str) -> str:
        if not v or len(v.strip()) == 0:
            raise ValueError("prefix 비어있음")
        if len(v) > 500:
            raise ValueError("prefix 길이 500 초과")
        if "\n" in v or "\r" in v:
            raise ValueError("줄바꿈 문자 금지")
        return v.strip()


@admin_router.post("/file-mode/add-allowed-prefix")
async def add_allowed_prefix(body: _AddPrefixBody):
    """39차: cloudium allowed_prefixes에 사용자 path 동적 추가 + 영구 저장.

    Cloudium 모드 전용 — local 모드는 400.
    영구 저장: config/cloudium_extra_prefixes.json (backend 재기동 후에도 유지).
    즉시 적용: 현재 resolver에 switch_mode 재호출로 반영.
    """
    from backend.services.cloudium_extra_prefixes import add_prefix
    from backend.services.file_resolver import CloudiumFileResolver, get_resolver

    if not isinstance(get_resolver(), CloudiumFileResolver):
        raise HTTPException(
            status_code=400,
            detail="cloudium 모드만 지원 — /api/file-mode 로 cloudium 모드 전환 후 사용",
        )

    try:
        result = add_prefix(body.prefix)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    # 영구 저장 후 즉시 resolver에 반영 (added=True일 때만)
    if result["added"]:
        resolver_config = _apply_extra_prefix_to_resolver(body.prefix)
    else:
        resolver_config = get_resolver().get_config()

    return {
        "ok": True,
        "added": result["added"],
        "prefix": result["prefix"],
        "extra_prefixes": result["prefixes"],
        **resolver_config,
    }


class _RemovePrefixBody(BaseModel):
    prefix: str

    @field_validator("prefix")
    @classmethod
    def _validate(cls, v: str) -> str:
        if not v or len(v.strip()) == 0:
            raise ValueError("prefix 비어있음")
        if len(v) > 500:
            raise ValueError("prefix 길이 500 초과")
        if "\n" in v or "\r" in v:
            raise ValueError("줄바꿈 문자 금지")
        return v.strip()


@admin_router.post("/file-mode/remove-allowed-prefix")
async def remove_allowed_prefix(body: _RemovePrefixBody):
    """39차: cloudium allowed_prefixes에서 사용자 path 제거 + 영구 저장.

    영구 저장소만 갱신 — 현재 resolver는 다음 switch_mode 또는 backend 재기동 시 반영.
    즉시 메모리에서 제거하려면 별도 set_file_mode 호출 또는 backend 재기동 권장.
    """
    from backend.services.cloudium_extra_prefixes import remove_prefix
    try:
        result = remove_prefix(body.prefix)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    return {
        "ok": True,
        "removed": result["removed"],
        "prefix": result["prefix"],
        "extra_prefixes": result["prefixes"],
    }


def _detect_preview_header_row(scan: list, is_test_spec: bool) -> int:
    """미리보기용 헤더 행 추정.

    - test-spec 시트: 기존 점수식(ID/상세 키워드 가중) 그대로 유지 → 회귀 0.
    - 일반 시트: row0가 병합 제목행(비어있는 셀 다수)일 때 더 풍부한 헤더행으로 교체.
      HSIS처럼 row0='Hardware Software'(1셀) / row1='Device,Pin,Signal...'(8셀+)인
      경우 header_row=0 고정 탓에 `r[:len(headers)]`로 데이터가 1열로 잘리던 버그 해소.
      row0보다 비어있지 않은 셀이 2개 이상 많은 행이 앞 8행 내에 있을 때만 교체(저위험).
    """
    if not scan:
        return 0
    if is_test_spec:
        best_row, best_score = 0, -1
        for ri, row in enumerate(scan[:15]):
            non_empty = sum(1 for c in row if c.strip())
            has_id = any('id' in c.lower() for c in row)
            has_detail = any(kw in ' '.join(row).lower()
                             for kw in ['title', 'description', 'method', 'environment', 'result', 'function'])
            score = non_empty + (3 if has_id else 0) + (3 if has_detail else 0)
            if score > best_score:
                best_score, best_row = score, ri
        return best_row
    # 일반 시트: 병합 헤더는 데이터행보다 셀 수가 적으므로 '셀 개수'가 아닌
    # '라벨 키워드 단어 수'로 헤더행을 판정한다. (HSIS row0=제목 1셀, row1=Device/
    # Pin/Signal… 라벨 다수 → row1 선택). 라벨 단어가 0개면 데이터만 있는 시트로
    # 보고 row0 유지(데이터행 오선택 방지). 1차=키워드수, 2차=비어있지 않은 셀 수.
    import re as _re
    HDR_WORDS = {
        'id', 'name', 'type', 'desc', 'description', 'signal', 'device', 'pin',
        'value', 'unit', 'address', 'addr', 'index', 'no', 'date', 'version',
        'author', 'method', 'result', 'function', 'parameter', 'param', 'mapping',
        'range', 'default', 'min', 'max', 'comment', 'remark', 'title', 'category',
        'environment', 'item', 'status', 'priority', 'requirement', 'req', 'class',
        '사용', '설명', '이름', '항목', '구분', '번호', '내용', '비고', '용도', '종류',
    }
    best_ri, best_key = 0, (-1, -1)
    for ri, row in enumerate(scan[:10]):
        non_empty = sum(1 for c in row if c.strip())
        kw = sum(1 for c in row
                 if c.strip() and any(w in HDR_WORDS for w in _re.findall(r'[a-z가-힣]+', c.lower())))
        key = (kw, non_empty)
        if key > best_key:
            best_key, best_ri = key, ri
    # 키워드가 전혀 없으면(데이터만 있는 시트) row0 유지.
    if best_key[0] <= 0:
        return 0
    # 깊은 행(4행+)이 헤더로 뽑혔는데 키워드가 빈약(<3)하면 데이터행 오승격으로 보고 row0 유지.
    # (row0이 약어 헤더라 키워드 0인데 데이터행이 우연히 라벨 단어를 가질 때의 오선택 방지)
    if best_ri >= 4 and best_key[0] < 3:
        return 0
    return best_ri


def _extract_docx_sheets(doc) -> List[Dict[str, Any]]:
    """docx에서 미리보기 시트를 '전체 행'으로 추출(페이지 슬라이싱은 호출부).

    반환 시트의 rows에는 전체 데이터가 담기고 total_rows/total_cols가 채워진다.
    paginate=True 시트만 호출부에서 row_start:row_end로 슬라이싱한다(other_tables는
    추출 시 100행으로 이미 캡됨 → paginate=False). 결과는 페이지 무관 → 캐시 안전.
    """
    ns_a = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
    func_tables: List[Dict[str, str]] = []
    comp_tables: List[Dict[str, str]] = []
    attr_tables: List[Dict[str, str]] = []
    other_tables: List[Dict[str, Any]] = []
    for table in doc.tables:
        first_text = table.rows[0].cells[0].text.strip() if table.rows else ""
        if "Software Component" in first_text and "Information" in first_text:
            comp_data: Dict[str, str] = {}
            for row in table.rows[1:]:
                cells = [c.text.strip() for c in row.cells]
                if len(cells) >= 3:
                    label = cells[0] or cells[1]
                    value = cells[2]
                    if label and "Software Component" not in label:
                        comp_data[label] = value
            if comp_data.get("ID") or comp_data.get("Name") or comp_data.get("SC ID") or comp_data.get("SC Name"):
                comp_tables.append(comp_data)
        elif first_text == "Attribute" and len(table.rows) <= 10:
            attr_data: Dict[str, str] = {}
            for row in table.rows:
                cells = [c.text.strip() for c in row.cells]
                if len(cells) >= 2 and cells[0]:
                    attr_data[cells[0]] = cells[1] if len(cells) > 1 else ""
            if attr_data:
                attr_tables.append(attr_data)
        elif "Function Information" in first_text:
            func_data: Dict[str, str] = {}
            for row in table.rows[1:]:
                cells_raw = row.cells
                cells = [c.text.strip() for c in cells_raw]
                if len(cells) >= 3:
                    label = cells[0] or cells[1]
                    value = cells[2]
                    if label and label != "[ Function Information ]":
                        func_data[label] = value
                    if label == "Logic Diagram":
                        for cell in cells_raw[2:]:
                            blips = cell._element.findall('.//a:blip', ns_a)
                            for b in blips:
                                embed = b.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                if embed:
                                    func_data["_image_id"] = embed
                                    break
                            if func_data.get("_image_id"):
                                break
            if func_data.get("ID") or func_data.get("Name"):
                func_tables.append(func_data)
        else:
            rows_data: List[List[str]] = []
            headers: List[str] = []
            for ri, row in enumerate(table.rows):
                raw = [c.text.strip() for c in row.cells]
                deduped: List[str] = []
                prev = None
                for c in raw:
                    if c != prev:
                        deduped.append(c)
                    prev = c
                if ri == 0:
                    headers = deduped
                else:
                    rows_data.append(deduped)
                if ri >= 100:
                    break
            if headers or rows_data:
                # other_tables는 페이지네이션 미구현 → 100행 캡 + paginate=False(무한 페이저 방지).
                other_tables.append({
                    "name": f"Table {len(other_tables)+1}",
                    "headers": headers,
                    "rows": rows_data[:100],
                    "paginate": False,
                    "has_more": False,
                    "total_rows": min(len(table.rows), 100),
                    "total_cols": len(headers),
                })

    sheets: List[Dict[str, Any]] = []

    if comp_tables:
        comp_keys = ["SC ID", "SC Name", "ID", "Name", "Description", "ASIL", "Related ID",
                     "Allocated Requirements", "Allocated Function", "Sub-Components", "Interface"]
        comp_headers = [k for k in comp_keys if any(c.get(k) for c in comp_tables)]
        extra = sorted(k for k in {k for c in comp_tables for k in c.keys()} - set(comp_headers)
                       if k and not k.isdigit() and not k.startswith("[") and not k.startswith("N/A") and len(k) > 2)
        comp_headers.extend(extra[:10])
        comp_rows = [[c.get(k, "") for k in comp_headers] for c in comp_tables]
        sheets.append({
            "name": f"Components ({len(comp_tables)})",
            "headers": comp_headers,
            "rows": comp_rows,
            "paginate": True,
            "total_rows": len(comp_tables),
            "total_cols": len(comp_headers),
        })

    if attr_tables:
        attr_keys = sorted({k for a in attr_tables for k in a.keys()})
        attr_rows = [[a.get(k, "") for k in attr_keys] for a in attr_tables]
        sheets.append({
            "name": f"Attributes ({len(attr_tables)})",
            "headers": attr_keys,
            "rows": attr_rows,
            "paginate": True,
            "total_rows": len(attr_tables),
            "total_cols": len(attr_keys),
        })

    if func_tables:
        func_keys = ["ID", "Name", "Prototype", "Description", "ASIL", "Related ID",
                     "Input Parameters", "Output Parameters", "Called Function", "Calling Function"]
        func_headers = [k for k in func_keys if any(f.get(k) for f in func_tables)]
        has_images = any(f.get("_image_id") for f in func_tables)
        if has_images:
            func_headers.append("Logic Diagram")
        func_rows = []
        for f in func_tables:
            # "Logic Diagram"만 제외(이미지 열은 별도 append). [:-1] 슬라이스는 has_images=False
            # 시 마지막 실데이터 열을 누락시켜 열 수 불일치를 유발하므로 사용하지 않는다.
            row_data = [f.get(k, "") for k in func_headers if k != "Logic Diagram"]
            if has_images:
                img_id = f.get("_image_id", "")
                row_data.append(f"__IMG__{img_id}" if img_id else "")
            func_rows.append(row_data)
        sheets.append({
            "name": f"Functions ({len(func_tables)})",
            "headers": func_headers,
            "rows": func_rows,
            "paginate": True,
            "total_rows": len(func_tables),
            "total_cols": len(func_headers),
        })

    sheets.extend(other_tables[:10])

    all_paras = [pg.text for pg in doc.paragraphs if pg.text.strip()]
    sheets.append({
        "name": "Content",
        "headers": ["Text"],
        "rows": [[t] for t in all_paras],
        "paginate": True,
        "total_rows": len(all_paras),
        "total_cols": 1,
    })

    # Table N(generic) 제외 — 의미 있는 명명 시트만 남긴다.
    useful = [s for s in sheets if not s["name"].startswith("Table ")]
    return useful if useful else sheets[:3]


def _docx_relmap(data: bytes) -> Dict[str, str]:
    """docx의 word/_rels/document.xml.rels에서 이미지 rId→Target 맵을 추출(작음)."""
    import io as _io
    import zipfile
    import xml.etree.ElementTree as ET
    out: Dict[str, str] = {}
    try:
        with zipfile.ZipFile(_io.BytesIO(data)) as z:
            rels_xml = z.read('word/_rels/document.xml.rels')
    except (KeyError, zipfile.BadZipFile):
        return out
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError:
        return out
    for rel in root:
        rid = rel.get('Id')
        target = rel.get('Target')
        rtype = (rel.get('Type') or '').lower()
        if rid and target and 'image' in rtype:
            out[rid] = target
    return out


def _docx_image_member(target: str) -> str:
    """rels Target(예: 'media/image1.png')을 zip 멤버 경로('word/media/image1.png')로."""
    import posixpath
    t = (target or '').replace('\\', '/')
    if t.startswith('/'):
        return t.lstrip('/')
    return posixpath.normpath(posixpath.join('word', t))


@router.post("/preview-excel")
def preview_excel_file(body: PreviewExcelRequest):
    """범용 문서 미리보기 — Cloudium 모드에서는 worker IPC로 read 위임.

    Cloudium 모드에서는 backend python.exe가 OS open 권한이 없으므로
    resolver.read_bytes(IPC) → BytesIO로 메모리에서 처리한다. backend가
    직접 file system을 stat/open하면 WinError 5 (액세스 거부)가 발생.
    """
    import io
    from backend.services.file_resolver import get_resolver

    file_path = body.path.strip()
    page = body.page
    page_size = body.page_size
    if not file_path:
        raise APIError(status_code=400, message="path required", code="MISSING_PATH")
    _enforce_resolver_access(file_path)

    resolver = get_resolver()
    # path 정규화만 (OS 호출 없음). resolver를 거쳐 stat — local은 직접, cloudium은 IPC
    p = Path(file_path).expanduser()
    try:
        exists = resolver.exists(file_path)
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")
    if not exists:
        raise HTTPException(status_code=404, detail=f"파일 없음: {file_path}")
    row_start = page * page_size
    row_end = row_start + page_size

    ext = p.suffix.lower()

    try:
        if ext in ('.xlsx', '.xls', '.xlsm'):
            import openpyxl
            # 바이트 캐시 → 페이지 이동 시 원본 재IPC/재read 제거(윈도우 파싱은 유지).
            data = _read_bytes_cached(resolver, file_path)
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheets = []

            for name in wb.sheetnames:
                ws = wb[name]
                # 헤더 탐지용으로 앞 15행만 스캔(윈도우 read — 7900행 시트 전체 파싱 회피).
                head_scan = [[str(c or '') for c in row]
                             for row in ws.iter_rows(min_row=1, max_row=15, values_only=True)]

                is_test_spec = any(kw in name.lower() for kw in ['test spec', 'test case', 'traceability', 'unit test'])
                header_row = _detect_preview_header_row(head_scan, is_test_spec)

                headers = list(head_scan[header_row]) if header_row < len(head_scan) else []
                while headers and not headers[-1].strip():
                    headers.pop()
                ncols = len(headers)

                # 병합 헤더 직후의 빈 행(서브헤더 병합 공백)은 데이터가 아니므로 건너뜀.
                # head_scan 기반이라 페이지 무관하게 일관(데이터 시작 오프셋 고정).
                data_offset = header_row
                for r in head_scan[header_row + 1:]:
                    if any(c.strip() for c in r):
                        break
                    data_offset += 1

                # ws.max_row(차원 태그)는 빈 행으로 부풀 수 있어(예: 7904인데 실데이터 3행)
                # 페이지 네비게이션은 total이 아닌 has_more(윈도우 뒤를 peek)로 판정한다.
                # max_row=None은 차원 태그 없는 xlsx(pandas/xlsxwriter 등) — iter_rows는
                # max_row=None을 허용(시트 끝까지)하므로 None을 0으로 강제하지 않는다.
                max_row = ws.max_row
                # 데이터는 1-based로 data_offset+2 행부터. 요청 윈도우 + peek 한 번에 read.
                win_min = data_offset + 2 + row_start
                win_max = data_offset + 1 + row_end
                PEEK = 200
                read_max = (win_max + PEEK) if max_row is None else min(win_max + PEEK, max_row)
                buf = []
                if max_row is None or win_min <= max_row:
                    for row in ws.iter_rows(min_row=win_min, max_row=read_max, values_only=True):
                        buf.append([str(c or '') for c in row])
                page_len = row_end - row_start
                window = buf[:page_len]
                peek = buf[page_len:]
                # 후행 빈 행 제거(마지막 페이지/유령 페이지 깔끔하게).
                while window and not any(c.strip() for c in window[-1]):
                    window.pop()
                data_rows = [(r[:ncols] if ncols else r) for r in window]
                has_more = any(any(c.strip() for c in r) for r in peek)
                # 시트 표시 여부(페이지 무관): head_scan에 헤더 다음 데이터가 있는가.
                # 단, 머리말(병합 제목/개정이력)이 길어 데이터가 16행+에서 시작하면 head_scan만으론
                # false-negative → 이미 읽은 buf 첫 행도 함께 검사(추가 I/O 없음).
                has_head_data = (
                    any(any(c.strip() for c in r) for r in head_scan[data_offset + 1:])
                    or (len(buf) > 0 and any(c.strip() for c in buf[0]))
                )

                sheets.append({
                    "name": name,
                    "headers": headers,
                    "rows": data_rows,
                    "has_more": has_more,
                    # total_rows: max_row 기반 상한(부정확할 수 있음 — 탭 힌트 용도).
                    "total_rows": max(0, (max_row or 0) - (data_offset + 1)),
                    "total_cols": ncols,
                    "_has_head_data": has_head_data,
                })

            wb.close()
            # 시트 목록은 페이지와 무관하게 안정적이어야 함(프론트 previewSheet 인덱스 유효 유지).
            # → head_scan 기반 데이터 유무로 판정(윈도우 행이 아님). Cover/History만 제외.
            skip = {'cover', 'history'}
            useful = [s for s in sheets if s["name"].lower() not in skip and s["_has_head_data"]]
            chosen = useful if useful else sheets
            for s in chosen:
                s.pop("_has_head_data", None)
            return {"ok": True, "filename": p.name, "sheets": chosen, "sheet_names": [s["name"] for s in chosen]}

        elif ext == '.docx':
            # 전체 시트 추출(전체 행)은 페이지 무관 → TTL 캐시. 페이지 이동 시
            # 36MB 재IPC + python-docx 전체 재파싱을 제거하고 캐시 슬라이스만 수행.
            payload = _preview_cache_get(_preview_payload, file_path)
            if payload is None:
                from report_gen.requirements import _safe_docx_open
                data = _read_bytes_cached(resolver, file_path)
                # 일부 SwUDS/SwDS docx는 임베드 이미지 파트가 깨져(BadZipFile) raw
                # docx.Document가 실패 → 500. _safe_docx_open이 손상 멤버만 우회 복원해 연다.
                doc = _safe_docx_open(io.BytesIO(data))
                payload = _extract_docx_sheets(doc)
                _preview_cache_put(_preview_payload, file_path, payload,
                                   _estimate_payload_size(payload), _path_sig(file_path))

            # 페이지 슬라이싱(캐시된 전체 시트에서). 캐시 객체는 변형하지 않는다(슬라이스=복사).
            chosen = []
            for fs in payload:
                if fs.get("paginate", True):
                    rows = fs["rows"][row_start:row_end]
                    has_more = fs["total_rows"] > row_end
                else:
                    rows = fs["rows"]            # other_tables: 추출 시 100행 캡, 페이지네이션 없음
                    has_more = False
                chosen.append({
                    "name": fs["name"],
                    "headers": fs["headers"],
                    "rows": rows,
                    "has_more": has_more,
                    "total_rows": fs["total_rows"],
                    "total_cols": fs["total_cols"],
                })
            return {"ok": True, "filename": p.name, "sheets": chosen, "sheet_names": [s["name"] for s in chosen]}

        elif ext == '.txt':
            text = resolver.read_text(file_path, encoding='utf-8')
            all_lines = text.splitlines()
            lines = all_lines[row_start:row_end]
            return {"ok": True, "filename": p.name, "sheets": [{
                "name": "Content",
                "headers": ["Line"],
                "rows": [[l] for l in lines],
                "has_more": len(all_lines) > row_end,
                "total_rows": len(all_lines),
                "total_cols": 1,
            }], "sheet_names": ["Content"]}

        elif ext in ('.csv', '.tsv'):
            import csv
            data = _read_bytes_cached(resolver, file_path)
            # 인코딩 자동 감지 — utf-8(BOM 포함) 우선, 실패 시 cp949
            try:
                text = data.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = data.decode('cp949', errors='replace')
            delim = '\t' if ext == '.tsv' else ','
            reader = csv.reader(io.StringIO(text), delimiter=delim)
            all_rows = list(reader)
            headers = all_rows[0] if all_rows else []
            data_rows = all_rows[1 + row_start:1 + row_end]
            return {"ok": True, "filename": p.name, "sheets": [{
                "name": "Sheet1",
                "headers": headers,
                "rows": data_rows,
                "has_more": max(0, len(all_rows) - 1) > row_end,
                "total_rows": max(0, len(all_rows) - 1),
                "total_cols": len(headers),
            }], "sheet_names": ["Sheet1"]}

        else:
            raise HTTPException(status_code=400, detail=f"지원하지 않는 형식: {ext}")

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/preview-image")
def preview_image(path: str, image_id: str):
    """docx 문서에서 이미지 추출 반환.

    Cloudium 모드에서는 worker IPC로 read 위임 (backend 직접 open 시
    WinError 5 발생).
    """
    import io
    from fastapi.responses import Response
    from backend.services.file_resolver import get_resolver
    _enforce_resolver_access(path)

    resolver = get_resolver()
    try:
        if not resolver.exists(path):
            raise HTTPException(status_code=404)
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")
    try:
        import mimetypes
        import zipfile
        # 바이트 캐시 + rId 맵 캐시 → 이미지 1장마다 36MB 재IPC + python-docx 전체
        # 재파싱하던 비용 제거. zip에서 해당 이미지 멤버 1개만 read(전 blob 적재 회피).
        data = _read_bytes_cached(resolver, path)
        relmap = _preview_cache_get(_preview_relmap, path)
        if relmap is None:
            relmap = _docx_relmap(data)
            _preview_cache_put(_preview_relmap, path, relmap, 0, _path_sig(path))
        target = relmap.get(image_id)
        if not target:
            raise HTTPException(status_code=404, detail="image not found")
        member = _docx_image_member(target)
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                blob = z.read(member)
        except Exception:  # noqa: BLE001 — 이미지는 비필수: 멤버 누락/CRC 손상/zlib 디코드
            # 실패(KeyError·BadZipFile·zlib.error 등) → 유효한 1x1 PNG로 graceful 대체
            # (깨진 이미지 아이콘 방지, _safe_docx_open과 동일 정책).
            from report_gen.requirements import _DOCX_PNG_1x1
            return Response(content=_DOCX_PNG_1x1, media_type='image/png')
        ct = mimetypes.guess_type(member)[0] or 'image/png'
        return Response(content=blob, media_type=ct)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@admin_router.post("/file-mode/browse-file")
async def browse_file(body: BrowseFileRequest = BrowseFileRequest()):
    """OS 파일 선택 다이얼로그.

    - cloudium 모드: worker IPC로 위임 (worker GUI에서 다이얼로그 → 클라우디움 폴더 보임)
    - local 모드: backend 자체 tkinter (local_service.pick_file)
    """
    from backend.services.file_resolver import CloudiumFileResolver, get_resolver
    resolver = get_resolver()

    if isinstance(resolver, CloudiumFileResolver):
        # worker가 권한 보유 — worker GUI에서 다이얼로그
        try:
            if body.kind == "directory":
                path = resolver.browse_directory(body.title, body.initialdir)
            else:
                path = resolver.browse_file(body.title, body.initialdir)
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except OSError as e:
            raise HTTPException(status_code=502, detail=f"worker IPC 실패: {e}")
        return {
            "ok": bool(path),
            "path": path or "",
            "via": "cloudium_worker",
            "error": None if path else "cancelled",
        }

    # LOCAL 모드 — backend 자체 tkinter
    from backend.services.local_service import pick_directory, pick_file
    if body.kind == "directory":
        path, error = pick_directory(body.title or "폴더 선택")
    else:
        path, error = pick_file(body.title or "파일 선택")
    return {
        "ok": bool(path),
        "path": path,
        "via": "local",
        "error": error or None,
    }


@router.post("/file-mode/check-access")
async def check_cloudium_access(body: CheckAccessRequest = CheckAccessRequest()):
    """경로 접근 가능 여부 확인.

    cloudium 모드는 게이트 프로세스(excel_rename_gui_v2.exe) 실행 여부도 함께 검사.
    """
    from backend.services.file_resolver import (
        CloudiumFileResolver,
        get_resolver,
        is_gate_running,
    )
    resolver = get_resolver()
    cfg = resolver.get_config()
    gate = {}
    if isinstance(resolver, CloudiumFileResolver):
        # **W7 fix**: force=True 제거. TTL 캐시 사용으로 다중 사용자 5초 polling
        # 환경에서 worker 부하 절감. 1초 stale은 사용자 인지 불가.
        gate = {
            "gate_process": resolver.gate_process,
            "gate_running": is_gate_running(resolver.gate_process),
        }
    test_path = body.path
    if test_path:
        try:
            if isinstance(resolver, CloudiumFileResolver):
                resolver._ensure_gate()
                resolver._check_allowed(test_path)
            accessible = Path(test_path).exists()
            return {"ok": True, "accessible": accessible, "path": test_path,
                    "mode": resolver.mode, **gate}
        except PermissionError as e:
            return {"ok": False, "accessible": False, "error": str(e),
                    "mode": resolver.mode, **gate}
    return {"ok": True, "mode": resolver.mode, **cfg, **gate}


@router.get("/metrics")
async def metrics():
    """Detailed system metrics for monitoring."""
    if psutil is None:
        return {
            "cpu_percent": None,
            "memory": {"total_mb": None, "used_percent": None},
            "disk": {"free_gb": None},
            "process": {"pid": os.getpid(), "threads": None},
            "note": "psutil not installed — install for full metrics",
        }
    return {
        "cpu_percent": psutil.cpu_percent(interval=0.1),
        "memory": {
            "total_mb": psutil.virtual_memory().total // (1024 * 1024),
            "used_percent": psutil.virtual_memory().percent,
        },
        "disk": {
            "free_gb": psutil.disk_usage("/").free // (1024**3),
        },
        "process": {
            "pid": os.getpid(),
            "threads": len(psutil.Process(os.getpid()).threads()),
        },
    }


@router.post("/cache/clear")
async def clear_cache():
    """Clear all in-memory caches."""
    from backend import state
    state.jenkins_progress.clear()
    state.uds_view_cache.clear()
    state.source_sections_cache.clear()
    state.session_list_cache.clear()
    clear_preview_cache()  # 리뷰 W6/W7: 미리보기 캐시도 비워 "All caches" 계약 충족 + 수동 무효화 레버
    return {"ok": True, "message": "All caches cleared"}
