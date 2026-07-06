"""Auto-generated router: jenkins"""
import json
import logging
import os
import re
import tempfile
import threading
import time
import traceback
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import unquote, urlparse

from fastapi import APIRouter, File, Form, HTTPException, Query, Response, UploadFile
from fastapi.responses import FileResponse

import config
from backend.helpers import (
    _apply_uds_view_filters,
    _build_excel_artifact_payload,
    _build_excel_artifact_summary,
    _compute_uds_mapping_summary,
    _create_jenkins_zip_file,
    _generate_docx_with_retry,
    _get_progress,
    _get_uds_view_payload_cached,
    _is_allowed_req_doc,
    _jenkins_exports_dir,
    _jenkins_logic_dir,
    _jenkins_report_publish_impl,
    _jenkins_sts_dir,
    _jenkins_suts_dir,
    _jenkins_templates_dir,
    _load_uds_meta,
    _load_vectorcast_rag,
    _normalize_jenkins_cache_root,
    _parse_component_map_file,
    _parse_path_list,
    _read_excel_artifact_sidecar,
    _resolve_cached_build_root,
    _run_impact_analysis_for_uds,
    _run_report_with_timeout,
    _safe_extract_zip,
    _safe_int,
    _save_uds_meta,
    _set_progress,
    _split_csv,
    _uds_generate_from_paths,
    _write_excel_artifact_sidecar,
    _write_residual_tbd_report,
    _write_upload_to_temp,
    build_vectorcast_metadata,
    evaluate_vectorcast_readiness,
    load_vectorcast_project_config,
)
from backend.schemas import (
    CallTreePreviewRequest,
    CodeSonarRequest,
    JenkinsBuildInfoRequest,
    JenkinsBuildsRequest,
    JenkinsCacheRequest,
    JenkinsCallTreeRequest,
    JenkinsImpactTriggerRequest,
    JenkinsJobsRequest,
    JenkinsPublishRequest,
    JenkinsRagQueryRequest,
    JenkinsReportRequest,
    JenkinsReportZipRequest,
    JenkinsScmInfoRequest,
    JenkinsServerFilesRequest,
    JenkinsSourceDownloadRequest,
    JenkinsSyncLocalRequest,
    JenkinsSyncRequest,
    ReportZipRequest,
    UdsDeleteRequest,
    UdsDiffRequest,
    UdsLabelRequest,
    UdsPublishRequest,
    UdsTraceabilityMatrixRequest,
)
from backend.services.assistant_service import read_report_bundle
from backend.services.call_tree import (
    build_call_tree,
    build_call_tree_precise,
    call_tree_to_csv,
    call_tree_to_html,
)
from backend.services.files import list_log_candidates, read_csv_rows, tail_text
from backend.services.jenkins_client import JenkinsClient
from backend.services.jenkins_helpers import _detect_reports_dir, _job_slug, _safe_artifact_path
from backend.services.jenkins_service import (
    _source_is_complete,  # 체크아웃 완전성 판정 단일 출처(.source_complete 센티널) — 콜트리 폴백에서 재사용
    ensure_source_checkout,
    get_build_info,
    list_builds,
    list_cached_builds,
    list_jobs,
    sync_jenkins_artifacts,
    sync_local_reports,
)
from backend.services.local_service import run_svn, svn_info_url
from backend.services.paths import is_under_any, safe_resolve_under
from backend.services.report_parsers import (
    build_report_summary,
    find_jenkins_source_root,
)
from backend.user_context import wrap_with_user

# 명시 RelatedID 링크 테이블 파생(P1) — 기존 빌더/생성기 수정 없이 그 출력만 소비.
from report_gen.trace_link_table import build_link_table

# 명시 RelatedID 링크 테이블 파생(P1) — 기존 빌더/생성기를 수정하지 않고 그 출력만 소비.
from report_generator import (
    _build_req_map_from_doc_paths,
    enrich_function_details_with_docs,
    generate_asil_related_confidence_report,
    generate_called_calling_accuracy_report,
    generate_swcom_context_report,
    generate_uds_constraints_report,
    generate_uds_field_quality_gate_report,
    generate_uds_function_mapping,
    generate_uds_preview_html,
    generate_uds_requirements_compare,
    generate_uds_requirements_from_docs,
    generate_uds_requirements_mapping,
    generate_uds_requirements_preview,
    generate_uds_source_sections,
    generate_uds_traceability_matrix,
    generate_uds_validation_report,
)

try:
    from workflow.rag import _read_text_from_file, get_kb, ingest_external_sources
except ImportError:
    _read_text_from_file = None
    get_kb = None
    ingest_external_sources = None
try:
    from workflow.uds_ai import generate_uds_ai_sections
except ImportError:
    generate_uds_ai_sections = None
from workflow.change_trigger import build_registry_trigger
from workflow.impact_jobs import start_impact_job
from workflow.impact_orchestrator import run_impact_update

repo_root = Path(__file__).resolve().parents[2]


router = APIRouter()
_logger = logging.getLogger("devops_api")
_api_logger = _logger


def _write_uds_payload_sidecar(out_path: Path, uds_payload: Dict[str, Any]) -> Optional[Path]:
    try:
        details = uds_payload.get("function_details")
        if not isinstance(details, dict):
            return None
        summary = uds_payload.get("summary")
        if not isinstance(summary, dict):
            summary = {}
        summary["mapping"] = _compute_uds_mapping_summary(details)
        uds_payload["summary"] = summary
        sidecar = out_path.with_suffix(".payload.json")
        payload = {
            "docx_path": str(out_path),
            "summary": summary,
            "function_details": details,
        }
        sidecar.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return sidecar
    except Exception as exc:
        _logger.warning("jenkins uds payload sidecar write skipped: %s", exc)
        return None


def _build_jenkins_excel_output(cache_root: str, category: str, stem: str, template_path: Optional[str]) -> Tuple[str, Path]:
    target_dir = _jenkins_sts_dir(cache_root) if category == "sts" else _jenkins_suts_dir(cache_root)
    target_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = Path(template_path).suffix.lower() if template_path and Path(template_path).suffix.lower() in {".xlsx", ".xlsm"} else ".xlsx"
    filename = f"{stem}_{ts}{suffix}"
    return filename, target_dir / filename


def _excel_media_type(file_path: Path) -> str:
    if file_path.suffix.lower() == ".xlsm":
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_label(job_url: str, cache_root: str, build_selector: str) -> str:
    build_root = _resolve_cached_build_root(job_url, cache_root, build_selector)
    if not build_root:
        return str(build_selector or "").strip()
    m = re.search(r"build_(\d+)$", build_root.name, flags=re.I)
    if m:
        return f"Build {m.group(1)}"
    return build_root.name


def _build_jenkins_vectorcast_response(
    *,
    job_url: str,
    cache_root: str,
    build_selector: str,
    package_dir: Path,
    package_name: str,
    manifest: Dict[str, Any],
    project_config: Dict[str, Any],
    units: List[str],
) -> Dict[str, Any]:
    metadata = build_vectorcast_metadata(
        project_config=project_config,
        source_root=str(project_config.get("source_root") or ""),
        units=units,
    )
    readiness = evaluate_vectorcast_readiness(metadata)
    return {
        "ok": True,
        "job_url": job_url,
        "cache_root": cache_root,
        "build_label": _build_label(job_url, cache_root, build_selector),
        "package_dir": str(package_dir),
        "package_name": package_name,
        "manifest": manifest,
        "files": sorted(str(p.name) for p in package_dir.iterdir() if p.is_file()),
        "project_config": metadata,
        "readiness": readiness,
    }


def _infer_build_label_for_artifact(job_url: str, cache_root: str, artifact_path: Path, build_selector: str) -> str:
    direct = _build_label(job_url, cache_root, build_selector)
    if direct and direct != str(build_selector or "").strip():
        return direct
    base = Path(cache_root).expanduser().resolve() if cache_root else (Path.home() / ".devops_pro_cache").resolve()
    job_root = (base / "jenkins" / _job_slug(job_url)).resolve()
    if not job_root.exists():
        return direct
    build_dirs = [p for p in job_root.glob("build_*") if p.is_dir()]
    if not build_dirs:
        return direct
    if len(build_dirs) == 1:
        m = re.search(r"build_(\d+)$", build_dirs[0].name, flags=re.I)
        return f"Build {m.group(1)}" if m else build_dirs[0].name
    artifact_mtime = artifact_path.stat().st_mtime
    def _score(path: Path) -> Tuple[float, float]:
        delta = abs(path.stat().st_mtime - artifact_mtime)
        prefer_past = 0.0 if path.stat().st_mtime <= artifact_mtime else 1.0
        return (prefer_past, delta)
    best = sorted(build_dirs, key=_score)[0]
    m = re.search(r"build_(\d+)$", best.name, flags=re.I)
    return f"Build {m.group(1)}" if m else best.name


def _load_excel_artifact_payload(
    file_path: Path,
    artifact_type: str,
    *,
    download_url: str,
    preview_url: str,
    build_label: str = "",
) -> Dict[str, Any]:
    payload = _read_excel_artifact_sidecar(file_path)
    if not payload:
        payload = _repair_excel_artifact_payload(
            file_path,
            artifact_type,
            download_url=download_url,
            preview_url=preview_url,
        )
    payload["filename"] = file_path.name
    payload["output_path"] = str(file_path)
    payload["download_url"] = download_url
    payload["preview_url"] = preview_url
    if build_label:
        payload["build_label"] = build_label
    if not str(payload.get("validation_report_path") or "").strip():
        validation_path = file_path.with_suffix(".validation.md")
        payload["validation_report_path"] = str(validation_path) if validation_path.exists() else ""
    if not str(payload.get("residual_report_path") or "").strip():
        residual_path = file_path.with_suffix(".residual.md")
        if not residual_path.exists():
            residual_path = file_path.with_suffix(".residual_tbd.md")
        payload["residual_report_path"] = str(residual_path) if residual_path.exists() else ""
    if not isinstance(payload.get("summary"), dict):
        payload["summary"] = _build_excel_artifact_summary(artifact_type, payload.get("raw_result") or payload)
    if build_label:
        payload["summary"]["build_label"] = build_label
    return payload


def _repair_excel_artifact_payload(
    file_path: Path,
    artifact_type: str,
    *,
    download_url: str,
    preview_url: str,
) -> Dict[str, Any]:
    kind = str(artifact_type or "").strip().lower()
    result: Dict[str, Any] = {
        "ok": True,
        "output_path": str(file_path),
        "filename": file_path.name,
        "download_url": download_url,
        "preview_url": preview_url,
    }
    validation_path = file_path.with_suffix(".validation.md")
    if validation_path.exists():
        result["validation_report_path"] = str(validation_path)
    residual_path = file_path.with_suffix(".residual_tbd.md")
    if residual_path.exists():
        result["residual_report_path"] = str(residual_path)
    try:
        if kind == "sts":
            from generators.suts import validate_sts_xlsm
            validation = validate_sts_xlsm(str(file_path))
            stats = validation.get("stats", {}) if isinstance(validation, dict) else {}
            result["validation"] = validation
            result["test_case_count"] = int(stats.get("tc_count") or 0)
        elif kind == "suts":
            from generators.suts import validate_suts_xlsm
            validation = validate_suts_xlsm(str(file_path))
            stats = validation.get("stats", {}) if isinstance(validation, dict) else {}
            result["validation"] = validation
            result["test_case_count"] = int(stats.get("tc_count") or 0)
            result["total_sequences"] = int(stats.get("seq_count") or 0)
            result["quality_report"] = {
                "avg_sequences_per_tc": float(stats.get("avg_seq_per_tc") or 0),
            }
    except Exception:
        result.setdefault("validation", {})
    payload = _build_excel_artifact_payload(
        kind,
        result,
        output_path=str(file_path),
        filename=file_path.name,
        download_url=download_url,
        preview_url=preview_url,
    )
    _write_excel_artifact_sidecar(file_path, kind, payload)
    return payload

@router.post("/api/jenkins/jobs")
def jenkins_jobs(req: JenkinsJobsRequest) -> Dict[str, Any]:
    base_url = (req.base_url or "").strip().rstrip("/")
    username = (req.username or "").strip()
    api_token = (req.api_token or "").strip()
    _api_logger.info("[jenkins/jobs] base_url=%s, username=%s, token_len=%d, verify_tls=%s", base_url, username, len(api_token), req.verify_tls)
    if not api_token:
        raise HTTPException(status_code=400, detail="API Token이 비어 있습니다. 토큰을 입력해주세요.")
    try:
        jobs = list_jobs(
            base_url=base_url,
            username=username,
            api_token=api_token,
            recursive=req.recursive,
            max_depth=req.max_depth,
            verify_tls=req.verify_tls,
        )
        return {"jobs": jobs}
    except Exception as e:
        error_msg = str(e)
        import traceback
        traceback.print_exc()  # 상세한 스택 트레이스 출력
        _api_logger.error("[jenkins/jobs] error: %s", error_msg)
        
        # 에러 타입에 따라 적절한 HTTP 상태 코드 반환
        error_lower = error_msg.lower()
        if "401" in error_lower or "unauthorized" in error_lower:
            status_code = 401
        elif "403" in error_lower or "forbidden" in error_lower:
            status_code = 403
        elif "404" in error_lower or "not found" in error_lower:
            status_code = 404
        elif "timeout" in error_lower or "timed out" in error_lower:
            status_code = 504
        elif "connection" in error_lower or "refused" in error_lower:
            status_code = 503
        else:
            status_code = 500
            
        raise HTTPException(
            status_code=status_code,
            detail=f"Jenkins 프로젝트 목록 조회 실패: {error_msg}"
        )


@router.post("/api/jenkins/builds")
def jenkins_builds(req: JenkinsBuildsRequest) -> Dict[str, Any]:
    job_url = (req.job_url or "").strip().rstrip("/")
    username = (req.username or "").strip()
    api_token = (req.api_token or "").strip()
    _api_logger.info("[jenkins/builds] job_url=%s, username=%s, token_len=%d", job_url, username, len(api_token))
    if not job_url:
        raise HTTPException(status_code=400, detail="Job URL이 비어 있습니다. Job을 선택해주세요.")
    if not api_token:
        raise HTTPException(status_code=400, detail="API Token이 비어 있습니다.")
    try:
        builds = list_builds(
            job_url=job_url,
            username=username,
            api_token=api_token,
            limit=req.limit,
            verify_tls=req.verify_tls,
        )
        _api_logger.info("[jenkins/builds] success: builds=%d", len(builds))
        return {"builds": builds}
    except Exception as e:
        err = str(e)
        _logger.error("[jenkins/builds] 오류: %s", err)
        import traceback
        traceback.print_exc()
        if "401" in err.lower() or "unauthorized" in err.lower():
            raise HTTPException(status_code=401, detail=f"Jenkins 빌드 목록 조회 실패: {err}")
        if "403" in err.lower() or "forbidden" in err.lower():
            raise HTTPException(status_code=403, detail=f"Jenkins 빌드 목록 조회 실패: {err}")
        if "404" in err.lower() or "not found" in err.lower():
            raise HTTPException(status_code=404, detail=f"Jenkins 빌드 목록 조회 실패: {err}")
        if "timeout" in err.lower() or "timed out" in err.lower():
            raise HTTPException(status_code=504, detail=f"Jenkins 빌드 목록 조회 실패: {err}")
        raise HTTPException(status_code=500, detail=f"Jenkins 빌드 목록 조회 실패: {err}")


@router.post("/api/jenkins/build-info")
def jenkins_build_info(req: JenkinsBuildInfoRequest) -> Dict[str, Any]:
    data = get_build_info(
        job_url=req.job_url,
        username=req.username,
        api_token=req.api_token,
        build_selector=req.build_selector,
        verify_tls=req.verify_tls,
    )
    return data


@router.get("/api/jenkins/progress")
def jenkins_progress(
    action: str,
    job_url: str,
    build_selector: str = "lastSuccessfulBuild",
    job_id: str = "",
) -> Dict[str, Any]:
    data = _get_progress(action, job_url, build_selector, job_id)
    return {"ok": bool(data), "progress": data}


@router.post("/api/jenkins/sync")
def jenkins_sync(req: JenkinsSyncRequest) -> Dict[str, Any]:
    job_url = req.job_url
    build_selector = req.build_selector
    _set_progress(
        "sync",
        job_url,
        build_selector,
        {
            "stage": "start",
            "percent": 1,
            "message": "동기화 준비 중",
            "done": False,
            "error": "",
        },
    )

    def _progress_cb(stage: str, data: Dict[str, Any]) -> None:
        if not isinstance(data, dict):
            data = {}
        percent = 5
        message = ""
        if stage == "list_artifacts":
            percent = 8
            message = f"아티팩트 목록 조회 ({data.get('count', 0)}개)"
        elif stage == "download_start":
            percent = 12
            message = f"아티팩트 다운로드 시작 ({data.get('total', 0)}개)"
        elif stage == "download":
            cur = int(data.get("current") or 0)
            total = max(1, int(data.get("total") or 1))
            percent = 12 + int((cur / total) * 60)
            message = f"다운로드 {cur}/{total}: {data.get('file') or ''}".strip()
        elif stage == "download_console":
            percent = 75
            message = "콘솔 로그 다운로드"
        elif stage == "scan_prepare":
            percent = 80
            message = "리포트 스캔 준비"
        elif stage == "scan_files":
            cur = int(data.get("current") or 0)
            total = max(1, int(data.get("max_files") or 1))
            percent = 80 + int(min(cur, total) / total * 15)
            message = f"리포트 스캔 {cur}개 파일"
        elif stage == "scan_cached":
            percent = 90
            message = "리포트 스캔 캐시 사용"
        elif stage == "scm_clone":
            percent = 78
            message = "SCM 소스 체크아웃"
        elif stage == "scm_done":
            percent = 79
            message = "SCM 소스 체크아웃 완료"
        elif stage == "scm_failed":
            percent = 79
            message = "SCM 소스 체크아웃 실패"
        elif stage == "scan_start":
            percent = 82
            message = "리포트 스캔/요약 생성"
        elif stage == "scan_done":
            percent = 95
            message = "리포트 요약 완료"
        _set_progress(
            "sync",
            job_url,
            build_selector,
            {
                "stage": stage,
                "percent": percent,
                "message": message,
                "current": data.get("current"),
                "total": data.get("total"),
                "file": data.get("file"),
            },
        )

    try:
        build_info, build_root, reports_dir, downloaded, artifacts = sync_jenkins_artifacts(
            job_url=req.job_url,
            username=req.username,
            api_token=req.api_token,
            cache_root=_normalize_jenkins_cache_root(req.cache_root),
            verify_tls=req.verify_tls,
            build_selector=req.build_selector,
            patterns=req.patterns,
            progress_cb=_progress_cb,
            scan_mode=req.scan_mode,
            scan_max_files=req.scan_max_files,
            scm_username=req.scm_username,
            scm_id=req.scm_id,
            force=req.force,
        )
        checkout = build_info.get("checkout", {}) if isinstance(build_info, dict) else {}
        checkout_ok = bool(checkout.get("ok"))
        checkout_err = str(checkout.get("error") or "")
        done_message = "동기화 완료" if checkout_ok else (
            f"동기화 완료 (SCM 체크아웃 실패: {checkout_err or 'unknown'})"
        )
        _set_progress(
            "sync",
            job_url,
            build_selector,
            {
                "stage": "done",
                "percent": 100,
                "message": done_message,
                "done": True,
                "checkout_ok": checkout_ok,
                "checkout_error": checkout_err,
            },
        )
        return {
            "build_info": build_info,
            "build_root": str(build_root),
            "reports_dir": str(reports_dir),
            "downloaded": downloaded,
            "artifacts": artifacts,
            "data": read_report_bundle(reports_dir),
        }
    except Exception as exc:
        tb = traceback.format_exc()
        # Log full traceback to server console for debugging
        _api_logger.error("[sync] %s failed: %s\n%s", job_url, exc, tb)
        # Include last 2 frames of traceback in error for frontend visibility
        tb_lines = [ln for ln in tb.splitlines() if ln.strip()]
        tail = " | ".join(tb_lines[-4:]) if len(tb_lines) > 2 else str(exc)
        _set_progress(
            "sync",
            job_url,
            build_selector,
            {
                "stage": "error",
                "percent": 100,
                "message": "동기화 실패",
                "done": True,
                "error": f"{exc} [at: {tail}]",
                "error_detail": tb,
            },
        )
        raise


@router.post("/api/jenkins/sync-async")
def jenkins_sync_async(req: JenkinsSyncRequest) -> Dict[str, Any]:
    job_url = req.job_url
    build_selector = req.build_selector
    job_id = uuid.uuid4().hex
    _set_progress(
        "sync",
        job_url,
        build_selector,
        {
            "stage": "start",
            "percent": 1,
            "message": "동기화 준비 중",
            "done": False,
            "error": "",
        },
        job_id=job_id,
    )

    def _progress_cb(stage: str, data: Dict[str, Any]) -> None:
        if not isinstance(data, dict):
            data = {}
        percent = 5
        message = ""
        if stage == "list_artifacts":
            percent = 8
            message = f"아티팩트 목록 조회 ({data.get('count', 0)}개)"
        elif stage == "download_start":
            percent = 12
            message = f"아티팩트 다운로드 시작 ({data.get('total', 0)}개)"
        elif stage == "download":
            cur = int(data.get("current") or 0)
            total = max(1, int(data.get("total") or 1))
            percent = 12 + int((cur / total) * 60)
            message = f"다운로드 {cur}/{total}: {data.get('file') or ''}".strip()
        elif stage == "download_console":
            percent = 75
            message = "콘솔 로그 다운로드"
        elif stage == "scan_prepare":
            percent = 80
            message = "리포트 스캔 준비"
        elif stage == "scan_files":
            cur = int(data.get("current") or 0)
            total = max(1, int(data.get("max_files") or 1))
            percent = 80 + int(min(cur, total) / total * 15)
            message = f"리포트 스캔 {cur}개 파일"
        elif stage == "scan_cached":
            percent = 90
            message = "리포트 스캔 캐시 사용"
        elif stage == "scm_clone":
            percent = 78
            message = "SCM 소스 체크아웃"
        elif stage == "scm_done":
            percent = 79
            message = "SCM 소스 체크아웃 완료"
        elif stage == "scm_failed":
            percent = 79
            message = "SCM 소스 체크아웃 실패"
        elif stage == "scan_start":
            percent = 82
            message = "리포트 스캔/요약 생성"
        elif stage == "scan_done":
            percent = 95
            message = "리포트 요약 완료"
        _set_progress(
            "sync",
            job_url,
            build_selector,
            {
                "stage": stage,
                "percent": percent,
                "message": message,
                "current": data.get("current"),
                "total": data.get("total"),
                "file": data.get("file"),
            },
            job_id=job_id,
        )

    def _run_sync() -> None:
        try:
            build_info, _br, _rd, _dl, _arts = sync_jenkins_artifacts(
                job_url=req.job_url,
                username=req.username,
                api_token=req.api_token,
                cache_root=_normalize_jenkins_cache_root(req.cache_root),
                verify_tls=req.verify_tls,
                build_selector=req.build_selector,
                patterns=req.patterns,
                progress_cb=_progress_cb,
                scan_mode=req.scan_mode,
                scan_max_files=req.scan_max_files,
                scm_username=req.scm_username,
                scm_id=req.scm_id,
                force=req.force,
            )
            checkout = build_info.get("checkout", {}) if isinstance(build_info, dict) else {}
            checkout_ok = bool(checkout.get("ok"))
            checkout_err = str(checkout.get("error") or "")
            done_message = "동기화 완료" if checkout_ok else (
                f"동기화 완료 (SCM 체크아웃 실패: {checkout_err})"
            )
            _set_progress(
                "sync",
                job_url,
                build_selector,
                {
                    "stage": "done",
                    "percent": 100,
                    "message": done_message,
                    "done": True,
                    "checkout_ok": checkout_ok,
                    "checkout_error": checkout_err,
                },
                job_id=job_id,
            )
        except Exception as exc:
            tb = traceback.format_exc()
            # Log full traceback to server console for debugging
            _api_logger.error("[sync-async] %s failed: %s\n%s", job_url, exc, tb)
            # Include last traceback frames in error message for frontend visibility
            tb_lines = [ln for ln in tb.splitlines() if ln.strip()]
            tail = " | ".join(tb_lines[-4:]) if len(tb_lines) > 2 else str(exc)
            _set_progress(
                "sync",
                job_url,
                build_selector,
                {
                    "stage": "error",
                    "percent": 100,
                    "message": "동기화 실패",
                    "done": True,
                    "error": f"{exc} [at: {tail}]",
                    "error_detail": tb,
                },
                job_id=job_id,
            )

    t = threading.Thread(target=wrap_with_user(_run_sync), daemon=True)
    t.start()
    return {"ok": True, "job_id": job_id}


@router.post("/api/jenkins/sync-local")
def jenkins_sync_local(req: JenkinsSyncLocalRequest) -> Dict[str, Any]:
    build_info, build_root, reports_dir, downloaded, artifacts = sync_local_reports(
        job_url=req.job_url,
        local_reports_dir=Path(req.local_reports_dir),
    )
    return {
        "build_info": build_info,
        "build_root": str(build_root),
        "reports_dir": str(reports_dir),
        "downloaded": downloaded,
        "artifacts": artifacts,
        "data": read_report_bundle(reports_dir),
    }


@router.post("/api/jenkins/cache")
def jenkins_cache(req: JenkinsCacheRequest) -> Dict[str, Any]:
    rows = list_cached_builds(job_url=req.job_url, cache_root=_normalize_jenkins_cache_root(req.cache_root))
    return {"builds": rows}


@router.post("/api/jenkins/report/data")
def jenkins_report_data(req: JenkinsReportRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    reports_dir = (build_root / "reports").resolve()
    return read_report_bundle(reports_dir)


@router.post("/api/jenkins/rag/query")
def jenkins_rag_query(req: JenkinsRagQueryRequest) -> Dict[str, Any]:
    query = str(req.query or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="query required")
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    report_dir = _detect_reports_dir(build_root)
    kb = get_kb(report_dir)
    top_k = max(1, int(req.top_k or 5))
    categories = [str(c).strip() for c in (req.categories or []) if str(c).strip()]
    rows = kb.search(query, top_k=top_k, categories=categories or None)
    items = []
    for row in rows:
        items.append(
            {
                "title": row.get("error_raw") or "",
                "category": row.get("category") or "",
                "source_file": row.get("source_file") or "",
                "score": float(row.get("score") or 0.0),
                "snippet": str(row.get("fix") or "")[:1200],
            }
        )
    return {"ok": True, "items": items}


@router.post("/api/jenkins/report/complexity")
def jenkins_complexity(req: JenkinsReportRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    rows: List[Dict[str, Any]] = []
    if build_root:
        report_dir = _detect_reports_dir(build_root)
        rows = read_csv_rows(report_dir / "complexity.csv")
    if rows:
        return {"rows": rows, "source": "jenkins"}
    # 빌드 산출물에 complexity.csv가 없으면(VectorCAST 결과가 cloudium SCM 경로에만 있는 흔한
    # 케이스) SCM 등록 VectorCAST 폴더의 AggregateCoverage per-function 복잡도로 폴백한다.
    # cloudium 폴더 파싱은 30분 TTL 캐시 — 동일 폴더를 vectorcast-rag(-async)로 한 번 로드했으면
    # 즉시 응답한다. (still-present sync /report/vectorcast-rag와 동일한 동기 계약)
    # 주의: 프론트 loadComplexity는 vcast_log_paths를 보내지 않아 UI에서는 이 폴백이 트리거되지
    # 않는다(UI는 비동기 vectorcast-rag-async가 실어 보낸 complexity_rows를 복잡도 표에 표시).
    # 이 동기 경로는 vcast_log_paths를 명시적으로 넘기는 프로그래매틱 호출 전용 — cold 캐시 시
    # ~100s 블로킹 가능하나 호출자 의도이며 이후 TTL 캐시로 즉시.
    cloud_paths = _collect_vcast_paths(req)
    if cloud_paths:
        cloud = _load_vectorcast_rag_from_cloudium_multi(cloud_paths)
        cr = (cloud.get("complexity_rows") if isinstance(cloud, dict) else None) or []
        if cr:
            return {"rows": cr, "source": "cloudium"}
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    return {"rows": rows, "source": "jenkins"}


@router.post("/api/jenkins/report/docs")
def jenkins_docs(req: JenkinsReportRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    report_dir = _detect_reports_dir(build_root)
    doc_path = (report_dir / "docs" / "html" / "index.html").resolve()
    if not doc_path.exists():
        return {"ok": False, "html": ""}
    return {"ok": True, "html": doc_path.read_text(encoding="utf-8", errors="ignore")}


@router.post("/api/jenkins/report/logs")
def jenkins_logs(req: JenkinsReportRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    report_dir = _detect_reports_dir(build_root)
    logs = list_log_candidates(report_dir)
    out = {k: [str(p.relative_to(report_dir)) for p in v] for k, v in logs.items()}
    return {"logs": out}


@router.post("/api/jenkins/report/logs/read")
def jenkins_logs_read(req: JenkinsReportRequest, path: str) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    report_dir = _detect_reports_dir(build_root)
    try:
        target = safe_resolve_under(report_dir, path)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid path")
    if not target.exists():
        raise HTTPException(status_code=404, detail="log not found")
    return {"path": str(target), "text": tail_text(target)}


@router.post("/api/jenkins/report/summary")
def jenkins_report_summary(req: JenkinsReportRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    report_dir = _detect_reports_dir(build_root)
    return build_report_summary(report_dir, project_root=repo_root)


def _load_vectorcast_rag_from_cloudium(path: str) -> Dict[str, Any]:
    """Jenkins 캐시에 VectorCAST RAG가 없을 때 Cloudium 경로에서 읽는다 (worker IPC).

    부트로더/FBL 등 테스트 결과가 Jenkins와 별도로 나올 수 있어, SwUT/SwIT 로그처럼
    사용자가 지정한 경로에서 resolver(local=직접 / cloudium=worker)로 vectorcast_rag.json
    을 읽는다. path가 .json이면 직접, 폴더면 표준 하위 경로들을 탐색.
    """
    import json as _json
    p = str(path or "").strip()
    if not p:
        return {}
    from backend.services.file_resolver import get_resolver
    from backend.services.resolver_helpers import enforce_resolver_access
    resolver = get_resolver()
    base = p.rstrip("/\\")
    cands = [p] if p.lower().endswith(".json") else [
        base + "/vectorcast_rag.json",
        base + "/vectorcast_rag/vectorcast_rag.json",
        base + "/report/vectorcast_rag/vectorcast_rag.json",
    ]
    for cand in cands:
        try:
            enforce_resolver_access(cand)
            if not resolver.exists(cand):
                continue
            obj = _json.loads(resolver.read_bytes(cand).decode("utf-8", "ignore"))
            if isinstance(obj, dict) and obj:
                return obj
        except (PermissionError, OSError):
            continue
        except Exception:
            continue
    # vectorcast_rag.json(우리 산출물)이 없으면 — cloudium 원본 VectorCAST 리포트
    # 폴더로 보고 직접 파싱한다 (SwUT/SwIT 빌더의 검증된 추출기 재사용). path가
    # .json이면 단일 파일 지정이므로 폴더 파싱 불가 → {}.
    if not p.lower().endswith(".json"):
        return _parse_vcast_logs_from_cloudium_folder(p)
    return {}


# 2026-06-24 — SCM(cloudium) IT 함수콜 보강. AggregateCoverage(구문/분기/MC-DC) 리포트에는
# 함수콜(Function Called) 커버리지가 없고 VectorCAST Metric report HTML에만 있다. 그동안
# 함수콜은 Jenkins 빌드 산출물에서만 나왔는데, SwITCV 빌더가 쓰는 parse_hmr_html을 재사용해
# SCM cloudium 경로에서도 동일하게 it_metrics.grand_totals(function_calls/functions)를 제공한다.
_MAX_METRIC_HTML_SCAN = 40   # 폴더당 Metric report 후보 HTML read 상한(runaway IPC 방지).


def _aggregate_it_function_calls(
    html_bytes_list: List[bytes],
) -> "tuple[Dict[str, Any], Dict[str, Dict[str, int]]]":
    """IT VectorCAST Metric report HTML들 → it_metrics.grand_totals + 함수명→함수콜 셀 map.

    각 HTML을 parse_hmr_html로 파싱(metric 양식 아니면 ok=False → skip), Function Called
    (covered_calls/total_calls)와 Functions(functions_covered/functions_total)를 전 함수 합산해
    grand_totals를 만든다. function_calls/functions 둘 다 total>0일 때만 키를 채운다(0% 위장 금지).

    Returns:
        (grand_totals, fc_by_name) — grand_totals는 {"function_calls": {covered,total,rate},
        "functions": {covered,total,rate}} 일부/전체. fc_by_name은 entries 병합용
        {함수명: {covered,total}}.
    """
    from backend.services.vcast_hmr_parser import parse_hmr_html
    fc_cov = fc_tot = fn_cov = fn_tot = 0
    fc_by_name: Dict[str, Dict[str, int]] = {}
    _seen_units: set = set()  # (함수명, unit_file) — multi-HTML 동일 함수 이중집계 방지
    for hb in html_bytes_list or []:
        try:
            hr = parse_hmr_html(hb)
        except Exception:  # noqa: BLE001 — 개별 HTML 파싱 실패는 skip
            continue
        if not getattr(hr, "ok", False):
            continue
        # metrics_by_name(유닛파일별 버킷)를 순회 — hr.metrics(이름 dedup, first-wins)는
        # 동명 다른-파일 static 함수(osif.c::Init vs canif.c::Init)를 첫 개만 남기고
        # 나머지를 합산에서 누락시켜 과소집계한다. by-name 버킷을 (함수,파일) 단위로 합산.
        for name, mlist in (getattr(hr, "metrics_by_name", None) or {}).items():
            nm = (name or "").strip()
            for m in (mlist or []):
                _uk = (nm, str(getattr(m, "unit_file", "") or ""))
                if _uk in _seen_units:  # 다른 HTML의 동일 (함수,파일) 중복은 skip
                    continue
                _seen_units.add(_uk)
                _cc = int(getattr(m, "covered_calls", 0) or 0)
                _tc = int(getattr(m, "total_calls", 0) or 0)
                fc_cov += _cc
                fc_tot += _tc
                fn_cov += int(getattr(m, "functions_covered", 0) or 0)
                fn_tot += int(getattr(m, "functions_total", 0) or 0)
                if nm:
                    _agg = fc_by_name.setdefault(nm, {"covered": 0, "total": 0})
                    _agg["covered"] += _cc
                    _agg["total"] += _tc
    grand: Dict[str, Any] = {}
    if fc_tot:
        grand["function_calls"] = {"covered": fc_cov, "total": fc_tot, "rate": round(fc_cov / fc_tot, 4)}
    if fn_tot:
        grand["functions"] = {"covered": fn_cov, "total": fn_tot, "rate": round(fn_cov / fn_tot, 4)}
    return grand, fc_by_name


# cloudium 원본 리포트 폴더 파싱은 무겁다(폴더당 수십 env × ExecutionResult HTML
# worker IPC read + BS4 — 실측 ~100s). cloudium은 read-only라 리포트가 릴리스 단위로
# 정적이므로 폴더 경로 기준 TTL 캐시로 반복 매트릭스 로드를 즉시화한다. 비어있는 결과
# ({})는 캐시하지 않아 worker 일시 장애 후 재시도를 허용한다.
_VCAST_CLOUDIUM_PARSE_CACHE: Dict[str, "tuple[float, Dict[str, Any]]"] = {}
_VCAST_CLOUDIUM_PARSE_LOCK = threading.Lock()
_VCAST_CLOUDIUM_PARSE_TTL = 1800.0  # 30분


def _parse_vcast_logs_from_cloudium_folder(path: str) -> Dict[str, Any]:
    """vectorcast_rag.json이 없는 cloudium 원본 리포트 폴더에서 실행결과를 직접 파싱.

    사용자가 설정 SCM '연결 문서 경로'에 등록하는 폴더는 우리 산출물(vectorcast_rag.json)
    이 아니라 VectorCAST 원본 리포트 폴더(`TestCaseData/`·`Execution/`·`Aggregate/`)다.
    SwUT/SwIT 빌더가 이미 cloudium에서 사용하는 `swut_input_adapter`의 레이아웃 자동
    감지(SWTE/VC2025) + 실행결과 추출기를 재사용해 pass/fail testcase 행을 만든다.

    무거운 TestCaseData(파일당 100+ 테이블)·coverage 파싱은 건너뛰고 ExecutionResult
    HTML만 읽어 매트릭스 join용 test_rows(subprogram/testcase/unit/result)를 조립한다.
    subprogram은 tc_name(`SwUFn_3401.001`)에서 함수 id(`SwUFn_3401`)를 도출 — ISO 26262
    추적성 체인이 SwUFn id로 매핑되므로 UDS mapping_pairs.source_ids와 join된다.
    """
    import re as _re
    from pathlib import Path as _P

    p = str(path or "").strip()
    if not p:
        return {}
    # TTL 캐시 조회 — parse는 락 밖에서 수행(동시 miss는 redundant parse 허용,
    # 락 점유 최소화). 캐시 객체 변형 방지 위해 사본 반환.
    _key = p.replace("\\", "/").rstrip("/").lower()
    _now = time.time()
    with _VCAST_CLOUDIUM_PARSE_LOCK:
        _cached = _VCAST_CLOUDIUM_PARSE_CACHE.get(_key)
    if _cached and (_now - _cached[0]) < _VCAST_CLOUDIUM_PARSE_TTL:
        return dict(_cached[1])
    try:
        from backend.services import swut_input_adapter as SA
        from backend.services.file_resolver import get_resolver
        from backend.services.jenkins_adapter import (
            _normalize_vcast_result,
            _summarize_vcast_tests,
        )
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).error(
            "vcast cloudium 파서 의존 모듈 import 실패: %s: %s", type(e).__name__, e
        )
        return {}

    resolver = get_resolver()
    # UT/IT 판별 — 등록 경로 문자열 기반 (source 태깅 + env_prefix; VC2025 suffix
    # 모드에선 env_prefix 미사용이라 무해, SWTE 변종 대비 정확 prefix 전달).
    # rank18: 'swit' bare substring은 무관 경로를 IT로 오분류할 수 있어 path-segment
    # (앞뒤가 경계/구분자) 매칭으로 제한.
    low = p.replace("\\", "/").lower()
    is_it = (
        ("통합" in p)
        or ("_it_" in low)
        or (_re.search(r"(?:^|[/_])swit(?:[/_]|$)", low) is not None)
    )
    kind = "IT" if is_it else "UT"
    env_prefix = "SwITC" if is_it else "SWTE"
    warnings: List[str] = []
    try:
        folder = SA._resolve_latest_release_folder(resolver, p, out_warnings=warnings)
        layout = SA._detect_log_layout(resolver, folder, warnings)
        sub_tc = os.path.join(folder, layout.tc_dir)
        if not resolver.exists(sub_tc):
            # silent-drop 방지(P1): tc 폴더 부재/worker 접근 실패를 warnings로 표면화한다.
            # 과거엔 bare {} 반환으로 사용자에게 사유 없이 VectorCAST 열이 비었다.
            return {"test_rows": [], "vcast_kind": kind,
                    "parse_warnings": warnings + [f"{kind}: 시험 TC 폴더 부재/접근 불가 ({sub_tc})"]}
        # ExecutionResult 폴더 — exec_dir 미존재 시 exec_dir_alts("Execution") 시도.
        sub_exec = os.path.join(folder, layout.exec_dir)
        if layout.exec_dir_alts and not SA._exists_quiet(resolver, sub_exec):
            for _alt in layout.exec_dir_alts:
                _cand = os.path.join(folder, _alt)
                if SA._exists_quiet(resolver, _cand):
                    sub_exec = _cand
                    break
        # TestCaseData 리포트 파일명에서 env 목록 빌드.
        tc_files = SA._list_dir_via_resolver(resolver, sub_tc, pattern="*.html")
        env_names = sorted({
            layout.extract_env(_P(f).name, env_prefix)
            for f in tc_files
            if layout.extract_env(_P(f).name, env_prefix)
        })
        test_rows: List[Dict[str, Any]] = []
        idx_cache: Dict[str, Dict[str, str]] = {}
        for env in env_names:
            exec_path = SA._resolve_report_path(
                resolver, sub_exec, env, layout.exec_suffix,
                idx_cache=idx_cache, out_warnings=warnings,
            )
            try:
                data = SA._read_via_resolver(resolver, exec_path)
                results = SA.extract_execution_results_with_actual(data)
            except (PermissionError, OSError) as e:
                warnings.append(f"{env}: ExecutionResult 접근 실패 ({type(e).__name__})")
                continue
            except Exception as e:  # noqa: BLE001 — 개별 env 파싱 실패는 skip + 누적
                warnings.append(f"{env}: ExecutionResult 파싱 실패 ({type(e).__name__})")
                continue
            for tc_name, row in (results or {}).items():
                tc = (tc_name or "").strip()
                # VectorCAST 의사 엔트리(<<COMPOUND>>/<<INIT>> 등)는 실제 함수가 아니므로
                # summary 과계상 방지를 위해 제외.
                if not tc or tc.startswith("<<"):
                    continue
                # subprogram: testcase명에 SwUFn/SwIFn ID가 박혀 있으면 그 ID로 정규화한다
                # (CTC_SwUFn_0431 / SwIT_SwUFn_0101_01 → SwUFn_0431/SwUFn_0101). UDS
                # extract-mapping이 source_ids에 동일 SwUFn ID를 노출하므로 join 성립.
                # SwUFn이 없으면 함수명 base(점 앞)로 fallback — UDS func_name과 매칭.
                m = _re.search(r"Sw[UI]Fn_\d+", tc)
                swufn = m.group(0) if m else ""
                func_base = tc.split(".")[0].strip()
                subprogram = (getattr(row, "subprogram", "") or swufn or func_base).strip()
                result = "PASS" if getattr(row, "passed", False) else "FAIL"
                test_rows.append({
                    "subprogram": subprogram,
                    "testcase": tc,
                    "unit": (getattr(row, "component", "") or env).strip(),
                    "result": result,
                    "function_id": swufn or func_base,
                    "source": kind,
                    "report": env,
                })
        if not test_rows:
            # silent-drop 방지(P1): env 스캔은 됐으나 결과 0건 — 누적 warnings 표면화.
            return {"test_rows": [], "vcast_kind": kind,
                    "parse_warnings": warnings + [f"{kind}: 시험 결과 0건 (폴더 스캔됨, env 파싱 실패 누적 가능)"]}
        summary = _summarize_vcast_tests(test_rows)
        failures: List[Dict[str, Any]] = [
            {
                "testcase": r.get("testcase"),
                "subprogram": r.get("subprogram"),
                "unit": r.get("unit"),
                "result": r.get("result"),
                "report": r.get("report"),
                "source": r.get("source"),
            }
            for r in test_rows
            if _normalize_vcast_result(r.get("result")) == "fail"
        ]
        top_n = int(getattr(config, "VCAST_FAILURES_TOP_N", 50))
        if top_n > 0:
            failures = failures[:top_n]

        # 커버리지(구문/분기/MC-DC) — env별 AggregateCoverage 리포트 단일 HTML에서 추출해
        # 합산한다. ExecutionResult와 동급 비용(env당 HTML 1개 + BS4 표 1개, 무거운
        # TestCaseData 전체 파싱은 여전히 skip). SwUT 빌더의 검증된 extract_aggregate_coverage
        # 재사용. 리포트 미존재/파싱 실패는 best-effort skip(테스트 결과는 그대로 반환).
        cov_acc: Dict[str, List[int]] = {"statement": [0, 0], "branch": [0, 0], "mcdc": [0, 0]}
        # per-function 상세(집계 폐기 방지). vcast_summary.{ut,it}_metrics.entries는 영향도
        # ASIL 차등 MC/DC delta(coverage_gap.load_function_coverage)의 입력이고, complexity_rows는
        # 복잡도 탭의 SCM 소스다. 추출기(extract_aggregate_coverage)가 이미 함수별로 다 주는데
        # 집계로만 접던 결함(audit D2/D3) 수정 — 함수별은 grand가 아니라 항상 funcs에서 모은다.
        fn_entries: List[Dict[str, Any]] = []
        complexity_rows: List[Dict[str, Any]] = []
        _comp_seen: set = set()

        def _cov_cell(cs: Any) -> Dict[str, Any]:
            """CoverageStats → {covered,total,rate}. total=0이면 rate=None(0% 위장 금지 — coverage_gap이 None을 '데이터 없음'으로 처리)."""
            c = int(getattr(cs, "covered", 0) or 0)
            t = int(getattr(cs, "total", 0) or 0)
            return {"covered": c, "total": t, "rate": (round(c / t, 4) if t else None)}

        sub_cov = os.path.join(folder, layout.cov_dir)
        cov_alt = ("_AggregateReport.html",) if layout.name == "vc2025" else ()
        if SA._exists_quiet(resolver, sub_cov):
            for env in env_names:
                try:
                    cov_path = SA._resolve_report_path(
                        resolver, sub_cov, env, layout.cov_suffix,
                        alt_suffixes=cov_alt, idx_cache=idx_cache, out_warnings=warnings,
                    )
                    cdata = SA._read_via_resolver(resolver, cov_path)
                    funcs, grand = SA.extract_aggregate_coverage(cdata)
                except (PermissionError, OSError) as e:
                    warnings.append(f"{env}: AggregateCoverage 접근 실패 ({type(e).__name__})")
                    continue
                except Exception as e:  # noqa: BLE001 — 개별 env 커버리지 실패는 skip + 누적
                    warnings.append(f"{env}: AggregateCoverage 파싱 실패 ({type(e).__name__})")
                    continue
                # env당 grand_total이 곧 그 env 함수들의 합 — 있으면 그것만(이중집계 방지),
                # 없으면 함수별 합산. 두 출처를 섞지 않는다.
                _has_grand = bool(grand and (grand.statement.total or grand.branch.total))
                for fc in ([grand] if _has_grand else funcs):
                    for _k in ("statement", "branch", "mcdc"):
                        cs = getattr(fc, _k, None)
                        if cs is not None and cs.total:
                            cov_acc[_k][0] += cs.covered
                            cov_acc[_k][1] += cs.total
                # 함수별 entries/complexity는 grand 합산이 아니라 항상 per-function(funcs)에서.
                for fc in funcs:
                    sub = (getattr(fc, "name", "") or getattr(fc, "unit_id", "") or "").strip()
                    if not sub:
                        continue
                    unit = (getattr(fc, "component_name", "") or env or "").strip()
                    fn_entries.append({
                        "unit": unit,
                        "subprogram": sub,
                        "ccn": int(getattr(fc, "complexity", 0) or 0),
                        "statements": _cov_cell(getattr(fc, "statement", None)),
                        "branches": _cov_cell(getattr(fc, "branch", None)),
                        "pairs": _cov_cell(getattr(fc, "mcdc", None)),
                    })
                    cplx = int(getattr(fc, "complexity", 0) or 0)
                    if cplx:
                        _ckey = (sub, unit)
                        if _ckey not in _comp_seen:
                            _comp_seen.add(_ckey)
                            complexity_rows.append({
                                "function": sub, "file": unit, "unit": unit, "complexity": cplx,
                            })
        coverage: Dict[str, Any] = {}
        for _k in ("statement", "branch", "mcdc"):
            _c, _t = cov_acc[_k]
            coverage[_k] = {"covered": _c, "total": _t, "rate": (round(_c / _t, 4) if _t else None)}

        # vcast_summary는 빌드 RAG와 동일 스키마({ut,it}_metrics.entries) — coverage_gap이
        # ut_metrics/it_metrics를 모두 읽으므로 폴더 종류(UT/IT)에 맞는 키 하나만 채운다.
        vcast_summary: Dict[str, Any] = {}
        if fn_entries:
            vcast_summary["it_metrics" if is_it else "ut_metrics"] = {"entries": fn_entries}

        # IT 폴더 — Metric report HTML에서 함수콜(Function Called)/함수 진입(Functions) 커버리지를
        # 추출해 it_metrics.grand_totals를 채운다. 함수콜은 AggregateCoverage(구문/분기/MC-DC)에
        # 없고 Metric report에만 있어, 그동안 Jenkins 빌드 산출물에서만 나오던 함수콜을 SCM
        # cloudium 경로에서도 동일하게 제공한다 (SwITCV 빌더 _discover_metric_report_bytes 대칭).
        # 후보 폴더: 등록 경로(p) + 해석된 릴리스 폴더(folder) — 양식별 Metric report 위치 차이 대비.
        # 파일명이 일정치 않아(APP `*_Metric_report_*`, BOOT `*_IT_*`) content-detect(parse_hmr_html.ok).
        if is_it and fn_entries:
            try:
                _html_list: List[bytes] = []
                _seen_html: set = set()
                _cand_dirs: List[str] = []
                for _d in (folder, p):
                    _dk = str(_d or "").replace("\\", "/").rstrip("/").lower()
                    if _d and _dk and _dk not in {str(x).replace(chr(92), '/').rstrip('/').lower() for x in _cand_dirs}:
                        _cand_dirs.append(_d)
                for _d in _cand_dirs:
                    if len(_html_list) >= _MAX_METRIC_HTML_SCAN:
                        break
                    try:
                        _htmls = SA._list_dir_via_resolver(resolver, _d, pattern="*.html")
                    except Exception:  # noqa: BLE001
                        continue
                    for _e in sorted(_htmls, key=lambda x: str(x).lower()):
                        if len(_html_list) >= _MAX_METRIC_HTML_SCAN:
                            break
                        _base = os.path.basename(str(_e).rstrip("\\/").replace("\\", "/"))
                        _full = _e if (str(_e).startswith("U:") or str(_e).startswith("/")) \
                            else os.path.join(_d, _base)
                        _fk = _full.replace("\\", "/").lower()
                        if _fk in _seen_html:
                            continue
                        _seen_html.add(_fk)
                        try:
                            _html_list.append(SA._read_via_resolver(resolver, _full))
                        except Exception:  # noqa: BLE001
                            continue
                _it_grand, _fc_by_name = _aggregate_it_function_calls(_html_list)
                if _it_grand:
                    vcast_summary.setdefault("it_metrics", {})["grand_totals"] = _it_grand
                    # per-function 함수콜을 entries에 병합(모듈 드릴다운 표시용) — 함수명/점앞 base 매칭.
                    for _ent in fn_entries:
                        _sp = (_ent.get("subprogram") or "").strip()
                        _hit = _fc_by_name.get(_sp) or _fc_by_name.get(_sp.split(".")[0])
                        if _hit and _hit.get("total"):
                            _ent["function_calls"] = {
                                "covered": _hit["covered"], "total": _hit["total"],
                                "rate": round(_hit["covered"] / _hit["total"], 4),
                            }
                    if _it_grand.get("function_calls"):
                        warnings.append(
                            "[metric-report] IT 함수콜 커버리지 보강 "
                            f"({_it_grand['function_calls']['covered']}/{_it_grand['function_calls']['total']})"
                        )
            except Exception:  # noqa: BLE001 — Metric report 부재/파싱 실패는 best-effort skip
                logging.getLogger(__name__).debug(
                    "IT Metric report 함수콜 집계 skip path=%s", p, exc_info=True
                )

        payload_out: Dict[str, Any] = {
            "build_root": folder,
            "source_folder": p,
            "vcast_kind": kind,
            "test_rows": test_rows,
            "test_rows_count": len(test_rows),
            "summary": summary,
            "failures": failures,
            "coverage": coverage,
            "vcast_summary": vcast_summary,      # P0: ASIL 차등 MC/DC delta(coverage_gap) 입력
            "complexity_rows": complexity_rows,  # P1: 복잡도 탭 SCM 소스
            "ut_reports": [] if is_it else [folder],
            "it_reports": [folder] if is_it else [],
            "parse_warnings": warnings,
            "parsed_from": "cloudium_raw_reports",
        }
        # 비어있지 않은 결과만 캐시 (worker 일시 장애 후 재시도 허용).
        with _VCAST_CLOUDIUM_PARSE_LOCK:
            # rank19: 무한 증가 방지 — 엔트리당 수 MB라 상한 도달 시 비운다.
            # 등록 경로는 통상 한 자릿수라 32 상한은 충분한 여유.
            if len(_VCAST_CLOUDIUM_PARSE_CACHE) >= 32:
                _VCAST_CLOUDIUM_PARSE_CACHE.clear()
            _VCAST_CLOUDIUM_PARSE_CACHE[_key] = (_now, payload_out)
        return dict(payload_out)
    except (PermissionError, OSError) as e:
        # rank11: 권한/IPC 장애를 '데이터 없음'으로 silent 은폐하지 말 것 — 누적 warning과
        # 함께 로그. 반환은 {}(graceful)이되 사유는 운영자가 추적 가능하게 남긴다.
        logging.getLogger(__name__).warning(
            "vcast cloudium 파싱 접근 실패 path=%s err=%s: %s warnings=%s",
            p, type(e).__name__, e, warnings[:5],
        )
        # silent-drop 방지(P1): worker 미응답/timeout(PermissionError)·IO 오류를 warnings로
        # 표면화. cloudium U: 유휴 후 첫 접근 timeout이 가장 흔한 실사용 통증(KJPDS02 PV).
        return {"test_rows": [], "vcast_kind": kind,
                "parse_warnings": warnings + [f"{kind}: cloudium 접근 실패 {type(e).__name__} (worker 미응답/timeout 가능): {str(e)[:80]}"]}
    except Exception:  # noqa: BLE001
        logging.getLogger(__name__).debug(
            "vcast cloudium 파싱 예외 path=%s", p, exc_info=True
        )
        # silent-drop 방지(P1): 예상 못한 파싱 예외도 warnings로 표면화(상세는 debug 로그).
        return {"test_rows": [], "vcast_kind": kind,
                "parse_warnings": warnings + [f"{kind}: 파싱 예외 (backend 로그 참조)"]}


def _split_vcast_summary_by_source(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """test_rows를 source(UT/IT)별로 나눠 합부 summary/카운트를 분리한다.

    각 row의 ``source``는 폴더 종류(``kind = "IT" if is_it else "UT"``, jenkins.py:1029)이며,
    coverage 분리 로직(``cov_it if _it else cov_ut``)과 동일하게 **IT가 아니면 UT로 귀속**해
    ut+it 합이 전체와 일치하도록 한다(미상 source는 UT로 흡수). 결합 summary와 별개로
    UT 패널은 UT만, IT 패널은 IT만 표시하기 위한 필드.
    """
    from backend.services.jenkins_adapter import _summarize_vcast_tests

    it_rows = [r for r in rows if str(r.get("source") or "").upper() == "IT"]
    ut_rows = [r for r in rows if str(r.get("source") or "").upper() != "IT"]
    return {
        "summary_ut": _summarize_vcast_tests(ut_rows) if ut_rows else None,
        "summary_it": _summarize_vcast_tests(it_rows) if it_rows else None,
        "test_rows_count_ut": len(ut_rows),
        "test_rows_count_it": len(it_rows),
    }


def _merge_vectorcast_payloads(payloads: List[Dict[str, Any]]) -> Dict[str, Any]:
    """여러 Cloudium VectorCAST RAG payload를 하나로 병합한다.

    부트로더/FBL/APP 등 결과가 별도 파일로 나올 때 각 payload의 test_rows를 합치고
    summary/failures를 재계산한다. 같은 경로를 중복 등록하거나 부모/자식 경로가 같은
    파일을 가리키는 경우 동일 row가 중복될 수 있어 (subprogram,testcase,unit,result,
    report) 키로 dedup하여 이중 집계를 막는다.
    """
    from backend.services.jenkins_adapter import (
        _normalize_vcast_result,
        _summarize_vcast_tests,
    )
    merged_rows: List[Dict[str, Any]] = []
    seen: set = set()
    ut_reports: List[Any] = []
    it_reports: List[Any] = []
    # 커버리지 합산 — 전체 + UT/IT 분리. 각 payload의 coverage는 폴더 단위 집계라 폴더별
    # 합산(row dedup과 무관). [covered, total].
    cov_all: Dict[str, List[int]] = {"statement": [0, 0], "branch": [0, 0], "mcdc": [0, 0]}
    cov_ut: Dict[str, List[int]] = {"statement": [0, 0], "branch": [0, 0], "mcdc": [0, 0]}
    cov_it: Dict[str, List[int]] = {"statement": [0, 0], "branch": [0, 0], "mcdc": [0, 0]}
    # per-function entries(vcast_summary)·complexity_rows 병합 — 폴더별 함수 집합은 보통
    # 서로소(부트로더/APP 등). entries는 coverage_gap이 메트릭별 max로 흡수하므로 단순 concat,
    # complexity는 (function,unit)로 dedup.
    merged_metrics: Dict[str, List[Dict[str, Any]]] = {}
    # {mk: {metric_key: {covered,total}}} — 폴더별 grand_totals 합산 누산기(버그2 fix)
    merged_grand: Dict[str, Dict[str, Dict[str, int]]] = {}
    merged_complexity: List[Dict[str, Any]] = []
    comp_seen: set = set()
    merged_warnings: List[str] = []
    for pl in payloads:
        if not isinstance(pl, dict):
            continue
        # silent-drop 방지(P1): 폴더별 파싱 실패/빈결과 사유(parse_warnings)를 병합 수집.
        _pw = pl.get("parse_warnings")
        if isinstance(_pw, list):
            merged_warnings.extend(str(w) for w in _pw)
        for r in (pl.get("test_rows") or []):
            if not isinstance(r, dict):
                continue
            key = (
                str(r.get("subprogram") or ""),
                str(r.get("testcase") or ""),
                str(r.get("unit") or ""),
                str(r.get("result") or ""),
                str(r.get("report") or ""),
                # rank9: source(UT/IT) 포함 — 같은 TC가 단위(UT)와 통합(IT) 양쪽에서
                # PASS로 나오는 정당한 케이스(ISO 26262 T4/T5 별도 캠페인 증거)에서
                # IT 증거가 조용히 드롭되는 것을 막는다.
                str(r.get("source") or ""),
            )
            if key in seen:
                continue
            seen.add(key)
            merged_rows.append(r)
        ut = pl.get("ut_reports")
        if isinstance(ut, list):
            ut_reports.extend(ut)
        it = pl.get("it_reports")
        if isinstance(it, list):
            it_reports.extend(it)
        cov = pl.get("coverage")
        if isinstance(cov, dict):
            _it = str(pl.get("vcast_kind") or "").upper() == "IT"
            for _m in ("statement", "branch", "mcdc"):
                _cell = cov.get(_m) or {}
                _c = int(_cell.get("covered") or 0)
                _t = int(_cell.get("total") or 0)
                cov_all[_m][0] += _c
                cov_all[_m][1] += _t
                _tgt = cov_it if _it else cov_ut
                _tgt[_m][0] += _c
                _tgt[_m][1] += _t
        vs = pl.get("vcast_summary")
        if isinstance(vs, dict):
            for _mk in ("ut_metrics", "it_metrics"):
                _blk = vs.get(_mk)
                if not isinstance(_blk, dict):
                    continue
                if isinstance(_blk.get("entries"), list):
                    merged_metrics.setdefault(_mk, []).extend(_blk["entries"])
                # grand_totals(함수콜/함수진입)는 폴더별 함수 집합이 disjoint(APP/BOOT 등)라
                # covered/total을 합산해 보존 — 드롭하면 2폴더+ 로드 시 카드가 사라진다.
                _gt = _blk.get("grand_totals")
                if isinstance(_gt, dict):
                    _acc = merged_grand.setdefault(_mk, {})
                    for _metric_key in ("function_calls", "functions"):
                        _mv = _gt.get(_metric_key)
                        if isinstance(_mv, dict) and int(_mv.get("total") or 0) > 0:
                            _sub = _acc.setdefault(_metric_key, {"covered": 0, "total": 0})
                            _sub["covered"] += int(_mv.get("covered") or 0)
                            _sub["total"] += int(_mv.get("total") or 0)
        for _cr in (pl.get("complexity_rows") or []):
            if not isinstance(_cr, dict):
                continue
            _ck = (str(_cr.get("function") or ""), str(_cr.get("unit") or _cr.get("file") or ""))
            if _ck in comp_seen:
                continue
            comp_seen.add(_ck)
            merged_complexity.append(_cr)

    summary = _summarize_vcast_tests(merged_rows)
    top_n = int(getattr(config, "VCAST_FAILURES_TOP_N", 50))
    failures = [
        {
            "testcase": r.get("testcase"),
            "requirement_id": r.get("requirement_id"),
            "unit": r.get("unit"),
            "subprogram": r.get("subprogram"),
            "result": r.get("result"),
            "report": r.get("report"),
            "source": r.get("source"),
        }
        for r in merged_rows
        if _normalize_vcast_result(r.get("result")) == "fail"
    ]
    if top_n > 0:
        failures = failures[:top_n]

    def _cov_dict(acc: Dict[str, List[int]]) -> Optional[Dict[str, Any]]:
        out: Dict[str, Any] = {}
        any_total = False
        for _m in ("statement", "branch", "mcdc"):
            _c, _t = acc[_m]
            out[_m] = {"covered": _c, "total": _t, "rate": (round(_c / _t, 4) if _t else None)}
            any_total = any_total or _t > 0
        return out if any_total else None

    def _vcast_summary_out() -> Dict[str, Any]:
        # entries + 합산된 grand_totals(함수콜/함수진입)를 모두 실어 카드 유실 방지(버그2).
        out: Dict[str, Any] = {}
        for _mk in set(merged_metrics) | set(merged_grand):
            _blk: Dict[str, Any] = {}
            if _mk in merged_metrics:
                _blk["entries"] = merged_metrics[_mk]
            _gt_acc = merged_grand.get(_mk) or {}
            _gt: Dict[str, Any] = {}
            for _metric_key, _sub in _gt_acc.items():
                _t = int(_sub.get("total") or 0)
                if _t > 0:
                    _cvd = int(_sub.get("covered") or 0)
                    _gt[_metric_key] = {"covered": _cvd, "total": _t, "rate": round(_cvd / _t, 4)}
            if _gt:
                _blk["grand_totals"] = _gt
            out[_mk] = _blk
        return out

    return {
        "test_rows": merged_rows,
        "test_rows_count": len(merged_rows),
        "summary": summary,
        **_split_vcast_summary_by_source(merged_rows),
        "failures": failures,
        "ut_reports": ut_reports,
        "it_reports": it_reports,
        "coverage": _cov_dict(cov_all),
        "coverage_ut": _cov_dict(cov_ut),
        "coverage_it": _cov_dict(cov_it),
        "vcast_summary": _vcast_summary_out(),
        "complexity_rows": merged_complexity,
        "merged_sources": len([p for p in payloads if isinstance(p, dict) and p]),
        # silent-drop 방지(P1): 폴더별 파싱 실패 사유를 병합해 상위로 전달(프론트 표면화).
        "parse_warnings": merged_warnings,
    }


def _load_vectorcast_rag_from_cloudium_multi(paths: List[str]) -> Dict[str, Any]:
    """복수 Cloudium 경로에서 VectorCAST RAG를 읽어 병합한다.

    단일 경로일 때는 원본 payload를 그대로 반환(모든 필드 보존), 2개 이상이면
    _merge_vectorcast_payloads로 test_rows/summary/failures를 합친다.
    """
    payloads: List[Dict[str, Any]] = []
    if len(paths) <= 1:
        for p in paths:
            pl = _load_vectorcast_rag_from_cloudium(p)
            if pl:
                payloads.append(pl)
    else:
        # rank4 fix: 폴더별 파싱은 독립적이고 cloudium worker는 동시 read를 정확히
        # 처리한다(검증: 3-thread 동시 read byte-identical). 직렬 시 5폴더 ~500s로
        # 브라우저/프록시 타임아웃 → 폴더 단위 병렬(max 4)로 ~max(단일폴더)≈100-150s로
        # 단축. 폴더별 TTL 캐시로 부분 진행도 보존된다.
        import concurrent.futures as _cf
        with _cf.ThreadPoolExecutor(max_workers=min(4, len(paths))) as _ex:
            # map은 입력 순서 보존 — merge dedup은 순서 무관.
            for pl in _ex.map(_load_vectorcast_rag_from_cloudium, paths):
                if pl:
                    payloads.append(pl)
    if not payloads:
        return {}
    if len(payloads) == 1:
        # 단일 폴더(UT 또는 IT 하나)도 UT/IT 분리 필드를 채워 프론트 패널이 결합값을
        # 오표시하지 않게 한다. 원본 캐시 dict를 변형하지 않도록 얕은 복사 후 갱신.
        pl = dict(payloads[0])
        pl.update(_split_vcast_summary_by_source(pl.get("test_rows") or []))
        return pl
    return _merge_vectorcast_payloads(payloads)


def _collect_vcast_paths(req: JenkinsReportRequest) -> List[str]:
    """요청에서 VectorCAST cloudium 경로 목록을 정규화/중복제거하여 모은다.

    vcast_log_paths(복수, 우선) + vcast_log_path(단일, 하위호환)를 합치고 순서를
    보존하며 대소문자/슬래시 방향 무시 dedup.
    """
    raw: List[str] = []
    for p in (req.vcast_log_paths or []):
        s = str(p or "").strip()
        if s:
            raw.append(s)
    legacy = (req.vcast_log_path or "").strip()
    if legacy:
        raw.append(legacy)
    seen: set = set()
    out: List[str] = []
    for p in raw:
        k = p.replace("\\", "/").rstrip("/").lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(p)
    return out


def _compute_vectorcast_rag(req: JenkinsReportRequest) -> Dict[str, Any]:
    """VectorCAST RAG 계산(빌드 산출물 → cloudium 등록 경로 폴백). 동기/비동기 라우트 공용.

    cloudium 원본 폴더 파싱은 무거우므로(수 분) 비동기 잡 경로(vectorcast-rag-async)에서도
    동일 로직을 재사용한다. 결과는 폴더 단위 TTL 캐시(_VCAST_CLOUDIUM_PARSE_CACHE)로 2회차+ 즉시.
    """
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    payload = _load_vectorcast_rag(build_root) if build_root else {}
    source = "jenkins"

    def _has_vcast_rows(p: Any) -> bool:
        return isinstance(p, dict) and bool(p.get("test_rows"))

    # Cloudium 폴백 — Jenkins 빌드 RAG가 없거나(빈 dict), "스캔은 됐지만 test_rows가 비어
    # 있는" 경우(VectorCAST 결과가 Jenkins 빌드엔 없고 cloudium U:/ 경로에만 있는 흔한
    # 케이스)에도 사용자 지정 경로(부트로더/APP 등 별도 결과)로 폴백한다. 과거엔 truthy-but-
    # rowless payload(예: build_20 vectorcast_rag.json은 키는 많지만 test_rows=[])를
    # 'present'로 오인해 `if not payload`가 폴백을 건너뛰어 → 등록 경로 결과가 끝까지
    # 안 쓰이고 VectorCAST/P&F 컬럼이 빈 채로 나왔다. test_rows 유무로 판정해 폴백한다.
    cloud_warnings: List[str] = []
    if not _has_vcast_rows(payload):
        cloud_paths = _collect_vcast_paths(req)
        if cloud_paths:
            cloud_payload = _load_vectorcast_rag_from_cloudium_multi(cloud_paths)
            if _has_vcast_rows(cloud_payload):
                payload = cloud_payload
                source = "cloudium"
            elif isinstance(cloud_payload, dict):
                # silent-drop 방지(P1): 폴백이 test_rows를 못 얻어도 실패 사유(worker
                # timeout/폴더 부재 등)는 살려 프론트가 표시하게 한다.
                _pw = cloud_payload.get("parse_warnings")
                if isinstance(_pw, list):
                    cloud_warnings.extend(str(w) for w in _pw)
    if not _has_vcast_rows(payload):
        _pw = payload.get("parse_warnings") if isinstance(payload, dict) else None
        all_pw = ([str(w) for w in _pw] if isinstance(_pw, list) else []) + cloud_warnings
        return {"ok": False, "error": "missing", "parse_warnings": all_pw}

    comparison: Dict[str, Any] = {}
    # 이전 빌드 delta 비교는 Jenkins 캐시 기반 — cloudium 폴백 소스에는 비적용.
    if build_root and source == "jenkins":
        try:
            builds = list_cached_builds(job_url=req.job_url, cache_root=_normalize_jenkins_cache_root(req.cache_root))
            summaries: List[Dict[str, Any]] = []
            for row in builds:
                cand_root = Path(row.get("build_root", ""))
                if not cand_root.exists():
                    continue
                rag = _load_vectorcast_rag(cand_root)
                if not isinstance(rag, dict) or not rag.get("summary"):
                    continue
                summaries.append({"summary": rag.get("summary") or {}, "build": row})
                if len(summaries) >= 2:
                    break
            if len(summaries) >= 2:
                cur = summaries[0]["summary"]
                prev = summaries[1]["summary"]
                comparison = {
                    "current": cur,
                    "previous": prev,
                    "delta": {
                        "total": (cur.get("total") or 0) - (prev.get("total") or 0),
                        "passed": (cur.get("passed") or 0) - (prev.get("passed") or 0),
                        "failed": (cur.get("failed") or 0) - (prev.get("failed") or 0),
                        "skipped": (cur.get("skipped") or 0) - (prev.get("skipped") or 0),
                        "pass_rate": (cur.get("pass_rate") or 0) - (prev.get("pass_rate") or 0),
                    },
                }
        except Exception:
            comparison = {}

    return {"ok": True, "data": payload, "comparison": comparison, "source": source}


@router.post("/api/jenkins/report/vectorcast-rag")
def jenkins_vectorcast_rag(req: JenkinsReportRequest) -> Dict[str, Any]:
    return _compute_vectorcast_rag(req)


@router.post("/api/jenkins/report/vectorcast-rag-async")
def jenkins_vectorcast_rag_async(req: JenkinsReportRequest) -> Dict[str, Any]:
    """무거운 cloudium VectorCAST 폴더 파싱(수 분)을 백그라운드 잡으로 실행한다.

    동기 호출은 원격 IPC 직렬 파싱으로 4~5분 블로킹 → 브라우저/프록시 타임아웃·탭 전환 abort로
    '에러처럼' 보였다. 잡으로 돌리고 기존 폴링(/api/scm/impact-job/{id}, /result)을 재사용한다.
    user context는 start_job의 wrap_with_user가 상속(cloudium worker 접근 필수).
    """
    from workflow.impact_jobs import start_job

    def _runner(_job_id: str) -> Dict[str, Any]:
        return _compute_vectorcast_rag(req)

    return start_job(
        scm_id=_job_slug(req.job_url) or "vcast",
        trigger_type="vectorcast",
        runner=_runner,
        # 내부 cloudium 경로를 잡 메타(폴링 응답에 노출)에 싣지 않는다 — 개수만 진단용 기록(W3).
        metadata={"job_url": req.job_url, "vcast_path_count": len(_collect_vcast_paths(req))},
    )


def _sa_module_of(path: str) -> str:
    """.../<TOOL>/<MODULE_날짜_버전>/file → 'MODULE_날짜_버전' (파일의 부모 폴더명)."""
    parts = re.split(r"[\\/]", (path or "").rstrip("\\/"))
    return parts[-2] if len(parts) >= 2 else ""


def _sa_module_label(path: str) -> str:
    """모듈 폴더명 → 표시 라벨. 'APP_260527_v0.05.37' → 'APP'. prefix 없으면 폴더명."""
    folder = _sa_module_of(path)
    prefix = folder.split("_")[0] if folder else ""
    return prefix or folder or "?"


def _sa_pmd_to_cpd(pmd: Any) -> Dict[str, Any]:
    """swsa PmdResult → 프론트 CPD 카드 shape(duplication_blocks/…/top_blocks)."""
    files: set = set()
    for b in pmd.blocks:
        files.update(b.basenames)
    top = pmd.blocks_sorted()[:20]
    return {
        "ok": pmd.total_blocks > 0,
        "duplication_blocks": pmd.total_blocks,
        "total_dup_lines": pmd.total_duplicated_lines,
        "total_tokens": sum(b.tokens for b in pmd.blocks),
        "files_involved": len(files),
        "top_blocks": [
            {"lines": b.lines, "tokens": b.tokens,
             "fragments": len(b.files), "files": b.basenames}
            for b in top
        ],
    }


def _sa_st201_to_qac(st: Any) -> Dict[str, Any]:
    """swsa St201Result(HMR) → 프론트 QAC HIS 카드 shape(함수 v(G) 분포)."""
    from backend.services.qac_parser import MatrixItem

    vgs = st.values_for(MatrixItem.V_G)
    summary: Dict[str, Any] = {"function_count": st.total_functions}
    if vgs:
        vs = sorted(vgs)
        summary["vg_max"] = max(vgs)
        summary["vg_mean"] = round(sum(vgs) / len(vgs), 2)
        summary["vg_p95"] = vs[min(len(vs) - 1, int(len(vs) * 0.95))]
        summary["vg_over_10"] = sum(1 for v in vgs if v > 10)
    mr = st.metric("ST201")
    top = [{"function": nm, "vg": v} for nm, v in (mr.worst_functions if mr else [])]
    return {"ok": bool(vgs), "summary": summary, "top_functions": top}


def _load_static_analysis(paths: List[str]) -> Dict[str, Any]:
    """SCM 등록 정적분석 폴더에서 4종 도구 산출물을 **모듈(APP/BOOT)별로** 찾아 파싱한다.

    회사 정적분석 4종 = CodeSonar(PDF)·CodeEye(OSS 종합 PDF)·QAC HIS(HMR HTML)·CPD(PMD TXT).
    각 도구는 ``<TOOL>/<MODULE_날짜_버전>/`` 하위에 모듈별로 존재 — 모듈 prefix(APP/BOOT)별
    **최신 분석 1개씩** 파싱해 ``modules`` 리스트로 반환한다. (과거 단일-파일 방식은
    사전순 마지막 1개만 남겨 APP 모듈을 누락시켰음.) 포맷 혼재 흡수: QAC는 html(HMR)+pdf(HIS),
    CPD는 txt(PMD)+xml(CPD)을 **합집합**으로 모아 파서를 확장자로 분기(존재-기반 폴백 아님).
    cloudium 모드는 worker IPC로 read(backend python은 권한 없음). 일부 도구/모듈만 있어도 graceful.

    응답 shape::
        {"ok", "codesonar"|"codeeye"|"qac"|"cpd": {"ok", "modules":[{label, module_folder, source, …}]},
         "warnings"?, "detail"?}
    """
    from backend.services.codesonar_pdf_parser import parse_codesonar_pdf
    from backend.services.file_resolver import get_resolver
    from backend.services.static_analysis_parsers import (
        parse_codeeye_pdf,
        parse_cpd_xml,
        parse_qac_his_pdf,
    )
    from backend.services.swsa_input_adapter import _select_latest_per_module
    from backend.services.swsa_pmd_parser import parse_pmd_cpd
    from backend.services.swsa_st201_binner import parse_st201_from_hmr

    resolver = get_resolver()
    all_files: List[str] = []
    for raw in paths or []:
        p = (raw or "").strip()
        if not p:
            continue
        if p.lower().endswith((".pdf", ".xml", ".txt", ".html", ".htm")):
            all_files.append(p)
            continue
        try:
            all_files.extend(str(f) for f in resolver.list_dir(p, "*", recursive=True))
        except Exception as e:  # cloudium 권한/미연결/경로부재 — graceful
            _logger.warning("[static-analysis] list_dir 실패 %s: %s", p[:80], e)

    def _read(path: Optional[str]) -> Optional[bytes]:
        if not path:
            return None
        try:
            return resolver.read_bytes(path)
        except Exception as e:
            _logger.warning("[static-analysis] read 실패 %s: %s", path[:80], e)
            return None

    warnings: List[str] = []

    def _modules(files: List[str], kind: str, parse_bytes: Any) -> Dict[str, Any]:
        """모듈 prefix별 최신 파일 1개씩 파싱 → {ok, modules:[…]}. 파서 예외는 모듈 단위 격리.

        parse_bytes(data, path): path 확장자로 포맷 분기(QAC html/pdf·CPD txt/xml 혼재 대응).
        """
        sel = _select_latest_per_module(files, kind, warnings) if files else []
        modules: List[Dict[str, Any]] = []
        for path in sorted(sel):
            data = _read(path)
            if not data:
                continue
            try:
                res = parse_bytes(data, path)
            except Exception as e:  # 한 모듈 파싱 실패가 전체를 무너뜨리지 않도록 격리
                _logger.warning("[static-analysis] %s 파싱 실패 %s: %s", kind, path[:80], e)
                continue
            if not (isinstance(res, dict) and res.get("ok")):
                continue
            res["label"] = _sa_module_label(path)
            res["module_folder"] = _sa_module_of(path)
            res["source"] = path
            modules.append(res)
        return {"ok": any(m.get("ok") for m in modules), "modules": modules}

    def _parse_qac(data: bytes, path: str) -> Dict[str, Any]:
        # 실 산출물이 html(HMR) 또는 pdf(HIS Metric) — 확장자로 분기(모듈별 혼재 허용).
        if path.lower().endswith((".html", ".htm")):
            return _sa_st201_to_qac(parse_st201_from_hmr(data))
        return parse_qac_his_pdf(data)

    def _parse_cpd(data: bytes, path: str) -> Dict[str, Any]:
        # PMD txt(회사 표준) 또는 PMD CPD xml — 확장자로 분기.
        if path.lower().endswith(".txt"):
            return _sa_pmd_to_cpd(parse_pmd_cpd(data))
        return parse_cpd_xml(data)

    lo = [(f, f.lower()) for f in all_files]

    cs_files = [f for f, fl in lo if fl.endswith(".pdf") and "codesonar" in fl]
    ce_files = [f for f, fl in lo if fl.endswith(".pdf") and "codeeye" in fl and "종합" in f]
    # QAC/CPD는 두 포맷을 **합집합**으로 모아 모듈별 최신 1개를 뽑고 파서를 확장자로 분기한다.
    # (과거 'html 있으면 pdf 무시' 식 존재-기반 폴백은 손상된 primary가 유효한 fallback을
    #  막거나 APP=html·BOOT=pdf 혼재 시 한쪽을 통째 누락시켰음.)
    qac_files = [
        f for f, fl in lo
        if (fl.endswith((".html", ".htm")) and "hmr" in fl)
        or (fl.endswith(".pdf") and "his" in fl and "metric" in fl
            and "codesonar" not in fl and "codeeye" not in fl)
    ]
    cpd_files = [
        f for f, fl in lo
        if (fl.endswith(".txt") and "pmd" in fl)
        or (fl.endswith(".xml") and ("cpd" in fl or "result_xml" in fl))
    ]

    out: Dict[str, Any] = {
        "codesonar": _modules(cs_files, "CodeSonar", lambda data, _p: parse_codesonar_pdf(data)),
        "codeeye": _modules(ce_files, "CodeEye", lambda data, _p: parse_codeeye_pdf(data)),
        "qac": _modules(qac_files, "QAC", _parse_qac),
        "cpd": _modules(cpd_files, "CPD", _parse_cpd),
    }
    out["ok"] = any(out[k].get("ok") for k in ("codesonar", "codeeye", "qac", "cpd"))
    if warnings:
        out["warnings"] = warnings
    if not out["ok"]:
        out["detail"] = "정적분석 산출물(CodeSonar/CodeEye/QAC HMR/PMD)을 찾지 못했습니다 (경로/권한 확인)"
    return out


@router.post("/api/jenkins/report/static-analysis")
def jenkins_static_analysis(req: CodeSonarRequest) -> Dict[str, Any]:
    """SCM 정적분석 폴더에서 CodeSonar/CPD/QAC HIS/CodeEye 4종 요약 지표를 추출한다.

    SCM 연결 문서 경로(linked_docs.codesonar) 또는 사용자 지정 폴더/파일 경로 목록을 받아
    각 도구별 최신 리포트를 파싱. 산출물은 정적분석 섹션의 도구별 카드/표에 표시된다.
    """
    return _load_static_analysis(req.paths)


@router.post("/api/jenkins/source-root")
def jenkins_source_root(req: JenkinsReportRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    return find_jenkins_source_root(build_root)


@router.post("/api/jenkins/source-root/download")
def jenkins_source_root_download(req: JenkinsSourceDownloadRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    source_dir = (Path(build_root) / "source").resolve()
    client = JenkinsClient(
        job_url=req.job_url,
        username=req.username,
        api_token=req.api_token,
        timeout_sec=30,
        verify_ssl=bool(req.verify_tls),
    )
    artifact_info: Dict[str, Any] | None = None

    def _download_artifact_url(artifact_url: str) -> Optional[Path]:
        if not artifact_url:
            return None
        parsed = urlparse(artifact_url)
        if not parsed.scheme or "/artifact/" not in parsed.path:
            return None
        rel = unquote(parsed.path.split("/artifact/", 1)[1]).lstrip("/")
        if not rel:
            return None
        dst = _safe_artifact_path(Path(build_root), rel)
        if not dst:
            dst = (Path(build_root) / "artifact_source" / Path(rel).name).resolve()
        dst.parent.mkdir(parents=True, exist_ok=True)
        req_obj = client._auth_req(artifact_url, accept="application/octet-stream")
        raw = client._open_bytes(req_obj)
        dst.write_bytes(raw)
        return dst
    checkout_result = ensure_source_checkout(
        build_root=build_root,
        client=client,
        build_selector=req.build_selector,
        progress_cb=None,
        scm_username=req.scm_username,
        scm_id=req.scm_id,
        force=req.force,
    )
    def _dir_has_entries(path: Path) -> bool:
        try:
            return path.exists() and path.is_dir() and any(path.iterdir())
        except Exception:
            return False
    if _dir_has_entries(source_dir):
        return {
            "ok": True,
            "path": str(source_dir),
            "build_root": str(build_root),
            "scm": checkout_result.get("scm"),
            "repo_url": checkout_result.get("repo_url"),
            "branch": checkout_result.get("branch"),
            "revision": checkout_result.get("revision"),
        }
    if req.scm_url:
        scm_type = (req.scm_type or "svn").lower()
        if scm_type == "svn":
            from backend.services.scm_registry import resolve_scm_credentials
            alt_user, alt_pass, _ = resolve_scm_credentials(
                repo_url=req.scm_url,
                scm_id=getattr(req, "scm_id", "") or "",
                override_username=req.scm_username or "",
            )
            alt = run_svn(
                project_root=str(build_root),
                workdir_rel="source",
                action="checkout",
                repo_url=req.scm_url,
                revision=req.scm_revision or "",
                username=alt_user,
                password=alt_pass,
            )
            if alt.get("rc") == 0 and _dir_has_entries(source_dir):
                return {
                    "ok": True,
                    "path": str(source_dir),
                    "build_root": str(build_root),
                    "source": "scm_manual",
                    "scm": "svn",
                    "repo_url": req.scm_url,
                    "revision": req.scm_revision,
                }
    if req.source_root:
        try:
            artifact_path = _download_artifact_url(req.source_root)
            if artifact_path:
                artifact_info = {"path": str(artifact_path), "source_root": req.source_root}
                if artifact_path.suffix.lower() == ".zip":
                    _safe_extract_zip(artifact_path, source_dir)
        except Exception as exc:
            artifact_info = {"error": str(exc), "source_root": req.source_root}
    if _dir_has_entries(source_dir):
        return {
            "ok": True,
            "path": str(source_dir),
            "build_root": str(build_root),
            "source": "artifact_zip",
            "scm": checkout_result.get("scm"),
            "repo_url": checkout_result.get("repo_url"),
            "branch": checkout_result.get("branch"),
            "revision": checkout_result.get("revision"),
            "artifact": artifact_info,
        }
    # fallback: source may exist in artifacts under a different path
    source_root_info = find_jenkins_source_root(Path(build_root))
    for cand in source_root_info.get("candidates", []):
        cand_path = Path(cand.get("path", ""))
        if _dir_has_entries(cand_path):
            return {
                "ok": True,
                "path": str(cand_path),
                "build_root": str(build_root),
                "source": "artifact",
                "scm": checkout_result.get("scm"),
                "repo_url": checkout_result.get("repo_url"),
                "branch": checkout_result.get("branch"),
                "revision": checkout_result.get("revision"),
                "candidates": source_root_info.get("candidates", []),
                "artifact": artifact_info,
            }
    return {
        "ok": False,
        "error": "source_dir_missing",
        "build_root": str(build_root),
        "source_dir": str(source_dir),
        "scm": checkout_result.get("scm"),
        "repo_url": checkout_result.get("repo_url"),
        "branch": checkout_result.get("branch"),
        "revision": checkout_result.get("revision"),
        "checkout_error": checkout_result.get("error"),
        "checkout_output": checkout_result.get("output"),
        "candidates": source_root_info.get("candidates", []),
        "artifact": artifact_info,
    }


@router.post("/api/jenkins/scm-info")
def jenkins_scm_info(req: JenkinsScmInfoRequest) -> Dict[str, Any]:
    if not req.scm_url:
        raise HTTPException(status_code=400, detail="scm_url required")
    scm_type = (req.scm_type or "svn").lower()
    if scm_type == "svn":
        from backend.services.scm_registry import resolve_scm_credentials
        info_user, info_pass, _ = resolve_scm_credentials(
            repo_url=req.scm_url,
            override_username=req.scm_username or "",
        )
        info = svn_info_url(
            repo_url=req.scm_url,
            username=info_user,
            password=info_pass,
        )
        if info.get("rc") != 0:
            raise HTTPException(status_code=500, detail=info.get("output") or "svn info failed")
        return {"ok": True, "scm": "svn", "revision": info.get("revision") or "", "output": info.get("output")}
    raise HTTPException(status_code=400, detail="unsupported scm_type")


def _try_svn_revision_range(req: JenkinsImpactTriggerRequest, build_rev: str):
    """현재 로컬 default 버전(A) ↔ 선택 빌드 버전(B) 사이의 svn 정밀 델타를 시도한다.

    A = `svn info <source_root>`의 로컬 작업본 revision(오프라인, 자격증명 불필요).
    B = build_rev(Jenkins lastBuiltRevision). 둘 다 정수 SVN revision일 때만
    `svn diff --summarize -r A:B <repo_url>`로 변경 파일 집합을 구한다.

    반환:
      - (files, True, meta{changed_files_source:'svn_revision_range', ...}) — 성공.
        A==B면 files=[]로 '변경 없음'을 명시(빈 changeSet 오등치와 구분).
      - None — svn 대상 아님/작업본 아님/revision 비정수/조회 실패 → 호출자가 단일
        빌드 changeSet 결과로 폴백(git·비-svn·회귀 0). repo_url은 registry(신뢰)에서만
        오고 revision은 정수검증하므로 SSRF/인자 주입 표면 없음.
    """
    try:
        if not str(build_rev or "").strip().isdigit():
            return None  # svn revision은 정수 — git SHA1/빈 값이면 대상 아님
        from backend.services.scm_registry import get_registry_entry, resolve_scm_credentials
        entry = get_registry_entry(req.scm_id)
        if entry is None or str(entry.scm_type or "").lower() != "svn":
            return None
        repo_url = str(entry.scm_url or "").strip()
        source_root = str(entry.source_root or "").strip()
        if not (repo_url and source_root):
            return None
        from backend.services.local_service import svn_info_url, svn_diff_summarize
        # A: 로컬 작업본 base revision. source_root는 콤마/세미콜론 구분 멀티패스일 수 있으므로
        # (예: 'C:\\...\\NE1AW_PORTING,C:\\...\\PDS128_FBL') 분리해, scm_url과 '같은 리포지토리'인
        # 작업본의 revision을 고른다. svn revision은 리포지토리-전역 정수라 리포 정합(작업본
        # Repository Root가 scm_url을 포함)이 맞는 경로만 A로 신뢰한다 — 불일치/미매칭이면 changeSet
        # 폴백(silent-wrong 방지). svn info는 작업본 대상이라 오프라인 조회.
        scm_norm = repo_url.rstrip("/")
        base_rev = ""
        for _raw in source_root.replace(";", ",").split(","):
            _p = _raw.strip()
            if not _p:
                continue
            _info = svn_info_url(repo_url=_p)
            _rev = str(_info.get("revision") or "").strip()
            if not _rev.isdigit():
                continue
            _root = str(_info.get("repo_root") or "").strip().rstrip("/")
            if _root and not (scm_norm == _root or scm_norm.startswith(_root + "/")):
                continue  # 다른 리포지토리 작업본 — A로 쓰면 무의미
            base_rev = _rev
            break
        if not base_rev.isdigit():
            _logger.warning(
                "svn revision-range skipped: no working copy in source_root matches scm_url (%s) — changeSet fallback",
                repo_url,
            )
            return None
        if base_rev == build_rev:
            # 로컬 default가 이미 빌드 revision과 동일 → 실제 변경 0건(확인됨).
            return [], True, {
                "changed_files_source": "svn_revision_range",
                "baseline_revision": base_rev,
                "build_revision": build_rev,
                "jenkins_changed_file_count": 0,
                "linkage_reason": f"local working copy already at build revision r{build_rev} (no changes)",
            }
        # A>B: 로컬 작업본(A)이 선택 빌드(B)보다 최신 → svn diff -r A:B가 역방향 델타(NEW/DELETE·
        # before/after 뒤집힘, '삭제 TC 제거' 가이드 오발동)를 낸다 → 명시적 changeSet 폴백.
        if int(base_rev) > int(build_rev):
            _logger.warning(
                "svn revision-range skipped: local rev %s newer than build rev %s — changeSet fallback",
                base_rev, build_rev,
            )
            return None
        username, password, _ = resolve_scm_credentials(scm_id=req.scm_id)
        diff = svn_diff_summarize(
            repo_url=repo_url,
            rev_a=base_rev,
            rev_b=build_rev,
            username=username,
            password=password,
        )
        if int(diff.get("rc", 1)) != 0:
            _logger.warning(
                "svn diff -r %s:%s failed (scm=%s): %s",
                base_rev, build_rev, req.scm_id, str(diff.get("output"))[:200],
            )
            return None  # 조회 실패 → changeSet 폴백
        files = [str(x) for x in (diff.get("files") or [])]
        meta: Dict[str, Any] = {
            "changed_files_source": "svn_revision_range",
            "baseline_revision": base_rev,
            "build_revision": build_rev,
            "jenkins_changed_file_count": len(files),
            "linkage_reason": f"svn diff --summarize -r {base_rev}:{build_rev}",
        }
        edit_types = diff.get("edit_types") or {}
        if edit_types:
            meta["changed_file_edit_types"] = edit_types
        return files, True, meta
    except Exception as exc:  # noqa: BLE001 — 어떤 실패든 changeSet로 graceful 폴백
        _logger.warning("svn revision-range diff failed (scm=%s): %s", req.scm_id, exc, exc_info=True)
        return None


def _resolve_jenkins_changed_files(req: JenkinsImpactTriggerRequest):
    """선택한 빌드의 changeSet에서 변경 .c/.h 파일을 가져온다.

    반환: (manual_changed_files, use_manual_only, meta_extra)
      - 성공: (files, True, {...jenkins_changeset...}) — files가 []여도 '빌드 변경 0건'으로
        간주(use_manual_only=True라 로컬 working-copy diff로 잘못 되돌아가지 않음).
      - 자격증명 없음/조회 실패: (None, False, {...local_diff_fallback...}) — 기존 로컬 SCM diff.
    서버 측 Jenkins 자격증명(config.get_jenkins_config)만 사용 — HTTP body로 토큰 받지 않음.
    """
    if not (req.build_number and str(req.job_url or "").strip()):
        return None, False, {"changed_files_source": "local_diff_fallback", "linkage_reason": "no build_number/job_url"}
    try:
        from backend.routers.config import get_jenkins_config
        cfg = get_jenkins_config()
        user = str(cfg.get("username") or "").strip()
        token = str(cfg.get("token") or "").strip()
        if not (user and token):
            return None, False, {"changed_files_source": "local_diff_fallback", "linkage_reason": "jenkins credentials not configured"}
        # SSRF/크리덴셜 유출 방지(fail-closed): 사용자 입력 job_url에 서버 토큰을 실어
        # 보내므로, baseUrl이 설정돼 있고 job_url이 그 하위(또는 동일)이며 '..'가 없을 때만
        # 호출한다. baseUrl 미설정이면 검증 불가 → 조회하지 않고 로컬 diff로 폴백한다.
        base_url = str(cfg.get("baseUrl") or "").strip().rstrip("/")
        job = str(req.job_url or "").strip()
        job_l = job.lower()
        under_base = bool(base_url) and (job_l == base_url.lower() or job_l.startswith(base_url.lower() + "/"))
        if (not under_base) or (".." in job):
            return None, False, {"changed_files_source": "local_diff_fallback", "linkage_reason": "job_url not under configured Jenkins baseUrl"}
        from backend.services.jenkins_service import get_build_changed_files
        res = get_build_changed_files(
            job_url=req.job_url,
            build_number=int(req.build_number),
            username=user,
            api_token=token,
            verify_tls=bool(cfg.get("verifyTls", True)),
        )
        files = [str(x) for x in (res.get("files") or [])]
        build_rev = str(res.get("revision") or "").strip()
        # ── baseline(로컬 default 버전 A) ↔ build 버전(B) 정밀 델타 (svn revision-range) ──
        # 영향은 '직전 빌드 대비 changeSet'이 아니라 '현재 로컬 작업본 revision(A) ↔ 선택
        # 빌드 revision(B)' 사이여야 한다. svn diff --summarize -r A:B로 빌드가 몇 번 끼어
        # 있든 A→B 전체 변경을 잡는다(빈 changeSet=0건 오등치 회피). svn 아님/작업본 아님/
        # 조회 실패 시 아래 단일 빌드 changeSet 결과로 graceful 폴백한다(git·회귀 0).
        svn_range = _try_svn_revision_range(req, build_rev)
        if svn_range is not None:
            return svn_range
        meta: Dict[str, Any] = {
            "changed_files_source": "jenkins_changeset",
            "build_revision": build_rev,
            "jenkins_changed_file_count": len(files),
        }
        # per-file editType(add/edit/delete) — cloudium/원격에서 NEW/DELETE 변경유형 분류의
        # 유일한 근거(로컬 working-copy diff 불가). 비어 있으면(affectedPaths만 제공) 생략.
        edit_types = res.get("edit_types") or {}
        if edit_types:
            meta["changed_file_edit_types"] = edit_types
        return files, True, meta
    except Exception as exc:  # noqa: BLE001 — 조회 실패는 로컬 diff로 graceful fallback
        _logger.warning("jenkins changeset fetch failed (scm=%s build=%s): %s", req.scm_id, req.build_number, exc, exc_info=True)
        return None, False, {"changed_files_source": "local_diff_fallback", "linkage_reason": f"changeset fetch failed: {exc}"[:200]}


def _make_jenkins_impact_trigger(req: JenkinsImpactTriggerRequest, *, source: str):
    """영향도를 '선택한 빌드의 실제 changeSet'에 묶어 ChangeTrigger를 만든다."""
    manual, use_manual_only, meta_extra = _resolve_jenkins_changed_files(req)
    return build_registry_trigger(
        trigger_type="jenkins",
        scm_id=req.scm_id,
        base_ref=req.base_ref,
        dry_run=req.dry_run,
        targets=req.targets or None,
        manual_changed_files=manual,
        use_manual_only=use_manual_only,
        metadata={
            "source": source,
            "build_number": req.build_number,
            "job_url": req.job_url,
            **meta_extra,
        },
    )


@router.post("/api/jenkins/impact/trigger")
def jenkins_impact_trigger(req: JenkinsImpactTriggerRequest) -> Dict[str, Any]:
    try:
        trigger = _make_jenkins_impact_trigger(req, source="api/jenkins/impact/trigger")
    except KeyError:
        raise HTTPException(status_code=404, detail="registry entry not found")
    return run_impact_update(trigger)


@router.post("/api/jenkins/impact/trigger-async")
def jenkins_impact_trigger_async(req: JenkinsImpactTriggerRequest) -> Dict[str, Any]:
    try:
        trigger = _make_jenkins_impact_trigger(req, source="api/jenkins/impact/trigger-async")
    except KeyError:
        raise HTTPException(status_code=404, detail="registry entry not found")
    return start_impact_job(trigger)


@router.post("/api/jenkins/uds/template-upload")
async def jenkins_uds_template_upload(
    file: UploadFile = File(...),
    job_url: str = Form(...),
    cache_root: str = Form(""),
    build_selector: str = Form("lastSuccessfulBuild"),
) -> Dict[str, Any]:
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    reject_upload_in_cloudium(file)
    if not file.filename:
        raise HTTPException(status_code=400, detail="template filename required")
    job_slug = _job_slug(job_url)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ext = Path(file.filename).suffix.lower() or ".docx"
    out_dir = _jenkins_templates_dir(cache_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"uds_template_{job_slug}_{ts}{ext}"
    content = await file.read()
    out_path.write_bytes(content)
    return {
        "ok": True,
        "template_path": str(out_path),
        "filename": out_path.name,
        "build_selector": build_selector,
    }


@router.post("/api/jenkins/uds/generate")
async def jenkins_uds_generate(
    job_url: str = Form(...),
    cache_root: str = Form(""),
    build_selector: str = Form("lastSuccessfulBuild"),
    template_path: str = Form(""),
    source_root: str = Form(""),
    source_only: bool = Form(False),
    req_files: List[UploadFile] = File(default_factory=list),
    req_paths: str = Form(""),
    logic_files: List[UploadFile] = File(default_factory=list),
    files: List[UploadFile] = File(default_factory=list),
    component_list: UploadFile = File(default=None),
    call_relation_mode: str = Form("code"),
    req_types: str = Form(""),
    show_mapping_evidence: bool = Form(False),
) -> Dict[str, Any]:
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    reject_upload_in_cloudium(*(req_files or []), *(logic_files or []), *(files or []), component_list)
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="source_root(코드 루트)가 필요합니다.")
    req_paths_list = _parse_path_list(req_paths)
    has_req_upload = any((f and f.filename) for f in (req_files or []))
    if not has_req_upload and not req_paths_list:
        raise HTTPException(status_code=400, detail="SRS/SDS 요구사항 문서를 최소 1개 이상 제공해주세요.")

    type_list = [t.strip().lower() for t in req_types.split(",") if t.strip()] if req_types else []

    build_root = _resolve_cached_build_root(job_url, cache_root, build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    report_dir = _detect_reports_dir(build_root)
    summary = build_report_summary(report_dir, project_root=repo_root)

    notes: List[str] = []
    for f in files:
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            text = _read_text_from_file(tmp_path)
        except Exception:
            text = ""
        if text:
            notes.append(text.strip())

    req_texts: List[str] = []
    component_map: Dict[str, Dict[str, str]] = {}
    if component_list and component_list.filename:
        tmp = _write_upload_to_temp(component_list, ".json")
        if tmp:
            try:
                component_map = _parse_component_map_file(tmp)
            except Exception:
                component_map = {}
            finally:
                try:
                    tmp.unlink(missing_ok=True)
                except Exception:
                    pass
    srs_texts: List[str] = []
    sds_texts: List[str] = []
    req_doc_paths: List[str] = []
    for idx, f in enumerate(req_files):
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            text = _read_text_from_file(tmp_path)
        except Exception:
            text = ""
        if tmp_path.suffix.lower() == ".docx":
            req_doc_paths.append(str(tmp_path))
        if text:
            req_texts.append(text.strip())
            ftype = type_list[idx] if idx < len(type_list) else ""
            if ftype == "srs":
                srs_texts.append(text.strip())
            elif ftype == "sds":
                sds_texts.append(text.strip())
    for path_str in req_paths_list:
        try:
            p = Path(path_str).expanduser().resolve()
            if not p.exists() or not p.is_file():
                continue
            if not _is_allowed_req_doc(p):
                continue
            text = _read_text_from_file(p)
        except Exception:
            text = ""
        if text:
            req_texts.append(text.strip())
            if p.suffix.lower() == ".docx":
                req_doc_paths.append(str(p))

    jenkins_meta = summary.get("jenkins") if isinstance(summary, dict) else {}
    if not isinstance(jenkins_meta, dict):
        jenkins_meta = {}
    summary_text = summary.get("summary_text", "") if isinstance(summary, dict) else ""
    source_sections: Dict[str, str] = {}
    if source_root_path and source_root_path.exists():
        source_sections = generate_uds_source_sections(
            str(source_root_path),
            component_map=component_map if component_map else None,
        )
    sds_doc_paths: List[str] = []
    for p in req_doc_paths:
        if "sds" in Path(p).name.lower():
            sds_doc_paths.append(str(p))
    if source_sections:
        details = source_sections.get("function_details", {})
        if isinstance(details, dict):
            enrich_function_details_with_docs(
                details,
                source_sections.get("function_table_rows", []),
                req_doc_paths=req_doc_paths,
                sds_doc_paths=sds_doc_paths,
            )
            source_sections["function_details"] = details
            rebuilt_by_name: Dict[str, Any] = {}
            for _, info in details.items():
                if not isinstance(info, dict):
                    continue
                name = str(info.get("name") or "").strip().lower()
                if name:
                    rebuilt_by_name[name] = info
            source_sections["function_details_by_name"] = rebuilt_by_name
    req_from_docs = generate_uds_requirements_from_docs(req_texts) if req_texts else ""
    req_map = _build_req_map_from_doc_paths(req_doc_paths, req_texts) if req_texts or req_doc_paths else {}
    logic_items: List[Dict[str, Any]] = []
    if logic_files:
        logic_dir = _jenkins_logic_dir(cache_root)
        logic_dir.mkdir(parents=True, exist_ok=True)
        ts_logic = datetime.now().strftime("%Y%m%d_%H%M%S")
        for f in logic_files:
            if not f or not f.filename:
                continue
            suffix = Path(f.filename).suffix.lower() or ".png"
            safe_name = "".join(c for c in Path(f.filename).stem if c.isalnum() or c in ("-", "_"))
            out_name = f"logic_{safe_name}_{ts_logic}{suffix}"
            out_path = logic_dir / out_name
            out_path.write_bytes(await f.read())
            logic_items.append(
                {
                    "title": f.filename,
                    "path": str(out_path),
                    "url": f"/api/jenkins/uds/logic?job_url={job_url}&cache_root={cache_root}&filename={out_name}",
                }
            )

    req_source = source_sections.get("requirements", "")
    if source_only:
        req_combined = req_source
    elif req_from_docs and req_source:
        req_combined = "\n".join([req_from_docs.strip(), req_source.strip()]).strip()
    else:
        req_combined = req_from_docs or req_source
    globals_order_list: List[str] = []
    globals_format_sep = ""
    logic_max_children = None
    logic_max_grandchildren = None
    logic_max_depth = None
    uds_payload = {
        "job_url": job_url,
        "build_number": jenkins_meta.get("build_number"),
        "project_name": summary.get("project") if isinstance(summary, dict) else "",
        "summary": summary,
        "overview": summary_text or source_sections.get("overview", ""),
        "requirements": req_combined,
        "interfaces": source_sections.get("interfaces", ""),
        "uds_frames": source_sections.get("uds_frames", ""),
        "notes": "\n".join(notes),
        "logic_diagrams": logic_items,
        "software_unit_design": source_sections.get("software_unit_design", ""),
        "unit_structure": source_sections.get("unit_structure", ""),
        "global_data": source_sections.get("global_data", ""),
        "interface_functions": source_sections.get("interface_functions", ""),
        "internal_functions": source_sections.get("internal_functions", ""),
        "function_table_rows": source_sections.get("function_table_rows", []),
        "global_vars": source_sections.get("global_vars", []),
        "static_vars": source_sections.get("static_vars", []),
        "macro_defs": source_sections.get("macro_defs", []),
        "calibration_params": source_sections.get("calibration_params", []),
        "function_details": source_sections.get("function_details", {}),
        "function_details_by_name": source_sections.get("function_details_by_name", {}),
        "call_map": source_sections.get("call_map", {}),
        "module_map": source_sections.get("module_map", {}),
        "req_map": req_map,
        "globals_info_map": source_sections.get("globals_info_map", {}),
        "common_macros": source_sections.get("common_macros", []),
        "type_defs": source_sections.get("type_defs", []),
        "param_defs": source_sections.get("param_defs", []),
        "version_defs": source_sections.get("version_defs", []),
        "globals_format_order": globals_order_list,
        "globals_format_sep": globals_format_sep,
        "logic_max_children": logic_max_children,
        "logic_max_grandchildren": logic_max_grandchildren,
        "logic_max_depth": logic_max_depth,
        "call_relation_mode": call_relation_mode,
        "show_mapping_evidence": bool(show_mapping_evidence),
        "srs_texts": srs_texts,
        "sds_texts": sds_texts,
    }
    impact_path = _run_impact_analysis_for_uds(
        source_root_path,
        os.getenv("UDS_CHANGED_FILES", ""),
    )
    if impact_path:
        notes_text = str(uds_payload.get("notes") or "").strip()
        uds_payload["notes"] = "\n".join([x for x in [notes_text, f"impact:{impact_path.name}"] if x])
    if source_only and source_sections.get("notes"):
        uds_payload["notes"] = (uds_payload.get("notes") or "").strip()
        uds_payload["notes"] = "\n".join(
            [x for x in [uds_payload["notes"], source_sections.get("notes")] if x]
        )
    job_slug = _job_slug(job_url)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = _jenkins_exports_dir(cache_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"uds_spec_{job_slug}_{ts}.docx"
    tpl = str(template_path).strip() or None
    _generate_docx_with_retry(tpl, uds_payload, out_path)
    _write_uds_payload_sidecar(out_path, uds_payload)
    residual_tbd_path = _write_residual_tbd_report(out_path, (uds_payload.get("summary") or {}).get("mapping") or {})
    validation_path = out_path.with_suffix(".validation.md")
    _jenkins_report_short = 300
    _jenkins_report_long = 600
    ok_validation, _ = _run_report_with_timeout(
        lambda: generate_uds_validation_report(str(out_path), str(validation_path)),
        timeout_seconds=_jenkins_report_short,
        report_name="validation report",
    )
    if not ok_validation:
        validation_path = None
    accuracy_path = out_path.with_suffix(".accuracy.md")
    src_root = str(source_root_path) if source_root_path else ""
    ok_accuracy, _ = _run_report_with_timeout(
        lambda: generate_called_calling_accuracy_report(
            str(out_path),
            src_root,
            str(accuracy_path),
            relation_mode=str(call_relation_mode or "code"),
        ),
        timeout_seconds=_jenkins_report_long,
        report_name="accuracy report",
    )
    if not ok_accuracy:
        accuracy_path = None
    swcom_context_path = out_path.with_suffix(".swcom_context.md")
    ok_swcom, _ = _run_report_with_timeout(
        lambda: generate_swcom_context_report(str(out_path), str(swcom_context_path)),
        timeout_seconds=_jenkins_report_short,
        report_name="swcom context report",
    )
    if not ok_swcom:
        swcom_context_path = None
    swcom_diff_path = None
    confidence_path = out_path.with_suffix(".field_confidence.md")
    ok_confidence, _ = _run_report_with_timeout(
        lambda: generate_asil_related_confidence_report(
            uds_payload,
            str(confidence_path),
            str(out_path),
        ),
        timeout_seconds=_jenkins_report_long,
        report_name="ASIL/Related confidence report",
    )
    if not ok_confidence:
        confidence_path = None
    constraints_path = out_path.with_suffix(".constraints.md")
    ok_constraints, _ = _run_report_with_timeout(
        lambda: generate_uds_constraints_report(uds_payload, str(constraints_path)),
        timeout_seconds=_jenkins_report_short,
        report_name="constraints report",
    )
    if not ok_constraints:
        constraints_path = None
    quality_gate_path = out_path.with_suffix(".quality_gate.md")
    ok_quality_gate, _ = _run_report_with_timeout(
        lambda: generate_uds_field_quality_gate_report(str(out_path), str(quality_gate_path)),
        timeout_seconds=_jenkins_report_short,
        report_name="field quality gate report",
    )
    if not ok_quality_gate:
        quality_gate_path = None

    # Quality DB recording (non-fatal)
    try:
        from backend.helpers import _compute_quick_quality_gate, _enrich_function_quality_fields
        from workflow.quality.recorder import record_uds_run
        # local 경로와 동일하게 enrich 후 quick_gate 계산 → 경로 간 점수 일관성.
        _enrich_function_quality_fields(uds_payload)
        record_uds_run(
            _compute_quick_quality_gate(uds_payload),
            output_path=str(out_path),
        )
    except Exception:
        pass

    preview_html = generate_uds_preview_html(uds_payload)
    preview_path = out_path.with_suffix(".html")
    preview_path.write_text(preview_html, encoding="utf-8")
    return {
        "ok": True,
        "filename": out_path.name,
        "download_url": f"/api/jenkins/uds/download?job_url={job_url}&cache_root={cache_root}&filename={out_path.name}",
        "preview_url": f"/api/jenkins/uds/preview?job_url={job_url}&cache_root={cache_root}&filename={preview_path.name}",
        "validation_path": str(validation_path) if validation_path else "",
        "accuracy_path": str(accuracy_path) if accuracy_path else "",
        "swcom_context_path": str(swcom_context_path) if swcom_context_path else "",
        "swcom_diff_path": str(swcom_diff_path) if swcom_diff_path else "",
        "confidence_path": str(confidence_path) if confidence_path else "",
        "constraints_path": str(constraints_path) if constraints_path else "",
        "quality_gate_path": str(quality_gate_path) if quality_gate_path else "",
        "impact_path": str(impact_path) if impact_path else "",
        "residual_tbd_report_path": str(residual_tbd_path) if residual_tbd_path else "",
    }


@router.post("/api/jenkins/uds/generate-async")
async def jenkins_uds_generate_async(
    job_url: str = Form(...),
    cache_root: str = Form(""),
    build_selector: str = Form("lastSuccessfulBuild"),
    template_path: str = Form(""),
    source_root: str = Form(""),
    source_only: bool = Form(False),
    req_files: List[UploadFile] = File(default_factory=list),
    req_paths: str = Form(""),
    logic_files: List[UploadFile] = File(default_factory=list),
    files: List[UploadFile] = File(default_factory=list),
    component_list: UploadFile = File(default=None),
    logic_source: str = Form(""),
    logic_max_children: Optional[int] = Form(None),
    logic_max_grandchildren: Optional[int] = Form(None),
    logic_max_depth: Optional[int] = Form(None),
    globals_format_order: str = Form(""),
    globals_format_sep: str = Form(""),
    globals_format_with_labels: bool = Form(True),
    ai_enable: bool = Form(False),
    ai_example_path: str = Form(""),
    ai_example_file: UploadFile = File(default=None),
    ai_detailed: bool = Form(True),
    rag_top_k: Optional[int] = Form(None),
    rag_categories: str = Form(""),
) -> Dict[str, Any]:
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    reject_upload_in_cloudium(
        *(req_files or []), *(logic_files or []), *(files or []),
        component_list, ai_example_file,
    )
    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증, 전체를 generate에 전달
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="source_root(코드 루트)가 필요합니다.")
    job_id = uuid.uuid4().hex
    _set_progress(
        "uds",
        job_url,
        build_selector,
        {
            "stage": "start",
            "percent": 1,
            "message": "UDS 생성 준비 중",
            "done": False,
            "error": "",
        },
        job_id=job_id,
    )

    req_paths_list = _parse_path_list(req_paths)
    has_req_upload = any((f and f.filename) for f in (req_files or []))
    if not has_req_upload and not req_paths_list:
        raise HTTPException(status_code=400, detail="SRS/SDS 요구사항 문서를 최소 1개 이상 제공해주세요.")
    globals_order_list = [
        x.strip()
        for x in re.split(r"[,\|;]+", globals_format_order or "")
        if x.strip()
    ]
    req_file_paths: List[Path] = []
    logic_file_paths: List[Path] = []
    note_file_paths: List[Path] = []
    ai_example_text = ""
    component_map: Dict[str, Dict[str, str]] = {}

    for f in req_files:
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            req_file_paths.append(Path(tmp.name))

    for f in logic_files:
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".png"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            logic_file_paths.append(Path(tmp.name))

    for f in files:
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            note_file_paths.append(Path(tmp.name))

    if component_list and component_list.filename:
        tmp = _write_upload_to_temp(component_list, ".json")
        if tmp:
            try:
                component_map = _parse_component_map_file(tmp)
            except Exception:
                component_map = {}
            finally:
                try:
                    tmp.unlink(missing_ok=True)
                except Exception:
                    pass

    if ai_example_path:
        try:
            p = Path(ai_example_path).expanduser().resolve()
            if p.exists() and p.is_file():
                ai_example_text = _read_text_from_file(p)
        except Exception:
            ai_example_text = ""
    if ai_example_file and ai_example_file.filename:
        try:
            suffix = Path(ai_example_file.filename).suffix.lower() or ".txt"
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(await ai_example_file.read())
                ai_example_text = _read_text_from_file(Path(tmp.name))
        except Exception:
            ai_example_text = ai_example_text or ""
    if ai_enable and not ai_example_text:
        for cand in [
            repo_root / "docs" / "UDSPDM01_UDS.txt",
            repo_root / "docs" / "HDPDM01_UDS.txt",
        ]:
            try:
                if cand.exists() and cand.is_file():
                    ai_example_text = _read_text_from_file(cand)
                    break
            except Exception:
                continue
    if ai_enable and not ai_example_text:
        try:
            ref_suds_path = Path(config.UDS_REF_SUDS_PATH)
            if ref_suds_path.exists() and ref_suds_path.is_file():
                ai_example_text = _read_text_from_file(ref_suds_path)
        except Exception:
            pass

    def _progress_cb(stage: str, data: Dict[str, Any]) -> None:
        if not isinstance(data, dict):
            data = {}
        _set_progress(
            "uds",
            job_url,
            build_selector,
            data,
            job_id=job_id,
        )

    def _worker() -> None:
        try:
            result = _uds_generate_from_paths(
                job_url=job_url,
                cache_root=cache_root,
                build_selector=build_selector,
                template_path=template_path,
                source_root=source_root,
                source_only=source_only,
                req_file_paths=req_file_paths,
                note_file_paths=note_file_paths,
                logic_file_paths=logic_file_paths,
                req_paths=req_paths_list,
                logic_source=logic_source,
                logic_max_children=logic_max_children,
                logic_max_grandchildren=logic_max_grandchildren,
                logic_max_depth=logic_max_depth,
                globals_format_order=",".join(globals_order_list),
                globals_format_sep=globals_format_sep,
                globals_format_with_labels=globals_format_with_labels,
                ai_enable=bool(ai_enable),
                ai_example_text=ai_example_text,
                ai_detailed=bool(ai_detailed),
                rag_top_k=_safe_int(rag_top_k, getattr(config, "AGENT_RAG_TOP_K_DEFAULT", 3))
                if rag_top_k is not None
                else None,
                rag_categories=_split_csv(rag_categories),
                progress_cb=_progress_cb,
                component_map=component_map if component_map else None,
            )
            _set_progress(
                "uds",
                job_url,
                build_selector,
                {
                    "stage": "done",
                    "percent": 100,
                    "message": "UDS 생성 완료",
                    "done": True,
                    "result": result,
                },
                job_id=job_id,
            )
        except Exception as exc:
            tb = traceback.format_exc()
            err_summary = str(exc)[:500]
            _logger.error("[UDS_ASYNC][%s] FAILED: %s\n%s", job_id, err_summary, tb)
            _set_progress(
                "uds",
                job_url,
                build_selector,
                {
                    "stage": "error",
                    "percent": 100,
                    "message": f"UDS 생성 실패: {err_summary}",
                    "done": True,
                    "error": err_summary,
                    "error_detail": tb,
                },
                job_id=job_id,
            )

    threading.Thread(target=wrap_with_user(_worker), daemon=True).start()
    return {"ok": True, "job_id": job_id}


@router.get("/api/jenkins/uds/download")
def jenkins_uds_download(job_url: str, cache_root: str, filename: str) -> FileResponse:
    if not filename:
        raise HTTPException(status_code=400, detail="filename required")
    out_dir = _jenkins_exports_dir(cache_root)
    target = (out_dir / filename).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="file not found")
    return FileResponse(str(target), filename=target.name)


@router.get("/api/jenkins/uds/preview")
def jenkins_uds_preview(job_url: str, cache_root: str, filename: str) -> Dict[str, Any]:
    if not filename:
        raise HTTPException(status_code=400, detail="filename required")
    out_dir = _jenkins_exports_dir(cache_root)
    try:
        target = safe_resolve_under(out_dir, filename)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid filename")
    if not target.exists():
        raise HTTPException(status_code=404, detail="file not found")
    try:
        text = target.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        text = ""
    if target.suffix.lower() == ".md":
        escaped = (
            "<pre>"
            + text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            + "</pre>"
        )
        return {"ok": True, "html": escaped}
    return {"ok": True, "html": text}


@router.get("/api/jenkins/uds/logic")
def jenkins_uds_logic(job_url: str, cache_root: str, filename: str) -> FileResponse:
    if not filename:
        raise HTTPException(status_code=400, detail="filename required")
    logic_dir = _jenkins_logic_dir(cache_root)
    try:
        target = safe_resolve_under(logic_dir, filename)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid filename")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="file not found")
    return FileResponse(str(target), filename=target.name)


@router.get("/api/jenkins/uds/list")
def jenkins_uds_list(job_url: str, cache_root: str) -> Dict[str, Any]:
    try:
        from report.constants import UDS_PLACEHOLDERS as _UDS_PH
    except ImportError:
        _UDS_PH = []
    job_slug = _job_slug(job_url)
    out_dir = _jenkins_exports_dir(cache_root)
    if not out_dir.exists():
        return {"ok": True, "items": [], "placeholders": _UDS_PH}
    meta = _load_uds_meta(out_dir, job_slug)
    labels = meta.get("labels") if isinstance(meta.get("labels"), dict) else {}
    items: List[Dict[str, Any]] = []
    for p in sorted(out_dir.glob(f"uds_spec_{job_slug}_*.docx"), reverse=True):
        try:
            stat = p.stat()
            preview_html = p.with_suffix(".html")
            preview_md = p.with_suffix(".md")
            preview = preview_html if preview_html.exists() else preview_md
            items.append(
                {
                    "filename": p.name,
                    "size": stat.st_size,
                    "mtime": stat.st_mtime,
                    "label": labels.get(p.name, ""),
                    "download_url": f"/api/jenkins/uds/download?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
                    "preview_url": f"/api/jenkins/uds/preview?job_url={job_url}&cache_root={cache_root}&filename={preview.name}"
                    if preview.exists()
                    else "",
                }
            )
        except Exception:
            continue
    return {"ok": True, "items": items, "placeholders": _UDS_PH}


@router.get("/api/jenkins/uds/view")
def jenkins_uds_view(
    job_url: str,
    cache_root: str,
    filename: str,
    q: str = Query(default=""),
    swcom: str = Query(default="all"),
    asil: str = Query(default="all"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    trace_q: str = Query(default=""),
    trace_page: int = Query(default=1, ge=1),
    trace_page_size: int = Query(default=100, ge=1, le=500),
) -> Dict[str, Any]:
    if not filename:
        raise HTTPException(status_code=400, detail="filename required")
    if Path(filename).suffix.lower() != ".docx":
        raise HTTPException(status_code=400, detail="filename must be .docx")
    out_dir = _jenkins_exports_dir(cache_root)
    try:
        docx_path = safe_resolve_under(out_dir, filename)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid filename")
    if not docx_path.exists():
        raise HTTPException(status_code=404, detail="file not found")
    accuracy_path = docx_path.with_suffix(".accuracy.md")
    quality_gate_path = docx_path.with_suffix(".quality_gate.md")
    payload = _get_uds_view_payload_cached(
        docx_path,
        accuracy_path if accuracy_path.exists() else None,
        quality_gate_path if quality_gate_path.exists() else None,
    )
    payload = _apply_uds_view_filters(
        payload,
        q=q,
        swcom=swcom,
        asil=asil,
        page=page,
        page_size=page_size,
        trace_q=trace_q,
        trace_page=trace_page,
        trace_page_size=trace_page_size,
    )
    payload["download_url"] = (
        f"/api/jenkins/uds/download?job_url={job_url}&cache_root={cache_root}&filename={docx_path.name}"
    )
    preview_candidate = docx_path.with_suffix(".html")
    if not preview_candidate.exists():
        preview_candidate = docx_path.with_suffix(".md")
    payload["preview_url"] = (
        f"/api/jenkins/uds/preview?job_url={job_url}&cache_root={cache_root}&filename={preview_candidate.name}"
        if preview_candidate.exists()
        else ""
    )
    payload["accuracy_path"] = str(accuracy_path) if accuracy_path.exists() else ""
    payload["quality_gate_path"] = str(quality_gate_path) if quality_gate_path.exists() else ""
    residual_tbd_path = docx_path.with_suffix(".residual_tbd.md")
    payload["residual_tbd_report_path"] = str(residual_tbd_path) if residual_tbd_path.exists() else ""
    return payload


def _parse_excel_preview(file_path: Path, max_rows: int = 30) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    wb = load_workbook(str(file_path), read_only=True, data_only=True, keep_vba=False)
    sheets: List[Dict[str, Any]] = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        mr = ws.max_row or 0
        mc = ws.max_column or 0
        headers: List[str] = []
        rows: List[List[Any]] = []
        if mr and mc:
            col_limit = min(mc, 20)
            for c in range(1, col_limit + 1):
                val = ws.cell(row=1, column=c).value
                headers.append(str(val) if val is not None else f"Col{c}")
            for r in range(1, min(mr + 1, 1 + max_rows)):
                row_data: List[Any] = []
                for c in range(1, col_limit + 1):
                    cell_val = ws.cell(row=r, column=c).value
                    if cell_val is None:
                        row_data.append("")
                    else:
                        row_data.append(cell_val if isinstance(cell_val, (int, float)) else str(cell_val).strip()[:200])
                if any(v != "" for v in row_data):
                    rows.append(row_data)
        sheets.append({"name": sname, "headers": headers, "rows": rows, "total_rows": mr, "total_cols": mc})
    names = list(wb.sheetnames)
    wb.close()
    return {"filename": file_path.name, "sheets": sheets, "sheet_names": names}


def _build_sts_function_details(source_root_path: Path, req_doc_paths: List[str], sds_doc_paths: List[str]) -> Dict[str, Any]:
    sections = generate_uds_source_sections(str(source_root_path))
    details = sections.get("function_details", {}) if isinstance(sections, dict) else {}
    if isinstance(details, dict):
        enrich_function_details_with_docs(
            details,
            sections.get("function_table_rows", []) if isinstance(sections, dict) else [],
            req_doc_paths=req_doc_paths,
            sds_doc_paths=sds_doc_paths,
        )
    return details if isinstance(details, dict) else {}


@router.post("/api/jenkins/sts/generate-async")
async def jenkins_sts_generate_async(
    job_url: str = Form(...),
    cache_root: str = Form(""),
    build_selector: str = Form("lastSuccessfulBuild"),
    source_root: str = Form(""),
    srs_path: str = Form(""),
    sds_path: str = Form(""),
    uds_path: str = Form(""),
    stp_path: str = Form(""),
    req_paths: str = Form(""),
    req_files: List[UploadFile] = File(default_factory=list),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_tc_per_req: int = Form(5),
) -> Dict[str, Any]:
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    from sts_generator import generate_sts
    reject_upload_in_cloudium(*(req_files or []))

    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="source_root is required")
    job_id = uuid.uuid4().hex
    req_paths_list = _parse_path_list(req_paths)
    req_texts: List[str] = []
    req_doc_paths: List[str] = []
    sds_doc_paths: List[str] = []
    srs_docx_path: Optional[str] = ""
    if srs_path:
        p = Path(srs_path).expanduser().resolve()
        if p.exists() and p.is_file():
            srs_docx_path = str(p)
    for path_str in req_paths_list:
        try:
            p = Path(path_str).expanduser().resolve()
            if not p.exists() or not p.is_file():
                continue
            text = _read_text_from_file(p)
            if text:
                req_texts.append(text.strip())
                if p.suffix.lower() == ".docx":
                    req_doc_paths.append(str(p))
                if "sds" in p.name.lower():
                    sds_doc_paths.append(str(p))
                if not srs_docx_path and "srs" in p.name.lower() and p.suffix.lower() == ".docx":
                    srs_docx_path = str(p)
        except Exception:
            continue
    for f in (req_files or []):
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            text = _read_text_from_file(tmp_path)
            if text:
                req_texts.append(text.strip())
                if tmp_path.suffix.lower() == ".docx":
                    req_doc_paths.append(str(tmp_path))
                if "sds" in f.filename.lower():
                    sds_doc_paths.append(str(tmp_path))
                if not srs_docx_path and "srs" in f.filename.lower() and suffix == ".docx":
                    srs_docx_path = str(tmp_path)
        except Exception:
            continue
    if not req_texts and not srs_docx_path:
        raise HTTPException(status_code=400, detail="SRS document is required")
    def _resolve_opt_j(val: str) -> Optional[str]:
        if not val:
            return None
        p2 = Path(val).expanduser().resolve()
        return str(p2) if p2.exists() and p2.is_file() else None

    sds_docx_path = _resolve_opt_j(sds_path)
    uds_file_path = _resolve_opt_j(uds_path)
    stp_docx_path = _resolve_opt_j(stp_path)

    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)
    out_filename, out_path = _build_jenkins_excel_output(cache_root, "sts", f"sts_{_job_slug(job_url)}", tpl_path)
    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}_STS",
        "version": version,
        "asil_level": asil_level,
        "max_tc_per_req": max_tc_per_req,
        "default_test_env": "SwTE_01",
    }
    _set_progress("jenkins_sts", job_url, build_selector, {"stage": "start", "percent": 1, "message": "STS start", "done": False, "error": ""}, job_id=job_id)

    def _on_progress(pct: int, msg: str):
        _set_progress("jenkins_sts", job_url, build_selector, {"stage": "generation", "percent": max(10, min(pct, 95)), "message": msg}, job_id=job_id)

    def _worker() -> None:
        try:
            _set_progress("jenkins_sts", job_url, build_selector, {"stage": "source_analysis", "percent": 5, "message": "Analyzing source"}, job_id=job_id)
            function_details = _build_sts_function_details(source_root_path, req_doc_paths, sds_doc_paths)
            result = generate_sts(
                requirements_text=req_texts,
                function_details=function_details,
                output_path=str(out_path),
                template_path=tpl_path,
                project_config=project_config,
                srs_docx_path=srs_docx_path,
                sds_docx_path=sds_docx_path,
                uds_path=uds_file_path,
                stp_path=stp_docx_path,
                on_progress=_on_progress,
            )
            download_url = f"/api/jenkins/sts/download?job_url={job_url}&cache_root={cache_root}&filename={out_filename}"
            preview_url = f"/api/jenkins/sts/preview?job_url={job_url}&cache_root={cache_root}&filename={out_filename}"
            payload = _build_excel_artifact_payload(
                "sts",
                {
                    "ok": True,
                    "output_path": str(out_path),
                    "filename": out_filename,
                    "download_url": download_url,
                    "test_case_count": result.get("test_case_count", 0),
                    "elapsed_seconds": result.get("elapsed_seconds", 0),
                    "quality_report": result.get("quality_report", {}),
                    "trace_coverage": result.get("trace_coverage", {}),
                    "validation": result.get("validation", {}),
                    "validation_report_path": result.get("validation_report_path", ""),
                    "build_label": _build_label(job_url, cache_root, build_selector),
                },
                output_path=str(out_path),
                filename=out_filename,
                download_url=download_url,
                preview_url=preview_url,
            )
            _write_excel_artifact_sidecar(out_path, "sts", payload)
            _set_progress("jenkins_sts", job_url, build_selector, {"stage": "done", "percent": 100, "message": "STS complete", "done": True, "error": "", "result": payload}, job_id=job_id)
        except Exception as exc:
            _set_progress("jenkins_sts", job_url, build_selector, {"stage": "error", "percent": 100, "message": str(exc)[:300], "done": True, "error": str(exc)[:500]}, job_id=job_id)

    threading.Thread(target=wrap_with_user(_worker), daemon=True).start()
    return {"ok": True, "job_id": job_id}


@router.get("/api/jenkins/sts/progress")
def jenkins_sts_progress(job_url: str, build_selector: str = "lastSuccessfulBuild", job_id: str = "") -> Dict[str, Any]:
    data = _get_progress("jenkins_sts", job_url, build_selector, job_id)
    return {"ok": bool(data), "progress": data}


@router.get("/api/jenkins/sts/download")
def jenkins_sts_download(job_url: str, cache_root: str, filename: str) -> FileResponse:
    target = (_jenkins_sts_dir(cache_root) / filename).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="STS file not found")
    return FileResponse(str(target), filename=target.name, media_type=_excel_media_type(target))


@router.get("/api/jenkins/sts/list")
def jenkins_sts_list(job_url: str, cache_root: str) -> Dict[str, Any]:
    out_dir = _jenkins_sts_dir(cache_root)
    if not out_dir.exists():
        return {"ok": True, "items": []}
    items: List[Dict[str, Any]] = []
    for p in sorted(out_dir.glob("*.xls*"), reverse=True):
        payload = _load_excel_artifact_payload(
            p,
            "sts",
            download_url=f"/api/jenkins/sts/download?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
            preview_url=f"/api/jenkins/sts/preview?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
            build_label=_infer_build_label_for_artifact(job_url, cache_root, p, "lastSuccessfulBuild"),
        )
        items.append({
            "filename": p.name,
            "size": p.stat().st_size,
            "mtime": p.stat().st_mtime,
            "download_url": f"/api/jenkins/sts/download?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
            "preview_url": f"/api/jenkins/sts/preview?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
            "validation_report_path": payload.get("validation_report_path", ""),
            "residual_report_path": payload.get("residual_report_path", ""),
            "summary": payload.get("summary", {}),
        })
    return {"ok": True, "items": items}


@router.get("/api/jenkins/sts/preview")
def jenkins_sts_preview(job_url: str, cache_root: str, filename: str, max_rows: int = 30) -> Dict[str, Any]:
    target = (_jenkins_sts_dir(cache_root) / filename).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="STS file not found")
    return _parse_excel_preview(target, max_rows)


@router.get("/api/jenkins/sts/view")
def jenkins_sts_view(job_url: str, cache_root: str, filename: str) -> Dict[str, Any]:
    target = (_jenkins_sts_dir(cache_root) / filename).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="STS file not found")
    return _load_excel_artifact_payload(
        target,
        "sts",
        download_url=f"/api/jenkins/sts/download?job_url={job_url}&cache_root={cache_root}&filename={target.name}",
        preview_url=f"/api/jenkins/sts/preview?job_url={job_url}&cache_root={cache_root}&filename={target.name}",
        build_label=_infer_build_label_for_artifact(job_url, cache_root, target, "lastSuccessfulBuild"),
    )


@router.post("/api/jenkins/suts/generate-async")
async def jenkins_suts_generate_async(
    job_url: str = Form(...),
    cache_root: str = Form(""),
    build_selector: str = Form("lastSuccessfulBuild"),
    source_root: str = Form(""),
    template_path: str = Form(""),
    project_id: str = Form(""),
    version: str = Form("v1.00"),
    asil_level: str = Form(""),
    max_sequences: int = Form(6),
) -> Dict[str, Any]:
    from suts_generator import generate_suts

    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.exists() or not source_root_path.is_dir():
        raise HTTPException(status_code=400, detail="source_root is required")
    tpl_path: Optional[str] = None
    if template_path:
        p = Path(template_path).expanduser().resolve()
        if p.exists() and p.is_file():
            tpl_path = str(p)
    out_filename, out_path = _build_jenkins_excel_output(cache_root, "suts", f"suts_{_job_slug(job_url)}", tpl_path)
    project_config = {
        "project_id": project_id or "PROJECT",
        "doc_id": f"{project_id or 'PROJECT'}-SUTS",
        "version": version,
        "asil_level": asil_level,
    }
    job_id = uuid.uuid4().hex
    _set_progress("jenkins_suts", job_url, build_selector, {"stage": "start", "percent": 1, "message": "SUTS start", "done": False, "error": ""}, job_id=job_id)

    def _on_progress(pct: int, msg: str):
        _set_progress("jenkins_suts", job_url, build_selector, {"stage": "generation", "percent": max(10, min(pct, 95)), "message": msg}, job_id=job_id)

    def _worker() -> None:
        try:
            _set_progress("jenkins_suts", job_url, build_selector, {"stage": "source_analysis", "percent": 5, "message": "Analyzing source"}, job_id=job_id)
            result = generate_suts(
                source_root=str(source_root_path),
                output_path=str(out_path),
                template_path=tpl_path,
                project_config=project_config,
                max_sequences=max_sequences,
                on_progress=_on_progress,
            )
            download_url = f"/api/jenkins/suts/download?job_url={job_url}&cache_root={cache_root}&filename={out_filename}"
            preview_url = f"/api/jenkins/suts/preview?job_url={job_url}&cache_root={cache_root}&filename={out_filename}"
            payload = _build_excel_artifact_payload(
                "suts",
                {
                    "ok": True,
                    "output_path": str(out_path),
                    "filename": out_filename,
                    "download_url": download_url,
                    "test_case_count": result.get("test_case_count", 0),
                    "total_sequences": result.get("total_sequences", 0),
                    "elapsed_seconds": result.get("elapsed_seconds", 0),
                    "quality_report": result.get("quality_report", {}),
                    "validation": result.get("validation", {}),
                    "validation_report_path": result.get("validation_report_path", ""),
                    "build_label": _build_label(job_url, cache_root, build_selector),
                },
                output_path=str(out_path),
                filename=out_filename,
                download_url=download_url,
                preview_url=preview_url,
            )
            _write_excel_artifact_sidecar(out_path, "suts", payload)
            _set_progress("jenkins_suts", job_url, build_selector, {"stage": "done", "percent": 100, "message": "SUTS complete", "done": True, "error": "", "result": payload}, job_id=job_id)
        except Exception as exc:
            _set_progress("jenkins_suts", job_url, build_selector, {"stage": "error", "percent": 100, "message": str(exc)[:300], "done": True, "error": str(exc)[:500]}, job_id=job_id)

    threading.Thread(target=wrap_with_user(_worker), daemon=True).start()
    return {"ok": True, "job_id": job_id}


@router.get("/api/jenkins/suts/progress")
def jenkins_suts_progress(job_url: str, build_selector: str = "lastSuccessfulBuild", job_id: str = "") -> Dict[str, Any]:
    data = _get_progress("jenkins_suts", job_url, build_selector, job_id)
    return {"ok": bool(data), "progress": data}


@router.get("/api/jenkins/suts/download")
def jenkins_suts_download(job_url: str, cache_root: str, filename: str) -> FileResponse:
    target = (_jenkins_suts_dir(cache_root) / filename).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="SUTS file not found")
    return FileResponse(str(target), filename=target.name, media_type=_excel_media_type(target))


@router.get("/api/jenkins/suts/list")
def jenkins_suts_list(job_url: str, cache_root: str) -> Dict[str, Any]:
    out_dir = _jenkins_suts_dir(cache_root)
    if not out_dir.exists():
        return {"ok": True, "items": []}
    items: List[Dict[str, Any]] = []
    for p in sorted(out_dir.glob("*.xls*"), reverse=True):
        payload = _load_excel_artifact_payload(
            p,
            "suts",
            download_url=f"/api/jenkins/suts/download?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
            preview_url=f"/api/jenkins/suts/preview?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
            build_label=_infer_build_label_for_artifact(job_url, cache_root, p, "lastSuccessfulBuild"),
        )
        items.append({
            "filename": p.name,
            "size": p.stat().st_size,
            "mtime": p.stat().st_mtime,
            "download_url": f"/api/jenkins/suts/download?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
            "preview_url": f"/api/jenkins/suts/preview?job_url={job_url}&cache_root={cache_root}&filename={p.name}",
            "validation_report_path": payload.get("validation_report_path", ""),
            "residual_report_path": payload.get("residual_report_path", ""),
            "summary": payload.get("summary", {}),
        })
    return {"ok": True, "items": items}


@router.get("/api/jenkins/suts/preview")
def jenkins_suts_preview(job_url: str, cache_root: str, filename: str, max_rows: int = 30) -> Dict[str, Any]:
    target = (_jenkins_suts_dir(cache_root) / filename).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="SUTS file not found")
    return _parse_excel_preview(target, max_rows)


@router.get("/api/jenkins/suts/view")
def jenkins_suts_view(job_url: str, cache_root: str, filename: str) -> Dict[str, Any]:
    target = (_jenkins_suts_dir(cache_root) / filename).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="SUTS file not found")
    return _load_excel_artifact_payload(
        target,
        "suts",
        download_url=f"/api/jenkins/suts/download?job_url={job_url}&cache_root={cache_root}&filename={target.name}",
        preview_url=f"/api/jenkins/suts/preview?job_url={job_url}&cache_root={cache_root}&filename={target.name}",
        build_label=_infer_build_label_for_artifact(job_url, cache_root, target, "lastSuccessfulBuild"),
    )


@router.post("/api/jenkins/suts/export-vectorcast")
def jenkins_suts_export_vectorcast(
    job_url: str = Form(...),
    cache_root: str = Form(""),
    build_selector: str = Form("lastSuccessfulBuild"),
    filename: str = Form(""),
    source_root: str = Form(""),
    project_id: str = Form(""),
    compiler: str = Form("CC"),
) -> Dict[str, Any]:
    """Generate a VectorCAST unit-test package from a Jenkins SUTS artifact."""
    from tools.export_suts_vectorcast import export_suts_to_vectorcast_model
    from tools.export_vectorcast_script import export_vectorcast_package

    out_dir = _jenkins_suts_dir(cache_root)
    if filename:
        xlsm_path = (out_dir / filename).resolve()
    else:
        candidates = sorted(out_dir.glob("*.xlsm"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not candidates:
            raise HTTPException(status_code=404, detail="No Jenkins SUTS file found")
        xlsm_path = candidates[0].resolve()
    if not xlsm_path.exists():
        raise HTTPException(status_code=404, detail="SUTS file not found")

    resolved_source_root = str(source_root or "").strip()
    cfg = load_vectorcast_project_config(project_id=project_id, source_root=resolved_source_root)
    effective_project_id = str(project_id or cfg.get("project_id") or "VECTORCAST").strip()
    effective_source_root = resolved_source_root or str(cfg.get("source_root") or "").strip()

    package_name = f"suts_vectorcast_{_job_slug(job_url)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    package_dir = _jenkins_exports_dir(cache_root) / "vectorcast" / package_name
    package_dir.mkdir(parents=True, exist_ok=True)
    intermediate_json = package_dir / "suts_vectorcast_model.json"
    warnings_md = package_dir / "suts_vectorcast_warnings.md"

    try:
        model = export_suts_to_vectorcast_model(
            str(xlsm_path),
            str(intermediate_json),
            warnings_md=str(warnings_md),
            project_id=effective_project_id,
        )
        manifest = export_vectorcast_package(
            str(intermediate_json),
            str(package_dir),
            package_name=package_name,
            source_root=effective_source_root,
            compiler=str(cfg.get("compiler") or compiler or "CC"),
            project_config=cfg,
        )
    except Exception as exc:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"VectorCAST package generation failed: {exc}")

    unit_names = [str(unit.get("unit_name") or "") for unit in model.get("units") or []]
    return _build_jenkins_vectorcast_response(
        job_url=job_url,
        cache_root=cache_root,
        build_selector=build_selector,
        package_dir=package_dir,
        package_name=package_name,
        manifest=manifest,
        project_config=cfg,
        units=unit_names,
    )


@router.post("/api/jenkins/uds/requirements-preview")
async def jenkins_uds_requirements_preview(
    req_files: List[UploadFile] = File(default_factory=list),
    req_paths: str = Form(""),
    source_root: str = Form(""),
) -> Dict[str, Any]:
    from backend.services.resolver_helpers import reject_upload_in_cloudium
    reject_upload_in_cloudium(*(req_files or []))
    req_texts: List[str] = []
    for f in req_files:
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower() or ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await f.read())
            tmp_path = Path(tmp.name)
        try:
            text = _read_text_from_file(tmp_path)
        except Exception:
            text = ""
        if text:
            req_texts.append(text.strip())
    # N20 fix: cloudium 모드에서 backend python.exe는 클라우디움 폴더 권한 없음
    # → Path.exists() / _read_text_from_file의 직접 read 모두 실패. resolver를
    # 통해 worker IPC로 read 후 임시 파일에 저장 → _read_text_from_file 호출.
    # local 모드에서는 resolver.read_bytes도 직접 read이라 동일 동작.
    from backend.services.file_resolver import get_resolver
    from backend.services.resolver_helpers import enforce_resolver_access
    _resolver = get_resolver()
    for path_str in _parse_path_list(req_paths):
        text = ""
        try:
            enforce_resolver_access(path_str)  # cloudium 게이트 + 화이트리스트
            if not _resolver.exists(path_str):
                continue
            suffix = Path(path_str).suffix or ".txt"
            if not _is_allowed_req_doc(Path(path_str)):
                continue
            data = _resolver.read_bytes(path_str)
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(data)
                tmp_p = Path(tmp.name)
            try:
                text = _read_text_from_file(tmp_p)
            except Exception:
                text = ""
            finally:
                try:
                    tmp_p.unlink()
                except Exception:
                    pass
        except (PermissionError, OSError):
            text = ""
        except Exception:
            text = ""
        if text:
            req_texts.append(text.strip())
    preview = generate_uds_requirements_preview(req_texts)
    mapping = generate_uds_requirements_mapping(preview.get("items") or [])
    compare = None
    function_mapping = None
    if source_root:
        try:
            compare = generate_uds_requirements_compare(preview.get("items") or [], source_root)
        except Exception:
            compare = None
        try:
            function_mapping = generate_uds_function_mapping(req_texts, source_root)
        except Exception:
            function_mapping = None
    return {
        "ok": True,
        "preview": preview,
        "mapping": mapping,
        "compare": compare,
        "function_mapping": function_mapping,
    }


@router.post("/api/jenkins/uds/diff")
def jenkins_uds_diff(req: UdsDiffRequest) -> Dict[str, Any]:
    out_dir = _jenkins_exports_dir(req.cache_root)
    try:
        a_path = safe_resolve_under(out_dir, req.filename_a)
        b_path = safe_resolve_under(out_dir, req.filename_b)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid filename")
    if a_path.suffix.lower() == ".docx":
        html = a_path.with_suffix(".html")
        if html.exists():
            a_path = html
    if b_path.suffix.lower() == ".docx":
        html = b_path.with_suffix(".html")
        if html.exists():
            b_path = html
    if not a_path.exists() or not b_path.exists():
        raise HTTPException(status_code=404, detail="file not found")
    a_html = a_path.read_text(encoding="utf-8", errors="ignore")
    b_html = b_path.read_text(encoding="utf-8", errors="ignore")
    a_sections = parse_uds_preview_html(a_html)
    b_sections = parse_uds_preview_html(b_html)
    diff: Dict[str, Any] = {}
    for key in sorted(set(a_sections.keys()) | set(b_sections.keys())):
        a_items = a_sections.get(key, [])
        b_items = b_sections.get(key, [])
        added = [x for x in b_items if x not in a_items]
        removed = [x for x in a_items if x not in b_items]
        diff[key] = {"added": added, "removed": removed}
    return {"ok": True, "diff": diff}


def _resolve_job_build_root(job_url: str, cache_root: str) -> Optional[Path]:
    """Locate the latest cached build directory for ``job_url``.

    Args:
        job_url: Jenkins job URL used to derive the cache slug.
        cache_root: Base cache directory (frontend-provided or default).

    Returns:
        Path to the newest ``build_*`` directory for the job, or ``None``
        if no cached build exists. Mirrors the lookup used by
        ``aggregate_stats``.
    """
    if not job_url:
        return None
    try:
        from backend.user_context import get_current_user
    except Exception:
        return None

    base = _normalize_jenkins_cache_root(cache_root)
    slug = _job_slug(job_url)
    current_user = get_current_user()

    user_base = base / current_user if (base / current_user).exists() else None
    candidates: List[Path] = []
    if user_base:
        candidates.append(user_base / "jenkins" / slug)
    candidates.append(base / "jenkins" / slug)
    if user_base:
        candidates.extend(
            d / "jenkins" / slug
            for d in user_base.iterdir()
            if d.is_dir() and d.name != "jenkins"
        )
    else:
        candidates.extend(
            d / "jenkins" / slug
            for d in base.iterdir()
            if d.is_dir() and d.name not in ("jenkins", "exports")
        )

    job_root = next(
        (c for c in candidates if c.exists() and list(c.glob("build_*"))),
        None,
    )
    if not job_root:
        return None

    build_dirs = sorted(job_root.glob("build_*"), reverse=True)
    return build_dirs[0] if build_dirs else None


def _cache_trace_summary(matrix: Dict[str, Any], req: UdsTraceabilityMatrixRequest) -> None:
    """Persist a compact traceability summary for dashboard quick-load.

    Writes ``report/trace_matrix_summary.json`` under the latest cached
    build directory so the dashboard can render a coverage overview
    without re-running the matrix generation.
    """
    job_url = (req.job_url or "").strip()
    if not job_url:
        return

    cache_root = req.cache_root or ".devops_pro_cache"
    build_root = _resolve_job_build_root(job_url, cache_root)
    if build_root is None:
        return

    report_dir = build_root / "report"
    report_dir.mkdir(parents=True, exist_ok=True)

    # generate_uds_traceability_matrix() returns rows/summary/total_requirements at top-level.
    # Some older callers wrap it as {"matrix": {...}} — accept both.
    if isinstance(matrix, dict) and "rows" in matrix:
        inner = matrix
    elif isinstance(matrix, dict) and isinstance(matrix.get("matrix"), dict):
        inner = matrix["matrix"]
    else:
        inner = matrix if isinstance(matrix, dict) else {}

    rows = inner.get("rows") if isinstance(inner, dict) else None
    rows = rows if isinstance(rows, list) else []
    summary_data = inner.get("summary") if isinstance(inner, dict) else None
    summary_data = summary_data if isinstance(summary_data, dict) else {}
    declared_total = inner.get("total_requirements") if isinstance(inner, dict) else None

    # Classification aligned with the detail view (SrsSdsSection deriveStatus):
    #   covered   : design (SDS or UDS source) AND tests both present
    #   partial   : exactly one of {design, tests} present
    #   uncovered : neither present
    covered = 0
    partial = 0
    uncovered = 0
    for row in rows:
        if not isinstance(row, dict):
            uncovered += 1
            continue
        has_tests = bool(
            row.get("tests")
            or row.get("sts_tests")
            or row.get("suts_tests")
            or row.get("sits_tests")
            or row.get("syts_tests")   # 시스템 시험(결정1) — 비기능/안전 요구의 시스템 레벨 검증
            or row.get("syits_tests")
            or row.get("vcast_tests")
            or row.get("test_ids")
        )
        has_design = bool(
            row.get("source_ids")
            or row.get("sds_components")
            or row.get("sds_functions")  # 추적 정화: 함수로만 추적된 요구사항 커버리지 회귀 방지(프론트 DESIGN_FIELDS lockstep)
            or row.get("hsis_signals")   # 시스템 인터페이스(HSIS) realization — SwEI 등 인터페이스 요구 커버(결정1)
            or row.get("functions")
            or row.get("mapping")
            or row.get("sds")
            or row.get("source_mapping")
        )
        # 비기능/안전 요구(SwNTR/SwNTSR)는 설계 분해 없이 시험으로 직접 검증된다(ISO 26262 요구사항기반
        # 시험). 따라서 설계 링크가 없어도 시험만 있으면 covered로 인정(결정1, 3-site lockstep).
        # 행 requirement_id는 RAW 철자(정규화 전)라 SyNTR_/SyNTSR_도 매칭(_normalize_req_id가 키에만
        # Sy→Sw collapse). 비기능/안전 요구 prefix 4종 모두 인정.
        _rid = str(row.get("requirement_id") or "").upper()
        is_nonfunctional = _rid.startswith(("SWNTR", "SWNTSR", "SYNTR", "SYNTSR"))
        if has_tests and (has_design or is_nonfunctional):
            covered += 1
        elif has_design or has_tests:
            partial += 1
        else:
            uncovered += 1

    # Prefer matrix-declared total; fall back to row count
    total = int(declared_total) if isinstance(declared_total, int) and declared_total > 0 else len(rows)
    # Normalize: if declared_total > len(rows), unclassified extras are "uncovered"
    if total > covered + partial + uncovered:
        uncovered += total - (covered + partial + uncovered)
    coverage_pct = round(covered / total * 100, 1) if total > 0 else 0.0

    # ASIL 결합(P5) — 대시보드 quick-load가 ASIL 갭/미상을 알 수 있게 전파(reviewer WARN-C:
    # 미전파 시 매트릭스 재생성해 detail 탭 진입해야만 갭이 보임). link_table에서 끌어옴.
    link_table = inner.get("link_table") if isinstance(inner, dict) else None
    asil_cov = link_table.get("asil_coverage") if isinstance(link_table, dict) else None
    asil_cov = asil_cov if isinstance(asil_cov, dict) else {}

    # ID 정합성 감사(trace_integrity) — 대시보드 quick-load가 충돌/dangling/placeholder를
    # 매트릭스 재생성 없이 알 수 있게 카운트만 전파(ASIL 패턴과 동일). 데이터 없으면 0/clean.
    integ = inner.get("integrity") if isinstance(inner, dict) else None
    integ_stats = integ.get("stats") if isinstance(integ, dict) else None
    integ_stats = integ_stats if isinstance(integ_stats, dict) else {}

    cache_payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "total_requirements": total,
        "covered": covered,
        "partial": partial,
        "uncovered": uncovered,
        "coverage_pct": coverage_pct,
        "summary_raw": summary_data,
        # ASIL 추적성 결합(P5) — 대시보드 표시용. 데이터 없으면 has=False/0.
        "asil_has": bool(asil_cov.get("has_asil")),
        "asil_gap_count": len(asil_cov.get("gaps") or []),
        "asil_unknown_count": int(asil_cov.get("unknown_count") or 0),
        # ID 정합성 감사(trace_integrity) — 충돌/dangling/placeholder 카운트 + clean 플래그.
        "integrity_clean": bool(integ_stats.get("clean", True)),
        "integrity_collision_count": int(integ_stats.get("collision_count") or 0),
        "integrity_dangling_count": int(integ_stats.get("dangling_count") or 0),
        "integrity_placeholder_count": int(integ_stats.get("placeholder_count") or 0),
    }

    (report_dir / "trace_matrix_summary.json").write_text(
        json.dumps(cache_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _cache_link_table(
    link_table: Optional[Dict[str, Any]], req: UdsTraceabilityMatrixRequest
) -> None:
    """명시 RelatedID 링크 테이블(P1)을 ``report/trace_link_table.json``에 영속화.

    ``build_link_table()`` 결과는 결정적(같은 입력 → 동일)이라, 이 파일은 빌드마다
    재계산되는 휴리스틱 bridge와 달리 **감사 가능한 추적성 baseline**이 된다.
    ``_cache_trace_summary``와 동일한 빌드 디렉토리 해석/쓰기 패턴을 따른다(best-effort).
    """
    if not isinstance(link_table, dict):
        return
    job_url = (req.job_url or "").strip()
    if not job_url:
        return

    cache_root = req.cache_root or ".devops_pro_cache"
    build_root = _resolve_job_build_root(job_url, cache_root)
    if build_root is None:
        return

    report_dir = build_root / "report"
    report_dir.mkdir(parents=True, exist_ok=True)

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        **link_table,
    }
    (report_dir / "trace_link_table.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


@router.post("/api/jenkins/uds/traceability-matrix")
def jenkins_uds_traceability_matrix(req: UdsTraceabilityMatrixRequest) -> Dict[str, Any]:
    try:
        matrix = generate_uds_traceability_matrix(
            req.requirement_items or [],
            mapping_pairs=req.mapping_pairs or [],
            vcast_rows=req.vcast_rows or [],
            sds_pairs=req.sds_pairs or [],
            sits_rows=req.sits_rows or [],
            uds_function_ids=req.uds_function_ids or [],
            component_asil=req.component_asil or {},
            hsis_pairs=req.hsis_pairs or [],
        )
        # 명시 RelatedID 링크 테이블 파생(P1) — hiMA식 매트릭스/감사 baseline.
        # additive: 기존 matrix dict를 변형하지 않고 새 키(link_table)만 더한다.
        try:
            matrix["link_table"] = build_link_table(matrix)
        except Exception as lt_exc:
            _api_logger.debug("Link table derivation skipped: %s", lt_exc)
        # Cache compact summary for dashboard quick-load (best-effort)
        try:
            _cache_trace_summary(matrix, req)
        except Exception as cache_exc:
            _api_logger.debug("Trace summary cache skipped: %s", cache_exc)
        # 링크 테이블 영속화(P1) — 빌드마다 재계산 대신 감사 가능 baseline으로 고정(best-effort)
        try:
            _cache_link_table(matrix.get("link_table"), req)
        except Exception as lt_cache_exc:
            _api_logger.debug("Link table cache skipped: %s", lt_cache_exc)
        return {"ok": True, "matrix": matrix}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/api/jenkins/uds/traceability-matrix/export-xlsx")
def jenkins_uds_traceability_matrix_xlsx(body: Dict[str, Any]) -> Response:
    """추적성 매트릭스(클라이언트 보유 matrix)를 감사용 xlsx로 렌더해 반환.

    hiMA TrMatrixReport(화면 그대로 xlsx) 대응 — 감사자가 가장 자주 요구하는 형식.
    프론트가 이미 생성한 matrix(rows+link_table)를 body로 보내면 재추출 없이 포맷만 한다.
    body: {"matrix": {...}, "meta": {project_name, job_url, build_selector, ...}}.
    """
    from report_gen.trace_matrix_xlsx import build_trace_xlsx

    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="invalid body")
    matrix = body.get("matrix") if isinstance(body.get("matrix"), dict) else body
    meta = body.get("meta") if isinstance(body.get("meta"), dict) else {}
    meta = dict(meta)
    # 생성시각은 서버에서 주입(클라 신뢰 불요) — 헤더 블록 표시용.
    meta.setdefault("generated_at", datetime.now().isoformat(timespec="seconds"))
    try:
        data = build_trace_xlsx(matrix, meta)
    except ImportError as exc:
        raise HTTPException(status_code=500, detail=f"openpyxl 미설치: {exc}")
    except Exception as exc:
        _api_logger.debug("Trace xlsx export failed: %s", exc)
        raise HTTPException(status_code=500, detail=f"xlsx 생성 실패: {exc}")
    fname = f"traceability_matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/api/jenkins/uds/trace-summary")
def jenkins_trace_summary(req: dict) -> Dict[str, Any]:
    """Fast lookup of the cached traceability matrix summary for a single job.

    Args:
        req: Dict with ``job_url`` (required) and optional ``cache_root``
            (defaults to ``.devops_pro_cache``).

    Returns:
        Cached summary payload with ``has_data=True`` when available,
        otherwise ``{"has_data": False, "reason": ...}``.
    """
    job_url = str(req.get("job_url", "")).strip()
    cache_root = str(req.get("cache_root", ".devops_pro_cache") or ".devops_pro_cache")

    if not job_url:
        return {"has_data": False, "reason": "job_url required"}

    build_root = _resolve_job_build_root(job_url, cache_root)
    if build_root is None:
        return {"has_data": False, "reason": "no cached build"}

    summary_path = build_root / "report" / "trace_matrix_summary.json"
    if not summary_path.exists():
        return {
            "has_data": False,
            "reason": "no cached summary — SRS & SDS 섹션에서 추적성 매트릭스를 먼저 생성하세요",
        }

    try:
        data = json.loads(summary_path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {"has_data": False, "reason": "invalid cache payload"}
        data["has_data"] = True
        return data
    except Exception as exc:
        return {"has_data": False, "reason": f"cache read failed: {exc}"}


# UDS 매핑 추출은 36MB read + 73MB document.xml 파싱(손상 docx fallback 시)이라 무겁다.
# 매트릭스 재로드/연타 시 같은 파싱이 쌓여 worker/CPU 경합 → 타임아웃을 유발하므로
# uds_path 기준 TTL 캐시로 재추출을 막는다. 파일은 세션 중 거의 안 바뀌므로 30분 TTL.
_UDS_MAPPING_CACHE: Dict[str, "tuple[float, Dict[str, Any]]"] = {}
_UDS_MAPPING_LOCK = threading.Lock()
_UDS_MAPPING_TTL = 1800.0


def _docx_tables_text(data: bytes) -> Optional[List[List[List[str]]]]:
    """docx bytes → tables[행[셀텍스트]]. 손상 docx 복구 fallback 포함.

    정상 파일은 python-docx로 읽는다. 임베디드 이미지 CRC 오류 등으로 python-docx가
    실패해도 추적성 매핑은 이미지와 무관한 '표'에서만 추출하므로, word/document.xml만
    직접 스트리밍 파싱해 표를 복구한다(손상 미디어 파트 우회). document.xml까지 손상돼
    파싱 불가하면 None.
    """
    import io as _io
    # 1) 정상 경로 — python-docx
    try:
        import docx as _docx
        doc = _docx.Document(_io.BytesIO(data))
        return [[[c.text for c in r.cells] for r in t.rows] for t in doc.tables]
    except Exception:
        pass
    # 2) 손상 fallback — document.xml만 직접 파싱 (이미지 등 손상 파트 우회).
    #    document.xml은 수십 MB일 수 있어 iterparse + elem.clear()로 메모리 방어.
    try:
        import xml.etree.ElementTree as _ET
        import zipfile as _zip
        W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
        tables: List[List[List[str]]] = []
        with _zip.ZipFile(_io.BytesIO(data)) as zf:
            with zf.open("word/document.xml") as f:
                for _event, elem in _ET.iterparse(f, events=("end",)):
                    if elem.tag != W + "tbl":
                        continue
                    rows: List[List[str]] = []
                    for tr in elem.findall(W + "tr"):
                        rows.append([
                            "".join(t.text or "" for t in tc.iter(W + "t"))
                            for tc in tr.findall(W + "tc")
                        ])
                    tables.append(rows)
                    elem.clear()
        return tables
    except Exception:
        return None


@router.post("/api/jenkins/uds/extract-mapping")
def jenkins_uds_extract_mapping(body: Dict[str, Any]) -> Dict[str, Any]:
    """UDS 문서에서 함수↔요구사항 매핑을 추출"""
    import re as _re

    from backend.services.file_resolver import get_resolver
    from backend.services.resolver_helpers import enforce_resolver_access
    uds_path = str(body.get("uds_path", "")).strip()
    if not uds_path:
        raise HTTPException(status_code=400, detail="uds_path required")
    enforce_resolver_access(uds_path)  # C3: health.py와 일관된 방어심층
    # TTL 캐시 — 손상 docx fallback(73MB 파싱) 재로드 pile-up 방지.
    _ck = uds_path.replace("\\", "/").rstrip("/").lower()
    _now = time.time()
    with _UDS_MAPPING_LOCK:
        _hit = _UDS_MAPPING_CACHE.get(_ck)
    if _hit and (_now - _hit[0]) < _UDS_MAPPING_TTL:
        return dict(_hit[1])
    resolver = get_resolver()
    try:
        if not resolver.exists(uds_path):
            raise HTTPException(status_code=400, detail=f"UDS 파일을 찾을 수 없습니다: {uds_path}")
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")

    try:
        data = resolver.read_bytes(uds_path)
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")
    # 손상된 docx(임베디드 이미지 CRC 오류 등)도 매핑은 표에서만 추출하므로 python-docx
    # 실패 시 document.xml 직접 파싱으로 복구. document.xml까지 손상되면 None.
    tables_text = _docx_tables_text(data)
    if tables_text is None:
        raise HTTPException(
            status_code=500,
            detail="UDS 파싱 실패: 문서를 열 수 없습니다(파일 손상 가능). 원본을 다시 저장 후 교체하세요.",
        )

    # Extract func_id → func_name → requirement_ids from Function Information tables
    req_to_sources: Dict[str, set] = {}
    # 전체 UDS 함수 인벤토리(함수명 + SwUFn ID) — 설계 req(SwSTR 등) 참조 유무와 무관하게
    # 모은다. UDS 함수표의 대부분(~95%)은 자기 SwUFn ID·Name만 있고 설계 req 참조가 없어
    # req_to_sources(=mapping_pairs)에서 누락되는데, 매트릭스의 SDS→UDS bridge는 함수명만
    # 매칭하면 되므로 전체 목록을 별도로 전달해 "UDS 함수" 컬럼이 채워지게 한다.
    all_funcs: set = set()
    for rows in tables_text:
        if len(rows) < 4:
            continue
        first_cell = (rows[0][0] if rows[0] else "").strip()
        if "Function Information" not in first_cell:
            continue

        func_id = ""
        func_name = ""
        req_refs = []
        for cells in rows:
            cells = [str(c).strip() for c in cells]
            label = cells[0] if cells else ""
            # python-docx는 병합셀을 grid로 펼쳐 값이 cells[2]에 오지만, document.xml
            # 직접 파싱(손상 docx fallback)은 실제 w:tc만 추출해 값이 cells[1]에 온다.
            # 두 경우 모두 커버: 3+셀이면 [2], 2셀이면 마지막 셀.
            if len(cells) > 2:
                value = cells[2]
            elif len(cells) > 1:
                value = cells[1]
            else:
                value = ""
            if label == "ID":
                func_id = value
            elif label == "Name":
                func_name = value
            # Find requirement IDs in all cells
            for c in cells:
                req_refs.extend(_re.findall(r"Sw[A-Z]{2,}_\d+", c))

        if func_name:
            all_funcs.add(func_name)
            if func_id:
                all_funcs.add(func_id)
            req_refs = sorted(set(req_refs) - {func_id})
            for rid in req_refs:
                if rid not in req_to_sources:
                    req_to_sources[rid] = set()
                req_to_sources[rid].add(func_name)
                # rank1 fix: func_id(SwUFn_NNNN)도 source로 등록한다. VectorCAST 원본
                # 리포트는 testcase를 SwUFn ID로 식별하므로(예 SwUFn_0133.001),
                # func_name(예 'main')만으로는 join이 0건이 된다. SwUFn ID를 함께
                # 노출해 vcast subprogram(SwUFn_NNNN 정규화)과 매칭되게 한다.
                if func_id:
                    req_to_sources[rid].add(func_id)

    # Convert to mapping_pairs format expected by traceability-matrix API
    mapping_pairs = []
    for rid, sources in sorted(req_to_sources.items()):
        mapping_pairs.append({
            "requirement_id": rid,
            "source_ids": sorted(sources),
        })

    result = {
        "ok": True,
        "mapping_pairs": mapping_pairs,
        "total_requirements": len(mapping_pairs),
        "total_functions": len({fn for fns in req_to_sources.values() for fn in fns}),
        # 전체 UDS 함수 인벤토리(매트릭스 uds_all_funcs 시드용) — 설계 req 참조 없는
        # 함수까지 포함. 매트릭스 SDS→UDS bridge가 이 목록으로 전체 함수를 매칭한다.
        "all_function_ids": sorted(all_funcs),
        "all_functions_count": len(all_funcs),
    }
    with _UDS_MAPPING_LOCK:
        if len(_UDS_MAPPING_CACHE) >= 16:
            _UDS_MAPPING_CACHE.clear()
        _UDS_MAPPING_CACHE[_ck] = (_now, result)
    return dict(result)


def _normalize_req_id(rid: str) -> str:
    """Normalize system-level (Sy*) requirement IDs to software-level (Sw*)
    and ensure consistent casing for matching with report_gen.requirements.

    Mapping: SyTR → SwTR, SyEIF → SwEI, SyTSR → SwTSR, SyNTR → SwNTR
    """
    import re as _re
    rid = "".join(rid.split())  # remove whitespace
    if rid.upper().startswith("SY"):
        rid = _re.sub(r"(?i)^SyEIF_", "SwEI_", rid)
        rid = _re.sub(r"(?i)^SyTR_", "SwTR_", rid)
        rid = _re.sub(r"(?i)^SyTSR_", "SwTSR_", rid)
        rid = _re.sub(r"(?i)^SyNTR_", "SwNTR_", rid)
        rid = _re.sub(r"(?i)^SyNTSR_", "SwNTSR_", rid)
        rid = _re.sub(r"(?i)^SyCNF_", "SwCNF_", rid)
    return rid.upper() if rid else rid


@router.post("/api/jenkins/sts/extract-traceability")
def jenkins_sts_extract_traceability(body: Dict[str, Any]) -> Dict[str, Any]:
    """STS/SUTS Excel에서 Traceability 시트의 요구사항↔TC 매핑 추출"""
    import io

    from backend.services.file_resolver import get_resolver
    from backend.services.resolver_helpers import enforce_resolver_access
    file_path = str(body.get("path", "")).strip()
    doc_type = str(body.get("doc_type", "")).strip().lower()  # "sts" or "suts"
    if not file_path:
        raise HTTPException(status_code=400, detail="path required")
    enforce_resolver_access(file_path)  # C3: 명시 endpoint-local 검증
    resolver = get_resolver()
    try:
        if not resolver.exists(file_path):
            raise HTTPException(status_code=400, detail=f"파일을 찾을 수 없습니다: {file_path}")
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")
    p = Path(file_path).expanduser()

    try:
        import openpyxl
        data = resolver.read_bytes(file_path)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Excel 읽기 실패: {exc}")

    # Find traceability sheet — 우선순위:
    # 1) body의 sheet_name이 명시되면 그 시트 사용 (외부 도구 생성 파일 대응)
    # 2) "traceability" / "trace" / "tc" / "test case" / "test spec" / "사양" / "트레이스" 등 자동 탐색
    # 3) 미발견 시 available_sheets와 함께 안내 — frontend가 사용자에게 시트 선택 노출 가능
    trace_ws = None
    trace_type = None
    sheet_name_arg = str(body.get("sheet_name", "")).strip()
    if sheet_name_arg and sheet_name_arg in wb.sheetnames:
        trace_ws = wb[sheet_name_arg]
        trace_type = "matrix" if "swrs" in sheet_name_arg.lower() else "list"
    else:
        # N23: keyword 좁힘 — "사양" 단독은 "기능 사양" 등 false positive 위험.
        # "테스트 사양" / "test 사양" 처럼 test 컨텍스트 결합한 패턴만 매칭.
        _trace_keywords = ("traceability", "trace", "test case", "testcase", "test spec",
                            "테스트 사양", "테스트사양", "트레이스")
        for name in wb.sheetnames:
            nl = name.lower()
            if any(kw in nl for kw in _trace_keywords) or nl.strip() == "tc":
                trace_ws = wb[name]
                trace_type = "matrix" if "swrs" in nl else "list"
                break

    if not trace_ws:
        all_sheets = list(wb.sheetnames)
        wb.close()
        return {
            "ok": False,
            "error": "Traceability 시트를 찾을 수 없습니다. sheet_name 인자로 명시하세요.",
            "available_sheets": all_sheets,
            "vcast_rows": [],
        }

    vcast_rows = []
    import re as _re

    # 헤더 기반 컬럼 탐지 (list 형식 전용) — KJPDS02 SwTS/SwUTS "3.SW Test Spec"처럼
    # TC ID/요구사항 컬럼이 고정 위치가 아니라 'Test Case ID' / 'SRS' 헤더로만 식별되는
    # 파일 대응. 헤더 행을 스캔해 TC 컬럼 + 요구사항(SRS/Related/SwRS) 컬럼을 찾는다.
    # 못 찾으면 기존 고정 컬럼(TC=5/req=6) 로직으로 fallback (하위호환).
    def _detect_header_cols(ws):
        # 상위 30행까지 스캔 (SwUTS는 preamble이 길어 헤더가 12행 이후). 먼저 그 행에
        # TC ID 컬럼이 있는지 보고, 있으면 같은 행에서 요구사항/Related 컬럼을 찾는다
        # (요구사항 컬럼을 TC 헤더 행으로 한정 → preamble/데이터 오탐 방지).
        max_c = min((ws.max_column or 1), 200)
        max_r = min((ws.max_row or 1), 30)
        for hr in range(1, max_r + 1):
            tc_col = None
            for c in range(1, max_c + 1):
                h = str(ws.cell(hr, c).value or "").strip().lower()
                if not h:
                    continue
                # TC ID 컬럼: 'Test Case ID'(SwTS) / 'TC_ID'(SwUTS) 등 표기 변형 흡수.
                # 'Test Case Generation Method'는 'testcaseid' 미포함 → 오탐 방지.
                _hn = h.replace(" ", "").replace("_", "")
                if "testcaseid" in _hn or _hn == "tcid":
                    tc_col = c
                    break
            if tc_col is None:
                continue
            # 멀티행 헤더 대응: 'Related ID'가 TC 상세헤더(예: SwUTS row4)보다 위 행
            # (병합 타이틀, row3)에 있을 수 있어 TC 행 포함 위로 3행 band를 스캔한다.
            req_id_cols = []
            for rr in range(max(1, hr - 3), hr + 1):
                for c in range(1, max_c + 1):
                    if c in req_id_cols:
                        continue
                    h = str(ws.cell(rr, c).value or "").strip().lower()
                    if not h:
                        continue
                    # 요구사항/추적 링크 컬럼만 매칭. 'srs'/'swrs'는 specific해서 substring
                    # 허용, 'related'/'requirement'는 'related functionality'/'requirement
                    # category' 같은 오탐 방지 위해 완전일치(공백 제거)로 한정. FS_REQ 제외.
                    hn = h.replace(" ", "")
                    if ("srs" in hn or "swrs" in hn
                            or hn in ("related", "relatedid", "relatedids",
                                      "relatedrequirement", "relatedreq", "requirement",
                                      "requirementid", "reqid", "trace", "traceid")):
                        req_id_cols.append(c)
            if req_id_cols:
                # unit(함수명) 컬럼 — SUTS/SITS를 SDS 함수명 bridge로 SRS에 연결하기
                # 위해 캡처 (SwUTS의 col4 'Unit' = 함수명). TC 헤더 행에서 완전일치 탐색.
                unit_col = None
                for c in range(1, max_c + 1):
                    h = str(ws.cell(hr, c).value or "").strip().lower()
                    if h in ("unit", "function", "function name",
                             "unit name", "function_name", "unit_name"):
                        unit_col = c
                        break
                return hr, tc_col, req_id_cols, unit_col
        return None

    # Determine source label from doc_type or auto-detect from sheet structure
    if trace_type == "matrix":
        source_label = doc_type.upper() if doc_type in ("sts", "suts") else "STS"
        # STS format: row 4 has req IDs as column headers, rows 5+ have TC IDs with markers
        req_cols = []
        for c in range(3, (trace_ws.max_column or 0) + 1):
            v = trace_ws.cell(4, c).value
            if v and ("Sw" in str(v) or "SW" in str(v).upper() or "Sy" in str(v)):
                req_cols.append((c, _normalize_req_id(str(v).strip())))

        for r in range(5, (trace_ws.max_row or 0) + 1):
            tc_id = str(trace_ws.cell(r, 3).value or "").strip()
            if not tc_id:
                continue
            for col, rid in req_cols:
                val = trace_ws.cell(r, col).value
                if val is not None and str(val).strip():
                    vcast_rows.append({
                        "requirement_id": rid,
                        "testcase": tc_id,
                        "source": source_label,
                        "result": "mapped",
                    })
    else:
        source_label = doc_type.upper() if doc_type in ("sts", "suts") else "SUTS"
        detected = _detect_header_cols(trace_ws)
        if detected:
            # 헤더 기반: TC ID 컬럼 ↔ 요구사항 컬럼. 병합셀/연속행은 직전 TC 유지,
            # 빈 행 50연속 시 조기 종료. 요구사항은 req 컬럼에서만 regex 추출.
            header_row, tc_col, req_id_cols, unit_col = detected
            empty_streak = 0
            current_tc = ""
            current_unit = ""
            for r in range(header_row + 1, (trace_ws.max_row or header_row) + 1):
                tc_v = str(trace_ws.cell(r, tc_col).value or "").strip()
                if tc_v:
                    current_tc = tc_v
                    current_unit = ""  # 새 TC 블록 시작 → unit 초기화
                if unit_col:
                    uv = str(trace_ws.cell(r, unit_col).value or "").strip()
                    if uv:
                        current_unit = uv
                found = []
                for rc in req_id_cols:
                    cv = str(trace_ws.cell(r, rc).value or "").strip()
                    if cv:
                        found += _re.findall(r"Sw[A-Za-z]{2,}_\d+|Sy[A-Za-z]{2,}_\d+", cv)
                if not tc_v and not found:
                    empty_streak += 1
                    if empty_streak >= 50:
                        break
                    continue
                empty_streak = 0
                if current_tc and found:
                    for rid in found:
                        vcast_rows.append({
                            "requirement_id": _normalize_req_id(rid),
                            "testcase": current_tc,
                            "unit": current_unit,
                            "source": source_label,
                            "result": "mapped",
                        })
        else:
            # Fallback: 기존 고정 컬럼 (TC=5, SRS req=6, func=4)
            for r in range(4, trace_ws.max_row + 1):
                tc_id = str(trace_ws.cell(r, 5).value or "").strip()
                req_raw = str(trace_ws.cell(r, 6).value or "").strip()
                func_name = str(trace_ws.cell(r, 4).value or "").strip()
                if not tc_id:
                    continue
                req_ids = _re.findall(r"Sw[A-Z]{2,}_\d+|Sy[A-Z]{2,}_\d+", req_raw)
                for rid in req_ids:
                    vcast_rows.append({
                        "requirement_id": _normalize_req_id(rid),
                        "testcase": tc_id,
                        "unit": func_name,
                        "source": source_label,
                        "result": "mapped",
                    })

    wb.close()

    # Summarize
    req_set = set(r["requirement_id"] for r in vcast_rows)
    return {
        "ok": True,
        "vcast_rows": vcast_rows,
        "total_mappings": len(vcast_rows),
        "requirements_covered": len(req_set),
    }


@router.post("/api/jenkins/sds/extract-mapping")
def jenkins_sds_extract_mapping(body: Dict[str, Any]) -> Dict[str, Any]:
    """SDS 문서에서 SwCom↔요구사항 매핑 추출 (추적성 매트릭스용)"""
    import re as _re
    import tempfile

    from backend.services.file_resolver import get_resolver
    from backend.services.resolver_helpers import enforce_resolver_access
    from report_gen.requirements import _extract_sds_partition_map, _normalize_req_id

    sds_path = str(body.get("sds_path", "")).strip()
    if not sds_path:
        raise HTTPException(status_code=400, detail="sds_path required")
    enforce_resolver_access(sds_path)  # C3
    resolver = get_resolver()
    try:
        if not resolver.exists(sds_path):
            raise HTTPException(status_code=400, detail=f"파일을 찾을 수 없습니다: {sds_path}")
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")

    # _extract_sds_partition_map은 file path를 받음 — cloudium에서 backend가 직접 open하면
    # 권한 없음. 임시 파일에 IPC bytes를 저장 후 helper가 그 임시 path를 사용.
    try:
        data = resolver.read_bytes(sds_path)
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        tmp.write(data)
        tmp_path = tmp.name
    try:
        partition_map = _extract_sds_partition_map(tmp_path)
    finally:
        try:
            Path(tmp_path).unlink()
        except OSError:
            pass
    if not partition_map:
        return {"ok": True, "sds_pairs": [], "total_components": 0, "total_requirements": 0}

    # Build requirement_id → [component_names] mapping
    # W3: Apply _normalize_req_id to handle whitespace/case in related IDs
    req_to_comps: Dict[str, list] = {}
    comp_set = set()
    # ASIL 결합(P5) — partition_map은 컴포넌트/함수별 ASIL을 이미 보유(_extract_sds_partition_map
    # 가 SwCom ASIL 컬럼·모듈 ASIL 헤더에서 채움). 매트릭스가 요구사항별 ASIL을 도출할 수
    # 있도록 {컴포넌트명(lower): ASIL} 맵을 그대로 노출(additive). related 없는 asil-only
    # 엔트리(SwCom 정의 행 등)도 포함 — req의 component_id가 그 행을 가리킬 수 있음.
    component_asil: Dict[str, str] = {}
    for comp_key, info in partition_map.items():
        casil = str(info.get("asil") or "").strip()
        if casil:
            component_asil[str(comp_key).strip().lower()] = casil
    # 함수/컴포넌트 분리(추적 정화): design_component_ids = 실 SwCom/모듈만(kind!='function').
    # SDS 밴드 집계가 인터페이스 함수 fan-out으로 24배 부풀려지던 것 차단. 함수는 component_ids에
    # 그대로 남아 SUTS/VCAST 브리지(sds_func_to_reqs)에는 영향 없음.
    req_to_design_comps: Dict[str, list] = {}
    design_comp_set = set()
    for comp_key, info in partition_map.items():
        related = info.get("related", "")
        if not related:
            continue
        # Match IDs allowing optional internal whitespace (e.g., "SwRS_ 001")
        raw_ids = _re.findall(r"Sw[A-Za-z]{2,}\s*_\s*\d+|Sy[A-Za-z]{2,}\s*_\s*\d+", related)
        req_ids = [_normalize_req_id(rid) for rid in raw_ids]
        if req_ids:
            comp_name = comp_key
            is_function = info.get("kind") == "function"
            comp_set.add(comp_name)
            if not is_function:
                design_comp_set.add(comp_name)
            for rid in req_ids:
                req_to_comps.setdefault(rid, [])
                if comp_name not in req_to_comps[rid]:
                    req_to_comps[rid].append(comp_name)
                if not is_function:
                    req_to_design_comps.setdefault(rid, [])
                    if comp_name not in req_to_design_comps[rid]:
                        req_to_design_comps[rid].append(comp_name)

    sds_pairs = [
        {
            "requirement_id": rid,
            "component_ids": comps,  # 컴포넌트+함수 전체(브리지 호환)
            "design_component_ids": req_to_design_comps.get(rid, []),  # 실 SwCom/모듈만(SDS 밴드용)
        }
        for rid, comps in sorted(req_to_comps.items())
    ]

    return {
        "ok": True,
        "sds_pairs": sds_pairs,
        "total_components": len(design_comp_set),  # 실 설계 컴포넌트 수(함수 fan-out 제외)
        "total_components_with_functions": len(comp_set),  # 참고: 함수 포함 전체
        "total_requirements": len(sds_pairs),
        # ASIL 결합(P5) — 컴포넌트/함수별 ASIL 맵. 매트릭스가 요구사항별 ASIL 도출에 사용.
        "component_asil": component_asil,
    }


@router.post("/api/jenkins/hsis/extract-mapping")
def jenkins_hsis_extract_mapping(body: Dict[str, Any]) -> Dict[str, Any]:
    """HSIS(HW-SW 인터페이스) xlsx에서 요구사항↔인터페이스 신호 매핑 추출.

    추적성 매트릭스의 인터페이스 밴드(design-arm)용. HSIS Related ID 컬럼의 시스템 요구
    ID(SyTR/SyTSR/SyEI…)를 `_normalize_req_id`로 SW namespace에 평탄화 → 요구사항별
    인터페이스 신호(HSI_xx / SW변수) 그룹. 인터페이스 요구(SwEI) 커버를 위해 SyEI→SwEI도
    로컬 보강(이 사내 템플릿은 시스템/SW 인터페이스 번호가 1:1 병행).
    """
    import re as _re
    import tempfile

    from backend.services.file_resolver import get_resolver
    from backend.services.resolver_helpers import enforce_resolver_access
    from generators.sts import parse_hsis_signals

    hsis_path = str(body.get("hsis_path", "")).strip()
    if not hsis_path:
        raise HTTPException(status_code=400, detail="hsis_path required")
    enforce_resolver_access(hsis_path)  # C3
    resolver = get_resolver()
    try:
        if not resolver.exists(hsis_path):
            raise HTTPException(status_code=400, detail=f"파일을 찾을 수 없습니다: {hsis_path}")
        data = resolver.read_bytes(hsis_path)
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")

    # parse_hsis_signals은 file path를 받음 — cloudium IPC bytes를 임시파일로 떨군 뒤 파싱(SDS 패턴).
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(data)
        tmp_path = tmp.name
    try:
        parsed = parse_hsis_signals(tmp_path)
    finally:
        try:
            Path(tmp_path).unlink()
        except OSError:
            pass

    signals = parsed.get("signals") or []

    def _norm_hsis_req(rid: str) -> str:
        n = _normalize_req_id(rid)  # SyTR→SwTR, SyTSR→SwTSR, SyEIF→SwEI 등 + 대문자
        n = _re.sub(r"^SYEI_", "SWEI_", n)  # HSIS는 SyEI(F 없음) 사용 → 인터페이스 요구(SwEI)에 연결
        return n

    req_to_sigs: Dict[str, list] = {}
    for s in signals:
        related = str(s.get("related_id") or "")
        raw_ids = _re.findall(r"Sw[A-Za-z]{2,}\s*_\s*\d+|Sy[A-Za-z]{2,}\s*_\s*\d+", related)
        if not raw_ids:
            continue
        sig_label = (str(s.get("id") or "").strip()
                     or str(s.get("sw_var_name") or "").strip()
                     or str(s.get("signal_name") or "").strip())
        if not sig_label:
            continue
        for rid in raw_ids:
            nrid = _norm_hsis_req(rid)
            if not nrid:
                continue
            req_to_sigs.setdefault(nrid, [])
            if sig_label not in req_to_sigs[nrid]:
                req_to_sigs[nrid].append(sig_label)

    hsis_pairs = [
        {"requirement_id": rid, "hsis_signals": sigs}
        for rid, sigs in sorted(req_to_sigs.items())
    ]
    return {
        "ok": True,
        "hsis_pairs": hsis_pairs,
        "total_signals": len(signals),
        "total_requirements": len(hsis_pairs),
    }


@router.post("/api/jenkins/sits/extract-traceability")
def jenkins_sits_extract_traceability(body: Dict[str, Any]) -> Dict[str, Any]:
    """SITS Excel에서 TC ID↔요구사항 매핑 추출"""
    import io
    import re as _re

    from backend.services.file_resolver import get_resolver
    from backend.services.resolver_helpers import enforce_resolver_access

    file_path = str(body.get("path", "")).strip()
    if not file_path:
        raise HTTPException(status_code=400, detail="path required")
    enforce_resolver_access(file_path)  # C3: cloudium 게이트 + 화이트리스트 검증
    resolver = get_resolver()
    try:
        if not resolver.exists(file_path):
            raise HTTPException(status_code=400, detail=f"파일을 찾을 수 없습니다: {file_path}")
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")

    try:
        import openpyxl
        data = resolver.read_bytes(file_path)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except (PermissionError, OSError) as exc:
        raise HTTPException(status_code=403, detail=f"파일 접근 거부: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Excel 읽기 실패: {exc}")

    vcast_rows = []
    _MAX_EMPTY_ROWS = 50  # C3: break after N consecutive empty rows
    # source 라벨 — 시스템 시험(SyTS/SyITS)이 동일 구조라 라벨만 갈아끼워 재사용(syts/syits 위임 엔드포인트).
    src_label = (str(body.get("source_label") or "SITS").strip() or "SITS")
    # 매핑 0건일 때 진단용 — 실제로 스캔한 시트명(silent-empty 표면화, W-SITS-fix#4).
    scanned_sheet = ""

    # Strategy 1: Look for Traceability sheet — sheet_name 명시 + 자동 탐색 keyword 확장
    # (외부 도구 생성 SITS 파일 대응 — N20 follow-up)
    trace_ws = None
    sheet_name_arg = str(body.get("sheet_name", "")).strip()
    if sheet_name_arg and sheet_name_arg in wb.sheetnames:
        trace_ws = wb[sheet_name_arg]
    else:
        # N23: keyword 좁힘 — "사양" 단독은 false positive 위험. test 컨텍스트 결합만 매칭.
        _trace_keywords = ("traceability", "trace", "test case", "testcase", "test spec",
                            "테스트 사양", "테스트사양", "트레이스")
        for name in wb.sheetnames:
            nl = name.lower()
            if any(kw in nl for kw in _trace_keywords) or nl.strip() == "tc":
                trace_ws = wb[name]
                break

    if trace_ws:
        scanned_sheet = trace_ws.title
        empty_streak = 0
        # PERF: read_only 모드에서 ws.cell(r,c) 랜덤 접근은 호출마다 시트 상단부터
        # 재파싱돼 O(행²·열)로 폭주한다(1874행×146열 실측 ~75분). 순차 iter_rows
        # 단일 패스는 O(행)이며 동일 파일에서 0.4초로 동작한다. 열은 4~199만 본다.
        # max_column은 read_only 모드에서 None일 수 있음 → 그 경우 199(상한)까지
        # 스캔해 cols 5+ req ID 누락 방지. 좁은 시트는 아래 len(row) 가드가 처리.
        max_col = min(trace_ws.max_column or 199, 199)
        for row in trace_ws.iter_rows(min_row=4, max_col=max_col, values_only=True):
            # 1-based 열 2/3 == 0-based 인덱스 1/2
            tc_id = str(row[1] or "").strip() if len(row) > 1 else ""
            if not tc_id or not _re.match(r"Sw\w+_\d+", tc_id, _re.I):
                tc_id = str(row[2] or "").strip() if len(row) > 2 else ""
            if not tc_id:
                empty_streak += 1
                if empty_streak >= _MAX_EMPTY_ROWS:
                    break
                continue
            empty_streak = 0
            # 1-based 열 4부터 == 0-based 인덱스 3부터
            for ci in range(3, len(row)):
                val = str(row[ci] or "").strip()
                req_ids = _re.findall(r"Sw[A-Za-z]{2,}_\d+|Sy[A-Za-z]{2,}_\d+", val)
                for rid in req_ids:
                    vcast_rows.append({
                        "requirement_id": _normalize_req_id(rid),
                        "testcase": tc_id,
                        "source": src_label,
                        "result": "mapped",
                    })
    else:
        # Strategy 2: Parse main test spec sheet — TC ID in col B, Related ID column
        spec_ws = None
        for name in wb.sheetnames:
            if "Integration Test" in name or "Test Spec" in name:
                spec_ws = wb[name]
                break
        # Fallback: look for sheet with most rows (likely the test spec)
        if not spec_ws:
            for name in wb.sheetnames:
                if "SW Integration" in name and "Strategy" not in name:
                    spec_ws = wb[name]
                    break

        # W2: Don't blindly use first sheet — return warning instead
        if not spec_ws:
            wb.close()
            return {
                "ok": True,
                "vcast_rows": [],
                "total_mappings": 0,
                "requirements_covered": 0,
                "warning": "SITS Traceability 또는 Integration Test 시트를 찾을 수 없습니다. sheet_name 인자로 명시하세요.",
                "available_sheets": list(wb.sheetnames),
            }
        scanned_sheet = spec_ws.title

        # Find the Related ID column by scanning header rows 5/6.
        # PERF: 본문 random .cell() 스캔도 Strategy 1과 동일한 O(행²) 폭주를
        # 유발하므로 header(2행)·본문 모두 iter_rows 순차 패스로 처리한다.
        related_col = -1
        hdr_scan_max = min(spec_ws.max_column or 199, 199)
        hdr_rows = list(spec_ws.iter_rows(
            min_row=5, max_row=6, max_col=hdr_scan_max, values_only=True))
        hdr5 = hdr_rows[0] if len(hdr_rows) > 0 else ()
        hdr6 = hdr_rows[1] if len(hdr_rows) > 1 else ()
        for ci in range(len(hdr5)):
            h5 = str(hdr5[ci] or "").strip().lower()
            h6 = str(hdr6[ci] or "").strip().lower() if ci < len(hdr6) else ""
            if "related" in h5 or "related" in h6 or "swds" in h6:
                related_col = ci + 1  # 0-based -> 1-based
                break
        if related_col < 0:
            related_col = 145  # default SITS layout

        empty_streak = 0
        # related_col(기본 145)까지 포함하도록 max_col 보장 (단 199 상한)
        body_max_col = min(max(spec_ws.max_column or 199, related_col), 199)
        for row in spec_ws.iter_rows(
                min_row=7, max_col=body_max_col, values_only=True):
            tc_id = str(row[1] or "").strip() if len(row) > 1 else ""
            if not tc_id:
                empty_streak += 1
                if empty_streak >= _MAX_EMPTY_ROWS:
                    break
                continue
            empty_streak = 0

            # Extract entry function name from description ("Verify integration: FuncName → ...")
            desc = str(row[2] or "").strip() if len(row) > 2 else ""
            entry_fn = ""
            if "integration:" in desc.lower():
                parts = desc.split(":", 1)
                if len(parts) > 1:
                    fn_part = parts[1].strip().split("→")[0].split("->")[0].strip()
                    if fn_part and not fn_part.startswith("("):
                        entry_fn = fn_part

            related_val = str(row[related_col - 1] or "").strip() if len(row) >= related_col else ""
            req_ids = _re.findall(r"Sw[A-Za-z]{2,}_\d+|Sy[A-Za-z]{2,}_\d+", related_val)
            for rid in req_ids:
                vcast_rows.append({
                    "requirement_id": _normalize_req_id(rid),
                    "testcase": tc_id,
                    "unit": entry_fn,
                    "source": src_label,
                    "result": "mapped",
                })
            # If no Sw* req IDs found but we have entry_fn, still add row
            # so function-based reverse mapping can work
            if not req_ids and entry_fn:
                vcast_rows.append({
                    "requirement_id": "",
                    "testcase": tc_id,
                    "unit": entry_fn,
                    "source": src_label,
                    "result": "mapped",
                })

    wb.close()

    # Dedup
    seen = set()
    deduped = []
    for row in vcast_rows:
        key = (row["requirement_id"], row["testcase"], row.get("unit", ""))
        if key not in seen:
            seen.add(key)
            deduped.append(row)

    # rid=""(testcase-only 행, W-SITS-fix#1)는 커버된 요구사항이 아니므로 covered에서 제외
    # (허위 커버리지 방지 — deep-review I1). total_mappings는 전체(2-hop 대기 행 포함).
    req_set = set(r["requirement_id"] for r in deduped if r["requirement_id"])
    direct_mapped = len(req_set)
    result: Dict[str, Any] = {
        "ok": True,
        "vcast_rows": deduped,
        "total_mappings": len(deduped),
        "requirements_covered": direct_mapped,
        "direct_mapped": direct_mapped,
    }
    # silent-empty 표면화(W-SITS-fix#4): 시트는 인식했으나 직접 요구 매핑 0건이면 warning+
    # available_sheets를 실어 프론트(SrsSdsSection)가 '요구열 없는 Test-Log 포맷' 등 원인을
    # 표기하게 한다. 과거엔 Strategy1이 시트를 찾고도 0행이면 아무 신호 없이 빈 배열만 반환해
    # SITS 밴드가 조용히 비었다(Strategy2의 '시트 없음' 경로만 warning을 실었음 — 비대칭 해소).
    if not deduped:
        result["warning"] = (
            f"{src_label} 시트('{scanned_sheet}')를 인식했으나 TC↔요구사항 매핑을 0건 추출했습니다. "
            "요구(Related) 열이 없는 Test-Log 포맷이거나 시트 레이아웃이 예상과 다를 수 있습니다."
        )
        result["available_sheets"] = list(wb.sheetnames)
    elif direct_mapped == 0:
        # 전부 testcase-only(직접 요구열 없는 Test-Log) — 직접 매핑 0, 매트릭스 2-hop 브리지
        # (TC의 SwUFn→SUTS→SDS→요구사항)에만 의존한다. fix#1(testcase-only emit)이 vcast_rows를
        # 채워 위 'not deduped' 경로를 우회하므로, 이 경우를 별도 표면화해 2-hop 미해소 시의
        # silent를 막는다(deep-review W6 — fix#1↔fix#4 상충 해소). 프론트는 이 warning을
        # 성공 표시와 함께 노출한다.
        result["warning"] = (
            f"{src_label} 시트('{scanned_sheet}')에서 직접 요구사항 매핑 없이 testcase만 "
            f"{len(deduped)}건 추출했습니다(요구 열 없는 Test-Log 포맷). 매트릭스의 2-hop 추적에 "
            "의존하므로, SITS 밴드가 비면 SUTS/SDS 연결을 확인하세요."
        )
        result["available_sheets"] = list(wb.sheetnames)
    return result


@router.post("/api/jenkins/syts/extract-traceability")
def jenkins_syts_extract_traceability(body: Dict[str, Any]) -> Dict[str, Any]:
    """시스템 시험(SyTS) 결과 xlsx — SITS와 동일 TC↔요구사항 구조. source 라벨만 'SyTS'.

    요구사항 ID는 SITS 경로의 `_normalize_req_id`가 SyTSR/SyNTSR→Sw*로 평탄화 →
    비기능/안전 요구가 SW 행에 join돼 시스템 레벨 검증으로 covered 승격(결정1)."""
    return jenkins_sits_extract_traceability({**body, "source_label": "SyTS"})


@router.post("/api/jenkins/syits/extract-traceability")
def jenkins_syits_extract_traceability(body: Dict[str, Any]) -> Dict[str, Any]:
    """시스템 통합시험(SyITS) 결과 xlsx — SITS와 동일 구조. source 라벨만 'SyITS'."""
    return jenkins_sits_extract_traceability({**body, "source_label": "SyITS"})


@router.post("/api/jenkins/uds/publish")
def jenkins_uds_publish(req: UdsPublishRequest) -> Dict[str, Any]:
    out_dir = _jenkins_exports_dir(req.cache_root)
    try:
        target = safe_resolve_under(out_dir, req.filename)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid filename")
    if not target.exists():
        raise HTTPException(status_code=404, detail="file not found")
    docs_dir = (repo_root / req.target_dir).resolve()
    docs_dir.mkdir(parents=True, exist_ok=True)
    out_path = docs_dir / target.name
    out_path.write_bytes(target.read_bytes())
    return {"ok": True, "path": str(out_path)}


@router.post("/api/jenkins/uds/label")
def jenkins_uds_label(req: UdsLabelRequest) -> Dict[str, Any]:
    job_slug = _job_slug(req.job_url)
    out_dir = _jenkins_exports_dir(req.cache_root)
    if not out_dir.exists():
        raise HTTPException(status_code=404, detail="export dir not found")
    try:
        target = safe_resolve_under(out_dir, req.filename)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid filename")
    if not target.exists():
        raise HTTPException(status_code=404, detail="file not found")
    meta = _load_uds_meta(out_dir, job_slug)
    labels = meta.get("labels")
    if not isinstance(labels, dict):
        labels = {}
    label = (req.label or "").strip()
    if label:
        labels[req.filename] = label
    else:
        labels.pop(req.filename, None)
    meta["labels"] = labels
    _save_uds_meta(out_dir, job_slug, meta)
    return {"ok": True, "filename": req.filename, "label": label}


@router.post("/api/jenkins/uds/delete")
def jenkins_uds_delete(req: UdsDeleteRequest) -> Dict[str, Any]:
    job_slug = _job_slug(req.job_url)
    out_dir = _jenkins_exports_dir(req.cache_root)
    if not out_dir.exists():
        raise HTTPException(status_code=404, detail="export dir not found")
    try:
        target = safe_resolve_under(out_dir, req.filename)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid filename")
    if not target.exists():
        raise HTTPException(status_code=404, detail="file not found")
    removed: List[str] = []
    for candidate in [target, target.with_suffix(".html"), target.with_suffix(".md")]:
        if candidate.exists():
            try:
                candidate.unlink()
                removed.append(candidate.name)
            except Exception:
                continue
    meta = _load_uds_meta(out_dir, job_slug)
    labels = meta.get("labels")
    if isinstance(labels, dict):
        labels.pop(req.filename, None)
        meta["labels"] = labels
    _save_uds_meta(out_dir, job_slug, meta)
    return {"ok": True, "removed": removed}


@router.post("/api/jenkins/call-tree")
def jenkins_call_tree(req: JenkinsCallTreeRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    # 신뢰 소스 경계 = 캐시 빌드 루트뿐 (path-traversal 리뷰 finding [3] 엄격 유지).
    # Jenkins sync가 SCM 소스를 build_root/source(.source_complete 센티널)에 체크아웃하므로
    # 외부 SCM 작업본(C:/Project/Ados/...)을 신뢰 목록에 넣을 필요가 없다. 프론트가 관성적으로
    # 보내는 외부 절대경로는 신뢰하지 않고(레지스트리는 비-admin 쓰기 가능 → 신뢰 원천 부적격),
    # build_root 하위의 체크아웃 사본을 스캔한다 — 두 소스는 동일(byte-identical checkout).
    checked_out = build_root / "source"
    # 완전성은 정본 센티널(.source_complete)로 판정 — bare exists()는 중단/진행중 체크아웃의
    # 부분 트리를 완료본처럼 스캔해 과소집계를 완료로 오인시킨다. 신호는 meta로 정직하게 노출.
    source_complete = _source_is_complete(checked_out)
    source_root = checked_out if checked_out.exists() else build_root
    raw_src = str(req.source_root or "").strip()
    if raw_src:
        # 클라가 build_root 하위 경로를 명시하면 그 부분만 스캔(신뢰 경계 안에서만 존중).
        # build_root 밖(외부 SCM 경로 포함)이거나 미존재면 무시하고 체크아웃 사본으로 폴백 —
        # 임의 외부 경로를 파일 read 대상으로 삼지 않으므로 traversal 우회 없음.
        cands = [Path(p.strip()).resolve() for p in raw_src.replace(";", ",").split(",") if p.strip()]
        picked = next((c for c in cands if is_under_any(c, [build_root]) and c.exists()), None)
        if picked is not None:
            source_root = picked
    # all_roots=True면 진입 함수를 백엔드가 자동 산출(in-degree 0 + 순환 대표) → entry 불필요.
    all_roots = bool(getattr(req, "all_roots", False))
    reverse = bool(getattr(req, "reverse", False))  # 역방향(called-by) 트리 — 누가 이 함수를 호출하나
    entries = [x.strip() for x in str(req.entry or "").replace("\n", ",").split(",") if x.strip()][:200]
    if not entries and not all_roots:
        raise HTTPException(status_code=400, detail="entry required")
    if not source_root.exists():
        raise HTTPException(status_code=404, detail="source_root not found")
    compile_db = Path(req.compile_commands_path).resolve() if req.compile_commands_path else None
    def _regex_engine() -> Dict[str, Any]:
        return build_call_tree(
            source_root,
            entries,
            include_paths=req.include_paths or [],
            exclude_paths=req.exclude_paths or [],
            max_depth=max(1, int(req.max_depth or 5)),
            max_files=max(1, int(req.max_files or 2000)),
            include_external=bool(req.include_external),
            compile_commands_path=compile_db,
            external_map=req.external_map or [],
            auto_roots=all_roots,
            reverse=reverse,
        )

    engine = str(getattr(req, "engine", "precise") or "precise").strip().lower()
    if engine == "precise":
        payload = build_call_tree_precise(
            source_root,
            entries,
            include_paths=req.include_paths or [],
            exclude_paths=req.exclude_paths or [],
            max_depth=max(1, int(req.max_depth or 5)),
            max_files=max(1, int(req.max_files or 2000)),
            include_external=bool(req.include_external),
            external_map=req.external_map or [],
            auto_roots=all_roots,
            reverse=reverse,
        )
        # tree-sitter 미가용(engine='unavailable') → regex 엔진 자동 폴백 (R1 완화)
        if (payload.get("stats") or {}).get("engine") == "unavailable":
            payload = _regex_engine()
            payload.setdefault("stats", {})["engine"] = "regex-fallback"
    else:
        payload = _regex_engine()
        payload.setdefault("stats", {}).setdefault("engine", "regex")
    payload["meta"] = {
        "job_url": req.job_url,
        "build_selector": req.build_selector,
        "build_root": str(build_root),
        # 실제 스캔된 경로와 체크아웃 완전성을 정직하게 노출 — 프론트가 외부 경로를 보냈어도
        # build_root/source(체크아웃 사본)를 스캔했음을, 그리고 그 체크아웃이 부분/미완이면
        # source_complete=false로 알려 undercounted 트리를 완료로 오인하지 않게 한다.
        "source_root": str(source_root),
        "source_complete": bool(source_complete),
    }
    return payload


@router.post("/api/jenkins/call-tree/save")
def jenkins_call_tree_save(req: JenkinsCallTreeRequest) -> Dict[str, Any]:
    payload = jenkins_call_tree(req)
    out_dir = _jenkins_exports_dir(req.cache_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    job_slug = _job_slug(req.job_url)
    sel = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in str(req.build_selector))
    fmt = str(req.output_format or "json").strip().lower()
    if fmt not in ("json", "html", "csv"):
        raise HTTPException(status_code=400, detail="invalid output_format")
    if fmt == "html":
        out_path = out_dir / f"jenkins_call_tree_{job_slug}_{sel}_{ts}.html"
        out_path.write_text(call_tree_to_html(payload, req.html_template), encoding="utf-8")
    elif fmt == "csv":
        out_path = out_dir / f"jenkins_call_tree_{job_slug}_{sel}_{ts}.csv"
        out_path.write_text(call_tree_to_csv(payload), encoding="utf-8")
    else:
        out_path = out_dir / f"jenkins_call_tree_{job_slug}_{sel}_{ts}.json"
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"filename": out_path.name, "path": str(out_path), "format": fmt}


@router.post("/api/jenkins/call-tree/preview-html")
def jenkins_call_tree_preview(req: CallTreePreviewRequest) -> Dict[str, Any]:
    payload = req.call_tree or {}
    html = call_tree_to_html(payload, req.html_template)
    return {"html": html}


@router.get("/api/jenkins/call-tree/download")
def jenkins_call_tree_download(job_url: str, cache_root: str, filename: str) -> FileResponse:
    out_dir = _jenkins_exports_dir(cache_root)
    try:
        target = safe_resolve_under(out_dir, filename)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid filename")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="file not found")
    media = "application/json"
    if target.suffix.lower() == ".html":
        media = "text/html"
    elif target.suffix.lower() == ".csv":
        media = "text/csv"
    return FileResponse(str(target), filename=target.name, media_type=media)


@router.post("/api/jenkins/report/files")
def jenkins_report_files(req: JenkinsReportRequest) -> Dict[str, Any]:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    # Jenkins 동기화된 빌드 루트 전체에서 파일 목록을 제공
    return list_report_files(build_root)


@router.post("/api/jenkins/server/files")
def jenkins_server_files(req: JenkinsServerFilesRequest) -> Dict[str, Any]:
    roots = getattr(config, "JENKINS_SERVER_ROOTS", [])
    allowed = [Path(p).expanduser().resolve() for p in roots if p]
    if not allowed:
        raise HTTPException(status_code=400, detail="jenkins server roots not configured")
    base = Path(req.root or "").expanduser().resolve()
    if not is_under_any(base, allowed) and base not in allowed:
        raise HTTPException(
            status_code=403,
            detail=f"server root not allowed: {base}",
        )
    rel = req.rel_path or "."
    try:
        scan_root = safe_resolve_under(base, rel)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail=f"invalid rel_path: {rel}",
        )
    if not scan_root.exists() or not scan_root.is_dir():
        raise HTTPException(status_code=404, detail=f"root not found: {scan_root}")

    exts = [e.lower().lstrip(".") for e in (req.exts or []) if str(e).strip()]
    if not exts:
        exts = [e.lower() for e in getattr(config, "JENKINS_SERVER_DOC_EXTS", [])]
    max_files = max(1, int(req.max_files or 5000))

    files = []
    scanned = 0
    for dirpath, _, filenames in os.walk(scan_root):
        for name in filenames:
            scanned += 1
            if scanned > max_files:
                return {
                    "ok": True,
                    "root": str(base),
                    "scan_root": str(scan_root),
                    "files": files,
                    "scanned": scanned,
                    "truncated": True,
                }
            p = Path(dirpath) / name
            ext = p.suffix.lower().lstrip(".")
            if exts and ext not in exts:
                continue
            try:
                rel_path = str(p.relative_to(base)).replace("\\", "/")
            except Exception:
                continue
            try:
                stat = p.stat()
                size = int(stat.st_size)
                mtime = int(stat.st_mtime)
            except Exception:
                size = 0
                mtime = 0
            files.append(
                {
                    "rel_path": rel_path,
                    "path": str(p),
                    "ext": ext,
                    "size": size,
                    "mtime": mtime,
                }
            )
    return {
        "ok": True,
        "root": str(base),
        "scan_root": str(scan_root),
        "files": files,
        "scanned": scanned,
        "truncated": False,
    }


@router.post("/api/jenkins/report/files/download")
def jenkins_report_files_download(req: JenkinsReportRequest, path: str) -> FileResponse:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    try:
        target = safe_resolve_under(build_root, path)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid path")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="file not found")
    return FileResponse(str(target), filename=target.name)


@router.post("/api/jenkins/report/files/download/zip")
def jenkins_report_files_download_zip(req: JenkinsReportZipRequest) -> FileResponse:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    report_dir = build_root
    scope = str(req.scope or "all").strip().lower()
    if scope in ("report", "reports"):
        report_dir = _detect_reports_dir(build_root)
    out_dir = _jenkins_exports_dir(req.cache_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    job_slug = _job_slug(req.job_url)
    sel = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in str(req.build_selector))
    out_path = out_dir / f"jenkins_reports_{job_slug}_{sel}_{ts}.zip"
    _create_jenkins_zip_file(
        report_dir,
        out_path,
        include_paths=req.include_paths,
        exclude_paths=req.exclude_paths,
        exts=req.exts,
    )
    return FileResponse(out_path, filename=out_path.name, media_type="application/zip")


@router.post("/api/jenkins/report/files/download/zip/select")
def jenkins_report_files_download_zip_select(req: JenkinsReportRequest, sel: ReportZipRequest) -> FileResponse:
    build_root = _resolve_cached_build_root(req.job_url, req.cache_root, req.build_selector)
    if not build_root:
        raise HTTPException(status_code=404, detail="cached build not found")
    report_dir = build_root
    out_dir = _jenkins_exports_dir(req.cache_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    job_slug = _job_slug(req.job_url)
    sel_key = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in str(req.build_selector))
    out_path = out_dir / f"jenkins_reports_{job_slug}_{sel_key}_{ts}.zip"
    paths = sel.paths or []
    if not paths:
        raise HTTPException(status_code=400, detail="paths required")
    with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for rel in paths:
            try:
                target = safe_resolve_under(report_dir, rel)
            except Exception:
                continue
            if target.exists() and target.is_file():
                zf.write(target, arcname=rel)
    return FileResponse(out_path, filename=out_path.name, media_type="application/zip")


@router.post("/api/jenkins/report/publish")
def jenkins_report_publish(req: JenkinsPublishRequest) -> Dict[str, Any]:
    return _jenkins_report_publish_impl(req)


@router.post("/api/jenkins/report/publish-async")
def jenkins_report_publish_async(req: JenkinsPublishRequest) -> Dict[str, Any]:
    job_url = req.job_url
    build_selector = req.build_selector
    job_id = uuid.uuid4().hex
    _set_progress(
        "publish",
        job_url,
        build_selector,
        {
            "stage": "start",
            "percent": 1,
            "message": "로컬 리포트 업로드 준비 중",
            "done": False,
            "error": "",
        },
        job_id=job_id,
    )

    def _run_publish() -> None:
        try:
            _jenkins_report_publish_impl(req, job_id=job_id)
        except Exception as exc:
            _set_progress(
                "publish",
                job_url,
                build_selector,
                {
                    "stage": "error",
                    "percent": 100,
                    "message": "로컬 리포트 업로드 실패",
                    "done": True,
                    "error": str(exc),
                },
                job_id=job_id,
            )
            return
        _set_progress(
            "publish",
            job_url,
            build_selector,
            {
                "stage": "done",
                "percent": 100,
                "message": "로컬 리포트 업로드 완료",
                "done": True,
            },
            job_id=job_id,
        )

    t = threading.Thread(target=wrap_with_user(_run_publish), daemon=True)
    t.start()
    return {"ok": True, "job_id": job_id}


# ──────────────────────────────────────────────────────────────────
# Aggregate stats across multiple projects
# ──────────────────────────────────────────────────────────────────

@router.post("/api/jenkins/aggregate-stats")
def aggregate_stats(req: dict) -> Dict[str, Any]:
    """Aggregate analysis_summary.json from latest builds of multiple jobs.

    Args:
        req: Dict with ``job_urls`` (list of Jenkins job URLs) and
             optional ``cache_root`` (defaults to .devops_pro_cache).

    Returns:
        Aggregated coverage, test, PRQA, and code metric statistics.
    """
    job_urls: List[str] = req.get("job_urls") or []
    cache_root: str = req.get("cache_root", ".devops_pro_cache")

    base = _normalize_jenkins_cache_root(cache_root)

    projects: List[Dict[str, Any]] = []

    # Accumulators
    cov_line_rates: List[float] = []
    cov_branch_rates: List[float] = []
    total_covered = 0
    total_lines = 0

    total_ut_cases = 0
    passed_ut_cases = 0
    total_it_cases = 0
    passed_it_cases = 0
    all_pass = True

    total_files = 0
    total_functions = 0
    total_nloc = 0

    total_diagnostics = 0
    total_loc = 0
    total_files_analyzed = 0

    # Get current user for multi-user cache isolation
    from backend.user_context import get_current_user
    current_user = get_current_user()

    for job_url in job_urls:
        slug = _job_slug(job_url)

        # Cache path patterns (user-isolated, in priority order):
        # 1. {base}/{user}/jenkins/{slug}/build_*            (current frontend defaultCacheRoot)
        # 2. {base}/jenkins/{slug}/build_*                    (legacy direct / shared)
        # 3. {base}/{user}/{fe_slug}/jenkins/{slug}/build_*  (legacy frontend with job-slug subdir)
        # 4. {base}/{fe_slug}/jenkins/{slug}/build_*          (legacy frontend w/o user)
        user_base = base / current_user if (base / current_user).exists() else None
        candidates: List[Path] = []
        if user_base:
            candidates.append(user_base / "jenkins" / slug)
        candidates.append(base / "jenkins" / slug)
        if user_base:
            # Legacy: job-slug subdirectories under the user's folder
            candidates.extend(
                d / "jenkins" / slug
                for d in user_base.iterdir()
                if d.is_dir() and d.name != "jenkins"
            )
        else:
            # Fallback: legacy structure without user directory
            candidates.extend(
                d / "jenkins" / slug
                for d in base.iterdir()
                if d.is_dir() and d.name not in ("jenkins", "exports")
            )

        job_root = None
        for candidate in candidates:
            if candidate.exists() and list(candidate.glob("build_*")):
                job_root = candidate
                break
        if not job_root:
            continue

        # Find latest build directory (highest build_N number)
        build_dirs = sorted(job_root.glob("build_*"), reverse=True)
        if not build_dirs:
            continue

        summary_path = build_dirs[0] / "report" / "analysis_summary.json"
        if not summary_path.exists():
            # Also check reports/ subdirectory
            summary_path = build_dirs[0] / "reports" / "analysis_summary.json"
        if not summary_path.exists():
            continue

        try:
            data = json.loads(summary_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        def _safe_int(val: Any, default: int = 0) -> int:
            try:
                return int(val) if val is not None else default
            except (TypeError, ValueError):
                return default

        def _safe_float(val: Any) -> float | None:
            try:
                return float(val) if val is not None else None
            except (TypeError, ValueError):
                return None

        # Coverage
        cov = data.get("coverage") or {}
        lr = _safe_float(cov.get("line_rate"))
        br = _safe_float(cov.get("branch_rate"))
        if lr is not None:
            cov_line_rates.append(lr)
        if br is not None:
            cov_branch_rates.append(br)
        total_covered += _safe_int(cov.get("covered"))
        total_lines += _safe_int(cov.get("total"))

        # Tests
        tests = data.get("tests") or {}
        details = tests.get("details") or {}
        ut = details.get("ut") or {}
        it = details.get("it") or {}
        ut_tc = ut.get("testcases") or {}
        it_tc = it.get("testcases") or {}
        total_ut_cases += _safe_int(ut_tc.get("total"))
        passed_ut_cases += _safe_int(ut_tc.get("ok"))
        total_it_cases += _safe_int(it_tc.get("total"))
        passed_it_cases += _safe_int(it_tc.get("ok"))
        if not tests.get("ok", True):
            all_pass = False

        # Code metrics
        cm = data.get("code_metrics") or {}
        total_files += _safe_int(cm.get("code_files"))
        total_functions += _safe_int(cm.get("functions"))
        total_nloc += _safe_int(cm.get("nloc"))

        # PRQA
        prqa = data.get("prqa") or {}
        crr = prqa.get("crr") or {}
        total_diagnostics += _safe_int(crr.get("diagnostic_count"))
        total_loc += _safe_int(crr.get("loc_source"))
        total_files_analyzed += _safe_int(crr.get("number_of_files"))

        # Jenkins info
        jenkins = data.get("jenkins") or {}
        ut_total = _safe_int(ut_tc.get("total"))

        # PRQA RCR summary for compliance metrics
        rcr = prqa.get("rcr") or {}
        rcr_summary = rcr.get("summary") or {}

        def _parse_fraction_first(val: Any) -> int:
            """Extract first number from 'N/M' string or return int."""
            if isinstance(val, (int, float)):
                return int(val)
            if isinstance(val, str):
                parts = val.split("/")
                try:
                    return int(parts[0].strip())
                except (ValueError, IndexError):
                    return 0
            return 0

        projects.append({
            "job_url": job_url,
            "name": job_url.rstrip("/").split("/")[-1],
            "build_number": jenkins.get("build_number"),
            "result": jenkins.get("result"),
            "line_rate": lr,
            "branch_rate": br,
            "ut_pass_rate": (
                round(_safe_int(ut_tc.get("ok")) / ut_total, 4)
                if ut_total > 0 else None
            ),
            "ut_total": _safe_int(ut_tc.get("total")),
            "it_total": _safe_int(it_tc.get("total")),
            "diagnostics": _safe_int(crr.get("diagnostic_count")),
            "loc": _safe_int(cm.get("nloc")),
            "functions": _safe_int(cm.get("functions")),
            "rcr_violated_rules": _parse_fraction_first(rcr_summary.get("Violated Rules", 0)),
            "rcr_compliance_index": _parse_fraction_first(rcr_summary.get("Project Compliance Index", 0)),
        })

    return {
        "project_count": len(projects),
        "coverage": {
            "avg_line_rate": (
                round(sum(cov_line_rates) / len(cov_line_rates), 4)
                if cov_line_rates else None
            ),
            "avg_branch_rate": (
                round(sum(cov_branch_rates) / len(cov_branch_rates), 4)
                if cov_branch_rates else None
            ),
            "total_covered": total_covered,
            "total_lines": total_lines,
        },
        "tests": {
            "total_ut_cases": total_ut_cases,
            "passed_ut_cases": passed_ut_cases,
            "total_it_cases": total_it_cases,
            "passed_it_cases": passed_it_cases,
            "all_pass": all_pass,
        },
        "code_metrics": {
            "total_files": total_files,
            "total_functions": total_functions,
            "total_nloc": total_nloc,
        },
        "prqa": {
            "total_diagnostics": total_diagnostics,
            "total_loc": total_loc,
            "total_files_analyzed": total_files_analyzed,
        },
        "projects": projects,
    }
