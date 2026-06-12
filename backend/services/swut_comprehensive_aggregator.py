"""SwUTCR (Software Unit Test Comprehensive Result) xlsm builder.

The comprehensive result workbook is a template-copy artifact.  It preserves the
company xlsm template, including merges, styles, borders, and VBA entries, then
fills every known SwUT evidence sheet from the same data sources used by SwUTCV
and SwUTR.
"""
from __future__ import annotations

import hashlib
import io
import re
from copy import copy
from dataclasses import dataclass, field, replace as dc_replace
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

from backend.services.excel_template_utils import (
    build_release_history_row,
    has_vba_macros,
    inspect_vba_refs,
    mark_user_input_required,
    safe_write,
    short_date,
    validate_build_meta,
    validate_xlsx_template_bytes,
)
from backend.services.swut_builder_helpers import (
    diagnose_asset_usage,
    extract_warnings_from_session,
)
from backend.services.swut_input_adapter import (
    CoverageStats,
    SwUTSession,
    aggregate_session,
)
from backend.services.swut_meta import BuildMetaBase


@dataclass
class SwutcrBuildMeta(BuildMetaBase):
    """Build metadata for SwUTCR."""

    doc_id_base: str = "HDPDM01-SwUTCR"
    target_coverage: float = 1.0
    target_pass_ratio: float = 1.0
    final_test_result: str = "OK"
    project_config: dict[str, Any] = field(default_factory=dict)


@dataclass
class SwutcrBuildResult:
    """SwUTCR build result."""

    ok: bool
    xlsm_io: io.BytesIO = field(default_factory=io.BytesIO)
    filename: str = ""
    warnings: list[str] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    incomplete_sheets: list[str] = field(default_factory=list)
    vba_macros_preserved: bool = False
    tool_qualification: dict[str, Any] = field(
        default_factory=lambda: {
            "evidence_class": "auto-generated draft",
            "asil_a_usage": "reviewer review required before evidence use",
            "asil_b_c_d_usage": "manual review required; do not use as standalone evidence",
        }
    )

    @property
    def xlsm_bytes(self) -> bytes:
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0)
        try:
            return self.xlsm_io.read()
        finally:
            self.xlsm_io.seek(pos)

    @property
    def result_size_bytes(self) -> int:
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0, 2)
        size = self.xlsm_io.tell()
        self.xlsm_io.seek(pos)
        return size

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "filename": self.filename,
            "result_size_bytes": self.result_size_bytes,
            "warnings": self.warnings,
            "incomplete_sheets": self.incomplete_sheets,
            "vba_macros_preserved": self.vba_macros_preserved,
            "summary": self.summary,
            "tool_qualification": self.tool_qualification,
        }


def _find_sheet(wb: Workbook, *needles: str, exclude: tuple[str, ...] = ()):
    lowered_exclude = tuple(e.lower() for e in exclude)
    for name in wb.sheetnames:
        low = name.lower()
        if any(e in low for e in lowered_exclude):
            continue
        if all(n.lower() in low for n in needles):
            return wb[name]
    return None


def _stamp_hmr_metrics(
    agg: dict[str, Any],
    hmr_html_bytes: bytes | None,
    warnings: list[str],
) -> None:
    """Stamp Function Calls coverage into aggregate function rows when HMR exists."""
    if not hmr_html_bytes:
        return
    from backend.services.vcast_hmr_parser import parse_hmr_html

    parse_warnings: list[str] = []
    hmr_result = parse_hmr_html(hmr_html_bytes, parse_warnings=parse_warnings)
    warnings.extend([f"[hmr] {w}" for w in parse_warnings])
    if not hmr_result.ok:
        warnings.append("[hmr] parse failed; Function Calls metric stamp skipped")
        return

    stamped = 0
    ambiguous = 0
    new_rows = []
    for fc in agg.get("function_rows") or []:
        candidates = hmr_result.metrics_by_name.get(fc.name, [])
        if len(candidates) > 1:
            ambiguous += 1
            warnings.append(
                f"[hmr] ambiguous function '{fc.name}' matched multiple HMR rows; stamp skipped"
            )
            new_rows.append(fc)
            continue
        metric = candidates[0] if candidates else None
        if metric and metric.total_calls > 0:
            new_rows.append(dc_replace(
                fc,
                function_calls_coverage=CoverageStats(
                    covered=metric.covered_calls,
                    total=metric.total_calls,
                    coverage_pct=metric.coverage_pct / 100.0,
                ),
            ))
            stamped += 1
        else:
            new_rows.append(fc)
    agg["function_rows"] = new_rows
    warnings.append(
        f"[hmr] Function Calls metric stamped: {stamped}/{len(new_rows)} "
        f"(ambiguous skipped: {ambiguous})"
    )


def _lookup_c_function(
    function_name: str,
    unit_id: str,
    c_function_map: dict[str, dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not c_function_map:
        return None
    for key in (function_name, unit_id):
        if key and key in c_function_map:
            return c_function_map[key]
    lowered = {k.lower(): v for k, v in c_function_map.items() if isinstance(k, str)}
    for key in (function_name, unit_id):
        if key and key.lower() in lowered:
            return lowered[key.lower()]
    return None


def _find_source_line(c_entry: dict[str, Any]) -> int | None:
    file_path = c_entry.get("file")
    name = str(c_entry.get("name") or "")
    signature = str(c_entry.get("signature") or "")
    if not file_path or not name:
        return None
    try:
        lines = Path(str(file_path)).read_text(encoding="utf-8", errors="ignore").splitlines()
    except OSError:
        return None

    normalized_signature = " ".join(signature.split())
    for idx, line in enumerate(lines, start=1):
        stripped = line.strip()
        if (
            normalized_signature
            and normalized_signature in " ".join(line.split())
            and not stripped.endswith(";")
        ):
            return idx
    needle = f"{name}("
    for idx, line in enumerate(lines, start=1):
        stripped = line.strip()
        if needle in line.replace(" ", "") and not stripped.endswith(";"):
            return idx
    return None


def _select_code_evidence_lines(body: str, function_name: str) -> list[str]:
    lines = [line.strip() for line in body.splitlines() if line.strip()]
    if not lines:
        return []
    lowered_name = function_name.lower()
    lowered_lines = [line.lower() for line in lines]

    priority_terms: list[str]
    if "workaround" in lowered_name or "adc0" in lowered_name:
        priority_terms = ["errata", "adc", "workaround", "adc0ctl", "adc0flwctl"]
    elif "register" in lowered_name or any("eccie" in line for line in lowered_lines):
        priority_terms = ["eccie", "register", "loopcnt"]
    elif "interpolate" in lowered_name:
        priority_terms = ["== (s32)0", "tempinterpolated", "adc2 - adc1"]
    elif "default:" in "\n".join(lowered_lines):
        priority_terms = ["default:", "switch"]
    elif "boundary" in "\n".join(lowered_lines) or "clamp" in "\n".join(lowered_lines):
        priority_terms = ["boundary", "<=", ">=", "clamp"]
    else:
        priority_terms = [
            "if", "else", "error", "err", "fail", "protect", "timeout",
            "crc", "checksum", "return",
        ]

    selected: list[str] = []
    for term in priority_terms:
        for idx, line in enumerate(lowered_lines):
            if term in line:
                start = max(idx - 1, 0)
                end = min(idx + 3, len(lines))
                selected.extend(lines[start:end])
                break
        if len(selected) >= 4:
            break
    if not selected:
        selected = lines[:4]

    deduped: list[str] = []
    for line in selected:
        if line not in deduped:
            deduped.append(line)
    return deduped[:6]


def _build_c_evidence(function_name: str, c_entry: dict[str, Any] | None) -> tuple[str, str]:
    if not c_entry:
        return "", ""
    source_file = Path(str(c_entry.get("file") or "")).name or "C source"
    line_no = _find_source_line(c_entry)
    location = f"{source_file}:{line_no}" if line_no else source_file
    code_lines = _select_code_evidence_lines(str(c_entry.get("body") or ""), function_name)
    if not code_lines:
        return location, location
    short_code = " / ".join(code_lines[:4])
    if len(short_code) > 520:
        short_code = short_code[:517] + "..."
    detail = location + "\n" + "\n".join(code_lines)
    return f"{location} | {short_code}", detail


def _build_full_function_text(c_entry: dict[str, Any] | None) -> str:
    if not c_entry:
        return ""
    signature = str(c_entry.get("signature") or "").strip()
    body = str(c_entry.get("body") or "").strip()
    if not (signature or body):
        return ""
    text = "\n".join(part for part in (signature, body) if part)
    # Excel cell text limit is 32,767 chars. Keep room for action text prefix.
    if len(text) > 30000:
        text = text[:30000] + "\n/* truncated: function body exceeds Excel cell limit */"
    return text


def _draft_failure_rationale(
    function_name: str,
    kind: str,
    value: str,
    c_entry: dict[str, Any] | None,
) -> tuple[str, str]:
    if not c_entry:
        return (
            f"{kind} coverage not completed by VectorCAST result ({value}); "
            "C source evidence unavailable.",
            "Review VectorCAST uncovered path, add TC if reachable, or document "
            "unreachable/deviation rationale.",
        )

    body = str(c_entry.get("body") or "").lower()
    source_file = Path(str(c_entry.get("file") or "")).name or "C source"
    calls = ", ".join(str(call) for call in (c_entry.get("calls") or [])[:3])
    globals_used = ", ".join(str(item) for item in (c_entry.get("used_globals") or [])[:3])
    context = f"{function_name} in {source_file}"
    lowered_name = function_name.lower()

    if "default:" in body:
        return (
            f"{kind} uncovered path appears to include switch/default defensive logic "
            f"({context}, {value}).",
            "Add a negative/default-path TC if reachable; otherwise record the "
            "unreachable defensive-branch rationale.",
        )
    if "null" in body or "nullptr" in body:
        return (
            f"{kind} uncovered path appears to include NULL guard logic "
            f"({context}, {value}).",
            "Add a NULL-input/error-path TC if the interface permits it; otherwise "
            "record why the guard is unreachable in this integration.",
        )
    if "workaround" in lowered_name or "adc0" in lowered_name or "errata" in body:
        return (
            f"{kind} uncovered path appears to include hardware errata/workaround "
            f"logic ({context}, {value}).",
            "Review whether the ADC/register workaround path is reachable in unit "
            "test; add hardware-state TC or document tool/environment limitation.",
        )
    if "register" in lowered_name or "eccie" in body:
        return (
            f"{kind} uncovered path appears to include hardware register self-test "
            f"logic ({context}, {value}).",
            "Add register success/failure TC coverage if the register can be "
            "stimulated; otherwise document hardware access limitation.",
        )
    if "interpolate" in lowered_name or "tempinterpolated" in body:
        return (
            f"{kind} uncovered path appears to include interpolation or divide-by-zero "
            f"guard logic ({context}, {value}).",
            "Add lookup-table boundary/interpolation TC coverage, including the "
            "equal-ADC guard if reachable.",
        )
    if any(token in body for token in (
        "error", "fail", "timeout", "crc", "checksum", "not_ok",
        "err_", "_err", "protect", "mucerror",
    )):
        return (
            f"{kind} uncovered path appears to include error-handling logic "
            f"({context}, {value}).",
            "Add fault-injection/error-path TC coverage, or document why the fault "
            "condition cannot be stimulated.",
        )
    if any(token in body for token in (
        "boundary", "clamp", "overflow", "underflow", "<=", ">=", "lookup",
    )):
        return (
            f"{kind} uncovered path appears to include range or boundary handling "
            f"({context}, {value}).",
            "Add boundary-value TC coverage, or document the excluded operating range "
            "with the linked requirement.",
        )
    if c_entry.get("is_static"):
        action_tail = f" Calls seen: {calls}." if calls else ""
        return (
            f"{kind} coverage not completed for static helper evidence "
            f"({context}, {value}).",
            "Review caller scenarios and add TC coverage through the public caller, "
            f"or justify helper-only unreachable code.{action_tail}",
        )

    detail = f" Uses globals: {globals_used}." if globals_used else ""
    return (
        f"{kind} coverage not completed by VectorCAST result with C source matched "
        f"({context}, {value}).",
        "Review uncovered statements/branches against the source and add TC coverage "
        f"or document unreachable/deviation rationale.{detail}",
    )


def _coverage_failures(
    function_rows: list[Any],
    c_function_map: dict[str, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    failures: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for fc in function_rows:
        for kind, stats in (("Statement", fc.statement), ("Branch", fc.branch)):
            if stats.total <= 0 or stats.covered >= stats.total:
                continue
            key = (fc.unit_id or fc.name, kind)
            if key in seen:
                continue
            seen.add(key)
            function_name = fc.name or fc.unit_id
            value = f"{stats.covered}/{stats.total}"
            c_entry = _lookup_c_function(function_name, fc.unit_id, c_function_map)
            reason, action = _draft_failure_rationale(function_name, kind, value, c_entry)
            evidence_short, evidence_detail = _build_c_evidence(function_name, c_entry)
            full_function = _build_full_function_text(c_entry)
            if evidence_short:
                reason = f"{reason}\nC code evidence: {evidence_short}"
                action = (
                    f"{action}\nReview basis: use the above C code branch/condition "
                    "as the TC design or unreachable/deviation rationale."
                )
            if full_function:
                action = f"{action}\n\nFull C function:\n{full_function}"
            failures.append({
                "function": function_name,
                "kind": kind,
                "value": value,
                "reason": reason,
                "action": action,
                "c_evidence": evidence_detail,
                "full_function_line_count": full_function.count("\n") + 1 if full_function else 0,
            })
    return failures


def _write_swutcr_metadata(ws, meta: SwutcrBuildMeta, cfg: dict[str, Any]) -> None:
    md = cfg.get("swutcr_metadata", {}) or {}
    safe_write(ws, 3, 5, md.get("project", meta.project_id))
    safe_write(ws, 4, 5, str(md.get("phase") or "").strip() or "DV")
    safe_write(ws, 5, 5, md.get("software_platform_ver", meta.release_sw_version))
    safe_write(ws, 6, 5, md.get("product", meta.project_id))
    safe_write(ws, 7, 5, md.get("verification_target", "MCU"))
    asil = (meta.asil_level or "").replace("ASIL", "").strip() or meta.asil_level
    safe_write(ws, 8, 5, md.get("asil", asil))
    safe_write(ws, 9, 5, md.get("compiler", ""))
    safe_write(ws, 10, 5, md.get("mcu", ""))
    # Requirement based unit and fault injection are generated from current SwUT data.
    # Back-to-back is kept disabled until a B2B evidence source is provided.
    safe_write(ws, 16, 7, md.get("ut101_enabled", "O"))
    safe_write(ws, 17, 7, md.get("ut201_enabled", "O"))
    safe_write(ws, 18, 7, md.get("ut301_enabled", "X"))


def _write_swutcr_sheet_header(ws, meta: SwutcrBuildMeta, cfg: dict[str, Any]) -> None:
    md = cfg.get("swutcr_metadata", {}) or {}
    safe_write(ws, 4, 3, md.get("test_iteration", "0.1"))
    safe_write(ws, 5, 3, md.get("software_platform_ver", meta.release_sw_version))
    safe_write(ws, 6, 3, meta.test_engineer or md.get("tester", meta.author))
    safe_write(ws, 7, 3, md.get("debugger", ""))
    safe_write(ws, 4, 6, md.get("prepare_hours", 0))
    safe_write(ws, 5, 6, md.get("execution_hours", 0))
    safe_write(ws, 6, 6, md.get("review_hours", 0))


def _write_ut101_long_text(ws, row: int, col: int, value: str) -> None:
    safe_write(ws, row, col, value)
    cell = ws.cell(row, col)
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    alignment.vertical = "top"
    cell.alignment = alignment


def _swutcr_function_count(agg: dict[str, Any]) -> int:
    return (
        agg.get("swutcr_qualified_function_count")
        or agg.get("function_count")
        or len(agg.get("function_rows") or [])
        or 0
    )


def _write_ut101(
    ws,
    meta: SwutcrBuildMeta,
    agg: dict[str, Any],
    cfg: dict[str, Any],
    warnings: list[str] | None = None,
) -> None:
    _write_swutcr_sheet_header(ws, meta, cfg)
    # W-8: 결함 ID prefix는 config phase를 따른다 (PV 산출물에 'DV' 고정 방지).
    md = cfg.get("swutcr_metadata", {}) or {}
    # H5 가드 — coverage 쪽 spec_phase와 동일 정규화 (''/None → 'DV').
    phase = str(md.get("phase") or "").strip() or "DV"
    function_rows = agg.get("function_rows") or []
    function_count = _swutcr_function_count(agg)
    failed_tcs = agg.get("failed", 0) or 0
    failures = _coverage_failures(function_rows, agg.get("c_function_map") or None)
    statement_fail = sum(1 for f in failures if f["kind"] == "Statement")
    branch_fail = sum(1 for f in failures if f["kind"] == "Branch")

    safe_write(ws, 73, 6, function_count)
    safe_write(ws, 73, 8, statement_fail)
    safe_write(ws, 73, 10, 0)
    safe_write(ws, 73, 11, 0)
    safe_write(ws, 73, 12, statement_fail)
    safe_write(ws, 74, 6, function_count)
    safe_write(ws, 74, 8, branch_fail)
    safe_write(ws, 74, 10, 0)
    safe_write(ws, 74, 11, 0)
    safe_write(ws, 74, 12, branch_fail)
    for col in range(6, 14):
        safe_write(ws, 75, col, "-")
    safe_write(ws, 76, 6, function_count)
    safe_write(ws, 76, 8, failed_tcs)
    safe_write(ws, 76, 9, "=G76/F76")
    safe_write(ws, 76, 10, 0)
    safe_write(ws, 76, 11, 0)
    safe_write(ws, 76, 12, failed_tcs)
    safe_write(ws, 76, 13, failed_tcs)
    for row in (77, 78):
        safe_write(ws, row, 6, function_count)
        safe_write(ws, row, 8, 0)
        safe_write(ws, row, 10, 0)
        safe_write(ws, row, 11, 0)
        safe_write(ws, row, 12, 0)

    start = 84
    max_failure_rows = 21  # Template-preformatted UT101 detail area: rows 84..104.
    if len(failures) > max_failure_rows and warnings is not None:
        warnings.append(
            f"UT101 coverage detail rows truncated to template capacity "
            f"{max_failure_rows}/{len(failures)}; summary counts retain full total."
        )
    for idx, failure in enumerate(failures[:max_failure_rows], start=1):
        row = start + idx - 1
        safe_write(ws, row, 2, f"UT-CVG-{phase}-{idx}")
        safe_write(ws, row, 3, "CVG")
        safe_write(ws, row, 4, failure["function"])
        safe_write(ws, row, 6, failure["kind"])
        safe_write(ws, row, 7, failure["value"])
        _write_ut101_long_text(ws, row, 8, failure["reason"])
        _write_ut101_long_text(ws, row, 12, failure["action"])
        line_count = max(
            str(failure["reason"]).count("\n") + 1,
            str(failure["action"]).count("\n") + 1,
        )
        full_function_lines = int(failure.get("full_function_line_count") or 0)
        target_height = min(max(78, line_count * 12, full_function_lines * 8), 409)
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, target_height)
        safe_write(ws, row, 15, "N/A")


def _collect_fi_fail_observations(
    session: SwUTSession,
) -> tuple[set[tuple[str, int]], list[tuple[str, int]]]:
    """라운드 108 MAJOR-2 — env 전수 관측으로 Fail 우선 병합 입력 수집.

    ``_build_fn_iteration_map`` 은 SUTR 공용(마지막 env가 같은
    'SwUFn_NNNN.MMM' 키를 덮어씀 — last-wins)이라 env1 Fail + env2 Pass면
    Fail이 가려진다. 공용 함수는 불변으로 두고, UT201 FI 교차 전용으로 전
    env를 다시 관측해 (a) Fail이 1회라도 관측된 키 집합, (b) env 간 상충
    (동일 키에 서로 다른 결과 상태) 키 목록을 돌려준다.

    Returns:
        (fail_keys, conflict_keys) — 키는 (정규화 num, iteration idx).
    """
    observed: dict[tuple[str, int], set[str]] = {}
    fail_keys: set[tuple[str, int]] = set()
    for env in session.environments:
        for tc_name in env.test_cases:
            m = re.match(r"SwUFn_(\d+)\.(\d+)", tc_name)
            if not m:
                continue
            key = (m.group(1).lstrip("0") or "0", int(m.group(2)))
            exec_r = env.test_results.get(tc_name)
            if exec_r is None:
                state = "miss"
            elif bool(exec_r.passed):
                state = "pass"
            else:
                state = "fail"
                fail_keys.add(key)
            observed.setdefault(key, set()).add(state)
    conflict_keys = sorted(k for k, states in observed.items() if len(states) > 1)
    return fail_keys, conflict_keys


def _cross_spec_fi_with_session(
    spec_fi: dict[str, Any],
    session: SwUTSession,
    out_warnings: list[str] | None = None,
) -> dict[str, Any]:
    """spec FI 추출 결과 × 실행결과 교차 (확정 규칙 2026-06-12).

    passed = FI 블록 중 (a) 모든 FI iteration이 실행결과에 존재하고 (b) 전부
    Pass인 블록 수 — Fail 또는 미실행 FI iteration이 있으면 미통과.

    ``not_executed_blocks`` = ≥1 FI iteration 미실행(또는 iteration index 파싱
    불가 — 검증 불가는 보수적으로 미실행 취급) 블록 수, ``failed_blocks`` =
    ≥1 Fail 블록 수. 한 블록이 양쪽 모두에 잡힐 수 있음 (집합 중첩 허용 —
    total ≠ passed + 미실행 + fail 일 수 있다).

    라운드 108 보수화 3종 (리뷰 findings fix):
        - MAJOR-2: env 간 동일 TC iteration 상충 결과는 Fail 우선 병합
          (passed=False 1회라도 관측되면 고정) + 상충 warning.
        - MINOR-3: 검증 가능 FI iteration 0개 블록(FI 세그먼트 anchor-only
          등)은 공허 통과 금지 — 미실행 분류 + 블록 수 warning.
        - MINOR-4: 중복 TC_ID 숫자 키(``fi_dup_keys``) 블록은 실행결과 귀속이
          모호하므로 보수적 미실행 처리 (블록 단위 카운트).

    ``stamped`` 는 ``_write_ut201`` 이 실제 E85/F85 stamp 분기에서만 True로
    갱신 (MINOR-6 — 시트 전무 템플릿에서 산출 위장 차단).
    """
    from backend.services.swut_sutr_spec_builder import _build_fn_iteration_map

    fn_iter_map = _build_fn_iteration_map(session)
    fail_keys, conflict_keys = _collect_fi_fail_observations(session)
    fi_keys: set[str] = spec_fi.get("fi_block_keys") or set()
    rows_per: dict[str, list[int]] = spec_fi.get("fi_iter_rows_per_block") or {}
    iters_per: dict[str, list[int]] = spec_fi.get("fi_iters_per_block") or {}
    dup_list: list[str] = list(spec_fi.get("fi_dup_keys") or [])
    dup_set: set[str] = set(dup_list)

    passed = 0
    not_executed = 0
    failed = 0
    zero_verifiable_blocks = 0
    dup_blocks = 0
    for key in sorted(fi_keys):
        if key in dup_set:
            # MINOR-4 — dup 키는 블록 단위 복원(1 + 추가 entry 수)해 전부
            # 미실행 처리 (total은 블록 단위라 counter도 블록 단위 정합).
            n_blocks = 1 + dup_list.count(key)
            dup_blocks += n_blocks
            not_executed += n_blocks
            continue
        rows = rows_per.get(key) or []
        iters = iters_per.get(key) or []
        rec_map = fn_iter_map.get(key) or {}
        if not iters:
            # MINOR-3 — 검증 가능 FI iteration 0개: any_missing=False·
            # any_fail=False 공허 통과를 차단하고 보수적 미실행 분류
            # (index 파싱 불가를 미실행 취급하는 기존 정책과 대칭).
            zero_verifiable_blocks += 1
            not_executed += 1
            continue
        any_missing = len(iters) < len(rows)  # index 파싱 불가 행 — 검증 불가
        any_fail = False
        for it in iters:
            if (key, it) in fail_keys:
                # MAJOR-2 — 어느 env든 Fail 1회 관측 시 고정 (last-wins 차단).
                any_fail = True
                continue
            rec = rec_map.get(it)
            if rec is None or rec.get("passed") is None:
                any_missing = True
            elif rec.get("passed") is False:
                any_fail = True
        if any_fail:
            failed += 1
        if any_missing:
            not_executed += 1
        if not any_fail and not any_missing:
            passed += 1

    if out_warnings is not None:
        if conflict_keys:
            sample = ", ".join(
                f"SwUFn_{num}.{idx:03d}" for num, idx in conflict_keys[:5]
            )
            out_warnings.append(
                "[swutcr] UT201 FI 교차 — env 간 동일 TC iteration 상충 결과 "
                f"{len(conflict_keys)}건 (예: {sample}) — Fail 관측 키는 Fail "
                "우선 고정"
            )
        if zero_verifiable_blocks:
            out_warnings.append(
                "[swutcr] UT201 FI 교차 — 검증 가능 FI iteration 0개 블록 "
                f"{zero_verifiable_blocks}건 (FI 세그먼트 anchor-only 등) — "
                "보수적 미실행 처리"
            )
        if dup_blocks:
            out_warnings.append(
                "[swutcr] UT201 FI 교차 — 중복 TC_ID 숫자 키 블록 "
                f"{dup_blocks}건 — 실행결과 귀속 모호로 보수적 미실행 처리"
            )

    return {
        "total": int(spec_fi.get("fi_block_total") or 0),
        "passed": passed,
        "not_executed_blocks": not_executed,
        "failed_blocks": failed,
        "iteration_total": int(spec_fi.get("fi_iteration_total") or 0),
        "spec_filename": str(spec_fi.get("spec_filename") or ""),
        "rule": str(spec_fi.get("rule") or "Test Method 'FI' 포함 TC 블록 수"),
        # MINOR-6 — 실제 stamp 분기(_write_ut201)에서만 True로 갱신.
        "stamped": False,
    }


def _write_ut201(
    ws,
    meta: SwutcrBuildMeta,
    agg: dict[str, Any],
    cfg: dict[str, Any],
    warnings: list[str] | None = None,
    spec_fi_auto: dict[str, Any] | None = None,
) -> None:
    _write_swutcr_sheet_header(ws, meta, cfg)
    md = cfg.get("swutcr_metadata", {}) or {}
    function_count = _swutcr_function_count(agg)
    safe_write(ws, 70, 4, md.get("tool_name", "VectorCAST"))
    safe_write(ws, 71, 4, md.get("tool_version", ""))
    safe_write(ws, 72, 3, md.get("excluded_scope", ""))
    safe_write(ws, 72, 6, md.get("excluded_size", ""))
    safe_write(ws, 77, 3, md.get("reference_document", "SwTP"))
    safe_write(ws, 78, 3, md.get("reference_id", "SwUTE_01"))
    safe_write(ws, 85, 3, function_count)

    # C-5: Fault Injection 실측치는 config 키가 있을 때만 stamp.
    # 이전 fallback(total=함수 수, passed=failed==0이면 total)은 FI 실측 부재 시
    # 측정한 것처럼 보이는 수치(예: 1014/1014)를 무표식 기입 — 24차 silent N/A
    # 제거 정책에 따라 노란 사용자입력 마킹으로 대체 (write_value_or_mark 패턴).
    fi_total = md.get("fault_injection_total")
    fi_passed = md.get("fault_injection_passed")
    has_fi_total = fi_total is not None and fi_total != ""
    has_fi_passed = fi_passed is not None and fi_passed != ""

    # 확정 규칙(2026-06-12) — config 키 **둘 다 부재** + spec 자동 산출값 존재
    # 시 spec 산출 FI stamp (config 하나라도 있으면 config 우선 — 기존 동작
    # 그대로). 둘 다 없고 spec도 없으면 기존 노란 마킹 (HDPDM01 — 무회귀).
    # warning 필수: 규칙·spec 파일명 명기 — DV 감사본 수기 402(재현 불가 stale
    # 카운트)와의 차이가 추적 가능해야 한다 (실측 위장 금지).
    # 라운드 108 MAJOR-1 — spec_fi_auto 분기는 미통과 N = 미실행 + Fail 혼합
    # (교차값)이므로 C90 비고에 breakdown을 보존해 미실행 블록의 Fail 위장 차단.
    spec_auto_breakdown: tuple[int, int] | None = None
    if not has_fi_total and not has_fi_passed and spec_fi_auto:
        try:
            _auto_total = int(spec_fi_auto.get("total"))  # type: ignore[arg-type]
            _auto_passed = int(spec_fi_auto.get("passed"))  # type: ignore[arg-type]
        except (TypeError, ValueError):
            _auto_total = _auto_passed = None  # type: ignore[assignment]
            if warnings is not None:
                warnings.append(
                    "[swutcr] UT201 FI spec 자동 산출값 비정수 — stamp skip, "
                    "노란 마킹 유지"
                )
        if _auto_total is not None and _auto_passed is not None:
            fi_total, fi_passed = _auto_total, _auto_passed
            has_fi_total = has_fi_passed = True
            _ne = int(spec_fi_auto.get("not_executed_blocks") or 0)
            _fl = int(spec_fi_auto.get("failed_blocks") or 0)
            spec_auto_breakdown = (_ne, _fl)
            # MINOR-6 — 실제 stamp 분기에서만 True (시트 전무/skip 시 False 유지).
            spec_fi_auto["stamped"] = True
            if warnings is not None:
                _spec_name = spec_fi_auto.get("spec_filename") or "(파일명 미상)"
                _rule = spec_fi_auto.get("rule") or "Test Method 'FI' 포함 TC 블록 수"
                warnings.append(
                    f"[swutcr] UT201 FI spec 자동 산출 — 규칙: {_rule}, "
                    f"spec={_spec_name}, total={_auto_total}/passed={_auto_passed}"
                    f" (미실행 {_ne}, fail {_fl})"
                )

    if has_fi_total:
        safe_write(ws, 85, 5, fi_total)
    else:
        mark_user_input_required(ws, 85, 5, hint="FI 실측 미제공")
    if has_fi_passed:
        safe_write(ws, 85, 6, fi_passed)
    else:
        mark_user_input_required(ws, 85, 6, hint="FI 실측 미제공")
    if has_fi_total and has_fi_passed:
        try:
            fi_failed = max(int(fi_total) - int(fi_passed), 0)
        except (TypeError, ValueError):
            fi_failed = 0
        safe_write(ws, 85, 7, "=E85-F85")
        safe_write(ws, 85, 8, "=E86")
        if fi_failed == 0:
            note = "해당 사항 없음 "
        elif spec_auto_breakdown is not None:
            # MAJOR-1 — spec 자동 산출 미통과는 미실행/Fail 혼합 — 'Fail TC
            # N건'으로 미실행 블록을 Fail로 위장하지 않는다 (breakdown 명기).
            note = (
                f"미통과 {fi_failed}건(미실행 {spec_auto_breakdown[0]}, "
                f"Fail {spec_auto_breakdown[1]})"
            )
        else:
            # config 분기(수기 실측) — 기존 문구 그대로.
            note = f"Fail TC {fi_failed}건"
        safe_write(ws, 90, 3, note)
    else:
        # 파생 셀(G85 수식/H85/90행 비고)도 fabricated 입력 기반 stamp 금지 —
        # placeholder 텍스트가 수식 operand가 되면 #VALUE! 가 되므로 수식 미기입.
        mark_user_input_required(ws, 90, 3, hint="FI 실측 미제공")
        if warnings is not None:
            missing = [
                key for key, present in (
                    ("fault_injection_total", has_fi_total),
                    ("fault_injection_passed", has_fi_passed),
                ) if not present
            ]
            warnings.append(
                "[swutcr] 2.UT201 fault injection 실측 미제공"
                f" ({', '.join(missing)}) — E85/F85/C90 사용자입력 마킹"
                " (함수 수 기반 fallback stamp 제거)"
            )


def _write_ut301(ws, meta: SwutcrBuildMeta, agg: dict[str, Any], cfg: dict[str, Any]) -> None:
    """Stamp Back-to-back test sheet with explicit not-applicable evidence state."""
    _write_swutcr_sheet_header(ws, meta, cfg)
    md = cfg.get("swutcr_metadata", {}) or {}
    enabled = str(md.get("ut301_enabled", "X")).strip().upper()
    note = md.get(
        "ut301_not_applicable_reason",
        "해당 사항 없음 - Back-to-back test evidence source not provided",
    )

    if enabled != "O":
        safe_write(ws, 75, 3, 0)
        safe_write(ws, 75, 5, 0)
        safe_write(ws, 75, 6, 0)
        safe_write(ws, 75, 7, 0)
        for col in (8, 9, 10):
            safe_write(ws, 75, col, "-")
        safe_write(ws, 80, 3, "N/A")
        safe_write(ws, 80, 5, "해당 사항 없음")
        safe_write(ws, 80, 8, note)
        safe_write(ws, 80, 12, "Summary에서 UT301 비적용(X) 처리")
        # W-6-①: 90:91행은 세로 머지된 라벨 헤더('SW Unit(함수)'/'미달성 사유'/
        # '대책'/'문장') — 91행에 쓰면 merge anchor redirect로 90행 라벨이 값으로
        # 덮어써진다. 라벨 보존을 위해 값은 헤더 아래 첫 값행(92행)에 기입.
        safe_write(ws, 92, 3, "N/A")
        safe_write(ws, 92, 5, "-")
        safe_write(ws, 92, 8, note)
        safe_write(ws, 92, 12, "추가 B2B evidence source 제공 시 재작성")


def _write_it801(ws, meta: SwutcrBuildMeta, cfg: dict[str, Any]) -> None:
    """Stamp heap memory analysis sheet and clear unsafe template defaults."""
    _write_swutcr_sheet_header(ws, meta, cfg)
    md = cfg.get("swutcr_metadata", {}) or {}
    safe_write(ws, 33, 4, md.get("heap_memory_tool", "mpatrol"))
    safe_write(ws, 34, 4, md.get("heap_memory_tool_version", "1.4.3"))
    safe_write(ws, 40, 3, md.get("heap_reference_document", "SwTP"))
    safe_write(ws, 41, 3, md.get("heap_reference_id", "SwTE_01"))

    leak_count = md.get("heap_memory_leak_count", 0)
    access_count = md.get("heap_access_violation_count", 0)
    safe_write(ws, 45, 5, leak_count)
    safe_write(ws, 46, 5, access_count)
    # W-6-②: 50행은 헤더('파일명'/'오류 내용'/'수정 여부') — 값은 헤더 아래
    # 첫 값행(51행)에 기입 (헤더 라벨 덮어쓰기 방지, +1 보정).
    safe_write(ws, 51, 3, "N/A")
    safe_write(ws, 51, 5, "해당 사항 없음")
    safe_write(ws, 51, 7, "O")
    safe_write(
        ws,
        51,
        10,
        md.get(
            "heap_detail_note",
            "Heap memory issue not detected in configured evidence; manual mpatrol log review required",
        ),
    )


def _write_btb_sheet(ws, meta: SwutcrBuildMeta, cfg: dict[str, Any]) -> None:
    """Stamp integration B2B template sheet with explicit not-applicable state."""
    md = cfg.get("swutcr_metadata", {}) or {}
    safe_write(ws, 4, 4, md.get("project", meta.project_id))
    safe_write(ws, 4, 11, md.get("mcu", ""))
    safe_write(ws, 5, 4, str(md.get("phase") or "").strip() or "DV")
    safe_write(ws, 5, 11, md.get("compiler", ""))
    safe_write(ws, 6, 4, md.get("product", meta.project_id))
    safe_write(ws, 6, 11, meta.test_engineer or md.get("tester", meta.author))
    safe_write(ws, 7, 4, md.get("debugger", ""))
    safe_write(ws, 7, 11, md.get("software_platform_ver", meta.release_sw_version))
    # W-6-③ (라운드 96-final QA fix): 10행은 머지 라벨 헤더(C10:D11 'SW 버전' /
    # E10:F11 'Test Period') — 10행에 쓰면 라벨이 값으로 덮어써진다. 값은 라벨
    # 아래 값행 12행(값칸 C12:D12 / E12:F12)에 기입. 시간값 G12:J12는 기존부터
    # 올바른 행이라 변경 없음.
    safe_write(ws, 12, 3, md.get("test_iteration", "0.1"))
    safe_write(ws, 12, 5, md.get("software_platform_ver", meta.release_sw_version))
    safe_write(ws, 12, 7, md.get("prepare_hours", 0))
    safe_write(ws, 12, 8, md.get("execution_hours", 0))
    safe_write(ws, 12, 9, md.get("review_hours", 0))
    safe_write(ws, 12, 10, "=SUM(G12:I12)")
    safe_write(ws, 32, 5, 0)
    safe_write(ws, 32, 8, 0)
    safe_write(ws, 32, 11, "-")
    safe_write(ws, 37, 3, "N/A")
    safe_write(ws, 37, 5, "해당 사항 없음")
    safe_write(ws, 37, 9, "통합 B2B evidence source not provided for SwUTCR")
    safe_write(ws, 37, 11, "SwIT/BTB 별도 근거 제공 시 재작성")
    safe_write(ws, 50, 3, "N/A")
    safe_write(ws, 50, 5, "-")
    safe_write(ws, 50, 7, "-")
    safe_write(ws, 50, 10, "해당 사항 없음")


def _write_swutcr_specific_sheets(
    wb: Workbook,
    meta: SwutcrBuildMeta,
    agg: dict[str, Any],
    cfg: dict[str, Any],
    warnings: list[str] | None = None,
    spec_fi_auto: dict[str, Any] | None = None,
) -> dict[str, Any]:
    written: dict[str, Any] = {}
    summary_ws = wb["Summary"] if "Summary" in wb.sheetnames else None
    if summary_ws is not None:
        _write_swutcr_metadata(summary_ws, meta, cfg)
        written["summary_metadata"] = True
    if "1.UT101" in wb.sheetnames:
        _write_ut101(wb["1.UT101"], meta, agg, cfg, warnings)
        written["ut101"] = True
    if "2.UT201" in wb.sheetnames:
        # C-5 배선 (라운드 96-final): FI 실측 미제공 시 _write_ut201이 emit하는
        # 사용자입력 마킹 warning이 산출물 warnings에 합류하도록 전달.
        # 확정 규칙(2026-06-12): spec_fi_auto = spec 산출 × 실행결과 교차값 —
        # config 키 부재 시 _write_ut201이 소비 (config 존재 시 무시).
        _write_ut201(wb["2.UT201"], meta, agg, cfg, warnings, spec_fi_auto)
        written["ut201"] = True
    if "3.UT301" in wb.sheetnames:
        _write_ut301(wb["3.UT301"], meta, agg, cfg)
        written["ut301"] = True
    if "21.IT801" in wb.sheetnames:
        _write_it801(wb["21.IT801"], meta, cfg)
        written["it801"] = True
    btb_ws = next((wb[name] for name in wb.sheetnames if "BTB" in name), None)
    if btb_ws is not None:
        _write_btb_sheet(btb_ws, meta, cfg)
        written["btb"] = True
    return written


def build_swutcr_report(
    session: SwUTSession,
    meta: SwutcrBuildMeta,
    template_bytes: bytes,
    *,
    deviation_cases: list[Any] | None = None,
    swuds_function_ids: set[str] | None = None,
    swuts_map: dict[str, Any] | None = None,
    hmr_html_bytes: bytes | None = None,
    spec_fi: dict[str, Any] | None = None,
) -> SwutcrBuildResult:
    """Build a SwUTCR xlsm from template and SwUT evidence data.

    Args:
        spec_fi: ``swut_sutr_spec_builder.extract_spec_fi_stats`` 결과 (확정
            규칙 2026-06-12 — UT201 FI spec 자동 산출). None이면 기존 동작
            (config 키 또는 노란 마킹). 존재 시 session 실행결과와 교차해
            ``_write_ut201`` 에 전달 — config 키가 있으면 config 우선.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for SwUTCR builder")

    validate_build_meta(
        meta.release_sw_version,
        meta.test_date,
        doc_id_sequence=meta.doc_id_sequence,
        test_engineer=meta.test_engineer,
        author=meta.default_author,
        approver=meta.default_approver or meta.approver_override,
        reviewer=meta.default_reviewer or meta.reviewer_override,
    )
    validate_xlsx_template_bytes(template_bytes, label="SwUTCR template")

    template_sha256_12 = hashlib.sha256(template_bytes).hexdigest()[:12]
    template_has_vba = has_vba_macros(template_bytes)
    warnings: list[str] = extract_warnings_from_session(session)
    warnings.extend(diagnose_asset_usage(
        swuts_map=swuts_map,
        c_function_map=session.c_function_map or None,
        swuds_function_map=session.swuds_function_map or None,
    ))
    if template_has_vba:
        warnings.append(
            "VBA macro execution NOT verified; open output xlsm in Excel before evidence use"
        )
        refs = inspect_vba_refs(template_bytes)
        if refs:
            warnings.append(f"VBA reference patterns found: {refs}; manual macro check required")

    from backend.services.excel_layout_resolver import inspect_swit_layout
    coverage_layout = inspect_swit_layout(template_bytes, "coverage")
    sitr_layout = inspect_swit_layout(template_bytes, "sitr")

    wb: Workbook = openpyxl.load_workbook(
        io.BytesIO(template_bytes),
        keep_vba=True,
        data_only=False,
    )
    is_swutcr_specific_template = any(
        name in wb.sheetnames for name in ("Summary", "1.UT101", "2.UT201")
    )
    if not is_swutcr_specific_template:
        warnings.extend([f"[coverage-layout] {w}" for w in coverage_layout.warnings])
        warnings.extend([f"[sutr-layout] {w}" for w in sitr_layout.warnings])

    agg = aggregate_session(session)
    if session.c_function_map:
        agg["c_function_map"] = session.c_function_map
    from backend.services.swut_coverage_aggregator import (
        _apply_template_swufn_order,
        _build_swuts_name_to_swufn_map,
        _compute_asil_distribution,
        _write_audit_log_sheet,
        _write_consistency_sheet,
        _write_coverage_sheet,
        _write_cover_sheet,
        _write_history_sheet,
        _write_test_summary_sheet,
        _write_traceability_sheet,
    )
    from backend.services.swut_sutr_aggregator import (
        _write_deviation,
        _write_test_log,
        _write_test_summary,
    )

    swuts_name_to_swufn = _build_swuts_name_to_swufn_map(swuts_map)
    if swuts_name_to_swufn:
        merged = dict(agg.get("function_name_to_swufn_from_suds") or {})
        merged.update(swuts_name_to_swufn)
        agg["function_name_to_swufn_from_suds"] = merged
        warnings.append(
            f"[swuts-map] unit_name->SwUFn mapping applied ({len(swuts_name_to_swufn)})"
        )

    _stamp_hmr_metrics(agg, hmr_html_bytes, warnings)
    _apply_template_swufn_order(wb, agg, warnings)
    agg["project_id"] = meta.project_id
    if meta.doc_filename_pattern:
        exception_filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version,
            date=short_date(meta.test_date),
        )
        agg["coverage_exception_note"] = exception_filename.split("_v", 1)[0]
    cfg: dict[str, Any] = getattr(meta, "project_config", {}) or {}
    swutcr_md = cfg.get("swutcr_metadata", {}) or {}
    raw_function_count = agg.get("function_count", 0) or len(agg.get("function_rows") or [])
    qualified_function_count = swutcr_md.get("qualified_function_total")
    if qualified_function_count not in (None, ""):
        try:
            qualified_function_count = int(qualified_function_count)
        except (TypeError, ValueError):
            warnings.append(
                "[swutcr] qualified_function_total is not an integer; raw function_count used"
            )
            qualified_function_count = None
        if qualified_function_count is not None:
            agg["swutcr_qualified_function_count"] = qualified_function_count
            if qualified_function_count != raw_function_count:
                warnings.append(
                    "[swutcr] qualified function total override applied: "
                    f"{qualified_function_count} (raw VectorCAST functions: {raw_function_count})"
                )

    asil_distribution, ids_by_asil, unmapped_fns = _compute_asil_distribution(
        agg.get("function_rows") or [],
        agg.get("function_asil_map") or {},
        function_asil_from_suds=agg.get("function_asil_from_suds"),
        component_asil_from_sds=agg.get("component_asil_from_sds"),
        function_asil_from_srs=agg.get("function_asil_from_srs"),
        function_name_to_swufn_from_suds=agg.get("function_name_to_swufn_from_suds"),
    )

    summary: dict[str, Any] = {
        "environments": len(session.environments),
        "total_tcs": agg.get("total_tcs", 0),
        "tested": agg.get("tested", 0),
        "passed": agg.get("passed", 0),
        "failed": agg.get("failed", 0),
        "function_rows": agg.get("function_count", 0),
        "swutcr_qualified_function_count": agg.get("swutcr_qualified_function_count"),
        "swutcr_raw_function_count": raw_function_count,
        "asil_distribution": asil_distribution,
        "asil_b_function_ids": ids_by_asil.get("B", []),
        "asil_c_function_ids": ids_by_asil.get("C", []),
        "asil_d_function_ids": ids_by_asil.get("D", []),
        "unmapped_function_names": unmapped_fns,
        "template_sha256_12": template_sha256_12,
        "build_timestamp": meta.build_timestamp,
        "swuts_name_to_swufn_used": len(swuts_name_to_swufn),
    }
    incomplete_sheets: list[str] = []

    cover_ws = _find_sheet(wb, "cover")
    if cover_ws is not None:
        _write_cover_sheet(cover_ws, meta, out_warnings=warnings, layout=coverage_layout)
    else:
        warnings.append("Cover sheet not found")
        incomplete_sheets.append("Cover")

    if not is_swutcr_specific_template:
        summary_ws = _find_sheet(wb, "test summary")
        if summary_ws is not None:
            _write_test_summary_sheet(
                summary_ws, meta, agg, out_warnings=warnings,
                layout=coverage_layout, summary=summary,
            )
            _write_test_summary(
                summary_ws, meta, agg, out_warnings=warnings,
                layout=sitr_layout, summary=summary,
            )
        else:
            warnings.append("Test Summary sheet not found")
            incomplete_sheets.append("Test Summary")

        trace_ws = _find_sheet(wb, "traceability")
        if trace_ws is not None:
            count = _write_traceability_sheet(
                trace_ws, session, out_warnings=warnings,
                layout=coverage_layout, agg=agg,
            )
            summary["traceability_o_cells"] = count
            if count == 0:
                incomplete_sheets.append(trace_ws.title)
        else:
            warnings.append("Traceability sheet not found")
            incomplete_sheets.append("Traceability")

        cons_ws = _find_sheet(wb, "consistency")
        if cons_ws is not None:
            count = _write_consistency_sheet(
                cons_ws, session,
                swuds_function_ids=swuds_function_ids,
                out_warnings=warnings,
                agg=agg,
            )
            summary["consistency_self_check_rows"] = count
            summary["consistency_swuds_compared"] = swuds_function_ids is not None
            if swuds_function_ids is None:
                incomplete_sheets.append(f"{cons_ws.title} (SwUDS compare partial)")
        else:
            warnings.append("Consistency sheet not found")

        coverage_ws = _find_sheet(
            wb, "coverage",
            exclude=("traceability", "consistency", "summary"),
        )
        if coverage_ws is not None:
            count = _write_coverage_sheet(
                coverage_ws, agg, layout=coverage_layout,
                out_warnings=warnings,
                c_function_map=session.c_function_map or None,
            )
            summary["coverage_rows_written"] = count
        else:
            warnings.append("Coverage sheet not found")
            incomplete_sheets.append("Coverage")

        deviation_ws = _find_sheet(wb, "deviation")
        if deviation_ws is not None:
            count = _write_deviation(deviation_ws, deviation_cases or [], warnings)
            summary["deviation_rows_written"] = count

        test_log_ws = _find_sheet(wb, "test log") or _find_sheet(wb, "test result")
        if test_log_ws is not None:
            count = _write_test_log(
                test_log_ws,
                session,
                function_asil_map=agg.get("function_asil_map") or {},
                out_warnings=warnings,
                layout=sitr_layout,
                swuts_map=swuts_map,
                c_function_map=session.c_function_map or None,
            )
            summary["test_log_rows_written"] = count
            if count == 0:
                incomplete_sheets.append(test_log_ws.title)
        else:
            warnings.append("Test Log sheet not found")

    history_ws = _find_sheet(wb, "history")
    if history_ws is not None:
        rows = build_release_history_row(meta, doc_kind="SwUTCR", out_warnings=warnings)
        # 라운드 96-final-2 — F5/G5 placeholder 보존은 _write_history_sheet 내부로
        # 호이스트(전 빌더 단일 적용, H5). 별도 사전 보정 불필요.
        count = _write_history_sheet(history_ws, rows, out_warnings=warnings)
        summary["history_rows_written"] = count
        if count == 0:
            incomplete_sheets.append("History")
    else:
        warnings.append("History sheet not found")

    if is_swutcr_specific_template and "AuditLog" in wb.sheetnames:
        del wb["AuditLog"]
        summary["audit_log_sheet_removed_for_template_fidelity"] = True
    elif not is_swutcr_specific_template and "AuditLog" not in wb.sheetnames:
        audit_ws = wb.create_sheet("AuditLog")
        count = _write_audit_log_sheet(audit_ws, meta, summary, agg, session, warnings)
        summary["audit_log_rows_written"] = count
        summary["audit_log_sheet_added"] = True

    # 확정 규칙(2026-06-12) — UT201 FI spec 산출 × 실행결과 교차. spec_fi
    # 부재 시 None → _write_ut201 기존 분기 (config 키 / 노란 마킹) 그대로.
    # 라운드 108 — 교차 보수화 warning(상충/공허/중복)은 산출물 warnings 합류.
    # summary['ut201_fi_auto'].stamped 는 _write_ut201 실제 stamp 시에만 True
    # (MINOR-6 — 2.UT201 시트 전무 템플릿에서 산출 위장 차단).
    spec_fi_auto: dict[str, Any] | None = None
    if spec_fi:
        spec_fi_auto = _cross_spec_fi_with_session(spec_fi, session, warnings)
        summary["ut201_fi_auto"] = spec_fi_auto

    if any(name in wb.sheetnames for name in ("Summary", "1.UT101", "2.UT201")):
        summary["swutcr_specific_written"] = _write_swutcr_specific_sheets(
            wb, meta, agg, cfg, warnings, spec_fi_auto,
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    wb.close()

    if meta.doc_filename_pattern:
        filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version,
            date=short_date(meta.test_date),
        )
    else:
        filename = (
            f"({meta.project_id}_DV_SwUTCR) Software Unit Test Comprehensive Result_"
            f"v{meta.release_sw_version}_{short_date(meta.test_date)}_R.xlsm"
        )

    return SwutcrBuildResult(
        ok=True,
        xlsm_io=out,
        filename=filename,
        warnings=warnings,
        incomplete_sheets=incomplete_sheets,
        vba_macros_preserved=template_has_vba,
        summary=summary,
    )


__all__ = [
    "SwutcrBuildMeta",
    "SwutcrBuildResult",
    "build_swutcr_report",
]
