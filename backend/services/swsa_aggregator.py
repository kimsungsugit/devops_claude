"""SwSA aggregator — 회사 SwSA 템플릿(.xlsm)에 정적분석 결과 채우기.

SwUT/SwIT template-copy 전략 동일: ``openpyxl.load_workbook(keep_vba=True)`` 로
회사 양식을 in-place 로 열어(셀병합/폰트/매크로 보존) 라벨 앵커로 셀만 덮어쓴다.

채우는 시트 (템플릿에 **존재하는 것만**):
  - Cover (META): doc_id/version/status/date/author
  - Summary 헤더 (META): project/phase/platform/asil/compiler/mcu
  - ST101 (AUTO): QAC results_data.xml M3CM → 위반 룰 개수 / 총 위반. 예외처리/
    수정대상은 로그 부재 → 노란 표시.
  - ST201 (AUTO): QAC HMR 메트릭 + PMD 중복 → Test-Info + (best-effort 결과)
  - ST1101 (AUTO): QAC HKCCM → Test-Info + 총 위반
  - 모든 실행 ST 시트: Test-Information(분석차수/SW Ver/Tester/Debugger)

판정 셀(예외처리/수정대상/P-F)은 QAC 에서 도출 불가 → ``write_value_or_mark`` 로
값 없으면 노란 '사용자 입력 필요'. extraction_failed 시에도 0 stamp 대신 노란 표시.

ISO 26262: 산출물은 evidence 'auto-generated draft'. reviewer 검토 의무.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any, List, Optional

from openpyxl import load_workbook

from backend.services.excel_template_utils import (
    has_vba_macros,
    is_formula_cell,
    mark_user_input_fill_only,
    safe_write,
    write_value_or_mark,
)
from backend.services.swsa_layout_resolver import (
    detect_st_layout,
    find_label_row,
    find_value_target,
)
from backend.services.swsa_meta import SwsaBuildMeta
from backend.services.swsa_pmd_parser import PmdResult
from backend.services.swsa_qac_xml_parser import (
    MISRA_MANDATORY,
    MISRA_REQUIRED,
    QacXmlResult,
)
from backend.services.swsa_st201_binner import (
    St201Result,
    bin_values_into_bands,
    metric_item_for_name,
    parse_band_predicate,
)

__all__ = ["SwsaBuildResult", "build_swsa_report"]

# 양식 시트명 (존재할 때만 채움)
SHEET_COVER = "Cover"
SHEET_HISTORY = "History"
SHEET_SUMMARY = "Summary"
SHEET_ST101 = "1.ST101"
SHEET_ST201 = "2.ST201"
SHEET_ST1101 = "11.ST1101"


@dataclass
class SwsaBuildResult:
    xlsm_io: io.BytesIO
    sheets_filled: List[str] = field(default_factory=list)
    filled_cells: int = 0
    user_input_cells: int = 0
    warnings: List[str] = field(default_factory=list)
    vba_preserved: bool = False


def _stamp(ws: Any, label: str, value: Any, *, max_row: int = 40,
           label_col: Optional[int] = None) -> bool:
    """라벨 옆 셀에 값 기입 (merge-aware). 라벨 없으면 False."""
    tgt = find_value_target(ws, label, max_row=max_row, label_col=label_col)
    if tgt is None:
        return False
    return safe_write(ws, tgt[0], tgt[1], value)


def _stamp_or_mark(ws: Any, label: str, value: Any, *, hint: str = "", max_row: int = 120,
                   label_col: Optional[int] = None) -> Optional[bool]:
    """라벨 옆 셀에 값/노란표시. 라벨 없으면 None."""
    tgt = find_value_target(ws, label, max_row=max_row, label_col=label_col)
    if tgt is None:
        return None
    return write_value_or_mark(ws, tgt[0], tgt[1], value, hint=hint)


def _write_value_safe(ws: Any, r: int, c: int, value: Any, res: "SwsaBuildResult",
                      what: str = "") -> bool:
    """수식 셀은 보존(미기입+경고), 비-수식 입력 셀만 값 기입 (C1/C2/C3 fix).

    v0.11 양식의 summary 셀은 detail 표를 참조하는 `=COUNTIF(...)` 등 audit
    교차검증 수식이다. literal 로 덮으면 summary↔detail 연결이 끊긴다. 또한 수식
    operand 셀에 string 을 쓰면 `#VALUE!` 가 된다. 따라서 수식 셀은 건드리지 않고
    경고만 남긴다 (v0.11 detail 표 채우기는 후속).
    """
    from openpyxl.utils import get_column_letter
    if is_formula_cell(ws, r, c):
        res.warnings.append(
            f"{ws.title}: 수식 셀 보존 — {what} 미기입 ({get_column_letter(c)}{r})"
        )
        return False
    if safe_write(ws, r, c, value):
        res.filled_cells += 1
        return True
    return False


# ─────────────────────────── Cover ───────────────────────────
def _write_cover(ws: Any, meta: SwsaBuildMeta, res: SwsaBuildResult) -> None:
    pairs = [
        ("Document ID", meta.doc_id),
        ("Version", meta.doc_version),
        ("Status", meta.doc_status),
        ("Date", meta.test_date),
        ("Author", meta.author),
    ]
    n = 0
    for label, val in pairs:
        # doc-block 라벨은 col C(3) — 사인오프 헤더 I2 'Author' 충돌 회피
        if _stamp(ws, label, val, max_row=40, label_col=3):
            n += 1
    # rank8: 사인오프 블록 (I2 Author/J2 Reviewer/K2 Approver 헤더 → 아래 행 이름)
    for label, val in (("Reviewer", meta.reviewer), ("Approver", meta.approver)):
        if not val:
            continue
        pos = find_label_row(ws, label, max_row=8)
        if pos and safe_write(ws, pos[0] + 1, pos[1], val):
            n += 1
    res.filled_cells += n
    if n:
        res.sheets_filled.append(SHEET_COVER)


# ─────────────────────────── History ───────────────────────────
def _write_history(ws: Any, meta: SwsaBuildMeta, res: SwsaBuildResult) -> None:
    """Revision History 최신 행 기입 (Version/Date/Description/Author/Reviewer/Approver).

    헤더(B='Version') 다음 데이터 행(템플릿 placeholder 'v0.10/2X.XX.XX/-Draft/XXX')에
    실제 빌드 정보를 기입. reviewer/approver 빈 값은 노란 표시 (Cover 정책 동일).
    """
    hdr = find_label_row(ws, "Version", max_row=20, label_col=2)
    if hdr is None:
        res.warnings.append(f"{ws.title}: 'Version' 헤더 미발견 — History 미기입")
        return
    hr, vcol = hdr  # B(Version) 컬럼
    # rank7 fix: 첫 데이터 행을 무조건 덮으면 기존 이력(DV: v0.10/v1.01/v1.02)을
    # 클로버. 첫 빈 행 또는 placeholder('2X.XX.XX'/'XXX') 행을 찾아 기입, 모두
    # 채워졌으면 마지막 다음 행에 append.
    r = hr + 1
    last_filled = hr
    for rr in range(hr + 1, hr + 31):
        bval = ws.cell(rr, vcol).value
        cval = ws.cell(rr, vcol + 1).value      # Date
        if bval is None or (isinstance(bval, str) and not bval.strip()):
            r = rr
            break
        if isinstance(cval, str) and "X" in cval.upper():  # placeholder 날짜
            r = rr
            break
        last_filled = rr
    else:
        r = last_filled + 1  # 모든 행이 실제 이력 → append
    # 컬럼: B=Version C=Date D=Description E=Author F=Reviewer G=Approver
    safe_write(ws, r, vcol, meta.doc_version)
    safe_write(ws, r, vcol + 1, meta.test_date)
    desc = meta.history_description or f"- {meta.doc_version} 정적분석 작성"
    safe_write(ws, r, vcol + 2, desc)
    safe_write(ws, r, vcol + 3, meta.author)
    res.filled_cells += 4
    # Reviewer / Approver — 빈 값이면 노란 (audit 필수)
    for off, val in ((4, meta.reviewer), (5, meta.approver)):
        if write_value_or_mark(ws, r, vcol + off, val, hint="audit 필수"):
            res.filled_cells += 1
        else:
            res.user_input_cells += 1
    res.sheets_filled.append(SHEET_HISTORY)


# ─────────────────────────── Summary ───────────────────────────
def _write_summary_header(ws: Any, meta: SwsaBuildMeta, res: SwsaBuildResult) -> None:
    asil = meta.asil_level.replace("ASIL", "").strip() or meta.asil_level  # 'ASIL A' → 'A'
    pairs = [
        ("Project", meta.project_id),
        ("Phase", meta.phase),
        ("Software Platform Ver.", meta.platform_version),
        ("Product", meta.product),
        ("검증 대상", meta.verification_target),
        ("ASIL 등급", asil),
        ("Complier", meta.compiler),   # 양식 철자 그대로 (Complier)
        ("MCU", meta.mcu),
    ]
    n = 0
    for label, val in pairs:
        # Summary 정보블록 라벨은 col B(2) — 검증대상 값 'MCU' 가 'MCU' 라벨과
        # 충돌하던 문제 회피 (값은 col E 에 쓰이므로 label_col=2 로 라벨만 매칭)
        r = _stamp_or_mark(ws, label, val, hint=label, max_row=14, label_col=2)
        if r is not None:
            n += 1
            if r is False:
                res.user_input_cells += 1
    res.filled_cells += n
    if n:
        res.sheets_filled.append(SHEET_SUMMARY)


# ─────────────────────── ST 공통 Test-Info ───────────────────────
def _write_st_test_info(ws: Any, meta: SwsaBuildMeta, res: SwsaBuildResult) -> bool:
    """모든 실행 ST 시트의 Test-Information 헤더 채우기. 라벨 앵커로 버전 흡수."""
    lay = detect_st_layout(ws)
    if lay.missing:
        res.warnings.append(f"{ws.title}: Test-Info 라벨 미발견 {lay.missing}")
    values = {
        "analysis_round": meta.analysis_round,
        "sw_version": meta.release_sw_version,
        "tester": meta.tester_name,
        "debugger": meta.debugger,
    }
    n = 0
    for key, (r, c) in lay.test_info.items():
        val = values.get(key, "")
        if write_value_or_mark(ws, r, c, val, hint=key):
            n += 1
        else:
            res.user_input_cells += 1
    res.filled_cells += n
    return bool(lay.test_info)


def _find_summary_rows(ws: Any, header_row: int, label_col_max: int = 4) -> dict:
    """ST101 Test Summary 표의 Mandatory/Required/Total 행 탐색.

    v0.11 은 3행(Mandatory/Required/Total), v0.10 은 단일행 → Total=header+1 default.
    """
    rows = {}
    for r in range(header_row + 1, header_row + 6):
        for c in range(1, label_col_max + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                vs = v.strip()
                if vs in ("Mandatory", "Required", "Total"):
                    rows[vs] = r
    if not rows:
        rows["Total"] = header_row + 1
    return rows


# ─────────────────────────── ST101 ───────────────────────────
def _write_st101(ws: Any, meta: SwsaBuildMeta, qac: Optional[QacXmlResult], res: SwsaBuildResult) -> None:
    _write_st_test_info(ws, meta, res)
    # 코딩룰 버전
    _stamp(ws, "코딩룰 버전", meta.misra_rule_version, max_row=12)

    misra = qac.misra if qac else None
    failed = (qac is None) or qac.extraction_failed or (misra is None)

    # rank5: deviation(제외) 노출 — 산출물은 active 만 표시하므로 제외 비율이 높으면
    # audit deviation report 의무를 경고로 surface (excluded = total - active).
    if misra is not None and misra.total > 0:
        pct = misra.excluded / misra.total * 100
        if pct >= 70:
            res.warnings.append(
                f"{ws.title}: 제외(deviation) {misra.excluded}/{misra.total}건 "
                f"({pct:.0f}%) — 산출물은 active {misra.active}만 표시, deviation report 의무"
            )

    # Test Summary 표: 헤더 '총 위반 건수' / '위반 룰 개수' 라벨 앵커
    h_total = find_label_row(ws, "총 위반 건수", max_row=120)
    h_rules = find_label_row(ws, "위반 룰 개수", max_row=120)
    h_exc = find_label_row(ws, "예외 처리 항목 수", max_row=120)
    if h_total is None:
        res.warnings.append(f"{ws.title}: '총 위반 건수' 헤더 미발견 — Summary 표 미기입")
    else:
        hr = h_total[0]
        srows = _find_summary_rows(ws, hr)
        # 카테고리별 값 (v0.11 Mandatory/Required/Total) — v0.10 은 Total 만
        cat_active = {}
        cat_rules = {}
        if misra is not None:
            cat_active["Total"] = misra.active
            cat_rules["Total"] = misra.distinct_rules()
            cat_active["Mandatory"] = misra.category(MISRA_MANDATORY).active
            cat_active["Required"] = misra.category(MISRA_REQUIRED).active
            cat_rules["Mandatory"] = sum(
                1 for r in misra.leaf_rules if r.severity == MISRA_MANDATORY and r.active > 0)
            cat_rules["Required"] = sum(
                1 for r in misra.leaf_rules if r.severity == MISRA_REQUIRED and r.active > 0)
            # H3: 표에 Mandatory/Required 행만 있고 둘 합이 Total과 다르면(Common 등) 경고
            if {"Mandatory", "Required"} & set(srows) and "Total" in srows:
                man_req = cat_active["Mandatory"] + cat_active["Required"]
                if man_req != misra.active:
                    res.warnings.append(
                        f"{ws.title}: 카테고리 합(Man+Req={man_req}) != 총위반({misra.active}) "
                        f"— 기타 카테고리(Common 등) 행 누락, audit 검토 필요"
                    )
        for cat, r in srows.items():
            # 총 위반 건수 / 위반 룰 개수 — 수식 셀(v0.11 =COUNTIF) 보존, 비-수식만 기입
            if failed:
                # 추출 실패: operand 파괴 없이 노란 배경만 (string 금지 → #VALUE! 방지)
                mark_user_input_fill_only(ws, r, h_total[1])
                res.user_input_cells += 1
                if h_rules is not None:
                    mark_user_input_fill_only(ws, r, h_rules[1])
                    res.user_input_cells += 1
            else:
                _write_value_safe(ws, r, h_total[1], cat_active.get(cat, 0), res, "총 위반 건수")
                if h_rules is not None:
                    _write_value_safe(ws, r, h_rules[1], cat_rules.get(cat, 0), res, "위반 룰 개수")
            # 예외 처리 항목 수 — 수식 operand(H77=D77-F77) 이므로 텍스트 금지, 노란 배경만
            if h_exc is not None:
                mark_user_input_fill_only(ws, r, h_exc[1])
                res.user_input_cells += 1

    # Test Environment: 도구명 / Version
    if misra is not None and qac is not None:
        _stamp(ws, "도구명", "Helix QAC", max_row=80)
        ver = qac.helix_qac_version.replace("Helix QAC", "").split("(")[0].strip()
        if ver:
            _stamp(ws, "Version", ver, max_row=80)

    # v0.11 5.1 Rule Violation detail (rank3) — 룰행 J/K 채우면 summary 수식 재계산
    if not failed and misra is not None:
        _write_rule_violation_detail(ws, misra.leaf_rules, res)

    res.sheets_filled.append(SHEET_ST101)


def _detect_app_boot(leaf_rules: List[Any]) -> tuple:
    """leaf 룰의 per_module 키에서 APP/BOOT prefix 감지. 단일 모듈이면 (mod, None)."""
    mods: set = set()
    for lr in leaf_rules:
        mods.update(lr.per_module.keys())
    app = next((m for m in mods if m.upper().startswith("APP")), None)
    boot = next((m for m in mods if m.upper().startswith("BOOT")), None)
    if app is None and mods:
        app = sorted(mods)[0]
    return app, boot


def _norm_rid(rid: str) -> str:
    """룰 ID 정규화 — 'C-INT-002'/'Rule-8.6' 의 'C-'/'Rule-' prefix 제거.

    ST101 detail 은 C='Rule-8.6'(parser 동일), ST1101 detail 은 D='INT-002'
    (parser 'C-INT-002') 라 prefix 차이를 흡수해 매칭한다.
    """
    s = (rid or "").strip()
    parts = s.split("-", 1)
    if len(parts) == 2 and parts[0] in ("C", "Rule"):
        return parts[1]
    return s


def _write_rule_violation_detail(ws: Any, leaf_rules: List[Any], res: SwsaBuildResult) -> None:
    """v0.11 detail 표: 템플릿 룰 행(Rule ID 열)에 J=APP/K=BOOT active 기입.

    ST101(C=Rule ID)/ST1101(D=Rule ID) 모두 — 'Rule ID' 헤더로 ID 열을 동적 탐색,
    '위반 건수' 헤더로 J(APP)/K(BOOT) 열을 탐색. prefix 정규화로 'INT-002'↔'C-INT-002'
    매칭. v0.10(룰 목록 미보유) → 헤더 미발견 graceful skip. B(=ROW)/L(=비율) 수식은
    보존하고 J/K 입력만 채운다 → summary COUNTIF/SUM 수식 자동 재계산. 0 은 공란.
    """
    j_hdr = find_label_row(ws, "위반 건수", max_row=220)
    if j_hdr is None:
        return
    hr, jcol = j_hdr
    kcol = jcol + 1  # BOOT (J=APP / K=BOOT 서브헤더)
    id_hdr = find_label_row(ws, "Rule ID", max_row=hr + 6)
    idcol = id_hdr[1] if id_hdr else 3  # default C
    app, boot = _detect_app_boot(leaf_rules)
    by_id: dict = {}
    for lr in leaf_rules:
        by_id[lr.rule_id.strip()] = lr
        by_id.setdefault(_norm_rid(lr.rule_id), lr)
    n = 0
    matched = 0
    for r in range(hr + 1, hr + 200):
        cval = ws.cell(r, idcol).value
        if not isinstance(cval, str) or not cval.strip():
            continue
        key = cval.strip()
        lr = by_id.get(key) or by_id.get(_norm_rid(key))
        if lr is None:
            continue
        matched += 1
        app_v = lr.active_for(app) if app else lr.active
        boot_v = lr.active_for(boot) if boot else 0
        if app_v and not is_formula_cell(ws, r, jcol) and safe_write(ws, r, jcol, app_v):
            n += 1
        if boot_v and not is_formula_cell(ws, r, kcol) and safe_write(ws, r, kcol, boot_v):
            n += 1
    res.filled_cells += n
    if matched:
        res.warnings.append(
            f"{ws.title}: v0.11 detail — {matched} 룰 매칭, {n}셀(J/K) 기입 (summary 수식 재계산)"
        )


# ─────────────────────────── ST201 ───────────────────────────
def _write_st201(ws: Any, meta: SwsaBuildMeta, st201: Optional[St201Result],
                 pmd: Optional[PmdResult], res: SwsaBuildResult) -> None:
    """ST201 Test Summary Report — 템플릿 주도 메트릭 밴드 채우기.

    템플릿 'Test Summary Report' 표를 스캔: D열 메트릭명 → MatrixItem, E열 밴드
    라벨 → 술어. 함수값을 그 밴드로 binning 해 F열(함수 개수)에 기입. H(예외처리)/
    J(수정대상=F-H)/L(결과)은 양식 수식이 자동 계산. 버전(v0.10/v0.11)·밴드 quirk
    무관 (라벨 그대로 파싱). 무소스 메트릭(Recursion/Stress)은 skip + 경고.
    """
    _write_st_test_info(ws, meta, res)
    # 'Test Summary Report' 섹션을 먼저 찾아 그 아래의 '함수 개수' 헤더만 매칭
    # (상단 Result 블록의 '함수 개수' I4 오매칭 방지).
    sec_row = 1
    for rr in range(1, min(ws.max_row, 140) + 1):
        bv = ws.cell(rr, 2).value
        if isinstance(bv, str) and "Test Summary Report" in bv:
            sec_row = rr
            break
    hdr = find_label_row(ws, "함수 개수", max_row=140, min_row=sec_row)
    if hdr is None:
        res.warnings.append(f"{ws.title}: Test Summary '함수 개수' 헤더 미발견 — 표 미기입")
        res.sheets_filled.append(SHEET_ST201)
        return
    hr, fcol = hdr  # F열(count) = 헤더 컬럼

    # C4: v0.11 양식은 F=APP / G=BOOT 분리. 현재는 병합 total 을 F 에 기입하므로
    # 분리 양식이면 투명성 경고 (모듈별 분리 기입은 후속 — St201Result.module 활용).
    sub = ws.cell(hr + 1, fcol).value
    sub_next = ws.cell(hr + 1, fcol + 1).value
    if isinstance(sub, str) and sub.strip().upper() == "APP" and \
            isinstance(sub_next, str) and sub_next.strip().upper() == "BOOT":
        res.warnings.append(
            f"{ws.title}: F=APP/G=BOOT 분리 양식 — 병합 total 을 F열에 기입 "
            "(모듈별 분리 미구현, G열 audit 검토 필요)"
        )

    filled = 0
    skipped: List[str] = []
    partial: List[str] = []

    def _flush(item: Any, name: str, rows: List[tuple]) -> None:
        nonlocal filled
        if not rows:
            return
        labels = [lbl for (_r, lbl) in rows]
        vals: Optional[List[int]] = None
        if item is not None and st201 is not None:
            vals = st201.values_for(item)
        elif "duplicat" in name.lower() and pmd is not None:
            vals = [b.lines for b in pmd.blocks]
        if not vals:
            # rank12: 무소스 메트릭(Recursion/Stress)은 F 밴드 셀을 노란 표시 —
            # 템플릿 잔여/수식 결과가 '검증됨'으로 오인되지 않도록 audit 신호.
            skipped.append(name.strip().replace("\n", " ")[:28])
            for (rr, _lbl) in rows:
                if not is_formula_cell(ws, rr, fcol):
                    mark_user_input_fill_only(ws, rr, fcol)
                    res.user_input_cells += 1
            return
        counts = bin_values_into_bands(vals, labels)
        # audit 투명성: 일부 값이 템플릿 밴드 밖이면(예: nesting=0 in '1~10') 기록
        if sum(counts) < len(vals):
            partial.append(f"{name.strip().replace(chr(10), ' ')[:20]}({sum(counts)}/{len(vals)})")
        for (rr, _lbl), cnt in zip(rows, counts):
            if safe_write(ws, rr, fcol, cnt):
                filled += 1

    current_item: Any = None
    current_name = ""
    pending: List[tuple] = []
    r = hr + 1
    while r <= min(ws.max_row, hr + 80):
        bval = ws.cell(r, 2).value
        if isinstance(bval, str) and bval.strip() == "Total":
            break
        dval = ws.cell(r, 4).value
        eval_ = ws.cell(r, 5).value
        if isinstance(dval, str) and dval.strip():
            _flush(current_item, current_name, pending)
            current_item = metric_item_for_name(dval)
            current_name = dval
            pending = []
        if isinstance(eval_, str) and parse_band_predicate(eval_) is not None:
            pending.append((r, eval_))
        r += 1
    _flush(current_item, current_name, pending)

    if skipped:
        res.warnings.append(f"{ws.title}: 무소스 메트릭 skip(수동 입력) {skipped}")
    if partial:
        res.warnings.append(f"{ws.title}: 일부 함수값이 템플릿 밴드 밖 {partial}")
    res.filled_cells += filled
    res.sheets_filled.append(SHEET_ST201)


# ─────────────────────────── ST1101 ───────────────────────────
def _write_st1101(ws: Any, meta: SwsaBuildMeta, qac: Optional[QacXmlResult], res: SwsaBuildResult) -> None:
    _write_st_test_info(ws, meta, res)
    _stamp(ws, "코딩룰 버전", meta.secure_rule_version, max_row=12)
    secure = qac.secure if qac else None
    failed = (qac is None) or (qac is not None and qac.extraction_failed) or (secure is None)
    # C3: Result 블록 헤더(J4=$E$79 수식) 오기입 방지 — Test Summary 섹션 앵커 후
    # 'Test Summary Report' 아래의 '총 위반 건수' 데이터 행만 타깃.
    sec_row = 1
    for rr in range(1, min(ws.max_row, 140) + 1):
        bv = ws.cell(rr, 2).value
        if isinstance(bv, str) and "Test Summary Report" in bv:
            sec_row = rr
            break
    h_total = find_label_row(ws, "총 위반 건수", max_row=140, min_row=sec_row)
    if h_total is None:
        res.warnings.append(f"{ws.title}: Test Summary '총 위반 건수' 헤더 미발견 — 미기입")
        res.sheets_filled.append(SHEET_ST1101)
        return
    srows = _find_summary_rows(ws, h_total[0])
    total_row = srows.get("Total", h_total[0] + 1)
    if failed or secure is None:
        mark_user_input_fill_only(ws, total_row, h_total[1])
        res.user_input_cells += 1
    else:
        _write_value_safe(ws, total_row, h_total[1], secure.active, res, "ST1101 총 위반")
        # rank4: 5.1 detail (D=Rule ID 'INT-002') J=APP/K=BOOT → summary 수식 재계산
        _write_rule_violation_detail(ws, secure.leaf_rules, res)
    res.sheets_filled.append(SHEET_ST1101)


def build_swsa_report(
    template_bytes: bytes,
    meta: SwsaBuildMeta,
    *,
    qac_xml: Optional[QacXmlResult] = None,
    st201: Optional[St201Result] = None,
    pmd: Optional[PmdResult] = None,
) -> SwsaBuildResult:
    """SwSA 템플릿(.xlsm) → 채워진 xlsm BytesIO.

    Args:
        template_bytes: 회사 SwSA 양식 .xlsm 바이트.
        meta: SwsaBuildMeta.
        qac_xml/st201/pmd: 파싱된 결과 (None 이면 해당 시트 노란 표시).

    Returns:
        SwsaBuildResult (xlsm_io 는 keep_vba 보존 스트림).
    """
    vba = has_vba_macros(template_bytes)
    wb = load_workbook(io.BytesIO(template_bytes), data_only=False, keep_vba=vba)
    res = SwsaBuildResult(xlsm_io=io.BytesIO(), vba_preserved=vba)
    sheets = set(wb.sheetnames)

    if SHEET_COVER in sheets:
        _write_cover(wb[SHEET_COVER], meta, res)
    if SHEET_HISTORY in sheets:
        _write_history(wb[SHEET_HISTORY], meta, res)
    if SHEET_SUMMARY in sheets:
        _write_summary_header(wb[SHEET_SUMMARY], meta, res)
    if SHEET_ST101 in sheets:
        _write_st101(wb[SHEET_ST101], meta, qac_xml, res)
    if SHEET_ST201 in sheets:
        _write_st201(wb[SHEET_ST201], meta, st201, pmd, res)
    if SHEET_ST1101 in sheets:
        _write_st1101(wb[SHEET_ST1101], meta, qac_xml, res)

    # 존재하지 않는 AUTO 시트 안내 (graceful)
    for sn in (SHEET_ST1101,):
        if sn not in sheets:
            res.warnings.append(f"{sn} 시트 없음 (템플릿 버전에 미포함) — graceful skip")

    # rank9: 파서 parse_warnings(unbinned/extraction/구버전 등)를 산출물 result 로
    # 전파 (audit reviewer 가 X-SwSA-Warnings 로 인지). 이전엔 silent 손실.
    for src, prefix in ((qac_xml, "QAC"), (st201, "HMR"), (pmd, "PMD")):
        for w in getattr(src, "parse_warnings", []) or []:
            res.warnings.append(f"[{prefix}] {w}")

    wb.save(res.xlsm_io)
    res.xlsm_io.seek(0)
    return res
