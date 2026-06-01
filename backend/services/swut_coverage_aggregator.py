"""SwUT Coverage Report v3.01 xlsx 빌더.

기존 v3.01 xlsx 템플릿을 BytesIO로 로드 → 셀 치환 → bytes 반환.
스타일/머지셀/색상 100% 보존 (template-copy 전략, 사용자 의사결정).

## 시트 매핑 (6 시트)

| 시트 | 출처 |
|------|------|
| Cover | meta + dialog (Doc ID / Project / ASIL / Author / Approver) |
| History | meta + git log 또는 사람 입력 |
| Test Summary | SwUTSession aggregate (Test Date/Engineer/Final/추적성/Stmt/Branch/MCDC) |
| 1.Traceability | TestCaseData에서 TC ID × Function ID O/X 매트릭스 |
| 2.Consistency | SwUDS↔SwUTS 정합성 — 본 라운드 placeholder (SwUDS docx 파싱 미연결) |
| 3. Coverage | per-function Statement/Branch/Exception (FunctionCoverage) |

## ISO 26262 Tool Qualification

- 출력 xlsx의 Cover 시트에 `[AUTO]` 라벨 부착하지 않음 (회사 표준 포맷 유지).
- 빌더 응답 메타에 `is_auto_generated=True` + `needs_review=True` 명시.
- ASIL A 한정 draft. B/C/D는 manual 재검토 의무 (모듈 docstring).
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from typing import Any

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

from backend.services.excel_template_utils import (
    BLANK_MARKUP,
    build_release_history_row,
    force_write_cell,
    mark_asil_a_function,
    mark_asil_b_function,
    mark_asil_c_function,
    mark_asil_d_function,
    mark_asil_qm_function,
    mark_fail_cell,
    mark_user_input_required,
    safe_write,
    short_date,
    validate_build_meta,
    validate_xlsx_template_bytes,
    write_label_or_mark,
    write_value_after_label,
)
from backend.services.swut_builder_helpers import extract_warnings_from_session
from backend.services.swut_meta import BuildMetaBase
from backend.services.swut_input_adapter import (
    EnvironmentData,
    FunctionCoverage,
    SwUTSession,
    aggregate_session,
)


# ---------------------------------------------------------------------------
# Dialog/meta payload
# ---------------------------------------------------------------------------

@dataclass
class CoverageBuildMeta(BuildMetaBase):
    """Coverage Report 빌드 메타 — `BuildMetaBase` 17 공통 필드 그대로 사용.

    Coverage Report는 base 외 추가 필드 없음. T137 (W3) BuildMetaBase 통합.
    """
    pass


@dataclass
class CoverageBuildResult:
    """Coverage Report 빌드 결과.

    14차 W1: 메모리 절약 — ``xlsx_io: BytesIO`` 가 주 저장소. ``xlsx_bytes`` 는
    backward compat property — 호출 시점에 ``getvalue()`` (1회 copy). router는
    ``xlsx_io`` 를 StreamingResponse로 직접 stream → bytes copy 회피 + chunk 전송.
    """
    ok: bool
    xlsx_io: io.BytesIO = field(default_factory=io.BytesIO)
    filename: str = ""
    warnings: list[str] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    # ISO 26262 audit hole 표시 (deep-reviewer W5/F3) — placeholder 시트 명시.
    incomplete_sheets: list[str] = field(default_factory=list)
    tool_qualification: dict[str, Any] = field(
        default_factory=lambda: {
            "evidence_class": "auto-generated draft",
            "asil_a_usage": "reviewer 승인 후 evidence로 사용 가능",
            "asil_b_c_d_usage": "단독 evidence 사용 금지 — manual review 의무",
        }
    )

    @property
    def xlsx_bytes(self) -> bytes:
        """Backward compat — BytesIO 전체를 bytes로 복사 (테스트/감사용)."""
        pos = self.xlsx_io.tell()
        self.xlsx_io.seek(0)
        try:
            return self.xlsx_io.read()
        finally:
            self.xlsx_io.seek(pos)

    @property
    def result_size_bytes(self) -> int:
        """BytesIO 크기 — len(xlsx_bytes) 회피 (full copy 없이 size만)."""
        pos = self.xlsx_io.tell()
        self.xlsx_io.seek(0, 2)  # SEEK_END
        size = self.xlsx_io.tell()
        self.xlsx_io.seek(pos)
        return size

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "filename": self.filename,
            "result_size_bytes": self.result_size_bytes,
            "warnings": self.warnings,
            "incomplete_sheets": self.incomplete_sheets,
            "summary": self.summary,
            "tool_qualification": self.tool_qualification,
        }


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

# helper 함수는 backend/services/excel_template_utils.py 로 이전 (reviewer 권고 X5).
# 두 빌더(Coverage / SUTR)가 동일 helper를 공유. 단일 출처로 유지보수.


# _aggregate_session 은 swut_input_adapter.aggregate_session 으로 통합 (deep-reviewer W3).


_OPTIONAL_LABELS = {
    "Build Timestamp", "Reviewer", "Doc. ID",
    # 라운드 F7 D11 fix: 회사 표준 양식 (★개발템플릿 V3)은 Cover에 이 라벨 부재.
    # 'Project Name'/'SW Version' 등은 1.Test Summary 시트로 이동 — Cover에 미발견
    # 시 silent OK (warning emit X).
    "Project", "ASIL Level", "Validation Date", "Status", "Version",
}


def _write_label(ws, label: str, value: Any, out_warnings: list[str] | None) -> None:
    """K1: 라벨 미발견 시 warnings 누적 (optional 라벨은 silent OK)."""
    ok = write_value_after_label(ws, label, value)
    if not ok and label not in _OPTIONAL_LABELS and out_warnings is not None:
        out_warnings.append(f"라벨 '{label}' 미발견 — 셀 쓰기 skip")


def _write_label_or_mark(
    ws, label: str, value: Any, hint: str,
    out_warnings: list[str] | None,
) -> None:
    """23차 T192/W12: excel_template_utils.write_label_or_mark 래퍼 — _OPTIONAL_LABELS 주입."""
    write_label_or_mark(
        ws, label, value, hint=hint,
        optional_labels=_OPTIONAL_LABELS,
        out_warnings=out_warnings,
    )


def _write_cover_sheet(
    ws, meta: CoverageBuildMeta, out_warnings: list[str] | None = None,
    *, layout: Any = None,
) -> None:
    """Cover 시트 — Doc ID/Project/ASIL/Author/Approver 등.

    54차 T282: layout 제공 시 cover_labels 매핑으로 v2.02 양식 동적 호환.
    layout=None이면 v3.01 hardcode label 사용 (SwUT backward compat).
    """
    if not ws:
        return
    labels = layout.cover_labels if layout else {}
    _write_label(ws, labels.get("project_full_name", "Project"),
                 meta.project_full_name, out_warnings)
    _write_label(ws, labels.get("asil_level", "ASIL Level"),
                 meta.asil_level, out_warnings)
    _write_label(ws, labels.get("status", "Status"),
                 "DRAFT — PENDING REVIEW", out_warnings)
    # 23차 T192: validation_date / reviewer / approver 비어있으면 노란 강조
    _write_label_or_mark(ws, labels.get("validation_date", "Validation Date"),
                         meta.validation_date,
                         "yyyy-mm-dd 형식 검증 완료일", out_warnings)
    _write_label_or_mark(ws, labels.get("author", "Author"), meta.author,
                         "test_engineer 또는 default_author", out_warnings)
    _write_label_or_mark(ws, labels.get("reviewer", "Reviewer"), meta.reviewer,
                         "검토자 이름", out_warnings)
    _write_label_or_mark(ws, labels.get("approver", "Approver"), meta.approver,
                         "승인자 이름 (필수)", out_warnings)
    if meta.doc_id_sequence:
        _write_label(ws, labels.get("doc_id", "Doc. ID"),
                     f"{meta.doc_id_base}-{meta.doc_id_sequence}", out_warnings)
    # 56차 T313: Version fill 추가 — SUTR `_write_cover`와 대칭 (이전 SUTR만 v2.02
    # G27 'Version'에 'v2.02' 기록, Coverage는 누락되어 template default 'v0.10' 유지)
    _write_label(ws, labels.get("version", "Version"),
                 f"v{meta.release_sw_version}", out_warnings)
    _write_label(ws, labels.get("build_timestamp", "Build Timestamp"),
                 meta.build_timestamp, out_warnings)


def _write_test_summary_sheet(
    ws, meta: CoverageBuildMeta, agg: dict[str, Any],
    out_warnings: list[str] | None = None,
    *, layout: Any = None, summary: dict[str, Any] | None = None,
) -> None:
    """Test Summary 시트 — 핵심 메트릭 표.

    54차 T282/T283: layout 제공 시 v2.02 양식 동적 호환:
        - test_summary_labels (SW Version / HW Version 등)
        - tc_stats_row B17-F17 (Total/Tested/Passed/Failed/Blocked)
        - requirements_row B22 SwITS 표기
    layout=None이면 v3.01 hardcode label 사용 (SwUT backward compat).
    """
    if not ws:
        return
    labels = layout.test_summary_labels if layout else {}
    _write_label(ws, labels.get("project_full_name", "Project Name"),
                 meta.project_full_name, out_warnings)
    _write_label(ws, labels.get("release_sw_version", "Release Name(SW)"),
                 meta.release_sw_version, out_warnings)
    _write_label(ws, labels.get("hw_version", "Test Target Version(HW)"),
                 meta.hw_version, out_warnings)
    _write_label(ws, labels.get("test_date", "Test Date"),
                 meta.test_date, out_warnings)
    # 24차: Test Engineer 빈 시 노란 강조 (사용자 입력 필요)
    _write_label_or_mark(ws, labels.get("test_engineer", "Test Engineer"),
                         meta.test_engineer,
                         "테스트 엔지니어 이름", out_warnings)
    _write_label(ws, labels.get("final_test_result", "Final Test Result"),
                 meta.final_test_result, out_warnings)

    # 54차 T283: v2.02 양식 신규 row — TC stats / Requirements
    _write_v202_extra_rows(ws, agg, layout, summary, out_warnings)


def _write_v202_tc_stats_row(
    ws, agg: dict[str, Any], layout: Any,
    summary: dict[str, Any] | None,
    out_warnings: list[str] | None = None,
    *,
    total_key: str = "total",
) -> None:
    """55-fix-3 W10 — TC stats data row fill helper (DRY 통합).

    swut_coverage `_write_v202_extra_rows` + swut_sutr `_write_test_summary` inline
    두 곳에서 동일 로직 ~35 lines 중복. 본 helper로 단일 진리 source 확보.

    Args:
        total_key: agg에서 total 값을 가져올 키. Coverage = "total_tcs" (없으면 "total"),
            SUTR = "total" — caller가 명시.
    """
    if layout is None or layout.tc_stats_row is None:
        return
    row = layout.tc_stats_row
    col = layout.tc_stats_col_start or 2

    # 56차 T306 — v2.02 Coverage 양식 label-missing fallback path: layout이 row 17이
    # 빈 row임을 감지 (tc_stats_label_missing=True)했으면 builder가 라벨도 stamp.
    # 회사 Coverage v2.02는 row 17이 사용자 수동 입력 영역으로 비어있음. SITR은
    # label 존재 → 본 branch 미실행 (기존 path 유지).
    if getattr(layout, "tc_stats_label_missing", False):
        label_row = row - 1  # data row 위 라인 = label row
        tc_stats_labels = (
            "Total Number of TCs",
            "Number of TCs Tested",
            "Number of TCs Passed",
            "Number of TCs Failed",
            "Number of TCs not executed",
        )
        for i, lbl in enumerate(tc_stats_labels):
            # label row가 비어있을 때만 stamp — 회사 양식이 일부 라벨만 있는
            # mixed case 방어 (다른 값 있으면 덮어쓰지 않음)
            existing_lbl = ws.cell(label_row, col + i).value
            if existing_lbl is None or (isinstance(existing_lbl, str) and existing_lbl.strip() == ""):
                safe_write(ws, label_row, col + i, lbl)
        if summary is not None:
            summary["tc_stats_fallback_used"] = True
        if out_warnings is not None:
            out_warnings.append(
                f"v2.02 Coverage fallback: TC stats label stamp row={label_row} "
                f"(56차 T306) — audit reviewer에 자동 라벨 채움 사전 통보 권장"
            )

    # 55-fix-2 W4 + 55-fix-3 W8: data row 비어있음 검증.
    # skip 시 산출물 cell에 노란 강조 + hint 추가 (audit silent 차단).
    existing = ws.cell(row, col).value
    if existing is not None and existing != "":
        msg = (
            f"TC stats data row (row={row}, col={col}) already has value "
            f"{existing!r} — 회사 양식 변형 가능성, fill skip (audit reviewer 확인 권장)"
        )
        if out_warnings is not None:
            out_warnings.append(msg)
        if summary is not None:
            summary["tc_stats_skipped_reason"] = msg
        # 55-fix-3 W8: 산출물 셀에 노란 강조 hint (audit reviewer가 X-* 헤더 안 봐도 인지)
        # 단 row 자체에 이미 값이 있어 col+5 (다음 col)에 hint
        mark_user_input_required(
            ws, row, col + 5,
            hint=(
                f"TC stats data row 변형 감지 — 회사 양식이 row {row}에 다른 값. "
                f"audit reviewer가 직접 검토 + 수동 입력 필요"
            ),
        )
        return

    # blocked는 모호 — 0 채움 + summary inferred 표시
    if total_key == "total_tcs":
        total = agg.get("total_tcs", agg.get("total", 0)) or 0
    else:
        total = agg.get(total_key, 0) or 0
    tested = agg.get("tested", 0) or 0
    passed = agg.get("passed", 0) or 0
    failed = agg.get("failed", 0) or 0
    safe_write(ws, row, col, total)
    safe_write(ws, row, col + 1, tested)
    safe_write(ws, row, col + 2, passed)
    safe_write(ws, row, col + 3, failed)
    safe_write(ws, row, col + 4, 0)
    # 54-fix W4: blocked=0 inferred 시각 안내
    mark_user_input_required(
        ws, row, col + 5,
        hint="Blocked TC 수 inferred=0 — VectorCAST blocked 데이터 미지원, 명시적 입력 필요",
    )
    if summary is not None:
        summary["tc_stats_blocked_inferred"] = True


def _write_v202_requirements_row(
    ws, layout: Any,
    summary: dict[str, Any] | None = None,
    out_warnings: list[str] | None = None,
) -> None:
    """55-fix-3 W10 — Requirements row fill helper (DRY 통합).

    55-fix-2 W5 가드: B 셀이 빈/'SwITS'면 fill, 다른 값 ('■' 헤더 등)이면 skip.
    55-fix-3 (deep-reviewer I5): skip 시 warning + summary reason 누적 (W4와 정책 통일).
    """
    if layout is None or layout.requirements_row is None:
        return
    row = layout.requirements_row

    # 56차 T306 — v2.02 Coverage 양식 label-missing fallback: layout이 row 20이
    # 빈 row임을 감지(requirements_label_missing=True)했으면 builder가 헤더+라벨 stamp.
    # SITR은 label/SwITS default 존재 → 본 branch 미실행 (기존 path 유지).
    if getattr(layout, "requirements_label_missing", False):
        # row 20 = 헤더, row 21 = "Source" 라벨, row 22 = "SwITS" 데이터
        header_existing = ws.cell(row, 2).value
        if header_existing is None or (
            isinstance(header_existing, str) and header_existing.strip() == ""
        ):
            safe_write(ws, row, 2, "■  Requirements/Design Coverage")
        source_existing = ws.cell(row + 1, 2).value
        if source_existing is None or (
            isinstance(source_existing, str) and source_existing.strip() == ""
        ):
            safe_write(ws, row + 1, 2, "Source")
        swits_existing = ws.cell(row + 2, 2).value
        if swits_existing is None or (
            isinstance(swits_existing, str) and swits_existing.strip() == ""
        ):
            safe_write(ws, row + 2, 2, "SwITS")
        if summary is not None:
            summary["requirements_fallback_used"] = True
        if out_warnings is not None:
            out_warnings.append(
                f"v2.02 Coverage fallback: Requirements 헤더+라벨 stamp row={row}~{row+2} "
                f"(56차 T306) — audit reviewer에 자동 채움 사전 통보 권장"
            )
        return

    existing = ws.cell(row, 2).value
    if existing is None or existing == "" or existing == "SwITS":
        safe_write(ws, row, 2, "SwITS")
    else:
        # I5 deferred fix — W5 skip 시 warning + summary reason 누적
        msg = (
            f"Requirements row (row={row}, col=B) already has value "
            f"{existing!r} — 회사 양식 변형 가능성, SwITS fill skip"
        )
        if out_warnings is not None:
            out_warnings.append(msg)
        if summary is not None:
            summary["requirements_row_skipped_reason"] = msg


def _write_v202_extra_rows(
    ws, agg: dict[str, Any], layout: Any, summary: dict[str, Any] | None,
    out_warnings: list[str] | None = None,
) -> None:
    """54차 T283 — v2.02 양식 신규 row fill (B17-F17 TC stats + B22 Requirements).

    55-fix-3 W10: helper로 DRY 통합. swut_sutr inline도 동일 helper 호출.
    layout이 None 또는 해당 row가 None (v3.01)이면 silent skip — SwUT 회귀 영향 zero.
    """
    _write_v202_tc_stats_row(
        ws, agg, layout, summary, out_warnings, total_key="total_tcs",
    )
    _write_v202_requirements_row(ws, layout, summary, out_warnings)


def _compute_asil_distribution(
    function_rows: list[FunctionCoverage],
    function_asil_map: dict[str, str],
    *,
    function_asil_from_suds: dict[str, str] | None = None,
    component_asil_from_sds: dict[str, str] | None = None,
    function_asil_from_srs: dict[str, str] | None = None,
    function_name_to_swufn_from_suds: dict[str, str] | None = None,
) -> tuple[dict[str, int], dict[str, list[str]], list[str]]:
    """30차 W21 + 31차 W29 + 라운드 84 T1801: function 별 ASIL 등급 분포 계산.

    라운드 84 T1801: SUDS/SDS/SRS source chain 통합 — `_write_coverage_sheet`/
    `_write_test_log` 의 fallback chain (라운드 80)과 priority 일관:
      1. c_source @asil (function_asil_map, 라운드 30 W21)
      2. SUDS docx function 직접 (라운드 80)
      3. SDS docx component (component_name 매칭, 라운드 80)
      4. SRS docx 보조 (함수명, 라운드 80)
      5. 미설정 → UNKNOWN

    이전 (라운드 30~83): function_asil_map만 사용 → SUDS/SDS/SRS 추출 값
    무시되어 AuditLog 시트 ASIL 분포 0건 (라운드 83 v17 검증).

    Args:
        function_rows: 집계된 함수 list (``FunctionCoverage``).
        function_asil_map: ``swut_asil_resolver`` 결과.
        function_asil_from_suds: SUDS docx 함수 단위 ASIL (옵션, 라운드 80).
        component_asil_from_sds: SDS docx 컴포넌트 ASIL (옵션, 라운드 80).
        function_asil_from_srs: SRS docx 보조 ASIL (옵션, 라운드 80).

    Returns:
        ``(distribution, function_ids_by_asil)`` — A/B/C/D/QM/UNKNOWN 5+1 bucket.
        function_ids_by_asil는 B/C/D만 누적 (audit 강조 대상).
    """
    distribution: dict[str, int] = {}
    ids_by_asil: dict[str, list[str]] = {"B": [], "C": [], "D": []}
    # 라운드 86 T2001: UNKNOWN bucket 함수의 fc.unit_id/name list (audit 진단용).
    unmapped: list[str] = []
    suds_map = function_asil_from_suds or {}
    sds_map = component_asil_from_sds or {}
    srs_map = function_asil_from_srs or {}
    name_to_swufn = function_name_to_swufn_from_suds or {}

    for fc in function_rows:
        candidate_keys = [fc.unit_id or "", fc.name or ""]
        asil = ""
        matched_id = ""
        # 1) function_asil_map 직접 (c_source @asil 결과)
        for key in candidate_keys:
            if not key:
                continue
            if key in function_asil_map:
                asil = function_asil_map[key]
                matched_id = key
                break
            m = _TC_FN_RE.search(key)
            if m and m.group(1) in function_asil_map:
                asil = function_asil_map[m.group(1)]
                matched_id = m.group(1)
                break
        # 라운드 84 T1801: 2) SUDS function 직접 매핑 (SwUFn_NNNN 추출)
        if not asil:
            for key in candidate_keys:
                if not key:
                    continue
                m = _TC_FN_RE.search(key)
                if m:
                    sw_fn_id = m.group(1)
                    if sw_fn_id in suds_map:
                        asil = suds_map[sw_fn_id]
                        matched_id = sw_fn_id
                        break
        # 라운드 85 T1903: 2b) SUDS reverse map (함수명 → SwUFn → ASIL).
        # fc.unit_id/name이 'main' 같은 함수명이면 reverse map으로 SwUFn 변환 후
        # SUDS ASIL lookup. 라이브 v1.07 unique 440건 매핑 — HDPDM01 UNKNOWN 98.3% 해소.
        if not asil and name_to_swufn and suds_map:
            for key in candidate_keys:
                if not key:
                    continue
                sw_fn_id = name_to_swufn.get(key)
                if sw_fn_id and sw_fn_id in suds_map:
                    asil = suds_map[sw_fn_id]
                    matched_id = sw_fn_id
                    break
        # 3) SDS component 매핑 (fc.component_name)
        if not asil and sds_map and getattr(fc, "component_name", ""):
            comp_raw = fc.component_name
            import re as _re_cn
            candidates_c = []
            m_swcom = _re_cn.search(r"SwCom_\d+", comp_raw)
            if m_swcom:
                candidates_c.append(m_swcom.group(0))
            for line in comp_raw.splitlines():
                stripped = line.strip().strip("()").strip()
                if stripped and stripped not in candidates_c:
                    candidates_c.append(stripped)
            candidates_c.append(comp_raw.strip())
            for k in candidates_c:
                if k in sds_map:
                    asil = sds_map[k]
                    matched_id = fc.unit_id or fc.name or ""
                    break
        # 4) SRS function 보조 (fc.name)
        if not asil and srs_map and fc.name:
            if fc.name in srs_map:
                asil = srs_map[fc.name]
                matched_id = fc.name

        bucket = f"ASIL_{asil}" if asil else "UNKNOWN"
        distribution[bucket] = distribution.get(bucket, 0) + 1
        if asil in ("B", "C", "D") and matched_id:
            ids_by_asil[asil].append(matched_id)
        # 라운드 86 T2001: UNKNOWN bucket 함수 list 누적 (audit 진단용).
        if not asil:
            fn_key = fc.unit_id or fc.name or "<no-id>"
            unmapped.append(fn_key)

    return (
        distribution,
        {k: sorted(set(v)) for k, v in ids_by_asil.items()},
        sorted(set(unmapped)),
    )


def _write_coverage_sheet(
    ws, agg: dict[str, Any], *, layout: Any = None,
    out_warnings: list[str] | None = None,
    is_swit_caller: bool = False,
    c_function_map: dict[str, dict[str, Any]] | None = None,
) -> int:
    """3. Coverage 시트 — per-function Statement/Branch/Exception 표.

    30차 W21: ``agg["function_asil_map"]`` 에 ASIL D 매핑된 함수는 row 전체에
    빨간 강조 (``mark_asil_d_function``). 색상은 FAIL과 동일 RGB이나 호출
    의미 분리.

    59차 F4-C: ``layout.coverage_metric_kind == "function_and_calls"`` 시
    KJPDS02 v1.01 양식 호환 — 추가 col에 Function Calls metric stamp
    (``FunctionCoverage.function_calls_coverage``). v2.02/v3.01은 단일 metric.

    라운드 74 T906: c_function_map 제공 시 vcast function_rows + c_parser 함수
    union dedup. c_parser only row는 빈 CoverageStats + 노란 마킹 + Note column에
    "[c_parser] coverage 미실측" 안내. 회사 KJPDS02 v1.01 양식 570 함수 row stamp
    대비 격차 해소 (HDPDM01은 60 vcast + 257 c_parser only ≈ 317).

    Returns:
        쓰여진 행 수.
    """
    if not ws:
        return 0
    # 라운드 74 T906 — c_parser merge (provided이면 union list 사용)
    # 라운드 76 T1107 — file 정보 사전 주입으로 dedup 정확성 향상.
    if c_function_map:
        from backend.services.swut_input_adapter import (
            enhance_function_coverage_with_file, merge_function_rows_with_c_parser,
        )
        # vcast function_rows에 c_parser file 정보 주입 → dedup key (name, file) 정확
        vcast_rows = list(agg.get("function_rows") or [])
        enhance_function_coverage_with_file(vcast_rows, c_function_map)
        agg["function_rows"] = vcast_rows
        function_rows = merge_function_rows_with_c_parser(
            agg, c_function_map, out_warnings=out_warnings,
        )
    else:
        function_rows = list(agg.get("function_rows") or [])
    if not function_rows:
        return 0

    # 헤더 행을 찾는다 — "Unit ID" 또는 "Function Name" 라벨 위치
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=False):
        labels = [
            str(c.value).strip() if c.value else ""
            for c in row
        ]
        if any(l in ("Unit ID", "Function Name", "Component") for l in labels):
            header_row = row[0].row
            break
    if header_row is None:
        return 0

    # 데이터 행 시작은 헤더 + 1 또는 + 2 (회사 포맷에 따라 hierarchical header)
    data_start = header_row + 2

    # 라운드 F7 D5 fix: 회사 표준 양식은 C2부터 'No' col 시작 (HDPDM01 release
    # 산출물은 C1부터). header row scan으로 'No' col 동적 감지 → 그 col부터 stamp.
    # 미발견 시 기존 hardcoded C1 fallback (backward-compat).
    no_col = 1  # fallback
    component_col_offset = 1  # No 다음 col이 Component (회사 표준) 또는 unit_id (HDPDM01)
    has_component_col = False
    for row in ws.iter_rows(min_row=max(1, header_row - 1),
                             max_row=header_row, values_only=False):
        for cell in row:
            v = str(cell.value or "").strip()
            if v == "No":
                no_col = cell.column
            elif v == "Component":
                has_component_col = True
    # 회사 표준: No=C2, Component=C3, Unit ID=C4, Name=C5, Statement Count=C6 ...
    # HDPDM01 release: No=C1, unit_id=C2, name=C3, Statement total=C4 ...
    if has_component_col:
        unit_id_col = no_col + 2  # skip No + Component
        # Statement: Count=+4, Total=+5, Pass=+6 (header에서 'Count'/'Total'/'Pass' 위치 그대로)
        stmt_count_col = no_col + 4
        # Branch: Count=+8 (Statement+Exception 사이 gap), Total=+9, Pass=+10
        branch_count_col = no_col + 8
    else:
        unit_id_col = no_col + 1
        stmt_count_col = no_col + 3
        branch_count_col = no_col + 6

    # 30차 W21: 함수별 ASIL 매핑 + ASIL D 식별.
    function_asil_map: dict[str, str] = agg.get("function_asil_map") or {}

    # 라운드 92 — KJPDS02 spec-based 양식 정렬 (회사 감사본 일치).
    #   회사 SwUTCV v1.01 4.Coverage 데이터 행은:
    #     C=SwCom_NN (컴포넌트 ID) / D=실 SwUFn_NNNN (SwUDS 함수 ID) /
    #     H,L=Pass/Fail (Statement/Branch 합격 여부)
    #   기존 (라운드 77~91) 동작은 D=순차 SwUFn_0001.. / C=env명 / H,L='O'/'X'.
    #   graceful — name→SwUFn 매핑(SwUDS 함수명→ID)이 agg에 존재할 때만 활성.
    #   매핑 없으면 (HDPDM01/SwIT) 기존 동작 100% 보존.
    name_to_swufn: dict[str, str] = agg.get("function_name_to_swufn_from_suds") or {}
    spec_based = bool(name_to_swufn) and has_component_col

    # F7 stage 8 T706 fix — SwITCV (회사 표준) layout 분기:
    # layout.coverage_metric_kind == "function_and_calls" → SwIT 양식
    #   회사 표준 SwITCV 4.Coverage R9 header: No(C2)/Component(C3)/Unit(C4-C5)/
    #     Functions(C6 — 단일 'O'/'X' Pass)/Exception(C7)/Function Called(C8 Count, C9 Total, C10 Pass)
    #   → Statement/Branch는 SwITCV에 없는 metric. Functions Pass + Function Calls만 stamp
    # default → SwUT 양식 (Statement + Branch + 옵션 Function Calls)
    # F7 stage 10 G3 fix: is_swit_caller 명시 — SwUT 호출 (build_coverage_report)는
    # SwUTCV (회사 표준 v1.01 layout)도 Statement/Branch 매핑. SwIT 호출
    # (build_swit_coverage_report)만 SwIT 분기 (Functions Pass + Function Called).
    is_swit_metric_layout = is_swit_caller and (
        layout is not None
        and getattr(layout, "coverage_metric_kind", "single") == "function_and_calls"
        and has_component_col
    )

    # 라운드 73 T803 — row 자동 확장 (1회 batch insert로 O(N²) 회피).
    # template은 고정 slot (회사 v3.01 SwUTCV 4.Coverage 15 slot / SwITCV 5 slot) —
    # 데이터(60+ 함수)가 slot 초과 시 stamp 잘림. 부족분을 사전 계산 후 단일
    # insert_rows + style/merge/dimension 복제.
    needed_last_row = data_start + len(function_rows) - 1
    if needed_last_row > ws.max_row:
        from backend.services.excel_template_utils import (
            auto_expand_row_block, push_sentinel_to_last_row,
            update_cross_refs_after_row_expansion,
        )
        # 라운드 76 자체평가 fix — old_totals_row를 ws.max_row가 아닌 양식 R5/R6
        # cross-ref formula에서 자동 detect. ws.max_row=69 (양식 sample slot 끝)와
        # cross-ref formula `=E25` (양식이 가정한 TOTALS row)가 일치하지 않을 때
        # 갱신 누락 방지. R5/R6 col 5~8 (E/F/G/H) 스캔 → `=<col>{row}` 패턴 첫 row.
        old_totals_row = ws.max_row
        for _r in (5, 6):
            for _c in range(5, 9):
                try:
                    _v = ws.cell(_r, _c).value
                except (AttributeError, IndexError):
                    continue
                if not isinstance(_v, str) or not _v.startswith("=") or "!" in _v:
                    continue
                _m = re.search(r"=([A-Z]+)(\d+)\b", _v)
                if _m:
                    _detected = int(_m.group(2))
                    # cell 자기 자신 ref(=E5 같은) 또는 calculated(R5/R6) 제외
                    if _detected > 10:  # 양식 TOTALS row는 보통 R20+
                        old_totals_row = _detected
                        break
            if old_totals_row != ws.max_row:
                break
        shortage = needed_last_row - ws.max_row
        # data_start + 1 위치에 신규 row 추가 → 기존 sentinel/footer 자동 downshift.
        inserted = auto_expand_row_block(
            ws,
            insert_at_row=data_start + 1,
            amount=shortage,
            template_row_idx=data_start,
            copy_style=True, copy_merge=True, copy_dimension=True,
        )
        if inserted < shortage and out_warnings is not None:
            out_warnings.append(
                f"[row_expand] Coverage 시트 row 부족 ({shortage}개 필요, {inserted}개 확장) — "
                "신규 stamp 일부 누락 가능"
            )
        # 라운드 76 T1107 — sentinel push + cross-ref formula 동적 갱신.
        # 양식 default `=E25` / `=H25` / `=I25` / `=L25` / `=M25` cross-ref가
        # row 폭증 후 R25 → R{new_totals_row} 자동 갱신. 안 하면 R25가 c_parser
        # 함수 row가 되어 Statement Total 같은 cross-ref formula 의미 깨짐.
        try:
            push_sentinel_to_last_row(ws)
            new_totals_row = data_start + len(function_rows)
            updated = update_cross_refs_after_row_expansion(
                ws,
                old_totals_row=old_totals_row,
                new_totals_row=new_totals_row,
            )
            if updated > 0 and out_warnings is not None:
                out_warnings.append(
                    f"[cross_ref] 양식 cross-ref formula {updated}건 동적 갱신 "
                    f"(R{old_totals_row} → R{new_totals_row}). audit reviewer 통보 의무."
                )
        except ImportError:
            pass

    # 기존 데이터 행을 덮어쓴다 (template이 기존 sample 데이터 가질 수 있음).
    # 라운드 76 자체평가 fix — has_component_col=True 시 C3 'Component' col에 stamp.
    # 회사 v3.01 SwUTCV 4.Coverage 양식 R8: C3='Component', C4='Unit', C5='Name'.
    # 이전: C3 미stamp (양식 default `SwCom_01: XXXX` 잔존 또는 빈). audit reviewer가
    # 함수의 소속 component (vcast component_name 또는 c_parser file) 인지 어려움.
    # 라운드 92 — spec_based 시 C열(Component) 데이터 영역 병합 해제. 회사 감사본은
    # C열 병합 없이 매 행 SwCom_NN 명시 (실측). 표준 v0.10 템플릿은 C10:C14 등
    # 세로 병합 잔존 → 비-anchor 행 stamp가 무시됨. 데이터 영역 병합만 해제 (헤더
    # C8:C9 보존). HDPDM01/SwIT (spec_based=False)는 영향 없음.
    if spec_based:
        comp_col_idx = no_col + 1
        for rng in list(ws.merged_cells.ranges):
            if (rng.min_col <= comp_col_idx <= rng.max_col
                    and rng.min_row >= data_start):
                try:
                    ws.unmerge_cells(str(rng))
                except (ValueError, KeyError):
                    pass
    written = 0
    # 라운드 92 — spec_based Statement/Branch Pass/Fail 집계 (TOTALS 섹션용) +
    # SwUFn 매핑 실패 함수 list (audit 진단).
    spec_stmt_fail = 0
    spec_stmt_total = 0
    spec_branch_fail = 0
    spec_branch_total = 0
    spec_unmatched: list[str] = []
    spec_unmatched_count = 0
    last_data_row = data_start - 1

    for i, fc in enumerate(function_rows):
        r = data_start + i
        last_data_row = r
        safe_write(ws, r, no_col, i + 1)

        # 라운드 74 T906 — c_parser only row 식별 (unit_id `SwUFn_C_<idx>` prefix).
        is_c_parser_only = bool(fc.unit_id and fc.unit_id.startswith("SwUFn_C_"))

        # 라운드 77 자체평가 fix — C4 'Unit ID' 회사 양식 호환 sequential SwUFn_NNNN.
        # 라운드 76 fix #4 후 vcast fc.unit_id가 함수명(`main`)으로 변경 → 회사 양식
        # 의도 (`SwUFn_0101` 형식 함수 식별자)와 mismatch. 사용자 검수: "ID가 함수이름이
        # 들어가 있네". vcast 함수 (unit_id == name인 경우)는 글로벌 sequential
        # `SwUFn_<i+1:04d>` 부여. sub_functions/c_parser only는 기존 unit_id 유지.
        # 라운드 92 — spec_based 시 D=실 SwUFn ID (SwUDS 함수명→ID 매핑).
        # 매핑 성공: SwUFn_0121 (회사 감사본 일치). 실패 (SUDS 미등재 함수):
        # 순차 SwUFn_NNNN fallback + spec_unmatched 누적 (audit 진단).
        resolved_swufn = ""
        if spec_based and not is_c_parser_only:
            resolved_swufn = name_to_swufn.get(fc.name, "") or name_to_swufn.get(
                fc.unit_id, ""
            )
            if not resolved_swufn:
                spec_unmatched_count += 1
                if len(spec_unmatched) < 60:
                    spec_unmatched.append(fc.name or fc.unit_id or "<no-id>")

        if resolved_swufn:
            display_unit_id = resolved_swufn
        elif (fc.unit_id == fc.name and fc.unit_id
                and not fc.unit_id.startswith("SwUFn_")):
            display_unit_id = f"SwUFn_{i + 1:04d}"
        else:
            display_unit_id = fc.unit_id

        # 라운드 77 T1204 — C3 'Component' stamp 정확화.
        # 라운드 76 fix #2 (b8eefea)에서 C3=fc.name이라 vcast row의 C3=C4=C5 중복 +
        # R10 anomaly. 라운드 77 T1201로 fc.component_name 신규 필드 — vcast row는
        # component name 추적, c_parser only는 file basename으로 주입 완료.
        if has_component_col:
            comp_col = no_col + 1  # No 다음 col
            comp_name = ""
            # 라운드 92 — spec_based 시 C=SwCom_NN (회사 감사본 일치).
            # 출처: 실 SwUFn_NNNN 앞 2자리 (SwUDS 'Related ID' SwCom 검증 결과 100%
            # 일치 — 라운드 92 .codex_tmp 조사).
            # 라운드 96 fix — resolved_swufn(SUDS 매핑 성공분) 대신 D열에 실제 표기되는
            # display_unit_id 기준으로 도출. 이전엔 SUDS 미매핑이나 vcast가 실 SwUFn을
            # 가진 행(예 SwUFn_3329)은 C 공란, 순차 fallback 행은 env명(Lib_sha256) leak.
            # display_unit_id가 SwUFn_NNNN면 항상 SwCom_NN 부여 → 일관성 확보.
            if spec_based:
                m_swufn = re.match(r"SwUFn_(\d{2})\d{2}", str(display_unit_id))
                if m_swufn:
                    comp_name = f"SwCom_{m_swufn.group(1)}"
            if not comp_name:
                # 우선순위: fc.component_name (vcast/sub_function/c_parser only 주입)
                # → fc.file.stem fallback (component_name 빈 string인 backward-compat)
                comp_name = fc.component_name
                if not comp_name and fc.file:
                    from pathlib import Path as _PathLocal2
                    comp_name = _PathLocal2(fc.file).stem
            if comp_name:
                # 라운드 96 — spec_based는 force_write_cell로 orphan MergedCell
                # (회사 양식 SwCom 그룹 병합 해제 잔존 셀) 강제 기록 → C 공란 해소.
                # HDPDM01/SwIT는 기존 safe_write 동작 보존.
                if spec_based:
                    # C열 테두리 통일은 written 루프 종료 후 최종 패스에서 수행
                    # (라운드 97 — 루프 중간 복사가 후속 단계에서 리셋되는 문제 회피).
                    force_write_cell(ws, r, comp_col, comp_name)
                else:
                    safe_write(ws, r, comp_col, comp_name)

        safe_write(ws, r, unit_id_col, display_unit_id)
        safe_write(ws, r, unit_id_col + 1, fc.name)

        # 라운드 92 — spec_based 매핑 실패 행(SwUDS 미등재)은 D 셀 노란 마킹.
        # 순차 SwUFn fallback은 추정 ID라 audit reviewer가 추적성 수동 검증 필요.
        # SwUDS 매핑 성공 행은 마킹 없음 (회사 감사본 동일).
        if spec_based and not is_c_parser_only and not resolved_swufn:
            from backend.services.excel_template_utils import _apply_fill
            from backend.services.design_tokens import USER_INPUT_FILL_RGB
            _apply_fill(ws, r, unit_id_col, USER_INPUT_FILL_RGB)

        if is_swit_metric_layout:
            # SwITCV — Functions Pass (C6) + Function Called metric (C8/C9/C10)
            # Functions Pass: function 매핑 여부 — 신규 session에 unit_id 있으면 'O'
            functions_pass_col = no_col + 4
            fcalls_count_col = no_col + 6
            if is_c_parser_only:
                # 라운드 74 T906 — c_parser only Functions Pass cell에 '[c_parser]' 안내.
                # 라운드 76 자체평가 fix — 안내 메시지 보강 + Name col에도 마킹.
                from backend.services.excel_template_utils import _apply_fill
                from backend.services.design_tokens import USER_INPUT_FILL_RGB
                safe_write(ws, r, functions_pass_col, "[c_parser only — 미실측]")
                _apply_fill(ws, r, functions_pass_col, USER_INPUT_FILL_RGB)
                # Name col(C5)도 [c_parser] suffix로 audit 마킹
                _apply_fill(ws, r, unit_id_col + 1, USER_INPUT_FILL_RGB)
            else:
                safe_write(ws, r, functions_pass_col, "O")
                fcc = getattr(fc, "function_calls_coverage", None)
                if fcc is not None and fcc.total > 0:
                    safe_write(ws, r, fcalls_count_col, fcc.covered)
                    safe_write(ws, r, fcalls_count_col + 1, fcc.total)
                    safe_write(ws, r, fcalls_count_col + 2, "O" if fcc.passed else "X")
        else:
            # SwUTCV / HDPDM01 — Statement + Branch metric
            if is_c_parser_only:
                # 라운드 76 자체평가 fix — c_parser only 안내 메시지 보강 + Name col 마킹.
                from backend.services.excel_template_utils import _apply_fill
                from backend.services.design_tokens import USER_INPUT_FILL_RGB
                safe_write(ws, r, stmt_count_col, "[c_parser only — 미실측]")
                safe_write(ws, r, stmt_count_col + 1, "-")
                safe_write(ws, r, stmt_count_col + 2, "-")
                safe_write(ws, r, branch_count_col, "-")
                safe_write(ws, r, branch_count_col + 1, "-")
                safe_write(ws, r, branch_count_col + 2, "-")
                _apply_fill(ws, r, stmt_count_col, USER_INPUT_FILL_RGB)
                _apply_fill(ws, r, branch_count_col, USER_INPUT_FILL_RGB)
                _apply_fill(ws, r, unit_id_col + 1, USER_INPUT_FILL_RGB)  # Name col
            else:
                # 라운드 92 — spec_based 시 Pass 셀(H/L) 표기 'Pass'/'Fail'
                # (회사 감사본). 기존 (HDPDM01/v3.01): 'O'/'X' 유지.
                stmt_mark = (
                    ("Pass" if fc.statement.passed else "Fail") if spec_based
                    else ("O" if fc.statement.passed else "X")
                )
                branch_mark = (
                    ("Pass" if fc.branch.passed else "Fail") if spec_based
                    else ("O" if fc.branch.passed else "X")
                )
                safe_write(ws, r, stmt_count_col, fc.statement.total)
                safe_write(ws, r, stmt_count_col + 1, fc.statement.covered)
                safe_write(ws, r, stmt_count_col + 2, stmt_mark)
                safe_write(ws, r, branch_count_col, fc.branch.total)
                safe_write(ws, r, branch_count_col + 1, fc.branch.covered)
                safe_write(ws, r, branch_count_col + 2, branch_mark)
                # 라운드 92 — spec_based TOTALS 집계 (Pass/Fail count).
                if spec_based:
                    spec_stmt_total += 1
                    spec_branch_total += 1
                    if not fc.statement.passed:
                        spec_stmt_fail += 1
                    if not fc.branch.passed:
                        spec_branch_fail += 1

            # 59차 F4-C — KJPDS02 v1.01 양식 (HDPDM01과 별도): Function Calls metric col stamp.
            if (
                layout is not None
                and getattr(layout, "coverage_metric_kind", "single") == "function_and_calls"
            ):
                fcc = getattr(fc, "function_calls_coverage", None)
                if fcc is not None and fcc.total > 0:
                    safe_write(ws, r, 10, fcc.total)
                    safe_write(ws, r, 11, fcc.covered)
                    safe_write(ws, r, 12, "O" if fcc.passed else "X")

        # 라운드 76 자체평가 fix — File col stamp.
        # SwUTCV v3.01: R8 C14 'File' / SwITCV v2.02: R9 C12 'File' — 양식별 mismatch.
        # 자동 detect: header row 8~10 scan으로 'File' label 위치 찾음. 미발견 시
        # caller 기준 fallback (SwIT=12, SwUT=14).
        if not hasattr(ws, "_file_col_cached"):
            _file_col_detected = None
            for _hr in range(max(1, header_row - 1), header_row + 2):
                for _hc in range(1, min(ws.max_column + 1, 20)):
                    try:
                        _hv = ws.cell(_hr, _hc).value
                    except (AttributeError, IndexError):
                        continue
                    if isinstance(_hv, str) and _hv.strip().lower() == "file":
                        _file_col_detected = _hc
                        break
                if _file_col_detected:
                    break
            ws._file_col_cached = _file_col_detected or (12 if is_swit_caller else 14)
        file_col = ws._file_col_cached
        if fc.file:
            from pathlib import Path as _PathLocal
            safe_write(ws, r, file_col, _PathLocal(fc.file).name)

        # 30차 W21 + 31차 W29: ASIL B/C/D 함수면 row의 핵심 컬럼 강조.
        # fc.unit_id 가 SwUFn_NNNN 패턴일 수 있고 또는 다른 ID. 둘 다 매칭 시도.
        asil = function_asil_map.get(fc.unit_id) or function_asil_map.get(fc.name)
        if not asil:
            # fc.name / fc.unit_id 에 SwUFn_NNNN 정규식 추출 fallback.
            m = _TC_FN_RE.search(fc.unit_id or "") or _TC_FN_RE.search(fc.name or "")
            if m:
                asil = function_asil_map.get(m.group(1))
        # 라운드 77 T1206 — c_function_map.comment_asil 직접 fallback.
        # 라운드 76 fix #4 후 vcast row unit_id가 함수명(`main`)으로 변경 → 기존
        # function_asil_map (SwUFn_NNNN key)로 매칭 안 됨 → ASIL 강조 누락. c_parser
        # comment_asil은 함수명 단위라 정확 매칭.
        if not asil and c_function_map:
            c_entry = c_function_map.get(fc.name)
            if c_entry:
                _ca = (c_entry.get("comment_asil") or "").strip().upper()
                if _ca in {"A", "B", "C", "D", "QM"}:
                    asil = _ca
        # 라운드 80 T1408 — ISO 26262 추적성 체인 fallback chain (SUDS → SDS → SRS).
        # agg에 주입된 ASIL maps 활용 (build_coverage_report가 session에서 복사).
        _sw_fn_id = fc.unit_id if (fc.unit_id or "").startswith("SwUFn_") else ""
        if not _sw_fn_id:
            _m_swufn = _TC_FN_RE.search(fc.unit_id or "") or _TC_FN_RE.search(fc.name or "")
            if _m_swufn:
                _sw_fn_id = _m_swufn.group(1)
                if not _sw_fn_id.startswith("SwUFn_"):
                    _sw_fn_id = f"SwUFn_{_sw_fn_id}" if _sw_fn_id.isdigit() else _sw_fn_id
        if not asil and _sw_fn_id:
            _suds_map = agg.get("function_asil_from_suds") or {}
            _sasil = _suds_map.get(_sw_fn_id)
            if _sasil:
                asil = _sasil
        if not asil:
            _sds_map = agg.get("component_asil_from_sds") or {}
            if _sds_map and fc.component_name:
                import re as _re_sds
                _candidates = []
                _m_swcom = _re_sds.search(r"SwCom_\d+", fc.component_name)
                if _m_swcom:
                    _candidates.append(_m_swcom.group(0))
                for _line in fc.component_name.splitlines():
                    _stripped = _line.strip().strip("()").strip()
                    if _stripped and _stripped not in _candidates:
                        _candidates.append(_stripped)
                _candidates.append(fc.component_name.strip())
                for _k in _candidates:
                    if _k in _sds_map:
                        asil = _sds_map[_k]
                        break
        if not asil:
            _srs_map = agg.get("function_asil_from_srs") or {}
            if _srs_map:
                _rasil = _srs_map.get(fc.name)
                if _rasil:
                    asil = _rasil
        # 라운드 81 T1503: ASIL 5단계 그라데이션 — A/QM 추가 (audit reviewer 친화).
        # D(빨강) > C(주황) > B(파랑) > A(녹색) > QM(회색) 위험도 ↓
        _marker = {
            "A": mark_asil_a_function,
            "B": mark_asil_b_function,
            "C": mark_asil_c_function,
            "D": mark_asil_d_function,
            "QM": mark_asil_qm_function,
        }.get(asil or "")
        if _marker:
            for col in (2, 3):  # Unit ID + Function Name 컬럼
                _marker(ws, r, col)

        written += 1

    # F7 자체평가 R1 C2 fix: clear policy — 신규 stamp 후 양식 default 함수 row
    # (Fun_B/Fun_C/Fun_D 등) clear. 회사 표준 SwUTCV 4.Coverage 양식이 R12+에
    # 5+ default 함수 보유 → 신규 session 2 함수만 R10/R11 stamp + R12+ default 잔존
    # → R25 `=SUM(F10:F24)` 수식이 default까지 sum하여 false coverage 산출.
    if written > 0:
        try:
            from backend.services.excel_template_utils import clear_data_range
            clear_start = data_start + written
            clear_end = ws.max_row
            if clear_end >= clear_start:
                cleared = clear_data_range(
                    ws,
                    start_row=clear_start, end_row=clear_end,
                    start_col=no_col, end_col=branch_count_col + 6,
                    preserve_formula=True, preserve_merged_anchor=True,
                    sentinel_patterns=[
                        "End of Document", "< End", "■ Appendix",
                        "Appendix", "※", "TOTALS", "GRAND TOTALS",
                    ],
                )
                if out_warnings is not None and cleared > 0:
                    out_warnings.append(
                        f"[clear] Coverage 시트 row {clear_start}~{clear_end} "
                        f"양식 default 함수 row {cleared} cell clear (false coverage 차단)"
                    )
        except ImportError:
            pass

    # 라운드 92 — spec_based TOTALS 섹션 + 상단 요약 (회사 감사본 일치).
    # clear 이후 stamp (clear가 default row 비운 뒤). 데이터 끝(last_data_row)
    # 다음 3행에 Fail/Pass/Total 카운트.
    if spec_based and written > 0 and not is_swit_metric_layout:
        _write_spec_totals(
            ws,
            data_start=data_start,
            last_data_row=last_data_row,
            unit_id_col=unit_id_col,
            stmt_label_col=stmt_count_col + 1,  # G (Fail/Pass/Total 라벨)
            stmt_value_col=stmt_count_col + 2,  # H (count 수식)
            branch_label_col=branch_count_col + 1,  # K
            branch_value_col=branch_count_col + 2,  # L
            no_col=no_col,
        )
        if out_warnings is not None:
            out_warnings.append(
                f"[spec-cov] 라운드 92 — Statement {spec_stmt_total} (Fail {spec_stmt_fail}) / "
                f"Branch {spec_branch_total} (Fail {spec_branch_fail}) TOTALS + 요약 stamp"
            )
            if spec_unmatched_count > 0:
                out_warnings.append(
                    f"[spec-cov] 함수명↔SwUFn 매핑 실패 {spec_unmatched_count}건 — "
                    f"SwUDS 미등재 함수 (순차 SwUFn fallback). 예: "
                    f"{', '.join(spec_unmatched[:15])}"
                )

    # 라운드 97 — spec_based C(Component)열 테두리 통일 (최종 패스).
    # 회사 빈 양식 C열은 세로 테두리만(b1100)이고 orphan 재생성 셀은 무테(b0000)
    # → REF는 사방 테두리(b1111). written 루프 중간 복사는 후속 단계(force_write/
    # ASIL fill 등)에서 리셋되므로, 모든 stamp/clear/totals 완료 후 D열(unit_id_col,
    # 정상 사방 테두리) border를 C열 데이터 행에 일괄 전파. spec_based(KJPDS02)만 —
    # HDPDM01/SwIT 보존.
    if spec_based and has_component_col and written > 0:
        import copy as _copy_border

        from openpyxl.cell.cell import MergedCell as _MC_border
        _comp_col_final = no_col + 1
        for _rr in range(data_start, last_data_row + 1):
            _cc = ws.cell(_rr, _comp_col_final)
            _dc = ws.cell(_rr, unit_id_col)
            if not isinstance(_cc, _MC_border) and not isinstance(_dc, _MC_border):
                _cc.border = _copy_border.copy(_dc.border)

    return written


def _write_spec_totals(
    ws, *, data_start: int, last_data_row: int,
    unit_id_col: int, no_col: int,
    stmt_label_col: int, stmt_value_col: int,
    branch_label_col: int, branch_value_col: int,
) -> None:
    """라운드 92 — 회사 감사본 4.Coverage TOTALS 3행 (Fail/Pass/Total) + 상단 요약.

    레퍼런스 (KJPDS02 v1.01) 실측 — H/L 열은 모두 수식:
      r580: D='Total'  G='Fail'  H==COUNTIF(H10:H579,"Fail")  K='Fail'  L=COUNTIF(...)
      r581:            G='Pass'  H==COUNTIF(H10:H579,"Pass")   K='Pass'  L=COUNTIF(...)
      r582:            G='Total' H==SUM(H580:H581)             K='Total' L=SUM(...)

    상단 요약 r5/r6 (양식 표준 v0.10 수식 `=E25`/`=H25` 등은 row 확장으로 cross-ref
    갱신되나 단일 TOTALS 행을 가정 → 3행 구조와 misalign). 레퍼런스 수식 패턴으로
    직접 덮어써 정합 보장:
      r5 Statement: E==B<last_no_row>(함수 수=No 마지막) F==H<row_fail> H==(E5-F5)/E5
      r6 Branch:    E==B<last_no_row>                   F==L<row_fail>

    `safe_write` 대신 직접 cell.value 할당 — 'Pass'/'Fail'/숫자 수식 모두 stamp.
    """
    row_fail = last_data_row + 1
    row_pass = last_data_row + 2
    row_total = last_data_row + 3

    def _set(rr, cc, val):
        try:
            ws.cell(rr, cc).value = val
        except (ValueError, AttributeError):
            pass

    # 라운드 93 fix — 레퍼런스는 H/L 열이 (Excel 캐시된) literal 값. openpyxl이 쓴
    # COUNTIF/SUM 수식은 캐시가 없어 파일 열기 전까지 공란("토탈결과 안 보임"). →
    # 데이터 행(H/L)에서 직접 Pass/Fail을 count해 **literal 값**을 stamp.
    def _count(value_col: int) -> tuple[int, int]:
        n_fail = n_pass = 0
        for rr in range(data_start, last_data_row + 1):
            v = ws.cell(rr, value_col).value
            s = str(v).strip().lower() if v is not None else ""
            if s == "fail":
                n_fail += 1
            elif s == "pass":
                n_pass += 1
        return n_fail, n_pass

    stmt_fail, stmt_pass = _count(stmt_value_col)
    br_fail, br_pass = _count(branch_value_col)
    stmt_total = stmt_fail + stmt_pass
    br_total = br_fail + br_pass

    # D 'Total' 라벨 (Unit ID col, 첫 row)
    _set(row_fail, unit_id_col, "Total")
    # Fail / Pass / Total row — literal count
    _set(row_fail, stmt_label_col, "Fail"); _set(row_fail, stmt_value_col, stmt_fail)
    _set(row_fail, branch_label_col, "Fail"); _set(row_fail, branch_value_col, br_fail)
    _set(row_pass, stmt_label_col, "Pass"); _set(row_pass, stmt_value_col, stmt_pass)
    _set(row_pass, branch_label_col, "Pass"); _set(row_pass, branch_value_col, br_pass)
    _set(row_total, stmt_label_col, "Total"); _set(row_total, stmt_value_col, stmt_total)
    _set(row_total, branch_label_col, "Total"); _set(row_total, branch_value_col, br_total)

    # 상단 요약 r5(Statement)/r6(Branch) — literal: E=Total, F=Fail Count,
    # G=Exception(=Fail, 레퍼런스 패턴), H=Coverage 비율((Total-Fail+Exception)/Total).
    # 레퍼런스: Fail이 전부 deviation/exception으로 간주되어 H=1.0.
    def _cov(total: int, fail: int) -> float:
        return round((total - fail + fail) / total, 4) if total else 0.0

    _set(5, 5, stmt_total); _set(5, 6, stmt_fail); _set(5, 7, stmt_fail); _set(5, 8, _cov(stmt_total, stmt_fail))
    _set(6, 5, br_total); _set(6, 6, br_fail); _set(6, 7, br_fail); _set(6, 8, _cov(br_total, br_fail))


# BLANK_MARKUP은 excel_template_utils에서 import (단일 출처).


_TC_FN_RE = re.compile(r"(SwUFn_\d+|SwITC_\d+(?:_\d+)?)")
# 라운드 73 P1 fix: SwIT TC name 'SwITC_NNNN' / 'SwITC_NNNN.NNN' 형식 매칭.
# 라운드 74 T901 fix: SwITC_NN_NN sub-index 보존 (`SwITC_3301_02` 형식). 회사
# KJPDS02 v1.01 SwITCV 2.Traceability가 sub-index 별 row stamp (R58 SwITC_3301_02 /
# R59 SwITC_3301_03 등 60 row × 110 col matrix). 이전 prefix 통합으로 12 row만 stamp
# → audit reviewer 검수에서 'sub-index 손실' 결함 발견.
# backward-compat: sub-index `_\d+` 선택 캡처. 'SwITC_01.001'은 그대로 SwITC_01,
# 'SwITC_01_02'는 SwITC_01_02로 sub-index 보존. SwUFn_ alternative는 그대로.


def _write_history_sheet(
    ws, history_rows: list[dict[str, str]], out_warnings: list[str] | None = None,
) -> int:
    """History 시트 — git log 자동 채움 (T134).

    회사 표준 History layout: Version / Date / Description / Author / Reviewer / Approver.
    헤더 행 다음부터 git log 결과를 row 단위로 작성.

    Returns:
        쓰여진 row 수.
    """
    if not ws or not history_rows:
        return 0
    # 헤더 행 찾기 — "Version" 라벨 위치
    header_pos = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "Version":
                header_pos = (row[0].row, cell.column)
                break
        if header_pos:
            break
    if header_pos is None:
        if out_warnings is not None:
            out_warnings.append("History 시트 'Version' 헤더 미발견 — git log 작성 skip")
        return 0

    start_row = header_pos[0] + 1
    col = header_pos[1]
    written = 0
    for i, h in enumerate(history_rows):
        r = start_row + i
        safe_write(ws, r, col, h.get("version", ""))
        safe_write(ws, r, col + 1, h.get("date", ""))
        safe_write(ws, r, col + 2, h.get("description", ""))
        safe_write(ws, r, col + 3, h.get("author", ""))
        safe_write(ws, r, col + 4, h.get("reviewer", ""))
        safe_write(ws, r, col + 5, h.get("approver", ""))
        written += 1
    return written


# 라운드 83 T1701: AuditLog 시트 layout 상수 — section header row 추적.
_AUDIT_LOG_TITLE = "ISO 26262 Audit Log — Auto-Generated by DevOps Release"
_AUDIT_LOG_MAX_WARNINGS = 20


def _build_audit_log_rows(
    meta: Any,
    summary: dict[str, Any],
    agg: dict[str, Any],
    session: SwUTSession,
    warnings: list[str] | None,
) -> list[tuple[str, str, str]]:
    """라운드 83 T1701: AuditLog 시트 row list 구축 (label, value, extra).

    6 섹션:
      1. 빌드 환경 (project_id/version/date/engineer/author/approver/timestamp/sha256)
      2. ISO 26262 추적성 체인 ASIL Source 활용도 (c_source/SUDS/SDS/SRS/SUTS/SwUDS)
      3. ASIL 등급 분포 (라운드 81 5단계 — D/C/B/A/QM/Total)
      4. 빌드 결과 통계 (envs/TCs/passed/failed/not_executed/function_count)
      5. Parse Warnings (top 20)
      6. Tool Qualification (evidence_class / ASIL_a_usage / round / CLAUDE.md ref)

    빈 tuple은 section separator (빈 row).
    """
    rows: list[tuple[str, str, str]] = []
    # Title
    rows.append((_AUDIT_LOG_TITLE, "", ""))
    rows.append(("", "", ""))

    # 1. 빌드 환경
    rows.append(("1. 빌드 환경", "", ""))
    rows.append(("Project ID", getattr(meta, "project_id", "") or "", ""))
    rows.append(("Project Full Name", getattr(meta, "project_full_name", "") or "", ""))
    rows.append(("Release SW Version", getattr(meta, "release_sw_version", "") or "", ""))
    rows.append(("Test Date", getattr(meta, "test_date", "") or "", ""))
    rows.append(("Validation Date", getattr(meta, "validation_date", "") or "", ""))
    rows.append(("Test Engineer", getattr(meta, "test_engineer", "") or "", ""))
    rows.append(("Author", getattr(meta, "default_author", "") or "", ""))
    rows.append(("Reviewer", getattr(meta, "default_reviewer", "") or "", ""))
    rows.append(("Approver", getattr(meta, "default_approver", "") or "", ""))
    rows.append(("Build Timestamp", str(summary.get("build_timestamp", "") or ""), ""))
    rows.append(("Template SHA256 (12)", str(summary.get("template_sha256_12", "") or ""), ""))
    rows.append(("", "", ""))

    # 2. ISO 26262 추적성 체인 ASIL Source 활용도
    rows.append(("2. ISO 26262 추적성 체인 ASIL Source 활용도", "", ""))
    rows.append((
        "C source @asil (라운드 30 W21)",
        str(_count_c_function_asil(getattr(session, "c_function_map", {}) or {})),
        "comment_asil 보유 함수",
    ))
    rows.append((
        "SUDS docx function ASIL (라운드 80)",
        str(len(agg.get("function_asil_from_suds") or {})),
        "SwUFn 직접 매핑",
    ))
    rows.append((
        "SDS docx component ASIL (라운드 80)",
        str(len(agg.get("component_asil_from_sds") or {})),
        "컴포넌트 단위 매핑",
    ))
    rows.append((
        "SRS docx 보조 ASIL (라운드 80)",
        str(len(agg.get("function_asil_from_srs") or {})),
        "함수명 보조 추출",
    ))
    rows.append((
        "SwUDS docx function map (라운드 32)",
        str(len(getattr(session, "swuds_function_map", {}) or {})),
        "heading+table 매핑",
    ))
    rows.append(("", "", ""))

    # 3. ASIL 등급 분포 (라운드 81 5단계)
    # 라운드 84 fix: _compute_asil_distribution은 "ASIL_A"/"UNKNOWN" key 사용 —
    # short key("A"/"B"...) fallback 적용.
    asil_dist = summary.get("asil_distribution") or {}
    total_asil = sum(asil_dist.values()) if asil_dist else 0
    rows.append(("3. ASIL 등급 분포 (라운드 81 5단계 그라데이션)", "", ""))
    for asil_key, label in (
        ("D", "ASIL D (MC/DC 필수, 빨강)"),
        ("C", "ASIL C (MC/DC 권장, 주황)"),
        ("B", "ASIL B (분기 필수, 파랑)"),
        ("A", "ASIL A (구문 충분, 녹색)"),
        ("QM", "QM (비안전, 회색)"),
    ):
        # _compute_asil_distribution → "ASIL_A"/"ASIL_QM" / fallback short key
        cnt = asil_dist.get(f"ASIL_{asil_key}", 0) or asil_dist.get(asil_key, 0)
        pct = f"{cnt / total_asil * 100:.1f}%" if total_asil else "0.0%"
        rows.append((label, str(cnt), pct))
    unknown = asil_dist.get("UNKNOWN", 0)
    if unknown:
        unknown_pct = f"{unknown / total_asil * 100:.1f}%" if total_asil else "0.0%"
        rows.append(("UNKNOWN (ASIL 미설정)", str(unknown), unknown_pct))
    rows.append(("Total", str(total_asil), "100.0%" if total_asil else "—"))
    rows.append(("", "", ""))

    # 라운드 86 T2002 + 87 T2101: 3-1. UNKNOWN 함수 list 분류 (c_only/stub/orphan).
    # 분류 규칙:
    #   c_only: c_function_map에 있음 → SUDS 등재 누락 (audit 추가 필요)
    #   stub:   `_` / `stub_` prefix → 자동 생성 (SUDS 등재 불필요)
    #   orphan: c source 부재 → vcast 결과만 (검토 필요)
    unmapped_fns = summary.get("unmapped_function_names") or []
    if unmapped_fns:
        c_fn_map = getattr(session, "c_function_map", {}) or {}
        classified = _classify_unmapped_functions(unmapped_fns, c_fn_map)
        c_only = classified["c_only"]
        stub = classified["stub"]
        orphan = classified["orphan"]
        rows.append((
            f"3-1. UNKNOWN 함수 list (audit 진단용, top {_AUDIT_LOG_MAX_WARNINGS})",
            f"총 {len(unmapped_fns)}건",
            f"c_only={len(c_only)} / stub={len(stub)} / orphan={len(orphan)}",
        ))
        idx = 1
        for cat, label in (("c_only", "SUDS 등재 누락"), ("stub", "자동 생성"), ("orphan", "c source 부재")):
            for fn in classified[cat][:_AUDIT_LOG_MAX_WARNINGS - idx + 1]:
                rows.append((f"U{idx}", f"[{cat}] {str(fn)[:160]}", label))
                idx += 1
                if idx > _AUDIT_LOG_MAX_WARNINGS:
                    break
            if idx > _AUDIT_LOG_MAX_WARNINGS:
                break
        if len(unmapped_fns) > _AUDIT_LOG_MAX_WARNINGS:
            rows.append((
                "...",
                f"외 {len(unmapped_fns) - _AUDIT_LOG_MAX_WARNINGS}건 생략",
                "",
            ))
        rows.append(("", "", ""))


    # 4. 빌드 결과 통계
    n_envs = summary.get("environments", 0)
    # SUTR summary는 'total' 키 사용 (Coverage는 'total_tcs') — robust fallback.
    total_tcs = summary.get("total_tcs") or summary.get("total") or agg.get("total_tcs", 0) or 0
    passed = summary.get("passed", 0)
    failed = summary.get("failed", 0)
    # SUTR summary는 function_count/function_rows 없음 — agg fallback.
    function_count = (
        summary.get("function_count")
        or agg.get("function_count")
        or len(agg.get("function_rows") or [])
        or 0
    )
    rows.append(("4. 빌드 결과 통계", "", ""))
    rows.append(("Environments", str(n_envs), ""))
    rows.append(("Total TCs", str(total_tcs), ""))
    pass_ratio = f"{passed / total_tcs * 100:.1f}%" if total_tcs else "—"
    fail_ratio = f"{failed / total_tcs * 100:.1f}%" if total_tcs else "—"
    rows.append(("Passed", str(passed), pass_ratio))
    rows.append(("Failed", str(failed), fail_ratio))
    rows.append(("Not Executed", str(summary.get("not_executed", 0)), ""))
    rows.append(("Function count", str(function_count), ""))
    rows.append(("", "", ""))

    # 5. Parse Warnings (top 20, silent skip 차단)
    warn_list = list(warnings or [])
    rows.append((
        f"5. Parse Warnings (top {_AUDIT_LOG_MAX_WARNINGS}, audit silent skip 차단)",
        f"총 {len(warn_list)}건", "",
    ))
    for i, w in enumerate(warn_list[:_AUDIT_LOG_MAX_WARNINGS], start=1):
        # 한 줄 cell 가독성 — 200자 절단
        rows.append((f"W{i}", str(w)[:200], ""))
    if len(warn_list) > _AUDIT_LOG_MAX_WARNINGS:
        rows.append((
            "...",
            f"외 {len(warn_list) - _AUDIT_LOG_MAX_WARNINGS}건 생략",
            "",
        ))
    rows.append(("", "", ""))

    # 6. Tool Qualification (ISO 26262)
    rows.append(("6. Tool Qualification (ISO 26262)", "", ""))
    rows.append(("Evidence Class", "auto-generated draft", ""))
    rows.append(("ASIL A Usage", "reviewer 검토 후 evidence 사용 가능", ""))
    rows.append(("ASIL B/C/D Usage", "단독 evidence 사용 금지 — manual review 의무", ""))
    rows.append((
        "Round",
        "R80 (SUDS/SDS/SRS) + R81 (5단계 그라데이션) + R82 (KJPDS02 자동 호환) + R83 (AuditLog)",
        "",
    ))
    rows.append(("CLAUDE.md Reference", "## ISO 26262 Safety Context + ## 시각 강조 정책", ""))

    return rows


def _classify_unmapped_functions(
    unmapped: list[str], c_function_map: dict[str, Any],
) -> dict[str, list[str]]:
    """라운드 87 T2101: UNKNOWN 함수 분류 — audit reviewer 진단 분류.

    Args:
        unmapped: unmapped 함수명 list (정렬 / dedup된 input).
        c_function_map: c_parser 결과 — comment_asil 없어도 c source 존재 확인용.

    Returns:
        ``{"c_only": [...], "stub": [...], "orphan": [...]}``
        - c_only: c source 존재 + SUDS 미등재 (SUDS docx 보강 필요)
        - stub: `_` / `stub_` prefix — 자동 생성 (정상 skip)
        - orphan: c source 부재 — vcast 결과만 (수동 검토)
    """
    result: dict[str, list[str]] = {"c_only": [], "stub": [], "orphan": []}
    for fn in unmapped:
        if fn.startswith(("_", "stub_")):
            result["stub"].append(fn)
        elif c_function_map and fn in c_function_map:
            result["c_only"].append(fn)
        else:
            result["orphan"].append(fn)
    return result


def _count_c_function_asil(c_function_map: dict[str, Any]) -> int:
    """라운드 83 T1701: c_function_map 중 comment_asil 보유 함수 수."""
    if not c_function_map:
        return 0
    cnt = 0
    for entry in c_function_map.values():
        if isinstance(entry, dict):
            asil = (entry.get("comment_asil") or "").strip().upper()
            if asil in {"A", "B", "C", "D", "QM"}:
                cnt += 1
    return cnt


def _write_audit_log_sheet(
    ws,
    meta: Any,
    summary: dict[str, Any],
    agg: dict[str, Any],
    session: SwUTSession,
    warnings: list[str] | None = None,
) -> int:
    """라운드 83 T1701: AuditLog 시트 — ISO 26262 audit metadata stamp.

    회사 양식 영향 0 (신규 시트 추가만, 기존 시트 보존). 6 섹션:
      1. 빌드 환경 / 2. ASIL source 활용도 / 3. ASIL 분포 /
      4. 빌드 통계 / 5. parse_warnings top 20 / 6. tool_qualification

    Args:
        ws: openpyxl Worksheet (신규 'AuditLog' 시트).
        meta: SutrBuildMeta / CoverageBuildMeta (project_id/version 등).
        summary: builder summary dict (asil_distribution / build_timestamp 등).
        agg: aggregate_session 결과 (ASIL maps).
        session: SwUTSession (c_function_map / swuds_function_map / parse_warnings).
        warnings: 외부 builder warnings 누적 list.

    Returns:
        쓰여진 row 수.
    """
    if ws is None:
        return 0
    # session.parse_warnings + builder warnings 통합
    all_warnings = list(getattr(session, "parse_warnings", []) or [])
    if warnings:
        all_warnings.extend(warnings)
    rows = _build_audit_log_rows(meta, summary, agg, session, all_warnings)
    for r_idx, (label, value, extra) in enumerate(rows, start=1):
        if label:
            safe_write(ws, r_idx, 1, label)
        if value:
            safe_write(ws, r_idx, 2, value)
        if extra:
            safe_write(ws, r_idx, 3, extra)
    return len(rows)


def _collect_tc_to_function(session: SwUTSession) -> dict[str, str]:
    """TC name → 함수 ID (`SwUFn_NNNN`).

    TC name 예:
      - SwUT: `SwUFn_0101.001` → `SwUFn_0101`
      - SwIT: `SwITC_SwUFn_0101.001` → `SwUFn_0101` (34차 deep-reviewer C1)

    34차 deep-reviewer C1 fix: `re.match`(^anchor)이 SwIT TC prefix `SwITC_`를
    거부해서 2.Consistency 시트 row 2가 항상 FAIL 반환 → 잘못된 audit evidence.
    `re.search`로 변경해 SwUT/SwIT TC 명명 모두 호환. SwUT TC도 prefix 없이
    시작하므로 search로 시작 위치 매칭 정상 (회귀 영향 없음).
    """
    out: dict[str, str] = {}
    for env in session.environments:
        for tc_name in env.test_cases:
            m = _TC_FN_RE.search(tc_name)
            if m:
                out[tc_name] = m.group(1)
    return out


def _compute_self_consistency(
    session: SwUTSession,
    swuds_function_ids: set[str] | None = None,
    *,
    test_kind: str = "SwUTS",
) -> list[dict[str, Any]]:
    """15차/16차: 자체 일관성 4가지 + SwUDS↔{test_kind} 매핑 (옵션) 검증.

    16차: ``swuds_function_ids`` 가 제공되면 row 5 'SwUDS↔{test_kind} 함수 ID 매핑' 추가.

    Args:
        test_kind: audit 라벨 — SwUT는 "SwUTS" (default), SwIT는 "SwIT" (34차 C2).
            row 5 item label과 _write_consistency_sheet intro 텍스트에 반영.

    Returns:
        list of {item, expected, actual, result, note}. result ∈ {PASS, FAIL}.
    """
    rows: list[dict[str, Any]] = []

    # 1. Function ID 일관성 — 모든 환경의 function_coverage가 비어있지 않음
    envs_with_fn = sum(1 for e in session.environments if e.function_coverage)
    total_envs = len(session.environments)
    rows.append({
        "item": "Function coverage data 완전성",
        "expected": f"{total_envs} 환경 모두 function_coverage 보유",
        "actual": f"{envs_with_fn}/{total_envs} 환경",
        "result": "PASS" if envs_with_fn == total_envs else "FAIL",
        "note": "비-zero function_coverage 환경 비율",
    })

    # 2. TC ↔ Function 매핑 — TC name이 SwUFn_NNNN.MMM 패턴 따름
    tc_to_fn = _collect_tc_to_function(session)
    all_tcs: set[str] = set()
    for env in session.environments:
        all_tcs.update(env.test_cases.keys())
    matched_pct = (len(tc_to_fn) / len(all_tcs) * 100) if all_tcs else 100.0
    rows.append({
        "item": "TC ↔ Function ID 패턴 일치",
        "expected": "100% TCs match SwUFn_NNNN.MMM",
        "actual": f"{len(tc_to_fn)}/{len(all_tcs)} ({matched_pct:.1f}%)",
        "result": "PASS" if len(tc_to_fn) == len(all_tcs) else "FAIL",
        "note": "패턴 불일치 시 trace 누락 위험",
    })

    # 3. TC execution coverage — test_results가 test_cases 전체를 cover
    tcs_with_result: set[str] = set()
    for env in session.environments:
        tcs_with_result.update(env.test_results.keys())
    missing = all_tcs - tcs_with_result
    rows.append({
        "item": "TC 실행 결과 완전성",
        "expected": f"{len(all_tcs)} TCs 모두 실행 결과 보유",
        "actual": f"{len(tcs_with_result)}/{len(all_tcs)} ({len(missing)} 누락)",
        "result": "PASS" if not missing else "FAIL",
        "note": "누락 TC는 SUTR Deviation 필요",
    })

    # 4. 환경별 TC 수 합 ↔ agg.total_tcs
    agg = aggregate_session(session)
    env_tc_sum = sum(len(e.test_cases) for e in session.environments)
    rows.append({
        "item": "TC 카운트 일관성 (env 합 ↔ aggregate)",
        "expected": str(env_tc_sum),
        "actual": str(agg["total_tcs"]),
        "result": "PASS" if env_tc_sum == agg["total_tcs"] else "FAIL",
        "note": "aggregate_session 무결성",
    })

    # 5. SwUDS ↔ {test_kind} 함수 ID 매핑 (16차) — swuds_function_ids 제공 시만.
    if swuds_function_ids is not None:
        swuts_fn_ids: set[str] = set()
        for env in session.environments:
            for fc in env.function_coverage:
                swuts_fn_ids.add(fc.unit_id)
        missing_in_swuts = swuds_function_ids - swuts_fn_ids  # SwUDS에 있고 {test_kind}에 없음
        extra_in_swuts = swuts_fn_ids - swuds_function_ids   # {test_kind}에 있고 SwUDS에 없음
        ok = not missing_in_swuts and not extra_in_swuts
        note_parts: list[str] = []
        if missing_in_swuts:
            note_parts.append(
                f"SwUDS 정의 미테스트: {sorted(missing_in_swuts)[:5]}"
                + (f" +{len(missing_in_swuts) - 5} more" if len(missing_in_swuts) > 5 else "")
            )
        if extra_in_swuts:
            note_parts.append(
                f"{test_kind} 추가 (SwUDS 미정의): {sorted(extra_in_swuts)[:5]}"
                + (f" +{len(extra_in_swuts) - 5} more" if len(extra_in_swuts) > 5 else "")
            )
        if not note_parts:
            note_parts.append("함수 ID 1:1 매칭")
        rows.append({
            "item": f"SwUDS ↔ {test_kind} 함수 ID 매핑",
            "expected": f"{len(swuds_function_ids)} 함수 (SwUDS)",
            "actual": f"{len(swuts_fn_ids)} 함수 ({test_kind})",
            "result": "PASS" if ok else "FAIL",
            "note": "; ".join(note_parts),
        })

    return rows


def _write_consistency_sheet_spec(
    ws,
    agg: dict[str, Any],
    name_to_swufn: dict[str, str],
    *,
    out_warnings: list[str] | None = None,
    test_kind: str = "SwUTS",
) -> int:
    """라운드 95 — KJPDS02 v1.01 SwUTCV 3.Consistency 레퍼런스 형식 작성 (spec_based).

    회사 레퍼런스((KJPDS02_DV_SwUTCV) Software Unit Test Coverage Result_v1.01) 구조:
      - r3~r7 요약: B3='Document Type' D3='SwUDS, SwUTS' / B4='Pass' D4=<O 개수> /
        B5='Fail' D5=<X 개수> / B6='Total' D6=<함수 수> / B7='Coverage' D7=<Pass/Total>.
        (라운드 94 정책: openpyxl 캐시 없음 회피 위해 수식 아닌 literal 값 stamp.)
      - r10 헤더: B=No / C=ID / D=Function Name / E='SwUDS와 SwUTS 항목 정합성 확인' /
        F=비고.
      - r11+: B=순번 / C='SwUTC_'+실 SwUFn ID (예 SwUTC_SwUFn_0121) / D=함수명 /
        E='O' (SwUDS 정의됨) 또는 'X' (미등재).

    함수 순서·SwUFn 매핑은 4.Coverage(_write_coverage_sheet 라운드 92) 및
    2.Traceability(_build_spec_swufn_order 라운드 93)와 **동일**하여 시트 간 일관성 유지.
    매핑 실패(SwUDS 미등재) 함수는 순차 SwUFn fallback + C셀 노란 마킹 (4.Coverage D셀과
    동일 추적성 — 추정 ID, audit reviewer 수동 검증 대상) + E='X'.

    Args:
        ws: 3.Consistency 시트.
        agg: ``aggregate_session`` 결과 dict (function_rows 보유).
        name_to_swufn: SwUDS 함수명→SwUFn ID 매핑 (4.Coverage/Traceability 공유).
        out_warnings: 진단 메시지 누적.
        test_kind: audit 라벨 (Document Type 표기용; KJPDS02는 'SwUDS, SwUTS').

    Returns:
        쓰여진 함수 list row 수.
    """
    function_rows = list(agg.get("function_rows") or [])
    if not function_rows:
        return 0

    # 안내문 (A1) — 레퍼런스 형식 spec_based 명시.
    safe_write(
        ws, 1, 1,
        f"본 시트는 SwUDS↔{test_kind} 함수 정합성 자동 검증 결과 (라운드 95 v1.01 "
        "레퍼런스 형식). 함수 순서·SwUFn ID는 4.Coverage / 2.Traceability 와 동일. "
        "'X' 행은 SwUDS 미등재 — reviewer 검토 + audit evidence 보강 필요.",
    )

    # 함수 순서·SwUFn ID 도출 — 4.Coverage 라운드 92 / _build_spec_swufn_order 와 동일 규칙.
    pass_count = 0
    fail_count = 0
    unmatched: list[str] = []
    rows_data: list[tuple[int, str, str, str, bool]] = []  # (no, swufn, name, result, matched)
    for i, fc in enumerate(function_rows):
        is_c_parser_only = bool(
            getattr(fc, "unit_id", "") and fc.unit_id.startswith("SwUFn_C_")
        )
        resolved = ""
        if not is_c_parser_only:
            resolved = name_to_swufn.get(getattr(fc, "name", ""), "") or name_to_swufn.get(
                getattr(fc, "unit_id", ""), ""
            )
        matched = bool(resolved)
        swufn_id = resolved or f"SwUFn_{i + 1:04d}"
        result = "O" if matched else "X"
        if matched:
            pass_count += 1
        else:
            fail_count += 1
            if len(unmatched) < 60:
                unmatched.append(getattr(fc, "name", "") or getattr(fc, "unit_id", "") or "<no-id>")
        rows_data.append((i + 1, swufn_id, getattr(fc, "name", "") or "", result, matched))

    total = len(rows_data)
    coverage = round(pass_count / total, 4) if total else 0

    # r3~r7 요약 (literal 값 — 라운드 94 정책, openpyxl 캐시 없음 회피).
    safe_write(ws, 3, 2, "Document Type")
    safe_write(ws, 3, 4, "SwUDS, SwUTS")
    safe_write(ws, 4, 2, "Pass")
    safe_write(ws, 4, 4, pass_count)
    safe_write(ws, 5, 2, "Fail")
    safe_write(ws, 5, 4, fail_count)
    safe_write(ws, 6, 2, "Total")
    safe_write(ws, 6, 4, total)
    safe_write(ws, 7, 2, "Coverage")
    safe_write(ws, 7, 4, coverage)

    # r10 헤더 (회사 v1.01 양식).
    header_row = 10
    data_start = 11
    safe_write(ws, header_row, 2, "No")
    safe_write(ws, header_row, 3, "ID")
    safe_write(ws, header_row, 4, "Function Name")
    safe_write(ws, header_row, 5, "SwUDS와 SwUTS 항목 정합성 확인")
    safe_write(ws, header_row, 6, "비고")

    # row 부족 시 자동 확장 (4.Coverage/Traceability 와 동일 헬퍼).
    needed_last_row = data_start + total - 1
    old_max = ws.max_row
    if needed_last_row > old_max:
        from backend.services.excel_template_utils import (
            auto_expand_row_block, push_sentinel_to_last_row,
            update_cross_refs_after_row_expansion,
        )
        shortage = needed_last_row - old_max
        inserted = auto_expand_row_block(
            ws,
            insert_at_row=data_start + 1,
            amount=shortage,
            template_row_idx=data_start,
            copy_style=True, copy_merge=True, copy_dimension=True,
        )
        if inserted < shortage and out_warnings is not None:
            out_warnings.append(
                f"[row_expand] 3.Consistency(spec) function list row 부족 "
                f"({shortage}개 필요, {inserted}개 확장) — 누락 가능"
            )
        try:
            push_sentinel_to_last_row(ws)
            update_cross_refs_after_row_expansion(
                ws, old_totals_row=old_max, new_totals_row=needed_last_row,
            )
        except Exception:  # noqa: BLE001 — graceful
            pass

    # data row stamp.
    written = 0
    for no, swufn_id, fn_name, result, matched in rows_data:
        r = data_start + (no - 1)
        safe_write(ws, r, 2, no)                       # B: No
        safe_write(ws, r, 3, f"SwUTC_{swufn_id}")      # C: SwUTC_<SwUFn>
        safe_write(ws, r, 4, fn_name)                  # D: Function Name
        safe_write(ws, r, 5, result)                   # E: 정합성 O/X
        # 매핑 실패(추정 SwUFn) → C셀 노란 마킹 (4.Coverage D셀과 동일 추적성).
        if not matched:
            from backend.services.excel_template_utils import _apply_fill
            from backend.services.design_tokens import USER_INPUT_FILL_RGB
            _apply_fill(ws, r, 3, USER_INPUT_FILL_RGB)
        written += 1

    if fail_count and out_warnings is not None:
        sample = ", ".join(unmatched[:10])
        more = f" 외 {fail_count - 10}건" if fail_count > 10 else ""
        out_warnings.append(
            f"[consistency_spec] SwUDS 미등재(X) {fail_count}건 (Pass {pass_count}/"
            f"{total}, Coverage {coverage}): {sample}{more}. SwUDS 버전 드리프트 가능 — "
            "audit reviewer 확인 의무."
        )

    return written


def _write_consistency_sheet(
    ws,
    session: SwUTSession,
    swuds_function_ids: set[str] | None = None,
    out_warnings: list[str] | None = None,
    *,
    test_kind: str = "SwUTS",
    agg: dict[str, Any] | None = None,
) -> int:
    """2.Consistency 시트 — {test_kind} 자체 일관성 + SwUDS↔{test_kind} 매핑 (16차).

    Args:
        swuds_function_ids: SwUDS docx에서 추출된 함수 ID set. 제공되면 row 5 추가.
        test_kind: audit 라벨 — SwUT는 "SwUTS" (default), SwIT는 "SwIT" (34차 C2 fix).
            intro 텍스트 + row 5 item label 동적 치환.
        agg: ``aggregate_session`` 결과 dict (라운드 95). spec_based(KJPDS02 SwUDS
            name→SwUFn 매핑 존재) 판정 + 4.Coverage/2.Traceability 와 동일 함수 순서
            도출에 사용. None이면 기존(HDPDM01/SwIT 자체-일관성) 동작 보존.

    Layout (비 spec_based): A1 = 안내, row 3 = 헤더, row 4부터 결과 row (4 또는 5개).
    Layout (spec_based, 라운드 95): 회사 KJPDS02 v1.01 SwUTCV 레퍼런스 형식 —
        r3~r7 요약(Document Type / Pass / Fail / Total / Coverage, literal 값) +
        r10 헤더(No/ID/Function Name/정합성/비고) + r11+ 함수 list (C=SwUTC_<SwUFn>,
        D=함수명, E=O/X). 함수 순서·SwUFn 매핑은 4.Coverage 와 동일.

    Returns:
        쓰여진 결과 row 수 (헤더 제외).
    """
    if not ws:
        return 0

    # 라운드 95 — KJPDS02 spec_based 게이트. SwUDS 함수명→SwUFn 매핑이 존재하면
    # 회사 레퍼런스(KJPDS02_DV_SwUTCV v1.01) 3.Consistency 형식으로 대체.
    # HDPDM01/SwIT(매핑 부재)는 기존 자체-일관성 형식 보존.
    _agg = agg if agg is not None else {}
    name_to_swufn_spec: dict[str, str] = _agg.get("function_name_to_swufn_from_suds") or {}
    spec_based_cons = bool(name_to_swufn_spec) and bool(_agg.get("function_rows"))
    if spec_based_cons:
        return _write_consistency_sheet_spec(
            ws, _agg, name_to_swufn_spec,
            out_warnings=out_warnings, test_kind=test_kind,
        )

    rows = _compute_self_consistency(
        session, swuds_function_ids=swuds_function_ids, test_kind=test_kind,
    )

    # 안내문 + 헤더 + data
    if swuds_function_ids is not None:
        intro = (
            f"본 시트는 {test_kind} 내부 자체 일관성 + SwUDS↔{test_kind} 함수 ID 매핑 "
            "자동 검증 결과 (16차 v3.02). FAIL 행은 reviewer 검토 + audit evidence 보강 필요."
        )
    else:
        intro = (
            f"본 시트는 {test_kind} 내부 자체 일관성 4 항목 자동 검증 결과. "
            f"SwUDS↔{test_kind} 함수 ID 매핑 비교는 swuds_docx_path 옵션 제공 시 "
            "자동 활성화 (16차)."
        )
    safe_write(ws, 1, 1, intro)
    safe_write(ws, 3, 1, "Item")
    safe_write(ws, 3, 2, "Expected")
    safe_write(ws, 3, 3, "Actual")
    safe_write(ws, 3, 4, "Result")
    safe_write(ws, 3, 5, "Note")

    written = 0
    for i, r in enumerate(rows):
        row_idx = 4 + i
        safe_write(ws, row_idx, 1, r["item"])
        safe_write(ws, row_idx, 2, r["expected"])
        safe_write(ws, row_idx, 3, r["actual"])
        safe_write(ws, row_idx, 4, r["result"])
        safe_write(ws, row_idx, 5, r["note"])
        # 23차 T192: FAIL row의 Result 셀 빨간 강조 — audit reviewer 가시성.
        if r["result"] == "FAIL":
            mark_fail_cell(ws, row_idx, 4)
        written += 1

    failed = [r["item"] for r in rows if r["result"] == "FAIL"]
    if failed and out_warnings is not None:
        out_warnings.append(
            f"2.Consistency 자체 일관성 FAIL {len(failed)}건: {', '.join(failed)}"
        )

    # 57차 T319 fix — 회사 v2.02 양식의 row 11+ SwUDS function list 자동 채움.
    # row 10 = 헤더 (No / ID / Function Name / SwUDS와 SwUTS 항목 정합성 확인 / 비고).
    # row 11+: 모든 환경의 function_coverage 추출 → fn_id + fn_name + 정합성 stamp.
    # 정확한 attr 이름: FunctionCoverage.unit_id (fn_id) + FunctionCoverage.name (fn_name).
    #
    # 라운드 76 자체평가 fix — 사용자 검수: "정합성 탭은 함수가 다 입력이 안 되어 있고".
    # 라운드 74 T907 롤백 후 4.Coverage만 c_parser merge 활성하고 3.Consistency
    # function list는 vcast 30 컴포넌트만 stamp. 회사 KJPDS02 v1.01 R578~R580 패턴
    # (`SwUTC_SwUFn_NNNN / s_FunctionName / O`)에 맞춰 ~377 함수 단위로 변경.
    # 라운드 76 enhance_function_coverage_with_file로 dedup key 정확성 향상 +
    # auto_expand + cross-ref formula 동적 갱신으로 양식 호환 유지.
    if session.environments:
        all_fns: dict[str, str] = {}  # unit_id → name
        for env in session.environments:
            for fc in env.function_coverage or []:
                unit_id = fc.unit_id  # 예: SwUFn_0101
                fn_name = fc.name      # 예: main
                if unit_id and unit_id not in all_fns:
                    all_fns[unit_id] = fn_name
        # 라운드 76 자체평가 fix — c_function_map 함수 자동 union (vcast에 없는 함수만).
        # SwUFn_C_<idx> synthetic unit_id 생성 (4.Coverage와 동일 패턴).
        c_fn_map_for_list = getattr(session, "c_function_map", None) or {}
        c_parser_added = 0
        if c_fn_map_for_list:
            next_idx = 9000
            existing_names = set(all_fns.values())
            for c_name in sorted(c_fn_map_for_list.keys()):
                if c_name in existing_names:
                    continue
                synthetic_uid = f"SwUFn_C_{next_idx}"
                all_fns[synthetic_uid] = c_name
                existing_names.add(c_name)
                next_idx += 1
                c_parser_added += 1
        swuds_set = swuds_function_ids or set()
        function_list_start = 11

        # 라운드 73 T802 — 2000 row hard limit 제거. auto_expand_row_block로 1회 batch.
        # 이전: `if row_idx_fn > 2000: break` 가 60+ 함수 silent truncate.
        sorted_fns = sorted(all_fns.items())
        needed_last_row = function_list_start + len(sorted_fns) - 1
        # 라운드 76 자체평가 fix — cross-ref formula 동적 갱신용 old_totals_row detect.
        old_totals_row_consistency = ws.max_row
        for _r in (4, 5, 6, 7):
            for _c in range(2, 8):
                try:
                    _v = ws.cell(_r, _c).value
                except (AttributeError, IndexError):
                    continue
                if not isinstance(_v, str) or not _v.startswith("=") or "!" in _v:
                    continue
                _m = re.search(r"=([A-Z]+)(\d+)\b", _v)
                if _m:
                    _detected = int(_m.group(2))
                    if _detected > 10:
                        old_totals_row_consistency = _detected
                        break
            if old_totals_row_consistency != ws.max_row:
                break

        if needed_last_row > ws.max_row:
            from backend.services.excel_template_utils import (
                auto_expand_row_block, push_sentinel_to_last_row,
                update_cross_refs_after_row_expansion,
            )
            shortage = needed_last_row - ws.max_row
            inserted = auto_expand_row_block(
                ws,
                insert_at_row=function_list_start + 1,
                amount=shortage,
                template_row_idx=function_list_start,
                copy_style=True, copy_merge=True, copy_dimension=True,
            )
            if inserted < shortage and out_warnings is not None:
                out_warnings.append(
                    f"[row_expand] Consistency 시트 function list row 부족 "
                    f"({shortage}개 필요, {inserted}개 확장) — 누락 가능"
                )
            # 라운드 76 자체평가 fix — cross-ref formula 동적 갱신 (3.Consistency도 동일).
            try:
                push_sentinel_to_last_row(ws)
                new_totals_row = function_list_start + len(sorted_fns)
                cr_updated = update_cross_refs_after_row_expansion(
                    ws,
                    old_totals_row=old_totals_row_consistency,
                    new_totals_row=new_totals_row,
                )
                if cr_updated > 0 and out_warnings is not None:
                    out_warnings.append(
                        f"[cross_ref] Consistency 시트 cross-ref formula {cr_updated}건 "
                        f"동적 갱신 (R{old_totals_row_consistency} → R{new_totals_row})."
                    )
            except ImportError:
                pass

        if c_parser_added > 0 and out_warnings is not None:
            out_warnings.append(
                f"[merge] 3.Consistency function list — c_parser only {c_parser_added}개 "
                "함수 추가 (SwUFn_C_9000~). coverage 미실측 — audit reviewer 확인 의무."
            )

        # 라운드 73 T812~T815 — 입력 자산 활용 stamp.
        # session에서 c_function_map / swuds_function_map 추출 (옵션).
        c_fn_map = getattr(session, "c_function_map", None) or {}
        swuds_fn_map = getattr(session, "swuds_function_map", None) or {}

        # 라운드 74 자체평가 fix — 헤더 stamp row 4 → 10 (양식 default function list
        # header row). 이전 R4 stamp는 자체 일관성 row 첫 번째 row와 섞여서 사용자
        # 검수에서 '헤더가 잘못된 row에 박혔다' 인지. R10이 'No/ID/Function Name/
        # SwUDS와 SwUTS 항목 정합성 확인/비고' 양식 header.
        if c_fn_map or swuds_fn_map:
            safe_write(ws, 10, 6, "Function Signature")  # F
            safe_write(ws, 10, 7, "C source desc")       # G
            safe_write(ws, 10, 8, "SwUDS heading")        # H
            safe_write(ws, 10, 9, "SwUDS desc")           # I

        for idx, (unit_id, fn_name) in enumerate(sorted_fns):
            row_idx_fn = function_list_start + idx
            safe_write(ws, row_idx_fn, 2, idx + 1)         # B: No
            safe_write(ws, row_idx_fn, 3, unit_id)         # C: Function ID
            safe_write(ws, row_idx_fn, 4, fn_name)         # D: Function Name
            # E: SwUDS↔SwUTS 정합성 — SwUDS function_ids set에 있으면 'O', 없으면 'X'
            in_swuds = unit_id in swuds_set if swuds_set else True
            safe_write(ws, row_idx_fn, 5, "O" if in_swuds else "X")

            # 라운드 73 T812 — C source signature stamp (F열).
            # unit_id 매칭 우선, fn_name fallback.
            c_fn = c_fn_map.get(unit_id) or c_fn_map.get(fn_name)
            if c_fn:
                sig = c_fn.get("signature") or ""
                if sig:
                    safe_write(ws, row_idx_fn, 6, sig[:200])  # F
                # 라운드 73 T814 — C comment_desc stamp (G열, 100자 truncate).
                desc = c_fn.get("comment_desc") or ""
                if desc:
                    safe_write(ws, row_idx_fn, 7, desc[:100] + ("..." if len(desc) > 100 else ""))

            # 라운드 73 T813/T815 — SwUDS heading + description stamp (H/I열).
            swuds_entry = swuds_fn_map.get(unit_id) or swuds_fn_map.get(fn_name)
            if swuds_entry:
                heading = swuds_entry.get("heading_text") or ""
                if heading:
                    safe_write(ws, row_idx_fn, 8, heading[:100])  # H
                swuds_desc = swuds_entry.get("description") or ""
                if swuds_desc:
                    safe_write(
                        ws, row_idx_fn, 9,
                        swuds_desc[:100] + ("..." if len(swuds_desc) > 100 else ""),
                    )  # I

    return written


def _build_spec_swufn_order(agg: dict[str, Any]) -> list[str]:
    """라운드 93 — KJPDS02 spec-based 2.Traceability 매트릭스 함수 순서 도출.

    4.Coverage 시트(_write_coverage_sheet)와 **동일한 함수 집합·순서**로 SwUFn
    ID list를 만든다. 두 시트가 같은 ``function_rows`` 를 공유해야 audit reviewer
    가 일관성을 확인할 수 있고, 회사 레퍼런스(KJPDS02_DV_SwUTCV v1.01)의 570
    diagonal 매트릭스와 차원이 일치한다.

    매핑 규칙 (4.Coverage 라운드 92 동작과 동일):
      1) SwUDS 함수명→ID 매핑(``function_name_to_swufn_from_suds``) 성공 → 실 SwUFn.
      2) 실패(SUDS 미등재) → 순차 ``SwUFn_NNNN`` fallback (4.Coverage D셀 노란
         마킹과 동일한 추정 ID — 추적성 수동 검증 대상).
      3) c_parser only row(unit_id ``SwUFn_C_`` prefix)는 fallback 순번 부여.

    Returns:
        ordered SwUFn ID list (function_rows 순서 보존). 빈 list면 spec-based 부적합.
    """
    function_rows = list(agg.get("function_rows") or [])
    name_to_swufn: dict[str, str] = agg.get("function_name_to_swufn_from_suds") or {}
    if not function_rows or not name_to_swufn:
        return []
    ordered: list[str] = []
    for i, fc in enumerate(function_rows):
        is_c_parser_only = bool(getattr(fc, "unit_id", "") and fc.unit_id.startswith("SwUFn_C_"))
        resolved = ""
        if not is_c_parser_only:
            resolved = name_to_swufn.get(getattr(fc, "name", ""), "") or name_to_swufn.get(
                getattr(fc, "unit_id", ""), ""
            )
        ordered.append(resolved or f"SwUFn_{i + 1:04d}")
    return ordered


def _write_traceability_spec_diagonal(
    ws, swufn_ids: list[str], *, out_warnings: list[str] | None = None,
) -> int:
    """라운드 93 — KJPDS02 v1.01 SwUTCV 2.Traceability diagonal 매트릭스 작성.

    회사 레퍼런스(KJPDS02_DV_SwUTCV v1.01) 2.Traceability 구조 (라이브 분석):
      - R12 col D.. : 함수별 헤더 (``SwUFn_NNNN``), 함수당 1 col.
      - R13 col D.. : count = 1 (``COUNTA`` 수식 채움).
      - R14.. : TC row. col A=순번 / col B=``SwUTC_<SwUFn>`` / col C=1 / 대각선 'O'.
    즉 함수 N개 → N col × N row 의 단위행렬(diagonal identity). 각 함수는 자신의
    단일 SwUTC TC를 가지며 대각 위치에 'O' 1개.

    템플릿(v0.10)은 419 함수만 정의 → SwUDS 전체(현재 데이터 571)보다 작아 ~143
    row / ~1300 cell 누락이 발생. 본 함수가 ``swufn_ids`` (4.Coverage와 동일 집합)
    로 헤더 col + diagonal row 를 **재작성**하여 레퍼런스 차원에 정렬한다.

    HDPDM01/SwIT 등 spec-based 가 아닌 호출은 본 함수를 타지 않는다(graceful).

    Args:
        ws: 2.Traceability 시트.
        swufn_ids: 함수 순서 SwUFn ID list (``_build_spec_swufn_order`` 출력).
        out_warnings: 진단 메시지 누적.

    Returns:
        쓰여진 'O' 셀 수 (= len(swufn_ids), 정상 시).
    """
    if not ws or not swufn_ids:
        return 0
    from backend.services.excel_template_utils import (
        auto_expand_row_block, clear_data_range, push_sentinel_to_last_row,
    )

    ws_title = getattr(ws, "title", "Traceability").strip()
    header_row = 12        # 회사 v1.01 양식 고정 (Cover/요약 R1~R11 보존)
    count_row = 13
    data_start = 14
    first_col = 4          # col D
    n = len(swufn_ids)
    needed_last_row = data_start + n - 1
    needed_last_col = first_col + n - 1

    # 1) 기존 template diagonal 영역 clear (헤더 col + data row + count row).
    #    R1~R11 (요약/Note) 및 col A~C 라벨 수식은 보존 — header/count/diagonal만 재작성.
    try:
        clear_data_range(
            ws,
            start_row=header_row, end_row=ws.max_row,
            start_col=first_col, end_col=max(ws.max_column or first_col, needed_last_col),
            preserve_formula=False, preserve_merged_anchor=True,
            sentinel_patterns=["End of Document", "Appendix", "TOTALS"],
        )
        # col A(순번)/B(라벨)/C(count) 의 data row 영역도 clear (template 419개 라벨 제거).
        clear_data_range(
            ws,
            start_row=data_start, end_row=ws.max_row,
            start_col=1, end_col=3,
            preserve_formula=True, preserve_merged_anchor=True,
            sentinel_patterns=["End of Document", "Appendix", "TOTALS"],
        )
    except Exception:  # noqa: BLE001 — clear 실패해도 stamp는 진행 (graceful)
        pass

    # 2) row 부족 시 자동 확장 (template 데이터 row 수 < n).
    if needed_last_row > ws.max_row:
        try:
            shortage = needed_last_row - ws.max_row
            auto_expand_row_block(
                ws,
                insert_at_row=ws.max_row + 1,
                amount=shortage,
                template_row_idx=data_start,
                copy_style=True, copy_merge=True, copy_dimension=True,
            )
        except Exception:  # noqa: BLE001
            pass

    # 3) 헤더 col(R12) + count(R13) + diagonal row(R14..) stamp.
    written = 0
    for i, fn_id in enumerate(swufn_ids):
        col = first_col + i
        r = data_start + i
        safe_write(ws, header_row, col, fn_id)     # R12 함수 헤더
        safe_write(ws, count_row, col, 1)          # R13 count
        safe_write(ws, r, 1, i + 1)                # col A 순번
        safe_write(ws, r, 2, f"SwUTC_{fn_id}")     # col B TC 라벨
        safe_write(ws, r, 3, 1)                    # col C count
        if safe_write(ws, r, col, "O"):            # 대각 'O'
            written += 1

    # 4) 요약 count 셀(C12=ID 수 / B13=TC 수) COUNTA 수식을 신규 차원으로 갱신.
    #    레퍼런스(KJPDS02 v1.01)는 C12=`=COUNTA(D12:VA12)` / B13=`=COUNTA(B14:B583)`.
    #    템플릿(419)의 stale 리터럴(419/418) 대신 신규 range 수식으로 재작성해
    #    함수 수가 자동 재계산되도록 한다.
    from openpyxl.utils import get_column_letter as _gcl
    last_col_letter = _gcl(needed_last_col)
    safe_write(ws, header_row, 3, f"=COUNTA({_gcl(first_col)}{header_row}:{last_col_letter}{header_row})")
    safe_write(ws, count_row, 2, f"=COUNTA(B{data_start}:B{needed_last_row})")

    try:
        push_sentinel_to_last_row(ws)
    except Exception:  # noqa: BLE001
        pass

    if out_warnings is not None:
        out_warnings.append(
            f"[trace] {ws_title} spec-based diagonal 재작성 — {n} 함수 × {n} TC "
            f"(레퍼런스 KJPDS02 v1.01 양식 정렬). 'O' {written}건. "
            "함수 집합·순서는 4.Coverage 와 동일(function_rows). 순차 SwUFn fallback "
            "행은 4.Coverage D셀 노란 마킹과 동일 추적성 검증 대상."
        )
    return written


def _write_traceability_sheet(
    ws, session: SwUTSession, out_warnings: list[str] | None = None,
    *, layout: Any = None, swits_tc_ids: list[str] | None = None,
    agg: dict[str, Any] | None = None,
) -> int:
    """1.Traceability 시트 — TC × Function 매트릭스 본격 작성 (T133).

    시트 헤더 행에서 `SwUFn_NNNN` 컬럼 위치 lookup → 각 TC 행에 'O' 표시.
    헤더 미발견 시 BLANK_MARKUP 유지.

    58차 F2: SwIT v2.02 양식은 헤더 row 위치가 v3.01 (1~10)보다 아래쪽 (20~25)에
    있을 수 있고 SwUFn_ 컬럼 개수도 더 작음. layout.traceability_header_row 제공
    시 그 row를 헤더로 강제. fallback 시 max_row=30 확장 + SwUFn_ 임계 50→5로 완화.

    Args:
        ws: 1.Traceability 시트.
        session: SwUTSession.
        out_warnings: 누락/실패 메시지 누적.
        layout: Optional[SwitLayout] — traceability_header_row 보유 시 우선.
        swits_tc_ids: 라운드 73 T807 — SwITCV switc_x_swst 분기 시 SwITS spec의
            TC ID list (예: 77 entries) 제공 시 session 12 TC만이 아닌 spec 전체
            stamp + Note column에 audit 안내. session에 없는 SwITS entry는
            'audit reviewer 수동 확인 — SwITS spec entry, vcast log 결과 미생성'
            메시지 stamp. None이면 session 기반 stamp만 (기존 동작).

    Returns:
        쓰여진 'O' 셀 수. 0이면 매트릭스 미작성.
    """
    if not ws:
        return 0

    # 59차 F4-C → 60차 F6-B 갱신 — KJPDS02 v1.01 양식 matrix kind 분기.
    # "switc_x_swst" 양식은 row = SwITC ID, col = SwST/SwSTR.
    # F6-B 라이브 분석 (T411) 결과: 검증된 양식 (KJPDS02 SwITS v1.01 + HDPDM01
    # SITS v2.02) 모두 'Integration Strategy' 시트가 SwST matrix가 아닌
    # **call graph (depth 1~14 tree)** 양식. SwITC×SwST traceability matrix 자체
    # 가 시방서에 존재하지 않음 — F4-C skip이 양식 설계상 정답.
    # 본 메시지는 audit reviewer가 'matrix 미stamp' 정상 동작임을 인지하도록 명시.
    # 다른 회사 양식 도입 시 재검증 필요.
    matrix_kind = (
        getattr(layout, "traceability_matrix_kind", "swufn_x_env") if layout is not None
        else "swufn_x_env"
    )
    if matrix_kind == "switc_x_swst":
        # 라운드 F7 stage 8 T705 부분 구현: SwITCV 2.Traceability 양식 default
        # SwITC row + 'O' 마킹 clear (false audit 차단) + 신규 session의 SwITC TC
        # 를 row stamp + count='1'. SwST × SwITC 'O' stamp는 SwITS spec 미제공
        # → skip + 명확한 warning (사용자에게 "T705 partial — SwST 매핑 부재" 안내).
        ws_title = getattr(ws, "title", "Traceability").strip()

        # 1) header row 찾기 — R11 SwST_01~SwSTR_NN header
        header_row_idx = None
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
            swst_cols = sum(
                1 for c in row
                if isinstance(c.value, str)
                and (str(c.value).strip().startswith(("SwST_", "SwSTR_")))
            )
            if swst_cols >= 3:
                header_row_idx = row[0].row
                break
        if header_row_idx is None:
            if out_warnings is not None:
                out_warnings.append(
                    f"{ws_title} switc_x_swst header (SwST_/SwSTR_) 미발견 — skip"
                )
            return 0

        # 2) data start = header + 2 (R12 count + R13~ SwITC row)
        data_start = header_row_idx + 2
        # session에서 unique SwITC ID 추출
        tc_to_fn = _collect_tc_to_function(session)
        # SwITC TC name 패턴 'SwITC_NN' 또는 'SwITC_SwUFn_NNNN.NNN' → SwITC ID prefix 추출
        import re as _re
        # 패턴 우선순위 (F7 stage 8 T705):
        # 1) SwITC_NN / SwITC_NN_NN (KJPDS02 회사 표준) — L975 inline regex 사용
        # 2) SwITC_SwUFn_NNNN.NNN (VectorCAST SwIT — SwUFn 부분 NNNN 추출 → SwITC_NNNN)
        # 3) SwUFn_NNNN.NNN (SwUT session — SwITC_NNNN 변환, F6-A 패턴)
        # F7 Round 3 N6 fix: dead code `_SWITC_DIRECT` 제거 (L975 inline 사용).
        _SWITC_WITH_FN = _re.compile(r"^SwITC_SwUFn_(\d+)")
        _SWUFN_ONLY = _re.compile(r"^SwUFn_(\d+)")
        switc_ids: list[str] = []
        seen: set[str] = set()
        for tc_name in tc_to_fn.keys():
            # 라운드 74 T902 — SwITC_NN_NN sub-index 보존. 회사 KJPDS02 v1.01 SwITCV
            # 2.Traceability는 sub-index별 row stamp (R58 'SwITC_3301_02' / R59
            # 'SwITC_3301_03' 등 unique row). 이전 prefix만 추출 (`SwITC_05_01` →
            # `SwITC_05`)은 다수 sub-index가 1 row로 합쳐져 60 row가 12로 줄어드는
            # 사용자 검수 결함.
            sid = None
            m = _re.match(r"^(SwITC_\d+(?:_\d+)?)", tc_name)
            if m:
                sid = m.group(1)
            else:
                fm = _SWITC_WITH_FN.match(tc_name)
                if fm:
                    sid = f"SwITC_{fm.group(1)}"
                else:
                    um = _SWUFN_ONLY.match(tc_name)
                    if um:
                        sid = f"SwITC_{um.group(1)}"
            if sid and sid not in seen:
                seen.add(sid)
                switc_ids.append(sid)
        switc_ids.sort()

        # 3) 양식 default clear (data_start ~ max_row, col 1~5: No/ID/Count + O 마킹 일부)
        try:
            from backend.services.excel_template_utils import clear_data_range
            cleared = clear_data_range(
                ws,
                start_row=data_start, end_row=ws.max_row,
                start_col=1, end_col=ws.max_column or 50,
                preserve_formula=True, preserve_merged_anchor=True,
                sentinel_patterns=["End of Document", "Appendix", "TOTALS"],
            )
            if out_warnings is not None and cleared > 0:
                out_warnings.append(
                    f"[clear] {ws_title} (SwITC×SwST matrix) 양식 default "
                    f"{cleared} cell clear — 신규 session SwITC {len(switc_ids)}건 stamp"
                )
        except ImportError:
            pass

        # 라운드 73 T807 — SwITS spec entries 활용 확장.
        # session에 미생성된 SwITS spec TC도 row stamp + Note column 안내.
        session_sid_set = set(switc_ids)
        spec_only_sids: list[str] = []
        if swits_tc_ids:
            for swits_tc in swits_tc_ids:
                # SwITS xlsm의 tc_id가 'SwITC_NN' / 'SwITC_NN_NN' / 'SwITC_SwUFn_NNNN' 패턴.
                # 라운드 74 T902 — sub-index 보존 (`SwITC_3301_02` 형식).
                m_swits = _re.match(r"^(SwITC_\d+(?:_\d+)?)", swits_tc)
                if m_swits:
                    sid_norm = m_swits.group(1)
                else:
                    fm = _SWITC_WITH_FN.match(swits_tc)
                    sid_norm = f"SwITC_{fm.group(1)}" if fm else swits_tc
                if sid_norm not in session_sid_set and sid_norm not in spec_only_sids:
                    spec_only_sids.append(sid_norm)
            spec_only_sids.sort()

        all_sids = switc_ids + spec_only_sids
        # row 자동 확장 — session + spec 합산이 ws.max_row 초과 시 신규 row.
        needed_last_row = data_start + len(all_sids) - 1
        if needed_last_row > ws.max_row:
            try:
                from backend.services.excel_template_utils import (
                    auto_expand_row_block, push_sentinel_to_last_row,
                )
                shortage = needed_last_row - ws.max_row
                auto_expand_row_block(
                    ws,
                    insert_at_row=data_start + 1,
                    amount=shortage,
                    template_row_idx=data_start,
                    copy_style=True, copy_merge=True, copy_dimension=True,
                )
                push_sentinel_to_last_row(ws)
            except ImportError:
                pass

        # 4) 신규 session의 SwITC row stamp (C2=No, C3=ID, C4=Count)
        # 회사 표준 SwITCV R13~: C2='SwITC_01' (ID), C3='3' (count)
        for i, sid in enumerate(switc_ids):
            r = data_start + i
            safe_write(ws, r, 1, i + 1)        # No
            safe_write(ws, r, 2, sid)          # SwITC ID
            safe_write(ws, r, 3, 1)            # Count (각 SwITC당 1)

        # 라운드 73 T807 — spec-only row (session 미생성) stamp + Note 안내.
        for i, sid in enumerate(spec_only_sids):
            r = data_start + len(switc_ids) + i
            safe_write(ws, r, 1, len(switc_ids) + i + 1)  # No 이어서
            safe_write(ws, r, 2, sid)                      # SwITC ID
            safe_write(ws, r, 3, 0)                        # Count = 0 (session 미실행)
            safe_write(
                ws, r, 4,
                "▶ audit reviewer 수동 확인 — SwITS spec entry, vcast log 결과 미생성",
            )

        if out_warnings is not None:
            note = (
                f"{ws_title} matrix kind 'switc_x_swst' partial stamp — session SwITC "
                f"{len(switc_ids)}건 + SwITS spec-only {len(spec_only_sids)}건 row stamp. "
                "SwST × SwITC 'O' 마킹은 SwITS spec에 SwST 매핑 부재 (T705 full — 회사 "
                "양식 SwITS spec에 SwST 컬럼 정의 부재). audit reviewer는 SwST 매핑 manual "
                "확인 의무. spec-only row는 vcast log 미생성 — 결과 누락 진단 필요."
            )
            out_warnings.append(note)
        return len(all_sids)

    # 라운드 93 — KJPDS02 spec-based 2.Traceability diagonal 재작성.
    #   회사 레퍼런스(KJPDS02_DV_SwUTCV v1.01)는 SwUDS 전체 함수(570) × TC diagonal
    #   매트릭스. 표준 템플릿(v0.10)은 419 함수만 정의 → ~143 row / ~1300 cell 누락.
    #   4.Coverage(라운드 92 spec_based)와 동일한 function_rows 집합으로 헤더 col +
    #   대각 row 를 재작성하여 레퍼런스 차원에 정렬한다.
    #   gate: agg에 function_name_to_swufn_from_suds 매핑 + function_rows 존재할 때만
    #   (HDPDM01/SwIT v3.01/v2.02 — 매핑 없음 — 은 기존 template-header 경로 100% 보존).
    if agg is not None and matrix_kind == "swufn_x_env":
        spec_swufn_order = _build_spec_swufn_order(agg)
        if spec_swufn_order:
            return _write_traceability_spec_diagonal(
                ws, spec_swufn_order, out_warnings=out_warnings,
            )

    # 58차 F2: layout 제공 시 traceability_header_row 강제. fallback은 자동 탐색.
    header_row_idx = None
    header_cols: dict[str, int] = {}
    layout_header_row = (
        getattr(layout, "traceability_header_row", None) if layout is not None else None
    )
    if layout_header_row is not None:
        # 강제 헤더 row — 그 row에서 SwUFn_/SwUTC_/SwITC_ prefix col 수집
        for cell in next(
            ws.iter_rows(min_row=layout_header_row, max_row=layout_header_row,
                         values_only=False),
            [],
        ):
            v = cell.value
            if isinstance(v, str):
                s = v.strip()
                if _TC_FN_RE.fullmatch(s) or s.startswith(("SwUTC_", "SwITC_", "SwUFn_")):
                    header_cols[s] = cell.column
        if header_cols:
            header_row_idx = layout_header_row

    if header_row_idx is None:
        # 자동 탐색 — max_row 20 → 30 확장, SwUFn_ 임계 50 → 5 완화 (SwIT v2.02 대응).
        for row in ws.iter_rows(min_row=1, max_row=30, values_only=False):
            cols: dict[str, int] = {}
            for cell in row:
                v = cell.value
                if isinstance(v, str) and _TC_FN_RE.fullmatch(v.strip()):
                    cols[v.strip()] = cell.column
            if len(cols) >= 5:
                header_row_idx = row[0].row
                header_cols = cols
                break

    if header_row_idx is None:
        if out_warnings is not None:
            out_warnings.append(
                "1.Traceability 헤더(SwUFn_xxxx 행) 미발견 — placeholder 유지"
            )
        safe_write(ws, 1, 1, BLANK_MARKUP)
        return 0

    # 2) 기존 TC 행 위치 인덱싱 — SwUTC_SwUFn_xxxx.NNN 또는 SwUFn_xxxx.NNN
    # 라운드 73 T804 — scan 한계 제거: data_start + 600 → ws.max_row.
    data_start = header_row_idx + 1
    tc_row_index: dict[str, int] = {}
    for row in ws.iter_rows(
        min_row=data_start, max_row=ws.max_row, values_only=False,
    ):
        for cell in row[:5]:
            v = cell.value
            if isinstance(v, str):
                stripped = v.strip()
                if stripped.startswith(("SwUTC_SwUFn_", "SwUFn_")):
                    tc_row_index[stripped] = row[0].row
                    break

    # 3) 우리 session TC name → 함수 매핑 + 'O' 표시.
    # T136: 회사 v3.01 row label은 `SwUTC_<fn_id>` (인덱스 `.NNN` 없음).
    # `SwUTC_<tc_name>` (인덱스 포함) 과 `<tc_name>` 도 fallback 시도.
    tc_to_fn = _collect_tc_to_function(session)

    # 라운드 73 T804 — header_cols에 매칭되는 fn_id 보유 신규 TC row 확장.
    # 양식에 미리 정의되지 않은 TC는 row insert + label stamp → 'O' 매트릭스 stamp 가능.
    missing_tcs_with_col: list[tuple[str, str]] = []  # [(tc_name, fn_id)]
    for tc_name, fn_id in tc_to_fn.items():
        if header_cols.get(fn_id) is None:
            continue
        # 기존 tc_row_index에 fn_id가 있는지 확인 (회사 표준 / 인덱스 포함 / native)
        if (f"SwUTC_{fn_id}" in tc_row_index
                or f"SwUTC_{tc_name}" in tc_row_index
                or tc_name in tc_row_index):
            continue
        missing_tcs_with_col.append((tc_name, fn_id))

    if missing_tcs_with_col:
        try:
            from backend.services.excel_template_utils import (
                auto_expand_row_block, push_sentinel_to_last_row,
            )
            last_existing_row = max(tc_row_index.values()) if tc_row_index else data_start
            inserted = auto_expand_row_block(
                ws,
                insert_at_row=last_existing_row + 1,
                amount=len(missing_tcs_with_col),
                template_row_idx=last_existing_row,
                copy_style=True, copy_merge=True, copy_dimension=True,
            )
            for i, (tc_name, fn_id) in enumerate(missing_tcs_with_col):
                new_row = last_existing_row + 1 + i
                if safe_write(ws, new_row, 2, f"SwUTC_{fn_id}"):
                    tc_row_index[f"SwUTC_{fn_id}"] = new_row
            # End-of-Document sentinel 마지막 row로 push
            push_sentinel_to_last_row(ws)
            if out_warnings is not None:
                out_warnings.append(
                    f"[row_expand] Traceability 신규 TC {inserted}개 row 추가 + label stamp"
                )
        except ImportError:
            pass

    written = 0
    matched_fn: set[str] = set()
    for tc_name, fn_id in tc_to_fn.items():
        col = header_cols.get(fn_id)
        if col is None:
            continue
        row_idx = (
            tc_row_index.get(f"SwUTC_{fn_id}")          # 회사 표준 (인덱스 없음)
            or tc_row_index.get(f"SwUTC_{tc_name}")     # 인덱스 포함 형식
            or tc_row_index.get(tc_name)                # 우리 빌더 native
        )
        if row_idx is None:
            continue
        # 같은 fn_id 가 여러 TC index를 가질 때 첫 매칭만 — 회사 시트 row 1개당 'O' 1개로 충분.
        if fn_id in matched_fn:
            continue
        if safe_write(ws, row_idx, col, "O"):
            written += 1
            matched_fn.add(fn_id)

    if written == 0:
        if out_warnings is not None:
            out_warnings.append(
                "1.Traceability — 헤더는 발견했으나 TC 매칭 0건 (회사 시트 row 명명 차이)"
            )
        safe_write(ws, 1, 1, BLANK_MARKUP)
    return written


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_coverage_report(
    session: SwUTSession,
    meta: CoverageBuildMeta,
    template_bytes: bytes,
    swuds_function_ids: set[str] | None = None,
    hmr_html_bytes: bytes | None = None,
) -> CoverageBuildResult:
    """Coverage Report v3.01 xlsx 생성.

    Args:
        session: SwUT 데이터 (input_adapter 출력).
        meta: 빌드 메타 (Project/ASIL/Author 등).
        template_bytes: 기존 v3.01 xlsx 파일 bytes (template).
        swuds_function_ids: 16차 — SwUDS 함수 ID set (옵션). 제공되면 2.Consistency에
            'SwUDS↔SwUTS 함수 ID 매핑' row 5 추가 + incomplete_sheets에서 partial 라벨 제거.
        hmr_html_bytes: 60차 F6-C — VectorCAST aggregate metrics report HTML
            (옵션, Jenkins_PDSM_UT/IT_metrics_report.html 양식). 제공 시 함수별
            Function Calls coverage를 추출하여 fc.function_calls_coverage 채움.
            KJPDS02 v1.01 양식의 row 6 'Function Calls' stamp source. None이면
            기존 빈 CoverageStats default 유지 (backward-compat, v2.02/v3.01 동작).

    Returns:
        CoverageBuildResult — xlsx_io 채워짐.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for SwUT Coverage Report builder")

    # deep-reviewer X3 + 5차 H1/H2: 입력 메타 종합 검증.
    validate_build_meta(
        meta.release_sw_version, meta.test_date,
        doc_id_sequence=meta.doc_id_sequence,
        test_engineer=meta.test_engineer,
        author=meta.default_author,
        approver=meta.default_approver or meta.approver_override,
        reviewer=meta.default_reviewer or meta.reviewer_override,
    )
    # Critical (reviewer S): ZIP bomb / magic byte 검증.
    validate_xlsx_template_bytes(template_bytes, label="Coverage Report template")

    # 5차 L1: 입력 template hash — audit 추적성.
    template_sha256_12 = hashlib.sha256(template_bytes).hexdigest()[:12]

    # 37차 fix → 38차 W1 DRY: extract_warnings_from_session helper로 추출.
    warnings: list[str] = extract_warnings_from_session(session)

    # 라운드 73 T816 — 입력 자산 활용도 진단 (c_function_map / swuds_function_map은 SwUTSession에서 추출).
    from backend.services.swut_builder_helpers import diagnose_asset_usage
    warnings.extend(diagnose_asset_usage(
        c_function_map=session.c_function_map or None,
        swuds_function_map=session.swuds_function_map or None,
    ))

    # 54-fix C1: SwUT 라우터에 v2.02 template 잘못 입력되더라도 silent 빈 셀 차단.
    # v3.01 양식은 fallback_to_v301=True로 기존 hardcode 동작과 동등 (회귀 zero 영향).
    # v2.02 양식 (사용자 실수 또는 의도된 mixed)이면 SW Version 등 v2.02 라벨 매핑.
    from backend.services.excel_layout_resolver import inspect_swit_layout
    layout = inspect_swit_layout(template_bytes, "coverage")
    if layout.warnings:
        warnings.extend([f"[layout] {w}" for w in layout.warnings])

    wb: Workbook = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=False)
    sheet_names = wb.sheetnames

    agg = aggregate_session(session)

    # 60차 F6-C — HMR HTML 제공 시 함수별 Function Calls coverage 채움.
    # VectorCAST aggregate metrics report (Jenkins_PDSM_UT/IT_metrics_report)
    # 양식 → 함수명 매칭으로 fc.function_calls_coverage stamp. 매칭 안 되면 skip
    # (graceful — v2.02/v3.01 빈 cell default 유지).
    if hmr_html_bytes:
        from backend.services.swut_input_adapter import CoverageStats
        from backend.services.vcast_hmr_parser import parse_hmr_html
        hmr_parse_warnings: list[str] = []
        hmr_result = parse_hmr_html(
            hmr_html_bytes, parse_warnings=hmr_parse_warnings,
        )
        if hmr_parse_warnings:
            warnings.extend([f"[hmr] {w}" for w in hmr_parse_warnings])
        if hmr_result.ok:
            # F6 자체평가 Round 1 W2 fix: dataclasses.replace + 새 list로 session
            # 객체 mutation 차단 (향후 session caching 도입 시 silent regression 방지).
            from dataclasses import replace as _dc_replace
            original_rows: list[FunctionCoverage] = agg.get("function_rows") or []
            new_function_rows: list[FunctionCoverage] = []
            stamped = 0
            ambiguous = 0
            for fc in original_rows:
                # F6 Round 1 C2 + W3 fix: metrics_by_name 사용 — 함수명 중복 시
                # ambiguous skip + warning. fc.unit_id dead fallback 제거 (HMR key는
                # 실제 C 함수명, fc.unit_id는 SwUFn_NNNN — 본질적으로 불일치).
                candidates = hmr_result.metrics_by_name.get(fc.name, [])
                if len(candidates) > 1:
                    ambiguous += 1
                    _files = ", ".join(sorted({c.unit_file for c in candidates}))
                    warnings.append(
                        f"[hmr] ambiguous function '{fc.name}' — 다중 unit_file "
                        f"({_files}) 매칭. silent wrong-pick 방지 위해 stamp skip"
                    )
                    new_function_rows.append(fc)
                    continue
                m = candidates[0] if candidates else None
                if m and m.total_calls > 0:
                    new_function_rows.append(_dc_replace(
                        fc,
                        function_calls_coverage=CoverageStats(
                            covered=m.covered_calls,
                            total=m.total_calls,
                            coverage_pct=m.coverage_pct / 100.0,
                        ),
                    ))
                    stamped += 1
                else:
                    new_function_rows.append(fc)
            warnings.append(
                f"[hmr] Function Calls metric stamped — {stamped}/{len(original_rows)} "
                f"functions matched (HMR metric count: {len(hmr_result.metrics)}, "
                f"ambiguous skipped: {ambiguous})"
            )
            # 새 list로 교체 — 이후 3.Coverage sheet writer가 stamped 값 사용,
            # session.environments[].function_coverage는 unchanged (W2 격리).
            agg["function_rows"] = new_function_rows

    # 30차 W21 + 31차 W29 + 라운드 84 T1801 + 85 T1903 + 86 T2001: unmapped fc list.
    asil_distribution, ids_by_asil, unmapped_fns = _compute_asil_distribution(
        agg.get("function_rows") or [],
        agg.get("function_asil_map") or {},
        function_asil_from_suds=agg.get("function_asil_from_suds"),
        component_asil_from_sds=agg.get("component_asil_from_sds"),
        function_asil_from_srs=agg.get("function_asil_from_srs"),
        function_name_to_swufn_from_suds=agg.get("function_name_to_swufn_from_suds"),
    )

    summary = {
        "environments": len(session.environments),
        "total_tcs": agg["total_tcs"],
        "passed": agg["passed"],
        "failed": agg["failed"],
        "function_rows": agg["function_count"],
        # 30차 W21 + 31차 W29: ASIL 등급 분포 + 등급별 함수 ID.
        "asil_distribution": asil_distribution,
        "asil_b_function_ids": ids_by_asil.get("B", []),
        "asil_c_function_ids": ids_by_asil.get("C", []),
        "asil_d_function_ids": ids_by_asil.get("D", []),
        # 라운드 86 T2002: UNKNOWN 함수 list (audit 진단용).
        "unmapped_function_names": unmapped_fns,
        # 31-fix D15: audit reviewer 공지 메타 — 회사 v3.01 표준 외 색상 확장 명시.
        "asil_highlight_policy": (
            "B=파랑(#E2F0FF) / C=주황(#FFE5CC) / D=빨강(#FFC7CE) — "
            "31차 비표준 audit 확장 (회사 v3.01 양식은 빨강만 사용)"
        ),
    }

    # Cover (Cover는 v2.02도 동일 시트명 "Cover" 사용 — 정확 매칭 유지)
    cover_ws = next((wb[n] for n in sheet_names if n.lower() == "cover"), None)
    if cover_ws is None:
        warnings.append("Cover 시트 미발견 — Doc ID/Author 등 미기록")
    else:
        # 54-fix C1: layout 전달 — v2.02 라벨 자동 매핑 + v3.01 fallback
        _write_cover_sheet(cover_ws, meta, out_warnings=warnings, layout=layout)

    # Test Summary — 54-fix C1: SwIT 53차 patern과 대칭. v2.02 "1.Test Summary" 호환.
    ts_ws = next((wb[n] for n in sheet_names if "test summary" in n.lower()), None)
    if ts_ws is None:
        warnings.append("Test Summary 시트 미발견")
    else:
        # 54-fix C1: layout + summary 전달
        _write_test_summary_sheet(
            ts_ws, meta, agg, out_warnings=warnings,
            layout=layout, summary=summary,
        )

    # 3. Coverage
    cov_ws = next((wb[n] for n in sheet_names
                   if "coverage" in n.lower() and "traceability" not in n.lower()
                   and "consistency" not in n.lower()), None)
    if cov_ws is None:
        warnings.append("Coverage 시트 미발견")
    else:
        # F7 자체평가 R2 N3 fix: layout + out_warnings 전달 — clear warning이
        # X-SwUT-Warnings 헤더로 propagate (이전 누락).
        # 라운드 76 T1105 — c_function_map 재활성. 라운드 74 롤백 사유 (row 폭증 +
        # 양식 cross-ref formula 깨짐)는 라운드 76 인프라로 해소:
        # (a) enhance_function_coverage_with_file로 dedup key 정확성 향상
        # (b) update_cross_refs_after_row_expansion으로 cross-ref formula 동적 갱신
        # (c) auto_expand_row_block + push_sentinel_to_last_row로 양식 row 확장 + sentinel push
        n_written = _write_coverage_sheet(
            cov_ws, agg, layout=layout, out_warnings=warnings,
            c_function_map=session.c_function_map or None,
        )
        summary["coverage_rows_written"] = n_written

    # 1.Traceability — T133 본격 작성 (TC×Function 매트릭스)
    # 라운드 F7 D1 fix: incomplete_sheets에 실제 시트 이름 보고 — 회사 표준은
    # '2.Traceability' (prefix 2.), HDPDM01 release는 '1.Traceability'. 시트 발견
    # 시 ws.title 사용.
    incomplete_sheets: list[str] = []
    trace_ws = next((wb[n] for n in sheet_names if "traceability" in n.lower()), None)
    if trace_ws is None:
        warnings.append("Traceability 시트 미발견")
    else:
        n_o = _write_traceability_sheet(
            trace_ws, session, out_warnings=warnings, layout=layout, agg=agg,
        )
        summary["traceability_o_cells"] = n_o
        if n_o == 0:
            incomplete_sheets.append(trace_ws.title.strip())

    # 2.Consistency — 15차: SwUTS 자체 일관성 4 row + 16차: SwUDS↔SwUTS 매핑 row (옵션).
    cons_ws = next((wb[n] for n in sheet_names if "consistency" in n.lower()), None)
    if cons_ws is not None:
        n_cons = _write_consistency_sheet(
            cons_ws, session,
            swuds_function_ids=swuds_function_ids,
            out_warnings=warnings,
            agg=agg,
        )
        summary["consistency_self_check_rows"] = n_cons
        # 라운드 95 — spec_based(KJPDS02)는 agg SwUDS name→SwUFn 매핑으로 정합성을
        # 시트에 직접 stamp → swuds_function_ids 없어도 비교 완료로 간주.
        _cons_spec_based = bool(
            (agg.get("function_name_to_swufn_from_suds") or {})
            and agg.get("function_rows")
        )
        if swuds_function_ids is not None or _cons_spec_based:
            summary["consistency_swuds_compared"] = True
            # SwUDS 매핑까지 자동 완료 — incomplete 표시 제거.
        else:
            summary["consistency_swuds_compared"] = False
            # 라운드 F7 D1 fix: ws.title 사용 (회사 표준 '3.Consistency')
            incomplete_sheets.append(
                f"{cons_ws.title.strip()} (SwUDS 비교 partial — v3.02)"
            )
    else:
        warnings.append("Consistency 시트 미발견")
        incomplete_sheets.append("Consistency")

    # History — 55-fix: 사용자 결정 B (single-row release entry).
    # 이전 git log 10건 → 산출물 release_sw_version + test_date 1 row만.
    hist_ws = next((wb[n] for n in sheet_names if n.lower() == "history"), None)
    if hist_ws is not None:
        # 55-fix-2 W2: SwUT prefix 명시 — audit reviewer가 산출물 식별 가능 (vs SwIT)
        # 55-fix-2 W6: out_warnings 전달 — release_sw_version/test_date 빈 시 누적
        release_rows = build_release_history_row(
            meta, doc_kind="SwUT Coverage Report", out_warnings=warnings,
        )
        n_h = _write_history_sheet(hist_ws, release_rows, out_warnings=warnings)
        summary["history_rows_written"] = n_h
        if n_h == 0:
            incomplete_sheets.append("History")

    summary["template_sha256_12"] = template_sha256_12
    summary["build_timestamp"] = meta.build_timestamp

    # 라운드 83 T1702: AuditLog 시트 신규 추가 (회사 양식 영향 0 — 신규 시트만).
    # build_timestamp / template_sha256_12 stamp 후 호출 (위 2개 필드 활용).
    try:
        if "AuditLog" not in wb.sheetnames:
            audit_ws = wb.create_sheet("AuditLog")
            n_audit = _write_audit_log_sheet(
                audit_ws, meta, summary, agg, session, warnings,
            )
            summary["audit_log_rows_written"] = n_audit
            summary["audit_log_sheet_added"] = True
    except Exception as _e:  # pragma: no cover — fail-safe
        warnings.append(
            f"AuditLog 시트 작성 실패 (산출물 영향 0): {type(_e).__name__}: {str(_e)[:80]}"
        )

    # 14차 W1: BytesIO 그대로 result에 저장 — getvalue() copy 회피.
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)  # router StreamingResponse가 처음부터 read
    wb.close()

    # 라운드 89: config doc_filenames[coverage] 패턴 우선 ({version}/{date} 치환).
    # 빈 값이면 HDPDM01 v3.01 하드코딩 default (backward compat).
    if meta.doc_filename_pattern:
        filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version, date=short_date(meta.test_date),
        )
    else:
        filename = (
            f"({meta.project_id})SwUT Coverage Report_"
            f"v{meta.release_sw_version}_{short_date(meta.test_date)}_R.xlsx"
        )

    return CoverageBuildResult(
        ok=True,
        xlsx_io=out,
        filename=filename,
        warnings=warnings,
        incomplete_sheets=incomplete_sheets,
        summary=summary,
    )


# `short_date`는 excel_template_utils에서 import — 모듈 하단 중복 정의 제거 (deep-reviewer C1).


# 38차 W2 — public API 명시. SwIT aggregator가 본 모듈의 _write_*_sheet private
# 함수들을 직접 import 중 (강결합 보류 — 35차 SwIT 라운드 채택). 본 __all__로
# 강결합 경계를 명시화: signature 변경 시 SwIT 회귀(test_swit_*_aggregator.py)
# 동시 검증 의무. 향후 W2 정리 시 public alias로 분리 권장.
__all__ = [
    # Public API
    "CoverageBuildMeta",
    "CoverageBuildResult",
    "build_coverage_report",
    # Private 함수 — SwIT가 import 중. 명시로 강결합 가시화.
    "_compute_asil_distribution",
    "_write_consistency_sheet",
    "_write_cover_sheet",
    "_write_coverage_sheet",
    "_write_history_sheet",
    "_write_test_summary_sheet",
    "_write_traceability_sheet",
]
