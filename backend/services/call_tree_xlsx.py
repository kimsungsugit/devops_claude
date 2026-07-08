"""함수 호출 트리(콜트리) → xlsx 내보내기 — 회사 SwITS "2.SW Integration Strategy" 대응.

`/api/jenkins/call-tree`(build_call_tree_precise) 결과 payload를 감사자가 익숙한
**depth 컬럼 들여쓰기** 형식의 xlsx로 렌더한다. 회사 SwITS 통합전략 시트가 함수 호출
계층을 컬럼(N depth)으로 펼치고 depth별 파스텔로 음영하던 레이아웃을 재현하되, 우리
강점(정의 파일 · ASIL 등급 · 함수포인터/간접호출/순환/깊이제한 마커)을 정보 열로 함께 담는다.

레이아웃(시트당):
  Row1  타이틀(병합)           예: "Software Integration Strategy"
  Row2  메타(병합)             엔진·함수·엣지·방향·생성시각·job/build
  Row3  설명(병합, wrap)
  Row5  헤더                   No · 진입 함수 · 1 depth · 2 depth · … · 정의 파일 · ASIL · 유형/비고
  Row6~ 데이터                 각 노드 1행, (C + depth) 컬럼에 함수명(depth 색), 진입 함수 블록 헤더는 노란/남색

입력 payload:
  - 단방향: ``{"trees":[node...], "stats":{...}, "meta":{...}, "missing":[...]}``
  - 양방향: ``{"bidir":True, "callers":{"trees":...}, "callees":{"trees":...}, "stats":{...}}``
  node: ``{"name", "calls":[...], "externals":[{"name","header","library"}], "asil"?, "file"?,
           "signature"?, "via_ref"?, "indirect"?:[...], "cycle"?, "truncated"?}``

순수 함수(부작용 없음, 시각 datetime은 호출자가 meta로 주입). openpyxl 미설치 시 ImportError.
색상은 backend.services.design_tokens 단일 출처(module-level hardcode 금지 정책 준수).
"""

from __future__ import annotations

import io
from typing import Any, Dict, List, Optional, Tuple

from backend.services.design_tokens import (
    ASIL_A_FILL_RGB,
    ASIL_B_FILL_RGB,
    ASIL_C_FILL_RGB,
    ASIL_D_FILL_RGB,
    ASIL_QM_FILL_RGB,
    CALL_TREE_DEPTH_FILLS,
    CALL_TREE_EXTERNAL_FILL_RGB,
    CALL_TREE_HEADER_FILL_RGB,
    CALL_TREE_ROOT_FILL_RGB,
    CALL_TREE_ROOT_FONT_RGB,
    CALL_TREE_SOURCE_FILL_RGB,
    MUTED_TEXT_FONT_RGB,
)

# DoS 가드 — 인증 뒤이고 클라이언트가 echo하는 자기 데이터지만, 단일 요청으로 거대한 시트
# 생성을 막는다(실데이터 상한의 넉넉한 배수: 콜트리 노드는 백엔드가 이미 forest 60K/루트 200
# 상한을 걸어 생성하므로 그 이하). 트리 컬럼 폭은 사용자 depth 설정(최대 20)의 2배로 clamp.
_MAX_TREES = 500        # 루트(진입 함수) 블록 수 상한
_MAX_NODES = 20000      # 시트당 렌더 행(노드) 누적 상한
_MAX_DEPTH = 40         # depth 컬럼 상한 — 초과 depth는 마지막 컬럼으로 clamp

_ASIL_FILL_RGB = {
    "A": ASIL_A_FILL_RGB,
    "B": ASIL_B_FILL_RGB,
    "C": ASIL_C_FILL_RGB,
    "D": ASIL_D_FILL_RGB,
    "QM": ASIL_QM_FILL_RGB,
}


def _cs(v: Any) -> str:
    """Excel 수식 주입 방지 — =,+,-,@ 로 시작하는 값에 ' 프리픽스.

    함수명/파일경로는 소스코드 유래라 위험이 낮지만, echo body를 신뢰하므로 export
    보안 표준(trace_matrix_xlsx._cs와 동일 규칙)을 방어적으로 적용.
    """
    s = "" if v is None else str(v)
    return ("'" + s) if s[:1] in ("=", "+", "-", "@") else s


def _node_note(node: Dict[str, Any], is_root: bool, is_external: bool) -> str:
    """노드 유형/플래그를 사람이 읽는 한 줄로 — 프론트 CallTreeNode 배지와 동일 의미."""
    parts: List[str] = []
    if is_root:
        parts.append("진입점")
    if is_external:
        header = str(node.get("header") or "").strip()
        library = str(node.get("library") or "").strip()
        if header or library:
            parts.append(f"외부 · {header or '?'}/{library or '?'}")
        else:
            parts.append("외부")
    if node.get("via_ref"):
        parts.append("↪ 참조(함수포인터)")
    indirect = node.get("indirect")
    if isinstance(indirect, list) and indirect:
        parts.append(f"⚡ 간접호출 {len(indirect)}")
    if node.get("cycle"):
        parts.append("↻ 순환")
    if node.get("truncated"):
        parts.append("… 깊이제한")
    return " · ".join(parts)


def _scan_tree(trees: List[Dict[str, Any]]) -> Tuple[int, bool]:
    """Pass 1 — 렌더 전 max_depth 계산 + 노드 수 캡 판정. (max_depth, truncated) 반환."""
    max_depth = 0
    count = 0
    truncated = False

    stack: List[Tuple[Dict[str, Any], int]] = [(t, 0) for t in reversed(trees)]
    while stack:
        node, depth = stack.pop()
        count += 1
        if count > _MAX_NODES:
            truncated = True
            break
        if depth > max_depth:
            max_depth = depth
        if depth >= _MAX_DEPTH:
            continue
        for ch in (node.get("calls") or []):
            if isinstance(ch, dict):
                stack.append((ch, depth + 1))
        for ex in (node.get("externals") or []):
            if isinstance(ex, dict):
                count += 1
                if depth + 1 > max_depth:
                    max_depth = depth + 1
    return min(max_depth, _MAX_DEPTH), truncated


def _meta_line(stats: Dict[str, Any], meta: Dict[str, Any], reverse: bool) -> str:
    """상단 메타 요약 한 줄. reverse는 호출자가 명시(양방향 시트는 stats.reverse 폴백에 의존 안 함)."""
    bits = [
        f"방향: {'역호출(caller)' if reverse else '호출(callee)'}",
        f"엔진: {stats.get('engine') or '?'}",
        f"함수: {stats.get('functions') or 0}",
        f"호출 엣지: {stats.get('edges') or 0}",
    ]
    if stats.get("roots"):
        bits.append(f"루트: {stats.get('roots')}")
    if stats.get("files_scanned"):
        bits.append(f"스캔 파일: {stats.get('files_scanned')}")
    if meta.get("generated_at"):
        bits.append(f"생성: {meta.get('generated_at')}")
    if meta.get("job_url"):
        bits.append(f"job: {meta.get('job_url')}")
    if meta.get("build_selector"):
        bits.append(f"build: {meta.get('build_selector')}")
    if meta.get("source_complete") is False:
        bits.append("⚠ 체크아웃 소스 미완(부분 집계)")
    return "   |   ".join(bits)


def build_call_tree_xlsx(payload: Any, meta: Optional[Dict[str, Any]] = None) -> bytes:
    """콜트리 payload dict → xlsx 바이트.

    Args:
        payload: ``/api/jenkins/call-tree`` 응답(단방향 ``{trees,stats,...}`` 또는
            양방향 ``{bidir,callers,callees,stats}``).
        meta: 헤더 블록용(generated_at, job_url, build_selector 등). 없으면 생략.

    Returns:
        xlsx 파일 바이트(openpyxl). openpyxl 미설치 시 ImportError.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    payload = payload if isinstance(payload, dict) else {}
    meta = dict(meta) if isinstance(meta, dict) else {}

    # ── 공통 스타일 (셀마다 재생성 방지, 워크북 전역 재사용) ──
    styles = {
        "title": Font(name="맑은 고딕", size=14, bold=True),
        "meta": Font(name="맑은 고딕", size=9, color=MUTED_TEXT_FONT_RGB),
        "desc": Font(name="맑은 고딕", size=10),
        "header": Font(name="맑은 고딕", size=10, bold=True),
        "root": Font(name="맑은 고딕", size=11, bold=True, color=CALL_TREE_ROOT_FONT_RGB),
        "node": Font(name="Consolas", size=10),
        "ext": Font(name="Consolas", size=9, italic=True, color=MUTED_TEXT_FONT_RGB),
        "note": Font(name="맑은 고딕", size=9, color=MUTED_TEXT_FONT_RGB),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "root_fill": PatternFill("solid", fgColor=CALL_TREE_ROOT_FILL_RGB),
        "source_fill": PatternFill("solid", fgColor=CALL_TREE_SOURCE_FILL_RGB),
        "header_fill": PatternFill("solid", fgColor=CALL_TREE_HEADER_FILL_RGB),
        "ext_fill": PatternFill("solid", fgColor=CALL_TREE_EXTERNAL_FILL_RGB),
        "depth_fills": [PatternFill("solid", fgColor=c) for c in CALL_TREE_DEPTH_FILLS],
        "asil_fills": {k: PatternFill("solid", fgColor=v) for k, v in _ASIL_FILL_RGB.items()},
        "get_col": get_column_letter,
    }

    wb = Workbook()
    root_ws = wb.active
    assert root_ws is not None  # Workbook()은 생성 시 active 시트 1개를 항상 보유

    if payload.get("bidir"):
        callees = payload.get("callees") if isinstance(payload.get("callees"), dict) else {}
        callers = payload.get("callers") if isinstance(payload.get("callers"), dict) else {}
        stats_all = payload.get("stats") if isinstance(payload.get("stats"), dict) else {}
        ws1 = root_ws
        ws1.title = "호출 트리 (callee)"
        _render_sheet(
            ws1, callees.get("trees"),
            "Software Integration Strategy — 호출(callee)",
            "- 각 진입 함수가 호출하는 하위 함수를 호출 계층(depth)에 따라 통합 순서로 정의함.",
            callees.get("stats") if isinstance(callees.get("stats"), dict) else stats_all,
            meta, styles, reverse=False,
        )
        ws2 = wb.create_sheet("역호출 트리 (caller)")
        _render_sheet(
            ws2, callers.get("trees"),
            "역호출 트리 — 누가 이 함수를 호출하나 (caller)",
            "- 각 함수를 호출하는 상위 함수(caller)를 역방향으로 추적함(영향 분석).",
            callers.get("stats") if isinstance(callers.get("stats"), dict) else stats_all,
            meta, styles, reverse=True,
        )
    else:
        trees = payload.get("trees")
        stats = payload.get("stats") if isinstance(payload.get("stats"), dict) else {}
        reverse = bool(stats.get("reverse"))
        ws = root_ws
        if reverse:
            ws.title = "역호출 트리 (caller)"
            _render_sheet(
                ws, trees, "역호출 트리 — 누가 이 함수를 호출하나 (caller)",
                "- 각 함수를 호출하는 상위 함수(caller)를 역방향으로 추적함(영향 분석).",
                stats, meta, styles, reverse=True,
            )
        else:
            ws.title = "SW Integration Strategy"
            _render_sheet(
                ws, trees, "Software Integration Strategy",
                "- 각 진입 함수가 호출하는 하위 함수를 호출 계층(depth)에 따라 통합 순서로 정의함.",
                stats, meta, styles, reverse=False,
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_sheet(
    ws: Any,
    trees_in: Any,
    title: str,
    desc: str,
    stats: Any,
    meta: Dict[str, Any],
    S: Dict[str, Any],
    reverse: Optional[bool] = None,
) -> None:
    """트리 리스트를 한 시트에 depth-컬럼 형식으로 렌더."""
    trees_all = [t for t in (trees_in if isinstance(trees_in, list) else []) if isinstance(t, dict)]
    tree_truncated = len(trees_all) > _MAX_TREES     # W-1: 루트(진입 함수) 상한 초과 표면화용
    trees = trees_all[:_MAX_TREES]
    stats = stats if isinstance(stats, dict) else {}
    # W-2: 방향은 호출자 명시 우선(양방향 시트는 sub-stats 폴백의 reverse 오표기 방지), 없으면 stats.reverse.
    is_reverse = reverse if reverse is not None else bool(stats.get("reverse"))

    max_depth, scan_truncated = _scan_tree(trees)
    depth_cols = max_depth + 1                 # depth 0..max_depth
    tree_start = 3                             # C열
    tree_end = tree_start + depth_cols - 1
    col_file = tree_end + 1
    col_asil = tree_end + 2
    col_note = tree_end + 3
    last_col = col_note
    get_col = S["get_col"]

    # ── Row 1: 타이틀 (병합) ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, title)
    c.font = S["title"]
    c.alignment = S["center"]

    # ── Row 2: 메타 (병합) ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws.cell(2, 1, _meta_line(stats, meta, is_reverse))
    c.font = S["meta"]
    c.alignment = S["left"]

    # ── Row 3: 설명 (병합, wrap) ──
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    c = ws.cell(3, 1, desc)
    c.font = S["desc"]
    c.alignment = S["wrap"]

    # ── Row 5: 헤더 ──
    hr = 5
    header_cells = {1: "No", 2: "진입 함수"}
    for d in range(depth_cols):
        header_cells[tree_start + d] = f"{d + 1} depth"
    header_cells[col_file] = "정의 파일"
    header_cells[col_asil] = "ASIL"
    header_cells[col_note] = "유형/비고"
    for col, label in header_cells.items():
        hc = ws.cell(hr, col, label)
        hc.font = S["header"]
        hc.fill = S["header_fill"]
        hc.alignment = S["center"] if col != col_note else S["left"]

    # ── 데이터 ──
    state = {"row": hr, "written": 0, "truncated": scan_truncated, "depth_capped": False}

    def _write(node: Dict[str, Any], depth: int, block_no: Optional[int],
               is_root: bool, is_external: bool) -> bool:
        """노드 1행 기록. 캡 도달 시 False(상위 순회 중단 신호)."""
        if state["written"] >= _MAX_NODES:
            state["truncated"] = True
            return False
        state["row"] += 1
        state["written"] += 1
        r = state["row"]
        d = depth if depth <= max_depth else max_depth
        name = node.get("name")

        # A열: 블록 번호(루트만)
        if is_root and block_no is not None:
            ac = ws.cell(r, 1, block_no)
            ac.font = S["note"]
            ac.alignment = S["center"]

        # B열: 진입 함수명(루트만, 노란/남색 강조)
        if is_root:
            bc = ws.cell(r, 2, _cs(name))
            bc.font = S["root"]
            bc.fill = S["root_fill"]
            bc.alignment = S["left"]

        # depth 트리 컬럼: 함수명 + depth 색(외부는 회색)
        tc = ws.cell(r, tree_start + d, _cs(name))
        tc.font = S["ext"] if is_external else S["node"]
        tc.fill = S["ext_fill"] if is_external else S["depth_fills"][min(d, len(S["depth_fills"]) - 1)]
        tc.alignment = S["left"]

        # 정의 파일(내부 노드만)
        if not is_external:
            f = str(node.get("file") or "").strip()
            if f:
                fc = ws.cell(r, col_file, _cs(f))
                fc.font = S["note"]
                fc.fill = S["source_fill"]
                fc.alignment = S["left"]

        # ASIL
        asil = str(node.get("asil") or "").strip().upper()
        if asil in S["asil_fills"]:
            acx = ws.cell(r, col_asil, asil)
            acx.fill = S["asil_fills"][asil]
            acx.font = S["header"]
            acx.alignment = S["center"]

        # 유형/비고
        note = _node_note(node, is_root, is_external)
        if note:
            nc = ws.cell(r, col_note, note)
            nc.font = S["note"]
            nc.alignment = S["left"]
        return True

    def _emit(node: Dict[str, Any], depth: int, block_no: Optional[int], is_root: bool) -> bool:
        if not _write(node, depth, block_no, is_root, False):
            return False
        # W1(재귀 깊이 가드): echo body(클라 조작 가능)의 비정상 깊은 체인이 CPython 재귀
        # 한계(≈980, sync-threadpool 얇은 스택은 더 낮음)를 넘겨 RecursionError→HTTP 500
        # 나는 것을 방지. pass1 _scan_tree(depth>=_MAX_DEPTH → 순회 중단)와 대칭 — depth를
        # 넘는 자식은 컬럼상 어차피 마지막에 clamp되므로 정보 손실도 최소. 실데이터는
        # 프론트가 max_depth≤20으로 상한하므로 이 경로에 절대 도달하지 않는다.
        if depth >= _MAX_DEPTH:
            # W-1: 깊이 상한으로 자식을 생략할 때 정직하게 표면화(silent 누락 방지).
            if node.get("calls") or node.get("externals"):
                state["depth_capped"] = True
            return True
        for ch in (node.get("calls") or []):
            if isinstance(ch, dict):
                if not _emit(ch, depth + 1, None, False):
                    return False
        for ex in (node.get("externals") or []):
            if isinstance(ex, dict):
                if not _write(ex, depth + 1, None, False, True):
                    return False
        return True

    stopped = False
    for idx, tree in enumerate(trees, start=1):
        if not _emit(tree, 0, idx, True):
            stopped = True
            break

    if not trees_all:
        state["row"] += 1
        ws.cell(state["row"], 1, "표시할 호출 트리가 없습니다. (진입 함수를 찾지 못했거나 트리가 비어 있음)").font = S["note"]
    else:
        # W-1: 세 캡(노드/루트/깊이) 어느 것이든 발동하면 한 행에 종합해 정직하게 표면화.
        # (실데이터는 프론트 depth≤20·root≤200 상한이라 미도달이나, echo body·미래 캡 변경 대비 audit 정직성.)
        cap_notes = []
        if stopped or state["truncated"]:
            cap_notes.append(f"노드 {_MAX_NODES:,}개 상한 도달")
        if tree_truncated:
            cap_notes.append(f"진입 함수 {_MAX_TREES}개 상한 초과({len(trees_all) - _MAX_TREES}개 생략)")
        if state["depth_capped"]:
            cap_notes.append(f"표시 깊이 {_MAX_DEPTH} 상한 도달(더 깊은 호출 생략)")
        if cap_notes:
            state["row"] += 1
            ws.cell(
                state["row"], 1,
                "⚠ " + " · ".join(cap_notes) + " — 진입 함수 지정 또는 깊이 축소를 권장합니다.",
            ).font = S["note"]

    # ── 열 너비 / 고정 창 ──
    ws.column_dimensions[get_col(1)].width = 5
    ws.column_dimensions[get_col(2)].width = 34
    for d in range(depth_cols):
        ws.column_dimensions[get_col(tree_start + d)].width = 28
    ws.column_dimensions[get_col(col_file)].width = 30
    ws.column_dimensions[get_col(col_asil)].width = 7
    ws.column_dimensions[get_col(col_note)].width = 26
    # 헤더행(5)까지 + 진입 함수(B)까지 고정 → 큰 트리 스크롤 시 맥락 유지.
    ws.freeze_panes = ws.cell(hr + 1, tree_start)
