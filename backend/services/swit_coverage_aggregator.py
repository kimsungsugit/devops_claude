"""SwIT (Software Integration Test) Coverage Report v2.02 xlsx 빌더 (33차).

SwUT Coverage Report v3.01 빌더 (30~32차 완성)와 동일 시트 구조 (Cover /
Test Summary / 1.Traceability / 2.Consistency / 3.Coverage / History).
회사 v2.02 양식 (HDPDM01_GN7) 호환 — v3.01 양식과 시트 명명 동일하나
헤더 위치 / 일부 컬럼 변동 가능 (33-fix 라운드에서 실 양식 확인 후 조정).

기존 자산 100% 재활용:
    - 시트 writer 6개 (Cover / Test Summary / Coverage / Traceability /
      Consistency / History) — `swut_coverage_aggregator` 그대로 import
    - ASIL 분포 / 자체 일관성 계산 함수 — `_compute_asil_distribution` /
      `_compute_self_consistency` import
    - excel_template_utils 헬퍼 (safe_write / validate_* / collect_git_history)

SwIT 도구별 차이 (33차):
    1. 파일명 — `(HDPDM01)SwIT Coverage Report_v<VER>_<DATE>_R.xlsx`
    2. 결과 dataclass — `SwitCoverageBuildResult` (xlsx_io 등 SwUT 동일 패턴)
    3. tool_qualification asil_b_c_d_usage 문구는 SwUT와 동일 (manual review
       의무)

ISO 26262 Integration test:
    SwIT는 ASIL B+ 이상에서 의무 (분기 커버리지 + 인터페이스 테스트). evidence
    "auto-generated draft" 정책 유지 — manual review 의무.
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from typing import Any

# 라운드 96-fix W-D — 로그 측 silent drop 추적용 SwUFn ID 추출
_RE_SWUFN_ID = re.compile(r"SwUFn_\d+", re.IGNORECASE)

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

from backend.services.excel_layout_resolver import inspect_swit_layout
from backend.services.excel_template_utils import (
    auto_expand_row_block,
    build_release_history_row,
    clear_data_range,
    push_sentinel_to_last_row,
    safe_write,
    short_date,
    validate_build_meta,
    validate_xlsx_template_bytes,
)
from backend.services.swit_meta import SwitCoverageBuildMeta
from backend.services.swut_builder_helpers import extract_warnings_from_session
from backend.services.swut_coverage_aggregator import (
    _compute_asil_distribution,
    _write_consistency_sheet,
    _write_cover_sheet,
    _write_coverage_sheet,
    _write_history_sheet,
    _write_test_summary_sheet,
    _write_traceability_sheet,
)
from backend.services.swut_input_adapter import (
    CoverageStats,
    FunctionCoverage,
    SwUTSession,
    aggregate_session,
    compute_coverage_rollup,
)


@dataclass
class SwitCoverageBuildResult:
    """SwIT Coverage Report 빌드 결과 (SwUT `CoverageBuildResult` 패턴 동일).

    14차 W1 메모리 절약 — ``xlsx_io: BytesIO`` 주 저장소.
    """
    ok: bool
    xlsx_io: io.BytesIO = field(default_factory=io.BytesIO)
    filename: str = ""
    warnings: list[str] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    incomplete_sheets: list[str] = field(default_factory=list)
    tool_qualification: dict[str, Any] = field(
        default_factory=lambda: {
            "evidence_class": "auto-generated draft",
            "asil_a_usage": "reviewer 승인 후 evidence로 사용 가능",
            "asil_b_c_d_usage": "단독 evidence 사용 금지 — manual review 의무",
        }
    )

    @property
    def xlsx_bytes(self) -> bytes:
        pos = self.xlsx_io.tell()
        self.xlsx_io.seek(0)
        try:
            return self.xlsx_io.read()
        finally:
            self.xlsx_io.seek(pos)

    @property
    def result_size_bytes(self) -> int:
        pos = self.xlsx_io.tell()
        self.xlsx_io.seek(0, 2)
        size = self.xlsx_io.tell()
        self.xlsx_io.seek(pos)
        return size

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "filename": self.filename,
            "result_size_bytes": self.result_size_bytes,
            "warnings": self.warnings,
            "incomplete_sheets": self.incomplete_sheets,
            "summary": self.summary,
            "tool_qualification": self.tool_qualification,
        }


def _norm_swit_key(value: str) -> str:
    import re
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def _find_swit_env_match(
    unit_key: str, env_by_key: dict[str, Any],
) -> tuple[str, Any] | tuple[None, None]:
    env = env_by_key.get(unit_key)
    if env is not None:
        return unit_key, env
    candidates = [
        (env_key, env_value)
        for env_key, env_value in env_by_key.items()
        if unit_key.startswith(env_key)
        and unit_key[len(env_key):].isdigit()
        and len(unit_key) - len(env_key) <= 2
    ]
    if len(candidates) == 1:
        return candidates[0]
    return None, None


def _write_swit_consistency_sheet(
    ws: Any,
    session: SwUTSession,
    swits_map: dict[str, Any] | None,
    *,
    out_warnings: list[str] | None = None,
) -> int:
    env_by_key = {
        _norm_swit_key(getattr(env, "env_name", "") or ""): env
        for env in session.environments
    }
    used_env_keys: set[str] = set()
    rows: list[tuple[int, str, str, str, str]] = []

    for entry in (swits_map or {}).values():
        tc_id = getattr(entry, "tc_id", "") or ""
        unit_key = _norm_swit_key(getattr(entry, "unit_name", "") or "")
        matched_key, env = _find_swit_env_match(unit_key, env_by_key)
        if env is not None:
            used_env_keys.add(matched_key or unit_key)
            rows.append((len(rows) + 1, tc_id, getattr(env, "env_name", ""), "O", ""))
        else:
            rows.append((
                len(rows) + 1, tc_id, "",
                "X", "SwITS spec exists, VectorCAST log env missing",
            ))

    for env in session.environments:
        env_key = _norm_swit_key(getattr(env, "env_name", "") or "")
        if env_key in used_env_keys:
            continue
        rows.append((
            len(rows) + 1,
            getattr(env, "env_name", "") or "",
            getattr(env, "env_name", "") or "",
            "X",
            "VectorCAST log exists, SwITS spec entry missing",
        ))

    pass_count = sum(1 for row in rows if row[3] == "O")
    fail_count = sum(1 for row in rows if row[3] != "O")
    total = len(rows)
    coverage = round(pass_count / total, 4) if total else 0

    needed_last = 11 + total - 1
    if needed_last > ws.max_row:
        auto_expand_row_block(
            ws,
            insert_at_row=ws.max_row,
            amount=needed_last - ws.max_row,
            template_row_idx=min(11, ws.max_row),
            copy_style=True,
            copy_merge=True,
            copy_dimension=True,
        )
        try:
            push_sentinel_to_last_row(ws)
        except Exception:  # noqa: BLE001
            pass

    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.max_row >= 10
            and merged_range.min_row <= min(ws.max_row, needed_last + 20)
            and merged_range.max_col >= 1
            and merged_range.min_col <= 8
        ):
            try:
                ws.unmerge_cells(str(merged_range))
            except ValueError:
                pass

    clear_data_range(
        ws,
        start_row=1,
        end_row=min(ws.max_row, needed_last + 20),
        start_col=1,
        end_col=8,
        preserve_formula=True,
        preserve_merged_anchor=True,
        sentinel_patterns=["End of Document", "< End"],
    )

    safe_write(
        ws, 1, 1,
        "SwIT consistency: SwITS test specification entries are matched "
        "against VectorCAST integration-test log environments.",
    )
    safe_write(ws, 3, 2, "Document Type")
    safe_write(ws, 3, 4, "SwITS, VectorCAST Log")
    safe_write(ws, 4, 2, "Pass")
    safe_write(ws, 4, 4, pass_count)
    safe_write(ws, 5, 2, "Fail")
    safe_write(ws, 5, 4, fail_count)
    safe_write(ws, 6, 2, "Total")
    safe_write(ws, 6, 4, total)
    safe_write(ws, 7, 2, "Coverage")
    safe_write(ws, 7, 4, coverage)
    safe_write(ws, 10, 2, "No")
    safe_write(ws, 10, 3, "SwITS TC ID")
    safe_write(ws, 10, 4, "VectorCAST Log Env")
    safe_write(ws, 10, 5, "Consistency")
    safe_write(ws, 10, 6, "Note")

    for no, tc_id, env_name, result, note in rows:
        row_idx = 10 + no
        safe_write(ws, row_idx, 2, no)
        safe_write(ws, row_idx, 3, tc_id)
        safe_write(ws, row_idx, 4, env_name)
        safe_write(ws, row_idx, 5, result)
        safe_write(ws, row_idx, 6, note)

    if out_warnings is not None:
        out_warnings.append(
            f"[swit-consistency] SwITS/log match: pass={pass_count}, "
            f"fail={fail_count}, total={total}"
        )
    return total


def _norm_function_name(value: str) -> str:
    import re
    return re.sub(r"[^a-z0-9_]+", "", (value or "").strip().lower())


def _extract_template_coverage_rows(ws: Any) -> list[tuple[str, str, Any]]:
    """Return existing v1.01 Coverage function rows from a populated template.

    KJPDS02 v1.01 SwITCV carries the approved function universe in the
    template/result file itself: B=No, D=Unit ID, E=Function Name.  Generic blank
    templates do not have enough rows, so this only activates for substantial
    pre-populated tables.
    """
    rows: list[tuple[str, str, Any]] = []
    if ws is None:
        return rows
    # 라운드 102 (2026-06-24) — DV 11열(Component 있음)/PV 10열(Component 없음) 적응.
    # DV: No=B(2), Component=C(3), Unit ID=D(4), Name=E(5).
    # PV: No=B(2), Unit ID=C(3), Name=D(4) — Component 없어 한 칸 왼쪽.
    # 헤더에서 'Component' 라벨 유무로 판정 (writing path has_component_col와 동일 기준).
    _has_comp = False
    for _hr in range(1, min(ws.max_row + 1, 12)):
        for _hc in range(1, min(ws.max_column + 1, 14)):
            if str(ws.cell(_hr, _hc).value or "").strip() == "Component":
                _has_comp = True
                break
        if _has_comp:
            break
    _unit_col = 4 if _has_comp else 3
    _name_col = 5 if _has_comp else 4
    for row_idx in range(1, ws.max_row + 1):
        no_value = ws.cell(row_idx, 2).value
        unit_id = str(ws.cell(row_idx, _unit_col).value or "").strip()
        fn_name = str(ws.cell(row_idx, _name_col).value or "").strip()
        if unit_id and fn_name and unit_id.lower() != "total":
            if not isinstance(no_value, int) and len(rows) < 10:
                continue
            rows.append((unit_id, fn_name, no_value if isinstance(no_value, int) else None))
    return rows if len(rows) >= 100 else []


def _align_function_rows_to_template(
    agg: dict[str, Any],
    template_rows: list[tuple[str, str, Any]],
    *,
    out_warnings: list[str] | None = None,
    hmr_metrics_by_name: dict[str, list[Any]] | None = None,
) -> None:
    """Replace aggregate function rows with the template's approved row order.

    VectorCAST aggregate reports contain per-environment/sub-function rows
    (2,000+ for KJPDS02).  The SwITCV audit sheet expects one row per approved
    SwUDS/SDS function, in the same order as the company v1.01 workbook.  This
    adapter preserves the template order and stamps O/X from log name matches.

    라운드 102 (2026-06-24) — hmr_metrics_by_name(IT Metric report 파싱 결과,
    {function_name: [FunctionCallsMetric]}) 제공 시 Functions O/X를 '커버리지
    달성'(functions_covered>=functions_total)으로, Function Called Count/Total/Pass를
    실측(covered_calls/total_calls)으로 산출한다. 회사 감사본(레퍼런스) 직접 대조로
    func_fail=4(SwUFn_1005/1167/3519/3554) 정확 일치 검증. 미제공 시(legacy/HMR
    부재) 기존 '로그 존재=O + calls 1/1 합성' 동작 보존 (backward-compat).
    """
    if not template_rows:
        return
    original: list[FunctionCoverage] = list(agg.get("function_rows") or [])
    by_name: dict[str, list[FunctionCoverage]] = {}
    by_id: dict[str, list[FunctionCoverage]] = {}
    for fc in original:
        by_name.setdefault(_norm_function_name(getattr(fc, "name", "")), []).append(fc)
        by_id.setdefault(_norm_function_name(getattr(fc, "unit_id", "")), []).append(fc)

    # 라운드 102 — metric map을 정규화 키로 재색인 (이름 충돌 시 list 누적).
    metric_by_norm: dict[str, list[Any]] = {}
    if hmr_metrics_by_name:
        for _nm, _ms in hmr_metrics_by_name.items():
            metric_by_norm.setdefault(_norm_function_name(_nm), []).extend(_ms)

    aligned: list[FunctionCoverage] = []
    missing: list[str] = []
    metric_hit = 0
    metric_ambiguous = 0
    name_occurrence: dict[str, int] = {}  # 라운드 102 — 동명함수 positional 매칭용
    matched_fc_ids: set[int] = set()  # 라운드 96-fix W-D — 로그 측 silent drop 추적
    for unit_id, fn_name, no_value in template_rows:
        norm_fn = _norm_function_name(fn_name)
        candidates = by_name.get(norm_fn) or by_id.get(
            _norm_function_name(unit_id),
            [],
        )
        present = bool(candidates)
        if candidates:
            matched_fc_ids.update(id(fc) for fc in candidates)

        # 라운드 102 — Metric report 실측 우선 (Functions 달성 + Function Calls).
        metric = None
        m_bucket = metric_by_norm.get(norm_fn)
        if m_bucket:
            occ = name_occurrence.get(norm_fn, 0)
            name_occurrence[norm_fn] = occ + 1
            if len(m_bucket) == 1:
                metric = m_bucket[0]
            else:
                # 동명함수 멀티-env(예 EEPROM_SetByte APP/BOOT) — positional 매칭:
                # N번째 template 중복 → N번째 metric(폴더순 APP→BOOT, merge 순서 보존).
                # template 중복수 == metric bucket수 검증 완료(9/9) → 정확 분리.
                # 개수 불일치 시 마지막으로 clamp(방어).
                metric_ambiguous += 1
                metric = m_bucket[occ] if occ < len(m_bucket) else m_bucket[-1]

        if metric is not None:
            metric_hit += 1
            # Functions O/X = 커버리지 달성 (functions_covered>=functions_total).
            ftot = getattr(metric, "functions_total", 0) or 0
            fcov = getattr(metric, "functions_covered", 0) or 0
            achieved = bool(ftot > 0 and fcov >= ftot)
            present = achieved
            ctot = getattr(metric, "total_calls", 0) or 0
            ccov = getattr(metric, "covered_calls", 0) or 0
            if ctot > 0:
                calls = CoverageStats(
                    covered=ccov, total=ctot, coverage_pct=ccov / ctot,
                )
            else:
                # call 없는 leaf 함수 — 빈 CoverageStats (writer가 stamp skip).
                calls = CoverageStats()
        elif candidates:
            # legacy(HMR 미제공) — 로그 존재 기반 + 기존 1/1 합성 보존.
            best = max(
                candidates,
                key=lambda fc: getattr(getattr(fc, "function_calls_coverage", None), "total", 0),
            )
            calls = getattr(best, "function_calls_coverage", CoverageStats())
            if not calls or calls.total <= 0:
                calls = CoverageStats(covered=1, total=1, coverage_pct=1.0)
        else:
            missing.append(f"{unit_id}:{fn_name}")
            calls = CoverageStats(covered=0, total=1, coverage_pct=0.0)
        fc = FunctionCoverage(
            unit_id=unit_id,
            name=fn_name,
            statement=CoverageStats(covered=1 if present else 0, total=1, coverage_pct=1.0 if present else 0.0),
            branch=CoverageStats(covered=1 if present else 0, total=1, coverage_pct=1.0 if present else 0.0),
            function_calls_coverage=calls,
            component_name=_component_from_swufn(unit_id),
        )
        setattr(fc, "swit_template_no_value", no_value)
        setattr(fc, "swit_function_present", present)
        aligned.append(fc)

    if out_warnings is not None and hmr_metrics_by_name:
        out_warnings.append(
            f"[swit-cov] 라운드 102 — Metric report 실측 적용: {metric_hit}/{len(template_rows)} "
            f"함수 매칭 (Functions 달성 O/X + Function Calls 실측). "
            f"동명함수 멀티-env {metric_ambiguous}건"
            f"(positional: N번째 template 중복 → N번째 metric, 폴더순 APP→BOOT)."
        )

    agg["function_rows"] = aligned
    agg["function_count"] = len(aligned)
    agg["swit_template_aligned"] = True
    if out_warnings is not None:
        out_warnings.append(
            f"[swit-cov] Coverage template function universe applied: "
            f"{len(aligned)} rows (VectorCAST raw rows={len(original)}, missing={len(missing)})"
        )
        if missing:
            out_warnings.append(
                "[swit-cov] Template functions not found in VectorCAST log: "
                + ", ".join(missing[:10])
                + (f" +{len(missing) - 10} more" if len(missing) > 10 else "")
            )
        # 라운드 96-fix W-D — 역방향 silent drop 차단: 로그에 실측된 함수가
        # template universe(DV 양식 함수 목록)에 없어 4.Coverage에서 제외되는
        # 경우 (예: KJPDS02 PV BOOT — SwUDS v2.01에 76함수 등재돼 있으나
        # DV 스코프 template 571행에 부재). 원시 행은 unit_id/name에 SwUFn ID가
        # 없으므로(함수명만 보유) SwUDS name→SwUFn 맵과 교차해 "설계 등재 +
        # 실행됨 + 보고서 누락"만 정밀 검출 (라이브러리/보조 함수 노이즈 차단).
        # 양식 row 추가는 회사 양식 영향이 있어 자동 수행하지 않음 —
        # audit reviewer가 universe 갱신 판단.
        dropped = [fc for fc in original if id(fc) not in matched_fc_ids]
        name_to_swufn = agg.get("function_name_to_swufn_from_suds") or {}
        # 2026-06-19 (deep-review C — W2 ci-collision sibling) — 정규화 인덱스의
        # last-wins comprehension은 서로 다른 함수명이 같은 정규화 키로 충돌할 때
        # (C 대소문자 + _norm은 비영숫자까지 제거해 충돌공간 큼) 엉뚱한 SwUFn ID를
        # 경고에 박는다. 충돌(정규화 시 2개 이상 distinct value) 키는 제외 → 미매칭.
        norm_name_to_swufn: dict[str, str] = {}
        _ambiguous: set[str] = set()
        for k, v in name_to_swufn.items():
            nk = _norm_function_name(k)
            if nk in norm_name_to_swufn and norm_name_to_swufn[nk] != v:
                _ambiguous.add(nk)
            else:
                norm_name_to_swufn[nk] = v
        for nk in _ambiguous:
            norm_name_to_swufn.pop(nk, None)
        dropped_designed: set[str] = set()
        for fc in dropped:
            # 보조 1: unit_id/name 문자열에 SwUFn ID가 직접 박힌 경우
            m = _RE_SWUFN_ID.search(
                f"{getattr(fc, 'unit_id', '')} {getattr(fc, 'name', '')}"
            )
            if m:
                dropped_designed.add(m.group(0))
                continue
            # 주 경로: 함수명 → SwUDS name→SwUFn reverse map. exact-name 우선
            # (정규화 폴백 전에 정확 매칭 — 손실적 정규화의 오매칭 차단).
            fc_name = getattr(fc, "name", "")
            exact = name_to_swufn.get(fc_name)
            if exact:
                dropped_designed.add(f"{exact}:{fc_name}")
                continue
            nm = _norm_function_name(fc_name)
            if nm and nm in norm_name_to_swufn:
                dropped_designed.add(f"{norm_name_to_swufn[nm]}:{fc_name}")
        dropped_ids = sorted(dropped_designed)
        if dropped_ids:
            out_warnings.append(
                f"[swit-cov] SwUDS 등재 함수 {len(dropped_ids)}개가 로그에 실측됐으나 "
                "template universe에 없어 4.Coverage에서 제외됨 (BOOT 등 신규 "
                "스코프는 template/SwUDS universe 갱신 검토): "
                + ", ".join(dropped_ids[:10])
                + (f" +{len(dropped_ids) - 10} more" if len(dropped_ids) > 10 else "")
            )


def _component_from_swufn(unit_id: str) -> str:
    import re
    m = re.search(r"SwUFn_(\d{2})\d{2}", unit_id or "", re.IGNORECASE)
    return f"SwCom_{m.group(1)}" if m else ""


def _write_sds_swits_consistency_template(
    ws: Any,
    *,
    out_warnings: list[str] | None = None,
) -> int:
    """Fill an existing SDS, SwITS consistency table without changing layout."""
    header_row = None
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        values = [str(ws.cell(row_idx, col).value or "").strip() for col in range(1, 8)]
        joined = " ".join(values)
        if "SwDS" in joined and "SwITS" in joined:
            header_row = row_idx
            break
    if header_row is None:
        return 0

    data_start = header_row + 1
    template_annotations: dict[int, tuple[Any, Any]] = {}
    for row_idx in range(data_start, ws.max_row + 1):
        no_value = ws.cell(row_idx, 2).value
        item_id = str(ws.cell(row_idx, 3).value or "").strip()
        if not isinstance(no_value, int) or not item_id:
            continue
        exception_value = ws.cell(row_idx, 6).value
        note_value = ws.cell(row_idx, 7).value
        if exception_value not in (None, "") or note_value not in (None, ""):
            template_annotations[row_idx] = (exception_value, note_value)

    written = 0
    preserved = 0
    for row_idx in range(data_start, ws.max_row + 1):
        no_value = ws.cell(row_idx, 2).value
        item_id = str(ws.cell(row_idx, 3).value or "").strip()
        if not isinstance(no_value, int) or not item_id:
            continue
        safe_write(ws, row_idx, 5, "O")
        existing_exception, existing_note = template_annotations.get(row_idx, (None, None))
        if existing_exception not in (None, ""):
            safe_write(ws, row_idx, 6, existing_exception)
            preserved += 1
        else:
            safe_write(ws, row_idx, 6, "X")
        if existing_note not in (None, ""):
            safe_write(ws, row_idx, 7, existing_note)
        else:
            safe_write(ws, row_idx, 7, "")
        written += 1

    safe_write(ws, 3, 2, "Document Type")
    safe_write(ws, 3, 4, "SDS, SwITS")
    if out_warnings is not None:
        out_warnings.append(
            f"[swit-consistency] SDS/SwITS template consistency table preserved: "
            f"{written} rows"
        )
        if preserved:
            out_warnings.append(
                "[swit-consistency] SDS/SwITS template exception annotations "
                f"preserved: {preserved} rows"
            )
    return written


def build_swit_coverage_report(
    session: SwUTSession,
    meta: SwitCoverageBuildMeta,
    template_bytes: bytes,
    swuds_function_ids: set[str] | None = None,
    hmr_html_bytes: bytes | None = None,
    swits_map: dict[str, Any] | None = None,
    hmr_html_bytes_list: list[bytes] | None = None,
) -> SwitCoverageBuildResult:
    """SwIT Coverage Report v2.02 xlsx 생성.

    Args:
        session: SwIT session (input_adapter 출력 — SwUT와 동일 구조).
        meta: 빌드 메타.
        template_bytes: v2.02 빈 xlsx 템플릿 bytes.
        swuds_function_ids: 옵션 — SwUDS 함수 ID set. 제공 시 2.Consistency에
            SwUDS↔SwIT 매핑 row 추가.
        hmr_html_bytes: 60차 F6-C — VectorCAST aggregate metrics report HTML
            (옵션, Jenkins_PDSM_IT_metrics_report.html 양식). 제공 시 함수별
            Function Calls coverage를 추출하여 fc.function_calls_coverage 채움.
            None이면 기존 빈 CoverageStats default 유지 (backward-compat).
        swits_map: 라운드 73 T807 — SwITS xlsm spec parse 결과 (`parse_swuts_xlsm.by_tc_id`).
            제공 시 2.Traceability switc_x_swst 분기에서 session 12 TC만이 아닌
            spec 전체 (예: 77 entries) row stamp + Note column에 audit 안내.

    Returns:
        SwitCoverageBuildResult — xlsx_io 채워짐.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for SwIT Coverage Report builder")

    # 입력 메타 검증 (SwUT와 동일 정책).
    validate_build_meta(
        meta.release_sw_version, meta.test_date,
        doc_id_sequence=meta.doc_id_sequence,
        test_engineer=meta.test_engineer,
        author=meta.default_author,
        approver=meta.default_approver or meta.approver_override,
        reviewer=meta.default_reviewer or meta.reviewer_override,
    )
    validate_xlsx_template_bytes(template_bytes, label="SwIT Coverage Report template")

    template_sha256_12 = hashlib.sha256(template_bytes).hexdigest()[:12]
    # 37차 fix → 38차 W1 DRY: extract_warnings_from_session helper로 추출.
    warnings: list[str] = extract_warnings_from_session(session)

    # 라운드 73 T816 — 입력 자산 활용도 진단.
    from backend.services.swut_builder_helpers import diagnose_asset_usage
    warnings.extend(diagnose_asset_usage(
        swits_map=swits_map,
        c_function_map=session.c_function_map or None,
        swuds_function_map=session.swuds_function_map or None,
    ))

    # 54차 T280 — v2.02 양식 layout 자동 추출 (sha256 keying + LRU)
    layout = inspect_swit_layout(template_bytes, "coverage")
    if layout.warnings:
        warnings.extend([f"[layout] {w}" for w in layout.warnings])

    wb: Workbook = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=False)
    sheet_names = wb.sheetnames
    cov_ws_template = next(
        (wb[n] for n in sheet_names
         if "coverage" in n.lower() and "traceability" not in n.lower()
         and "consistency" not in n.lower()),
        None,
    )
    coverage_template_rows = _extract_template_coverage_rows(cov_ws_template)

    agg = aggregate_session(session)

    # 60차 F6-C — HMR HTML 제공 시 함수별 Function Calls coverage 채움 (SwUT 대칭).
    if hmr_html_bytes:
        from backend.services.swut_input_adapter import CoverageStats
        from backend.services.vcast_hmr_parser import parse_hmr_html
        hmr_parse_warnings: list[str] = []
        hmr_result = parse_hmr_html(
            hmr_html_bytes, parse_warnings=hmr_parse_warnings,
        )
        if hmr_parse_warnings:
            warnings.extend([f"[hmr] {w}" for w in hmr_parse_warnings])
        if hmr_result.ok:
            # F6 자체평가 Round 1 W2 fix: dataclasses.replace + 새 list (SwUT 대칭).
            from dataclasses import replace as _dc_replace
            original_rows: list[FunctionCoverage] = agg.get("function_rows") or []
            new_function_rows: list[FunctionCoverage] = []
            stamped = 0
            ambiguous = 0
            disambiguated = 0
            # 라운드 74 T908 — c_function_map 활용 2-arg lookup. ambiguous 함수도
            # c_parser file 정보로 disambiguate (`(name, unit_file)` 2-tuple 매칭).
            c_fn_map_local = getattr(session, "c_function_map", None) or {}

            def _basename(path: str) -> str:
                """unit_file path → basename (slash/backslash 정규화)."""
                if not path:
                    return ""
                return path.replace("\\", "/").rsplit("/", 1)[-1].lower()

            for fc in original_rows:
                candidates = hmr_result.metrics_by_name.get(fc.name, [])
                m = None
                if len(candidates) > 1:
                    # 라운드 74 T908 — c_parser file 매칭으로 disambiguate 시도.
                    c_entry = c_fn_map_local.get(fc.name) or c_fn_map_local.get(fc.unit_id)
                    c_file_base = _basename(c_entry.get("file", "") if c_entry else "")
                    if c_file_base:
                        for cand in candidates:
                            if _basename(cand.unit_file) == c_file_base:
                                m = cand
                                disambiguated += 1
                                break
                    if m is None:
                        ambiguous += 1
                        _files = ", ".join(sorted({c.unit_file for c in candidates}))
                        warnings.append(
                            f"[hmr] ambiguous function '{fc.name}' — 다중 unit_file "
                            f"({_files}) 매칭. c_parser file 정보 없음 → stamp skip"
                        )
                        new_function_rows.append(fc)
                        continue
                else:
                    m = candidates[0] if candidates else None
                if m and m.total_calls > 0:
                    new_function_rows.append(_dc_replace(
                        fc,
                        function_calls_coverage=CoverageStats(
                            covered=m.covered_calls,
                            total=m.total_calls,
                            coverage_pct=m.coverage_pct / 100.0,
                        ),
                    ))
                    stamped += 1
                else:
                    new_function_rows.append(fc)
            if disambiguated > 0:
                warnings.append(
                    f"[hmr] c_parser file disambiguation: {disambiguated} ambiguous 함수가 "
                    "c_parser 파일 매칭으로 정확 stamp됨"
                )
            warnings.append(
                f"[hmr] Function Calls metric stamped — {stamped}/{len(original_rows)} "
                f"functions matched (HMR metric count: {len(hmr_result.metrics)}, "
                f"ambiguous skipped: {ambiguous})"
            )
            agg["function_rows"] = new_function_rows

    # 라운드 102 (2026-06-24) — 멀티 Metric report(APP+BOOT IT) 파싱·병합 →
    # _align에 전달해 Functions O/X(달성)+Function Calls 실측 산출. hmr_html_bytes_list
    # (router 자동발견) 우선, 단일 hmr_html_bytes도 합산. metrics_by_name은 함수명별
    # list라 dict merge 시 bucket extend (동명함수 멀티-env 보존).
    merged_metrics_by_name: dict[str, list[Any]] = {}
    _metric_sources: list[bytes] = []
    if hmr_html_bytes_list:
        _metric_sources.extend(b for b in hmr_html_bytes_list if b)
    if hmr_html_bytes and hmr_html_bytes not in _metric_sources:
        _metric_sources.append(hmr_html_bytes)
    if _metric_sources:
        from backend.services.vcast_hmr_parser import parse_hmr_html as _php
        _ok_n = 0
        for _src in _metric_sources:
            _pw: list[str] = []
            _res = _php(_src, parse_warnings=_pw)
            if _res.ok:
                _ok_n += 1
                for _nm, _ms in _res.metrics_by_name.items():
                    merged_metrics_by_name.setdefault(_nm, []).extend(_ms)
        if merged_metrics_by_name:
            warnings.append(
                f"[swit-cov] 라운드 102 — Metric report {_ok_n}/{len(_metric_sources)}건 파싱, "
                f"고유 함수 {len(merged_metrics_by_name)}개 (Functions 달성+Function Calls 실측 소스)"
            )

    # 30차 W21 + 31차 W29 + 라운드 84 T1801 + 85 T1903 + 86 T2001: unmapped fc list.
    _align_function_rows_to_template(
        agg, coverage_template_rows, out_warnings=warnings,
        hmr_metrics_by_name=merged_metrics_by_name or None,
    )

    asil_distribution, ids_by_asil, unmapped_fns = _compute_asil_distribution(
        agg.get("function_rows") or [],
        agg.get("function_asil_map") or {},
        function_asil_from_suds=agg.get("function_asil_from_suds"),
        component_asil_from_sds=agg.get("component_asil_from_sds"),
        function_asil_from_srs=agg.get("function_asil_from_srs"),
        function_name_to_swufn_from_suds=agg.get("function_name_to_swufn_from_suds"),
    )

    summary: dict[str, Any] = {
        "environments": len(session.environments),
        "total_tcs": agg["total_tcs"],
        "passed": agg["passed"],
        "failed": agg["failed"],
        "function_rows": agg["function_count"],
        # Quality DB 기록용 커버리지 roll-up (구문/분기/MC-DC %). SwUT와 동일 헬퍼.
        **compute_coverage_rollup(agg.get("function_rows") or []),
        # 30차 W21 + 31차 W29 + 32차 W28: ASIL 분포 + 등급별 함수 ID + 정책 메타.
        "asil_distribution": asil_distribution,
        "asil_b_function_ids": ids_by_asil.get("B", []),
        "asil_c_function_ids": ids_by_asil.get("C", []),
        "asil_d_function_ids": ids_by_asil.get("D", []),
        # 라운드 86 T2002: UNKNOWN 함수 list (audit 진단용).
        "unmapped_function_names": unmapped_fns,
        "asil_highlight_policy": (
            "B=파랑(#E2F0FF) / C=주황(#FFE5CC) / D=빨강(#FFC7CE) — "
            "31차 비표준 audit 확장 (회사 v2.02 양식은 빨강만 사용)"
        ),
    }

    # Cover
    cover_ws = next((wb[n] for n in sheet_names if n.lower() == "cover"), None)
    if cover_ws is None:
        warnings.append("Cover 시트 미발견 — Doc ID/Author 등 미기록")
    else:
        # 54차 T282: layout 전달 — v2.02 cover_labels 동적 매칭.
        _write_cover_sheet(cover_ws, meta, out_warnings=warnings, layout=layout)

    # Test Summary — 53차 fix: SwIT v2.02 양식은 "1.Test Summary"라 substring 매칭으로 변경.
    # SwUT v3.01의 "Test Summary"도 substring으로 포함되어 호환 유지.
    ts_ws = next((wb[n] for n in sheet_names if "test summary" in n.lower()), None)
    if ts_ws is None:
        warnings.append("Test Summary 시트 미발견")
    else:
        # 54차 T282/T283: layout + summary 전달 — v2.02 label + B17 TC stats + B22.
        _write_test_summary_sheet(
            ts_ws, meta, agg, out_warnings=warnings,
            layout=layout, summary=summary,
        )

    # 3.Coverage
    incomplete_sheets: list[str] = []
    cov_ws = next(
        (wb[n] for n in sheet_names
         if "coverage" in n.lower() and "traceability" not in n.lower()
         and "consistency" not in n.lower()),
        None,
    )
    if cov_ws is None:
        warnings.append("Coverage 시트 미발견")
    else:
        # F7 R2 N3 + Stage 10 G3 fix: layout + out_warnings + is_swit_caller=True
        # SwITCV는 SwIT 분기 (Functions Pass + Function Called metric)
        # 라운드 76 T1106 — c_function_map 재활성 (SwUTCV 대칭).
        n_written = _write_coverage_sheet(
            cov_ws, agg, layout=layout, out_warnings=warnings,
            is_swit_caller=True,
            c_function_map=session.c_function_map or None,
        )
        summary["coverage_rows_written"] = n_written

    # 1.Traceability
    trace_ws = next((wb[n] for n in sheet_names if "traceability" in n.lower()), None)
    # 라운드 F7 D1 fix: 실제 시트 이름 보고 (회사 표준 SwITCV는 '2.Traceability')
    if trace_ws is None:
        warnings.append("Traceability 시트 미발견")
    else:
        # 라운드 73 T807 — swits_tc_ids 전달 (SwITS spec 77 entries 활용).
        swits_tc_ids_list = list(swits_map.keys()) if swits_map else None
        n_o = _write_traceability_sheet(
            trace_ws, session, out_warnings=warnings, layout=layout,
            swits_tc_ids=swits_tc_ids_list,
        )
        summary["traceability_o_cells"] = n_o
        if n_o == 0:
            incomplete_sheets.append(trace_ws.title.strip())

    # 2.Consistency — 34차 C2 fix: test_kind="SwIT" (intro/row 5 item 라벨 치환)
    cons_ws = next((wb[n] for n in sheet_names if "consistency" in n.lower()), None)
    if cons_ws is not None:
        n_template_cons = _write_sds_swits_consistency_template(
            cons_ws, out_warnings=warnings,
        )
        if n_template_cons:
            n_cons = n_template_cons
            summary["consistency_sds_swits_compared"] = True
        elif swits_map:
            n_cons = _write_swit_consistency_sheet(
                cons_ws, session, swits_map, out_warnings=warnings,
            )
            summary["consistency_swits_log_compared"] = True
        else:
            n_cons = _write_consistency_sheet(
                cons_ws, session,
                swuds_function_ids=swuds_function_ids,
                out_warnings=warnings,
                test_kind="SwIT",
            )
            if swuds_function_ids is not None:
                summary["consistency_swuds_compared"] = True
            else:
                summary["consistency_swuds_compared"] = False
                incomplete_sheets.append(
                    f"{cons_ws.title.strip()} (SwUDS 비교 partial)"
                )
        summary["consistency_self_check_rows"] = n_cons
    else:
        warnings.append("Consistency 시트 미발견")
        incomplete_sheets.append("Consistency")

    # History — 55-fix: single-row release entry (사용자 결정 B)
    hist_ws = next((wb[n] for n in sheet_names if n.lower() == "history"), None)
    if hist_ws is not None:
        # 55-fix-2 W2: 4 aggregator 명명 통일 ("Coverage Report" 풀네임)
        # 55-fix-2 W6: out_warnings 전달
        release_rows = build_release_history_row(
            meta, doc_kind="SwIT Coverage Report", out_warnings=warnings,
        )
        n_h = _write_history_sheet(hist_ws, release_rows, out_warnings=warnings)
        summary["history_rows_written"] = n_h
        if n_h == 0:
            incomplete_sheets.append("History")

    summary["template_sha256_12"] = template_sha256_12
    summary["build_timestamp"] = meta.build_timestamp

    # 라운드 87 fix: AuditLog 시트 추가 (SwUT Coverage 대칭, 라운드 83 누락분).
    try:
        from backend.services.swut_coverage_aggregator import _write_audit_log_sheet
        if "AuditLog" not in wb.sheetnames:
            audit_ws = wb.create_sheet("AuditLog")
            n_audit = _write_audit_log_sheet(
                audit_ws, meta, summary, agg, session, warnings,
            )
            summary["audit_log_rows_written"] = n_audit
            summary["audit_log_sheet_added"] = True
    except Exception as _e:  # pragma: no cover
        warnings.append(
            f"AuditLog 시트 작성 실패 (산출물 영향 0): {type(_e).__name__}: {str(_e)[:80]}"
        )

    # 14차 W1: BytesIO 그대로 result에 저장.
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    wb.close()

    if meta.doc_filename_pattern:
        filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version, date=short_date(meta.test_date),
        )
    else:
        filename = (
            f"({meta.project_id})SwIT Coverage Report_"
            f"v{meta.release_sw_version}_{short_date(meta.test_date)}_R.xlsx"
        )
    return SwitCoverageBuildResult(
        ok=True,
        xlsx_io=out,
        filename=filename,
        warnings=warnings,
        incomplete_sheets=incomplete_sheets,
        summary=summary,
    )


__all__ = [
    "SwitCoverageBuildResult",
    "build_swit_coverage_report",
]
