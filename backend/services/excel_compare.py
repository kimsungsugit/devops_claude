"""Excel 파일 비교 기능

두 Excel 파일의 특정 시트를 비교하여 차이점을 찾습니다.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
    Workbook = load_workbook
except ImportError:
    Workbook = None


@dataclass
class ExcelCompareItem:
    """Excel 비교 아이템"""
    path_source: Path
    path_target: Path
    sheet_source: int = 1  # 1-based
    sheet_target: int = 1  # 1-based
    
    @property
    def valid(self) -> bool:
        """유효성 검사"""
        if not self.path_source.exists() or not self.path_target.exists():
            return False
        
        if self.sheet_source <= 0 or self.sheet_target <= 0:
            return False
        
        source_ext = self.path_source.suffix.lower()
        target_ext = self.path_target.suffix.lower()
        
        return source_ext in ['.xlsx', '.xlsm'] and target_ext in ['.xlsx', '.xlsm']


@dataclass
class ExcelDiffItem:
    """Excel 차이점 아이템"""
    row: int
    column: int
    source_data: str
    target_data: str


def read_excel_sheet(filepath: Path, sheet_index: int) -> Optional[List[List[str]]]:
    """Excel 파일의 특정 시트 읽기
    
    Args:
        filepath: Excel 파일 경로
        sheet_index: 시트 인덱스 (1-based)
    
    Returns:
        시트 데이터 (행 리스트)
    """
    if Workbook is None:
        raise ImportError("openpyxl is required")
    
    if not filepath.exists():
        return None
    
    try:
        wb = load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        
        if sheet_index < 1 or sheet_index > len(sheet_names):
            return None
        
        sheet = wb[sheet_names[sheet_index - 1]]
        data = []
        
        for row in sheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            data.append(row_data)
        
        wb.close()
        return data
    
    except Exception as e:
        print(f"Error reading Excel file {filepath}: {e}")
        return None


def compare_excel_files(compare_item: ExcelCompareItem) -> List[ExcelDiffItem]:
    """두 Excel 파일 비교
    
    Args:
        compare_item: Excel 비교 설정
    
    Returns:
        차이점 리스트
    """
    if not compare_item.valid:
        return []
    
    # 소스 파일 읽기
    source_data = read_excel_sheet(compare_item.path_source, compare_item.sheet_source)
    if source_data is None:
        return []
    
    # 타겟 파일 읽기
    target_data = read_excel_sheet(compare_item.path_target, compare_item.sheet_target)
    if target_data is None:
        return []
    
    # 행/열 최대값 계산
    max_rows = max(len(source_data), len(target_data))
    max_cols = max(
        max(len(row) for row in source_data) if source_data else 0,
        max(len(row) for row in target_data) if target_data else 0
    )
    
    # 열 개수 차이가 너무 크면 에러
    if abs((len(source_data[0]) if source_data else 0) - (len(target_data[0]) if target_data else 0)) > 10:
        return []
    
    diffs = []
    
    # 비교 수행
    for row_idx in range(max_rows):
        source_row = source_data[row_idx] if row_idx < len(source_data) else None
        target_row = target_data[row_idx] if row_idx < len(target_data) else None
        
        # 행이 없는 경우
        if source_row is None or target_row is None:
            source_val = "Null" if source_row is None else (source_row[0] if source_row else "")
            target_val = "Null" if target_row is None else (target_row[0] if target_row else "")
            diffs.append(ExcelDiffItem(
                row=row_idx + 1,
                column=0,
                source_data=source_val,
                target_data=target_val
            ))
            continue
        
        # 각 셀 비교
        for col_idx in range(max_cols):
            source_val = source_row[col_idx] if col_idx < len(source_row) else ""
            target_val = target_row[col_idx] if col_idx < len(target_row) else ""
            
            if source_val != target_val:
                diffs.append(ExcelDiffItem(
                    row=row_idx + 1,
                    column=col_idx + 1,
                    source_data=str(source_val),
                    target_data=str(target_val)
                ))
    
    return diffs
