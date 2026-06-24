"""SwIT (Software Integration Test) SITR v2.02 xlsm 빌더 (34차).

SwUT SUTR 빌더 (17차 + 31차 W27 + 31-fix D10/D15) 패턴 90% 차용. xlsm 매크로
보존 (keep_vba=True) — 회사 v2.02 양식 (HDPDM01_GN7) 호환.

기존 자산 100% 재활용:
    - 시트 writer 5개 (Cover / Test Summary / Deviation / Test Log / History)
      — `swut_sutr_aggregator` import (private 함수 그대로 사용)
    - History writer — `swut_coverage_aggregator._write_history_sheet`
    - ASIL 분포 — `swut_coverage_aggregator._compute_asil_distribution`
    - 자체 일관성 (2.Consistency) — `swut_coverage_aggregator._write_consistency_sheet`
    - VBA 보존 sanity — `excel_template_utils.has_vba_macros` + `inspect_vba_refs`

SwIT 도구별 차이 (34차):
    1. 파일명 — `(HDPDM01_SITR) Software Integration Test Result_v<VER>_<DATE>_R.xlsm`
       (사용자 레퍼런스 `(HDPDM01_SITR) Software Integration Test Result_v2.02_240219.xlsm`
       패턴 정확 매칭)
    2. 결과 dataclass — `SwitSitrBuildResult` (xlsm_io 등 SwUT SutrBuildResult 동일)
    3. tool_qualification — manual review 의무 동일

ISO 26262 Integration test:
    SwIT SITR은 ASIL B+ 이상의 evidence — 분기 커버리지 + 인터페이스 테스트 결과 +
    Deviation 기록. ASIL D Test Log row 시각 강조 (31차 W27 col+4/5).
    evidence "auto-generated draft" — manual review 의무.
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from typing import Any

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

from backend.services.excel_layout_resolver import inspect_swit_layout
from backend.services.excel_template_utils import (
    auto_expand_row_block,
    build_release_history_row,
    clear_data_range,
    find_kv_row,
    has_vba_macros,
    inspect_vba_refs,
    push_sentinel_to_last_row,
    safe_write,
    short_date,
    validate_build_meta,
    validate_xlsx_template_bytes,
)
from backend.services.swit_meta import SwitSitrBuildMeta
from backend.services.swut_builder_helpers import extract_warnings_from_session
from backend.services.swut_coverage_aggregator import (
    _compute_asil_distribution,
    _write_consistency_sheet,
    _write_history_sheet,
)
from backend.services.swut_input_adapter import (
    SwUTSession,
    aggregate_session,
)
from backend.services.swut_sutr_aggregator import (
    _write_cover,
    _write_deviation,
    _write_test_summary,
)


@dataclass
class SwitSitrBuildResult:
    """SwIT SITR 빌드 결과 (SwUT `SutrBuildResult` 패턴 동일).

    14차 W1 메모리 절약 — ``xlsm_io: BytesIO`` 주 저장소.
    """
    ok: bool
    xlsm_io: io.BytesIO = field(default_factory=io.BytesIO)
    filename: str = ""
    warnings: list[str] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    incomplete_sheets: list[str] = field(default_factory=list)
    # VBA 매크로 ZIP entry 존재 여부 (deep-reviewer W2) — 실제 실행은 사용자 의무.
    vba_macros_preserved: bool = False
    tool_qualification: dict[str, Any] = field(
        default_factory=lambda: {
            "evidence_class": "auto-generated draft",
            "asil_a_usage": "reviewer 승인 후 evidence로 사용 가능",
            "asil_b_c_d_usage": "단독 evidence 사용 금지 — manual review 의무",
        }
    )

    @property
    def xlsm_bytes(self) -> bytes:
        """Backward compat — BytesIO 전체를 bytes로 복사 (테스트/감사용)."""
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0)
        try:
            return self.xlsm_io.read()
        finally:
            self.xlsm_io.seek(pos)

    @property
    def result_size_bytes(self) -> int:
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0, 2)
        size = self.xlsm_io.tell()
        self.xlsm_io.seek(pos)
        return size

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "filename": self.filename,
            "result_size_bytes": self.result_size_bytes,
            "warnings": self.warnings,
            "incomplete_sheets": self.incomplete_sheets,
            "vba_macros_preserved": self.vba_macros_preserved,
            "summary": self.summary,
            "tool_qualification": self.tool_qualification,
        }


def _norm_swit_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def _find_swit_env_match(
    unit_key: str, env_by_key: dict[str, Any],
) -> tuple[str, Any] | tuple[None, None]:
    env = env_by_key.get(unit_key)
    if env is not None:
        return unit_key, env
    candidates = [
        (env_key, env_value)
        for env_key, env_value in env_by_key.items()
        if unit_key.startswith(env_key)
        and unit_key[len(env_key):].isdigit()
        and len(unit_key) - len(env_key) <= 2
    ]
    if len(candidates) == 1:
        return candidates[0]
    return None, None


def _swit_entry_method(entry: Any) -> str:
    test_method = (getattr(entry, "test_method", "") or "").strip()
    generation_method = (getattr(entry, "generation_method", "") or "").strip()
    if test_method and generation_method:
        return f"{test_method}, {generation_method}"
    return test_method or generation_method or "REQ, IFT, AOR, ABV"


def _swit_test_method(entry: Any) -> str:
    raw = (getattr(entry, "test_method", "") or "").strip() or "REQ, IFT"
    return re.sub(r"\s*,\s*", ", ", raw.replace("\n", " ")).strip()


def _swit_generation_method(entry: Any) -> str:
    raw = (getattr(entry, "generation_method", "") or "").strip() or "AOR, ABV "
    normalized = re.sub(r"\s*,\s*", ", ", raw.replace("\n", " ")).strip()
    return f"{normalized} " if normalized == "AOR, ABV" else normalized


def _swit_env_result(env: Any) -> str:
    results = list((getattr(env, "test_results", {}) or {}).values())
    if not results:
        return "N/A"
    return "Pass" if all(getattr(r, "passed", False) for r in results) else "Fail"


def _swit_value(value: Any) -> str:
    if isinstance(value, tuple):
        value = value[0] if value else ""
    return "" if value is None else str(value)


def _swit_numeric_value(value: str) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return int(text, 0)
    except ValueError:
        return None


def _swit_missing_value(*, reference_layout: bool) -> str:
    return "N/A" if reference_layout else ""


def _swit_lookup_value(
    data: dict[str, Any],
    name: str,
    *,
    reference_layout: bool,
) -> str:
    if name not in data:
        return _swit_missing_value(reference_layout=reference_layout)
    value = _swit_value(data.get(name))
    if value == "":
        return _swit_missing_value(reference_layout=reference_layout)
    return value


def _swit_actual_value(
    actual: dict[str, Any],
    expected: dict[str, Any],
    name: str,
    *,
    reference_layout: bool,
) -> str:
    if name not in actual:
        return _swit_missing_value(reference_layout=reference_layout)
    value = _swit_value(actual.get(name))
    if value == "":
        return _swit_missing_value(reference_layout=reference_layout)
    expected_value = _swit_value(expected.get(name, ""))
    if reference_layout and expected_value:
        actual_num = _swit_numeric_value(value)
        expected_num = _swit_numeric_value(expected_value)
        if actual_num is not None and actual_num == expected_num:
            return expected_value.strip()
    return value


def _swit_var_names(
    env: Any,
    tc_names: list[str] | None = None,
) -> tuple[list[str], list[str], list[str]]:
    input_names: list[str] = []
    expected_names: list[str] = []
    actual_names: list[str] = []
    seen_in: set[str] = set()
    seen_exp: set[str] = set()
    seen_act: set[str] = set()
    selected_tc_names = tc_names or sorted((getattr(env, "test_cases", {}) or {}).keys())
    for tc_name in selected_tc_names:
        items = env.test_cases.get(tc_name) or []
        item = items[0] if items else None
        if item is not None:
            for key in (getattr(item, "input_data", {}) or {}):
                if key not in seen_in:
                    seen_in.add(key)
                    input_names.append(key)
            for key in (getattr(item, "expected_result", {}) or {}):
                if key not in seen_exp:
                    seen_exp.add(key)
                    expected_names.append(key)
        result = (getattr(env, "test_results", {}) or {}).get(tc_name)
        actual = getattr(result, "actual_result", {}) or {} if result else {}
        if not actual:
            result_items = getattr(env, "tc_result_items", {}).get(tc_name, [])
            result_item = result_items[0] if result_items else None
            actual = getattr(result_item, "actual_result", {}) or {} if result_item else {}
        for key in actual:
            if key not in seen_act:
                seen_act.add(key)
                actual_names.append(key)
    if not actual_names:
        actual_names = list(expected_names)
    return input_names, expected_names, actual_names


def _swit_norm_description(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip().lower()


def _swit_tc_names_for_block(env: Any, block_description: str) -> list[str]:
    """Return TC names in an env that belong to one SwITS block description."""
    tc_map = getattr(env, "test_cases", {}) or {}
    all_tc_names = sorted(tc_map.keys())
    desc_key = _swit_norm_description(block_description)
    if not desc_key:
        return all_tc_names

    matched: list[str] = []
    for tc_name in all_tc_names:
        items = tc_map.get(tc_name) or []
        item = items[0] if items else None
        item_desc = _swit_norm_description(getattr(item, "description", "") if item else "")
        if not item_desc:
            continue
        if item_desc == desc_key or item_desc.startswith(desc_key) or desc_key.startswith(item_desc):
            matched.append(tc_name)
    return matched or all_tc_names


def _swit_tc_id_from_env(env: Any) -> str:
    env_name = getattr(env, "env_name", "") or ""
    match = re.search(r"SwUFn_(\d+(?:_\d+)?)", env_name, flags=re.IGNORECASE)
    if match:
        return f"SwITC_{match.group(1)}"
    return env_name or "SwITC_UNKNOWN"


def _merge_swit_block_columns(ws: Any, start_row: int, end_row: int, columns: list[int]) -> None:
    if end_row <= start_row:
        return
    existing_ranges = {str(rng) for rng in ws.merged_cells.ranges}
    for col_idx in sorted({col for col in columns if col and col > 0}):
        range_ref = (
            f"{ws.cell(start_row, col_idx).coordinate}:"
            f"{ws.cell(end_row, col_idx).coordinate}"
        )
        if range_ref in existing_ranges:
            continue
        ws.merge_cells(range_ref)


def _build_swit_test_log_blocks(
    session: SwUTSession,
    swits_map: dict[str, Any] | None,
    *,
    split_entries: bool = False,
) -> list[dict[str, Any]]:
    env_by_key = {
        _norm_swit_key(getattr(env, "env_name", "") or ""): env
        for env in session.environments
    }
    used_env_keys: set[str] = set()
    blocks: list[dict[str, Any]] = []
    missing_spec_blocks: list[dict[str, Any]] = []

    for tc_id, entry in (swits_map or {}).items():
        unit_key = _norm_swit_key(getattr(entry, "unit_name", "") or "")
        matched_key, env = _find_swit_env_match(unit_key, env_by_key)
        if env is not None:
            used_env_keys.add(matched_key or unit_key)
            blocks.append({
                "tc_id": getattr(entry, "tc_id", "") or tc_id,
                "description": (
                    getattr(entry, "description", "")
                    or getattr(entry, "unit_name", "")
                    or ""
                ),
                "method": _swit_test_method(entry),
                "generation_method": _swit_generation_method(entry),
                "precondition": getattr(entry, "precondition", "") or "",
                "env": env,
                "note": "",
            })
            continue
        missing_spec_blocks.append({
            "tc_id": getattr(entry, "tc_id", "") or tc_id,
            "description": getattr(entry, "description", "") or getattr(entry, "unit_name", "") or "",
            "method": _swit_test_method(entry),
            "generation_method": _swit_generation_method(entry),
            "precondition": getattr(entry, "precondition", "") or "",
            "env": None,
            "note": "SwITS spec entry - VectorCAST log missing",
        })

    if not split_entries and blocks:
        grouped_blocks: dict[str, dict[str, Any]] = {}
        for block in blocks:
            env = block["env"]
            env_key = _norm_swit_key(getattr(env, "env_name", "") or "")
            group = grouped_blocks.setdefault(env_key, {"env": env, "blocks": []})
            group["blocks"].append(block)
        blocks = []
        for group in grouped_blocks.values():
            group_blocks = group["blocks"]
            tc_ids = [b["tc_id"] for b in group_blocks]
            descriptions = list(dict.fromkeys(
                b["description"] for b in group_blocks if b["description"]
            ))
            methods = list(dict.fromkeys(b["method"] for b in group_blocks if b["method"]))
            generation_methods = list(dict.fromkeys(
                b["generation_method"] for b in group_blocks if b["generation_method"]
            ))
            preconditions = list(dict.fromkeys(
                b["precondition"] for b in group_blocks if b["precondition"]
            ))
            description = descriptions[0] if descriptions else ""
            if len(tc_ids) > 1:
                description = f"{description}\nGrouped SwITS: {', '.join(tc_ids)}"
            blocks.append({
                "tc_id": ", ".join(tc_ids),
                "description": description,
                "method": ", ".join(methods),
                "generation_method": ", ".join(generation_methods),
                "precondition": "\n".join(preconditions),
                "env": group["env"],
                "note": "",
            })

    blocks.extend(missing_spec_blocks)

    for env in session.environments:
        env_key = _norm_swit_key(getattr(env, "env_name", "") or "")
        if env_key in used_env_keys:
            continue
        blocks.append({
            "tc_id": _swit_tc_id_from_env(env),
            "description": getattr(env, "component_name", "") or getattr(env, "env_name", "") or "",
            "method": "REQ, IFT",
            "generation_method": "AOR, ABV ",
            "precondition": "",
            "env": env,
            "note": "VectorCAST log entry - SwITS spec match missing",
        })

    return blocks


def _write_swit_test_log(
    ws: Any,
    session: SwUTSession,
    *,
    layout: Any = None,
    swits_map: dict[str, Any] | None = None,
    out_warnings: list[str] | None = None,
) -> tuple[int, int]:
    pos = (
        find_kv_row(ws, "TC ID", max_row=15)
        or find_kv_row(ws, "TC_ID", max_row=15)
        or find_kv_row(ws, "Test Case ID", max_row=15)
    )
    if pos is None:
        if out_warnings is not None:
            out_warnings.append("SwITR Test Log TC ID header not found")
        return 0, 0

    start_row = pos[0] + 1
    is_reference_swit_log = (
        (ws.max_column or 0) >= 300
        and
        str(ws.cell(4, 6).value or "").strip().lower() == "tc_id"
        and str(ws.cell(4, 10).value or "").strip().lower() == "inpt[0]"
    )
    blocks = _build_swit_test_log_blocks(
        session, swits_map, split_entries=is_reference_swit_log,
    )
    reference_slots: list[tuple[int, int]] = []
    if is_reference_swit_log:
        for merged_range in sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
            if (
                merged_range.min_col == 2
                and merged_range.max_col == 2
                and merged_range.min_row >= start_row
            ):
                reference_slots.append((merged_range.min_row, merged_range.max_row))
    if is_reference_swit_log and reference_slots and len(blocks) > len(reference_slots):
        if out_warnings is not None:
            out_warnings.append(
                "[swit-test-log] reference template slot overflow: "
                f"blocks={len(blocks)}, slots={len(reference_slots)} — "
                "extra blocks skipped to preserve footer/template layout",
            )
        blocks = blocks[:len(reference_slots)]
    if is_reference_swit_log and reference_slots:
        rows_needed = max(end for _start, end in reference_slots) - start_row + 1
    else:
        rows_needed = sum(1 + len((getattr(b["env"], "test_cases", {}) or {})) for b in blocks)
    if rows_needed <= 0:
        return 0, 0

    sentinel_row = None
    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in range(1, min(ws.max_column, 40) + 1):
            value = ws.cell(row_idx, col_idx).value
            if isinstance(value, str) and "End of Document" in value:
                sentinel_row = row_idx
                break
        if sentinel_row:
            break
    data_slots = (sentinel_row - start_row) if sentinel_row else max(0, ws.max_row - start_row + 1)
    if rows_needed > data_slots:
        insert_at = sentinel_row or (ws.max_row + 1)
        shortage = rows_needed - data_slots
        auto_expand_row_block(
            ws,
            insert_at_row=insert_at,
            amount=shortage,
            template_row_idx=start_row,
            copy_style=True,
            copy_merge=True,
            copy_dimension=True,
        )
        try:
            push_sentinel_to_last_row(ws)
        except Exception:  # noqa: BLE001
            pass

    is_kjpds_swit_log = (
        str(ws.cell(5, 8).value or "").strip().lower() == "input"
        and str(ws.cell(5, 18).value or "").strip().lower() == "expected result"
        and str(ws.cell(5, 28).value or "").strip().lower() == "actual result"
    )

    if not is_reference_swit_log:
        for merged_range in list(ws.merged_cells.ranges):
            if (
                merged_range.max_row >= start_row
                and merged_range.min_row <= min(ws.max_row, start_row + rows_needed + 20)
                and merged_range.max_col >= 1
                and merged_range.min_col <= 40
            ):
                try:
                    ws.unmerge_cells(str(merged_range))
                except ValueError:
                    pass
    clear_end_col = 378 if is_reference_swit_log else 40
    preserved_reference_values: dict[tuple[int, int], Any] = {}
    if is_reference_swit_log:
        snapshot_end_row = min(ws.max_row, start_row + rows_needed + 20)
        snapshot_end_col = min(ws.max_column or clear_end_col, clear_end_col)
        for snap_row in range(start_row, snapshot_end_row + 1):
            for snap_col in range(1, snapshot_end_col + 1):
                snap_value = ws.cell(snap_row, snap_col).value
                if snap_value not in (None, ""):
                    preserved_reference_values[(snap_row, snap_col)] = snap_value
    clear_data_range(
        ws,
        start_row=start_row,
        end_row=min(ws.max_row, start_row + rows_needed + 20),
        start_col=1,
        end_col=min(ws.max_column or clear_end_col, clear_end_col),
        preserve_formula=True,
        preserve_merged_anchor=True,
        sentinel_patterns=["End of Document", "< End"],
    )

    if is_reference_swit_log:
        no_col = 2
        env_col = 3
        method_col = 4
        generation_col = 5
        tc_id_col = 6
        iter_index_col = 7
        desc_col = 8
        iter_tc_col = desc_col
        input_col = 10
        expected_col = 108
        actual_col = 242
        unit_result_col = 376
        total_result_col = 377
        log_data_col = 0
        precondition_col = 9
        input_max = 98
        expected_max = 134
        actual_max = 134
    elif is_kjpds_swit_log:
        no_col = 0
        env_col = 0
        generation_col = 0
        tc_id_col = 2
        iter_index_col = 3
        iter_tc_col = 4
        desc_col = 3
        method_col = 5
        input_col = 8
        expected_col = 18
        actual_col = 28
        unit_result_col = 38
        total_result_col = 39
        log_data_col = 40
        precondition_col = 6
        input_max = 10
        expected_max = 10
        actual_max = 10
    else:
        no_col = 0
        env_col = 0
        generation_col = 0
        tc_id_col = 2
        iter_index_col = 3
        iter_tc_col = 4
        desc_col = 3
        method_col = pos[1] + 2
        input_col = getattr(layout, "test_log_input_col", None) or pos[1] + 4
        expected_col = getattr(layout, "test_log_expected_col", None) or input_col + 10
        actual_col = getattr(layout, "test_log_actual_col", None) or expected_col + 10
        unit_result_col = getattr(layout, "test_log_pass_fail_col", None) or pos[1] + 3
        total_result_col = getattr(layout, "test_log_pass_fail_total_col", None) or 0
        log_data_col = getattr(layout, "test_log_log_data_col", None) or 0
        precondition_col = getattr(layout, "test_log_precondition_col", None) or 0
        input_max = 10
        expected_max = 10
        actual_max = 10

    preserved_reference_conflicts = 0

    def _write_reference_guarded(
        row: int,
        col: int,
        value: Any,
        *,
        preserve_reference: bool = False,
    ) -> None:
        nonlocal preserved_reference_conflicts
        if col <= 0:
            return
        if preserve_reference and is_reference_swit_log:
            template_value = preserved_reference_values.get((row, col))
            if (
                template_value not in (None, "")
                and value not in (None, "")
                and str(template_value) != str(value)
            ):
                safe_write(ws, row, col, template_value)
                preserved_reference_conflicts += 1
                return
        safe_write(ws, row, col, value)

    row_idx = start_row
    block_count = 0
    iteration_count = 0
    reference_env_used: dict[str, set[str]] = {}
    for block in blocks:
        env = block["env"]
        input_names: list[str] = []
        expected_names: list[str] = []
        actual_names: list[str] = []
        tc_names: list[str] = []
        slot_end_row: int | None = None
        if is_reference_swit_log and block_count < len(reference_slots):
            row_idx, slot_end_row = reference_slots[block_count]
        if env is not None:
            if is_reference_swit_log and slot_end_row is not None:
                env_key = _norm_swit_key(getattr(env, "env_name", "") or "")
                all_tc_names = sorted((getattr(env, "test_cases", {}) or {}).keys())
                used_tc_names = reference_env_used.setdefault(env_key, set())
                slot_count = max(0, slot_end_row - row_idx)
                matched_tc_names = [
                    name for name in _swit_tc_names_for_block(env, block["description"])
                    if name not in used_tc_names
                ]
                fallback_tc_names = [
                    name for name in all_tc_names
                    if name not in used_tc_names and name not in matched_tc_names
                ]
                tc_names = (matched_tc_names + fallback_tc_names)[:slot_count]
                used_tc_names.update(tc_names)
            else:
                tc_names = sorted((getattr(env, "test_cases", {}) or {}).keys())
            input_names, expected_names, actual_names = _swit_var_names(env, tc_names)
        input_names = input_names[:input_max]
        expected_names = expected_names[:expected_max]
        if is_reference_swit_log:
            actual_names = list(expected_names)
        actual_names = actual_names[:actual_max]

        block_no = str(block_count + 1) if is_reference_swit_log else block_count + 1
        if no_col > 0:
            safe_write(ws, row_idx, no_col, block_no)
        if env_col > 0 and env is not None:
            env_name = getattr(env, "env_name", "")
            safe_write(ws, row_idx, env_col, env_name.upper() if is_reference_swit_log else env_name)
        safe_write(ws, row_idx, tc_id_col, block["tc_id"])
        if not is_reference_swit_log:
            safe_write(ws, row_idx, desc_col, block["description"])
        safe_write(ws, row_idx, method_col, block["method"])
        if generation_col > 0:
            safe_write(ws, row_idx, generation_col, block.get("generation_method", block["method"]))
        if precondition_col > 0:
            safe_write(ws, row_idx, precondition_col, block["precondition"])
        for offset, name in enumerate(input_names):
            _write_reference_guarded(
                row_idx, input_col + offset, name,
                preserve_reference=is_reference_swit_log,
            )
        for offset, name in enumerate(expected_names):
            _write_reference_guarded(
                row_idx, expected_col + offset, name,
                preserve_reference=is_reference_swit_log,
            )
        for offset, name in enumerate(actual_names):
            _write_reference_guarded(
                row_idx, actual_col + offset, name,
                preserve_reference=is_reference_swit_log,
            )

        total_result = _swit_env_result(env) if env is not None else "N/A"
        if is_reference_swit_log and unit_result_col > 0:
            safe_write(ws, row_idx, unit_result_col, total_result)
        if total_result_col > 0:
            safe_write(ws, row_idx, total_result_col, total_result)
        if env is not None and log_data_col > 0:
            safe_write(ws, row_idx, log_data_col, getattr(env, "env_name", ""))
        elif block["note"] and log_data_col > 0:
            safe_write(ws, row_idx, log_data_col, block["note"])

        for iter_idx, tc_name in enumerate(tc_names, start=1):
            iter_row = row_idx + iter_idx
            items = env.test_cases.get(tc_name) or [] if env is not None else []
            item = items[0] if items else None
            result = env.test_results.get(tc_name) if env is not None else None
            actual = getattr(result, "actual_result", {}) or {} if result else {}
            if not actual and env is not None:
                result_items = getattr(env, "tc_result_items", {}).get(tc_name, [])
                result_item = result_items[0] if result_items else None
                actual = getattr(result_item, "actual_result", {}) or {} if result_item else {}
            safe_write(ws, iter_row, iter_index_col, iter_idx)
            if is_reference_swit_log:
                iter_description = (
                    getattr(item, "description", "")
                    or block["description"]
                    or tc_name
                )
                if iter_description.startswith("Interface :"):
                    iter_description = f"                {iter_description}"
                safe_write(ws, iter_row, iter_tc_col, iter_description)
            else:
                safe_write(ws, iter_row, iter_tc_col, tc_name)
            input_data = getattr(item, "input_data", {}) or {} if item else {}
            expected_data = getattr(item, "expected_result", {}) or {} if item else {}
            for offset, name in enumerate(input_names):
                _write_reference_guarded(
                    iter_row,
                    input_col + offset,
                    _swit_lookup_value(
                        input_data, name, reference_layout=is_reference_swit_log,
                    ),
                    preserve_reference=is_reference_swit_log,
                )
            for offset, name in enumerate(expected_names):
                _write_reference_guarded(
                    iter_row,
                    expected_col + offset,
                    _swit_lookup_value(
                        expected_data, name, reference_layout=is_reference_swit_log,
                    ),
                    preserve_reference=is_reference_swit_log,
                )
            for offset, name in enumerate(actual_names):
                _write_reference_guarded(
                    iter_row,
                    actual_col + offset,
                    _swit_actual_value(
                        actual, expected_data, name,
                        reference_layout=is_reference_swit_log,
                    ),
                    preserve_reference=is_reference_swit_log,
                )
            iter_result = "Pass" if result and getattr(result, "passed", False) else ("Fail" if result else "N/A")
            safe_write(ws, iter_row, unit_result_col, iter_result)
            iteration_count += 1

        if not is_reference_swit_log and tc_names:
            per_iteration_cols = {
                iter_index_col,
                iter_tc_col,
                unit_result_col,
                *range(input_col, input_col + input_max),
                *range(expected_col, expected_col + expected_max),
                *range(actual_col, actual_col + actual_max),
            }
            merge_cols = [
                col for col in (
                    no_col,
                    env_col,
                    tc_id_col,
                    desc_col,
                    method_col,
                    generation_col,
                    precondition_col,
                    total_result_col,
                    log_data_col,
                )
                if col and col not in per_iteration_cols
            ]
            _merge_swit_block_columns(ws, row_idx, row_idx + len(tc_names), merge_cols)

        if is_reference_swit_log and slot_end_row is not None:
            row_idx = slot_end_row + 1
        else:
            row_idx += 1 + len(tc_names)
        block_count += 1

    restored_reference_parameters = 0
    if is_reference_swit_log and preserved_reference_values:
        parameter_columns = (
            set(range(input_col, input_col + input_max))
            | set(range(expected_col, expected_col + expected_max))
            | set(range(actual_col, actual_col + actual_max))
        )
        for (ref_row, ref_col), template_value in preserved_reference_values.items():
            if ref_col not in parameter_columns:
                continue
            if ws.cell(ref_row, ref_col).value in (None, ""):
                safe_write(ws, ref_row, ref_col, template_value)
                restored_reference_parameters += 1

    if out_warnings is not None:
        if preserved_reference_conflicts:
            out_warnings.append(
                "[swit-test-log] reference template parameter cells preserved on "
                f"{preserved_reference_conflicts} parser conflicts "
                "(duplicate variable-name ambiguity guard)"
            )
        if restored_reference_parameters:
            out_warnings.append(
                "[swit-test-log] reference template parameter cells restored on "
                f"{restored_reference_parameters} blank parser slots "
                "(wide template slot guard)"
            )
        out_warnings.append(
            f"[swit-test-log] blocks={block_count}, iteration_rows={iteration_count}, "
            f"swits_entries={len(swits_map or {})}, log_envs={len(session.environments)}"
        )
    return block_count, iteration_count


def build_swit_sitr_report(
    session: SwUTSession,
    meta: SwitSitrBuildMeta,
    template_bytes: bytes,
    deviation_cases: list[Any] | None = None,
    swuds_function_ids: set[str] | None = None,
    swuts_map: dict[str, Any] | None = None,
) -> SwitSitrBuildResult:
    """SwIT SITR v2.02 xlsm 생성.

    Args:
        session: SwIT session (input_adapter 출력 — SwUT와 동일 구조).
        meta: 빌드 메타 (doc_id_base="HDPDM01-SITR").
        template_bytes: 회사 v2.02 빈 xlsm 템플릿 bytes (VBA 매크로 포함 가능).
        deviation_cases: deviation 결과 (None이면 빈 Deviation 시트).
        swuds_function_ids: 옵션 — SwUDS 함수 ID set. 제공 시 2.Consistency에
            SwUDS↔SwIT 매핑 row 추가 (시트 존재 시).
        swuts_map: 60차 F6-A — SwITS xlsm parser 결과 (옵션, SwUT swuts_map 인자명 재사용).
            제공 시 Test Log B/C/D + Precondition col에 spec 데이터 stamp.

    Returns:
        SwitSitrBuildResult — xlsm_io 채워짐. 매크로 ZIP entry는 보존되나
        실행 동작은 사용자가 Excel에서 확인 필요 (deep-reviewer W2).
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for SwIT SITR builder")

    # 입력 메타 검증 (SwUT SUTR과 동일 정책).
    validate_build_meta(
        meta.release_sw_version, meta.test_date,
        doc_id_sequence=meta.doc_id_sequence,
        test_engineer=meta.test_engineer,
        author=meta.default_author,
        approver=meta.default_approver or meta.approver_override,
        reviewer=meta.default_reviewer or meta.reviewer_override,
    )
    # Critical (reviewer S): ZIP bomb / magic byte 검증 (xlsm도 ZIP 기반).
    validate_xlsx_template_bytes(template_bytes, label="SwIT SITR template")

    template_sha256_12 = hashlib.sha256(template_bytes).hexdigest()[:12]
    # 37차 fix → 38차 W1 DRY: extract_warnings_from_session helper로 추출.
    warnings: list[str] = extract_warnings_from_session(session)

    # 라운드 73 T816 — 입력 자산 활용도 진단.
    from backend.services.swut_builder_helpers import diagnose_asset_usage
    warnings.extend(diagnose_asset_usage(
        swits_map=swuts_map,  # SITR는 swuts_map kwarg로 SwITS 받음
        c_function_map=session.c_function_map or None,
        swuds_function_map=session.swuds_function_map or None,
    ))

    # 54차 T280 — v2.02 양식 layout 자동 추출 (sha256 keying + LRU).
    layout = inspect_swit_layout(template_bytes, "sitr")
    if layout.warnings:
        warnings.extend([f"[layout] {w}" for w in layout.warnings])

    # deep-reviewer W2: VBA 매크로 ZIP entry 존재 여부 사전 측정.
    template_has_vba = has_vba_macros(template_bytes)
    if template_has_vba:
        warnings.append(
            "VBA macro execution NOT verified — open output xlsm in Excel and verify "
            "macros before submitting as evidence (ZIP entry preserved but stale ref 위험)"
        )
        # 5차 reviewer I1: VBA stale ref 의심 패턴 grep.
        vba_refs_found = inspect_vba_refs(template_bytes)
        if vba_refs_found:
            warnings.append(
                f"VBA stale ref 위험 패턴 발견 — {vba_refs_found} 패턴이 vbaProject.bin에 "
                "존재하며 셀/시트 이동 시 매크로 깨질 위험 (수동 검증 의무)"
            )

    # keep_vba=True — .xlsm 매크로 보존
    wb: Workbook = openpyxl.load_workbook(
        io.BytesIO(template_bytes), keep_vba=True, data_only=False,
    )
    sheet_names = wb.sheetnames

    agg = aggregate_session(session)

    # 30차 W21 + 31차 W29 + 라운드 84 T1801 + 85 T1903 + 86 T2001: unmapped fc list.
    asil_distribution, ids_by_asil, unmapped_fns = _compute_asil_distribution(
        agg.get("function_rows") or [],
        agg.get("function_asil_map") or {},
        function_asil_from_suds=agg.get("function_asil_from_suds"),
        component_asil_from_sds=agg.get("component_asil_from_sds"),
        function_asil_from_srs=agg.get("function_asil_from_srs"),
        function_name_to_swufn_from_suds=agg.get("function_name_to_swufn_from_suds"),
    )

    summary: dict[str, Any] = {
        "environments": len(session.environments),
        "total": agg["total"],
        "tested": agg["tested"],
        "passed": agg["passed"],
        "failed": agg["failed"],
        "deviation_cases_written": 0,
        "test_log_rows_written": 0,
        # 30차 W21 + 31차 W29: SwUT Coverage/SUTR과 동일 키 — UI 노출 통일.
        "asil_distribution": asil_distribution,
        "asil_b_function_ids": ids_by_asil.get("B", []),
        "asil_c_function_ids": ids_by_asil.get("C", []),
        "asil_d_function_ids": ids_by_asil.get("D", []),
        # 라운드 86 T2002: UNKNOWN 함수 list (audit 진단용).
        "unmapped_function_names": unmapped_fns,
        "asil_highlight_policy": (
            "B=파랑(#E2F0FF) / C=주황(#FFE5CC) / D=빨강(#FFC7CE) — "
            "31차 비표준 audit 확장 (회사 v2.02 양식은 빨강만 사용)"
        ),
    }

    # Cover
    cover_ws = next((wb[n] for n in sheet_names if n.lower() == "cover"), None)
    if cover_ws is None:
        warnings.append("Cover 시트 미발견")
    else:
        # 54차 T282: layout 전달
        _write_cover(cover_ws, meta, out_warnings=warnings, layout=layout)

    # Test Summary — 53차 fix: SwIT v2.02 양식의 "1.Test Summary" 등 prefix 호환 substring 매칭.
    ts_ws = next((wb[n] for n in sheet_names if "test summary" in n.lower()), None)
    if ts_ws is None:
        warnings.append("Test Summary 시트 미발견")
    else:
        # 54차 T282/T283: layout + summary → v2.02 label + B17 TC stats + B22
        _write_test_summary(
            ts_ws, meta, agg, out_warnings=warnings,
            layout=layout, summary=summary,
        )

    # Deviation — 53차 fix: substring 매칭
    dev_ws = next((wb[n] for n in sheet_names if "deviation" in n.lower()), None)
    if dev_ws is None:
        # 라운드 74 T903 — Deviation 시트 fallback warning 톤 분리.
        # 회사 KJPDS02 v1.01 양식은 SwITR 4 시트만 (Cover/History/1.Test Summary/2.Test Log)
        # → Deviation 시트 미정의가 정상. WARN 톤 → INFO 톤으로 분리 (audit reviewer 혼동 해소).
        # layout.deviation_sheet_present=False는 inspect_swit_layout이 v1.01 양식 인식한 결과.
        if getattr(layout, "deviation_sheet_present", True) is False:
            warnings.append(
                "[양식정상] Deviation 시트 미발견 — 회사 v1.01 양식 표준 (4 시트). "
                "audit reviewer는 deviation 발생 시 별도 첨부 필요."
            )
        else:
            warnings.append(
                "[양식손상] Deviation 시트 미발견 — v2.02/v3.01 양식은 Deviation 시트 정의 필수. "
                "template 손상 가능성 — 입력 template 확인 의무."
            )
    elif deviation_cases:
        n = _write_deviation(dev_ws, deviation_cases, out_warnings=warnings)
        summary["deviation_cases_written"] = n

    # Test Log — 53차 fix: substring 매칭. 57차 T314: 'test result'도 포함 (v2.02
    # SUTR/SITR 회사 양식 시트명 'Test Result' 호환).
    log_ws = next(
        (wb[n] for n in sheet_names
         if "test log" in n.lower() or "test result" in n.lower()),
        None,
    )
    if log_ws is None:
        warnings.append("Test Log/Result 시트 미발견")
    else:
        n, n_iter = _write_swit_test_log(
            log_ws, session, layout=layout, swits_map=swuts_map,
            out_warnings=warnings,
        )
        summary["test_log_rows_written"] = n
        summary["test_log_iteration_rows_written"] = n_iter

    incomplete_sheets: list[str] = []

    # History — 55-fix: single-row release entry (사용자 결정 B)
    hist_ws = next((wb[n] for n in sheet_names if n.lower() == "history"), None)
    if hist_ws is not None:
        # 55-fix-2 W2: 4 aggregator 명명 통일
        # 55-fix-2 W6: out_warnings 전달
        release_rows = build_release_history_row(
            meta, doc_kind="SwIT SITR", out_warnings=warnings,
        )
        n_h = _write_history_sheet(hist_ws, release_rows, out_warnings=warnings)
        summary["history_rows_written"] = n_h
        if n_h == 0:
            incomplete_sheets.append("History")

    # 2.Consistency — SUTR v3.01과 마찬가지로 옵션 (양식에 없으면 silent skip).
    # 34차 C2 fix: test_kind="SwIT" — intro 텍스트 + row 5 item label 치환.
    cons_ws = next((wb[n] for n in sheet_names if "consistency" in n.lower()), None)
    if cons_ws is not None:
        n_cons = _write_consistency_sheet(
            cons_ws, session,
            swuds_function_ids=swuds_function_ids,
            out_warnings=warnings,
            test_kind="SwIT",
        )
        summary["consistency_self_check_rows"] = n_cons
        if swuds_function_ids is not None:
            summary["consistency_swuds_compared"] = True
        else:
            summary["consistency_swuds_compared"] = False
            incomplete_sheets.append("2.Consistency (SwUDS 비교 partial)")

    # 라운드 83 T1703: AuditLog 시트 신규 추가 (SwUT 대칭).
    try:
        from backend.services.swut_coverage_aggregator import _write_audit_log_sheet
        if "AuditLog" not in wb.sheetnames:
            audit_ws = wb.create_sheet("AuditLog")
            n_audit = _write_audit_log_sheet(
                audit_ws, meta, summary, agg, session, warnings,
            )
            summary["audit_log_rows_written"] = n_audit
            summary["audit_log_sheet_added"] = True
    except Exception as _e:  # pragma: no cover — fail-safe
        warnings.append(
            f"AuditLog 시트 작성 실패 (산출물 영향 0): {type(_e).__name__}: {str(_e)[:80]}"
        )

    # 라운드 107 — 템플릿/기입 수식을 openpyxl이 캐시 미저장(cached=None) → 재계산
    # 안 하는 뷰어에서 공백. fullCalcOnLoad로 열 때 자동 재계산(SwITCV 라운드 102 정합).
    # 캐시 미저장은 불변이라 data_only 다운스트림 영향 0.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # pragma: no cover — openpyxl 버전 차 방어
        pass

    # 14차 W1: BytesIO 그대로 result에 저장.
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    wb.close()

    if meta.doc_filename_pattern:
        filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version, date=short_date(meta.test_date),
        )
    else:
        # 사용자 레퍼런스 파일명 패턴 정확 매칭:
        # `(HDPDM01_SITR) Software Integration Test Result_v2.02_240219.xlsm`
        filename = (
            f"({meta.project_id}_SITR) Software Integration Test Result_"
            f"v{meta.release_sw_version}_{short_date(meta.test_date)}_R.xlsm"
        )

    summary["template_sha256_12"] = template_sha256_12
    summary["build_timestamp"] = meta.build_timestamp
    return SwitSitrBuildResult(
        ok=True,
        xlsm_io=out,
        filename=filename,
        warnings=warnings,
        incomplete_sheets=incomplete_sheets,
        vba_macros_preserved=template_has_vba,
        summary=summary,
    )


__all__ = [
    "SwitSitrBuildResult",
    "build_swit_sitr_report",
]
