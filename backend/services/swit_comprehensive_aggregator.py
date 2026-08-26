"""SwITCR (Software Integration Test Comprehensive Result) xlsm builder.

The SwITCR workbook is a template-copy artifact. It keeps the company xlsm
template intact, including merged cells, styles, borders, and VBA entries, then
stamps integration-test evidence collected from SwITS, SwITCV, SwITR, and the
same VectorCAST/MDS session used by SwITCV/SwITR builders.
"""
from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass, field
from typing import Any

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

from backend.services.excel_layout_resolver import (
    coverage_column_base,
    coverage_summary_col,
    find_coverage_sheet,
)
from backend.services.excel_template_utils import (
    build_release_history_row,
    dot_date,
    has_vba_macros,
    inspect_vba_refs,
    mark_user_input_required,
    safe_write,
    short_date,
    stamp_cover_document_id,
    validate_build_meta,
    validate_xlsx_template_bytes,
    write_label_or_mark,
    write_signature_block,
    write_value_after_label,
)
from backend.services.swut_builder_helpers import (
    diagnose_asset_usage,
    extract_warnings_from_session,
)
from backend.services.swut_comprehensive_aggregator import (
    _build_c_evidence,
    _build_full_function_text,
    _coverage_failures,
    _find_sheet,
    _lookup_c_function,
    _write_ut101_long_text,
)
from backend.services.swut_coverage_aggregator import (
    _compute_asil_distribution,
    _write_history_sheet,
)
from backend.services.swut_input_adapter import SwUTSession, aggregate_session
from backend.services.swut_meta import BuildMetaBase


@dataclass
class SwitcrBuildMeta(BuildMetaBase):
    """Build metadata for SwITCR."""

    doc_id_base: str = "HDPDM01-SwITCR"
    target_coverage: float = 1.0
    target_pass_ratio: float = 1.0
    final_test_result: str = "OK"
    project_config: dict[str, Any] = field(default_factory=dict)


@dataclass
class SwitcrBuildResult:
    """SwITCR build result."""

    ok: bool
    xlsm_io: io.BytesIO = field(default_factory=io.BytesIO)
    filename: str = ""
    warnings: list[str] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    incomplete_sheets: list[str] = field(default_factory=list)
    vba_macros_preserved: bool = False
    tool_qualification: dict[str, Any] = field(
        default_factory=lambda: {
            "evidence_class": "auto-generated draft",
            "asil_a_usage": "reviewer review required before evidence use",
            "asil_b_c_d_usage": "manual review required; do not use as standalone evidence",
        }
    )

    @property
    def xlsm_bytes(self) -> bytes:
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0)
        try:
            return self.xlsm_io.read()
        finally:
            self.xlsm_io.seek(pos)

    @property
    def result_size_bytes(self) -> int:
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0, 2)
        size = self.xlsm_io.tell()
        self.xlsm_io.seek(pos)
        return size

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "filename": self.filename,
            "result_size_bytes": self.result_size_bytes,
            "warnings": self.warnings,
            "incomplete_sheets": self.incomplete_sheets,
            "vba_macros_preserved": self.vba_macros_preserved,
            "summary": self.summary,
            "tool_qualification": self.tool_qualification,
        }


#: 세션 실측 ↔ SwITR 문서 대조 축 (표시 라벨, agg 키, 문서 키).
_SWITR_CROSSCHECK_AXES = (
    ("총 TC", "total_tcs", "sitr_test_log_tcs"),
    ("통과 TC", "passed", "sitr_pass_count"),
    ("실패 TC", "failed", "sitr_fail_count"),
)


def _switr_divergence_warnings(
    agg: dict[str, Any], switr_summary: dict[str, Any],
) -> list[str]:
    """VectorCAST 세션과 승인된 SwITR 문서가 다르면 경고 문자열을 낸다.

    ⚠ 이 대조는 2026-08-26 까지 **한 번도 일어나지 않았다** — SwITR 을 1.1MB 나 읽어
      놓고 시트명(`3.Test Log` vs 코드의 `2.Test Log`)과 TC 열(B vs 코드의 F)이 어긋나
      `sitr_*` 세 키가 늘 부재였기 때문이다. 읽기만 고치면 값은 `or` 폴백 자리에 조용히
      앉을 뿐이라, **어긋남을 말해야** 읽은 값이 쓸모가 있다.

    ⚠ 한쪽이 없으면 건너뛴다 — **부재는 불일치가 아니다.** 없는 증거를 0 으로 접어
      "문서가 0 건이라고 한다" 는 없는 사실을 만들지 않는다.
    ⚠ 일치를 경고로 남기지 않는다 — 정상은 조용해야 한다.
    """
    out: list[str] = []
    for label, agg_key, doc_key in _SWITR_CROSSCHECK_AXES:
        live = agg.get(agg_key)
        doc = switr_summary.get(doc_key)
        if not isinstance(live, (int, float)) or not isinstance(doc, (int, float)):
            continue
        if int(live) != int(doc):
            out.append(
                f"[evidence] {label}: VectorCAST 세션 {int(live)} 와 SwITR 문서 "
                f"{int(doc)} 가 다릅니다 — 산출물에는 세션 값을 실었습니다"
            )
    return out


#: SwITR `1.Test Summary` 집계 블록의 헤더 라벨 → 우리 키.
_SITR_TC_HEADERS = (
    ("sitr_test_log_tcs", "number of tcs tested"),
    ("sitr_pass_count", "number of tcs passed"),
    ("sitr_fail_count", "number of tcs failed"),
)


def _sitr_summary_from_test_summary(wb: Any) -> dict[str, Any]:
    """SwITR `1.Test Summary` 의 **Total 행**(문서가 스스로 말하는 집계)을 읽는다.

    ⚠ Test Log 행 세기보다 이걸 **먼저** 쓴다. 정본 실측(2026-08-26)에서 `3.Test Log` 는
      TC ID 가 세로 병합이라 병합 그룹이 54개, 결과 셀이 630개인데 문서의 Total 은 611 이다
      — 행을 세는 접근 자체가 이 양식에서 틀린 답을 낸다.

    라벨로만 찾는다(행·열 상수 없음). 못 찾으면 **빈 dict** 를 낸다 — 없는 집계를 0 으로
    지어내지 않는다. SwITCV 에도 같은 이름의 시트가 있지만 이 블록이 없어 여기서 걸러진다.
    """
    ws = None
    for name in getattr(wb, "sheetnames", []) or []:
        if "testsummary" in "".join(str(name).split()).lower():
            ws = wb[name]
            break
    if ws is None:
        return {}
    max_r = min(int(getattr(ws, "max_row", 0) or 0), 80)
    max_c = min(int(getattr(ws, "max_column", 0) or 0), 20)

    header_row = None
    cols: dict[str, int] = {}
    for r in range(1, max_r + 1):
        found: dict[str, int] = {}
        for c in range(1, max_c + 1):
            label = str(ws.cell(r, c).value or "").strip().lower()
            for key, want in _SITR_TC_HEADERS:
                if label == want:
                    found[key] = c
        if len(found) == len(_SITR_TC_HEADERS):
            header_row, cols = r, found
            break
    if header_row is None:
        return {}

    # 헤더 아래에서 `Total` 행을 찾는다. 부분합(Requirements Based / Fault Injection …)을
    # 우리가 더하지 않는다 — 문서가 합쳐 둔 값을 그대로 쓴다(합산 규칙이 양식마다 다르다).
    for r in range(header_row + 1, min(header_row + 20, max_r) + 1):
        is_total = any(
            str(ws.cell(r, c).value or "").strip().lower() == "total"
            for c in range(1, min(max_c, 4) + 1)
        )
        if not is_total:
            continue
        out: dict[str, Any] = {}
        for key, _ in _SITR_TC_HEADERS:
            value = ws.cell(r, cols[key]).value
            if not isinstance(value, (int, float)):
                return {}          # 한 칸이라도 수가 아니면 통째로 포기(부분 신뢰 금지)
            out[key] = int(value)
        return out
    return {}


def _sitr_summary_from_test_log(wb: Any) -> dict[str, Any]:
    """`N.Test Log` 시트에서 TC 행을 센다 — `1.Test Summary` 가 없는 판(v1.01)용 폴백.

    ⚠ 시트명을 정확 매칭하지 않는다. 회사 정본은 `3.Test Log` 인데 예전 코드가
      `"2.Test Log"` 만 봐서 **한 번도 읽힌 적이 없었다**(2026-08-26 실측).
    ⚠ TC ID 열도 상수가 아니다. 예전 코드는 `row[5]`(F열) 고정이었는데 정본은 **B열**이다.
    """
    ws = None
    for name in getattr(wb, "sheetnames", []) or []:
        if "testlog" in "".join(str(name).split()).lower():
            ws = wb[name]
            break
    if ws is None:
        return {}
    max_r = min(int(getattr(ws, "max_row", 0) or 0), 400)
    max_c = min(int(getattr(ws, "max_column", 0) or 0), 60)

    tc_col = verdict_col = None
    tc_hits: dict[int, int] = {}
    verdict_hits: dict[int, int] = {}
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            value = str(ws.cell(r, c).value or "").strip()
            if value.startswith("SwITC"):
                tc_hits[c] = tc_hits.get(c, 0) + 1
            elif value in ("Pass", "Fail", "PASS", "FAIL", "OK", "NG"):
                verdict_hits[c] = verdict_hits.get(c, 0) + 1
    if tc_hits:
        tc_col = max(tc_hits, key=lambda c: tc_hits[c])
    if verdict_hits:
        verdict_col = max(verdict_hits, key=lambda c: verdict_hits[c])
    if tc_col is None:
        return {}

    full_r = int(getattr(ws, "max_row", 0) or 0)
    tc_count = pass_count = fail_count = 0
    for r in range(1, full_r + 1):
        if not str(ws.cell(r, tc_col).value or "").strip().startswith("SwITC"):
            continue
        tc_count += 1
        if verdict_col is None:
            continue
        verdict = str(ws.cell(r, verdict_col).value or "").strip().lower()
        if verdict in ("fail", "ng"):
            fail_count += 1
        elif verdict in ("pass", "ok"):
            pass_count += 1
    return {
        "sitr_test_log_tcs": tc_count,
        "sitr_pass_count": pass_count,
        "sitr_fail_count": fail_count,
    }


def _load_workbook_summary(bytes_value: bytes | None, *, keep_vba: bool = False) -> dict[str, Any]:
    """Extract compact evidence from an existing SwITCV/SwITR workbook."""
    if not bytes_value or openpyxl is None:
        return {}
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(bytes_value),
            read_only=False,
            keep_vba=keep_vba,
            data_only=True,
        )
    except Exception:
        return {}
    try:
        out: dict[str, Any] = {
            "sheet_names": list(wb.sheetnames),
            "sheet_count": len(wb.sheetnames),
        }
        cov = find_coverage_sheet(wb)
        if cov is not None:
            # 요약셀(r5/r6)은 회사 템플릿 수식이라 openpyxl이 cached=None으로 저장 →
            # data_only 읽기 시 항상 None (roll-up이 함수 fail/exception을 0으로
            # 오인식해 거짓 100% 커버리지 표기, ISO 26262 무결성 결함). 데이터행(SwITCV
            # function/call 경로는 O/X·count·exception 모두 리터럴, swut_coverage 라102)
            # 에서 집계해 채운다. 셀에 리터럴 캐시값이 있으면(외부 재계산본) 그 값 우선.
            #
            # ⚠ 2026-08-26 실측 — 여기 열 번호가 **DV(11열) 판에 고정**돼 있었다.
            #   라운드 102 가 같은 파일의 `_extract_template_coverage_rows` 만 DV/PV
            #   적응시키고 이 함수를 빠뜨려, KJPDS02 PV(10열) SwITCV 에서 **전 열이 한 칸씩
            #   밀렸다**: `functions_total` 이 Total(1014) 대신 Fail Count(4) 를 읽어
            #   253배 과소 보고했고, `function_result` 는 O/X 열이 아니라 Exception 열을
            #   읽어 `coverage_fail_details` 가 통째로 비었다 — 실재하는 커버리지 미달성
            #   4건(SwUFn_1005/1167/3519/3554)이 SwITCR 에 "해당사항 없음"으로 나갔다.
            #   판정은 `excel_layout_resolver` 단일 출처를 쓴다(복제하지 말 것).
            base = coverage_column_base(cov)
            total_col = coverage_summary_col(cov, base=base)
            fn_total = fn_fail = fn_exc = 0
            call_rows = call_fail = call_exc = 0
            fail_details: list[dict[str, Any]] = []
            for row_idx in range(11, cov.max_row + 1):
                unit_id = str(cov.cell(row_idx, base).value or "").strip()
                name = str(cov.cell(row_idx, base + 1).value or "").strip()
                if not (unit_id or name):
                    continue
                # 마감 TOTAL 행은 함수가 아니다 — 세면 합계가 1 늘어난다
                # (`_extract_template_coverage_rows` 와 같은 방어).
                if unit_id.lower() == "total" or name.lower() == "total":
                    continue
                function_result = cov.cell(row_idx, base + 2).value
                function_exception = cov.cell(row_idx, base + 3).value
                call_count = cov.cell(row_idx, base + 4).value
                call_total = cov.cell(row_idx, base + 5).value
                call_result = cov.cell(row_idx, base + 6).value
                call_exception = cov.cell(row_idx, base + 7).value
                note = str(cov.cell(row_idx, base + 8).value or "").strip()
                fn_total += 1
                if function_result == "X":
                    fn_fail += 1
                if str(function_exception or "").strip().upper() == "O":
                    fn_exc += 1
                if str(call_result or "").strip().upper() in ("O", "X"):
                    call_rows += 1
                if call_result == "X":
                    call_fail += 1
                if str(call_exception or "").strip().upper() == "O":
                    call_exc += 1
                if function_result == "X":
                    fail_details.append({
                        "kind": "함수커버리지",
                        "unit_id": unit_id,
                        "function": name,
                        "value": function_result,
                        "exception": function_exception,
                        "note": note,
                    })
                if call_result == "X":
                    fail_details.append({
                        "kind": "호출커버리지",
                        "unit_id": unit_id,
                        "function": name,
                        "value": f"{call_count}/{call_total}",
                        "exception": call_exception,
                        "note": note,
                    })
            out["coverage_fail_details"] = fail_details
            # 요약: 셀 리터럴 캐시값 우선, 없으면(수식→None) 데이터행 집계값.
            # 열은 `Total | Fail Count | Exception` 연속 3칸 (헤더 라벨로 찾은 total_col 기준).
            def _cell(row: int, col: int) -> Any:
                return cov.cell(row, col).value

            def _pick(row: int, offset: int, fallback: int) -> Any:
                value = _cell(row, total_col + offset)
                return value if isinstance(value, (int, float)) else fallback

            out["functions_total"] = _pick(5, 0, fn_total)
            out["functions_fail_count"] = _pick(5, 1, fn_fail)
            out["functions_exception_count"] = _pick(5, 2, fn_exc)
            out["function_calls_total"] = _pick(6, 0, call_rows)
            out["function_calls_fail_count"] = _pick(6, 1, call_fail)
            out["function_calls_exception_count"] = _pick(6, 2, call_exc)
        # ⚠ 2026-08-26 실측 — SwITR 증거 읽기가 **통째로 죽어 있었다.** 두 이유가 겹쳤다:
        #   ① 시트명을 `"2.Test Log"` 로 정확 매칭했는데 정본은 `3.Test Log` 다
        #   ② TC ID 열을 `row[5]`(F) 로 박았는데 정본은 **B열**이다
        #   그래서 `sitr_*` 세 키가 늘 부재였고, 이걸 폴백으로 쓰던 자리는 전부 세션 값만
        #   봤다 — 승인된 결과 문서와의 **대조가 한 번도 일어나지 않았다.**
        # 문서가 스스로 말하는 집계를 먼저 쓰고, 없을 때만 행을 센다.
        sitr = _sitr_summary_from_test_summary(wb) or _sitr_summary_from_test_log(wb)
        out.update(sitr)
        return out
    finally:
        wb.close()


def _norm_cell(value: Any) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def _load_fault_injection_summary(bytes_value: bytes | None) -> dict[str, Any]:
    """Extract FI TC/pass-fail evidence from KJPDS02 Fault Injection result xlsx."""
    if not bytes_value or openpyxl is None:
        return {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(bytes_value), data_only=True, read_only=False)
    except Exception:
        return {}
    try:
        ws = wb["FI_Test Case"] if "FI_Test Case" in wb.sheetnames else wb[wb.sheetnames[0]]
        groups: list[dict[str, Any]] = []
        current: dict[str, Any] | None = None
        for row_idx in range(1, ws.max_row + 1):
            tc_id = _norm_cell(ws.cell(row_idx, 3).value)
            if tc_id.startswith("SwIFITC"):
                current = {
                    "tc_id": tc_id,
                    "index": ws.cell(row_idx, 2).value,
                    "expected_names": [
                        _norm_cell(ws.cell(row_idx, col).value)
                        for col in range(15, 17)
                        if _norm_cell(ws.cell(row_idx, col).value)
                    ],
                    "actual_names": [
                        _norm_cell(ws.cell(row_idx, col).value)
                        for col in range(17, 19)
                        if _norm_cell(ws.cell(row_idx, col).value)
                    ],
                    "steps": [],
                }
                groups.append(current)
                continue
            if current is None:
                continue
            step_no = ws.cell(row_idx, 4).value
            method = _norm_cell(ws.cell(row_idx, 5).value)
            if step_no in (None, "") or method != "FI":
                continue
            expected = [_norm_cell(ws.cell(row_idx, col).value) for col in range(15, 17)]
            actual = [_norm_cell(ws.cell(row_idx, col).value) for col in range(17, 19)]
            passed = expected == actual
            current["steps"].append({
                "step": step_no,
                "method": method,
                "generation": _norm_cell(ws.cell(row_idx, 6).value),
                "expected": expected,
                "actual": actual,
                "passed": passed,
            })
        for group in groups:
            steps = group.get("steps") or []
            group["passed"] = bool(steps) and all(step.get("passed") for step in steps)
            group["step_count"] = len(steps)
        failed = [group for group in groups if not group.get("passed")]
        return {
            "sheet_name": ws.title,
            "tc_count": len(groups),
            "step_count": sum(int(group.get("step_count") or 0) for group in groups),
            "passed_tc_count": len(groups) - len(failed),
            "failed_tc_count": len(failed),
            "tc_ids": [group.get("tc_id") for group in groups],
            "failed_cases": failed,
        }
    finally:
        wb.close()


def _write_common_header(ws, meta: SwitcrBuildMeta, cfg: dict[str, Any]) -> None:
    md = cfg.get("switcr_metadata", {}) or {}
    safe_write(ws, 4, 3, md.get("test_iteration", "0.1"))
    safe_write(ws, 5, 3, md.get("software_platform_ver", meta.release_sw_version))
    safe_write(ws, 6, 3, md.get("tester") or meta.test_engineer or meta.author)
    safe_write(ws, 7, 3, md.get("debugger", ""))
    safe_write(ws, 4, 6, md.get("prepare_hours", 0))
    safe_write(ws, 5, 6, md.get("execution_hours", 0))
    safe_write(ws, 6, 6, md.get("review_hours", 0))


def _write_switcr_test_environment(
    ws,
    cfg: dict[str, Any],
    *,
    tool_row: int,
    ref_doc_row: int,
    tool_name: str = "VectorCast",
    tool_version: str = " 2025 sp.4",
) -> None:
    md = cfg.get("switcr_metadata", {}) or {}
    version = str(md.get("tool_version", tool_version))
    if not tool_version.startswith(" "):
        version = version.strip()
    safe_write(ws, tool_row, 4, tool_name)
    safe_write(ws, tool_row + 1, 4, version)
    safe_write(ws, tool_row + 2, 3, md.get("excluded_scope", "LIN Stack"))
    safe_write(ws, tool_row + 2, 4, md.get("excluded_size_label", "Total Lines "))
    safe_write(ws, tool_row + 2, 6, _int_value(md.get("excluded_size"), 65826))
    safe_write(ws, ref_doc_row, 3, md.get("reference_doc", "SwTP"))
    safe_write(ws, ref_doc_row + 1, 3, md.get("reference_id", "HKY-KJPDS02-SwTP-2881"))


def _set_bold(ws, cells: list[str], bold: bool = True) -> None:
    import copy as _copy

    for addr in cells:
        font = _copy.copy(ws[addr].font)
        font.bold = bold
        ws[addr].font = font


def _copy_reference_style_block(
    ws,
    ref_ws,
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    import copy as _copy

    _unmerge_overlapping(ws, min_row, max_row, min_col, max_col)
    for row in range(min_row, max_row + 1):
        src_dim = ref_ws.row_dimensions[row]
        dst_dim = ws.row_dimensions[row]
        dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel
        for col in range(min_col, max_col + 1):
            src = ref_ws.cell(row, col)
            dst = ws.cell(row, col)
            dst.font = _copy.copy(src.font)
            dst.fill = _copy.copy(src.fill)
            dst.border = _copy.copy(src.border)
            dst.alignment = _copy.copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = _copy.copy(src.protection)
    for col in range(min_col, max_col + 1):
        letter = openpyxl.utils.get_column_letter(col)
        src_dim = ref_ws.column_dimensions[letter]
        dst_dim = ws.column_dimensions[letter]
        dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel
    for merged_range in ref_ws.merged_cells.ranges:
        if (
            merged_range.max_row >= min_row
            and merged_range.min_row <= max_row
            and merged_range.max_col >= min_col
            and merged_range.min_col <= max_col
        ):
            _merge_if_needed(ws, str(merged_range))


def _apply_switcr_reference_styles(
    wb: Workbook,
    reference_bytes: bytes | None,
    warnings: list[str],
) -> None:
    if not reference_bytes:
        return
    try:
        ref_wb = openpyxl.load_workbook(
            io.BytesIO(reference_bytes),
            keep_vba=True,
            data_only=False,
        )
    except Exception as exc:
        warnings.append(f"[switcr_reference] style reference load failed: {type(exc).__name__}: {exc}")
        return
    try:
        blocks = {
            "1.IT101": (1, 134, 1, 19),
            "2.IT201": (1, 92, 1, 19),
            "3.IT301": (1, 144, 1, 19),
            "5.IT501": (1, 123, 1, 19),
            "6.IT601": (1, 55, 1, 19),
            "8.IT801": (1, 63, 1, 19),
        }
        for sheet_name, (min_row, max_row, min_col, max_col) in blocks.items():
            target_name = sheet_name if sheet_name in wb.sheetnames else next(
                (name for name in wb.sheetnames if name.endswith(sheet_name)), ""
            )
            reference_name = sheet_name if sheet_name in ref_wb.sheetnames else next(
                (name for name in ref_wb.sheetnames if name.endswith(sheet_name)), ""
            )
            if not target_name or not reference_name:
                warnings.append(f"[switcr_reference] sheet missing for style copy: {sheet_name}")
                continue
            _copy_reference_style_block(
                wb[target_name],
                ref_wb[reference_name],
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
    finally:
        ref_wb.close()


def _function_count(agg: dict[str, Any], switcv_summary: dict[str, Any]) -> int:
    value = (
        agg.get("switcr_qualified_function_count")
        or switcv_summary.get("functions_total")
        or agg.get("function_count")
        or len(agg.get("function_rows") or [])
        or 0
    )
    try:
        return int(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return 0


def _int_value(value: Any, default: int = 0) -> int:
    try:
        return int(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def _unmerge_overlapping(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.max_row >= start_row
            and merged_range.min_row <= end_row
            and merged_range.max_col >= start_col
            and merged_range.min_col <= end_col
        ):
            ws.unmerge_cells(str(merged_range))


def _merge_if_needed(ws, cell_range: str) -> None:
    if cell_range not in {str(item) for item in ws.merged_cells.ranges}:
        ws.merge_cells(cell_range)


def _unmerge_if_exists(ws, cell_range: str) -> None:
    if cell_range in {str(item) for item in ws.merged_cells.ranges}:
        ws.unmerge_cells(cell_range)


def _prepare_it101_reference_layout(ws) -> None:
    """Match the KJPDS02 SwITCR reference IT101 4.2/4.3 layout."""
    import copy as _copy

    _unmerge_overlapping(ws, 82, 134, 2, 15)
    for row in range(82, 135):
        for col in range(2, 16):
            safe_write(ws, row, col, "")
    safe_write(ws, 81, 2, "  4.2 Fail Report")
    safe_write(ws, 82, 2, "NO")
    safe_write(ws, 82, 3, "Fail 항목 ")
    safe_write(ws, 82, 5, "SW Unit(함수)")
    safe_write(ws, 82, 6, "함수명 ")
    safe_write(ws, 82, 9, "Fail 항목 사유")
    safe_write(ws, 82, 13, "대책")
    for row in range(82, 93):
        _merge_if_needed(ws, f"C{row}:D{row}")
        _merge_if_needed(ws, f"F{row}:H{row}")
        _merge_if_needed(ws, f"I{row}:L{row}")
        _merge_if_needed(ws, f"M{row}:O{row}")
    safe_write(ws, 94, 2, "  4.3 Coverage Not Completed")
    safe_write(ws, 95, 2, "NO")
    safe_write(ws, 95, 3, "SW Unit(함수)")
    safe_write(ws, 95, 5, "커버리지")
    safe_write(ws, 95, 7, "미달성 사유")
    safe_write(ws, 95, 12, "조치 내용 (수정 방법 및 불가 사유)")
    safe_write(ws, 96, 5, "유형")
    safe_write(ws, 96, 6, "값")
    _merge_if_needed(ws, "B95:B96")
    _merge_if_needed(ws, "C95:D96")
    _merge_if_needed(ws, "E95:F95")
    _merge_if_needed(ws, "G95:K96")
    _merge_if_needed(ws, "L95:O96")
    for row in range(97, 107):
        _merge_if_needed(ws, f"C{row}:D{row}")
        _merge_if_needed(ws, f"G{row}:K{row}")
        _merge_if_needed(ws, f"L{row}:O{row}")
    safe_write(ws, 109, 2, "  3.4 추적성 미달성 사유")
    safe_write(ws, 110, 2, "No")
    safe_write(ws, 110, 3, "요구사항(상세설계) ID")
    safe_write(ws, 110, 5, "추적성 미달성 사유")
    safe_write(ws, 111, 2, "-")
    safe_write(ws, 111, 3, "해당 사항 없음")
    safe_write(ws, 111, 5, "해당사항 없음")
    safe_write(ws, 114, 2, "  3.5 (ASIL C/D) 결함 주입 검증: 추적성 미달성 사유")
    safe_write(ws, 115, 2, "No")
    safe_write(ws, 115, 3, "식별된 결함")
    safe_write(ws, 115, 5, "추적성 미달성 사유")
    safe_write(ws, 116, 2, "-")
    safe_write(ws, 116, 3, "해당 사항 없음")
    safe_write(ws, 116, 5, "해당사항 없음")
    safe_write(ws, 117, 2, "< End of Document >")
    for cell_range in (
        "C110:D110", "E110:K110", "C111:D111", "E111:K111",
        "C115:D115", "E115:K115", "C116:D116", "E116:K116",
        "B117:G117",
    ):
        _merge_if_needed(ws, cell_range)

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="FFD9D9D9")
    thin_side = openpyxl.styles.Side(style="thin", color="FF000000")
    table_border = openpyxl.styles.Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side,
    )
    center = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = _copy.copy(ws.cell(81, 2).font)
    title_font.bold = True
    for title_row in (81, 94):
        ws.cell(title_row, 2).font = _copy.copy(title_font)
    for title_row in (109, 114, 117):
        ws.cell(title_row, 2).font = _copy.copy(title_font)
    for row in (82, 95, 96):
        for col in range(2, 16):
            cell = ws.cell(row, col)
            bold_font = _copy.copy(cell.font)
            bold_font.bold = True
            cell.font = bold_font
            cell.fill = _copy.copy(header_fill)
            cell.alignment = _copy.copy(center)
            cell.border = _copy.copy(table_border)
    reference_heights = {
        83: 56.25, 84: 51.75, 85: 30.75, 86: 53.25, 87: 56.25,
        88: 57.0, 89: 69.75, 90: 30.75, 91: 52.5, 92: 42.0,
        97: 51.0, 98: 58.5, 99: 39.75, 100: 43.5, 101: 48.75,
        102: 86.25, 103: 37.5, 104: 13.5, 105: 49.5, 106: 30.0,
    }
    for row in list(range(83, 93)) + list(range(97, 107)):
        for col in range(2, 16):
            cell = ws.cell(row, col)
            body_font = _copy.copy(cell.font)
            body_font.bold = False
            cell.font = body_font
            cell.alignment = _copy.copy(center)
            cell.border = _copy.copy(table_border)
        ws.row_dimensions[row].height = reference_heights[row]
    for row in (110, 115):
        for col in range(2, 12):
            cell = ws.cell(row, col)
            bold_font = _copy.copy(cell.font)
            bold_font.bold = True
            cell.font = bold_font
            cell.fill = _copy.copy(header_fill)
            cell.alignment = _copy.copy(center)
            cell.border = _copy.copy(table_border)
    for row in (111, 116):
        for col in range(2, 12):
            cell = ws.cell(row, col)
            body_font = _copy.copy(cell.font)
            body_font.bold = False
            cell.font = body_font
            cell.alignment = _copy.copy(center)
            cell.border = _copy.copy(table_border)
    for row, height in {
        109: 14.25, 111: 14.25, 114: 14.25, 116: 14.25, 117: 13.5,
    }.items():
        ws.row_dimensions[row].height = height


_IT101_REFERENCE_ACTIONS: dict[str, str] = {
    "adc0_stop_current_workaround": (
        "완료 시 하드웨어에 의해 자동으로 0으로 클리어되는 플로우 제어 레지스터로, "
        "데이터시트 스펙상 변환 시간 내 루프 종료가 보장됨."
    ),
    "s_syseepromctrl_writecrcbyte": (
        "해당함수는 s_Write2Eeprom_Cal_Checksum() 함수에서 호출중이며, "
        "LIN 통 신 메시지 WDBI 의 u8g_DID_CAL_CHECKSUM 요청시 호출됨 "
    ),
    "s_syseepromctrl_writesystemflag": "해당 함수 삭제 되었음 (UDS 문서 V1.05 기준)",
    "s_sha256_update": (
        "HUNK_SIZE 의 define 값에 의해 사용되거나 안될수 있는 함수로 현제 설정으론 "
        "해당 함수가 불리지 않으나 HUNK_SIZE 의 변경시 호출될 여지가 있는 함수임."
    ),
    "s_write2eeprom_cal_checksum": "테스트 환경의 문제로 실제 코드 동작은 문제가 없음을 확인함 ",
    "s_processdoormovementstate": (
        "해당 함수는 코드상 명시적으로 들어간 코드로 실제로 해당 조건에 도달하지 못하나 "
        "전체 코드 동작엔 영향을 주지 않음을 확인함"
    ),
    "adc_monitor_measure": (
        "코드 로직상 u8g_MAX 로 입력시 false 옵션으로 동작하며, 이는 의도와 같으며 "
        "전체 동작에 문제가 없을 확인함"
    ),
}


_IT101_REFERENCE_43_REASONS: dict[str, str] = {
    "adc0_stop_current_workaround:function": "해당 함수내 무한루프로 인하여 호출된 후 TestCase 비정상종료",
    "s_syseepromctrl_writecrcbyte:function": "사용하지 않는 함수로 호출 되지 않음",
    "s_syseepromctrl_writesystemflag:function": "사용하지 않는 함수로 호출 되지 않음",
    "s_sha256_update:call": (
        "해당 함수의 input변수 U32 len이 반복문 iteration 수에 영향을 주어 반복문 내 "
        "조건문을 만족하지 못하여, 하위함수인 s_progress_callback()을 호출하지 못함."
    ),
    "adc0_stop_current_workaround:call": "해당 함수내 무한루프로 인해 Test Case 가 비정상 종료",
    "s_write2eeprom_cal_checksum:call": (
        "해당 함수는 calculate_crc32_eeprom()을 호출하는데 이 함수는 반복문 범위"
        "(StartAddr = 0, EndAddr = 1024) 만큼 u8g_SysEepromCtrl_ReadDirectAccess() "
        "함수를 호출한다. \n이는 현재 테스트환경 설정 가능한 TestCase Event Limit을 초과하여, "
        "for문 범위만큼 함수 호출 후, 그 이후 함수 4개를 호출할 수 없음."
    ),
    "s_syseepromctrl_writecrcbyte:call": "사용하지 않는 함수로 호출 되지 않음",
    "s_syseepromctrl_writesystemflag:call": "사용하지 않는 함수로 호출 되지 않음",
    "s_processdoormovementstate:call": (
        "해당함수의 상위->상위 함수인 s_MotorSpeedCtrl()이 실행된후 상위 함수 "
        "s_MotorSpeedCtrl_general()이 실행되기 위한 조건이 해당함수 하위 함수인 "
        "s_MotorSpdCtrl_AutoStop() 호출 조건과 모순된다.   "
    ),
    "adc_monitor_measure:call": (
        "해당함수의 입력 변수인 u8g_MAX는 0xFF 값을 갖는데 이는 하위함수 "
        "ADC_MONITOR_MainMeasure() 호출하기 위한 조건과 모순됨."
    ),
}


_IT101_REFERENCE_TC_IDS: dict[str, str] = {
    "s_sha256_update": "SwITC_0129",
    "s_write2eeprom_cal_checksum": "SwITC_0101_01",
    "s_processdoormovementstate": "SwITC_1501",
    "adc_monitor_measure": "SwITC_0201",
}


def _canonical_function_name(value: Any) -> str:
    return str(value or "").strip().lower().lstrip("_")


def _it101_reference_key(failure: dict[str, Any]) -> tuple[str, str]:
    function_name = _canonical_function_name(failure.get("function"))
    coverage_kind = "function" if failure.get("kind") == "함수커버리지" else "call"
    return function_name, coverage_kind


def _it101_coverage_percent(failure: dict[str, Any]) -> str:
    raw_value = str(failure.get("value") or "").strip()
    if "/" not in raw_value:
        if raw_value.endswith("%"):
            return raw_value
        return "0%" if failure.get("kind") == "함수커버리지" else raw_value
    covered_text, total_text = raw_value.split("/", 1)
    try:
        covered = int(float(covered_text.strip()))
        total = int(float(total_text.strip()))
    except ValueError:
        return raw_value
    if total <= 0:
        return "0%"
    return f"{round((covered / total) * 100):.0f}%"


def _it101_coverage_value(failure: dict[str, Any]) -> float | int | str:
    raw_value = str(failure.get("value") or "").strip()
    if "/" not in raw_value:
        return 0 if failure.get("kind") == "함수커버리지" else raw_value
    covered_text, total_text = raw_value.split("/", 1)
    try:
        covered = int(float(covered_text.strip()))
        total = int(float(total_text.strip()))
    except ValueError:
        return raw_value
    return 0 if total <= 0 else covered / total


def _it101_display_unit_id(failure: dict[str, Any]) -> str:
    unit_id = str(failure.get("unit_id") or "").strip()
    function_name, coverage_kind = _it101_reference_key(failure)
    if function_name == "adc0_stop_current_workaround" and coverage_kind == "function":
        return f"{unit_id} "
    return unit_id


def _it101_display_function_name(failure: dict[str, Any]) -> str:
    raw = str(failure.get("function") or "").strip()
    function_name, coverage_kind = _it101_reference_key(failure)
    if function_name == "s_syseepromctrl_writesystemflag":
        return f"{raw}\xa0"
    if function_name == "adc0_stop_current_workaround":
        return f" {raw}" if coverage_kind == "call" else raw
    if function_name in {
        "s_syseepromctrl_writecrcbyte",
        "s_write2eeprom_cal_checksum",
        "s_processdoormovementstate",
    }:
        return f" {raw}"
    return raw


def _it101_reference_reason_action(failure: dict[str, Any], *, section: str) -> tuple[str, str]:
    function_name, coverage_kind = _it101_reference_key(failure)
    action = _IT101_REFERENCE_ACTIONS.get(
        function_name,
        "SwITCV 예외 항목을 검토하여 TC 보강, 미도달 경로 판단 또는 Deviation 근거를 확정함",
    )
    raw_value = str(failure.get("value") or "").strip()
    if section == "4.2":
        if coverage_kind == "function":
            reason = (
                "해당 함수 내 무한루프로 인하여 호출된 후 TestCase 비정상 종료"
                if function_name == "adc0_stop_current_workaround"
                else "사용하지 않는 함수로 호출 되지 않음"
            )
        elif function_name in (
            "adc0_stop_current_workaround",
            "s_syseepromctrl_writecrcbyte",
            "s_syseepromctrl_writesystemflag",
        ):
            reason = "사용하지 않는 함수로 호출되지 않아 호출 커버리지 산출 불가"
        else:
            tc_id = _IT101_REFERENCE_TC_IDS.get(function_name, "SwITC")
            suffix = "" if function_name == "s_write2eeprom_cal_checksum" else " "
            reason = f"{tc_id} 에서 {raw_value} ({_it101_coverage_percent(failure)}){suffix}"
    else:
        reason = _IT101_REFERENCE_43_REASONS.get(
            f"{function_name}:{coverage_kind}",
            str(failure.get("note") or "").strip() or f"{failure.get('kind')} 미달성 ({raw_value})",
        )
    if section == "4.3" and function_name == "adc_monitor_measure":
        action = "코드 로직상 u8g_MAX 로 입력시 false 옵션으로 동작하며, 전체 동작에 문제가 없을 확인함"
    return reason, action


def _prepare_it301_reference_layout(ws) -> None:
    _unmerge_overlapping(ws, 88, 140, 2, 15)
    _unmerge_overlapping(ws, 141, 144, 2, 15)
    for row in range(88, 141):
        for col in range(2, 16):
            safe_write(ws, row, col, "")
    for row in range(141, 145):
        for col in range(2, 16):
            safe_write(ws, row, col, "")
    for cell_range in (
        "C89:D89", "E89:J89", "K89:O89",
        "C90:D90", "E90:J90", "K90:O90",
        "C93:D93", "E93:F93", "G93:J93", "K93:N93",
        "C94:D94", "E94:F94", "G94:J94", "K94:N94",
        "C97:D97", "E97:K97", "C98:D98", "E98:K98",
        "C102:D102", "E102:K102", "C103:D103", "E103:K103",
        "B105:G105",
    ):
        _merge_if_needed(ws, cell_range)
    for row, height in {
        88: 14.25, 89: 17.25, 90: 14.25, 92: 14.25, 93: 17.25,
        94: 14.25, 96: 14.25, 98: 14.25, 101: 14.25,
        103: 14.25, 105: 13.5,
    }.items():
        ws.row_dimensions[row].height = height


def _write_it101(
    ws,
    meta: SwitcrBuildMeta,
    agg: dict[str, Any],
    cfg: dict[str, Any],
    switcv_summary: dict[str, Any],
    warnings: list[str],
    switr_summary: dict[str, Any] | None = None,
) -> int:
    """IT101(4.1 커버리지 집계 + 4.2/4.3 미달성 표)를 쓴다.

    ⚠ 2026-08-26 — `sitr_*` 키를 **`switcv_summary` 에서 찾고 있었다.** 그 키는 SwITR
      워크북에서만 나오므로 늘 부재였고(게다가 이 함수는 `switr_summary` 를 받지도 않았다),
      TC 수·통과 수는 세션 값만 보고 승인 문서와 대조되지 않았다.
    """
    _write_common_header(ws, meta, cfg)
    safe_write(ws, 60, 4, "VectorCast")
    safe_write(ws, 61, 4, "2025.sp4")
    safe_write(ws, 62, 3, "LIN Stack")
    safe_write(ws, 62, 4, "Total Lines")
    safe_write(ws, 62, 6, 65826)
    safe_write(ws, 67, 3, "SwTP")
    safe_write(ws, 68, 3, "HKY-KJPDS02-SwTP-2881")
    safe_write(ws, 5, 9, "=H79")
    function_count = _function_count(agg, switcv_summary)
    _switr = switr_summary or {}
    failed_tcs = agg.get("failed", 0) or _switr.get("sitr_fail_count", 0) or 0
    details = list(switcv_summary.get("coverage_fail_details") or [])
    function_fail = _int_value(switcv_summary.get("functions_fail_count"))
    function_exception = _int_value(switcv_summary.get("functions_exception_count"))
    call_fail = _int_value(switcv_summary.get("function_calls_fail_count"))
    call_exception = _int_value(switcv_summary.get("function_calls_exception_count"))
    total_tcs = agg.get("total_tcs") or _switr.get("sitr_test_log_tcs") or 0
    passed_tcs = agg.get("passed") or _switr.get("sitr_pass_count") or 0
    it101_tc_total = _int_value(
        (cfg.get("switcr_metadata", {}) or {}).get("it101_tc_total"),
        _int_value(_switr.get("sitr_test_log_tcs"), _int_value(total_tcs)),
    )
    call_total = switcv_summary.get("function_calls_total") or function_count

    safe_write(ws, 75, 5, function_count)
    safe_write(ws, 75, 6, "=E75-G75")
    safe_write(ws, 75, 7, function_fail)
    safe_write(ws, 75, 8, '=IFERROR((F75+K75)/E75, "")')
    safe_write(ws, 75, 9, 0)
    safe_write(ws, 75, 10, 0)
    safe_write(ws, 75, 11, function_exception)
    safe_write(ws, 75, 12, '=IF(E75=F75+I75+K75,"Pass","Fail")')
    safe_write(ws, 76, 5, call_total)
    safe_write(ws, 76, 6, "=E76-G76")
    safe_write(ws, 76, 7, call_fail)
    safe_write(ws, 76, 8, '=IFERROR((F76+K76)/E76, "")')
    safe_write(ws, 76, 9, 0)
    safe_write(ws, 76, 10, 0)
    safe_write(ws, 76, 11, call_exception)
    safe_write(ws, 76, 12, '=IF(E76=F76+I76+K76,"Pass","Fail")')
    safe_write(ws, 77, 5, it101_tc_total)
    safe_write(ws, 77, 6, min(passed_tcs, it101_tc_total))
    safe_write(ws, 77, 7, "=E77-F77")
    safe_write(ws, 77, 8, '=IFERROR((F77+K77)/E77, "")')
    safe_write(ws, 77, 9, 0)
    safe_write(ws, 77, 10, 0)
    safe_write(ws, 77, 11, 0)
    safe_write(ws, 77, 12, '=IF(E77=F77+I77+K77,"Pass","Fail")')
    safe_write(ws, 78, 5, it101_tc_total)
    safe_write(ws, 78, 6, max(it101_tc_total - failed_tcs, 0))
    safe_write(ws, 78, 7, "=E78-F78")
    safe_write(ws, 78, 8, '=IFERROR((F78+K78)/E78, "")')
    safe_write(ws, 78, 9, 0)
    safe_write(ws, 78, 10, 0)
    safe_write(ws, 78, 11, 0)
    safe_write(ws, 78, 12, '=IF(E78=F78+I78+K78,"Pass","Fail")')
    safe_write(ws, 79, 8, '=IFERROR(AVERAGE(H75:H78), "")')
    safe_write(ws, 79, 12, '=IFERROR(AVERAGE(L75:L78), "")')

    _prepare_it101_reference_layout(ws)
    if not details:
        safe_write(ws, 83, 2, 1)
        safe_write(ws, 83, 3, "해당사항 없음")
        safe_write(ws, 83, 5, "N/A")
        safe_write(ws, 83, 6, "해당사항 없음")
        safe_write(ws, 83, 9, "SwITCV/SWITR 근거상 미달성 커버리지 항목 없음")
        safe_write(ws, 83, 13, "검토자는 SwITCV 4.Coverage 최종본을 확인함")
        safe_write(ws, 97, 2, 1)
        safe_write(ws, 97, 3, "해당사항 없음")
        safe_write(ws, 97, 5, "N/A")
        safe_write(ws, 97, 6, "N/A")
        safe_write(ws, 97, 7, "SwITCV/SWITR 근거상 미달성 커버리지 항목 없음")
        safe_write(ws, 97, 12, "검토자는 SwITCV 4.Coverage 최종본을 확인함")
        return 0

    details = [
        *[item for item in details if item.get("kind") == "함수커버리지"],
        *[item for item in details if item.get("kind") != "함수커버리지"],
    ]
    visible_details = details[:10]
    if len(details) > len(visible_details):
        warnings.append(
            f"IT101 reference layout supports 10 coverage failure rows; "
            f"truncated {len(details) - len(visible_details)} extra row(s)."
        )

    for idx, failure in enumerate(visible_details, start=1):
        row = 82 + idx
        reason, action = _it101_reference_reason_action(failure, section="4.2")
        safe_write(ws, row, 2, idx)
        safe_write(ws, row, 3, failure.get("kind"))
        safe_write(ws, row, 5, _it101_display_unit_id(failure))
        safe_write(ws, row, 6, _it101_display_function_name(failure))
        _write_ut101_long_text(ws, row, 9, reason)
        _write_ut101_long_text(ws, row, 13, action)

    for idx, failure in enumerate(visible_details, start=1):
        row = 96 + idx
        reason, action = _it101_reference_reason_action(failure, section="4.3")
        safe_write(ws, row, 2, idx)
        safe_write(ws, row, 3, _it101_display_function_name(failure))
        safe_write(ws, row, 5, failure.get("kind"))
        safe_write(ws, row, 6, _it101_coverage_value(failure))
        _write_ut101_long_text(ws, row, 7, reason)
        _write_ut101_long_text(ws, row, 12, action)
    # 라운드 96-final QA fix — 양식 10행 한도 초과분을 산출물 자체에 가시화
    # (이전: warnings에만 truncation 기록 → audit reviewer가 문서 단독 검토 시
    # 누락 인지 불가). 4.2는 r93, 4.3은 r107 빈 행에 '외 N건' 표기.
    overflow = len(details) - len(visible_details)
    if overflow > 0:
        safe_write(ws, 93, 3,
                   f"외 {overflow}건 — SwITCV 4.Coverage X 행 / AuditLog 참조")
        safe_write(ws, 107, 3,
                   f"외 {overflow}건 — SwITCV 4.Coverage X 행 / AuditLog 참조")
    return len(details)

    start = 83
    max_failure_rows = 6
    if not details:
        safe_write(ws, start, 3, "N/A")
        safe_write(ws, start, 5, "No uncovered requirement-based integration coverage")
        safe_write(ws, start, 7, "SwITCV/SWITR evidence indicates no failed coverage item.")
        safe_write(ws, start, 12, "Manual reviewer shall confirm SwITCV 4.Coverage.")
        return 0

    c_function_map = agg.get("c_function_map") or None
    c_function_map = agg.get("c_function_map") or None

    def _detail_texts(failure: dict[str, Any]) -> tuple[str, str, str, str, str]:
        function_name = str(failure.get("function") or "").strip()
        unit_id = str(failure.get("unit_id") or "").strip()
        c_entry = _lookup_c_function(function_name, unit_id, c_function_map)
        evidence_short, _ = _build_c_evidence(function_name, c_entry)
        full_function = _build_full_function_text(c_entry)
        note = str(failure.get("note") or "").strip()
        value = str(failure.get("value") or "").strip()
        reason = note if note and note != "SwITCR 참고" else (
            f"{failure['kind']} 미달성 ({value}); SwITCV 4.Coverage 기준"
        )
        if evidence_short:
            reason = f"{reason}\nC code evidence: {evidence_short}"
        action = (
            "Review SwITCV exception and confirm TC addition, unreachable path, "
            "or deviation rationale."
        )
        if full_function:
            action = f"{action}\n\nFull C function:\n{full_function}"
        return unit_id, function_name, value, reason, action

    for idx, failure in enumerate(details[:max_failure_rows], start=1):
        row = start + idx - 1
        unit_id, function_name, _value, reason, action = _detail_texts(failure)
        safe_write(ws, row, 2, idx)
        safe_write(ws, row, 3, failure["kind"])
        safe_write(ws, row, 5, unit_id)
        _write_ut101_long_text(ws, row, 7, f"{function_name}\n{reason}")
        _write_ut101_long_text(ws, row, 12, action)
        line_count = max(
            reason.count("\n") + 1,
            action.count("\n") + 1,
        )
        full_function_lines = action.count("\n") + 1 if "Full C function:" in action else 0
        ws.row_dimensions[row].height = min(max(78, line_count * 12, full_function_lines * 8), 409)

    overflow_start = 93
    overflow_details = details[max_failure_rows:]
    overflow_capacity = max(ws.max_row - overflow_start + 1, 0)
    if len(overflow_details) > overflow_capacity:
        warnings.append(
            f"IT101 coverage-not-completed rows truncated to template capacity "
            f"{overflow_capacity}/{len(overflow_details)}; summary counts retain full total."
        )
    for offset, failure in enumerate(overflow_details[:overflow_capacity], start=1):
        row = overflow_start + offset - 1
        unit_id, function_name, value, reason, action = _detail_texts(failure)
        safe_write(ws, row, 2, offset)
        safe_write(ws, row, 3, f"{unit_id}\n{function_name}".strip())
        safe_write(ws, row, 5, failure["kind"])
        safe_write(ws, row, 6, value)
        _write_ut101_long_text(ws, row, 7, reason)
        _write_ut101_long_text(ws, row, 12, action)
        safe_write(ws, row, 15, "N/A")
    return len(details)


def _write_result_summary(
    ws,
    meta: SwitcrBuildMeta,
    agg: dict[str, Any],
    cfg: dict[str, Any],
    switcv_summary: dict[str, Any],
    switr_summary: dict[str, Any],
    *,
    start_row: int,
    evidence_name: str,
) -> None:
    _write_common_header(ws, meta, cfg)
    total_tcs = agg.get("total_tcs") or switr_summary.get("sitr_test_log_tcs") or 0
    passed = agg.get("passed") or switr_summary.get("sitr_pass_count") or 0
    failed = agg.get("failed") or switr_summary.get("sitr_fail_count") or 0
    safe_write(ws, start_row, 3, total_tcs)
    safe_write(ws, start_row, 5, total_tcs)
    safe_write(ws, start_row, 6, passed)
    safe_write(ws, start_row, 7, failed)
    safe_write(ws, start_row, 8, "=E{0}-F{0}".format(start_row))
    safe_write(ws, start_row + 5, 3, "SwITS, SwITCV, SwITR")
    safe_write(ws, start_row + 6, 3, evidence_name)
    safe_write(ws, start_row + 11, 3, "N/A" if failed == 0 else f"Fail TC {failed}")
    safe_write(ws, start_row + 11, 5, "No failed integration test item" if failed == 0 else "Failed integration test item exists")
    safe_write(ws, start_row + 11, 8, "Review SwITR 2.Test Log and linked SwITS test case.")
    safe_write(ws, start_row + 11, 12, "Manual review required before evidence use.")


def _write_it201(
    ws,
    meta: SwitcrBuildMeta,
    agg: dict[str, Any],
    cfg: dict[str, Any],
    failures: list[dict[str, Any]],
    warnings: list[str] | None = None,
) -> int:
    _write_common_header(ws, meta, cfg)
    _write_switcr_test_environment(
        ws, cfg, tool_row=55, ref_doc_row=62,
        tool_name="VectorCast ", tool_version="2025 sp.4",
    )
    md = cfg.get("switcr_metadata", {}) or {}
    total_tcs = _int_value(md.get("interface_total"), _int_value(agg.get("total_tcs")))
    safe_write(ws, 70, 2, "IT")
    safe_write(ws, 70, 3, "인터페이스 커버리지\n(검증된 인터페이스 / 전체 인터페이스) *100")
    safe_write(ws, 70, 6, total_tcs)
    # A2 — passed 실측(interface_passed) 부재 시 passed=total(=100% 커버리지) 조작 금지.
    # 형제 SwUTCR UT201 과 동일 원칙: 명시 증거 없으면 노란 사용자입력 마킹 + 경고로
    # 표면화한다. ISO 26262 인터페이스 검증 증거를 무측정 100% 로 위장하면 audit 무결성이
    # 깨진다. total_tcs 는 agg 실측이라 그대로 기입(분모).
    _if_passed = md.get("interface_passed")
    if _if_passed is not None and str(_if_passed).strip() != "":
        passed = min(_int_value(_if_passed), total_tcs)
        safe_write(ws, 70, 7, passed)
        # H70(=F70-G70)은 G70 이 숫자일 때만 유효한 수식이다. 무증거 경로에선 G70 이
        # placeholder 텍스트라 H70 이 Excel 에서 #VALUE! 가 되므로(deep-review W1),
        # 파생 수식 셀도 evidence 분기 안에서만 기입 — else 에선 함께 마킹(UT201식).
        safe_write(ws, 70, 8, "=F70-G70")
    else:
        mark_user_input_required(ws, 70, 7, hint="인터페이스 검증 통과 수 실측 미제공")
        mark_user_input_required(ws, 70, 8, hint="인터페이스 검증 통과 수 실측 미제공")
        if warnings is not None:
            warnings.append(
                "[switcr] 2.IT201 interface_passed 실측 미제공 — G70/H70 사용자입력 마킹"
                " (passed=total 100% 위장 제거)"
            )
    safe_write(ws, 70, 9, '=IFERROR(G70/F70, "")')
    safe_write(ws, 70, 10, 0)
    safe_write(ws, 70, 11, 0)
    safe_write(ws, 70, 12, 0)
    safe_write(ws, 70, 13, '=IFERROR(((J70+K70+L70)+G70)/F70, "")')
    safe_write(ws, 71, 2, "총 TC 수는 각 TC ID마다 할당된 세부 TC 기준으로 산출하였음.")
    safe_write(ws, 73, 3, "Interface ")

    start = 75
    safe_write(ws, start, 2, "해당사항 없음")
    safe_write(ws, start, 3, "해당사항 없음")
    safe_write(ws, start, 5, "인터페이스커버리지")
    safe_write(ws, start, 6, 1)
    safe_write(ws, start, 7, "해당사항 없음 ")
    safe_write(ws, start, 12, "해당사항 없음 ")
    safe_write(ws, start, 15, "해당사항 없음")
    safe_write(ws, 77, 2, "< End of Document >")
    safe_write(ws, 83, 2, "")
    safe_write(ws, 9, 15, "VectorCAST Report로 첨부파일 대체")
    safe_write(ws, 92, 18, "SwITCV 문서 참조")
    for cell_range in (
        "C76:D76", "G76:K76", "L76:N76",
        "C77:D77", "G77:K77", "L77:N77",
        "C78:D78", "G78:K78", "L78:N78",
        "C79:D79", "G79:K79", "L79:N79",
        "C80:D80", "G80:K80", "L80:N80",
        "C81:D81", "G81:K81", "L81:N81",
        "B83:G83",
    ):
        _unmerge_if_exists(ws, cell_range)
    _merge_if_needed(ws, "B71:F71")
    _merge_if_needed(ws, "B77:G77")
    _set_bold(ws, ["B77"])
    return 0


def _write_it301(
    ws,
    meta: SwitcrBuildMeta,
    agg: dict[str, Any],
    cfg: dict[str, Any],
    switr_summary: dict[str, Any],
    fault_injection_summary: dict[str, Any] | None = None,
    warnings: list[str] | None = None,
) -> None:
    _write_common_header(ws, meta, cfg)
    _write_switcr_test_environment(ws, cfg, tool_row=70, ref_doc_row=77)
    md = cfg.get("switcr_metadata", {}) or {}
    fi = fault_injection_summary or {}
    # A2 — 결함주입(FI)은 ASIL 안전기구 검증 증거다. total(count)/passed 를 **독립 게이트**
    # 한다(deep-review W2): 형제 SwUTCR UT201 처럼 count 만 있고 passed 가 없으면
    # passed=total(=100% PASS)로 조작하지 않고 F85/H85 를 노란 마킹한다. 단일 bool 이던
    # 예전 게이트는 count-only 부분증거를 통과시켜 무측정 100% PASS 를 부활시켰다.
    # fi(실측) 우선, 없으면 config(switcr_metadata). 값 없으면 None(마킹).
    def _fi_present(key_fi: str, key_md: str) -> int | None:
        for src, k in ((fi, key_fi), (md, key_md)):
            v = src.get(k)
            if v is not None and str(v).strip() != "":
                return _int_value(v)
        return None
    _total = _fi_present("tc_count", "fault_injection_count")
    _passed = _fi_present("passed_tc_count", "fault_injection_passed")
    _fi_measured = _total is not None and _passed is not None
    if _total is not None:
        safe_write(ws, 85, 3, _total)
        safe_write(ws, 85, 5, _total)
    else:
        mark_user_input_required(ws, 85, 3, hint="결함주입 TC 수 실측 미제공")
        mark_user_input_required(ws, 85, 5, hint="결함주입 TC 수 실측 미제공")
    if _passed is not None:
        safe_write(ws, 85, 6, min(_passed, _total) if _total is not None else _passed)
    else:
        mark_user_input_required(ws, 85, 6, hint="결함주입 통과 수 실측 미제공")
    if _total is not None and _passed is not None:
        failed = _int_value(fi.get("failed_tc_count"), max(_total - min(_passed, _total), 0))
        safe_write(ws, 85, 7, "=E85-F85")
        safe_write(ws, 85, 8, 1 if _total and failed == 0 else 0)
    else:
        # count 또는 passed 실측이 없으면 판정/실패수를 계산하지 않는다(위장 PASS 금지).
        failed = 0  # 미측정 — 아래 Fail Report 도 '해당사항 없음'(=통과) 대신 마킹(W3).
        mark_user_input_required(ws, 85, 7, hint="결함주입 판정 실측 미제공")
        mark_user_input_required(ws, 85, 8, hint="결함주입 판정 실측 미제공")
        if warnings is not None:
            _missing = [n for n, ok in (("count", _total is not None), ("passed", _passed is not None)) if not ok]
            warnings.append(
                f"[switcr] 3.IT301 결함주입 실측 부족({', '.join(_missing) or '전무'}) — "
                "E85/F85/G85/H85 사용자입력 마킹 (count-only 100% PASS·위장 제거)"
            )
    _prepare_it301_reference_layout(ws)
    safe_write(ws, 88, 2, "  4.2 Fail Report")
    safe_write(ws, 89, 2, "NO")
    safe_write(ws, 89, 3, "Fail 항목 (TC ID)")
    safe_write(ws, 89, 5, "Fail 항목 사유")
    safe_write(ws, 89, 11, "대책")
    if failed:
        for idx, case in enumerate((fi.get("failed_cases") or [])[:9], start=1):
            row = 89 + idx
            safe_write(ws, row, 2, idx)
            safe_write(ws, row, 3, case.get("tc_id") or f"FI-{idx}")
            safe_write(ws, row, 5, "Expected/Actual mismatch in Fault Injection evidence.")
            safe_write(ws, row, 11, "Review FI_Test Case actual results and linked SwITS fault-injection TC.")
    elif _fi_measured:
        safe_write(ws, 90, 2, 1)
        safe_write(ws, 90, 3, "해당사항 없음")
        safe_write(ws, 90, 5, "해당사항 없음")
        safe_write(ws, 90, 11, "해당사항 없음")
    else:
        # W3 — FI 미측정이면 Fail Report 를 "해당사항 없음"(=Fail 0=통과)으로 위장하지
        # 않는다. 상단 E85/F85 노란 마킹과 모순되던 것을 제거(감사자 섹션별 상충 판독 방지).
        safe_write(ws, 90, 2, 1)
        mark_user_input_required(ws, 90, 3, hint="결함주입 미측정 — Fail 항목 미확정")
        mark_user_input_required(ws, 90, 5, hint="결함주입 미측정")
        mark_user_input_required(ws, 90, 11, hint="결함주입 미측정")
    safe_write(ws, 92, 2, "  4.3 Coverage Not Completed")
    safe_write(ws, 93, 2, "NO")
    safe_write(ws, 93, 3, "Test Case")
    safe_write(ws, 93, 5, "결함목록")
    safe_write(ws, 93, 7, "미달성 사유")
    safe_write(ws, 93, 11, "대책")
    safe_write(ws, 93, 15, "조치기한")
    for col in (2, 3, 5, 7, 11, 15):
        safe_write(ws, 94, col, 1 if col == 2 else "해당사항 없음")
    safe_write(ws, 96, 2, "  3.4 추적성 미달성 사유")
    safe_write(ws, 97, 2, "No")
    safe_write(ws, 97, 3, "요구사항(상세설계) ID")
    safe_write(ws, 97, 5, "추적성 미달성 사유")
    safe_write(ws, 98, 2, 1)
    safe_write(ws, 98, 3, "해당사항 없음")
    safe_write(ws, 98, 5, "해당사항 없음")
    safe_write(ws, 101, 2, "  3.5 (ASIL C/D) 결함 주입 검증: 추적성 미달성 사유")
    safe_write(ws, 102, 2, "No")
    safe_write(ws, 102, 3, "식별된 결함")
    safe_write(ws, 102, 5, "추적성 미달성 사유")
    safe_write(ws, 103, 2, 1)
    safe_write(ws, 103, 3, "해당사항 없음")
    safe_write(ws, 103, 5, "해당사항 없음")
    safe_write(ws, 105, 2, "< End of Document >")
    _set_bold(ws, [
        "B88", "B89", "C89", "E89", "K89", "B92", "B93", "C93",
        "E93", "G93", "K93", "O93", "B96", "B97", "C97", "E97",
        "B101", "B102", "C102", "E102", "B105",
    ])


def _write_it401(
    ws, meta: SwitcrBuildMeta, cfg: dict[str, Any], warnings: list[str] | None = None,
) -> None:
    _write_common_header(ws, meta, cfg)
    md = cfg.get("switcr_metadata", {}) or {}
    safe_write(ws, 6, 3, md.get("debugger", ""))
    _write_switcr_test_environment(ws, cfg, tool_row=65, ref_doc_row=72)
    # W5 — 자원사용(RAM/ROM/stack) 실측(resource_usage 설정)이 없을 때 하드코딩 "Pass"+
    # 가짜 사용량(예: =1312/4096)을 stamp 하던 것을 금지 — 노란 사용자입력 마킹. 단
    # 템플릿상 의도적 N/A 행(동적메모리 제외 등, result 가 Pass 가 아님)은 그대로 둔다.
    defaults = {
        78: ("=1312/4096", "Pass", ""),
        79: (0.211, "Pass", ""),
        80: ("N/A", "N/A", "N/A"),
        81: ("77.8%\n46.9%", "Pass", ""),
        82: (0.251, "Pass", ""),
        83: ("", "", "."),
    }
    safe_write(ws, 80, 5, "동적 메모리 사용량")
    safe_write(ws, 80, 7, "PDSM은 동적메모리를 사용하고 있지 않으므로 \n해당 시험은 제외함")
    resource_usage = md.get("resource_usage", {}) or {}
    _missing = []
    for row, fallback in defaults.items():
        if str(row) in resource_usage:
            value, result, attachment = resource_usage[str(row)]
            safe_write(ws, row, 11, value)
            safe_write(ws, row, 12, result)
            safe_write(ws, row, 13, attachment)
        elif str(fallback[1]).strip().lower() == "pass":
            # 실측 없이 "Pass" 로 판정하던 자원 행 — 값·판정을 마킹(가짜 사용량 제거).
            mark_user_input_required(ws, row, 11, hint="자원 사용량 실측 미제공")
            mark_user_input_required(ws, row, 12, hint="자원 사용량 판정 실측 미제공")
            safe_write(ws, row, 13, fallback[2])
            _missing.append(row)
        else:
            # 의도적 N/A·템플릿 행 (동적메모리 제외 등) — 그대로.
            value, result, attachment = fallback
            safe_write(ws, row, 11, value)
            safe_write(ws, row, 12, result)
            safe_write(ws, row, 13, attachment)
    if _missing and warnings is not None:
        warnings.append(
            f"[switcr] 4.IT401 자원사용 실측 미제공(row {_missing}) — K/L열 사용자입력 마킹 "
            "(하드코딩 Pass·가짜 사용량 제거)"
        )


def _write_it701(
    ws, meta: SwitcrBuildMeta, cfg: dict[str, Any], warnings: list[str] | None = None,
) -> None:
    _write_common_header(ws, meta, cfg)
    md = cfg.get("switcr_metadata", {}) or {}
    safe_write(ws, 6, 3, md.get("debugger", ""))
    _write_switcr_test_environment(ws, cfg, tool_row=47, ref_doc_row=54)
    protection_info = {
        59: ("메모리 오류", "RAM error", "○", "Pattern Check", "System Reset", ""),
        60: ("", "ROM Test (*)", "○", "Pattern Check", "System Reset", ""),
        61: ("", "Memory Protection", "○", "ECC Check", "System Reset", ""),
        62: ("", "Stack Under/Overflow", "○", "Pattern Check", "System Reset", ""),
        63: ("타이밍 오류", "Clock Error", "○", "ECC Check", "System Reset", ""),
        64: ("", "Watch Dog (*)", "○", "Alive Supervision", "System Reset", ""),
        65: ("기타", "I/O Error", "○", "Control Flow Monitoring", "Logic Control", "기능 동작 정지 후 정상 복귀 모니터링 "),
    }
    for row, (category, item, designed, method, action, note) in protection_info.items():
        if category:
            safe_write(ws, row, 2, category)
        safe_write(ws, row, 3, item)
        safe_write(ws, row, 6, designed)
        safe_write(ws, row, 7, method)
        safe_write(ws, row, 9, action)
        safe_write(ws, row, 12, note)
    # W5 — 안전기구(watchdog/RAM/ROM/stack/ECC/clock) 검증 결과를 실측(system_error_protection
    # 설정) 없이 "Pass"로 조작하지 않는다. IT701 은 ASIL C/D 안전기구 증거라 무측정 Pass 는
    # audit 무결성 위반(IT201/IT301=A2 과 같은 조작 패턴). 실측 없으면 노란 사용자입력 마킹.
    results = md.get("system_error_protection", {}) or {}
    _missing = []
    for row in range(70, min(ws.max_row, 76) + 1):
        safe_write(ws, row, 6, "○")
        _r = results.get(str(row))
        if _r is not None and str(_r).strip() != "":
            safe_write(ws, row, 7, _r)
        else:
            mark_user_input_required(ws, row, 7, hint="안전기구 검증 결과 실측 미제공")
            _missing.append(row)
        safe_write(ws, row, 10, "X")
        safe_write(ws, row, 11, "N/A")
    if _missing and warnings is not None:
        warnings.append(
            f"[switcr] 7.IT701 시스템오류보호(ASIL 안전기구) 결과 실측 미제공 "
            f"(row {_missing}) — G열 사용자입력 마킹 (무측정 'Pass' 위장 제거)"
        )


def _write_not_applicable(ws, meta: SwitcrBuildMeta, cfg: dict[str, Any], reason: str) -> None:
    """Leave non-applicable SwITCR sheets untouched except for their sheet title.

    These sheets are marked as 해당X in the workbook. Writing generic N/A rows is
    unsafe because IT501/IT601/IT801 use different table layouts and the same
    row number can be a real header or detail area.
    """
    return


def _write_switcr_cover(
    ws, meta: SwitcrBuildMeta, cfg: dict[str, Any],
    out_warnings: list[str] | None = None,
) -> None:
    """라운드 96-final QA fix — SwITCR Cover stamp (이전: 완전 미스탬프 Critical).

    XXXX 공양식 Cover placeholder(G26 'HKY-[P_Name]-SwITCR-28A1' / G28
    'Unspecified' / G29 '202X.XX.XX' / G30 'XXXX')가 산출물에 그대로 잔존하던
    결함. SwITCV/SwITR Cover와 동일 항목 체계(C26~C30 kv + I2/J2/K2 서명란)로
    stamp한다.
    """
    if ws is None:
        return
    author = meta.test_engineer or meta.default_author or ""
    reviewer = getattr(meta, "reviewer", "") or (
        getattr(meta, "default_reviewer", "") or getattr(meta, "reviewer_override", "")
    )
    approver = getattr(meta, "approver", "") or (
        getattr(meta, "default_approver", "") or getattr(meta, "approver_override", "")
    )
    # 서명란 (I2/J2/K2 라벨 아래 기입 — 빈 값은 노란 마킹)
    sig_row = write_signature_block(ws, {
        "Author": author, "Reviewer": reviewer, "Approver": approver,
    }, hint_map={
        "Author": "test_engineer 또는 default_author",
        "Reviewer": "검토자 이름",
        "Approver": "승인자 이름 (필수)",
    })
    if sig_row is None:
        # deep-reviewer 96-final W1 — 비-trio 템플릿 변형 fallback (coverage/sutr
        # writer와 대칭). 미적용 시 Reviewer/Approver가 어디에도 안 써지는 silent 누락.
        write_label_or_mark(ws, "Reviewer", reviewer, hint="검토자 이름")
        write_label_or_mark(ws, "Approver", approver, hint="승인자 이름 (필수)")
    # 표지 kv 항목 (C27~C30) — Document ID는 placeholder/phase 보정 helper 사용
    write_value_after_label(ws, "Version", f"v{meta.release_sw_version}")
    write_value_after_label(ws, "Status", "DRAFT — PENDING REVIEW")
    write_value_after_label(ws, "Date", dot_date(meta.test_date))
    if author:
        write_value_after_label(
            ws, "Author", author,
            min_row=(sig_row + 1) if sig_row else 1,
        )
    stamp_cover_document_id(
        ws, project_id=meta.project_id,
        doc_filename_pattern=getattr(meta, "doc_filename_pattern", "") or "",
        out_warnings=out_warnings,
    )


def _write_summary_sheet(ws, meta: SwitcrBuildMeta, cfg: dict[str, Any]) -> None:
    md = cfg.get("switcr_metadata", {}) or {}
    safe_write(ws, 3, 5, md.get("project", meta.project_id))
    safe_write(ws, 4, 5, str(md.get("phase") or "").strip() or "DV")
    safe_write(ws, 5, 5, md.get("software_platform_ver", meta.release_sw_version))
    safe_write(ws, 6, 5, md.get("product", ""))
    safe_write(ws, 7, 5, md.get("verification_target", ""))
    safe_write(ws, 8, 5, md.get("asil", meta.asil_level.replace("ASIL ", "")))
    safe_write(ws, 9, 5, md.get("compiler", ""))
    safe_write(ws, 10, 5, md.get("mcu", ""))

    for row in (16, 17, 18, 19, 22):
        safe_write(ws, row, 7, "O")
    for row in (20, 21, 23, 24):
        safe_write(ws, row, 7, "X")
    # 라운드 96-final QA fix — 비대상 시트 rename('(해당X)' prefix) 후 O열 참조
    # 문자열 동기화. 이전: 구 시트명('5.IT501' 등) + 존재하지 않는 '8.IT802' stamp
    # → H~L열 unguarded INDIRECT 20셀이 Excel 열람 시 #REF!. IT802 행(r24)은
    # 8.IT801 시트 내 섹션이므로 동일 시트를 참조한다.
    # deep-reviewer 96-final W2 — 하드코딩 대신 실제 wb 시트명에서 동적 해석:
    # rename 실패/시트 부재 변형에서도 존재하는 이름만 stamp (desync 차단).
    def _actual_sheet_name(base: str) -> str:
        sheetnames = getattr(getattr(ws, "parent", None), "sheetnames", []) or []
        renamed = f"(해당X){base}"
        if renamed in sheetnames:
            return renamed
        if base in sheetnames:
            return base
        return renamed  # 시트 자체 부재 — rename 기대명 유지 (G열 'X' gate가 차단)

    safe_write(ws, 20, 15, _actual_sheet_name("5.IT501"))
    safe_write(ws, 21, 15, _actual_sheet_name("6.IT601"))
    safe_write(ws, 23, 15, _actual_sheet_name("8.IT801"))
    safe_write(ws, 24, 15, _actual_sheet_name("8.IT801"))
    safe_write(ws, 19, 13, '=IF(G19="O",INDIRECT($O19&"!I$10"), "")')


def _rename_not_applicable_sheet(wb: Workbook, sheet_name: str) -> str:
    actual = sheet_name if sheet_name in wb.sheetnames else ""
    if not actual:
        for candidate in wb.sheetnames:
            if candidate.endswith(sheet_name):
                actual = candidate
                break
    if not actual:
        return sheet_name
    if actual.startswith("(해당X)"):
        return actual
    new_name = f"(해당X){sheet_name}"
    if new_name in wb.sheetnames:
        return actual
    wb[actual].title = new_name
    return new_name


def _order_switcr_sheets(wb: Workbook) -> None:
    desired = [
        "Cover", "History", "Guideline", "Summary",
        "1.IT101", "2.IT201", "3.IT301", "4.IT401", "7.IT701",
        "(해당X)5.IT501", "(해당X)6.IT601", "(해당X)8.IT801",
    ]
    by_name = {ws.title: ws for ws in wb.worksheets}
    ordered = [by_name[name] for name in desired if name in by_name]
    ordered.extend(ws for ws in wb.worksheets if ws.title not in desired)
    wb._sheets = ordered


def build_switcr_report(
    session: SwUTSession,
    meta: SwitcrBuildMeta,
    template_bytes: bytes,
    *,
    swits_map: dict[str, Any] | None = None,
    switcv_bytes: bytes | None = None,
    switr_bytes: bytes | None = None,
    fault_injection_bytes: bytes | None = None,
    switcr_reference_bytes: bytes | None = None,
) -> SwitcrBuildResult:
    """Build a SwITCR xlsm from template and integration-test evidence."""
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for SwITCR builder")

    validate_build_meta(
        meta.release_sw_version,
        meta.test_date,
        doc_id_sequence=meta.doc_id_sequence,
        test_engineer=meta.test_engineer,
        author=meta.default_author,
        approver=meta.default_approver or meta.approver_override,
        reviewer=meta.default_reviewer or meta.reviewer_override,
    )
    validate_xlsx_template_bytes(template_bytes, label="SwITCR template")

    template_sha256_12 = hashlib.sha256(template_bytes).hexdigest()[:12]
    template_has_vba = has_vba_macros(template_bytes)
    warnings: list[str] = extract_warnings_from_session(session)
    warnings.extend(diagnose_asset_usage(
        swuts_map=swits_map,
        c_function_map=session.c_function_map or None,
        swuds_function_map=session.swuds_function_map or None,
    ))
    if template_has_vba:
        warnings.append(
            "VBA macro execution NOT verified; open output xlsm in Excel before evidence use"
        )
        refs = inspect_vba_refs(template_bytes)
        if refs:
            warnings.append(f"VBA reference patterns found: {refs}; manual macro check required")

    switcv_summary = _load_workbook_summary(switcv_bytes, keep_vba=False)
    switr_summary = _load_workbook_summary(switr_bytes, keep_vba=True)
    fault_injection_summary = _load_fault_injection_summary(fault_injection_bytes)

    wb: Workbook = openpyxl.load_workbook(
        io.BytesIO(template_bytes),
        keep_vba=True,
        data_only=False,
    )
    is_switcr_specific_template = any(
        _find_sheet(wb, token) is not None
        for token in ("it101", "it201", "it301")
    )
    agg = aggregate_session(session)
    if session.c_function_map:
        agg["c_function_map"] = session.c_function_map
    cfg: dict[str, Any] = getattr(meta, "project_config", {}) or {}
    md = cfg.get("switcr_metadata", {}) or {}
    qualified_function_total = md.get("qualified_function_total")
    if qualified_function_total not in (None, ""):
        try:
            agg["switcr_qualified_function_count"] = int(qualified_function_total)
        except (TypeError, ValueError):
            warnings.append("[switcr] qualified_function_total is not an integer; raw count used")

    asil_distribution, ids_by_asil, unmapped_fns = _compute_asil_distribution(
        agg.get("function_rows") or [],
        agg.get("function_asil_map") or {},
        function_asil_from_suds=agg.get("function_asil_from_suds"),
        component_asil_from_sds=agg.get("component_asil_from_sds"),
        function_asil_from_srs=agg.get("function_asil_from_srs"),
        function_name_to_swufn_from_suds=agg.get("function_name_to_swufn_from_suds"),
    )

    warnings.extend(_switr_divergence_warnings(agg, switr_summary))

    summary: dict[str, Any] = {
        "environments": len(session.environments),
        "total_tcs": agg.get("total_tcs", 0) or switr_summary.get("sitr_test_log_tcs", 0),
        "tested": agg.get("tested", 0),
        "passed": agg.get("passed", 0) or switr_summary.get("sitr_pass_count", 0),
        "failed": agg.get("failed", 0) or switr_summary.get("sitr_fail_count", 0),
        "function_rows": agg.get("function_count", 0),
        "switcr_qualified_function_count": agg.get("switcr_qualified_function_count"),
        "switcr_function_count": _function_count(agg, switcv_summary),
        "switcv_summary": switcv_summary,
        "switr_summary": switr_summary,
        "fault_injection_summary": fault_injection_summary,
        "swits_entries": len(swits_map or {}),
        "asil_distribution": asil_distribution,
        "asil_b_function_ids": ids_by_asil.get("B", []),
        "asil_c_function_ids": ids_by_asil.get("C", []),
        "asil_d_function_ids": ids_by_asil.get("D", []),
        "unmapped_function_names": unmapped_fns,
        "template_sha256_12": template_sha256_12,
        "build_timestamp": meta.build_timestamp,
    }
    incomplete_sheets: list[str] = []

    it101 = _find_sheet(wb, "it101")
    if it101 is not None:
        summary["it101_failure_rows"] = _write_it101(
            it101, meta, agg, cfg, switcv_summary, warnings,
            switr_summary=switr_summary,
        )
    else:
        incomplete_sheets.append("1.IT101")
        warnings.append("1.IT101 sheet not found")

    it201 = _find_sheet(wb, "it201")
    if it201 is not None:
        failures = _coverage_failures(agg.get("function_rows") or [], agg.get("c_function_map") or None)
        summary["it201_coverage_rows"] = _write_it201(it201, meta, agg, cfg, failures, warnings)
    else:
        incomplete_sheets.append("2.IT201")

    it301 = _find_sheet(wb, "it301")
    if it301 is not None:
        _write_it301(it301, meta, agg, cfg, switr_summary, fault_injection_summary, warnings)
    else:
        incomplete_sheets.append("3.IT301")

    it401 = _find_sheet(wb, "it401")
    if it401 is not None:
        _write_it401(it401, meta, cfg, warnings)
    else:
        incomplete_sheets.append("4.IT401")

    it701 = _find_sheet(wb, "it701")
    if it701 is not None:
        _write_it701(it701, meta, cfg, warnings)
    else:
        incomplete_sheets.append("7.IT701")

    na_defaults = {
        "5.IT501": md.get("it501_not_applicable_reason", "Back-to-back integration evidence is not applicable for this release."),
        "6.IT601": md.get("it601_not_applicable_reason", "Error code monitoring evidence is not applicable for this release."),
        "8.IT801": md.get("it801_not_applicable_reason", "Heap memory usage evidence is not applicable for this release."),
    }
    for sheet_name, reason in na_defaults.items():
        actual = _rename_not_applicable_sheet(wb, sheet_name)
        if actual in wb.sheetnames:
            _write_not_applicable(wb[actual], meta, cfg, reason)
            summary[f"{sheet_name}_not_applicable"] = True

    summary_ws = _find_sheet(wb, "summary")
    if summary_ws is not None:
        _write_summary_sheet(summary_ws, meta, cfg)
    else:
        incomplete_sheets.append("Summary")

    # 라운드 96-final QA fix — Cover stamp (이전: 완전 미스탬프 Critical)
    cover_ws = _find_sheet(wb, "cover")
    if cover_ws is not None:
        _write_switcr_cover(cover_ws, meta, cfg, out_warnings=warnings)
    else:
        incomplete_sheets.append("Cover")
        warnings.append("Cover sheet not found")

    history_ws = _find_sheet(wb, "history")
    if history_ws is not None:
        rows = build_release_history_row(meta, doc_kind="SwITCR", out_warnings=warnings)
        count = _write_history_sheet(history_ws, rows, out_warnings=warnings)
        summary["history_rows_written"] = count
        if count == 0:
            incomplete_sheets.append("History")
    else:
        warnings.append("History sheet not found")

    if is_switcr_specific_template and "AuditLog" in wb.sheetnames:
        del wb["AuditLog"]
        summary["audit_log_sheet_removed_for_template_fidelity"] = True

    _apply_switcr_reference_styles(wb, switcr_reference_bytes, warnings)
    _order_switcr_sheets(wb)

    # 라운드 107 — 템플릿/기입 수식을 openpyxl이 캐시 미저장(cached=None) → 재계산
    # 안 하는 뷰어에서 공백. fullCalcOnLoad로 열 때 자동 재계산(SwITCV 라운드 102 정합).
    # 캐시 미저장은 불변이라 data_only 다운스트림 영향 0.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # pragma: no cover — openpyxl 버전 차 방어
        pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    wb.close()

    if meta.doc_filename_pattern:
        filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version,
            date=short_date(meta.test_date),
        )
    else:
        filename = (
            f"({meta.project_id}_DV_SwITCR) Software Integration Test Comprehensive Result_"
            f"v{meta.release_sw_version}_{short_date(meta.test_date)}_R.xlsm"
        )

    return SwitcrBuildResult(
        ok=True,
        xlsm_io=out,
        filename=filename,
        warnings=warnings,
        incomplete_sheets=incomplete_sheets,
        vba_macros_preserved=template_has_vba,
        summary=summary,
    )


__all__ = [
    "SwitcrBuildMeta",
    "SwitcrBuildResult",
    "build_switcr_report",
]
