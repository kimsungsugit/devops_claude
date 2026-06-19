"""SwUT SUTR (Software Unit Test Result) — **spec-based** Test Log 빌더 (라운드 91).

회사 감사본 양식(KJPDS02 v1.01 등)은 SUTR '3.Test Log' 시트가 SwUTS spec의
'2.SW Unit Test Spec' 시트(Test Case + Input + Expected 이미 완비)와 **동일 구조** +
Actual Result / Pass-Fail / Log Data 섹션 추가다.

기존 ``swut_sutr_aggregator.build_sutr`` (표준 Version3 좁은 38열 양식 + 함수별 블록
재구성)와는 양식이 근본적으로 다르므로 **신규 경로**로 분리한다.

## 핵심 전략 — 표준 SUTR 템플릿 베이스 + spec 시트 이식 (라운드 92 전환)

라운드 91은 SwUTS spec xlsm 자체를 베이스로 복사하여 출력 시트가 spec 구성
(Cover / History / Introduction / 1.Test Environment / 3.Test Log)으로 잘못 나왔다.
레퍼런스 SUTR 구성은 Cover / History / 1.Test Summary / 2.Deviation / 3.Test Log 이다.

라운드 92는 베이스를 **표준 SUTR 템플릿** (회사 ★개발템플릿 Version3 `(XXXX_SwUTR)
...v0.10`) 으로 전환한다:

1. 표준 SUTR 템플릿 xlsm 을 keep_vba 로 로드 (Cover / History / 1.Test Summary /
   2.Deviation / 3.Test Result(좁은 38열) 보유).
2. Cover / 1.Test Summary / 2.Deviation / History 를 기존 표준 writer
   (`swut_sutr_aggregator._write_cover` / `_write_test_summary` / `_write_deviation`
   + `_write_history_sheet`)로 채움 — meta + TC 카운트(Total/Tested/Passed/Failed/
   not-exec) + Requirements coverage.
3. 표준의 좁은 '3.Test Result' 시트를 **제거**.
4. SwUTS spec '2.SW Unit Test Spec' (와이드 268열) 시트를 표준 템플릿 wb 에
   **크로스-워크북 풀 카피** (`copy_sheet_across_workbooks` — value/style/merge/
   column width/row height 보존) 하여 '3.Test Log' 로 이식.
5. 레퍼런스 SUTR 레이아웃에 맞춰 이식 시트에 Actual/Pass-Fail/Log 섹션의 **헤더(r3
   병합 + r4 서브헤더)** 추가 + anchor 스캔 → VectorCAST 매칭으로 Actual 값 /
   iteration Pass-Fail(JF) / 함수 Total(JG) / Log Data(JH) 채움 (라운드 91 로직 유지).

## 레이아웃 (KJPDS02 v1.01 레퍼런스 실측, max_col=268)

| 영역 | 열 | 출처 |
|------|----|----|
| Test Case (Index/TC_ID/Unit/Method/Generation) | B(2)~F(6) | spec 보존 |
| Input (Inpt[0]~) | G(7)~BE(57) | spec 보존 |
| Expected (ExpR[0]~) | BF(58)~FE(161) | spec 보존 |
| Related ID (spec 전용) | FF(162) | → Actual로 대체 |
| Actual Result (Param 1~) | FF(162)~JE(265) | VectorCAST 추가 |
| Pass/Fail (iteration) | JF(266) | VectorCAST 추가 |
| Pass (함수 Total, 세로병합) | JG(267) | VectorCAST 추가 |
| Log Data | JH(268) | VectorCAST 추가 |

## 매칭 키 (라운드 91 실측)

spec TC_ID 숫자(``SwUTC_NNNN``) == VectorCAST ``SwUFn_NNNN`` 숫자. 직접 4-digit
숫자 매칭. iteration: spec G index(1..N) ↔ vcast ``.MMM`` 정렬 순서.

## Actual 채우기 정책

VectorCAST ``actual_result`` 는 dotted 변수명 분리(예 ``_LP0DR.Byte`` →
``_LP0DR``/``Byte``)로 부정확. 레퍼런스 감사본은 Pass iteration의 Actual = Expected
값과 동일하게 기록(실측). 따라서:

- iteration이 Pass면 Actual = 해당 iteration의 Expected 값 복제 (감사본 패턴 일치).
- iteration이 Fail이면 vcast actual_result best-effort + 부재 시 노란 마킹.
- iteration이 미실행(N/A)면 Actual 비움 + Pass/Fail = "N/A".

## ISO 26262 Tool Qualification
ASIL A 한정 draft. B/C/D는 manual review 의무. Input/Expected는 spec 그대로 보존
(audit truth 불변), Actual/Pass-Fail만 추가.

## 라운드 105 — spec 컬럼 레이아웃 동적화 (PV WIP spec 공존)

PV SwUTS spec(작성중 v0.10)은 'Safety Related' 열 삽입(col 5)으로 Test Method=6/
Generation=7/iteration index=8/Inpt[0]=9, Expected=ExpR[0..83](105~188), Related
ID(r4='SUDS')=189 — DV 고정 상수(Expected 58~161, Related ID=162)를 그대로 쓰면
Actual stamp가 Expected 헤더/데이터(105~188)와 SUDS(189)를 덮어쓰고 겹침 병합을
만든다(실측 깨짐 9건). 따라서 위 레이아웃 표는 **DV 레퍼런스 기본값**이며, 실제
빌드는 ``_detect_spec_layout`` 이 r3 그룹 헤더('Expected Result'/'Related ID') +
r4 'ExpR[..]'/'Inpt[..]' 서브헤더를 스캔해 ``SpecLayout`` 을 동적 산출한다:

- Actual 시작 = Related ID 열 (대체 — DV 레퍼런스 SUTR과 동일 규칙).
- Actual 폭 = Expected 폭, Pass/Fail·Pass·Log Data = Actual 끝 +1/+2/+3.
- iteration index 열 = Inpt[0] 직전 열 (DV=G7, PV=8).
- **DV 호환 게이트**: DV spec에서 동적 산출값 == 기존 상수 (58/162/266/267/268).
  스캔 실패 시 DV 상수 fallback + warning.
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass
from typing import Any

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

from backend.services.design_tokens import (
    USER_INPUT_FILL_RGB,
    USER_INPUT_PLACEHOLDER,
)
from backend.services.excel_template_utils import (
    copy_sheet_across_workbooks,
    has_vba_macros,
    inspect_vba_refs,
    mark_asil_a_function,
    mark_asil_b_function,
    mark_asil_c_function,
    mark_asil_d_function,
    mark_asil_qm_function,
    safe_write,
    compact_empty_styled_cells,
    sanitize_xlsm_external_links,
    verify_xlsx_integrity,
    short_date,
    validate_build_meta,
    validate_xlsx_template_bytes,
    write_value_after_label,
)
from backend.services.swut_builder_helpers import extract_warnings_from_session
from backend.services.swut_input_adapter import SwUTSession, aggregate_session
from backend.services.swut_sutr_aggregator import SutrBuildMeta, SutrBuildResult
from backend.services.swuts_excel_parser import XLSM_MAX_BYTES as _XLSM_MAX_BYTES

# ---------------------------------------------------------------------------
# 레이아웃 상수 (KJPDS02 DV v1.01 레퍼런스 실측 — 라운드 105부터 **기본값/fallback**.
# 실제 빌드는 _detect_spec_layout 이 SpecLayout 을 동적 산출. DV spec에서는 산출값
# == 아래 상수 (하위 호환 게이트 — 기존 268열 테스트 단언 무수정 통과 근거).)
# ---------------------------------------------------------------------------

SPEC_SHEET_RE = re.compile(r"(Unit|Integration)\s*Test\s*Spec", re.IGNORECASE)
LOG_SHEET_NAME = "3.Test Log"

# 행/열 위치 (1-based)
HEADER_SECTION_ROW = 3       # 병합 섹션 헤더 (Test Case / Input / Expected / Actual ...)
SUBHEADER_ROW = 4            # 서브헤더 (Index / TC_ID / Inpt[0] / ExpR[0] / Param 1 ...)
DATA_START_ROW = 5           # 첫 함수 anchor

# 라운드 105 — Safety Related 열 삽입(PV col 5) 전수 점검 결과:
#   - COL_INDEX/COL_TC_ID/COL_UNIT(2/3/4)는 DV·PV 동일 (Safety Related는 Unit 뒤
#     삽입 — PV r4 실측 Index(2)/TC_ID(3)/Unit(4)/Safety Related(5)) → 고정 유지.
#   - COL_METHOD/COL_GENERATION은 DV 위치(5/6) 참고용 — 빌더 로직 미사용 (PV는 6/7).
#   - COL_ITER_INDEX/COL_INPUT_START는 PV에서 8/9로 시프트 → SpecLayout 동적
#     (iter_index = Inpt[0] 직전 열). 아래 값은 DV 기본값.
#   - HEADER_SECTION_ROW/SUBHEADER_ROW/DATA_START_ROW(3/4/5)는 DV·PV 동일 → 고정.
COL_INDEX = 2                # B
COL_TC_ID = 3                # C
COL_UNIT = 4                 # D
COL_METHOD = 5               # E (DV 전용 참고 — 로직 미사용)
COL_GENERATION = 6           # F (DV 전용 참고 — 로직 미사용)
COL_ITER_INDEX = 7           # G (DV 기본값 — PV=8, SpecLayout.iter_index 사용)
COL_INPUT_START = 8          # H (DV 기본값 — PV=9, SpecLayout.input_start 사용)

# Expected 끝 = spec 시트 마지막 데이터 열 직전(FE=161). spec FF(162)=Related ID.
# 레퍼런스 SUTR: Actual=FF(162)~JE(265), JF(266)=Pass/Fail, JG(267)=Total, JH(268)=Log.
COL_RELATED_ID = 162         # FF (spec 전용) — Actual로 대체
COL_ACTUAL_START = 162       # FF
COL_PASS_FAIL = 266          # JF
COL_PASS_TOTAL = 267         # JG
COL_LOG_DATA = 268           # JH
ACTUAL_MAX = COL_PASS_FAIL - COL_ACTUAL_START  # 104 (Param 1~104)
# 라운드 104 — Actual Result 서식 미러 소스. Expected Result(BF=58~FE=161, 104열)는
# spec graft로 서식 보유하나 Actual(FF=162~)은 우리 추가 열이라 무서식 → Expected를
# 1:1 미러(같은 변수의 기대/실제값, 동일 레이아웃). 헤더 r3 'Expected Result'@58 확인.
COL_EXPECTED_START = 58      # BF

_FILL_RGB = USER_INPUT_FILL_RGB


# ---------------------------------------------------------------------------
# 라운드 105 — spec 컬럼 레이아웃 동적 산출
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class SpecLayout:
    """spec 시트 컬럼 레이아웃 (라운드 105 — 헤더 스캔 동적 산출).

    기본값 = DV 레퍼런스(KJPDS02 v1.01, 268열) 상수와 동일 — 스캔 실패 fallback /
    레거시 직접 호출(layout 미전달) 시에도 기존 동작 보존. 파생 좌표는 전부
    property 로 단일 진리원화:

    - ``actual_start`` = Related ID 열 (대체) — DV 162 / PV 189.
    - ``actual_max`` = Expected 폭 — DV 104 / PV 84.
    - ``pass_fail``/``pass_total``/``log_data`` = Actual 끝 +1/+2/+3 —
      DV 266/267/268, PV 273/274/275.
    - ``iter_index`` = Inpt[0] 직전 열 — DV 7(G) / PV 8.
    """

    expected_start: int = COL_EXPECTED_START   # 58 (DV) / 105 (PV)
    related_id: int = COL_RELATED_ID           # 162 (DV) / 189 (PV)
    input_start: int = COL_INPUT_START         # 8 (DV) / 9 (PV)
    detected: bool = False                     # True = 헤더 스캔 성공

    @property
    def actual_start(self) -> int:
        return self.related_id

    @property
    def actual_max(self) -> int:
        return self.related_id - self.expected_start

    @property
    def pass_fail(self) -> int:
        return self.actual_start + self.actual_max

    @property
    def pass_total(self) -> int:
        return self.pass_fail + 1

    @property
    def log_data(self) -> int:
        return self.pass_total + 1

    @property
    def iter_index(self) -> int:
        return self.input_start - 1


def _detect_spec_layout(ws, out_warnings: list[str] | None = None) -> SpecLayout:
    """spec 시트 헤더 스캔 → :class:`SpecLayout` 동적 산출 (라운드 105).

    r3 그룹 헤더('Expected Result' 시작 / 'Related ID' 위치)를 1차 소스로,
    r4 'ExpR[..]' 범위를 r3 미발견 시 폴백 소스로 사용한다 (병합 r3 헤더는
    read-write 로드에서 anchor 셀에 값이 있으므로 직접 읽기 가능). iteration
    index 열은 r4 'Inpt[..]' 첫 위치 직전 열로 산출 (DV=7, PV=8 — PV col 7은
    Generation Method 세그먼트 병합이라 고정 상수로는 iteration 행 67% 누락).

    **반드시 ``_write_log_headers`` 전에 호출** — 헤더 stamp가 r3 'Related ID'
    를 'Actual Result'로 덮어쓰므로 이후에는 스캔 불가.

    스캔 실패(헤더 미발견/순서 모순) 시 DV 상수 fallback + warning.
    """
    max_col = min(int(ws.max_column or 0), 4096)

    expected_start: int | None = None
    related_id: int | None = None
    for c in range(1, max_col + 1):
        v = ws.cell(HEADER_SECTION_ROW, c).value
        if not isinstance(v, str):
            continue
        t = " ".join(v.split()).lower()
        if expected_start is None and t == "expected result":
            expected_start = c
        elif related_id is None and t == "related id":
            related_id = c

    expr_first: int | None = None
    expr_last: int | None = None
    inpt_first: int | None = None
    for c in range(1, max_col + 1):
        v = ws.cell(SUBHEADER_ROW, c).value
        if not isinstance(v, str):
            continue
        t = v.strip().lower()
        if t.startswith("expr["):
            if expr_first is None:
                expr_first = c
            expr_last = c
        elif t.startswith("inpt[") and inpt_first is None:
            inpt_first = c

    # r3 미발견 시 r4 ExpR[..] 범위 폴백 (Expected 시작=ExpR[0], Related ID=마지막+1).
    if expected_start is None:
        expected_start = expr_first
    if related_id is None and expr_last is not None:
        related_id = expr_last + 1

    if (
        expected_start is None
        or related_id is None
        or not (COL_UNIT < expected_start < related_id)
    ):
        if out_warnings is not None:
            out_warnings.append(
                "[spec-sutr] spec 헤더 스캔 실패 ('Expected Result'/'Related ID'/"
                "'ExpR[..]' 미발견 또는 순서 모순) — DV 레퍼런스 상수 레이아웃 "
                f"fallback (Actual={COL_ACTUAL_START}, Pass/Fail={COL_PASS_FAIL})"
            )
        return SpecLayout()

    # iteration index = Inpt[0] 직전 열. Inpt 미발견/비정상(<7 — Test Case 메타
    # 열 침범)이면 DV 기본값 8 유지.
    if inpt_first is not None and COL_GENERATION < inpt_first < expected_start:
        input_start = inpt_first
    else:
        input_start = COL_INPUT_START

    return SpecLayout(
        expected_start=expected_start,
        related_id=related_id,
        input_start=input_start,
        detected=True,
    )


# ---------------------------------------------------------------------------
# 라운드 107 — UT201 Fault Injection spec 자동 산출 (확정 규칙 2026-06-12)
#
# 규칙: FI 수량 = spec '2.SW Unit Test Spec'에서 Test Method 열에 'FI' 세그먼트를
# 1개 이상 포함한 TC(함수) 블록 수. Test Method 열은 method 세그먼트 단위
# **세로병합**(anchor만 값)이라 merged range 전개 없이는 FI iteration 행 귀속이
# 누락된다 (PV 실측: 병합 1,865 range).
#
# ground truth 재현 검증 (2026-06-12 실측):
#   - DV bk_SwUTS_v0.11_251126.xlsm → FI 블록 405 / FI iteration 1,598
#   - PV wip_pv_SwUTS_v0.10_260608.xlsm → FI 블록 808 / FI iteration 3,229
# (DV 감사본 수기 402는 전 가용 산출물로 재현 불가한 stale 카운트로 판명 —
#  산출물 warning에 규칙·spec 파일명을 명기해 차이를 추적 가능하게 한다.)
# ---------------------------------------------------------------------------

_TEST_METHOD_LABEL = "test method"
_FI_METHOD_VALUE = "FI"
# 라운드 108 INFO-8 — swuts_excel_parser.XLSM_MAX_BYTES(64MB) 재사용 (리터럴
# 중복 정의 제거 — DoS 가드 단일 진리원).
_SPEC_FI_MAX_BYTES = _XLSM_MAX_BYTES


def _find_test_method_col(ws) -> int | None:
    """'Test Method' 헤더 열 탐색 — r4(서브헤더) 우선, r3 폴백.

    DV=5(E) / PV=6(F — 'Safety Related' 삽입 시프트, 실측). 공백 collapse +
    case-insensitive 비교. 미발견 시 None.
    """
    max_col = min(int(ws.max_column or 0), 64)
    for row in (SUBHEADER_ROW, HEADER_SECTION_ROW):
        for c in range(1, max_col + 1):
            v = ws.cell(row, c).value
            if (
                isinstance(v, str)
                and " ".join(v.split()).strip().lower() == _TEST_METHOD_LABEL
            ):
                return c
    return None


def _method_values_by_row(ws, method_col: int) -> dict[int, Any]:
    """Test Method 열 세로병합 세그먼트 전개 — {row: anchor 값}.

    read-write 로드에서 병합 range의 비-anchor 셀은 None이므로, method_col을
    덮는 모든 merged range의 anchor(min_row, min_col) 값을 range 전 행에 전파.
    map 미포함 행은 caller가 셀 직접 읽기 (비병합 행 — DV는 대부분 비병합 실측).
    """
    merged: dict[int, Any] = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= method_col <= rng.max_col:
            anchor_val = ws.cell(rng.min_row, rng.min_col).value
            for rr in range(rng.min_row, rng.max_row + 1):
                merged[rr] = anchor_val
    return merged


def _extract_fi_from_sheet(
    ws,
    layout: SpecLayout,
    out_warnings: list[str] | None = None,
) -> dict[str, Any] | None:
    """spec 시트에서 FI 블록/iteration 추출 (확정 규칙 — docstring 위 참조).

    Returns:
        {fi_block_keys: set[str(정규화 num)], fi_iter_rows_per_block,
         fi_iters_per_block(iteration index), fi_block_total,
         fi_iteration_total, blocks_total, fi_dup_keys, method_col} 또는
        None (Test Method 열 미발견).

    블록 키는 ``_build_fn_iteration_map`` 과 동일 정규화 (``lstrip('0')``).
    iteration index 파싱 불가 행은 ``fi_iter_rows_per_block`` 에만 남고
    ``fi_iters_per_block`` 에는 빠진다 — 교차검증에서 '미실행' 취급 (보수적).
    """
    method_col = _find_test_method_col(ws)
    if method_col is None:
        if out_warnings is not None:
            out_warnings.append(
                "[swutcr] UT201 FI spec 자동 산출 불가 — 'Test Method' 헤더 열 "
                "미발견 (r3/r4 스캔). 노란 마킹 유지"
            )
        return None

    blocks = _scan_spec_blocks(ws, layout=layout)
    merged_method = _method_values_by_row(ws, method_col)

    def _method_at(row: int) -> str:
        v = merged_method.get(row)
        if v is None and row not in merged_method:
            v = ws.cell(row, method_col).value
        return str(v).strip().upper() if v is not None else ""

    fi_block_keys: set[str] = set()
    fi_iter_rows_per_block: dict[str, list[int]] = {}
    fi_iters_per_block: dict[str, list[int]] = {}
    fi_block_total = 0
    fi_iteration_total = 0
    dup_keys: list[str] = []

    for blk in blocks:
        key = (blk["num"] or "").lstrip("0") or "0"
        block_has_fi = False
        fi_rows: list[int] = []
        fi_iters: list[int] = []
        # anchor 포함 — 실측상 anchor는 method 빈 값 + iteration index 없음이나,
        # FI 세그먼트가 anchor를 덮는 변종도 블록 FI 판정에는 포함 (보수적).
        for r in [blk["anchor"]] + blk["iter_rows"]:
            if _method_at(r) != _FI_METHOD_VALUE:
                continue
            block_has_fi = True
            raw_idx = ws.cell(r, layout.iter_index).value
            if raw_idx in (None, ""):
                continue  # iteration index 없는 행 — FI iteration 아님
            fi_rows.append(r)
            try:
                fi_iters.append(int(str(raw_idx).strip()))
            except (TypeError, ValueError):
                pass  # 행은 남기고 index만 누락 → 교차검증에서 미실행 취급
        if not block_has_fi:
            continue
        fi_block_total += 1
        fi_iteration_total += len(fi_rows)
        if key in fi_block_keys:
            dup_keys.append(key)
            fi_iter_rows_per_block[key].extend(fi_rows)
            fi_iters_per_block[key].extend(fi_iters)
        else:
            fi_block_keys.add(key)
            fi_iter_rows_per_block[key] = fi_rows
            fi_iters_per_block[key] = fi_iters

    if dup_keys and out_warnings is not None:
        out_warnings.append(
            "[swutcr] UT201 FI spec 자동 산출 — 중복 TC_ID 숫자 키 "
            f"{sorted(set(dup_keys))[:5]} (FI 블록 {len(dup_keys)}건 병합 귀속, "
            "교차검증 정확도 영향 가능)"
        )

    return {
        "rule": "Test Method 'FI' 포함 TC 블록 수",
        "blocks_total": len(blocks),
        "fi_block_total": fi_block_total,
        "fi_iteration_total": fi_iteration_total,
        "fi_block_keys": fi_block_keys,
        "fi_iter_rows_per_block": fi_iter_rows_per_block,
        "fi_iters_per_block": fi_iters_per_block,
        # 라운드 108 MINOR-4 — 중복 TC_ID 숫자 키 (추가 블록당 1 entry — 키별
        # 블록 수 복원용). 교차검증이 dup 키 블록을 보수적 미실행 처리하는 입력.
        # DV/PV 실파일 실측 dup 0 — 실파일 가드(warns==[]) 무영향.
        "fi_dup_keys": dup_keys,
        "method_col": method_col,
        "layout_detected": layout.detected,
    }


def extract_spec_fi_stats(
    spec_xlsm_bytes: bytes,
    *,
    spec_filename: str = "",
    out_warnings: list[str] | None = None,
) -> dict[str, Any] | None:
    """SwUTS spec xlsm bytes → UT201 FI 자동 산출 통계 (확정 규칙 2026-06-12).

    SwUTCR `_write_ut201` 의 config ``fault_injection_total/passed`` 부재 시
    spec 산출 경로 입력. 실패(파일 깨짐/시트·헤더 미발견)는 None + warning —
    caller는 기존 노란 마킹 경로 유지 (실측 위장 금지).

    Args:
        spec_xlsm_bytes: SwUTS spec xlsm bytes ('2.SW Unit Test Spec' 보유).
        spec_filename: 산출물 warning 추적용 spec 파일명 (결과 dict에 동봉).
        out_warnings: 실패/모호 사유 push (None이면 silent).

    Returns:
        ``_extract_fi_from_sheet`` 결과 + ``spec_filename`` 키, 또는 None.
    """
    if openpyxl is None:
        if out_warnings is not None:
            out_warnings.append(
                "[swutcr] UT201 FI spec 자동 산출 불가 — openpyxl 미설치"
            )
        return None
    if not spec_xlsm_bytes:
        if out_warnings is not None:
            out_warnings.append(
                "[swutcr] UT201 FI spec 자동 산출 불가 — spec bytes 비어있음"
            )
        return None
    if len(spec_xlsm_bytes) > _SPEC_FI_MAX_BYTES:
        if out_warnings is not None:
            out_warnings.append(
                f"[swutcr] UT201 FI spec 자동 산출 불가 — spec 크기 "
                f"{len(spec_xlsm_bytes):,} > 한도 {_SPEC_FI_MAX_BYTES:,} (DoS 방지)"
            )
        return None

    wb = None
    try:
        # 병합 range 전개가 필요해 read_only 불가(ReadOnlyWorksheet는
        # merged_cells 미보장) — read-write 로드. keep_vba=False (읽기 전용 용도).
        wb = openpyxl.load_workbook(
            io.BytesIO(spec_xlsm_bytes), keep_vba=False, data_only=False,
        )
        ws, _name = _find_spec_sheet(wb)
        if ws is None:
            if out_warnings is not None:
                out_warnings.append(
                    "[swutcr] UT201 FI spec 자동 산출 불가 — SwUTS spec 시트 "
                    "('2.SW Unit Test Spec' 류) 미발견"
                )
            return None
        layout = _detect_spec_layout(ws, out_warnings)
        stats = _extract_fi_from_sheet(ws, layout, out_warnings)
        if stats is not None:
            stats["spec_filename"] = spec_filename
        return stats
    except Exception as exc:  # noqa: BLE001 — 산출 실패는 노란 마킹 폴백 (정직 보고)
        if out_warnings is not None:
            out_warnings.append(
                "[swutcr] UT201 FI spec 자동 산출 실패 — "
                f"{type(exc).__name__}: {exc} (노란 마킹 유지)"
            )
        return None
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:  # noqa: BLE001 — close 실패 무해
                pass


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_spec_sheet(wb: Workbook):
    """SwUTS spec 시트('2.SW Unit Test Spec' 류)를 찾아 반환. 미발견 시 None."""
    for name in wb.sheetnames:
        if SPEC_SHEET_RE.search(name):
            return wb[name], name
    return None, None


def _scan_spec_blocks(ws, layout: SpecLayout | None = None) -> list[dict[str, Any]]:
    """spec 시트 anchor 스캔 → 함수 블록 list.

    각 블록: anchor row(B=Index digit, C=TC_ID, D=Unit) + iteration row들.

    TC_ID는 DV ``SwUTC_NNNN`` / PV ``SwUFn_NNNN`` 직접 표기 모두 지원 — 숫자만
    추출(``(\\d+)``)하므로 prefix·대소문자('SwUfn_1361' WIP 오타 포함) 무관하게
    VectorCAST ``SwUFn_NNNN.MMM`` 매칭 키가 나온다.

    iteration index 열은 ``layout.iter_index`` (DV=7/G, PV=8 — PV col 7은
    Generation Method 세그먼트 병합이라 고정 7 사용 시 행 67% 누락, 라운드 105).

    Returns:
        [{anchor, tc_id, unit, num(4-digit), iter_rows: [row,...]}, ...]
    """
    lo = layout or SpecLayout()
    blocks: list[dict[str, Any]] = []
    rr = DATA_START_ROW
    max_row = ws.max_row
    while rr <= max_row:
        b = ws.cell(rr, COL_INDEX).value
        c = ws.cell(rr, COL_TC_ID).value
        if str(b if b is not None else "").strip().isdigit() and c:
            tc_id = str(c).strip()
            num_m = re.search(r"(\d+)", tc_id)
            num = num_m.group(1) if num_m else ""
            # iteration row: anchor 다음부터, 다음 anchor 전까지, iter index 보유.
            iter_rows: list[int] = []
            k = rr + 1
            while k <= max_row:
                nb = ws.cell(k, COL_INDEX).value
                nc = ws.cell(k, COL_TC_ID).value
                if str(nb if nb is not None else "").strip().isdigit() and nc:
                    break  # 다음 anchor
                g = ws.cell(k, lo.iter_index).value
                if g not in (None, ""):
                    iter_rows.append(k)
                k += 1
            blocks.append({
                "anchor": rr,
                "tc_id": tc_id,
                "unit": str(ws.cell(rr, COL_UNIT).value or "").strip(),
                "num": num,
                "iter_rows": iter_rows,
            })
            rr = k
        else:
            rr += 1
    return blocks


def _expected_var_cols(
    ws, anchor_row: int, layout: SpecLayout | None = None,
) -> list[int]:
    """anchor row의 Expected 섹션에서 변수명이 있는 열 리스트.

    범위는 ``layout.expected_start`` ~ ``layout.related_id - 1`` (DV 58~161,
    PV 105~188 — 고정 58 사용 시 PV에서 Input 열을 expected로 오인 + ExpR[57..83]
    누락, 라운드 105). Actual 변수명/값 stamp 시 동일 열 offset 사용 (감사본은
    Actual 변수 = Expected 변수와 동일 — 레퍼런스 실측: anchor r5 BF='return',
    FF='return').
    """
    lo = layout or SpecLayout()
    cols: list[int] = []
    for c in range(lo.expected_start, lo.related_id):
        if ws.cell(anchor_row, c).value not in (None, ""):
            cols.append(c)
    return cols


def _build_fn_iteration_map(session: SwUTSession) -> dict[str, dict[int, Any]]:
    """SwUFn 숫자(4-digit) → {iteration_index(int): {"passed":bool|None,
    "actual": dict, "env": env}}.

    iteration_index 는 ``.MMM`` suffix(1-based). 정렬 보장.
    """
    fn_map: dict[str, dict[int, Any]] = {}
    for env in session.environments:
        for tc_name in env.test_cases:
            # 2026-06-18 fix — case-insensitive. spec TC_ID에 'SwUfn_NNNN'(소문자 fn)
            # WIP 오타가 섞이면(레퍼런스 SUTR도 동일 오타 보유) 기존 case-sensitive
            # match가 탈락 → 해당 함수 iteration 전부 누락 → Total N/A(미실행) 오표기.
            # 레퍼런스는 동일 소문자에도 Pass 렌더되므로 robust 매칭이 정합.
            m = re.match(r"SwUFn_(\d+)\.(\d+)", tc_name, re.IGNORECASE)
            if not m:
                continue
            num = m.group(1).lstrip("0") or "0"
            idx = int(m.group(2))
            exec_r = env.test_results.get(tc_name)
            actual_dict: dict = {}
            passed = None
            if exec_r is not None:
                passed = bool(exec_r.passed)
                actual_dict = getattr(exec_r, "actual_result", {}) or {}
            if not actual_dict:
                tr_items = getattr(env, "tc_result_items", {}).get(tc_name, [])
                tr_item = tr_items[0] if tr_items else None
                if tr_item is not None:
                    actual_dict = getattr(tr_item, "actual_result", {}) or {}
            fn_map.setdefault(num, {})[idx] = {
                "passed": passed,
                "actual": actual_dict,
                "env": env,
                "tc_name": tc_name,
            }
    return fn_map


# ---------------------------------------------------------------------------
# Header writer (Actual / Pass-Fail / Log 섹션)
# ---------------------------------------------------------------------------

def _apply_actual_result_style(ws, layout: SpecLayout | None = None) -> int:
    """라운드 104 — Actual Result(+Pass/Fail·Pass) 열에 템플릿 서식 적용.

    spec graft는 Test Case/Input/Expected(BF=58~FE=161)까지만 서식을 가져오고
    Actual(FF=162~JE=265)·Pass/Fail(JF=266)·Pass(JG=267)는 우리 추가 열이라
    무서식(무테/default 폰트) → 사용자 보고 "Actual만 템플릿 미적용".

    Expected를 1:1 미러: Actual col (162+i) ← Expected col (58+i). **같은 wb 내**이므로
    ``_style`` 인덱스 직접 복사가 정확+빠름 (cross-wb는 라운드 103처럼 객체복사 필요하나
    여기는 동일 wb). value는 보존(``_style`` 만 복제). 서브헤더/헤더 행
    (1~SUBHEADER_ROW)은 건드리지 않음.

    라운드 106 — Pass/Fail·Pass·Log Data 는 Expected 마지막 열 border 복제
    (L:thin R:double)가 DV 감사본 패턴과 불일치 실측되어 **명시 Border 교체**:

    - Pass/Fail: L:double R:thin B:thin, T:thin (첫 데이터행만 T:medium —
      헤더 경계선).
    - Pass(Total): L:thin R:thin T:thin B:thin (함수별 병합 outline 동일 렌더).
    - Log Data: L:thin 만, 첫 데이터행 T:medium (DV 레퍼런스 실측 — 데이터
      영역은 그 외 무테, 값도 비워 둠).

    font 는 기존대로 Expected 마지막 열 미러(맑은 고딕 10 통일), fill 은 보존
    (ASIL/Pass 강조 마킹 유지).

    라운드 105 — 열 좌표는 ``layout`` 동적 값 사용 (DV offset 104 / PV offset 84.
    고정 offset 사용 시 PV에서 미러 소스 앞 47열이 Input 영역 + ExpR[57..83]/SUDS
    서식 오염).

    Returns:
        서식 적용한 셀 수.
    """
    import copy as _copy

    from openpyxl.cell.cell import MergedCell as _MC
    from openpyxl.styles import Border, Font, Side

    def _result_font(src_font, *, bold: bool | None = None) -> Font:
        return Font(
            name="맑은 고딕",
            size=10,
            bold=src_font.bold if bold is None else bold,
            italic=src_font.italic,
            underline=src_font.underline,
            strike=src_font.strike,
            color=_copy.copy(src_font.color),
            vertAlign=src_font.vertAlign,
            charset=src_font.charset,
            family=src_font.family,
            scheme=src_font.scheme,
        )

    lo = layout or SpecLayout()
    offset = lo.actual_start - lo.expected_start  # DV 104 / PV 84
    _thin = Side(style="thin")
    _double = Side(style="double")
    _medium = Side(style="medium")
    restyled = 0
    for r in range(SUBHEADER_ROW + 1, ws.max_row + 1):
        # Actual ← Expected 미러 (_style 통째 — Actual엔 마킹 없음).
        for ec in range(lo.expected_start, lo.actual_start):
            exp = ws.cell(r, ec)
            if not getattr(exp, "has_style", False):
                continue
            act = ws.cell(r, ec + offset)
            if isinstance(act, _MC) or isinstance(exp, _MC):
                continue
            try:
                act._style = _copy.copy(exp._style)
                restyled += 1
            except (AttributeError, TypeError):
                pass
        # Pass/Fail·Pass·Log Data: DV 감사본 패턴 명시 Border (라운드 106 —
        # docstring 참조. Expected 마지막 열 복제는 L:thin R:double 로 불일치).
        # font 는 Expected 마지막 열 미러, fill 보존(ASIL/Pass 강조 마킹 유지).
        ref = ws.cell(r, lo.actual_start - 1)  # DV FE=161 / PV 188
        if getattr(ref, "has_style", False) and not isinstance(ref, _MC):
            first_data = r == DATA_START_ROW
            col_borders = (
                (lo.pass_fail, Border(
                    left=_double, right=_thin,
                    top=(_medium if first_data else _thin), bottom=_thin,
                )),
                (lo.pass_total, Border(
                    left=_thin, right=_thin, top=_thin, bottom=_thin,
                )),
            )
            for tc, tc_border in col_borders:
                dst = ws.cell(r, tc)
                if isinstance(dst, _MC):
                    continue
                try:
                    dst.border = tc_border
                    dst.font = _result_font(ref.font)
                    restyled += 1
                except (AttributeError, TypeError):
                    pass
            log_cell = ws.cell(r, lo.log_data)
            if not isinstance(log_cell, _MC):
                try:
                    log_cell.border = Border(
                        left=_thin, top=(_medium if first_data else None),
                    )
                    log_cell.font = _result_font(ref.font, bold=False)
                    restyled += 1
                except (AttributeError, TypeError):
                    pass
    return restyled


def _write_log_headers(
    ws, out_warnings: list[str] | None, layout: SpecLayout | None = None,
) -> None:
    """레퍼런스 SUTR 레이아웃의 Actual/Pass-Fail/Log 헤더를 추가.

    좌표는 ``layout`` 동적 값 (DV: Actual=FF162~JE265, JF/JG/JH=266/267/268.
    PV: Actual=189~272, 273/274/275 — 라운드 105). 고정 상수 사용 시 PV에서
    r3 'Expected Result'(105) 병합 anchor 덮어쓰기 + ExpR[57..83]/SUDS 서브헤더
    파괴 + 겹침 병합 생성(Excel 복구 경고)이 실측됨.

    - r3: Actual 시작='Actual Result' (Actual 폭 병합), +'Pass/Fail'/'Pass'/'Log Data'.
    - r4: 'Param N' 서브헤더 + Log Data (r3:r4 병합).
    - spec Related ID(r4 'SUDS' 포함) 컬럼 헤더는 Actual로 대체됨 (DV 레퍼런스 규칙).
    - r2: Pass/Fail·Pass COUNTIF 요약 수식 (레퍼런스 복제, 범위는 데이터 끝까지 확장).
    """
    lo = layout or SpecLayout()
    # spec의 Related ID 컬럼은 함수 블록마다 세로 병합 + r3/r4 헤더 병합 보유.
    # Actual을 iteration별로 채우려면 Actual 영역 범위에 걸친 모든 병합을 해제해야
    # 한다 (병합 셀의 비-anchor 셀 쓰기는 openpyxl이 무시 → iteration Actual 값이
    # anchor row로 흘러가 손실됨). Total(pass_total)은 후속 함수별 재병합.
    # 데이터 병합 + r3/r4 헤더 병합 모두 해제. Expected 영역 병합(min_col <
    # actual_start — PV r3 DA3:GF3 등)은 보존된다.
    for rng in list(ws.merged_cells.ranges):
        if lo.actual_start <= rng.min_col <= (lo.pass_fail - 1):
            try:
                ws.unmerge_cells(str(rng))
            except (ValueError, KeyError):
                pass

    safe_write(ws, 1, 1, "Software Unit Test Log")

    # r3 섹션 헤더.
    # 2026-06-18 fix — 레퍼런스 실측: r3 Pass/Fail(col273), Pass열(col274=pass_total)은
    # r3 공란, Log Data(col275). 기존엔 pass_total r3에 'Pass'를 써 레퍼런스(공란)와 불일치.
    safe_write(ws, HEADER_SECTION_ROW, lo.actual_start, "Actual Result")
    safe_write(ws, HEADER_SECTION_ROW, lo.pass_fail, "Pass/Fail")
    safe_write(ws, HEADER_SECTION_ROW, lo.log_data, "Log Data")
    try:
        ws.merge_cells(
            start_row=HEADER_SECTION_ROW, end_row=HEADER_SECTION_ROW,
            start_column=lo.actual_start, end_column=lo.pass_fail - 1,
        )
        ws.merge_cells(
            start_row=HEADER_SECTION_ROW, end_row=SUBHEADER_ROW,
            start_column=lo.log_data, end_column=lo.log_data,
        )
    except (ValueError, AttributeError):
        pass

    # r4 서브헤더 — ActR[0..actual_max-1] (레퍼런스 0-base 인덱싱 스킴).
    # 2026-06-18 fix — 기존 'Param N'(1-base)은 레퍼런스 'ActR[N]'(0-base, Expected의
    # ExpR[N]·Input의 Inpt[N]과 동일 스킴)과 전 84열 불일치였음. JN(pass_total) r4도
    # 'Pass'→'Total'로 교정(레퍼런스 실측 JN4='Total').
    for i in range(lo.actual_max):
        safe_write(ws, SUBHEADER_ROW, lo.actual_start + i, f"ActR[{i}]")
    safe_write(ws, SUBHEADER_ROW, lo.pass_fail, "Unit")
    safe_write(ws, SUBHEADER_ROW, lo.pass_total, "Total")

    try:
        from copy import copy as _copy_style

        from openpyxl.styles import Alignment, Font

        header_font = Font(name="맑은 고딕", size=10, bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        for row_idx in (HEADER_SECTION_ROW, SUBHEADER_ROW):
            for col_idx in range(lo.actual_start, lo.log_data + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.font = _copy_style(header_font)
                cell.alignment = _copy_style(header_alignment)
    except (AttributeError, TypeError):
        pass

    # r2 COUNTIF 요약 수식 (레퍼런스 복제 — 범위는 데이터 끝까지).
    last_row = ws.max_row
    from openpyxl.utils import get_column_letter
    jf = get_column_letter(lo.pass_fail)
    jg = get_column_letter(lo.pass_total)
    try:
        ws.cell(2, lo.pass_fail).value = (
            f'=COUNTIF({jf}{DATA_START_ROW}:{jf}{last_row}, "Fail")'
        )
        ws.cell(2, lo.pass_total).value = (
            f'=COUNTIF({jg}{DATA_START_ROW}:{jg}{last_row},"Fail")'
            f'+COUNTIF({jg}{DATA_START_ROW}:{jg}{last_row},"N/A")'
        )
    except (ValueError, AttributeError):
        if out_warnings is not None:
            out_warnings.append("[spec-sutr] r2 COUNTIF 수식 stamp 실패 (산출물 영향 경미)")


# ---------------------------------------------------------------------------
# Actual / Pass-Fail / Log 채우기
# ---------------------------------------------------------------------------

def _fill_actual_and_result(
    ws,
    blocks: list[dict[str, Any]],
    fn_iter_map: dict[str, dict[int, Any]],
    asil_map: dict[str, str],
    out_warnings: list[str] | None,
    layout: SpecLayout | None = None,
) -> dict[str, Any]:
    """spec 함수 블록에 Actual/Pass-Fail/Log 채움.

    좌표는 ``layout`` 동적 값 (라운드 105) — 고정 상수 사용 시 PV에서 Actual
    stamp가 ExpR[57..83] 실데이터(2,355셀) + SUDS 추적성(1,014셀)을 덮어쓴다.

    Returns: 통계 dict (functions/iterations/matched_fn/unmatched_fn/
        inexact_actual + ``na_tc_list`` — Total N/A 함수 대표 목록, W-4 기재용).
    """
    from openpyxl.utils import get_column_letter  # noqa: F401

    lo = layout or SpecLayout()

    stats: dict[str, Any] = {
        "functions": 0, "iterations": 0,
        "matched_fn": 0, "unmatched_fn": 0,
        "fn_pass": 0, "fn_fail": 0, "fn_na": 0,
        "iter_pass": 0, "iter_fail": 0, "iter_na": 0,
        "actual_from_expected": 0, "actual_from_vcast": 0, "actual_missing": 0,
    }
    unmatched_list: list[str] = []
    # 라운드 96-final W-4 — Total 'N/A' 함수(미매칭 + 매칭됐으나 실행 0) 대표 목록.
    # '■ List of Test Case not Executed' 기재용 — cap으로 메모리 상한 (전체 카운트는
    # fn_na, 잔여는 '외 N건' 요약).
    _NA_LIST_MAX = 20
    na_tc_list: list[str] = []

    for blk in blocks:
        stats["functions"] += 1
        anchor = blk["anchor"]
        num = (blk["num"] or "").lstrip("0") or "0"
        iter_data = fn_iter_map.get(num)

        # Expected 변수 열 (anchor) = Actual 변수 열 offset.
        exp_cols = _expected_var_cols(ws, anchor, lo)
        # Actual 변수명 = Expected 변수명 (감사본 패턴) — anchor에 stamp.
        for off, ec in enumerate(exp_cols):
            if off >= lo.actual_max:
                break
            var_name = ws.cell(anchor, ec).value
            safe_write(ws, anchor, lo.actual_start + off, var_name)

        if iter_data is None:
            stats["unmatched_fn"] += 1
            if len(unmatched_list) < 40:
                unmatched_list.append(f"{blk['tc_id']}({blk['unit']})")
            if len(na_tc_list) < _NA_LIST_MAX:
                na_tc_list.append(f"{blk['tc_id']} ({blk['unit']})")
            # 미매칭 함수 — Pass/Fail N/A 표기 + Total N/A.
            for ir in blk["iter_rows"]:
                safe_write(ws, ir, lo.pass_fail, "N/A")
                stats["iter_na"] += 1
                stats["iterations"] += 1
            safe_write(ws, anchor, lo.pass_total, "N/A")
            stats["fn_na"] += 1
            if blk["iter_rows"]:
                try:
                    ws.merge_cells(
                        start_row=anchor, end_row=blk["iter_rows"][-1],
                        start_column=lo.pass_total, end_column=lo.pass_total,
                    )
                except (ValueError, AttributeError):
                    pass
            _apply_asil_mark(ws, anchor, num, blk, asil_map, lo)
            continue

        stats["matched_fn"] += 1
        any_exec = False
        all_pass = True

        # 2026-06-18 fix — Log Data(JO) 채움. 레퍼런스 실측: 함수 블록당 1회,
        # anchor+1행(iter_rows[0])에 '{env_name}_TestCaseDataReport' 기록
        # (env_name은 VC2025 레이아웃에서 'SwUT_NN_<comp>' 형식 → 레퍼런스 값과 동일).
        # 기존엔 JO 전 행 공란(레퍼런스 622건 미채움). 미매칭 함수(iter_data None)는
        # 보고서 매핑 불가라 공란 유지 (레퍼런스도 미실행 블록 공란).
        if blk["iter_rows"]:
            # reviewer W2 — 다중 env 함수: 실제 실행된(passed 결정됨) iteration의
            # env 우선 선택. 함수가 두 번째 env에서만 실행됐을 때 첫 env명이 잘못
            # 라벨되는 것 방지. 실행 레코드 없으면 임의 env(공란 fallback)로 강등.
            _env_obj = next(
                (r.get("env") for r in iter_data.values()
                 if r.get("env") and r.get("passed") is not None),
                None,
            ) or next(
                (r.get("env") for r in iter_data.values() if r.get("env")), None
            )
            _env_nm = getattr(_env_obj, "env_name", "") if _env_obj else ""
            if _env_nm:
                safe_write(
                    ws, blk["iter_rows"][0], lo.log_data,
                    f"{_env_nm}_TestCaseDataReport",
                )
                stats["log_data_written"] = stats.get("log_data_written", 0) + 1

        for fallback_idx, ir in enumerate(blk["iter_rows"], start=1):
            stats["iterations"] += 1
            raw_iter_idx = ws.cell(ir, lo.iter_index).value
            try:
                it_idx = int(str(raw_iter_idx).strip())
            except (TypeError, ValueError):
                it_idx = fallback_idx
            rec = iter_data.get(it_idx)
            if rec is None:
                safe_write(ws, ir, lo.pass_fail, "N/A")
                stats["iter_na"] += 1
                continue
            passed = rec["passed"]
            if passed is None:
                safe_write(ws, ir, lo.pass_fail, "N/A")
                stats["iter_na"] += 1
                continue
            any_exec = True
            # Actual 값 채우기 — Pass면 Expected 복제 (감사본 패턴), Fail이면 vcast.
            for off, ec in enumerate(exp_cols):
                if off >= lo.actual_max:
                    break
                ac = lo.actual_start + off
                if passed:
                    exp_val = ws.cell(ir, ec).value
                    safe_write(ws, ir, ac, exp_val)
                    stats["actual_from_expected"] += 1
                else:
                    var_name = ws.cell(anchor, ec).value
                    av = _lookup_vcast_actual(rec["actual"], var_name)
                    if av is not None:
                        safe_write(ws, ir, ac, av)
                        stats["actual_from_vcast"] += 1
                    else:
                        _mark_cell(ws, ir, ac)
                        stats["actual_missing"] += 1
            if passed:
                safe_write(ws, ir, lo.pass_fail, "Pass")
                stats["iter_pass"] += 1
            else:
                all_pass = False
                safe_write(ws, ir, lo.pass_fail, "Fail")
                stats["iter_fail"] += 1
        total_str = "Pass" if (any_exec and all_pass) else ("Fail" if any_exec else "N/A")
        if total_str == "Pass":
            stats["fn_pass"] += 1
        elif total_str == "Fail":
            stats["fn_fail"] += 1
        else:
            stats["fn_na"] += 1
            # 매칭됐으나 실행 iteration 0 — 미실행 목록에 등재 (W-4).
            if len(na_tc_list) < _NA_LIST_MAX:
                na_tc_list.append(f"{blk['tc_id']} ({blk['unit']})")
        safe_write(ws, anchor, lo.pass_total, total_str)
        # 라운드 91 fix — 레퍼런스 감사본은 anchor 행 JF(Pass/Fail)에도 함수 결과를
        # 표기(첫 iteration 겸용 양식). anchor JF=total로 정합 (이전: anchor JF 공란).
        safe_write(ws, anchor, lo.pass_fail, total_str)
        # 함수 Total 세로 병합 (anchor ~ 마지막 iteration).
        if blk["iter_rows"]:
            try:
                ws.merge_cells(
                    start_row=anchor, end_row=blk["iter_rows"][-1],
                    start_column=lo.pass_total, end_column=lo.pass_total,
                )
            except (ValueError, AttributeError):
                pass
        _apply_asil_mark(ws, anchor, num, blk, asil_map, lo)

    if unmatched_list and out_warnings is not None:
        out_warnings.append(
            f"[spec-sutr] 함수↔SwUFn 매칭 실패 {stats['unmatched_fn']}건 — "
            f"Pass/Fail N/A 표기. 예: {', '.join(unmatched_list[:15])}"
        )
    stats["na_tc_list"] = na_tc_list
    return stats


def _lookup_vcast_actual(actual_dict: dict, var_name: Any) -> Any:
    """vcast actual_result(dict{var:(actual,exp)})에서 var_name best-effort 조회.

    dotted 변수명(예 ``_LP0DR.Byte``)이 분리될 수 있어 여러 후보 시도.
    매칭 0이면 None.
    """
    if not actual_dict or not var_name:
        return None
    vn = str(var_name).strip()
    cand = [vn]
    if "." in vn:
        cand.extend(vn.split("."))
    cand.append(vn.lstrip("_"))
    for k in cand:
        if k in actual_dict:
            t = actual_dict[k]
            if isinstance(t, tuple) and t:
                return t[0] if t[0] not in (None, "") else None
            return t if t not in (None, "") else None
    return None


def _apply_asil_mark(
    ws, anchor: int, num: str, blk: dict, asil_map: dict[str, str],
    layout: SpecLayout | None = None,
) -> None:
    """anchor의 Total(DV JG=267 / PV 274) 셀에 ASIL 등급 시각 강조."""
    lo = layout or SpecLayout()
    fn_id = f"SwUFn_{num.zfill(4)}" if num else ""
    asil = asil_map.get(fn_id, "") if fn_id else ""
    marker = {
        "A": mark_asil_a_function, "B": mark_asil_b_function,
        "C": mark_asil_c_function, "D": mark_asil_d_function,
        "QM": mark_asil_qm_function,
    }.get((asil or "").strip().upper())
    if marker:
        marker(ws, anchor, lo.pass_total)


def _mark_cell(ws, row: int, col: int) -> None:
    """Actual 부재 셀 노란 마킹 (audit 가시성)."""
    try:
        from openpyxl.styles import PatternFill
        ws.cell(row, col).fill = PatternFill(
            start_color=_FILL_RGB, end_color=_FILL_RGB, fill_type="solid",
        )
    except (ValueError, AttributeError):
        pass


def _collect_coverage_gaps(
    agg: dict[str, Any],
) -> list[tuple[str, str, Any, Any, str]]:
    """agg function_coverage에서 커버리지<100% 함수 추출 (2026-06-18 Fix 4).

    Returns: [(swufn_id, name, stmt_gap|None, branch_gap|None, resolve_kind)]
      — swufn 숫자→name 순. gap = (covered, total, coverage_pct). 둘 다 None인 함수는 제외.
      resolve_kind ∈ {'exact','ci','fallback'} (W2 — 'ci'/'fallback'은 B열 노란마킹).
    같은 함수(name 기준)가 다중 env로 중복 시 더 낮은 커버리지(보수적) 유지.
    """
    # 2026-06-18 Item 1 — 해결 우선순위: SwUTS 스펙(권위) > SwUDS 이름맵 >
    # case-insensitive(casing 차이 흡수) > fc.unit_id > 함수명 원문(fallback).
    spec_map: dict[str, str] = agg.get("spec_name_to_swufn") or {}
    suds_map: dict[str, str] = agg.get("function_name_to_swufn_from_suds") or {}

    # 2026-06-19 deep-review W2 — case-insensitive 인덱스는 모호 키(소문자화 시 서로
    # 다른 SwUFn으로 충돌하는 키, 예 'Calc'→0010 / 'calc'→0021)를 제외한다. 모호 키로
    # CI 폴백하면 둘 중 임의 함수에 오매칭되므로 아예 폴백 금지(→ fallback 노란마킹).
    def _ci_index(m: dict[str, str]) -> dict[str, str]:
        idx: dict[str, str] = {}
        ambiguous: set[str] = set()
        for k, v in m.items():
            lk = k.lower()
            if lk in idx and idx[lk] != v:
                ambiguous.add(lk)
            else:
                idx[lk] = v
        for lk in ambiguous:
            idx.pop(lk, None)
        return idx

    spec_ci = _ci_index(spec_map)
    suds_ci = _ci_index(suds_map)

    def _resolve_swufn(name: str, unit_id: str) -> tuple[str, str]:
        """(swufn, kind) — kind ∈ {'exact','ci','fallback'}.

        'ci'(case-insensitive 폴백)는 C 대소문자 충돌쌍(Foo vs foo)을 다른 함수에
        오매칭할 수 있어, 호출자가 valid-looking SwUFn이어도 노란마킹하도록 구분한다.
        """
        nl = name.lower()
        exact = spec_map.get(name) or suds_map.get(name)
        if exact:
            return (exact, "exact")
        ci = spec_ci.get(nl) or suds_ci.get(nl)
        if ci:
            return (ci, "ci")
        return (unit_id or name, "fallback")

    # 2026-06-19 deep-review W3 — acc 키를 함수 식별자(name)로 둔다. 이전엔 resolve된
    # swufn을 키로 써서, 서로 다른 두 함수가 같은 SwUFn으로 resolve되면(CI 오매칭 또는
    # spec num 중복) 1행으로 붕괴해 한 함수의 커버리지 갭이 산출물에서 누락됐다
    # (안전 산출물에서 갭 under-reporting은 위험). swufn은 표시값으로만 보관. 동일
    # 함수의 다중 env 병합(보수적 최소 pct)은 name 키로 그대로 유지.
    acc: dict[str, dict[str, Any]] = {}
    for fc in agg.get("function_rows", []) or []:
        name = getattr(fc, "name", "") or ""
        swufn, kind = _resolve_swufn(name, getattr(fc, "unit_id", "") or name)
        s = getattr(fc, "statement", None)
        b = getattr(fc, "branch", None)
        s_gap = (
            (s.covered, s.total, s.coverage_pct)
            if s and s.total > 0 and s.covered < s.total else None
        )
        b_gap = (
            (b.covered, b.total, b.coverage_pct)
            if b and b.total > 0 and b.covered < b.total else None
        )
        if not s_gap and not b_gap:
            continue
        ident = name or swufn
        cur = acc.get(ident)
        if cur is None:
            acc[ident] = {
                "swufn": swufn, "name": name, "kind": kind,
                "stmt": s_gap, "branch": b_gap,
            }
        else:
            if s_gap and (cur["stmt"] is None or s_gap[2] < cur["stmt"][2]):
                cur["stmt"] = s_gap
            if b_gap and (cur["branch"] is None or b_gap[2] < cur["branch"][2]):
                cur["branch"] = b_gap

    def _num(sw: str) -> int:
        m = re.search(r"(\d+)", sw or "")
        return int(m.group(1)) if m else 0

    return [
        (info["swufn"], info["name"], info["stmt"], info["branch"], info["kind"])
        for info in sorted(
            acc.values(), key=lambda v: (_num(v["swufn"]), v["name"])
        )
    ]


def _compress_line_numbers(nums: list[int]) -> str:
    """미커버 line 번호 리스트 → 압축 문자열 (2026-06-18 Item 2).

    연속 구간은 ``a~b`` 로, 단독은 그대로. 레퍼런스 G 스타일 정합.
    예: [202] → '202', [152..169] → '152~169', [1,2,5,6,9] → '1~2, 5~6, 9'.
    """
    s = sorted({int(n) for n in nums if n is not None})
    if not s:
        return ""
    runs: list[tuple[int, int]] = []
    start = prev = s[0]
    for n in s[1:]:
        if n == prev + 1:
            prev = n
        else:
            runs.append((start, prev))
            start = prev = n
    runs.append((start, prev))
    return ", ".join(f"{a}~{b}" if a != b else f"{a}" for a, b in runs)


# Excel(.xlsx) 불법 제어문자 — C 소스/주석에 form-feed(\x0c)·\x07·\x1a 등이 섞이면
# openpyxl이 셀 .value 대입 시 IllegalCharacterError를 raise. \t\n\r은 합법이라 보존.
# (deep-review C1 — safe_write에도 동일 방어 있으나 source 단에서도 정제해 이중 안전.)
_EXCEL_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _format_gap_source(pairs: list[tuple[int, str]], max_len: int = 240) -> str:
    """미커버 (line, source) → G(Description) 셀 텍스트 (2026-06-18 Item 2).

    중복 소스 제거 후 줄바꿈 연결 (레퍼런스 G 스타일). 셀 가독성 위해 길이 제한.
    Excel 불법 제어문자는 제거(2026-06-19 deep-review C1 — IllegalCharacterError 방지).
    """
    seen: list[str] = []
    for _ln, src in pairs:
        t = _EXCEL_ILLEGAL_CHARS_RE.sub("", src or "").strip()
        if t and t not in seen:
            seen.append(t)
    out = "\n".join(seen)
    return out[: max_len - 1] + "…" if len(out) > max_len else out


def _write_spec_deviation(
    ws, agg: dict[str, Any], out_warnings: list[str],
) -> int:
    """spec-based SUTR 2.Deviation — 레퍼런스 감사본 스키마 (2026-06-18 Fix 4).

    표준 ``_write_deviation``(Test Case ID/Issue/Status 스키마)와 분리된 신규 경로.
    표준 SUTR(HDPDM01/SwIT)은 기존 writer 유지 — 본 함수는 spec-based만 사용.

    레퍼런스 실측 스키마 (KJPDS02 PV SwUTR 260615 2.Deviation):
      r1: B='Deviation Report'
      r3: B='■ Deviation List'
      r4: B='Unit'(B:C) | D='Coverage'(D:E) | F='미달성 사유'(F:G)
      r5: B='ID' C='Name' D='Type' E='Value' F='line' G='Description'
      r6+: 커버리지<100% 함수 — B=SwUFn C=name D=Statements/Branches E='N / M (X%)'
           F=미달 line G=소스 발췌+사유

    자동/수기 경계 (2026-06-18 Item 2 갱신): B/C/D/E는 커버리지 집계로 자동 산출.
    F(미달 line)·G(소스 발췌)는 AggregateCoverageReport annotated source('Code
    Coverage for <unit>')에서 추출한 ``agg["function_gap_lines"]``로 **자동 채움**.
    annotated source 미가용/함수명 미스매치 함수만 노란 placeholder 유지(fail-safe).

    Returns: 쓰여진 gap 행 수.
    """
    gaps = _collect_coverage_gaps(agg)
    # 2026-06-18 Item 2 — 함수명→{statements/branches:[(line,src)]} 미커버 line 맵.
    # extract_uncovered_lines가 채움. case-insensitive 폴백 인덱스 동반(metrics 표
    # 함수명과 annotated source 함수명 casing 차이 흡수).
    gap_lines_map: dict[str, dict] = agg.get("function_gap_lines") or {}
    gap_lines_ci = {str(k).lower(): v for k, v in gap_lines_map.items()}
    auto_fg = 0
    placeholder_fg = 0
    nonexact_fn = 0  # W2 — exact가 아닌(ci/fallback) SwUFn 해결 함수 수(노란마킹)

    # 1) 기존 템플릿(표준 Test Case ID 스키마) 영역 clear — 병합 해제 후 값 제거.
    # reviewer W3 — ws.max_row가 merged-only 행을 누락할 수 있어(openpyxl 특성)
    # 하한 100으로 상향. 표준 deviation 템플릿(72행 + Appendix)의 잔여 행까지 확실히
    # clear (잔존 시 신규 스키마와 혼재). 신규 데이터는 항상 100행 미만이라 안전.
    clear_rows = max(ws.max_row or 0, 100)
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= clear_rows and rng.min_col <= 8:
            try:
                ws.unmerge_cells(str(rng))
            except (ValueError, KeyError):
                pass
    for r in range(1, clear_rows + 1):
        for c in range(1, 9):
            try:
                ws.cell(r, c).value = None
            except (ValueError, AttributeError):
                pass

    # 2) 헤더 (레퍼런스 스키마).
    safe_write(ws, 1, 2, "Deviation Report")
    safe_write(ws, 3, 2, "■ Deviation List")
    safe_write(ws, 4, 2, "Unit")
    safe_write(ws, 4, 4, "Coverage")
    safe_write(ws, 4, 6, "미달성 사유")
    safe_write(ws, 5, 2, "ID")
    safe_write(ws, 5, 3, "Name")
    safe_write(ws, 5, 4, "Type")
    safe_write(ws, 5, 5, "Value")
    safe_write(ws, 5, 6, "line")
    safe_write(ws, 5, 7, "Description")
    for a, b in (("B4", "C4"), ("D4", "E4"), ("F4", "G4")):
        try:
            ws.merge_cells(f"{a}:{b}")
        except (ValueError, AttributeError):
            pass
    try:
        from copy import copy as _copy
        from openpyxl.styles import Alignment, Font
        hf = Font(name="맑은 고딕", size=10, bold=True)
        ha = Alignment(horizontal="center", vertical="center")
        for rr in (4, 5):
            for cc in range(2, 8):
                cell = ws.cell(rr, cc)
                cell.font = _copy(hf)
                cell.alignment = _copy(ha)
    except (AttributeError, TypeError):
        pass

    # 3) gap 행 — B/C/D/E 자동, F/G annotated source 자동(미가용 시 placeholder).
    r = 6
    written = 0
    if not gaps:
        safe_write(ws, r, 2, "해당 사항 없음")
        r += 1
    for swufn, name, s_gap, b_gap, kind in gaps:
        first = True
        gl = gap_lines_map.get(name) or gap_lines_ci.get(str(name).lower()) or {}
        for typ, gap, bucket in (
            ("Statements", s_gap, "statements"),
            ("Branches", b_gap, "branches"),
        ):
            if gap is None:
                continue
            covered, total, pct = gap
            if first:
                safe_write(ws, r, 2, swufn)
                safe_write(ws, r, 3, name)
                # W1 — SwUFn 미해결(fallback) 시 B에 함수명 원문. 'SwUFn_' 형식이
                # 아니면 노란 마킹. 2026-06-19 deep-review W2 — kind=='ci'(대소문자
                # 폴백)는 valid-looking 'SwUFn_NNNN'이어도 다른 함수 오매칭 가능성이
                # 있어 노란 마킹(audit 수기 검증 유도). 'exact'만 무마킹.
                if (kind != "exact"
                        or not re.match(r"SwUFn_\d", str(swufn), re.IGNORECASE)):
                    _mark_cell(ws, r, 2)
                    nonexact_fn += 1
                first = False
            safe_write(ws, r, 4, typ)
            safe_write(ws, r, 5, f"{covered} / {total} ({round(pct * 100)}%)")
            # F(미달 line)·G(소스 발췌) — annotated source 자동 산출. 해당 type의
            # gap line이 있으면 자동 채움, 없으면(파싱 미스/미가용) 노란 placeholder.
            pairs = (gl.get(bucket) if isinstance(gl, dict) else None) or []
            if pairs:
                safe_write(ws, r, 6, _compress_line_numbers([p[0] for p in pairs]))
                g_txt = _format_gap_source(pairs)
                if g_txt:
                    safe_write(ws, r, 7, g_txt)
                else:
                    safe_write(ws, r, 7, USER_INPUT_PLACEHOLDER)
                    _mark_cell(ws, r, 7)
                auto_fg += 1
            else:
                safe_write(ws, r, 6, USER_INPUT_PLACEHOLDER)
                _mark_cell(ws, r, 6)
                safe_write(ws, r, 7, USER_INPUT_PLACEHOLDER)
                _mark_cell(ws, r, 7)
                placeholder_fg += 1
            r += 1
            written += 1

    # 4) Appendix sentinel + End (레퍼런스 정합).
    r += 1
    safe_write(ws, r, 2, "■ Appendix - 발생 가능 값")
    r += 1
    safe_write(ws, r, 2, "Related Test Case ID")
    safe_write(ws, r, 3, "Parameter")
    safe_write(ws, r, 5, "Value")
    safe_write(ws, r, 7, "Note")
    r += 1
    safe_write(ws, r, 2, "해당 사항 없음")
    r += 2
    safe_write(ws, r, 2, "■ Appendix - 첨부자료")
    r += 2
    safe_write(ws, r, 2, "< End of Document >")

    if out_warnings is not None and written:
        out_warnings.append(
            f"[spec-sutr] 2.Deviation 커버리지 미달 {written}행 기재 — "
            f"F(line)/G(소스) 자동 {auto_fg}행 (annotated source), "
            f"placeholder {placeholder_fg}행 (annotated source 미가용)"
        )
        if nonexact_fn:
            out_warnings.append(
                f"[spec-sutr] 2.Deviation SwUFn 해결 비-exact {nonexact_fn}함수 "
                f"(B열 노란마킹) — 스펙/SwUDS 미등재 또는 case-insensitive 폴백. "
                f"audit 수기 검증 필요."
            )
        # 2026-06-19 deep-review W1 — E(커버리지 값)는 _collect_coverage_gaps의
        # metric별 최저 coverage env, F/G(line/소스)는 aggregate_session의 gap-line
        # 최다 env 기준이라, 동일 함수가 다중 env에 등장하면 E와 F/G가 서로 다른 env
        # 스냅샷일 수 있다(라인 수와 미달 카운트 불일치 가능). 단일 env 함수(현 KJPDS02
        # PV 실측: env당 함수 unique)에서는 미발현이나, 다중 env 일반 경로에서 잠재 →
        # 정직 disclosure.
        out_warnings.append(
            "[spec-sutr] 2.Deviation 주의: 다중 env에 동일 함수가 등장할 경우 "
            "E(커버리지 값)와 F/G(미달 line/소스)가 서로 다른 env 스냅샷일 수 있음 "
            "(라인 수↔미달 카운트 불일치 가능) — 다중 env 함수는 audit 수기 확인 권장."
        )
    return written


def _write_cover_meta_legacy(
    wb: Workbook, meta: SutrBuildMeta, out_warnings: list[str] | None,
) -> None:
    """라운드 91 호환 — spec wb 베이스 Cover 시트 best-effort label stamp.

    표준 SUTR 템플릿 미제공(template_xlsm_bytes=None) fallback 경로 전용.
    """
    cover = next((wb[n] for n in wb.sheetnames if n.lower() == "cover"), None)
    if cover is None:
        if out_warnings is not None:
            out_warnings.append("[spec-sutr] Cover 시트 미발견 — meta stamp skip")
        return
    write_value_after_label(cover, "Project", meta.project_full_name)
    write_value_after_label(cover, "ASIL Level", meta.asil_level)
    write_value_after_label(cover, "Version", f"v{meta.release_sw_version}")
    write_value_after_label(cover, "Test Date", meta.test_date)
    if meta.author:
        write_value_after_label(cover, "Author", meta.author)
    if meta.approver:
        write_value_after_label(cover, "Approver", meta.approver)


def _fill_standard_aux_sheets(
    wb: Workbook,
    meta: SutrBuildMeta,
    session: SwUTSession,
    agg: dict[str, Any],
    summary: dict[str, Any],
    deviation_cases: list[Any] | None,
    out_warnings: list[str],
) -> None:
    """표준 SUTR 템플릿의 Cover / 1.Test Summary / 2.Deviation / History 채움 (R92).

    표준 ``swut_sutr_aggregator`` writer 재사용 — `build_sutr` (표준 양식) 와 동일
    로직으로 meta + TC 카운트 + Requirements coverage 를 stamp. 3.Test Log 는 spec
    시트 이식 (별도 처리) 이므로 여기서 건드리지 않는다.
    """
    from backend.services.excel_template_utils import build_release_history_row
    from backend.services.swut_coverage_aggregator import _write_history_sheet
    from backend.services.swut_sutr_aggregator import (
        _write_cover,
        _write_test_summary,
    )

    sheet_names = wb.sheetnames

    # layout=None — 표준 v3.01 양식의 라벨 기반 find_kv_row 경로 사용. _write_test_summary
    # 의 TC stats(layout 의존) 분기는 layout None 시 skip되고, 함수 단위 카운트는
    # _fill_test_summary_counts 가 별도로 stamp (레퍼런스 감사본 정합).

    cover_ws = next((wb[n] for n in sheet_names if n.lower() == "cover"), None)
    if cover_ws is None:
        out_warnings.append("[spec-sutr] Cover 시트 미발견")
    else:
        _write_cover(cover_ws, meta, out_warnings=out_warnings)

    ts_ws = next((wb[n] for n in sheet_names if "test summary" in n.lower()), None)
    if ts_ws is None:
        out_warnings.append("[spec-sutr] 1.Test Summary 시트 미발견")
    else:
        _write_test_summary(
            ts_ws, meta, agg, out_warnings=out_warnings,
            layout=None, summary=summary,
        )

    dev_ws = next((wb[n] for n in sheet_names if "deviation" in n.lower()), None)
    if dev_ws is None:
        out_warnings.append("[spec-sutr] 2.Deviation 시트 미발견")
    else:
        # 2026-06-18 Fix 4 — spec-based 전용 deviation writer. 레퍼런스 감사본
        # 2.Deviation은 'Test Case ID/Issue/Status'(표준 _write_deviation)가 아니라
        # 커버리지 미달 함수 목록(Unit/Coverage/미달성 사유) 스키마. 표준 writer는
        # 표준 SUTR(HDPDM01/SwIT)에서 그대로 사용되므로 spec 경로만 신규 writer로
        # 분기 (회귀 없음). deviation_cases(수기 입력 deviation)는 spec 양식엔
        # Appendix '발생 가능 값'으로 분리되므로 본 목록과 무관.
        n = _write_spec_deviation(dev_ws, agg, out_warnings=out_warnings)
        summary["deviation_cases_written"] = n

    hist_ws = next((wb[n] for n in sheet_names if n.lower() == "history"), None)
    if hist_ws is not None:
        release_rows = build_release_history_row(
            meta, doc_kind="SwUT SUTR (spec-based)", out_warnings=out_warnings,
        )
        n_h = _write_history_sheet(hist_ws, release_rows, out_warnings=out_warnings)
        summary["history_rows_written"] = n_h


def _fill_test_summary_counts(
    wb: Workbook, summary: dict[str, Any], fill_stats: dict[str, int],
    out_warnings: list[str],
) -> None:
    """1.Test Summary TC 카운트를 spec 매칭 결과로 보정 (R92).

    표준 `_write_test_summary` 는 session aggregate (VectorCAST TC 단위) 기준 카운트를
    stamp 한다. 레퍼런스 감사본 1.Test Summary 의 카운트는 **함수 단위** (Total Number
    of TCs = 함수 570) 이므로, spec 함수 블록 매칭 결과로 r17/r18 영역을 덮어쓴다.

    레퍼런스 실측: Total=570 / Tested=569 / Passed=569 / Failed=0 / not-exec=1.
    = spec_function_blocks / matched_fn / iter pass·fail 기반 산출.
    """
    ts_ws = next((wb[n] for n in wb.sheetnames if "test summary" in n.lower()), None)
    if ts_ws is None:
        return
    from backend.services.excel_template_utils import find_kv_row

    total = fill_stats.get("functions", 0)
    # 함수 단위 카운트 (레퍼런스 감사본 1.Test Summary 기준):
    #   not_exec = Total 'N/A' 함수 (미매칭 + 매칭됐으나 실행 iteration 0).
    #   passed = Total 'Pass', failed = Total 'Fail', tested = passed + failed.
    not_exec = fill_stats.get("fn_na", 0)
    passed = fill_stats.get("fn_pass", 0)
    failed = fill_stats.get("fn_fail", 0)
    tested = passed + failed

    pos = find_kv_row(ts_ws, "Total Number of TCs", max_row=30)
    if pos is None:
        out_warnings.append(
            "[spec-sutr] 1.Test Summary 'Total Number of TCs' 헤더 미발견 — TC 카운트 stamp skip"
        )
        return
    data_row = pos[0] + 1
    col = pos[1]
    safe_write(ts_ws, data_row, col, total)
    safe_write(ts_ws, data_row, col + 1, tested)
    safe_write(ts_ws, data_row, col + 2, passed)
    safe_write(ts_ws, data_row, col + 3, failed)
    safe_write(ts_ws, data_row, col + 4, not_exec)
    summary["test_summary_tc_total"] = total
    summary["test_summary_tested"] = tested
    summary["test_summary_passed"] = passed
    summary["test_summary_failed"] = failed
    summary["test_summary_not_executed"] = not_exec

    # Requirements/Design Coverage (SWUDS row) — 함수 수 기반.
    req_pos = find_kv_row(ts_ws, "System Design", max_row=30)
    if req_pos is None:
        req_pos = find_kv_row(ts_ws, "SWUDS", max_row=30)
    if req_pos is not None:
        rr, rc = req_pos
        safe_write(ts_ws, rr, rc + 1, total)     # can be tested
        safe_write(ts_ws, rr, rc + 2, tested)    # tested
        safe_write(ts_ws, rr, rc + 3, not_exec)  # not tested
        summary["requirements_swuds_total"] = total

        # 라운드 96-final C-4 — coverage 수식 통일 (모순 차단).
        # 템플릿 잔존 '=IFERROR(D22/C22,"")'는 N/A(not tested)를 미가산(92.3%),
        # 'Actual Coverage'(C10)는 리터럴 1(100%)이라 열람 시 두 수치가 모순.
        # DV 감사본 형식 '=IFERROR((D22+E22)/C22,"")'(tested+N/A 가산)로 교체하고
        # C10은 리터럴 대신 이 셀을 참조하는 수식으로 재기입 — 항상 동일 산출.
        # 좌표는 req_pos 기준 동적 산출 (데이터 양에 따라 행이 이동해도 정합).
        from openpyxl.utils import get_column_letter
        col_total = get_column_letter(rc + 1)    # 예: C (can be tested)
        col_tested = get_column_letter(rc + 2)   # 예: D (tested)
        col_na = get_column_letter(rc + 3)       # 예: E (not tested)
        cov_col = rc + 4                         # 예: F (coverage 수식)
        cov_letter = get_column_letter(cov_col)

        def _dv_cov_formula(r: int) -> str:
            return (
                f'=IFERROR(({col_tested}{r}+{col_na}{r})/{col_total}{r},"")'
            )

        safe_write(ts_ws, rr, cov_col, _dv_cov_formula(rr))
        # 동일 블록의 다른 requirement row에 남은 템플릿 수식(Dn/Cn 형식)도 통일.
        _tmpl_cov_re = re.compile(
            rf"^=\s*IFERROR\(\s*{col_tested}(\d+)\s*/\s*{col_total}\1",
            re.IGNORECASE,
        )
        for r2 in range(max(1, rr - 3), rr + 4):
            if r2 == rr:
                continue
            v2 = ts_ws.cell(r2, cov_col).value
            if isinstance(v2, str) and _tmpl_cov_re.match(v2):
                safe_write(ts_ws, r2, cov_col, _dv_cov_formula(r2))
        summary["requirements_coverage_formula"] = _dv_cov_formula(rr)

        # 'Actual Coverage' — _write_test_summary가 stamp한 리터럴(agg 기반)을
        # 위 coverage 수식 셀 참조로 재기입 (두 표시 단일 진리원).
        cov_label_pos = find_kv_row(ts_ws, "Actual Coverage", max_row=30)
        if cov_label_pos is not None:
            lr, lc = cov_label_pos
            target_col = lc + 1
            for mr in ts_ws.merged_cells.ranges:
                if mr.min_row <= lr <= mr.max_row and mr.min_col <= lc <= mr.max_col:
                    target_col = mr.max_col + 1
                    break
            safe_write(ts_ws, lr, target_col, f"={cov_letter}{rr}")
            summary["actual_coverage_formula"] = f"={cov_letter}{rr}"


def _write_not_executed_list(
    wb: Workbook, fill_stats: dict[str, Any], summary: dict[str, Any],
    out_warnings: list[str],
) -> int:
    """'■ List of Test Case not Executed' 목록 stamp (라운드 96-final W-4).

    검증 발견: 미실행 카운트(예: F18=44)는 stamp되는데 목록 영역은 공백 —
    카운트·목록 불일치로 audit reviewer가 미실행 TC를 문서 단독으로 식별 불가.
    가용 빈 행 한도 내 대표 항목 + 잔여 '외 N건 — spec 등재, 실행 로그 미발견'
    요약 기재 (전체 목록은 session 데이터/AuditLog). spec 버전 문구는 소스가
    바뀔 수 있어 중립 표기 (라운드 106 리뷰 minor).

    헤더/가용 행은 동적 탐색 — 좌표 하드코딩 금지 (데이터 양에 따라 이동).

    Returns: 기재한 행 수 (요약 행 포함).
    """
    ts_ws = next((wb[n] for n in wb.sheetnames if "test summary" in n.lower()), None)
    if ts_ws is None:
        return 0
    total_na = int(fill_stats.get("fn_na", 0) or 0)
    if total_na <= 0:
        return 0
    na_list: list[str] = list(fill_stats.get("na_tc_list") or [])

    # 헤더 탐색 — '■ List of Test Case not Executed' (v1.01 변형 흡수 substring).
    header_pos: tuple[int, int] | None = None
    for row in ts_ws.iter_rows(min_row=1, max_row=min(ts_ws.max_row, 80)):
        for cell in row:
            v = cell.value
            if (isinstance(v, str) and "list" in v.lower()
                    and "not executed" in v.lower()):
                header_pos = (cell.row, cell.column)
                break
        if header_pos is not None:
            break
    if header_pos is None:
        out_warnings.append(
            f"[spec-sutr] 미실행 TC {total_na}건 — '■ List of Test Case not "
            "Executed' 헤더 미발견으로 목록 기재 skip (양식 확인)"
        )
        return 0
    hr, hc = header_pos

    # 가용 빈 행 스캔 — 다음 섹션/비어있지 않은 행 전까지, 최대 8행.
    avail: list[int] = []
    for r in range(hr + 1, min(hr + 9, ts_ws.max_row + 1)):
        row_vals = [
            ts_ws.cell(r, c).value for c in range(max(1, hc - 1), hc + 8)
        ]
        if any(v not in (None, "") for v in row_vals):
            # 라운드 97 재검증 fix — 섹션 컬럼 sub-header 행('Test Case ID' /
            # 'Rationale why ...')은 목록 영역의 일부다. KJPDS02 v1.01 실측:
            # 헤더(R26) 바로 아래 R27이 sub-header라 여기서 break하면 가용 행
            # 0 → 미실행 목록이 영영 미기재 (W-4 fix 무효화). sub-header는
            # skip하고 그 아래 빈 행을 가용 행으로 쓴다. 다음 섹션 헤더
            # ('■ ...' 등 일반 비어있지 않은 행)는 기존대로 break.
            if any(
                isinstance(v, str) and v.strip().lower() == "test case id"
                for v in row_vals
            ):
                continue
            break
        avail.append(r)
    if not avail:
        out_warnings.append(
            f"[spec-sutr] 미실행 TC {total_na}건 — 목록 가용 행 0 (양식 확인)"
        )
        return 0

    # 대표 항목 + '외 N건' 요약 — 전부 들어가면 요약 행 생략.
    if total_na <= len(avail) and len(na_list) >= total_na:
        n_show = total_na
    else:
        n_show = min(len(na_list), len(avail) - 1)
    written = 0
    for i in range(n_show):
        safe_write(ts_ws, avail[i], hc, na_list[i])
        written += 1
    rest = total_na - n_show
    if rest > 0:
        safe_write(
            ts_ws, avail[written], hc,
            f"외 {rest}건 — spec 등재, 실행 로그 미발견",
        )
        written += 1
    summary["not_executed_list_rows"] = written
    summary["not_executed_total"] = total_na
    return written


def _apply_phase_platform_meta(
    wb: Workbook, meta: SutrBuildMeta, summary: dict[str, Any],
    out_warnings: list[str],
) -> None:
    """1.Test Summary 프로젝트/버전 항목 phase 보정 (라운드 96-final W-11).

    KJPDS02 v1.01 1.Test Summary 상단 kv (좌표는 라벨 동적 탐색):
      - Project Name(C4): ``{project_id}_{phase}`` (예: KJPDS02_PV) — phase는
        config ``swutcr_metadata.phase``. 키 부재 시 기존 동작 유지 (full name).
      - SW Version(C5): config ``swutcr_metadata.software_platform_ver``(예: 25A1)
        우선 — 부재 시 기존 동작.
      - HW Version(C6): SwUT는 HW 비대상 — cfg ``hw_version`` 부재 시 'N/A' 기본.
        v3.01 양식 라벨('Test Target Version(HW)')은 건드리지 않음 (라벨 미발견
        silent — _write_test_summary의 meta.hw_version stamp 유지).

    cfg 접근은 기존 config 헬퍼(``load_meta_from_config`` — mtime 캐시) 경유,
    빌더 시그니처 불변. config 미존재/키 부재 시 모든 항목 기존 동작.
    """
    ts_ws = next((wb[n] for n in wb.sheetnames if "test summary" in n.lower()), None)
    if ts_ws is None:
        return
    from backend.services.swut_meta_resolver import load_meta_from_config
    md = (load_meta_from_config(meta.project_id) or {}).get(
        "swutcr_metadata", {}) or {}
    if not isinstance(md, dict):
        md = {}

    phase = str(md.get("phase", "") or "").strip()
    if phase and meta.project_id:
        project_phase = f"{meta.project_id}_{phase}"
        if write_value_after_label(ts_ws, "Project Name", project_phase):
            summary["test_summary_project"] = project_phase

    platform_ver = str(md.get("software_platform_ver", "") or "").strip()
    if platform_ver:
        ok = write_value_after_label(ts_ws, "SW Version", platform_ver)
        if not ok:
            ok = write_value_after_label(ts_ws, "Release Name(SW)", platform_ver)
        if ok:
            summary["test_summary_sw_version"] = platform_ver

    hw = str(md.get("hw_version", "") or "").strip() or "N/A"
    if write_value_after_label(ts_ws, "HW Version", hw):
        summary["test_summary_hw_version"] = hw


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_sutr_from_spec(
    session: SwUTSession,
    meta: SutrBuildMeta,
    spec_xlsm_bytes: bytes,
    *,
    template_xlsm_bytes: bytes | None = None,
    function_asil_map: dict[str, str] | None = None,
    deviation_cases: list[Any] | None = None,
) -> SutrBuildResult:
    """SUTR 생성 — 표준 SUTR 템플릿 베이스 + SwUTS spec '3.Test Log' 이식 (라운드 92).

    라운드 91은 spec wb 자체를 베이스로 써 출력 시트 구성이 틀렸다 (Introduction /
    1.Test Environment 포함, 1.Test Summary / 2.Deviation 누락). 라운드 92는 표준 SUTR
    템플릿을 베이스로 하고, spec '2.SW Unit Test Spec' (와이드 268열) 시트만 '3.Test
    Log'로 이식하여 레퍼런스 시트 구성 [Cover / History / 1.Test Summary / 2.Deviation
    / 3.Test Log]을 맞춘다.

    Args:
        session: collect_swut_session 출력 (VectorCAST Actual/Pass-Fail source).
        meta: 빌드 메타 (Cover / Test Summary / History stamp).
        spec_xlsm_bytes: SwUTS spec xlsm bytes ('2.SW Unit Test Spec' — Input/Expected/
            TC 보존, '3.Test Log'로 이식).
        template_xlsm_bytes: 표준 SUTR 템플릿 xlsm bytes (베이스 — Cover/History/
            1.Test Summary/2.Deviation/3.Test Result 보유). None이면 backward-compat
            으로 spec wb 베이스 (라운드 91 동작) — audit 비권장, warning 누적.
        function_asil_map: SwUFn_NNNN → ASIL 등급 (옵션, anchor 시각 강조).
        deviation_cases: 2.Deviation 시트 stamp용 (옵션).

    Returns:
        SutrBuildResult — xlsm_io에 표준 템플릿 + spec 이식 '3.Test Log' + Actual/
        Pass-Fail/Log + Cover/Test Summary/Deviation/History stamp.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for spec-based SUTR builder")

    validate_build_meta(
        meta.release_sw_version, meta.test_date,
        doc_id_sequence=meta.doc_id_sequence,
        test_engineer=meta.test_engineer,
        author=meta.default_author,
        approver=meta.default_approver or meta.approver_override,
        reviewer=meta.default_reviewer or meta.reviewer_override,
    )
    validate_xlsx_template_bytes(spec_xlsm_bytes, label="SwUTS spec xlsm")

    spec_sha256_12 = hashlib.sha256(spec_xlsm_bytes).hexdigest()[:12]
    warnings: list[str] = extract_warnings_from_session(session)
    asil_map = function_asil_map or {}

    # VBA 검사 — spec + (있으면) 표준 템플릿 둘 다 keep_vba.
    spec_has_vba = has_vba_macros(spec_xlsm_bytes)
    template_has_vba = spec_has_vba
    if template_xlsm_bytes is not None:
        validate_xlsx_template_bytes(template_xlsm_bytes, label="표준 SUTR 템플릿 xlsm")
        template_has_vba = has_vba_macros(template_xlsm_bytes)
    if template_has_vba or spec_has_vba:
        warnings.append(
            "VBA macro execution NOT verified — open output xlsm in Excel and verify "
            "macros before submitting as evidence"
        )
        refs = inspect_vba_refs(template_xlsm_bytes or spec_xlsm_bytes)
        if refs:
            warnings.append(
                f"VBA stale ref 위험 패턴 — {refs} (셀/시트 이동 시 매크로 깨질 위험)"
            )

    # spec wb 로드 (Test Log source).
    spec_wb: Workbook = openpyxl.load_workbook(
        io.BytesIO(spec_xlsm_bytes), keep_vba=True, data_only=False,
    )
    spec_ws, spec_name = _find_spec_sheet(spec_wb)
    if spec_ws is None:
        warnings.append("[spec-sutr] SwUTS spec 시트 미발견 — 빌드 불가")
        spec_wb.close()
        return SutrBuildResult(ok=False, warnings=warnings)

    agg = aggregate_session(session)

    if template_xlsm_bytes is not None:
        # 라운드 92 — 표준 SUTR 템플릿을 베이스로 로드.
        wb: Workbook = openpyxl.load_workbook(
            io.BytesIO(template_xlsm_bytes), keep_vba=True, data_only=False,
        )
        # 표준의 좁은 '3.Test Result' 시트 제거.
        for nm in list(wb.sheetnames):
            low = nm.lower()
            if "test result" in low or low == LOG_SHEET_NAME.lower():
                del wb[nm]
        # spec 와이드 시트를 '3.Test Log'로 풀 카피 이식 (끝에 추가).
        log_ws = copy_sheet_across_workbooks(
            spec_ws, wb, new_title=LOG_SHEET_NAME, insert_index=None,
        )
    else:
        # backward-compat (라운드 91) — spec wb 자체 베이스. audit 비권장.
        warnings.append(
            "[spec-sutr] 표준 SUTR 템플릿 미제공 — spec wb 베이스 fallback (라운드 91 호환). "
            "시트 구성이 레퍼런스(Cover/History/1.Test Summary/2.Deviation/3.Test Log)와 "
            "다를 수 있음 — config sutr_template 등록 권장"
        )
        wb = spec_wb
        if spec_name != LOG_SHEET_NAME and LOG_SHEET_NAME not in wb.sheetnames:
            spec_ws.title = LOG_SHEET_NAME
        log_ws = spec_ws

    # 라운드 105 — spec 컬럼 레이아웃 동적 산출 (DV 268열 / PV 'Safety Related'
    # 삽입 + ExpR[0..83](105~188) + SUDS(189) 레이아웃 공존). 반드시 헤더 stamp
    # **전에** 스캔 — _write_log_headers가 r3 'Related ID'를 'Actual Result'로
    # 덮어쓰므로 이후에는 위치 식별 불가. DV spec에서는 산출값 == 기존 상수
    # (162/266/267/268 — 하위 호환 게이트), 스캔 실패 시 DV 상수 fallback+warning.
    layout = _detect_spec_layout(log_ws, warnings)

    # 헤더 추가 (Actual/Pass-Fail/Log).
    _write_log_headers(log_ws, warnings, layout=layout)

    # anchor 스캔 → 함수 블록 (이식된 '3.Test Log' 시트 기준).
    blocks = _scan_spec_blocks(log_ws, layout=layout)
    fn_iter_map = _build_fn_iteration_map(session)

    # 2026-06-18 Item 1 — SwUTS 스펙 자체가 함수명↔SwUFn 권위 소스. blk["unit"](D열)=
    # VectorCAST 함수명, blk["num"]=SwUFn 4-digit. SwUDS 이름맵은 정적/private 헬퍼
    # (s_*, prv_*, _FC 등)를 다수 누락하나, 시험 대상 함수는 전부 스펙에 정의돼 있어
    # 2.Deviation B열 SwUFn 해결률이 14/22 → 22/22로 개선. agg에 주입 →
    # _collect_coverage_gaps가 SwUDS보다 우선 사용.
    spec_name_to_swufn: dict[str, str] = {}
    for _blk in blocks:
        _u = (_blk.get("unit") or "").strip()
        _n = (_blk.get("num") or "").strip()
        if _u and _n:
            spec_name_to_swufn.setdefault(_u, f"SwUFn_{_n.zfill(4)}")
    agg["spec_name_to_swufn"] = spec_name_to_swufn

    fill_stats = _fill_actual_and_result(
        log_ws, blocks, fn_iter_map, asil_map, warnings, layout=layout,
    )

    # 라운드 104 — Actual Result 열(FF=162~JE=265) 서식 적용. spec graft는 Test Case/
    # Input/Expected 까지만 서식 보유, Actual 이후는 우리 추가 열이라 무서식(무테/
    # default 폰트) → 사용자 보고 "Actual만 템플릿 미적용". Expected(BF=58~FE=161)를
    # 1:1 미러. **같은 wb 내**이므로 cross-wb 객체복사가 아닌 ``_style`` 인덱스 직접
    # 복사가 정확+빠름 (라운드 103 cross-wb 문제와 구분). value는 보존(_style만 복제).
    _restyled = _apply_actual_result_style(log_ws, layout=layout)

    summary = {
        "builder": "spec-based-r92" if template_xlsm_bytes is not None else "spec-based-r91",
        "spec_sheet": spec_name,
        "spec_sha256_12": spec_sha256_12,
        # 라운드 105 — 동적 산출 레이아웃 (관측성: DV=162/266/267/268, PV=189/273/274/275).
        "spec_layout": {
            "detected": layout.detected,
            "expected_start": layout.expected_start,
            "actual_start": layout.actual_start,
            "actual_max": layout.actual_max,
            "pass_fail": layout.pass_fail,
            "pass_total": layout.pass_total,
            "log_data": layout.log_data,
            "iter_index": layout.iter_index,
        },
        "build_timestamp": meta.build_timestamp,
        "environments": len(session.environments),
        "total": agg["total"],
        "tested": agg["tested"],
        "passed": agg["passed"],
        "failed": agg["failed"],
        "spec_function_blocks": fill_stats["functions"],
        "matched_functions": fill_stats["matched_fn"],
        "unmatched_functions": fill_stats["unmatched_fn"],
        "fn_pass": fill_stats["fn_pass"],
        "fn_fail": fill_stats["fn_fail"],
        "fn_na": fill_stats["fn_na"],
        "iteration_rows": fill_stats["iterations"],
        "iter_pass": fill_stats["iter_pass"],
        "iter_fail": fill_stats["iter_fail"],
        "iter_na": fill_stats["iter_na"],
        "actual_from_expected": fill_stats["actual_from_expected"],
        "actual_from_vcast": fill_stats["actual_from_vcast"],
        "actual_missing": fill_stats["actual_missing"],
        "actual_cells_restyled": _restyled,
    }

    # 보조 시트 채움 (Cover / 1.Test Summary / 2.Deviation / History).
    if template_xlsm_bytes is not None:
        _fill_standard_aux_sheets(
            wb, meta, session, agg, summary, deviation_cases, warnings,
        )
        # 1.Test Summary TC 카운트를 함수 단위로 보정 (레퍼런스 정합).
        _fill_test_summary_counts(wb, summary, fill_stats, warnings)
        # 라운드 96-final W-4 — 미실행 TC 목록 (not-exec 카운트와 목록 정합).
        _write_not_executed_list(wb, fill_stats, summary, warnings)
        # 라운드 96-final W-11 — phase/platform ver 보정 (KJPDS02_PV / 25A1 / N/A).
        _apply_phase_platform_meta(wb, meta, summary, warnings)
    else:
        # backward-compat — 라운드 91 Cover label stamp만.
        _write_cover_meta_legacy(wb, meta, warnings)

    # 시트 순서 정리 — 3.Test Log가 2.Deviation 뒤에 오도록 (표준 베이스만).
    if template_xlsm_bytes is not None and LOG_SHEET_NAME in wb.sheetnames:
        names = [n for n in wb.sheetnames if n != "AuditLog"]
        # 표준 순서: Cover, History, 1.Test Summary, 2.Deviation, 3.Test Log
        # copy_sheet가 끝에 넣었으므로 이미 마지막 (AuditLog 미존재 시) — 명시 정렬.
        summary["output_sheet_order"] = names

    out = io.BytesIO()
    wb.save(out)
    # 라운드 106 — save 직후 무결성 검증 + 손상 시 gc 후 1회 재시도. 거대 '3.Test Log'
    #   save 가 일시적 메모리 압박으로 잘리면(unclosed token / CRC 손상) 첫 시도가
    #   손상될 수 있어, 가비지 수거 후 재 save 로 복구 시도 (wb close 전이어야 재 save 가능).
    _save_ok, _save_err = verify_xlsx_integrity(out.getvalue())
    if not _save_ok:
        import gc as _gc
        _gc.collect()
        out = io.BytesIO()
        wb.save(out)
        summary["save_retried"] = True
        warnings.append(
            f"[spec-sutr] 첫 save 무결성 실패({_save_err}) → gc 후 재 save (거대 "
            "Test Log 메모리 압박 의심). 최종 상태는 아래 무결성 검증으로 재확인, "
            "동시 빌드 수 축소 권장."
        )
    if wb is not spec_wb:
        spec_wb.close()
    wb.close()

    # 라운드 101 — 회사 ★개발템플릿 잔재(독일어 HARA externalLink + 외부참조
    # defined names) 제거 → Excel "연결 업데이트/복구" 경고 차단. keep_vba 로드 시
    # 외부링크 파트가 raw archive로 보존돼 openpyxl 객체 조작이 무효 → save된
    # bytes를 zip 레벨에서 직접 정화.
    # 라운드 106 — 빈 양식 셀 self-closing 정규화 (openpyxl 3.1.5 비효율 제거).
    #   거대 '3.Test Log'(7899행×275열 중 92%가 빈 양식 셀)의 비압축 XML ~24% 절감
    #   → Excel 열기 메모리/속도 개선 + save 중 손상(unclosed token) 위험 완화.
    #   값/스타일/병합 보존, worksheet 파트만 대상 (PV SwUTR '파일 안 열림' 근인 대응).
    _compacted, _n_compact = compact_empty_styled_cells(out.getvalue())
    if _n_compact:
        out = io.BytesIO(_compacted)
        summary["empty_cells_compacted"] = _n_compact
        warnings.append(
            f"[spec-sutr] 빈 양식 셀 {_n_compact}개 self-closing 정규화 "
            "(openpyxl 3.1.5 비효율 — Test Log 비대/손상 완화)"
        )

    _sanitized, _ext_removed = sanitize_xlsm_external_links(out.getvalue())
    if _ext_removed:
        out = io.BytesIO(_sanitized)
        summary["external_links_stripped"] = _ext_removed
        warnings.append(
            f"[spec-sutr] 템플릿 외부링크 파트 {_ext_removed}건 + 외부참조 defined "
            "name 제거 (독일어 HARA 양식 잔재 — Excel 연결 경고 차단)"
        )
    out.seek(0)

    # 라운드 106 — save 무결성 검증 (거대 Test Log XML 잘림 손상 배포 차단).
    #   save 가 메모리 압박으로 중단되면 worksheet XML 이 </worksheet> 미완결로
    #   잘려 Excel 에서 안 열린다 (PV SwUTR 실측). 손상 감지 시 critical warning.
    _ok, _err = verify_xlsx_integrity(out.getvalue())
    summary["integrity_check"] = "ok" if _ok else f"FAILED: {_err}"
    if not _ok:
        warnings.append(
            f"[spec-sutr] ⚠️ 산출물 무결성 검증 실패: {_err} — 거대 '3.Test Log' "
            "save 가 메모리 압박으로 잘렸을 수 있음. 재생성 권장 (동시 빌드 축소 시 회복)."
        )
    out.seek(0)

    if meta.doc_filename_pattern:
        filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version, date=short_date(meta.test_date),
        )
    else:
        filename = (
            f"({meta.project_id}_DV_SwUTR) Software Unit Test Result_"
            f"v{meta.release_sw_version}_{short_date(meta.test_date)}_R.xlsm"
        )

    return SutrBuildResult(
        ok=True,
        xlsm_io=out,
        filename=filename,
        warnings=warnings,
        vba_macros_preserved=template_has_vba,
        summary=summary,
    )


__all__ = [
    "build_sutr_from_spec",
    "extract_spec_fi_stats",
    "LOG_SHEET_NAME",
    "COL_ACTUAL_START",
    "COL_PASS_FAIL",
    "COL_PASS_TOTAL",
    "COL_LOG_DATA",
    "SPEC_SHEET_RE",
    "SpecLayout",
]
