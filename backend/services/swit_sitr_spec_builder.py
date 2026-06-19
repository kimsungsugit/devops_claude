"""SwIT SITR (Software Integration Test Result) — **spec-based** Test Log 빌더.

회사 KJPDS02 PV SwITR v0.10 양식(5시트: Cover / History / 1.Test Summary /
2.Deviation / 3.Test Log)에 정합하는 spec-based 빌더. SwUT SUTR의
``swut_sutr_spec_builder.build_sutr_from_spec`` (라운드 92~110) 를 미러링하되
SwIT 델타만 적용한다.

## SwUTR 대비 핵심 차이 (SwIT 델타)

1. **2.Deviation 7열 스키마** (``_write_swit_spec_deviation``) — SwUTR 6열
   (Unit/Coverage/미달성 사유)에 **TC 컬럼(B)** 이 추가된 7열. r6+ B열에
   ``SwITC_SwUFn_NNNN`` 형식 TC-id 기재. gap에는 TC 정보가 없어 best-effort
   파생(``agg`` 역매핑 → ``SwITC_{swufn}`` 파생 → 노란 placeholder).

2. **base 템플릿 처리** — v0.10 SwITR 템플릿은 4시트(Cover/History/
   1.Test Summary/**2.Test Log** 40열)이고 separate 2.Deviation 시트가 **없다**.
   따라서: (a) 좁은 '2.Test Log' 시트 삭제 → (b) spec sheet를 '3.Test Log'로
   이식 → (c) '2.Deviation' 시트 **신규 생성** 후 작성.

3. **TC-id 매칭** — SwITS spec F열 TC_ID는 ``SwITC_0101_01`` (SwUFn_ 없음),
   VectorCAST env test_case는 ``SwITC_SwUFn_0101.MMM`` / ``SwUFn_0101.MMM``.
   SwUTR ``_build_fn_iteration_map`` 은 ``re.match(r"SwUFn_...")`` 로 anchor
   되어 ``SwITC_SwUFn_`` prefix를 탈락시킨다. 따라서 SwIT 전용
   ``_build_swit_fn_iteration_map`` (``SwITC_SwUFn_`` / ``SwUFn_`` 양쪽 prefix
   허용) 을 작성한다. anchor 스캔/Actual stamp/Total 산출은 SwUTR 헬퍼 재사용.

## 재사용 (SwUTR 제네릭 헬퍼 — SwUFn/TC-id 무관 또는 SwITS에도 동작)

``swut_sutr_spec_builder`` 에서 직접 import:
    _detect_spec_layout / _write_log_headers / _scan_spec_blocks /
    _expected_var_cols / _find_spec_sheet / _fill_actual_and_result /
    _apply_actual_result_style / _lookup_vcast_actual / _apply_asil_mark /
    _mark_cell / _collect_coverage_gaps / _compress_line_numbers /
    _format_gap_source / SpecLayout / SPEC_SHEET_RE / DATA_START_ROW /
    COL_INDEX / COL_TC_ID / COL_UNIT / USER_INPUT_PLACEHOLDER

aux 시트:
    swut_sutr_aggregator._write_cover / _write_test_summary
    swut_coverage_aggregator._write_history_sheet

## ISO 26262 Integration test
SwIT SITR은 ASIL B+ 이상의 evidence (분기 커버리지 + 인터페이스 테스트 +
Deviation 기록). evidence "auto-generated draft" — manual review 의무.
"""
from __future__ import annotations

import hashlib
import io
import re
from typing import Any

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

from backend.services.design_tokens import USER_INPUT_PLACEHOLDER
from backend.services.excel_template_utils import (
    build_release_history_row,
    compact_empty_styled_cells,
    copy_sheet_across_workbooks,
    has_vba_macros,
    inspect_vba_refs,
    safe_write,
    sanitize_xlsm_external_links,
    short_date,
    validate_build_meta,
    validate_xlsx_template_bytes,
    verify_xlsx_integrity,
)
from backend.services.swit_meta import SwitSitrBuildMeta
from backend.services.swit_sitr_aggregator import SwitSitrBuildResult
from backend.services.swut_builder_helpers import extract_warnings_from_session
from backend.services.swut_coverage_aggregator import _write_history_sheet
from backend.services.swut_input_adapter import (
    SwUTSession,
    aggregate_session,
)
from backend.services.swut_sutr_aggregator import (
    _write_cover,
    _write_test_summary,
)
from backend.services.swut_sutr_spec_builder import (
    DATA_START_ROW,
    HEADER_SECTION_ROW,
    SUBHEADER_ROW,
    _apply_actual_result_style,
    _collect_coverage_gaps,
    _compress_line_numbers,
    _detect_spec_layout,
    _fill_actual_and_result,
    _find_spec_sheet,
    _format_gap_source,
    _mark_cell,
    _write_log_headers,
)

# SwIT SITR 이식 시트 제목 — SwUTR과 동일 ('3.Test Log').
LOG_SHEET_NAME = "3.Test Log"
DEVIATION_SHEET_NAME = "2.Deviation"
# v0.10 SwITR 템플릿의 좁은 Test Log 시트 (이식 전 삭제 대상).
_NARROW_LOG_SHEET_NAME = "2.Test Log"


# ---------------------------------------------------------------------------
# SwIT 전용 iteration map (env_name 기준 — SwUFn id는 tc_name이 아닌 env_name에 있음)
# ---------------------------------------------------------------------------


def _swit_tc_key(value: str) -> str:
    """SwIT TC 식별자 정규화 — spec TC_ID와 VectorCAST env_name을 동일 키로 정렬.

    실측(2026-06-19): VectorCAST SwIT는 **env_name**에 SwUFn id가 있고
    (``SwIT_SwUFn_0101_01`` / ``SwIT_FI_SwFn_34``), **tc_name**은 서브프로그램명+
    iteration(``g_Ap_Main.001``)이다. spec TC_ID(``SwITC_SwUFn_0101_01`` /
    ``SwITC_FI_SwFn_34``)와 env_name을 매칭하려면 선두 ``SwITC_`` / ``SwIT_``
    prefix를 떼고 대문자화한다:

      spec ``SwITC_SwUFn_0101_01`` / env ``SwIT_SwUFn_0101_01`` → ``SWUFN_0101_01``
      spec ``SwITC_FI_SwFn_34``    / env ``SwIT_FI_SwFn_34``    → ``FI_SWFN_34``

    영문 시작이라 ``_fill_actual_and_result`` 의 ``num.lstrip('0')`` 무영향.
    실측 44/54 블록 매칭(나머지 10 SWUFN_1701~1710은 로그에 env 부재=미실행 N/A).
    """
    s = str(value or "").strip()
    s = re.sub(r"^Sw(?:ITC|IT)_", "", s, count=1, flags=re.IGNORECASE)
    return s.upper()


# tc_name 끝의 .MMM iteration suffix (서브프로그램명.NNN).
_SWIT_ITER_RE = re.compile(r"\.(\d+)\s*$")


def _build_swit_fn_iteration_map(
    session: SwUTSession,
) -> dict[str, dict[int, Any]]:
    """env_name 정규화 키 → {iteration_index(int): {"passed","actual","env",
    "tc_name"}} (SwIT 델타 — 2026-06-19 env_name 기준 재작성).

    구버전은 tc_name에서 ``SwUFn_(\\d+)\\.(\\d+)`` 매칭을 시도했으나, 실측상 SwIT
    tc_name은 서브프로그램명(``g_Ap_Main.001``)이라 전부 미스(맵 0개) → fill
    미동작. SwUFn id는 ``env.env_name``(``SwIT_SwUFn_0101_01``)에 있으므로 env
    단위로 키잉하고 iteration index는 tc_name ``.MMM`` suffix에서 추출한다.

    동일 idx 다중 서브프로그램 충돌 시 첫 등록 보존하되, passed 미결정 슬롯은
    passed 결정된 레코드로 승격(실행 누락 라벨 방지).
    """
    fn_map: dict[str, dict[int, Any]] = {}
    for env in session.environments:
        key = _swit_tc_key(getattr(env, "env_name", ""))
        if not key:
            continue
        slot = fn_map.setdefault(key, {})
        for tc_name in env.test_cases:
            m = _SWIT_ITER_RE.search(str(tc_name))
            if not m:
                continue
            idx = int(m.group(1))
            exec_r = env.test_results.get(tc_name)
            actual_dict: dict = {}
            passed = None
            if exec_r is not None:
                passed = bool(exec_r.passed)
                actual_dict = getattr(exec_r, "actual_result", {}) or {}
            if not actual_dict:
                tr_items = getattr(env, "tc_result_items", {}).get(tc_name, [])
                tr_item = tr_items[0] if tr_items else None
                if tr_item is not None:
                    actual_dict = getattr(tr_item, "actual_result", {}) or {}
            existing = slot.get(idx)
            if existing is not None and not (
                existing.get("passed") is None and passed is not None
            ):
                continue  # 첫 등록 보존 (passed 미결정→결정 승격만 허용)
            slot[idx] = {
                "passed": passed,
                "actual": actual_dict,
                "env": env,
                "tc_name": tc_name,
            }
    return fn_map


# ---------------------------------------------------------------------------
# spec 함수명 → TC-id 역매핑 (2.Deviation B열 best-effort)
# ---------------------------------------------------------------------------

def _build_swit_unit_to_tc_map(
    blocks: list[dict[str, Any]],
) -> dict[str, str]:
    """spec 함수 블록에서 {Unit(D열) → TC-id(C열)} 역매핑 (SwIT 2.Deviation B열).

    ``_collect_coverage_gaps`` gap은 함수명(Unit)만 알고 TC 정보가 없으므로,
    spec 블록의 Unit↔TC-id 매핑으로 B열 TC-id를 best-effort 복원한다. 동일
    Unit이 다중 TC면 첫 TC-id 유지. 키는 Unit 원문 + lowercase 양쪽 등재
    (casing 차이 흡수).
    """
    unit_to_tc: dict[str, str] = {}
    for blk in blocks:
        unit = (blk.get("unit") or "").strip()
        tc_id = (blk.get("tc_id") or "").strip()
        if not unit or not tc_id:
            continue
        unit_to_tc.setdefault(unit, tc_id)
        unit_to_tc.setdefault(unit.lower(), tc_id)
    return unit_to_tc


def _derive_swit_tc_from_swufn(swufn: str) -> str:
    """SwUFn-id(``SwUFn_0101``)에서 SwIT TC-id(``SwITC_SwUFn_0101``) 파생.

    spec Unit↔TC 역매핑 미스 시 fallback. 레퍼런스 Test Log B열 형식
    (``SwITC_SwUFn_NNNN``) 정합. ``SwUFn_`` 형식이 아니면 빈 string.
    """
    m = re.search(r"SwUFn_(\d+)", swufn or "", re.IGNORECASE)
    if not m:
        return ""
    return f"SwITC_SwUFn_{m.group(1)}"


# ---------------------------------------------------------------------------
# SwIT 전용 spec 블록 스캐너 (SwUTS `_scan_spec_blocks` 미러 — SwITS 레이아웃 델타)
# ---------------------------------------------------------------------------

# SwITS spec B열 anchor TC_ID 패턴. SwUTS는 B=숫자 index(`isdigit`)이나 SwITS PV는
# B열에 TC_ID(`SwITC_SwUFn_NNNN_NN` / WIP 'SwUfn' 오타)가 직접 들어가고 블록 전체가
# 세로병합된다. 'SwITC_FI_SwFn_NN'(Fault Injection) 블록도 anchor로 포착(매칭은
# `_build_swit_fn_iteration_map`의 SwUFn 정규식이 거르므로 N/A 처리됨).
_SWIT_ANCHOR_RE = re.compile(r"Sw(?:ITC|UFn|Ufn)", re.IGNORECASE)


def _scan_swit_spec_blocks(ws, layout=None) -> list[dict[str, Any]]:
    """SwITS spec anchor 스캔 → 함수 블록 list (SwIT 델타).

    SwUTR ``_scan_spec_blocks``(B=숫자 index, C=TC_ID, D=Unit)는 SwITS PV
    레이아웃과 불일치해 anchor 0개 → fill 미동작. SwITS PV 실측 구조:

      - **B열(2) = TC_ID**(``SwITC_SwUFn_NNNN_NN``), 블록 전체 **세로병합**
        (예 B5:B40). 병합 비-anchor 행은 openpyxl이 None 반환 → anchor 행만 값.
      - **C열(3) = iteration index**(1..N). anchor 행은 C 공란(변수명 행).
      - Input은 I열(9)~ (anchor 행=변수명, iteration 행=값).

    anchor 행(변수명) + 이후 다음 anchor 전까지 C열 digit(iteration index) 보유
    행을 ``iter_rows``로 수집. ``num``은 TC_ID 첫 숫자열(4-digit) — VectorCAST
    ``(?:SwITC_)?SwUFn_NNNN.MMM`` 매칭 키와 동일(``lstrip('0')`` 후).

    iteration index 컬럼이 layout.iter_index(SwITS는 'Param 1' 헤더라 'Inpt['
    미발견 → 기본 G열로 오탐, 공란)와 어긋나도, ``_fill_actual_and_result``가
    iter index 파싱 실패 시 순번 fallback(enumerate)으로 강등 — iter_rows가 C열
    기준 1..N 연속이므로 VectorCAST .001..00N과 순번 정합한다.

    ``layout`` 인자는 시그니처 호환용(현재 미사용 — anchor/iter 판정이 B/C열
    고정). Returns: ``[{anchor, tc_id, unit, num, iter_rows}, ...]``.
    """
    blocks: list[dict[str, Any]] = []
    max_row = int(ws.max_row or 0)
    rr = DATA_START_ROW
    while rr <= max_row:
        b = ws.cell(rr, 2).value  # B = TC_ID (SwITS)
        bs = str(b).strip() if b is not None else ""
        if bs and _SWIT_ANCHOR_RE.match(bs):
            # num = fill 매칭 키. SwUFn id는 env_name과 동일 정규화(_swit_tc_key)로
            # 정렬해야 하므로 (B열 TC_ID 전체 → 'SWUFN_0101_01' / 'FI_SWFN_34'),
            # 4-digit 첫 숫자만 쓰면 _NN suffix 손실로 0101_01~05가 모두 충돌한다.
            swufn_m = re.search(r"SwUFn_(\d+)", bs, re.IGNORECASE)
            swufn_num = swufn_m.group(1) if swufn_m else ""
            iter_rows: list[int] = []
            k = rr + 1
            while k <= max_row:
                nb = ws.cell(k, 2).value
                nbs = str(nb).strip() if nb is not None else ""
                if nbs and _SWIT_ANCHOR_RE.match(nbs):
                    break  # 다음 anchor
                # iteration index 열: regular 블록은 C(3)·H(8) 둘 다, FI 블록
                # (SwITC_FI_SwFn_NN)은 C 공란 + H(8)에만 index → C-or-H로 양쪽 포착
                # (H만 쓰면 무방하나 robust하게 OR). anchor 행은 둘 다 공란이라 제외.
                ci = ws.cell(k, 3).value  # C = iteration index (regular)
                hi = ws.cell(k, 8).value  # H = iteration index (regular+FI)
                if (ci is not None and str(ci).strip().isdigit()) or (
                    hi is not None and str(hi).strip().isdigit()
                ):
                    iter_rows.append(k)
                k += 1
            blocks.append({
                "anchor": rr,
                "tc_id": bs,
                # SwITS spec엔 별도 Unit 열이 없음 — TC_ID에서 SwUFn_NNNN 파생
                # (deviation 비움 모드라 fill엔 비핵심, spec_name_to_swufn/경고용).
                "unit": f"SwUFn_{swufn_num.zfill(4)}" if swufn_num else bs,
                # 매칭 키 — env_name 정규화와 동일 (TC_ID 전체 prefix-strip+대문자).
                "num": _swit_tc_key(bs),
                "iter_rows": iter_rows,
            })
            rr = k if k > rr else rr + 1
        else:
            rr += 1
    return blocks


def _apply_swit_legend_offset(ws, layout, out_warnings=None) -> bool:
    """Test Log 상단에 2행 삽입 + B2:B4 색범례 — 회사 PV 레퍼런스 정합.

    spec 복사본은 헤더 r3(HEADER_SECTION_ROW)/데이터 r5(DATA_START_ROW)이나,
    레퍼런스 SwITR Test Log는 상단에 Expected/Actual/Pass-Fail 색범례(B2:B4) 3행을
    두어 **헤더 r5 / 서브헤더 r6 / 데이터 r7**이다(2행 하향).

    ``move_range(translate=True)`` 로 A2:끝을 2행 하향(셀 값·_style·수식 자동 보정
    — 실측: COUNTIF LE5:LE669 → LE7:LE671 자동 변환), 병합 range는 move_range가
    이동 안 하므로 사전 unmerge → 사후 +2 재매핑. 범례 텍스트는 레퍼 실측 그대로
    (오타 'Acual' 포함), fill은 각 컬럼 그룹 헤더(Expected/Actual/Pass-Fail)에서
    복제해 색 코딩을 일치시킨다.

    **반드시 모든 Test Log 쓰기(fill·style) 완료 후 호출** — 이후 행 좌표가 +2.
    실패 시(비치명) 시프트 없이 원본 레이아웃 유지하고 False 반환.
    """
    from copy import copy as _copy

    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import range_boundaries

    try:
        max_col = int(ws.max_column or 0)
        max_row = int(ws.max_row or 0)
        if max_col < 2 or max_row < DATA_START_ROW:
            return False
        merges = [str(rng) for rng in ws.merged_cells.ranges]
        for m in merges:
            ws.unmerge_cells(m)
        # A2:끝 2행 하향 (행1 상단 빈행 유지). translate=True → 수식 좌표 자동 +2.
        ws.move_range(
            f"A2:{get_column_letter(max_col)}{max_row}", rows=2, cols=0, translate=True,
        )
        # 병합 재매핑 (+2행, min_row==1 병합은 미시프트).
        for m in merges:
            c1, r1, c2, r2 = range_boundaries(m)
            if None in (c1, r1, c2, r2):
                continue
            ws.merge_cells(
                start_row=(r1 + 2 if r1 >= 2 else r1), start_column=c1,
                end_row=(r2 + 2 if r2 >= 2 else r2), end_column=c2,
            )
        # 범례 B2:B4 (+ C열 '-'). 헤더는 이제 HEADER_SECTION_ROW+2 (=r5).
        hdr_row = HEADER_SECTION_ROW + 2
        legend = (
            ("Expected Result", layout.expected_start),
            ("Acual Result", layout.actual_start),  # 레퍼 실측 철자 그대로(Acual)
            ("Pass//Fail", layout.pass_fail),
        )
        for i, (label, src_col) in enumerate(legend):
            row = 2 + i
            safe_write(ws, row, 2, label)
            safe_write(ws, row, 3, "-")
            try:
                src_fill = ws.cell(hdr_row, src_col).fill
                if src_fill is not None:
                    ws.cell(row, 2).fill = _copy(src_fill)
            except (AttributeError, ValueError):
                pass
        return True
    except (ValueError, AttributeError, TypeError, KeyError, IndexError) as e:
        # 범례는 cosmetic — 실패해도 데이터/구조는 정상이므로 시프트 없이 진행.
        if out_warnings is not None:
            out_warnings.append(
                f"[spec-sitr] Test Log 범례 offset 실패(비치명, 원본 레이아웃 유지): "
                f"{type(e).__name__}: {e}"
            )
        return False


# ---------------------------------------------------------------------------
# 2.Deviation 7열 (SwIT 델타 — SwUTR `_write_spec_deviation` 6열 + TC 컬럼)
# ---------------------------------------------------------------------------

def _write_swit_spec_deviation(
    ws,
    agg: dict[str, Any],
    out_warnings: list[str],
    *,
    empty: bool = False,
    unit_to_tc: dict[str, str] | None = None,
) -> int:
    """spec-based SITR 2.Deviation — 레퍼런스 KJPDS02 PV SwITR 7열 스키마.

    SwUTR ``_write_spec_deviation`` (6열: Unit/Coverage/미달성 사유) 베이스에
    **TC 컬럼(B)** 이 추가된 7열. 레퍼런스 실측 스키마:

      r1: B='Deviation Report'
      r3: B='■ Deviation List'
      r4: B='TC'(B:B) | C='Unit'(C:D) | E='Coverage'(E:F) | G='미달성 사유'(G:H)
      r5: B='ID' C='ID' D='Name' E='Type' F='Value' G='line' H='Description'
      r6+: B=TC-id(``SwITC_SwUFn_NNNN``) C=Unit(SwUFn_NNNN) D=name
           E=Statements/Branches F='N / M (X%)' G=미달 line H=소스+사유

    자동/수기 경계: C/D/E/F는 커버리지 집계로 자동 산출. G(미달 line)·H(소스
    발췌)는 annotated source(``agg["function_gap_lines"]``)에서 자동 채움,
    미가용 시 노란 placeholder. **B(TC-id)** 는 spec Unit↔TC 역매핑 →
    ``SwITC_SwUFn_NNNN`` 파생 → 노란 placeholder 순으로 best-effort.

    empty=True (meta.deviation_empty) → '해당 사항 없음' (SwUTR과 동일).

    Returns: 쓰여진 gap 행 수.
    """
    gaps = [] if empty else _collect_coverage_gaps(agg)
    gap_lines_map: dict[str, dict] = agg.get("function_gap_lines") or {}
    gap_lines_ci = {str(k).lower(): v for k, v in gap_lines_map.items()}
    u2tc = unit_to_tc or {}
    auto_fg = 0
    placeholder_fg = 0
    nonexact_fn = 0
    tc_placeholder = 0

    if empty and out_warnings is not None:
        out_warnings.append(
            "[spec-sitr] 2.Deviation 비움(deviation_empty=true) — 커버리지 미달 "
            "목록 미기재, '해당 사항 없음' 표기. 미달 상세는 SwITCV 산출물 참조."
        )

    # 1) 신규 생성 시트(또는 잔존 양식) clear — 병합 해제 후 값 제거.
    clear_rows = max(ws.max_row or 0, 100)
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= clear_rows and rng.min_col <= 8:
            try:
                ws.unmerge_cells(str(rng))
            except (ValueError, KeyError):
                pass
    for r in range(1, clear_rows + 1):
        for c in range(1, 9):
            try:
                ws.cell(r, c).value = None
            except (ValueError, AttributeError):
                pass

    # 2) 헤더 (레퍼런스 7열 스키마 — SwUTR 6열 + TC 컬럼 시프트).
    safe_write(ws, 1, 2, "Deviation Report")
    safe_write(ws, 3, 2, "■ Deviation List")
    safe_write(ws, 4, 2, "TC")
    safe_write(ws, 4, 3, "Unit")
    safe_write(ws, 4, 5, "Coverage")
    safe_write(ws, 4, 7, "미달성 사유")
    safe_write(ws, 5, 2, "ID")
    safe_write(ws, 5, 3, "ID")
    safe_write(ws, 5, 4, "Name")
    safe_write(ws, 5, 5, "Type")
    safe_write(ws, 5, 6, "Value")
    safe_write(ws, 5, 7, "line")
    safe_write(ws, 5, 8, "Description")
    # r4 그룹 병합: TC(B:B 단일), Unit(C:D), Coverage(E:F), 미달성 사유(G:H).
    for a, b in (("C4", "D4"), ("E4", "F4"), ("G4", "H4")):
        try:
            ws.merge_cells(f"{a}:{b}")
        except (ValueError, AttributeError):
            pass
    try:
        from copy import copy as _copy

        from openpyxl.styles import Alignment, Font
        hf = Font(name="맑은 고딕", size=10, bold=True)
        ha = Alignment(horizontal="center", vertical="center")
        for rr in (4, 5):
            for cc in range(2, 9):
                cell = ws.cell(rr, cc)
                cell.font = _copy(hf)
                cell.alignment = _copy(ha)
    except (AttributeError, TypeError):
        pass

    # 3) gap 행 — B=TC-id, C=SwUFn, D=name, E=Type, F=Value, G=line, H=Desc.
    r = 6
    written = 0
    if not gaps:
        safe_write(ws, r, 2, "해당 사항 없음")
        r += 1
    for swufn, name, s_gap, b_gap, kind in gaps:
        first = True
        gl = gap_lines_map.get(name) or gap_lines_ci.get(str(name).lower()) or {}
        # B(TC-id) best-effort — spec Unit↔TC 역매핑 → SwITC_SwUFn 파생 → placeholder.
        tc_id = u2tc.get(name) or u2tc.get(str(name).lower()) or ""
        tc_derived = False
        if not tc_id:
            tc_id = _derive_swit_tc_from_swufn(swufn)
            tc_derived = bool(tc_id)
        for typ, gap, bucket in (
            ("Statements", s_gap, "statements"),
            ("Branches", b_gap, "branches"),
        ):
            if gap is None:
                continue
            covered, total, pct = gap
            if first:
                # B(TC-id) — 역매핑/파생/placeholder.
                if tc_id:
                    safe_write(ws, r, 2, tc_id)
                    if tc_derived:
                        _mark_cell(ws, r, 2)  # SwUFn 파생 — 수기 확인 유도.
                        tc_placeholder += 1
                else:
                    safe_write(ws, r, 2, USER_INPUT_PLACEHOLDER)
                    _mark_cell(ws, r, 2)
                    tc_placeholder += 1
                # C(SwUFn-id) — SwUTR B열과 동일 처리(비-exact 노란마킹).
                safe_write(ws, r, 3, swufn)
                safe_write(ws, r, 4, name)
                if (kind != "exact"
                        or not re.match(r"SwUFn_\d", str(swufn), re.IGNORECASE)):
                    _mark_cell(ws, r, 3)
                    nonexact_fn += 1
                first = False
            safe_write(ws, r, 5, typ)
            safe_write(ws, r, 6, f"{covered} / {total} ({round(pct * 100)}%)")
            # G(미달 line)·H(소스 발췌) — annotated source 자동, 미가용 시 placeholder.
            pairs = (gl.get(bucket) if isinstance(gl, dict) else None) or []
            if pairs:
                safe_write(ws, r, 7, _compress_line_numbers([p[0] for p in pairs]))
                g_txt = _format_gap_source(pairs)
                if g_txt:
                    safe_write(ws, r, 8, g_txt)
                else:
                    safe_write(ws, r, 8, USER_INPUT_PLACEHOLDER)
                    _mark_cell(ws, r, 8)
                auto_fg += 1
            else:
                safe_write(ws, r, 7, USER_INPUT_PLACEHOLDER)
                _mark_cell(ws, r, 7)
                safe_write(ws, r, 8, USER_INPUT_PLACEHOLDER)
                _mark_cell(ws, r, 8)
                placeholder_fg += 1
            r += 1
            written += 1

    # 4) Appendix sentinel + End (레퍼런스 정합 — 7열 시프트).
    r += 1
    safe_write(ws, r, 2, "■ Appendix - 발생 가능 값")
    r += 1
    safe_write(ws, r, 2, "Related Test Case ID")
    safe_write(ws, r, 3, "Parameter")
    safe_write(ws, r, 6, "Value")
    safe_write(ws, r, 8, "Note")
    r += 1
    safe_write(ws, r, 2, "해당 사항 없음")
    r += 2
    safe_write(ws, r, 2, "■ Appendix - 첨부자료")
    r += 2
    safe_write(ws, r, 2, "< End of Document >")

    if out_warnings is not None and written:
        out_warnings.append(
            f"[spec-sitr] 2.Deviation 커버리지 미달 {written}행 기재 — "
            f"G(line)/H(소스) 자동 {auto_fg}행 (annotated source), "
            f"placeholder {placeholder_fg}행 (annotated source 미가용)"
        )
        if tc_placeholder:
            out_warnings.append(
                f"[spec-sitr] 2.Deviation B열 TC-id 추정/placeholder {tc_placeholder}건 "
                "(노란마킹) — spec Unit↔TC 역매핑 미스로 SwITC_SwUFn 파생 또는 미상. "
                "audit 수기 확인 필요."
            )
        if nonexact_fn:
            out_warnings.append(
                f"[spec-sitr] 2.Deviation SwUFn 해결 비-exact {nonexact_fn}함수 "
                "(C열 노란마킹) — 스펙/SwUDS 미등재 또는 case-insensitive 폴백. "
                "audit 수기 검증 필요."
            )
        out_warnings.append(
            "[spec-sitr] 2.Deviation 주의: 다중 env에 동일 함수가 등장할 경우 "
            "F(커버리지 값)와 G/H(미달 line/소스)가 서로 다른 env 스냅샷일 수 있음 "
            "(다중 env 함수는 audit 수기 확인 권장)."
        )
    return written


# ---------------------------------------------------------------------------
# 보조 시트 채움 (Cover / 1.Test Summary / 2.Deviation / History)
# ---------------------------------------------------------------------------

def _fill_swit_aux_sheets(
    wb: Workbook,
    meta: SwitSitrBuildMeta,
    agg: dict[str, Any],
    summary: dict[str, Any],
    unit_to_tc: dict[str, str],
    out_warnings: list[str],
) -> None:
    """SITR 템플릿의 Cover / 1.Test Summary / 2.Deviation / History 채움.

    표준 ``swut_sutr_aggregator`` writer 재사용 (SwUTR ``_fill_standard_aux_sheets``
    미러). 단 2.Deviation 만 SwIT 전용 신규 writer(``_write_swit_spec_deviation``)
    사용. 3.Test Log 는 spec 이식 (별도 처리)이라 건드리지 않는다.
    """
    sheet_names = wb.sheetnames

    cover_ws = next((wb[n] for n in sheet_names if n.lower() == "cover"), None)
    if cover_ws is None:
        out_warnings.append("[spec-sitr] Cover 시트 미발견")
    else:
        _write_cover(cover_ws, meta, out_warnings=out_warnings)

    ts_ws = next((wb[n] for n in sheet_names if "test summary" in n.lower()), None)
    if ts_ws is None:
        out_warnings.append("[spec-sitr] 1.Test Summary 시트 미발견")
    else:
        _write_test_summary(
            ts_ws, meta, agg, out_warnings=out_warnings,
            layout=None, summary=summary,
        )

    dev_ws = next(
        (wb[n] for n in sheet_names if "deviation" in n.lower()), None,
    )
    if dev_ws is None:
        out_warnings.append("[spec-sitr] 2.Deviation 시트 미발견 — 신규 생성 실패")
    else:
        n = _write_swit_spec_deviation(
            dev_ws, agg, out_warnings,
            empty=bool(getattr(meta, "deviation_empty", False)),
            unit_to_tc=unit_to_tc,
        )
        summary["deviation_cases_written"] = n

    hist_ws = next((wb[n] for n in sheet_names if n.lower() == "history"), None)
    if hist_ws is not None:
        release_rows = build_release_history_row(
            meta, doc_kind="SwIT SITR (spec-based)", out_warnings=out_warnings,
        )
        n_h = _write_history_sheet(hist_ws, release_rows, out_warnings=out_warnings)
        summary["history_rows_written"] = n_h


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_sitr_from_spec(
    session: SwUTSession,
    meta: SwitSitrBuildMeta,
    spec_xlsm_bytes: bytes,
    *,
    template_xlsm_bytes: bytes | None = None,
    function_asil_map: dict[str, str] | None = None,
    deviation_cases: list[Any] | None = None,
) -> SwitSitrBuildResult:
    """SITR 생성 — v0.10 SwITR 템플릿 베이스 + SwITS spec '3.Test Log' 이식.

    SwUT ``build_sutr_from_spec`` 미러. SwIT 델타:
      - 좁은 '2.Test Log' 시트 삭제 (v0.10 템플릿은 4시트 — separate Deviation 없음).
      - spec '3. SW Integration Test Spec ' 시트를 '3.Test Log' 로 이식.
      - '2.Deviation' 시트 **신규 생성** 후 7열 스키마로 작성.
      - VectorCAST test_case 'SwITC_SwUFn_NNNN.MMM' prefix 흡수
        (``_build_swit_fn_iteration_map``).

    Args:
        session: collect_swit_session 출력 (VectorCAST Actual/Pass-Fail source).
        meta: SwIT SITR 빌드 메타 (doc_id_base="HDPDM01-SITR").
        spec_xlsm_bytes: SwITS spec xlsm bytes ('3. SW Integration Test Spec ' —
            Input/Expected/TC 보존, '3.Test Log'로 이식).
        template_xlsm_bytes: v0.10 SwITR 템플릿 xlsm bytes (베이스 — Cover/History/
            1.Test Summary/2.Test Log 보유). None이면 spec wb 베이스 fallback
            (audit 비권장, warning 누적).
        function_asil_map: SwUFn_NNNN → ASIL 등급 (옵션, anchor 시각 강조).
        deviation_cases: 현재 미사용 (SwUTR 시그니처 호환 — Appendix '발생 가능
            값'은 본 목록과 무관, 후속 라운드 연동).

    Returns:
        SwitSitrBuildResult — xlsm_io에 템플릿 + spec 이식 '3.Test Log' + Actual/
        Pass-Fail/Log + Cover/Test Summary/2.Deviation/History stamp.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for spec-based SITR builder")

    validate_build_meta(
        meta.release_sw_version, meta.test_date,
        doc_id_sequence=meta.doc_id_sequence,
        test_engineer=meta.test_engineer,
        author=meta.default_author,
        approver=meta.default_approver or meta.approver_override,
        reviewer=meta.default_reviewer or meta.reviewer_override,
    )
    validate_xlsx_template_bytes(spec_xlsm_bytes, label="SwITS spec xlsm")

    spec_sha256_12 = hashlib.sha256(spec_xlsm_bytes).hexdigest()[:12]
    warnings: list[str] = extract_warnings_from_session(session)
    asil_map = function_asil_map or {}

    # VBA 검사 — spec + (있으면) 템플릿 둘 다 keep_vba.
    spec_has_vba = has_vba_macros(spec_xlsm_bytes)
    template_has_vba = spec_has_vba
    if template_xlsm_bytes is not None:
        validate_xlsx_template_bytes(template_xlsm_bytes, label="SwITR 템플릿 xlsm")
        template_has_vba = has_vba_macros(template_xlsm_bytes)
    if template_has_vba or spec_has_vba:
        warnings.append(
            "VBA macro execution NOT verified — open output xlsm in Excel and verify "
            "macros before submitting as evidence"
        )
        refs = inspect_vba_refs(template_xlsm_bytes or spec_xlsm_bytes)
        if refs:
            warnings.append(
                f"VBA stale ref 위험 패턴 — {refs} (셀/시트 이동 시 매크로 깨질 위험)"
            )

    # spec wb 로드 (Test Log source).
    spec_wb: Workbook = openpyxl.load_workbook(
        io.BytesIO(spec_xlsm_bytes), keep_vba=True, data_only=False,
    )
    spec_ws, spec_name = _find_spec_sheet(spec_wb)
    if spec_ws is None:
        warnings.append("[spec-sitr] SwITS spec 시트 미발견 — 빌드 불가")
        spec_wb.close()
        return SwitSitrBuildResult(ok=False, warnings=warnings)

    agg = aggregate_session(session)

    if template_xlsm_bytes is not None:
        # v0.10 SwITR 템플릿을 베이스로 로드.
        wb: Workbook = openpyxl.load_workbook(
            io.BytesIO(template_xlsm_bytes), keep_vba=True, data_only=False,
        )
        # 좁은 '2.Test Log'(40열) 시트 제거 — spec 와이드 이식으로 대체.
        removed_narrow = False
        for nm in list(wb.sheetnames):
            low = nm.lower()
            if (low == _NARROW_LOG_SHEET_NAME.lower()
                    or low == "2.test result"
                    or ("test log" in low and "3." not in nm)
                    or "test result" in low):
                del wb[nm]
                removed_narrow = True
        if not removed_narrow:
            warnings.append(
                "[spec-sitr] 좁은 '2.Test Log/Result' 시트 미발견 — 템플릿 양식 확인 "
                "(v0.10 SwITR 4시트 가정과 불일치 가능)"
            )
        # spec 와이드 시트를 '3.Test Log'로 풀 카피 이식 (끝에 추가).
        log_ws = copy_sheet_across_workbooks(
            spec_ws, wb, new_title=LOG_SHEET_NAME, insert_index=None,
        )
        # '2.Deviation' 시트 신규 생성 (v0.10 템플릿엔 없음) — index=3 (Cover/
        # History/1.Test Summary 뒤). 시트 수가 부족하면 끝에 추가(create_sheet가
        # index 범위 밖이면 append).
        if DEVIATION_SHEET_NAME in wb.sheetnames:
            del wb[DEVIATION_SHEET_NAME]
        wb.create_sheet(DEVIATION_SHEET_NAME, index=3)
    else:
        # backward-compat — spec wb 자체 베이스. audit 비권장.
        warnings.append(
            "[spec-sitr] SwITR 템플릿 미제공 — spec wb 베이스 fallback. "
            "시트 구성이 레퍼런스(Cover/History/1.Test Summary/2.Deviation/3.Test Log)"
            "와 다를 수 있음 — config sitr_template 등록 권장"
        )
        wb = spec_wb
        if spec_name != LOG_SHEET_NAME and LOG_SHEET_NAME not in wb.sheetnames:
            spec_ws.title = LOG_SHEET_NAME
        log_ws = spec_ws
        if DEVIATION_SHEET_NAME not in wb.sheetnames:
            wb.create_sheet(DEVIATION_SHEET_NAME)

    # spec 컬럼 레이아웃 동적 산출 (반드시 _write_log_headers 전 — r3 'Related ID'를
    # 'Actual Result'로 덮어쓰므로 이후 위치 식별 불가).
    layout = _detect_spec_layout(log_ws, warnings)

    # 헤더 추가 (Actual/Pass-Fail/Log).
    _write_log_headers(log_ws, warnings, layout=layout)

    # Actual 서브헤더를 Expected 서브헤더(Param N)와 동일하게 미러 — `_write_log_headers`
    # 는 'ActR[i]'를 stamp하나, 회사 PV 레퍼런스는 Expected와 동일한 'Param N' 라벨을
    # 사용(Actual=Expected 1:1 미러). `_apply_actual_result_style`는 _style만 복제(값 보존)
    # 라 이 값이 유지된다. expected_start~ 라벨이 비면 ActR[i] 유지(fallback).
    for _i in range(layout.actual_max):
        _exp_label = log_ws.cell(SUBHEADER_ROW, layout.expected_start + _i).value
        if _exp_label not in (None, ""):
            safe_write(log_ws, SUBHEADER_ROW, layout.actual_start + _i, _exp_label)

    # anchor 스캔 → 함수 블록 (이식된 '3.Test Log' 기준). SwIT 전용 스캐너 —
    # SwITS PV는 B열 TC_ID 세로병합 + C열 iteration index 구조라 SwUTS
    # `_scan_spec_blocks`(B=숫자 index)로는 anchor 0개 (2026-06-19 fix).
    blocks = _scan_swit_spec_blocks(log_ws, layout=layout)
    fn_iter_map = _build_swit_fn_iteration_map(session)

    # spec 자체가 함수명↔SwUFn 권위 소스 — agg 주입 (_collect_coverage_gaps 우선 사용).
    # SwITS spec엔 별도 함수명 열이 없어 unit(SwUFn_NNNN)이 곧 SwUFn id (num은
    # 2026-06-19부터 env_name 정렬용 정규화 키라 SwUFn 파생 불가 → unit 직접 사용).
    spec_name_to_swufn: dict[str, str] = {}
    for _blk in blocks:
        _u = (_blk.get("unit") or "").strip()
        if _u:
            spec_name_to_swufn.setdefault(_u, _u)
    agg["spec_name_to_swufn"] = spec_name_to_swufn

    # 2.Deviation B열 TC-id 역매핑 (Unit → TC-id).
    unit_to_tc = _build_swit_unit_to_tc_map(blocks)

    fill_stats = _fill_actual_and_result(
        log_ws, blocks, fn_iter_map, asil_map, warnings, layout=layout,
    )

    # Actual Result 열 서식 적용 (Expected 1:1 미러).
    _restyled = _apply_actual_result_style(log_ws, layout=layout)

    # Test Log 상단 2행 색범례 offset — 회사 PV 레퍼런스 정합 (헤더 r5/데이터 r7).
    # **모든 Test Log 쓰기 완료 후** (이후 행 좌표 +2). 비치명 — 실패해도 진행.
    _legend_ok = _apply_swit_legend_offset(log_ws, layout, out_warnings=warnings)

    summary: dict[str, Any] = {
        "builder": "swit-spec-based" if template_xlsm_bytes is not None
        else "swit-spec-based-fallback",
        "spec_sheet": spec_name,
        "spec_sha256_12": spec_sha256_12,
        "spec_layout": {
            "detected": layout.detected,
            "expected_start": layout.expected_start,
            "actual_start": layout.actual_start,
            "actual_max": layout.actual_max,
            "pass_fail": layout.pass_fail,
            "pass_total": layout.pass_total,
            "log_data": layout.log_data,
            "iter_index": layout.iter_index,
        },
        "build_timestamp": meta.build_timestamp,
        "environments": len(session.environments),
        "total": agg["total"],
        "tested": agg["tested"],
        "passed": agg["passed"],
        "failed": agg["failed"],
        "spec_function_blocks": fill_stats["functions"],
        "matched_functions": fill_stats["matched_fn"],
        "unmatched_functions": fill_stats["unmatched_fn"],
        "fn_pass": fill_stats["fn_pass"],
        "fn_fail": fill_stats["fn_fail"],
        "fn_na": fill_stats["fn_na"],
        "iteration_rows": fill_stats["iterations"],
        "iter_pass": fill_stats["iter_pass"],
        "iter_fail": fill_stats["iter_fail"],
        "iter_na": fill_stats["iter_na"],
        "actual_from_expected": fill_stats["actual_from_expected"],
        "actual_from_vcast": fill_stats["actual_from_vcast"],
        "actual_missing": fill_stats["actual_missing"],
        "actual_cells_restyled": _restyled,
        "legend_offset_applied": _legend_ok,
    }

    # spec 매칭 0건 — TC-id 매칭 실패 가능성 (SwITS↔레퍼런스 불일치) 정직 보고.
    if fill_stats["functions"] > 0 and fill_stats["matched_fn"] == 0:
        warnings.append(
            "[spec-sitr] ⚠️ spec 함수 블록 ↔ VectorCAST SwUFn 매칭 0건 — SwITS spec "
            "TC_ID(예 'SwITC_0101_01')와 VectorCAST test_case(예 'SwITC_SwUFn_0101"
            ".001')의 숫자 키 불일치 의심. Pass/Fail 전부 N/A. 정합 미세조정 필요 "
            "(후속 라운드)."
        )

    # 보조 시트 채움 (Cover / 1.Test Summary / 2.Deviation / History).
    if template_xlsm_bytes is not None:
        _fill_swit_aux_sheets(
            wb, meta, agg, summary, unit_to_tc, warnings,
        )
    else:
        # backward-compat — Cover label stamp (best-effort).
        cover_ws = next(
            (wb[n] for n in wb.sheetnames if n.lower() == "cover"), None,
        )
        if cover_ws is not None:
            _write_cover(cover_ws, meta, out_warnings=warnings)
        dev_ws = next(
            (wb[n] for n in wb.sheetnames if "deviation" in n.lower()), None,
        )
        if dev_ws is not None:
            n = _write_swit_spec_deviation(
                dev_ws, agg, warnings,
                empty=bool(getattr(meta, "deviation_empty", False)),
                unit_to_tc=unit_to_tc,
            )
            summary["deviation_cases_written"] = n

    if LOG_SHEET_NAME in wb.sheetnames:
        summary["output_sheet_order"] = list(wb.sheetnames)

    out = io.BytesIO()
    wb.save(out)
    # save 직후 무결성 검증 + 손상 시 gc 후 1회 재시도 (거대 '3.Test Log' 메모리 압박).
    _save_ok, _save_err = verify_xlsx_integrity(out.getvalue())
    if not _save_ok:
        import gc as _gc
        _gc.collect()
        out = io.BytesIO()
        wb.save(out)
        summary["save_retried"] = True
        warnings.append(
            f"[spec-sitr] 첫 save 무결성 실패({_save_err}) → gc 후 재 save (거대 "
            "Test Log 메모리 압박 의심). 최종 상태는 아래 무결성 검증으로 재확인, "
            "동시 빌드 수 축소 권장."
        )
    if wb is not spec_wb:
        spec_wb.close()
    wb.close()

    # 빈 양식 셀 self-closing 정규화 (openpyxl 3.1.5 비효율 제거 — Test Log 비대/손상 완화).
    _compacted, _n_compact = compact_empty_styled_cells(out.getvalue())
    if _n_compact:
        out = io.BytesIO(_compacted)
        summary["empty_cells_compacted"] = _n_compact
        warnings.append(
            f"[spec-sitr] 빈 양식 셀 {_n_compact}개 self-closing 정규화 "
            "(openpyxl 3.1.5 비효율 — Test Log 비대/손상 완화)"
        )

    # 외부링크 파트 + 외부참조 defined name 제거 (Excel 연결 경고 차단).
    _sanitized, _ext_removed = sanitize_xlsm_external_links(out.getvalue())
    if _ext_removed:
        out = io.BytesIO(_sanitized)
        summary["external_links_stripped"] = _ext_removed
        warnings.append(
            f"[spec-sitr] 템플릿 외부링크 파트 {_ext_removed}건 + 외부참조 defined "
            "name 제거 (Excel 연결 경고 차단)"
        )
    out.seek(0)

    # 최종 무결성 검증 (거대 Test Log XML 잘림 손상 배포 차단).
    _ok, _err = verify_xlsx_integrity(out.getvalue())
    summary["integrity_check"] = "ok" if _ok else f"FAILED: {_err}"
    if not _ok:
        warnings.append(
            f"[spec-sitr] ⚠️ 산출물 무결성 검증 실패: {_err} — 거대 '3.Test Log' "
            "save 가 메모리 압박으로 잘렸을 수 있음. 재생성 권장 (동시 빌드 축소 시 회복)."
        )
    out.seek(0)

    if meta.doc_filename_pattern:
        filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version, date=short_date(meta.test_date),
        )
    else:
        # 레퍼런스 파일명 패턴:
        # `(KJPDS02_PV_SwITR) Software Integration Test Result_v{ver}_{date}_R.xlsm`
        filename = (
            f"({meta.project_id}_PV_SwITR) Software Integration Test Result_"
            f"v{meta.release_sw_version}_{short_date(meta.test_date)}_R.xlsm"
        )

    return SwitSitrBuildResult(
        ok=True,
        xlsm_io=out,
        filename=filename,
        warnings=warnings,
        vba_macros_preserved=template_has_vba,
        summary=summary,
    )


__all__ = [
    "build_sitr_from_spec",
    "LOG_SHEET_NAME",
    "DEVIATION_SHEET_NAME",
]
