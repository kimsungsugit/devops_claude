"""Excel artifact comparison helpers for template fidelity validation.

The existing ``excel_compare`` module is intentionally value-only.  SwUT/SwIT
audit artifacts also need merge/style/formula checks, so this module keeps the
stricter comparison separate from the older API.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook


@dataclass
class ExcelArtifactDiff:
    sheet: str
    category: str
    coordinate: str
    expected: str
    actual: str

    def to_dict(self) -> dict[str, str]:
        return {
            "sheet": self.sheet,
            "category": self.category,
            "coordinate": self.coordinate,
            "expected": self.expected,
            "actual": self.actual,
        }


@dataclass
class ExcelArtifactCompareReport:
    ok: bool
    summary: dict[str, int] = field(default_factory=dict)
    diffs: list[ExcelArtifactDiff] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "summary": self.summary,
            "diffs": [diff.to_dict() for diff in self.diffs],
            "warnings": self.warnings,
        }


def _load(source: bytes | bytearray | io.BytesIO | str | Path | Workbook) -> Workbook:
    if isinstance(source, Workbook):
        return source
    if isinstance(source, bytes | bytearray):
        return load_workbook(io.BytesIO(source), data_only=False, keep_vba=False)
    if isinstance(source, io.BytesIO):
        pos = source.tell()
        source.seek(0)
        try:
            return load_workbook(source, data_only=False, keep_vba=False)
        finally:
            source.seek(pos)
    return load_workbook(Path(source), data_only=False, keep_vba=False)


def _blank_equiv(value: Any) -> Any:
    return "" if value is None else value


def _value_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _color_sig(color: Any) -> tuple[Any, ...]:
    if color is None:
        return ("", "", "", "")
    return (
        getattr(color, "type", ""),
        getattr(color, "rgb", ""),
        getattr(color, "indexed", ""),
        getattr(color, "theme", ""),
    )


def _side_sig(side: Any) -> tuple[Any, ...]:
    if side is None:
        return ("", ("", "", "", ""))
    return (getattr(side, "style", ""), _color_sig(getattr(side, "color", None)))


def _font_sig(cell: Any) -> tuple[Any, ...]:
    font = cell.font
    return (
        font.name,
        float(font.sz) if font.sz is not None else None,
        bool(font.bold),
        bool(font.italic),
        font.underline,
        _color_sig(font.color),
    )


def _border_sig(cell: Any) -> tuple[Any, ...]:
    border = cell.border
    return (
        _side_sig(border.left),
        _side_sig(border.right),
        _side_sig(border.top),
        _side_sig(border.bottom),
    )


def _fill_sig(cell: Any) -> tuple[Any, ...]:
    fill = cell.fill
    return (
        fill.fill_type,
        _color_sig(fill.fgColor),
        _color_sig(fill.bgColor),
    )


def _alignment_sig(cell: Any) -> tuple[Any, ...]:
    alignment = cell.alignment
    return (
        alignment.horizontal,
        alignment.vertical,
        bool(alignment.wrap_text),
        bool(alignment.shrink_to_fit),
    )


def _cell_is_default_blank(cell: Any) -> bool:
    return (
        _blank_equiv(cell.value) == ""
        and not cell.has_style
        and not cell.hyperlink
        and not cell.comment
    )


def _sheet_names(reference: Workbook, generated: Workbook, sheets: list[str] | None) -> list[str]:
    if sheets is not None:
        return [name for name in sheets if name in reference.sheetnames and name in generated.sheetnames]
    return [name for name in reference.sheetnames if name in generated.sheetnames]


def compare_excel_artifacts(
    reference_source: bytes | bytearray | io.BytesIO | str | Path | Workbook,
    generated_source: bytes | bytearray | io.BytesIO | str | Path | Workbook,
    *,
    sheets: list[str] | None = None,
    max_diffs_per_category: int = 50,
    compare_values: bool = True,
    compare_styles: bool = True,
    blank_equivalent: bool = True,
) -> ExcelArtifactCompareReport:
    """Compare workbook values, formulas, merges, and core cell styles.

    ``reference_source`` is the expected template/reference artifact.  The
    generated workbook is compared against it without mutating either workbook.
    """
    reference = _load(reference_source)
    generated = _load(generated_source)
    summary: dict[str, int] = {}
    diffs: list[ExcelArtifactDiff] = []
    warnings: list[str] = []

    def add(sheet: str, category: str, coordinate: str, expected: Any, actual: Any) -> None:
        summary[category] = summary.get(category, 0) + 1
        if summary[category] <= max_diffs_per_category:
            diffs.append(ExcelArtifactDiff(
                sheet=sheet,
                category=category,
                coordinate=coordinate,
                expected=_value_text(expected),
                actual=_value_text(actual),
            ))

    missing_ref = [name for name in (sheets or generated.sheetnames) if name not in reference.sheetnames]
    missing_gen = [name for name in (sheets or reference.sheetnames) if name not in generated.sheetnames]
    for name in missing_ref:
        add(name, "sheet", name, "present in generated", "missing in reference")
    for name in missing_gen:
        add(name, "sheet", name, "present in reference", "missing in generated")

    for sheet_name in _sheet_names(reference, generated, sheets):
        ref_ws = reference[sheet_name]
        gen_ws = generated[sheet_name]
        if ref_ws.max_row != gen_ws.max_row or ref_ws.max_column != gen_ws.max_column:
            add(
                sheet_name,
                "dimension",
                "used-range",
                f"{ref_ws.max_row}x{ref_ws.max_column}",
                f"{gen_ws.max_row}x{gen_ws.max_column}",
            )

        ref_merges = {str(rng) for rng in ref_ws.merged_cells.ranges}
        gen_merges = {str(rng) for rng in gen_ws.merged_cells.ranges}
        for merge_ref in sorted(ref_merges - gen_merges):
            add(sheet_name, "merge", merge_ref, "merged", "missing")
        for merge_gen in sorted(gen_merges - ref_merges):
            add(sheet_name, "merge", merge_gen, "missing", "merged")

        max_row = max(ref_ws.max_row, gen_ws.max_row)
        max_col = max(ref_ws.max_column, gen_ws.max_column)
        for row_idx in range(1, max_row + 1):
            ref_height = ref_ws.row_dimensions[row_idx].height
            gen_height = gen_ws.row_dimensions[row_idx].height
            if ref_height != gen_height:
                add(sheet_name, "dimension", str(row_idx), ref_height, gen_height)
        for col_idx in range(1, max_col + 1):
            ref_letter = get_column_letter(col_idx)
            ref_width = ref_ws.column_dimensions[ref_letter].width
            gen_width = gen_ws.column_dimensions[ref_letter].width
            if ref_width != gen_width:
                add(sheet_name, "dimension", ref_letter, ref_width, gen_width)

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                ref_cell = ref_ws.cell(row_idx, col_idx)
                gen_cell = gen_ws.cell(row_idx, col_idx)
                if _cell_is_default_blank(ref_cell) and _cell_is_default_blank(gen_cell):
                    continue

                ref_value = _blank_equiv(ref_cell.value) if blank_equivalent else ref_cell.value
                gen_value = _blank_equiv(gen_cell.value) if blank_equivalent else gen_cell.value
                if compare_values and ref_value != gen_value:
                    category = "formula" if _is_formula(ref_value) or _is_formula(gen_value) else "value"
                    add(sheet_name, category, ref_cell.coordinate, ref_value, gen_value)

                if not compare_styles:
                    continue
                style_checks = (
                    ("font", _font_sig(ref_cell), _font_sig(gen_cell)),
                    ("border", _border_sig(ref_cell), _border_sig(gen_cell)),
                    ("fill", _fill_sig(ref_cell), _fill_sig(gen_cell)),
                    ("alignment", _alignment_sig(ref_cell), _alignment_sig(gen_cell)),
                )
                for category, expected, actual in style_checks:
                    if expected != actual:
                        add(sheet_name, category, ref_cell.coordinate, expected, actual)

    for category, count in summary.items():
        if count > max_diffs_per_category:
            warnings.append(
                f"{category}: {count} diffs found; stored first {max_diffs_per_category}."
            )

    return ExcelArtifactCompareReport(ok=not summary, summary=summary, diffs=diffs, warnings=warnings)


__all__ = [
    "ExcelArtifactCompareReport",
    "ExcelArtifactDiff",
    "compare_excel_artifacts",
]
