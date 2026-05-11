"""SUTR Deviation rationale 자동 생성기.

ISO 26262 단위테스트 산출물에서 가장 빈번한 deviation 유형인
**Divide-by-zero**에 대해 다음 자동 분석을 수행한다.

1. SUTR `Deviation` 시트의 Issue 컬럼에서 'Divide by zero' 패턴 검출
2. Issue 텍스트에서 분모 변수 추출 (①, ②, ... 또는 줄바꿈 단위)
3. 각 분모 변수에 대해 C 소스 디렉토리에서 `#define` 매크로 추적
4. 매크로 값을 enumerate해서 "0 불가" 자동 rationale 문자열 생성

## ISO 26262 Tool Qualification 사용 제약 (의무 명시)

본 도구의 출력은 **자동 생성 draft** 이며 manual review evidence와
동등하게 취급되지 않는다. ASIL 등급별 사용 범위:

- **ASIL A** : draft 보조용. 사람 reviewer가 결과 검토 후 승인 시 OK.
- **ASIL B/C/D** : **단독 evidence 사용 금지**. 사람 reviewer가
  본 도구가 추적하지 못하는 함수 호출 chain / 복합 식 / 외부 입력
  range 등 모든 차원의 분석을 수동으로 완료해야 한다. 본 도구 출력은
  reviewer가 놓친 매크로 후보 발견 목적 (보조 검출기)으로만 사용.

## 알려진 한계 (사용자 인지 의무)

1. **1-hop 매크로 추적만**: 함수 호출 chain (예: `var = func()` 후
   `func()` 내부의 매크로) 미추적. SUTR 원본에서 사람이 작성한
   rationale의 추적 깊이를 따라가지 못함.
2. **복합 식 분모 분석 불가**: `(a - b)` / `(2*a - b - c)` 같은
   차이/곱 분모는 individual 변수 enum만 수행, "차이값=0" 분석은 미지원.
3. **줄당 1 변수**: Issue 텍스트는 줄당 1 변수 가정 (`① a`, `② b`).
   `① a - b` 같은 줄에서는 `a`만 enum.
4. **포맷 의존**: Hyundai/Mobis 스타일 헤더(`Test Case ID`, `Issue`,
   `Deviation`, `Status`)를 가정. 다른 회사 산출물에서 헤더 미발견 시
   결과에 `ConsistencyIssue(parse_error)`로 명시 보고.
"""
from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass, field
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]


# ---------------------------------------------------------------------------
# Patterns
# ---------------------------------------------------------------------------

# Issue 텍스트의 패턴 분류 키워드
_PATTERN_DIVIDE = re.compile(
    r"(divide\s*by\s*zero|분모\s*자리|divide\s*[-_]?\s*zero)",
    re.IGNORECASE,
)

# 원형 숫자 마커 (Issue 텍스트의 ① ② ③ ④ ⑤ … 항목 구분자)
_CIRCLED_DIGITS = "①②③④⑤⑥⑦⑧⑨⑩"

# 식별자 휴리스틱: 보통 Hyundai/Mobis 스타일 prefix (sNg / u8g / s16g / u16g / s32s / u32s)
_TYPED_IDENT_RE = re.compile(
    r"\b((?:[us][0-9]+[gst])_[A-Za-z_][A-Za-z0-9_]+|[A-Za-z_][A-Za-z0-9_]+)\b"
)

# C 소스의 #define 매크로
def _compact_row(row: tuple) -> list[tuple[int, str]]:
    """머지셀(None leading)을 걸러서 (column_index, value) 페어만 반환."""
    return [
        (i, str(c).strip()) for i, c in enumerate(row)
        if c is not None and str(c).strip()
    ]


_DEFINE_RE = re.compile(
    r"#define\s+([A-Za-z_][A-Za-z0-9_]*)\s+\(?\s*\(?\s*\(?\s*"
    r"\(?\s*[A-Za-z_][A-Za-z0-9_]*\s*\)?\s*"
    r"\(?\s*(-?\d+\.?\d*)\s*[Uu]?\s*\)?\s*\)?\s*\)?\s*\)?"
)

# 할당 형태: var = MACRO 또는 var = number
_ASSIGN_RE = re.compile(
    r"\b([A-Za-z_][A-Za-z0-9_]*)\s*=\s*([A-Za-z_][A-Za-z0-9_]*|\(\s*[A-Za-z_][A-Za-z0-9_]*\s*\)?\s*-?\d+)"
)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class MacroEvidence:
    macro_name: str
    value: str  # raw 값 (정수/실수 문자열)
    file: str = ""
    line: int = 0


@dataclass
class DenominatorAnalysis:
    name: str
    macros: list[MacroEvidence] = field(default_factory=list)
    is_zero_possible: bool | None = None  # True/False/None(unknown)
    rationale: str = ""
    # ISO 26262 Tool Qualification — 모든 rationale은 자동 생성물이며
    # 사람 reviewer 승인 전까지 evidence로 사용 불가. frontend가 표 형식으로
    # 렌더링하더라도 이 플래그를 통해 audit가 자동/수동을 구분할 수 있다.
    is_auto_generated: bool = True
    needs_review: bool = True


@dataclass
class DeviationCase:
    tc_id: str
    tc_no: str  # "TC2", "TC1~5", "TC10" 등
    pattern: str  # "divide_by_zero" | "unknown"
    issue_text: str
    denominators: list[DenominatorAnalysis] = field(default_factory=list)
    auto_rationale: str = ""

    def to_dict(self, src_root: str | None = None) -> dict[str, Any]:
        """JSON 직렬화. src_root 지정 시 매크로 파일 경로를 상대화 (X8: 절대경로 노출 방지)."""
        return {
            "tc_id": self.tc_id,
            "tc_no": self.tc_no,
            "pattern": self.pattern,
            "issue_text": self.issue_text,
            "denominators": [
                {
                    "name": d.name,
                    "macros": [
                        {
                            "macro": m.macro_name,
                            "value": m.value,
                            "file": _normalize_path(m.file, src_root),
                            "line": m.line,
                        }
                        for m in d.macros
                    ],
                    "is_zero_possible": d.is_zero_possible,
                    "rationale": d.rationale,
                    "is_auto_generated": d.is_auto_generated,
                    "needs_review": d.needs_review,
                }
                for d in self.denominators
            ],
            "auto_rationale": self.auto_rationale,
        }


def _normalize_path(fpath: str, src_root: str | None) -> str:
    """X8: 외부 응답에서 절대 파일시스템 경로 노출 방지.

    src_root 지정 시 상대 경로, 미지정 시 basename만 반환.
    """
    if not fpath:
        return ""
    try:
        if src_root:
            return os.path.relpath(fpath, src_root).replace("\\", "/")
        return os.path.basename(fpath)
    except (ValueError, OSError):
        return os.path.basename(fpath)


# ---------------------------------------------------------------------------
# C source scanner
# ---------------------------------------------------------------------------

class CSourceIndex:
    """C 소스 디렉토리를 1회 스캔해서 #define 및 할당 인덱스를 만든다.

    Args:
        src_root: 스캔 대상 디렉토리.
        allowed_roots: 신뢰 가능한 root prefix 화이트리스트 (e.g.
            ``["D:/Project/Ados/", "D:/Project/devops/"]``). 본 인덱스가
            API endpoint를 통해 외부 입력 path로 호출될 경우, 시스템
            디렉토리(`/etc/`, `C:\\Windows\\`) / UNC(`\\\\server\\share`)
            traversal을 막기 위해 endpoint 단에서 지정해야 한다.
            ``None`` 이면 비검증 (CLI/내부 호출 가정).

    Raises:
        ValueError: ``allowed_roots`` 지정 시 ``src_root``가 그 어느
            prefix에도 속하지 않으면 즉시 거부.
    """

    def __init__(self, src_root: str, allowed_roots: list[str] | None = None):
        # path traversal 방어 (deep-reviewer 시나리오 4): endpoint 노출 시점에
        # allowed_roots 전달 의무.
        if allowed_roots is not None:
            abs_src = os.path.abspath(src_root).replace("\\", "/").lower()
            ok = False
            for root in allowed_roots:
                abs_root = os.path.abspath(root).replace("\\", "/").lower()
                if abs_src.startswith(abs_root.rstrip("/") + "/") or abs_src == abs_root:
                    ok = True
                    break
            if not ok:
                raise ValueError(
                    f"src_root '{src_root}' is not within allowed_roots {allowed_roots}"
                )
        self.src_root = src_root
        self.macros: dict[str, list[MacroEvidence]] = {}
        self.assignments: dict[str, set[str]] = {}  # var_name -> {macro_name, ...}
        self._scanned = False

    def scan(self) -> None:
        if self._scanned:
            return
        self._scanned = True
        if not os.path.isdir(self.src_root):
            return
        # S1 fix: followlinks=False — symlink 무한 루프/외부 경로 traversal 방지.
        for dirpath, _, files in os.walk(self.src_root, followlinks=False):
            for fname in files:
                if not fname.lower().endswith((".c", ".h")):
                    continue
                fpath = os.path.join(dirpath, fname)
                try:
                    with open(fpath, encoding="utf-8", errors="ignore") as f:
                        for lineno, line in enumerate(f, 1):
                            # #define MACRO ( ( type ) ( value ) )
                            m = _DEFINE_RE.search(line)
                            if m:
                                name, val = m.group(1), m.group(2)
                                self.macros.setdefault(name, []).append(
                                    MacroEvidence(macro_name=name, value=val,
                                                  file=fpath, line=lineno)
                                )
                            # var = MACRO_NAME or var = ( type )0
                            am = _ASSIGN_RE.search(line)
                            if am:
                                lhs, rhs = am.group(1), am.group(2)
                                # rhs may be "( type )-1" → extract identifier만
                                rhs_id = re.search(r"\b([A-Za-z_][A-Za-z0-9_]*)\b", rhs)
                                if rhs_id:
                                    self.assignments.setdefault(lhs, set()).add(rhs_id.group(1))
                except (OSError, UnicodeDecodeError):
                    continue

    def find_macros_for(self, var_name: str) -> list[MacroEvidence]:
        """변수 이름에 직접 #define이 있거나, 할당된 매크로의 #define을 모두 반환."""
        self.scan()
        found = list(self.macros.get(var_name, []))
        for macro_name in self.assignments.get(var_name, set()):
            found.extend(self.macros.get(macro_name, []))
        # 변수명 자체가 매크로 prefix를 가지는 경우 prefix-stripped 검색
        # 예: s16g_SysOptCtrl_OverPos → s16s_FRONT_OVER_POS / s16s_REAR_OVER_POS
        # → 휴리스틱: var_name 토큰 일부가 들어간 매크로 찾기
        base_tokens = [t for t in re.split(r"[_]", var_name) if len(t) > 3 and t.isalpha()]
        if base_tokens and not found:
            for macro_name, evidences in self.macros.items():
                tokens = re.split(r"[_]", macro_name)
                if any(bt.upper() in (t.upper() for t in tokens) for bt in base_tokens):
                    found.extend(evidences)
        # de-dup
        seen: set[tuple[str, str, str, int]] = set()
        unique: list[MacroEvidence] = []
        for e in found:
            key = (e.macro_name, e.value, e.file, e.line)
            if key not in seen:
                seen.add(key)
                unique.append(e)
        return unique


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _extract_denominators(issue_text: str) -> list[str]:
    """Issue 텍스트에서 분모 변수명 후보를 추출."""
    candidates: list[str] = []
    # 원형 숫자 이후 줄 또는 줄 단위 라인에서 식별자 추출.
    for line in issue_text.splitlines():
        line = line.strip()
        # 원형 숫자 마커 제거
        for c in _CIRCLED_DIGITS:
            line = line.replace(c, " ")
        # 'variables', '분모' 같은 안내 텍스트가 들어있는 줄은 제외 (헤더성)
        if not line or any(kw in line for kw in ("Divide", "분모", "변수")):
            continue
        # ( ... ) 같은 표현식: 가장 처음 등장 식별자만 후보
        for m in _TYPED_IDENT_RE.finditer(line):
            ident = m.group(1)
            if ident in ("U", "S", "U8", "S8", "U16", "S16", "U32", "S32"):
                continue
            candidates.append(ident)
            break  # 한 줄 당 1개만
    # 중복 제거, 순서 유지
    return list(dict.fromkeys(candidates))


def _parse_tc_label(label: str) -> tuple[str, str]:
    """'SwUTC_SwUFn_407 (TC2)' → ('SwUTC_SwUFn_407', 'TC2')."""
    m = re.match(r"^(SwUTC_SwUFn_\d+)\s*\(?\s*(TC[\d~]+)?\s*\)?", label.strip())
    if not m:
        return (label.strip(), "")
    return (m.group(1), m.group(2) or "")


_INIT_MACRO_RE = re.compile(
    r"_(CLR|CLEAR|RESET|NULL|INIT|ZERO|DEFAULT|OFF|INV|INVALID|NA|NONE|MIN)$",
    re.IGNORECASE,
)
# ISO 26262 Tool Qualification 라벨 — 자동 생성 산출물임을 명시 (ISO1).
_AUTO_DRAFT_PREFIX = "[AUTO-GENERATED DRAFT — 검토자 승인 필요] "


def _is_init_macro(macro: MacroEvidence) -> bool:
    """`s16g_CLR`, `*_RESET` 같은 초기화/clear 매크로 식별.

    이런 매크로는 보통 0 또는 sentinel 값을 가지며, 변수에 reset 시점에만
    대입되고 산술 연산 시점의 가능 값 enumerate에는 포함되면 안 됨.
    """
    return bool(_INIT_MACRO_RE.search(macro.macro_name))


def _build_rationale(name: str, macros: list[MacroEvidence]) -> tuple[bool | None, str]:
    """매크로 evidence 기반 "0 불가" 판정 + rationale 문자열 생성.

    분류:
      - value_macros: 산술 시점 값으로 사용되는 정상 매크로 (예: FRONT_OVER_POS=2472)
      - init_macros: clear/reset 시점 매크로 (예: s16g_CLR=0) — evidence로만 보존
    """
    if not macros:
        return (None, f"{name}: 관련 매크로/할당 정의를 소스에서 찾지 못함 — 수동 검토 필요")

    value_macros = [m for m in macros if not _is_init_macro(m)]
    init_macros = [m for m in macros if _is_init_macro(m)]

    if not value_macros:
        init_names = ", ".join(m.macro_name for m in init_macros[:3])
        return (
            None,
            f"{name}: 초기화 매크로({init_names})만 발견 — 산술 시점 값 enumerate 불가, "
            f"수동 검토 필요"
        )

    has_zero = any(_is_zero_value(m.value) for m in value_macros)
    macro_list = ", ".join(f"{m.macro_name}={m.value}" for m in value_macros[:6])
    if len(value_macros) > 6:
        macro_list += f" (+{len(value_macros) - 6} more)"

    init_note = ""
    if init_macros:
        init_names = ", ".join(m.macro_name for m in init_macros[:2])
        init_note = f" (참고: 초기화 매크로 {init_names}는 산술 시점 값 분석에서 제외)"

    if has_zero:
        zero_macros = [m for m in value_macros if _is_zero_value(m.value)]
        zero_names = ", ".join(m.macro_name for m in zero_macros)
        return (
            True,
            f"{name}: 0 가능 — 산술 값 매크로 {zero_names} 중 0 존재 — "
            f"deviation 부적절, 수동 재검토 필요{init_note}"
        )

    rationale = (
        f"{name}: 입력 가능한 모든 경우 매크로 정의값({macro_list})으로 한정되며 "
        f"어느 것도 0이 아니므로 분모가 0이 될 수 없음 — Divide-by-zero 발생 불가{init_note}"
    )
    return (False, rationale)


def _is_zero_value(v: str) -> bool:
    try:
        return abs(float(v.strip())) < 1e-9
    except (ValueError, TypeError):
        return False


def _load_workbook(source: Any) -> Any:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for SwUT deviation generator")
    if isinstance(source, (bytes, bytearray)):
        return openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    if isinstance(source, str):
        return openpyxl.load_workbook(source, read_only=True, data_only=True)
    return source


def generate_deviation_rationales(
    sutr_source: Any,
    c_source_root: str | os.PathLike[str] | None = None,
    out_warnings: list[str] | None = None,
    allowed_roots: list[str] | None = None,
) -> list[DeviationCase]:
    """SUTR Deviation 시트를 분석해 자동 rationale을 생성.

    Args:
        sutr_source: SUTR xlsx (path / bytes / Workbook).
        c_source_root: C 소스 디렉토리 root. None이면 매크로 추적 생략.
        out_warnings: 호출자 전달 list. 파싱 실패(시트 누락 / 헤더 미발견 등) 시
            warning 문자열이 append됨. silent empty 방지 (deep-reviewer 시나리오 3).
        allowed_roots: ``CSourceIndex`` path traversal 방어용 화이트리스트.
            endpoint 노출 시 endpoint 단에서 의무 주입 (시나리오 4).

    Returns:
        DeviationCase 리스트. 빈 리스트는 (a) deviation 없음 또는 (b) 파싱 실패.
        구분이 필요하면 out_warnings 인자를 사용해 차이를 확인할 것.
    """
    wb = _load_workbook(sutr_source)

    deviation_sheet = None
    for name in wb.sheetnames:
        if name.lower() == "deviation":
            deviation_sheet = wb[name]
            break
    if deviation_sheet is None:
        if out_warnings is not None:
            out_warnings.append(
                "Deviation 시트가 워크북에 없음 — SUTR 포맷 검증 필요"
            )
        return []

    # Deviation 시트 파싱 — 헤더(Test Case ID/Issue/Deviation/Status) 행 위치 찾기
    rows = list(deviation_sheet.iter_rows(values_only=True))
    header_idx = None
    for i, r in enumerate(rows[:20]):
        pairs = _compact_row(r)
        if any(v == "Test Case ID" for _, v in pairs):
            header_idx = i
            break
    if header_idx is None:
        if out_warnings is not None:
            out_warnings.append(
                "Deviation 시트의 'Test Case ID' 헤더를 첫 20행에서 찾지 못함 — "
                "포맷 검증 필요 (Hyundai/Mobis 스타일 가정)"
            )
        return []

    src_index = (
        CSourceIndex(str(c_source_root), allowed_roots=allowed_roots)
        if c_source_root else None
    )

    cases: list[DeviationCase] = []
    for r in rows[header_idx + 1:]:
        pairs = _compact_row(r)
        if not pairs:
            continue
        label = pairs[0][1]

        # Q5 fix: Appendix / End of Document 진입 시 조기 탈출 (불필요한 행 순회 제거)
        if "Appendix" in label or "End of Document" in label:
            break

        # TC 행만 처리 (SwUTC_*)
        if not label.startswith("SwUTC_"):
            continue

        tc_id, tc_no = _parse_tc_label(label)
        # Issue 텍스트: TC 라벨 셀 다음 셀
        issue_text = pairs[1][1] if len(pairs) > 1 else ""

        pattern = "divide_by_zero" if _PATTERN_DIVIDE.search(issue_text) else "unknown"

        denom_analyses: list[DenominatorAnalysis] = []
        if pattern == "divide_by_zero":
            denominators = _extract_denominators(issue_text)
            for denom in denominators:
                macros = src_index.find_macros_for(denom) if src_index else []
                is_zero, rationale = _build_rationale(denom, macros)
                denom_analyses.append(DenominatorAnalysis(
                    name=denom, macros=macros,
                    is_zero_possible=is_zero, rationale=rationale,
                ))

        # 통합 rationale — ISO 26262 Tool Qualification 관점에서 자동 도구 출력은
        # manual review evidence와 동일 취급될 수 없으므로 명시적 draft 라벨 부착.
        if pattern == "divide_by_zero" and denom_analyses:
            parts = [d.rationale for d in denom_analyses]
            auto_rationale = (
                f"{_AUTO_DRAFT_PREFIX}[{tc_id} {tc_no}] Divide-by-zero deviation 자동 분석:\n  - "
                + "\n  - ".join(parts)
            )
        else:
            auto_rationale = ""

        cases.append(DeviationCase(
            tc_id=tc_id, tc_no=tc_no, pattern=pattern,
            issue_text=issue_text.strip(),
            denominators=denom_analyses, auto_rationale=auto_rationale,
        ))

    return cases


def generate_deviation_rationales_dict(
    sutr_source: Any,
    c_source_root: str | os.PathLike[str] | None = None,
    allowed_roots: list[str] | None = None,
) -> dict[str, Any]:
    """JSON 직렬화 가능한 형태로 반환.

    X8: c_source_root 기준으로 매크로 evidence 파일 경로 상대화.
    파싱 실패 시 ``parse_warnings`` 필드에 사유 명시 (deep-reviewer 시나리오 3).
    """
    warnings: list[str] = []
    cases = generate_deviation_rationales(
        sutr_source, c_source_root,
        out_warnings=warnings, allowed_roots=allowed_roots,
    )
    root = str(c_source_root) if c_source_root else None
    return {
        "ok": True,
        "deviation_count": len(cases),
        "pattern_counts": _count_by_pattern(cases),
        "cases": [c.to_dict(src_root=root) for c in cases],
        "parse_warnings": warnings,
        # ISO 26262 Tool Qualification 메타데이터
        "tool_qualification": {
            "evidence_class": "auto-generated draft",
            "asil_a_usage": "draft 보조용, 사람 reviewer 승인 후 사용 가능",
            "asil_b_c_d_usage": "단독 evidence 사용 금지 — manual review 의무",
            "limitations": [
                "1-hop 매크로 추적 (함수 호출 chain 미추적)",
                "복합 식 분모 분석 불가 (차이/곱 형태)",
                "줄당 1 변수 가정",
                "Hyundai/Mobis 포맷 의존",
            ],
        },
    }


def _count_by_pattern(cases: list[DeviationCase]) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in cases:
        out[c.pattern] = out.get(c.pattern, 0) + 1
    return out
