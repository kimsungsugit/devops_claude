"""SwUTS / SwITS xlsm parser (F6-A 라운드, 60차).

SwUTS / SwITS는 시험 명세 문서. 회사/버전별 4가지 양식이 라이브 분석에서 확인됨
(KJPDS02 SwUTS v1.01 / KJPDS02 SwITS v1.01 / HDPDM01 SUTS v3.01 /
HDPDM01 SITS v2.02). 양식별로 TC_ID 형식, header row 위치, 열 매핑이 모두 다름.

본 파서는 **양식 dispatch 없이 헤더 라벨 자동 매칭** 방식으로 동작:

  1. 시트명에 ``r"(Unit|Integration)\\s*Test\\s*Spec"`` 매칭되는 시트만 처리
     (KJPDS02 SwITS는 시트명 끝 공백 포함 — ``"3. SW Integration Test Spec "``).
  2. 시트의 첫 ~15 row 중 normalized cell text가 라벨 후보 set과 3개 이상 매칭
     되는 row를 header row로 인식.
  3. col mapping 추출 → data row를 ``SwUTSEntry`` 로 변환.
  4. 누락 col은 graceful skip — 빈 string 또는 None.

ISO 26262 Tool Qualification:
    - evidence_class: "auto-generated draft"
    - ASIL A: reviewer 검토 후 evidence 사용 가능
    - ASIL B/C/D: 단독 evidence 사용 금지 — manual review 의무

설계 원칙:
    - openpyxl read_only=True / data_only=True / keep_vba=False
      (매크로 무시, 메모리 최소화)
    - XLSM_MAX_BYTES = 64MB (zip bomb 방지)
    - 단일 시트 entry 0건 → 해당 시트 skip + parse_warnings emit
    - 모든 시트 entry 0건 → ok=False

사용:
    >>> from backend.services.swuts_excel_parser import parse_swuts_xlsm
    >>> with open(path, "rb") as f:
    ...     result = parse_swuts_xlsm(f.read())
    >>> result.by_tc_id["SwUTC_0121"].test_method
    'REQ'
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

try:
    from openpyxl import load_workbook  # type: ignore
    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - hook fail-safe
    load_workbook = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False


XLSM_MAX_BYTES = 64 * 1024 * 1024  # 64MB — DoS 방지
_TC_SHEET_RE = re.compile(r"(Unit|Integration)\s*Test\s*Spec", re.IGNORECASE)
_HEADER_SCAN_MAX_ROWS = 15  # 첫 15 row에서 header 자동 감지
_DATA_SCAN_MAX_ROWS = 2000  # data row max (KJPDS02 ~140 TC, 안전 마진)
_MAX_COL = 30  # 회사 양식 최대 컬럼 추정 (KJPDS02 SwUTS는 20+)
_MIN_HEADER_LABELS = 3  # header row 인식 임계

# TC_ID 내 함수 식별자 substring (SwUFn_NNNN). KJPDS02 SwUTS는 함수명을 별도
# Unit col에 두지만 HDPDM01 SUTS는 TC_ID 자체에 ``SwUTC_SwUFn_0101`` 형태로 포함.
_FUNCTION_ID_RE = re.compile(r"SwUFn_(\d+)")


@dataclass
class SwUTSEntry:
    """SwUTS / SwITS의 단위 TC 항목.

    여러 양식을 단일 dataclass로 통합:
      - KJPDS02 SwUTS (Description/Precondition 없음) — 두 field 빈 string
      - HDPDM01 SUTS/SITS (Description/Precondition 있음) — 모두 채움
      - sub-TC index가 별도 row로 표현되는 양식 — sub_index 필드 채움
    """
    tc_id: str                # 'SwUTC_0121' / 'SwUTC_SwUFn_0101' / 'SwITC_01'
    sheet_name: str = ""      # 출처 시트명 (debugging용)
    row_index: int = 0        # 1-based row (debugging용)
    sub_index: str = ""       # sub-TC index (예: '1', '2'). header row의 'Sequence' col 값
    unit_name: str = ""       # KJPDS02 SwUTS의 'Unit' col (s_safe_rotr) — Description fallback
    description: str = ""     # 'Interface : main -> s_System_I'
    precondition: str = ""
    test_method: str = ""     # 'REQ' / 'IFT' / 'FI'
    generation_method: str = ""  # 'ABV' / 'AEC' / 'AOR'
    function_id: str = ""     # 'SwUFn_0121' — TC_ID 또는 unit_name에서 substring 추출
    raw_inputs: dict[str, str] = field(default_factory=dict)
    # 'Inpt[0]': '0x0', ... — 디버깅/추적용 (raw input col 값)


@dataclass
class SwUTSParseResult:
    """SwUTS / SwITS xlsm 파싱 결과."""
    ok: bool
    entries: list[SwUTSEntry] = field(default_factory=list)
    parse_warnings: list[str] = field(default_factory=list)
    processed_sheets: list[str] = field(default_factory=list)

    @property
    def tc_ids(self) -> set[str]:
        return {e.tc_id for e in self.entries if e.tc_id}

    @property
    def by_tc_id(self) -> dict[str, SwUTSEntry]:
        """tc_id → 첫 매칭 entry (sub-TC가 여러 개면 첫 row 선택).

        SUTR Test Log stamp 시 caller는 normalize 후 lookup.
        """
        result: dict[str, SwUTSEntry] = {}
        for e in self.entries:
            if e.tc_id and e.tc_id not in result:
                result[e.tc_id] = e
        return result

    @property
    def by_function_id(self) -> dict[str, list[SwUTSEntry]]:
        """function_id (SwUFn_NNNN) → SwUTC_* entry list. 1:N 매핑.

        TC_ID 직접 매칭 실패 시 fallback chain에서 사용.
        """
        result: dict[str, list[SwUTSEntry]] = {}
        for e in self.entries:
            if e.function_id:
                result.setdefault(e.function_id, []).append(e)
        return result

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "tc_count": len(self.entries),
            "processed_sheets": self.processed_sheets,
            "entries": [
                {
                    "tc_id": e.tc_id,
                    "sheet": e.sheet_name,
                    "function_id": e.function_id,
                    "description": e.description[:200],
                }
                for e in self.entries[:50]
            ],
            "parse_warnings": self.parse_warnings,
            "tool_qualification": {
                "evidence_class": "auto-generated draft",
                "asil_a_usage": "reviewer 검토 후 evidence 사용 가능",
                "asil_b_c_d_usage": "단독 evidence 사용 금지 — manual review 의무",
                "format_assumption": (
                    "Hyundai/Mobis 양식 — TC 시트 자동 감지 + header row 라벨 매칭. "
                    "KJPDS02 SwUTS v1.01 / KJPDS02 SwITS v1.01 / "
                    "HDPDM01 SUTS v3.01 / HDPDM01 SITS v2.02 검증."
                ),
            },
        }


# 라벨 후보 — normalize (lower + space/punct 제거) 후 비교.
# 각 field의 정확한 라벨 변종을 1:N 매핑. 첫 매칭이 우선.
_LABEL_MAP: dict[str, tuple[str, ...]] = {
    "tc_id": ("tcid", "testcaseid"),
    "description": ("description",),
    "precondition": ("precondition",),
    "test_method": ("testmethod",),
    "generation_method": (
        "testcasegenerationmethod",
        "tcgenerationmethod",
        "generationmethod",
    ),
    "unit_name": ("unit", "name", "대응환경명", "componentunit"),
    "sub_index": ("sequence",),  # HDPDM01 SUTS의 sub-TC seq col (col 11)
}


def _normalize_label(s: str) -> str:
    """라벨 normalize — lower + non-alphanumeric 제거 (공백/'.'/'-'/'_' 등)."""
    if not s:
        return ""
    s = s.lower()
    return re.sub(r"[^a-z0-9가-힣]", "", s)


def _scan_header_row(
    rows: list[tuple[Any, ...]],
) -> tuple[int, dict[str, list[int]]] | None:
    """rows (첫 15 row, 0-based list).

    각 row에서 normalized cell text와 _LABEL_MAP 비교. 매칭 수 ≥ _MIN_HEADER_LABELS
    이면 header row. col mapping {field_name: [1-based_col_idx, ...]} 반환.

    **multiple col 지원**: 같은 라벨이 여러 col에 있으면 모두 저장 (HDPDM01
    SUTS의 'Test Case Generation Method' col 9 + col 12 케이스). data 추출
    시 첫 nonempty col 선택.

    Returns:
        (header_row_index_0based, col_map) or None
    """
    for row_idx, row in enumerate(rows):
        col_map: dict[str, list[int]] = {}
        for col_idx, val in enumerate(row, 1):
            if val is None:
                continue
            norm = _normalize_label(str(val))
            if not norm:
                continue
            for field_name, candidates in _LABEL_MAP.items():
                if norm in candidates:
                    col_map.setdefault(field_name, []).append(col_idx)
                    break
        if len(col_map) >= _MIN_HEADER_LABELS:
            return row_idx, col_map
    return None


def _extract_function_id(tc_id: str, unit_name: str) -> str:
    """TC_ID 또는 unit_name에서 SwUFn_NNNN substring 추출.

    1. tc_id 안에 SwUFn_NNNN 있으면 우선 (HDPDM01 SUTS의 'SwUTC_SwUFn_0101')
    2. unit_name에서 시도 (드물지만 안전)
    3. tc_id가 SwUTC_NNNN 형식이면 SwUFn_NNNN로 변환 시도 (KJPDS02 SwUTS)
       — 단, KJPDS02는 SwUTC index가 SwUFn과 1:1 대응 보장 안 됨. 추측만.
    """
    for src in (tc_id, unit_name):
        if not src:
            continue
        m = _FUNCTION_ID_RE.search(src)
        if m:
            return m.group(0)
    # KJPDS02 SwUTS fallback — SwUTC_NNNN → SwUFn_NNNN. 1:1 대응 가정.
    if tc_id and tc_id.startswith("SwUTC_"):
        m = re.match(r"SwUTC_(\d+)$", tc_id)
        if m:
            return f"SwUFn_{m.group(1)}"
    return ""


def _merge_subrow_into_entry(target: SwUTSEntry, sub: SwUTSEntry) -> None:
    """TC_ID 없는 sub-TC row 정보를 직전 TC entry로 merge.

    빈 field만 채움 (첫 sub-TC 값 우선). description/precondition/test_method/
    generation_method/sub_index 대상. raw_inputs는 첫 sub-TC 것만 보존.
    """
    if not target.description and sub.description:
        target.description = sub.description
    if not target.precondition and sub.precondition:
        target.precondition = sub.precondition
    if not target.test_method and sub.test_method:
        target.test_method = sub.test_method
    if not target.generation_method and sub.generation_method:
        target.generation_method = sub.generation_method
    if not target.sub_index and sub.sub_index:
        target.sub_index = sub.sub_index
    if not target.raw_inputs and sub.raw_inputs:
        target.raw_inputs = sub.raw_inputs


def _build_entry_from_row(
    row: tuple[Any, ...],
    col_map: dict[str, list[int]],
    sheet_name: str,
    row_idx_1based: int,
    last_unit_name: str,
) -> SwUTSEntry | None:
    """data row → SwUTSEntry. col_map의 col은 1-based, list (multiple col 지원).

    last_unit_name: 직전 row의 unit_name (sub-TC row가 unit_name col 비울 때 상속).

    같은 field에 multiple col 매핑 시 첫 nonempty cell 선택 (HDPDM01 SUTS의
    'Test Case Generation Method' col 9 + col 12 케이스 — 양식이 col 12에만
    값을 채움).
    """
    def _get_cell(field_name: str) -> str:
        cols = col_map.get(field_name) or []
        for col in cols:
            if col > len(row):
                continue
            v = row[col - 1]
            if v is None:
                continue
            s = str(v).strip()
            if s:
                return s
        return ""

    tc_id = _get_cell("tc_id")
    unit_name = _get_cell("unit_name") or last_unit_name
    description = _get_cell("description")
    precondition = _get_cell("precondition")
    test_method = _get_cell("test_method")
    generation_method = _get_cell("generation_method")
    sub_index = _get_cell("sub_index")

    # 모든 의미 있는 field 비어있으면 skip
    has_content = any([
        tc_id, description, test_method, generation_method,
    ])
    if not has_content:
        return None

    function_id = _extract_function_id(tc_id, unit_name)

    # raw_inputs — col_map에 매핑되지 않은 col들의 값 (debug용, 첫 12개만)
    raw_inputs: dict[str, str] = {}
    mapped_cols: set[int] = set()
    for cols in col_map.values():
        mapped_cols.update(cols)
    for col_idx, val in enumerate(row, 1):
        if col_idx in mapped_cols or val is None:
            continue
        s = str(val).strip()
        if s and len(raw_inputs) < 12:
            raw_inputs[f"col{col_idx}"] = s[:80]

    return SwUTSEntry(
        tc_id=tc_id,
        sheet_name=sheet_name,
        row_index=row_idx_1based,
        sub_index=sub_index,
        unit_name=unit_name,
        description=description[:500],
        precondition=precondition[:500],
        test_method=test_method[:50],
        generation_method=generation_method[:50],
        function_id=function_id,
        raw_inputs=raw_inputs,
    )


def parse_swuts_xlsm(
    xlsm_bytes: bytes,
    *,
    parse_warnings: list[str] | None = None,
) -> SwUTSParseResult:
    """SwUTS/SwITS xlsm bytes → SwUTSParseResult.

    Fail-safe:
        - openpyxl 미설치 → ok=False
        - XLSM_MAX_BYTES 초과 → ok=False
        - load_workbook 실패 → ok=False
        - 시트별 처리 중 예외 → 해당 시트 skip + parse_warnings + 다음 시트 진행
        - 모든 시트 entry 0건 → ok=False
    """
    warnings = parse_warnings if parse_warnings is not None else []

    if not _HAS_OPENPYXL:
        return SwUTSParseResult(
            ok=False,
            parse_warnings=warnings + ["openpyxl 미설치 — xlsm 파싱 불가"],
        )

    if not xlsm_bytes:
        return SwUTSParseResult(
            ok=False,
            parse_warnings=warnings + ["xlsm bytes 비어있음 — read 실패 추정"],
        )

    if len(xlsm_bytes) > XLSM_MAX_BYTES:
        return SwUTSParseResult(
            ok=False,
            parse_warnings=warnings + [
                f"xlsm 크기 {len(xlsm_bytes):,} > 한도 {XLSM_MAX_BYTES:,} — DoS 방지",
            ],
        )

    try:
        wb = load_workbook(  # type: ignore[misc]
            io.BytesIO(xlsm_bytes),
            read_only=True,
            data_only=True,
            keep_vba=False,
        )
    except Exception as exc:
        return SwUTSParseResult(
            ok=False,
            parse_warnings=warnings + [
                f"openpyxl load_workbook 실패 — {type(exc).__name__}: {exc}",
            ],
        )

    tc_sheets = [n for n in wb.sheetnames if _TC_SHEET_RE.search(n)]
    if not tc_sheets:
        try:
            wb.close()
        except Exception:
            pass
        return SwUTSParseResult(
            ok=False,
            parse_warnings=warnings + [
                f"TC 시트 미발견 — sheetnames={wb.sheetnames}. "
                "회사 양식이 'Unit Test Spec' / 'Integration Test Spec' 패턴 아님 추정",
            ],
        )

    all_entries: list[SwUTSEntry] = []
    processed: list[str] = []

    for sheet_name in tc_sheets:
        try:
            ws = wb[sheet_name]
            # 1. header scan용 첫 15 row 수집
            header_rows: list[tuple[Any, ...]] = []
            for r in ws.iter_rows(
                min_row=1, max_row=_HEADER_SCAN_MAX_ROWS,
                max_col=_MAX_COL, values_only=True,
            ):
                header_rows.append(r)

            scan = _scan_header_row(header_rows)
            if scan is None:
                warnings.append(
                    f"시트 {sheet_name!r}: header row 자동 감지 실패 "
                    f"(라벨 매칭 < {_MIN_HEADER_LABELS}개). skip."
                )
                continue
            header_row_0based, col_map = scan
            header_row_1based = header_row_0based + 1
            processed.append(sheet_name)

            # 2. data row 순회 — header row + 1부터 _DATA_SCAN_MAX_ROWS까지.
            # 양식 특성: TC_ID가 있는 row(메타) + 그 다음 sub-TC row들에
            # test_method / generation_method / description 등이 분산. TC_ID
            # 누락 row는 별도 entry 안 만들고 직전 entry로 merge (첫 sub-TC 값 우선).
            last_unit_name = ""
            current_entry: SwUTSEntry | None = None
            entries_this_sheet = 0
            for row_idx, row in enumerate(
                ws.iter_rows(
                    min_row=header_row_1based + 1,
                    max_row=header_row_1based + _DATA_SCAN_MAX_ROWS,
                    max_col=_MAX_COL,
                    values_only=True,
                ),
                start=header_row_1based + 1,
            ):
                entry = _build_entry_from_row(
                    row, col_map, sheet_name, row_idx, last_unit_name,
                )
                if entry is None:
                    continue

                # TC_ID 있으면 신규 entry 시작
                if entry.tc_id:
                    if entry.unit_name:
                        last_unit_name = entry.unit_name
                    current_entry = entry
                    all_entries.append(entry)
                    entries_this_sheet += 1
                else:
                    # TC_ID 없음 — sub-TC row. 직전 entry로 merge (빈 field만 채움).
                    if current_entry is not None:
                        _merge_subrow_into_entry(current_entry, entry)
                    else:
                        # F6 Round 5 NF1 fix: 첫 data row가 TC_ID 없는 sub-TC면
                        # 직전 메타 row 없어 silent drop. 양식 변종 가능성 — warning emit.
                        warnings.append(
                            f"시트 {sheet_name!r} row {row_idx}: TC_ID 없는 sub-TC "
                            "row지만 직전 메타 row 없음 — sub-TC 정보 drop"
                        )

            if entries_this_sheet == 0:
                warnings.append(
                    f"시트 {sheet_name!r}: data row 0건 — header 다음 row가 모두 빈 cell"
                )
        except Exception as exc:
            warnings.append(
                f"시트 {sheet_name!r} 처리 중 예외 — {type(exc).__name__}: {exc}"
            )

    try:
        wb.close()
    except Exception:
        pass

    if not all_entries:
        return SwUTSParseResult(
            ok=False,
            parse_warnings=warnings + [
                "모든 TC 시트에서 entry 0건 — 양식 불일치 추정"
            ],
            processed_sheets=processed,
        )

    return SwUTSParseResult(
        ok=True,
        entries=all_entries,
        parse_warnings=warnings,
        processed_sheets=processed,
    )


__all__ = [
    "SwUTSEntry",
    "SwUTSParseResult",
    "parse_swuts_xlsm",
    "XLSM_MAX_BYTES",
]
