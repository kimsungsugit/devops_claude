"""SwUT SUTR (Software Unit Test Result) v3.01 xlsm 빌더.

기존 v3.01 xlsm 템플릿을 BytesIO로 로드 → 5시트 셀 치환 → bytes 반환.
스타일/머지셀/매크로 보존 (template-copy 전략, keep_vba=True).

## 시트 매핑 (5 시트)

| 시트 | 출처 |
|------|------|
| Cover | meta (Doc ID / Project / ASIL / Author) |
| History | git log + 사람 입력 (다음 라운드, 본 라운드 placeholder) |
| Test Summary | SwUTSession 집계 (Total/Tested/Passed/Failed/Deviated/NotExec) |
| Deviation | swut_deviation_generator 자동 호출 (DRAFT 라벨) |
| Test Log | per-TC input/expected/actual/pass (각 환경 TestCaseData 통합) |

## ISO 26262 Tool Qualification
ASIL A 한정 draft. B/C/D는 manual review 의무.
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from typing import Any

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

from backend.services.array_collapse import build as _collapse_build
from backend.services.excel_template_utils import (
    build_release_history_row,
    dot_date,
    find_kv_row,
    has_vba_macros,
    inspect_vba_refs,
    mark_asil_a_function,
    mark_asil_b_function,
    mark_asil_c_function,
    mark_asil_d_function,
    mark_asil_qm_function,
    safe_write,
    short_date,
    stamp_cover_document_id,
    truncate_cell_text,
    validate_build_meta,
    validate_xlsx_template_bytes,
    write_label_or_mark,
    write_signature_block,
    write_value_after_label,
)
from backend.services.swut_builder_helpers import extract_warnings_from_session
from backend.services.swut_input_adapter import (
    SwUTSession,
    aggregate_session,
    compute_final_result,
)
from backend.services.swut_meta import BuildMetaBase

# 31차 W27: TC name에서 SwUFn_NNNN 함수 ID 추출 (Coverage builder와 동일 패턴).
# 위치는 import 블록 **아래** — 예전엔 import 사이에 끼어 뒤쪽 import 가 전부 E402 였다.
_TC_FN_RE = re.compile(r"(SwUFn_\d+)")


def _collapsed_value_cell(grp, data: dict) -> str:
    """DC-1: collapsed 다차원 배열 컬럼의 input/expected 셀 텍스트 (ArrayGroup.format_values).

    그룹 멤버(원본 변수명)를 ``data`` dict에서 조회 — 키 없으면 None(집계 제외).
    """
    return grp.format_values(lambda k: str(data[k]) if k in data else None)


def _collapsed_actual_cell(grp, actual_dict: dict) -> str:
    """DC-1: collapsed actual 컬럼 셀 텍스트 (ArrayGroup.format_actual → OK / NG(k/N)).

    actual_dict 값은 (actual, expected) 튜플 — 일치 여부로 OK/NG 집계. 튜플 아니거나
    부재면 매치 판정 불가(None)로 집계에서 제외.
    """
    def _match(k):
        t = actual_dict.get(k)
        # reviewer W2: 부재(None)만 집계 제외. falsy 실제값('0','')은 정상 튜플이면 비교.
        if t is None:
            return None
        if isinstance(t, tuple) and len(t) >= 2:
            return str(t[0]) == str(t[1])
        return None  # 튜플 아님(스칼라 등) → 매치 판정 불가

    text, _ = grp.format_actual(_match)
    return text


@dataclass
class SutrBuildMeta(BuildMetaBase):
    """SUTR 빌드 메타 — base에 SUTR 전용 2 필드 + final_test_result default override.

    T137 (W3 fix): `CoverageBuildMeta` 와 17 공통 필드를 `BuildMetaBase` 단일 출처로.
    """
    doc_id_base: str = "HDPDM01-SUTR"
    target_coverage: float = 1.0
    target_pass_ratio: float = 1.0
    final_test_result: str = "OK"  # Coverage는 "PASS", SUTR은 "OK"
    # 2026-06-19 — spec-based 2.Deviation 모드. True면 커버리지 미달 함수 목록을
    # 기재하지 않고 표준 "해당 사항 없음"(회사 DV ref v1.01 형식)으로 비워둔다.
    # config `sutr_deviation_empty` (per-project)에서 주입. False=기존 미달 목록 유지.
    deviation_empty: bool = False


@dataclass
class SutrBuildResult:
    """SUTR 빌드 결과.

    14차 W1: 메모리 절약 — ``xlsm_io: BytesIO`` 가 주 저장소. ``xlsm_bytes`` 는
    backward compat property — 호출 시점에 ``getvalue()`` (1회 copy).
    """
    ok: bool
    xlsm_io: io.BytesIO = field(default_factory=io.BytesIO)
    filename: str = ""
    warnings: list[str] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    incomplete_sheets: list[str] = field(default_factory=list)
    # VBA 매크로 ZIP entry 존재 여부 (deep-reviewer W2) — 실제 실행 검증은 사용자 의무.
    vba_macros_preserved: bool = False
    tool_qualification: dict[str, Any] = field(
        default_factory=lambda: {
            "evidence_class": "auto-generated draft",
            "asil_a_usage": "reviewer 승인 후 evidence로 사용 가능",
            "asil_b_c_d_usage": "단독 evidence 사용 금지 — manual review 의무",
        }
    )

    @property
    def xlsm_bytes(self) -> bytes:
        """Backward compat — BytesIO 전체를 bytes로 복사 (테스트/감사용)."""
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0)
        try:
            return self.xlsm_io.read()
        finally:
            self.xlsm_io.seek(pos)

    @property
    def result_size_bytes(self) -> int:
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0, 2)
        size = self.xlsm_io.tell()
        self.xlsm_io.seek(pos)
        return size

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "filename": self.filename,
            "result_size_bytes": self.result_size_bytes,
            "warnings": self.warnings,
            "incomplete_sheets": self.incomplete_sheets,
            "vba_macros_preserved": self.vba_macros_preserved,
            "summary": self.summary,
            "tool_qualification": self.tool_qualification,
        }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# helper 함수는 backend/services/excel_template_utils.py 로 통합 (reviewer 권고 X5).


# _aggregate 는 swut_input_adapter.aggregate_session 으로 통합 (deep-reviewer W3).


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

_OPTIONAL_LABELS = {
    "Build Timestamp", "Reviewer", "Doc. ID",
    # 라운드 F7 D11 fix: 회사 표준 양식 (★개발템플릿 V3) Cover에 부재 — 1.Test
    # Summary로 이동 완료. silent OK (warning emit X).
    "Project", "ASIL Level", "Validation Date", "Status", "Version",
    # 라운드 96-final: KJPDS02 v1.01 Cover 항목 (C29 'Date') — 타 양식 부재 시 silent.
    "Date",
}


def _write_label(ws, label: str, value: Any, out_warnings: list[str] | None) -> None:
    """K1 (reviewer): 라벨 미발견 시 warnings 누적 — 단 optional 라벨은 silent OK."""
    ok = write_value_after_label(ws, label, value)
    if not ok and label not in _OPTIONAL_LABELS and out_warnings is not None:
        out_warnings.append(f"라벨 '{label}' 미발견 — 셀 쓰기 skip")


def _write_label_or_mark(
    ws, label: str, value: Any, hint: str,
    out_warnings: list[str] | None,
) -> None:
    """23차 T192/W12: excel_template_utils.write_label_or_mark 래퍼 — _OPTIONAL_LABELS 주입."""
    write_label_or_mark(
        ws, label, value, hint=hint,
        optional_labels=_OPTIONAL_LABELS,
        out_warnings=out_warnings,
    )


def _write_cover(
    ws, meta: SutrBuildMeta, out_warnings: list[str] | None = None,
    *, layout: Any = None,
) -> None:
    """54차 T282: layout 제공 시 cover_labels 매핑으로 v2.02 양식 동적 호환."""
    labels = layout.cover_labels if layout else {}
    _write_label(ws, labels.get("project_full_name", "Project"),
                 meta.project_full_name, out_warnings)
    _write_label(ws, labels.get("asil_level", "ASIL Level"),
                 meta.asil_level, out_warnings)
    _write_label(ws, labels.get("status", "Status"),
                 "DRAFT — PENDING REVIEW", out_warnings)
    # 23차 T192: 비어있으면 노란 강조 (audit reviewer 가시성)
    _write_label_or_mark(ws, labels.get("validation_date", "Validation Date"),
                         meta.validation_date,
                         "yyyy-mm-dd 형식 검증 완료일", out_warnings)
    # 라운드 96-final QA fix — 가로 연속 서명란(KJPDS02 v1.01 I2/J2/K2) 감지 시
    # 라벨 '아래' 셀에 이름 기입 ('Author' 값이 'Reviewer' 라벨을 덮어쓰던 결함).
    _reviewer = getattr(meta, "reviewer", "") or (
        getattr(meta, "default_reviewer", "") or getattr(meta, "reviewer_override", "")
    )
    _sig_row = write_signature_block(ws, {
        labels.get("author", "Author"): meta.author,
        labels.get("reviewer", "Reviewer"): _reviewer,
        labels.get("approver", "Approver"): meta.approver,
    }, hint_map={
        labels.get("author", "Author"): "test_engineer 또는 default_author",
        labels.get("reviewer", "Reviewer"): "검토자 이름",
        labels.get("approver", "Approver"): "승인자 이름 (필수)",
    })
    if _sig_row is None:
        _write_label_or_mark(ws, labels.get("author", "Author"), meta.author,
                             "test_engineer 또는 default_author", out_warnings)
        _write_label_or_mark(ws, labels.get("approver", "Approver"), meta.approver,
                             "승인자 이름 (필수)", out_warnings)
    else:
        # 서명란 아래 별도 표지 항목 'Author' kv (예: KJPDS02 C30) — G30에 기입.
        if meta.author:
            write_value_after_label(
                ws, labels.get("author", "Author"), meta.author,
                min_row=_sig_row + 1,
            )
    if meta.doc_id_sequence:
        _write_label(ws, labels.get("doc_id", "Doc. ID"),
                     f"{meta.doc_id_base}-{meta.doc_id_sequence}", out_warnings)
    _write_label(ws, labels.get("version", "Version"),
                 f"v{meta.release_sw_version}", out_warnings)
    # 라운드 96-final QA fix — Cover 'Date' DV 잔존 차단 + 'Document ID' 보정.
    _write_label(ws, labels.get("date", "Date"),
                 dot_date(meta.test_date), out_warnings)
    stamp_cover_document_id(
        ws, project_id=meta.project_id,
        doc_filename_pattern=getattr(meta, "doc_filename_pattern", "") or "",
        out_warnings=out_warnings,
    )
    # optional — 회사 v3.01 template에 라벨이 없을 수 있어 silent skip 허용.
    _write_label(ws, labels.get("build_timestamp", "Build Timestamp"),
                 meta.build_timestamp, out_warnings)


def _write_test_summary(
    ws, meta: SutrBuildMeta, agg: dict[str, Any],
    out_warnings: list[str] | None = None,
    *, layout: Any = None, summary: dict[str, Any] | None = None,
) -> None:
    """54차 T282/T283: layout 제공 시 v2.02 양식 동적 호환."""
    labels = layout.test_summary_labels if layout else {}
    _write_label(ws, labels.get("project_full_name", "Project Name"),
                 meta.project_full_name, out_warnings)
    _write_label(ws, labels.get("release_sw_version", "Release Name(SW)"),
                 meta.release_sw_version, out_warnings)
    _write_label(ws, labels.get("hw_version", "Test Target Version(HW)"),
                 meta.hw_version, out_warnings)
    _write_label(ws, labels.get("test_date", "Test Date"),
                 meta.test_date, out_warnings)
    # 24차: Test Engineer 빈 시 노란 강조 (Coverage와 대칭)
    _write_label_or_mark(ws, labels.get("test_engineer", "Test Engineer"),
                         meta.test_engineer,
                         "테스트 엔지니어 이름", out_warnings)
    _write_label(ws, labels.get("target_coverage", "Target Coverage"),
                 meta.target_coverage, out_warnings)
    # deep-reviewer X7: 0/N의 silent wrong-pick 회피 — N=0이면 "N/A" 명시.
    # 24차: agg.total == 0 → N/A는 input 데이터 부재 의미, 노란 강조로 reviewer 가시화.
    if agg["total"] > 0:
        _write_label(ws, labels.get("actual_coverage", "Actual Coverage"),
                     agg["tested"] / agg["total"], out_warnings)
    else:
        _write_label_or_mark(ws, labels.get("actual_coverage", "Actual Coverage"), "",
                             "VectorCAST 데이터 부재 — log_folder 재확인", out_warnings)
    _write_label(ws, labels.get("target_pass_ratio", "Target Pass ratio"),
                 meta.target_pass_ratio, out_warnings)
    if agg["tested"] > 0:
        _write_label(ws, labels.get("actual_pass_ratio", "Actual Pass ratio"),
                     agg["passed"] / agg["tested"], out_warnings)
    else:
        _write_label_or_mark(ws, labels.get("actual_pass_ratio", "Actual Pass ratio"), "",
                             "실행된 TC 없음 — log 또는 deviation 확인", out_warnings)
    # ⚠ 예전엔 `meta.final_test_result`(정적 기본값 "OK")를 그대로 썼다 — 실패 5/5 인
    # 산출물도 "OK" 로 찍혔다. 이제 집계에서 도출한다. 판정은 swut_input_adapter 단일 출처.
    # meta 값은 이 문서의 **긍정 토큰**으로만 쓴다(전부 통과면 오늘과 같은 값 → 하위호환).
    _fr = compute_final_result(agg, positive_token=meta.final_test_result)
    _write_label(ws, labels.get("final_test_result", "Final Test Result"),
                 _fr["display"], out_warnings)
    if _fr["verdict"] != "ok" and out_warnings is not None:
        out_warnings.append(
            f"[final-result] Final Test Result = '{_fr['display']}' — {_fr['reason']}"
        )
    if summary is not None:
        # xlsx 를 열어보지 않고도 판정을 알 수 있게 API 응답에도 싣는다.
        summary["final_result"] = _fr["display"]
        summary["final_result_verdict"] = _fr["verdict"]
        summary["final_result_reason"] = _fr["reason"]

    # 55-fix-3 W10: helper 추출 — swut_coverage와 단일 source. inline 중복 제거.
    from backend.services.swut_coverage_aggregator import (
        _write_v202_requirements_row,
        _write_v202_tc_stats_row,
    )
    _write_v202_tc_stats_row(
        ws, agg, layout, summary, out_warnings, total_key="total",
    )
    _write_v202_requirements_row(ws, layout, summary, out_warnings)


def _deviation_case_fields(case: Any) -> tuple[str, str, str, str] | None:
    """case → (tc_id, tc_no, issue_text, auto_rationale).

    deep-reviewer W6/X7: dict/DeviationCase 외 shape는 명시적 거부.
    5차 H3 Critical: issue_text/auto_rationale은 xlsx 셀 한도 방어용 truncate.
    """
    if isinstance(case, dict):
        tc_id_v = str(case.get("tc_id", "") or "")
        tc_no_v = str(case.get("tc_no", "") or "")
        if not tc_id_v:
            return None
        issue, _ = truncate_cell_text(case.get("issue_text", ""))
        rationale, _ = truncate_cell_text(case.get("auto_rationale", ""))
        return (tc_id_v, tc_no_v, issue, rationale)
    tc_id = getattr(case, "tc_id", None)
    if not tc_id:
        return None
    issue, _ = truncate_cell_text(getattr(case, "issue_text", ""))
    rationale, _ = truncate_cell_text(getattr(case, "auto_rationale", ""))
    return (str(tc_id), str(getattr(case, "tc_no", "") or ""), issue, rationale)


def _write_deviation(
    ws, deviation_cases: list[Any], out_warnings: list[str] | None = None,
    *, none_text: str | None = None,
) -> int:
    """Deviation 시트 — swut_deviation_generator 결과 기록.

    라운드 F7 T707: clear policy — deviation_cases 빈 list 또는 신규 stamp 후
    양식 default deviation 데이터 clear (Appendix sentinel 보존).

    라운드 96-final: ``none_text`` (keyword-only, 기본 None — backward compat) —
    0건일 때 첫 데이터 행에 기재할 DV 관례 문구 (예: '해당사항 없음'). None이면
    기존 동작(공백 유지). SwIT(swit_sitr_aggregator) 등 기존 호출처 무영향.

    Returns: 쓰여진 행 수 (none_text 기재는 카운트 제외).
    """
    pos = find_kv_row(ws, "Test Case ID", max_row=10)
    if pos is None:
        if out_warnings is not None:
            out_warnings.append("Deviation 시트 'Test Case ID' 헤더 미발견 — skip")
        return 0
    start_row = pos[0] + 1
    written = 0
    skipped = 0
    for case in deviation_cases or []:
        fields = _deviation_case_fields(case)
        if fields is None:
            skipped += 1
            continue
        tc_id_v, tc_no_v, issue, rationale = fields
        tc_label = f"{tc_id_v} ({tc_no_v})" if tc_no_v else tc_id_v
        r = start_row + written
        safe_write(ws, r, pos[1], tc_label)
        safe_write(ws, r, pos[1] + 1, issue)
        safe_write(ws, r, pos[1] + 2, rationale)
        safe_write(ws, r, pos[1] + 3, "Auto-Generated")
        written += 1
    if skipped and out_warnings is not None:
        out_warnings.append(
            f"Deviation case shape 검증 실패 — {skipped}건 skip (dict 또는 DeviationCase 필요)"
        )

    # 라운드 F7 T707: clear policy — 신규 stamp 후 다음 row부터 양식 default clear.
    # 'Appendix' sentinel ('■ Appendix' 등)을 만나면 그 직전까지만 clear. 양식의
    # 기존 deviation 4건 default (다른 release 데이터)가 신규 빌드에 보존되는
    # partial overwrite 결함 차단.
    try:
        from backend.services.excel_template_utils import clear_data_range
        clear_start = start_row + written
        clear_end = ws.max_row
        # Appendix sentinel 위치 탐지 — 그 직전까지 clear
        appendix_row = None
        for r in range(clear_start, min(clear_end + 1, clear_start + 100)):
            cell_value = ws.cell(r, 2).value
            if isinstance(cell_value, str) and "Appendix" in cell_value:
                appendix_row = r
                break
        actual_end = (appendix_row - 1) if appendix_row else min(clear_end, clear_start + 50)
        if actual_end >= clear_start:
            cleared = clear_data_range(
                ws,
                start_row=clear_start, end_row=actual_end,
                start_col=pos[1], end_col=pos[1] + 4,
                preserve_formula=True, preserve_merged_anchor=True,
            )
            if out_warnings is not None and cleared > 0:
                out_warnings.append(
                    f"[clear] Deviation row {clear_start}~{actual_end} 양식 default "
                    f"{cleared} cell clear (partial overwrite 차단)"
                )
    except ImportError:
        pass
    # 라운드 96-final — 0건 시 DV 관례 '해당사항 없음' 기재 (옵션). clear 블록
    # 이후에 기입해야 함 (clear_start=start_row+0 — 먼저 쓰면 clear가 지움).
    if written == 0 and none_text:
        safe_write(ws, start_row, pos[1], none_text)
    return written


def _write_variable_name_header_row(
    ws,
    layout: Any,
    session: SwUTSession,
    *,
    input_col: int,
    expected_col: int,
    actual_col: int,
    input_max: int,
    expected_max: int,
    actual_max: int,
    out_warnings: list[str] | None = None,
) -> tuple[list[str], list[str], list[str]]:
    """59차 F4-A — Test Log 변수명 헤더 row stamp + 환경 합집합 sorted list 반환.

    KJPDS02 v1.01 양식은 row 5 col 10~ 에 변수명 (예: u16g_SysDiag_SystemStatus)
    stamp 후 row 6~ 에 값. 우리 추출은 환경별 dict[변수명, 값] 보유 — env가 여러
    개일 때 같은 변수가 일부 env에만 있을 수 있어 **전체 환경 합집합 + sorted**
    를 헤더로 stamp하고 TC stamp는 그 col 순서대로 lookup.

    Args:
        ws: openpyxl worksheet.
        layout: SwitLayout — test_log_variable_header_row None이면 skip.
        session: SwUTSession — environments 순회.
        input_col / expected_col / actual_col: column 시작 위치.
        input_max / expected_max / actual_max: block 최대 stamp 수.
        out_warnings: diag 누적용.

    Returns:
        (input_var_names, expected_var_names, actual_var_names) — sorted union list,
        max_count 적용 후. empty 3-tuple이면 backward-compat (caller가 dict.values()
        순서 사용).
    """
    if layout is None or getattr(layout, "test_log_variable_header_row", None) is None:
        return [], [], []

    header_row = layout.test_log_variable_header_row

    input_keys: set[str] = set()
    expected_keys: set[str] = set()
    actual_keys: set[str] = set()

    for env in session.environments:
        # TestCaseItem.input_data / expected_result 추출 — env.test_cases는
        # dict[str, List[TestCaseItem]] (TCBank.test_cases carry forward 패턴).
        for tc_items in env.test_cases.values():
            items = tc_items if isinstance(tc_items, list) else [tc_items]
            for tc_item in items:
                input_keys.update(getattr(tc_item, "input_data", {}) or {})
                expected_keys.update(getattr(tc_item, "expected_result", {}) or {})
        # ExecutionRow.actual_result (58차 F1 BeautifulSoup 추출) 우선
        for exec_r in env.test_results.values():
            actual_keys.update(getattr(exec_r, "actual_result", {}) or {})
        # TestResultItem.actual_result (vcast_parser TestResultItem) fallback
        for tr_items in getattr(env, "tc_result_items", {}).values():
            items = tr_items if isinstance(tr_items, list) else [tr_items]
            for tr_item in items:
                actual_keys.update(getattr(tr_item, "actual_result", {}) or {})

    input_list = sorted(input_keys)[: max(0, input_max)]
    expected_list = sorted(expected_keys)[: max(0, expected_max)]
    actual_list = sorted(actual_keys)[: max(0, actual_max)]

    for i, name in enumerate(input_list):
        safe_write(ws, header_row, input_col + i, name)
    for i, name in enumerate(expected_list):
        safe_write(ws, header_row, expected_col + i, name)
    for i, name in enumerate(actual_list):
        safe_write(ws, header_row, actual_col + i, name)

    if out_warnings is not None:
        msg = (
            f"F4-A: 변수명 헤더 row {header_row} stamp — "
            f"input/expected/actual = "
            f"{len(input_list)}/{len(expected_list)}/{len(actual_list)} variables "
            f"(union sorted, max {input_max}/{expected_max}/{actual_max})"
        )
        out_warnings.append(f"[diag] {msg}")

    return input_list, expected_list, actual_list


def _build_fn_id_to_component_map(session: SwUTSession) -> dict[str, str]:
    """라운드 80 T1407: 세션 단위 fn_id → component_name 캐시.

    환경별 function_coverage iterate 1회로 ``SwUFn_NNNN`` → ``SwCom_NN`` (또는
    별칭) 매핑 dict 구축. SDS 컴포넌트 ASIL fallback에서 사용.

    component_name 값에 ``"SwCom_01\\n(System OS)"`` 같은 multi-line 양식 가능 —
    첫 줄 (SwCom_NN 부분) 우선, 다음 줄 (별칭) 둘 다 등록.
    """
    cache: dict[str, str] = {}
    for env in session.environments:
        for fc in env.function_coverage:
            fn_id = getattr(fc, "unit_id", "") or getattr(fc, "function_id", "")
            comp_raw = getattr(fc, "component_name", "") or ""
            if not fn_id or not comp_raw:
                continue
            # SwUFn_NNNN 부분 추출
            import re as _re_fn
            m = _re_fn.search(r"SwUFn_\d+", fn_id)
            if m:
                cache.setdefault(m.group(0), comp_raw)
    return cache


def _resolve_component_asil(
    fn_id: str,
    fn_to_comp: dict[str, str],
    sds_map: dict[str, str],
) -> str:
    """라운드 80 T1407: fn_id → component_name → SDS ASIL lookup.

    component_name 양식이 ``"SwCom_01\\n(System OS)"`` 다양 — 후보 키 여러 개
    시도 (SwCom_NN / strip / 별칭). 매칭 0 시 빈 string.
    """
    if not fn_id or not sds_map:
        return ""
    comp_raw = fn_to_comp.get(fn_id, "")
    if not comp_raw:
        return ""
    # 후보 키들 — SwCom_NN, 첫 줄, 전체 strip
    import re as _re_cn
    candidates = []
    m_swcom = _re_cn.search(r"SwCom_\d+", comp_raw)
    if m_swcom:
        candidates.append(m_swcom.group(0))
    for line in comp_raw.splitlines():
        stripped = line.strip().strip("()").strip()
        if stripped and stripped not in candidates:
            candidates.append(stripped)
    candidates.append(comp_raw.strip())
    for k in candidates:
        if k in sds_map:
            return sds_map[k]
    return ""


def _write_test_log(
    ws,
    session: SwUTSession,
    function_asil_map: dict[str, str] | None = None,
    out_warnings: list[str] | None = None,
    *,
    layout: Any = None,
    swuts_map: dict[str, Any] | None = None,
    c_function_map: dict[str, dict[str, Any]] | None = None,
) -> int:
    """Test Log/Test Result 시트 — 함수(SwUFn_NNNN)당 1블록.

    라운드 90 재작성 — 회사 감사본(SUTR v3.01 / SITR v2.02) 양식 정합.
    이전(57~89차)은 iteration key(``SwUFn_1901.001``)를 TC 1개로 보고 per-TC 고정
    6 row 블록을 만들고 한 iteration의 변수들을 anchor row의 열에 채웠다 → 회사
    양식과 불일치. 신규 양식:

    - **anchor row**: B=TC ID, C=Title, D=Generation Method, 그리고 Input/Expected/
      Actual 섹션 각 열에 **변수명**을 나열. AK(Total)=함수 전체 Pass/Fail (블록
      세로 병합). ASIL 등급 시각 강조도 anchor row에 적용.
    - **iteration row** (anchor+1부터): E=iteration index(1..N), 각 변수 열에 그
      iteration의 값(Input/Expected/Actual). AJ(Unit)=그 iteration의 Pass/Fail.
    - 블록 높이 = 1 anchor + iteration 수 (가변).

    column 위치는 layout (test_log_input_col 등) 우선, 없으면 v3.01 hardcode
    (Input=F/6, Expected=P/16, Actual=Z/26, AJ=36, AK=37, AL=38).

    Generation Method(D열)는 swuts_map(SwUTS spec)의 test_method/generation_method
    우선, 없으면 "AEC, ABV". VectorCAST 로그에는 method 정보 미존재(vcast_parser
    확인) → 본 라운드는 spec fallback 유지.

    변수 > layout max_count 시 truncate + out_warnings에 경고 누적.
    """
    if not ws:
        return 0
    # 헤더 찾기
    pos = find_kv_row(ws, "Test Case ID", max_row=10)
    if pos is None:
        pos = find_kv_row(ws, "TC ID", max_row=10)
    if pos is None:
        return 0
    start_row = pos[0] + 1
    col = pos[1]
    asil_map = function_asil_map or {}

    # 31차 W27 reviewer W3 + 31-fix D10: col+4/5 빈 영역 가정 검증.
    # 헤더 row의 col+4/5 셀이 비어있는지 확인 — 회사 v3.01 양식 업그레이드 시
    # 데이터 덮어쓰기 위험 가시화. logger.warning은 backend 로그 전용 → 사용자
    # UI에 표시 안 됨. out_warnings에도 누적해서 X-SwUT-Warnings 헤더 통해
    # frontend "Warnings" 패널에 노출.
    header_col4 = ws.cell(pos[0], col + 4).value
    header_col5 = ws.cell(pos[0], col + 5).value
    if header_col4 or header_col5:
        msg = (
            f"SUTR Test Log col+4/5 not empty (col+4={header_col4!r}, "
            f"col+5={header_col5!r}) — 회사 양식 업그레이드 가능성, ASIL "
            "컬럼 덮어쓰기 진행. audit reviewer 확인 권장"
        )
        import logging
        logging.getLogger(__name__).warning(msg)
        if out_warnings is not None:
            out_warnings.append(msg)

    # 57차 T314 — Coverage Traceability와 동일 TC source 사용 (환경 union → unique TC).
    # 환경별 iterate 대신 1941 unique TC name을 한 번에 stamp + 회사 v2.02 양식의
    # 1 TC당 6 row step 자동 적용. SITR도 동일 함수 import로 동시 효과.
    from backend.services.swut_coverage_aggregator import _collect_tc_to_function
    tc_to_fn_id = _collect_tc_to_function(session)  # {tc_name: fn_id}

    # 라운드 80 T1407: 세션 단위 fn_id → component_name cache (SDS ASIL fallback용).
    _fn_to_comp_cache = _build_fn_id_to_component_map(session)

    # tc_name → 첫 매칭 env (component_name + test_results 조회용).
    # 환경별 동일 TC가 중복 정의되면 첫 환경 우선 (Coverage source semantic 일치).
    tc_to_env: dict[str, Any] = {}
    total_tcs_in_envs = 0
    for env in session.environments:
        for tc_name in env.test_cases:
            total_tcs_in_envs += 1
            if tc_name not in tc_to_env:
                tc_to_env[tc_name] = env

    tc_row_step = (
        getattr(layout, "test_log_tc_row_step", 1) if layout is not None else 1
    )

    # 57차 T314 diag: 진단용 — backend log 및 out_warnings에 데이터 흐름 출력
    diag_msg = (
        f"SUTR _write_test_log diag: environments={len(session.environments)}, "
        f"total_tcs_in_envs={total_tcs_in_envs}, "
        f"tc_to_fn_id_size={len(tc_to_fn_id)}, "
        f"tc_row_step={tc_row_step}, start_row={start_row}, col={col}"
    )
    import logging as _logging
    _logging.getLogger(__name__).info(diag_msg)
    if out_warnings is not None:
        out_warnings.append(f"[diag] {diag_msg}")

    # 57차 T319 fix — Template block (row 5~base+row_step*template_count) 의
    # merge pattern + style 추출 → 새 TC block (template 영역 밖)에 동일 확장.
    # 회사 v2.02 SUTR 양식: 6 TC × 6 row × 38 col merge/style pattern.
    # audit reviewer가 일관 양식으로 산출물 검토.
    import copy as _copy
    template_merges_local: list[tuple[int, int, int, int]] = []
    for _mc in list(ws.merged_cells.ranges):
        if (start_row <= _mc.min_row <= start_row + tc_row_step - 1 and
                _mc.max_row <= start_row + tc_row_step - 1):
            # template block 1번 내부 merge — offset 0 row 시작
            template_merges_local.append(
                (_mc.min_row - start_row, _mc.max_row - start_row,
                 _mc.min_col, _mc.max_col)
            )

    # 라운드 89 perf — template block 스타일을 1회 precompute (StyleArray 캐시).
    # 이전: 신규 TC block마다 cell별 font/border/fill/alignment/protection 5종을
    # _copy.copy() → openpyxl serialisable __copy__/to_tree/from_tree 머신이 TC당
    # ~980회 호출 (3771 TC = 7분+, endpoint 비실용적). openpyxl은 cell._style
    # (StyleArray = 공유 style table 인덱스)로 스타일을 표현하므로 _style만 복사하면
    # 시각 결과 100% 동일 + 직렬화 회피로 극적 가속. (offset, col) → _style.
    _tpl_block_styles: dict[tuple[int, int], Any] = {}
    _tpl_max_col = ws.max_column or 38
    for _off in range(tc_row_step):
        _src_r = start_row + _off
        for _c in range(1, _tpl_max_col + 1):
            _sc = ws.cell(_src_r, _c)
            if _sc.has_style:
                _tpl_block_styles[(_off, _c)] = _copy.copy(_sc._style)

    # 58차 F3 — column 매핑 layout-aware (layout 제공 시 동적, None이면 v3.01 hardcode):
    #   회사 v3.01 SUTR (SwUT): F=Input, P=Expected, Z=Actual, AJ=Pass/Fail Unit,
    #       AK=Pass/Fail Total, AL=Log Data
    #   회사 v2.02 SITR (SwIT): H=Input, R=Expected, AB=Actual, AL=Pass/Fail,
    #       AN=Log Data (Pass/Fail Total 없음)
    # SwitLayout.test_log_input_col 등 6 field 우선 사용 → None이면 v3.01 hardcode.
    # 57차 T319 fix: 이전 col+3/+4/+5 잘못 stamp 정정 후 hardcode 36/37/38.
    # layout이 6개 column 중 1개라도 보유하면 layout-aware (v2.02). 모두 None이면 v3.01 hardcode.
    _layout_has_col = layout is not None and any(
        getattr(layout, attr, None)
        for attr in (
            "test_log_input_col", "test_log_expected_col", "test_log_actual_col",
            "test_log_pass_fail_col", "test_log_log_data_col",
        )
    )
    if _layout_has_col:
        # v3.01 hardcode default을 fallback으로 사용 (각 col이 None이면 적용).
        INPUT_COL = layout.test_log_input_col or 6
        EXPECTED_COL = layout.test_log_expected_col or 16
        ACTUAL_COL = layout.test_log_actual_col or 26
        PASS_FAIL_UNIT_COL = layout.test_log_pass_fail_col or 36
        PASS_FAIL_TOTAL_COL = layout.test_log_pass_fail_total_col or 0  # 0 = skip stamp
        LOG_DATA_COL = layout.test_log_log_data_col or 38
    else:
        # v3.01 backward-compat hardcode (SwUT SUTR 양식)
        INPUT_COL = 6     # F
        EXPECTED_COL = 16  # P
        ACTUAL_COL = 26    # Z
        PASS_FAIL_UNIT_COL = 36   # AJ
        PASS_FAIL_TOTAL_COL = 37  # AK
        LOG_DATA_COL = 38         # AL

    # 59차 F4-A — block max counts (layout 기반 동적). layout None backward-compat 10.
    input_max = (
        getattr(layout, "test_log_input_max_count", 10) if layout is not None else 10
    )
    expected_max = (
        getattr(layout, "test_log_expected_max_count", 10) if layout is not None else 10
    )
    actual_max = (
        getattr(layout, "test_log_actual_max_count", 10) if layout is not None else 10
    )

    # 라운드 90 T1601 — 회사 감사본 양식: 함수(SwUFn_NNNN)당 1블록.
    #   anchor row: B=TC ID / C=Title / D=Method + Input/Expected/Actual 섹션에
    #     **변수명을 열에 나열**. AK(Total)=함수 전체 Pass/Fail (블록 세로 병합).
    #   iteration row (anchor+1부터): E=iteration index(1..N), 각 변수 열에 그
    #     iteration의 값. AJ=그 iteration의 Pass/Fail.
    # 이전(57차~89차) "iteration=행 1개, 변수=열" per-TC 6 row 고정 블록을 폐기하고
    # 함수별 가변 높이 블록으로 재작성. _write_variable_name_header_row(글로벌 단일
    # 헤더 row stamp)는 본 양식과 비호환이라 호출 제거 — 변수명은 함수별 anchor row에.
    # 라운드 90 T1601 — 함수별 그룹핑.
    #   iteration key 'SwUFn_1901.001' → 함수 id 'SwUFn_1901' (suffix '.NNN' 제거).
    #   SwIT 'SwITC_SwUFn_0101.001' → 함수 id 'SwUFn_0101' (_collect_tc_to_function이
    #     이미 SwUFn_NNNN 추출 — fn_id 그대로 group key 사용).
    # group: fn_id → [iteration_key, ...] (정렬). iteration 정렬은 '.NNN' 숫자 우선.
    def _iter_sort_key(tc_name: str) -> tuple[int, str]:
        m = re.search(r"\.(\d+)\s*$", tc_name)
        if m:
            return (int(m.group(1)), tc_name)
        return (0, tc_name)

    fn_groups: dict[str, list[str]] = {}
    for tc_name, fn_id in tc_to_fn_id.items():
        fn_groups.setdefault(fn_id, []).append(tc_name)
    for fn_id in fn_groups:
        fn_groups[fn_id].sort(key=_iter_sort_key)
    sorted_fn_ids = sorted(fn_groups.keys())

    # 블록 높이 = 1 anchor + iteration 수. 전체 needed row 미리 산출.
    block_heights: dict[str, int] = {
        fn_id: 1 + len(fn_groups[fn_id]) for fn_id in sorted_fn_ids
    }
    total_rows_needed = sum(block_heights.values())

    # MERGE-06 가드 — tc_row_step 오검출 방어. 세로 병합(B/C/D 등)은 template_merges_local
    # (양식 첫 블록 사전 병합)을 블록마다 복제해 재구성한다. tc_row_step이 1로 잘못 감지되면
    # (예: v2.02 멀티행 양식인데 step=1) 첫 블록 병합 capture 범위가 1행으로 좁아져
    # template_merges_local이 비고, 다중행(block_height>1) 블록이 세로 병합 없이 stamp된다.
    # 즉시 오류는 아니나 '셀병합 불일치' 추가 실패면이라 audit reviewer가 식별하도록 경고.
    if (out_warnings is not None and not template_merges_local
            and sorted_fn_ids and max(block_heights.values()) > 1):
        out_warnings.append(
            f"[merge_guard] 함수 블록 세로 병합 양식 미검출 (tc_row_step={tc_row_step}, "
            f"template 병합 0개, 최대 블록 {max(block_heights.values())}행) — 다중행 함수 "
            "블록이 세로 병합 없이 stamp될 수 있음. 양식 헤더/tc_row_step 감지 확인 권장."
        )

    # 변수>10 truncate 발생 함수 수 집계 (보고용).
    _truncated_fn_count = 0

    # 라운드 90 — 함수별 가변 높이 블록에 맞춰 row 확장.
    # 이전(73차) needed = start + (TC수-1)*step + (step-1) (고정 step) 폐기.
    # 신규 needed = start + (1 anchor + iter수 합) - 1.
    if sorted_fn_ids:
        needed_last_row = start_row + total_rows_needed - 1
        if needed_last_row > ws.max_row:
            try:
                from backend.services.excel_template_utils import (
                    auto_expand_row_block,
                    push_sentinel_to_last_row,
                )
                shortage = needed_last_row - ws.max_row
                # template block 1번 (start_row ~ start_row + tc_row_step - 1) 다음에 row 확장.
                inserted = auto_expand_row_block(
                    ws,
                    insert_at_row=start_row + tc_row_step,
                    amount=shortage,
                    template_row_idx=start_row,
                    copy_style=True, copy_merge=True, copy_dimension=True,
                )
                push_sentinel_to_last_row(ws)
                if inserted < shortage and out_warnings is not None:
                    out_warnings.append(
                        f"[row_expand] Test Log row 부족 ({shortage}개 필요, "
                        f"{inserted}개 확장) — stamp 일부 누락 가능"
                    )
            except ImportError:
                pass

    # 라운드 90 — anchor row TC ID/Title/Method를 함수의 첫 iteration + swuts_map으로
    # 산출하는 helper (이전 per-TC 로직 재사용).
    def _resolve_anchor_meta(rep_tc_name: str, component_name: str):
        """함수 anchor row의 (TC ID, Title, Method, precondition_col, precondition,
        swuts_entry) 산출. swuts_map 우선, 없으면 fallback."""
        swuts_entry = None
        if swuts_map:
            swuts_entry = swuts_map.get(rep_tc_name)
            if swuts_entry is None:
                _fn_match = re.search(r"SwUFn_(\d+)", rep_tc_name)
                if _fn_match:
                    _fn_id_full = _fn_match.group(0)
                    swuts_entry = swuts_map.get(f"SwUTC_{_fn_match.group(1)}")
                    if swuts_entry is None:
                        swuts_entry = swuts_map.get(f"SwUTC_{_fn_id_full}")
                    if swuts_entry is None:
                        swuts_entry = swuts_map.get(f"SwITC_{_fn_match.group(1)}")
        if swuts_entry is not None:
            tc_id = getattr(swuts_entry, "tc_id", "") or rep_tc_name
            description = (
                getattr(swuts_entry, "description", "")
                or getattr(swuts_entry, "unit_name", "")
                or component_name
            )
            _method = getattr(swuts_entry, "test_method", "")
            _gen_method = getattr(swuts_entry, "generation_method", "")
            if _method and _gen_method:
                method = f"{_method}, {_gen_method}"
            elif _method:
                method = _method
            elif _gen_method:
                method = _gen_method
            else:
                method = "AEC, ABV"
            precond_col = (
                getattr(layout, "test_log_precondition_col", None)
                if layout is not None else None
            )
            precond = getattr(swuts_entry, "precondition", "")
            return tc_id, description, method, precond_col, precond, swuts_entry
        # fallback — SwIT 'SwITC_SwUFn_NNNN.NNN' → 'SwITC_NNNN', SwUT은 함수 id 그대로.
        tc_id = rep_tc_name
        _switc_fn = re.match(r"^SwITC_SwUFn_(\d+)", rep_tc_name)
        if _switc_fn:
            tc_id = f"SwITC_{_switc_fn.group(1)}"
        return tc_id, component_name, "AEC, ABV", None, "", None

    # 라운드 90 — 함수의 ASIL 등급 산출 (이전 per-TC 로직을 함수 단위로 1회).
    def _resolve_fn_asil(fn_id: str, rep_tc_name: str, swuts_entry) -> str:
        asil = asil_map.get(fn_id, "") if fn_id else ""
        _cand_name = ""
        if not asil and c_function_map:
            _se = swuts_entry
            if _se is None and swuts_map:
                _se = swuts_map.get(rep_tc_name)
                if _se is None:
                    _fm = re.search(r"SwUFn_(\d+)", rep_tc_name)
                    if _fm:
                        _se = swuts_map.get(f"SwUTC_{_fm.group(1)}")
            if _se:
                _sig = getattr(_se, "unit_name", "") or ""
                _match = re.search(r"(\w+)\s*\(", _sig)
                _cand_name = _match.group(1) if _match else _sig.strip()
            if _cand_name:
                _c_entry = c_function_map.get(_cand_name)
                if _c_entry:
                    _ca = (_c_entry.get("comment_asil") or "").strip().upper()
                    if _ca in {"A", "B", "C", "D", "QM"}:
                        asil = _ca
        if not asil and fn_id:
            _suds_map = getattr(session, "function_asil_from_suds", {}) or {}
            _sasil = _suds_map.get(fn_id)
            if _sasil:
                asil = _sasil
        if not asil and fn_id:
            _sds_map = getattr(session, "component_asil_from_sds", {}) or {}
            if _sds_map:
                _casil = _resolve_component_asil(fn_id, _fn_to_comp_cache, _sds_map)
                if _casil:
                    asil = _casil
        if not asil:
            _srs_map = getattr(session, "function_asil_from_srs", {}) or {}
            _srs_key = _cand_name
            if not _srs_key and swuts_map:
                _entry_x = swuts_entry or swuts_map.get(rep_tc_name)
                if _entry_x:
                    _sig_x = getattr(_entry_x, "unit_name", "") or ""
                    _m_srs = re.search(r"(\w+)\s*\(", _sig_x)
                    if _m_srs:
                        _srs_key = _m_srs.group(1)
            if _srs_map and _srs_key:
                _rasil = _srs_map.get(_srs_key)
                if _rasil:
                    asil = _rasil
        return asil

    # 라운드 90 — 함수의 Input/Expected/Actual 변수명 union 산출.
    # 첫 iteration 순서 보존 + 이후 iteration의 신규 var append (insertion-ordered).
    def _collect_fn_var_union(iter_keys: list[str], envs_by_key: dict):
        input_names: list[str] = []
        expected_names: list[str] = []
        actual_names: list[str] = []
        seen_in: set[str] = set()
        seen_exp: set[str] = set()
        seen_act: set[str] = set()
        for ik in iter_keys:
            ev = envs_by_key.get(ik)
            if ev is None:
                continue
            tc_items = ev.test_cases.get(ik) or []
            tc_item = tc_items[0] if tc_items else None
            if tc_item is not None:
                for k in (getattr(tc_item, "input_data", {}) or {}):
                    if k not in seen_in:
                        seen_in.add(k)
                        input_names.append(k)
                for k in (getattr(tc_item, "expected_result", {}) or {}):
                    if k not in seen_exp:
                        seen_exp.add(k)
                        expected_names.append(k)
            # actual 변수명 — actual_result key (부정확 가능, 가능 범위). 없으면 후처리.
            exec_r = ev.test_results.get(ik)
            ad = getattr(exec_r, "actual_result", {}) or {} if exec_r else {}
            if not ad:
                tr_items = getattr(ev, "tc_result_items", {}).get(ik, [])
                tr_item = tr_items[0] if tr_items else None
                if tr_item is not None:
                    ad = getattr(tr_item, "actual_result", {}) or {}
            for k in ad:
                if k not in seen_act:
                    seen_act.add(k)
                    actual_names.append(k)
        return input_names, expected_names, actual_names

    written = 0          # stamp된 함수 블록 수
    iteration_rows = 0   # stamp된 iteration row 수 (진단/보고용)
    cur_row = start_row
    # 라운드 89 perf — per-block merge defer (resolve_merge_anchor 캐시 안정화).
    _deferred_merges: list[tuple[int, int, int, int]] = []

    for fn_id in sorted_fn_ids:
        iter_keys = fn_groups[fn_id]
        rep_tc_name = iter_keys[0]
        env0 = tc_to_env.get(rep_tc_name)
        component_name = env0.component_name if env0 is not None else ""

        # 함수 전체 Pass/Fail Total — 모든 iteration passed면 Pass, 하나라도 Fail이면
        # Fail, exec 전무면 N/A.
        any_exec = False
        all_pass = True
        for ik in iter_keys:
            ev = tc_to_env.get(ik)
            er = ev.test_results.get(ik) if ev is not None else None
            if er is not None:
                any_exec = True
                if not er.passed:
                    all_pass = False
        total_result_str = "Pass" if (any_exec and all_pass) else ("Fail" if any_exec else "N/A")

        # 변수명 union 산출.
        in_raw, exp_raw, act_raw = _collect_fn_var_union(iter_keys, tc_to_env)
        # actual 변수명이 비었거나 부정확하면 expected 변수명 재사용 (spec 지침).
        if not act_raw:
            act_raw = list(exp_raw)
        # DC-1 + collapse_all (2026-06-24 실데이터 검증/사용자 결정): 모든 배열(단일차원·
        # sparse 포함)을 base당 1열로 접는다. 실 HDPDM01 SUTR에서 45개 함수가 단일차원 배열
        # (lin_pFrameBuf[0..21] 등)로 10열 절단·silent 손실 중이었음 — collapse_all로 45/45 해소.
        # 배열 없는 함수는 columns == 원본 순서 그대로(no-op). C# gold(단일차원=별도 컬럼)와는
        # 다른 표현이며 ASIL audit deviation으로 보고됨.
        in_info = _collapse_build(in_raw, collapse_all=True)
        exp_info = _collapse_build(exp_raw, collapse_all=True)
        act_info = _collapse_build(act_raw, collapse_all=True)
        in_names, exp_names, act_names = (
            in_info.columns, exp_info.columns, act_info.columns
        )
        # truncate (이제 접기 후 컬럼 기준 — raw 대비 대폭 감소).
        _orig_in, _orig_exp, _orig_act = len(in_names), len(exp_names), len(act_names)
        # DC-1 W2: ASIL audit가 손실 데이터를 식별하도록 잘리는 컬럼명을 보존.
        _dropped = (
            in_names[input_max:] + exp_names[expected_max:] + act_names[actual_max:]
        )
        in_names = in_names[:input_max]
        exp_names = exp_names[:expected_max]
        act_names = act_names[:actual_max]
        if (_orig_in > input_max or _orig_exp > expected_max
                or _orig_act > actual_max):
            _truncated_fn_count += 1
            if out_warnings is not None:
                _drop_show = ", ".join(_dropped[:8])
                if len(_dropped) > 8:
                    _drop_show += f" …(+{len(_dropped) - 8})"
                out_warnings.append(
                    f"[truncate] 함수 {fn_id} 변수 컬럼 한도 초과(접기 후) — "
                    f"input {_orig_in}→{len(in_names)}, "
                    f"expected {_orig_exp}→{len(exp_names)}, "
                    f"actual {_orig_act}→{len(act_names)} (한도 "
                    f"{input_max}/{expected_max}/{actual_max}; "
                    f"raw {len(in_raw)}/{len(exp_raw)}/{len(act_raw)}; "
                    f"누락 컬럼: {_drop_show})"
                )

        anchor_r = cur_row
        block_height = block_heights[fn_id]

        # --- anchor row ---
        tc_id, title, method, precond_col, precond, swuts_entry = _resolve_anchor_meta(
            rep_tc_name, component_name
        )
        safe_write(ws, anchor_r, col, tc_id)          # B = TC ID
        safe_write(ws, anchor_r, col + 1, title)      # C = Title
        safe_write(ws, anchor_r, col + 2, method)     # D = Generation Method
        if precond_col and precond:
            safe_write(ws, anchor_r, precond_col, precond)
        # 변수명을 각 섹션 열에 나열.
        for pi, var_name in enumerate(in_names):
            safe_write(ws, anchor_r, INPUT_COL + pi, var_name)
        for pi, var_name in enumerate(exp_names):
            safe_write(ws, anchor_r, EXPECTED_COL + pi, var_name)
        for pi, var_name in enumerate(act_names):
            safe_write(ws, anchor_r, ACTUAL_COL + pi, var_name)
        # AK(Total) — 함수 전체 Pass/Fail. anchor row에 1회 stamp 후 블록 세로 병합.
        if PASS_FAIL_TOTAL_COL > 0:
            safe_write(ws, anchor_r, PASS_FAIL_TOTAL_COL, total_result_str)
            if block_height > 1:
                _deferred_merges.append(
                    (anchor_r, anchor_r + block_height - 1,
                     PASS_FAIL_TOTAL_COL, PASS_FAIL_TOTAL_COL)
                )
        # Log Data — env/함수 단위 log path 추정 (anchor row).
        if env0 is not None and getattr(env0, "env_name", ""):
            safe_write(ws, anchor_r, LOG_DATA_COL, f"{env0.env_name}/{fn_id}.log")

        # ASIL 시각 강조 — anchor row Total col (AK) 또는 Unit col (AJ).
        asil = _resolve_fn_asil(fn_id, rep_tc_name, swuts_entry)
        _asil_marker = {
            "A": mark_asil_a_function,
            "B": mark_asil_b_function,
            "C": mark_asil_c_function,
            "D": mark_asil_d_function,
            "QM": mark_asil_qm_function,
        }.get(asil)
        if _asil_marker:
            _mark_col = PASS_FAIL_TOTAL_COL if PASS_FAIL_TOTAL_COL > 0 else PASS_FAIL_UNIT_COL
            _asil_marker(ws, anchor_r, _mark_col)

        # --- iteration rows ---
        for it_idx, ik in enumerate(iter_keys, start=1):
            ir = anchor_r + it_idx
            ev = tc_to_env.get(ik)
            tc_items = ev.test_cases.get(ik) or [] if ev is not None else []
            tc_item = tc_items[0] if tc_items else None
            exec_r = ev.test_results.get(ik) if ev is not None else None

            safe_write(ws, ir, 5, it_idx)  # E = iteration index

            input_data = getattr(tc_item, "input_data", {}) or {} if tc_item else {}
            expected_data = (
                getattr(tc_item, "expected_result", {}) or {} if tc_item else {}
            )
            actual_dict: dict = {}
            if exec_r is not None:
                actual_dict = getattr(exec_r, "actual_result", {}) or {}
            if not actual_dict and ev is not None:
                tr_items = getattr(ev, "tc_result_items", {}).get(ik, [])
                tr_item = tr_items[0] if tr_items else None
                if tr_item is not None:
                    actual_dict = getattr(tr_item, "actual_result", {}) or {}

            # DC-1: collapsed 컬럼은 group 멤버를 모아 1셀로(format_values/_actual),
            # 일반 컬럼은 기존대로 dict 단일 조회. 접기 없으면 grp=None → 기존 동작 그대로.
            for pi, col_name in enumerate(in_names):
                grp = in_info.get_group(col_name)
                if grp is not None:
                    val = _collapsed_value_cell(grp, input_data)
                else:
                    v = input_data.get(col_name, "")
                    val = str(v) if v else ""
                safe_write(ws, ir, INPUT_COL + pi, val)
            for pi, col_name in enumerate(exp_names):
                grp = exp_info.get_group(col_name)
                if grp is not None:
                    val = _collapsed_value_cell(grp, expected_data)
                else:
                    v = expected_data.get(col_name, "")
                    val = str(v) if v else ""
                safe_write(ws, ir, EXPECTED_COL + pi, val)
            for pi, col_name in enumerate(act_names):
                grp = act_info.get_group(col_name)
                if grp is not None:
                    val = _collapsed_actual_cell(grp, actual_dict)
                else:
                    t = actual_dict.get(col_name, "")
                    v = t[0] if isinstance(t, tuple) and t else (str(t) if t else "")
                    val = str(v) if v else ""
                safe_write(ws, ir, ACTUAL_COL + pi, val)

            # AJ(Unit) — 그 iteration의 Pass/Fail.
            iter_result = (
                "Pass" if exec_r and exec_r.passed else
                "Fail" if exec_r else
                "N/A"
            )
            safe_write(ws, ir, PASS_FAIL_UNIT_COL, iter_result)

            iteration_rows += 1

        # --- 신규 블록 style/merge 복제 (template 영역 밖) ---
        # 이전 per-TC 6 row 고정 → 함수 블록 가변 높이. template block 1행
        # (start_row)의 style/merge를 블록 각 row에 복제.
        if anchor_r > start_row + tc_row_step - 1 and tc_row_step >= 1:
            for off in range(block_height):
                dst_row = anchor_r + off
                # template block 내 대응 offset (블록 높이 > template step이면 wrap).
                tpl_off = off % tc_row_step
                for c_n in range(1, _tpl_max_col + 1):
                    tpl_style = _tpl_block_styles.get((tpl_off, c_n))
                    if tpl_style is not None:
                        ws.cell(dst_row, c_n)._style = _copy.copy(tpl_style)
            for off_start, off_end, mc_min_col, mc_max_col in template_merges_local:
                _deferred_merges.append(
                    (anchor_r + off_start, anchor_r + off_end, mc_min_col, mc_max_col)
                )

        cur_row += block_height
        written += 1

    # 진단 로그.
    if out_warnings is not None:
        out_warnings.append(
            f"[diag] R90 함수 블록 stamp: functions={written}, "
            f"iteration_rows={iteration_rows}, truncated_fn={_truncated_fn_count}, "
            f"last_row={cur_row - 1}"
        )

    # 라운드 89: defer한 merge 일괄 적용 (write 루프 종료 후).
    for new_min_r, new_max_r, mc_min_col, mc_max_col in _deferred_merges:
        try:
            ws.merge_cells(
                start_row=new_min_r, end_row=new_max_r,
                start_column=mc_min_col, end_column=mc_max_col,
            )
        except (ValueError, AttributeError):
            pass

    # 라운드 F7 T707: clear policy — 신규 stamp 후 다음 row부터 양식 default clear.
    # 라운드 90: clear_start는 함수 블록 누적 높이 기준 (cur_row).
    if written > 0:
        try:
            from backend.services.excel_template_utils import clear_data_range
            clear_start = cur_row
            clear_end = ws.max_row
            if clear_end >= clear_start:
                # F7 자체평가 R1 C1/C3 fix: col range 1~20 → 1~40 확장 (양식 default
                # 'Pass' 텍스트가 col 36/37에 prefill되어 잔존). sentinel_patterns로
                # '< End of Document >' / '■ Appendix' 같은 양식 끝 마커 보존.
                cleared = clear_data_range(
                    ws,
                    start_row=clear_start, end_row=clear_end,
                    start_col=1, end_col=40,
                    preserve_formula=True, preserve_merged_anchor=True,
                    sentinel_patterns=[
                        "End of Document", "< End", "■ Appendix",
                        "Appendix", "※",
                    ],
                )
                if out_warnings is not None and cleared > 0:
                    out_warnings.append(
                        f"[clear] Test Log/Result row {clear_start}~{clear_end} "
                        f"양식 default {cleared} cell clear (partial overwrite 차단)"
                    )
        except ImportError:
            pass
    return written


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_sutr(
    session: SwUTSession,
    meta: SutrBuildMeta,
    template_bytes: bytes,
    deviation_cases: list[Any] | None = None,
    swuds_function_ids: set[str] | None = None,
    swuts_map: dict[str, Any] | None = None,
) -> SutrBuildResult:
    """SUTR v3.01 xlsm 생성.

    Args:
        session: input_adapter 출력.
        meta: 빌드 메타.
        template_bytes: 기존 v3.01 xlsm 파일 bytes.
        deviation_cases: swut_deviation_generator 결과 (None이면 빈 Deviation 시트).
        swuds_function_ids: 17차 — SwUDS 함수 ID set (옵션). 제공되면 2.Consistency에
            SwUDS↔SwUTS 매핑 row 추가. Coverage builder와 대칭.
        swuts_map: 60차 F6-A — SwUTS xlsm parser 결과 (옵션). 제공되면 Test Log의
            col B/C/D + Precondition col에 spec docx 데이터 stamp. None이면 기존
            하드코딩 ("AEC, ABV") fallback (backward-compat).
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for SwUT SUTR builder")

    # deep-reviewer X3 + 5차 H1/H2: 입력 메타 종합 검증.
    validate_build_meta(
        meta.release_sw_version, meta.test_date,
        doc_id_sequence=meta.doc_id_sequence,
        test_engineer=meta.test_engineer,
        author=meta.default_author,
        approver=meta.default_approver or meta.approver_override,
        reviewer=meta.default_reviewer or meta.reviewer_override,
    )
    # Critical (reviewer S): ZIP bomb / magic byte 검증.
    validate_xlsx_template_bytes(template_bytes, label="SUTR template")

    # 5차 L1 (ISO F3 추적성): 입력 template hash — audit 시 입력 동일성 검증용.
    template_sha256_12 = hashlib.sha256(template_bytes).hexdigest()[:12]

    # 37차 fix → 38차 W1 DRY: extract_warnings_from_session helper로 추출.
    warnings: list[str] = extract_warnings_from_session(session)

    # 라운드 73 T816 — 입력 자산 활용도 진단.
    from backend.services.swut_builder_helpers import diagnose_asset_usage
    warnings.extend(diagnose_asset_usage(
        swuts_map=swuts_map,
        c_function_map=session.c_function_map or None,
        swuds_function_map=session.swuds_function_map or None,
    ))

    # 54-fix C1: SwUT 라우터에 v2.02 template 잘못 입력 시 silent 빈 셀 차단.
    # v3.01 SUTR 양식은 fallback_to_v301=True로 hardcode 동작과 동등.
    from backend.services.excel_layout_resolver import inspect_swit_layout
    layout = inspect_swit_layout(template_bytes, "sitr")
    if layout.warnings:
        warnings.extend([f"[layout] {w}" for w in layout.warnings])

    # deep-reviewer W2: VBA 매크로 ZIP entry 존재 여부 사전 측정.
    template_has_vba = has_vba_macros(template_bytes)
    vba_refs_found: list[str] = []
    if template_has_vba:
        warnings.append(
            "VBA macro execution NOT verified — open output xlsm in Excel and verify "
            "macros before submitting as evidence (ZIP entry preserved but stale ref 위험)"
        )
        # 5차 reviewer I1: VBA stale ref 의심 패턴 grep.
        vba_refs_found = inspect_vba_refs(template_bytes)
        if vba_refs_found:
            warnings.append(
                f"VBA stale ref 위험 패턴 발견 — {vba_refs_found} 패턴이 vbaProject.bin에 "
                "존재하며 셀/시트 이동 시 매크로 깨질 위험 (수동 검증 의무)"
            )

    # keep_vba=True — .xlsm 매크로 보존
    wb: Workbook = openpyxl.load_workbook(
        io.BytesIO(template_bytes), keep_vba=True, data_only=False,
    )
    sheet_names = wb.sheetnames

    agg = aggregate_session(session)

    # 30차 W21 + 31차 W29 + 라운드 84 T1801 + 85 T1903 + 86 T2001: unmapped fc list.
    from backend.services.swut_coverage_aggregator import _compute_asil_distribution
    asil_distribution, ids_by_asil, unmapped_fns = _compute_asil_distribution(
        agg.get("function_rows") or [],
        agg.get("function_asil_map") or {},
        function_asil_from_suds=agg.get("function_asil_from_suds"),
        component_asil_from_sds=agg.get("component_asil_from_sds"),
        function_asil_from_srs=agg.get("function_asil_from_srs"),
        function_name_to_swufn_from_suds=agg.get("function_name_to_swufn_from_suds"),
    )

    summary = {
        "environments": len(session.environments),
        "total": agg["total"],
        "tested": agg["tested"],
        "passed": agg["passed"],
        "failed": agg["failed"],
        "deviation_cases_written": 0,
        "test_log_rows_written": 0,
        # 30차 W21 + 31차 W29: Coverage builder와 동일 키 — UI 노출 통일.
        "asil_distribution": asil_distribution,
        "asil_b_function_ids": ids_by_asil.get("B", []),
        "asil_c_function_ids": ids_by_asil.get("C", []),
        "asil_d_function_ids": ids_by_asil.get("D", []),
        # 라운드 86 T2002: UNKNOWN 함수 list (audit 진단용).
        "unmapped_function_names": unmapped_fns,
        # 31-fix D15: audit 공지 메타 — Coverage builder와 대칭.
        "asil_highlight_policy": (
            "B=파랑(#E2F0FF) / C=주황(#FFE5CC) / D=빨강(#FFC7CE) — "
            "31차 비표준 audit 확장 (회사 v3.01 양식은 빨강만 사용)"
        ),
    }

    cover_ws = next((wb[n] for n in sheet_names if n.lower() == "cover"), None)
    if cover_ws is None:
        warnings.append("Cover 시트 미발견")
    else:
        # 54-fix C1: layout 전달
        _write_cover(cover_ws, meta, out_warnings=warnings, layout=layout)

    # 54-fix C1: 53차 SwIT 패턴과 대칭 — v2.02 "1.Test Summary" 등 prefix 호환 substring.
    ts_ws = next((wb[n] for n in sheet_names if "test summary" in n.lower()), None)
    if ts_ws is None:
        warnings.append("Test Summary 시트 미발견")
    else:
        # 54-fix C1: layout + summary 전달
        _write_test_summary(
            ts_ws, meta, agg, out_warnings=warnings,
            layout=layout, summary=summary,
        )

    # 54-fix C1: substring 대칭
    # 59차 F4-C: KJPDS02 v1.01 양식은 Deviation 시트 없음 — layout.deviation_sheet_present
    # False면 시트 미발견을 정상으로 처리 + parse_warnings 안내 메시지 변경.
    dev_ws = next((wb[n] for n in sheet_names if "deviation" in n.lower()), None)
    deviation_required = (
        layout is None or getattr(layout, "deviation_sheet_present", True)
    )
    if dev_ws is None:
        if deviation_required:
            warnings.append("Deviation 시트 미발견")
        else:
            warnings.append(
                "Deviation 시트 미발견 — v1.01 양식 (정상, layout.deviation_sheet_present=False)"
            )
    elif deviation_cases:
        n = _write_deviation(dev_ws, deviation_cases, out_warnings=warnings)
        summary["deviation_cases_written"] = n

    # 57차 T314 fix: 회사 v2.02 SUTR 양식 시트명이 '3.Test Result' (Log 아닌 Result).
    # 54-fix C1 substring 매칭에 'test result'도 포함 — SwUT SUTR (v2.02 Result) /
    # SwIT SITR (Log) 모두 호환.
    log_ws = next(
        (wb[n] for n in sheet_names
         if "test log" in n.lower() or "test result" in n.lower()),
        None,
    )
    if log_ws is None:
        warnings.append("Test Log/Result 시트 미발견")
    else:
        # 54-fix C1: layout 전달 — AL marker
        # 라운드 78 T1303: c_function_map 전달 — ASIL fallback (asil_map 빈 dict 시
        # c_parser comment_asil 매핑으로 ASIL 강조 적용).
        n = _write_test_log(
            log_ws, session,
            function_asil_map=agg.get("function_asil_map"),
            out_warnings=warnings,
            layout=layout,
            swuts_map=swuts_map,
            c_function_map=session.c_function_map or None,
        )
        summary["test_log_rows_written"] = n

    incomplete_sheets: list[str] = []
    # History — 55-fix: 사용자 결정 B (single-row release entry).
    hist_ws = next((wb[n] for n in sheet_names if n.lower() == "history"), None)
    if hist_ws is not None:
        from backend.services.swut_coverage_aggregator import _write_history_sheet
        # 55-fix-2 W2: SwUT prefix 명시 (vs SwIT SITR)
        # 55-fix-2 W6: out_warnings 전달
        release_rows = build_release_history_row(
            meta, doc_kind="SwUT SUTR", out_warnings=warnings,
        )
        n_h = _write_history_sheet(hist_ws, release_rows, out_warnings=warnings)
        summary["history_rows_written"] = n_h
        if n_h == 0:
            incomplete_sheets.append("History")

    # 17차 T171: 2.Consistency 시트 — Coverage builder와 대칭.
    # SUTR 템플릿에 시트가 없으면 silent skip (Hyundai 양식 변형 안전).
    cons_ws = next((wb[n] for n in sheet_names if "consistency" in n.lower()), None)
    if cons_ws is not None:
        from backend.services.swut_coverage_aggregator import _write_consistency_sheet
        n_cons = _write_consistency_sheet(
            cons_ws, session,
            swuds_function_ids=swuds_function_ids,
            out_warnings=warnings,
        )
        summary["consistency_self_check_rows"] = n_cons
        if swuds_function_ids is not None:
            summary["consistency_swuds_compared"] = True
        else:
            summary["consistency_swuds_compared"] = False
            incomplete_sheets.append("2.Consistency (SwUDS 비교 partial — v3.02)")

    summary["template_sha256_12"] = template_sha256_12
    summary["build_timestamp"] = meta.build_timestamp

    # 라운드 83 T1703: AuditLog 시트 신규 추가 (Coverage 대칭, 회사 양식 영향 0).
    try:
        from backend.services.swut_coverage_aggregator import _write_audit_log_sheet
        if "AuditLog" not in wb.sheetnames:
            audit_ws = wb.create_sheet("AuditLog")
            n_audit = _write_audit_log_sheet(
                audit_ws, meta, summary, agg, session, warnings,
            )
            summary["audit_log_rows_written"] = n_audit
            summary["audit_log_sheet_added"] = True
    except Exception as _e:  # pragma: no cover — fail-safe
        warnings.append(
            f"AuditLog 시트 작성 실패 (산출물 영향 0): {type(_e).__name__}: {str(_e)[:80]}"
        )

    # 라운드 107 — 템플릿/기입 수식을 openpyxl이 캐시 미저장(cached=None) → 재계산
    # 안 하는 뷰어에서 공백. fullCalcOnLoad로 열 때 자동 재계산(SwITCV 라운드 102 정합).
    # 캐시 미저장은 불변이라 data_only 다운스트림 영향 0.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # pragma: no cover — openpyxl 버전 차 방어
        pass

    # 14차 W1: BytesIO 그대로 result에 저장 — getvalue() copy 회피.
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    wb.close()

    # 라운드 89: config doc_filenames[sutr] 패턴 우선 ({version}/{date} 치환).
    # 빈 값이면 HDPDM01 v3.01 하드코딩 default (backward compat).
    if meta.doc_filename_pattern:
        filename = meta.doc_filename_pattern.format(
            version=meta.release_sw_version, date=short_date(meta.test_date),
        )
    else:
        filename = (
            f"({meta.project_id}_SUTR) Software Unit Test Result_"
            f"v{meta.release_sw_version}_{short_date(meta.test_date)}_R.xlsm"
        )

    return SutrBuildResult(
        ok=True,
        xlsm_io=out,
        filename=filename,
        warnings=warnings,
        incomplete_sheets=incomplete_sheets,
        vba_macros_preserved=template_has_vba,
        summary=summary,
    )


# `short_date`는 excel_template_utils에서 import — 모듈 하단 중복 정의 제거 (deep-reviewer C1).


# 38차 W2 — public API 명시. SwIT SITR aggregator가 본 모듈의 _write_cover /
# _write_test_summary / _write_deviation / _write_test_log private 함수들을 직접
# import 중 (35차 SwIT 라운드 강결합 보류). 본 __all__로 강결합 경계를 명시화:
# signature 변경 시 SwIT 회귀(test_swit_sitr_aggregator.py) 동시 검증 의무.
__all__ = [
    # Public API
    "SutrBuildMeta",
    "SutrBuildResult",
    "build_sutr",
    # Private 함수 — SwIT SITR이 import 중. 명시로 강결합 가시화.
    "_write_cover",
    "_write_deviation",
    "_write_test_log",
    "_write_test_summary",
]
