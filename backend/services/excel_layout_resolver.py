"""54차 — SwIT v2.02 / SwUT v3.01 양식 layout inspect 모듈.

회사 양식 (HDPDM01 NE_GN7) Coverage Report / SUTR / SITR 등 xlsx/xlsm template을
openpyxl로 inspect하여 label↔cell 매핑 + 신규 row(B17 TC stats, B22 Requirements,
Test Log AL marker) 위치를 자동 추출. v2.02 양식 빈 cell 누락 fix (54차 사용자 보고).

## 매칭 전략

각 정보 항목은 v2.02 후보 label을 우선 매칭 → 못 찾으면 v3.01 후보 label 매칭 →
모두 실패면 warnings 누적 + fallback_to_v301=True.

| 정보 | v2.02 후보 | v3.01 후보 |
|------|------------|------------|
| release_sw_version | "SW Version", "SW 버전" | "Release Name(SW)" |
| hw_version | "HW Version", "HW 버전" | "Test Target Version(HW)" |
| project_full_name | "Project Name", "Project" | 동일 |
| test_date | "Test Date", "테스트 일자" | 동일 |
| test_engineer | "Test Engineer", "테스트 엔지니어" | 동일 |
| final_test_result | "Final Test Result", "최종 결과" | 동일 |
| tc_stats_row (v2.02) | "Total TC", "Test Case Count", "TC Count" | (없음) |
| requirements_row (v2.02) | "Requirements/Design Coverage", "Requirements Coverage" | (없음) |
| test_log_extra_marker_col | header row "Marker" / "Pass/Fail Marker" | (없음) |

## 캐싱

sha256 keying + maxsize=4 LRU. 회사 양식 4개 (SwUT Cov / SUTR / SwIT Cov / SITR)
범위 내. 동일 template 반복 빌드 시 openpyxl load_workbook 1회만.

## ISO 26262 audit 영향

본 모듈은 audit 산출물 정확도 향상 (v2.02 빈 cell fill) — evidence_class 변경 없음.
manual review 의무 동일.
"""
from __future__ import annotations

import hashlib
import io
import re
import threading
from dataclasses import dataclass, field
from typing import Literal, Optional

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

# 54-fix C2: ZIP bomb 방어 backstop — caller에 의존하지 않고 본 모듈에서도 검증.
from backend.services.excel_template_utils import (
    TemplateValidationError,
    validate_xlsx_template_bytes,
)


@dataclass(frozen=True)
class CellRef:
    """label 셀의 (row, col) — 1-based. value cell은 머지 보정 후 col+1."""
    row: int
    col: int


@dataclass
class SwitLayout:
    """v2.02/v3.01 template inspect 결과.

    Attributes:
        detected_version: 매칭된 라벨 비율에 따른 추정 ("v2.02" / "v3.01" / "unknown").
        fallback_to_v301: True면 v2.02 후보 다수 실패 → v3.01 hardcode fallback 사용 권장.
        cover_labels: {"project_full_name": "Project", "asil_level": "ASIL Level", ...}.
            값은 template에서 실제 매칭된 label text — writer 함수가 이 라벨로 find_kv_row.
        test_summary_labels: 동일 패턴. v2.02면 "SW Version", v3.01이면 "Release Name(SW)".
        tc_stats_row: **55-fix 이후 의미 = data row** (label_row + 1). 회사 v2.02 SITR
            양식은 라벨이 row 17 가로 배치 (B17~F17), 데이터는 row 18에 위치. writer가
            본 row에 직접 fill. None이면 v3.01 (해당 row 없음). 이전 (55-fix 이전)
            의미는 "label row"였음 — semantic breaking change. 55-fix-2 W1 docstring 갱신.
        tc_stats_col_start: TC stats row의 첫 데이터 셀 col. **55-fix 이후 = label_col**
            (가로 배치라 라벨 col부터 데이터 시작). 이전 = label_col + 1.
        requirements_row: v2.02의 Requirements/Design Coverage row index. None이면 v3.01.
            **주의**: 회사 v2.02 SITR 양식은 row 22에 default 'SwITS' 채워져 있어
            우리 코드가 추가 fill 필요 없음. `'■  Requirements/Design Coverage'` 후보
            추가 시 row 20 (헤더) 매칭되어 덮어쓰기 위험 — 추가 금지 (W5).
        deviation_header_cell: SITR Deviation 시트 헤더 위치 (참고용, 현재 미사용).
        test_log_header_cell: SITR Test Log 시트 헤더 위치 (참고용).
        test_log_extra_marker_col: v2.02의 추가 marker col (예: AL col). None이면 미적용.
        tc_stats_label_missing: **56차 신규** — v2.02 Coverage 양식 fallback 감지. 회사
            Coverage Report v2.02 양식은 row 17이 빈 row (사용자 수동 입력 영역)라
            label 매칭 실패. 본 flag True면 writer가 라벨도 함께 stamp. SITR은 label
            존재 → 항상 False (기존 path 유지).
        requirements_label_missing: 동일 패턴. v2.02 Coverage가 B20에 헤더 부재 시 True.
        test_log_tc_row_step: **57차 T314 신규** — Test Log/Test Result 시트의 1 TC당
            row 수. 회사 v2.02 SUTR/SITR 양식은 6 (TC ID + Params 1~5 sub-row).
            v3.01 또는 기본은 1. inspect 시 template의 TC ID 열 (B)에서 첫 2개 TC
            row 위치 차이로 동적 감지. SUTR/SITR _write_test_log가 이 step 적용해
            row 5, 11, 17, ... 에 TC ID stamp.
        test_log_input_col: **58차 F3 신규** — Test Log 'Input' / 'Input Parameters'
            라벨 col. SwIT SITR v2.02 = 8 (H), SwUT SUTR v3.01 = 6 (F). None이면
            v3.01 hardcode fallback (6).
        test_log_expected_col: 'Expected Result' col. v2.02 = 18 (R), v3.01 = 16 (P).
        test_log_actual_col: 'Actual Result' col. v2.02 = 28 (AB), v3.01 = 26 (Z).
        test_log_pass_fail_col: 'Pass/Fail' / 'Pass/Fail Unit' col. v2.02 SITR = 38
            (AL), v3.01 SUTR = 36 (AJ).
        test_log_pass_fail_total_col: 'Pass/Fail Total' col. v3.01 SUTR only = 37 (AK).
            v2.02 SITR은 단일 'Pass/Fail' col만 사용 → None.
        test_log_log_data_col: 'Log Data' / 'Log' col. v2.02 SITR = 40 (AN), v3.01 SUTR
            = 38 (AL).
        traceability_header_row: **58차 F2 신규** — 1.Traceability 시트 헤더 row
            (SwUFn_ 또는 SwUTC_ prefix 컬럼이 다수 수평 배열된 row). SwIT v2.02 양식
            은 v3.01보다 더 아래쪽 (row 20~25) 위치. None이면 writer가 자동 탐색.
        warnings: inspect 중 누락 라벨 / 시트 미발견 등 메시지.
    """
    detected_version: Literal["v1.01", "v2.02", "v3.01", "unknown"] = "unknown"
    fallback_to_v301: bool = False
    cover_labels: dict[str, str] = field(default_factory=dict)
    test_summary_labels: dict[str, str] = field(default_factory=dict)
    tc_stats_row: Optional[int] = None
    tc_stats_col_start: Optional[int] = None
    requirements_row: Optional[int] = None
    deviation_header_cell: Optional[CellRef] = None
    test_log_header_cell: Optional[CellRef] = None
    test_log_extra_marker_col: Optional[int] = None
    tc_stats_label_missing: bool = False
    requirements_label_missing: bool = False
    test_log_tc_row_step: int = 1
    # 58차 F3 — Test Log 시트 column 위치 (None이면 v3.01 hardcode fallback)
    test_log_input_col: Optional[int] = None
    test_log_expected_col: Optional[int] = None
    test_log_actual_col: Optional[int] = None
    test_log_pass_fail_col: Optional[int] = None
    test_log_pass_fail_total_col: Optional[int] = None
    test_log_log_data_col: Optional[int] = None
    # 60차 F6-A — Test Log Precondition 컬럼 위치 (SwUTS/SwITS spec stamp용).
    # KJPDS02 SwITS v1.01 = col 9, HDPDM01 SUTS v3.01 = col 10, SITS v2.02 = col 6.
    # KJPDS02 SwUTS v1.01 양식은 precondition col 없음 → None.
    # None이면 _write_test_log에서 Precondition stamp skip (backward-compat).
    test_log_precondition_col: Optional[int] = None
    # 58차 F2 — Coverage 1.Traceability 시트 헤더 row 위치
    traceability_header_row: Optional[int] = None
    # 59차 F4-A — Test Log 변수명 헤더 row (KJPDS02 v1.01 = row 5).
    # 매 row col block에 input/expected/actual 변수명을 stamp하는 row 위치.
    # None이면 양식이 변수명 헤더를 요구하지 않음 — v2.02/v3.01 기본 동작 유지 (skip).
    test_log_variable_header_row: Optional[int] = None
    # 59차 F4-A — Input/Expected/Actual 각 column block 최대 변수 수.
    # _write_test_log truncate [:10] 대신 layout 기반 동적 사용. _scan_test_log_max_counts
    # 가 expected_col - input_col 등으로 산출. v3.01 기본 10. KJPDS02 v1.01은 양식
    # max_col=378 으로 ~80~100 변수 stamp 가능.
    test_log_input_max_count: int = 10
    test_log_expected_max_count: int = 10
    test_log_actual_max_count: int = 10
    # 59차 F4-A — Test Log step 배치 형태.
    # "step_in_rows": v2.02/v1.01 양식 — TC당 N row 분배 (tc_row_step > 1)
    # "single_row":   v3.01 양식 — TC당 1 row (tc_row_step == 1)
    # _inspect_internal에서 tc_row_step 값에 따라 자동 결정.
    test_log_step_layout: str = "single_row"
    # 59차 F4-C — KJPDS02 v1.01 양식 시트 구성 분기 (Deviation/Traceability matrix/
    # Coverage metric 다른 양식 패밀리). 모든 default는 v2.02/v3.01 backward-compat.
    sitr_sheet_count: int = 5
    """SITR (xlsm) 시트 수. v1.01=4 (Deviation 시트 없음), v2.02/v3.01=5."""
    coverage_sheet_count: int = 4
    """Coverage (xlsx) 시트 수. v1.01=6 (Cover/History/1.Test Summary/2.Traceability/
    3.Consistency/4.Coverage), v2.02/v3.01=4."""
    traceability_matrix_kind: Literal["swufn_x_env", "switc_x_swst"] = "swufn_x_env"
    """Traceability matrix 차원. v1.01=switc_x_swst (SwITC ID × SwST/SwSTR 항목),
    v2.02/v3.01=swufn_x_env (SwUFn × 환경 매트릭스)."""
    coverage_metric_kind: Literal["single", "function_and_calls"] = "single"
    """Coverage metric 분리. v1.01=function_and_calls (Functions row 5 / Function Calls
    row 6 별도), v2.02/v3.01=single (단일 Coverage row)."""
    deviation_sheet_present: bool = True
    """SUTR/SITR Deviation 시트 존재 여부. v1.01=False (Test Log만), v2.02/v3.01=True."""
    test_summary_coverage_breakdown: int = 1
    """1.Test Summary의 Coverage breakdown count. v1.01=4 (추적성/정합성/Function/
    FunctionCalls 4가지 별도 표시), v2.02/v3.01=1 (단일 Coverage)."""
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Candidate label 후보 (v2.02 → v3.01 fallback)
# ---------------------------------------------------------------------------

_COVER_CANDIDATES: dict[str, tuple[str, ...]] = {
    # field_name → (v2.02 후보, ..., v3.01 후보)
    "project_full_name": ("Project", "프로젝트"),
    "asil_level": ("ASIL Level", "ASIL"),
    "status": ("Status", "상태"),
    "validation_date": ("Validation Date", "검증 일자", "검증일"),
    "author": ("Author", "작성자"),
    "reviewer": ("Reviewer", "검토자"),
    "approver": ("Approver", "승인자"),
    "doc_id": ("Doc. ID", "Doc ID", "문서 번호"),
    "version": ("Version", "버전"),
    "build_timestamp": ("Build Timestamp", "빌드 시간"),
}

_TEST_SUMMARY_CANDIDATES: dict[str, tuple[str, ...]] = {
    "project_full_name": ("Project Name", "Project", "프로젝트명"),
    # v2.02 "SW Version" 우선, v3.01 "Release Name(SW)" fallback
    "release_sw_version": ("SW Version", "SW 버전", "Release Name(SW)"),
    "hw_version": ("HW Version", "HW 버전", "Test Target Version(HW)"),
    "test_date": ("Test Date", "테스트 일자", "시험 일자"),
    "test_engineer": ("Test Engineer", "테스트 엔지니어", "시험 담당"),
    "final_test_result": ("Final Test Result", "최종 결과", "최종 시험 결과"),
    # SUTR/SITR 전용
    "target_coverage": ("Target Coverage", "목표 커버리지"),
    "actual_coverage": ("Actual Coverage", "실측 커버리지"),
    "target_pass_ratio": ("Target Pass ratio", "Target Pass Ratio", "목표 Pass 비율"),
    "actual_pass_ratio": ("Actual Pass ratio", "Actual Pass Ratio", "실측 Pass 비율"),
}

# v2.02 양식 신규 row label 후보 — 55-fix: 사용자 실 산출물 보고로 'Total Number of TCs'
# 추가. 회사 v2.02 SITR 양식의 실 라벨이 candidate-tuple에 없어 row 검출 실패했음.
_TC_STATS_LABELS: tuple[str, ...] = (
    "Total Number of TCs",  # v2.02 SITR 실 양식 (55-fix 사용자 보고)
    "Total TC", "Test Case Count", "TC Count",
    "전체 TC", "TC 수",
)
_REQUIREMENTS_LABELS: tuple[str, ...] = (
    "Requirements/Design Coverage",
    "Requirements Coverage",
    "Design Coverage",
    "요구사항/설계 커버리지",
)
# 55-fix 주의: 회사 v2.02 SITR 양식은 row 20에 '■  Requirements/Design Coverage'
# 헤더 + row 22에 이미 'SwITS' default 채워져 있음. ■ prefix 후보를 추가하면
# requirements_row=20이 검출되어 우리 writer가 라벨 row를 덮어씀. 따라서 추가 금지.
# 만약 향후 회사 양식이 SwITS default 안 갖는 양식 도입 시 별도 검토.
_TEST_LOG_MARKER_LABELS: tuple[str, ...] = (
    "Marker", "Pass/Fail Marker", "검수 표시", "통과 표시",
)


# ---------------------------------------------------------------------------
# Inspect helpers
# ---------------------------------------------------------------------------

def _find_sheet(wb, name_predicate) -> object | None:
    """workbook에서 predicate(name) True인 첫 시트 반환."""
    for n in wb.sheetnames:
        if name_predicate(n):
            return wb[n]
    return None


def _scan_label_cell(
    ws, candidates: tuple[str, ...], *, max_row: int = 60,
) -> tuple[Optional[CellRef], Optional[str]]:
    """시트의 첫 N행에서 candidates 중 하나 매칭되는 cell + 매칭 label 반환.

    매칭은 cell.value.strip() == candidate 정확 비교 (대소문자 구분).
    """
    if ws is None:
        return (None, None)
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue
            for cand in candidates:
                if s == cand:
                    return (CellRef(row=cell.row, col=cell.column), cand)
    return (None, None)


def _scan_labels_in_sheet(
    ws, candidates_map: dict[str, tuple[str, ...]], *, max_row: int = 60,
) -> tuple[dict[str, str], dict[str, CellRef]]:
    """시트에서 여러 field_name별 candidate matching.

    Returns:
        (labels_found, cells_found):
            labels_found: {field_name: matched_label_text}
            cells_found: {field_name: CellRef(row, col)} (label cell — value는 col+1)
    """
    labels: dict[str, str] = {}
    cells: dict[str, CellRef] = {}
    for field_name, cands in candidates_map.items():
        cell_ref, matched = _scan_label_cell(ws, cands, max_row=max_row)
        if cell_ref and matched:
            labels[field_name] = matched
            cells[field_name] = cell_ref
    return labels, cells


def _detect_version_from_labels(
    test_summary_labels: dict[str, str],
) -> Literal["v2.02", "v3.01", "unknown"]:
    """Test Summary 라벨로 양식 버전 추정.

    "SW Version" / "HW Version" 매칭 → v2.02. "Release Name(SW)" / "Test Target
    Version(HW)" 매칭 → v3.01.
    """
    release_label = test_summary_labels.get("release_sw_version", "")
    hw_label = test_summary_labels.get("hw_version", "")
    v202_signals = sum(
        1 for lbl in (release_label, hw_label)
        if lbl in ("SW Version", "SW 버전", "HW Version", "HW 버전")
    )
    v301_signals = sum(
        1 for lbl in (release_label, hw_label)
        if lbl in ("Release Name(SW)", "Test Target Version(HW)")
    )
    if v202_signals > v301_signals:
        return "v2.02"
    if v301_signals > v202_signals:
        return "v3.01"
    return "unknown"


def _scan_tc_stats_row(ws) -> tuple[Optional[int], Optional[int]]:
    """Test Summary 시트에서 TC stats data row 위치 탐색.

    55-fix: 회사 v2.02 SITR 양식은 라벨 row가 헤더 (B17~F17 가로 배치)이고
    data는 label_row + 1 (B18~F18)에 위치. label_row를 그대로 반환하면 writer가
    label을 덮어씀. 본 함수는 **data row** 반환.

    또한 col_start = label_col (라벨이 가로 배치라 첫 라벨 col부터 데이터 시작).

    Returns:
        (tc_stats_data_row, tc_stats_col_start) — 둘 다 None이면 v3.01 (없음).
    """
    cell_ref, _ = _scan_label_cell(ws, _TC_STATS_LABELS, max_row=60)
    if cell_ref is None:
        return (None, None)
    # data row = 라벨 row + 1, 데이터 시작 col = 라벨 col (가로 배치)
    return (cell_ref.row + 1, cell_ref.col)


def _scan_requirements_row(ws) -> Optional[int]:
    """Test Summary 시트에서 Requirements/Design Coverage row 탐색."""
    cell_ref, _ = _scan_label_cell(ws, _REQUIREMENTS_LABELS, max_row=60)
    return cell_ref.row if cell_ref else None


def _scan_test_log_marker_col(ws) -> Optional[int]:
    """Test Log 시트의 header row에서 marker column 탐색.

    회사 v2.02 양식이 AL column에 별도 marker label을 둔다고 보고됨 (53차 사용자).
    candidates에 정확 매칭되는 cell의 column 반환.
    """
    cell_ref, _ = _scan_label_cell(ws, _TEST_LOG_MARKER_LABELS, max_row=10)
    return cell_ref.col if cell_ref else None


def _scan_test_log_columns(ws) -> dict[str, Optional[int]]:
    """58차 F3 — Test Log 시트 헤더 row에서 column 위치 라벨 매칭.

    회사 v2.02 SITR 양식 헤더 (row 1~10 범위):
        B='Test Case', H='Input', R='Expected Result', AB='Actual Result',
        AL='Pass/Fail', AN='Log Data'
    회사 v3.01 SUTR 양식 헤더:
        B='Test Case ID', F='Input', P='Expected Result', Z='Actual Result',
        AJ='Pass/Fail Unit', AK='Pass/Fail Total', AL='Log Data'

    탐색 범위 max_row=12, max_col=50. 같은 라벨 여러 cell 매칭 시 첫 cell 우선.

    Returns:
        dict[label_key, Optional[int]]:
            - input_col / expected_col / actual_col / pass_fail_col /
              pass_fail_total_col / log_data_col
            미발견은 None. 모두 None이면 v3.01 hardcode fallback.
    """
    result: dict[str, Optional[int]] = {
        "input_col": None, "expected_col": None, "actual_col": None,
        "pass_fail_col": None, "pass_fail_total_col": None, "log_data_col": None,
        # 60차 F6-A — SwUTS/SwITS spec Precondition stamp col 감지.
        "precondition_col": None,
    }
    if ws is None:
        return result
    max_row = min(ws.max_row + 1, 12) if ws.max_row else 12
    # 59차 F4-A — KJPDS02 v1.01 양식 max_col=378 호환. v2.02/v3.01 양식의 라벨은
    # col 50 안에 있지만 v1.01은 더 넓은 범위 가능 — 500 확장 (수십 ms 비용).
    max_col = min(ws.max_column + 1, 500) if ws.max_column else 500
    # 라운드 74 T904 — KJPDS02 v1.01 양식 column 헤더 패턴 추가.
    # 회사 KJPDS02 SwITR v1.01 Test Log row 3에 `Input` 라벨이 cell column 10에 있고,
    # row 4에 `Inpt[0] / Inpt[1] / ...` 변수명 stamp. 비슷하게 Exp[N]/Act[N]도 가능.
    # 첫 `Inpt[0]` cell column을 input_col로 인식 (별도 'Input' 라벨이 인접 cell에 없을 때).
    _inpt_re = _INPT_LABEL_RE
    _exp_re = re.compile(r"^Exp\[\d+\]$")
    _act_re = re.compile(r"^Act\[\d+\]$")

    for r in range(1, max_row):
        for c in range(1, max_col):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            s = v.strip().lower()
            if not s:
                continue
            # 라운드 74 T904 — Inpt[0]/Exp[0]/Act[0] 첫 인덱스 매칭 (v1.01 양식)
            raw = v.strip()
            if result["input_col"] is None and _inpt_re.match(raw) and raw.endswith("[0]"):
                result["input_col"] = c
                continue
            if result["expected_col"] is None and _exp_re.match(raw) and raw.endswith("[0]"):
                result["expected_col"] = c
                continue
            if result["actual_col"] is None and _act_re.match(raw) and raw.endswith("[0]"):
                result["actual_col"] = c
                continue
            if result["input_col"] is None and s in (
                "input", "input parameters", "input parameter",
            ):
                result["input_col"] = c
            elif result["expected_col"] is None and s in (
                "expected result", "expected", "expected results",
            ):
                result["expected_col"] = c
            elif result["actual_col"] is None and s in (
                "actual result", "actual", "actual results",
            ):
                result["actual_col"] = c
            elif result["pass_fail_total_col"] is None and s in (
                "pass/fail total", "pass fail total",
            ):
                # 'Pass/Fail Total'은 'Pass/Fail'보다 longer match — 먼저 검사.
                result["pass_fail_total_col"] = c
            elif result["pass_fail_col"] is None and s in (
                "pass/fail", "pass/fail unit", "pass fail", "pass/fail (unit)",
            ):
                result["pass_fail_col"] = c
            elif result["log_data_col"] is None and s in (
                "log data", "log", "log file",
            ):
                result["log_data_col"] = c
            elif result["precondition_col"] is None and s in (
                "precondition", "pre-condition", "전제 조건", "전제조건",
            ):
                # 60차 F6-A — SwUTS/SwITS Precondition stamp col
                result["precondition_col"] = c
    return result


# 59차 F4-A — 변수명 헤더 row 감지용 regex.
# 일반 영문 변수 식별자 (예: u16g_SysDiag_SystemStatus, s_System_I) 또는
# 양식 라벨 패턴 (예: Inpt[0], Inpt[1], ...).
_VAR_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
_INPT_LABEL_RE = re.compile(r"^Inpt\[\d+\]$")


def _scan_test_log_variable_header_row(
    ws,
    *,
    input_col: Optional[int],
    expected_col: Optional[int],
    max_row_limit: int = 16,
) -> Optional[int]:
    """59차 F4-A — Test Log 변수명 헤더 row 자동 감지.

    KJPDS02 v1.01 양식은 row 5의 col 10~ 영역에
    ``u16g_SysDiag_SystemStatus``, ``u8g_SysEepromCtrl_InLineMod_F``, ...
    같은 변수명을 stamp한다. _write_test_log가 이 row에 환경별 input/expected/
    actual 변수의 합집합 + sorted 결과를 stamp하려면 row 위치가 필요.

    헤더 row 정의: ``input_col`` ~ ``expected_col-1`` 범위에서 영문 식별자
    (``_VAR_NAME_RE``) 또는 ``Inpt[N]`` 패턴(``_INPT_LABEL_RE``)이 3개 이상
    연속 발견되는 row.

    Args:
        ws: openpyxl worksheet (Test Log 시트).
        input_col: ``_scan_test_log_columns`` 결과 input_col (None이면 scan skip).
        expected_col: 동일. 없으면 ``input_col + 30`` 까지 scan.
        max_row_limit: 헤더 스캔 최대 row (default 16 — KJPDS02 row 5 / v2.02
            row 4~5 / 여유분 포함).

    Returns:
        헤더 row (1-indexed) 또는 None (미발견). None일 때 _write_test_log는
        변수명 헤더 stamp skip (backward-compat v2.02/v3.01 동작 유지).
    """
    if ws is None or input_col is None:
        return None
    end_col = expected_col if expected_col else input_col + 30
    if end_col <= input_col:
        return None
    max_row_attr = getattr(ws, "max_row", 0) or 0
    max_row = min(max_row_attr + 1, max_row_limit) if max_row_attr else max_row_limit
    for r in range(1, max_row):
        match_count = 0
        for c in range(input_col, end_col):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue
            if _VAR_NAME_RE.match(s) or _INPT_LABEL_RE.match(s):
                match_count += 1
                if match_count >= 3:
                    return r
    return None


def _scan_test_log_max_counts(
    cols: dict[str, Optional[int]],
    *,
    ws_max_col: int = 0,
) -> dict[str, int]:
    """59차 F4-A — Test Log Input/Expected/Actual 각 column block 최대 변수 수 산출.

    ``_scan_test_log_columns`` 결과로 input_col / expected_col / actual_col /
    pass_fail_col 위치를 받아, 인접 col 차이로 block 크기를 결정한다.

    block 종단:
        - input_max  = expected_col - input_col
        - expected_max = actual_col - expected_col
        - actual_max = pass_fail_col - actual_col (없으면 ws_max_col - actual_col,
          그도 없으면 default 10)

    Args:
        cols: ``_scan_test_log_columns`` 반환 dict.
        ws_max_col: ws.max_column — pass_fail_col 부재 시 fallback. 0 또는
            음수면 default 10.

    Returns:
        ``{"input_max_count": int, "expected_max_count": int, "actual_max_count": int}``
        — 모든 값 ``max(1, diff)`` 보장. 미산출 시 default 10 유지 (backward-compat).
    """
    input_col = cols.get("input_col")
    expected_col = cols.get("expected_col")
    actual_col = cols.get("actual_col")
    pass_fail_col = cols.get("pass_fail_col")

    def _safe_diff(a: Optional[int], b: Optional[int], default: int = 10) -> int:
        if a is None or b is None:
            return default
        diff = b - a
        return max(1, diff)

    actual_end: Optional[int]
    if pass_fail_col is not None:
        actual_end = pass_fail_col
    elif ws_max_col > 0 and actual_col is not None and ws_max_col > actual_col:
        actual_end = ws_max_col + 1
    else:
        actual_end = None

    return {
        "input_max_count": _safe_diff(input_col, expected_col, 10),
        "expected_max_count": _safe_diff(expected_col, actual_col, 10),
        "actual_max_count": _safe_diff(actual_col, actual_end, 10),
    }


def _scan_traceability_header(ws) -> Optional[int]:
    """58차 F2 — 1.Traceability 시트 헤더 row 동적 감지.

    헤더 row 정의: SwUFn_ 또는 SwUTC_ prefix 컬럼이 다수 (>=5개) 수평 배열된 row.
    SwIT v2.02 양식은 헤더가 row 20~25 부근에 있을 수 있고 SwUT v3.01은 더 위
    (row 3~10) 가능. max_row=30 스캔, max_col=500 (1941 TC × 30 환경).

    Returns:
        헤더 row (1-indexed) 또는 None (미발견).
    """
    if ws is None:
        return None
    max_row = min(ws.max_row + 1, 35) if ws.max_row else 35
    max_col = min(ws.max_column + 1, 500) if ws.max_column else 500
    for r in range(1, max_row):
        swufn_count = 0
        for c in range(1, max_col):
            v = ws.cell(r, c).value
            if isinstance(v, str) and (
                v.startswith("SwUFn_") or v.startswith("SwUTC_") or v.startswith("SwITC_")
            ):
                swufn_count += 1
                if swufn_count >= 5:
                    return r
    return None


def _scan_test_log_tc_row_step(ws) -> int:
    """57차 T314 — Test Log/Test Result 시트의 1 TC당 row step 동적 감지.

    회사 v2.02 SUTR/SITR 양식은 TC ID + Params 1~5 sub-row pattern (6 row step).
    B 열에서 SwUTC_ / SwITC_ prefix를 가진 첫 2개 TC row 위치 차이를 반환.

    Returns:
        int: row step (회사 v2.02 = 6, v3.01 또는 미발견 = 1).
    """
    if ws is None:
        return 1
    tc_rows: list[int] = []
    max_row = min(ws.max_row + 1, 60) if ws.max_row else 60
    for r in range(1, max_row):
        v = ws.cell(r, 2).value
        if isinstance(v, str):
            s = v.strip()
            if s.startswith("SwUTC_") or s.startswith("SwITC_"):
                tc_rows.append(r)
                if len(tc_rows) >= 2:
                    break
    if len(tc_rows) >= 2:
        step = tc_rows[1] - tc_rows[0]
        if step > 0:
            return step
    return 1


# ---------------------------------------------------------------------------
# Internal inspect (캐시 entry)
# ---------------------------------------------------------------------------

def _inspect_internal(
    template_bytes: bytes, kind: Literal["coverage", "sitr"],
) -> SwitLayout:
    """openpyxl load_workbook → 시트별 inspect → SwitLayout 반환.

    54-fix C2: ZIP bomb 방어 — 외부에서 validate_xlsx_template_bytes 호출하지 않고
    바로 inspect 호출 시도 거부. 압축비 / 단일 파일 크기 / 총 크기 한도는
    excel_template_utils 정책 동일.

    openpyxl 미설치 또는 load 실패 시 fallback_to_v301=True + warnings 누적하여 반환.
    """
    if openpyxl is None:
        return SwitLayout(
            detected_version="unknown",
            fallback_to_v301=True,
            warnings=["openpyxl 미설치 — layout inspect 불가, v3.01 fallback"],
        )
    # 54-fix C2: ZIP bomb / magic byte 사전 검증.
    try:
        validate_xlsx_template_bytes(template_bytes, label=f"layout inspect ({kind})")
    except TemplateValidationError as e:
        return SwitLayout(
            detected_version="unknown",
            fallback_to_v301=True,
            warnings=[f"template 입력 검증 실패 — {e!s:.150}"],
        )
    try:
        # SITR도 xlsm이라 keep_vba=True (data_only=False 기본).
        wb = openpyxl.load_workbook(
            io.BytesIO(template_bytes),
            keep_vba=(kind == "sitr"),
            data_only=False,
        )
    except Exception as e:
        return SwitLayout(
            detected_version="unknown",
            fallback_to_v301=True,
            warnings=[f"template load 실패 ({type(e).__name__}: {e!s:.80}) — v3.01 fallback"],
        )

    warnings: list[str] = []
    try:
        # Cover 시트
        cover_ws = _find_sheet(wb, lambda n: n.lower() == "cover")
        cover_labels: dict[str, str] = {}
        if cover_ws is not None:
            cover_labels, _ = _scan_labels_in_sheet(cover_ws, _COVER_CANDIDATES)
        else:
            warnings.append("Cover 시트 미발견 — cover_labels 비어 있음")

        # Test Summary 시트 — 53차 fix substring 매칭과 일치 (v2.02 "1.Test Summary")
        ts_ws = _find_sheet(wb, lambda n: "test summary" in n.lower())
        test_summary_labels: dict[str, str] = {}
        tc_stats_row: Optional[int] = None
        tc_stats_col_start: Optional[int] = None
        requirements_row: Optional[int] = None
        if ts_ws is not None:
            test_summary_labels, _ = _scan_labels_in_sheet(
                ts_ws, _TEST_SUMMARY_CANDIDATES,
            )
            # v2.02 신규 row 탐색
            tc_stats_row, tc_stats_col_start = _scan_tc_stats_row(ts_ws)
            requirements_row = _scan_requirements_row(ts_ws)
        else:
            warnings.append("Test Summary 시트 미발견 — test_summary_labels 비어 있음")

        # Test Log 시트 (SITR only)
        test_log_extra_marker_col: Optional[int] = None
        test_log_header_cell: Optional[CellRef] = None
        deviation_header_cell: Optional[CellRef] = None
        test_log_tc_row_step = 1
        # 58차 F3 — Test Log column 위치 (None이면 v3.01 hardcode fallback)
        test_log_input_col: Optional[int] = None
        test_log_expected_col: Optional[int] = None
        test_log_actual_col: Optional[int] = None
        test_log_pass_fail_col: Optional[int] = None
        test_log_pass_fail_total_col: Optional[int] = None
        test_log_log_data_col: Optional[int] = None
        # 60차 F6-A — SwUTS/SwITS spec Precondition stamp col
        test_log_precondition_col: Optional[int] = None
        # 59차 F4-A — 변수명 헤더 row + Input/Expected/Actual 각 block max count.
        test_log_variable_header_row: Optional[int] = None
        test_log_input_max_count = 10
        test_log_expected_max_count = 10
        test_log_actual_max_count = 10
        test_log_step_layout = "single_row"
        if kind == "sitr":
            log_ws = _find_sheet(wb, lambda n: "test log" in n.lower() or "test result" in n.lower())
            if log_ws is not None:
                test_log_extra_marker_col = _scan_test_log_marker_col(log_ws)
                # header cell 위치 참고 (writer가 직접 find_kv_row 사용하지만 inspect 결과 참고용)
                tc_header, _ = _scan_label_cell(
                    log_ws, ("Test Case ID", "TC ID", "TC name"), max_row=10,
                )
                test_log_header_cell = tc_header
                # 57차 T314 — 1 TC당 row step 동적 감지 (회사 v2.02 = 6)
                test_log_tc_row_step = _scan_test_log_tc_row_step(log_ws)
                # 58차 F3 — column 위치 자동 감지
                _cols = _scan_test_log_columns(log_ws)
                test_log_input_col = _cols.get("input_col")
                test_log_expected_col = _cols.get("expected_col")
                test_log_actual_col = _cols.get("actual_col")
                test_log_pass_fail_col = _cols.get("pass_fail_col")
                test_log_pass_fail_total_col = _cols.get("pass_fail_total_col")
                test_log_log_data_col = _cols.get("log_data_col")
                test_log_precondition_col = _cols.get("precondition_col")
                # 59차 F4-A — 변수명 헤더 row + block max counts + step layout.
                test_log_variable_header_row = _scan_test_log_variable_header_row(
                    log_ws,
                    input_col=test_log_input_col,
                    expected_col=test_log_expected_col,
                )
                ws_max_col = getattr(log_ws, "max_column", 0) or 0
                _max_counts = _scan_test_log_max_counts(_cols, ws_max_col=ws_max_col)
                test_log_input_max_count = _max_counts["input_max_count"]
                test_log_expected_max_count = _max_counts["expected_max_count"]
                test_log_actual_max_count = _max_counts["actual_max_count"]
                # tc_row_step > 1 → step_in_rows (v2.02 6 row 분배), else single_row (v3.01).
                test_log_step_layout = (
                    "step_in_rows" if test_log_tc_row_step > 1 else "single_row"
                )
            dev_ws = _find_sheet(wb, lambda n: "deviation" in n.lower())
            if dev_ws is not None:
                dev_header, _ = _scan_label_cell(
                    dev_ws, ("Test Case ID", "TC ID"), max_row=10,
                )
                deviation_header_cell = dev_header

        # 58차 F2 — Coverage Traceability 시트 헤더 row dynamic 감지
        traceability_header_row: Optional[int] = None
        if kind == "coverage":
            trace_ws = _find_sheet(
                wb, lambda n: "traceability" in n.lower(),
            )
            if trace_ws is not None:
                traceability_header_row = _scan_traceability_header(trace_ws)

        # 양식 버전 추정
        detected_version = _detect_version_from_labels(test_summary_labels)

        # 59차 F4-C — KJPDS02 v1.01 양식 자동 감지 (시트 구성 패턴).
        # v1.01 SITR 시트: Cover/History/1.Test Summary/2.Test Log (4 시트, Deviation 없음)
        # v1.01 Coverage 시트: Cover/History/1.Test Summary/2.Traceability/3.Consistency/
        #                      4.Coverage (6 시트)
        # v2.02/v3.01 SITR: Cover/History/1.Test Summary/Deviation/Test Log/...
        # v2.02/v3.01 Coverage: Cover/History/Test Summary/1.Traceability/2.Consistency/
        #                      3.Coverage
        sheet_names_lower = [s.lower() for s in wb.sheetnames]
        sitr_sheet_count_actual = len(wb.sheetnames)
        coverage_sheet_count_actual = len(wb.sheetnames)
        # v1.01 signature: "4.coverage" 시트명 (3.Coverage가 아닌 4.Coverage)
        has_4_coverage = any("4.coverage" in n for n in sheet_names_lower)
        has_3_consistency = any("3.consistency" in n for n in sheet_names_lower)
        has_2_traceability = any("2.traceability" in n for n in sheet_names_lower)
        v101_signals = sum([has_4_coverage, has_3_consistency, has_2_traceability])
        if v101_signals >= 2:
            # v1.01 양식 감지 — detected_version override (test_summary_labels 보존)
            detected_version = "v1.01"
        deviation_sheet_present_v = any(
            "deviation" in n for n in sheet_names_lower
        )
        # 라운드 F7 D2 fix: 시트명 prefix만으로 판단하지 않고 traceability 시트의
        # 실제 header row 내용 inspect. 회사 표준 SwUTCV는 v1.01 signature
        # (4.coverage / 3.consistency / 2.traceability) 갖지만 R12 header는
        # SwUFn_01xx (SwUFn matrix) — switc_x_swst 아님. SwST_01 / SwSTR_NN 등이
        # 다수 발견되면 switc_x_swst matrix로 분류 (SwITCV 양식).
        traceability_matrix_kind_v: Literal["swufn_x_env", "switc_x_swst"] = "swufn_x_env"
        if kind == "coverage":
            trace_ws_for_kind = _find_sheet(wb, lambda n: "traceability" in n.lower())
            if trace_ws_for_kind is not None:
                swst_count = 0
                swufn_count = 0
                # 첫 15 row × 30 col scan — header row 위치 fixed 아님
                try:
                    for row in trace_ws_for_kind.iter_rows(
                        min_row=1, max_row=15, max_col=30, values_only=True,
                    ):
                        for cell in row:
                            if not isinstance(cell, str):
                                continue
                            s = cell.strip()
                            if s.startswith(("SwST_", "SwSTR_")):
                                swst_count += 1
                            elif s.startswith("SwUFn_"):
                                swufn_count += 1
                except (AttributeError, ValueError):
                    pass
                # F7 자체평가 R1 W4 fix: 임계값 보강 — SwST 3건 이상 또는 SwST 1건+ +
                # SwUFn 0건이면 switc_x_swst. 회사 표준 SwITCV (14 SwST/SwSTR) +
                # MVP/소규모 SwITCV (3 SwST) + Edge case (SwST 1건 + SwUFn 0건) 대응.
                if (swst_count >= 3 and swst_count > swufn_count) or (
                    swst_count > 0 and swufn_count == 0
                ):
                    traceability_matrix_kind_v = "switc_x_swst"
        coverage_metric_kind_v: Literal["single", "function_and_calls"] = (
            "function_and_calls" if detected_version == "v1.01" else "single"
        )
        test_summary_coverage_breakdown_v = 4 if detected_version == "v1.01" else 1

        # fallback 판정: v2.02 신규 row가 모두 없고 v3.01 라벨 매칭이 다수면 v3.01
        # → fallback_to_v301 False (정상 v3.01 양식). v2.02 라벨 매칭 0이면 fallback.
        v202_label_count = sum(
            1 for v in test_summary_labels.values()
            if v in ("SW Version", "SW 버전", "HW Version", "HW 버전")
        )
        fallback_to_v301 = detected_version == "v3.01" or (
            detected_version == "unknown" and v202_label_count == 0
        )

        # 56차 T306 — v2.02 Coverage 양식 label-missing fallback.
        # 회사 Coverage Report v2.02는 row 17 (TC stats) + row 20 (Requirements)을
        # 사용자 수동 입력 영역으로 두어 template에 라벨 부재 → label 매칭 실패 →
        # 기존엔 silent skip. fallback path: B17/B20 cell이 None이면 default position
        # 사용 + label_missing=True 표시 → writer가 라벨도 함께 stamp.
        tc_stats_label_missing = False
        requirements_label_missing = False
        if detected_version == "v2.02" and ts_ws is not None:
            if tc_stats_row is None:
                # row 17 B열이 빈 cell이면 fallback 활성
                try:
                    b17 = ts_ws.cell(17, 2).value  # type: ignore[attr-defined]
                except Exception:  # pragma: no cover — openpyxl edge case
                    b17 = "_"
                if b17 is None or (isinstance(b17, str) and b17.strip() == ""):
                    tc_stats_row = 18  # data row = label row + 1
                    tc_stats_col_start = 2  # B 열
                    tc_stats_label_missing = True
                    warnings.append(
                        "v2.02 양식 Coverage: TC stats row label 미발견 → "
                        "default position (row=17 label / row=18 data, col=B) fallback. "
                        "writer가 라벨도 stamp (56차 T306)."
                    )
                else:
                    warnings.append(
                        "v2.02 양식 추정되나 TC stats row label 미발견 + B17 비어있지 않음 — "
                        f"fallback 미적용. 후보 {_TC_STATS_LABELS}"
                    )
            if requirements_row is None:
                try:
                    b20 = ts_ws.cell(20, 2).value  # type: ignore[attr-defined]
                except Exception:  # pragma: no cover
                    b20 = "_"
                if b20 is None or (isinstance(b20, str) and b20.strip() == ""):
                    requirements_row = 20
                    requirements_label_missing = True
                    warnings.append(
                        "v2.02 양식 Coverage: Requirements row label 미발견 → "
                        "default position (row=20, col=B) fallback. writer가 헤더+"
                        "라벨도 stamp (56차 T306)."
                    )
                else:
                    warnings.append(
                        "v2.02 양식 추정되나 Requirements row label 미발견 + B20 비어있지 않음 — "
                        f"fallback 미적용. 후보 {_REQUIREMENTS_LABELS}"
                    )

        return SwitLayout(
            detected_version=detected_version,
            fallback_to_v301=fallback_to_v301,
            cover_labels=cover_labels,
            test_summary_labels=test_summary_labels,
            tc_stats_row=tc_stats_row,
            tc_stats_col_start=tc_stats_col_start,
            requirements_row=requirements_row,
            deviation_header_cell=deviation_header_cell,
            test_log_header_cell=test_log_header_cell,
            test_log_extra_marker_col=test_log_extra_marker_col,
            tc_stats_label_missing=tc_stats_label_missing,
            requirements_label_missing=requirements_label_missing,
            test_log_tc_row_step=test_log_tc_row_step,
            # 58차 F3 — Test Log column 위치
            test_log_input_col=test_log_input_col,
            test_log_expected_col=test_log_expected_col,
            test_log_actual_col=test_log_actual_col,
            test_log_pass_fail_col=test_log_pass_fail_col,
            test_log_pass_fail_total_col=test_log_pass_fail_total_col,
            test_log_log_data_col=test_log_log_data_col,
            # 60차 F6-A — Test Log Precondition col (SwUTS/SwITS spec stamp)
            test_log_precondition_col=test_log_precondition_col,
            # 58차 F2 — Coverage Traceability 헤더 row
            traceability_header_row=traceability_header_row,
            # 59차 F4-A — Test Log 변수명 헤더 row + block max counts + step layout
            test_log_variable_header_row=test_log_variable_header_row,
            test_log_input_max_count=test_log_input_max_count,
            test_log_expected_max_count=test_log_expected_max_count,
            test_log_actual_max_count=test_log_actual_max_count,
            test_log_step_layout=test_log_step_layout,
            # 59차 F4-C — KJPDS02 v1.01 양식 시트 구성 분기
            sitr_sheet_count=sitr_sheet_count_actual,
            coverage_sheet_count=coverage_sheet_count_actual,
            traceability_matrix_kind=traceability_matrix_kind_v,
            coverage_metric_kind=coverage_metric_kind_v,
            deviation_sheet_present=deviation_sheet_present_v,
            test_summary_coverage_breakdown=test_summary_coverage_breakdown_v,
            warnings=warnings,
        )
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Public API + LRU cache
# ---------------------------------------------------------------------------

# 54-fix W1/W3: cache stampede + StopIteration 방어 + maxsize 4→8 (thrashing 여유).
# 자체 LRU dict → threading.Lock 보호. Semaphore(3) SwUT + Semaphore(2) SwIT 동시
# 빌드 + 향후 C1 fix로 SwUT도 layout 호출 시 5건 동시 thread 가능.
_LAYOUT_CACHE: dict[tuple[str, str], SwitLayout] = {}
_MAX_CACHE_SIZE = 8  # W3 — 회사 양식 4개 + 시험용/향후 추가 여유
_CACHE_LOCK = threading.Lock()


def inspect_swit_layout(
    template_bytes: bytes,
    kind: Literal["coverage", "sitr"],
) -> SwitLayout:
    """v2.02 또는 v3.01 양식 template을 inspect하여 SwitLayout 추출.

    sha256(template_bytes) + kind 키로 LRU 캐시. 동일 template 반복 빌드 시
    openpyxl load_workbook 1회만 호출.

    54-fix W1: cache stampede 방어 — _CACHE_LOCK으로 hit/miss + insertion 직렬화.
    동일 key 동시 miss 시 두 번째 thread는 첫 번째의 결과를 cache hit으로 받음.

    Args:
        template_bytes: 회사 양식 xlsx (coverage) 또는 xlsm (sitr) bytes.
        kind: "coverage" 또는 "sitr".

    Returns:
        SwitLayout — detected_version + 라벨 매핑 + 신규 row 위치 + warnings.
        openpyxl 미설치 또는 load 실패 시 fallback_to_v301=True + warnings 누적.
    """
    sha = hashlib.sha256(template_bytes).hexdigest()
    key = (sha, kind)
    # Fast path — lock 밖 확인
    with _CACHE_LOCK:
        cached = _LAYOUT_CACHE.get(key)
        if cached is not None:
            return cached
    # Miss path — inspect (load_workbook 비용 ~수십 ms, lock 밖)
    layout = _inspect_internal(template_bytes, kind)
    # Insertion + LRU evict — lock 안 (stampede 시 dup work 가능하나 결과 동일)
    with _CACHE_LOCK:
        # Stampede 검사 — 다른 thread가 먼저 set 했으면 그 값을 반환 (메모리 일관성)
        existing = _LAYOUT_CACHE.get(key)
        if existing is not None:
            return existing
        if len(_LAYOUT_CACHE) >= _MAX_CACHE_SIZE:
            # 가장 오래된 entry 제거 (dict insertion order, Python 3.7+).
            # 빈 dict 진입 방어 (W1 StopIteration) — len check가 보장
            try:
                oldest = next(iter(_LAYOUT_CACHE))
                del _LAYOUT_CACHE[oldest]
            except (StopIteration, KeyError):  # pragma: no cover — race-safe guard
                pass
        _LAYOUT_CACHE[key] = layout
        return layout


def clear_layout_cache() -> None:
    """테스트/관리 용 — 캐시 초기화."""
    with _CACHE_LOCK:
        _LAYOUT_CACHE.clear()


def cache_size() -> int:
    """테스트 용 — 현재 캐시 entry 수."""
    with _CACHE_LOCK:
        return len(_LAYOUT_CACHE)


__all__ = [
    "CellRef",
    "SwitLayout",
    "inspect_swit_layout",
    "clear_layout_cache",
    "cache_size",
]
