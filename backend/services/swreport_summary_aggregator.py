"""SW Test Result Report — 전 레벨 통합 Summary 빌더 (ES95411 master).

목적
----
플랫폼이 레벨별로 산출한 **완성 결과 산출물**(SwUTCR/SwITCR/SwSA 등 — 회사
ES95411 양식의 ``NN.<TestID>`` detail 시트를 가진 xlsm)을 파싱하여, 마스터
리포트(ES95411)의 ``Summary`` 시트(ST 정적·UT 단위·IT 통합·ET 시스템 전 레벨)를
한 표로 채운 ``.xlsm`` 를 생성한다. 현재 수기로 유지되는 cross-level 통합표를
자동화한다.

핵심 모델 (v1.02 ground-truth 검증 완료, MISMATCH 0)
----------------------------------------------------
``Summary`` 표는 각 detail 시트의 **Result 블록**을 roll-up한 것이다:

* 매칭 키 = test ID (UT101/IT101/ST101…). Summary 행의 ``SheetName``(P열) 또는
  ``ID``(E열)에서 선행 ``NN.`` 접두어를 제거해 source 시트와 매칭한다.
* **수행(O) primary 행**: detail Result 블록에서 분석차수·SW Ver.·Tester·
  Debugger·총분석시간·P/F 를 읽어 Summary 행(H/I/J/L/M/N)에 stamp.
* **미수행(X) 행**: 결과 컬럼을 공란으로 둔다 (hidden detail의 stale 값 무시).
* **sub 행**(ST202~211, IT802 등 — parent 시트 공유): meta 컬럼 공란, P/F만 parent
  에서 승계 + 경고.
* **Total 행**: 점검대상(G)=수행 개수, 총분석시간(M)=합, P/F(N)=Fail 개수.

openpyxl 제약
-------------
* source(Result 블록 값 읽기)는 ``data_only=True`` — Excel이 캐시한 계산값 필요.
* template(출력 쓰기)은 ``data_only=False, keep_vba=True`` — 수식/VBA/스타일 보존.
* 따라서 source 워크북과 template 워크북을 **분리 로드**한다 (동일 파일이어도 2회).

제약 (backend/services/CLAUDE.md)
--------------------------------
* Cloudium worker는 read-only — 본 빌더는 **bytes in → bytes out**, 파일을 직접
  쓰지 않는다 (라우터가 resolver로 read, Response로 반환).
* 시각 강조 RGB/placeholder는 ``design_tokens`` 단일 출처 (excel_template_utils의
  ``mark_*``/``write_value_or_mark`` 래퍼 경유).
* ISO 26262: 본 산출물은 auto-generated draft — ASIL B/C/D 단독 evidence 금지,
  manual review 의무 (tool_qualification 메타 승계).
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from backend.services.design_tokens import FAIL_FILL_RGB
from backend.services.excel_template_utils import (
    has_vba_macros,
    mark_fail_cell,
    resolve_merge_anchor,
    safe_write,
    sanitize_xlsm_external_links,
    short_date,
    validate_xlsx_template_bytes,
)
from backend.services.swut_meta import BuildMetaBase

# --- Summary 표 열 (1-indexed) — ES95411 v1.02 레이아웃 ------------------------
COL_NO = 2          # B  No
COL_CATEGORY = 3    # C  Test Category (그룹별 세로 머지)
COL_SUBNOTE = 4     # D  부가 표기 (Code Metric / Heap Memory / -)
COL_ID = 5          # E  Test ID (ST101, UT101, IT101…)
COL_NAME = 6        # F  Test Name
COL_PERFORMED = 7   # G  점검 대상 (O/X — 수행 여부, planning)
COL_ITER = 8        # H  분석 차수
COL_SWVER = 9       # I  SW Ver.
COL_TESTER = 10     # J  Tester
COL_TOOL = 11       # K  Tool (정적: catalog, 동적 미사용)
COL_DEBUGGER = 12   # L  Debugger
COL_HOURS = 13      # M  총 분석시간 (HR)
COL_PF = 14         # N  P/F
COL_NOTE = 15       # O  Note
COL_SHEETNAME = 16  # P  Sheet Name (NN.<TestID>)

_PREFIX_RE = re.compile(r"^\s*\d+\.\s*")


@dataclass
class SwReportBuildMeta(BuildMetaBase):
    """ES95411 Summary 헤더 블록 메타. 값이 있으면 template 위에 덮어쓴다.

    doc_id_base 등 BuildMetaBase 필드는 사용하지 않으나 공통 base를 따른다
    (author/reviewer/approver property 재사용).
    """
    phase: str = ""            # DV
    product: str = ""          # PDS
    test_target: str = ""      # 검증 대상 (MCU)
    compiler: str = ""
    mcu: str = ""
    software_platform_ver: str = ""  # E5 (release_sw_version과 별개일 수 있음)


@dataclass
class ResultBlock:
    """detail 시트 1개의 Result 블록에서 추출한, Summary가 필요로 하는 필드."""
    test_id: str
    source_sheet: str
    source_label: str
    test_iteration: str = ""
    sw_version: str = ""
    tester: str = ""
    debugger: str = ""
    total_hours: Any = ""
    pf: str = ""

    @property
    def is_fail(self) -> bool:
        return "fail" in (self.pf or "").strip().lower()

    @property
    def has_result(self) -> bool:
        return bool((self.pf or "").strip())


@dataclass
class SwReportBuildResult:
    """``build_summary_report`` 산출물 — bytes out + 진단."""
    ok: bool
    xlsm_io: io.BytesIO
    filename: str
    warnings: list[str] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    incomplete_rows: list[str] = field(default_factory=list)
    vba_macros_preserved: bool = False
    tool_qualification: dict[str, Any] = field(
        default_factory=lambda: {
            "evidence_class": "auto-generated draft",
            "manual_review_required": True,
            "note": (
                "Cross-level Summary는 레벨별 산출물 Result 블록의 기계적 roll-up이다. "
                "ASIL B/C/D 단독 evidence로 사용 금지 — reviewer 검토 필수."
            ),
        }
    )

    @property
    def xlsm_bytes(self) -> bytes:
        pos = self.xlsm_io.tell()
        self.xlsm_io.seek(0)
        data = self.xlsm_io.read()
        self.xlsm_io.seek(pos)
        return data

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "filename": self.filename,
            "warnings": self.warnings,
            "summary": self.summary,
            "incomplete_rows": self.incomplete_rows,
            "vba_macros_preserved": self.vba_macros_preserved,
            "tool_qualification": self.tool_qualification,
        }


# --- 값 정규화 ---------------------------------------------------------------
def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().replace("\r", " ").replace("\n", " ")


def strip_id(name: Any) -> str:
    """``21.UT101`` / ``2. ST201`` → ``UT101`` / ``ST201`` (대문자 정규화)."""
    return _PREFIX_RE.sub("", _norm(name)).strip().upper()


def normalize_pf(raw: str) -> str:
    """detail P/F 값을 Summary 표기(Pass/Fail/원본)로 정규화."""
    s = _norm(raw).lower()
    if not s:
        return ""
    if "fail" in s or s in ("ng", "x"):
        return "Fail"
    if "pass" in s or s in ("ok", "o"):
        return "Pass"
    return _norm(raw)


# --- detail 시트 Result 블록 추출 (라벨 기반, 레이아웃 변종 무관) ----------------
def _scan_labels(ws: Worksheet, max_row: int = 14, max_col: int = 12) -> dict:
    """(label_text → (row,col)) 첫 출현 위치. 라벨은 strip 후 정확 일치."""
    pos: dict[str, tuple[int, int]] = {}
    for r in range(1, min(max_row, ws.max_row or max_row) + 1):
        for c in range(1, min(max_col, ws.max_column or max_col) + 1):
            t = _norm(ws.cell(row=r, column=c).value)
            if t and t not in pos:
                pos[t] = (r, c)
    return pos


# Result 블록 라벨 — _value_right가 인접 블록으로 넘어가 다른 라벨을 값으로 오인하지
# 않도록 stop-word로 사용 (라운드 리뷰 F1 fix: hours wrong-pick 차단).
_KNOWN_LABELS = frozenset({
    "분석차수", "SW Ver.", "Tester", "Debugger",
    "준비", "수행", "검토", "Total",
    "P/F", "TC Pass율", "커버리지",
})


def _num_or_none(v: Any):
    """숫자(또는 숫자문자열)면 float, 아니면 None. bool은 숫자로 보지 않음."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def _value_right(ws: Worksheet, row: int, col: int, *, max_scan: int = 3) -> str:
    """(row,col) 우측 값 1개 추출 — 인접 라벨/블록 헤더 침범 차단 (F1 fix).

    max_scan칸 내 첫 비어있지 않은 셀을 반환하되, 그 셀이 다른 Result 라벨
    (``_KNOWN_LABELS``)이거나 블록 헤더(``■``)면 '값 없음'으로 본다. 빈 셀을 지나
    엉뚱한 라벨(예: 빈 hours 셀 우측의 'P/F')까지 긁던 무한 우측 스캔 버그 제거.
    """
    for cc in range(col + 1, col + 1 + max_scan):
        v = ws.cell(row=row, column=cc).value
        if v is None:
            continue
        s = _norm(v)
        if s == "":
            continue
        if s in _KNOWN_LABELS or s.startswith("■"):
            return ""
        return s
    return ""


def extract_result_block(ws: Worksheet, test_id: str, source_label: str = "") -> ResultBlock:
    """detail 시트의 Result 블록에서 Summary 필드 추출.

    라벨(분석차수/SW Ver./Tester/Debugger/P/F)의 오른쪽 값을 읽는다. 총분석시간은
    '분석시간' 블록(준비/수행/검토/Total)의 Total 값을 — 준비 라벨 열에 정렬된
    Total을 anchor로 — 읽어 다른 'Total'과 혼동을 방지한다.
    """
    pos = _scan_labels(ws)
    blk = ResultBlock(test_id=test_id, source_sheet=ws.title, source_label=source_label)

    def grab(label: str) -> str:
        rc = pos.get(label)
        return _value_right(ws, rc[0], rc[1]) if rc else ""

    blk.test_iteration = grab("분석차수")
    blk.sw_version = grab("SW Ver.")
    blk.tester = grab("Tester")
    blk.debugger = grab("Debugger")
    blk.pf = normalize_pf(grab("P/F"))
    blk.total_hours = _extract_hours(ws, pos)
    return blk


def _extract_hours(ws: Worksheet, pos: dict):
    """분석시간 블록 Total(HR). Total 라벨 '바로 우측 1칸'만 읽어 숫자면 반환.

    수식 캐시 미스 등으로 Total 셀이 None/비숫자이면 준비+수행+검토 합산으로 fallback
    (F1/F2 fix — 빈 hours 셀에서 우측 'P/F' 라벨을 긁던 wrong-pick 차단).
    반환: float | "" (확정 불가 시 빈 문자열 → 호출자가 incomplete 처리).
    """
    time_col = pos.get("준비", (0, 0))[1]
    if time_col:
        for r in range(1, (ws.max_row or 14) + 1):
            if _norm(ws.cell(row=r, column=time_col).value) == "Total":
                n = _num_or_none(ws.cell(row=r, column=time_col + 1).value)
                if n is not None:
                    return n
                break  # Total 셀이 비었거나 비숫자 → 컴포넌트 합산으로
        comp = 0.0
        got = False
        for lbl in ("준비", "수행", "검토"):
            rc = pos.get(lbl)
            if rc:
                n = _num_or_none(ws.cell(row=rc[0], column=rc[1] + 1).value)
                if n is not None:
                    comp += n
                    got = True
        if got:
            return comp
    # 분석시간 블록 자체가 없으면 일반 'Total' 라벨 우측 1칸 시도
    rc = pos.get("Total")
    if rc:
        n = _num_or_none(ws.cell(row=rc[0], column=rc[1] + 1).value)
        if n is not None:
            return n
    return ""


# --- source 워크북에서 detail 시트 index 구축 --------------------------------
def index_detail_blocks(
    source_workbooks: list[tuple[str, bytes]],
    *,
    warnings: list[str],
) -> dict[str, ResultBlock]:
    """source xlsm bytes들을 파싱해 {stripped_test_id → ResultBlock} index 구축.

    여러 source가 같은 ID를 가지면 첫 source가 이기고 경고. detail 시트 판별은
    '분석차수' 또는 'P/F' 라벨 존재 + 시트명이 ``NN.<ID>`` 패턴인 것으로 한다.
    """
    index: dict[str, ResultBlock] = {}
    for label, data in source_workbooks:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=False, data_only=True)
        except Exception as exc:  # noqa: BLE001 — 손상 source는 경고 후 skip
            warnings.append(f"[{label}] source 워크북 로드 실패: {type(exc).__name__}: {exc}")
            continue
        try:
            for ws in wb.worksheets:
                key = strip_id(ws.title)
                if not key or not _PREFIX_RE.match(_norm(ws.title)):
                    continue  # NN. 접두 없는 시트(Cover/Summary 등)는 detail 아님
                pos = _scan_labels(ws)
                if "분석차수" not in pos and "P/F" not in pos:
                    continue  # Result 블록 없는 시트
                blk = extract_result_block(ws, key, source_label=label)
                if key in index:
                    warnings.append(
                        f"detail 시트 ID 중복 '{key}': [{index[key].source_label}] 우선, "
                        f"[{label}] 무시"
                    )
                    continue
                index[key] = blk
        finally:
            wb.close()
    return index


# --- Summary 표 영역 탐지 -----------------------------------------------------
def _find_summary_sheet(wb: openpyxl.Workbook) -> Worksheet | None:
    for ws in wb.worksheets:
        if _norm(ws.title).lower() == "summary":
            return ws
    return None


@dataclass
class SummaryCols:
    """Summary 표 컬럼 위치. default는 ES95411 v1.02 레이아웃이고 ``_locate_table``이
    헤더 라벨로 override한다 — 공식 템플릿/타 프로젝트가 컬럼을 옮겨도 적응(컬럼
    하드코딩 의존 제거). 라벨을 못 찾은 컬럼만 default 유지(graceful)."""
    no: int = COL_NO
    category: int = COL_CATEGORY
    subnote: int = COL_SUBNOTE
    id: int = COL_ID
    name: int = COL_NAME
    performed: int = COL_PERFORMED
    iter: int = COL_ITER
    swver: int = COL_SWVER
    tester: int = COL_TESTER
    tool: int = COL_TOOL
    debugger: int = COL_DEBUGGER
    hours: int = COL_HOURS
    pf: int = COL_PF
    note: int = COL_NOTE
    sheetname: int = COL_SHEETNAME
    detected: tuple[str, ...] = ()  # 헤더에서 실제로 매핑된 필드명(진단/보고용)

    @property
    def result_primary(self) -> tuple[int, ...]:
        return (self.iter, self.swver, self.tester, self.debugger, self.hours, self.pf)


# 헤더 라벨(정규화 lower) → SummaryCols 필드명. 회사 ES95411 표준 표 헤더.
_HEADER_LABEL_MAP = {
    "no": "no",
    "category": "category",
    "id": "id",
    "test name": "name",
    "점검 대상": "performed", "점검대상": "performed",
    "분석 차수": "iter", "분석차수": "iter",
    "sw ver.": "swver", "sw ver": "swver",
    "tester": "tester",
    "tool": "tool",
    "debugger": "debugger",
    "총 분석시간": "hours", "총분석시간": "hours", "총 분석시간 (hr)": "hours",
    "p/f": "pf",
    "note": "note",
    "sheet name": "sheetname", "sheetname": "sheetname",
}


def _locate_table(ws: Worksheet) -> tuple[SummaryCols, int, int]:
    """헤더 행(ID 라벨)·컬럼맵·데이터 시작/Total 행 자동 탐지.

    컬럼을 하드코딩하지 않고 헤더 라벨에서 위치를 찾는다. 'ID' 라벨이 있는 행을
    헤더로 잡고, 그 행 + 위 1행(2단 헤더의 No/Sheet Name)에서 알려진 라벨을 매핑.
    못 찾은 라벨은 v1.02 default. 반환: (SummaryCols, 첫 데이터 행, Total 행).
    """
    header_row, id_col = 0, 0
    for r in range(1, (ws.max_row or 60) + 1):
        for c in range(1, (ws.max_column or 20) + 1):
            if _norm(ws.cell(row=r, column=c).value).lower() == "id":
                header_row, id_col = r, c
                break
        if header_row:
            break
    if not header_row:
        header_row, id_col = 15, COL_ID  # fallback: v1.02 레이아웃

    cols = SummaryCols()
    found: dict[str, int] = {}
    for r in (header_row - 1, header_row):
        if r < 1:
            continue
        for c in range(1, (ws.max_column or 20) + 1):
            field = _HEADER_LABEL_MAP.get(_norm(ws.cell(row=r, column=c).value).lower())
            if field and field not in found:
                found[field] = c
    found.setdefault("id", id_col)
    for field, c in found.items():
        setattr(cols, field, c)
    cols.detected = tuple(sorted(found))

    total_row = 0
    for r in range(header_row + 1, (ws.max_row or 60) + 1):
        for c in range(1, min(8, (ws.max_column or 8)) + 1):
            if _norm(ws.cell(row=r, column=c).value).lower() == "total":
                total_row = r
                break
        if total_row:
            break
    if not total_row:
        total_row = ws.max_row or 60
    return cols, header_row + 1, total_row


# --- 메인 빌더 ---------------------------------------------------------------
def build_summary_report(
    template_bytes: bytes,
    source_workbooks: list[tuple[str, bytes]],
    meta: SwReportBuildMeta | None = None,
    *,
    filename: str | None = None,
) -> SwReportBuildResult:
    """ES95411 template + source 산출물들 → 통합 Summary가 채워진 .xlsm.

    Args:
        template_bytes: ES95411 양식 xlsm (Summary 표 골격 + detail 시트 보유).
        source_workbooks: [(label, bytes)] — ES95411-style detail 시트를 가진 산출물.
            단일파일 refresh 모드면 template 자체를 source로 한 번 더 넘긴다.
        meta: 헤더 블록 override (없으면 template 값 유지).
        filename: 출력 파일명 override.
    """
    meta = meta or SwReportBuildMeta()
    warnings: list[str] = []
    incomplete: list[str] = []

    validate_xlsx_template_bytes(template_bytes, label="ES95411 template")
    template_has_vba = has_vba_macros(template_bytes)

    # source index (data_only=True)
    index = index_detail_blocks(source_workbooks, warnings=warnings)

    # template (data_only=False, keep_vba) — 쓰기 대상
    wb = openpyxl.load_workbook(
        io.BytesIO(template_bytes), keep_vba=template_has_vba, data_only=False,
    )
    summ = _find_summary_sheet(wb)
    if summ is None:
        wb.close()
        raise ValueError("template에 'Summary' 시트가 없습니다 (ES95411 양식 아님).")

    cols, first_row, total_row = _locate_table(summ)

    matched = 0
    stamped = 0
    fail_ids: list[str] = []
    performed_count = 0
    hours_sum = 0.0
    rows_seen = 0

    for r in range(first_row, total_row):
        rid = strip_id(summ.cell(row=r, column=cols.id).value)
        if not rid:
            continue
        rows_seen += 1
        planned = _norm(summ.cell(row=r, column=cols.performed).value).upper()
        sheet_key = strip_id(summ.cell(row=r, column=cols.sheetname).value)
        is_primary = bool(sheet_key) and (sheet_key == rid)

        # 미수행(X) → 결과 컬럼 공란 보장, Total 카운트 제외
        if planned == "X":
            _blank_result_cols(summ, r, cols, primary=is_primary)
            continue

        # 매칭: primary는 자기 ID, sub는 parent(sheetname) 키
        lookup_key = rid if is_primary else (sheet_key or rid)
        blk = index.get(lookup_key)

        if planned == "O":
            performed_count += 1
        elif planned == "":
            # 계획열 공란 → 결과 매칭되면 수행으로 간주, 아니면 skip(공란 유지)
            if blk and blk.has_result:
                safe_write(summ, r, cols.performed, "O")
                performed_count += 1
            else:
                if blk is not None:  # 매칭은 됐으나 P/F 미캐시(수식?) → silent 미수행 방지
                    incomplete.append(
                        f"row{r} {rid}: source P/F 미캐시(수식 추정) — 수행 판정 불가. "
                        f"source를 Excel로 1회 열어 저장 후 재시도 권장."
                    )
                _blank_result_cols(summ, r, cols, primary=is_primary)
                continue

        if blk is None:
            incomplete.append(f"row{r} {rid}: 매칭 detail 시트 없음 (key={lookup_key})")
            _blank_result_cols(summ, r, cols, primary=is_primary)
            continue

        matched += 1
        if is_primary:
            _stamp_primary(summ, r, blk, cols)
            if not blk.pf:
                incomplete.append(
                    f"row{r} {rid}: source P/F 미캐시(수식 추정) — Excel로 1회 열어 저장 권장."
                )
            if blk.total_hours == "":
                incomplete.append(
                    f"row{r} {rid}: source 분석시간 미캐시 — Excel로 1회 열어 저장 권장."
                )
        else:
            # sub 행: meta 컬럼(분석차수/SWVer/Tester/hours)은 parent와 세로 병합되거나
            # (ES95411 H17:H27 등) Debugger처럼 행별로 양식에 채워져 있으므로 빌더가
            # 건드리지 않고 template에 위임한다. P/F만 parent에서 승계 + 경고.
            safe_write(summ, r, cols.pf, blk.pf)
            _apply_pf_fill(summ, r, cols.pf, blk.is_fail)
            warnings.append(
                f"row{r} {rid}: sub 항목 — parent '{blk.test_id}' P/F({blk.pf}) 승계. "
                f"개별 P/F 확인 권장."
            )
        stamped += 1

        # Total 집계
        if blk.is_fail:
            fail_ids.append(rid)
        if is_primary:
            hours_sum += _to_float(blk.total_hours)

    # Total 행
    safe_write(summ, total_row, cols.performed, performed_count)
    safe_write(summ, total_row, cols.hours, _fmt_num(hours_sum))
    safe_write(summ, total_row, cols.pf, len(fail_ids))

    # 헤더 블록: Fail 개수 / 종합 결과 + meta override
    _apply_header(summ, meta, cols, first_row, fail_count=len(fail_ids))

    # 저장 + 외부링크 정화
    # 라운드 107 — ES95411 템플릿 보존 수식(집계 셀)이 캐시 미저장으로 재계산 안 하는
    # 뷰어에서 공백이 되지 않도록 fullCalcOnLoad. sanitize_xlsm_external_links는
    # calcPr 미접촉(검증)이라 정화 후에도 플래그 보존. 캐시 불변→다운스트림 영향 0.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # pragma: no cover — openpyxl 버전 차 방어
        pass
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    cleaned, removed = sanitize_xlsm_external_links(out.getvalue())
    out = io.BytesIO(cleaned)
    out.seek(0)
    if removed:
        warnings.append(f"외부 링크 잔재 {removed}건 제거")

    if filename is None:
        ver = meta.release_sw_version or "draft"
        date = short_date(meta.test_date) if meta.test_date else ""
        suffix = f"_v{ver}" + (f"_{date}" if date else "")
        filename = f"({meta.project_id}) Test Result Report{suffix}.xlsm"

    summary = {
        "rows_seen": rows_seen,
        "performed_count": performed_count,
        "matched_rows": matched,
        "stamped_rows": stamped,
        "fail_count": len(fail_ids),
        "fail_ids": fail_ids,
        "total_hours": _fmt_num(hours_sum),
        "source_count": len(source_workbooks),
        "indexed_detail_sheets": len(index),
        "incomplete_count": len(incomplete),
        "detected_columns": list(cols.detected),
        "overall_result": "Fail" if fail_ids else "Pass",
    }
    return SwReportBuildResult(
        ok=True,
        xlsm_io=out,
        filename=filename,
        warnings=warnings,
        summary=summary,
        incomplete_rows=incomplete,
        vba_macros_preserved=template_has_vba,
    )


# --- preview (JSON, Excel 미빌드) --------------------------------------------
def preview_summary_report(
    template_bytes: bytes,
    source_workbooks: list[tuple[str, bytes]],
    meta: SwReportBuildMeta | None = None,
) -> dict[str, Any]:
    """build 없이 통합 표 행 데이터 + 집계만 JSON으로 반환 (프론트 미리보기용).

    Excel을 쓰지 않으므로 template은 catalog(E/G/P/F열) 읽기만 — ``data_only=True``.
    매칭 로직/집계는 build와 동일 규칙.
    """
    meta = meta or SwReportBuildMeta()
    warnings: list[str] = []
    validate_xlsx_template_bytes(template_bytes, label="ES95411 template")
    index = index_detail_blocks(source_workbooks, warnings=warnings)

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), read_only=True, data_only=True)
    try:
        summ = _find_summary_sheet(wb)
        if summ is None:
            raise ValueError("template에 'Summary' 시트가 없습니다 (ES95411 양식 아님).")
        cols, first_row, total_row = _locate_table(summ)
        rows: list[dict[str, Any]] = []
        performed = 0
        fail_ids: list[str] = []
        hours_sum = 0.0
        matched = 0
        for r in range(first_row, total_row):
            rid = strip_id(summ.cell(row=r, column=cols.id).value)
            if not rid:
                continue
            planned = _norm(summ.cell(row=r, column=cols.performed).value).upper()
            sheet_key = strip_id(summ.cell(row=r, column=cols.sheetname).value)
            is_primary = bool(sheet_key) and (sheet_key == rid)
            lookup_key = rid if is_primary else (sheet_key or rid)
            blk = None if planned == "X" else index.get(lookup_key)
            row_d: dict[str, Any] = {
                "row": r,
                "id": rid,
                "name": _norm(summ.cell(row=r, column=cols.name).value),
                "category": _norm(summ.cell(row=r, column=cols.category).value),
                "tool": _norm(summ.cell(row=r, column=cols.tool).value),
                "sheet_name": _norm(summ.cell(row=r, column=cols.sheetname).value),
                "planned": planned or "(auto)",
                "is_primary": is_primary,
                "matched": blk is not None,
            }
            if blk is not None:
                row_d.update({
                    "iteration": blk.test_iteration,
                    "sw_version": blk.sw_version,
                    "tester": blk.tester,
                    "debugger": blk.debugger,
                    "hours": blk.total_hours,
                    "pf": blk.pf,
                    "source": blk.source_label,
                })
                matched += 1
                if planned != "X":
                    performed += 1
                    if blk.is_fail:
                        fail_ids.append(rid)
                    if is_primary:
                        hours_sum += _to_float(blk.total_hours)
            elif planned == "O":
                performed += 1
            rows.append(row_d)
    finally:
        wb.close()

    return {
        "rows": rows,
        "warnings": warnings,
        "summary": {
            "rows_total": len(rows),
            "performed_count": performed,
            "matched_rows": matched,
            "fail_count": len(fail_ids),
            "fail_ids": fail_ids,
            "total_hours": _fmt_num(hours_sum),
            "source_count": len(source_workbooks),
            "indexed_detail_sheets": len(index),
            "detected_columns": list(cols.detected),
            "overall_result": "Fail" if fail_ids else "Pass",
        },
    }


# --- stamp 헬퍼 --------------------------------------------------------------
def _stamp_primary(ws: Worksheet, row: int, blk: ResultBlock, cols: SummaryCols) -> None:
    safe_write(ws, row, cols.iter, blk.test_iteration)
    safe_write(ws, row, cols.swver, blk.sw_version)
    safe_write(ws, row, cols.tester, blk.tester)
    safe_write(ws, row, cols.debugger, blk.debugger)
    # hours는 숫자일 때만 stamp (수식 캐시 미스로 빈/비숫자면 미기록 + 호출자 incomplete).
    if isinstance(blk.total_hours, (int, float)):
        safe_write(ws, row, cols.hours, _fmt_num(blk.total_hours))
    safe_write(ws, row, cols.pf, blk.pf)
    _apply_pf_fill(ws, row, cols.pf, blk.is_fail)


def _apply_pf_fill(ws: Worksheet, row: int, col: int, is_fail: bool) -> None:
    """P/F 셀 강조 — 데이터 변경을 충실히 반영.

    Fail이면 빨강(``mark_fail_cell``). Pass이면, 셀에 **우리 FAIL 색**(``FAIL_FILL_RGB``)이
    남아 있을 때만 제거한다 — populated 템플릿/직전 출력에서 Fail→Pass로 바뀌었을 때
    빨강 잔존을 방지하되, 양식의 정당한 배경(다른 색)은 건드리지 않는다.
    """
    if is_fail:
        mark_fail_cell(ws, row, col)
        return
    ar, ac = resolve_merge_anchor(ws, row, col)
    cell = ws.cell(row=ar, column=ac)
    rgb = str(getattr(getattr(getattr(cell, "fill", None), "fgColor", None), "rgb", "") or "").upper()
    if rgb.endswith(FAIL_FILL_RGB[-6:]):
        cell.fill = PatternFill(fill_type=None)


def _blank_result_cols(ws: Worksheet, row: int, cols: SummaryCols, *, primary: bool) -> None:
    """미수행/미매칭 행의 결과 컬럼을 공란화 (stale 제거).

    primary 행: 결과 6열(분석차수/SWVer/Tester/Debugger/시간/P-F) 전부. 단일·비병합
    행이라 안전. sub 행: meta는 template에 위임(병합/행별 Debugger 보존)하고 P/F만 공란화.
    """
    if primary:
        for c in cols.result_primary:
            safe_write(ws, row, c, None)
    else:
        safe_write(ws, row, cols.pf, None)
    _apply_pf_fill(ws, row, cols.pf, False)  # 미수행/미매칭 → 빨강 잔존 제거


def _apply_header(
    ws: Worksheet, meta: SwReportBuildMeta, cols: SummaryCols, first_row: int,
    *, fail_count: int,
) -> None:
    """헤더 블록 — meta 값 있으면 override + Fail/Result 계산값 기록.

    헤더 라벨(Project/Phase/…/Fail/Result)을 라벨 열(No 열)에서 찾고, 값 열은 라벨
    우측 첫 비어있지 않은 셀로 자동 감지(머지된 라벨 폭 무관) — 양식이 달라도 견고.
    """
    label_to_value = {
        "Project": meta.project_full_name or meta.project_id,
        "Phase": meta.phase,
        "Software Platform Ver.": meta.software_platform_ver or meta.release_sw_version,
        "Product": meta.product,
        "검증 대상": meta.test_target,
        "ASIL 등급": meta.asil_level,
        "Complier": meta.compiler,
        "MCU": meta.mcu,
    }
    # 라벨 열(No 열)에서 헤더 라벨 행 찾기 (표 header 이전 영역 한정)
    label_row: dict[str, int] = {}
    for r in range(1, first_row):
        b = _norm(ws.cell(row=r, column=cols.no).value)
        if b:
            label_row.setdefault(b, r)

    # 값 열 자동 감지: 첫 라벨 행에서 라벨 우측 첫 비어있지 않은(머지 anchor) 셀 열
    value_col = cols.id  # default: ID 열 = 표준 양식의 값 열(E)
    for lab in ("Project", "ASIL 등급", "Result", "Software Platform Ver."):
        if lab in label_row:
            rr = label_row[lab]
            for c in range(cols.no + 1, (ws.max_column or 17) + 1):
                if _norm(ws.cell(row=rr, column=c).value) != "":
                    value_col = c
                    break
            break

    for label, val in label_to_value.items():
        if val and label in label_row:
            safe_write(ws, label_row[label], value_col, val)

    # Fail 개수 / 종합 Result
    if "Fail" in label_row:
        safe_write(ws, label_row["Fail"], value_col, fail_count)
    if "Result" in label_row:
        safe_write(ws, label_row["Result"], value_col, "Fail" if fail_count else "Pass")


# --- 숫자 유틸 ---------------------------------------------------------------
def _to_float(v: Any) -> float:
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return 0.0


def _fmt_num(f: float):
    return int(f) if float(f).is_integer() else round(f, 2)


__all__ = [
    "SwReportBuildMeta",
    "SwReportBuildResult",
    "ResultBlock",
    "build_summary_report",
    "preview_summary_report",
    "extract_result_block",
    "index_detail_blocks",
    "strip_id",
    "normalize_pf",
]
