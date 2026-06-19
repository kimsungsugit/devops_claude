"""Excel xlsx/xlsm 템플릿 빌더 공통 helper.

회사 표준 SwUT 빌더 (Coverage / SUTR) 가 공통으로 사용하는:
- 머지셀 anchor 보정 (회사 xlsx는 거의 모든 셀이 머지)
- label-value KV 쓰기
- 짧은 날짜 변환
- **xlsx/xlsm bytes 입력 검증** (ZIP bomb / 헤더 위조 방어)

검증 helper는 reviewer 권고에 따라 빌더 진입점에서 호출 의무.
"""
from __future__ import annotations

import io
import re
import weakref
import zipfile
from typing import Any

from .design_tokens import (
    ASIL_A_FILL_RGB as _ASIL_A_FILL_RGB,
    ASIL_B_FILL_RGB as _ASIL_B_FILL_RGB,
    ASIL_C_FILL_RGB as _ASIL_C_FILL_RGB,
    ASIL_D_FILL_RGB as _ASIL_D_FILL_RGB,
    ASIL_QM_FILL_RGB as _ASIL_QM_FILL_RGB,
    FAIL_FILL_RGB as _FAIL_FILL_RGB,
    USER_INPUT_FILL_RGB as _USER_INPUT_FILL_RGB,
    USER_INPUT_PLACEHOLDER,
)

try:
    import openpyxl  # type: ignore
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # type: ignore
    from openpyxl.styles import PatternFill  # type: ignore
    from openpyxl.utils.exceptions import IllegalCharacterError  # type: ignore
    _HAS_PATTERN_FILL = True
except ImportError:  # pragma: no cover - openpyxl 미설치 fail-safe
    openpyxl = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")  # type: ignore
    IllegalCharacterError = ValueError  # type: ignore[assignment,misc]
    _HAS_PATTERN_FILL = False

# ZIP bomb 한계 — 압축 해제 후 100MB 초과 시 거부. 일반 xlsx/xlsm은 수 MB.
_MAX_DECOMPRESSED_SIZE = 100 * 1024 * 1024
# 단일 파일이 의심스러울 정도로 큰 경우 — 보통 sharedStrings.xml도 수 MB 이내
_MAX_SINGLE_FILE_SIZE = 50 * 1024 * 1024
# 압축비 상한 (decompressed / compressed). ZIP bomb은 보통 1000+.
_MAX_COMPRESSION_RATIO = 200

# Office Open XML (xlsx/xlsm/docx) 매직 바이트 — ZIP 헤더와 동일.
_XLSX_MAGIC = b"PK\x03\x04"


class TemplateValidationError(ValueError):
    """xlsx/xlsm bytes 입력이 유효하지 않을 때 — ZIP bomb 방어용."""


class BuildMetaValidationError(ValueError):
    """빌더 입력 메타가 유효하지 않을 때 — endpoint 노출 전 입력 신뢰 경계 (deep-reviewer X3)."""


# Hyundai/Mobis 표준 release 버전 형식. 미충족 시 파일명 깨짐(`_v__240219_R.xlsx`).
_RELEASE_VERSION_RE = re.compile(r"^\d+\.\d+(\.\d+)?$")
# yyyy-mm-dd 또는 yyyy/mm/dd 또는 yy-mm-dd
_TEST_DATE_RE = re.compile(r"^\d{2,4}[-/]\d{1,2}[-/]\d{1,2}")


_MAX_NAME_LEN = 100   # 사람 이름 / Test Engineer / Author 등 최대 길이
_MAX_ISSUE_TEXT_LEN = 2000  # deviation issue_text 셀 한도 (xlsx 32767 한도보다 훨씬 작게)
_DOC_ID_SEQ_RE = re.compile(r"^\d+$")


def _safe_repr(v: str, limit: int = 40) -> str:
    """reviewer X8: 에러 메시지에 사용자 입력 그대로 노출 안 되도록 truncate + repr."""
    return repr(v)[:limit]


def validate_name_field(value: str, field_name: str, *, allow_empty: bool = True) -> None:
    """test_engineer/author/approver/reviewer 등 사람 이름 필드 검증 (H1 Critical).

    - 길이 100자 이내
    - 줄바꿈(`\\n`/`\\r`) 금지 — xlsx 셀 깨짐 방지

    Raises:
        BuildMetaValidationError
    """
    if not value:
        if allow_empty:
            return
        raise BuildMetaValidationError(f"{field_name} is empty")
    if len(value) > _MAX_NAME_LEN:
        raise BuildMetaValidationError(
            f"{field_name} 길이 {len(value)} > {_MAX_NAME_LEN} 한도"
        )
    if "\n" in value or "\r" in value:
        raise BuildMetaValidationError(
            f"{field_name}에 줄바꿈 문자 포함 — 단일 라인 필요 (값={_safe_repr(value)})"
        )


def validate_build_meta(
    release_sw_version: str,
    test_date: str,
    *,
    doc_id_sequence: str = "",
    test_engineer: str = "",
    author: str = "",
    approver: str = "",
    reviewer: str = "",
) -> None:
    """빌더 진입에서 필수 입력 검증 (deep-reviewer X3 + 5차 H1/H2).

    Raises:
        BuildMetaValidationError: 빈 string / 형식 미충족 시.
    """
    if not release_sw_version or not release_sw_version.strip():
        raise BuildMetaValidationError(
            "release_sw_version is empty — 빈 string은 파일명 `_v__...` 깨짐"
        )
    if not _RELEASE_VERSION_RE.match(release_sw_version.strip()):
        raise BuildMetaValidationError(
            f"release_sw_version {_safe_repr(release_sw_version)} 형식 미충족 — "
            "'\\d+.\\d+(.\\d+)?' 필요"
        )
    if not test_date or not test_date.strip():
        raise BuildMetaValidationError("test_date is empty")
    if not _TEST_DATE_RE.match(test_date.strip()):
        raise BuildMetaValidationError(
            f"test_date {_safe_repr(test_date)} 형식 미충족 — yyyy-mm-dd / yyyy/mm/dd 필요"
        )
    # H2: doc_id_sequence는 digit만 허용 (빈 string OK)
    if doc_id_sequence and not _DOC_ID_SEQ_RE.match(doc_id_sequence):
        raise BuildMetaValidationError(
            f"doc_id_sequence {_safe_repr(doc_id_sequence)} 비-digit — Doc ID 체계 위반"
        )
    # H1: 사람 이름 필드 4개 검증
    validate_name_field(test_engineer, "test_engineer")
    validate_name_field(author, "author")
    validate_name_field(approver, "approver")
    validate_name_field(reviewer, "reviewer")


def truncate_cell_text(value: Any, max_len: int = _MAX_ISSUE_TEXT_LEN) -> tuple[str, bool]:
    """xlsx 셀 한도(32,767자) 훨씬 이내로 truncate (H3 — uncaught IllegalCharacterError 방어).

    Returns:
        (truncated_value, was_truncated)
    """
    s = str(value) if value is not None else ""
    if len(s) <= max_len:
        return (s, False)
    return (s[:max_len] + " …(truncated)", True)


def validate_xlsx_template_bytes(data: bytes, *, label: str = "template") -> None:
    """xlsx/xlsm template bytes 검증 (Critical S — ZIP bomb / 헤더 위조 방어).

    Raises:
        TemplateValidationError: bytes가 유효한 xlsx/xlsm 아니거나 ZIP bomb 의심.
    """
    if not data:
        raise TemplateValidationError(f"{label} bytes empty")
    if data[:4] != _XLSX_MAGIC:
        raise TemplateValidationError(
            f"{label} magic bytes mismatch — 'PK\\x03\\x04' 헤더 부재"
        )

    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as e:
        raise TemplateValidationError(f"{label} not a valid ZIP: {e}")

    try:
        # xlsx/xlsm 필수 마커 (Open Office XML 구조) 확인
        namelist = zf.namelist()
        if not any(n.startswith("xl/") or n == "[Content_Types].xml" for n in namelist):
            raise TemplateValidationError(
                f"{label} Open Office XML 구조 미발견 — xl/ 또는 [Content_Types].xml 없음"
            )

        # ZIP bomb 검증: 압축비 + 총 크기 + 단일 파일 크기
        total_decompressed = 0
        for info in zf.infolist():
            if info.file_size > _MAX_SINGLE_FILE_SIZE:
                raise TemplateValidationError(
                    f"{label} 단일 entry 크기 초과: {info.filename} = "
                    f"{info.file_size:,} bytes (한도 {_MAX_SINGLE_FILE_SIZE:,})"
                )
            if info.compress_size > 0:
                ratio = info.file_size / info.compress_size
                if ratio > _MAX_COMPRESSION_RATIO:
                    raise TemplateValidationError(
                        f"{label} 압축비 ZIP bomb 의심: {info.filename} ratio={ratio:.0f}"
                    )
            total_decompressed += info.file_size
            if total_decompressed > _MAX_DECOMPRESSED_SIZE:
                raise TemplateValidationError(
                    f"{label} 총 압축 해제 크기 초과: {total_decompressed:,} bytes"
                )
    finally:
        # I3 (deep-reviewer): infolist 한도 초과로 raise 시에도 zf 명시적 close.
        zf.close()


# 빌더가 placeholder 시트 (예: 1.Traceability TC×Function 매트릭스 미구현) 셀에 부착하는
# 명시적 마커. consistency_checker / audit 도구가 같은 string으로 감지해서 evidence로
# 잘못 분류하지 않도록 한다 (시나리오 A self-validation false positive 방어).
BLANK_MARKUP = (
    "BLANK — auto-generated placeholder (TC×Function matrix pending). "
    "ISO 26262 evidence 사용 전 수동 작성 필수."
)


def sheet_is_blank_placeholder(ws: Any) -> bool:
    """시트 anchor 영역(A1~A3, B1~B3)에 BLANK_MARKUP이 있는지 확인.

    Returns:
        True면 빌더가 작성한 placeholder 시트 — consistency_checker는 해당 시트의
        데이터 추출을 skip하고 parse_warnings에 등록해야 함.
    """
    if ws is None:
        return False
    for r in range(1, 4):
        for c in range(1, 4):
            try:
                v = ws.cell(row=r, column=c).value
            except (AttributeError, IndexError):
                continue
            if v and isinstance(v, str) and v.startswith("BLANK"):
                return True
    return False


def has_vba_macros(data: bytes) -> bool:
    """xlsx/xlsm 안에 ``xl/vbaProject.bin`` entry 존재 여부.

    ``True``면 SUTR 빌더 출력 시 매크로가 ZIP entry로 보존됨. 단, **매크로 실행 가능성을
    보장하지는 않음** — VBA 코드가 시트 ref / Defined Name을 가리키는 경우 빌더의 셀
    변경으로 stale ref 깨질 위험 (deep-reviewer W2).
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return "xl/vbaProject.bin" in zf.namelist()
    except (zipfile.BadZipFile, OSError):
        return False


_VBA_REF_PATTERNS = (
    re.compile(rb"Cells\s*\(", re.IGNORECASE),
    re.compile(rb"Sheets\s*\(", re.IGNORECASE),
    re.compile(rb"Names\s*\(", re.IGNORECASE),
    re.compile(rb"Range\s*\(", re.IGNORECASE),
)


def build_release_history_row(
    meta: Any, *, doc_kind: str = "",
    out_warnings: list[str] | None = None,
) -> list[dict[str, str]]:
    """55-fix — 산출물별 release entry single row (사용자 결정 B).

    이전: `collect_git_history`로 git log 10건 채워서 audit reviewer 혼동
        (commit hash + 자동 commit + 매일 snapshot이 산출물 history로 보임).
    현재: 산출물의 release_sw_version + test_date + author 1 row만.
        reviewer/approver는 audit 입력으로 빈칸 둠.

    55-fix-2 W6: release_sw_version 또는 test_date 빈 시 warning 누적 +
        빈 cell silent fill 방지. router는 Pydantic으로 검증되지만 내부 호출
        (회귀, PoC, dataclass default)에서 빈 값 위험 — audit 추적성 보호.

    Args:
        meta: BuildMetaBase 또는 sub class (release_sw_version, test_date, author 속성).
        doc_kind: 옵션 — "SwUT Coverage Report" / "SwUT SUTR" / "SwIT Coverage Report"
            / "SwIT SITR" — description 식별 (W2 표준화).
        out_warnings: 옵션 — warning 누적 list. 빈 입력 사유 추가 (W6).

    Returns:
        list[1 row dict]. _write_history_sheet에 전달.
    """
    release = (getattr(meta, "release_sw_version", "") or "").strip()
    test_date = (getattr(meta, "test_date", "") or "").strip()
    author = getattr(meta, "author", "") or ""

    # 55-fix-2 W6 + 55-fix-3 W9: 빈 입력 audit 추적성 — silent fill 차단.
    # ISO 26262 audit 책임자 식별 필수 — author 누락도 검증.
    if out_warnings is not None:
        ctx = f" [{doc_kind}]" if doc_kind else ""  # I6: doc_kind context 부착
        if not release:
            out_warnings.append(
                f"History row release_sw_version 빈 string — meta.release_sw_version 누락 "
                f"(audit reviewer가 산출물에서 빈 version cell을 데이터 누락으로 오해 가능){ctx}"
            )
        if not test_date:
            out_warnings.append(
                f"History row test_date 빈 string — meta.test_date 누락{ctx}"
            )
        # 55-fix-3 W9: author 누락 — ISO 26262 audit 책임자 식별 필수
        if not author:
            out_warnings.append(
                f"History row author 빈 string — meta.author 누락 "
                f"(audit reviewer 책임자 식별 불가){ctx}"
            )

    # date format: yyyy-mm-dd → yy.mm.dd (collect_git_history와 동일)
    date_short = ""
    if len(test_date) >= 10:
        # "2026-05-14" → "26.05.14"
        try:
            normalized = test_date.replace("/", "-")
            parts = normalized.split("-")
            if len(parts) >= 3:
                yy = parts[0][-2:]
                mm = f"{int(parts[1]):02d}"
                dd = f"{int(parts[2][:2]):02d}"
                date_short = f"{yy}.{mm}.{dd}"
        except (ValueError, IndexError):
            date_short = test_date[:8]

    description = f"Initial release v{release}" if release else "Initial release"
    if doc_kind:
        description = f"{description} ({doc_kind})"

    return [{
        "version": f"v{release}" if release else "",
        "date": date_short,
        "description": description,
        "author": str(author)[:50],
        "reviewer": "",
        "approver": "",
    }]


def collect_git_history(
    repo_root: str | None = None,
    *,
    limit: int = 10,
) -> list[dict[str, str]]:
    """git log → History 시트 자동 채움용 row list (T134).

    55-fix 이후 SwUT/SwIT 빌더는 `build_release_history_row` 사용. git log 전체 표기는
    audit reviewer가 산출물 history로 혼동했음 (commit hash + 자동 commit + snapshot).
    본 함수는 유지 (backward compat) — 별도 호출처 또는 테스트 회귀에서 사용.

    Returns:
        list[{version, date, description, author, reviewer, approver}].
        git 명령 실패 시 빈 list.
    """
    import os
    import subprocess
    cwd = repo_root or os.getcwd()
    try:
        result = subprocess.run(
            ["git", "log", f"--max-count={limit}",
             "--pretty=format:%h\x1f%ai\x1f%an\x1f%s"],
            cwd=cwd, capture_output=True, text=True, encoding="utf-8",
            timeout=10,
        )
    except (subprocess.SubprocessError, OSError, FileNotFoundError):
        return []
    if result.returncode != 0:
        return []

    out: list[dict[str, str]] = []
    for line in result.stdout.splitlines():
        parts = line.split("\x1f")
        if len(parts) < 4:
            continue
        sha, iso_date, author, subject = parts[0], parts[1], parts[2], parts[3]
        # iso_date "2024-02-19 10:44:13 +0900" → "24.02.19"
        date_short = ""
        if len(iso_date) >= 10:
            yy = iso_date[2:4]
            mm = iso_date[5:7]
            dd = iso_date[8:10]
            date_short = f"{yy}.{mm}.{dd}"
        out.append({
            "version": f"v{sha[:7]}",
            "date": date_short,
            "description": subject[:200],
            "author": author[:50],
            "reviewer": "",
            "approver": "",
        })
    return out


def inspect_vba_refs(data: bytes) -> list[str]:
    """VBA 매크로의 stale ref 위험 패턴 감지 (5차 reviewer I1).

    ``vbaProject.bin`` 안 raw bytes에서 ``Cells(`` / ``Sheets(`` / ``Names(`` / ``Range(``
    같은 cell/sheet 참조 호출을 grep. 발견되면 빌더의 셀 변경으로 stale ref가 될 수 있어
    warning 권장.

    Returns:
        발견된 패턴 이름 list (예: ``["Cells(", "Sheets("]``). 빈 list면 의심 패턴 없음.
        VBA entry 없거나 읽기 실패 시도 빈 list.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            if "xl/vbaProject.bin" not in zf.namelist():
                return []
            vba_bytes = zf.read("xl/vbaProject.bin")
    except (zipfile.BadZipFile, OSError, KeyError):
        return []

    found: list[str] = []
    for pat in _VBA_REF_PATTERNS:
        if pat.search(vba_bytes):
            # 패턴 source에서 시각화용 이름만 추출 (예: ``rb"Cells\s*\("`` → ``"Cells("``).
            label = (
                pat.pattern.decode("ascii", errors="replace")
                .replace("\\s*", "")
                .replace("\\(", "(")
            )
            found.append(label)
    return found


# ---------------------------------------------------------------------------
# Sheet helpers (머지셀 보정)
# ---------------------------------------------------------------------------

def find_kv_row(
    ws: Any, label: str, max_row: int = 50, *, min_row: int = 1,
) -> tuple[int, int] | None:
    """시트의 첫 N행에서 label 셀 위치(row,col) 찾기.

    라운드 96-final: ``min_row`` 추가 — 같은 라벨이 서명란(상단)과 표지 항목
    (하단, 예: KJPDS02 v1.01 Cover의 I2 'Author'와 C30 'Author')에 중복될 때
    두 번째 occurrence를 찾기 위한 시작 행 (default 1 — backward compat).
    """
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.strip() == label:
                return (cell.row, cell.column)
    return None


# 라운드 89 perf — worksheet별 row-indexed merge anchor 캐시.
# 이전 resolve_merge_anchor는 ws.merged_cells.ranges 전체를 매 호출 선형 스캔(O(merges)).
# auto_expand_row_block(copy_merge)가 확장 행마다 template merge를 복제해 merge가
# 행 수에 비례 누적 → safe_write마다 O(merges) → 전체 O(n²) (py-spy n=24: safe_write의
# resolve_merge_anchor가 최대 hotspot). row→[(min_col,max_col,anchor)] 인덱스로 해당
# row의 merge만 검사(O(K)). merge 개수가 바뀔 때만 재색인. WeakKeyDictionary로 ws GC 시
# 자동 정리 + ws 객체 비오염.
_MERGE_ANCHOR_CACHE: "weakref.WeakKeyDictionary[Any, tuple[int, dict[int, list]]]" = (
    weakref.WeakKeyDictionary()
)


def resolve_merge_anchor(ws: Any, row: int, col: int) -> tuple[int, int]:
    """좌표가 머지 영역 안이면 top-left anchor 좌표로 보정 (row-indexed 캐시)."""
    merges = ws.merged_cells.ranges
    n = len(merges)
    cached = _MERGE_ANCHOR_CACHE.get(ws)
    if cached is None or cached[0] != n:
        by_row: dict[int, list] = {}
        for mr in merges:
            span = (mr.min_col, mr.max_col, (mr.min_row, mr.min_col))
            for r in range(mr.min_row, mr.max_row + 1):
                by_row.setdefault(r, []).append(span)
        cached = (n, by_row)
        try:
            _MERGE_ANCHOR_CACHE[ws] = cached
        except TypeError:  # ws가 weakref 불가하면 캐시 skip (정확성 유지)
            for mr in merges:
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    return (mr.min_row, mr.min_col)
            return (row, col)
    for minc, maxc, anchor in cached[1].get(row, ()):
        if minc <= col <= maxc:
            return anchor
    return (row, col)


def safe_write(ws: Any, row: int, col: int, value: Any) -> bool:
    """머지 영역 anchor 보정 후 쓰기. 머지된 비-anchor 셀이면 silent skip + False.

    2026-06-19 (deep-review C1) — C 소스/로그 유래 문자열(예: 2.Deviation G열 소스
    발췌)에 Excel 불법 제어문자(form-feed \\x0c, \\x07, \\x1a 등)가 섞이면 openpyxl이
    셀 .value 대입 시 ``IllegalCharacterError``를 raise한다. 이전엔 ``AttributeError``
    만 잡아 예외가 빌더 전체로 전파 → 산출물 생성 자체가 중단됐다. 불법문자를 제거
    (``\\t``\\n``\\r`` 합법 문자는 보존)하고 1회 재시도해 단일 셀 이상이 전체 빌드를
    깨지 않도록 한다(모든 cell-write 경로 공통 방어).
    """
    anchor_row, anchor_col = resolve_merge_anchor(ws, row, col)
    try:
        ws.cell(row=anchor_row, column=anchor_col).value = value
        return True
    except AttributeError:
        return False
    except IllegalCharacterError:
        try:
            ws.cell(row=anchor_row, column=anchor_col).value = (
                ILLEGAL_CHARACTERS_RE.sub("", str(value))
            )
            return True
        except (AttributeError, IllegalCharacterError, ValueError):
            return False


def force_write_cell(
    ws: Any, row: int, col: int, value: Any,
    *, style_ref: tuple[int, int] | None = None,
) -> bool:
    """라운드 96 — orphan ``MergedCell``에도 강제 기록.

    ``safe_write`` 상위 호환: live 머지면 anchor 보정 후 기록(동일 동작),
    추가로 **orphan MergedCell**(``ws.merged_cells.ranges``에 없는데 셀 객체가
    ``MergedCell``로 남은 상태 — openpyxl 머지 해제 quirk)도 처리한다.
    이 경우 stale 객체를 ``_cells``에서 제거 → fresh ``Cell`` 재생성 후 기록.

    배경: 회사 양식의 일부 SwCom 그룹 C 셀이 세로 병합돼 있었고, 파이프라인
    앞단계에서 해제되며 anchor만 normal Cell로 복원, continuation 행은
    ``MergedCell`` 객체로 잔존 → ``safe_write``가 ``AttributeError``로 silent
    skip하여 C(Component) 공란 발생(라운드 96 진단). live 머지가 아니므로
    셀 재생성은 머지 무결성에 영향 없음.

    Args:
        style_ref: ``(row, col)`` 지정 시, orphan을 fresh ``Cell``로 재생성한
            **경우에 한해** 그 좌표 셀의 border/fill/font/alignment를 복사한다
            (라운드 97 fix). 재생성된 Cell은 기본 스타일(테두리 無)이라 회사
            양식 세로 테두리가 끊김 → 같은 행의 정상 셀(예 D열) 스타일을
            전파해 양식 일관성 보존. orphan이 아니면 기존 스타일 유지(복사 안 함).
    """
    import copy as _copy

    from openpyxl.cell.cell import MergedCell as _MergedCell
    anchor_row, anchor_col = resolve_merge_anchor(ws, row, col)
    cell = ws.cell(row=anchor_row, column=anchor_col)
    if isinstance(cell, _MergedCell):
        # live 머지 range 없는 orphan → stale 객체 제거 후 fresh Cell 생성.
        ws._cells.pop((anchor_row, anchor_col), None)
        cell = ws.cell(row=anchor_row, column=anchor_col)
        # fresh Cell은 기본 스타일(테두리/배경 無) → 양식 일관성 위해 인접
        # 정상 셀 스타일 전파 (라운드 97).
        if style_ref is not None:
            ref = ws.cell(row=style_ref[0], column=style_ref[1])
            if not isinstance(ref, _MergedCell):
                cell.border = _copy.copy(ref.border)
                cell.fill = _copy.copy(ref.fill)
                cell.font = _copy.copy(ref.font)
                cell.alignment = _copy.copy(ref.alignment)
    try:
        cell.value = value
        return True
    except AttributeError:
        return False
    except IllegalCharacterError:
        # 2026-06-19 (deep-review C1 sibling) — safe_write와 동일 방어. 현재 호출처는
        # 통제된 수식만 기록하나, shared helper 미래 오용 대비 불법 제어문자 sanitize.
        try:
            cell.value = ILLEGAL_CHARACTERS_RE.sub("", str(value))
            return True
        except (AttributeError, IllegalCharacterError, ValueError):
            return False


def strip_external_links(wb: Any) -> int:
    """라운드 101 — 끊긴 외부 워크북 링크 + 외부참조 defined names 제거.

    회사 ★개발템플릿(SUTR/Coverage 공용)에 다른 양식(독일어 HARA
    ``FSE_HARA_v0.1_DE.xlsx``, 네트워크 경로 ``\\\\kffile1\\...``)에서 복사된
    externalLink 잔재가 있어, 산출물을 열 때 Excel이 "연결 업데이트/복구" 경고를
    띄운다. 외부 워크북이 우리 환경에 없으므로 끊긴 참조 — 산출물엔 불필요한 잔재.

    제거 대상:
      1) ``wb._external_links`` (externalLink 파트 — save 시 재생성 안 됨).
      2) workbook-scoped defined names 중 외부참조(``[N]...`` 형식) — externalLink가
         사라지면 깨지므로 함께 제거. 시트별 ``_xlnm.Print_Area`` 등 정상 이름은 보존
         (값에 ``[`` 외부 인덱스 없음).

    Returns:
        제거한 항목 수 (external link + defined name).
    """
    removed = 0
    ext = getattr(wb, "_external_links", None)
    if ext:
        removed += len(ext)
        wb._external_links = []
    # 외부참조 defined names (openpyxl 3.1 dict-like API).
    try:
        for _nm in list(wb.defined_names.keys()):
            _val = getattr(wb.defined_names[_nm], "value", "") or ""
            if "[" in _val and "]" in _val:
                del wb.defined_names[_nm]
                removed += 1
    except (AttributeError, TypeError, KeyError):
        pass
    return removed


def compact_empty_styled_cells(data: bytes) -> tuple[bytes, int]:
    """라운드 106 — openpyxl 3.1.5가 빈 스타일 셀을 ``<c s=".." t="n"></c>``
    (값 None인데 닫는 태그 명시)로 저장하는 비효율을 self-closing ``<c s=".."/>`` 로
    정규화 (저장된 xlsx/xlsm **bytes** zip 레벨 후처리).

    배경: 회사 감사본 양식은 격자 테두리를 위해 데이터 영역 전체에 스타일을 깔아,
    값 없는 셀이 시트의 90%+ 를 차지한다 (KJPDS02 PV SwUTR '3.Test Log' = 7899행
    × 275열 중 **92%(199만)가 빈 양식 셀**). openpyxl 3.1.5 는 이 빈 셀을
    self-closing 하지 않고 ``t="n"`` 타입 + 닫는 태그로 써서 비압축 XML 이 ~24%
    부푼다 (73MB → 56MB). 거대 시트의 비대는 Excel 열기 파싱 메모리/속도를
    악화시키고, save 중 메모리 압박으로 XML 이 잘리는(unclosed token) 손상 위험을
    키운다 (PV SwUTR '파일 안 열림' 보고의 근인).

    값 셀(``<v>``/``<is>`` 내용 보유)·스타일(``s`` 속성)·병합은 보존하고, **값 없는**
    빈 셀의 불필요한 ``t="n"></c>`` / ``t="inlineStr"></c>`` 만 self-closing 으로 압축.
    worksheet XML 만 대상 (vbaProject/media 등 바이너리 파트 무영향).

    Returns:
        ``(정규화된 bytes, 압축한 빈 셀 수)``. 대상이 없으면 원본 그대로 (count 0).
    """
    import io as _io
    import re as _re
    import zipfile as _zip

    _empty_n = _re.compile(rb'<c([^>]*) t="n"></c>')
    _empty_s = _re.compile(rb'<c([^>]*) t="inlineStr"></c>')
    zin = _zip.ZipFile(_io.BytesIO(data))
    total = 0
    out = _io.BytesIO()
    with _zip.ZipFile(out, "w", _zip.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            n = item.filename
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
                raw, c1 = _empty_n.subn(rb"<c\1/>", raw)
                raw, c2 = _empty_s.subn(rb"<c\1/>", raw)
                total += c1 + c2
            zout.writestr(item, raw)
    zin.close()
    if total == 0:
        return data, 0
    return out.getvalue(), total


def verify_xlsx_integrity(data: bytes) -> tuple[bool, str]:
    """라운드 106 — 저장된 xlsx/xlsm **bytes** 무결성 검증 (배포 전 손상 차단).

    거대 시트('3.Test Log' 등) save 가 메모리 압박으로 중단되면 worksheet XML 이
    ``</worksheet>`` 로 닫히지 못한 채 잘리거나(unclosed token) zip CRC 가 깨진다.
    이 손상은 Excel 에서 "파일을 열 수 없음/복구"로 나타난다 (PV SwUTR 실측:
    sheet5.xml CRC 손상 + unclosed token). 본 검증은 worksheet 파트의 (1) CRC
    (zipfile read 가 자동 검증), (2) ``</worksheet>`` 종료 태그 완결성을 확인해
    손상 산출물이 사용자에게 배포되기 전에 감지한다.

    Returns:
        ``(ok, error_message)``. ok=True 면 error 는 빈 문자열.
    """
    import io as _io
    import zipfile as _zip

    try:
        z = _zip.ZipFile(_io.BytesIO(data))
    except _zip.BadZipFile as e:
        return False, f"zip 손상: {e}"
    for item in z.infolist():
        n = item.filename
        if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
            try:
                raw = z.read(n)  # zipfile.read 가 CRC 자동 검증
            except _zip.BadZipFile as e:
                return False, f"{n} CRC 손상: {e}"
            if not raw.rstrip().endswith(b"</worksheet>"):
                return False, f"{n} XML 미완결(truncated — save 중단 의심)"
    return True, ""


def sanitize_xlsm_external_links(data: bytes) -> tuple[bytes, int]:
    """라운드 101 — 저장된 xlsx/xlsm **bytes**에서 외부링크 파트/참조를 zip 레벨 제거.

    ``strip_external_links(wb)`` (openpyxl 객체 레벨)는 ``keep_vba=True`` 로드 시
    외부링크 파트가 raw archive로 보존돼 save 시 그대로 재출력되므로 무효였다
    (라운드 101 진단). 따라서 **save 후 bytes**를 zip 레벨에서 직접 정화한다:

      1) ``xl/externalLinks/**`` 파트 전체 제거.
      2) ``[Content_Types].xml`` 의 externalLink Override 제거.
      3) ``xl/workbook.xml`` 의 ``<externalReferences>`` 블록 + 외부참조
         (``[N]...``) defined name 제거 (시트별 Print_Area 등 정상 이름 보존).
      4) ``xl/_rels/workbook.xml.rels`` 의 externalLink relationship 제거.

    외부링크 파트가 없으면 원본 bytes 그대로 반환 (변경 0).

    Returns:
        ``(정화된 bytes, 제거한 externalLink 파트 수)``.
    """
    import io as _io
    import re as _re
    import zipfile as _zip

    zin = _zip.ZipFile(_io.BytesIO(data))
    names = zin.namelist()
    ext_parts = [n for n in names if n.startswith("xl/externalLinks/")]
    # 라운드 102 — vbaProject dangling 검출: keep_vba=True save가 workbook.xml.rels
    # 에 vbaProject relationship은 쓰면서 vbaProject.bin 파트는 안 써서 끊긴 참조
    # 발생 (Excel "우리 산출물만 복구" 경고의 진짜 원인 — 라운드 102 진단). bin이
    # 없는데 rels에 vbaProject 참조가 있으면 그 relationship 제거 (회사 원본 동일 상태).
    _has_vba_bin = "xl/vbaProject.bin" in names
    _wb_rels_name = "xl/_rels/workbook.xml.rels"
    _vba_dangling = False
    if not _has_vba_bin and _wb_rels_name in names:
        _vba_dangling = "vbaproject" in zin.read(_wb_rels_name).decode(
            "utf-8", "replace").lower()
    if not ext_parts and not _vba_dangling:
        zin.close()
        return data, 0

    out = _io.BytesIO()
    zout = _zip.ZipFile(out, "w", _zip.ZIP_DEFLATED)
    for item in zin.infolist():
        n = item.filename
        if n.startswith("xl/externalLinks/"):
            continue  # 1) 파트 제거
        raw = zin.read(n)
        if n == "[Content_Types].xml":
            raw = _re.sub(
                rb'<Override PartName="/xl/externalLinks/[^"]*"[^>]*/>', b"", raw,
            )
        elif n == "xl/workbook.xml":
            txt = raw.decode("utf-8")
            # 3a) <externalReferences>...</externalReferences> 제거.
            txt = _re.sub(r"<externalReferences>.*?</externalReferences>", "", txt,
                          flags=_re.DOTALL)
            # 3b) 외부참조([N]...) defined name 제거 (정상 이름 보존).
            txt = _re.sub(
                r'<definedName\b[^>]*>\s*\[\d+\][^<]*</definedName>', "", txt,
            )
            raw = txt.encode("utf-8")
        elif n == "xl/_rels/workbook.xml.rels":
            raw = _re.sub(
                rb'<Relationship\b[^>]*externalLink[^>]*/>', b"", raw,
            )
            if _vba_dangling:
                # vbaProject.bin 없는데 남은 끊긴 relationship 제거 (대소문자 무관).
                raw = _re.sub(
                    rb'<Relationship\b[^>]*[Vv][Bb][Aa][Pp]roject[^>]*/>', b"", raw,
                )
        zout.writestr(item, raw)
    zout.close()
    zin.close()
    return out.getvalue(), len(ext_parts) + (1 if _vba_dangling else 0)


def clear_data_range(
    ws: Any,
    *,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    preserve_formula: bool = True,
    preserve_merged_anchor: bool = True,
    sentinel_patterns: list[str] | None = None,
) -> int:
    """라운드 D T601: 시트의 data row 영역 clear (template-copy partial overwrite 결함 fix).

    SwUT/SwIT builder가 template-copy 전략 사용 시 양식의 default 데이터
    (예: 이전 release의 419 함수, 4 deviation 등)를 clear하지 않고 신규 데이터를
    일부만 덮어쓰기 → audit 신뢰성 무너짐. 본 helper로 builder가 stamp 전에
    data row 영역을 명시 clear.

    Args:
        ws: openpyxl Worksheet.
        start_row / end_row: clear 대상 row range (inclusive, 1-based).
        start_col / end_col: clear 대상 col range (inclusive, 1-based).
        preserve_formula: True면 `=` 시작 cell (수식) 보존 — Test Summary 같은
            cross-sheet reference 수식 보호.
        preserve_merged_anchor: True면 머지 영역의 비-anchor cell은 건드리지 않음
            (anchor만 clear) — 머지 깨짐 방지.
        sentinel_patterns: 이 substring 중 하나를 포함한 cell 발견 시 그 row
            직전까지만 clear. 예: ['End of Document', '< End', '■ Appendix']
            — 양식의 끝 마커/Appendix 보호 (F7 자체평가 Round 1 C1/C3 fix).

    Returns:
        cleared cell count.
    """
    if start_row > end_row or start_col > end_col:
        return 0
    # F7 자체평가 R1 C1/C3 + R2 N2: sentinel 사전 탐지 — end_row 조기 종료.
    # N2 fix: substring 매칭 (`pat in v`)이 false positive — 함수명 'AppendixHelper'
    # 등이 'Appendix' substring 매칭하여 잘못 clear 무력화. strict 매칭으로 변경:
    # - 양식 끝 마커는 '<', '■', '※' prefix (예: '< End of Document >', '■ Appendix')
    # - 또는 strip 후 pat과 exact / startswith 매칭
    # - 'TOTALS' / 'GRAND TOTALS' 단독 cell만 매칭 (포함 부분 매칭 X)
    if sentinel_patterns:
        actual_end = end_row
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                try:
                    v = ws.cell(row=r, column=c).value
                except (AttributeError, IndexError):
                    continue
                if not isinstance(v, str):
                    continue
                v_stripped = v.strip()
                for pat in sentinel_patterns:
                    # 양식 마커는 보통 cell 단독으로 존재 (다른 텍스트 미혼합)
                    # strict 매칭 3가지: exact / startswith '<'/'■'/'※' + pat / pat이 단독 단어
                    if v_stripped == pat:
                        actual_end = r - 1
                        break
                    if v_stripped.startswith(("<", "■", "※")) and pat in v_stripped:
                        actual_end = r - 1
                        break
                if actual_end != end_row:
                    break
            if actual_end != end_row:
                break
        end_row = actual_end
        if start_row > end_row:
            return 0
    cleared = 0
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            try:
                cell = ws.cell(row=r, column=c)
            except (AttributeError, IndexError):
                continue
            if cell.value is None:
                continue
            # 수식 보존
            if preserve_formula and isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            # 머지 영역 비-anchor 보존
            if preserve_merged_anchor:
                anchor_r, anchor_c = resolve_merge_anchor(ws, r, c)
                if (anchor_r, anchor_c) != (r, c):
                    continue
            try:
                cell.value = None
                cleared += 1
            except AttributeError:
                continue
    return cleared


def auto_expand_row_block(
    ws: Any,
    *,
    insert_at_row: int,
    amount: int,
    template_row_idx: int,
    copy_style: bool = True,
    copy_merge: bool = True,
    copy_dimension: bool = True,
    max_col_scan: int | None = None,
) -> int:
    """라운드 73 T801: row 자동 확장 — template row의 style/merge/dimension 복제.

    회사 v3.01/v2.02 양식은 고정 sample slot (Coverage 5~15 row, Consistency 30 row 등).
    실데이터(1941 TC × 60 fn / 5000+ TC 미래)가 slot 초과 시 stamp 잘림 → 본 helper로
    insert_rows + template 단일 row의 style/merge/dimension 복제하여 양식 일관성 유지.

    Args:
        ws: openpyxl Worksheet.
        insert_at_row: 신규 row가 들어갈 위치 (1-based). 기존 이 row는 amount만큼 downshift.
        amount: 삽입할 row 수.
        template_row_idx: style/merge/dimension 복제할 template row (1-based).
            ⚠️ template_row_idx >= insert_at_row면 insert 후 template_row_idx += amount.
            본 함수는 insert 전에 style을 capture하므로 호출자가 인지 필요.
        copy_style: cell._style (font/fill/border/number_format/alignment 등) 복제.
        copy_merge: template row의 single-row merge range를 신규 row마다 동일 col span으로 추가.
            multi-row merge는 skip (복잡도 차단).
        copy_dimension: row_dimensions.height 복제.
        max_col_scan: style 복제 시 최대 col (None이면 ws.max_column).

    Returns:
        실제 삽입된 row 수 (실패 시 0).

    Side effects:
        - ws.merged_cells.ranges가 openpyxl 내부에서 자동 offset 보정 (insert_rows 표준 동작).
        - End-of-Document sentinel ('< End of Document >')는 insert_at_row 이상이면 자동 downshift.

    Raises:
        없음. 실패 시 silent 0 반환 + 호출자가 warning 누적.

    Performance:
        single-row insert가 N회 호출되면 O(N²) cost (openpyxl insert_rows는 매번
        전체 cell shift). 대량 row 필요 시 amount를 N으로 묶어 1회 호출 권장
        (raw insert_rows 1회 + 그 후 style replication N회 = O(N) cost).
    """
    if openpyxl is None or amount <= 0 or insert_at_row < 1 or template_row_idx < 1:
        return 0

    # 1) Capture template state BEFORE insert (template_row_idx 위치가 shift되기 전)
    template_merges: list[tuple[int, int]] = []  # single-row (row, col_min, col_max)
    template_multi_row_merges: list[tuple[int, int, int, int]] = []  # (row_offset, height, col_min, col_max)
    if copy_merge:
        # template_row_idx가 multi-row merge의 첫 row 또는 안쪽일 때 그 블록 전체 capture.
        # 회사 v3.01 SUTR Test Result는 1 TC당 6-row 단위 block merge (B17:B22 등).
        # auto_expand이 6-row block 단위로 호출되면 block 전체 multi-row merge 복제.
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row == template_row_idx and mr.max_row == template_row_idx:
                template_merges.append((mr.min_col, mr.max_col))
            elif mr.min_row == template_row_idx and mr.max_row > template_row_idx:
                # multi-row merge — template_row를 첫 row로 한 block.
                height = mr.max_row - mr.min_row + 1
                template_multi_row_merges.append((0, height, mr.min_col, mr.max_col))

    template_height: float | None = None
    if copy_dimension:
        rd = ws.row_dimensions.get(template_row_idx)
        if rd is not None:
            template_height = getattr(rd, "height", None)

    max_c = max_col_scan or ws.max_column
    template_styles: list[tuple[int, Any]] = []
    if copy_style:
        import copy as _copy
        for c in range(1, max_c + 1):
            try:
                src_cell = ws.cell(row=template_row_idx, column=c)
                if getattr(src_cell, "has_style", False):
                    template_styles.append((c, _copy.copy(src_cell._style)))
            except (AttributeError, IndexError):
                continue

    # 2) Insert rows — openpyxl 내장 merged_cells offset 보정 적용
    try:
        ws.insert_rows(insert_at_row, amount=amount)
    except (AttributeError, ValueError, TypeError):
        return 0

    # 3) Apply captured style/merge/dimension to each new row
    import copy as _copy
    try:
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError:
        get_column_letter = None  # type: ignore[assignment]

    for offset in range(amount):
        new_row = insert_at_row + offset

        if copy_style:
            for col_idx, style in template_styles:
                try:
                    ws.cell(row=new_row, column=col_idx)._style = _copy.copy(style)
                except (AttributeError, IndexError):
                    continue

        if copy_dimension and template_height is not None:
            try:
                ws.row_dimensions[new_row].height = template_height
            except (AttributeError, KeyError):
                pass

        if copy_merge and get_column_letter is not None:
            for min_col, max_col in template_merges:
                try:
                    start_ref = f"{get_column_letter(min_col)}{new_row}"
                    end_ref = f"{get_column_letter(max_col)}{new_row}"
                    ws.merge_cells(f"{start_ref}:{end_ref}")
                except (AttributeError, ValueError):
                    continue

    # multi-row merge 복제 (block 단위) — caller가 block size N의 정확한 배수로 amount 지정 시.
    # 회사 v3.01 SUTR Test Result는 6-row block (B17:B22 → B23:B28 → ...). amount % height == 0 시
    # 신규 row 전체에 동일 col span으로 multi-row merge 적용.
    if copy_merge and get_column_letter is not None and template_multi_row_merges:
        for _row_off, height, min_col, max_col in template_multi_row_merges:
            if height <= 1 or amount % height != 0:
                continue
            blocks = amount // height
            for b in range(blocks):
                block_start = insert_at_row + b * height
                block_end = block_start + height - 1
                try:
                    start_ref = f"{get_column_letter(min_col)}{block_start}"
                    end_ref = f"{get_column_letter(max_col)}{block_end}"
                    ws.merge_cells(f"{start_ref}:{end_ref}")
                except (AttributeError, ValueError):
                    continue

    return amount


def push_sentinel_to_last_row(
    ws: Any,
    *,
    sentinel_text: str = "< End of Document >",
    search_max_row: int | None = None,
) -> int | None:
    """라운드 73 T804 helper: '< End of Document >' sentinel이 데이터 row 중간에 있으면
    실제 마지막 데이터 row 뒤로 이동.

    insert_rows 사용 시 sentinel이 row 중간에 박힐 수 있어 audit 시각 깨짐 — 본 helper로
    sentinel을 ws.max_row 위치로 push.

    Args:
        ws: Worksheet.
        sentinel_text: 찾을 sentinel text (exact match, strip 후).
        search_max_row: scan 한계 (None이면 ws.max_row).

    Returns:
        sentinel이 이동된 새 row 번호 (None이면 sentinel 미발견).
    """
    if openpyxl is None:
        return None

    scan_end = search_max_row or ws.max_row
    # 라운드 89 perf — ws.max_column을 루프 밖으로 1회 hoist. openpyxl의 max_column은
    # property로 매 호출마다 전체 셀을 순회(O(cells))하므로, for r 루프 안에서 호출하면
    # O(rows × cells) = O(n²)가 된다 (py-spy: _write_test_log 시간의 99.7%가 여기).
    # 본 함수는 읽기 후 sentinel 값만 이동 — 컬럼 수 불변이라 hoist는 동작 동일.
    max_col = ws.max_column
    found_row: int | None = None
    found_col: int | None = None
    for r in range(1, scan_end + 1):
        for c in range(1, max_col + 1):
            try:
                v = ws.cell(row=r, column=c).value
            except (AttributeError, IndexError):
                continue
            if isinstance(v, str) and v.strip() == sentinel_text:
                found_row, found_col = r, c
                break
        if found_row is not None:
            break

    if found_row is None:
        return None

    # 실제 데이터 last row 찾기 (sentinel 위치 제외, 그 위에서 마지막 non-empty)
    last_data_row = found_row
    for r in range(found_row + 1, scan_end + 1):
        for c in range(1, max_col + 1):
            try:
                v = ws.cell(row=r, column=c).value
            except (AttributeError, IndexError):
                continue
            if v is not None and (not isinstance(v, str) or v.strip()):
                last_data_row = r
                break

    if last_data_row == found_row:
        return found_row  # 이미 마지막

    # sentinel을 last_data_row + 1로 이동
    try:
        ws.cell(row=found_row, column=found_col).value = None
        ws.cell(row=last_data_row + 1, column=found_col).value = sentinel_text
        return last_data_row + 1
    except (AttributeError, IndexError):
        return found_row


def update_cross_refs_after_row_expansion(
    ws: Any,
    *,
    old_totals_row: int,
    new_totals_row: int,
    scan_max_row: int = 30,
    scan_max_col: int = 30,
) -> int:
    """라운드 76 T1101: 양식 cross-ref formula의 hardcoded row reference 동적 갱신.

    회사 ★개발템플릿 V3 SwUTCV 4.Coverage R5 `=E25` / R6 `=L25` 같은 hardcoded row
    reference가 양식 default 가정 TOTALS row (예: 25 — 15 함수 slot용). c_parser
    primary merge 활성 시 row 폭증 (60→377) 후 R25는 c_parser 함수 row가 되어
    cross-ref formula 의미 깨짐. 본 helper로 `=E25` → `=E{new_totals_row}` 자동 갱신.

    Args:
        ws: openpyxl Worksheet.
        old_totals_row: 양식 default TOTALS row (보통 25).
        new_totals_row: row 폭증 후 실제 TOTALS row 위치.
        scan_max_row: cross-ref formula 탐색 row 한계 (보통 양식 헤더 30 row 안).
        scan_max_col: cross-ref formula 탐색 col 한계.

    Returns:
        갱신된 cell 수.

    Policy:
        - regex `=([A-Z]+){old_totals_row}\\b` (단어 경계) 매칭 — `=E25` ✓, `=E250` ✗
        - cross-sheet ref (`'2.Traceability'!H9` 등 `!` 포함)는 변경 안 함
        - calculated formula (`=(E5-F5)/E5` 같은 R{old}가 아닌 R5/R6) 영향 없음
        - merge anchor 보정 후 safe_write
        - idempotent: 이미 new_totals_row 참조하는 cell은 변경 안 함 (regex 매칭 0)

    backward-compat:
        old_totals_row == new_totals_row 시 변경 0 (auto_expand 미가동 케이스).
    """
    if openpyxl is None or old_totals_row == new_totals_row:
        return 0
    pattern = re.compile(rf"=([A-Z]+){old_totals_row}\b")
    updated = 0
    max_r = min(ws.max_row, scan_max_row)
    max_c = min(ws.max_column, scan_max_col)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            try:
                v = ws.cell(row=r, column=c).value
            except (AttributeError, IndexError):
                continue
            if not isinstance(v, str) or not v.startswith("="):
                continue
            # cross-sheet ref skip — `'시트명'!cell` 또는 `시트명!cell` 모두
            if "!" in v:
                continue
            # `\g<1>` 명시 — Python re가 `\1396`을 group 139로 잘못 해석하는 것 방지.
            new_v = pattern.sub(rf"=\g<1>{new_totals_row}", v)
            if new_v == v:
                continue
            if safe_write(ws, r, c, new_v):
                updated += 1
    return updated


def normalize_expanded_data_block(
    ws: Any,
    *,
    data_start: int,
    total: int,
    col_lo: int = 2,
    col_hi: int = 6,
    copy_fill: bool = False,
    finalize_medium_cols: list[int] | tuple[int, ...] | None = None,
) -> int:
    """라운드 101: auto_expand insert로 데이터 영역 끝에 밀려든 빈 양식 footer
    잔재(본문끝 무테 / Totals medium+강조fill / 빈 / sentinel)를 본문 표준 행
    스타일로 교정한다.

    배경: 회사 빈 양식은 고정 sample slot + Totals/sentinel footer를 갖는다.
    함수 수가 slot을 초과하면 ``auto_expand_row_block``이 ``insert_at_row``
    (=data_start+1)에 행을 삽입하면서 원래 footer 행들이 데이터 영역 끝으로
    downshift되고, 그 위에 데이터가 stamp되어 마지막 N행의 테두리/음영이 본문과
    불균일해진다 (KJPDS02 PV 보고 — 3.Consistency/4.Coverage 마지막 3~4행).

    기존 시트별 패스는 'No열에 테두리가 하나라도 있으면 정상 행'(``any(border)``)으로
    판정해 footer 잔재(bottom 없음·medium)를 정상으로 **오판**했다 (footer 행도
    top/left/right 테두리는 보유). 본 헬퍼는 데이터 영역 border 시그니처의 **다수결**로
    본문 표준을 정하고, 표준에서 벗어난 행만 표준 행의 border/font/alignment/
    (옵션)fill/height로 교정한다 — footer 잔재(소수)는 반드시 표준과 다른
    시그니처이므로 누락 없이 잡힌다.

    Args:
        ws: openpyxl Worksheet.
        data_start: 첫 데이터 행 (1-based).
        total: 데이터 행 수. 마지막 데이터 행 = ``data_start + total - 1``.
        col_lo, col_hi: 정규화 대상 열 범위 (inclusive, 1-based).
        copy_fill: True면 fill도 표준 행에서 복제 (footer 강조 fill 잔재 제거).
            B열 음영·노란 마킹 등 규칙 fill은 호출자가 헬퍼 호출 **후** 재적용해야 한다.
        finalize_medium_cols: 지정 시 마지막 데이터 행의 해당 열 bottom을 medium으로
            마감 (회사 양식 표끝 굵은 선). None이면 마감 생략 (별도 마감행 존재 시).

    Returns:
        교정된 행 수 (footer 잔재 없으면 0).

    Policy:
        - auto_expand 미가동(함수 ≤ slot) 시 데이터 영역에 footer 잔재가 없어
          전 행이 단일 시그니처 → 교정 0 (no-op, backward-compat).
        - MergedCell anchor가 아닌 셀은 skip (openpyxl write 금지).
        - fill을 건드리지 않는 기본(copy_fill=False) 모드는 노란 마킹 등 stamp 시점
          fill을 보존한다.
    """
    if openpyxl is None or total <= 0 or data_start < 1 or col_hi < col_lo:
        return 0
    import copy as _cp
    from collections import Counter

    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Border, Side

    def _sig(r: int) -> tuple:
        out = []
        for c in range(col_lo, col_hi + 1):
            b = ws.cell(r, c).border
            out.append(tuple(
                (getattr(b, s).style if getattr(b, s) else None)
                for s in ("top", "bottom", "left", "right")
            ))
        return tuple(out)

    rows = [data_start + i for i in range(total)]
    counter: Counter = Counter()
    row_sig: dict[int, tuple] = {}
    for r in rows:
        s = _sig(r)
        row_sig[r] = s
        counter[s] += 1
    if not counter:
        return 0
    std_sig, _ = counter.most_common(1)[0]
    std_row = next((r for r in rows if row_sig[r] == std_sig), data_start)
    std_h = ws.row_dimensions[std_row].height

    fixed = 0
    for r in rows:
        if row_sig[r] == std_sig:
            continue
        for c in range(col_lo, col_hi + 1):
            dst = ws.cell(r, c)
            src = ws.cell(std_row, c)
            if isinstance(dst, MergedCell) or isinstance(src, MergedCell):
                continue
            dst.border = _cp.copy(src.border)
            dst.font = _cp.copy(src.font)
            dst.alignment = _cp.copy(src.alignment)
            if copy_fill:
                dst.fill = _cp.copy(src.fill)
        if std_h is not None:
            ws.row_dimensions[r].height = std_h
        fixed += 1

    if finalize_medium_cols:
        last = data_start + total - 1
        med = Side(style="medium")
        for c in finalize_medium_cols:
            lc = ws.cell(last, c)
            if isinstance(lc, MergedCell):
                continue
            b = lc.border
            lc.border = Border(
                top=b.top, left=b.left, right=b.right, bottom=med,
                diagonal=b.diagonal, diagonalUp=b.diagonalUp,
                diagonalDown=b.diagonalDown,
            )
    return fixed


# 23차 T192 / 29차 W17: 시각 강조 RGB + placeholder 텍스트는
# ``design_tokens`` 단일 출처에서 import (위 import 블록 참조).


def _apply_fill(ws: Any, row: int, col: int, rgb: str) -> bool:
    """openpyxl PatternFill 적용 — openpyxl 미설치 / 머지셀 비-anchor면 silent False."""
    if not _HAS_PATTERN_FILL:
        return False
    anchor_row, anchor_col = resolve_merge_anchor(ws, row, col)
    try:
        ws.cell(row=anchor_row, column=anchor_col).fill = PatternFill(  # type: ignore[misc]
            start_color=rgb, end_color=rgb, fill_type="solid",
        )
        return True
    except AttributeError:
        return False


def mark_user_input_required(
    ws: Any, row: int, col: int, hint: str = "",
) -> bool:
    """23차 T192: 사용자 입력 필요 셀에 노란 배경 + placeholder 텍스트.

    Args:
        hint: 텍스트 뒤에 추가할 안내 (예: "Approver 이름").

    Returns:
        쓰기 + 색칠 모두 성공 시 True.
    """
    label = USER_INPUT_PLACEHOLDER + (f" — {hint}" if hint else "")
    wrote = safe_write(ws, row, col, label)
    filled = _apply_fill(ws, row, col, _USER_INPUT_FILL_RGB)
    return wrote and filled


def write_value_or_mark(
    ws: Any, row: int, col: int, value: Any, hint: str = "",
) -> bool:
    """23차 T192: value 있으면 그대로 쓰고, 빈 string/None이면 사용자 입력 표시.

    Returns:
        True if value written; False if marked as user-input-required.
    """
    if value:
        return safe_write(ws, row, col, value)
    mark_user_input_required(ws, row, col, hint=hint)
    return False


def mark_user_input_fill_only(ws: Any, row: int, col: int) -> bool:
    """노란 배경만 적용 (텍스트 미기입).

    수식의 operand 이거나 수식 셀에 placeholder STRING 을 쓰면 `=A-B` 같은 식이
    `#VALUE!` 가 되거나 audit 교차검증 수식이 파괴된다. 그런 셀에는 텍스트 대신
    배경색만으로 '사용자 입력 필요'를 표시한다 (값은 보존).
    """
    return _apply_fill(ws, row, col, _USER_INPUT_FILL_RGB)


def is_formula_cell(ws: Any, row: int, col: int) -> bool:
    """대상 셀(머지 anchor 보정)이 수식(`=...`)인지. 수식 셀 덮어쓰기 방지용."""
    anchor_row, anchor_col = resolve_merge_anchor(ws, row, col)
    try:
        v = ws.cell(row=anchor_row, column=anchor_col).value
    except (AttributeError, ValueError):
        return False
    return isinstance(v, str) and v.startswith("=")


def mark_fail_cell(ws: Any, row: int, col: int) -> bool:
    """23차 T192: 2.Consistency FAIL row 등 강조용 빨간 배경."""
    return _apply_fill(ws, row, col, _FAIL_FILL_RGB)


def mark_asil_d_function(ws: Any, row: int, col: int) -> bool:
    """30차 W21: 3.Coverage 시트의 ASIL D 함수 row 강조용 빨간 배경.

    색상 RGB는 ``mark_fail_cell`` 과 동일하나 호출 의미를 분리.
    - ``mark_fail_cell``: TC 실행 결과 FAIL 표시
    - ``mark_asil_d_function``: audit 검토 우선순위 (MC/DC 커버리지 필수) 표시

    동일 셀에 두 의미가 겹치면 호출 순서 보장으로 마지막 호출이 우선.
    빌더는 ASIL 강조를 FAIL 강조보다 나중에 호출 권장.
    """
    return _apply_fill(ws, row, col, _ASIL_D_FILL_RGB)


def mark_asil_b_function(ws: Any, row: int, col: int) -> bool:
    """31차 W29: 3.Coverage 시트의 ASIL B 함수 row 강조 — 연한 파랑.

    audit 검토 우선순위: 분기 커버리지 필수. ASIL D (빨간)보다 낮은 시인성
    으로 단계 구분.
    """
    return _apply_fill(ws, row, col, _ASIL_B_FILL_RGB)


def mark_asil_c_function(ws: Any, row: int, col: int) -> bool:
    """31차 W29: 3.Coverage 시트의 ASIL C 함수 row 강조 — 연한 주황.

    audit 검토 우선순위: MC/DC 커버리지 권장. ASIL D (빨간)과 ASIL B (파랑)
    사이 단계 — 색상으로 시각적 등급 차이 명시.
    """
    return _apply_fill(ws, row, col, _ASIL_C_FILL_RGB)


def mark_asil_a_function(ws: Any, row: int, col: int) -> bool:
    """라운드 81 T1502: 3.Coverage / SUTR Test Log ASIL A 함수 row 강조 — 연한 녹색.

    audit 검토 우선순위: 구문 커버리지로 충분 (가장 약한 안전 등급).
    HDPDM01 NE_GN7 환경처럼 A/QM 함수가 압도적인 경우에도 audit reviewer가
    분포를 한눈에 인지 가능. ASIL B (파랑)보다 낮은 시인성으로 단계 구분.
    """
    return _apply_fill(ws, row, col, _ASIL_A_FILL_RGB)


def mark_asil_qm_function(ws: Any, row: int, col: int) -> bool:
    """라운드 81 T1502: 3.Coverage / SUTR Test Log QM 함수 row 강조 — 연한 회색.

    Quality Management — 비안전, 정보성. 함수가 ISO 26262 안전 요구사항 외부
    (애플리케이션 로직 등)에 있음을 audit reviewer에게 명시. ASIL A 이상
    (녹색/파랑/주황/빨강)과 색조 자체 분리 (회색) — 안전 vs 비안전 한눈 구분.
    """
    return _apply_fill(ws, row, col, _ASIL_QM_FILL_RGB)


def write_label_or_mark(
    ws: Any,
    label: str,
    value: Any,
    hint: str = "",
    optional_labels: set[str] | None = None,
    out_warnings: list[str] | None = None,
    max_row: int = 50,
) -> bool:
    """23차 T192/W12: 라벨 옆 셀에 value 쓰기 — value 빈 경우 노란 placeholder.

    Coverage / SUTR builder 양쪽에서 공유 (이전 ``_write_label_or_mark`` 중복 제거).

    동작:
      1. ``find_kv_row``로 라벨 위치 찾기
      2. 머지 영역 보정 후 target_col 결정
      3. value 있으면 ``safe_write``, 없으면 ``mark_user_input_required`` (노란 강조)

    Args:
        optional_labels: 라벨 미발견 시 warnings 누적 skip 대상 (예: {"Reviewer"}).
        out_warnings: 미발견 라벨 사유 누적용.

    Returns:
        value가 실제로 쓰였으면 True (mark는 False).
    """
    pos = find_kv_row(ws, label, max_row)
    if not pos:
        if (optional_labels is None or label not in optional_labels) and out_warnings is not None:
            out_warnings.append(f"라벨 '{label}' 미발견 — 셀 쓰기 skip")
        return False
    row, col = pos
    target_col = col + 1
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            target_col = mr.max_col + 1
            break
    return write_value_or_mark(ws, row, target_col, value, hint=hint)


def write_value_after_label(
    ws: Any, label: str, value: Any, max_row: int = 50, *, min_row: int = 1,
) -> bool:
    """`label` 셀 옆 컬럼에 value 쓰기 (머지 영역 보정)."""
    pos = find_kv_row(ws, label, max_row, min_row=min_row)
    if not pos:
        return False
    row, col = pos
    target_col = col + 1
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            target_col = mr.max_col + 1
            break
    return safe_write(ws, row, target_col, value)


def write_signature_block(
    ws: Any,
    names: "dict[str, Any]",
    *,
    hint_map: "dict[str, str] | None" = None,
    max_row: int = 10,
) -> int | None:
    """라운드 96-final — 가로 연속 서명란 감지 후 라벨 '아래' 셀에 이름 기입.

    KJPDS02 v1.01 Cover 등은 서명란 라벨(Author/Reviewer/Approver)이 같은 행에
    가로로 인접(I2/J2/K2)하고 이름 칸은 라벨 아래 머지 셀(I3:I4 등)이다.
    기존 ``write_value_after_label``(라벨 우측 기입)을 이 레이아웃에 쓰면
    'Author' 값이 'Reviewer' **라벨을 덮어쓰는** 결함 발생 (2026-06-11 QA 확정).

    같은 행에 ``names``의 라벨이 2개 이상 발견되면 서명란 블록으로 판정하고,
    각 라벨 바로 아래 셀(머지 anchor 보정)에 값 기입 — 빈 값은 노란 마킹.

    Returns:
        처리한 라벨 행 번호 (블록 미발견 시 None — caller는 기존 우측 기입 fallback).
    """
    label_cells: dict[str, tuple[int, int]] = {}
    sig_row: int | None = None
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
        found: dict[str, tuple[int, int]] = {}
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip() in names:
                found.setdefault(v.strip(), (cell.row, cell.column))
        if len(found) >= 2:
            label_cells = found
            sig_row = next(iter(found.values()))[0]
            break
    if sig_row is None:
        return None
    for label, (r, c) in label_cells.items():
        tr, tc = resolve_merge_anchor(ws, r + 1, c)
        write_value_or_mark(
            ws, tr, tc, names.get(label),
            hint=(hint_map or {}).get(label, f"{label} 이름"),
        )
    return sig_row


def dot_date(s: str) -> str:
    """`2026-06-04` / `2026/06/04` → `2026.06.04` (회사 Cover Date 표기)."""
    return re.sub(r"[-/]", ".", (s or "").strip())


def stamp_cover_document_id(
    ws: Any,
    *,
    project_id: str,
    doc_filename_pattern: str = "",
    out_warnings: "list[str] | None" = None,
    label: str = "Document ID",
) -> bool:
    """라운드 96-final — Cover 'Document ID' 셀의 placeholder/phase 토큰 보정.

    보정 규칙 (값 변경 시 노란 마킹 + warning — serial은 사용자 검증 의무):
      1. ``[P_Name]`` placeholder (XXXX 공양식) → ``{project_id}_{phase}`` 치환.
      2. phase 토큰 불일치 (예: 빌드는 PV인데 ``_DV-`` 잔존) → ``_{phase}-`` 치환.
    phase는 ``doc_filename_pattern``의 ``_PV_``/``_DV_`` 토큰에서 도출 — 없으면 skip.

    Returns:
        값을 변경했으면 True.
    """
    pos = find_kv_row(ws, label, 50)
    if not pos:
        return False
    row, col = pos
    target_col = col + 1
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            target_col = mr.max_col + 1
            break
    tr, tc = resolve_merge_anchor(ws, row, target_col)
    cur = str(ws.cell(tr, tc).value or "")
    if not cur:
        return False
    m = re.search(r"_(PV|DV)_", doc_filename_pattern or "")
    phase = m.group(1) if m else ""
    new = cur
    if "[P_Name]" in new and project_id:
        new = new.replace(
            "[P_Name]", f"{project_id}_{phase}" if phase else project_id,
        )
    if phase:
        other = "DV" if phase == "PV" else "PV"
        if f"_{other}-" in new:
            new = new.replace(f"_{other}-", f"_{phase}-")
    if new == cur:
        return False
    safe_write(ws, tr, tc, new)
    _apply_fill(ws, tr, tc, _USER_INPUT_FILL_RGB)
    if out_warnings is not None:
        out_warnings.append(
            f"[cover] Document ID 자동 보정 '{cur}' → '{new}' — serial(끝 토큰)은 "
            "회사 문서 채번 규칙으로 검증 필요 (노란 마킹)"
        )
    return True


def short_date(s: str) -> str:
    """`2024-02-19` → `240219` (회사 표준 파일명용)."""
    if not s:
        return ""
    m = re.match(r"(\d{2,4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if not m:
        return s.replace("-", "").replace("/", "")[:6]
    yy = m.group(1)[-2:]
    return f"{yy}{int(m.group(2)):02d}{int(m.group(3)):02d}"


def copy_sheet_across_workbooks(
    src_ws: Any,
    dst_wb: Any,
    *,
    new_title: str,
    insert_index: int | None = None,
) -> Any:
    """src 워크북의 시트를 dst 워크북에 풀 카피 (라운드 92).

    openpyxl 은 워크북 간 ``copy_worksheet`` 를 native 지원하지 않으므로 cell value +
    ``cell._style`` (StyleArray 복제) + merged_cells + column_dimensions +
    row_dimensions + sheet_view/dimension 을 수동 복제한다. 회사 감사본 SUTR
    '3.Test Log' (4340r × 268c, 병합/폭/높이 보존)를 표준 SUTR 템플릿 wb 에 이식하는
    용도. R91 spec wb 베이스 → R92 표준 템플릿 wb 베이스 전환의 핵심 helper.

    Args:
        src_ws: 복사할 원본 worksheet (예: SwUTS spec '2.SW Unit Test Spec').
        dst_wb: 대상 Workbook (표준 SUTR 템플릿 로드본).
        new_title: 대상 시트 제목 (예: '3.Test Log').
        insert_index: 대상 wb 의 sheet 위치 (None이면 끝에 추가). 0-based.

    Returns:
        생성된 대상 worksheet.

    Notes:
        - 머지셀은 ``src_ws.merged_cells.ranges`` 를 그대로 string 으로 재적용.
        - 셀 스타일은 ``copy.copy(src_cell._style)`` (StyleArray 는 immutable index
          tuple 이라 shallow copy 로 충분). 폰트/채움/테두리/정렬/표시형식 보존.
        - 수식/문자열/숫자 value 모두 ``cell.value`` 로 그대로 이전 (data_only=False
          로 로드한 원본이어야 수식 보존).
        - column_dimensions (width/hidden) + row_dimensions (height/hidden) 복제.
    """
    import copy as _copy

    if new_title in dst_wb.sheetnames:
        # 동명 시트 충돌 방지 — 기존 제거 후 재생성.
        del dst_wb[new_title]

    dst_ws = dst_wb.create_sheet(title=new_title)

    # 1) cell value + style 복제.
    # 라운드 103 — cross-workbook 스타일은 ``_style``(인덱스 튜플) 직접 복사 금지.
    # src wb의 fontId/fillId/borderId 인덱스를 dst wb에 그대로 쓰면 dst의 fonts/
    # fills/borders 배열 길이를 초과해 끊긴 참조 → Excel "styles.xml 복구" 경고
    # (진단: fontId 38 > fonts 35). 실제 스타일 **객체**를 복사하면 openpyxl이 dst
    # wb 스타일 테이블에 등록하고 인덱스를 재할당 → 정합 보장.
    # 성능: 셀마다 객체 6개 복사는 O(cells×styles)로 268열×수천행에서 timeout.
    # src ``_style`` (StyleArray) → dst ``_style`` 매핑 캐시로 **unique 스타일만**
    # 변환 (수만 셀 → 수백 unique). 같은 dst wb 내 인덱스 재사용은 안전.
    _style_map: dict = {}
    max_row = src_ws.max_row
    max_col = src_ws.max_column
    for row in src_ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for src_cell in row:
            if src_cell.value is None and not getattr(src_cell, "has_style", False):
                continue
            dst_cell = dst_ws.cell(
                row=src_cell.row, column=src_cell.column, value=src_cell.value,
            )
            if getattr(src_cell, "has_style", False):
                try:
                    _key = tuple(src_cell._style)
                except (TypeError, AttributeError):
                    _key = None
                if _key is not None and _key in _style_map:
                    dst_cell._style = _style_map[_key]
                else:
                    try:
                        dst_cell.font = _copy.copy(src_cell.font)
                        dst_cell.fill = _copy.copy(src_cell.fill)
                        dst_cell.border = _copy.copy(src_cell.border)
                        dst_cell.alignment = _copy.copy(src_cell.alignment)
                        dst_cell.number_format = src_cell.number_format
                        dst_cell.protection = _copy.copy(src_cell.protection)
                        if _key is not None:
                            _style_map[_key] = dst_cell._style
                    except (AttributeError, TypeError):
                        pass

    # 2) merged_cells 복제.
    for mr in list(src_ws.merged_cells.ranges):
        try:
            dst_ws.merge_cells(str(mr))
        except (ValueError, AttributeError):
            pass

    # 3) column_dimensions 복제 (width / hidden / outline / 범위 min~max).
    # 라운드 96-final W-14 — openpyxl은 `<col min="18" max="57" hidden="1"/>` 같은
    # 범위 col 정의를 첫 열 key 1개에 min/max로 보존한다. min/max 미복제 시 dst가
    # 단일 열로 축소돼 DV spec 시트의 hidden 범위(cols 18-57/74-161)·폭 설정 소실.
    for key, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[key]
        for attr in ("min", "max", "width", "hidden", "outlineLevel",
                     "bestFit", "customWidth", "collapsed"):
            val = getattr(dim, attr, None)
            if val is not None:
                try:
                    setattr(dst_dim, attr, val)
                except (AttributeError, ValueError):
                    pass

    # 4) row_dimensions 복제 (height / hidden).
    for key, dim in src_ws.row_dimensions.items():
        dst_dim = dst_ws.row_dimensions[key]
        for attr in ("height", "hidden", "outlineLevel", "customHeight"):
            val = getattr(dim, attr, None)
            if val is not None:
                try:
                    setattr(dst_dim, attr, val)
                except (AttributeError, ValueError):
                    pass

    # 5) sheet view / freeze panes / dimension 복제 (best-effort).
    try:
        dst_ws.freeze_panes = src_ws.freeze_panes
    except (AttributeError, ValueError):
        pass
    try:
        dst_ws.sheet_format.defaultColWidth = src_ws.sheet_format.defaultColWidth
        dst_ws.sheet_format.defaultRowHeight = src_ws.sheet_format.defaultRowHeight
    except (AttributeError, ValueError):
        pass

    # 6) 위치 재배치.
    if insert_index is not None:
        cur_idx = dst_wb.sheetnames.index(new_title)
        dst_wb.move_sheet(new_title, offset=insert_index - cur_idx)

    return dst_ws
