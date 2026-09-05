"""VectorCAST Excel 리포트 생성기

TResultParser C# 프로그램의 Excel 생성 로직을 Python으로 포팅한 생성기입니다.
openpyxl을 사용하여 TResultParser와 동일한 형식의 Excel 리포트를 생성합니다.
"""

from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, Dict, Optional

try:
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    import re as _re_fallback
    Workbook = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    Comment = None
    get_column_letter = None
    # openpyxl 미설치 fail-safe — excel_template_utils.py와 동일 정의.
    ILLEGAL_CHARACTERS_RE = _re_fallback.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
    IllegalCharacterError = ValueError  # type: ignore[assignment,misc]
    BarChart = None
    Reference = None
    DataPoint = None

from backend.services.vcast_parser import (
    CoverageItem,
    MatixDataBank,
    MatricFunCallItem,
    MatricStatementItem,
    MetricsBank,
    TCBank,
)


class XlsCellStyle(Enum):
    """Excel 셀 스타일"""
    Title = 0
    Caption = 1
    General = 2
    Fixed = 3
    BgSkyBlue = 4
    BgYellow = 5
    BgOrange = 6
    BgRed = 7
    BgPink = 8
    BgSkyBlueL = 9
    BgPurpleL = 10
    BgPeachL = 11
    BgLightBlue = 12
    BgLightGray = 13
    BgPinkFRed = 14
    BgBlueFRed = 15
    FgRed = 16
    NONE = 17  # 예약어 None 대신 NONE 사용


class BorderEdge(Enum):
    """테두리 위치"""
    Left = 0
    Right = 1
    Top = 2
    Bottom = 3


@dataclass
class ExcelStyle:
    """Excel 스타일 정보"""
    font_color: str = "000000"
    font_bold: bool = False
    bg_color: Optional[str] = None
    border: bool = True


class XlsxManager:
    """Excel 파일 관리 클래스 (TResultParser의 XlsxManager 포팅)"""
    
    FONTSIZE_DEFAULT = 10
    XLS_TITLE_COLCOUNT = 14
    
    def __init__(self):
        self.workbook: Optional[Workbook] = None
        self.worksheet = None
        self.filepath: Optional[Path] = None
        self._styles = self._init_styles()
    
    def _init_styles(self) -> Dict[XlsCellStyle, ExcelStyle]:
        """스타일 초기화"""
        return {
            XlsCellStyle.Title: ExcelStyle(
                font_color="FFFFFF",
                font_bold=True,
                bg_color="203764"
            ),
            XlsCellStyle.Caption: ExcelStyle(
                font_color="000000",
                font_bold=True,
                bg_color="DDEBF7"
            ),
            XlsCellStyle.General: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color=None
            ),
            XlsCellStyle.Fixed: ExcelStyle(
                font_color="000000",
                font_bold=True,
                bg_color="EEEEEE"
            ),
            XlsCellStyle.BgYellow: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="FFFF00"
            ),
            XlsCellStyle.BgOrange: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="FFA500"
            ),
            XlsCellStyle.BgRed: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="FF0000"
            ),
            XlsCellStyle.BgPink: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="FFC0CB"
            ),
            XlsCellStyle.BgSkyBlueL: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="C5D9F1"
            ),
            XlsCellStyle.BgPurpleL: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="E4DFEC"
            ),
            XlsCellStyle.BgPeachL: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="F2DCDB"
            ),
            XlsCellStyle.BgLightBlue: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="ADD8E6"
            ),
            XlsCellStyle.BgLightGray: ExcelStyle(
                font_color="000000",
                font_bold=False,
                bg_color="D3D3D3"
            ),
            XlsCellStyle.FgRed: ExcelStyle(
                font_color="FF0000",
                font_bold=False,
                bg_color=None
            ),
        }
    
    def create(self, filepath: Path) -> bool:
        """Excel 파일 생성"""
        if Workbook is None:
            raise ImportError("openpyxl is required")
        
        try:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Sheet1"
            self.filepath = Path(filepath)
            self._hide_gridlines()
            return True
        except Exception as e:
            print(f"Excel create error: {e}")
            return False
    
    def select_sheet(self, sheet_index: int, sheet_name: str, add_if_missing: bool = False) -> bool:
        """시트 선택"""
        if not self.workbook:
            return False
        
        try:
            if add_if_missing and sheet_index > len(self.workbook.worksheets):
                for i in range(len(self.workbook.worksheets), sheet_index):
                    self.workbook.create_sheet(f"Sheet{i + 1}")
            
            if sheet_index <= len(self.workbook.worksheets):
                self.worksheet = self.workbook.worksheets[sheet_index - 1]
                if sheet_name:
                    self.worksheet.title = sheet_name
                return True
        except Exception as e:
            print(f"Select sheet error: {e}")
            return False
    
    def write_data(self, row: int, col: int, data: Any) -> None:
        """셀에 데이터 쓰기"""
        if not self.worksheet or row < 1 or col < 1:
            return
        
        cell = self.worksheet.cell(row=row, column=col)
        if isinstance(data, bool):
            cell.value = data
        elif isinstance(data, (int, float)):
            cell.value = data
        else:
            # 2026-06-19 (deep-review C1 sibling) — vcast/qac 파싱 텍스트(C 소스/덤프/
            # 로그)에 Excel 불법 제어문자(\x0c form-feed/\x07/\x1a)가 섞이면 openpyxl이
            # cell.value 대입 시 IllegalCharacterError를 raise → 단일 셀이 generate 전체
            # 크래시. sanitize 후 재시도(\t\n\r 보존). qac_excel_generator도 본 sink 공유.
            text = str(data) if data is not None else ""
            try:
                cell.value = text
            except IllegalCharacterError:
                cell.value = ILLEGAL_CHARACTERS_RE.sub("", text)
    
    def apply_style(self, row_start: int, col_start: int, row_end: int, col_end: int, style: XlsCellStyle) -> None:
        """스타일 적용"""
        if not self.worksheet or style == XlsCellStyle.NONE:
            return
        
        style_info = self._styles.get(style)
        if not style_info:
            return
        
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = self.worksheet.cell(row=row, column=col)
                
                # 폰트 설정
                font = Font(
                    name="Arial",
                    size=self.FONTSIZE_DEFAULT,
                    bold=style_info.font_bold,
                    color=style_info.font_color
                )
                cell.font = font
                
                # 배경색 설정
                if style_info.bg_color:
                    fill = PatternFill(start_color=style_info.bg_color, end_color=style_info.bg_color, fill_type="solid")
                    cell.fill = fill
                
                # 정렬 설정
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # 테두리 설정
                if style_info.border:
                    thin_border = Border(
                        left=Side(style="thin", color="000000"),
                        right=Side(style="thin", color="000000"),
                        top=Side(style="thin", color="000000"),
                        bottom=Side(style="thin", color="000000")
                    )
                    cell.border = thin_border
    
    def merge(self, row_start: int, col_start: int, row_end: int, col_end: int) -> None:
        """셀 병합"""
        if not self.worksheet:
            return
        
        start_cell = f"{get_column_letter(col_start)}{row_start}"
        end_cell = f"{get_column_letter(col_end)}{row_end}"
        self.worksheet.merge_cells(f"{start_cell}:{end_cell}")
    
    def set_column_width(self, col: int, width: int) -> None:
        """열 너비 설정"""
        if not self.worksheet:
            return
        
        col_letter = get_column_letter(col)
        # C# ClosedXML uses width * 0.2; openpyxl uses character width
        # Match C# output: pixel * 0.2 + 0.71 (ClosedXML padding)
        self.worksheet.column_dimensions[col_letter].width = width * 0.2 + 0.71
    
    def set_row_height(self, row: int, height: int) -> None:
        """행 높이 설정"""
        if not self.worksheet:
            return
        
        self.worksheet.row_dimensions[row].height = height
    
    def set_wrap_text(self, row_start: int, col_start: int, row_end: int, col_end: int, wrap: bool = True) -> None:
        """텍스트 줄바꿈 설정"""
        if not self.worksheet:
            return
        
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=wrap, horizontal="center", vertical="center")
    
    def add_comment(self, row: int, col: int, message: str) -> None:
        """주석 추가"""
        if not self.worksheet or not message:
            return
        
        cell = self.worksheet.cell(row=row, column=col)
        cell.comment = Comment(message, "TResultParser")
    
    def draw_double_border(self, row_start: int, col_start: int, row_end: int, col_end: int, edge: BorderEdge, color: str = "000000") -> None:
        """이중 테두리 그리기"""
        if not self.worksheet:
            return
        
        double_side = Side(style="double", color=color)
        
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = self.worksheet.cell(row=row, column=col)
                old = cell.border or Border()
                kw = {"left": old.left, "right": old.right, "top": old.top, "bottom": old.bottom}
                if edge == BorderEdge.Top:
                    kw["top"] = double_side
                elif edge == BorderEdge.Bottom:
                    kw["bottom"] = double_side
                elif edge == BorderEdge.Left:
                    kw["left"] = double_side
                elif edge == BorderEdge.Right:
                    kw["right"] = double_side
                cell.border = Border(**kw)
    
    def draw_thick_border(self, row_start: int, col_start: int, row_end: int, col_end: int, edge: BorderEdge, color: str = "000000") -> None:
        """굵은 테두리 그리기"""
        if not self.worksheet:
            return

        thick_side = Side(style="thick", color=color)

        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = self.worksheet.cell(row=row, column=col)
                old = cell.border or Border()
                kw = {"left": old.left, "right": old.right, "top": old.top, "bottom": old.bottom}
                if edge == BorderEdge.Top:
                    kw["top"] = thick_side
                elif edge == BorderEdge.Bottom:
                    kw["bottom"] = thick_side
                elif edge == BorderEdge.Left:
                    kw["left"] = thick_side
                elif edge == BorderEdge.Right:
                    kw["right"] = thick_side
                cell.border = Border(**kw)
    
    def close(self, save: bool = True) -> bool:
        """Excel 파일 저장 및 닫기"""
        if not self.workbook or not self.filepath:
            return False
        
        try:
            if save:
                self.workbook.save(str(self.filepath))
            return True
        except Exception as e:
            print(f"Excel close error: {e}")
            return False
    
    def _hide_gridlines(self) -> None:
        """그리드 라인 숨기기"""
        if self.worksheet:
            self.worksheet.sheet_view.showGridLines = False


def generate_testcase_excel(tcbank: TCBank, output_path: Path, mode: str = "TestCase") -> bool:
    """테스트 케이스 Excel 리포트 생성
    
    Args:
        tcbank: 파싱된 테스트 케이스 데이터
        output_path: 출력 파일 경로
        mode: 리포트 모드 ("TestCase", "TestResult", "TestReport")
    
    Returns:
        성공 여부
    """
    if Workbook is None:
        raise ImportError("openpyxl is required")
    
    excel = XlsxManager()
    if not excel.create(output_path):
        return False
    
    col_offset = 2
    row_offset = 6
    
    # 제목 생성
    title = f"VectorCAST {mode} Report"
    if tcbank.component_name:
        title = f"{title} - {tcbank.component_name}"
    
    excel.write_data(1, 1, title)
    excel.apply_style(1, 1, 1, excel.XLS_TITLE_COLCOUNT, XlsCellStyle.Title)
    excel.merge(1, 1, 1, excel.XLS_TITLE_COLCOUNT)
    excel.set_row_height(1, 40)
    excel.set_column_width(1, 1)
    
    # 데이터 행 생성
    current_row = row_offset
    col_count = 0
    
    if mode == "TestCase":
        # TestCase 모드: TC Index, TC ID, Unit Name, TC Gen Method, Input, Expected Result, Related ID
        headers = ["TC Index", "TC ID", "Unit Name", "TC Gen Method"]
        col_count = len(headers) + len(tcbank.input_names) + len(tcbank.exp_result_names) + 1
        
        # 헤더 행
        col = col_offset
        for h in headers:
            excel.write_data(current_row, col, h)
            col += 1
        
        # Input 헤더
        if tcbank.input_names:
            excel.write_data(current_row, col, "Input")
            excel.merge(current_row, col, current_row, col + len(tcbank.input_names) - 1)
            col += len(tcbank.input_names)
        
        # Expected Result 헤더
        if tcbank.exp_result_names:
            excel.write_data(current_row, col, "Expected Result")
            excel.merge(current_row, col, current_row, col + len(tcbank.exp_result_names) - 1)
            col += len(tcbank.exp_result_names)
        
        # Related ID 헤더
        related_id_col = col
        excel.write_data(current_row, col, "Related ID")

        excel.apply_style(current_row, col_offset, current_row, col_offset + col_count - 1, XlsCellStyle.Caption)

        # Sub-header row (input/expected names)
        current_row += 1
        
        # 데이터 행
        tc_index = 0
        for tc_name, tc_items in sorted(tcbank.test_cases.items()):
            for tc_item in tc_items:
                tc_index += 1
                col = col_offset
                
                excel.write_data(current_row, col, tc_index)
                col += 1
                excel.write_data(current_row, col, tc_item.header.test_case_name)
                col += 1
                excel.write_data(current_row, col, tc_item.header.unit_name)
                col += 1
                excel.write_data(current_row, col, "")  # TC Gen Method
                col += 1
                
                # Input 데이터
                for input_name in tcbank.input_names:
                    value = tc_item.input_data.get(input_name, "")
                    excel.write_data(current_row, col, value)
                    col += 1
                
                # Expected Result 데이터
                for exp_name in tcbank.exp_result_names:
                    value = tc_item.expected_result.get(exp_name, "")
                    excel.write_data(current_row, col, value)
                    col += 1
                
                # Related ID
                excel.write_data(current_row, col, "")
                
                excel.set_wrap_text(current_row, col_offset, current_row, col_offset + col_count - 1, True)
                current_row += 1
    
    elif mode == "TestResult":
        # TestResult 모드: TC Index, TC ID, Actual Result, Pass/Fail, Memo
        headers = ["TC Index", "TC ID"]
        col_count = len(headers) + len(tcbank.act_result_names) + 3  # Actual Result + Pass/Fail + Memo
        
        # 헤더 행
        col = col_offset
        for h in headers:
            excel.write_data(current_row, col, h)
            col += 1
        
        # Actual Result 헤더
        if tcbank.act_result_names:
            excel.write_data(current_row, col, "Actual Result")
            excel.merge(current_row, col, current_row, col + len(tcbank.act_result_names) - 1)
            col += len(tcbank.act_result_names)
        
        # Pass/Fail 헤더
        excel.write_data(current_row, col, "Pass/Fail")
        excel.merge(current_row, col, current_row, col + 1)
        col += 2
        
        # Memo 헤더
        excel.write_data(current_row, col, "Memo")
        
        excel.apply_style(current_row, col_offset, current_row, col_offset + col_count - 1, XlsCellStyle.Caption)
        current_row += 1
        
        # 데이터 행
        tc_index = 0
        for tc_name, tr_items in sorted(tcbank.test_results.items()):
            for tr_item in tr_items:
                tc_index += 1
                col = col_offset
                
                excel.write_data(current_row, col, tc_index)
                col += 1
                excel.write_data(current_row, col, tr_item.header.test_case_name)
                col += 1
                
                # Actual Result 데이터
                for act_name in tcbank.act_result_names:
                    if act_name in tr_item.actual_result:
                        actual, expected = tr_item.actual_result[act_name]
                        excel.write_data(current_row, col, actual)
                    else:
                        excel.write_data(current_row, col, "")
                    col += 1
                
                # Pass/Fail
                pass_fail = "PASS" if tr_item.passed else "FAIL"
                excel.write_data(current_row, col, pass_fail)
                excel.merge(current_row, col, current_row, col + 1)
                col += 2
                
                # Memo
                excel.write_data(current_row, col, "")
                
                excel.set_wrap_text(current_row, col_offset, current_row, col_offset + col_count - 1, True)
                current_row += 1
    
    # 테이블 포맷 설정
    _set_table_format(excel, col_offset, row_offset, current_row - 1, col_count)

    # 세로 셀 병합 (MERGE-01/02) — C# AllowMerging 컬럼 포팅.
    #   레퍼런스(UCTestCaseList.cs:1147-1182)는 setTableFormat 테두리 후 병합을 발행하므로
    #   여기서도 _set_table_format 다음에 적용한다. 데이터 행만 대상(캡션/제목 제외).
    first_data_row = row_offset + 1
    last_data_row = current_row - 1
    if last_data_row > first_data_row:
        # generic 값-기준 세로 병합: TC ID(연속 동일 케이스), Unit Name(같은 unit 묶음).
        #   GenMethod/Pass-Fail 등은 generic 대상에서 제외(아래 별도 처리/가로병합 보존).
        _merge_equal_runs(excel, col_offset + 1, first_data_row, last_data_row)  # TC ID
        if mode == "TestCase":
            _merge_equal_runs(excel, col_offset + 2, first_data_row, last_data_row)  # Unit Name
            # TC Gen Method 경계 병합(현재 공란 → no-op이나 충실 포팅, 과병합 방지).
            _merge_genmethod_runs(excel, col_offset + 3, first_data_row, last_data_row, tm_col=None)

    # 열 너비 설정 (C# 기준: base=[1,60,100,150,20] + data=80 each + last=150)
    excel.set_column_width(1, 1)  # spacer col A
    base_widths = [60, 100, 150, 20]  # TC Index, TC ID, Unit, TC Gen Method
    for i, w in enumerate(base_widths):
        excel.set_column_width(col_offset + i, w)
    # Data columns (Input, Expected, Actual, etc.) = 80 each
    for col in range(col_offset + len(base_widths), col_offset + col_count - 1):
        excel.set_column_width(col, 80)
    # Last column (Related ID / Log) = 150
    excel.set_column_width(col_offset + col_count - 1, 150)

    return excel.close(True)


def _set_table_format(excel: XlsxManager, col_offset: int, row_offset: int, last_row: int, col_count: int) -> None:
    """테이블 포맷 설정"""
    row_start = row_offset - 1
    row_last = last_row
    col_last = col_offset + col_count - 1
    
    # 테두리 그리기
    excel.draw_thick_border(row_start, col_offset, row_start, col_last, BorderEdge.Top)
    excel.draw_double_border(row_start + 2, col_offset, row_start + 2, col_last, BorderEdge.Top)
    excel.draw_double_border(row_last, col_offset, row_last, col_last, BorderEdge.Top)
    
    excel.draw_thick_border(row_start, col_offset, row_last, col_offset, BorderEdge.Left)
    excel.draw_thick_border(row_start, col_last, row_last, col_last, BorderEdge.Right)
    excel.draw_thick_border(row_last, col_offset, row_last, col_last, BorderEdge.Bottom)


def _merge_equal_runs(
    excel: XlsxManager, col: int, first_row: int, last_row: int, *, skip_empty: bool = True
) -> int:
    """단일 컬럼에서 연속 동일값 run을 세로 병합 (MERGE-01).

    C# 레퍼런스의 ``FlexgidLib.getMergedCellsOnColumns`` + 적용 루프
    (``UCTestCaseList.cs:1147-1158``) 등가 포팅. C1FlexGrid는 ``AllowMerging`` 컬럼
    (TC_ID/UnitName 등)의 인접 동일값을 자동으로 하나의 세로 병합 셀로 묶는데,
    openpyxl에는 그 기능이 없어 데이터 값에서 직접 run을 찾아 ``merge``한다.

    - 빈값/None은 경계로 처리(merge 시작/확장 안 함) — C# GenMethod 처리(1165 ``IsNullOrEmpty``)와
      동일하게 보수적. 무관한 공란 블록이 한 셀로 뭉치는 것을 방지.
    - run 길이 1(단일 셀)은 merge하지 않음(openpyxl 단일셀 merge 무의미·위험).

    반환: 발행한 병합 개수.
    """
    ws = excel.worksheet
    if ws is None or last_row <= first_row or col < 1:
        return 0

    def _norm(v: Any) -> Optional[str]:
        if v is None:
            return None
        s = str(v)
        return s if s.strip() != "" else None

    merged = 0
    r = first_row
    while r <= last_row:
        v = _norm(ws.cell(row=r, column=col).value)
        if v is None and skip_empty:
            r += 1
            continue
        run_end = r
        while run_end + 1 <= last_row and _norm(ws.cell(row=run_end + 1, column=col).value) == v:
            run_end += 1
        if run_end > r:
            excel.merge(r, col, run_end, col)
            merged += 1
        r = run_end + 1
    return merged


def _merge_genmethod_runs(
    excel: XlsxManager, gm_col: int, first_row: int, last_row: int, tm_col: Optional[int] = None
) -> int:
    """GenMethod 컬럼 경계 인식 세로 병합 (MERGE-02).

    C# ``UCTestCaseList.cs:1160-1182`` 포팅. GenMethod(TC Gen Method)는 generic
    값-기준 병합(``_merge_equal_runs``)에서 **의도적으로 제외**하고, 별도로
    "(GenMethod, TestMethod) 쌍이 연속으로 같은 구간"만 병합한다.
    예) REQ-ABV 2행 + FI-ABV 2행 → AOR/ABV가 4행으로 과병합되지 않고 2+2로 분리.

    현재 ``generate_testcase_excel`` 레이아웃은 TC Gen Method가 공란이고 별도
    Test Method 컬럼이 없어(``tm_col=None``) 사실상 no-op이다. 그러나 추후 GenMethod가
    채워질 때 generic 병합의 과병합을 막기 위한 경계 처리를 미리 충실히 포팅해 둔다.
    빈 GenMethod는 run을 시작하지 않는다.

    반환: 발행한 병합 개수.
    """
    ws = excel.worksheet
    if ws is None or last_row <= first_row or gm_col < 1:
        return 0

    def _s(row: int, c: int) -> str:
        v = ws.cell(row=row, column=c).value
        return str(v) if v is not None else ""

    merged = 0
    r = first_row
    while r <= last_row:
        gm = _s(r, gm_col)
        if gm.strip() == "":
            r += 1
            continue
        tm = _s(r, tm_col) if tm_col is not None else None
        run_end = r
        while (
            run_end + 1 <= last_row
            and _s(run_end + 1, gm_col) == gm
            and (tm_col is None or _s(run_end + 1, tm_col) == tm)
        ):
            run_end += 1
        if run_end > r:
            excel.merge(r, gm_col, run_end, gm_col)
            merged += 1
        r = run_end + 1
    return merged


def generate_metrics_excel(metrics_bank: MetricsBank, output_path: Path, unit_bank: Optional[Dict[str, str]] = None) -> bool:
    """Metrics Excel 리포트 생성
    
    Args:
        metrics_bank: 파싱된 Metrics 데이터
        output_path: 출력 파일 경로
        unit_bank: Unit ID 매핑 (선택사항)
    
    Returns:
        성공 여부
    """
    if Workbook is None:
        raise ImportError("openpyxl is required")
    
    excel = XlsxManager()
    if not excel.create(output_path):
        return False
    
    col_offset = 1
    row_offset = 3
    
    # UT Matrics 시트 (Statement)
    if metrics_bank.statement_data:
        excel.select_sheet(1, "UT Matrics", True)
        _generate_ut_metrics_sheet(excel, metrics_bank.statement_data, unit_bank, col_offset, row_offset)
    
    # IT Matrics 시트 (Functions)
    if metrics_bank.functions_data:
        sheet_num = 2 if metrics_bank.statement_data else 1
        excel.select_sheet(sheet_num, "IT Matrics", True)
        _generate_it_metrics_sheet(excel, metrics_bank.functions_data, unit_bank, col_offset, row_offset)
    
    return excel.close(True)


def _generate_ut_metrics_sheet(
    excel: XlsxManager,
    statement_data: Dict[str, MatixDataBank],
    unit_bank: Optional[Dict[str, str]],
    col_offset: int,
    row_offset: int
) -> None:
    """UT Matrics 시트 생성"""
    # 헤더
    headers = [
        "No", "TestID", "UnitID", "SubProgram", "Complexity",
        "Stat(Cnt)", "Stat(TTL)", "Stat(%)",
        "Branch(Cnt)", "Branch(TTL)", "Branch(%)",
        "ITS Called", "FCalls(Cnt)", "FCalls(TTL)", "FCalls(%)"
    ]
    
    current_row = row_offset
    col_count = len(headers)
    
    # 헤더 행
    for col, header in enumerate(headers, start=col_offset):
        excel.write_data(current_row, col, header)
    
    excel.apply_style(current_row, col_offset, current_row, col_offset + col_count - 1, XlsCellStyle.Caption)
    current_row += 1
    
    # 데이터 행
    row_num = 1
    statement_total = MatricStatementItem()
    statement_total.statements = CoverageItem("")
    statement_total.branches = CoverageItem("")
    statement_total.functions_call = CoverageItem("")
    
    for unit_name, bank in sorted(statement_data.items()):
        count = 0
        root_row = current_row
        
        for subprogram, item in sorted(bank.dic_data.items()):
            if not isinstance(item, MatricStatementItem):
                continue
            
            is_root = count == 0
            excel.write_data(current_row, col_offset, row_num)
            
            # TestID
            excel.write_data(current_row, col_offset + 1, item.id if is_root else "")
            
            # UnitID
            unit_id = ""
            if unit_bank and item.subprogram:
                func_name_lower = item.subprogram.lower()
                for uid, fname in unit_bank.items():
                    if fname.lower() == func_name_lower:
                        unit_id = uid
                        break
            excel.write_data(current_row, col_offset + 2, unit_id)
            
            # SubProgram
            excel.write_data(current_row, col_offset + 3, item.subprogram)
            
            # Complexity
            excel.write_data(current_row, col_offset + 4, item.complexity)
            
            # Statements
            if item.statements:
                excel.write_data(current_row, col_offset + 5, item.statements.count)
                excel.write_data(current_row, col_offset + 6, item.statements.total)
                excel.write_data(current_row, col_offset + 7, item.statements.percentage)
                
                statement_total.statements.count += item.statements.count
                statement_total.statements.total += item.statements.total
            
            # Branches
            if item.branches:
                excel.write_data(current_row, col_offset + 8, item.branches.count)
                excel.write_data(current_row, col_offset + 9, item.branches.total)
                excel.write_data(current_row, col_offset + 10, item.branches.percentage)
                
                statement_total.branches.count += item.branches.count
                statement_total.branches.total += item.branches.total
            
            # ITS Called
            excel.write_data(current_row, col_offset + 11, "O" if item.is_function else "X")
            
            # Function Calls
            if item.functions_call:
                excel.write_data(current_row, col_offset + 12, item.functions_call.count)
                excel.write_data(current_row, col_offset + 13, item.functions_call.total)
                excel.write_data(current_row, col_offset + 14, item.functions_call.percentage)
                
                statement_total.functions_call.count += item.functions_call.count
                statement_total.functions_call.total += item.functions_call.total
            
            excel.set_wrap_text(current_row, col_offset, current_row, col_offset + col_count - 1, True)
            current_row += 1
            row_num += 1
            count += 1
    
    # Total 행
    excel.write_data(current_row, col_offset + 1, "Total")
    if statement_total.statements:
        excel.write_data(current_row, col_offset + 5, statement_total.statements.count)
        excel.write_data(current_row, col_offset + 6, statement_total.statements.total)
        excel.write_data(current_row, col_offset + 7, statement_total.statements.percentage)
    if statement_total.branches:
        excel.write_data(current_row, col_offset + 8, statement_total.branches.count)
        excel.write_data(current_row, col_offset + 9, statement_total.branches.total)
        excel.write_data(current_row, col_offset + 10, statement_total.branches.percentage)
    if statement_total.functions_call:
        excel.write_data(current_row, col_offset + 12, statement_total.functions_call.count)
        excel.write_data(current_row, col_offset + 13, statement_total.functions_call.total)
        excel.write_data(current_row, col_offset + 14, statement_total.functions_call.percentage)
    
    # 제목 행
    excel.write_data(1, 1, "UT Matrics")
    excel.apply_style(1, 1, 1, col_count, XlsCellStyle.Title)
    excel.merge(1, 1, 1, col_count)
    
    # 열 너비 설정 (C# 기준: [40,100,100,250,80, 80,80,80, 80,80,80, 80, 80,80,80])
    ut_widths = [40, 100, 100, 250, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80]
    for i, w in enumerate(ut_widths):
        excel.set_column_width(col_offset + i, w)


def _generate_it_metrics_sheet(
    excel: XlsxManager,
    functions_data: Dict[str, MatixDataBank],
    unit_bank: Optional[Dict[str, str]],
    col_offset: int,
    row_offset: int
) -> None:
    """IT Matrics 시트 생성"""
    # 헤더
    headers = ["No", "Unit", "UnitID", "SubProgram", "Complexity", "Functions", "Function Calls"]
    
    current_row = row_offset
    col_count = len(headers)
    
    # 헤더 행
    for col, header in enumerate(headers, start=col_offset):
        excel.write_data(current_row, col, header)
    
    excel.apply_style(current_row, col_offset, current_row, col_offset + col_count - 1, XlsCellStyle.Caption)
    current_row += 1
    
    # 데이터 행
    row_num = 1
    funcall_total = MatricFunCallItem()
    funcall_total.functions = CoverageItem("")
    funcall_total.functions_call = CoverageItem("")
    
    for unit_name, bank in sorted(functions_data.items()):
        for subprogram, item in sorted(bank.dic_data.items()):
            if not isinstance(item, MatricFunCallItem):
                continue
            
            excel.write_data(current_row, col_offset, row_num)
            
            # Unit
            excel.write_data(current_row, col_offset + 1, item.unit_name)
            
            # UnitID
            unit_id = ""
            if unit_bank and item.subprogram:
                func_name_lower = item.subprogram.lower()
                for uid, fname in unit_bank.items():
                    if fname.lower() == func_name_lower:
                        unit_id = uid
                        break
            excel.write_data(current_row, col_offset + 2, unit_id)
            
            # SubProgram
            excel.write_data(current_row, col_offset + 3, item.subprogram)
            
            # Complexity
            excel.write_data(current_row, col_offset + 4, item.complexity)
            
            # Functions
            if item.functions:
                excel.write_data(current_row, col_offset + 5, item.functions.coverage)
                funcall_total.functions.count += item.functions.count
                funcall_total.functions.total += item.functions.total
            
            # Function Calls
            if item.functions_call:
                excel.write_data(current_row, col_offset + 6, item.functions_call.coverage)
                funcall_total.functions_call.count += item.functions_call.count
                funcall_total.functions_call.total += item.functions_call.total
            
            excel.set_wrap_text(current_row, col_offset, current_row, col_offset + col_count - 1, True)
            current_row += 1
            row_num += 1
    
    # Total 행
    excel.write_data(current_row, col_offset + 1, "Total")
    if funcall_total.functions:
        excel.write_data(current_row, col_offset + 5, funcall_total.functions.coverage)
    if funcall_total.functions_call:
        excel.write_data(current_row, col_offset + 6, funcall_total.functions_call.coverage)
    
    # 제목 행
    excel.write_data(1, 1, "IT Matrics")
    excel.apply_style(1, 1, 1, col_count, XlsCellStyle.Title)
    excel.merge(1, 1, 1, col_count)

    # 열 너비 설정 (C# 기준: [40,250,100,300,80,140,140])
    it_widths = [40, 250, 100, 300, 80, 140, 140]
    for i, w in enumerate(it_widths):
        if col_offset + i <= col_offset + col_count - 1:
            excel.set_column_width(col_offset + i, w)


def add_coverage_charts(wb, metrics_bank: MetricsBank) -> None:
    """Coverage Summary 시트에 커버리지 차트를 생성합니다.

    Args:
        wb: openpyxl Workbook 인스턴스
        metrics_bank: 파싱된 Metrics 데이터
    """
    if BarChart is None:
        return

    statement_data = metrics_bank.statement_data
    if not statement_data:
        return

    ws = wb.create_sheet("Coverage Summary")
    ws.sheet_view.showGridLines = False

    # ------------------------------------------------------------------
    # 1. 타이틀
    # ------------------------------------------------------------------
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Coverage Summary"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ------------------------------------------------------------------
    # 2. 데이터 테이블: 함수별 Statement% / Branch% 커버리지
    # ------------------------------------------------------------------
    data_start_row = 4
    ws.cell(row=data_start_row, column=1, value="Function")
    ws.cell(row=data_start_row, column=2, value="Statement %")
    ws.cell(row=data_start_row, column=3, value="Branch %")
    ws.cell(row=data_start_row, column=4, value="Stmt Covered")
    ws.cell(row=data_start_row, column=5, value="Stmt Uncovered")
    for c in range(1, 6):
        ws.cell(row=data_start_row, column=c).font = Font(bold=True)

    # 요약 카운터
    total_stmt_count = 0
    total_stmt_total = 0
    total_branch_count = 0
    total_branch_total = 0

    data_row = data_start_row + 1
    for unit_name, bank in sorted(statement_data.items()):
        for subprogram, item in sorted(bank.dic_data.items()):
            if not isinstance(item, MatricStatementItem):
                continue

            ws.cell(row=data_row, column=1, value=item.subprogram or subprogram)

            # Statement coverage percentage
            stmt_pct = 0.0
            stmt_covered = 0
            stmt_uncovered = 0
            if item.statements:
                try:
                    stmt_pct = float(item.statements.percentage)
                except (ValueError, TypeError):
                    stmt_pct = 0.0
                stmt_covered = item.statements.count
                stmt_uncovered = item.statements.total - item.statements.count
                total_stmt_count += item.statements.count
                total_stmt_total += item.statements.total

            # Branch coverage percentage
            branch_pct = 0.0
            if item.branches:
                try:
                    branch_pct = float(item.branches.percentage)
                except (ValueError, TypeError):
                    branch_pct = 0.0
                total_branch_count += item.branches.count
                total_branch_total += item.branches.total

            ws.cell(row=data_row, column=2, value=stmt_pct)
            ws.cell(row=data_row, column=3, value=branch_pct)
            ws.cell(row=data_row, column=4, value=stmt_covered)
            ws.cell(row=data_row, column=5, value=stmt_uncovered)
            data_row += 1

    data_end_row = data_row - 1

    if data_end_row < data_start_row + 1:
        # 데이터가 없으면 차트를 생성하지 않음
        return

    # 요약 통계
    ws.cell(row=3, column=7, value="Total Statements").font = Font(bold=True)
    ws.cell(row=3, column=8, value=total_stmt_total)
    ws.cell(row=3, column=9, value="Covered").font = Font(bold=True)
    ws.cell(row=3, column=10, value=total_stmt_count)

    # ------------------------------------------------------------------
    # 3. Bar Chart: Statement% / Branch% per function
    # ------------------------------------------------------------------
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Statement / Branch Coverage (%)"
    bar_chart.y_axis.title = "Coverage %"
    bar_chart.x_axis.title = "Function"
    bar_chart.y_axis.scaling.max = 100
    bar_chart.width = 30
    bar_chart.height = 15

    cats = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
    stmt_vals = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
    branch_vals = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_end_row)

    bar_chart.add_data(stmt_vals, titles_from_data=True)
    bar_chart.add_data(branch_vals, titles_from_data=True)
    bar_chart.set_categories(cats)

    # 계열 색상 설정
    bar_chart.series[0].graphicalProperties.solidFill = "4472C4"  # 파랑: Statement
    bar_chart.series[1].graphicalProperties.solidFill = "ED7D31"  # 주황: Branch

    ws.add_chart(bar_chart, "A" + str(data_end_row + 3))

    # ------------------------------------------------------------------
    # 4. Stacked Bar Chart: Covered vs Uncovered statements per function
    # ------------------------------------------------------------------
    stacked_chart = BarChart()
    stacked_chart.type = "col"
    stacked_chart.grouping = "stacked"
    stacked_chart.style = 10
    stacked_chart.title = "Covered vs Uncovered Statements"
    stacked_chart.y_axis.title = "Statements"
    stacked_chart.x_axis.title = "Function"
    stacked_chart.width = 30
    stacked_chart.height = 15

    cats2 = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
    covered_vals = Reference(ws, min_col=4, min_row=data_start_row, max_row=data_end_row)
    uncovered_vals = Reference(ws, min_col=5, min_row=data_start_row, max_row=data_end_row)

    stacked_chart.add_data(covered_vals, titles_from_data=True)
    stacked_chart.add_data(uncovered_vals, titles_from_data=True)
    stacked_chart.set_categories(cats2)

    # 계열 색상: 초록(covered), 빨강(uncovered)
    stacked_chart.series[0].graphicalProperties.solidFill = "00B050"
    stacked_chart.series[1].graphicalProperties.solidFill = "FF4444"

    ws.add_chart(stacked_chart, "A" + str(data_end_row + 20))

    # 열 너비 조정
    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 16
