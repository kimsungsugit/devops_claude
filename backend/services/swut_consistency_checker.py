"""SwUT Coverage Report ↔ Software Unit Test Result(SUTR) 일관성 검증.

두 산출물은 같은 단위테스트 캠페인에서 별도 시트/파일로 생성되므로
다음 4가지 항목의 cross-validation이 필요하다.

1. **미커버 Function ↔ 미실행 TC**:
   Coverage Report `1.Traceability` 시트에서 한 번도 'O' 표시되지 않은 함수와
   SUTR `Test Summary`의 'List of Test Case not Executed' 항목이 일대일 매칭.

2. **Exception 카운트 ↔ Deviation 카운트**:
   Coverage Report `3.Coverage`의 Statement+Branch Exception 합 ≥
   SUTR `Test Summary`의 Deviation List 행 수.
   (Exception은 분기/구문 단위, Deviation은 TC 단위이므로 등호가 아닌 ≥)

3. **Total TC 일관성**:
   Coverage Report Traceability TC 행 수 == SUTR Total Number of TCs.

4. **Final Result 용어 통일**:
   Coverage Report 'Final Test Result'와 SUTR 'Final Test Result'가
   동일 의미(PASS/OK) 표기인지.

ISO 26262 ASIL A 이상 단위테스트 산출물 심사 시 두 문서 간 불일치가
잦은 지적 사항이라 자동 검증이 필요하다.

## ISO 26262 Tool Qualification 사용 제약

- **ASIL A**: 본 도구의 issue 결과를 reviewer가 검토 후 그대로 활용 가능.
- **ASIL B/C/D**: 본 도구 결과는 reviewer가 놓친 불일치 후보 발견용으로만
  사용. 모든 issue는 사람이 검증 후 evidence 확정 의무.

## 알려진 한계

- Hyundai/Mobis 스타일 시트명/헤더 가정 (`1.Traceability`, `3. Coverage`,
  `Test Summary`, `Final Test Result`, `Total Number of TCs`).
- 다른 포맷(LG/Bosch/Continental 등)에서 헤더 미발견 시
  `ConsistencyReport.parse_warnings` 에 사유 명시.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover - hook fail-safe
    openpyxl = None  # type: ignore[assignment]

from backend.services.excel_template_utils import sheet_is_blank_placeholder

# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ConsistencyIssue:
    severity: str  # "critical" | "warning" | "info"
    category: str  # "uncovered_mismatch" | "exception_deviation" | "total_tc" | "final_result"
    message: str
    expected: Any = None
    actual: Any = None


@dataclass
class ConsistencyReport:
    ok: bool
    issues: list[ConsistencyIssue] = field(default_factory=list)
    coverage_summary: dict[str, Any] = field(default_factory=dict)
    sutr_summary: dict[str, Any] = field(default_factory=dict)
    parse_warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "issues": [
                {
                    "severity": i.severity,
                    "category": i.category,
                    "message": i.message,
                    "expected": i.expected,
                    "actual": i.actual,
                }
                for i in self.issues
            ],
            "coverage_summary": self.coverage_summary,
            "sutr_summary": self.sutr_summary,
            "parse_warnings": self.parse_warnings,
            # ISO 26262 Tool Qualification 메타데이터
            "tool_qualification": {
                "evidence_class": "auto-generated draft",
                "asil_a_usage": "issue 결과를 reviewer 검토 후 활용",
                "asil_b_c_d_usage": "단독 evidence 사용 금지 — manual review 의무",
                "format_assumption": "Hyundai/Mobis 스타일 시트명/헤더",
            },
        }


# ---------------------------------------------------------------------------
# Extractors
# ---------------------------------------------------------------------------

_RE_SWUFN = re.compile(r"^SwUFn_\d+$")
# 라운드 96-fix W-B — Traceability 매트릭스 헤더 ID 일반화. HDPDM01은 TC×SwUFn
# 매트릭스지만 KJPDS02 SwIT v1.01은 TC×SDS 설계 항목 매트릭스로 ID 계열이 혼합
# (실측 r11: SwST_01~09 + SwSTR_* + SwFn_* + SwTK_* = 108 items) — SwUFn만
# 인정하면 헤더 탐지 실패로 total_tcs=0 (cross-validation 무력).
# deep-reviewer 96-fix W#1/W#2: 광역 `Sw[A-Za-z]{1,6}` 대신 실측 계열 whitelist —
# TC ID alt-prefix(SwIT_0201/SwUT_0101) 및 SwReq_/SwArch_ 과잉 매칭 차단
# (ISO 26262 cross-validation 도구 결정성). 신규 계열 발견 시 여기에 명시 추가.
# KJPDS02 v1.01 매트릭스 실측 108 = SwST 9 + SwSTR 22 + SwCom 33 + SwFn 41 + SwTK 3.
_RE_ITEM_ID = re.compile(r"^(?:SwUFn|SwSTR|SwST|SwCom|SwFn|SwTK)_\d+$")
# 35차: _RE_SWUTC를 module-level 고정에서 함수-local 동적 compile로 변경. tc_prefix
# kwarg (SwUT="SwUTC" default, SwIT="SwITC")로 SwUT/SwIT 양쪽 호환.


def _tc_id_patterns(tc_prefix: str) -> tuple[re.Pattern, re.Pattern]:
    """라운드 96-fix W-B — TC ID 명명 변형 일반화 (KJPDS02 실측 호환).

    지원 명명:
      - HDPDM01: ``{P}_SwUFn_12`` (P=SwUTC/SwITC — 기존 유일 지원형)
      - KJPDS02 DV spec/Test Log: ``{P}_0201`` / ``{P}_0101_01``
      - KJPDS02 PV VectorCAST: ``SwIT_SwUFn_0101_01`` (P의 trailing 'C' 탈락형)
        + compound iteration suffix ``.001``

    Returns:
        (match_re, fn_re):
            match_re — TC ID 전체 매칭 판정 ($ 앵커).
            fn_re — 함수 ID 추출 (group 1=SwUFn_id 또는 group 2=숫자 → SwUFn_{n}).
    """
    alt = tc_prefix[:-1] if tc_prefix.endswith("C") else tc_prefix
    base = rf"(?:{re.escape(tc_prefix)}|{re.escape(alt)})"
    match_re = re.compile(rf"^{base}_(?:SwUFn_)?\d+(?:_\d+)*(?:\.\d+)?$")
    fn_re = re.compile(rf"^{base}_(?:(SwUFn_\d+)|(\d{{2,}}))")
    return match_re, fn_re


def _extract_function_id(fn_re: re.Pattern, tc_id: str) -> str | None:
    """fn_re 매칭 결과에서 SwUFn 함수 ID 정규화 (KJPDS02 숫자형은 SwUFn_ 부여)."""
    m = fn_re.match(tc_id)
    if not m:
        return None
    return m.group(1) or f"SwUFn_{m.group(2)}"


def _compact_row(row: tuple) -> list[tuple[int, str]]:
    """머지셀(None leading)을 걸러서 (column_index, value) 페어만 반환."""
    return [
        (i, str(c).strip()) for i, c in enumerate(row)
        if c is not None and str(c).strip()
    ]


def _find_value_after_label(pairs: list[tuple[int, str]], label: str) -> str | None:
    """compact_row 결과에서 label에 매칭되는 셀의 다음 셀 값 반환."""
    for i, (_, v) in enumerate(pairs):
        if v == label and i + 1 < len(pairs):
            return pairs[i + 1][1]
    return None


def _row_has_label(pairs: list[tuple[int, str]], label: str) -> bool:
    return any(v == label for _, v in pairs)


def _row_contains(pairs: list[tuple[int, str]], substr: str) -> bool:
    return any(substr in v for _, v in pairs)


def _extract_coverage_summary(
    wb: Any,
    out_warnings: list[str] | None = None,
    *,
    tc_prefix: str = "SwUTC",
) -> dict[str, Any]:
    """Coverage Report 워크북에서 핵심 지표 추출.

    시나리오 A 방어: 우리 빌더가 작성한 placeholder 시트(`BLANK_MARKUP`)면 데이터
    추출 skip + out_warnings에 등록 — self-validation false positive 차단.
    """
    summary: dict[str, Any] = {
        "total_tcs": 0,
        "total_functions": 0,
        "uncovered_functions": [],
        "exception_statement": 0,
        "exception_branch": 0,
        "final_result": "",
        "trace_sheet_is_placeholder": False,
    }

    # 1.Traceability 시트: TC × Function 매트릭스
    trace_sheet = None
    for name in wb.sheetnames:
        if "traceability" in name.lower():
            trace_sheet = wb[name]
            break

    # placeholder 감지 (시나리오 A) — BLANK_MARKUP 있으면 데이터 추출 skip
    if trace_sheet is not None and sheet_is_blank_placeholder(trace_sheet):
        summary["trace_sheet_is_placeholder"] = True
        if out_warnings is not None:
            out_warnings.append(
                "1.Traceability 시트가 자동 생성 placeholder — TC×Function 매트릭스 미작성, "
                "데이터 추출 skip"
            )
        trace_sheet = None

    if trace_sheet is not None:
        rows = list(trace_sheet.iter_rows(values_only=True))
        # Q1 fix: header 탐지를 임계치 > 50 → "최다 SwUFn 행"으로 변경.
        # 소규모 ASIL A 보조 모듈(함수 51개 미만)에서도 동작하도록 함.
        header_idx = None
        best_count = 0
        for i, r in enumerate(rows[:20]):
            # 라운드 96-fix W-B: SwUFn(HDPDM01) 외 SwST(KJPDS02 SwIT) 헤더 허용
            sw_count = sum(1 for c in r if isinstance(c, str) and _RE_ITEM_ID.match(c))
            if sw_count > best_count:
                best_count = sw_count
                header_idx = i
        # 적어도 매트릭스 ID 1개라도 있어야 함 — 0이면 Traceability 시트로 인정 안 함.
        if best_count == 0:
            header_idx = None

        if header_idx is not None:
            header = rows[header_idx]
            func_cols: dict[int, str] = {
                i: c for i, c in enumerate(header)
                if isinstance(c, str) and _RE_ITEM_ID.match(c)
            }
            o_count_per_func: dict[str, int] = dict.fromkeys(func_cols.values(), 0)
            tc_count = 0

            # 35차: tc_prefix 동적 적용 + 라운드 96-fix W-B 명명 변형 일반화
            _tc_re, _ = _tc_id_patterns(tc_prefix)
            for r in rows[header_idx + 1:]:
                tc_id = next(
                    (c for c in r[:5] if isinstance(c, str) and _tc_re.match(c)),
                    None,
                )
                if not tc_id:
                    continue
                tc_count += 1
                for col_idx, fn_id in func_cols.items():
                    if col_idx < len(r):
                        v = r[col_idx]
                        if isinstance(v, str) and v.strip().upper() == "O":
                            o_count_per_func[fn_id] += 1

            summary["total_tcs"] = tc_count
            summary["total_functions"] = len(func_cols)
            summary["uncovered_functions"] = sorted(
                fn for fn, cnt in o_count_per_func.items() if cnt == 0
            )

    # 3. Coverage 시트: Exception 카운트
    cov_sheet = None
    for name in wb.sheetnames:
        lname = name.lower()
        # '3. Coverage' / '3.Coverage' / 'Coverage' 매칭. '1.Traceability', '2.Consistency'는 제외.
        if "coverage" in lname and "traceability" not in lname and "consistency" not in lname:
            cov_sheet = wb[name]
            break

    if cov_sheet is not None:
        # 'Coverage' 헤더 행(Total / Fail / Exception / Coverage)을 찾고 Statement/Branch 행에서 Exception(4번째 값) 추출.
        for r in cov_sheet.iter_rows(values_only=True, max_row=10):
            pairs = _compact_row(r)
            if not pairs:
                continue
            label = pairs[0][1]
            # Statement / Branch 행: [label, Total, Fail Count, Exception, Coverage]
            if label == "Statement" and len(pairs) >= 4:
                try:
                    summary["exception_statement"] = int(pairs[3][1])
                except (ValueError, TypeError):
                    pass
            elif label == "Branch" and len(pairs) >= 4:
                try:
                    summary["exception_branch"] = int(pairs[3][1])
                except (ValueError, TypeError):
                    pass

    # Test Summary 시트: Final Test Result
    for name in wb.sheetnames:
        lname = name.lower()
        if lname == "test summary" or (lname.endswith("summary") and "test" in lname):
            ws = wb[name]
            for r in ws.iter_rows(values_only=True):
                pairs = _compact_row(r)
                v = _find_value_after_label(pairs, "Final Test Result")
                if v:
                    summary["final_result"] = v
                    break
            if summary["final_result"]:
                break

    return summary


# Test Summary 의 TC 통계 라벨 → summary 키. 회사 v2.02 양식은 'Deviated' 열이
# **없어서 5열**이고, 구 양식/SITR 은 6열이다. 위치로 읽으면 5열 문서에서
# not_executed 값이 deviated 자리로 한 칸 밀린다 → 라벨로 열을 찾는다.
_TC_STAT_LABEL_KEYS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("total_tcs", ("total number of tcs",)),
    ("tested", ("number of tcs tested",)),
    ("passed", ("number of tcs passed",)),
    ("failed", ("number of tcs failed",)),
    ("deviated", ("number of deviated tcs", "number of tcs deviated")),
    ("not_executed", ("number of tcs not executed", "number of tcs notexecuted")),
)


def _tc_stat_key(label: str) -> str | None:
    norm = " ".join((label or "").strip().lower().split())
    for key, variants in _TC_STAT_LABEL_KEYS:
        if norm in variants:
            return key
    return None


def _fill_tc_stats(
    summary: dict[str, Any],
    header_pairs: list[tuple[int, str]],
    value_pairs: list[tuple[int, str]],
    out_warnings: list[str] | None,
) -> None:
    """헤더 라벨의 **열 인덱스**로 값을 집는다.

    ⚠ 이전 판은 ``zip(keys, nxt_pairs[:6])`` 위치 매핑이었다. 이 저장소가 직접
      찍어내는 회사 v2.02 양식(``swut_coverage_aggregator`` 가 stamp 하는 라벨은
      **5개** — Deviated 없음)에서는 미실행 수가 ``deviated`` 에 들어가고
      ``not_executed`` 는 대입조차 되지 않아 0 으로 남았다. 경고도 없었다.
      그 결과 정합한 산출물 쌍에 "Total TC 불일치" Warning 이 뜨고, 프론트는
      미실행 N 건을 Deviation N 건으로 표시했다.

    ``_compact_row`` 가 빈 셀을 버리므로 **위치는 어긋나도 열 인덱스는 보존**된다 —
    그래서 열로 맞추는 것이 위치보다 강하다.
    """
    by_col = dict(value_pairs)
    resolved: dict[str, int] = {}
    for col, label in header_pairs:
        key = _tc_stat_key(label)
        if key is None or key in resolved:
            continue
        raw = by_col.get(col)
        if raw is None:
            continue
        try:
            resolved[key] = int(raw) if raw else 0
        except (ValueError, TypeError):
            continue
    if resolved:
        summary.update(resolved)
        return

    # 라벨을 하나도 못 읽었다 — 구형 양식일 수 있다. 열 수가 정확히 맞을 때만
    # 위치 폴백을 허용하고, 아니면 **채우지 않는다**(0 으로 위장하지 않는다).
    keys = tuple(k for k, _ in _TC_STAT_LABEL_KEYS)
    if len(value_pairs) == len(keys):
        for k, (_, val) in zip(keys, value_pairs):
            try:
                summary[k] = int(val) if val else 0
            except (ValueError, TypeError):
                pass
        return
    if out_warnings is not None:
        out_warnings.append(
            f"Test Summary TC 통계: 라벨을 못 읽었고 열 수도 {len(value_pairs)}"
            f"(기대 {len(keys)}) 라 매핑 불가 — deviated/not_executed 를 채우지 않았다"
        )


def _extract_sutr_summary(
    wb: Any, *, tc_prefix: str = "SwUTC", out_warnings: list[str] | None = None,
) -> dict[str, Any]:
    """SUTR 워크북에서 핵심 지표 추출."""
    summary: dict[str, Any] = {
        "total_tcs": 0,
        "tested": 0,
        "passed": 0,
        "failed": 0,
        "deviated": 0,
        "not_executed": 0,
        "not_executed_tcs": [],
        "deviation_tcs": [],
        "final_result": "",
    }

    # Test Summary 시트
    summary_sheet = None
    for name in wb.sheetnames:
        if "test summary" == name.lower():
            summary_sheet = wb[name]
            break
        if "summary" in name.lower() and summary_sheet is None:
            summary_sheet = wb[name]

    if summary_sheet is not None:
        rows = list(summary_sheet.iter_rows(values_only=True))
        section = None
        # 35차 reviewer W1: 루프 진입 전 한 번만 compile — _extract_coverage_summary
        # 패턴과 일관성. 수천 행 SITR에서 매 반복 compile 회피 (개선).
        # 라운드 96-fix W-B: KJPDS02 trailing 'C' 탈락형(SwIT_/SwUT_)도 허용
        _alt_prefix = tc_prefix[:-1] if tc_prefix.endswith("C") else tc_prefix
        _tc_prefixes = (f"{tc_prefix}_", f"{_alt_prefix}_")
        _, _deviation_re = _tc_id_patterns(tc_prefix)
        for idx, r in enumerate(rows):
            pairs = _compact_row(r)
            if not pairs:
                continue
            first = pairs[0][1]

            # Section transitions (이 행은 헤더로만 쓰고 next row부터 컨텐츠)
            if _row_contains(pairs, "List of Test Case not Executed"):
                section = "not_executed"
                continue
            if _row_contains(pairs, "Deviation List"):
                section = "deviation"
                continue
            if _row_contains(pairs, "Test Defects List") or _row_contains(pairs, "Test Defect List"):
                section = "defects"
                continue
            # Test Case ID 헤더 행은 컨텐츠 시작 직전 → skip
            if first == "Test Case ID":
                continue

            # Key-value: Final Test Result
            v = _find_value_after_label(pairs, "Final Test Result")
            if v:
                summary["final_result"] = v

            # Total Number of TCs 헤더 행 발견 시 다음 행에서 값 추출
            if _row_has_label(pairs, "Total Number of TCs"):
                if idx + 1 < len(rows):
                    _fill_tc_stats(summary, pairs, _compact_row(rows[idx + 1]), out_warnings)

            # Section content: TC ID로 시작하는 행
            # 35차: SwUT는 "SwUTC_" prefix, SwIT는 "SwITC_" — 위 루프 진입 전 한 번 compile.
            if section == "not_executed" and first.startswith(_tc_prefixes):
                summary["not_executed_tcs"].append(first)
            elif section == "deviation" and first.startswith(_tc_prefixes):
                if _deviation_re.match(first):
                    summary["deviation_tcs"].append(first)

        summary["not_executed_tcs"] = list(dict.fromkeys(summary["not_executed_tcs"]))
        summary["deviation_tcs"] = list(dict.fromkeys(summary["deviation_tcs"]))

    return summary


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

_RESULT_EQUIV = {"pass": "ok", "passed": "ok", "ok": "ok",
                 "fail": "fail", "failed": "fail", "ng": "fail"}


def _normalize_result(s: str) -> str:
    return _RESULT_EQUIV.get(s.strip().lower(), s.strip().lower())


def _load_workbook(source: Any) -> Any:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for SwUT consistency check")
    if isinstance(source, (bytes, bytearray)):
        return openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    if isinstance(source, str):
        return openpyxl.load_workbook(source, read_only=True, data_only=True)
    # assume already a Workbook
    return source


def summarize_coverage_report(
    coverage_source: Any, *, tc_prefix: str = "SwUTC",
) -> dict[str, Any]:
    """단일 Coverage Report(.xlsx)만 파싱해 coverage_summary를 반환(정합성 비교 없이).

    check_swut_consistency가 내부적으로 쓰는 `_extract_coverage_summary`를 단독 호출 —
    빌더 산출물 하나만 있어도 미커버 함수·Exception·Total TC·Final Result를 바로 본다.
    SwIT는 thin wrapper(swit_consistency_checker)에서 tc_prefix="SwITC"로 호출.

    Returns: {coverage_summary: dict, parse_warnings: list[str]}. 파싱 실패는
    parse_warnings에 누적되며 coverage_summary는 부분/빈 dict일 수 있다(예외 안 던짐).
    """
    parse_warnings: list[str] = []
    # 손상/암호화/비-xlsx는 openpyxl이 BadZipFile/InvalidFileException 등을 던진다 —
    # docstring 계약(예외 안 던짐)대로 잡아 빈 요약+사유로 환원(500 대신 200+warning).
    try:
        wb = _load_workbook(coverage_source)
    except Exception as e:  # noqa: BLE001 — 로드 실패 전체를 계약대로 흡수
        parse_warnings.append(
            f"Coverage Report 로드 실패 — 손상/암호화/비-xlsx 의심 ({type(e).__name__})"
        )
        return {"coverage_summary": {}, "parse_warnings": parse_warnings}
    _owns_wb = wb is not coverage_source  # passthrough Workbook(호출자 소유)은 닫지 않음
    try:
        names = [n.lower() for n in wb.sheetnames]
        if not any("traceability" in n for n in names):
            parse_warnings.append(
                "Coverage Report에 'Traceability' 시트 없음 — Hyundai/Mobis 포맷 가정 검증 필요"
            )
        if not any(
            "coverage" in n and "traceability" not in n and "consistency" not in n
            for n in names
        ):
            parse_warnings.append("Coverage Report에 'Coverage' 시트(3.Coverage 등) 없음")
        cov = _extract_coverage_summary(
            wb, out_warnings=parse_warnings, tc_prefix=tc_prefix,
        )
    finally:
        if _owns_wb and hasattr(wb, "close"):
            wb.close()   # read_only 워크북의 ZipFile/파일핸들 해제
    return {"coverage_summary": cov, "parse_warnings": parse_warnings}


def summarize_test_report(
    sutr_source: Any, *, tc_prefix: str = "SwUTC",
) -> dict[str, Any]:
    """단일 SUTR/SITR(.xlsm)만 파싱해 sutr_summary를 반환(정합성 비교 없이).

    `_extract_sutr_summary` 단독 호출. not_executed는 빌더가 Test Summary 행에
    숫자로 stamp하지 않는 알려진 결함이 있어(노란 hint 셀) 추출 시 0으로 남는다 —
    여기서만 not_executed_tcs 목록 길이로 read-side 보정한다(공유 추출기/정합성 경로는
    불변, 단일 문서 표시값만 정확화). Returns {sutr_summary, parse_warnings}.
    """
    parse_warnings: list[str] = []
    try:
        wb = _load_workbook(sutr_source)
    except Exception as e:  # noqa: BLE001 — 손상/암호화/비-xlsx를 계약대로 흡수
        parse_warnings.append(
            f"Test Report 로드 실패 — 손상/암호화/비-xlsx 의심 ({type(e).__name__})"
        )
        return {"sutr_summary": {}, "parse_warnings": parse_warnings}
    _owns_wb = wb is not sutr_source
    try:
        names = [n.lower() for n in wb.sheetnames]
        if not any(
            "test summary" == n or ("summary" in n and "test" in n) for n in names
        ):
            parse_warnings.append("Test Report에 'Test Summary' 시트 없음")
        sutr = _extract_sutr_summary(wb, tc_prefix=tc_prefix, out_warnings=parse_warnings)
    finally:
        if _owns_wb and hasattr(wb, "close"):
            wb.close()
    # not_executed read-side 보정 — 목록은 있는데 카운트가 0이면 목록 길이로 채움.
    # 상한 clamp: 섹션 파싱 드리프트로 목록이 Total TC를 넘으면 total로 보정 + 경고
    # (표시 전용 — 공유 _extract_sutr_summary/정합성 verdict는 raw 값 그대로 사용).
    ne_tcs = sutr.get("not_executed_tcs") or []
    if not sutr.get("not_executed") and ne_tcs:
        total = sutr.get("total_tcs") or 0
        n = len(ne_tcs)
        if total and n > total:
            parse_warnings.append(
                f"미실행 TC 목록 {n}건이 Total TC {total} 초과 — 섹션 파싱 드리프트 의심, {total}로 보정"
            )
            n = total
        sutr["not_executed"] = n
    return {"sutr_summary": sutr, "parse_warnings": parse_warnings}


def check_swut_consistency(
    coverage_source: Any,
    sutr_source: Any,
    *,
    tc_prefix: str = "SwUTC",
) -> ConsistencyReport:
    """Coverage Report ↔ SUTR(SITR) 일관성 검증.

    Args:
        coverage_source: Coverage Report xlsx (path / bytes / Workbook).
        sutr_source: SUTR (SwUT) 또는 SITR (SwIT) xlsm (path / bytes / Workbook).
        tc_prefix: TC name prefix — SwUT="SwUTC" (default), SwIT="SwITC" (35차).
            미실행 TC name에서 함수 ID 추출 시 regex `^{tc_prefix}_(SwUFn_\\d+)` 사용.

    Returns:
        ConsistencyReport — issues 비어 있으면 ok=True. 파싱 실패는 parse_warnings에 기록.
    """
    cov_wb = _load_workbook(coverage_source)
    sutr_wb = _load_workbook(sutr_source)

    # 파싱 사전 검증 — 필수 시트/헤더 미발견 시 silent empty 막고 명시 warning.
    parse_warnings: list[str] = []
    cov_sheet_names_lower = [n.lower() for n in cov_wb.sheetnames]
    if not any("traceability" in n for n in cov_sheet_names_lower):
        parse_warnings.append(
            "Coverage Report에 'Traceability' 시트 없음 — Hyundai/Mobis 포맷 가정 검증 필요"
        )
    if not any("coverage" in n and "traceability" not in n and "consistency" not in n
               for n in cov_sheet_names_lower):
        parse_warnings.append(
            "Coverage Report에 'Coverage' 시트(3.Coverage 등) 없음"
        )
    if not any("test summary" == n or ("summary" in n and "test" in n)
               for n in cov_sheet_names_lower):
        parse_warnings.append("Coverage Report에 'Test Summary' 시트 없음")

    sutr_sheet_names_lower = [n.lower() for n in sutr_wb.sheetnames]
    if not any("test summary" == n or ("summary" in n and "test" in n)
               for n in sutr_sheet_names_lower):
        parse_warnings.append("SUTR에 'Test Summary' 시트 없음")

    cov = _extract_coverage_summary(
        cov_wb, out_warnings=parse_warnings, tc_prefix=tc_prefix,
    )
    sutr = _extract_sutr_summary(sutr_wb, tc_prefix=tc_prefix, out_warnings=parse_warnings)

    issues: list[ConsistencyIssue] = []

    # ---- (1) 미커버 Function ↔ 미실행 TC 매칭 ----
    # 미커버 SwUFn_X 에는 대응 TC SwUTC_SwUFn_X 가 미실행 목록에 있어야 함.
    cov_uncov = set(cov.get("uncovered_functions") or [])
    sutr_ne = sutr.get("not_executed_tcs") or []
    sutr_ne_func: set[str] = set()
    # 35차: tc_prefix 동적 적용 + 라운드 96-fix W-B 명명 변형 일반화
    _, _tc_fn_re = _tc_id_patterns(tc_prefix)
    for tc in sutr_ne:
        fid = _extract_function_id(_tc_fn_re, tc)
        if fid:
            sutr_ne_func.add(fid)

    only_in_cov = cov_uncov - sutr_ne_func
    only_in_sutr = sutr_ne_func - cov_uncov
    if only_in_cov:
        issues.append(ConsistencyIssue(
            severity="warning",
            category="uncovered_mismatch",
            message=f"Coverage Report 미커버 Function이 SUTR 미실행 목록에 없음: {sorted(only_in_cov)}",
            expected=sorted(cov_uncov),
            actual=sorted(sutr_ne_func),
        ))
    if only_in_sutr:
        issues.append(ConsistencyIssue(
            severity="warning",
            category="uncovered_mismatch",
            message=f"SUTR 미실행 TC가 Coverage Report에는 커버된 Function: {sorted(only_in_sutr)}",
            expected=sorted(cov_uncov),
            actual=sorted(sutr_ne_func),
        ))

    # ---- (2) Exception 카운트 ↔ Deviation 카운트 ----
    exc_total = (cov.get("exception_statement") or 0) + (cov.get("exception_branch") or 0)
    dev_total = len(sutr.get("deviation_tcs") or [])
    # Exception은 분기/구문 단위, Deviation은 TC 단위 → exc_total >= dev_total 기대.
    if exc_total < dev_total:
        issues.append(ConsistencyIssue(
            severity="warning",
            category="exception_deviation",
            message=f"Coverage Exception ({exc_total}) < SUTR Deviation TC ({dev_total}) — 분기/구문이 TC 수보다 적은 비정상 상태",
            expected=f">= {dev_total}",
            actual=exc_total,
        ))
    elif dev_total > 0 and exc_total == 0:
        issues.append(ConsistencyIssue(
            severity="warning",
            category="exception_deviation",
            message="SUTR Deviation은 존재하나 Coverage Exception이 0 — Coverage Report 미반영 의심",
            expected=">0",
            actual=0,
        ))

    # ---- (3) Total TC 일관성 ----
    cov_total = cov.get("total_tcs") or 0
    sutr_total = sutr.get("total_tcs") or 0
    if cov_total and sutr_total and cov_total != sutr_total:
        # 미실행 TC가 Coverage Traceability에 누락된 경우 차이 1 이상 발생 가능.
        diff = sutr_total - cov_total
        ne_count = sutr.get("not_executed") or 0
        if diff == ne_count and ne_count > 0:
            issues.append(ConsistencyIssue(
                severity="info",
                category="total_tc",
                message=f"Coverage Traceability TC({cov_total}) vs SUTR Total({sutr_total}) 차이 {diff} — 미실행 TC 수와 일치, 정상",
                expected=sutr_total,
                actual=cov_total,
            ))
        else:
            issues.append(ConsistencyIssue(
                severity="warning",
                category="total_tc",
                message=f"Total TC 불일치: Coverage={cov_total}, SUTR={sutr_total} (미실행 {ne_count}개 고려해도 차이 {diff})",
                expected=sutr_total,
                actual=cov_total,
            ))

    # ---- (4) Final Result 용어 통일 ----
    cov_fr = cov.get("final_result", "")
    sutr_fr = sutr.get("final_result", "")
    if cov_fr and sutr_fr:
        if _normalize_result(cov_fr) != _normalize_result(sutr_fr):
            issues.append(ConsistencyIssue(
                severity="warning",
                category="final_result",
                message=f"Final Test Result 의미 불일치: Coverage='{cov_fr}', SUTR='{sutr_fr}'",
                expected=cov_fr,
                actual=sutr_fr,
            ))
        elif cov_fr.strip().upper() != sutr_fr.strip().upper():
            issues.append(ConsistencyIssue(
                severity="info",
                category="final_result",
                message=f"Final Test Result 표기만 다름 (의미 동일): Coverage='{cov_fr}', SUTR='{sutr_fr}'",
                expected=cov_fr,
                actual=sutr_fr,
            ))

    # critical 0건이면 ok=True (warning만 있어도 보고용으로는 통과 가능)
    # 단, parse_warnings 존재 시 일관성 검증 결과 자체의 신뢰도가 떨어지므로 ok=False.
    critical_present = any(i.severity == "critical" for i in issues)
    return ConsistencyReport(
        ok=not critical_present and not parse_warnings,
        issues=issues,
        coverage_summary=cov,
        sutr_summary=sutr,
        parse_warnings=parse_warnings,
    )
