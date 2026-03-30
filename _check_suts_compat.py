import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'D:\Project\devops\260105\reports\suts\suts_local_20260310_164646.xlsm',
    read_only=True, data_only=True
)
ws = wb['2.SW Unit Test Spec']
count = 0
related_count = 0
io_count = 0
sample_rows = []
for r in range(7, min((ws.max_row or 7) + 1, 5000)):
    tc_id = ws.cell(row=r, column=3).value
    if tc_id and str(tc_id).startswith('SwUTC'):
        count += 1
        related = str(ws.cell(row=r, column=149).value or '').strip()
        if related and related != 'None':
            related_count += 1
        n_inp = sum(1 for c in range(14, 63) if ws.cell(row=r, column=c).value is not None)
        n_out = sum(1 for c in range(63, 149) if ws.cell(row=r, column=c).value is not None)
        if n_inp > 0 or n_out > 0:
            io_count += 1
        if len(sample_rows) < 5:
            sample_rows.append((str(tc_id), related, n_inp, n_out))
wb.close()

with open(r'D:\Project\devops\260105\_suts_result.txt', 'w', encoding='utf-8') as f:
    f.write(f'Total SUTS TCs: {count}\n')
    f.write(f'With related_fid: {related_count} ({100*related_count/max(count,1):.1f}%)\n')
    f.write(f'With I/O: {io_count} ({100*io_count/max(count,1):.1f}%)\n')
    f.write(f'Sample TCs:\n')
    for tc_id, related, ni, no in sample_rows:
        f.write(f'  {tc_id} -> related={related}, inp={ni}, out={no}\n')
print('Done')
