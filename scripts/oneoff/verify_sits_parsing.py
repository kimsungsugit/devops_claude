"""SITS 문서 파싱 완전성 검증 (U: 드라이브 실데이터 환경에서 실행).

목적: 추적성 매트릭스의 SITS 밴드가 낮은 원인이
  (A) SwITS 스펙 문서에 통합시험 케이스가 실제로 적어서인지, vs
  (B) 파서가 문서를 절단/누락(조기종료·시트선택·컬럼)해서인지
를 판별한다. 또한 (C) VectorCAST IT 로그가 매트릭스에 반영되는지 교차 확인.

실행:  cd <repo>  &&  backend/.venv/Scripts/python.exe scripts/oneoff/verify_sits_parsing.py [scm_id]
       (scm_id 생략 시 kjpds02_pv 사용. 필요시 config/file_mode.json이 local/cloudium 어느쪽인지 확인)

출력을 그대로 복사해 공유하면 A/B/C 판정이 가능하다.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))

SWITC_RE = re.compile(r"SwITC\w*_\w+", re.I)
SWUFN_RE = re.compile(r"Sw[UI]Fn_\d+", re.I)


def _load_scm(scm_id: str):
    from backend.services.scm_registry import get_registry_entry
    e = get_registry_entry(scm_id)
    if not e:
        print(f"[ERR] scm_id '{scm_id}' 미등록")
        sys.exit(1)
    return e


def step1_raw_document(sits_path: str) -> None:
    """[진실] SwITS 실파일을 파서 무관하게 직접 열어 시트별 행수·SwITC 케이스 총수 카운트."""
    print("\n========== STEP 1: SwITS 문서 직접 파싱 (파서 무관 = 진실값) ==========")
    print("path:", sits_path)
    try:
        import openpyxl
        from backend.services.file_resolver import get_resolver
        import io
        data = get_resolver().read_bytes(sits_path)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        print(f"[ERR] 문서 열기 실패: {exc}")
        return
    print("시트 목록:", wb.sheetnames)
    total_switc: set = set()
    for name in wb.sheetnames:
        ws = wb[name]
        rows = 0
        switc: set = set()
        swufn: set = set()
        for row in ws.iter_rows(values_only=True):
            rows += 1
            for cell in row:
                s = str(cell or "")
                if not s:
                    continue
                for m in SWITC_RE.findall(s):
                    switc.add(m.upper())
                for m in SWUFN_RE.findall(s):
                    swufn.add(m.upper())
        total_switc |= switc
        flag = "  <-- 통합시험 스펙 시트 후보" if switc else ""
        print(f"  [{name}] rows={rows}  SwITC={len(switc)}  SwUFn={len(swufn)}{flag}")
    print(f">>> 문서 전체 distinct SwITC 케이스: {len(total_switc)}")
    wb.close()


def step2_parser_output(sits_path: str) -> None:
    """[파서] SITS extract 엔드포인트가 같은 파일에서 실제로 몇 케이스를 뽑는지."""
    print("\n========== STEP 2: SITS extract 파서 결과 (백엔드가 뽑는 값) ==========")
    try:
        from backend.services import file_resolver as fr
        # 로컬 파일이면 local resolver 강제(파싱 로직만 검증). cloudium 실경로면 주석 처리.
        # fr._resolver = fr.LocalFileResolver()
        from backend.routers.jenkins import jenkins_sits_extract_traceability
        res = jenkins_sits_extract_traceability({"path": sits_path})
    except Exception as exc:  # noqa: BLE001
        print(f"[ERR] 파서 호출 실패: {exc}")
        return
    rows = res.get("vcast_rows", [])
    tcs = {r.get("testcase") for r in rows}
    print(f"파서 추출: {len(rows)} 행, distinct testcase={len(tcs)}")
    print("distinct testcases:", sorted(tcs))
    print("direct_mapped:", res.get("direct_mapped"), "| warning:", res.get("warning"))
    print("available_sheets:", res.get("available_sheets"))
    print(">>> STEP1 문서 SwITC 수와 이 값이 크게 다르면 = 파서 절단/누락 확정")


def step3_vectorcast_it(entry) -> None:
    """[교차] VectorCAST IT 로그(APP_IT/BOOT_IT)가 실제로 몇 통합시험 함수를 담는지."""
    print("\n========== STEP 3: VectorCAST IT 로그 규모 (실제 통합시험) ==========")
    vlist = list(entry.linked_docs.vectorcast or [])
    it_paths = [p for p in vlist if "_IT_" in p.upper() or "/IT" in p.upper() or "IT_REPORT" in p.upper()]
    print("linked_docs.vectorcast:", len(vlist), "개")
    for p in vlist:
        tag = "  <-- IT(통합)" if p in it_paths else "  (UT/기타)"
        print("   -", p.split("/")[-1], tag)
    if not it_paths:
        print("[!] IT 로그 경로가 linked_docs.vectorcast에서 식별 안 됨 — 경로 네이밍 확인 필요")
    print(">>> IT 로그의 통합시험 함수 수 >> SITS 스펙 케이스(STEP2)면 = 스펙 파싱 누락 강한 증거")
    print("    (IT 로그 상세 파싱은 vectorcast-rag 엔드포인트/서버 필요 — 경로 존재 여부만 우선 확인)")
    import os
    for p in it_paths:
        exists = "존재" if os.path.exists(p) else "접근불가/부재"
        print(f"    IT 경로 {exists}: {p}")


def step4_cloudium_read_completeness(entry) -> None:
    """[worker] cloudium worker가 U: 문서를 '다' 읽는지 진단.

    개별 파일 read(read_bytes)는 chunking으로 완전하나, 디렉토리 리스팅(list_dir)은
    30s timeout·실패 시 silent 빈 목록 `[]`을 반환한다(실 로그에 정적분석 폴더 entries=0
    이력). 폴더 스캔이 필요한 IT 로그류는 여기서 통째 누락될 수 있다. cold(유휴 후 첫)와
    warm 리스팅 소요시간을 재 timeout 근접 여부를 본다.
    """
    print("\n========== STEP 4: cloudium worker read 완전성 진단 ==========")
    import time
    from backend.services import file_resolver as fr
    try:
        mode = fr.get_file_mode()
    except Exception:  # noqa: BLE001
        mode = "?"
    print("file_mode:", mode)
    resolver = fr.get_resolver()
    is_cloud = resolver.__class__.__name__ == "CloudiumFileResolver"
    print("resolver:", resolver.__class__.__name__)
    if is_cloud:
        try:
            alive = fr.is_gate_running()
            print("worker ping(8765):", "응답" if alive else "무응답 — worker 미실행!")
        except Exception as exc:  # noqa: BLE001
            print("worker ping 실패:", exc)

    def _timed_list(path: str, label: str) -> None:
        for tag in ("cold", "warm"):
            t0 = time.time()
            try:
                items = resolver.list_dir(path, pattern="*")
                dt = time.time() - t0
                warn = "  <-- timeout 근접(30s)!" if dt > 20 else ("  <-- 빈 목록(누락 의심)" if not items else "")
                print(f"  [{label}/{tag}] {dt:6.2f}s  entries={len(items)}{warn}")
            except Exception as exc:  # noqa: BLE001
                dt = time.time() - t0
                print(f"  [{label}/{tag}] {dt:6.2f}s  ERROR: {type(exc).__name__}: {str(exc)[:80]}")

    # SITS 문서 read_bytes 완전성(크기)
    sits_path = entry.linked_docs.sits
    if sits_path:
        try:
            t0 = time.time()
            data = resolver.read_bytes(sits_path)
            print(f"  SwITS read_bytes: {len(data):,} bytes in {time.time()-t0:.2f}s (완전 수신=chunking+eof)")
        except Exception as exc:  # noqa: BLE001
            print(f"  SwITS read_bytes ERROR: {type(exc).__name__}: {str(exc)[:100]}")
    # IT 로그 폴더 + 정적분석 폴더 리스팅 timing
    for p in list(entry.linked_docs.vectorcast or []):
        if "_IT_" in p.upper() or "통합" in p:
            _timed_list(p, "IT로그")
    for p in list(entry.linked_docs.codesonar or []):
        _timed_list(p, "정적분석")
    print(">>> read_bytes는 크면 완전, list_dir이 timeout/빈목록이면 그 폴더 파일 통째 누락(silent)")


def main() -> None:
    scm_id = sys.argv[1] if len(sys.argv) > 1 else "kjpds02_pv"
    entry = _load_scm(scm_id)
    sits_path = entry.linked_docs.sits
    print(f"SCM: {entry.id} ({entry.name})")
    if not sits_path:
        print("[ERR] linked_docs.sits 비어있음")
        sys.exit(1)
    step1_raw_document(sits_path)
    step2_parser_output(sits_path)
    step3_vectorcast_it(entry)
    step4_cloudium_read_completeness(entry)
    print("\n========== 판정 가이드 ==========")
    print("A. STEP1 SwITC ≈ STEP2 파서 케이스  → 파싱 정상, SITS 낮음은 문서 특성(v0.10 작성중)")
    print("B. STEP1 SwITC >> STEP2 파서 케이스  → 파서 절단/누락 (시트선택·empty_streak=50·컬럼)")
    print("C. STEP3 IT 함수 >> STEP2 스펙 케이스 → 통합시험은 실행됐으나 스펙/IT 반영 누락")
    print("D. STEP4 list_dir이 30s 근접/빈목록  → worker 리스팅 timeout으로 폴더 데이터 silent 누락")


if __name__ == "__main__":
    main()
