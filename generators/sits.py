"""SITS (Software Integration Test Specification) auto-generation engine.

Generates XLSM output matching the reference SITS structure:
  - TC 행(SwITC_xx) + 서브케이스 행
  - Columns: TC ID | Description | Call chain | Safety | Test Method | Gen Method |
             Input Param 1-82 | Expected Param 1-113 | Related ID
  - Sheets: Cover, History, 1.Introduction, 2.Test Environment,
            2.SW Integration Strategy(`_STRATEGY_SHEET_NAME`), `_SPEC_SHEET_NAME`

⚠ 시트 이름은 **상수 `_SPEC_SHEET_NAME` 하나**가 출처다. 라이터와 검증기가 각자
문자열을 들고 있으면 한쪽만 고쳐진다 — 실제로 그랬다(라이터는 `3.…` 로 옮겼는데
`validate_sits_xlsm` 은 `4.…` 를 계속 찾아 **자기 산출물을 한 번도 못 읽었다**:
TC 0 · sub-case 0 으로 보고하면서 "미검증" 처리, 2026-08-14 게이트 실측).
"""
from __future__ import annotations

import json
import logging
import re
import time
from collections import deque
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from backend.services.iso26262_doc_asil_extractor import _RELATED_PREFIX_CANON
from generators._artifact_check import apply_write_back_check
from generators.safety_marks import resolve_safety_related
from generators.uds_design_ids import load_uds_design_ids, resolve_design_id
from report_gen.doc_kind import is_sds_filename

_logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column layout constants (1-based, matching reference SITS XLSM)
# ---------------------------------------------------------------------------
# ⚠ **납품 정본**(KJPDS02_PV_SwITS v1.02, 205열) 기준이다. 표준 템플릿 v0.10 이
#   아니다 — 템플릿은 Param 10개 고정이라 실제 파라미터(입력 82 / 기대 113)를 못 담는다.
#
# 정본 실측(2026-08-11):
#   시트    : `3.SW Integration Test Spec`  (예전엔 시트 `4.…` 를 따로 만들어 붙였다)
#   r3 밴드 : B3:G3 'Test Case' · I3:CL3 'Input' · CM3:GU3 'Expected Result' · GV3 'Related ID'
#   r4 헤더 : B TC ID · C4:D4 'Description' · E Safety Related · F Test Method
#             · G Test Case Generation Method · H (서브케이스 번호) · I~ Param n · GV 'SwDS'
#   r5~     : TC 행(변수명) 1개 + 서브케이스 행 N개. B/E/F/G/GV 는 **블록 전체 병합**
#             (SUTS 와 다르다 — SUTS 는 Method 가 시퀀스 그룹 단위다)
_BAND_ROW = 3
_HEADER_ROW = 4
_DATA_START_ROW = 5

# 시험 규격 시트 이름 — 라이터·검증기의 **단일 출처**(위 모듈 docstring 참조).
_SPEC_SHEET_NAME = "3.SW Integration Test Spec"

# 통합 전략 시트 — 정본은 `2.SW`, 등록 템플릿은 `2. SW`(공백) 다. 이름은 **템플릿이 준
# 것을 그대로 둔다**(사용자 매크로가 시트명을 참조한다). 대신 찾을 때 공백·대소문자를
# 무시해 두 표기를 같은 시트로 본다 — 이름이 어긋난 채로 새 시트를 만들면 한 파일에
# 빈 정본 시트와 채워진 사본이 공존한다(`3.…`/`4.…` 에서 이미 겪은 결함).
_STRATEGY_SHEET_NAME = "2.SW Integration Strategy"
_RELID_CHECK_SHEET = "Related_ID 확인"
_RELID_TIDY_SHEET = "Related_ID 정리"
_END_OF_DOC = "< End of Document >"

# 전략 시트 트리 상한. 체인(`_MAX_CHAIN_NODES`=100)과 **분리한다** — 체인은 한 칸에
# `a -> b -> …` 문자열로 들어가 100 이면 이미 읽기 한계지만, 전략 시트는 행으로 펼치므로
# 더 담을 수 있다(정본 최대 427노드/블록). 시트 전체 행 상한은 따로 둔다.
_STRATEGY_MAX_NODES = 400
_STRATEGY_MAX_ROWS = 20000
_STRATEGY_MAX_DEPTH = 20

# 통합 항목 ID 접두 — 정본은 시험 케이스(`SwITC_…`)와 통합 항목(`SwIT_…`)을 구별해
# 근거 시트에는 후자를 쓴다. 두 시트를 연결하려면 이 변환이 **한 곳**에 있어야 한다.
_TC_ID_PREFIX = "SwITC_"
_INTEGRATION_ID_PREFIX = "SwIT_"


def _integration_id(tc_id: str) -> str:
    """`SwITC_SwUFn_0101_01` → `SwIT_SwUFn_0101_01`(정본 근거 시트 표기)."""
    t = str(tc_id or "")
    return _INTEGRATION_ID_PREFIX + t[len(_TC_ID_PREFIX):] if t.startswith(_TC_ID_PREFIX) else t

# Related ID 를 진입점에서 몇 홉까지 모을 것인가. 정본 50 TC · 원소 340 대조(2026-08-14):
#
#     홉  재현율            과잉    정확일치
#     0   142 (41.8%)         4      13     ← 진입 함수만(어휘만 넓힌 판)
#     1   174 (51.2%)        17      22
#     2   225 (66.2%)       134      24     ← 채택
#     3   253 (74.4%)       302      26
#     5   313 (92.1%)       557      27
#
# 3홉부터 한계효용이 뒤집힌다(재현 +8%p 에 과잉 +168). 과잉 302 중 293 은 정본이 **다른
# TC 에서는 쓰는** ID 라 허위 추적은 아니지만, 한 칸에 그만큼 실리면 "이 통합 지점의 설계
# 요소" 가 아니라 "이 서브트리 어딘가" 가 된다. 2홉이 그 경계다.
_RELATED_CHAIN_DEPTH = 2
_RELATED_CHAIN_NODES = 60
# 한 칸에 실을 ID 상한(정본 최대 28개). 넘치면 자르되 **몇 개를 잘랐는지** 통계로 낸다.
_MAX_RELATED_IDS = 40

_TCID_COL = 2          # B  — TC ID (SwITC_xx)
_DESC_COL = 3          # C  — 서브케이스 번호
_CHAIN_COL = 4         # D  — call chain (정본도 여기에 'Interface : main -> …' 를 쓴다)
_SAFETY_COL = 5        # E  — Safety Related (O/X)
_METHOD_COL = 6        # F  — Test Method
_GEN_COL = 7           # G  — Test Case Generation Method
_SEQ_COL = 8           # H  — 서브케이스 번호(반복)
_INPUT_COL_START = 9   # I  — Input Param 1
_INPUT_COL_END = 90    # CL — Input Param 82
_EXP_COL_START = 91    # CM — Expected Param 1
_EXP_COL_END = 203     # GU — Expected Param 113
_RELATED_COL = 204     # GV — Related ID (SwDS)

# ── 값 어휘 — SwITS 정본 실측 ────────────────────────────────────────────────
# ⚠ 결합자가 SwUTS 와 다르다. SwUTS 는 슬래시(`AOR/ABV`), SwITS 는 쉼표(`AOR, AEC`).
#   실측: Method `REQ, IFT` 49 · `FI` 5 / Gen `AOR, AEC` 49 · `AOR/ABV` 5.
_SITS_METHOD_DEFAULT = "REQ, IFT"   # 통합시험 기본 — 요구 기반 + 인터페이스 시험
_SITS_METHOD_FAULT = "FI"           # 고장 주입
_SITS_GEN_DEFAULT = "AOR, AEC"
_SITS_GEN_BOUNDARY = "AOR/ABV"

# ── Introduction 1.5 / 1.6 — 이 문서가 **자기 어휘를 정의하는 곳** ──────────
#
# ⚠ 감사자는 3번 시트의 `REQ, IFT` · `FI` 를 보고 1.5 표에서 그 코드를 찾는다.
#   그런데 저장소 템플릿의 1.5 는 **하드웨어 통합시험 판**(`FNCT`/`FIT`/`ELCT`)이라,
#   우리가 내는 코드가 **문서 안에 정의되지 않은 값**이었다(실측: 정본 v1.02 는
#   REQ/IFT/RUT/SEP/FI). 같은 결함을 SwUTS 가 먼저 겪고 3번 시트만 고쳤는데
#   (`suts.py` `_METHOD_REQ` 주석), SITS 는 그 반대로 3번 시트만 고치고 Introduction 을
#   두었다. 두 곳이 갈라지지 않게 **여기 하나**를 두고 양쪽이 참조한다.
#
# 정본(KJPDS02_PV_SwITS v1.02 Introduction) 원문 그대로. ASIL 열은 그 문서의 값이며
# 우리가 판정에 쓰지 않는다(문서가 선언하는 적용 범위일 뿐 — 지어내지 않는다).
_INTRO_TEST_METHODS: Tuple[Tuple[str, str, str], ...] = (
    ("Requirements-based Test", "REQ", "A,B,C,D"),
    ("Interface Test", "IFT", "A,B,C,D"),
    ("Resource Usage Test", "RUT", "A,B,C,D"),
    ("System Error Protection Analysis", "SEP", "A,B,C,D"),
    ("Fault Injection Test", "FI", "B,C,D"),
)
_INTRO_GEN_METHODS: Tuple[Tuple[str, str, str], ...] = (
    ("Analysis Of Requirement", "AOR", "A,B,C,D"),
    ("Analysis Of internal and external Interface", "AOI", "B,C,D"),
    ("generation and Analysis of Equivalence Classes", "AEC", "C,D"),
    ("Analysis of Boundary Values", "ABV", "C,D"),
    ("knowledge or experience based ERror guessing", "ERG", "A,B,C,D"),
    ("Analysis of Funtional Depedency", "AFD", "C,D"),
    ("Analysis of common limit conditions, sequences and sources of Dependent Failures",
     "ADF", "C,D"),
    ("Analysis of environment codition and operational Use Cases", "AUC", "B,C,D"),
    ("STAndard if existing", "STA", "-"),
    ("Analysis of Significant Variables", "ASV", "A,B,C,D"),
)


def _intro_codes(rows: Tuple[Tuple[str, str, str], ...]) -> frozenset:
    return frozenset(a for _n, a, _l in rows)


def _split_method_codes(value: Any) -> List[str]:
    """`"REQ, IFT"` · `"AOR/ABV"` → `["REQ","IFT"]` · `["AOR","ABV"]`.

    결합자는 문서마다 다르다(SwITS 는 Test Method 가 쉼표, Gen Method 가 슬래시).
    판정은 **코드 단위**로 해야 하므로 둘 다 분리한다.
    """
    return [p for p in (t.strip() for t in re.split(r"[,/]", str(value or ""))) if p]


# `Safety Related` 칸 — 구현은 `generators/safety_marks.py` 가 단일 출처다.
# ⚠ 여기에 다시 쓰지 말 것. 예전엔 STS·SUTS·SITS 가 각자 한 벌씩 들고 있었고, 안전 판정
#   수정 커밋 3건이 그중 `sts.py` 에는 한 번도 안 닿았다.
_safety_mark = resolve_safety_related


def _resolve_flow_asil(
    fn_name: str,
    info: Dict[str, Any],
    uds_asil_map: Optional[Dict[str, str]] = None,
    uds_swcom_map: Optional[Dict[str, List[str]]] = None,
    sds_map: Optional[Dict[str, Dict[str, str]]] = None,
) -> str:
    """흐름의 ASIL — **입력문서 우선**, 근거가 없으면 빈 문자열.

    우선순위:
      ① SwUDS 가 함수마다 적어 둔 ASIL (`extract_function_asil_from_kv_tables`)
      ② 그 함수가 속한 SwCom 의 ASIL 을 SDS 에서 상속(ISO 26262 — 함수는 소속 SW
         컴포넌트의 등급을 상속한다). 여러 컴포넌트면 **가장 높은 등급**.
      ③ 소스 주석(`@asil`)
      ④ 없으면 **빈칸** — `QM` 으로 강등하지 않는다.

    ## 왜 이 순서인가 (2026-08-14 실측, KJPDS02_PV · 흐름 367)

        축                              O     X   빈칸
        소스 주석만 · TBD→QM 강등(현행)   86   281     0
        소스 주석만 · 강등 없이           86    21   260
        SwUDS 함수 ASIL                211    94    62
        SDS 컴포넌트 상속               215     0   152
        **① → ② → ③**                 213    94    60
        (정본: O 43(79.6%) · X 11 · 빈칸 0)

    현행은 **260건을 근거 없이 `X`(비안전)로** 찍고 있었다. SwUDS 에 함수별 ASIL 이
    1,003건(A 701 · QM 302) 있는데 한 건도 안 봤기 때문이다. `_safety_mark` 는 근거
    부재를 빈칸으로 두도록 짜여 있었지만 그 앞의 `or "QM"` 이 그 경로를 죽였다 —
    안전 등급을 **지어내지 않는다**는 이 저장소의 규약(SUTS·STS 와 동일)에 어긋난다.
    """
    name = str(fn_name or "")
    got = str((uds_asil_map or {}).get(name.lower()) or "").strip()
    if got:
        return got
    # ② 컴포넌트 상속 — SwUDS 가 준 SwCom 을 SDS 에서 찾아 가장 높은 등급을 취한다.
    best, best_rank = "", len(_ASIL_RANK)
    for com in (uds_swcom_map or {}).get(name.lower()) or []:
        entry = (sds_map or {}).get(str(com).lower()) or (sds_map or {}).get(str(com))
        cand = str((entry or {}).get("asil") or "").strip()
        if cand and _asil_rank(cand) < best_rank:
            best, best_rank = cand, _asil_rank(cand)
    if best:
        return best
    # ③ 소스 주석. `TBD` 는 "정해지지 않았다" 이지 "비안전" 이 아니다 — 그대로 둔다.
    src = str(info.get("asil") or "").strip()
    return "" if src.upper() == "TBD" else src


def _sits_test_method(itc: Dict[str, Any]) -> str:
    """통합 TC 의 Test Method — **전용 FI TC 인가**로 판정한다.

    ⚠ 예전 판은 "오류 전파 서브케이스를 하나라도 가지면 TC 전체가 FI" 였다.
      그런데 정본(KJPDS02_PV_SwITS v1.02)에서 FI 는 **전용 TC** 다:

          REQ, IFT ↔ AOR, AEC   49건   (동등분할 — 무효 등가류 EC1/EC7 포함)
          FI       ↔ AOR/ABV     5건   (경계값분석 · `SwITC_FI_SwFn_NN`)

      즉 정본은 무효 경계 서브케이스를 가진 TC 를 FI 로 올리지 **않는다** —
      Test Method 와 Gen Method 가 짝이라 AEC 를 쓰면서 FI 인 조합이 0건이다.

    ⚠ 옛 판정은 **살아 있는 지뢰**였다. `_generate_sub_cases` 의 오류전파 블록
      (`ERR_PROP_*`)은 `len(sub_cases) < max_cases` 가드에 막혀 있는데, 기본
      경계값이 정확히 7개라 `max_subcases=7` 이면 예산이 꽉 차 한 번도 안 돈다.
      예산을 조금만 올리면(기본값 `_DEFAULT_SUBCASES=14` 포함) **전 TC 가 FI 로
      뒤집힌다** — 실측: max 7 → `REQ, IFT` · max 10/14 → `FI`.
      영향도 재생성(`impact_orchestrator._run_sits_generation`)은 이 인자를 안 넘겨
      기본값 14 를 쓰므로, 그 경로의 산출물은 **전 TC 가 FI** 였다.
    """
    m = str(itc.get("test_method") or "").strip()
    return m or _SITS_METHOD_DEFAULT


def _sits_gen_method(gen: Any, test_method: Optional[str] = None) -> str:
    """정본 어휘의 Gen Method. **Test Method 와 짝**이다.

    정본 실측(KJPDS02_PV_SwITS v1.02, 54건 — 다른 조합 **0건**):

        REQ, IFT  ↔  AOR, AEC      49
        FI        ↔  AOR/ABV        5

    ⚠ 예전엔 생성기 내부 라벨(`gen`)만 보고 독립으로 정했는데, 그 라벨을 만드는
    `_determine_gen_method_for_flow` 의 **네 분기가 모두 `ABV` 를 포함**한다
    (`AOR, ABV` · `ABV, AEC` · `ABV` · `ABV, AEC`). 그래서 `AOR, AEC` 가지는
    **한 번도 발생한 적이 없고**, 실 산출물 367건이 전부 `REQ, IFT` × `AOR/ABV` —
    정본에 0건인 조합 — 이었다(2026-08-14 실측).

    ⚠ 기존 단위 테스트가 `_sits_gen_method("AEC") == "AOR, AEC"` 를 단언하며 초록이었다.
    생성기가 순수 `"AEC"` 를 **낼 수 없으므로** 그 경로는 도달 불가였다 — 실 입력이
    아닌 값으로 검증하면 "동작한다" 는 인상만 남는다.

    `test_method` 를 주면 그 짝을 따른다. 생략하면 구 동작(라벨 기반)이라 하위 호환은
    되지만 정본 짝을 보장하지 못한다 — 라이터는 반드시 넘긴다.
    """
    if test_method is not None:
        m = str(test_method).strip().upper()
        return _SITS_GEN_BOUNDARY if m.startswith(_SITS_METHOD_FAULT) else _SITS_GEN_DEFAULT
    g = str(gen or "").strip().upper()
    return _SITS_GEN_BOUNDARY if ("ABV" in g or "BV" in g) else _SITS_GEN_DEFAULT

# ⚠ 주석의 숫자는 **옛 레이아웃**(67/70)이었다. 실제 값은 정본과 같은 82/113 이다
#   — 배열 펼침 예산이 이 값이라 숫자를 잘못 읽으면 15+43칸을 안 쓰는 줄 안다.
_MAX_INPUT_PARAMS = _INPUT_COL_END - _INPUT_COL_START + 1   # 82 (I~CL)
_MAX_EXP_PARAMS = _EXP_COL_END - _EXP_COL_START + 1         # 113 (CM~GU)

# Row 6 상세 헤더(열 번호 → 라벨). `generate_sits_xlsm`이 시트에 쓰는 값이자, 영향도 탭의
# 문서 초안이 Excel 붙여넣기 TSV 열 순서를 얻는 **단일 출처**다(복제 금지 — suts와 동일 원칙).
_DETAIL_HEADERS = {
    _TCID_COL: "TC ID",
    _DESC_COL: "Description",
    _CHAIN_COL: "Call Chain",
    _SAFETY_COL: "Safety Related",
    _METHOD_COL: "Test Method",
    _GEN_COL: "Test Case Generation Method",
    _RELATED_COL: "SwDS",
}
# ⚠ `Precondition` 열은 정본에 **없다**(그건 SwTS 정본의 열이다). 예전 판은 F열에
#   Precondition 을 써서 정본의 `Test Method` 자리를 차지하고 있었다.
# (`_MAX_SUBCASES = 16` 은 어디서도 참조되지 않는 죽은 상수였다 — 옆의 14 와 값이 달라
#  "상한 16 이 걸린다" 는 오해를 부른다. 실제 상한은 아래 _DEFAULT_SUBCASES 뿐이다.)
_DEFAULT_SUBCASES = 14  # 7 BV + 4 COND_COMB + 2 ERR_PROP + 2 GLOBAL

# Boundary value sets for common C types — 7 values per type:
#   min_inv | min_valid | low_mid | mid | high_mid | max_valid | max_inv
# This lets max_subcases=7 produce 7 distinct sub-cases per TC.
_BOUNDARY_SETS: Dict[str, List[Any]] = {
    "uint8":  [-1,    0,    63,    127,   191,   255,   256],
    "uint16": [-1,    0,   16383, 32767, 49151, 65535, 65536],
    "uint32": [-1,    0,   0x3FFFFFFF, 0x7FFFFFFF, 0xBFFFFFFF, 0xFFFFFFFF, 0x100000000],
    "int8":   [-129, -128,  -64,    0,    63,   127,   128],
    "int16":  [-32769, -32768, -16384, 0, 16383, 32767, 32768],
    "int32":  [-2147483649, -2147483648, -1073741824, 0, 1073741823, 2147483647, 2147483648],
    "float":  [-1.0,  0.0,   0.25,  0.5,  0.75,  1.0,   1.001],
    "bool":   [-1,    0,     0,     0,    1,     1,     2],
    "default": [-1,   0,    63,    127,   191,   255,   256],
}

_SDS_MAP_CACHE: Optional[Dict[str, Dict[str, str]]] = None
_SDS_MAP_CACHE_MTIME: float = 0.0

# ---------------------------------------------------------------------------
# STP document parsing
# ---------------------------------------------------------------------------

def _parse_stp_document(stp_path: str) -> Dict[str, Any]:
    """Load and parse an STP file (.docx/.pdf/.txt) into a structured context dict.

    Returns:
        {
            "raw":                 str   — full extracted text,
            "doc_id":              str   — 문서번호 (e.g. "HDPDM01-STP-0825"),
            "version":             str   — 개정번호 (e.g. "v1.01"),
            "environments":        List[str] — test environment labels,
            "regression_strategy": str   — regression strategy excerpt,
        }
    """
    try:
        from generators.sts import _load_stp_context
        raw = _load_stp_context(stp_path)
    except Exception:
        raw = ""

    if not raw:
        return {}

    ctx: Dict[str, Any] = {
        "raw": raw,
        "doc_id": "",
        "version": "",
        "environments": [],
        "regression_strategy": "",
    }

    # 문서번호
    m = re.search(r"문서번호\s+([\w\-./]+)", raw)
    if m:
        ctx["doc_id"] = m.group(1).strip()

    # 개정번호 / 버전
    m = re.search(r"(?:개정번호|버전|Version|Rev\.?)\s+(v[\d.]+|\d+\.\d+)", raw, re.IGNORECASE)
    if m:
        ctx["version"] = m.group(1).strip()

    # 테스트 환경 — look for known environment keywords per line
    _ENV_PAT = re.compile(
        r"(HW.?in.?the.?loop|Hardware.?in.?the.?loop|HiL|"
        r"ECU\s*네트워크|ECU\s*network|"
        r"차량(?:\s*환경)?|Vehicle|MiL|SiL|TargetHW)",
        re.IGNORECASE,
    )
    seen_envs: set = set()
    for line in raw.splitlines():
        line = line.strip()
        m = _ENV_PAT.search(line)
        if m:
            # Use the matched token as the canonical environment label
            label = m.group(0).strip()
            if label.lower() not in seen_envs:
                seen_envs.add(label.lower())
                ctx["environments"].append(label)
        if len(ctx["environments"]) >= 6:
            break

    # 회귀 전략
    m = re.search(r"회귀\s*전략[^\n]*\n(.*?)(?=\n\n|\Z)", raw, re.DOTALL)
    if m:
        ctx["regression_strategy"] = m.group(0).strip()[:300]

    _logger.info(
        "SITS: STP parsed — doc_id=%s version=%s envs=%s",
        ctx["doc_id"], ctx["version"], ctx["environments"],
    )
    return ctx


# ---------------------------------------------------------------------------
# Shared helpers (re-used from sts / suts patterns)
# ---------------------------------------------------------------------------

# SwUDS 파생 맵 캐시. (경로, mtime_ns, size) → {"swcom": …, "asil": …}.
# 53MB docx 를 매번, 그것도 소비처마다 따로 훑을 수 없다(`_load_uds_maps`).
_UDS_SWCOM_CACHE: Dict[str, Tuple[Tuple[int, int], Dict[str, Any]]] = {}


def load_uds_swcom_map(uds_path: Optional[str]) -> Dict[str, List[str]]:
    """SwUDS 가 함수마다 적어 둔 `Related ID` 에서 `{함수명(소문자): [SwCom_NN]}`.

    ## 왜 이게 Related 칸의 소스인가

    정본 실측(KJPDS02_PV_SwITS v1.02) — Related 칸의 어휘는 **설계/시험 요소 ID** 다:
    `SwCom` 170회(33종) · `SwFn` 69 · `SwSTR` 62 · `SwST` 38 · `SwTK` 8.
    요구 ID(`SwTR_`/`SwTSR_`/`SwNTR_`/`SwEI_`)는 **0 건**이다.

    그리고 이 로더가 SwUDS 에서 뽑은 SwCom 33종은 **정본의 33종과 완전히 같다**
    (교집합 33 · 차집합 양쪽 0 · 함수 매핑 1,025건, 2026-08-14 실측). 즉 정본은
    바로 이 표를 근거로 Related 칸을 채운다.

    ⚠ 대비 — SDS 파티션 맵은 이 칸의 소스가 **아니다**. 값 스키마 실측은
    `{asil, canonical, component_description, description, kind, related}` 이고
    함수 항목(588개)의 `related` 에 든 것은 전부 요구 ID다(설계 ID 토큰 0건).

    실패는 빈 맵이다 — 그 경우 호출부는 합성 ID 로 내려가되 **합성임을 표시**해야 한다.
    """
    raw = str(uds_path or "").strip()
    if not raw:
        return {}
    p = Path(raw)
    try:
        st = p.stat()
        sig = (st.st_mtime_ns, st.st_size)
    except OSError as exc:
        _logger.warning("SITS: SwUDS 접근 실패 — SwCom 보강 생략: %s (%s)", raw, exc)
        return {}
    return _load_uds_maps(p, sig).get("swcom") or {}


def load_uds_asil_map(uds_path: Optional[str]) -> Dict[str, str]:
    """SwUDS 가 함수마다 적어 둔 ASIL — `{함수명(소문자): 'A'|'QM'|…}`.

    Safety Related 칸의 **1순위 근거**다(`_resolve_flow_asil`). 실측 1,003건
    (A 701 · QM 302)이 들어 있는데 예전엔 한 건도 안 봤고, 소스 주석이 없는 260건이
    `QM` 으로 강등돼 `X`(비안전)로 문서에 실렸다.
    """
    raw = str(uds_path or "").strip()
    if not raw:
        return {}
    p = Path(raw)
    try:
        st = p.stat()
        sig = (st.st_mtime_ns, st.st_size)
    except OSError as exc:
        _logger.warning("SITS: SwUDS 접근 실패 — ASIL 보강 생략: %s (%s)", raw, exc)
        return {}
    return _load_uds_maps(p, sig).get("asil") or {}


def load_uds_related_map(uds_path: Optional[str]) -> Dict[str, List[str]]:
    """SwUDS `Related ID` 칸의 **토큰 전체** — `{함수명(소문자): [SwCom_NN, SwFn_NN, …]}`.

    `load_uds_swcom_map` 이 같은 칸에서 `SwCom_` 만 걸러 쓰는 좁은 판이다. Related 칸의
    정본 어휘는 다섯 종(SwCom·SwFn·SwSTR·SwST·SwTK)이라 SwCom 만 쓰면 **19% 를 버린다**
    (실측 1,298 토큰 중 246). SwCom 축은 SDS 컴포넌트 맵 키와 맞춰야 해서 계속 좁게 쓰고,
    산출물 `Related ID` 칸은 이 넓은 판을 쓴다.
    """
    raw = str(uds_path or "").strip()
    if not raw:
        return {}
    p = Path(raw)
    try:
        st = p.stat()
        sig = (st.st_mtime_ns, st.st_size)
    except OSError as exc:
        _logger.warning("SITS: SwUDS 접근 실패 — Related ID 보강 생략: %s (%s)", raw, exc)
        return {}
    return _load_uds_maps(p, sig).get("related") or {}


def _load_uds_maps(p: Path, sig: Tuple[int, int]) -> Dict[str, Any]:
    """SwUDS 를 **한 번만** 읽어 Related·SwCom·ASIL 세 맵을 함께 뽑는다.

    ⚠ 로더가 각자 `read_bytes()` 하면 53MB 문서를 그 수만큼 읽는다. 소비처가 늘 때마다
    비용이 배로 붙으므로 캐시를 문서 단위로 둔다.
    """
    key = str(p.resolve()).lower()
    cached = _UDS_SWCOM_CACHE.get(key)
    if cached and cached[0] == sig:
        return cached[1]
    out: Dict[str, Any] = {"swcom": {}, "asil": {}, "related": {}}
    try:
        from backend.services.iso26262_doc_asil_extractor import (
            extract_function_asil_from_kv_tables,
            extract_function_related_ids_from_kv_tables,
        )
        data = p.read_bytes()
        # SwCom 판은 Related 판에서 거른다 — 문서를 두 번 훑지 않고, 두 축이 갈라지지도 않는다.
        out["related"] = extract_function_related_ids_from_kv_tables(data) or {}
        out["swcom"] = {n: coms for n, toks in out["related"].items()
                        if (coms := [t for t in toks if t.startswith("SwCom_")])}
        out["asil"] = extract_function_asil_from_kv_tables(data) or {}
    except Exception as exc:  # noqa: BLE001 - 보강 실패는 보고하고 빈 맵으로 계속한다
        _logger.warning("SITS: SwUDS 추출 실패(%s) — 합성 ID·소스 ASIL 로 내려간다: %s",
                        type(exc).__name__, exc)
        return out
    _UDS_SWCOM_CACHE[key] = (sig, out)
    _logger.info("SITS: SwUDS 로드 (%s) — Related ID %d함수/%d토큰 · SwCom %d건 · ASIL %d건",
                 p.name, len(out["related"]),
                 sum(len(v) for v in out["related"].values()),
                 len(out["swcom"]), len(out["asil"]))
    return out


def _load_default_sds_map() -> Dict[str, Dict[str, str]]:
    global _SDS_MAP_CACHE, _SDS_MAP_CACHE_MTIME
    docs_dir = Path(__file__).resolve().parents[1] / "docs"
    try:
        # `*SDS*` 글롭은 `SwDS` 표기를 놓친다("swds" 에 "sds" 없음) — 전량 글롭 후 단일 출처로 거른다.
        sds_files = sorted(p for p in docs_dir.glob("*.docx") if is_sds_filename(p.name))
        if sds_files:
            current_mtime = sds_files[0].stat().st_mtime
            # Return cached copy if file hasn't changed
            if _SDS_MAP_CACHE is not None and current_mtime == _SDS_MAP_CACHE_MTIME:
                return _SDS_MAP_CACHE
            from report_gen.requirements import _extract_sds_partition_map
            for f in sds_files:
                m = _extract_sds_partition_map(str(f))
                if m:
                    _SDS_MAP_CACHE = m
                    _SDS_MAP_CACHE_MTIME = f.stat().st_mtime
                    _logger.info("SITS: SDS map loaded from %s (%d entries)", f.name, len(m))
                    return _SDS_MAP_CACHE
    except Exception as e:
        _logger.debug("SITS: SDS map load failed: %s", e)
    # No SDS file or load failed — cache empty dict to avoid re-attempting every call
    if _SDS_MAP_CACHE is None:
        _SDS_MAP_CACHE = {}
    return _SDS_MAP_CACHE


def _infer_boundary_values(var_name: str) -> List[Any]:
    """Infer boundary values from annotated variable string or variable name.

    Supports two forms:
      - Annotated: '[IN] U8 u8t_Data' or '[OUT] return U16 (range: 0 ~ 65535)'
        → type is extracted from the explicit C type token (U8, S16, U32, …)
      - Plain name: 'u8Speed', 'u16Voltage'
        → type inferred from naming prefix (u8, u16, s32, …)
    """
    # ── 1. Explicit type token from annotated '[IN] TYPE varname' format ────
    type_match = re.search(
        r"\b(U8|U16|U32|U64|S8|S16|S32|S64|BOOL|BOOLEAN|FLOAT|FLOAT32|DOUBLE)\b",
        var_name,
        re.IGNORECASE,
    )
    if type_match:
        tok = type_match.group(1).upper()
        _type_map = {
            "U8": "uint8", "U16": "uint16", "U32": "uint32", "U64": "uint32",
            "S8": "int8",  "S16": "int16",  "S32": "int32",  "S64": "int32",
            "BOOL": "bool", "BOOLEAN": "bool",
            "FLOAT": "float", "FLOAT32": "float", "DOUBLE": "float",
        }
        return _BOUNDARY_SETS[_type_map[tok]]

    # ── 2. Naming-convention prefix / suffix (plain variable names) ─────────
    name = var_name.lower().lstrip("_")
    if re.search(r"\bu8|uint8|byte", name):
        return _BOUNDARY_SETS["uint8"]
    if re.search(r"\bu16|uint16|word", name):
        return _BOUNDARY_SETS["uint16"]
    if re.search(r"\bu32|uint32|dword", name):
        return _BOUNDARY_SETS["uint32"]
    if re.search(r"\bs8\b|int8", name):
        return _BOUNDARY_SETS["int8"]
    if re.search(r"\bs16\b|int16", name):
        return _BOUNDARY_SETS["int16"]
    if re.search(r"\bs32\b|int32", name):
        return _BOUNDARY_SETS["int32"]
    if re.search(r"float|flt|f32", name):
        return _BOUNDARY_SETS["float"]
    if re.search(r"flag|enable|active|bool|b_", name):
        return _BOUNDARY_SETS["bool"]
    return _BOUNDARY_SETS["default"]


def _clean_var_name(raw: str) -> str:
    """`[IN] u8g_Speed` → `u8g_Speed`. 정본 표기 규칙은 SUTS 와 **같은 출처**를 쓴다.

    ## 예전 판은 두 줄로 정반대 일을 했다

        s = re.sub(r"\\[.*?\\]", "", raw)   # `[IN]` 태그와 함께 **배열 첨자까지** 지웠다
        s = re.sub(r"\\s+", "_", s)         # 타입을 버리는 대신 **이름에 이어붙였다**

    그래서 `const UINT8 * data` 가 `const_UINT8_*_data` 로, `return U8` 이
    `return_UINT8` 로, `p->m` 이 `p->m` 그대로 산출물에 실렸다. 정본(VectorCAST)은
    각각 `data` · `return` · `p[0].m` 이라고 적는다.

    실측(2026-08-14, KJPDS02_PV): 정본 기대 656칸 중 **일치 0** · 입력 496 중 26(5.2%).
    미달의 최대 축이 배열이나 수집 범위가 아니라 **이름 표기 자체**였다.

    ## 복제하지 않는다

    SUTS 가 다섯 라운드에 걸쳐 정본과 맞춰 놓은 규칙(`return` 슬롯 · 포인터 `[0].` ·
    타입 한정자 제거 · 주석 제거)을 그대로 **호출**한다. 여기에 같은 로직을 다시 쓰면
    한쪽만 고쳐지는 이 저장소의 반복 실패 모드가 된다(`_resolve_sds_map` 을 같은 이유로
    이미 재사용하고 있다).

    ⚠ 이름을 뽑지 못하면 **빈 문자열**이다(예전엔 `raw[:40]` 으로 원문 조각을 흘렸다).
       호출부는 빈 값을 걸러야 한다.
    """
    from generators.suts import _extract_var_names
    names = _extract_var_names([str(raw or "")])
    return names[0] if names else ""


def _clean_global_var_name(raw: str) -> str:
    """전역 엔트리용 — `[INDIRECT] u8s_Flag` → `u8s_Flag`.

    ⚠ 파라미터용(`_clean_var_name`)과 **다른 함수**여야 한다. 전역의 방향 태그는
    `[IN]`/`[OUT]`/`[INOUT]` 말고도 `[INDIRECT]`·`[INDIRECT2]` 로 온다. 파라미터
    정제기는 그 세 개만 벗기므로 `[INDIRECT] …` 는 남은 대괄호 때문에 형태 검사에서
    **통째로 버려진다**.

    실측(2026-08-14): 두 경로를 한 함수로 합쳤더니 정본과 맞던 입력 **9칸이 사라졌다**
    — `u8s_E2EInitFlag_SBCM0`·`u8s_PrevCounter_SBCM0`·`g_DoorState`·
    `u32s_SecuritySeed` 처럼 전부 전역이었다. SUTS 도 같은 이유로 두 함수를 나눠 두고
    있고(`_clean_global_name` 은 태그 목록을 `_DIR_TAG_PAT` 단일 출처로 본다),
    그 주석은 태그 하나를 빼먹어 같은 실패를 두 번 겪었다고 적어 두었다.
    """
    from generators.suts import _clean_global_name, _vc_pointer_notation
    return _vc_pointer_notation(_clean_global_name(str(raw or "")))


def _get_module_name(info: Dict[str, Any]) -> str:
    """Derive module/component name from function info."""
    file_path = info.get("file") or info.get("source_file") or ""
    if file_path:
        stem = Path(file_path).stem
        # Strip trailing _PDS, _Main suffixes to get component
        stem = re.sub(r"(_PDS|_Main|_main)$", "", stem, flags=re.IGNORECASE)
        return stem
    return info.get("module_name") or info.get("component") or "Unknown"


def _infer_swcom_id(module_name: str, swcom_counter: Dict[str, int]) -> str:
    """Map module name to SwCom_XX ID, assigning new IDs incrementally."""
    key = module_name.lower()
    if key not in swcom_counter:
        swcom_counter[key] = len(swcom_counter) + 1
    return f"SwCom_{swcom_counter[key]:02d}"


def _parse_req_ids(text: str) -> List[str]:
    """Extract SwXX_NN requirement IDs from text."""
    return re.findall(r"\bSw(?:TR|TSR|NTR|NTSR|ST|STR|Fn|Com)_\d+\b", text or "")


def _expand_arrays(
    names: List[str],
    raws: List[str],
    budget: int,
    globals_info: Dict[str, Dict[str, str]],
    struct_members: Dict[str, Dict[str, str]],
    root_sizes: Dict[str, Tuple[int, ...]],
) -> Tuple[List[str], List[str], int, int, int]:
    """배열 이름을 원소로 펼친다 — 판정은 **전부 SUTS 자산**에 위임한다.

    정본(KJPDS02_PV_SwITS v1.02)은 배열을 원소 단위로 적는다:
    `g_sys_error_his[0]`…`[15]` · `u8g_SysEepromCtrl_PartNoInfo[0]`…`[9]`.
    우리는 base 한 칸으로 내고 있었다 — 같은 대상을 **다른 입도로** 부르는 것이라
    과다와 미달이 동시에 생긴다(실측: 정본 `[N]` 셀 414 vs 우리 179).

    ⚠ 크기 규칙을 여기서 다시 쓰지 않는다. `_array_sizes`/`_expand_array_entries` 는
      SUTS 가 여러 라운드에 걸쳐 맞춰 둔 것이고(선언 크기 vs 관찰 첨자, 다차원,
      구조체 배열의 root 첨자, 예산 부족 시 **펼치지 않고 base 유지**), 복제하면
      한쪽만 고쳐진다. `uds_unit_io.py` 가 못 박은 규약(문서의 `[N]` 은 선언 크기라
      떼고 실제 펼침은 소스 크기로)도 그 안에 이미 들어 있다.

    ⚠ `raws` 를 **함께** 펼친다. SITS 는 이름과 원문을 인덱스로 짝짓는다
      (`expected_raws[ev_idx]`) — 이름만 늘리면 원소마다 **다른 변수의** 타입·경계값이
      붙는다. 값이 틀리는 게 아니라 짝이 어긋나는 것이라 눈으로 안 보인다.

    Returns: `(names, raws, 펼친 이름 수, 예산부족 건너뜀, 방출 원소 수)`
    """
    if not names:
        return names, raws, 0, 0, 0
    from generators.suts import (
        _array_sizes,
        _declared_type_map,
        _expand_array_entries,
        _mid_member_sizes,
        _root_type_hints,
    )

    sizes = _array_sizes(raws, globals_info=globals_info, struct_members=struct_members)
    mid: Dict[str, Tuple[int, Tuple[int, ...]]] = {}
    if struct_members:
        _rtypes = _root_type_hints(raws, declared_types=_declared_type_map(globals_info))
        mid = _mid_member_sizes(names, _rtypes, struct_members)
    out, st = _expand_array_entries(
        names, sizes, budget, root_sizes=root_sizes or None, mid_sizes=mid or None,
        parallel=raws,
    )
    par = st.get("parallel")
    # 계약 위반은 조용히 넘기지 않는다 — 짝이 어긋난 채로 나가면 원소마다 엉뚱한
    # 타입이 붙고, 그건 산출물만 봐서는 절대 안 보인다.
    if par is None or len(par) != len(out):
        raise RuntimeError(
            f"배열 펼침 부속 리스트 불일치: names={len(out)} raws={len(par or [])}")
    return (out, par, len(st.get("expanded") or []), len(st.get("skipped") or []),
            max(0, len(out) - len(names)))


def _reach_cross_module(
    entry: str,
    calls_map: Dict[str, List[str]],
    module_of: Dict[str, str],
    max_hop: int,
) -> List[str]:
    """`entry` 에서 `max_hop` 홉 이내에 닿는 **다른 모듈** 함수들(폭 우선 방문 순).

    통합 시험의 대상은 "모듈 경계를 넘는 실행 경로"다. 그런데 계층 진입점
    (`main` · `*_Main` · ISR · 프로토콜 파서)은 대개 **같은 모듈의 내부 함수만 직접
    호출**하고, 실제 경계 횡단은 두세 홉 아래에서 일어난다. 직접 callee 만 보면 이런
    진입점이 통째로 탈락한다.

    실측(2026-08-14, KJPDS02_PV): 정본이 시험하는 통합지점 15개가 전부 이 경로로
    빠져 있었다 — `main`(직접 호출 2개가 둘 다 같은 모듈) · `s_SysMain_Init` ·
    `s_System_MainLoop` · `g_DrvIn_Main` · `g_DrvOut_Main` · `g_UDS_RDBI_Paser` ·
    `g_UDS_WDBI_Paser` · `g_UDS_SessionCtrl` · `SCI0_ISR` · `LinRawToTp` 등.
    """
    my = (module_of.get(entry) or "").lower()
    if not my:
        return []
    found: List[str] = []
    seen = {entry}
    frontier = [entry]
    for _ in range(max(1, max_hop)):
        nxt: List[str] = []
        for fn in frontier:
            for callee in calls_map.get(fn, []):
                if callee in seen:
                    continue
                seen.add(callee)
                if (module_of.get(callee) or "").lower() != my:
                    found.append(callee)
                else:
                    # 같은 모듈이면 더 내려가 본다 — 경계는 그 아래에 있다.
                    nxt.append(callee)
        frontier = nxt
        if not frontier:
            break
    return found


def _bfs_call_order(
    entry: str,
    calls_map: Dict[str, List[str]],
    limit: int,
    max_depth: Optional[int] = None,
) -> List[str]:
    """`entry` 로부터 **거리 순**(가까운 함수 먼저)으로 호출 그래프를 훑은 방문 순서.

    ## 왜 체인(DFS)과 다른 순서를 쓰나

    체인은 "이 통합 경로가 어디를 지나는가" 를 **서술**하는 것이라 깊이 우선이 정본과
    같은 모양이다. 반면 관측 대상(입력·기대)은 열 상한(입력 82 · 기대 113)에 묶여
    있어 **무엇을 먼저 담느냐가 곧 회수**다 — 후보가 상한의 3~8배나 된다
    (`s_SysMain_Init` 636개 · `main` 417개).

    실측(2026-08-14, KJPDS02_PV · 정본 입력 716 · 기대 910):

        수집 범위·순서          입력 회수   기대 회수
        현행(entry+직접 callee)   108        129
        경로 전체 · 깊이 우선      195        253
        경로 전체 · **거리 순**    213        281
        상한 없음(이론 최대)       324        393

    깊이 우선은 한 갈래로 멀리 내려가 상한을 써 버린다. 거리 순은 진입점 주변을 먼저
    채우는데, 정본이 적는 관측 대상이 거기 몰려 있다.
    """
    seen = {entry}
    order = [entry]
    queue = deque([(entry, 0)])
    while queue and len(order) < limit:
        cur, depth = queue.popleft()
        if max_depth is not None and depth >= max_depth:
            continue
        for callee in calls_map.get(cur, []):
            if callee in seen:
                continue
            seen.add(callee)
            order.append(callee)
            queue.append((callee, depth + 1))
            if len(order) >= limit:
                break
    return order


def _build_call_chain_nodes(
    entry: str,
    calls_map: Dict[str, List[str]],
    max_nodes: int,
) -> Tuple[List[str], int]:
    """호출 트리를 깊이 우선으로 편 방문 순서와, **상한에 막혀 못 넣은 함수 수**.

    정본 실측(KJPDS02_PV_SwITS v1.02): 체인은
    `Interface : main -> s_System_InitSequence -> s_SysMain_Init -> …` 처럼 **경로
    전체**이고 길이 분포는 1~5홉 22 · 6~20홉 12 · 21~50홉 8 · 50홉 초과 5 (최대 92),
    같은 함수가 두 번 나오는 TC 는 **0/54** 다(= visited 기반 전개).
    우리는 `entry + 직접 cross callee 4개` = 최대 5홉만 적고 있었다.

    ⚠ 상한을 넘으면 **자르되 몇 개를 못 실었는지 돌려준다**. 조용히 자르면 읽는 쪽이
    "이 경로가 전부" 로 읽는다(이 저장소가 반복해서 물린 절단-침묵 패턴).
    """
    seen = {entry}
    order = [entry]
    dropped: set = set()
    stack: List[Any] = [iter(calls_map.get(entry, []))]
    while stack:
        try:
            callee = next(stack[-1])
        except StopIteration:
            stack.pop()
            continue
        if callee in seen:
            continue
        if len(order) >= max_nodes:
            # 상한 초과분은 **고유 함수 단위**로 센다(같은 함수를 여러 번 만나도 1).
            dropped.add(callee)
            continue
        seen.add(callee)
        order.append(callee)
        stack.append(iter(calls_map.get(callee, [])))
    return order, len(dropped)


# ---------------------------------------------------------------------------
# Core: integration flow collection
# ---------------------------------------------------------------------------

# 통합 흐름 상한. 폭주 방지용 안전밸브이지 "이만큼만 시험하면 된다"는 뜻이 아니다.
#
# ⚠ 이 값은 전이 판정(`_reach_cross_module`) 이전에 정해졌다. 실측(2026-08-14,
#   KJPDS02_PV 1,157함수): 후보가 239 → **367** 로 늘었고 그중 **306개가 SwUDS 등재**다.
#   즉 기본값 120 이면 설계가 인정한 통합 지점의 절반 이상이 규격에서 빠진다.
#   값을 여기서 올리지 않는 이유는 프로젝트마다 규모가 다르기 때문이다 — 대신 잘린
#   내역을 **보이게** 만들어 두었다: 경고 2줄(총량 / 그중 등재분) + 품질 리포트
#   `integration_flow_coverage.dropped_in_design_doc_count`. 그 수치가 0 이 아니면
#   호출자가 `max_flows` 를 올릴 근거가 된다.
_DEFAULT_MAX_FLOWS = 120

# 직접 callee 가 전부 같은 모듈일 때 **몇 홉까지 내려가** 경계를 찾을지.
# (`_reach_cross_module` — 계층 진입점은 자기 모듈 안쪽만 직접 호출한다)
_CROSS_REACH_HOPS = 3
# 호출 체인 한 줄에 담을 함수 수 상한. 정본 실측 최대는 92 홉이다.
_MAX_CHAIN_NODES = 100

# 관측 대상(입력·기대)을 모을 때 훑는 경로 함수 수 상한.
_VAR_SCAN_NODES = 200

# 관측 대상을 모을 **깊이** 상한. 회수와 정밀도의 균형점이다.
#
# 실측(2026-08-14, KJPDS02_PV · 정본 입력 716 · 기대 910 · 열 상한 82/113):
#
#     깊이   입력 일치 / 총량 (정밀도)     기대 일치 / 총량 (정밀도)
#      1        108 /  471 (22.9%)          129 /  475 (27.2%)
#    **2**    **187 /  888 (21.1%)**      **219 /  951 (23.0%)**
#      3        205 / 1161 (17.7%)          268 / 1274 (21.0%)
#      5+       213 / 1295 (16.4%)          281 / 1408 (20.0%)
#
# 깊이 2 는 회수의 88%(입력)·78%(기대)를 얻으면서 총량은 무제한 대비 31% 적고,
# 정밀도는 1홉 수준을 지킨다. 더 멀리 가면 **먼 전역이 가까운 것을 열에서 밀어낸다**
# — 열이 82칸뿐이라 "많이 담기" 는 곧 "잘못 담기" 다.
# ⚠ 정밀도 20%대가 상한인 이유: 정본은 VectorCAST 실행 결과로 관측 대상을 고르고
#   우리는 정적 호출 그래프만 본다. 실행 데이터 없이 더 좁히면 회수가 먼저 무너진다.
_VAR_SCAN_DEPTH = 2

# `collect_integration_flows` 가 `stats_out` 에 싣는 **흐름 축** 키.
# ⚠ 생산자와 품질 리포트가 이 목록 **하나**를 본다. 예전엔 리포트가 자기 화이트리스트를
#   따로 들고 있어 생산자에 키를 추가해도 조용히 버려졌다 — 같은 결함을 두 번 겪었다.
#   여기에 없는 흐름 키는 리포트에 도달하지 않으므로, 키를 늘리면 **여기부터** 고칠 것.
#   (가드: `test_generators_sits.py::TestFlowStatsReachTheReport`)
_FLOW_COV_KEYS: Tuple[str, ...] = (
    "total_flows_found", "flows_emitted", "flows_dropped", "flow_emit_pct", "max_flows",
    "dropped_safety_related_count", "dropped_asil_distribution", "dropped_entry_fns",
    "dropped_in_design_doc_count",
    "transitive_entries", "transitive_entries_emitted", "cross_reach_hops",
    "chain_truncated_flows", "chain_max_nodes",
    # TC ID 축 — 설계 ID(SwUDS) 유래인지 파싱 순번인지
    "design_id_hits", "design_id_lookups", "design_id_map_entries",
    # Related ID 축 — 진입 함수 자신인지 호출 트리 아래인지, 칸 상한에 잘렸는지
    "related_chain_flows", "related_chain_ids", "related_chain_depth",
    "related_truncated_ids",
    # 요구 ID 축 — 이 칸에는 **안 실리는** 링크. 실리지 않는다고 세지 않으면
    # "그런 링크가 없다" 와 구별되지 않는다.
    "req_id_flows", "req_id_total",
    # 배열 원소 펼침 축 — 정본과 같은 입도로 냈는가, 예산에 걸려 못 펼쳤는가
    "array_expanded_inputs", "array_expanded_expected", "array_elements_emitted",
    "array_skipped_budget", "array_size_map_entries", "array_struct_types",
    # 관측 대상 선별 축 — 열이 찬 것과 후보를 못 담은 것이 구별되게
    "var_selection_basis", "var_scan_depth", "var_scan_nodes_max",
    "var_candidates_input", "var_candidates_expected",
    "var_budget_cut_input", "var_budget_cut_expected",
    # FI 축 — 0 이 "요청 없음"인지 "요청했는데 못 냄"인지 구분되게
    "fi_emitted", "fi_requested", "fi_unresolved",
    # 근거 시트(전략 / Related_ID) 산출 실적 — 시트가 비어도 **왜 비었는지** 보이게
    "strategy_blocks", "strategy_nodes", "strategy_nodes_dropped",
    "strategy_blocks_truncated",
    "relid_check_rows", "relid_tidy_rows", "relid_index_rows",
)

# 그중 **손실 축** — "몇 개를 못 실었나". 요약 표면(영향도 카드)은 전 키를 싣기엔
# 좁으므로 이 부분집합만 싣는다.
#
# ⚠ 새 손실 키를 만들면 **여기에도 넣는다**. 소비처가 손으로 고르면 키를 늘려도 조용히
#   빠진다 — 이 저장소가 같은 결함을 세 층에서 겪었다(생산자→리포트, 리포트→평가기,
#   리포트→영향도). 이름 규약(`dropped`/`truncated`/`skipped`/`unresolved`/`cut`)을
#   따르는 키가 여기 없으면 `test_sits_var_selection_axis.py` 가 실패한다.
_FLOW_LOSS_KEYS: Tuple[str, ...] = (
    "flows_dropped", "dropped_safety_related_count", "dropped_in_design_doc_count",
    # ⚠ 아래 둘은 스칼라가 아니다(분포 dict · 함수명 list). 그래도 **손실 축이라**
    #   여기 있어야 한다 — 소비처가 모양 때문에 빼면 "무엇을 잃었는지" 가 사라진다.
    #   요약 표면은 dict 는 `k=v` 로, list 는 건수로 접어서 싣는다.
    "dropped_asil_distribution", "dropped_entry_fns",
    "chain_truncated_flows", "related_truncated_ids", "array_skipped_budget",
    "fi_unresolved", "strategy_nodes_dropped", "strategy_blocks_truncated",
    "var_budget_cut_input", "var_budget_cut_expected",
)
# 위 규약을 검사할 때 쓰는 어휘 — 가드와 문서가 같은 목록을 본다.
_FLOW_LOSS_NAME_MARKERS: Tuple[str, ...] = (
    "dropped", "truncated", "skipped", "unresolved", "_cut_",
)

_ASIL_RANK: Dict[str, int] = {"D": 0, "C": 1, "B": 2, "A": 3, "QM": 4}


def _asil_rank(asil: Any) -> int:
    """ASIL 선별 우선순위 — 값이 작을수록 먼저 남긴다. 미상 등급은 QM 뒤로 보낸다."""
    return _ASIL_RANK.get(str(asil or "").strip().upper(), len(_ASIL_RANK))


def _select_flows_within_cap(
    candidates: List[Dict[str, Any]],
    max_flows: Optional[int],
    stats_out: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """`max_flows` 캡을 적용하되 **무엇이 잘렸는지 남기고**, 안전등급 높은 쪽을 살린다.

    예전엔 수집 루프가 캡에 닿으면 그냥 `break` 했다. 정렬 키가 함수명 알파벳순이라
    어느 흐름이 살아남는지가 **안전등급과 무관**하게 정해졌고, 잘렸다는 사실 자체가
    어디에도 안 남았다(로그·품질 리포트 모두). 실측(KJPDS02 계열 900함수):
    통합 흐름 145개 중 25개가 조용히 사라졌고 그 중 7개가 ASIL A 였다 —
    같은 모듈(Sys_UDS_LinComp)이 알파벳 경계에서 두 동강 났다.

    출력 순서는 알파벳 그대로 둔다(선별만 안전우선). 문서 행 순서를 흔들지 않기 위해서다.
    """
    total = len(candidates)
    kept = candidates
    dropped: List[Dict[str, Any]] = []

    if max_flows is not None and 0 <= max_flows < total:
        indexed = list(enumerate(candidates))
        # (등급, **설계 문서 등재 여부**, 알파벳 순번).
        #
        # ⚠ 가운데 항이 없으면 이 프로젝트에서는 사실상 **알파벳순 = 임의**다(후보가
        #   거의 전부 QM 이다). 실측(2026-08-14, KJPDS02_PV): 후보 367 · 캡 200 에서
        #   정본이 시험하는 통합지점 34개 중 **30개만** 살아남았고, 밀려난 것 중에는
        #   직전 라운드까지 잘 나오던 `s_Ap_ExecuteControlFunctions`·
        #   `s_SystemHashCalculate` 가 있었다 — 후보가 늘면 캡이 **기존 정답을 밀어낸다**.
        #
        #   SwUDS 등재 여부를 키에 넣으면 같은 캡에서 **34/34** 가 생존한다. 근거:
        #   정본 통합지점은 34개 전부가 SwUDS 에 등재돼 있고(100%), 후보 전체로는
        #   83.4%(306/367)라 미등재 61개가 먼저 밀린다. 필터로 쓰면 약하지만
        #   **정렬 키로는 정확**하다. (호출 트리 크기·이름 패턴은 정본 보존율이
        #   각각 35%·56% 로 오히려 나빠 기각했다.)
        ranked = sorted(indexed, key=lambda t: (
            _asil_rank(t[1].get("asil")),
            0 if t[1].get("in_design_doc") else 1,
            t[0],
        ))
        keep_idx = {i for i, _ in ranked[:max_flows]}
        kept = [c for i, c in indexed if i in keep_idx]
        dropped = [c for i, c in indexed if i not in keep_idx]

    dist: Dict[str, int] = {}
    for c in dropped:
        # ⚠ 근거 없는 것을 `QM`(= 안전요구 면제)으로 세면 "안전 관련은 안 잘렸다" 는
        #   거짓 안심이 된다. 모르는 건 모른다고 센다.
        key = str(c.get("asil") or "").strip() or "(근거없음)"
        dist[key] = dist.get(key, 0) + 1
    safety_dropped = sum(
        n for a, n in dist.items() if a.strip().upper() in ("A", "B", "C", "D")
    )

    if stats_out is not None:
        stats_out.update({
            "total_flows_found": total,
            "max_flows": max_flows,
            "flows_emitted": len(kept),
            "flows_dropped": len(dropped),
            "flow_emit_pct": round(len(kept) / max(total, 1) * 100, 1),
            "dropped_entry_fns": [str(c.get("fn_name") or "") for c in dropped],
            "dropped_asil_distribution": dist,
            "dropped_safety_related_count": safety_dropped,
            # 잘린 것 중 **설계 문서에 등재된** 함수 수. 0 이 아니면 캡이 설계가
            # 인정한 단위를 먹고 있다는 뜻이므로 캡 값을 다시 볼 신호다.
            "dropped_in_design_doc_count": sum(1 for c in dropped if c.get("in_design_doc")),
        })

    if dropped:
        _logger.warning(
            "SITS: 통합 흐름 %d개 중 %d개만 생성한다 — max_flows=%s 캡으로 %d개 제외"
            "(안전관련 ASIL A~D %d개 포함). 제외된 흐름은 시험 규격에 **존재하지 않는다**. "
            "예: %s",
            total, len(kept), max_flows, len(dropped), safety_dropped,
            ", ".join(str(c.get("fn_name") or "") for c in dropped[:5]),
        )
    _doc_dropped = [c for c in dropped if c.get("in_design_doc")]
    if _doc_dropped:
        # 등재분까지 먹었다는 건 캡이 **설계가 인정한 단위**를 자르고 있다는 뜻이다.
        # 실측(2026-08-14, KJPDS02_PV): 후보 367 중 등재 306, 캡 200 → 등재분 106개가
        # 잘렸고 그 안에 정본 통합지점 3개(`s_System_MainLoop`·`s_SystemHashCalculate`·
        # `s_SysEepromCtrl_WriteData_Direct`)가 있었다. 캡 값을 다시 볼 신호다.
        _logger.warning(
            "SITS: 그중 %d개는 **SwUDS 에 등재된** 통합 지점이다 — 캡(%s)이 설계가 인정한 "
            "단위를 자르고 있다. 예: %s",
            len(_doc_dropped), max_flows,
            ", ".join(str(c.get("fn_name") or "") for c in _doc_dropped[:5]),
        )

    return kept


def collect_integration_flows(
    function_details: Dict[str, Dict[str, Any]],
    # ⚠ `None` = 캡 없음. `_select_flows_within_cap` 은 처음부터 `Optional[int]` 를
    #   받았는데 여기만 `int` 로 좁아, 캡 **전** 총량을 재려는 호출자가 타입상 막혔다.
    #   (총량은 캡 전에만 보인다 — 결과 길이로 되짚으면 절단을 못 본다.)
    max_flows: Optional[int] = _DEFAULT_MAX_FLOWS,
    stats_out: Optional[Dict[str, Any]] = None,
    sds_map: Optional[Dict[str, Dict[str, str]]] = None,
    uds_swcom_map: Optional[Dict[str, List[str]]] = None,
    uds_asil_map: Optional[Dict[str, str]] = None,
    uds_related_map: Optional[Dict[str, List[str]]] = None,
    # ⚠ 신규 인자는 **맨 끝**에 붙인다(위치 인자 호출부가 조용히 다른 값에 바인딩된다).
    globals_info_map: Optional[Dict[str, Dict[str, str]]] = None,
    struct_members: Optional[Dict[str, Dict[str, str]]] = None,
) -> List[Dict[str, Any]]:
    """Identify cross-module integration flows from function call graph.

    통합 흐름 = **모듈 경계를 넘는 실행 경로의 진입점**이다. 경계를 직접 넘든
    (`cross_via="direct"`), 같은 모듈 안쪽을 몇 홉 지나 넘든(`"transitive"`) 자격은
    같다 — 계층 진입점(`main`·`*_Main`·ISR·프로토콜 파서)은 후자의 모양이고, 직접
    callee 만 보던 판정에서 **정본 통합지점 15개가 통째로 빠졌다**(`_reach_cross_module`).

    Returns list of flow dicts:
      { flow_id, entry_fn, call_chain, cross_via, cross_calls, chain_dropped,
        functions, module_name, swcom_id, input_vars, expected_vars, asil,
        related_ids, synthetic_related_ids }

    `call_chain` 은 정본과 같이 **경로 전체**다(`_build_call_chain_nodes`).
    `chain_dropped` 가 0 이 아니면 그 경로는 상한에 걸려 잘린 것이다.

    Args:
        sds_map: Related ID 보강용 SDS 파티션 맵. None 이면 저장소 `docs/` 글롭
            (`_load_default_sds_map`)으로 폴백하는데 이는 **프로젝트 무관**이다
            — `sts.py`/`suts.py` 는 이미 같은 파라미터를 갖고 있었고 여기만 없어서
            **호출자가 대상 프로젝트의 SDS 를 줄 방법 자체가 없었다.**
        uds_swcom_map: `{함수명(소문자): [SwCom_NN]}` — Related 칸 SwCom 의 **유일한
            문서 근거**(`load_uds_swcom_map`). 주지 않으면 순번 합성 ID 로 내려간다.
        globals_info_map / struct_members: 배열 **선언 크기**의 출처
            (`generate_uds_source_sections` 의 `globals_info_map`/`struct_member_arrays`).
            정본은 배열을 `g_sys_error_his[0]`…`[15]` 처럼 **원소 단위**로 적는다 —
            안 주면 base 한 칸으로 나가 정본 열과 입도가 어긋난다.

    `stats_out` 를 주면 캡 절단 내역(총 후보 수·제외 수·제외분 ASIL 분포)과
    **SDS 보강 실적**(`sds_*` 키)을 채운다. 소비처에서 결과 길이로 되짚으면 절단을 못
    본다 — 캡 **전** 총량이 여기서만 보인다.
    """
    # ── SDS 보강 계측 ──────────────────────────────────────────────────────
    # ⚠ 실측(2026-07-31): 이 보강은 **한 건도 산출한 적이 없다.**
    #   `_load_default_sds_map()` 이 주는 맵의 값 스키마는
    #   `{kind, description, related, asil, component_description, canonical}` 인데
    #   여기서는 `entry.get("swcom") or entry.get("component")` 를 읽었다 — **없는
    #   필드**라 항상 None 이고, 그 사실이 `except Exception: pass` 에 묻혀 있었다.
    #   (같은 맵을 쓰는 `sts.py::_lookup_sds_related_ids` 는 실재 필드 `related` 를 읽는다.)
    #   대체 필드를 **추측하지 않는다** — 틀린 SwCom 을 추적성 열에 넣는 건 0 건보다 나쁘다.
    #   대신 0 을 **보이게** 만든다: 아래 카운터가 `stats_out` 으로 나간다.
    _sds_lookups = 0        # 조회 시도한 함수 수
    _sds_key_hits = 0       # 맵에서 키가 잡힌 수
    _sds_swcom_hits = 0     # 실제로 SwCom 을 얻은 수 — 이 맵으로는 **구조적으로 0**(위 참조)
    # ── SwUDS 축 — Related 칸의 실제 소스(`load_uds_related_map` docstring 참조) ──
    _uds_swcom_lookups = 0  # 조회 시도한 함수 수
    _uds_swcom_hits = 0     # SwUDS 에서 Related ID 를 실제로 얻은 함수 수
    _uds_swcom_ids = 0      # 그렇게 얻은 ID 총 개수(함수당 다중 가능)
    # 진입 함수 자신이 아니라 **호출 트리 아래**에서 온 ID — 근거의 거리가 다르므로
    # 따로 센다(칸만 보면 둘을 구별할 수 없다).
    _chain_rel_flows = 0
    _chain_rel_ids = 0
    _related_truncated = 0  # 칸 상한에 걸려 잘라낸 ID 수
    # 요구 ID 축 — Related 칸의 값이 아니라서 빼지만(아래 ②) 버리지는 않는다.
    _req_id_flows = 0       # 요구 ID 링크를 가진 흐름 수
    _req_id_total = 0       # 그렇게 얻은 요구 ID 총 개수
    # 배열 원소 펼침 — 예산에 걸려 **못 펼친** 것도 센다(0 만 보면 "펼칠 게 없었다"
    # 인지 "배선이 끊겼다"인지 구분이 안 된다).
    _arr_expanded_in = 0
    _arr_expanded_exp = 0
    _arr_skipped = 0
    _arr_elements = 0
    # 관측 대상 선별 축 — **후보가 몇 개였는지**를 센다.
    #
    # 정본은 관측 대상을 VectorCAST 실행 결과에서 고르고 우리는 정적 호출 그래프만
    # 본다. 깊이는 이미 최적점에서 멈춰 있고(`_VAR_SCAN_DEPTH` 주석의 깊이별 실측 표),
    # 남은 격차는 **더 담아서** 줄지 않는다 — 후보가 열 상한의 3~8배라 더 담는 건 곧
    # 잘못 담는 것이다. 그래서 이 축은 닫지 않고 **보이게** 만든다: 산출물만 보면
    # 82칸이 찬 것과 후보 400개 중 82개만 실린 것이 구별되지 않는다.
    _var_cand_in = 0
    _var_cand_exp = 0
    _var_budget_cut_in = 0
    _var_budget_cut_exp = 0
    # 배열 선언 크기는 흐름마다 안 바뀐다 — 루프 **밖에서** 한 번만 만든다
    # (SUTS 가 같은 이유로 unit 루프 밖에 둔다: 전역 1,525 × 흐름 367 헛돔 방지).
    _gim = globals_info_map or {}
    _smem = struct_members or {}
    _root_sizes: Dict[str, Tuple[int, ...]] = {}
    if _gim:
        from generators.suts import _decl_dims_from_array_field, _dim_product
        for _rn, _ri in _gim.items():
            _rd = _decl_dims_from_array_field(str((_ri or {}).get("array") or ""))
            if _rd and _dim_product(_rd) > 1:
                _root_sizes[str(_rn)] = _rd
    _sds_source = "argument" if sds_map is not None else "repo_docs_glob"
    if sds_map is None:
        sds_map = _load_default_sds_map()
    # Build name → info lookup
    name_to_info: Dict[str, Dict[str, Any]] = {}
    for fid, info in function_details.items():
        if isinstance(info, dict):
            name_to_info[str(info.get("name") or "")] = info

    # Set of all project function names (lower-case) for ISR-artefact filtering
    _fn_name_set: set = {n.lower() for n in name_to_info if n}

    # 전이 판정·체인 전개용 호출 그래프. **프로젝트 함수끼리의 호출만** 남긴다
    # (memset/printf 같은 외부 심볼은 파싱 그래프에 없고 통합 경로도 아니다).
    _module_of: Dict[str, str] = {n: _get_module_name(i) for n, i in name_to_info.items() if n}
    # ⚠ 전략 시트도 같은 그래프를 쓴다 — 필터가 갈라지면 규격 시트의 체인과 전략 시트의
    #   트리가 서로 다른 그래프가 되므로 **단일 빌더**를 공유한다(`_build_calls_map`).
    _calls_map: Dict[str, List[str]] = _build_calls_map(function_details)
    _transitive_entries = 0       # 직접 경계 없이 전이로만 자격을 얻은 진입점 수
    _chain_truncated_total = 0    # 체인이 상한에 걸려 잘린 흐름 수(침묵 금지)

    swcom_counter: Dict[str, int] = {}
    flows: List[Dict[str, Any]] = []
    seen_entries: set = set()

    # Sort by name for deterministic output
    sorted_items = sorted(
        [(fid, info) for fid, info in function_details.items() if isinstance(info, dict)],
        key=lambda x: str(x[1].get("name") or ""),
    )

    # ── Pass 1: 자격 판정만 (싸다) — 후보를 **전부** 모은다 ──────────────────
    # 예전엔 이 루프가 `len(flows) >= max_flows` 에서 break 했다. 그러면 캡 이후의
    # 후보는 세어지지도 않아 "몇 개가 잘렸는지" 를 아무도 알 수 없다. 자격 판정
    # (calls_list 유무 + cross-module callee 유무)은 dict 조회뿐이라 전량 수행해도 싸고,
    # 비싼 변수/기대값 구성은 Pass 2 에서 선별된 것에만 한다 = 기존 비용과 동일.
    candidates: List[Dict[str, Any]] = []
    for fid, info in sorted_items:
        fn_name = str(info.get("name") or "")
        if not fn_name or fn_name in seen_entries:
            continue

        calls_list = list(info.get("calls_list") or [])
        if not calls_list:
            continue

        my_module = _get_module_name(info)

        # Find calls that cross module boundaries.
        # Only include callees that are known project functions (present in name_to_info).
        # External library / OS calls (memset, printf, …) are excluded because they are
        # not in the parsed function graph and do not represent software integration flows.
        cross_calls: List[str] = []
        for callee in calls_list:
            callee_info = name_to_info.get(callee)
            if callee_info:
                callee_module = _get_module_name(callee_info)
                if callee_module and callee_module.lower() != my_module.lower():
                    cross_calls.append(callee)

        # 직접 callee 가 전부 같은 모듈이어도, **몇 홉 아래에서** 경계를 넘으면 그건
        # 통합 지점이다 — 계층 진입점(`main`·`*_Main`·ISR·파서)의 전형적인 모양이다.
        # 이 폴백이 없을 때 정본 통합지점 15개가 통째로 빠졌다(`_reach_cross_module`).
        _cross_via = "direct"
        if not cross_calls:
            cross_calls = _reach_cross_module(
                fn_name, _calls_map, _module_of, _CROSS_REACH_HOPS)
            if not cross_calls:
                continue
            _cross_via = "transitive"
            _transitive_entries += 1

        seen_entries.add(fn_name)

        # ⚠ 예전엔 `str(info.get("asil") or "QM")` + `TBD → QM` 이었다. 소스 주석만 보고,
        #   근거가 없으면 **비안전으로 단정**한 것이다(실측 260건). 입력문서를 본다.
        _cand_asil = _resolve_flow_asil(
            fn_name, info, uds_asil_map=uds_asil_map,
            uds_swcom_map=uds_swcom_map, sds_map=sds_map)

        # ⚠ SwCom 은 **후보 전체**(알파벳순)에 대해 여기서 부여한다. 예전엔 캡 안쪽
        # 루프에서 부여돼 **ID 가 캡 값에 의존**했다 — 캡을 바꾸면 같은 모듈이 다른
        # SwCom 을 받는다. 후보 전체 기준이면 캡·선별 정책이 바뀌어도 ID 가 고정된다.
        # (실측: 이 프로젝트는 캡 120/무제한 어느 쪽도 모듈 29개·ID 변동 0건 = 무해한 변경)
        candidates.append({
            "fid": fid,
            "info": info,
            "fn_name": fn_name,
            "my_module": my_module,
            "cross_calls": cross_calls,
            # 경계를 **직접** 넘는지, 몇 홉 아래에서 넘는지. 같은 `cross_calls` 라도
            # 근거가 다르므로 산출물·리포트가 구별할 수 있어야 한다.
            "cross_via": _cross_via,
            # 설계 문서(SwUDS)에 등재된 함수인가 — 캡 선별 키
            # (`_select_flows_within_cap` 주석의 실측 근거 참조).
            "in_design_doc": fn_name.lower() in (uds_swcom_map or {}),
            "asil": _cand_asil,
            "swcom_id": _infer_swcom_id(my_module, swcom_counter),
        })

    # ── 캡 적용: 안전등급 높은 쪽을 남기고, 잘린 내역을 stats_out 에 남긴다 ────
    selected = _select_flows_within_cap(candidates, max_flows, stats_out)

    # ── Pass 2: 선별된 후보만 비싼 구성 ──────────────────────────────────────
    for _cand in selected:
        fid = _cand["fid"]
        info = _cand["info"]
        fn_name = _cand["fn_name"]
        my_module = _cand["my_module"]
        cross_calls = _cand["cross_calls"]

        # ── Call chain — 정본은 **경로 전체**를 적는다 ────────────────────────
        # 정본 실측(KJPDS02_PV_SwITS v1.02): `Interface : main -> s_System_InitSequence
        # -> …` 로 최대 92 홉, 같은 함수 재등장 0/54(= visited 전개). 우리는
        # `entry + 직접 cross callee 4개` = 최대 5 홉만 적어, 통합 경로가 아니라
        # **첫 갈림길 몇 개**만 보여 주고 있었다.
        _chain_nodes, _chain_dropped = _build_call_chain_nodes(
            fn_name, _calls_map, _MAX_CHAIN_NODES)
        call_chain = " -> ".join(_chain_nodes)
        if _chain_dropped:
            _chain_truncated_total += 1

        # Collect variables
        # Each entry stored as (display_name, annotated_raw) so that
        # _infer_boundary_values can use the explicit C type token.
        inputs_raw = list(info.get("inputs") or [])
        outputs_raw = list(info.get("outputs") or [])
        globals_g = list(info.get("globals_global") or [])
        globals_s = list(info.get("globals_static") or [])

        # Build (var_name, annotated_raw) pairs — filter out entries whose
        # cleaned name matches a known function name (ISR stub artefact).
        input_pairs: List[Tuple[str, str]] = []
        # Pointer parameters of the entry function are observable I/O.
        # _lw_parse_params strips '*' from var names, so detect via prototype instead.
        ptr_out_pairs: List[Tuple[str, str]] = []
        _proto = str(info.get("prototype") or "")
        _ptr_params: set = set()
        if _proto and "(" in _proto:
            _param_str = _proto.split("(", 1)[1].rsplit(")", 1)[0]
            for _pp in _param_str.split(","):
                _pp = _pp.strip()
                if "*" in _pp and "const" not in _pp.lower():
                    # Extract variable name (last token, stripped of *)
                    _pparts = _pp.split()
                    if _pparts:
                        _pname = _pparts[-1].strip("*&;")
                        if _pname:
                            _ptr_params.add(_pname.lower())
        for raw in inputs_raw[:20]:
            vn = _clean_var_name(raw)
            # ⚠ 빈 이름 가드. 예전 `_clean_var_name` 은 무엇을 받든 문자열을 냈기에
            #   (`raw[:40]` 폴백) 이 검사가 필요 없었다 — 이제는 못 뽑으면 빈 값이다.
            if vn and vn.lower() not in _fn_name_set and vn not in {p[0] for p in input_pairs}:
                input_pairs.append((vn, raw))
                # Pointer param (*) is also an out-parameter
                if vn.lower() in _ptr_params:
                    ptr_out_pairs.append((vn, raw))

        # If entry has no inputs, aggregate callee inputs as integration-level inputs
        if not input_pairs:
            for callee in cross_calls[:4]:
                callee_info = name_to_info.get(callee)
                if callee_info:
                    # Build pointer param set from callee prototype
                    _cproto = str(callee_info.get("prototype") or "")
                    _c_ptr_params: set = set()
                    if _cproto and "(" in _cproto:
                        _cps = _cproto.split("(", 1)[1].rsplit(")", 1)[0]
                        for _cpp in _cps.split(","):
                            _cpp = _cpp.strip()
                            if "*" in _cpp and "const" not in _cpp.lower():
                                _cpparts = _cpp.split()
                                if _cpparts:
                                    _cpname = _cpparts[-1].strip("*&;")
                                    if _cpname:
                                        _c_ptr_params.add(_cpname.lower())
                    for craw in (callee_info.get("inputs") or [])[:6]:
                        cvn = _clean_var_name(craw)
                        if cvn and cvn.lower() not in _fn_name_set and cvn not in {p[0] for p in input_pairs}:
                            input_pairs.append((cvn, craw))
                            if cvn.lower() in _c_ptr_params:
                                ptr_out_pairs.append((cvn, craw))
                if len(input_pairs) >= _MAX_INPUT_PARAMS:
                    break

        # Globals as additional observed inputs
        for g in (globals_g + globals_s)[:15]:
            gn = _clean_global_var_name(g)
            if gn and gn.lower() not in _fn_name_set and gn not in {p[0] for p in input_pairs}:
                input_pairs.append((gn, g))

        # ── 관측 대상을 **경로 전체**로 넓힌다 ────────────────────────────────
        # 체인은 R2 에서 경로 전체(최대 100홉)로 폈는데 변수 수집은 `entry + 직접
        # callee 4개` 에 그대로 묶여 있었다 — 두 축이 어긋난 채였다. 실측(2026-08-14):
        # 정본 미달의 **98.7%(입력)·94.5%(기대)** 가 "뿌리조차 없음" 이었고, 그 대부분이
        # 경로상 함수들의 전역이다(`main` TC 에 정본이 적는 `u8g_Cpu_OnLvdStatusChanged_F`
        # 처럼). 회수: 입력 108 → 213 · 기대 129 → 281 (정본 716 / 910 기준).
        #
        # 순서는 **거리 순**(`_bfs_call_order`)이다 — 후보가 상한의 3~8배라 무엇을 먼저
        # 담느냐가 곧 회수이고, 깊이 우선은 한 갈래로 멀리 내려가 상한을 써 버린다.
        _var_nodes = _bfs_call_order(
            fn_name, _calls_map, _VAR_SCAN_NODES, max_depth=_VAR_SCAN_DEPTH)
        #
        # ⚠ 예산이 차도 **멈추지 않고 후보를 끝까지 센다**. 담는 규칙은 그대로다
        #   (`input_pairs` 결과는 이전과 동일) — 세기만 추가한다. 예전엔 상한에서
        #   `break` 해 버려 "82칸을 채웠다" 와 "후보 400 중 82 만 담았다" 가 산출물에서
        #   같은 모양이었다.
        _cand_in: set = {p[0] for p in input_pairs}
        for _node in _var_nodes:
            _ni = name_to_info.get(_node)
            if not _ni or _node == fn_name:
                continue      # entry 자신은 위에서 이미 훑었다
            for _g in ((_ni.get("globals_global") or []) + (_ni.get("globals_static") or [])):
                _gn = _clean_global_var_name(_g)
                if _gn and _gn.lower() not in _fn_name_set and _gn not in _cand_in:
                    _cand_in.add(_gn)
                    if len(input_pairs) < _MAX_INPUT_PARAMS:
                        input_pairs.append((_gn, _g))
        _var_cand_in += len(_cand_in)
        _var_budget_cut_in += max(0, len(_cand_in) - len(input_pairs))

        input_vars: List[str] = [p[0] for p in input_pairs[:_MAX_INPUT_PARAMS]]
        # Keep annotated raws for type inference
        input_raws: List[str] = [p[1] for p in input_pairs[:_MAX_INPUT_PARAMS]]
        input_vars, input_raws, _n_in, _n_skip, _n_el = _expand_arrays(
            input_vars, input_raws, _MAX_INPUT_PARAMS, _gim, _smem, _root_sizes)
        _arr_expanded_in += _n_in
        _arr_skipped += _n_skip
        _arr_elements += _n_el

        # Expected: own outputs + pointer out-params + callee outputs + callee globals
        exp_pairs: List[Tuple[str, str]] = []
        for raw in outputs_raw[:10]:
            vn = _clean_var_name(raw)
            if vn and vn.lower() not in _fn_name_set:
                exp_pairs.append((vn, raw))
        # Pointer out-params of entry function are expected observables
        for vn, raw in ptr_out_pairs:
            if vn not in {p[0] for p in exp_pairs}:
                exp_pairs.append((vn, raw))
        # ⚠ 정본은 이 칸에 **변수 이름만** 적는다 — `함수명() 변수` 같은 접두를 쓰지
        #   않는다(정본 기대 1,172칸 중 괄호 접두 0건). 접두를 붙이면 같은 변수가 부르는
        #   함수마다 다른 이름이 되어, 정본과 대조할 때 **한 칸도 맞지 않는다**.
        #   어느 함수가 그 전역을 건드리는지는 Interface 체인이 이미 말해 준다.
        for callee in cross_calls[:5]:
            callee_info = name_to_info.get(callee)
            if callee_info:
                for v in (callee_info.get("outputs") or [])[:5]:
                    vn = _clean_var_name(v)
                    if vn and vn.lower() not in _fn_name_set and vn not in {p[0] for p in exp_pairs}:
                        exp_pairs.append((vn, v))
                # Callee globals as observable side-effect outputs
                for g in ((callee_info.get("globals_global") or []) + (callee_info.get("globals_static") or []))[:4]:
                    gn = _clean_global_var_name(g)
                    if gn and gn.lower() not in _fn_name_set and gn not in {p[0] for p in exp_pairs}:
                        exp_pairs.append((gn, g))

        # 입력과 같은 이유로 기대도 **경로 전체**를 본다(위 입력부 주석 참조).
        # 회수: 기대 129 → 281 (정본 910 기준).
        # 입력과 같은 이유로 후보를 끝까지 센다(위 입력부 주석 참조).
        _cand_exp: set = {p[0] for p in exp_pairs}
        for _node in _var_nodes:
            _ni = name_to_info.get(_node)
            if not _ni or _node == fn_name:
                continue
            for _raw, _cleaner in (
                [(x, _clean_var_name) for x in (_ni.get("outputs") or [])]
                + [(x, _clean_global_var_name)
                   for x in ((_ni.get("globals_global") or []) + (_ni.get("globals_static") or []))]
            ):
                _nm = _cleaner(_raw)
                if _nm and _nm.lower() not in _fn_name_set and _nm not in _cand_exp:
                    _cand_exp.add(_nm)
                    if len(exp_pairs) < _MAX_EXP_PARAMS:
                        exp_pairs.append((_nm, _raw))
        _var_cand_exp += len(_cand_exp)
        _var_budget_cut_exp += max(0, len(_cand_exp) - len(exp_pairs))

        # If still no expected vars, mine global writes from logic_flow conditions
        if not exp_pairs:
            _GLOBAL_WRITE_RE = re.compile(
                r"\b(g_\w+|gs_\w+|g[A-Z]\w+)\s*(?:\[[\w\s+\-*]+\])?\s*=",
            )
            for src_fn in [fn_name] + list(cross_calls[:4]):
                src_info = name_to_info.get(src_fn) if src_fn != fn_name else info
                if not src_info:
                    continue
                for node in (src_info.get("logic_flow") or [])[:20]:
                    for m in _GLOBAL_WRITE_RE.finditer(str(node.get("text", "") + node.get("condition", ""))):
                        gname = m.group(1)
                        # 위와 같은 이유로 `함수명()` 접두를 붙이지 않는다(정본은 0건).
                        if gname not in {p[0] for p in exp_pairs}:
                            exp_pairs.append((gname, gname))
                if len(exp_pairs) >= _MAX_EXP_PARAMS:
                    break

        expected_vars: List[str] = [p[0] for p in exp_pairs[:_MAX_EXP_PARAMS]]
        expected_raws: List[str] = [p[1] for p in exp_pairs[:_MAX_EXP_PARAMS]]
        # ⚠ 입력·기대 **양쪽** 이다. 한쪽만 펼치면 한 행 안에서 같은 변수가 다른
        #   이름으로 두 번 나온다(SUTS 가 실측 120건으로 확인한 함정).
        expected_vars, expected_raws, _n_ex, _n_skip2, _n_el2 = _expand_arrays(
            expected_vars, expected_raws, _MAX_EXP_PARAMS, _gim, _smem, _root_sizes)
        _arr_expanded_exp += _n_ex
        _arr_skipped += _n_skip2
        _arr_elements += _n_el2

        # ASIL — Pass 1 에서 정규화한 값을 그대로 쓴다. 여기서 다시 계산하면
        # 선별 기준(등급)과 방출 값이 갈라질 수 있다.
        asil = _cand["asil"]

        # ── Related IDs ──────────────────────────────────────────────────────
        # 정본 실측: 이 칸의 어휘는 SwCom 170 · SwFn 69 · SwSTR 62 · SwST 38 · SwTK 8 —
        # **설계/시험 요소 ID** 다. 요구 ID(SwTR_ 계열)는 0 건이다.
        #
        # ⚠ 두 가지를 좁게 잡고 있었다(2026-08-14 정본 50 TC · 원소 340개 대조):
        #   ① 어휘 — `SwCom_` 만 남겨 SwFn/SwSTR/SwST/SwTK 를 버렸다
        #   ② 범위 — **진입 함수 한 개**만 조회했다. 정본 340 원소 중 entry 자신으로
        #      설명되는 건 142(41.8%) 뿐이고 나머지 198 은 **호출 트리 아래에서만** 온다
        #      (정본은 `Related_ID 확인` 시트에서 트리 전체를 조회해 합집합을 만든다).
        #   둘을 합친 재현율: 18.8% → 74.4%(트리 3홉). 아래 두 축이 그 수정이다.
        related_parts: List[str] = []
        # ① 진입 함수 자신 — SwUDS 가 그 함수에 적어 둔 Related ID 전체
        _uds_swcom_lookups += 1
        _rel_map = uds_related_map if uds_related_map is not None else (uds_swcom_map or {})
        _uds_hit = list(_rel_map.get(fn_name.lower()) or [])
        if _uds_hit:
            _uds_swcom_hits += 1
            _uds_swcom_ids += len(_uds_hit)
            related_parts.extend(_uds_hit)
        # ①-b 호출 트리 아래 함수들 — 진입 함수 뒤에 붙인다(앞쪽이 더 직접적인 근거다).
        #     ⚠ 순서는 **거리 순**이다(체인의 깊이 우선 순서가 아니라). 상한에 걸릴 때
        #       무엇이 먼저 담기느냐가 곧 회수이고, 진입점에 가까운 함수의 설계 요소가
        #       그 통합 지점과 더 관련 있다(관측 대상에서 이미 검증된 같은 이유).
        _chain_rel: List[str] = []
        for _node in _bfs_call_order(fn_name, _calls_map, _RELATED_CHAIN_NODES,
                                     max_depth=_RELATED_CHAIN_DEPTH)[1:]:
            for _tok in _rel_map.get(_node.lower()) or []:
                if _tok not in related_parts and _tok not in _chain_rel:
                    _chain_rel.append(_tok)
        if _chain_rel:
            _chain_rel_flows += 1
            _chain_rel_ids += len(_chain_rel)
        related_parts.extend(_chain_rel)
        # 여기까지가 **SwUDS 문서에 적혀 있는 것**이다. 아래 균형 조정(`_balance_related_ids`)
        # 이 이 값을 지우지 못하게 표시해 둔다 — 그 함수는 "한 ID 가 흐름의 20% 를 넘으면
        # 과집중" 으로 보고 지우는데, 정본 실측에서 `SwFn_42` 는 50 TC 중 15(30%)에 정당하게
        # 쓰인다. 어휘를 넓힌 이번 라운드가 아니면 대상이 SwCom 뿐이라 드러나지 않던 구멍이다.
        _doc_related: List[str] = list(related_parts)
        # ② 소스 주석/SRS 경로가 실어 준 ID — **요구 ID 는 이 칸의 값이 아니다.**
        #    정본(KJPDS02_PV v1.02)에서 이 칸의 부제는 문자 그대로 `SwDS` 이고,
        #    49 TC · 342 토큰 중 요구 ID(SwTR/SwTSR/SwNTR…)는 **0 건**이다. 어휘
        #    판정은 `_RELATED_PREFIX_CANON`(SwUDS `Related ID` 칸 어휘의 단일 출처)
        #    하나로 한다 — 여기에 접두 목록을 다시 적으면 한쪽만 고쳐진다.
        #    ⚠ 칸에서 뺀다고 **버리지는 않는다**. 이 링크 자체는 진짜다: 실측 13건이
        #    전부 SRS 설명문이 함수를 명시적으로 부르는 경우였다("s_SysMain_Init( )
        #    초기화 함수 호출" · "Cpu_SRAM_ECC( ) 함수를 호출한다"). 칸이 틀렸을 뿐
        #    이라 `req_ids` 로 분리해 내보낸다 — 지금까지는 설계 칸에 섞여 있어서
        #    **요구 링크로 세지도, 보이지도 않았다**.
        req_ids: List[str] = []
        for field in ("srs_req_ids", "related", "related_id"):
            for _tok in _parse_req_ids(str(info.get(field) or "")):
                if _tok.split("_")[0].upper() in _RELATED_PREFIX_CANON:
                    related_parts.append(_tok)      # 설계 어휘 — 이 칸이 맞다
                elif _tok not in req_ids:
                    req_ids.append(_tok)
        if req_ids:
            _req_id_flows += 1
            _req_id_total += len(req_ids)
        # ③ from SDS map — 결과가 0 이어도 **왜 0 인지** 셀 수 있어야 한다(위 주석 참조).
        _sds_lookups += 1
        try:
            for cand in [fn_name, fn_name.lower()]:
                entry = sds_map.get(cand)
                if entry:
                    _sds_key_hits += 1
                    swcom_cand = entry.get("swcom") or entry.get("component")
                    if swcom_cand:
                        related_parts.append(swcom_cand)
                        _sds_swcom_hits += 1
                    break
        except Exception as e:  # noqa: BLE001 - 조회 실패는 보고하고 계속한다
            _logger.warning("SITS: SDS Related 보강 조회 실패(%s) — 이 함수는 건너뛴다: %s",
                            type(e).__name__, fn_name)
        # Assign SwCom.
        # ⚠ _infer_swcom_id는 **모듈 등장 순번**으로 만든 합성 ID다(실제 SDS component ID가
        # 아니다). 모든 flow에 무조건 들어가므로 related_ids는 절대 비지 않는다 — 이 값을
        # 요구 추적성 분자로 세면 항상 100%가 된다. 어느 ID가 합성인지 **삽입 지점에서**
        # 기록해 두어 품질 지표가 추측 없이 걸러낼 수 있게 한다.
        # (위 SDS map이 같은 ID를 이미 넣었다면 그건 문서 유래이므로 합성으로 치지 않는다.)
        # ⚠ SwUDS 에서 **진짜** SwCom 을 얻었으면 합성값을 덧붙이지 않는다. 덧붙이면 한
        #   칸에 문서 유래 SwCom 과 다른 컴포넌트를 가리키는 순번 합성값이 나란히 실리고,
        #   셀만 보는 쪽은 둘을 구별할 방법이 없다(합성 표시는 `synthetic_related_ids`
        #   에만 있고 산출물 셀에는 없다). 그 상태로 정본과 대조하면 모양이 같아 '일치'
        #   로도 잡힌다 — 가장 나쁜 종류의 거짓 추적성이다.
        swcom_id = _cand["swcom_id"]   # Pass 1 에서 후보 전체 기준으로 부여됨
        synthetic_related: List[str] = []
        if not _uds_hit and swcom_id not in related_parts:
            related_parts.insert(0, swcom_id)
            synthetic_related.append(swcom_id)

        # Deduplicate while preserving order
        seen_rel: set = set()
        deduped_related: List[str] = []
        for r in related_parts:
            if r and r not in seen_rel:
                seen_rel.add(r)
                deduped_related.append(r)
        # 칸 상한 — 넘치면 자르되 **몇 개를 잘랐는지** 남긴다. 앞쪽(진입 함수 자신 →
        # 가까운 호출)이 먼저 담기므로 잘리는 건 항상 가장 먼 근거다.
        if len(deduped_related) > _MAX_RELATED_IDS:
            _related_truncated += len(deduped_related) - _MAX_RELATED_IDS
            deduped_related = deduped_related[:_MAX_RELATED_IDS]

        # Collect indirect (global) vars for GLOBAL strategy
        indirect_vars_list: List[str] = []
        for g in globals_g + globals_s:
            tag = str(g).upper()
            gn = _clean_global_var_name(g)
            if gn and "[INDIRECT]" in tag and gn not in {p[0] for p in input_pairs}:
                if gn not in indirect_vars_list and len(indirect_vars_list) < 5:
                    indirect_vars_list.append(gn)
        # Also collect from callees
        for callee in cross_calls[:4]:
            callee_info = name_to_info.get(callee)
            if callee_info:
                for g in (callee_info.get("globals_global") or [])[:5]:
                    tag = str(g).upper()
                    gn = _clean_global_var_name(g)
                    if gn and "[INDIRECT]" in tag and gn not in indirect_vars_list:
                        if len(indirect_vars_list) < 5:
                            indirect_vars_list.append(gn)

        flows.append({
            "flow_id": fid,
            "entry_fn": fn_name,
            "call_chain": call_chain,
            # 경계를 직접 넘는지(`direct`) 몇 홉 아래에서 넘는지(`transitive`).
            "cross_via": _cand.get("cross_via", "direct"),
            "cross_calls": cross_calls,
            # 체인이 상한에 걸려 못 실은 함수 수. 0 이 아니면 이 경로는 **전부가 아니다**.
            "chain_dropped": _chain_dropped,
            "functions": _chain_nodes,
            "module_name": my_module,
            "swcom_id": swcom_id,
            "input_vars": input_vars,
            "input_raws": input_raws,   # annotated originals for type inference
            "expected_vars": expected_vars,
            "expected_raws": expected_raws,
            "indirect_vars": indirect_vars_list,
            "asil": asil,
            "related_ids": deduped_related,
            # related_ids 중 순번 기반 합성분(요구 추적성 분자에서 제외 — 위 삽입부 주석)
            "synthetic_related_ids": synthetic_related,
            # SwUDS 문서에 실제로 적혀 있던 ID — 균형 조정이 지우면 안 되는 것들
            "doc_related_ids": [r for r in _doc_related if r in seen_rel],
            # SRS/주석 유래 **요구** ID. Related 칸에는 안 싣는다(정본 0건) —
            # 칸이 아니라 데이터로 남겨 하류가 제대로 된 브리지를 만들 수 있게 한다.
            "req_ids": req_ids,
            "logic_flow": info.get("logic_flow") or [],
        })

    # 보강 실적을 **반드시** 내보낸다. 0 을 침묵시키면 "보강이 동작한다" 로 읽힌다.
    if stats_out is not None:
        stats_out.update({
            "sds_source": _sds_source,
            "sds_map_entries": len(sds_map or {}),
            "sds_lookups": _sds_lookups,
            "sds_key_hits": _sds_key_hits,
            "sds_swcom_hits": _sds_swcom_hits,
            # SwUDS 축 — Related 칸의 실제 소스
            "uds_swcom_map_entries": len(uds_swcom_map or {}),
            "uds_swcom_lookups": _uds_swcom_lookups,
            "uds_swcom_hits": _uds_swcom_hits,
            "uds_swcom_ids": _uds_swcom_ids,
            # 경계 판정 축 — 직접 vs 전이. 전이분이 0 이면 계층 진입점이 다시 빠지고
            # 있다는 뜻이므로 이 값이 회귀 신호가 된다.
            "transitive_entries": _transitive_entries,
            "cross_reach_hops": _CROSS_REACH_HOPS,
            # 체인 절단 — 0 이 아니면 그 흐름의 경로는 **전부가 아니다**.
            "chain_truncated_flows": _chain_truncated_total,
            "chain_max_nodes": _MAX_CHAIN_NODES,
            # Related 축 — 진입 함수 자신 vs 호출 트리 아래. 근거의 거리가 다르므로
            # 한 칸에 섞여 있어도 몇 개가 어디서 왔는지 여기서 보인다.
            "related_chain_flows": _chain_rel_flows,
            "related_chain_ids": _chain_rel_ids,
            "related_chain_depth": _RELATED_CHAIN_DEPTH,
            "related_truncated_ids": _related_truncated,
            # 요구 ID 축 — 칸에서 뺀 만큼 여기서 보여야 "없다" 와 구별된다.
            "req_id_flows": _req_id_flows,
            "req_id_total": _req_id_total,
            # 배열 원소 펼침 — 0 이면 배선이 끊긴 것인지 펼칠 게 없었던 것인지
            # `array_size_map_entries` 로 구분된다.
            "array_expanded_inputs": _arr_expanded_in,
            "array_expanded_expected": _arr_expanded_exp,
            "array_elements_emitted": _arr_elements,
            "array_skipped_budget": _arr_skipped,
            "array_size_map_entries": len(_root_sizes) + len(_gim),
            # ⚠ 구조체 멤버 배열은 **별도 축**이다. 위 값에 합치면 0 인지 아닌지 안 보이고,
            #   실제로 이 키가 캐시에 없어 0 인 채로 돌던 것을 진단하는 데 한 라운드가 들었다
            #   (`_SOURCE_SECTIONS_SCHEMA_VERSION` v12 참조).
            "array_struct_types": len(_smem),
            # 관측 대상 선별 축 — 정본과의 남은 격차가 **여기**다. 닫지 않고 보이게 한다.
            #   basis 가 static_call_graph 인 한 정본(VectorCAST 실행 관측)과는
            #   원리적으로 다른 집합이 나온다. 값을 지어내지 않고 사실을 적는다.
            "var_selection_basis": "static_call_graph",
            "var_scan_depth": _VAR_SCAN_DEPTH,
            "var_scan_nodes_max": _VAR_SCAN_NODES,
            "var_candidates_input": _var_cand_in,
            "var_candidates_expected": _var_cand_exp,
            "var_budget_cut_input": _var_budget_cut_in,
            "var_budget_cut_expected": _var_budget_cut_exp,
        })
    if _sds_lookups and not _sds_swcom_hits:
        # ⚠ 이 맵으로는 **구조적으로 0** 이다(값 스키마에 swcom/component 필드가 없고
        #   함수 항목의 `related` 는 요구 ID 뿐 — 실측). 그래서 문장을 "설정이 잘못됐다"
        #   가 아니라 사실대로 적고, SwUDS 축이 채웠는지를 같이 말한다.
        _logger.info(
            "SITS: SDS 맵에는 SwCom 축이 없다(조회 %d · 키매칭 %d · 맵 %d항목 · 출처=%s) "
            "— Related 의 SwCom 은 SwUDS 축에서 온다: %d/%d 함수 · ID %d개.",
            _sds_lookups, _sds_key_hits, len(sds_map or {}), _sds_source,
            _uds_swcom_hits, _uds_swcom_lookups, _uds_swcom_ids,
        )
    # ⚠ `None`(맵을 **주지 않은** 호출)과 `{}`(주려다 **비어서 온** 호출)는 다르다.
    #   전자는 SwCom 보강을 의도하지 않은 경로(영향도 dry-run 등)라 경고하면 노이즈가
    #   되고, 후자는 SwUDS 를 읽으려다 실패한 것이라 반드시 보여야 한다.
    #   `generate_sits` 는 실패해도 `{}` 를 넘기므로 진짜 실패는 항상 잡힌다.
    if uds_swcom_map is not None and _uds_swcom_lookups and not _uds_swcom_hits:
        _logger.warning(
            "SITS: SwUDS 기반 SwCom 보강이 %d회 조회에서 **0건** 산출했다(맵 %d항목). "
            "Related ID 는 순번 합성 SwCom 만 남는다 — 추적성 지표를 그대로 믿지 말 것.",
            _uds_swcom_lookups, len(uds_swcom_map),
        )
    if _chain_truncated_total:
        _logger.warning(
            "SITS: 호출 체인이 상한(%d)에 걸려 흐름 %d개에서 잘렸다 — 그 TC 의 "
            "Interface 칸은 **경로 전부가 아니다**.",
            _MAX_CHAIN_NODES, _chain_truncated_total,
        )
    # ⚠ `_transitive_entries` 는 **후보** 기준(Pass 1)이고 `flows` 는 캡 적용 후다.
    #   방출된 쪽을 따로 세지 않으면 캡에 잘린 전이 후보가 직접분으로 둔갑한다.
    _emitted_transitive = sum(1 for f in flows if f.get("cross_via") == "transitive")
    _logger.info(
        "SITS: collected %d integration flows (직접 경계 %d · 전이 경계 %d · %d홉까지 탐색"
        " · 후보 단계 전이 %d)",
        len(flows), len(flows) - _emitted_transitive, _emitted_transitive,
        _CROSS_REACH_HOPS, _transitive_entries,
    )
    if stats_out is not None:
        stats_out["transitive_entries_emitted"] = _emitted_transitive
    return flows


def _balance_related_ids(
    flows: List[Dict[str, Any]],
    max_freq_pct: float = 0.20,
) -> List[Dict[str, Any]]:
    """Redistribute over-concentrated Related IDs across flows.

    A req_id that appears in more than ``max_freq_pct`` of all flows is
    considered "over-used".  For flows that reference an over-used req_id
    *and* have at least one other (non-SwCom) req_id available, the
    over-used req_id is dropped so that SwCom IDs and less-frequent
    req_ids are surfaced instead.  SwCom_xx structural IDs are never
    removed.
    """
    total = len(flows)
    if total == 0:
        return flows

    max_count = max(1, int(total * max_freq_pct))

    # Count how many flows use each req_id
    usage: Dict[str, int] = {}
    for flow in flows:
        for rid in (flow.get("related_ids") or []):
            usage[rid] = usage.get(rid, 0) + 1

    # ⚠ **문서에 적혀 있는 ID 는 지우지 않는다.** 이 함수의 목적은 순번 합성 ID·소스 주석
    #   유래가 전 흐름에 붙어 추적성이 100% 로 보이는 것을 막는 것이지, SwUDS 가 여러 함수에
    #   정당하게 배정한 ID 를 걷어내는 게 아니다 — 정본 실측(50 TC)에서 `SwFn_42` 는 15건
    #   (30%)에 쓰이므로 이 임계(20%)면 **정본에 있는 값이 지워진다**.
    protected = {r for f in flows for r in (f.get("doc_related_ids") or [])}
    over_used = {rid for rid, cnt in usage.items()
                 if cnt > max_count and not rid.startswith("SwCom_") and rid not in protected}
    if not over_used:
        return flows

    _logger.info(
        "_balance_related_ids: %d over-used IDs (threshold %d/%d): %s",
        len(over_used), max_count, total, sorted(over_used),
    )

    trimmed = 0
    for flow in flows:
        rids = flow.get("related_ids") or []
        non_swcom = [r for r in rids if not r.startswith("SwCom_")]
        # Only drop over-used IDs when there are other non-SwCom alternatives
        if len(non_swcom) > 1:
            filtered = [r for r in rids if r not in over_used or r.startswith("SwCom_")]
            if len(filtered) < len(rids):
                flow["related_ids"] = filtered
                trimmed += 1

    _logger.info("_balance_related_ids: trimmed %d flows", trimmed)
    return flows


# ---------------------------------------------------------------------------
# Core: ITC generation
# ---------------------------------------------------------------------------

def _determine_gen_method_for_flow(flow: Dict[str, Any]) -> str:
    """Select ABV / AEC / AOR based on flow characteristics."""
    logic = flow.get("logic_flow") or []
    has_cond = any(
        isinstance(n, dict) and n.get("type") in ("if", "switch")
        for n in logic
    )
    n_inputs = len(flow.get("input_vars", []))
    n_cross = len(flow.get("cross_calls", []))

    if n_cross >= 3:
        return "AOR, ABV"
    if has_cond and n_inputs > 0:
        return "ABV, AEC"
    if n_inputs > 2:
        return "ABV"
    return "ABV, AEC"


def _generate_sub_cases(
    flow: Dict[str, Any],
    max_cases: int = _DEFAULT_SUBCASES,
    stp_environments: Optional[List[str]] = None,
    gen_method: str = "ABV",
) -> List[Dict[str, Any]]:
    """Generate sub-cases (boundary value rows) for an integration flow.

    Each sub-case has:
      case_num, call_chain, precondition, inputs {var: value}, expected {var: value}

    If ``stp_environments`` is provided (parsed from STP document), each sub-case
    precondition cycles through the defined test environments (HW-in-the-loop, ECU
    network, etc.) instead of a plain numeric index.
    """
    input_vars = flow.get("input_vars") or []
    expected_vars = flow.get("expected_vars") or []
    # Annotated originals carry explicit C type tokens (e.g. '[IN] U16 u16Speed')
    input_raws = flow.get("input_raws") or input_vars
    expected_raws = flow.get("expected_raws") or expected_vars
    call_chain = flow.get("call_chain", "")

    use_aec = "AEC" in str(gen_method).upper()

    # AEC equivalence class labels aligned to the 7-value boundary set:
    #   [min_inv, min_valid, low_mid, mid, high_mid, max_valid, max_inv]
    _AEC_LABELS: List[str] = [
        "EC1:무효-하한",   # min_inv    — invalid below minimum
        "EC2:유효-하한",   # min_valid  — valid lower boundary
        "EC3:유효-정상-L", # low_mid    — valid nominal low
        "EC4:유효-중간",   # mid        — valid mid
        "EC5:유효-정상-H", # high_mid   — valid nominal high
        "EC6:유효-상한",   # max_valid  — valid upper boundary
        "EC7:무효-상한",   # max_inv    — invalid above maximum
    ]

    def _precondition(case_idx: int) -> str:
        if stp_environments:
            return stp_environments[case_idx % len(stp_environments)]
        return str(case_idx + 1)

    def _case_label(case_idx: int) -> str:
        """Case number with optional AEC equivalence class label."""
        num = case_idx + 1
        if use_aec and case_idx < len(_AEC_LABELS):
            return f"{num} [{_AEC_LABELS[case_idx]}]"
        return str(num)

    if not input_vars:
        # No explicit inputs: generate scenario-based sub-cases using environment cycling.
        # Even without I/O data, integration flows can be exercised in multiple test
        # environments / scenarios (normal, boundary, error) per ISTQB integration test.
        _SCENARIO_LABELS = [
            "Normal operation",
            "Boundary condition",
            "Error / fault injection",
            "Post-initialization state",
            "Concurrent invocation",
            "Recovery sequence",
            "Stress / extended run",
        ]
        n_no_io = min(max_cases, len(_SCENARIO_LABELS)) if max_cases > 1 else 1
        # If STP environments available, cap to realistic count
        if stp_environments:
            n_no_io = min(n_no_io, max(max_cases, len(stp_environments)))
        result_cases: List[Dict[str, Any]] = []
        for i in range(n_no_io):
            scenario = _SCENARIO_LABELS[i]
            label = _case_label(i)
            precond = _precondition(i)
            result_cases.append({
                "case_num": i + 1,
                "case_label": label,
                "call_chain": call_chain if i == 0 else "",
                "precondition": precond,
                "inputs": {"Scenario": scenario},
                "expected": {v: "N/A" for v in (expected_vars[:5] or ["Result"])},
            })
        return result_cases

    # Determine boundary value sets using annotated raws first (type-token priority),
    # then fall back to name-prefix heuristic for plain variable names.
    bv_sets = [_infer_boundary_values(r) for r in input_raws]
    n_cases = min(max_cases, len(bv_sets[0]))

    sub_cases: List[Dict[str, Any]] = []
    for case_idx in range(n_cases):
        inputs: Dict[str, Any] = {}
        for var_idx, var_name in enumerate(input_vars):
            bv = bv_sets[var_idx]
            inputs[var_name] = bv[case_idx] if case_idx < len(bv) else bv[-1]

        # Expected: boundary-aware values using annotated raws
        expected: Dict[str, Any] = {}
        is_boundary = (case_idx == 0 or case_idx == n_cases - 1)
        for ev_idx, ev in enumerate(expected_vars):
            ev_raw = expected_raws[ev_idx] if ev_idx < len(expected_raws) else ev
            bv_exp = _infer_boundary_values(ev_raw)
            if is_boundary:
                # Error boundary → clamp to nearest valid value
                expected[ev] = bv_exp[1] if case_idx == 0 else bv_exp[3]
            else:
                expected[ev] = bv_exp[case_idx] if case_idx < len(bv_exp) else bv_exp[-1]

        sub_cases.append({
            "case_num": case_idx + 1,
            "case_label": _case_label(case_idx),
            "call_chain": call_chain if case_idx == 0 else "",
            "precondition": _precondition(case_idx),
            "inputs": inputs,
            "expected": expected,
        })

    # ── Additional strategies for branch coverage ──
    next_num = len(sub_cases) + 1

    # GAP A: Condition combination — toggle each input while others at mid
    if len(input_vars) >= 2 and len(sub_cases) < max_cases:
        for toggle_idx in range(min(4, len(input_vars))):
            if len(sub_cases) >= max_cases:
                break
            comb_inputs: Dict[str, Any] = {}
            for vi, vname in enumerate(input_vars):
                bv = bv_sets[vi] if vi < len(bv_sets) else _infer_boundary_values(vname)
                if vi == toggle_idx:
                    comb_inputs[vname] = bv[1] if toggle_idx % 2 == 0 else bv[5]  # min or max
                else:
                    comb_inputs[vname] = bv[3]  # mid
            comb_expected: Dict[str, Any] = {}
            for ev_idx, ev in enumerate(expected_vars):
                ev_raw = expected_raws[ev_idx] if ev_idx < len(expected_raws) else ev
                bv_exp = _infer_boundary_values(ev_raw)
                comb_expected[ev] = bv_exp[3]  # mid expected
            toggle_var = input_vars[toggle_idx] if toggle_idx < len(input_vars) else f"var{toggle_idx}"
            direction = "최솟값" if toggle_idx % 2 == 0 else "최댓값"
            sub_cases.append({
                "case_num": next_num,
                "case_label": f"COND_{toggle_idx+1} [{toggle_var}={direction}]",
                "call_chain": "",
                "precondition": f"조건 조합: {toggle_var}={direction}, 나머지=중간값",
                "inputs": comb_inputs,
                "expected": comb_expected,
            })
            next_num += 1

    # GAP C: Error propagation — inject boundary errors and check chain behavior
    if input_vars and len(sub_cases) < max_cases:
        for err_idx, err_key in enumerate(["min_inv", "max_inv"]):
            if len(sub_cases) >= max_cases:
                break
            err_inputs: Dict[str, Any] = {}
            for vi, vname in enumerate(input_vars):
                bv = bv_sets[vi]
                err_inputs[vname] = bv[0] if err_key == "min_inv" else bv[-1]
            err_expected: Dict[str, Any] = {}
            for ev_idx, ev in enumerate(expected_vars):
                ev_raw = expected_raws[ev_idx] if ev_idx < len(expected_raws) else ev
                bv_exp = _infer_boundary_values(ev_raw)
                err_expected[ev] = bv_exp[1] if err_key == "min_inv" else bv_exp[5]
            direction = "하한 초과" if err_key == "min_inv" else "상한 초과"
            sub_cases.append({
                "case_num": next_num,
                "case_label": f"ERR_PROP_{err_idx+1} [{direction}]",
                "call_chain": "",
                "precondition": f"에러 전파: 입력 {direction} → 콜체인 방어 처리 확인",
                "inputs": err_inputs,
                "expected": err_expected,
            })
            next_num += 1

    # GAP D: Global state combination — toggle indirect (global) vars
    indirect_vars = flow.get("indirect_vars") or []
    if indirect_vars and input_vars and len(sub_cases) < max_cases:
        for gv_idx, gv in enumerate(indirect_vars[:2]):
            if len(sub_cases) >= max_cases:
                break
            gstate_inputs: Dict[str, Any] = {}
            for vi, vname in enumerate(input_vars):
                bv = bv_sets[vi]
                gstate_inputs[vname] = bv[3]  # mid
            gv_bv = _infer_boundary_values(gv)
            gstate_inputs[gv] = gv_bv[1]  # global at min
            gstate_expected: Dict[str, Any] = {}
            for ev_idx, ev in enumerate(expected_vars):
                ev_raw = expected_raws[ev_idx] if ev_idx < len(expected_raws) else ev
                bv_exp = _infer_boundary_values(ev_raw)
                gstate_expected[ev] = bv_exp[3]
            gstate_expected[gv] = gv_bv[1]  # expect global stays at min (no change by function)
            sub_cases.append({
                "case_num": next_num,
                "case_label": f"GLOBAL_{gv_idx+1} [{gv}=min]",
                "call_chain": "",
                "precondition": f"글로벌 상태: {gv}=최솟값, 입력=중간값 → 상태 의존 분기 커버",
                "inputs": gstate_inputs,
                "expected": gstate_expected,
            })
            next_num += 1

    return sub_cases


def generate_itc_list(
    flows: List[Dict[str, Any]],
    max_subcases: int = _DEFAULT_SUBCASES,
    stp_environments: Optional[List[str]] = None,
    design_ids: Optional[Dict[str, Any]] = None,
    stats_out: Optional[Dict[str, Any]] = None,
    # ⚠ 신규 인자는 **맨 끝**에 붙인다(위치 인자 호출부가 조용히 다른 값에 바인딩된다).
    fi_flows: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """Generate list of Integration Test Cases from flows.

    Each ITC has:
      tc_id, gen_method, input_vars, expected_vars, related_ids, sub_cases

    ## TC ID 는 **SwUDS 설계 ID** 를 쓴다

    정본 실측(KJPDS02_PV_SwITS v1.02, 54건): `SwITC_SwUFn_0101_01` 12 ·
    `SwITC_SwUFn_0110` 37 · `SwITC_FI_SwFn_07` 5 — 전부 설계 ID 기반이다.
    우리는 파싱 순번(`SwITC_01`)을 써서 정본과 **교집합 0** 이었다(2026-08-14 실측).
    SUTS 가 같은 결함을 이미 고쳐 뒀고(`TC_ID = "SwUTC_" + SUDS` 로 1,013/1,014 일치)
    그 해석기(`generators.uds_design_ids`)를 **그대로 재사용**한다.

    ⚠ 설계 ID 를 못 찾으면 **순번으로 돌아간다**(`SwITC_NN`). SUTS 는 그 칸을 비우지만
    여기서는 TC ID 가 문서 식별자라 비울 수 없다 — 대신 모양이 달라 구별되고,
    `stats_out["design_id_hits"]` 로 몇 건이 문서 근거인지 셀 수 있다.
    """
    itcs: List[Dict[str, Any]] = []
    _design_hits = 0
    for idx, flow in enumerate(flows, start=1):
        _sid = resolve_design_id(design_ids, flow.get("entry_fn")) if design_ids else ""
        if _sid:
            _design_hits += 1
            tc_id = f"SwITC_{_sid}"
        else:
            tc_id = f"SwITC_{idx:02d}"
        gen_method = _determine_gen_method_for_flow(flow)
        sub_cases = _generate_sub_cases(
            flow, max_cases=max_subcases,
            stp_environments=stp_environments,
            gen_method=gen_method,
        )
        # If scenario-based sub-cases were generated (no real IO), expose the "Scenario"
        # pseudo-input so the XLSM writer renders the column header + values.
        effective_input_vars = list(flow["input_vars"])
        effective_expected_vars = list(flow["expected_vars"])
        if not effective_input_vars and sub_cases and "Scenario" in (sub_cases[0].get("inputs") or {}):
            effective_input_vars = ["Scenario"]
        # If expected_vars is empty but sub-cases carry result, add "Result" header
        if not effective_expected_vars and sub_cases:
            first_exp = sub_cases[0].get("expected") or {}
            if first_exp:
                effective_expected_vars = list(first_exp.keys())[:_MAX_EXP_PARAMS]
        itcs.append({
            "tc_id": tc_id,
            # Test Method 는 **여기서 확정**한다. 예전엔 라이터가 서브케이스 라벨을
            # 훑어 추론했는데, 오류전파 서브케이스 하나로 TC 전체가 FI 로 뒤집혔다
            # (`_sits_test_method` 주석의 지뢰). 정본은 FI 를 전용 TC 로만 쓴다.
            "test_method": _SITS_METHOD_DEFAULT,
            "gen_method": gen_method,
            "entry_fn": flow["entry_fn"],
            "call_chain": flow["call_chain"],
            "module_name": flow["module_name"],
            "input_vars": effective_input_vars,
            "expected_vars": effective_expected_vars,
            "related_ids": flow["related_ids"],
            "synthetic_related_ids": flow.get("synthetic_related_ids") or [],
            # Related 칸에는 안 나가지만 중간 JSON 으로는 내보낸다(아래 참조).
            "req_ids": flow.get("req_ids") or [],
            "sub_cases": sub_cases,
            "asil": flow["asil"],
        })
    # ── 전용 FI(고장 주입) TC ────────────────────────────────────────────
    # ⚠ **선별은 유도하지 않고 입력으로 받는다.** 정본의 FI 5건
    #   (`SwITC_FI_SwFn_07/34/36/37/41`)이 무엇으로 골렸는지 8가지 근거를 전부
    #   실측했는데 하나도 재현하지 못했다:
    #     logic_flow 오류어휘 0/5 · SySM 안전메커니즘 추적 0/5(집합 자체가 어긋남)
    #     ASIL 5/5지만 39/41 선택(무용) · 이름 휴리스틱 5/5지만 24선택(4.8배)
    #     오류/진단 전역 3/5 · 상태값 반환 3/5 · 둘의 교집합 2/5
    #     SwUDS 문서에 "fault injection" 어휘 0회 · 소스에 `@fault`/`@fi` 태그 0건
    #       (`@fi` 108건은 전부 **`@file`** — substring 위양성)
    #   정본의 5건은 사람의 안전분석(FMEA/FTA)에서 나온 선택이다. 추측해서
    #   Test Method 칸에 근거 없는 FI 를 붙이면 감사 문서가 검증 유형을 거짓 진술한다.
    _fi_emitted = 0
    for _ff in (fi_flows or []):
        _did = str(_ff.get("fi_design_id") or "").strip()
        if not _did:
            continue
        _fi_sub = _generate_sub_cases(
            _ff, max_cases=max_subcases, stp_environments=stp_environments,
            gen_method=_SITS_GEN_BOUNDARY,
        )
        itcs.append({
            # 정본 형태: `SwITC_FI_SwFn_07`. (`_integration_id` 는 `SwITC_`→`SwIT_`
            #  변환기라 여기 쓰면 접두가 통째로 빠진다.)
            "tc_id": f"{_TC_ID_PREFIX}FI_{_did}",
            "test_method": _SITS_METHOD_FAULT,
            # 정본 실측: FI 는 **항상** `AOR/ABV`(경계값분석)와 짝이다.
            "gen_method": _SITS_GEN_BOUNDARY,
            "entry_fn": _ff.get("entry_fn", ""),
            "call_chain": _ff.get("call_chain", ""),
            "module_name": _ff.get("module_name", ""),
            "input_vars": list(_ff.get("input_vars") or []),
            "expected_vars": list(_ff.get("expected_vars") or []),
            "related_ids": list(_ff.get("related_ids") or []),
            "synthetic_related_ids": list(_ff.get("synthetic_related_ids") or []),
            "req_ids": list(_ff.get("req_ids") or []),
            "sub_cases": _fi_sub,
            "asil": _ff.get("asil", ""),
        })
        _fi_emitted += 1
    if stats_out is not None:
        stats_out.update({
            "design_id_hits": _design_hits,
            "design_id_lookups": len(itcs),
            "design_id_map_entries": len((design_ids or {}).get("by_name") or {}),
            # FI 축 — 0 이 "요청이 없었다"인지 "요청했는데 못 냈다"인지 구분되게.
            "fi_emitted": _fi_emitted,
            "fi_requested": len(fi_flows or []),
        })
    if not _fi_emitted:
        # 침묵 금지: 정본에는 FI 축이 있다. 0 인 이유를 산출 시점에 남긴다.
        _logger.info(
            "SITS: 전용 FI TC 0건 — 고장주입 대상(설계 ID)이 지정되지 않았다. "
            "정본은 SwFn 설계 요소 5건을 FI 로 두는데, 그 선별은 안전분석 산출물이라 "
            "소스/SwUDS/SRS 에서 유도되지 않는다(8가지 근거 실측 — 코드 주석 참조).",
        )
    _logger.info(
        "SITS: generated %d ITCs, %d total sub-cases (TC ID 중 설계 ID 유래 %d · 순번 %d)",
        len(itcs), sum(len(t["sub_cases"]) for t in itcs), _design_hits,
        len(itcs) - _design_hits,
    )
    if itcs and not _design_hits:
        # 전부 순번이면 TC ID 축이 정본과 **한 건도** 안 맞는다. 조용히 넘기지 않는다.
        _logger.warning(
            "SITS: TC ID 가 전부 파싱 순번이다(설계 ID 0/%d · 맵 %d항목). 정본은 "
            "`SwITC_SwUFn_xxxx` 형태이므로 이 축은 대조 불가다 — SwUDS 경로를 확인할 것.",
            len(itcs), len((design_ids or {}).get("by_name") or {}),
        )
    return itcs


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def _find_ws(wb, *candidates: str):
    """여러 표기 후보 중 먼저 맞는 시트(`Introduction` ↔ `1.Introduction`).

    ⚠ 이름 정규화는 `_find_sheet` **단일 출처**를 쓴다. 여기서 다시 쓰면 라벨 변종
      대응이 한쪽에만 반영돼 두 축의 답이 갈라진다(이 파일이 이미 겪은 형태).
    """
    for c in candidates:
        ws = _find_sheet(wb, c)
        if ws is not None:
            return ws
    return None


def _fill_sits_front_matter(
    wb,
    project_id: str,
    doc_id: str,
    version: str,
    front_matter: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Cover / Introduction / Test Environment 를 실제 값으로 채운다.

    ⚠ 저장소 템플릿의 앞부분은 **하드웨어 통합시험 판**이 그대로 남아 있다. 실측
      (2026-08-18, 등록 템플릿 재생성본):

          1.5 Test Method   `FUnctional Testing/FNCT` · `Fault Injection Testing/FIT`
                            · `Electrical Testing/ELCT`
          1.3 용어사전       `"HDPDM01 Glossary"` ← **다른 프로젝트 이름**
          1.4 Reference     File Name 열 전부 빈칸 · Note 가 "HW 요구사항 명세서"
          1.1 Purpose       `XXXX 프로젝트`
          Cover             `HKY-[P_Name]-SwITS-2895` · `v0.10` · `Unspecified`
                            · `202X.XX.XX` · `XXXX`

      그런데 3번 시트는 `REQ, IFT` · `FI` 를 쓴다 — **문서가 자기 안에서 정의하지 않은
      코드를 쓰고 있었다.** 감사자가 1.5 표를 펴면 그 코드가 없다.
      어휘는 `_INTRO_TEST_METHODS`/`_INTRO_GEN_METHODS` **하나**에서 온다.

    ⚠ 1.2 Scope 의 "하드웨어 통합 테스트 사양" 문구는 **정본도 그대로**다(그쪽 템플릿
      상속). 정본과 같은 것을 '고치면' 오히려 멀어지므로 손대지 않는다.

    Returns: 채운 항목 통계(무엇을 못 채웠는지도 보이게).
    """
    fm = front_matter or {}
    stats: Dict[str, Any] = {"cover": 0, "intro": 0, "refs": 0,
                             "test_methods": 0, "gen_methods": 0, "missing": []}

    # ── Cover ────────────────────────────────────────────────────────────
    ws = _find_ws(wb, "Cover")
    if ws is not None:
        for cell, val in (("G26", doc_id), ("G27", version),
                          ("G28", fm.get("status") or "Open"),
                          ("G29", fm.get("date") or ""), ("G30", fm.get("author") or "")):
            if val:
                ws[cell] = val
                stats["cover"] += 1
    else:
        stats["missing"].append("Cover")

    # ── Introduction ─────────────────────────────────────────────────────
    ws = _find_ws(wb, "Introduction", "1.Introduction")
    if ws is None:
        stats["missing"].append("Introduction")
        return stats

    # 1.1 Purpose — 프로젝트 이름. 템플릿은 `XXXX` 를 박아 둔다.
    b4 = str(ws["B4"].value or "")
    if b4:
        ws["B4"] = re.sub(r"XXXX|\[P_Name\]", project_id, b4)
        stats["intro"] += 1
    # 1.3 용어사전 — 템플릿엔 **다른 프로젝트** 이름이 들어 있다.
    b15 = str(ws["B15"].value or "")
    if b15:
        ws["B15"] = f'"{fm.get("glossary") or (project_id + " Glossary")}"을 참조한다'
        stats["intro"] += 1

    # 1.4 Reference — 실제로 읽은 입력 문서. 없으면 그 줄은 비워 둔다(지어내지 않는다).
    for i, (fname, note) in enumerate((fm.get("references") or [])[:4]):
        r = 20 + i
        ws.cell(row=r, column=2, value=fname)
        # Note 열은 템플릿 D / 정본 E — 이미 값이 있는 열에 쓴다(둘 다 비면 D).
        note_col = 5 if str(ws.cell(row=19, column=5).value or "").strip() else 4
        ws.cell(row=r, column=note_col, value=note)
        stats["refs"] += 1

    # 1.5 / 1.6 — 어휘 표를 **문서가 실제로 쓰는 값**으로. 헤더 행을 찾아 그 아래에 쓴다.
    def _fill_table(header_text: str, rows) -> int:
        hdr = None
        for r in range(1, min(ws.max_row, 80) + 1):
            if str(ws.cell(row=r, column=2).value or "").strip().startswith(header_text):
                hdr = r
                break
        if hdr is None:
            stats["missing"].append(header_text)
            return 0
        # 헤더(`Method | 약어 | ASIL`) 바로 아래부터
        start = hdr + 2
        from copy import copy as _copy
        n = 0
        for i, (name, abbr, asil) in enumerate(rows):
            r = start + i
            for col, val in ((2, name), (4, abbr), (5, asil)):
                c = ws.cell(row=r, column=col)
                if i > 0:  # 늘어난 행은 첫 행 서식을 따라간다(무서식 행이 튀지 않게)
                    src = ws.cell(row=start, column=col)
                    c.font, c.border = _copy(src.font), _copy(src.border)
                    c.alignment, c.fill = _copy(src.alignment), _copy(src.fill)
                c.value = val
            n += 1
        return n

    stats["test_methods"] = _fill_table("1.5 Test Method", _INTRO_TEST_METHODS)
    stats["gen_methods"] = _fill_table("1.6 Test Case Generation Method", _INTRO_GEN_METHODS)

    # ── Test Environment — 정본은 환경 ID(`SwITE_01`)를 적는다 ────────────
    ws_env = _find_ws(wb, "1.Test Environment", "2.Test Environment")
    env_id = str(fm.get("test_env_id") or "").strip()
    if ws_env is not None and env_id:
        for r in range(1, min(ws_env.max_row, 20) + 1):
            v = str(ws_env.cell(row=r, column=2).value or "")
            if "통합테스트 환경" in v and env_id not in v:
                ws_env.cell(row=r, column=2,
                            value=f"SwTP에 정의된 {env_id} SW 통합테스트 환경을 사용함.")
                stats["intro"] += 1
                break
    return stats


def _create_sits_cover(
    wb, project_id: str, doc_id: str, version: str, asil_level: str,
    stp_context: Optional[Dict[str, Any]] = None,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    ws = wb.active
    ws.title = "Cover"

    title_font = Font(name="맑은 고딕", size=24, bold=True)
    label_font = Font(name="맑은 고딕", size=9, bold=True)
    data_font = Font(name="맑은 고딕", size=9)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    hdr_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    col_widths = {"A": 2.875, "B": 6.875, "C": 13.0, "D": 13.0, "E": 13.0,
                  "F": 13.0, "G": 13.0, "H": 4.625, "I": 6.875, "J": 13.0, "K": 10.625}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B5:K5")
    ws["B5"] = "Software Integration Test Specification\n(소프트웨어 통합테스트 명세서)"
    ws["B5"].font = title_font
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 123.0

    ws["I2"] = "Doc. ID"
    ws["I2"].font = label_font
    ws["I2"].alignment = center
    ws.merge_cells("J2:K2")
    ws["J2"] = doc_id
    ws["J2"].font = data_font
    ws["J2"].alignment = center

    ws["I3"] = "Version"
    ws["I3"].font = label_font
    ws["I3"].alignment = center
    ws.merge_cells("J3:K3")
    ws["J3"] = version
    ws["J3"].font = data_font
    ws["J3"].alignment = center

    stp_doc_id = (stp_context or {}).get("doc_id", "")
    stp_ver = (stp_context or {}).get("version", "")
    stp_ref = stp_doc_id + (f" {stp_ver}" if stp_ver else "")
    info_rows = [
        ("Project", project_id),
        ("ASIL Level", asil_level),
        ("STP Ref.", stp_ref or "-"),
        ("Status", "Draft"),
        ("Date", datetime.now().strftime("%Y-%m-%d")),
    ]
    for i, (label, value) in enumerate(info_rows):
        r = 21 + i
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=11)
        ws.cell(row=r, column=2, value=label).font = label_font
        ws.cell(row=r, column=2).fill = hdr_fill
        ws.cell(row=r, column=2).border = thin
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=6, value=value).font = data_font
        ws.cell(row=r, column=6).border = thin
        ws.cell(row=r, column=6).alignment = left


def _create_sits_history(wb, version: str) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    ws = wb.create_sheet("History")
    hdr_font = Font(name="맑은 고딕", size=10, bold=True)
    data_font = Font(name="맑은 고딕", size=9)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    hdr_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col, w in {"A": 1.25, "B": 8.375, "C": 9.125, "D": 35.5,
                   "E": 8.625, "F": 13.0, "G": 13.0, "H": 1.25}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:G2")
    ws["B2"] = "▶ Revision History"
    ws["B2"].font = Font(name="맑은 고딕", size=12, bold=True)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    for i, h in enumerate(["Version", "Date", "Description", "Author", "Reviewer", "Approver"]):
        c = ws.cell(row=4, column=2 + i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin
        c.alignment = center

    for ci, val in enumerate([version, datetime.now().strftime("%Y.%m.%d"),
                               "- Auto-generated", "Auto", "-", "-"]):
        cell = ws.cell(row=5, column=2 + ci, value=val)
        cell.font = data_font
        cell.border = thin


def _create_sits_intro(wb) -> None:
    from openpyxl.styles import Font
    ws = wb.create_sheet("1.Introduction")
    ws["A1"] = "Introduction"
    ws["A1"].font = Font(name="맑은 고딕", size=12, bold=True)
    ws["B3"] = "1.1 Purpose"
    ws["B3"].font = Font(name="맑은 고딕", size=10, bold=True)
    ws["B4"] = (
        "본 문서는 소프트웨어 통합테스트 명세를 기술하는 문서이며, "
        "소프트웨어 통합테스트 수행자에 의해서 작성된다."
    )
    ws["B6"] = "1.2 Scope"
    ws["B6"].font = Font(name="맑은 고딕", size=10, bold=True)
    ws["B7"] = (
        "본 문서는 소프트웨어 컴포넌트 간 통합 인터페이스 및 "
        "통합 테스트 케이스를 정의한다."
    )
    # ⚠ 1.3~1.6 이 통째로 없었다. 특히 **1.5 는 이 문서가 자기 Test Method 어휘를
    #   정의하는 곳**이라, 없으면 3번 시트의 `REQ, IFT`·`FI` 가 문서 안에서 정의되지
    #   않은 코드가 된다(템플릿 경로는 하드웨어 판이 남아 있어 같은 결과였다).
    #   행 배치는 정본(KJPDS02_PV_SwITS v1.02)과 같게 둔다 — 채움 함수는 헤더 텍스트로
    #   찾으므로 배치가 흔들려도 동작하지만, 사람이 두 문서를 나란히 볼 때가 있다.
    _hdr = Font(name="맑은 고딕", size=10, bold=True)
    _body = Font(name="맑은 고딕", size=9)
    ws["B14"] = "1.3 Terms, Abbreviations and Definitions"
    ws["B14"].font = _hdr
    ws["B15"] = '"Glossary"을 참조한다'
    ws["B15"].font = _body
    ws["B18"] = "1.4 Reference"
    ws["B18"].font = _hdr
    ws["B19"], ws["E19"] = "File Name", "Note"
    ws["B19"].font = ws["E19"].font = _hdr

    def _table(row0: int, title: str, rows) -> None:
        ws.cell(row=row0, column=2, value=title).font = _hdr
        for ci, lab in ((2, "Method"), (4, "약어"), (5, "ASIL")):
            ws.cell(row=row0 + 1, column=ci, value=lab).font = _hdr
        for i, (name, abbr, asil) in enumerate(rows):
            for ci, val in ((2, name), (4, abbr), (5, asil)):
                ws.cell(row=row0 + 2 + i, column=ci, value=val).font = _body

    _table(26, "1.5 Test Method", _INTRO_TEST_METHODS)
    _table(35, "1.6 Test Case Generation Method", _INTRO_GEN_METHODS)


def _create_sits_test_env(wb, stp_context: Optional[Dict[str, Any]] = None) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    ws = wb.create_sheet("2.Test Environment")
    ws["A1"] = "Test Environments"
    ws["A1"].font = Font(name="맑은 고딕", size=12, bold=True)

    stp_doc_id = (stp_context or {}).get("doc_id", "")
    envs = (stp_context or {}).get("environments", [])

    if envs:
        stp_ref = f"STP 참조: {stp_doc_id}" if stp_doc_id else "STP 참조"
        ws["B3"] = f"통합 테스트는 {stp_ref}에서 정의된 환경을 기준으로 수행된다."
        ws["B3"].font = Font(name="맑은 고딕", size=9)

        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
        hdr_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        hdr_font = Font(name="맑은 고딕", size=9, bold=True)
        data_font = Font(name="맑은 고딕", size=9)
        center = Alignment(horizontal="center", vertical="center")

        ws.cell(row=5, column=2, value="SwITE ID").font = hdr_font
        ws.cell(row=5, column=2).fill = hdr_fill
        ws.cell(row=5, column=2).border = thin
        ws.cell(row=5, column=2).alignment = center
        ws.cell(row=5, column=3, value="Test Environment").font = hdr_font
        ws.cell(row=5, column=3).fill = hdr_fill
        ws.cell(row=5, column=3).border = thin
        ws.cell(row=5, column=3).alignment = center

        for i, env in enumerate(envs, start=1):
            r = 5 + i
            ws.cell(row=r, column=2, value=f"SwITE_{i:02d}").font = data_font
            ws.cell(row=r, column=2).border = thin
            ws.cell(row=r, column=2).alignment = center
            ws.cell(row=r, column=3, value=env).font = data_font
            ws.cell(row=r, column=3).border = thin
    else:
        ws["B3"] = (
            "통합 테스트는 SwITE_01에서 정의된 환경을 기준으로 수행된다.\n"
            "- SwITE_01은 STP에서 정의되어 있다."
        )


def _norm_sheet_key(name: Any) -> str:
    """시트 이름 비교용 정규화 — 공백 제거 + 소문자(`2.SW…` ↔ `2. SW…`)."""
    return re.sub(r"\s+", "", str(name or "")).lower()


def _find_sheet(wb, name: str):
    """정규화 이름이 같은 시트를 돌려준다(없으면 None). 표기 흔들림 흡수."""
    want = _norm_sheet_key(name)
    for sn in wb.sheetnames:
        if _norm_sheet_key(sn) == want:
            return wb[sn]
    return None


def _build_strategy_tree(
    entry: str,
    calls_map: Dict[str, List[str]],
    max_nodes: int,
    max_depth: int,
) -> Tuple[List[Tuple[int, str]], int]:
    """호출 트리를 **경로별로** 편 `(상대깊이, 함수명)` 목록과 못 실은 함수 수.

    ⚠ `_build_call_chain_nodes`(체인)와 **의도적으로 다르다**. 체인은 visited 기반이라
    같은 함수가 한 번만 나오고(정본 3번 시트 실측: 중복 0/54), 전략 시트의 트리는 호출
    경로마다 다시 펼쳐진다(정본 2번 시트 실측: `s_HistoryPushDoorState` 가 한 블록에서
    5회 — 서로 다른 부모 밑에 각각). 한쪽 규칙을 다른 쪽에 쓰면 둘 다 정본과 어긋난다.

    사이클은 **현재 경로**를 기준으로만 끊는다(형제 가지에서 같은 함수가 다시 나오는 건
    재귀가 아니라 정상 호출이다).
    """
    out: List[Tuple[int, str]] = []
    dropped: set = set()
    stack: List[Tuple[str, int, Tuple[str, ...]]] = [(entry, 0, (entry,))]
    while stack:
        fn, depth, path = stack.pop()
        if len(out) >= max_nodes:
            dropped.add(fn)
            continue
        out.append((depth, fn))
        if depth >= max_depth:
            continue
        for callee in reversed(calls_map.get(fn, [])):
            if callee in path:      # 재귀 — 이 경로에서만 끊는다
                continue
            stack.append((callee, depth + 1, path + (callee,)))
    return out, len(dropped)


def _build_calls_map(function_details: Dict[str, Dict[str, Any]]) -> Dict[str, List[str]]:
    """`{함수명: [프로젝트 내부 callee]}`. 외부 심볼(memset·printf 등)은 뺀다.

    ⚠ `collect_integration_flows` 안쪽에도 같은 그래프가 필요하다 — 두 벌로 만들면 한쪽만
      필터가 바뀌었을 때 전략 시트의 트리와 규격 시트의 체인이 **서로 다른 그래프**가 된다.
    """
    names = {str(i.get("name") or "") for i in function_details.values()
             if isinstance(i, dict)}
    names.discard("")
    out: Dict[str, List[str]] = {}
    for info in function_details.values():
        if not isinstance(info, dict):
            continue
        n = str(info.get("name") or "")
        if not n or n in out:
            continue
        out[n] = [c for c in (info.get("calls_list") or []) if c in names]
    return out


def _build_file_map(function_details: Dict[str, Dict[str, Any]]) -> Dict[str, str]:
    """`{함수명: 파일명}` — 전략 시트 B열의 파일 목록에 쓴다(정본과 같은 자리)."""
    out: Dict[str, str] = {}
    for info in function_details.values():
        if not isinstance(info, dict):
            continue
        n = str(info.get("name") or "")
        if not n or n in out:
            continue
        raw = str(info.get("file") or info.get("file_name") or info.get("source_file") or "")
        if raw:
            out[n] = Path(raw).name
    return out


# 절대 깊이의 기준점. 정본 전략 시트의 `N depth` 는 진입점 기준 상대값이 아니라
# **실행 시작점 기준**이다(`s_Ap_ExecuteControlFunctions` = 5 depth).
_STRATEGY_ROOTS = ("main", "_EntryPoint")


def _absolute_depth_map(
    calls_map: Dict[str, List[str]],
    roots: Sequence[str],
    limit: int = 20000,
) -> Dict[str, int]:
    """루트(`main` 류)로부터의 **호출 깊이**. 정본 전략 시트의 `N depth` 가 이 값이다.

    정본 실측: 블록 시작 depth 는 1 이 22개지만 `s_Ap_ExecuteControlFunctions` 는
    `5 depth` 로 시작한다 — 진입점 기준 상대값이 아니라 **main 기준 절대값**이다.
    루트에서 닿지 않는 진입점(ISR·`_EntryPoint` 등)은 그 자체가 실행 시작점이므로 1.
    """
    depth: Dict[str, int] = {}
    queue: deque = deque()
    for r in roots:
        if r in calls_map and r not in depth:
            depth[r] = 1
            queue.append(r)
    while queue and len(depth) < limit:
        cur = queue.popleft()
        for callee in calls_map.get(cur, []):
            if callee not in depth:
                depth[callee] = depth[cur] + 1
                queue.append(callee)
    return depth


def _write_strategy_sheet(
    wb,
    flows: List[Dict[str, Any]],
    calls_map: Dict[str, List[str]],
    file_of: Dict[str, str],
    depth_of: Dict[str, int],
    stats_out: Optional[Dict[str, Any]] = None,
    id_of_entry: Optional[Dict[str, str]] = None,
) -> None:
    """정본 `2.SW Integration Strategy` 를 채운다 — 통합 지점별 호출 트리.

    정본 실측(v1.02): 39블록 · 2,303행 · 2,765셀. 블록 하나는

        헤더행   C~ : `5 depth` … `15 depth`     ← 열 = 깊이
        첫 행    B  : 통합 ID(`SwIT_SwUFn_0504_01`) · 깊이열 = 진입 함수
        이후     B  : 그 트리가 걸치는 **파일 목록** · 깊이열 = 하위 함수(경로별 전개)

    우리는 이 시트를 **템플릿 그대로**(제목·설명 2셀) 내보내고 있었다. 트리는 이미
    `call_chain` 을 만들 때 갖고 있는 정보라 새로 계산할 것이 없다 — 안 쓰고 있었을 뿐이다.
    """
    from openpyxl.styles import Alignment, Font

    ws = _find_sheet(wb, _STRATEGY_SHEET_NAME)
    if ws is None:
        ws = wb.create_sheet(_STRATEGY_SHEET_NAME)
        ws.cell(row=1, column=1, value="Software Integration Strategy").font = Font(
            name="맑은 고딕", size=12, bold=True)
        ws.cell(row=4, column=1, value=(
            "- 소프트웨어 테스트 계획서(STP)에서 정의한 SW Integration Strategy 에 따라 "
            "다음과 같이 통합 순서를 정의한다."))

    # 템플릿이 남긴 `< End of Document >` 는 지운다 — 트리를 그 위/아래 어디에 써도
    # 문서 끝 표시가 본문 중간에 남으면 읽는 사람이 거기서 끝난 줄 안다.
    end_row = 0
    for r in range(1, min(int(ws.max_row or 1), 4000) + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and _END_OF_DOC in str(v):
            ws.cell(row=r, column=1).value = None
            end_row = r
    hdr_font = Font(name="맑은 고딕", size=9, bold=True)
    data_font = Font(name="맑은 고딕", size=9)
    left = Alignment(horizontal="left", vertical="center")

    row = max(6, end_row and 6 or 6)
    blocks = truncated_blocks = 0
    node_total = node_dropped = 0
    for f in flows:
        entry = str(f.get("entry_fn") or "")
        if not entry:
            continue
        nodes, dropped = _build_strategy_tree(
            entry, calls_map, _STRATEGY_MAX_NODES, _STRATEGY_MAX_DEPTH)
        if not nodes:
            continue
        if row + len(nodes) + 2 > _STRATEGY_MAX_ROWS:
            truncated_blocks += 1
            continue
        base = depth_of.get(entry, 1)
        depths = sorted({base + d for d, _ in nodes})
        # ── 헤더행: 열 = 깊이 ────────────────────────────────────────────────
        col_of = {}
        for i, d in enumerate(depths):
            col = 3 + i          # C 부터
            col_of[d] = col
            c = ws.cell(row=row, column=col, value=f"{d} depth")
            c.font = hdr_font
            c.alignment = left
        row += 1
        # ── B열: 통합 ID + 이 트리가 걸치는 파일 목록 ────────────────────────
        # ⚠ 여기 쓰는 ID 는 3번 시트의 TC ID 와 **연결되는 값**이어야 한다. 예전엔 내부
        #   `flow_id`(`SwUFn_3596`)를 썼는데, 그건 규격 시트 어디에도 없는 이름이라 두 시트를
        #   맞춰 볼 방법이 없었다(정본은 `SwITC_…` ↔ `SwIT_…` 로 접두만 다르다).
        side: List[str] = [(id_of_entry or {}).get(entry)
                           or str(f.get("flow_id") or entry)]
        for _d, fn in nodes:
            fp = str(file_of.get(fn) or "")
            if fp and fp not in side:
                side.append(fp)
        for i, (d, fn) in enumerate(nodes):
            if i < len(side):
                ws.cell(row=row + i, column=2, value=side[i]).font = data_font
            c = ws.cell(row=row + i, column=col_of[base + d], value=fn)
            c.font = data_font
            c.alignment = left
        # 파일 목록이 트리보다 길면 남는 항목도 마저 적는다(조용히 버리지 않는다).
        for j in range(len(nodes), len(side)):
            ws.cell(row=row + j, column=2, value=side[j]).font = data_font
        row += max(len(nodes), len(side)) + 1
        blocks += 1
        node_total += len(nodes)
        node_dropped += dropped
    ws.cell(row=row + 1, column=1, value=_END_OF_DOC).font = data_font

    if stats_out is not None:
        stats_out["strategy_blocks"] = blocks
        stats_out["strategy_nodes"] = node_total
        stats_out["strategy_nodes_dropped"] = node_dropped
        stats_out["strategy_blocks_truncated"] = truncated_blocks
    _logger.info("SITS: 전략 시트 — 블록 %d · 노드 %d(상한 초과 %d) · 행 상한으로 뺀 블록 %d",
                 blocks, node_total, node_dropped, truncated_blocks)


def _write_relid_sheets(
    wb,
    flows: List[Dict[str, Any]],
    calls_map: Dict[str, List[str]],
    file_of: Dict[str, str],
    uds_related_map: Dict[str, List[str]],
    design_ids: Optional[Dict[str, Any]] = None,
    stats_out: Optional[Dict[str, Any]] = None,
    id_of_entry: Optional[Dict[str, str]] = None,
) -> None:
    """정본 `Related_ID 확인` / `Related_ID 정리` 를 채운다 — Related 칸의 **계산 근거**.

    정본은 이 두 시트에서 Related ID 를 만들어 3번 시트로 옮긴다(실측: 3번 시트 55건 중
    **43건이 `정리` D열과 문자 그대로 같고** 6건은 포함관계). 즉 이 시트들은 장식이 아니라
    산출물의 근거표다 — 없으면 Related 칸이 어디서 왔는지 문서 안에서 확인할 방법이 없다.

    `확인`  좌측: 트리 평탄화(A=통합ID·파일 / B=함수 / C=그 함수의 Related ID / D=블록 합집합)
            우측: SwUDS 함수 인덱스(F=No / G=설계ID / H=함수명 / I=Related ID)
    `정리`      : B=통합ID / C=진입 함수 / D=합집합

    ⚠ 정본 `정리` 시트의 F~H 열(대조·`O` 표시)은 **사람이 검증한 흔적**이라 만들지 않는다.
      생성기가 자기 출력을 자기가 `O` 로 표시하면 검증한 적 없는 것이 검증된 것처럼 보인다
      (이 저장소가 `[[project_provenance_laundering]]` 에서 이미 겪은 형태다).
    """
    from openpyxl.styles import Font

    hdr_font = Font(name="맑은 고딕", size=9, bold=True)
    data_font = Font(name="맑은 고딕", size=9)

    for name in (_RELID_CHECK_SHEET, _RELID_TIDY_SHEET):
        old = _find_sheet(wb, name)
        if old is not None:
            del wb[old.title]
    ws = wb.create_sheet(_RELID_CHECK_SHEET)
    ws2 = wb.create_sheet(_RELID_TIDY_SHEET)

    for col, label in ((1, "통합 ID / 파일"), (2, "함수"), (3, "Related ID"),
                       (4, "블록 합집합")):
        ws.cell(row=1, column=col, value=label).font = hdr_font
    for col, label in ((6, "No"), (7, "ID"), (8, "Name"), (9, "RelatedID")):
        ws.cell(row=1, column=col, value=label).font = hdr_font
    for col, label in ((2, "통합 ID"), (3, "진입 함수"), (4, "Related ID(합집합)")):
        ws2.cell(row=1, column=col, value=label).font = hdr_font

    row = 2
    trow = 2
    for f in flows:
        entry = str(f.get("entry_fn") or "")
        if not entry:
            continue
        nodes, _dropped = _build_strategy_tree(
            entry, calls_map, _STRATEGY_MAX_NODES, _STRATEGY_MAX_DEPTH)
        if not nodes:
            continue
        if row + len(nodes) + 2 > _STRATEGY_MAX_ROWS:
            continue
        flow_id = (id_of_entry or {}).get(entry) or str(f.get("flow_id") or entry)
        union = [x for x in (f.get("related_ids") or []) if x]
        side = [flow_id]
        for _d, fn in nodes:
            fp = str(file_of.get(fn) or "")
            if fp and fp not in side:
                side.append(fp)
        for i, (_d, fn) in enumerate(nodes):
            if i < len(side):
                ws.cell(row=row + i, column=1, value=side[i]).font = data_font
            ws.cell(row=row + i, column=2, value=fn).font = data_font
            own = ", ".join(uds_related_map.get(fn.lower()) or [])
            if own:
                ws.cell(row=row + i, column=3, value=own).font = data_font
        ws.cell(row=row, column=4, value=", ".join(union)).font = data_font
        row += max(len(nodes), len(side)) + 1

        ws2.cell(row=trow, column=2, value=flow_id).font = data_font
        ws2.cell(row=trow, column=3, value=entry).font = data_font
        ws2.cell(row=trow, column=4, value=", ".join(union)).font = data_font
        trow += 1

    # ── 우측: SwUDS 함수 인덱스(정본과 같은 4열) ───────────────────────────
    # ⚠ 이 맵의 키는 **소문자**다(문서 표기 흔들림 흡수용). 그대로 쓰면 두 가지가 깨진다:
    #   ① 설계 ID 조회 — `design_ids["by_name"]` 은 원문 케이스 키라 전부 미스(G열 전멸)
    #   ② 이름 표기 — 문서에 없는 `_entrypoint` 같은 이름이 산출물에 실린다
    #   그래서 소문자 키로 조회하되 **원문 이름을 되살려** 적는다.
    by_name = (design_ids or {}).get("by_name") or {}
    design_by_lower: Dict[str, str] = {}
    orig_name: Dict[str, str] = {}
    for k, v in by_name.items():
        design_by_lower.setdefault(str(k).lower(), str(v))
        orig_name.setdefault(str(k).lower(), str(k))
    for k in calls_map:
        orig_name.setdefault(str(k).lower(), str(k))
    idx_row = 2
    for no, fn in enumerate(sorted(uds_related_map), start=1):
        ws.cell(row=idx_row, column=6, value=no).font = data_font
        ws.cell(row=idx_row, column=7, value=design_by_lower.get(fn, "")).font = data_font
        ws.cell(row=idx_row, column=8, value=orig_name.get(fn, fn)).font = data_font
        ws.cell(row=idx_row, column=9,
                value=", ".join(uds_related_map.get(fn) or [])).font = data_font
        idx_row += 1

    if stats_out is not None:
        stats_out["relid_check_rows"] = row - 2
        stats_out["relid_tidy_rows"] = trow - 2
        stats_out["relid_index_rows"] = idx_row - 2
    _logger.info("SITS: Related_ID 시트 — 확인 %d행 · 정리 %d행 · SwUDS 인덱스 %d행",
                 row - 2, trow - 2, idx_row - 2)


def generate_sits_xlsm(
    template_path: Optional[str],
    itcs: List[Dict[str, Any]],
    output_path: str,
    project_config: Optional[Dict[str, Any]] = None,
    flows: Optional[List[Dict[str, Any]]] = None,
    stp_context: Optional[Dict[str, Any]] = None,
    strategy_context: Optional[Dict[str, Any]] = None,
    # ⚠ 신규 인자는 **맨 끝**에 붙인다(위치 인자 호출부가 조용히 다른 값에 바인딩된다).
    front_matter: Optional[Dict[str, Any]] = None,
) -> str:
    """Generate SITS XLSM file matching the reference structure.

    `strategy_context` 를 주면 근거 시트 두 벌(`2.SW Integration Strategy`,
    `Related_ID 확인`/`정리`)을 함께 채운다:
    `{calls_map, file_of, depth_of, uds_related_map, design_ids, stats_out}`.
    안 주면 그 시트들은 템플릿 상태 그대로 나간다 — 조용히 빈 시트를 만들지는 않는다.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        _logger.error("openpyxl not installed")
        raise

    cfg = project_config or {}
    project_id = cfg.get("project_id", "PROJECT")
    doc_id = cfg.get("doc_id", f"{project_id}-SITS")
    version = cfg.get("version", "v1.00")
    asil_level = cfg.get("asil_level", "")

    if template_path and Path(template_path).is_file():
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        _logger.info("Loaded SITS template: %s", template_path)
    else:
        wb = openpyxl.Workbook()
        _create_sits_cover(wb, project_id, doc_id, version, asil_level, stp_context=stp_context)
        _create_sits_history(wb, version)
        _create_sits_intro(wb)
        _create_sits_test_env(wb, stp_context=stp_context)
        # ⚠ 무템플릿 경로가 만드는 시트 이름은 정본과 다르다(`1.Introduction` ·
        #   `2.Test Environment` vs 정본 `Introduction` · `1.Test Environment`).
        #   아래 채움은 두 표기를 모두 찾으므로 어느 경로든 같은 값이 들어간다.
        # 통합 전략 시트는 아래 `_write_strategy_sheet` 가 만든다(`strategy_context` 를 준
        # 경우). 예전엔 여기서 `3-1.SW Integration Strategy` 라는 **다른 이름의 시트**를
        # cross-calls 표로 따로 만들었는데, 정본 시트(`2.SW…`)를 채우게 된 지금은 한 파일에
        # 같은 주제의 시트가 두 벌 남는다 — `3.`/`4.` 에서 이미 겪은 결함이라 제거했다.
        _logger.info("Created new SITS workbook (no template)")

    # 템플릿이든 신규든 **앞부분은 반드시 채운다**. 템플릿의 1.5 표는 하드웨어
    # 통합시험 어휘(FNCT/FIT/ELCT)라, 두지 않으면 3번 시트가 쓰는 REQ/IFT/FI 가
    # 이 문서 안에서 정의되지 않은 코드가 된다(`_fill_sits_front_matter` 참조).
    try:
        _fm_stats = _fill_sits_front_matter(wb, project_id, doc_id, version,
                                            front_matter=front_matter)
        if _fm_stats.get("missing"):
            _logger.warning("SITS: 앞부분 시트 일부를 못 찾았다 — %s",
                            ", ".join(str(x) for x in _fm_stats["missing"]))
        _logger.info(
            "SITS: 앞부분 채움 — Cover %d · Introduction %d · Reference %d행 · "
            "1.5 %d행 · 1.6 %d행",
            _fm_stats.get("cover", 0), _fm_stats.get("intro", 0),
            _fm_stats.get("refs", 0), _fm_stats.get("test_methods", 0),
            _fm_stats.get("gen_methods", 0),
        )
        if strategy_context is not None and strategy_context.get("stats_out") is not None:
            strategy_context["stats_out"]["front_matter"] = _fm_stats
    except Exception as _fm_exc:  # noqa: BLE001 — 앞부분 실패로 규격 시트를 잃지 않는다
        _logger.warning("SITS: 앞부분 채움 실패(%s) — 규격 시트는 그대로 진행: %s",
                        type(_fm_exc).__name__, _fm_exc)

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    hdr_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    hdr_font = Font(name="맑은 고딕", size=9, bold=True)
    data_font = Font(name="맑은 고딕", size=8)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ⚠ 정본의 시트는 `3.SW Integration Test Spec` 이다. 예전엔 템플릿의 3번 시트를
    #   놔둔 채 `4.…` 를 **따로 붙여** 한 파일에 빈 정본 시트와 채워진 사본이 공존했다.
    sheet_name = _SPEC_SHEET_NAME
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── Row 1: title ────────────────────────────────────────────────────────
    title_font = Font(name="맑은 고딕", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_RELATED_COL)
    ws.cell(row=1, column=1, value="Software Integration Test Specification").font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Helper: fill + merge ────────────────────────────────────────────────
    def _fill_and_merge(row: int, c_start: int, c_end: int, label: str) -> None:
        for c in range(c_start, c_end + 1):
            ws.cell(row=row, column=c).fill = hdr_fill
            ws.cell(row=row, column=c).border = thin
            ws.cell(row=row, column=c).alignment = center
        ws.cell(row=row, column=c_start, value=label).font = hdr_font
        if c_end > c_start:
            try:
                ws.merge_cells(start_row=row, start_column=c_start,
                                end_row=row, end_column=c_end)
            except Exception:
                pass

    # ── Row 5: group headers ────────────────────────────────────────────────
    _fill_and_merge(_BAND_ROW, _TCID_COL, _GEN_COL, "Test Case")
    _fill_and_merge(_BAND_ROW, _INPUT_COL_START, _INPUT_COL_END, "Input")
    _fill_and_merge(_BAND_ROW, _EXP_COL_START, _EXP_COL_END, "Expected Result")
    _fill_and_merge(_BAND_ROW, _RELATED_COL, _RELATED_COL, "Related ID")
    ws.row_dimensions[_BAND_ROW].height = 18

    # ── Row 6: detail headers ── 정의는 모듈 상수 `_DETAIL_HEADERS`(단일 출처) ──
    # 별칭이 아니라 사본 — 이후 누가 여기서 헤더를 덧쓰더라도 모듈 상수가 오염되지 않게.
    detail_headers: Dict[int, str] = dict(_DETAIL_HEADERS)
    for col_i in range(1, _RELATED_COL + 1):
        cell = ws.cell(row=_HEADER_ROW, column=col_i)
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = center
        cell.font = hdr_font
        if col_i in detail_headers:
            cell.value = detail_headers[col_i]
        elif _INPUT_COL_START <= col_i <= _INPUT_COL_END:
            cell.value = f"Param {col_i - _INPUT_COL_START + 1}"
        elif _EXP_COL_START <= col_i <= _EXP_COL_END:
            cell.value = f"Param {col_i - _EXP_COL_START + 1}"
    ws.row_dimensions[_HEADER_ROW].height = 30

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 1.0
    ws.column_dimensions[get_column_letter(_TCID_COL)].width = 14
    ws.column_dimensions[get_column_letter(_DESC_COL)].width = 10
    ws.column_dimensions[get_column_letter(_CHAIN_COL)].width = 40
    ws.column_dimensions[get_column_letter(_SAFETY_COL)].width = 9
    ws.column_dimensions[get_column_letter(_METHOD_COL)].width = 12
    ws.column_dimensions[get_column_letter(_GEN_COL)].width = 14
    ws.column_dimensions[get_column_letter(_SEQ_COL)].width = 5
    for ci in range(_INPUT_COL_START, _INPUT_COL_END + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    for ci in range(_EXP_COL_START, _EXP_COL_END + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.column_dimensions[get_column_letter(_RELATED_COL)].width = 35

    # ── Data rows ────────────────────────────────────────────────────────────
    current_row = _DATA_START_ROW
    for itc in itcs:
        tc_id = itc["tc_id"]
        input_vars = itc.get("input_vars") or []
        expected_vars = itc.get("expected_vars") or []
        related_str = ", ".join(itc.get("related_ids") or [])
        gen_method = itc.get("gen_method", "ABV")

        # TC header row
        tc_desc = (
            f"Verify integration: {itc.get('entry_fn', '')} → "
            + " → ".join((itc.get("call_chain") or "").split(" -> ")[1:3])
        ).rstrip(" →")

        ws.cell(row=current_row, column=_TCID_COL, value=tc_id).font = Font(name="맑은 고딕", size=9, bold=True)
        ws.cell(row=current_row, column=_TCID_COL).border = thin
        ws.cell(row=current_row, column=_DESC_COL, value=tc_desc).font = data_font
        ws.cell(row=current_row, column=_DESC_COL).border = thin
        ws.cell(row=current_row, column=_DESC_COL).alignment = wrap
        # Safety Related / Test Method — 정본은 TC 행에 쓰고 블록 전체로 병합한다
        # (SUTS 와 다르다: SUTS 는 Method 가 시퀀스 그룹 단위다).
        ws.cell(row=current_row, column=_SAFETY_COL,
                value=_safety_mark(itc.get("asil"))).font = data_font
        ws.cell(row=current_row, column=_SAFETY_COL).alignment = center
        ws.cell(row=current_row, column=_SAFETY_COL).border = thin
        # ⚠ Gen 은 Method 에서 유도한다 — 정본에서 둘은 짝이다(`_sits_gen_method`).
        #   독립으로 정하면 정본에 없는 조합이 나온다(실측: 367건 전부 그랬다).
        _test_method = _sits_test_method(itc)
        ws.cell(row=current_row, column=_METHOD_COL,
                value=_test_method).font = data_font
        ws.cell(row=current_row, column=_METHOD_COL).alignment = center
        ws.cell(row=current_row, column=_METHOD_COL).border = thin
        ws.cell(row=current_row, column=_GEN_COL,
                value=_sits_gen_method(gen_method, _test_method)).font = data_font
        ws.cell(row=current_row, column=_GEN_COL).alignment = center
        ws.cell(row=current_row, column=_GEN_COL).border = thin
        ws.cell(row=current_row, column=_RELATED_COL, value=related_str).font = data_font
        ws.cell(row=current_row, column=_RELATED_COL).border = thin
        ws.cell(row=current_row, column=_RELATED_COL).alignment = wrap

        # Input param name headers in TC row
        for vi, var_name in enumerate(input_vars[:_MAX_INPUT_PARAMS]):
            col = _INPUT_COL_START + vi
            ws.cell(row=current_row, column=col, value=var_name).font = data_font
            ws.cell(row=current_row, column=col).border = thin

        # Expected param name headers in TC row
        for vi, var_name in enumerate(expected_vars[:_MAX_EXP_PARAMS]):
            col = _EXP_COL_START + vi
            ws.cell(row=current_row, column=col, value=var_name).font = data_font
            ws.cell(row=current_row, column=col).border = thin

        ws.row_dimensions[current_row].height = 18
        _tc_row = current_row
        current_row += 1

        # Sub-case rows
        for sc in itc.get("sub_cases") or []:
            desc_val = sc.get("case_label") or sc["case_num"]
            ws.cell(row=current_row, column=_DESC_COL, value=desc_val).font = data_font
            ws.cell(row=current_row, column=_DESC_COL).border = thin
            ws.cell(row=current_row, column=_DESC_COL).alignment = wrap

            chain_val = sc.get("call_chain") or ""
            if chain_val:
                ws.cell(row=current_row, column=_CHAIN_COL, value=chain_val).font = data_font
                ws.cell(row=current_row, column=_CHAIN_COL).alignment = wrap
            ws.cell(row=current_row, column=_CHAIN_COL).border = thin

            # 서브케이스 번호는 정본에서 C(Description 앞칸)와 H 두 곳에 온다.
            ws.cell(row=current_row, column=_SEQ_COL, value=sc["case_num"]).font = data_font
            ws.cell(row=current_row, column=_SEQ_COL).alignment = center
            ws.cell(row=current_row, column=_SEQ_COL).border = thin

            # Input values
            sc_inputs = sc.get("inputs") or {}
            for vi, var_name in enumerate(input_vars[:_MAX_INPUT_PARAMS]):
                col = _INPUT_COL_START + vi
                val = sc_inputs.get(var_name, "")
                ws.cell(row=current_row, column=col, value=val).font = data_font
                ws.cell(row=current_row, column=col).border = thin
                ws.cell(row=current_row, column=col).alignment = center

            # Expected values
            sc_expected = sc.get("expected") or {}
            for vi, var_name in enumerate(expected_vars[:_MAX_EXP_PARAMS]):
                col = _EXP_COL_START + vi
                val = sc_expected.get(var_name, "")
                ws.cell(row=current_row, column=col, value=val).font = data_font
                ws.cell(row=current_row, column=col).border = thin
                ws.cell(row=current_row, column=col).alignment = center

            ws.row_dimensions[current_row].height = 14
            current_row += 1

        # TC 메타 열은 블록 전체 병합 — 정본이 그렇다(B5:B40 · E5:E40 · F5:F40 · G5:G40 · GV5:GV40).
        _block_end = current_row - 1
        if _block_end > _tc_row:
            for mc in (_TCID_COL, _SAFETY_COL, _METHOD_COL, _GEN_COL, _RELATED_COL):
                try:
                    ws.merge_cells(start_row=_tc_row, start_column=mc,
                                   end_row=_block_end, end_column=mc)
                except Exception as exc:  # noqa: BLE001
                    _logger.debug("SITS TC merge skipped (col %d, %d:%d): %s",
                                  mc, _tc_row, _block_end, exc)

    # Freeze panes — 헤더 아래 첫 데이터 행에서 고정
    ws.freeze_panes = f"C{_DATA_START_ROW}"

    # ── 근거 시트 — 통합 전략 트리 + Related ID 산출 근거 ────────────────────
    # 시험 규격 시트를 다 쓴 뒤에 채운다(같은 flows 를 두 벌 쓰는 것이므로 순서 의존은
    # 없지만, 실패해도 규격 시트는 이미 완성돼 있어야 한다).
    if strategy_context:
        _sctx = strategy_context
        try:
            # 근거 시트의 통합 ID 는 규격 시트 TC ID 에서 파생한다 — 두 시트를 맞춰 볼 수
            # 없는 ID 를 적으면 근거표 구실을 못한다(정본: `SwITC_…` ↔ `SwIT_…`).
            _id_of_entry: Dict[str, str] = {}
            for _t in itcs or []:
                _e = str(_t.get("entry_fn") or "")
                if _e and _e not in _id_of_entry:
                    _id_of_entry[_e] = _integration_id(str(_t.get("tc_id") or ""))
            _write_strategy_sheet(
                wb, flows or [],
                _sctx.get("calls_map") or {}, _sctx.get("file_of") or {},
                _sctx.get("depth_of") or {}, _sctx.get("stats_out"),
                id_of_entry=_id_of_entry)
            _write_relid_sheets(
                wb, flows or [],
                _sctx.get("calls_map") or {}, _sctx.get("file_of") or {},
                _sctx.get("uds_related_map") or {}, _sctx.get("design_ids"),
                _sctx.get("stats_out"), id_of_entry=_id_of_entry)
        except Exception as exc:  # noqa: BLE001 — 근거 시트 실패가 규격 시트를 못 죽인다
            _logger.warning("SITS: 근거 시트 생성 실패(%s) — 규격 시트는 그대로 저장한다: %s",
                            type(exc).__name__, exc)
            if isinstance(_sctx.get("stats_out"), dict):
                _sctx["stats_out"]["strategy_error"] = f"{type(exc).__name__}: {exc}"

    # Save
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    _logger.info("SITS XLSM saved: %s (rows=%d)", out_path.name, current_row - _DATA_START_ROW)
    return str(out_path)


# ---------------------------------------------------------------------------
# Quality report
# ---------------------------------------------------------------------------

def generate_sits_quality_report(
    itcs: List[Dict[str, Any]],
    total_source_functions: int = 0,
    flow_stats: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    total_tc = len(itcs)
    total_sub = sum(len(t.get("sub_cases") or []) for t in itcs)
    avg_sub = round(total_sub / max(total_tc, 1), 1)

    gen_dist: Dict[str, int] = {}
    for itc in itcs:
        for m in re.split(r"[,\s]+", itc.get("gen_method") or "ABV"):
            m = m.strip()
            if m:
                gen_dist[m] = gen_dist.get(m, 0) + 1

    with_related = sum(1 for t in itcs if t.get("related_ids"))
    related_pct = round(with_related / max(total_tc, 1) * 100, 1)

    # ── 요구 추적성은 Related ID 보유율과 다른 축이다 ──
    # collect_integration_flows가 모든 flow에 순번 기반 합성 SwCom_XX를 삽입하므로
    # related_ids는 절대 비지 않는다 → related_pct는 사실상 항상 100%다. 그 값을 요구
    # 추적성으로 쓰면 요구 링크가 하나도 없어도 게이트를 통과한다. 여기서는 **삽입 지점이
    # 기록한 synthetic_related_ids를 뺀 실제 ID**로만 분자를 센다(문자열 prefix 추측 아님 —
    # SDS 문서에서 온 진짜 SwCom ID는 합성으로 분류되지 않는다).
    def _real_related(t: Dict[str, Any]) -> List[str]:
        synth = set(t.get("synthetic_related_ids") or [])
        return [r for r in (t.get("related_ids") or []) if r not in synth]

    with_req_trace = sum(1 for t in itcs if _real_related(t))
    req_trace_pct = round(with_req_trace / max(total_tc, 1) * 100, 1)
    # related_ids는 있으나 전부 합성인 TC — "링크 있음"으로 보이지만 추적 근거는 0이다.
    synthetic_only_count = sum(
        1 for t in itcs if (t.get("related_ids") and not _real_related(t))
    )

    swcom_dist: Dict[str, int] = {}
    for t in itcs:
        rids = t.get("related_ids") or []
        for r in rids:
            if r.startswith("SwCom_"):
                swcom_dist[r] = swcom_dist.get(r, 0) + 1

    with_io = sum(
        1 for t in itcs
        if t.get("input_vars") or t.get("expected_vars")
    )
    io_pct = round(with_io / max(total_tc, 1) * 100, 1)

    # ── 통합 흐름 캡 절단 (있으면) ──────────────────────────────────────────
    # TC 수(total_test_cases)만 보면 "흐름 120개 전부 시험함" 으로 읽힌다. 분모는
    # 생성된 흐름 수가 아니라 **소스에서 찾은 흐름 수**다 — 캡에 잘린 만큼 규격에
    # 아예 없는 흐름이 생기므로 그 사실을 리포트에 남긴다.
    fs = flow_stats or {}
    flow_cov: Dict[str, Any] = {}
    if fs.get("total_flows_found") is not None:
        # ⚠ 키 목록은 `_FLOW_COV_KEYS` **하나**가 출처다. 예전엔 여기에 이름을 손으로
        #   나열했는데, 생산자에 키를 추가해도 이 목록에 없으면 **조용히 버려졌다**.
        #   같은 결함을 두 번 겪었다(2026-07-31 `sds_*`, 2026-08-14 전이/체인 축) —
        #   두 번째에는 "전이 판정이 동작했는가" 를 리포트에서 확인할 수 없어,
        #   캡에 잘린 정본 지점 3개의 원인 규명이 한 라운드 늦어졌다.
        flow_cov = {k: fs[k] for k in _FLOW_COV_KEYS if k in fs}

    # SDS 기반 Related 보강 실적. **조건 없이** 싣는다 — 0 건이야말로 실어야 하는 값이다.
    # ⚠ 이걸 빠뜨렸다가 자체 감사에서 잡혔다: `collect_integration_flows` 가 `stats_out`
    #   으로 `sds_*` 를 내보내도 여기서 **이름 지정한 8개 키만** 골라 담아 전부 버려졌고,
    #   그래서 보강 실적은 로그에만 남았다(품질 리포트는 API 로 나가지만 로그는 안 나간다).
    #   "보고를 추가했다" 와 "보고가 도달한다" 는 다른 문제다.
    sds_enrich: Dict[str, Any] = {}
    if fs.get("sds_lookups") is not None:
        _lk = int(fs.get("sds_lookups") or 0)
        _hit = int(fs.get("sds_swcom_hits") or 0)
        sds_enrich = {
            "source": fs.get("sds_source"),            # argument | repo_docs_glob
            "map_entries": int(fs.get("sds_map_entries") or 0),
            "lookups": _lk,
            "key_hits": int(fs.get("sds_key_hits") or 0),
            "swcom_hits": _hit,
            # 분모 0 = 미측정(0% 아님) — 이 저장소 규약
            "yield_pct": round(100.0 * _hit / _lk, 2) if _lk else None,
        }

    # SwUDS 축 — Related 칸의 **실제** SwCom 소스. 위 sds_enrich 가 0 인 것은 결함이
    # 아니라 그 맵의 성질이고(스키마에 SwCom 이 없다), 실적은 여기서 봐야 한다.
    uds_swcom_enrich: Dict[str, Any] = {}
    if fs.get("uds_swcom_lookups") is not None:
        _ulk = int(fs.get("uds_swcom_lookups") or 0)
        _uhit = int(fs.get("uds_swcom_hits") or 0)
        uds_swcom_enrich = {
            "map_entries": int(fs.get("uds_swcom_map_entries") or 0),
            "lookups": _ulk,
            "hits": _uhit,
            "ids": int(fs.get("uds_swcom_ids") or 0),
            "yield_pct": round(100.0 * _uhit / _ulk, 2) if _ulk else None,
        }

    return {
        "total_test_cases": total_tc,
        "total_sub_cases": total_sub,
        "avg_sub_cases_per_tc": avg_sub,
        # 캡에 잘린 흐름이 있으면 비지 않는다(없으면 {} — 소비처는 .get 으로 읽는다).
        "integration_flow_coverage": flow_cov,
        # SDS 보강이 **어느 문서로 몇 건** 산출했는지. 저장소 폴백(프로젝트 무관)이면
        # source 로 드러난다.
        "sds_related_enrichment": sds_enrich,
        # SwUDS 축 — Related 칸의 SwCom 은 여기서 온다(정본과 같은 표).
        "uds_swcom_enrichment": uds_swcom_enrich,
        # Related ID **필드 보유율**(합성 포함) — 서식 채움 지표이지 추적성이 아니다.
        "with_related_count": with_related,
        "related_coverage_pct": related_pct,
        # 실제 요구/설계 ID 기준 추적성(합성 SwCom 제외) — 품질 게이트가 쓰는 값.
        "with_requirement_trace_count": with_req_trace,
        "requirement_traceability_pct": req_trace_pct,
        "synthetic_only_related_count": synthetic_only_count,
        "with_io_count": with_io,
        "io_coverage_pct": io_pct,
        "gen_method_distribution": gen_dist,
        "swcom_distribution": swcom_dist,
        "total_source_functions": total_source_functions,
    }


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_sits_xlsm(xlsm_path: str) -> Dict[str, Any]:
    """Validate generated SITS XLSM for structural and data quality."""
    issues: List[str] = []
    stats: Dict[str, Any] = {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"valid": False, "issues": ["openpyxl not installed"], "stats": {}}

    p = Path(xlsm_path)
    if not p.exists():
        return {"valid": False, "issues": [f"File not found: {xlsm_path}"], "stats": {}}

    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        return {"valid": False, "issues": [f"Cannot open: {e}"], "stats": {}}

    stats["sheets"] = wb.sheetnames
    stats["sheet_count"] = len(wb.sheetnames)

    # ⚠ 시트 이름·데이터 시작행은 **라이터와 같은 상수**를 쓴다. 문자열/숫자를 여기에
    #   복제하면 라이터가 옮겨갈 때 리더만 뒤에 남는다 — 실제로 시트는 `4.…`(라이터는
    #   `3.…` 로 이동), 시작행은 7(라이터는 5)로 굳어 있었고, 그래서 이 검증기는
    #   **자기 산출물을 한 줄도 못 읽으면서** TC 0 · sub-case 0 을 보고했다.
    required_sheets = [_SPEC_SHEET_NAME]
    for s in required_sheets:
        if s not in wb.sheetnames:
            issues.append(f"Missing required sheet: {s}")

    tc_count = 0
    sub_count = 0

    if _SPEC_SHEET_NAME in wb.sheetnames:
        ws = wb[_SPEC_SHEET_NAME]
        for row in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
            if not row:
                continue
            tc_id_val = row[_TCID_COL - 1] if len(row) >= _TCID_COL else None
            desc_val = row[_DESC_COL - 1] if len(row) >= _DESC_COL else None
            if tc_id_val and str(tc_id_val).startswith("SwITC_"):
                tc_count += 1
            elif desc_val is not None and str(desc_val).strip():
                # ⚠ 예전엔 `re.match(r"^\d", desc)` 였다 — desc 가 숫자로 시작할 때만
                # sub-case 로 셌다. 그런데 라이터는 `case_label or case_num` 을 쓰고
                # case_label 은 `COND_1 [...]`·`ERR_PROP_1 [...]`·`GLOBAL_*` 처럼 문자로
                # 시작한다. 라이터 포맷이 바뀌었는데 리더 휴리스틱이 안 따라간 것이다.
                # 실측(실 프로젝트 120 TC): 파일에 1288행이 있는데 840 만 세어 34.8% 과소,
                # avg_sub_per_tc 도 7.0(실제 10.7)이었다. 그런데 valid 는 True 였다.
                # 판정을 프리픽스 추측이 아니라 **구조**로 바꾼다: 라이터는 sub-case 행에
                # TC ID 를 절대 안 쓰고 _DESC_COL 은 항상 채운다(위 writer 참조).
                # 이 시트는 template 이 있어도 통째로 지우고 다시 만들므로 잔여행이 없다.
                sub_count += 1

        stats["tc_count"] = tc_count
        stats["flow_count"] = tc_count  # 1 flow per ITC in SITS
        stats["sub_case_count"] = sub_count
        stats["avg_sub_per_tc"] = round(sub_count / max(tc_count, 1), 1)

        if tc_count == 0:
            issues.append("No test cases (SwITC_*) found")
        if sub_count == 0:
            issues.append("No sub-cases found")

    wb.close()
    return {"valid": len(issues) == 0, "issues": issues, "stats": stats}


# ---------------------------------------------------------------------------
# Validation report
# ---------------------------------------------------------------------------

def generate_sits_validation_report(
    xlsm_path: str,
    quality_report: Optional[Dict[str, Any]] = None,
    validation: Optional[Dict[str, Any]] = None,
) -> str:
    """Write .validation.md file next to XLSM and return its path."""
    validation_data = validation if isinstance(validation, dict) else validate_sits_xlsm(xlsm_path)
    stats = validation_data.get("stats", {})
    issues = validation_data.get("issues", [])
    qr = quality_report or {}

    lines = [
        "# SITS 생성 문서 자동 검증 리포트",
        "",
        f"**파일**: `{Path(xlsm_path).name}`  ",
        f"**검증 시각**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ",
        f"**결과**: {'PASS' if validation_data.get('valid') else 'FAIL'}",
        "",
        "---",
        "",
        "## 1. 구조 검증",
        "",
        "| 항목 | 값 |",
        "|------|-----|",
        f"| 시트 수 | {stats.get('sheet_count', 0)} |",
        f"| 시트 목록 | {', '.join(stats.get('sheets', []))} |",
        f"| TC 수 (SwITC_*) | {stats.get('tc_count', 0)} |",
        f"| Sub-case 수 | {stats.get('sub_case_count', 0)} |",
        f"| TC당 평균 Sub-case | {stats.get('avg_sub_per_tc', 0)} |",
        "",
    ]

    if qr:
        lines += [
            "## 2. 품질 지표",
            "",
            "| 항목 | 값 |",
            "|------|-----|",
            f"| 총 TC 수 | {qr.get('total_test_cases', 0)} |",
            f"| 총 Sub-case 수 | {qr.get('total_sub_cases', 0)} |",
            f"| Related ID 보유 TC | {qr.get('with_related_count', 0)} |",
            f"| Related ID 커버리지 (합성 포함) | {qr.get('related_coverage_pct', 0)}% |",
            f"| 요구 추적성 (합성 SwCom 제외) | {qr.get('requirement_traceability_pct', 0)}% |",
            f"| 합성 ID만 있는 TC | {qr.get('synthetic_only_related_count', 0)} |",
            f"| I/O 파라미터 보유 TC | {qr.get('with_io_count', 0)} |",
            f"| I/O 커버리지 | {qr.get('io_coverage_pct', 0)}% |",
            f"| 생성 방법 분포 | {qr.get('gen_method_distribution', {})} |",
            "",
        ]

    if issues:
        lines += ["## 3. 이슈", ""]
        for iss in issues:
            lines.append(f"- ❌ {iss}")
    else:
        lines += ["## 3. 이슈", "", "- 이슈 없음"]

    report_path = Path(xlsm_path).with_suffix(".validation.md")
    report_path.write_text("\n".join(lines), encoding="utf-8")
    _logger.info("SITS validation report: %s", report_path.name)
    return str(report_path)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_sits(
    source_root: str,
    output_path: str,
    template_path: Optional[str] = None,
    project_config: Optional[Dict[str, Any]] = None,
    ai_config: Optional[Dict[str, Any]] = None,
    max_subcases: int = _DEFAULT_SUBCASES,
    on_progress: Optional[Any] = None,
    srs_docx_path: Optional[str] = None,
    sds_docx_path: Optional[str] = None,
    uds_path: Optional[str] = None,
    hsis_path: Optional[str] = None,
    stp_path: Optional[str] = None,
    # ⚠ 신규 인자는 **맨 끝**에 붙인다. 중간에 끼우면 위치 인자로 부르는 호출부가
    #    조용히 다른 값에 바인딩된다(현재 호출부 4곳은 전부 키워드지만 계약은 지킨다).
    max_flows: int = _DEFAULT_MAX_FLOWS,
    # 고장 주입 대상 **설계 ID**(예: `["SwFn_07", "SwFn_34"]`). 이 선별은 안전분석
    # 산출물이라 소스/문서에서 유도되지 않는다 — 8가지 근거 실측은
    # `generate_itc_list` 의 FI 블록 주석 참조. 주지 않으면 전용 FI TC 는 0건이고
    # 그 사실이 로그와 `fi_requested`(=0)로 남는다.
    fi_design_ids: Optional[Sequence[str]] = None,
) -> Dict[str, Any]:
    """Top-level SITS generation pipeline.

    Args:
        source_root: Root directory of C source code
        output_path: Path for output XLSM file
        template_path: Optional SITS template XLSM
        project_config: Optional config dict (project_id, version, asil_level, doc_id)
        ai_config: Optional AI config dict (reserved, not used yet)
        max_subcases: Maximum sub-cases per TC (default _DEFAULT_SUBCASES = 14)
            — 중복 기재돼 있었고 "default 5"·"default 7" 둘 다 실제 값과 달랐다.
        on_progress: Optional callback(pct: int, message: str)
        srs_docx_path: Optional SRS DOCX for requirement ID enrichment
        sds_docx_path: Optional SDS DOCX for component context
        uds_path: Optional UDS DOCX/XLSM for function descriptions
        hsis_path: Optional HSIS XLSX for hardware signal context
        max_flows: 통합 흐름 상한(default _DEFAULT_MAX_FLOWS = 120). 걸리면 안전등급
            높은 흐름부터 남기고, 잘린 내역이 로그 + quality_report
            ["integration_flow_coverage"] 에 남는다. 실측 프로젝트에서 145개 중
            25개가 이 값에 걸린다 — 규격에 없는 흐름이 그만큼 생긴다는 뜻이다.
        stp_path: Optional STP DOCX for test strategy context

    Returns:
        Dict with: output_path, quality_report, test_case_count, total_sub_cases,
                   elapsed_seconds, validation, validation_report_path
    """
    def _progress(pct: int, msg: str) -> None:
        _logger.info("[%d%%] %s", pct, msg)
        if on_progress:
            try:
                on_progress(pct, msg)
            except Exception:
                pass

    _logger.info("=== SITS Generation Start ===")
    t0 = time.time()

    _progress(5, "SITS 생성 시작")

    # ── Stage 1-4: document context loading ─────────────────────────────────
    # ⚠ SITS 는 `sds_docx_path` 를 받고도 **Related ID 보강에는 쓰지 않았다** — 흐름
    #   수집이 저장소 `docs/` 글롭(프로젝트 무관, 현재 HDPDM01)만 봤다. SUTS 가 정확히
    #   같은 결함을 이미 고쳐 뒀고(`suts._resolve_sds_map` docstring 참조) 그 헬퍼를
    #   **재사용**한다 — 복제하면 한쪽만 고쳐지는 이 저장소의 반복 실패 모드가 된다.
    _project_sds_map: Optional[Dict[str, Dict[str, str]]] = None
    if sds_docx_path:
        _progress(7, "SDS 설계 컨텍스트 로드 중")
        try:
            from generators.suts import _resolve_sds_map
            _project_sds_map = _resolve_sds_map(sds_docx_path)
        except Exception as e:  # noqa: BLE001 - 확보 실패는 폴백 사유로 보고만 한다
            _logger.warning("SITS: 프로젝트 SDS 맵 확보 실패(%s) — 저장소 docs/ 폴백으로 "
                            "넘어간다(프로젝트 무관): %s", type(e).__name__, e)
        try:
            from generators.sts import _load_sds_summary
            sds_summary = _load_sds_summary(sds_docx_path)
            if sds_summary:
                _logger.info("SITS: SDS summary loaded (%d chars)", len(sds_summary))
        except Exception as e:
            _logger.debug("SITS: SDS load skipped: %s", e)

    if uds_path:
        _progress(8, "UDS 함수 설명 로드 중")
        try:
            from generators.sts import _load_uds_descriptions
            _uds_descs = _load_uds_descriptions(uds_path)
            if _uds_descs:
                _logger.info("SITS: UDS descriptions loaded (%d entries)", len(_uds_descs))
        except Exception as e:
            _logger.debug("SITS: UDS load skipped: %s", e)

    stp_context: Dict[str, Any] = {}
    if stp_path:
        _progress(9, "STP 시험 전략 로드 중")
        try:
            stp_context = _parse_stp_document(stp_path)
        except Exception as e:
            _logger.debug("SITS: STP load skipped: %s", e)

    if hsis_path:
        _progress(10, "HSIS 신호 로드 중")
        try:
            from generators.sts import _load_hsis_signals
            _hsis = _load_hsis_signals(hsis_path)
            if _hsis:
                _logger.info("SITS: HSIS signals loaded")
        except Exception as e:
            _logger.debug("SITS: HSIS load skipped: %s", e)

    # ── Stage 5: source parsing ──────────────────────────────────────────────
    _progress(15, "소스 코드 파싱 시작")
    # 콤마 구분 복수 경로 지원: 첫 번째 경로로 검증
    _first_root = source_root.split(",")[0].strip() if source_root else ""
    source_root_path = Path(_first_root).resolve() if _first_root else None
    if not source_root_path or not source_root_path.is_dir():
        return {
            "output_path": "",
            "quality_report": {},
            "test_case_count": 0,
            "total_sub_cases": 0,
            "elapsed_seconds": round(time.time() - t0, 1),
            "error": "유효한 소스 코드 루트 경로가 없습니다.",
        }

    function_details: Dict[str, Dict[str, Any]] = {}
    total_source_functions = 0
    try:
        try:
            from backend.helpers import _get_source_sections_cached
            report_data = _get_source_sections_cached(source_root)  # 콤마 구분 그대로 전달
        except Exception:
            from report_generator import generate_uds_source_sections
            report_data = generate_uds_source_sections(source_root)  # 콤마 구분 그대로 전달
        function_details = report_data.get("function_details", {})
        # 배열 **선언 크기**의 출처. 이걸 안 꺼내면 흐름 수집이 `(size: N)` 꼬리가
        # 붙은 엔트리만 펼치고, 문서 유래 이름(꼬리 없음)은 base 한 칸으로 나간다
        # — SUTS 가 7차 라운드에서 겪은 사각과 같은 형태다.
        _globals_info_map = report_data.get("globals_info_map", {}) or {}
        _struct_members = report_data.get("struct_member_arrays", {}) or {}
        total_source_functions = len(function_details)
        if not function_details:
            raise ValueError("No function_details in source parse result")
    except Exception as e:
        _logger.warning("SITS: full source parse failed, trying lightweight: %s", e)
        try:
            from generators.suts import _lightweight_parse
            function_details = _lightweight_parse(_first_root)
            # 경량 파서는 전역 선언표를 안 만든다 — 없는 것을 빈 맵으로 둔다
            # (그러면 `(size: N)` 꼬리가 있는 엔트리만 펼쳐지고, 그 사실은
            #  `array_size_map_entries` 0 으로 보인다).
            _globals_info_map = {}
            _struct_members = {}
            total_source_functions = len(function_details)
        except Exception as e2:
            _logger.error("SITS: lightweight parse also failed: %s", e2)
            return {
                "output_path": "",
                "quality_report": {},
                "test_case_count": 0,
                "total_sub_cases": 0,
                "elapsed_seconds": round(time.time() - t0, 1),
                "error": f"소스 파싱 실패: {e2}",
            }

    _progress(30, f"소스 파싱 완료 — {total_source_functions}개 함수 발견")

    # SRS requirement ID enrichment — per-function mapping
    if srs_docx_path:
        _progress(32, "SRS 요구사항 ID 매핑 중")
        try:
            from generators.sts import parse_srs_docx_tables
            reqs = parse_srs_docx_tables(srs_docx_path)
            if reqs:
                _logger.info("SITS: SRS reqs loaded (%d)", len(reqs))

                # Build a map: fn_name_lower → [req_ids] by scanning each requirement's
                # description for function names.  Only exact word-boundary matches count
                # to avoid "get" matching "get_speed", "get_torque", etc.
                _fn_names_lower = {
                    str(info.get("name") or "").lower(): fid
                    for fid, info in function_details.items()
                    if isinstance(info, dict) and len(str(info.get("name") or "")) >= 4
                }
                fn_to_req_ids: Dict[str, List[str]] = {}
                for req in reqs:
                    req_id = str(req.get("id") or "").strip()
                    if not req_id:
                        continue
                    req_desc = str(req.get("description") or "").lower()
                    for fn_lower in _fn_names_lower:
                        # Word-boundary match: function name must appear as whole word
                        if re.search(r"\b" + re.escape(fn_lower) + r"\b", req_desc):
                            fn_to_req_ids.setdefault(fn_lower, [])
                            if req_id not in fn_to_req_ids[fn_lower]:
                                fn_to_req_ids[fn_lower].append(req_id)

                # Annotate function_details
                matched = 0
                for fid, info in function_details.items():
                    if not isinstance(info, dict):
                        continue
                    fn_lower = str(info.get("name") or "").lower()
                    ids = fn_to_req_ids.get(fn_lower)
                    if ids:
                        info.setdefault("srs_req_ids", ", ".join(ids[:3]))
                        matched += 1
                _logger.info("SITS: SRS enrichment: %d functions matched", matched)
        except Exception as e:
            _logger.debug("SITS: SRS enrichment skipped: %s", e)

    # UDS description enrichment
    if uds_path:
        try:
            from generators.sts import _load_uds_descriptions, _merge_uds_into_function_details
            uds_descs = _load_uds_descriptions(uds_path)
            if uds_descs:
                _merge_uds_into_function_details(function_details, uds_descs)
        except Exception as e:
            _logger.debug("SITS: UDS enrichment skipped: %s", e)

    # ── Stage 6: collect integration flows ───────────────────────────────────
    _progress(40, "통합 흐름 수집 중")
    # Related ID 의 SwCom 축은 **SwUDS** 에서 온다(`load_uds_swcom_map` docstring — 정본과
    # 같은 표다). 못 얻으면 순번 합성 ID 로 내려가되 합성임이 산출물에 표시된다.
    _uds_swcom_map = load_uds_swcom_map(uds_path)
    # Safety Related 의 1순위 근거도 SwUDS 다(`_resolve_flow_asil`) — 소스 주석만 보면
    # 근거 없는 260건이 `QM`→`X`(비안전)로 실린다.
    _uds_asil_map = load_uds_asil_map(uds_path)
    # Related 칸은 SwCom 만이 아니다(SwFn·SwSTR·SwST·SwTK) — 넓은 판을 따로 받는다.
    _uds_related_map = load_uds_related_map(uds_path)
    flow_stats: Dict[str, Any] = {}
    flows = collect_integration_flows(
        function_details, max_flows=max_flows, stats_out=flow_stats,
        sds_map=_project_sds_map, uds_swcom_map=_uds_swcom_map,
        uds_asil_map=_uds_asil_map, uds_related_map=_uds_related_map,
        globals_info_map=_globals_info_map, struct_members=_struct_members)

    if not flows:
        _logger.warning("SITS: No integration flows found — check cross-module calls in source")
        return {
            "output_path": "",
            "quality_report": {},
            "test_case_count": 0,
            "total_sub_cases": 0,
            "elapsed_seconds": round(time.time() - t0, 1),
            "error": "통합 흐름을 찾을 수 없습니다. 소스 파싱 결과를 확인해주세요.",
        }

    # ⚠ "수집 완료" 는 캡에 잘렸을 때 완결을 주장하는 거짓말이 된다. 잘렸으면 그렇게 쓴다.
    _dropped_flows = int(flow_stats.get("flows_dropped") or 0)
    if _dropped_flows:
        _progress(
            50,
            f"{len(flows)}개 통합 흐름 수집 — 전체 {flow_stats.get('total_flows_found')}개 중 "
            f"{_dropped_flows}개는 max_flows 캡으로 제외(규격에 미포함)",
        )
    else:
        _progress(50, f"{len(flows)}개 통합 흐름 수집 완료")

    # ── Stage 6b: balance over-concentrated Related IDs ──────────────────────
    flows = _balance_related_ids(flows)

    # ── Stage 7: generate ITCs ───────────────────────────────────────────────
    _progress(60, "통합 테스트 케이스 생성 중")
    stp_envs = stp_context.get("environments") or []
    # TC ID 는 **SwUDS 설계 ID** 를 쓴다(`generate_itc_list` docstring — 정본 실측).
    # 못 얻으면 순번으로 내려가되 그 사실이 로그·리포트에 남는다.
    _design_ids = load_uds_design_ids(uds_path) if uds_path else None
    # ── 전용 FI TC 의 대상 흐름 ────────────────────────────────────────
    # 지정된 설계 ID 를 SwUDS Related 맵으로 역인덱스해 그 설계 요소를 실현하는
    # 함수들의 흐름을 찾는다. 못 찾은 ID 는 **조용히 버리지 않고** 센다
    # (요청은 있었는데 0 건 나가는 것과 요청이 없던 것은 다른 말이다).
    _fi_flows: List[Dict[str, Any]] = []
    _fi_unresolved: List[str] = []
    if fi_design_ids:
        _by_design: Dict[str, List[str]] = {}
        for _fn, _toks in (_uds_related_map or {}).items():
            for _tk in _toks or []:
                _by_design.setdefault(str(_tk), []).append(str(_fn))
        _flow_by_fn = {str(f.get("entry_fn") or "").lower(): f for f in flows}
        for _did in fi_design_ids:
            _did = str(_did or "").strip()
            if not _did:
                continue
            _hit = next((_flow_by_fn[_f.lower()] for _f in _by_design.get(_did, [])
                         if _f.lower() in _flow_by_fn), None)
            if _hit is None:
                _fi_unresolved.append(_did)
                continue
            _fi_flows.append({**_hit, "fi_design_id": _did})
        if _fi_unresolved:
            _logger.warning(
                "SITS: FI 대상 %d건이 통합 흐름에 없어 TC 를 못 냈다: %s "
                "(SwUDS Related 맵 %d함수 · 흐름 %d개)",
                len(_fi_unresolved), ", ".join(_fi_unresolved[:8]),
                len(_uds_related_map or {}), len(flows),
            )
    itcs = generate_itc_list(flows, max_subcases=max_subcases,
                             stp_environments=stp_envs or None,
                             design_ids=_design_ids, stats_out=flow_stats,
                             fi_flows=_fi_flows or None)
    if flow_stats is not None:
        flow_stats["fi_unresolved"] = len(_fi_unresolved)

    _progress(65, f"{len(itcs)}개 TC, {sum(len(t['sub_cases']) for t in itcs)}개 sub-case 생성 완료")

    # ── Stage 8: quality report ──────────────────────────────────────────────
    _progress(70, "품질 보고서 생성 중")
    quality_report = generate_sits_quality_report(
        itcs, total_source_functions, flow_stats=flow_stats)

    # ── Stage 9: XLSM generation ─────────────────────────────────────────────
    _progress(80, "XLSM 파일 생성 중")
    # 근거 시트용 컨텍스트. 호출 그래프·파일 귀속·절대 깊이는 이미 파싱해 둔 것에서
    # 나오므로 새 IO 는 없다. `stats_out` 은 **같은 dict** 를 넘겨 시트 실적이 아래에서
    # 품질 리포트로 합류하게 한다.
    _calls_map_all = _build_calls_map(function_details)
    _file_of = _build_file_map(function_details)
    _strategy_ctx = {
        "calls_map": _calls_map_all,
        "file_of": _file_of,
        "depth_of": _absolute_depth_map(_calls_map_all, _STRATEGY_ROOTS),
        "uds_related_map": _uds_related_map,
        "design_ids": _design_ids,
        "stats_out": flow_stats,
    }
    try:
        # ── Introduction 1.4 Reference — **실제로 읽은** 입력 문서만 싣는다 ──
        #   정본은 Glossary/SwRS/SDS/SwTP 4행을 파일명+버전까지 적는다. 우리 템플릿은
        #   File Name 열이 통째로 비어 있고 Note 는 "HW 요구사항 명세서" 였다.
        #   ⚠ 못 읽은 문서의 줄은 **비워 둔다** — 참조했다고 적으면 문서가 거짓말한다.
        _refs: List[Tuple[str, str]] = []
        for _p, _note in ((srs_docx_path, "SW 요구사항 명세서"),
                          (sds_docx_path, "SW 아키텍처 설계서"),
                          (uds_path, "SW 상세 설계서"),
                          (stp_path, "SW 테스트 계획서")):
            if _p:
                _refs.append((Path(str(_p)).name, _note))
        _front = {
            "date": datetime.now().astimezone().strftime("%Y.%m.%d"),
            "author": (project_config or {}).get("author") or "",
            "status": (project_config or {}).get("status") or "Open",
            "glossary": f"{(project_config or {}).get('project_id') or 'PROJECT'} Glossary",
            "references": _refs,
            # 정본은 환경 ID(`SwITE_01`)를 적는다. STP 에서 못 얻으면 안 적는다.
            "test_env_id": (stp_context or {}).get("integration_env_id") or "",
        }
        actual_output = generate_sits_xlsm(
            template_path=template_path,
            itcs=itcs,
            output_path=output_path,
            project_config=project_config,
            flows=flows,
            stp_context=stp_context,
            strategy_context=_strategy_ctx,
            front_matter=_front,
        )
    except Exception as e:
        _logger.error("SITS: XLSM generation failed: %s", e)
        return {
            "output_path": "",
            "quality_report": quality_report,
            "test_case_count": len(itcs),
            "total_sub_cases": sum(len(t["sub_cases"]) for t in itcs),
            "elapsed_seconds": round(time.time() - t0, 1),
            "error": f"XLSM 생성 실패: {e}",
        }

    # ⚠ 근거 시트 실적은 품질 리포트를 만든 **뒤에** 채워진다(시트를 그때 쓰므로).
    #   그대로 두면 `_FLOW_COV_KEYS` 에 키를 넣어도 리포트엔 영영 안 실린다 — 같은 필터를
    #   한 번 더 적용해 합류시킨다("보고를 추가했다" 와 "보고가 도달한다" 는 다른 문제다).
    _late = {k: flow_stats[k] for k in _FLOW_COV_KEYS if k in flow_stats}
    if _late:
        quality_report.setdefault("integration_flow_coverage", {}).update(_late)
    if flow_stats.get("strategy_error"):
        quality_report["strategy_sheet_error"] = flow_stats["strategy_error"]

    # ── Stage 9.5: save intermediate JSON for VectorCAST export ─────────────
    try:
        _intermediate: Dict[str, Any] = {
            "schema_version": "1.0",
            "project_id": (project_config or {}).get("project_id", "PROJECT"),
            "source": {
                "source_root": source_root,
                "sits_path": actual_output,
                "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            },
            "integrations": [
                {
                    "tc_id": itc["tc_id"],
                    "entry_fn": itc["entry_fn"],
                    "call_chain": itc["call_chain"],
                    "module_name": itc["module_name"],
                    "gen_method": itc["gen_method"],
                    "asil": itc.get("asil", "QM"),
                    # `related_ids` 는 설계 요소(SwDS 칸), `req_ids` 는 SRS 유래 요구
                    # 링크다. 둘을 한 칸에 섞으면 감사자가 구별할 수 없어 분리했다.
                    "metadata": {"related_ids": itc["related_ids"],
                                 "req_ids": itc.get("req_ids") or []},
                    "sub_cases": [
                        {
                            "case_num": sc.get("case_num", i + 1),
                            "case_label": sc.get("case_label", str(i + 1)),
                            "precondition": sc.get("precondition", ""),
                            "inputs": sc.get("inputs") or {},
                            "expected": sc.get("expected") or {},
                        }
                        for i, sc in enumerate(itc.get("sub_cases") or [])
                    ],
                }
                for itc in itcs
            ],
            "export_warnings": [],
        }
        _intermediate_path = Path(actual_output).with_name(
            Path(actual_output).stem + "_vectorcast.json"
        )
        _intermediate_path.write_text(
            json.dumps(_intermediate, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        _logger.info("SITS: intermediate JSON saved → %s", _intermediate_path)
    except Exception as _e:
        _logger.warning("SITS: intermediate JSON save failed: %s", _e)

    # ── Stage 10: validation ─────────────────────────────────────────────────
    _progress(90, "XLSM 검증 중")
    validation = validate_sits_xlsm(actual_output)
    # 파일에서 되읽은 수가 실제로 만든 수와 같은지 대조한다. 이게 없으면 라이터가
    # 흘려도 `valid: True` 가 나오고, 호출자에게 가는 test_case_count 는 파일이 아니라
    # 생성기가 세어준 값이라 아무도 눈치채지 못한다.
    validation = apply_write_back_check(validation, {
        "tc_count": len(itcs),
        "sub_case_count": sum(len(t.get("sub_cases") or []) for t in itcs),
    })
    if not validation.get("valid"):
        _logger.warning("SITS validation issues: %s", validation.get("issues"))

    # ── Stage 11: validation report ──────────────────────────────────────────
    _progress(95, "검증 보고서 생성 중")
    validation_report_path = ""
    try:
        validation_report_path = generate_sits_validation_report(
            actual_output, quality_report, validation
        )
    except Exception as e:
        _logger.warning("SITS: validation report generation failed: %s", e)

    elapsed = round(time.time() - t0, 1)
    _progress(100, f"SITS 생성 완료 ({elapsed}s)")
    _logger.info("=== SITS Generation Done: %d TCs, %d sub-cases, %.1fs ===",
                 len(itcs), sum(len(t["sub_cases"]) for t in itcs), elapsed)

    # Quality DB recording (non-fatal)
    try:
        from workflow.quality.recorder import record_run
        record_run(
            "sits", quality_report,
            project_root=str(source_root or ""),
            elapsed_sec=elapsed,
            output_path=actual_output,
            ai_model=str((ai_config or {}).get("model", "")),
        )
    except Exception:
        # non-fatal 은 유지하되 침묵은 금지 (sts.py 의 동일 블록이 NameError 를
        # 몇 년간 삼켜 품질 기록이 통째로 유실된 전례).
        _logger.exception("SITS quality record skipped (non-fatal)")

    return {
        "output_path": actual_output,
        "quality_report": quality_report,
        "test_case_count": len(itcs),
        "total_sub_cases": sum(len(t["sub_cases"]) for t in itcs),
        "elapsed_seconds": elapsed,
        "validation": validation,
        "validation_report_path": validation_report_path,
    }
