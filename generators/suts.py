"""SUTS (Software Unit Test Specification) auto-generation engine.

Generates XLSM output from UDS function details and source code analysis.
Each unit function gets a dedicated TC with input/output variable columns
and multiple test sequences (boundary values, error conditions, etc.).
"""
from __future__ import annotations

import logging
import os
import re
import tempfile
import time
from contextlib import contextmanager
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from generators._artifact_check import apply_write_back_check
from report_gen.doc_kind import is_sds_filename
from report_gen.requirements import _extract_sds_partition_map
from report_gen.source_parser import is_const_type
from workflow.code_parser.c_parser import blank_c_comments

_logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# ── 시트 레이아웃 — **납품 정본 기준** (KJPDS02_SwUTS v1.02, 189열) ──────────
#
# ⚠ 회사 표준 템플릿(v0.10)이 아니라 **정본**을 따른다. 실측(2026-08-11):
#   - 표준 템플릿 v0.10 = 28열. Input/Expected 가 `Param 1~10` 고정이고
#     Safety Related·Test Method 열이 **없다**. 실제 함수는 파라미터가 최대 96개라
#     템플릿 폭으로는 담기지 않는다.
#   - 정본 v1.02 = 189열. 프로젝트가 템플릿을 확장한 형태이고, 이것이 납품물이다.
#
# ⚠ 이전 판은 셋 중 **어느 것도 아닌 제3의 레이아웃**이었다(149열). `Description`
#   `Test Environment` `Precondition` `Sequence` 4열은 **STS 정본의 열**이라
#   SUTS 에 와 있으면 안 된다 — 열이 밀려 정본 파서가 전부 잘못 읽는다.
#
# 정본 실측 구조:
#   r3 밴드 : B3:G3 'Test Case' · H3:CZ3 'Input' · DA3:GF3 'Expected Result' · GG3 'Related ID'
#   r4 헤더 : B Index · C TC_ID · D Unit · E Safety Related · F Test Method
#             · G Test Case Generation Method · H ' '(시퀀스 번호) · I~ Inpt[n] · DA~ ExpR[n] · GG SUDS
#   r5~     : TC 블록 = 변수명 행 1개 + 시퀀스 행 N개 (B/C/D/E/GG 는 블록 전체 병합)
_BAND_ROW = 3
_HEADER_ROW = 4
_DATA_START_ROW = 5

_COL_INDEX = 2             # B   Index (연번 — 정본은 1..1014 연속)
_COL_TC_ID = 3             # C   TC_ID
_COL_UNIT = 4              # D   Unit (함수명)
_COL_SAFETY = 5            # E   Safety Related (O/X)
_COL_METHOD = 6            # F   Test Method (REQ/FI) — **시퀀스 그룹 단위**
_COL_GEN = 7               # G   Test Case Generation Method — 시퀀스 그룹 단위
_SEQ_COL = 8               # H   시퀀스 번호 (헤더는 공백 한 칸 — 정본 그대로)
_INPUT_COL_START = 9       # I   Inpt[0]
_INPUT_COL_END = 104       # CZ  Inpt[95]
_OUTPUT_COL_START = 105    # DA  ExpR[0]
_OUTPUT_COL_END = 188      # GF  ExpR[83]
_RELATED_COL = 189         # GG  SUDS

# 헤더 행(열 번호 → 라벨). `generate_suts_xlsm`이 시트에 쓰는 값이자, 영향도 탭의
# 문서 초안이 Excel 붙여넣기 TSV 열 순서를 얻는 **단일 출처**다(복제 금지).
_FIXED_HEADERS = {
    _COL_INDEX: "Index",
    _COL_TC_ID: "TC_ID",
    _COL_UNIT: "Unit",
    _COL_SAFETY: "Safety Related",
    _COL_METHOD: "Test Method",
    _COL_GEN: "Test Case Generation Method",
    _SEQ_COL: " ",
}
_RELATED_HEADER = "SUDS"   # Related ID 컬럼 라벨

# ── 값 어휘 — 정본과 Introduction(1.5/1.6)에서 온다 ─────────────────────────
#
# ⚠ 이전 판은 `FIT`/`FNCT`/`RVW` 를 썼다. 그건 **STS 어휘**이고 SwUTS Introduction
#   1.5 표에 아예 없는 값이다. 정본 실측: REQ 1,437 · FI 815 (그 둘뿐).
_METHOD_REQ = "REQ"        # Requirements based test — 유효 범위 시험
_METHOD_FI = "FI"          # Fault Injection Test — 유효 범위 밖 시험
# 유효 범위를 벗어나는 값을 넣는 전략 = 고장 주입. 정본도 경계 초과 시퀀스를 FI 로 묶는다
# (첫 TC: seq 1~3 REQ / 4~7 FI).
_FI_STRATEGIES = frozenset({"BV_MIN_INV", "BV_MAX_INV", "ERROR_PATH"})

# ⚠ 결합자는 **문서마다 다르다**. SwUTS 정본은 슬래시(`AOR/ABV`), SwITS 정본은
#   쉼표(`AOR, AEC`). 통일하지 말 것 — 각 정본을 따른다.
#   정본 실측: AOR/ABV 1,638 · AOR/AEC 636 (그 둘뿐).
_GEN_BOUNDARY = "AOR/ABV"  # 경계값 분석
_GEN_EQUIV = "AOR/AEC"     # 등가 분할(조건·분기 조합)

_MAX_SEQUENCES = 10
_DEFAULT_SEQ_COUNT = 24  # 6 BV + 4 COND + 6 SWITCH + 3 LOOP + 3 GLOBAL + 1 VOID + 6 MC/DC

_GEN_METHODS = {"AEC, ABV", "ABV, AOR", "AOR", "ABV"}
_DEFAULT_GEN_METHOD = "AEC, ABV"
_DEFAULT_TEST_ENV = "SwTE_01"

_SDS_MAP_CACHE: Optional[Dict[str, Dict[str, str]]] = None


def _merge_sds_partition_map(
    merged: Dict[str, Dict[str, str]], data: Dict[str, Dict[str, str]]
) -> None:
    """first-wins 병합 — 이미 값이 있는 필드는 덮어쓰지 않는다."""
    for key, value in data.items():
        if key not in merged:
            merged[key] = dict(value)
            continue
        for field in ("asil", "related", "description"):
            if value.get(field) and not merged[key].get(field):
                merged[key][field] = value[field]


def load_sds_map_from(sds_docx_path: str) -> Dict[str, Dict[str, str]]:
    """사용자가 지정한 SDS 문서 하나에서 파티션 맵(ASIL/related/description)을 읽는다.

    `_load_default_sds_map`(저장소 `docs/` 글롭)과 달리 **경로를 그대로 존중**한다.
    SUTS 생성기는 오래도록 `sds_docx_path` 인자를 받고도 본문에서 쓰지 않아,
    프로젝트가 무엇이든 저장소 `docs/`에 들어있는 SDS(현재 HDPDM01)로 ASIL을 채웠다
    — 다른 프로젝트의 안전 등급이 조용히 섞이는 경로였다.
    """
    if not sds_docx_path:
        return {}
    merged: Dict[str, Dict[str, str]] = {}
    try:
        _merge_sds_partition_map(merged, _extract_sds_partition_map(sds_docx_path))
    except Exception as exc:
        _logger.warning("SDS 파티션 맵 파싱 실패 — ASIL 보강 생략: %s (%s)", sds_docx_path, exc)
        return {}
    return merged


def _resolve_sds_map(sds_docx_path: Optional[str]) -> Optional[Dict[str, Dict[str, str]]]:
    """SUTS ASIL 보강에 쓸 SDS 맵을 확보한다. None이면 호출자가 폴백을 쓴다.

    입력은 resolver 경유(`_resolved_doc_input`)라 cloudium worker-only 경로도 잡는다.
    지정했는데 못 쓰게 된 경우는 **반드시 경고를 남긴다** — 폴백(저장소 `docs/` 글롭)이
    조용히 대신하면, 다른 프로젝트의 ASIL로 채워진 산출물을 정상으로 오인한다.
    """
    if not sds_docx_path:
        return None
    with _resolved_doc_input(sds_docx_path, "SDS") as local:
        if not local:
            _logger.warning(
                "SUTS: SDS 입력을 확보하지 못해 ASIL 보강이 저장소 docs/ 폴백(프로젝트 무관)으로 "
                "넘어간다: %s", sds_docx_path)
            return None
        sds_map = load_sds_map_from(local)
    if not sds_map:
        _logger.warning(
            "SUTS: SDS를 지정했으나 파티션 0건 — ASIL 보강이 저장소 docs/ 폴백(프로젝트 무관)으로 "
            "넘어간다: %s", sds_docx_path)
        return None
    _logger.info("SUTS: SDS 파티션 %d건 로드 — ASIL 출처=%s", len(sds_map), sds_docx_path)
    return sds_map


def _load_default_sds_map() -> Dict[str, Dict[str, str]]:
    """저장소 `docs/`의 SDS 글롭 폴백.

    ⚠ 프로젝트 무관이다 — 호출자가 SDS 경로를 알고 있으면 `load_sds_map_from`을 쓸 것.
    """
    global _SDS_MAP_CACHE
    if _SDS_MAP_CACHE is not None:
        return _SDS_MAP_CACHE
    docs_dir = Path(__file__).resolve().parents[1] / "docs"
    merged: Dict[str, Dict[str, str]] = {}
    picked: List[str] = []
    if docs_dir.exists():
        for path in docs_dir.glob("*.docx"):
            # `"sds" in name` 은 `SwDS` 표기를 놓친다("swds" 에 "sds" 없음) — 단일 출처 사용.
            if not is_sds_filename(path.name):
                continue
            picked.append(path.name)
            _merge_sds_partition_map(merged, _extract_sds_partition_map(str(path)))
    if merged:
        # ⚠ 침묵 금지 — 이 맵으로 단위 ASIL 을 채우는데 출처가 **다른 프로젝트**일 수 있다.
        _logger.warning(
            "SDS 미지정 — 저장소 docs/ 글롭 폴백 사용(**프로젝트 무관**): %s (%d 엔트리). "
            "대상 프로젝트의 SDS 를 `load_sds_map_from` 으로 넘기면 이 폴백은 쓰이지 않는다",
            ", ".join(picked) or "(없음)", len(merged))
    _SDS_MAP_CACHE = merged
    return merged


def _resolve_unit_asil(info: Dict[str, Any], sds_map: Dict[str, Dict[str, str]]) -> str:
    def _norm(value: str) -> str:
        return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

    module_name = str(info.get("module_name") or "").strip()
    candidates: List[str] = []
    if module_name:
        candidates.append(module_name)
        base = re.sub(r"_pds$", "", module_name, flags=re.I)
        candidates.append(base)
        tokenized = re.sub(r"([a-z])([A-Z])", r"\1 \2", base.replace("_", " "))
        tokenized = re.sub(r"\bctrl\b", "control", tokenized, flags=re.I)
        tokenized = re.sub(r"\bdiag\b", "diagnostic", tokenized, flags=re.I)
        words = [w for w in tokenized.split() if w.lower() not in {"ap", "drv", "sys", "pds", "main", "func"}]
        if words:
            candidates.append(" ".join(words))
    for candidate in candidates:
        direct = sds_map.get(candidate.lower())
        if direct and direct.get("asil"):
            return str(direct["asil"]).strip()
    for candidate in candidates:
        nc = _norm(candidate)
        if not nc:
            continue
        for key, value in sds_map.items():
            nk = _norm(key)
            if not nk:
                continue
            if nc == nk or nc in nk or nk in nc:
                return str(value.get("asil") or "").strip()
    return ""

_SRS_REQ_ID_PAT = re.compile(
    r"\b(?:SW[_R]?|SRS|Sw|HDPDM\d*|SWR|SWS|SYSRS)[_-]?\d[\w_-]*",
    re.I,
)


def _resolve_srs_req_ids_for_function(
    func_name: str,
    sds_map: Dict[str, Dict[str, str]],
) -> str:
    """Resolve SRS requirement IDs for a function via SDS partition map `related` field."""
    if not sds_map or not func_name:
        return ""
    candidates = [func_name.lower(), func_name.lower().replace("_", " ")]
    for candidate in candidates:
        entry = sds_map.get(candidate)
        if entry:
            related = str(entry.get("related", "") or "")
            ids = _SRS_REQ_ID_PAT.findall(related)
            if ids:
                return ", ".join(ids[:4])
    # Fuzzy: partial name match
    fn_lower = func_name.lower()
    for key, entry in sds_map.items():
        if fn_lower in key or key in fn_lower:
            related = str(entry.get("related", "") or "")
            ids = _SRS_REQ_ID_PAT.findall(related)
            if ids:
                return ", ".join(ids[:4])
    return ""


# C type boundary values (min_invalid, min_valid, zero, mid, max_valid, max_invalid)
_TYPE_BOUNDARIES: Dict[str, Dict[str, Any]] = {
    "uint8_t":  {"min_inv": -1,     "min": 0,      "mid": 127,   "max": 255,     "max_inv": 256},
    "uint8":    {"min_inv": -1,     "min": 0,      "mid": 127,   "max": 255,     "max_inv": 256},
    "uint16_t": {"min_inv": -1,     "min": 0,      "mid": 32767, "max": 65535,   "max_inv": 65536},
    "uint16":   {"min_inv": -1,     "min": 0,      "mid": 32767, "max": 65535,   "max_inv": 65536},
    "uint32_t": {"min_inv": -1,     "min": 0,      "mid": 2**15, "max": 2**32-1, "max_inv": 2**32},
    "int8_t":   {"min_inv": -129,   "min": -128,   "mid": 0,     "max": 127,     "max_inv": 128},
    "int16_t":  {"min_inv": -32769, "min": -32768, "mid": 0,     "max": 32767,   "max_inv": 32768},
    "int16":    {"min_inv": -32769, "min": -32768, "mid": 0,     "max": 32767,   "max_inv": 32768},
    "int32_t":  {"min_inv": -(2**31)-1, "min": -(2**31), "mid": 0, "max": 2**31-1, "max_inv": 2**31},
    "float":    {"min_inv": -1001.0, "min": -1000.0, "mid": 0.0,  "max": 1000.0,  "max_inv": 1001.0},
    "bool":     {"min_inv": -1,     "min": 0,       "mid": 0,    "max": 1,       "max_inv": 2},
    "bit":      {"min_inv": -1,     "min": 0,       "mid": 0,    "max": 1,       "max_inv": 2},
}
_DEFAULT_BOUNDARY = {"min_inv": -1, "min": 0, "mid": 127, "max": 255, "max_inv": 256}

# Known C types where out-of-range input defaults to saturation (no "[검증 필요]")
# Unsigned types: deterministic wrap/saturation.
# Fixed-width signed (int8/16/32): 임베디드 환경에서 포화 처리 일반적 (컴파일러 -fwrapv 또는 HW saturation).
# C 표준 signed (char, short, long, int): overflow = UB → "[검증 필요]" 유지.
_KNOWN_SATURATE_TYPES = frozenset({
    "uint8", "uint16", "uint32", "int8", "int16", "int32",
    "float", "bit", "bool", "byte", "word", "dword",
    "unsignedchar", "unsignedshort", "unsignedlong", "unsignedint",
})

# Strategy labels for boundary-value test sequences (module-level constant)
_STRAT_LABEL: Dict[str, str] = {
    "BV_MIN_INV": "유효 하한 초과 (경계-1): 에러/포화 처리 확인",
    "BV_MIN":     "최솟값 경계 입력: 최솟값에서 정상 처리 확인",
    "BV_MID":     "정상 중간값 입력: 정상 동작 범위 확인",
    "BV_MAX":     "최댓값 경계 입력: 최댓값에서 정상 처리 확인",
    "BV_MAX_INV": "유효 상한 초과 (경계+1): 에러/포화 처리 확인",
    "MIXED":      "혼합 경계값: 짝수 인수=최솟값, 홀수 인수=최댓값 조합",
}

def _get_strategy_label(strat_name: str, input_vars: Optional[List[str]] = None,
                        switch_cases: Optional[List[Tuple[str, Any, str]]] = None,
                        loop_var: str = "",
                        global_vars: Optional[List[str]] = None) -> str:
    """Get human-readable label for any strategy including COND_COMB, SWITCH, LOOP, GLOBAL."""
    input_vars = input_vars or []
    switch_cases = switch_cases or []
    global_vars = global_vars or []
    if strat_name in _STRAT_LABEL:
        return _STRAT_LABEL[strat_name]
    if strat_name.startswith("COND_COMB_"):
        idx = int(strat_name.split("_")[-1])
        var = input_vars[idx] if idx < len(input_vars) else f"var{idx}"
        return f"조건 조합: {var}=최솟값, 나머지=중간값 → 분기 커버리지 향상"
    if strat_name.startswith("SWITCH_"):
        idx = int(strat_name.split("_")[-1])
        if idx < len(switch_cases):
            sw_var, sw_val, sw_label = switch_cases[idx]
            return f"Switch-case: {sw_var}={sw_val} ({sw_label}) → case 분기 커버"
        return f"Switch-case: case {idx}"
    if strat_name == "LOOP_ZERO":
        return f"루프 경계: {loop_var or 'counter'}=0 → 루프 미실행 경로 확인"
    if strat_name == "LOOP_ONE":
        return f"루프 경계: {loop_var or 'counter'}=1 → 루프 1회 실행 경로"
    if strat_name == "LOOP_MAX":
        return f"루프 경계: {loop_var or 'counter'}=최댓값 → 루프 최대 반복 경로"
    if strat_name.startswith("GLOBAL_"):
        idx = int(strat_name.split("_")[-1])
        gv = global_vars[idx] if idx < len(global_vars) else f"global{idx}"
        return f"글로벌 상태: {gv}=최솟값 → 글로벌 의존 분기 커버"
    if strat_name == "VOID_SIDE_EFFECT":
        return "Void 부작용: 입력 경계 초과 → 글로벌 변수 상태 변화 검증"
    if strat_name == "MCDC_BASE":
        return "MC/DC baseline: 모든 조건 True → 결정 True 확인"
    if strat_name.startswith("MCDC_"):
        return f"MC/DC: 개별 조건 토글 → 결정 결과 변화 확인 (ASIL D)"
    return strat_name

# Domain-keyword based float boundaries for physical/engineering signals
_FLOAT_DOMAIN_BOUNDS: List[Tuple[List[str], Dict[str, Any]]] = [
    (["voltage", "volt", "_v_", "_vbat", "_vcc"],
     {"min_inv": -1.0, "min": 0.0, "mid": 12.0, "max": 60.0, "max_inv": 61.0}),
    (["temperature", "temp", "_temp", "_t_"],
     {"min_inv": -41.0, "min": -40.0, "mid": 25.0, "max": 150.0, "max_inv": 151.0}),
    (["speed", "_spd", "velocity", "_vel"],
     {"min_inv": -1.0, "min": 0.0, "mid": 60.0, "max": 300.0, "max_inv": 301.0}),
    (["pressure", "_pres", "_press"],
     {"min_inv": -0.1, "min": 0.0, "mid": 2.5, "max": 10.0, "max_inv": 10.1}),
    (["current", "_cur", "_amp"],
     {"min_inv": -0.1, "min": 0.0, "mid": 5.0, "max": 50.0, "max_inv": 51.0}),
    (["angle", "_ang", "degree", "_deg"],
     {"min_inv": -1.0, "min": 0.0, "mid": 90.0, "max": 360.0, "max_inv": 361.0}),
    (["percent", "_pct", "ratio", "_ratio"],
     {"min_inv": -1.0, "min": 0.0, "mid": 50.0, "max": 100.0, "max_inv": 101.0}),
]


def _get_float_bounds_for_var(var_name: str) -> Dict[str, Any]:
    """Return domain-specific float boundaries based on variable name keywords."""
    name_lower = var_name.lower()
    for keywords, bounds in _FLOAT_DOMAIN_BOUNDS:
        if any(kw in name_lower for kw in keywords):
            return bounds
    return _TYPE_BOUNDARIES["float"]


# Patterns for inferring types from variable names
_TYPE_NAME_PATTERNS = [
    (re.compile(r"\bu8[gs]?_|uint8|U8|BYTE", re.I), "uint8_t"),
    (re.compile(r"\bu16[gs]?_|uint16|U16|WORD", re.I), "uint16_t"),
    (re.compile(r"\bu32[gs]?_|uint32|U32|DWORD", re.I), "uint32_t"),
    (re.compile(r"\bs8[gs]?_|int8[^_]|S8", re.I), "int8_t"),
    (re.compile(r"\bs16[gs]?_|int16|S16", re.I), "int16_t"),
    (re.compile(r"\bs32[gs]?_|int32|S32", re.I), "int32_t"),
    (re.compile(r"\bBits\.|_F\b|_Flag|_Sta\b|_Enable|_Disable", re.I), "bit"),
    (re.compile(r"\bf32|float|FLOAT", re.I), "float"),
    (re.compile(r"\bbool\b|BOOL|boolean", re.I), "bool"),
]


# ---------------------------------------------------------------------------
# Phase 1: Data extraction
# ---------------------------------------------------------------------------

_TYPE_NAMES = {
    "U8", "U16", "U32", "S8", "S16", "S32",
    "uint8_t", "uint16_t", "uint32_t", "int8_t", "int16_t", "int32_t",
    "BOOL", "void", "char", "int", "float", "double", "long",
    "unsigned", "signed", "short", "const", "volatile", "static",
    # LIN 스택·Processor Expert 타입. 반환값 슬롯 교정이 주 경로지만, 참조 SUDS
    # 문서 경유분처럼 태그 없이 들어오는 입구를 위해 안전망으로 둔다.
    "U64", "S64", "l_u8", "l_u16", "l_u32", "l_bool", "byte", "word", "bool", "dword",
}

# Local temp variable prefixes — these live on stack, not meaningful for unit test I/O
_LOCAL_TEMP_PATS = re.compile(
    r"^(u8t_|u16t_|u32t_|s8t_|s16t_|s32t_|sf_t|tmpVal|temp_|tmp_|loop_|idx_|cnt_|i$|j$|k$|n$)",
    re.I,
)

# Prefixes that are clearly global reads (function inputs)
_INPUT_PREFIXES = ("u8g_", "u16g_", "u32g_", "s16g_", "s8g_", "s32g_")
# Prefixes that are clearly module-static writes (function outputs)
_OUTPUT_PREFIXES = ("u8s_", "u16s_", "u32s_", "s16s_", "s32s_")
# Hardware registers — typically both read and written
_REG_PAT = re.compile(r"^REG_|^lin_|^PS\.|^DiagData\.")

# 파서가 붙이는 방향 태그. **앵커 매칭이어야 한다** — 예전엔 `"[IN]" in tag` 였는데
# `"[IN]" in "[INOUT] x"` 도 `"[OUT]" in "[INOUT] x"` 도 **둘 다 False** 다(`[INOUT]` 안에
# `[IN]`·`[OUT]` 이 연속으로 들어있지 않다). 그래서 파서가 가장 정확하게 아는 축인
# `[INOUT]` 이 통째로 "태그 없음"으로 떨어져 아래 프리픽스 휴리스틱을 타고, 대부분
# `elif not is_in_global: role_out = True` 에 걸려 **출력 전용**이 됐다.
# 실측(KJPDS02 파서 산출 750함수 중 전역 보유 556개): [IN] 1,423 · [OUT] 1,052 ·
# **[INOUT] 305** · [INDIRECT] 1,114 · 무태그 529. `LinSend` 는 `[INOUT] s_LinFrame …` 를
# 받고도 입력이 0개였고 같은 이름이 기대결과에만 실렸다(정본은 입력에 둔다).
#
# ⚠ **`INDIRECT2` 를 빼먹으면 안 된다** — 같은 함정이 한 번 더 있었다. 2홉 전파는
#   `[INDIRECT2]` 로 태그되는데(`report_gen/uds_generator.py:2107`), 이 정규식이
#   `INDIRECT` 만 알면 `[INDIRECT2]` 는 매칭에 실패해 **"태그 없음"** 으로 떨어지고
#   아래 프리픽스 휴리스틱을 탄다. 그 결과 **1홉(`[INDIRECT]`)은 입력에서 빼는데
#   2홉은 입력으로 올리는**, 증거가 멀수록 느슨해지는 뒤집힌 판정이 됐다.
#   실측(2026-08-12, KJPDS02): SPI 레지스터가 살아나자 `g_DrvIn_DRV8706SQ_Init` ·
#   `..._Left` · `s_IIM20670_Init` 3건이 정본엔 입력 0개인데 `_SPI0SR` 계열을
#   입력으로 냈다 — 읽기는 2홉 아래 `u16g_DrvIn_SPI_DataTransfer` 안에서 일어난다.
_DIR_TAG_PAT = re.compile(r"^\s*\[(IN|OUT|INOUT|INDIRECT2|INDIRECT)\]", re.I)


def dir_tag(entry: Any) -> str:
    """전역 엔트리의 방향 태그(대문자). 태그가 없으면 빈 문자열.

    **방향 태그 판정의 단일 출처.** 소비처가 각자 정규식을 들고 있으면 태그가
    하나 늘 때 한쪽만 고쳐진다 — 이 저장소가 `[INOUT]`(A-1)과 `[INDIRECT2]`
    두 번 겪은 실패다.
    """
    m = _DIR_TAG_PAT.match(str(entry or ""))
    return m.group(1).upper() if m else ""


def _is_const_global(name: str, gim: Optional[Dict[str, Dict[str, str]]]) -> bool:
    """`const` 전역은 시험 입력으로 **설정할 수 없고** 기대결과로 **변하지도 않는다**.

    실측(KJPDS02_PV 정본 1,005 unit): 정본 SUTS 는 const 전역을 입력 **0칸** · 기대
    **0칸** — 어느 입도로도 단 한 번도 적지 않는다. 우리는 419칸(입력 160 · 기대 259)
    을 냈고 그중 정본과 일치한 건 **0** 이다. 즉 억제의 대가가 0 이다.
    `au32_Sha256RoundConstants[0..63]` 처럼 배열이면 원소 확장이 노이즈를 배로 불린다.

    ⚠ 파라미터의 `const`(`const U8 *p`)는 **대상이 아니다** — 가리키는 곳이 읽기
      전용일 뿐 그 버퍼는 시험이 채워 넣어야 하는 입력이다. 이 판정은 전역 루프에서만
      쓴다.
    ⚠ `gim` 이 비면 판정할 근거가 없어 **억제하지 않는다**. 호출부가 안 넘기면 산출물이
      달라지므로, 아래 요약 로그가 그 사실을 명시한다(조용한 분기 금지).
    """
    return is_const_type(((gim or {}).get(name) or {}).get("type"))


def collect_unit_functions(
    function_details: Dict[str, Dict[str, Any]],
    globals_info_map: Optional[Dict[str, Dict[str, str]]] = None,
    sds_map: Optional[Dict[str, Dict[str, str]]] = None,
) -> List[Dict[str, Any]]:
    """Collect and structure unit functions from report_generator output.

    Matching reference SUTS patterns: variables can appear as BOTH input and
    output (read-modify-write). Local temps are excluded. REG_ and state
    vars are placed in output. Caps at reasonable counts per function.

    Args:
        sds_map: ASIL/related 보강에 쓸 SDS 파티션 맵. None이면 저장소 `docs/` 글롭
            폴백(`_load_default_sds_map`)을 쓴다 — **프로젝트 무관**이므로 호출자가
            대상 프로젝트의 SDS를 알고 있으면 `load_sds_map_from`으로 만들어 넘길 것.
    """
    gim = globals_info_map or {}
    if sds_map is None:
        sds_map = _load_default_sds_map()
    units: List[Dict[str, Any]] = []
    _const_skipped = 0

    for fid, info in function_details.items():
        if not isinstance(info, dict):
            continue
        name = info.get("name", "")
        if not name:
            continue

        prototype = info.get("prototype") or f"void {name}(void)"
        inputs_raw = info.get("inputs") or []
        outputs_raw = info.get("outputs") or []
        globals_g = info.get("globals_global") or []
        globals_s = info.get("globals_static") or []

        input_vars: List[str] = _extract_var_names(inputs_raw)
        output_vars: List[str] = _extract_var_names(outputs_raw)

        inp_set = set(input_vars)
        out_set = set(output_vars)

        globals_g_set = set(globals_g)

        for g in globals_g + globals_s:
            gn = _clean_global_name(g)
            if not gn or gn in _TYPE_NAMES:
                continue
            if len(gn) <= 2 or not re.match(r"[A-Za-z_]", gn):
                continue
            if _is_const_global(gn, gim):
                _const_skipped += 1
                continue
            if _LOCAL_TEMP_PATS.match(gn):
                continue

            tag = str(g).upper()
            is_in_global = g in globals_g_set

            role_in = False
            role_out = False

            # 방향 태그 — 앵커 매칭. `[INOUT]` 은 읽기·쓰기 **둘 다**다.
            _dir_tag = dir_tag(g)
            # ⚠ 간접 판정을 **파싱된 태그에서** 뽑는다. 예전의 `"[INDIRECT]" in tag` 는
            #   `[INDIRECT2]` 를 못 봐서 2홉이 직접 사용처럼 통과했다.
            is_indirect = _dir_tag.startswith("INDIRECT")
            if _dir_tag in {"IN", "INOUT"}:
                role_in = True
            if _dir_tag in {"OUT", "INOUT"}:
                role_out = True

            # 키워드 폴백 — 참조 SUDS 문서 경유 등 태그를 안 붙이는 생산자를 위해 남긴다.
            if any(k in tag for k in ["READ", "RHS"]):
                role_in = True
            if any(k in tag for k in ["WRITE", "LHS"]):
                role_out = True

            if not role_in and not role_out:
                if gn.startswith(_OUTPUT_PREFIXES):
                    role_out = True
                elif gn.startswith(_INPUT_PREFIXES):
                    if not is_indirect:
                        role_in = True
                elif _REG_PAT.match(gn):
                    if not is_indirect:
                        role_in = True
                    role_out = True
                elif gn.startswith(("g_", "r_")):
                    if not is_indirect:
                        role_in = True
                    role_out = True
                elif not is_in_global:
                    role_out = True
                elif not is_indirect:
                    role_in = True

            # ⚠ 표기 정합은 **맨 마지막**에만 한다. 위의 `_is_const_global`·
            #   `_LOCAL_TEMP_PATS`·프리픽스 판정은 전부 C 이름(`gn`)을 키로 쓰므로
            #   먼저 바꾸면 그 조회들이 조용히 빗나간다.
            gd = _vc_pointer_notation(gn)
            if role_in and gd not in inp_set:
                input_vars.append(gd)
                inp_set.add(gd)
            if role_out and gd not in out_set:
                output_vars.append(gd)
                out_set.add(gd)

        component = ""
        module = info.get("module_name", "")
        if fid and re.match(r"SwUFn_\d+", fid):
            comp_num = fid.replace("SwUFn_", "")[:2]
            component = f"SwCom_{comp_num}"
            if module:
                component = f"{component}\n({module})"

        # Attempt to resolve SRS requirement IDs via SDS partition map
        srs_req_ids = _resolve_srs_req_ids_for_function(name, sds_map)

        if not output_vars:
            ret_type = _infer_return_type(prototype)
            if ret_type and ret_type.lower() != "void":
                # 정본 표기와 같은 이름을 쓴다 — `return_<함수명>` 은 정본 어디에도 없다.
                output_vars.append(_RETURN_VAR)
                out_set.add(_RETURN_VAR)

        max_inp = _INPUT_COL_END - _INPUT_COL_START + 1
        max_out = _OUTPUT_COL_END - _OUTPUT_COL_START + 1

        # 배열을 원소 단위로 펼친다(정본과 같은 입도). 입력·기대 **양쪽** 이다 —
        # 실측상 같은 unit 에서 양쪽에 펼쳐진 배열이 120건이라, 한쪽만 펼치면 한 행
        # 안에서 같은 변수가 다른 이름으로 두 번 나온다.
        _sizes = _array_sizes(inputs_raw, outputs_raw, globals_g, globals_s)
        input_vars, _in_exp = _expand_array_entries(input_vars, _sizes, max_inp)
        output_vars, _out_exp = _expand_array_entries(output_vars, _sizes, max_out)

        asil = str(info.get("asil") or "TBD").strip()
        if not asil or asil.upper() == "TBD":
            asil = _resolve_unit_asil(info, sds_map) or asil

        # Collect indirect (global) vars for GLOBAL/VOID strategies
        indirect_vars: List[str] = []
        for g in globals_g + globals_s:
            gn = _clean_global_name(g)
            # ⚠ 2홉(`[INDIRECT2]`)도 간접이다 — 여기서 빠지면 GLOBAL/VOID 전략이
            #   간접 변수를 하나도 못 받는다.
            if gn and dir_tag(g).startswith("INDIRECT") and gn not in inp_set and gn not in out_set:
                if gn not in indirect_vars and len(indirect_vars) < 5:
                    indirect_vars.append(gn)

        units.append({
            "fid": fid,
            "name": name,
            "prototype": prototype,
            "component": component,
            "input_vars": input_vars[:max_inp],
            "output_vars": output_vars[:max_out],
            # 무엇을 펼쳤고 무엇을 예산 때문에 못 펼쳤나. 건너뛴 배열은 정본보다
            # **입도가 낮은** 칸이 되므로 조용히 두면 "정본과 다르다"의 원인을 못 짚는다.
            "array_expansion": {"input": _in_exp, "output": _out_exp},
            "indirect_vars": indirect_vars,
            "logic_flow": info.get("logic_flow") or [],
            "calls_list": info.get("calls_list") or [],
            "description": info.get("description", ""),
            "asil": asil,
            "srs_req_ids": srs_req_ids,
            "precondition": info.get("precondition", ""),
        })

    units.sort(key=lambda u: u["fid"])
    # 배열 확장 집계. 건너뛴 게 있으면 WARNING 으로 올린다 — 예산 때문에 정본보다
    # 입도가 낮아진 칸이 있다는 뜻이고, 그건 조용하면 안 된다.
    _exp_n = sum(len(u["array_expansion"]["input"]["expanded"])
                 + len(u["array_expansion"]["output"]["expanded"]) for u in units)
    _skip = [(u["name"], s["name"], s["elements"], s["remaining"])
             for u in units
             for axis in ("input", "output")
             for s in u["array_expansion"][axis]["skipped"]]
    # ⚠ `globals_info_map` 이 없으면 const 판정 자체를 못 한다 — 같은 소스라도 산출물이
    #   달라지므로 **명시**한다(조용한 분기는 이 저장소가 여러 번 데었다).
    _const_note = (
        f" | const 전역 억제 {_const_skipped}칸" if gim
        else " | ⚠globals_info_map 없음 → const 억제 안 함"
    )
    (_logger.warning if _skip else _logger.info)(
        "Collected %d unit functions | 배열 확장 %d건%s%s",
        len(units), _exp_n,
        ("  ⚠예산 부족으로 미확장 %d건: %s" % (
            len(_skip),
            ", ".join(f"{u}::{n}({k}원소, 여유 {r})" for u, n, k, r in _skip[:3]),
        )) if _skip else "",
        _const_note,
    )
    return units


def _infer_return_type(prototype: str) -> str:
    """Extract the return type from a C function prototype string."""
    proto = prototype.strip()
    m = re.match(r"^([\w\s\*]+?)\s+\w+\s*\(", proto)
    if not m:
        return "void"
    ret = m.group(1).strip()
    ret = re.sub(r"\b(static|inline|extern|const|volatile)\b", "", ret).strip()
    return ret if ret else "void"


# 파서가 이름 **뒤에** 붙이는 주석형 꼬리(`_format_param_entry`).
#   `u8g_Hash (idx: u8t_Index)` · `ctx (range: 0x0 ~ 0xFFFFFFFF)` · `div (divisor: no 0)`
# ⚠ 이름은 마지막 토큰에서 뽑는데 이 꼬리를 안 떼면 **꼬리가 이름이 된다**:
#   · `(idx: u8t_Index)` → 이름이 `u8t_Index)` → `_LOCAL_TEMP_PATS`(`u8t_`)에 걸려 **전역이 통째로 사라진다**
#   · `(range: … 0xFFFFFFFF)` → 이름이 `0xFFFFFFFF)` → 식별자가 아니라 **파라미터가 통째로 사라진다**
# 실측(2026-08-12, KJPDS02 750함수): 입력 0개 unit 221 건 중 **57 건**이 이 경로였다.
# `s_sha256_transform` 은 정본이 입력 9개를 적는데 우리는 0개였다.
# ⚠ 꼬리 키워드는 **한 곳**에서만 정의한다. 아래 두 정규식이 같은 목록을 각자 들고
#   있으면 새 꼬리를 추가할 때 하나만 고쳐지고, 그 꼬리가 그대로 **이름이 된다**.
#   (같은 부류의 실패를 이 저장소가 `[INOUT]`·`[INDIRECT2]` 로 두 번 겪었다.)
_PARAM_ANNOT_KEYS = "idx|range|divisor|size"
_PARAM_ANNOT_TAIL = re.compile(rf"\s*\((?:{_PARAM_ANNOT_KEYS})\s*:[^)]*\)\s*$", re.I)

# 꼬리의 **시작**만 찾는다. 끝은 괄호를 세어서 찾아야 한다 — `idx:` 안에 괄호가 중첩되기
# 때문이다(`_normalize_bracket_expr` 가 매크로를 못 접으면 원문이 그대로 실린다):
#   `u8g_SysEepromCtrl_PartNoInfo (idx: ( ( U8 )( 2U ) ), ( ( U8 )( 8U ) ), …)`
# `[^)]*\)` 는 첫 `)` 에서 멈춰 `\s*$` 가 안 맞으므로 **꼬리가 하나도 안 떨어진다**.
# 그러면 마지막 토큰이 `))` 가 되고, 두 글자라 이름 필터에서 탈락해 **진짜 전역이 사라진다**
# (실측 2026-08-12: `u8s_DeviceTypeChk_*` 2건).
_PARAM_ANNOT_HEAD = re.compile(rf"\s*\((?:{_PARAM_ANNOT_KEYS})\s*:", re.I)


def _strip_param_annotations(s: str) -> str:
    """이름 뒤 주석형 꼬리를 **전부** 뗀다(여러 개가 이어 붙고, 안에 괄호가 중첩된다)."""
    out = str(s or "").strip()
    while True:
        stripped = _PARAM_ANNOT_TAIL.sub("", out)
        if stripped != out:
            out = stripped
            continue
        # 중첩 괄호가 있어 위 `$` 앵커 정규식이 못 잡은 경우. **마지막** 꼬리 후보를
        # 잡아 괄호를 세어 끝을 찾는다. 꼬리는 이름 **뒤에만** 붙으므로, 잘라낸 자리가
        # 문자열 끝이 아니면 꼬리가 아니다 — 건드리지 않는다(이름 중간을 지우면 다른
        # 이름이 된다).
        matches = list(_PARAM_ANNOT_HEAD.finditer(out))
        if not matches:
            return out
        m = matches[-1]
        depth = 0
        end = len(out)
        for i in range(out.index("(", m.start()), len(out)):
            if out[i] == "(":
                depth += 1
            elif out[i] == ")":
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break
        if out[end:].strip():
            return out
        out = out[: m.start()].strip()


# 파라미터 문자열이 **선언이 아닌** 경우. 상위 파서가 주석 블록을 통째로 파라미터 하나로
# 딸려보내는 일이 있다(Processor Expert 계열 `*_GetVal` 실측 40건):
#   `[IN] void) ** This method is implemented as a macro. … // if (Val == (U8) TRUE (range: …)`
# ⚠ 이걸 그냥 두면 마지막 토큰인 **`TRUE` 가 변수명이 된다** — 없는 입력을 지어내는 것이라
#   빈 칸보다 나쁘다(꼬리 주석 제거를 넣자마자 실제로 3건 발생했다). 선언이 아니면 **버린다**.
#   버려진 건 게이트가 `param_string_unusable` 로 보고한다 — 침묵시키지 않는다.
# **선언자 모양** 검사. 타입·이름 토큰과 `*` `[]` `.` `->` 만 허용한다.
# 괄호·대입·세미콜론이 있으면 선언이 아니다(주석 잔해나 코드 조각이 딸려온 것).
#   OK : `l_u8 msg_length` · `const l_u8* const data` · `q * queue->queue_tail` · `buf[8]`
#   NG : `void) ** This method is implemented as a macro. …` · `if positive = 0 l_u8 err`
# ⚠ 여기서 통과시키면 **마지막 토큰이 이름이 된다** — 즉 없는 입력을 지어낸다. 빈 칸보다
#   나쁘다(꼬리 주석 제거를 처음 넣었을 때 실제로 `TRUE` 3건이 그렇게 들어갔다).
#   버려진 건 게이트가 `param_string_unusable` 로 보고한다 — 침묵시키지 않는다.
_PARAM_DECL_SHAPE = re.compile(r"[A-Za-z_][\w\s\*\[\]\.>-]*")
_PARAM_DECL_MAX_LEN = 120


# 반환값 슬롯. 생산자 5곳이 `[OUT] return <타입>` 형태로 낸다(`function_analyzer`·
# `backend/helpers/common`·`uds_generator`·`tools/generate_uds_local`·아래 3179행) —
# **그 계약은 건드리지 않는다**. 문제는 소비처였다: `^return\s+` 를 지우고 마지막 토큰을
# 취해 **타입 이름을 변수로** 냈다(실측 KJPDS02_PV 기대열 287건: `U8` 144 · `U16` 62 ·
# `S16` 32 · `l_u8` 19 …). 정본은 반환값을 **`return`** 이라고 적는다(기대 엔트리 5,389
# 중 `return` 290 · `return[0]` 7).
_RETURN_SLOT_RE = re.compile(r"^return\b", re.I)
_RETURN_VAR = "return"


# 포인터 표기. 정본(VectorCAST)은 포인터 뒤에 **1원소 이상의 배열**을 잡아주므로
# `p[0]` · `p[0].m` 으로 적는다. 우리 생산자는 C 문법 그대로 `p` · `p->m` 을 낸다 —
# **같은 대상을 다르게 부르는 것**이라, 표기만 맞추면 과다와 미달이 동시에 닫힌다.
# 실측(KJPDS02_PV 시뮬): 입력 163칸 · 기대 124칸이 과다→일치로 이동, **잃은 일치 0**.
# ⚠ 생산자 계약(`[IN] word * Values`)은 건드리지 않는다 — UDS 상세설계엔 C 표기가 맞다.
#   `return` 슬롯과 같은 방식이다(소비처에서만 교정).
_ARROW_RE = re.compile(r"\s*->\s*")


def _vc_pointer_notation(name: str) -> str:
    """``p->m`` → ``p[0].m``. 화살표가 없으면 원본 그대로."""
    s = str(name or "")
    return _ARROW_RE.sub("[0].", s) if "->" in s else s


def _extract_var_names(raw_list: List[str]) -> List[str]:
    """Extract clean variable names from [IN]/[OUT] tagged param strings."""
    names: List[str] = []
    for raw in raw_list:
        s = str(raw).strip()
        s = re.sub(r"^\[(?:IN|OUT|INOUT)\]\s*", "", s)
        if _RETURN_SLOT_RE.match(s):
            if _RETURN_VAR not in names:
                names.append(_RETURN_VAR)
            continue
        s = re.sub(r"^return\s+", "", s, flags=re.I)
        # 파라미터 앞에 붙은 설명 주석(`/* [IN] … */ l_u8 msg_length`)을 지운다.
        # 생산자(`_parse_signature_params`)가 2026-08-12부터 안 붙이지만, **캐시된 산출물과
        # 참조 SUDS 문서 경유분에는 남아 있다** — 소비처에서도 한 번 더 지워야 회복된다.
        s = _strip_param_annotations(blank_c_comments(s).strip())
        if len(s) > _PARAM_DECL_MAX_LEN or not _PARAM_DECL_SHAPE.fullmatch(s):
            continue
        # Remove type qualifiers, keep only the symbol name
        parts = s.split()
        if not parts:
            continue
        candidate = parts[-1].strip("*&;,")
        candidate = re.sub(r"\[.*?\]$", "", candidate)
        if candidate and re.match(r"[A-Za-z_]", candidate):
            candidate = _vc_pointer_notation(candidate)
            if candidate not in names:
                names.append(candidate)
    return names


# 배열 원소 확장.
#
# 정본 SUTS 는 배열을 **원소 단위로** 적는다(실측 KJPDS02_PV):
#   입력 엔트리 6,014 중 `name[N]` 3,023(50.3%) · 기대 5,389 중 2,716(50.4%)
#   base 134 중 **120개가 모든 unit 에서 같은 개수** = 관찰 첨자가 아니라 선언 크기
#   최대 원소 60 · 입력 unit당 최대 **96 = 열 상한 정확히**(초과 0) · 기대 84 = 상한
#
# ⚠ 상한을 넘기면 **펼치지 않고 base 이름을 그대로 둔다**. 원소를 잘라 넣으면
#   "이 배열은 앞 k칸만 시험한다"는 없는 사실을 적게 되고, 뒤에 오는 **다른 변수**가
#   통째로 밀려난다. 변수는 하나도 잃지 않고 입도만 낮추는 쪽이 정직하다.
#   건너뛴 것은 `array_expansion` 으로 보고한다 — 침묵시키지 않는다.
# 다차원은 `9x8` 로 실린다 — 차원을 곱해 버리면 `[i][j]` 를 복원할 수 없다.
_SIZE_TAIL_RE = re.compile(r"\(\s*size\s*:\s*(\d+(?:\s*x\s*\d+)*)\s*\)", re.I)
# 파라미터 표시엔 선언 차원이 `buf[10]`·`t[3][4]` 로 이미 들어 있다(`_format_param_entry`).
# 이 필드에서 `[N]` 은 **항상 선언 크기**다(원소 표기를 내는 생산자가 없다).
_PARAM_DIM_RE = re.compile(r"((?:\[\d+\])+)\s*$")


def _dim_product(dims: Tuple[int, ...]) -> int:
    n = 1
    for d in dims:
        n *= int(d)
    return n


def _array_sizes(*raw_groups: List[str]) -> Dict[str, Tuple[int, ...]]:
    """원시 엔트리에서 `이름 → 차원 튜플` 을 모은다(1차원도 `(60,)` 로 담는다)."""
    sizes: Dict[str, Tuple[int, ...]] = {}
    for group in raw_groups:
        for raw in group or []:
            s = str(raw or "")
            dims: Tuple[int, ...] = ()
            m = _SIZE_TAIL_RE.search(s)
            if m:
                dims = tuple(int(x) for x in re.findall(r"\d+", m.group(1)))
            if not dims:
                head = _strip_param_annotations(
                    re.sub(r"^\[(?:IN|OUT|INOUT|INDIRECT2|INDIRECT)\]\s*", "", s.strip())
                )
                m2 = _PARAM_DIM_RE.search(head)
                if m2:
                    dims = tuple(int(x) for x in re.findall(r"\d+", m2.group(1)))
            if not dims or _dim_product(dims) <= 1:
                continue
            name = _clean_global_name(s)
            name = re.sub(r"(?:\[\d+\])+$", "", name)
            if name and _dim_product(dims) > _dim_product(sizes.get(name) or ()):
                sizes[name] = dims
    return sizes


def _elem_suffixes(dims: Tuple[int, ...]) -> List[str]:
    """`(9, 8)` → `['[0][0]', '[0][1]', …]` — 정본과 같은 row-major 순서."""
    out = [""]
    for d in dims:
        out = [f"{pre}[{i}]" for pre in out for i in range(int(d))]
    return out


def _expand_array_entries(
    names: List[str], sizes: Dict[str, Tuple[int, ...]], budget: int
) -> Tuple[List[str], Dict[str, Any]]:
    """배열 이름을 원소로 펼친다. 예산이 모자라면 **펼치지 않고 그대로 둔다**."""
    out: List[str] = []
    expanded: List[str] = []
    skipped: List[Dict[str, Any]] = []
    total = len(names)
    for pos, nm in enumerate(names):
        dims = sizes.get(nm) or ()
        n = _dim_product(dims) if dims else 0
        if n <= 1:
            out.append(nm)
            continue
        remaining = budget - len(out)
        # 남은 이름들도 최소 한 칸씩은 자리가 있어야 한다.
        # ⚠ `names.index(nm)` 는 같은 이름이 두 번 들어오면 **첫 위치**를 돌려줘
        #   예약분을 과다 계산한다 — 위치는 열거로 받는다.
        reserve = total - pos - 1
        if n > remaining - reserve:
            out.append(nm)
            skipped.append({"name": nm, "elements": n, "remaining": max(0, remaining - reserve)})
            continue
        out.extend(nm + sfx for sfx in _elem_suffixes(dims))
        expanded.append(nm)
    return out, {
        "expanded": expanded,
        "skipped": skipped,
        "budget": budget,
        "used": len(out),
    }


def _clean_global_name(g: str) -> str:
    s = str(g).strip()
    s = re.sub(r"^\[INDIRECT\]\s*", "", s)
    s = re.sub(r"^\[(?:IN|OUT|INOUT)\]\s*", "", s)
    s = _strip_param_annotations(s)
    parts = s.split()
    return parts[-1].strip("*&;,") if parts else ""


# ---------------------------------------------------------------------------
# Phase 2: Variable type analysis
# ---------------------------------------------------------------------------

_globals_type_cache: Dict[str, str] = {}


def _gim_to_type_map(gim: Dict[str, Any]) -> Dict[str, str]:
    """globals_info_map({var:{type:...}})에서 {var: raw_type} 타입맵을 추출한다.

    ``set_globals_type_cache``(프로세스 전역 시딩)와 ``_build_doc_proposal``(로컬 type_cache
    주입, workflow/impact_orchestrator.py)의 **단일 출처**다 — 두 곳이 독립 구현이면 한쪽만
    고쳐 드리프트하는 전례가 있어 통합한다(비-dict/빈 타입 값 방어 포함).
    """
    out: Dict[str, str] = {}
    for var_name, info in (gim or {}).items():
        if not isinstance(info, dict):
            continue
        vtype = str(info.get("type") or "").strip()
        if vtype:
            out[str(var_name)] = vtype
    return out


def set_globals_type_cache(gim: Dict[str, Dict[str, str]]) -> None:
    """Populate type cache from globals_info_map for precise type resolution."""
    _globals_type_cache.clear()
    _globals_type_cache.update(_gim_to_type_map(gim))


def infer_variable_type(var_name: str, type_cache: Optional[Dict[str, str]] = None) -> str:
    """Infer C type from variable naming convention or globals_info_map.

    type_cache: 명시적 {var: raw_type} 맵. 주어지면 프로세스 전역 ``_globals_type_cache``
      대신 이것을 읽는다 — 호출자(예: 영향도 문서 초안 합성)가 전역을 변이시키지 않고
      정확한 타입 해상도를 얻게 해 write-race/타 프로젝트 오염을 원천 차단한다.
      None이면 기존대로 전역 캐시를 읽는다(실 문서생성 경로는 무변경).
    """
    cache = type_cache if type_cache is not None else _globals_type_cache
    if var_name in cache:
        raw = cache[var_name]
        mapped = _normalize_type(raw)
        if mapped:
            return mapped
    for pat, typename in _TYPE_NAME_PATTERNS:
        if pat.search(var_name):
            return typename
    return "uint8_t"


def _normalize_type(raw: str) -> str:
    """Map raw C type string to a known boundary type."""
    r = raw.strip().lower()
    for key in _TYPE_BOUNDARIES:
        if key in r:
            return key
    alias = {"u8": "uint8_t", "u16": "uint16_t", "u32": "uint32_t",
             "s8": "int8_t", "s16": "int16_t", "s32": "int32_t"}
    for k, v in alias.items():
        if r == k:
            return v
    return ""


def get_boundary_values(typename: str) -> Dict[str, Any]:
    normalized = typename.lower().replace(" ", "").replace("_t", "_t")
    return _TYPE_BOUNDARIES.get(normalized, _DEFAULT_BOUNDARY)


# ---------------------------------------------------------------------------
# Phase 3: Test sequence generation
# ---------------------------------------------------------------------------

def determine_gen_method(unit: Dict[str, Any]) -> str:
    """Determine TC generation method based on function characteristics."""
    logic = unit.get("logic_flow") or []
    has_conditions = any(n.get("type") in ("if", "switch") for n in logic if isinstance(n, dict))
    has_loops = any(n.get("type") == "loop" for n in logic if isinstance(n, dict))
    n_inputs = len(unit.get("input_vars", []))

    if has_conditions and n_inputs > 0:
        return "AEC, ABV"
    if n_inputs > 2:
        return "ABV, AOR"
    if n_inputs > 0:
        return "ABV"
    return "AOR"


def resolve_safety_related(asil: Any) -> str:
    """정본의 `Safety Related` 칸 값 — `O`(안전 관련) / `X`(비안전) / 빈칸(근거 없음).

    ⚠ 이전 판은 `"X" if is_safety else ""` 였다. **의미가 정반대**다 — 정본은
    `O` 566 · `X` 311 로 두 값을 다 쓰고(실측), `O` 가 안전 관련이다. ASIL 을 가진
    단위가 문서상 "비안전"으로 읽히고 있었다.

    ⚠ 근거가 없을 때(`""`/`TBD`) **`X` 로 단정하지 않는다.** `X` 는 "확인했고 안전
    관련이 아니다" 라는 주장이고, 모르는 것을 그렇게 적으면 under-classification 이다.
    정본에도 빈칸이 137개 있다 — 빈칸이 정직한 표기다.
    """
    val = str(asil or "").strip().upper()
    if val in ("A", "B", "C", "D") or val.startswith("ASIL"):
        return "O"
    if val == "QM":
        return "X"
    return ""


def resolve_seq_test_method(strategy: Any) -> str:
    """시퀀스 하나의 Test Method — 정본은 **시퀀스 그룹 단위**로 REQ/FI 를 나눈다."""
    return _METHOD_FI if str(strategy or "").strip() in _FI_STRATEGIES else _METHOD_REQ


def resolve_seq_gen_method(strategy: Any) -> str:
    """시퀀스 하나의 TC Generation Method — 경계값이면 `AOR/ABV`, 조건 조합이면 `AOR/AEC`."""
    s = str(strategy or "").strip()
    if s.startswith("COND_COMB_") or s.startswith("SWITCH_") or s.startswith("MCDC"):
        return _GEN_EQUIV
    return _GEN_BOUNDARY


def determine_test_method(unit: Dict[str, Any]) -> str:
    """Infer a unit-test method label for the fixed SUTS columns.

    ⚠ 정본 시트에는 쓰이지 않는다(정본은 시퀀스 그룹별 `resolve_seq_test_method`).
    Traceability 시트·요약 통계가 아직 쓰므로 남겨 둔다.
    """
    logic = unit.get("logic_flow") or []
    has_conditions = any(n.get("type") in ("if", "switch") for n in logic if isinstance(n, dict))
    has_loops = any(n.get("type") == "loop" for n in logic if isinstance(n, dict))
    n_inputs = len(unit.get("input_vars", []))

    if has_conditions or has_loops:
        return "FNCT"
    if n_inputs > 0:
        return "FIT"
    return "RVW"


def generate_sequences(
    unit: Dict[str, Any],
    max_seq: int = _DEFAULT_SEQ_COUNT,
    type_cache: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    """Generate test sequences for a unit function.

    Produces boundary-value and error-condition test sequences matching
    the reference SUTS patterns:
      Seq 1: all inputs at error-low boundary (min_inv)
      Seq 2: all inputs at minimum valid (0 for unsigned)
      Seq 3: normal mid-range values
      Seq 4: all inputs at maximum valid
      Seq 5: all inputs at error-high boundary (max_inv)
      Seq 6: mixed combination (alternating valid/boundary)
    """
    input_vars = unit.get("input_vars") or []
    output_vars = unit.get("output_vars") or []

    if not input_vars and not output_vars:
        fn_name = unit.get("name", "function")
        prototype = unit.get("prototype", "")
        is_void_return = "void" in prototype.split("(")[0].lower() if "(" in prototype else True
        calls_list = unit.get("calls_list") or []
        logic_flow = unit.get("logic_flow") or []

        # Build calls summary for description
        if calls_list:
            calls_short = calls_list[:5]
            calls_str = ", ".join(f"{c}()" for c in calls_short)
            if len(calls_list) > 5:
                calls_str += f" 외 {len(calls_list) - 5}개"
            calls_note = f" 하위 함수 [{calls_str}] 순차 호출 확인"
        else:
            calls_note = " 예외 없이 완료되며 호출 후 상태 이상 없음"

        # Extract guard condition from logic_flow for ERROR_PATH description
        guard_cond = ""
        for node in logic_flow:
            cond = str(node.get("condition", "") or node.get("text", "")).strip()
            if cond and any(op in cond for op in ("<", ">", "==", "!=", "NULL", "null")):
                guard_cond = cond[:80]
                break
        if guard_cond:
            error_desc = (
                f"{fn_name}() 에러 경로: 조건 [{guard_cond}] 위반 상태에서 호출, "
                f"에러 처리 루틴 진입 또는 안전 상태 유지 확인"
            )
        else:
            error_desc = (
                f"{fn_name}() 에러 경로: 의존 모듈/전역 변수 비정상 상태에서 호출, "
                f"에러 처리 루틴 진입 또는 안전 상태 유지 확인"
            )

        # Build inputs/expected from indirect_vars (callee global vars) if available
        indirect_vars: List[str] = unit.get("indirect_vars") or []
        normal_inputs: Dict[str, Any] = {}
        error_inputs: Dict[str, Any] = {}
        normal_expected: Dict[str, Any] = {}
        error_expected: Dict[str, Any] = {}
        if indirect_vars:
            for _iv in indirect_vars[:4]:
                _vtype = infer_variable_type(_iv, type_cache)
                _bounds = (
                    _get_float_bounds_for_var(_iv) if _vtype == "float"
                    else get_boundary_values(_vtype)
                )
                normal_inputs[_iv] = _bounds.get("mid", 0)
                normal_expected[_iv] = _bounds.get("mid", 0)
                error_inputs[_iv] = _bounds.get("max_inv", _bounds.get("max", 255) + 1)
                error_expected[_iv] = _bounds.get("max", 255)  # clamped/saturated

        seqs = [
            {"seq_num": 1, "inputs": normal_inputs, "expected": normal_expected,
             "strategy": "NORMAL_CALL",
             "description": f"{fn_name}() 정상 호출:{calls_note}"},
            {"seq_num": 2, "inputs": error_inputs, "expected": error_expected,
             "strategy": "ERROR_PATH",
             "description": error_desc},
            {"seq_num": 3, "inputs": {}, "expected": {},
             "strategy": "REPEAT_CALL",
             "description": (
                 f"{fn_name}() 반복 호출 안정성: 100회 연속 호출 후 메모리 누수 없음, "
                 f"시스템 상태 일관성 유지"
             )},
        ]
        if not is_void_return:
            seqs.append({"seq_num": 4, "inputs": {}, "expected": {},
                         "strategy": "RETURN_CHECK",
                         "description": f"{fn_name}() 반환값 검증: 반환값이 정의된 범위 내 유효한 값임을 확인"})
        return seqs[:max_seq]

    var_types = {v: infer_variable_type(v, type_cache) for v in input_vars}
    var_bounds = {
        v: (_get_float_bounds_for_var(v) if t == "float" else get_boundary_values(t))
        for v, t in var_types.items()
    }

    out_types = {v: infer_variable_type(v, type_cache) for v in output_vars}
    out_bounds = {
        v: (_get_float_bounds_for_var(v) if t == "float" else get_boundary_values(t))
        for v, t in out_types.items()
    }

    logic_flow = unit.get("logic_flow") or []

    strategies = [
        ("BV_MIN_INV", "min_inv"),
        ("BV_MIN",     "min"),
        ("BV_MID",     "mid"),
        ("BV_MAX",     "max"),
        ("BV_MAX_INV", "max_inv"),
        ("MIXED",      None),
    ]

    # ── Additional strategies for branch coverage ──
    # GAP 1: Condition combination — toggle each input while others stay at mid
    if len(input_vars) >= 2:
        for toggle_idx in range(min(4, len(input_vars))):
            strategies.append((f"COND_COMB_{toggle_idx}", f"_cond_{toggle_idx}"))

    # GAP 2: Switch-case — generate TC per enum/case value from logic_flow
    _extra_switch = _extract_switch_cases(logic_flow, input_vars)[:6]
    for sw_idx in range(len(_extra_switch)):
        strategies.append((f"SWITCH_{sw_idx}", f"_switch_{sw_idx}"))

    # GAP 3: Loop boundary — 0/1/max iterations for loop-containing functions
    _has_loop = any(
        str(n.get("type", "")).lower() == "loop" for n in logic_flow if isinstance(n, dict)
    )
    _loop_var = ""
    if _has_loop:
        for n in logic_flow:
            if str(n.get("type", "")).lower() == "loop":
                cond = str(n.get("condition", ""))
                for iv in input_vars:
                    if iv.lower() in cond.lower():
                        _loop_var = iv
                        break
                break
        if not _loop_var and input_vars:
            _loop_var = input_vars[0]
        if _loop_var:
            strategies.append(("LOOP_ZERO", "_loop_0"))
            strategies.append(("LOOP_ONE", "_loop_1"))
            strategies.append(("LOOP_MAX", "_loop_max"))

    # GAP 4: Global state combination — toggle indirect (global) vars
    indirect_vars: List[str] = unit.get("indirect_vars") or []
    _extra_globals: List[str] = []
    if indirect_vars:
        for gv in indirect_vars[:3]:
            _extra_globals.append(gv)
            strategies.append((f"GLOBAL_{len(_extra_globals)-1}", f"_global_{len(_extra_globals)-1}"))

    # GAP 5: Void side-effect — for functions with inputs but no outputs,
    # add sequence using indirect_vars as expected outputs
    if input_vars and not output_vars and indirect_vars:
        strategies.append(("VOID_SIDE_EFFECT", "_void_se"))

    # GAP 6: MC/DC — Modified Condition/Decision Coverage
    # Extract conditions from logic_flow and generate True/False toggle per condition
    _mcdc_conditions = _extract_mcdc_conditions(logic_flow, input_vars, type_cache)
    # Add baseline FIRST (all conditions at true values)
    if _mcdc_conditions:
        strategies.append(("MCDC_BASE", "_mcdc_base"))
    for mc_idx in range(len(_mcdc_conditions[:6])):
        strategies.append((f"MCDC_{mc_idx}", f"_mcdc_{mc_idx}"))

    # Pre-compute clamp/guard analysis once (avoid repeated DFS per strategy)
    check_vars = output_vars or input_vars
    _has_any_clamp = any(_flow_has_clamp_pattern(logic_flow, v) for v in check_vars)
    _has_any_guard = any(_flow_has_guard_clause(logic_flow, v) for v in check_vars)

    def _resolve_inv_label(sname: str) -> str:
        """Resolve 'error/saturation' ambiguity using pre-computed flow analysis."""
        if sname not in ("BV_MIN_INV", "BV_MAX_INV"):
            return _STRAT_LABEL.get(sname, sname)
        direction = "하한 초과 (경계-1)" if sname == "BV_MIN_INV" else "상한 초과 (경계+1)"
        if _has_any_clamp and _has_any_guard:
            return f"유효 {direction}: 포화(clamp) 및 가드 조건 처리 확인"
        elif _has_any_clamp:
            return f"유효 {direction}: 포화(saturation) 처리 확인"
        elif _has_any_guard:
            return f"유효 {direction}: 가드 조건에 의한 에러 처리 확인"
        else:
            return f"유효 {direction}: 방어 처리 확인 (포화 추정)"

    sequences: List[Dict[str, Any]] = []
    for idx, (strat_name, bound_key) in enumerate(strategies[:max_seq]):
        seq_num = idx + 1
        inp_vals: Dict[str, Any] = {}
        exp_vals: Dict[str, Any] = {}

        if bound_key == "_mcdc_base":
            # MC/DC baseline: all conditions at true values
            for v in input_vars:
                bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                # Check if this var has an MCDC condition
                mc_true = None
                for mc in _mcdc_conditions:
                    if mc[0] == v:
                        mc_true = mc[3]  # true_val
                        break
                if mc_true is not None:
                    inp_vals[v] = _format_test_value(mc_true, var_types.get(v, "uint8_t"))
                else:
                    inp_vals[v] = _format_test_value(bnd.get("mid", 0), var_types.get(v, "uint8_t"))
            for v in output_vars:
                bnd = out_bounds.get(v, _DEFAULT_BOUNDARY)
                exp_vals[v] = _format_test_value(bnd.get("mid", 0), out_types.get(v, "uint8_t"))
        elif bound_key and bound_key.startswith("_mcdc_"):
            # MC/DC: toggle one condition to flip the decision outcome
            mc_idx = int(bound_key.split("_")[-1])
            if mc_idx < len(_mcdc_conditions):
                mc_var, _, _, _, mc_false_val = _mcdc_conditions[mc_idx]
                # Clamp mc_false_val to type boundary to prevent overflow (e.g. 256 for uint8)
                try:
                    _mc_type = var_types.get(mc_var, "uint8_t")
                    _mc_bnd = var_bounds.get(mc_var, _DEFAULT_BOUNDARY)
                    mc_false_val = max(_mc_bnd.get("min", 0), min(mc_false_val, _mc_bnd.get("max", 255)))
                except (TypeError, ValueError):
                    pass
                for v in input_vars:
                    bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                    if v == mc_var:
                        # Set to the false-side value (toggle from baseline true)
                        inp_vals[v] = _format_test_value(mc_false_val, var_types.get(v, "uint8_t"))
                    else:
                        inp_vals[v] = _format_test_value(bnd.get("mid", 0), var_types.get(v, "uint8_t"))
                for v in output_vars:
                    bnd = out_bounds.get(v, _DEFAULT_BOUNDARY)
                    exp_vals[v] = _format_test_value(bnd.get("mid", 0), out_types.get(v, "uint8_t"))
        elif bound_key and bound_key.startswith("_loop_"):
            # Loop boundary: set loop counter var to 0 / 1 / max
            loop_key = bound_key.split("_")[-1]
            for v in input_vars:
                bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                if v == _loop_var:
                    if loop_key == "0":
                        inp_vals[v] = _format_test_value(0, var_types.get(v, "uint8_t"))
                    elif loop_key == "1":
                        inp_vals[v] = _format_test_value(1, var_types.get(v, "uint8_t"))
                    else:  # max
                        inp_vals[v] = _format_test_value(min(bnd.get("max", 255), 255), var_types.get(v, "uint8_t"))
                else:
                    inp_vals[v] = _format_test_value(bnd.get("mid", 0), var_types.get(v, "uint8_t"))
            for v in output_vars:
                bnd = out_bounds.get(v, _DEFAULT_BOUNDARY)
                exp_vals[v] = _format_test_value(bnd.get("mid", 0), out_types.get(v, "uint8_t"))
        elif bound_key and bound_key.startswith("_global_"):
            # Global state: toggle indirect var to min, inputs at mid
            gv_idx = int(bound_key.split("_")[-1])
            for v in input_vars:
                bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                inp_vals[v] = _format_test_value(bnd.get("mid", 0), var_types.get(v, "uint8_t"))
            # Add the global var as input with boundary value
            if gv_idx < len(_extra_globals):
                gv = _extra_globals[gv_idx]
                gv_type = infer_variable_type(gv, type_cache)
                gv_bnd = _get_float_bounds_for_var(gv) if gv_type == "float" else get_boundary_values(gv_type)
                inp_vals[gv] = _format_test_value(gv_bnd.get("min", 0), var_types.get(gv, gv_type) if gv in var_types else gv_type)
            for v in output_vars:
                bnd = out_bounds.get(v, _DEFAULT_BOUNDARY)
                exp_vals[v] = _format_test_value(bnd.get("mid", 0), out_types.get(v, "uint8_t"))
        elif bound_key == "_void_se":
            # Void side-effect: input at boundary, check globals as expected
            for v in input_vars:
                bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                inp_vals[v] = _format_test_value(bnd.get("max_inv", bnd.get("max", 255) + 1), var_types.get(v, "uint8_t"))
            for gv in _extra_globals:
                gv_type = infer_variable_type(gv, type_cache)
                gv_bnd = _get_float_bounds_for_var(gv) if gv_type == "float" else get_boundary_values(gv_type)
                exp_vals[gv] = _format_test_value(gv_bnd.get("mid", 0), gv_type)
        elif bound_key and bound_key.startswith("_cond_"):
            # Condition combination: toggle one input to min, others stay at mid
            toggle_idx = int(bound_key.split("_")[-1])
            for i, v in enumerate(input_vars):
                bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                if i == toggle_idx:
                    raw = bnd.get("min", 0) if toggle_idx % 2 == 0 else bnd.get("max", 0)
                else:
                    raw = bnd.get("mid", 0)  # others at mid
                inp_vals[v] = _format_test_value(raw, var_types.get(v, "uint8_t"))
            for v in output_vars:
                bnd = out_bounds.get(v, _DEFAULT_BOUNDARY)
                exp_vals[v] = _format_test_value(bnd.get("mid", 0), out_types.get(v, "uint8_t"))
        elif bound_key and bound_key.startswith("_switch_"):
            # Switch-case: set target var to specific case value
            sw_idx = int(bound_key.split("_")[-1])
            if sw_idx < len(_extra_switch):
                sw_var, sw_val, _ = _extra_switch[sw_idx]
                for v in input_vars:
                    bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                    if v == sw_var:
                        inp_vals[v] = _format_test_value(sw_val, var_types.get(v, "uint8_t"))
                    else:
                        inp_vals[v] = _format_test_value(bnd.get("mid", 0), var_types.get(v, "uint8_t"))
                for v in output_vars:
                    bnd = out_bounds.get(v, _DEFAULT_BOUNDARY)
                    exp_vals[v] = _format_test_value(bnd.get("mid", 0), out_types.get(v, "uint8_t"))
        elif bound_key:
            for v in input_vars:
                bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                raw = bnd.get(bound_key, 0)
                inp_vals[v] = _format_test_value(raw, var_types.get(v, "uint8_t"))
            for v in output_vars:
                bnd = out_bounds.get(v, _DEFAULT_BOUNDARY)
                raw = _infer_expected_for_strategy(
                    bnd, bound_key, out_types.get(v, "uint8_t"), logic_flow, v
                )
                exp_vals[v] = _format_test_value(raw, out_types.get(v, "uint8_t"))
        else:
            for i, v in enumerate(input_vars):
                bnd = var_bounds.get(v, _DEFAULT_BOUNDARY)
                key = "min" if i % 2 == 0 else "max"
                raw = bnd.get(key, 0)
                inp_vals[v] = _format_test_value(raw, var_types.get(v, "uint8_t"))
            for v in output_vars:
                bnd = out_bounds.get(v, _DEFAULT_BOUNDARY)
                raw = bnd.get("mid", 0)
                exp_vals[v] = _format_test_value(raw, out_types.get(v, "uint8_t"))

        # Build human-readable description showing actual variable names and values
        label = _resolve_inv_label(strat_name) if strat_name in _STRAT_LABEL else (
            _get_strategy_label(strat_name, input_vars, _extra_switch,
                                _loop_var if _has_loop else "", _extra_globals)
        )
        inp_parts = [f"{v}={inp_vals[v]}" for v in input_vars if v in inp_vals]
        exp_parts = [f"{v}={exp_vals[v]}" for v in output_vars if v in exp_vals]
        # Include extra expected vars (e.g., globals from VOID_SIDE_EFFECT)
        for v in exp_vals:
            if v not in output_vars:
                exp_parts.append(f"{v}={exp_vals[v]}")
        desc_lines = [label]
        if inp_parts:
            desc_lines.append("Input: " + ", ".join(inp_parts))
        if exp_parts:
            desc_lines.append("Expected: " + ", ".join(exp_parts))
        description = "\n".join(desc_lines)

        sequences.append({
            "seq_num": seq_num,
            "inputs": inp_vals,
            "expected": exp_vals,
            "strategy": strat_name,
            "description": description,
        })

    return sequences


def _format_test_value(value: Any, typename: str) -> Any:
    """Format test values to match reference document patterns.

    - bit/bool: use hex (0x0, 0x1)
    - REG/hardware: use hex for small ints
    - others: plain integer
    """
    if value is None:
        return 0
    if isinstance(value, float) and value == int(value):
        value = int(value)
    if typename in ("bit", "bool"):
        if isinstance(value, (int, float)):
            iv = int(value)
            if 0 <= iv <= 0xFF:
                return f"0x{iv:X}"
    return value


def _infer_expected_for_strategy(
    bounds: Dict[str, Any],
    strategy_key: str,
    typename: str,
    logic_flow: List[Dict[str, Any]],
    var_name: str,
) -> Any:
    """Infer expected output value based on input strategy and logic analysis.

    Uses logic_flow branch analysis to determine more accurate expected outputs:
    - Error-boundary inputs (min_inv/max_inv): saturation, clamping, or error flag
    - Valid boundary inputs (min/max): the boundary value itself (pass-through or capped)
    - Mid-range inputs: normal processing result
    """
    is_bit = typename in ("bit", "bool")
    has_guard = _flow_has_guard_clause(logic_flow, var_name)
    has_clamp = _flow_has_clamp_pattern(logic_flow, var_name)
    is_enable_flag = _is_enable_disable_var(var_name)
    is_counter = _is_counter_accumulator_var(var_name)
    is_state_var = _is_state_machine_var(var_name)
    bmin = bounds.get("min", 0)
    bmax = bounds.get("max", 0)
    bmid = bounds.get("mid", 0)
    # Pre-compute type normalization once (used by min_inv/max_inv fallback)
    _normalized = typename.lower().replace(" ", "").replace("_t", "")
    _is_known_type = _normalized in _KNOWN_SATURATE_TYPES

    # Enable/disable flag: output toggles between 0/1 on valid input
    if is_enable_flag and strategy_key in ("min", "BV_MIN"):
        return 0
    if is_enable_flag and strategy_key in ("max", "BV_MAX"):
        return 1

    # Counter/accumulator: mid-range or clamp at max on overflow
    if is_counter:
        if strategy_key == "max_inv":
            return bmax  # saturates/wraps at max
        if strategy_key == "min_inv":
            return bmin  # saturates at min

    # State machine variable: invalid input → stays in safe/init state
    if is_state_var and strategy_key in ("min_inv", "max_inv"):
        return bmin  # remain in initial/safe state on invalid transition

    if strategy_key == "min_inv":
        if is_bit:
            return bmax
        if has_clamp:
            return bmin   # clamped to lower bound
        if has_guard:
            return bmin   # guarded: stays at safe min value
        if _is_known_type:
            return bmin   # type-inferred saturation to lower bound
        raw = bounds.get("min_inv", bmin)
        return f"[검증 필요] {raw}"

    if strategy_key == "max_inv":
        if is_bit:
            return bmax
        if has_clamp:
            return bmax   # clamped to upper bound
        if has_guard:
            return bmax   # guarded: stays at safe max value
        if _is_known_type:
            return bmax   # type-inferred saturation to upper bound
        raw = bounds.get("max_inv", bmax)
        return f"[검증 필요] {raw}"

    if strategy_key == "min":
        if is_bit:
            return 0
        return bmin

    if strategy_key == "max":
        return bmax

    return bmid


def _is_enable_disable_var(var_name: str) -> bool:
    """Check if variable name indicates an enable/disable flag or activation signal."""
    name = var_name.lower()
    keywords = ("enable", "disable", "active", "flag", "en_", "_en", "on_", "_on",
                 "inhibit", "valid", "allowed", "permit")
    return any(kw in name for kw in keywords)


def _is_counter_accumulator_var(var_name: str) -> bool:
    """Check if variable name indicates a counter or accumulator."""
    name = var_name.lower()
    keywords = ("count", "cnt", "accum", "sum", "total", "index", "idx",
                 "tick", "timer", "elapsed", "delta")
    return any(kw in name for kw in keywords)


def _is_state_machine_var(var_name: str) -> bool:
    """Check if variable name indicates a state machine variable."""
    name = var_name.lower()
    keywords = ("state", "_st_", "_sts", "status", "mode", "phase", "stage",
                 "step", "fsm", "_sm_")
    return any(kw in name for kw in keywords)


def _extract_mcdc_conditions(
    logic_flow: List[Dict[str, Any]],
    input_vars: List[str],
    type_cache: Optional[Dict[str, str]] = None,
) -> List[Tuple[str, str, Any, Any, Any]]:
    """Extract MC/DC-relevant conditions from logic_flow.

    Returns list of (variable, operator, threshold, true_value, false_value) tuples.
    For 'if (A > 10)': variable=A, op='>', threshold=10, true_val=11, false_val=10 (boundary value)
    """
    conditions: List[Tuple[str, str, Any, Any, Any]] = []
    seen_keys: set = set()
    _OPS = {"<": ("<", lambda t: t - 1, lambda t: t),
            ">": (">", lambda t: t + 1, lambda t: t),
            "<=": ("<=", lambda t: t, lambda t: t + 1),
            ">=": (">=", lambda t: t, lambda t: t - 1),
            "==": ("==", lambda t: t, lambda t: t + 1),
            "!=": ("!=", lambda t: t + 1, lambda t: t)}

    for node in logic_flow:
        ntype = str(node.get("type", "")).lower()
        if ntype != "if":
            # Recurse
            for child in node.get("children", []):
                conditions.extend(_extract_mcdc_conditions([child], input_vars, type_cache))
            continue

        cond = str(node.get("condition", "")).strip()
        if not cond:
            for child in node.get("children", []):
                conditions.extend(_extract_mcdc_conditions([child], input_vars, type_cache))
            continue

        # Parse conditions: "var > 10", "var >= other_var", "var == CONST"
        # Also extract condition variables NOT in input_vars (locals, constants)
        _cond_vars = re.findall(r"[a-zA-Z_]\w+", cond)
        _all_vars = list(input_vars)
        for cv in _cond_vars:
            if cv.lower() not in {v.lower() for v in _all_vars} and len(cv) > 2:
                if cv.lower() not in ("if", "else", "true", "false", "null", "void", "return"):
                    _all_vars.append(cv)

        for iv in _all_vars:
            for op_str, (op_label, true_fn, false_fn) in _OPS.items():
                # Pattern 1: var OP numeric_constant ("var > 10", "var>=0x0A")
                pat_num = re.compile(
                    rf"(?:^|[^a-zA-Z_]){re.escape(iv)}\s*{re.escape(op_str)}\s*([\-]?(?:0[xX][0-9a-fA-F]+|\d+))",
                    re.IGNORECASE,
                )
                m = pat_num.search(cond)
                if m:
                    try:
                        threshold = int(m.group(1), 0)
                        true_val = true_fn(threshold)
                        false_val = false_fn(threshold)
                        key = (iv, op_str, threshold)
                        if key not in seen_keys:
                            seen_keys.add(key)
                            conditions.append((iv, op_label, threshold, true_val, false_val))
                    except (ValueError, TypeError):
                        pass
                    continue

                # Pattern 2: var OP other_variable ("var >= other_var")
                # Also try: other_var OP input_var (reversed operand order)
                pat_var = re.compile(
                    rf"(?:^|[^a-zA-Z_]){re.escape(iv)}\s*{re.escape(op_str)}\s*([a-zA-Z_]\w+)",
                    re.IGNORECASE,
                )
                # Also match when input_var is on the RIGHT side: "local_var >= input_var"
                _reverse_ops = {">": "<", "<": ">", ">=": "<=", "<=": ">=", "==": "==", "!=": "!="}
                pat_rev = re.compile(
                    rf"([a-zA-Z_]\w+)\s*{re.escape(op_str)}\s*{re.escape(iv)}(?:[^a-zA-Z_]|$)",
                    re.IGNORECASE,
                )
                m2 = pat_var.search(cond) or pat_rev.search(cond)
                if m2:
                    rhs_var = m2.group(1)
                    # Use boundary values of the input variable for MC/DC toggle
                    iv_type = infer_variable_type(iv, type_cache)
                    iv_bnd = (
                        _get_float_bounds_for_var(iv) if iv_type == "float"
                        else get_boundary_values(iv_type)
                    )
                    mid = iv_bnd.get("mid", 127)
                    bmin = iv_bnd.get("min", 0)
                    bmax = iv_bnd.get("max", 255)
                    # For "var >= other": true when var is high, false when var is low
                    if op_str in (">", ">="):
                        true_val = bmax
                        false_val = bmin
                    elif op_str in ("<", "<="):
                        true_val = bmin
                        false_val = bmax
                    elif op_str == "==":
                        true_val = mid
                        false_val = bmin if mid != bmin else bmax
                    else:  # !=
                        true_val = bmin if mid != bmin else bmax
                        false_val = mid
                    key = (iv, op_str, rhs_var)
                    if key not in seen_keys:
                        seen_keys.add(key)
                        conditions.append((iv, op_label, rhs_var, true_val, false_val))

        # Recurse into children
        for child in node.get("children", []):
            conditions.extend(_extract_mcdc_conditions([child], input_vars, type_cache))

    return conditions


def _extract_switch_cases(
    logic_flow: List[Dict[str, Any]],
    input_vars: List[str],
) -> List[Tuple[str, Any, str]]:
    """Extract switch-case values from logic_flow for branch coverage.

    Returns list of (variable_name, case_value, case_label) tuples.
    """
    cases: List[Tuple[str, Any, str]] = []
    for node in logic_flow:
        ntype = str(node.get("type", "")).lower()
        # switch-case nodes
        if ntype == "switch":
            sw_var = str(node.get("variable", "") or node.get("condition", "")).strip()
            # Match to input_vars
            matched_var = ""
            for iv in input_vars:
                if iv.lower() in sw_var.lower() or sw_var.lower() in iv.lower():
                    matched_var = iv
                    break
            if not matched_var and input_vars:
                matched_var = input_vars[0]
            for child in node.get("children", []) or node.get("cases", []):
                case_val = child.get("value")
                if case_val is None:
                    case_val = child.get("case")
                case_label = str(child.get("label", "") or child.get("text", "") or f"case_{case_val}")
                if case_val is not None:
                    cases.append((matched_var, case_val, case_label))
        # if-else chains that look like enum comparisons (e.g., "var == ENUM_VAL")
        elif ntype == "if":
            cond = str(node.get("condition", "")).strip()
            for iv in input_vars:
                if iv.lower() in cond.lower() and "==" in cond:
                    # Extract the compared value
                    parts = cond.split("==")
                    if len(parts) == 2:
                        val_str = parts[1].strip().strip("() ")
                        try:
                            val = int(val_str, 0)  # supports 0x hex
                            cases.append((iv, val, f"조건 {iv}=={val_str}"))
                        except ValueError:
                            pass  # Skip enum names — can't use as numeric test input
        # Recurse into children (for non-switch nodes only)
        if ntype != "switch":
            for child in node.get("children", []):
                cases.extend(_extract_switch_cases([child], input_vars))
    # Deduplicate by (var, val)
    seen = set()
    unique: List[Tuple[str, Any, str]] = []
    for c in cases:
        key = (c[0], str(c[1]))
        if key not in seen:
            seen.add(key)
            unique.append(c)
    return unique


def _flow_has_guard_clause(logic_flow: List[Dict[str, Any]], var_name: str) -> bool:
    """Check if logic_flow contains an if-guard referencing var_name (range check)."""
    clean = var_name.lower().strip()
    for node in logic_flow:
        cond = str(node.get("condition", "") or node.get("text", "")).lower()
        if clean in cond:
            for kw in ("<", ">", "<=", ">=", "==", "!=", "min", "max", "limit"):
                if kw in cond:
                    return True
        for child in node.get("children", []):
            if _flow_has_guard_clause([child], var_name):
                return True
    return False


def _flow_has_clamp_pattern(logic_flow: List[Dict[str, Any]], var_name: str) -> bool:
    """Check if logic_flow contains a clamp/saturation pattern for var_name."""
    clean = var_name.lower().strip()
    for node in logic_flow:
        text = str(node.get("text", "") or node.get("condition", "")).lower()
        if clean in text:
            for kw in ("clamp", "saturate", "limit", "cap", "bound", "clip"):
                if kw in text:
                    return True
            if ("=" in text) and any(w in text for w in ("max", "min", "0xff", "0xffff")):
                return True
        for child in node.get("children", []):
            if _flow_has_clamp_pattern([child], var_name):
                return True
    return False


# ---------------------------------------------------------------------------
# Phase 4: AI Enhancement (optional)
# ---------------------------------------------------------------------------

_SUTS_AI_SYSTEM_PROMPT = (
    "You are a software unit test engineer writing SUTS for automotive ECU software (ISO 26262).\n"
    "Given a C function context and test sequences, provide accurate expected output values.\n"
    "Rules:\n"
    "- Analyze the function name, description, calls, and logic conditions to infer behavior.\n"
    "- For void/no-param functions: use 'Indirect variables' as testable state variables.\n"
    "  NORMAL_CALL: expected = typical post-call values (e.g., initialized/reset state).\n"
    "  ERROR_PATH: expected = safe/default state after error (0 or initial value).\n"
    "  REPEAT_CALL: expected = same stable state (idempotent).\n"
    "- For functions with inputs: boundary-exceeding inputs → clamped/saturated expected output.\n"
    "- Return ONLY a JSON array: [{\"seq_num\":1, \"expected\":{\"var\":value,...}}, ...]\n"
    "- Only set expected values for variables that appear in 'Indirect variables' or 'Output variables'.\n"
    "- Values must be numeric (int or float). Use 0 for unknown/safe defaults."
)


_AI_TIMEOUT_SEC = 30
_AI_MAX_RETRIES = 2


def _ai_call_with_retry(agent_call_fn, ai_config, messages, *,
                         stage: str, max_retries: int = _AI_MAX_RETRIES,
                         timeout: int = _AI_TIMEOUT_SEC,
                         temperature: float = 0.2) -> str:
    """Wrapper around agent_call with timeout and retry logic."""
    import json as _json
    import threading

    last_err: Optional[Exception] = None
    for attempt in range(1, max_retries + 1):
        result_holder: Dict[str, Any] = {}
        exc_holder: List[Exception] = []

        def _invoke():
            try:
                r = agent_call_fn(
                    ai_config, messages,
                    role="writer", stage=stage,
                    settings={"temperature": temperature},
                )
                result_holder["val"] = r
            except Exception as ex:
                exc_holder.append(ex)

        t = threading.Thread(target=_invoke, daemon=True)
        t.start()
        t.join(timeout=timeout)

        if t.is_alive():
            _logger.warning("AI call timed out (attempt %d/%d, %ds)", attempt, max_retries, timeout)
            last_err = TimeoutError(f"AI call timed out after {timeout}s")
            continue

        if exc_holder:
            last_err = exc_holder[0]
            _logger.warning("AI call error (attempt %d/%d): %s", attempt, max_retries, last_err)
            continue

        raw = result_holder.get("val")
        reply = raw.get("output", "") if isinstance(raw, dict) else ""
        if reply:
            return reply
        _logger.warning("AI returned empty response (attempt %d/%d)", attempt, max_retries)
        last_err = ValueError("Empty AI response")

    if last_err:
        _logger.warning("AI call exhausted retries: %s", last_err)
    return ""


def _parse_ai_json(reply: str, expect_list: bool = True) -> Any:
    """Parse AI response as JSON with fallback regex extraction."""
    import json as _json
    if not reply:
        return None
    try:
        payload = _json.loads(reply) if isinstance(reply, str) else reply
        if expect_list and isinstance(payload, list):
            return payload
        if not expect_list and isinstance(payload, dict):
            return payload
        return payload
    except Exception:
        pattern = r"\[[\s\S]*\]" if expect_list else r"\{[\s\S]*\}"
        m = re.search(pattern, reply)
        if m:
            try:
                return _json.loads(m.group())
            except Exception:
                pass
    return None


def _validate_ai_sequence_item(item: Any, valid_seq_nums: set) -> bool:
    """Validate a single AI-enhanced sequence item."""
    if not isinstance(item, dict):
        return False
    if "seq_num" not in item or "expected" not in item:
        return False
    if item["seq_num"] not in valid_seq_nums:
        return False
    if not isinstance(item["expected"], dict):
        return False
    return True


def enhance_sequences_with_ai(
    unit: Dict[str, Any],
    sequences: List[Dict[str, Any]],
    ai_config: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """Enhance expected output values using AI with timeout and retry."""
    if not ai_config:
        return sequences

    try:
        from workflow.ai import agent_call
    except ImportError:
        _logger.warning("workflow.ai not available; skipping AI enhancement")
        return sequences

    _inp_vars = unit.get("input_vars") or []
    _out_vars = unit.get("output_vars") or []
    _indirect = unit.get("indirect_vars") or []
    _calls = unit.get("calls_list") or []
    _lf = unit.get("logic_flow") or []

    # Summarise logic_flow conditions for AI context
    _cond_lines: List[str] = []
    def _collect_conds(nodes: List[Any], depth: int = 0) -> None:
        for _n in nodes:
            if not isinstance(_n, dict):
                continue
            _c = str(_n.get("condition") or _n.get("text") or "").strip()
            if _c and depth < 3:
                _cond_lines.append(_c[:100])
            for _key in ("true_body", "false_body", "body"):
                _sub = _n.get(_key)
                if isinstance(_sub, list):
                    _collect_conds(_sub, depth + 1)
    _collect_conds(_lf)

    func_ctx = (
        f"Function: {unit.get('prototype', '')}\n"
        f"Description: {unit.get('description', '')}\n"
        f"Input variables: {_inp_vars}\n"
        f"Output variables: {_out_vars}\n"
        f"Calls: {_calls[:8]}\n"
    )
    if _indirect:
        func_ctx += f"Indirect variables (from callees): {_indirect}\n"
    if _cond_lines:
        func_ctx += f"Logic conditions: {_cond_lines[:6]}\n"

    seq_info = "Current sequences:\n"
    for s in sequences:
        seq_info += f"  Seq {s['seq_num']} ({s['strategy']}): inputs={s['inputs']}, expected={s['expected']}\n"

    reply = _ai_call_with_retry(
        agent_call, ai_config,
        [
            {"role": "system", "content": _SUTS_AI_SYSTEM_PROMPT},
            {"role": "user", "content": func_ctx + "\n" + seq_info},
        ],
        stage="suts_enhance",
        temperature=0.2,
    )

    payload = _parse_ai_json(reply, expect_list=True)
    if isinstance(payload, list):
        valid_nums = {s["seq_num"] for s in sequences}
        seq_map = {s["seq_num"]: s for s in sequences}
        applied = 0
        for item in payload:
            if _validate_ai_sequence_item(item, valid_nums):
                seq_map[item["seq_num"]]["expected"].update(item["expected"])
                applied += 1
        if applied:
            _logger.info("AI enhanced %d/%d sequences for %s", applied, len(payload), unit.get("name"))

    return sequences


# ---------------------------------------------------------------------------
# Phase 5: XLSM output
# ---------------------------------------------------------------------------

def generate_suts_xlsm(
    template_path: Optional[str],
    units: List[Dict[str, Any]],
    all_sequences: Dict[str, List[Dict[str, Any]]],
    output_path: str,
    project_config: Optional[Dict[str, Any]] = None,
) -> str:
    """Generate SUTS XLSM file."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        _logger.error("openpyxl not installed")
        raise

    cfg = project_config or {}
    project_id = cfg.get("project_id", "PROJECT")
    doc_id = cfg.get("doc_id", f"{project_id}-SUTS")
    version = cfg.get("version", "v1.00")
    asil_level = cfg.get("asil_level", "")

    if template_path and Path(template_path).is_file():
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        _logger.info("Loaded SUTS template: %s", template_path)
    else:
        wb = openpyxl.Workbook()
        _create_suts_cover(wb, project_id, doc_id, version, asil_level)
        _create_suts_history(wb, version)
        _create_suts_intro(wb)
        _create_suts_test_env(wb)
        _logger.info("Created new SUTS workbook (no template)")

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    hdr_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    hdr_font = Font(name="맑은 고딕", size=9, bold=True)
    data_font = Font(name="맑은 고딕", size=8)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 정본을 템플릿으로 쓰면 **과거 개정 이력이 그대로 딸려온다**. 지우지 않고
    # 다음 행에 이번 개정을 덧붙인다(사용자 결정, 2026-08-12) — 그게 개정 이력의
    # 본래 쓰임이고, 지우면 문서가 어디서 왔는지 사라진다.
    from generators.history_row import append_history_row
    append_history_row(wb, version=version, description=str(cfg.get("history_note") or ""))

    sheet_name = "2.SW Unit Test Spec"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # --- Title row (row 1, merged A1 to last column — matches reference A1:EG1) ---
    title_font = Font(name="맑은 고딕", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_RELATED_COL)
    ws.cell(row=1, column=1, value="Software Unit Test Specification").font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # --- Header rows (5-6) ---
    # Row 5: group headers — merged spans
    def _fill_and_merge(row, c_start, c_end, label):
        for c in range(c_start, c_end + 1):
            ws.cell(row=row, column=c).fill = hdr_fill
            ws.cell(row=row, column=c).border = thin
            ws.cell(row=row, column=c).alignment = center
        ws.cell(row=row, column=c_start, value=label).font = hdr_font
        if c_end > c_start:
            try:
                ws.merge_cells(
                    start_row=row, start_column=c_start,
                    end_row=row, end_column=c_end,
                )
            except Exception:
                pass

    # 밴드 행 — 정본 실측: B3:G3 · H3:CZ3 · DA3:GF3 · GG3
    # ⚠ 'Input' 밴드는 시퀀스 번호 열(H)부터 시작한다. 정본 그대로다.
    _fill_and_merge(_BAND_ROW, _COL_INDEX, _COL_GEN, "Test Case")
    _fill_and_merge(_BAND_ROW, _SEQ_COL, _INPUT_COL_END, "Input")
    _fill_and_merge(_BAND_ROW, _OUTPUT_COL_START, _OUTPUT_COL_END, "Expected Result")
    _fill_and_merge(_BAND_ROW, _RELATED_COL, _RELATED_COL, "Related ID")
    ws.row_dimensions[_BAND_ROW].height = 18

    # 헤더 행 — 정의는 모듈 상수 `_FIXED_HEADERS`(단일 출처).
    for c, h in _FIXED_HEADERS.items():
        cell = ws.cell(row=_HEADER_ROW, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = center

    ws.cell(row=_HEADER_ROW, column=_RELATED_COL, value=_RELATED_HEADER).font = hdr_font
    ws.cell(row=_HEADER_ROW, column=_RELATED_COL).fill = hdr_fill
    ws.cell(row=_HEADER_ROW, column=_RELATED_COL).border = thin
    ws.cell(row=_HEADER_ROW, column=_RELATED_COL).alignment = center

    # Inpt[n] / ExpR[n] 슬롯 라벨 — 정본은 헤더 행에 이 이름을 둔다(변수명은 TC 행).
    for idx, col in enumerate(range(_INPUT_COL_START, _INPUT_COL_END + 1)):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=f"Inpt[{idx}]")
        cell.font, cell.fill, cell.border, cell.alignment = hdr_font, hdr_fill, thin, center
    for idx, col in enumerate(range(_OUTPUT_COL_START, _OUTPUT_COL_END + 1)):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=f"ExpR[{idx}]")
        cell.font, cell.fill, cell.border, cell.alignment = hdr_font, hdr_fill, thin, center
    ws.row_dimensions[_HEADER_ROW].height = 34.5

    # Column widths — 정본 기준(고정 열만; Inpt/ExpR 는 변수명 길이로 아래에서 잡는다)
    ws.column_dimensions["B"].width = 7       # Index
    ws.column_dimensions["C"].width = 24      # TC_ID
    ws.column_dimensions["D"].width = 34      # Unit
    ws.column_dimensions["E"].width = 9       # Safety Related
    ws.column_dimensions["F"].width = 11      # Test Method
    ws.column_dimensions["G"].width = 14      # TC Generation Method
    ws.column_dimensions["H"].width = 5       # 시퀀스 번호
    ws.column_dimensions[get_column_letter(_RELATED_COL)].width = 16  # SUDS

    # --- Data rows ---
    row_num = _DATA_START_ROW
    tc_count = 0
    total_seq = 0

    for unit in units:
        fid = unit["fid"]
        seqs = all_sequences.get(fid, [])
        if not seqs:
            seqs = [{"seq_num": 1, "inputs": {}, "expected": {}, "strategy": "N/A"}]

        # SUDS(설계 ID) — 확보하지 못하면 **빈칸**으로 둔다.
        # ⚠ 이전 판은 `fid`(= 소스 파싱 순번 `SwUFn_{n:04d}`)를 그대로 적었다. 그건
        #   SwUDS 가 부여한 설계 ID 가 아니라 이 실행에서 만든 번호라, 모양만 맞고
        #   **다른 설계 요소를 가리킨다**(정본과 교집합 178/251 — 나머지는 오조준).
        #   틀린 ID 가 추적성으로 보이는 것이 빈칸보다 나쁘다.
        suds_id = str(unit.get("suds_id") or unit.get("design_id") or "").strip()
        # 정본은 `TC_ID = "SwUTC_" + SUDS`(1,013/1,014). 설계 ID 를 확보한 경우에만
        # 그 규칙을 따르고, 못 찾았으면 종전대로 내부 fid 로 만든다 — TC_ID 는 시트의
        # 키라 비울 수 없기 때문이다(비우면 행을 식별할 수 없다).
        tc_id = f"SwUTC_{suds_id or fid}"
        start_row = row_num

        # TC 정의 행 — 정본에서 이 행은 **변수명 행**이다. 시퀀스 번호·Test Method·
        # TC Gen Method 는 여기 쓰지 않는다(아래 시퀀스 그룹에서 쓴다).
        ws.cell(row=row_num, column=_COL_INDEX, value=tc_count + 1).font = data_font
        ws.cell(row=row_num, column=_COL_INDEX).alignment = center
        ws.cell(row=row_num, column=_COL_TC_ID, value=tc_id).font = data_font
        ws.cell(row=row_num, column=_COL_UNIT, value=unit["name"]).font = data_font
        ws.cell(row=row_num, column=_COL_SAFETY,
                value=resolve_safety_related(unit.get("asil"))).font = data_font
        ws.cell(row=row_num, column=_COL_SAFETY).alignment = center
        if suds_id:
            ws.cell(row=row_num, column=_RELATED_COL, value=suds_id).font = data_font
            ws.cell(row=row_num, column=_RELATED_COL).alignment = center

        # Input variable names in TC row
        input_vars = unit.get("input_vars", [])
        for vi, vname in enumerate(input_vars):
            col = _INPUT_COL_START + vi
            if col > _INPUT_COL_END:
                break
            cell = ws.cell(row=row_num, column=col, value=vname)
            cell.font = hdr_font
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col)].width = max(
                12, min(len(vname) + 2, 24)
            )

        # Output variable names in TC row
        output_vars = unit.get("output_vars", [])
        for vi, vname in enumerate(output_vars):
            col = _OUTPUT_COL_START + vi
            if col > _OUTPUT_COL_END:
                break
            cell = ws.cell(row=row_num, column=col, value=vname)
            cell.font = hdr_font
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col)].width = max(
                12, min(len(vname) + 2, 24)
            )

        # Apply borders to TC row
        max_data_col = max(
            12,
            _INPUT_COL_START + len(input_vars) - 1,
            _OUTPUT_COL_START + len(output_vars) - 1,
            _RELATED_COL,
        )
        for c in range(2, max_data_col + 1):
            ws.cell(row=row_num, column=c).border = thin
            ws.cell(row=row_num, column=c).alignment = wrap

        row_num += 1

        # 시퀀스 행 — Test Method / TC Gen Method 는 **연속된 같은 값끼리 묶어** 쓴다.
        # 정본이 그렇게 돼 있다(첫 TC: seq 1~3 = REQ, 4~7 = FI 로 F열이 두 번 병합).
        # 값이 바뀌는 지점에서만 쓰고, 아래 `_seq_groups` 로 병합한다.
        _seq_groups: List[Tuple[int, int, str, str]] = []   # (start_row, end_row, method, gen)
        for seq in seqs:
            ws.cell(row=row_num, column=_SEQ_COL, value=seq["seq_num"]).font = data_font
            ws.cell(row=row_num, column=_SEQ_COL).alignment = center
            ws.cell(row=row_num, column=_SEQ_COL).border = thin

            strategy_val = str(seq.get("strategy", "") or "")
            s_method = resolve_seq_test_method(strategy_val)
            s_gen = resolve_seq_gen_method(strategy_val)
            if _seq_groups and _seq_groups[-1][2] == s_method and _seq_groups[-1][3] == s_gen:
                g = _seq_groups[-1]
                _seq_groups[-1] = (g[0], row_num, g[2], g[3])
            else:
                _seq_groups.append((row_num, row_num, s_method, s_gen))

            # Input values
            for vi, vname in enumerate(input_vars):
                col = _INPUT_COL_START + vi
                if col > _INPUT_COL_END:
                    break
                val = seq.get("inputs", {}).get(vname)
                if val is not None:
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.font = data_font
                    cell.alignment = center
                    cell.border = thin

            # Expected output values
            for vi, vname in enumerate(output_vars):
                col = _OUTPUT_COL_START + vi
                if col > _OUTPUT_COL_END:
                    break
                val = seq.get("expected", {}).get(vname)
                if val is not None:
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.font = data_font
                    cell.alignment = center
                    cell.border = thin

            row_num += 1
            total_seq += 1

        # Test Method / TC Gen Method — 그룹의 첫 행에 쓰고 그룹 범위로 병합한다.
        for g_start, g_end, g_method, g_gen in _seq_groups:
            for col, val in ((_COL_METHOD, g_method), (_COL_GEN, g_gen)):
                cell = ws.cell(row=g_start, column=col, value=val)
                cell.font = data_font
                cell.alignment = center
                cell.border = thin
                if g_end > g_start:
                    try:
                        ws.merge_cells(start_row=g_start, start_column=col,
                                       end_row=g_end, end_column=col)
                    except Exception as exc:  # noqa: BLE001
                        _logger.debug("seq group merge skipped (%s%d:%d): %s",
                                      get_column_letter(col), g_start, g_end, exc)

        # TC 메타 열은 블록 전체 병합 — 정본과 같다(B/C/D/E/GG 가 5:12 처럼 덮인다).
        # ⚠ Test Method(F)·TC Gen Method(G)는 **제외**한다. 시퀀스 그룹 단위라
        #   블록 전체로 병합하면 REQ/FI 구분이 사라진다.
        end_row = row_num - 1
        tc_def_row = start_row
        merge_cols = [_COL_INDEX, _COL_TC_ID, _COL_UNIT, _COL_SAFETY, _RELATED_COL]
        if end_row > tc_def_row:
            for mc in merge_cols:
                try:
                    ws.merge_cells(
                        start_row=tc_def_row, start_column=mc,
                        end_row=end_row, end_column=mc,
                    )
                except Exception as exc:  # noqa: BLE001
                    _logger.debug("TC meta merge skipped (col %d, %d:%d): %s",
                                  mc, tc_def_row, end_row, exc)

        tc_count += 1

    _logger.info("Wrote %d TCs, %d sequences to sheet", tc_count, total_seq)

    # --- Traceability sheet: Component → Function → TC ---
    _write_suts_traceability_sheet(wb, units, thin, hdr_fill, hdr_font, data_font)

    # --- Remove default sheet if we created new workbook ---
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    _logger.info("SUTS saved: %s", out)
    return str(out)


def _write_suts_traceability_sheet(wb, units, border, hdr_fill, hdr_font, data_font):
    """Write traceability sheet mapping Components → Functions → SUTS TCs."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    sheet_name = "3.Traceability"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    covered_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.cell(row=1, column=1, value="Traceability Between [SUDS] and [SUTS]").font = hdr_font

    headers = ["#", "Component", "Function ID", "Function Name", "TC ID",
               "SRS Req ID", "Input Vars", "Output Vars", "Sequences", "Gen Method", "Status"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = center

    widths = [5, 16, 16, 32, 22, 28, 10, 10, 10, 14, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    row = 4
    for idx, u in enumerate(units, 1):
        tc_id = f"SwUTC_{u['fid']}"
        n_inp = len(u.get("input_vars", []))
        n_out = len(u.get("output_vars", []))
        has_io = n_inp > 0 or n_out > 0
        has_req = bool(u.get("srs_req_ids", ""))
        status = "Covered" if has_io else "No I/O"

        ws.cell(row=row, column=1, value=idx).font = data_font
        ws.cell(row=row, column=2, value=(u.get("component") or "").split("\n")[0]).font = data_font
        ws.cell(row=row, column=3, value=u["fid"]).font = data_font
        ws.cell(row=row, column=4, value=u["name"]).font = data_font
        ws.cell(row=row, column=5, value=tc_id).font = data_font
        ws.cell(row=row, column=6, value=u.get("srs_req_ids", "")).font = data_font
        ws.cell(row=row, column=7, value=n_inp).font = data_font
        ws.cell(row=row, column=8, value=n_out).font = data_font
        ws.cell(row=row, column=9, value=len(u.get("logic_flow", []))).font = data_font
        ws.cell(row=row, column=10, value=determine_gen_method(u)).font = data_font
        ws.cell(row=row, column=11, value=status).font = data_font

        for ci in range(1, 12):
            ws.cell(row=row, column=ci).border = border
            ws.cell(row=row, column=ci).alignment = wrap
            if has_io:
                ws.cell(row=row, column=ci).fill = covered_fill

        row += 1

    # Summary at bottom
    row += 1
    total = len(units)
    with_io = sum(1 for u in units if u.get("input_vars") or u.get("output_vars") or u.get("indirect_vars"))
    with_req = sum(1 for u in units if u.get("srs_req_ids"))
    ws.cell(row=row, column=1, value="Summary").font = hdr_font
    row += 1
    ws.cell(row=row, column=1, value="Total Functions").font = data_font
    ws.cell(row=row, column=2, value=total).font = data_font
    row += 1
    ws.cell(row=row, column=1, value="With I/O (Covered)").font = data_font
    ws.cell(row=row, column=2, value=with_io).font = data_font
    row += 1
    ws.cell(row=row, column=1, value="Coverage %").font = data_font
    ws.cell(row=row, column=2, value=f"{round(with_io / max(total, 1) * 100, 1)}%").font = data_font
    row += 1
    ws.cell(row=row, column=1, value="With SRS Req ID").font = data_font
    ws.cell(row=row, column=2, value=with_req).font = data_font
    row += 1
    ws.cell(row=row, column=1, value="SRS Traceability %").font = data_font
    ws.cell(row=row, column=2, value=f"{round(with_req / max(total, 1) * 100, 1)}%").font = data_font


def _create_suts_cover(wb, project_id, doc_id, version, asil_level):
    ws = wb.active
    ws.title = "Cover"
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    title_font = Font(name="맑은 고딕", size=24, bold=True)
    label_font = Font(name="맑은 고딕", size=9, bold=True)
    data_font = Font(name="맑은 고딕", size=9)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    hdr_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column widths matching reference
    col_widths = {"A": 2.875, "B": 6.875, "C": 13.0, "D": 13.0, "E": 13.0,
                  "F": 13.0, "G": 13.0, "H": 4.625, "I": 6.875, "J": 13.0, "K": 10.625}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # B5:K5 merged — main title block (height=123 matching reference)
    ws.merge_cells("B5:K5")
    ws["B5"] = "Software Unit Test Specification\n(소프트웨어 단위테스트 명세서)"
    ws["B5"].font = title_font
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 123.0

    # I2 = "Doc. ID" label, J2:K2 merged = value
    ws["I2"] = "Doc. ID"
    ws["I2"].font = label_font
    ws["I2"].alignment = center
    ws.merge_cells("J2:K2")
    ws["J2"] = doc_id
    ws["J2"].font = data_font
    ws["J2"].alignment = center

    # I3 = "Version" label, J3:K3 merged = value
    ws["I3"] = "Version"
    ws["I3"].font = label_font
    ws["I3"].alignment = center
    ws.merge_cells("J3:K3")
    ws["J3"] = version
    ws["J3"].font = data_font
    ws["J3"].alignment = center

    info_rows = [
        ("Project", project_id),
        ("ASIL Level", asil_level),
        ("Status", "Draft"),
        ("Date", datetime.now().strftime("%Y-%m-%d")),
    ]
    for i, (label, value) in enumerate(info_rows):
        r = 21 + i
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=11)
        ws.cell(row=r, column=2, value=label).font = label_font
        ws.cell(row=r, column=2).fill = hdr_fill
        ws.cell(row=r, column=2).border = thin
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=6, value=value).font = data_font
        ws.cell(row=r, column=6).border = thin
        ws.cell(row=r, column=6).alignment = left


def _create_suts_history(wb, version):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    ws = wb.create_sheet("History")
    hdr_font = Font(name="맑은 고딕", size=10, bold=True)
    data_font = Font(name="맑은 고딕", size=9)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    hdr_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Column widths matching reference: A:1.25, B:8.375, C:9.125, D:35.5, E:8.625, F:13.0, G:13.0, H:1.25
    ws.column_dimensions["A"].width = 1.25
    ws.column_dimensions["B"].width = 8.375
    ws.column_dimensions["C"].width = 9.125
    ws.column_dimensions["D"].width = 35.5
    ws.column_dimensions["E"].width = 8.625
    ws.column_dimensions["F"].width = 13.0
    ws.column_dimensions["G"].width = 13.0
    ws.column_dimensions["H"].width = 1.25
    ws.row_dimensions[2].height = 18.0
    ws.row_dimensions[3].height = 14.25

    ws.merge_cells("B2:G2")
    ws["B2"] = "▶ Revision History"
    ws["B2"].font = Font(name="맑은 고딕", size=12, bold=True)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Version", "Date", "Description", "Author", "Reviewer", "Approver"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=2 + i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin
        c.alignment = center

    row_data = [
        (version, datetime.now().strftime("%Y.%m.%d"), "- Auto-generated", "Auto", "-", "-"),
    ]
    for ri, (ver, date, desc, author, reviewer, approver) in enumerate(row_data):
        r = 5 + ri
        for ci, val in enumerate([ver, date, desc, author, reviewer, approver]):
            cell = ws.cell(row=r, column=2 + ci, value=val)
            cell.font = data_font
            cell.border = thin


def _create_suts_intro(wb):
    ws = wb.create_sheet("1.Introduction")
    from openpyxl.styles import Font
    ws["A1"] = "Introduction"
    ws["A1"].font = Font(name="맑은 고딕", size=12, bold=True)
    ws["B3"] = "1.1 Purpose"
    ws["B3"].font = Font(name="맑은 고딕", size=10, bold=True)
    ws["B4"] = (
        "본 문서는 소프트웨어 유닛테스트 명세를 기술하는 문서이며, "
        "소프트웨어 유닛테스트 수행자에 의해서 작성된다."
    )
    ws["B6"] = (
        "유닛 소프트웨어 테스트의 근거가 되는 문서로서 정의며 "
        "유닛 소프트웨어 테스트 수행자에게 제공된다."
    )
    ws["B8"] = "1.2 Scope"
    ws["B8"].font = Font(name="맑은 고딕", size=10, bold=True)
    ws["B9"] = "본 문서는 유닛테스트 테스트 대상의 정의를 포함하며, 소프트웨어 단위 테스트의 사양을 정의하고 있다."


def _create_suts_test_env(wb):
    ws = wb.create_sheet("1.Test Environment")
    from openpyxl.styles import Font
    ws["A1"] = "Test Environments"
    ws["A1"].font = Font(name="맑은 고딕", size=12, bold=True)
    ws["B3"] = "STP의 SwUTE_01 과 테스트 환경으로 동일하다."
    ws["B6"] = "< End of Document >"


# ---------------------------------------------------------------------------
# Quality report
# ---------------------------------------------------------------------------

def generate_suts_quality_report(
    units: List[Dict[str, Any]],
    all_sequences: Dict[str, List[Dict[str, Any]]],
    total_source_functions: int = 0,
) -> Dict[str, Any]:
    total_tc = len(units)
    total_seq = sum(len(s) for s in all_sequences.values())
    total_inp = sum(len(u.get("input_vars", [])) for u in units)
    total_out = sum(len(u.get("output_vars", [])) for u in units)
    avg_seq = round(total_seq / max(total_tc, 1), 1)
    with_io = sum(
        1 for u in units
        if u.get("input_vars") or u.get("output_vars") or u.get("indirect_vars")
        or any(
            bool(s.get("inputs")) or bool(s.get("expected"))
            for s in all_sequences.get(u["fid"], [])
        )
    )
    with_logic = sum(1 for u in units if u.get("logic_flow"))

    gen_methods: Dict[str, int] = {}
    for u in units:
        gm = determine_gen_method(u)
        gen_methods[gm] = gen_methods.get(gm, 0) + 1

    components: Dict[str, int] = {}
    for u in units:
        comp = (u.get("component") or "Unknown").split("\n")[0]
        components[comp] = components.get(comp, 0) + 1

    # ⚠ 분모를 TC 수로 떨어뜨리지 않는다. 예전엔 `total_source_functions or total_tc`
    #   라, 소스 함수 수를 못 받으면 **자기 자신을 분모로** 써서 언제나 100.0% 가
    #   나왔다. 실측(KJPDS02_PV)에서는 함수 목록이 251개로 잘린 뒤 251/251 = 100% 로
    #   보고됐다 — 정본 1,014 함수 중 77.8% 를 버리고 "완전 커버" 라고 말한 것이다.
    #   재지 못했으면 **`None`**(미측정)이지 100% 가 아니다.
    src_total = int(total_source_functions or 0)
    func_coverage_pct: Optional[float] = (
        round(total_tc / src_total * 100, 1) if src_total > 0 else None
    )
    io_coverage_pct = round(with_io / max(total_tc, 1) * 100, 1)

    return {
        "total_test_cases": total_tc,
        "total_sequences": total_seq,
        "avg_sequences_per_tc": avg_seq,
        "total_input_vars": total_inp,
        "total_output_vars": total_out,
        "with_io_count": with_io,
        "with_logic_count": with_logic,
        "function_coverage_pct": func_coverage_pct,
        "io_coverage_pct": io_coverage_pct,
        "total_source_functions": src_total,
        "gen_method_distribution": gen_methods,
        "component_distribution": components,
    }


# ---------------------------------------------------------------------------
# Document validation
# ---------------------------------------------------------------------------

def validate_suts_xlsm(
    xlsm_path: str,
    expected_tc_range: Optional[tuple] = None,
    expected_seq_range: Optional[tuple] = None,
) -> Dict[str, Any]:
    """Validate generated SUTS XLSM for structural and data quality.

    Returns dict with 'valid' bool, 'issues' list, 'warnings' list, and 'stats' dict.
    """
    issues: List[str] = []
    warnings: List[str] = []
    stats: Dict[str, Any] = {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"valid": False, "issues": ["openpyxl not installed"], "warnings": [], "stats": {}}

    p = Path(xlsm_path)
    if not p.exists():
        return {"valid": False, "issues": [f"File not found: {xlsm_path}"], "warnings": [], "stats": {}}

    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        return {"valid": False, "issues": [f"Cannot open: {e}"], "stats": {}}

    required_sheets = ["2.SW Unit Test Spec"]
    for s in required_sheets:
        if s not in wb.sheetnames:
            issues.append(f"Missing required sheet: {s}")

    expected_sheets = ["Cover", "History", "1.Introduction", "1.Test Environment",
                       "2.SW Unit Test Spec", "3.Traceability"]
    stats["sheets"] = wb.sheetnames
    stats["sheet_count"] = len(wb.sheetnames)
    for s in expected_sheets:
        if s not in wb.sheetnames:
            issues.append(f"Optional sheet missing: {s}")

    if "2.SW Unit Test Spec" in wb.sheetnames:
        ws = wb["2.SW Unit Test Spec"]
        # ⚠ 레이아웃을 **하드코딩하지 않는다**. 예전엔 `min_row=7`·`max_col=149`·
        #   `row[12]`(옛 Seq.No)·`row[13:62]`(옛 Input) 가 박혀 있었다. 정본 레이아웃
        #   으로 바꾸자 검증기가 시퀀스 7,267건을 **1,576건으로** 셌다(-5,691).
        #   파일은 멀쩡한데 검증기만 틀려서 정상 산출물을 결함으로 신고했다.
        #   상수에서 파생하면 레이아웃이 또 바뀌어도 같이 따라간다.
        max_col = min(int(ws.max_column or _RELATED_COL), _RELATED_COL)
        stats["max_col"] = max_col

        tc_count = 0
        seq_count = 0
        empty_io_tcs = 0
        last_row = _HEADER_ROW
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=_DATA_START_ROW, max_col=max_col, values_only=True),
            start=_DATA_START_ROW,
        ):
            last_row = row_idx
            tc_id = row[_COL_TC_ID - 1] if len(row) >= _COL_TC_ID else None
            if tc_id and str(tc_id).startswith("SwUTC"):
                tc_count += 1
                has_input = any(
                    v not in (None, "")
                    for v in row[_INPUT_COL_START - 1:min(_INPUT_COL_END, len(row))]
                )
                has_output = any(
                    v not in (None, "")
                    for v in row[_OUTPUT_COL_START - 1:min(_OUTPUT_COL_END, len(row))]
                )
                if not has_input and not has_output:
                    empty_io_tcs += 1
            seq_val = row[_SEQ_COL - 1] if len(row) >= _SEQ_COL else None
            if seq_val is not None and str(seq_val).strip():
                seq_count += 1

        stats["max_row"] = last_row

        stats["tc_count"] = tc_count
        stats["seq_count"] = seq_count
        stats["empty_io_tc_count"] = empty_io_tcs
        stats["avg_seq_per_tc"] = round(seq_count / max(tc_count, 1), 1)

        if tc_count == 0:
            issues.append("No test cases (SwUTC_*) found")
        if seq_count == 0:
            issues.append("No test sequences found")
        if empty_io_tcs > tc_count * 0.5:
            issues.append(f"Over 50% TCs lack I/O variables ({empty_io_tcs}/{tc_count})")

        if expected_tc_range:
            lo, hi = expected_tc_range
            if tc_count < lo or tc_count > hi:
                issues.append(f"TC count {tc_count} outside expected range [{lo}, {hi}]")

        if expected_seq_range:
            lo, hi = expected_seq_range
            if seq_count < lo or seq_count > hi:
                issues.append(f"Sequence count {seq_count} outside expected range [{lo}, {hi}]")

    wb.close()
    return {"valid": len(issues) == 0, "issues": issues, "stats": stats}


@contextmanager
def _resolved_doc_input(path: Optional[str], label: str):
    """문서 입력(SRS/UDS/HSIS)을 **로컬에서 열 수 있는 경로**로 확보한다.

    과거엔 `Path(p).is_file()`만 봤다. cloudium 모드에서 U:\\ 같은 경로는 backend 프로세스에
    권한이 없어(worker exe만 접근 가능) 항상 False가 되고, 보강 블록이 **경고 한 줄 없이**
    통째로 skip됐다 — 산출물엔 "요구 ID 없음"으로만 남아 원인을 알 수 없었다.

    로컬에 있으면 원래 경로를 그대로 돌려주고(추가 I/O 0), worker에만 있으면 resolver로
    bytes를 읽어 임시 파일로 materialize한 뒤 종료 시 지운다. 어느 쪽도 아니면 None을
    yield하되 **사유를 warning으로 남긴다**.

    Yields: 열 수 있는 로컬 경로(str) 또는 None.
    """
    raw = str(path or "").strip()
    if not raw:
        yield None
        return
    try:
        if Path(raw).is_file():
            yield raw
            return
    except OSError as exc:      # 권한 거부(U:\ 등) — 로컬 판정 불가일 뿐 부재는 아니다
        _logger.debug("%s: 로컬 stat 실패(%s) — resolver로 재시도", label, exc)

    resolver = None
    try:
        from backend.services.file_resolver import get_resolver
        resolver = get_resolver()
    except Exception as exc:    # standalone 실행 등 backend 미가용
        _logger.warning("%s 입력을 건너뜀 — 로컬에 없고 resolver도 불가: %s (%s)", label, raw, exc)
        yield None
        return

    try:
        if not resolver.is_file(raw):
            _logger.warning("%s 입력을 건너뜀 — resolver(mode=%s)에도 없음: %s",
                            label, getattr(resolver, "mode", "?"), raw)
            yield None
            return
        data = resolver.read_bytes(raw)
    except Exception as exc:
        _logger.warning("%s 입력 읽기 실패 — 보강 생략: %s (%s)", label, raw, exc)
        yield None
        return

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(
            suffix=Path(raw).suffix or ".bin", prefix=f"{label}_", delete=False,
        ) as fh:
            fh.write(data)
            tmp_path = fh.name
        _logger.info("%s: worker 경로를 임시 파일로 materialize (%d bytes)", label, len(data))
        yield tmp_path
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                _logger.debug("%s: 임시 파일 정리 실패 %s", label, tmp_path)


def validate_sts_xlsm(xlsm_path: str) -> Dict[str, Any]:
    """STS 검증 — 구현은 generators.sts로 이관됐다(하위호환 re-export).

    이 함수가 여기 있던 동안 SUTS 레이아웃 상수(5/6/4열)로 STS를 읽어 Action·Expected가
    비어도 통과시켰다. 열 스키마는 산출물을 쓰는 모듈이 소유해야 한다 — generators.sts의
    `_STS_SCHEMA`가 writer·validator 공통 출처다. 기존 import 경로는 그대로 둔다.
    """
    from generators.sts import validate_sts_xlsm as _impl
    return _impl(xlsm_path)


# ---------------------------------------------------------------------------
# Top-level pipeline
# ---------------------------------------------------------------------------

def generate_suts(
    source_root: str,
    output_path: str,
    template_path: Optional[str] = None,
    project_config: Optional[Dict[str, Any]] = None,
    ai_config: Optional[Dict[str, Any]] = None,
    max_sequences: int = _DEFAULT_SEQ_COUNT,
    on_progress: Optional[Any] = None,
    srs_docx_path: Optional[str] = None,
    sds_docx_path: Optional[str] = None,
    uds_path: Optional[str] = None,
    hsis_path: Optional[str] = None,
    target_function_names: Optional[List[str]] = None,
    scope: str = "suds",
) -> Dict[str, Any]:
    """Top-level SUTS generation pipeline.

    Args (추가):
        scope: `"suds"`(기본) = SwUDS 설계 ID 가 있는 함수만 — **정본과 같은 범위**.
            `"source"` = 소스에서 찾은 함수 전부. SwUDS 문서가 없으면 `"suds"` 여도
            좁히지 않고 그 사실을 보고한다.

    Args:
        source_root: Root directory of C source code
        output_path: Path for output XLSM file
        template_path: Optional SUTS template XLSM
        project_config: Optional config dict
        ai_config: Optional AI config dict for Gemini enhancement
        max_sequences: Maximum sequences per TC
        on_progress: Optional callback(pct: int, message: str) for progress updates
        srs_docx_path: Optional path to SRS DOCX for requirement ID enrichment
        sds_docx_path: Optional path to SDS DOCX for ASIL/design context
        uds_path: Optional path to UDS DOCX/XLSM for function descriptions

    Returns:
        Dict with keys: output_path, quality_report, test_case_count, etc.
    """
    def _progress(pct: int, msg: str):
        _logger.info("[%d%%] %s", pct, msg)
        if on_progress:
            try:
                on_progress(pct, msg)
            except Exception:
                pass

    _logger.info("=== SUTS Generation Start ===")
    t0 = time.time()
    target_name_set = {
        str(name or "").strip().lower()
        for name in (target_function_names or [])
        if str(name or "").strip()
    }

    _progress(5, "소스 코드 파싱 시작")
    globals_info_map: Dict[str, Dict[str, str]] = {}
    try:
        try:
            from backend.helpers import _get_source_sections_cached

            report_data = _get_source_sections_cached(source_root)
        except Exception:
            from report_generator import generate_uds_source_sections

            report_data = generate_uds_source_sections(source_root)
        function_details = report_data.get("function_details", {})
        globals_info_map = report_data.get("globals_info_map", {}) or {}
        if not function_details:
            raise ValueError("generate_uds_source_sections returned no function_details")
    except Exception as e:
        _logger.warning("Full UDS source parse failed, trying lightweight: %s", e)
        function_details = _lightweight_parse(source_root)

    if target_name_set:
        function_details = {
            fid: info
            for fid, info in function_details.items()
            if isinstance(info, dict) and str(info.get("name") or "").strip().lower() in target_name_set
        }

    _progress(25, f"소스 파싱 완료 - {len(function_details)}개 함수 발견")

    # 함수 단위 override 맵 — **보강·보충 전용**. 필터로 쓰지 않는다.
    #
    # ⚠ 예전엔 이 목록에 **없는 함수를 전부 버렸다**("레퍼런스에 있는 함수만 포함").
    #   그 결과가 실측으로 드러났다(2026-08-11, KJPDS02_PV):
    #     · 소스 파싱 900~1,153 함수 → override 251개로 잘림 → TC 251개
    #     · 정본 SwUTS 는 1,014 함수. 그중 **782개(77.8%)가 목록에 없어 침묵 탈락**
    #     · override 251개 중 정본에 실재하는 건 223개뿐 — 28개는 없는 함수다
    #   이 파일은 **저장소**(`docs/`)에 있어 프로젝트가 바뀌어도 같은 251개로 자른다.
    #   과거 어느 스냅샷의 목록이 지금 프로젝트의 시험 범위를 정하고 있었다.
    #
    #   게다가 커버리지 분모가 **필터 후 값**이라 언제나 251/251 = "함수 커버리지
    #   100.0%" 로 보고됐다 — 77.8% 를 버리고 100% 라고 말하는 fail-open 이다.
    #
    #   그래서 필터를 걷어내고 **탈락 대신 보충만** 한다(목록에 있는데 소스에 없는
    #   함수는 종전대로 빈 엔트리로 추가한다 — 그건 정보를 더하지 빼지 않는다).
    _ovr_only_names: List[str] = []
    try:
        import json as _json
        for _ovr_path in [
            Path(__file__).resolve().parent.parent / "docs" / "uds_function_swcom_override.json",
            Path(__file__).resolve().parent / "docs" / "uds_function_swcom_override.json",
        ]:
            if _ovr_path.exists():
                _ovr_data = _json.loads(_ovr_path.read_text(encoding="utf-8"))
                _ovr_names = set(_ovr_data.keys())
                _ovr_names_lower = {n.lower() for n in _ovr_names}
                if _ovr_names:
                    before2 = len(function_details)
                    # ⚠ 여기서 걸러내지 않는다(위 주석 참조). 목록 밖 함수도 그대로 둔다.
                    _in_list = sum(
                        1 for info in function_details.values()
                        if isinstance(info, dict) and (
                            str(info.get("name") or "") in _ovr_names
                            or str(info.get("name") or "").lower() in _ovr_names_lower
                        )
                    )
                    # override에 있지만 파서에 없는 함수를 빈 엔트리로 추가
                    _existing_names = {str(info.get("name") or "").lower() for info in function_details.values() if isinstance(info, dict)}
                    _added = 0
                    for _ovr_name, _ovr_info in _ovr_data.items():
                        if _ovr_name.lower() not in _existing_names:
                            _sc = _ovr_info.get("swcom", 0)
                            _fid = f"SwUFn_{_sc:02d}{99 - _added:02d}"
                            function_details[_fid] = {
                                "id": _fid,
                                "name": _ovr_name,
                                "prototype": f"void {_ovr_name}(void)",
                                "description": "",
                                "asil": _ovr_info.get("asil", "TBD"),
                                "related": _ovr_info.get("related", "TBD"),
                                "inputs": [],
                                "outputs": [],
                                "logic_flow": [],
                                "calls_list": [],
                                "file": "",
                                "module_name": f"SwCom_{_sc:02d}",
                            }
                            _added += 1
                            _ovr_only_names.append(_ovr_name)
                    # 침묵 금지 — 목록과 소스가 얼마나 어긋나는지 그대로 보고한다.
                    _progress(
                        28,
                        f"override 보강: 소스 {before2}개 중 목록 일치 {_in_list}개 "
                        f"(+{_added} 보충 → {len(function_details)}개). 필터 아님",
                    )
                    _logger.info(
                        "uds override: source=%d in_list=%d added=%d total=%d "
                        "(목록 %d개 — 필터로 쓰지 않는다)",
                        before2, _in_list, _added, len(function_details), len(_ovr_names),
                    )
                break
    except Exception:
        pass

    if sds_docx_path:
        _progress(29, "SDS 설계 컨텍스트 로드 중")
    _sds_map = _resolve_sds_map(sds_docx_path)

    _progress(30, "유닛 함수 수집 중")
    units = collect_unit_functions(function_details, globals_info_map, sds_map=_sds_map)

    if not units:
        _logger.warning("No unit functions found!")
        return {
            "output_path": "",
            "quality_report": {},
            "test_case_count": 0,
            "elapsed_seconds": round(time.time() - t0, 1),
            "error": "No functions found in source code",
        }

    _progress(35, f"{len(units)}개 유닛 함수 수집 완료")

    # ── SRS requirement ID enrichment ────────────────────────────────────
    # 입력 경로는 resolver 경유로 확보한다(worker-only 입력의 침묵 skip 차단 —
    # _resolved_doc_input 주석 참조). 아래 UDS/HSIS 블록도 같은 규약.
    with _resolved_doc_input(srs_docx_path, "SRS") as _srs_local:
        if _srs_local:
            _progress(36, "SRS 요구사항 ID 보강 중")
            try:
                from generators.sts import parse_srs_docx_tables
                srs_reqs = parse_srs_docx_tables(_srs_local)
                if srs_reqs:
                    # Build function_name → req_ids map from SRS data
                    fn_to_reqs: Dict[str, List[str]] = {}
                    for req in srs_reqs:
                        req_id = req.get("id", "")
                        if not req_id:
                            continue
                        related = str(req.get("related_id") or req.get("verification") or "")
                        desc = str(req.get("description") or req.get("name") or "")
                        # Find function name references in requirement text
                        for m in re.finditer(r"\b([A-Za-z_]\w*(?:_pds|_init|_main|_run|_update|_check|_calc|_set|_get|_proc))\b", related + " " + desc):
                            fn_key = m.group(1).lower()
                            if fn_key not in fn_to_reqs:
                                fn_to_reqs[fn_key] = []
                            if req_id not in fn_to_reqs[fn_key]:
                                fn_to_reqs[fn_key].append(req_id)
                    # Enrich units that have no srs_req_ids yet
                    for unit in units:
                        if unit.get("srs_req_ids"):
                            continue
                        fn_lower = unit["name"].lower()
                        direct = fn_to_reqs.get(fn_lower)
                        if direct:
                            unit["srs_req_ids"] = ", ".join(direct[:4])
                    _logger.info("SRS enrichment: %d reqs parsed, %d units have req IDs now",
                                 len(srs_reqs),
                                 sum(1 for u in units if u.get("srs_req_ids")))
            except Exception as _e:
                _logger.warning("SRS enrichment skipped: %s", _e)

    # ── UDS function description enrichment ──────────────────────────────
    with _resolved_doc_input(uds_path, "UDS") as _uds_local:
        if _uds_local:
            _progress(37, "UDS 함수 설명 보강 중")
            try:
                from generators.sts import _load_uds_descriptions
                uds_descs = _load_uds_descriptions(_uds_local)
                if uds_descs:
                    enriched_count = 0
                    for unit in units:
                        fn_lower = unit["name"].lower()
                        uds_desc = uds_descs.get(fn_lower)
                        if uds_desc and len(uds_desc) > len(unit.get("description") or ""):
                            unit["description"] = uds_desc
                            enriched_count += 1
                    _logger.info("UDS descriptions enriched for %d units", enriched_count)
            except Exception as _e:
                _logger.warning("UDS description enrichment skipped: %s", _e)

            # ── 설계 ID(SwUFn_xxxx) — SUDS 칸과 TC_ID 의 근거 ──────────────
            # 정본 실측: `TC_ID = "SwUTC_" + SUDS` 가 1,013/1,014 에서 성립한다.
            # ⚠ 못 찾으면 **비운다**. 예전엔 소스 파싱 순번(`SwUFn_{n:04d}`)을 넣어
            #   모양만 맞고 다른 설계 요소를 가리켰다(정본과 교집합 178/251).
            try:
                from generators.uds_design_ids import load_uds_design_ids, resolve_design_id
                _design = load_uds_design_ids(_uds_local)
                if _design.get("by_name"):
                    _hit = 0
                    for unit in units:
                        did = resolve_design_id(_design, unit.get("name"))
                        if did:
                            unit["suds_id"] = did
                            _hit += 1
                    _logger.info(
                        "UDS design IDs: %d/%d units matched (문서 %d ids · 동명이인 %d 제외)",
                        _hit, len(units), len(_design["by_name"]), len(_design.get("ambiguous") or []),
                    )
            except Exception as _e:  # noqa: BLE001
                _logger.warning("UDS design-id enrichment skipped: %s", _e)

    # ── 시험 범위 — 기본은 **SwUDS 기반**(정본과 같은 범위) ────────────────────
    #
    # SUTS 는 SwUDS(단위 설계서)를 근거로 만드는 문서다. 정본도 그렇다 — 정본 1,005
    # 함수는 SwUDS 설계 ID 1,026 과 교집합 1,001 로 사실상 일치한다(실측 2026-08-11).
    # 소스에는 그보다 많은 함수가 있고(실측 1,160), 그중 155개는 정본이 시험 대상으로
    # 삼지 않는다(부트로더 계열 등).
    #
    # ⚠ 이건 앞서 걷어낸 `docs/uds_function_swcom_override.json` 필터와 **성질이 다르다**.
    #   그건 저장소에 박힌 251개 목록이라 프로젝트가 바뀌어도 같은 걸로 잘랐다.
    #   이건 **그 프로젝트의 SwUDS 문서**가 근거이고, 문서가 없으면 필터도 걸지 않는다.
    #
    # 범위를 좁힌 사실은 **반드시 보고한다** — 조용히 자르면 커버리지가 또 자기 자신을
    # 분모로 삼게 된다.
    _scope = str(scope or "suds").strip().lower()
    _scope_note = ""
    if _scope == "suds":
        _with_id = [u for u in units if str(u.get("suds_id") or "").strip()]
        if not _with_id:
            _scope_note = (
                "SwUDS 설계 ID 를 하나도 확보하지 못해 범위를 좁히지 않았습니다 "
                "(SwUDS 문서가 없거나 읽지 못했습니다)."
            )
            _logger.warning("SUTS scope=suds: %s", _scope_note)
        elif len(_with_id) < len(units):
            _dropped = len(units) - len(_with_id)
            _scope_note = (
                f"SwUDS 기반 범위: 소스 {len(units)}개 중 설계 ID 가 있는 "
                f"{len(_with_id)}개만 시험합니다 ({_dropped}개 제외 — SwUDS 에 없는 함수)."
            )
            _logger.info("SUTS scope=suds: %s", _scope_note)
            units = _with_id
    else:
        _scope_note = f"소스 전체 범위: {len(units)}개 함수 전부를 시험합니다(SwUDS 미대조)."
        _logger.info("SUTS scope=source: %s", _scope_note)
    if _scope_note:
        _progress(40, _scope_note)

    # ── HSIS signal enrichment ────────────────────────────────────────────
    # Uses HSIS xlsx to enrich: srs_req_ids (from related_id), variable
    # boundary hints from characteristics (e.g. "0...255"), and srs_req_ids
    # for units that read/write HSIS signal SW variables.
    # 파일 접근만 with 안에서 끝낸다 — 아래 가공은 메모리 데이터라 임시 파일이 필요 없다.
    _hsis_data: Optional[Dict[str, Any]] = None
    with _resolved_doc_input(hsis_path, "HSIS") as _hsis_local:
        if _hsis_local:
            _progress(38, "HSIS 신호 보강 중")
            try:
                from generators.sts import _load_hsis_signals
                _hsis_data = _load_hsis_signals(_hsis_local)
            except Exception as _hsis_exc:
                _logger.warning("HSIS 파싱 실패 — 보강 생략: %s", _hsis_exc)
    if _hsis_data:
        try:
            _hsis_signals = _hsis_data.get("signals", [])
            if _hsis_signals:
                # Build sw_var_name → signal dict (one var can split by \n/,)
                _hsis_var_map: Dict[str, Dict[str, Any]] = {}
                for _sig in _hsis_signals:
                    _sw_raw = str(_sig.get("sw_var_name") or "")
                    for _tok in re.split(r"[\n,\s]+", _sw_raw):
                        _tok = _tok.strip()
                        if _tok and re.match(r"^[A-Za-z_]\w+$", _tok):
                            _hsis_var_map[_tok] = _sig

                # Parse "min...max" or "min - max" from characteristics
                def _parse_hsis_range(chars: str):
                    if not chars:
                        return None, None
                    m = re.search(r"([-\d.]+)\s*\.{2,3}\s*([-\d.]+)", chars)
                    if not m:
                        m = re.search(r"([-\d.]+)\s*[-~]\s*([-\d.]+)", chars)
                    if m:
                        try:
                            return float(m.group(1)), float(m.group(2))
                        except ValueError:
                            pass
                    return None, None

                enriched_hsis = 0
                for unit in units:
                    # Collect all variable names used by this unit
                    # unit dict uses "input_vars"/"output_vars" (string lists),
                    # not "inputs"/"outputs" (dicts).
                    _unit_vars: List[str] = list(unit.get("input_vars") or [])
                    _unit_vars += list(unit.get("output_vars") or [])

                    _matched: List[Dict[str, Any]] = [
                        _hsis_var_map[v] for v in _unit_vars if v in _hsis_var_map
                    ]
                    if not _matched:
                        continue

                    # 1) enrich srs_req_ids from HSIS related_id
                    if not unit.get("srs_req_ids"):
                        _hsis_req_ids = [
                            s["related_id"] for s in _matched
                            if s.get("related_id") and str(s["related_id"]).strip()
                        ]
                        if _hsis_req_ids:
                            unit["srs_req_ids"] = ", ".join(
                                list(dict.fromkeys(_hsis_req_ids))[:4]
                            )

                    # 2) store HSIS boundary hints on the unit for sequence generation
                    _hsis_bounds: Dict[str, tuple] = {}
                    for _vname in _unit_vars:
                        if _vname in _hsis_var_map:
                            _chars = _hsis_var_map[_vname].get("characteristics", "")
                            _lo, _hi = _parse_hsis_range(_chars)
                            if _lo is not None and _hi is not None:
                                _hsis_bounds[_vname] = (_lo, _hi)
                    if _hsis_bounds:
                        unit.setdefault("hsis_bounds", {}).update(_hsis_bounds)

                    enriched_hsis += 1

                _logger.info("HSIS enrichment: %d units enriched from %d signals",
                             enriched_hsis, len(_hsis_signals))
        except Exception as _hsis_exc:
            _logger.warning("HSIS enrichment skipped: %s", _hsis_exc)

    if globals_info_map:
        set_globals_type_cache(globals_info_map)

    # ── Indirect variable enrichment for void/no-param functions ─────────────
    # For units with no input/output vars, derive testable variables from
    # the global variables of their callee functions (indirect side effects).
    _fn_name_to_info: Dict[str, Dict[str, Any]] = {
        info.get("name", ""): info
        for info in function_details.values()
        if isinstance(info, dict) and info.get("name")
    }
    for _void_unit in units:
        if _void_unit.get("input_vars") or _void_unit.get("output_vars"):
            continue
        _indirect: List[str] = []
        for _callee_name in (_void_unit.get("calls_list") or [])[:8]:
            _callee_info = _fn_name_to_info.get(_callee_name)
            if not _callee_info:
                continue
            # Prefer callee outputs (side effects), then inputs
            _callee_outs = _extract_var_names(_callee_info.get("outputs") or [])
            _callee_ins = _extract_var_names(_callee_info.get("inputs") or [])
            for _v in _callee_outs + _callee_ins:
                if _v not in _indirect:
                    _indirect.append(_v)
            if len(_indirect) >= 6:
                break
        if _indirect:
            _void_unit["indirect_vars"] = _indirect[:6]
            _logger.debug("void unit %s: indirect_vars=%s", _void_unit["name"], _indirect[:6])

    # ── logic_flow variable extraction for remaining void functions ──────────
    # For units still without any testable variable (calls_list was also empty),
    # extract C identifier names from logic_flow condition/text strings.
    # Filters out: C keywords, all-caps macro names, single-letter tokens.
    _C_KEYWORDS = frozenset({
        "if", "else", "while", "for", "return", "switch", "case", "break",
        "continue", "do", "null", "true", "false", "void", "int", "char",
        "uint", "uint8", "uint16", "uint32", "int8", "int16", "int32",
    })
    for _void_unit in units:
        if (_void_unit.get("input_vars") or _void_unit.get("output_vars")
                or _void_unit.get("indirect_vars")):
            continue
        _lf_vars: List[str] = []

        def _walk_flow_nodes(nodes: List[Any]) -> None:
            for _n in nodes:
                if not isinstance(_n, dict):
                    continue
                _text = str(_n.get("condition") or _n.get("text") or "")
                # Extract C identifiers from condition/text
                for _tok in re.findall(r"\b([A-Za-z_][A-Za-z0-9_]{2,})\b", _text):
                    if (_tok.lower() not in _C_KEYWORDS
                            and not _tok.isupper()  # skip ALL_CAPS macros
                            and _tok not in _lf_vars
                            and len(_lf_vars) < 6):
                        _lf_vars.append(_tok)
                # Recurse into sub-bodies
                for _key in ("true_body", "false_body", "body"):
                    _sub = _n.get(_key)
                    if isinstance(_sub, list):
                        _walk_flow_nodes(_sub)

        _walk_flow_nodes(_void_unit.get("logic_flow") or [])
        if _lf_vars:
            _void_unit["indirect_vars"] = _lf_vars
            _logger.debug("void unit %s: logic_flow vars=%s", _void_unit["name"], _lf_vars)

    # Identify void functions that still lack all variable info — these are
    # the only ones that need AI enhancement (limits API calls to ~12 units).
    _void_no_vars = {
        u["fid"] for u in units
        if not u.get("input_vars") and not u.get("output_vars")
        and not u.get("indirect_vars")
    }
    _logger.info("Units needing AI enhancement: %d", len(_void_no_vars))

    _progress(40, "테스트 시퀀스 생성 시작")
    all_sequences: Dict[str, List[Dict[str, Any]]] = {}
    ai_enhanced = 0
    for i, unit in enumerate(units):
        seqs = generate_sequences(unit, max_sequences)
        if ai_config and unit["fid"] in _void_no_vars:
            seqs = enhance_sequences_with_ai(unit, seqs, ai_config)
            ai_enhanced += 1
        all_sequences[unit["fid"]] = seqs
        if (i + 1) % 50 == 0 or i == len(units) - 1:
            pct = 40 + int(35 * (i + 1) / len(units))
            _progress(pct, f"시퀀스 생성 {i+1}/{len(units)}")
    if ai_enhanced:
        _logger.info("AI enhanced %d void-function units", ai_enhanced)

    total_seq = sum(len(s) for s in all_sequences.values())
    _progress(80, f"시퀀스 생성 완료 - {total_seq}개")

    quality = generate_suts_quality_report(units, all_sequences, len(function_details))

    _progress(85, "XLSM 파일 생성 중")
    out = generate_suts_xlsm(template_path, units, all_sequences, output_path, project_config)

    _progress(90, "생성 문서 자동 검증 중")
    validation = validate_suts_xlsm(out)
    # 생성 수 ↔ 파일 기록 수 대조. `validate_suts_xlsm` 에 `expected_tc_range`/
    # `expected_seq_range` 인자가 **있는데도** 호출부 4곳이 전부 기본값 None 이라
    # 그 대조는 한 번도 실행된 적이 없었다 — 정답(total_seq)이 바로 윗줄에 있는데도.
    # 판정 로직은 세 생성기 공용 단일 출처(`_artifact_check`)로 통일한다.
    validation = apply_write_back_check(validation, {
        "tc_count": len(units),
        "seq_count": total_seq,
    })
    if validation.get("issues"):
        _logger.warning("SUTS validation issues: %s", validation["issues"])

    validation_report_path = ""
    try:
        validation_report_path = generate_suts_validation_report(out, quality, validation=validation)
        _logger.info("SUTS validation report: %s", validation_report_path)
    except Exception as _vr:
        _logger.warning("SUTS validation report generation skipped: %s", _vr)

    elapsed = time.time() - t0
    _progress(100, f"SUTS 생성 완료 ({elapsed:.1f}초)")

    # Quality DB recording (non-fatal)
    try:
        from workflow.quality.recorder import record_run
        record_run(
            "suts", quality,
            project_root=str(source_root or ""),
            elapsed_sec=elapsed,
            output_path=out,
            ai_model=str((ai_config or {}).get("model", "")),
        )
    except Exception:
        # non-fatal 은 유지하되 침묵은 금지 (sts.py 의 동일 블록이 NameError 를
        # 몇 년간 삼켜 품질 기록이 통째로 유실된 전례).
        _logger.exception("SUTS quality record skipped (non-fatal)")

    return {
        "output_path": out,
        "quality_report": quality,
        "test_case_count": len(units),
        "total_sequences": total_seq,
        "elapsed_seconds": round(elapsed, 1),
        "validation": validation,
        "validation_report_path": validation_report_path,
    }


def _lightweight_parse(source_root: str) -> Dict[str, Dict[str, Any]]:
    """Lightweight C source parsing when full report_generator is unavailable."""
    from report.c_parsing import (
        _extract_c_definitions,
        _extract_c_function_bodies,
        _extract_simple_call_names,
        _strip_c_comments,
    )

    root = Path(source_root)
    c_files = list(root.rglob("*.c"))
    function_details: Dict[str, Dict[str, Any]] = {}
    fn_counter = 0

    for cf in c_files:
        try:
            raw = cf.read_text(encoding="utf-8", errors="replace")
        except Exception:
            continue

        stripped = _strip_c_comments(raw)
        defs = _extract_c_definitions(stripped)
        bodies = _extract_c_function_bodies(stripped)

        for d in defs:
            # _extract_c_definitions returns Tuple[name, params, is_static]
            if isinstance(d, tuple):
                name = d[0] if len(d) > 0 else ""
                params = d[1] if len(d) > 1 else ""
            elif isinstance(d, dict):
                name = d.get("name", "")
                params = d.get("params", "")
            else:
                continue
            if not name:
                continue
            fn_counter += 1
            fid = f"SwUFn_{fn_counter:04d}"
            sig = f"void {name}({params})" if params else f"void {name}(void)"
            body = bodies.get(name, "")
            calls = _extract_simple_call_names(body) if body else []

            function_details[fid] = {
                "id": fid,
                "name": name,
                "prototype": sig,
                "inputs": [f"[IN] {p}" for p in _lw_parse_params(sig)],
                "outputs": _lw_parse_outputs(sig, name),
                "calls_list": calls,
                "logic_flow": _lw_extract_logic_flow(body),
                "globals_global": [],
                "globals_static": [],
                "module_name": cf.stem,
                "file": str(cf),
                "description": "",
                "asil": "TBD",
                "precondition": "",
            }

    return function_details


def _lw_parse_params(sig: str) -> List[str]:
    if "(" not in sig:
        return []
    params = sig.split("(", 1)[1].rsplit(")", 1)[0].strip()
    if not params or params.lower() == "void":
        return []
    result = []
    for p in params.split(","):
        p = p.strip()
        parts = p.split()
        if parts:
            result.append(parts[-1].strip("*&"))
    return result


def _lw_parse_outputs(sig: str, name: str) -> List[str]:
    if not sig:
        return []
    head = sig.split(name, 1)[0] if name in sig else sig
    head = re.sub(r"\b(static|extern|inline)\b", "", head).strip()
    if head and "void" not in head.lower():
        return [f"[OUT] return {head.strip()}"]
    return []


_LW_BRANCH_RE = re.compile(
    r'\b(if|else\s+if|else|switch|case|for|while)\b\s*(\([^)]*\))?',
    re.IGNORECASE,
)


def _lw_extract_logic_flow(body: str) -> List[Dict[str, Any]]:
    """Extract simplified logic flow nodes from a C function body."""
    if not body:
        return []
    nodes: List[Dict[str, Any]] = []
    for m in _LW_BRANCH_RE.finditer(body):
        keyword = m.group(1).strip().lower()
        cond = (m.group(2) or "").strip("() \t")
        node: Dict[str, Any] = {"type": keyword, "text": m.group(0).strip()}
        if cond:
            node["condition"] = cond
        nodes.append(node)
        if len(nodes) >= 40:
            break
    return nodes


# ---------------------------------------------------------------------------
# Document validation
# ---------------------------------------------------------------------------

def generate_suts_validation_report(
    xlsm_path: str,
    quality_report: Optional[Dict[str, Any]] = None,
    validation: Optional[Dict[str, Any]] = None,
) -> str:
    """Generate a validation report markdown for SUTS XLSM.

    Writes a .validation.md file next to the XLSM and returns its path.
    """
    validation_data = validation if isinstance(validation, dict) else validate_suts_xlsm(xlsm_path)
    stats = validation_data.get("stats", {})
    issues = validation_data.get("issues", [])
    qr = quality_report or {}

    tc_count = stats.get("tc_count", 0)
    seq_count = stats.get("seq_count", 0)
    empty_io = stats.get("empty_io_tc_count", 0)

    lines = [
        "# SUTS 생성 문서 자동 검증 리포트",
        "",
        f"**파일**: `{Path(xlsm_path).name}`  ",
        f"**검증 시각**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ",
        f"**결과**: {'PASS' if validation_data.get('valid') else 'FAIL'}",
        "",
        "---",
        "",
        "## 1. 구조 검증",
        "",
        "| 항목 | 값 |",
        "|------|-----|",
        f"| 시트 수 | {stats.get('sheet_count', 0)} |",
        f"| 시트 목록 | {', '.join(stats.get('sheets', []))} |",
        f"| TC 수 | {tc_count} |",
        f"| 시퀀스 수 | {seq_count} |",
        f"| TC당 평균 시퀀스 | {stats.get('avg_seq_per_tc', 0)} |",
        f"| I/O 없는 TC 수 | {empty_io} |",
        "",
    ]

    if qr:
        lines.extend([
            "## 2. 품질 지표",
            "",
            "| 항목 | 값 |",
            "|------|-----|",
            f"| 총 TC 수 | {qr.get('total_test_cases', 0)} |",
            f"| 총 시퀀스 수 | {qr.get('total_sequences', 0)} |",
            f"| TC당 평균 시퀀스 | {qr.get('avg_sequences_per_tc', 0)} |",
            f"| 총 입력 변수 | {qr.get('total_input_vars', 0)} |",
            f"| 총 출력 변수 | {qr.get('total_output_vars', 0)} |",
            f"| I/O 보유 TC | {qr.get('with_io_count', 0)} ({qr.get('io_coverage_pct', 0)}%) |",
            f"| 로직 보유 TC | {qr.get('with_logic_count', 0)} |",
            # 미측정은 `0%` 가 아니라 `—` 다. 0% 로 그리면 "한 함수도 안 덮였다" 로 읽힌다.
            "| 함수 커버리지 | "
            + (f"{qr['function_coverage_pct']}%" if qr.get("function_coverage_pct") is not None else "— (미측정)")
            + f" (소스 함수 {qr.get('total_source_functions') or '—'}개 기준) |",
            "",
        ])
        if qr.get("gen_method_distribution"):
            lines.extend([
                "### 생성 방법 분포",
                "",
                "| 방법 | 수 |",
                "|------|-----|",
            ])
            for k, v in qr["gen_method_distribution"].items():
                lines.append(f"| {k} | {v} |")
            lines.append("")

    gate_items = [
        ("TC 존재", tc_count > 0),
        ("시퀀스 존재", seq_count > 0),
        ("I/O 없는 TC < 50%", empty_io <= tc_count * 0.5 if tc_count else True),
        ("TC당 평균 시퀀스 >= 2", stats.get("avg_seq_per_tc", 0) >= 2 if tc_count else True),
        # ⚠ 미측정(None)을 통과로 접지 않는다 — 못 잰 것을 "이상 없음" 으로 만들면
        #   게이트가 fail-open 이 된다. TC 가 있으면 그건 그것대로 별도 항목이 본다.
        ("함수 커버리지 측정됨", (qr or {}).get("function_coverage_pct") is not None),
    ]
    passed = sum(1 for _, ok in gate_items if ok)

    lines.extend([
        f"## 3. Quality Gate ({passed}/{len(gate_items)})",
        "",
        "| 항목 | 결과 |",
        "|------|------|",
    ])
    for name, ok in gate_items:
        lines.append(f"| {name} | {'PASS' if ok else 'FAIL'} |")
    lines.append("")

    if issues:
        lines.extend(["## 4. Issues", ""])
        for i in issues:
            lines.append(f"- {i}")
        lines.append("")

    out_path = Path(xlsm_path).with_suffix(".validation.md")
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return str(out_path)


def validate_suts_output(xlsm_path: str) -> Dict[str, Any]:
    """Validate a generated SUTS XLSM for structural completeness."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    issues: List[str] = []
    warnings: List[str] = []
    stats: Dict[str, Any] = {"sheets": wb.sheetnames, "sheet_count": len(wb.sheetnames)}

    expected_sheets = ["Cover", "History", "1.Introduction", "1.Test Environment", "2.SW Unit Test Spec"]
    for s in expected_sheets:
        if s not in wb.sheetnames:
            issues.append(f"Missing sheet: {s}")

    if "2.SW Unit Test Spec" in wb.sheetnames:
        ws = wb["2.SW Unit Test Spec"]
        tc_count = 0
        seq_count = 0
        total_inp = 0
        total_out = 0
        tc_no_inp = 0
        tc_no_out = 0
        # ⚠ 열 번호를 하드코딩하지 않는다 — 레이아웃 상수에서 파생한다. 예전엔
        #   `range(14,63)`·`column=11` 이 박혀 있어, 레이아웃이 바뀌면 검증기가
        #   조용히 0을 세고 그 0이 "이슈 없음"으로 통과했다(fail-open).
        #
        # ⚠ `read_only=True` 에서 `ws.cell(row, col)` 랜덤 접근을 쓰지 말 것.
        #   순차 스트리밍이라 되짚으면 **빈 셀을 돌려준다**. 실측(2026-08-11):
        #   시퀀스 7,267건이 파일에 멀쩡히 있는데 검증기는 1,576건으로 셌다
        #   (-5,691). 행이 1,975 → 8,219 로 늘자 증상이 드러났다. 파일이 아니라
        #   **검증기가 틀린 것**이라, 그대로 뒀으면 정상 산출물을 결함으로 신고한다.
        #   전체 스캔은 `iter_rows` 가 유일한 정답이다
        #   (`[[reference_openpyxl_readonly_cell_perf]]` — 성능만이 아니라 정확성 문제).
        _max_col = max(_OUTPUT_COL_END, _RELATED_COL)
        for row in ws.iter_rows(min_row=_DATA_START_ROW, max_col=_max_col, values_only=True):
            tc_id = row[_COL_TC_ID - 1] if len(row) >= _COL_TC_ID else None
            if tc_id and str(tc_id).startswith("SwUTC"):
                tc_count += 1
                _n_inp = sum(
                    1 for v in row[_INPUT_COL_START - 1:_INPUT_COL_END] if v not in (None, "")
                )
                _n_out = sum(
                    1 for v in row[_OUTPUT_COL_START - 1:_OUTPUT_COL_END] if v not in (None, "")
                )
                total_inp += _n_inp
                total_out += _n_out
                # ⚠ 평균은 0 을 숨긴다. 실측(2026-08-12): 948 TC 중 **338 건이 입력 0개**인데
                #   평균은 2.0 이라 `avg_inp < 1` 게이트를 그대로 통과했다. 입력이 없는
                #   시퀀스는 시험이 성립하지 않으므로 건수를 따로 센다.
                if _n_inp == 0:
                    tc_no_inp += 1
                if _n_out == 0:
                    tc_no_out += 1
            elif len(row) >= _SEQ_COL and row[_SEQ_COL - 1] not in (None, ""):
                seq_count += 1

        stats["tc_count"] = tc_count
        stats["seq_count"] = seq_count
        stats["avg_inp"] = round(total_inp / max(tc_count, 1), 1)
        stats["avg_out"] = round(total_out / max(tc_count, 1), 1)
        stats["avg_seq"] = round(seq_count / max(tc_count, 1), 1)
        # 항상 싣는다 — 0 건이어도 키가 있어야 "재지 않았다" 와 "0 이었다" 가 구분된다.
        stats["tc_without_input"] = tc_no_inp
        stats["tc_without_expected"] = tc_no_out

        if tc_count == 0:
            issues.append("No test cases found")
        if seq_count == 0:
            issues.append("No sequences found")
        if stats["avg_inp"] < 1:
            issues.append(f"Low avg input vars: {stats['avg_inp']}")
        if stats["avg_out"] < 1:
            issues.append(f"Low avg output vars: {stats['avg_out']}")
        # 경고는 `issues` 와 분리한다 — 입력 0개 TC 는 **정상일 수도 있다**(파라미터도
        # 전역도 없는 함수. 정본도 1,005 중 172 건이 그렇다). `valid` 를 뒤집으면
        # 정상 산출물이 실패로 신고된다. 다만 숨기지도 않는다.
        if tc_count and tc_no_inp:
            warnings.append(
                f"입력 변수가 없는 TC {tc_no_inp}건 "
                f"({tc_no_inp * 100.0 / tc_count:.1f}%) — 해당 시퀀스는 실행 값이 없다"
            )
        if tc_count and tc_no_out:
            warnings.append(
                f"기대 결과가 없는 TC {tc_no_out}건 ({tc_no_out * 100.0 / tc_count:.1f}%)"
            )

    wb.close()
    stats["issues"] = issues
    stats["warnings"] = warnings
    stats["valid"] = len(issues) == 0
    return stats
