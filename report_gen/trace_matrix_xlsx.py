"""추적성 매트릭스 → 감사용 xlsx 내보내기 (hiMA TrMatrixReport 대응).

`generate_uds_traceability_matrix()` + `build_link_table()` 결과를 감사자가 가장 자주
요구하는 **xlsx** 형식으로 렌더한다. hiMA가 화면 그대로 xlsx로 내보내던 기능에 대응하되,
우리 강점(부동소수 커버리지%, ASIL 셀 결합, 역방향 공백)을 함께 담는다.

시트 구성:
  1. 교차표   — 행=요구사항 × 열=SDS 컴포넌트, 셀 O/공백 + UDS/STS/SUTS/SITS/VC 카운트.
               추적 0건 행 핑크, ASIL 갭 행 앰버, ASIL 등급 색. 상단 문서메타·커버리지 헤더.
  2. 링크테이블 — 평면 (target_id, related_id, related_type, source, confidence) 감사 baseline.
  3. 커버리지  — 밴드별 부동소수%, ASIL 등급별 충족/갭/미상.

순수 함수(부작용 없음, 시각 datetime은 호출자가 meta로 주입). openpyxl 미설치 시 ImportError.
"""

from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

# DoS 가드 — 인증 뒤이고 클라이언트가 echo하는 자기 데이터지만, 단일 요청으로 거대한
# 시트 생성을 막는다(실데이터 상한의 넉넉한 배수: req ~1k, SDS열 ~1k, 링크 ~50k).
_MAX_ROWS = 20000
_MAX_COLS = 2000
_MAX_LINKS = 200000
# 시트4(정합성 감사) 누적 캡 — export는 클라이언트 echo body를 신뢰하므로(producer의
# trace_integrity 캡은 이 경로에서 안 돈다) 시트1~3과 동일하게 xlsx 측 외부 상한을 둔다.
_MAX_SHEET4_ROWS = 20000   # 시트4 전체 데이터 행 누적 상한
_MAX_BANDS = 500           # dangling/placeholder 밴드(dict 키) 수 상한

# ASIL 등급 셀 글자색(교차표) — 등급 인지성. QM/미상은 muted.
_ASIL_FONT = {"D": "FF991B1B", "C": "FFDC2626", "B": "FFB45309", "A": "FF2563EB", "QM": "FF6B7280"}

# 행 배경 — 추적 0건(핑크)·ASIL 갭(앰버). hiMA 0카운트 핑크밴드 대응 + 우리 ASIL 차별점.
_PINK = "FFFEE2E2"
_AMBER = "FFFEF9C3"
_HEADER_BG = "FFE5E7EB"


def _cs(v: Any) -> str:
    """Excel 수식 주입 방지 — =,+,-,@ 로 시작하는 문서유래 텍스트에 ' 프리픽스.

    링크테이블/요구사항ID/컴포넌트명 등 외부 문서에서 온 값이 셀에서 수식으로
    해석되는 것을 차단(export 보안 표준 완화책). 숫자/고정라벨엔 적용 안 함.
    """
    s = "" if v is None else str(v)
    return ("'" + s) if s[:1] in ("=", "+", "-", "@") else s


def _unwrap(matrix: Any) -> Dict[str, Any]:
    if isinstance(matrix, dict):
        if "rows" in matrix:
            return matrix
        inner = matrix.get("matrix")
        if isinstance(inner, dict):
            return inner
    return matrix if isinstance(matrix, dict) else {}


def _test_id(t: Any) -> str:
    if not isinstance(t, dict):
        return ""
    for k in ("testcase", "subprogram", "unit", "id"):
        v = str(t.get(k) or "").strip()
        if v:
            return v
    return ""


def _row_bands(row: Dict[str, Any]) -> Dict[str, List[str]]:
    """행에서 밴드별 related_id 목록 — build_link_table/CrossMatrixView와 동일 규칙."""
    def tids(arr: Any) -> List[str]:
        return [x for x in (_test_id(t) for t in (arr if isinstance(arr, list) else [])) if x]

    return {
        "SDS": [str(c).strip() for c in (row.get("sds_components") or []) if str(c).strip()],
        "UDS": [str(s).strip() for s in (row.get("source_ids") or []) if str(s).strip()],
        "STS": tids(row.get("sts_tests")),
        "SUTS": tids(row.get("suts_tests")),
        "SITS": tids(row.get("sits_tests")),
        "VectorCAST": [
            _test_id(t) for t in (row.get("tests") or [])
            if isinstance(t, dict) and t.get("source") == "VectorCAST" and _test_id(t)
        ],
    }


def build_trace_xlsx(matrix: Any, meta: Optional[Dict[str, Any]] = None) -> bytes:
    """추적성 매트릭스 dict → xlsx 바이트.

    Args:
        matrix: ``generate_uds_traceability_matrix()`` 결과(top-level 또는 ``{"matrix":...}``)
            — rows 및 (있으면) link_table 포함.
        meta: 헤더 블록용 — project_name, generated_at, doc_versions 등(없으면 생략).

    Returns:
        xlsx 파일 바이트(openpyxl).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    inner = _unwrap(matrix)
    rows = inner.get("rows") if isinstance(inner, dict) else None
    rows = [r for r in (rows if isinstance(rows, list) else []) if isinstance(r, dict)][:_MAX_ROWS]
    lt = inner.get("link_table") if isinstance(inner, dict) else None
    lt = lt if isinstance(lt, dict) else {}
    asil_cov = lt.get("asil_coverage") if isinstance(lt.get("asil_coverage"), dict) else {}
    meta = meta if isinstance(meta, dict) else {}

    # ASIL 갭 set(앰버 강조용) — link_table 우선.
    gap_ids = {
        str(g.get("target_id")) for g in (asil_cov.get("gaps") or [])
        if isinstance(g, dict) and g.get("target_id")
    }

    # 행 데이터 파생(교차표) + SDS 열 수집
    built: List[Dict[str, Any]] = []
    sds_cols: List[str] = []
    sds_seen = set()
    for row in rows:
        rid = str(row.get("requirement_id") or "").strip()
        if not rid:
            continue
        bands = _row_bands(row)
        for c in bands["SDS"]:
            if c not in sds_seen:
                sds_seen.add(c)
                sds_cols.append(c)
        total = sum(len(bands[b]) for b in bands)
        built.append({
            "rid": rid,
            "name": str(row.get("requirement_name") or "").strip(),
            "asil": str(row.get("asil") or "").strip().upper(),
            "bands": bands,
            "total": total,
        })
    sds_cols = sorted(sds_cols)[:_MAX_COLS]

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    pink_fill = PatternFill("solid", fgColor=_PINK)
    amber_fill = PatternFill("solid", fgColor=_AMBER)

    wb = Workbook()

    # ── 시트 1: 교차표 ──
    ws = wb.active
    ws.title = "교차표"
    r = 1
    # 문서 메타 헤더 블록
    ws.cell(r, 1, "추적성 매트릭스 (Traceability Matrix)").font = Font(bold=True, size=13)
    r += 1
    for label, key in (("프로젝트", "project_name"), ("생성시각", "generated_at"),
                       ("Job", "job_url"), ("빌드", "build_selector")):
        v = str(meta.get(key) or "").strip()
        if v:
            ws.cell(r, 1, label).font = bold
            ws.cell(r, 2, v)
            r += 1
    # 커버리지 요약 한 줄
    by_band = (lt.get("coverage") or {}).get("by_band") or {}
    if by_band:
        cov_txt = " · ".join(
            f"{b} {(by_band.get(b) or {}).get('pct', 0)}%" for b in ("SDS", "UDS", "STS", "SUTS", "SITS", "VectorCAST")
        )
        ws.cell(r, 1, "커버리지").font = bold
        ws.cell(r, 2, cov_txt)
        r += 1
    total_reqs = len(built)
    uncovered_n = sum(1 for b in built if b["total"] == 0)
    ws.cell(r, 1, "요구사항 수").font = bold
    ws.cell(r, 2, f"{total_reqs}건 (추적 0건 {uncovered_n}건, ASIL 갭 {len(gap_ids)}건, 미상 {int(asil_cov.get('unknown_count') or 0)}건)")
    r += 2

    has_asil = bool(asil_cov.get("has_asil"))
    # 표 헤더
    head_row = r
    headers = ["요구사항", "제목"]
    if has_asil:
        headers.append("ASIL")
    headers += list(sds_cols) + ["UDS", "STS", "SUTS", "SITS", "VC", "합계"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(head_row, ci, _cs(h))
        c.font = bold
        c.fill = header_fill
        c.alignment = center
    ws.freeze_panes = ws.cell(head_row + 1, (3 if has_asil else 2) + 1)

    for bi, b in enumerate(built):
        rr = head_row + 1 + bi
        uncovered = b["total"] == 0
        gaprow = b["rid"] in gap_ids
        row_fill = pink_fill if uncovered else (amber_fill if gaprow else None)
        ws.cell(rr, 1, _cs(b["rid"])).font = bold
        ws.cell(rr, 2, _cs(b["name"]))
        col = 3
        if has_asil:
            ac = ws.cell(rr, col, (b["asil"] + (" ⚠" if gaprow else "")) if b["asil"] else "–")
            ac.alignment = center
            if b["asil"] in _ASIL_FONT:
                ac.font = Font(bold=True, color=_ASIL_FONT[b["asil"]])
            col += 1
        sds_set = set(b["bands"]["SDS"])
        for sc in sds_cols:
            cc = ws.cell(rr, col, "O" if sc in sds_set else "")
            cc.alignment = center
            col += 1
        ws.cell(rr, col, len(b["bands"]["UDS"]) or "").alignment = center
        col += 1
        for band in ("STS", "SUTS", "SITS", "VectorCAST"):
            ws.cell(rr, col, "O" if b["bands"][band] else "").alignment = center
            col += 1
        tc = ws.cell(rr, col, b["total"])
        tc.alignment = center
        tc.font = bold
        # 행 배경(고정 좌측 + 합계까지)
        if row_fill is not None:
            for cidx in range(1, col + 1):
                ws.cell(rr, cidx).fill = row_fill

    # 열 너비(고정 좌측만 — SDS 열은 좁게)
    ws.column_dimensions[get_column_letter(1)].width = 16
    ws.column_dimensions[get_column_letter(2)].width = 30

    # ── 시트 2: 링크테이블 ──
    ws2 = wb.create_sheet("링크테이블")
    lh = ["target_id", "related_id", "related_type", "source", "confidence"]
    for ci, h in enumerate(lh, 1):
        c = ws2.cell(1, ci, h)
        c.font = bold
        c.fill = header_fill
    links = lt.get("links") or []
    for li, lk in enumerate(links[:_MAX_LINKS], start=2):
        if not isinstance(lk, dict):
            continue
        for ci, k in enumerate(lh, 1):
            ws2.cell(li, ci, _cs(lk.get(k)))
    ws2.freeze_panes = "A2"
    for ci, w in enumerate((22, 28, 18, 12, 12), 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── 시트 3: 커버리지 ──
    ws3 = wb.create_sheet("커버리지")
    ws3.cell(1, 1, "밴드별 커버리지").font = Font(bold=True, size=12)
    ws3.cell(2, 1, "밴드").font = bold
    ws3.cell(2, 2, "추적 요구사항").font = bold
    ws3.cell(2, 3, "전체").font = bold
    ws3.cell(2, 4, "비율(%)").font = bold
    rr = 3
    for b in ("SDS", "UDS", "STS", "SUTS", "SITS", "VectorCAST"):
        bd = by_band.get(b) or {}
        ws3.cell(rr, 1, b)
        ws3.cell(rr, 2, bd.get("linked_targets", 0))
        ws3.cell(rr, 3, bd.get("total_targets", 0))
        ws3.cell(rr, 4, bd.get("pct", 0))
        rr += 1
    # ASIL 등급별
    rr += 1
    ws3.cell(rr, 1, "ASIL 등급별 (충족/갭/미상)").font = Font(bold=True, size=12)
    rr += 1
    ws3.cell(rr, 1, "등급").font = bold
    ws3.cell(rr, 2, "요구사항").font = bold
    ws3.cell(rr, 3, "시험추적").font = bold
    ws3.cell(rr, 4, "갭").font = bold
    rr += 1
    for lvl, v in sorted((asil_cov.get("by_level") or {}).items()):
        if not isinstance(v, dict):
            continue
        ws3.cell(rr, 1, lvl)
        ws3.cell(rr, 2, v.get("targets", 0))
        ws3.cell(rr, 3, v.get("test_covered", 0))
        ws3.cell(rr, 4, v.get("gap", 0))
        rr += 1
    # ASIL 갭 상세
    gaps = asil_cov.get("gaps") or []
    if gaps:
        rr += 1
        ws3.cell(rr, 1, "ASIL 갭 상세 (안전등급 대비 시험추적 부족)").font = Font(bold=True, size=12)
        rr += 1
        ws3.cell(rr, 1, "요구사항").font = bold
        ws3.cell(rr, 2, "ASIL").font = bold
        ws3.cell(rr, 3, "부족 밴드").font = bold
        rr += 1
        for g in gaps:
            if not isinstance(g, dict):
                continue
            ws3.cell(rr, 1, _cs(g.get("target_id")))
            ws3.cell(rr, 2, str(g.get("asil") or ""))
            ws3.cell(rr, 3, ", ".join(g.get("missing") or []))
            rr += 1
    for ci, w in enumerate((24, 14, 14, 18), 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ── 시트 4: 정합성 감사 (trace_integrity — hiMA WrongRelatedID/WrongName 대응) ──
    integ = inner.get("integrity") if isinstance(inner, dict) else None
    if isinstance(integ, dict):
        istats = integ.get("stats")
        istats = istats if isinstance(istats, dict) else {}
        ws4 = wb.create_sheet("정합성 감사")
        ws4.cell(1, 1, "ID 정합성 감사 (Traceability ID Integrity)").font = Font(bold=True, size=12)
        clean = bool(istats.get("clean", True))
        ws4.cell(2, 1, "결과").font = bold
        ws4.cell(
            2, 2,
            "이상 없음(clean)" if clean else (
                f"정규화 충돌 {int(istats.get('collision_count') or 0)} · "
                f"오참조 의심 {int(istats.get('dangling_suspect_count') or 0)} · "
                f"계층참조 {int(istats.get('dangling_foreign_count') or 0)} · "
                f"placeholder {int(istats.get('placeholder_count') or 0)}"
            ),
        )
        rr4 = 4
        # 1) 정규화 충돌 — raw 철자 N개가 같은 canonical로 붕괴(hiMA exact-match가 silent 오인).
        collisions = integ.get("id_collisions") or []
        if collisions:
            ws4.cell(rr4, 1, "정규화 충돌 (raw 철자 → 단일 canonical 병합)").font = Font(bold=True, size=11)
            rr4 += 1
            for ci, h in enumerate(("canonical", "변형 수", "raw 변형(전체)", "표시 유지"), 1):
                cc = ws4.cell(rr4, ci, h)
                cc.font = bold
                cc.fill = header_fill
            rr4 += 1
            for c in collisions:
                if rr4 >= _MAX_SHEET4_ROWS:
                    break
                if not isinstance(c, dict):
                    continue
                ws4.cell(rr4, 1, _cs(c.get("canonical"))).fill = amber_fill
                ws4.cell(rr4, 2, int(c.get("variant_count") or 0)).alignment = center
                ws4.cell(rr4, 3, _cs(" | ".join(str(v) for v in (c.get("variants") or []))))
                ws4.cell(rr4, 4, _cs(c.get("kept")))
                rr4 += 1
            rr4 += 1
        # 2) 상향 dangling — 설계/단위설계가 SRS 부분집합에 없는 요구사항 참조(namespace별).
        dangling = integ.get("dangling_refs") or {}
        if dangling:
            ws4.cell(rr4, 1, "상향 dangling 참조 (SRS 부분집합에 없는 대상 인용)").font = Font(bold=True, size=11)
            rr4 += 1
            ns_map = integ.get("dangling_by_namespace") or {}
            ns_txt = "; ".join(
                f"{band}: " + ", ".join(f"{k}×{v}" for k, v in (ns_map.get(band) or {}).items())
                for band in sorted(ns_map) if ns_map.get(band)
            )
            if ns_txt:
                ws4.cell(rr4, 1, "namespace 분포").font = bold
                ws4.cell(rr4, 2, _cs(ns_txt))
                rr4 += 1
            for ci, h in enumerate(("출처", "참조 ID(raw)", "정규화", "namespace", "심각도"), 1):
                cc = ws4.cell(rr4, ci, h)
                cc.font = bold
                cc.fill = header_fill
            rr4 += 1
            for band in sorted(dangling)[:_MAX_BANDS]:
                if rr4 >= _MAX_SHEET4_ROWS:
                    break
                for d in (dangling.get(band) or []):
                    if rr4 >= _MAX_SHEET4_ROWS:
                        break
                    if not isinstance(d, dict):
                        continue
                    sev = str(d.get("severity") or "")
                    ws4.cell(rr4, 1, band)
                    ws4.cell(rr4, 2, _cs(d.get("ref_id")))
                    ws4.cell(rr4, 3, _cs(d.get("normalized")))
                    ws4.cell(rr4, 4, _cs(d.get("namespace")))
                    sc = ws4.cell(rr4, 5, "오참조 의심" if sev == "suspect" else "계층참조")
                    # 오참조 의심만 앰버 강조 — foreign(계층참조)은 무강조(정보성).
                    if sev == "suspect":
                        sc.fill = amber_fill
                    rr4 += 1
            rr4 += 1
        # 3) placeholder 참조 ID — 미완성 템플릿 토큰(SwCom_XX/TBD/??).
        placeholders = integ.get("placeholder_ids") or {}
        if placeholders:
            ws4.cell(rr4, 1, "placeholder 참조 ID (미완성 템플릿)").font = Font(bold=True, size=11)
            rr4 += 1
            for ci, h in enumerate(("출처", "placeholder ID"), 1):
                cc = ws4.cell(rr4, ci, h)
                cc.font = bold
                cc.fill = header_fill
            rr4 += 1
            for band in sorted(placeholders)[:_MAX_BANDS]:
                if rr4 >= _MAX_SHEET4_ROWS:
                    break
                for pid in (placeholders.get(band) or []):
                    if rr4 >= _MAX_SHEET4_ROWS:
                        break
                    ws4.cell(rr4, 1, band)
                    ws4.cell(rr4, 2, _cs(pid)).fill = amber_fill
                    rr4 += 1
        for ci, w in enumerate((20, 40, 20, 14, 14), 1):
            ws4.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
