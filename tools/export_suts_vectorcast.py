from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


_TC_SHEET = "2.SW Unit Test Spec"
_DATA_START_ROW = 7
_COMPONENT_COL = 2
_TC_ID_COL = 3
_NAME_COL = 4
_DESCRIPTION_COL = 5
_TEST_METHOD_COL = 8
_GEN_METHOD_COL = 9
_PRECONDITION_COL = 10
_SEQUENCE_TEXT_COL = 11
_TC_GEN_METHOD_COL = 12
_SEQ_NO_COL = 13
_INPUT_COL_START = 14
_INPUT_COL_END = 62
_OUTPUT_COL_START = 63
_OUTPUT_COL_END = 148
_RELATED_COL = 149
_REQ_PAT = re.compile(r"\b(?:Sw|Sy)[A-Za-z_]*_\d+\b")

# 헤더 스캔 범위 — SUTS 템플릿이 2종(레이아웃 상이)이라 위 상수만으론 한쪽만 맞다.
#   HDPDM01 v3.01 : 헤더 rows5-6, SeqNo=13 Input=14.. Expected=63.. Related=149  ← 위 상수와 일치
#   KJPDS02_PV v1.02: 헤더 rows3-4, SeqNo=8  Inpt[0]=9.. ExpR[0]=105.. Related=189
# 하드코딩 상수는 KJPDS02_PV에서 입력컬럼(col13=Inpt[4])을 SeqNo 게이트로 오용 → 입력 ≤4개 함수의
# 전 시퀀스행이 '빈 블록'으로 드롭(1013중 704), 파싱된 것도 입력/기대값이 컬럼 밀림으로 오염됐다.
_HEADER_SCAN_ROWS = 8
_MAX_SCAN_COLS = 1024


def _detect_columns(ws: Any) -> Dict[str, Any]:
    """문서 헤더에서 실제 컬럼 위치를 탐지(레이아웃 적응형). 모듈 상수는 HDPDM01 기본값·폴백.

    불변식(관측된 3개 레이아웃 모두 성립): SeqNo=입력시작-1, 입력끝=출력시작-1, 출력끝=Related-1.
    전역 Input/Inpt[0]/Expected/Related 헤더가 없는 문서(구 단위테스트 픽스처)는 밴드를 탐지하지
    못하므로 모듈 상수를 그대로 유지한다(무회귀). TC_ID(3)/Name(4)은 3개 레이아웃 공통이라 상수 유지.
    ⚠ 'Input' 병합앵커가 KJPDS02_PV에선 SeqNo 컬럼(c8)에 걸리므로 Inpt[0]를 우선 신뢰한다."""
    cols: Dict[str, Any] = {
        "component": _COMPONENT_COL, "tc_id": _TC_ID_COL, "name": _NAME_COL,
        "description": _DESCRIPTION_COL, "test_method": _TEST_METHOD_COL,
        "gen_method": _GEN_METHOD_COL, "precondition": _PRECONDITION_COL,
        "seq_text": _SEQUENCE_TEXT_COL, "tc_gen_method": _TC_GEN_METHOD_COL,
        "seq_no": _SEQ_NO_COL, "input_start": _INPUT_COL_START, "input_end": _INPUT_COL_END,
        "output_start": _OUTPUT_COL_START, "output_end": _OUTPUT_COL_END, "related": _RELATED_COL,
    }
    inpt0 = input_hdr = expected = related = None
    found: Dict[str, int] = {}
    gen_cols: List[int] = []
    maxc = min(int(ws.max_column or 0), _MAX_SCAN_COLS)
    for r in range(1, _HEADER_SCAN_ROWS + 1):
        for c in range(1, maxc + 1):
            raw = ws.cell(row=r, column=c).value
            if raw is None:
                continue
            t = str(raw).strip().lower()
            if not t:
                continue
            # ⚠ 개행/탭/연속공백을 단일 공백으로 접는다 — generators/suts.py 가 실제로 쓰는 헤더는
            # "Test\nMethod"·"Gen.\nMethod"·"Test Case\nGen.Method" 처럼 **개행 삽입**이라(deep-review C1)
            # strip/replace만으론 라벨이 안 맞아 header_driven 문서에서 test_method/gen_method 를
            # 조용히 None(→"")으로 떨궜다(HDPDM01 생성본 재파싱 시 provenance 침묵 손실).
            tn = " ".join(t.replace("_", " ").split())
            if inpt0 is None and re.fullmatch(r"(?:inpt|input)\[0\]", tn):
                inpt0 = c
            if input_hdr is None and tn == "input":
                input_hdr = c
            # "expected result"로 좁힌다(W2 오탐 방지) — 스캔창(rows 1-8)이 데이터행과 겹쳐
            # "Expected coverage…" 같은 산문이 bare "expected"에 걸려 output_start를 탈취하던 벡터 차단.
            # 생성기/라이브 2템플릿 모두 헤더는 "Expected Result"(개행은 정규화가 접음).
            if expected is None and tn.startswith("expected result"):
                expected = c
            if related is None and "related id" in tn:
                related = c
            if "description" not in found and tn == "description":
                found["description"] = c
            if "precondition" not in found and tn == "precondition":
                found["precondition"] = c
            if "seq_text" not in found and tn == "sequence":
                found["seq_text"] = c
            if "test_method" not in found and tn == "test method":
                found["test_method"] = c
            # "generation method"(정식) 과 "gen. method"(생성본 약어) 를 모두 포용. 'test method'는
            # 'gen'을 포함하지 않으므로 오탐 없음.
            if "gen" in tn and "method" in tn and c not in gen_cols:
                gen_cols.append(c)
    if gen_cols:
        found["gen_method"] = gen_cols[0]
    input_start = inpt0 or input_hdr
    # 밴드 원자성(W1): input과 expected를 **함께** 탐지해야 밴드를 신뢰한다. input만 탐지하고 상수
    # output/related와 섞으면 역전/혼합 밴드(예 [63..29] 공집합, 입력밴드가 expected 열 흡수)를 만들어
    # expected 값이 input으로 오분류되는 침묵 손상이 난다 → 부분탐지는 전부 상수 폴백(구 파서 동작).
    header_driven = input_start is not None and expected is not None
    if input_start is not None and expected is not None:
        cols["input_start"] = input_start
        cols["seq_no"] = input_start - 1
        cols["output_start"] = expected
        cols["input_end"] = expected - 1
        if related is not None:
            cols["related"] = related
            cols["output_end"] = related - 1
    elif input_start is not None:
        # 입력 헤더는 있는데 Expected 헤더가 없다 = 미지/변형 레이아웃. 침묵 손상 대신 사유를 남긴다(X8).
        cols["_detect_warning"] = ("SUTS 헤더탐지: 입력 헤더는 찾았으나 'Expected Result' 미탐지 — "
                                   "상수 레이아웃으로 폴백(밴드 오검출 방지). 컬럼 정합 수동 확인 필요")
    # 필드 컬럼: 라벨을 찾으면 그 컬럼. 못 찾았는데 header_driven이면 그 템플릿에 컬럼이 없는 것
    # (KJPDS02_PV는 Description/Precondition/Sequence-text 부재) → None 반환해 파서가 입력컬럼을
    # 오독(쓰레기값)하는 대신 "" 를 쓰게 한다. header_driven이 아니면(픽스처/부분탐지) 상수 유지.
    for key in ("description", "precondition", "seq_text", "test_method", "gen_method"):
        if key in found:
            cols[key] = found[key]
        elif header_driven:
            cols[key] = None
    cols["tc_gen_method"] = gen_cols[1] if len(gen_cols) >= 2 else (None if header_driven else _TC_GEN_METHOD_COL)
    return cols


def _clean_text(value: Any) -> str:
    text = str(value or "").strip()
    return text


def _normalize_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return None
    if text.startswith("[검증 필요]"):
        return {"verification_required": True, "raw": text}
    if re.fullmatch(r"-?\d+", text):
        try:
            return int(text)
        except Exception:
            return text
    if re.fullmatch(r"-?\d+\.\d+", text):
        try:
            return float(text)
        except Exception:
            return text
    return text


def _extract_related_ids(*texts: str) -> List[str]:
    ids: List[str] = []
    for text in texts:
        for req_id in _REQ_PAT.findall(str(text or "")):
            if req_id not in ids:
                ids.append(req_id)
    return ids


def _iter_tc_blocks(ws: Any, cols: Dict[str, Any]) -> Iterable[Tuple[int, int]]:
    tc_col = cols["tc_id"] or _TC_ID_COL
    row = _DATA_START_ROW
    max_row = ws.max_row
    while row <= max_row:
        if _clean_text(ws.cell(row=row, column=tc_col).value):
            start = row
            row += 1
            while row <= max_row and not _clean_text(ws.cell(row=row, column=tc_col).value):
                row += 1
            yield start, row - 1
        else:
            row += 1


def _header_names(ws: Any, row: int, col_start: int, col_end: int) -> List[Tuple[int, str]]:
    headers: List[Tuple[int, str]] = []
    for col in range(col_start, col_end + 1):
        name = _clean_text(ws.cell(row=row, column=col).value)
        if name:
            headers.append((col, name))
    return headers


def _parse_sequence_row(
    ws: Any,
    row: int,
    input_headers: List[Tuple[int, str]],
    output_headers: List[Tuple[int, str]],
    unit_meta: Dict[str, Any],
    cols: Dict[str, Any],
) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    warnings: List[Dict[str, str]] = []
    seq_no = ws.cell(row=row, column=cols["seq_no"] or _SEQ_NO_COL).value
    seq_no_text = _clean_text(seq_no)
    sequence_no = int(seq_no) if isinstance(seq_no, int) else seq_no_text
    base_tc_id = str(unit_meta["base_tc_id"])
    _seq_text_col = cols.get("seq_text")
    sequence = {
        "name": f"{base_tc_id}__SEQ_{int(sequence_no):02d}" if str(sequence_no).isdigit() else f"{base_tc_id}__SEQ_{seq_no_text or row}",
        "base_tc_id": base_tc_id,
        "sequence_no": sequence_no,
        "description": _clean_text(ws.cell(row=row, column=_seq_text_col).value) if _seq_text_col else "",
        "precondition": unit_meta.get("precondition", ""),
        "inputs": {},
        "expected": {},
        "notes": {
            "strategy": unit_meta.get("gen_method", ""),
            "test_method": unit_meta.get("test_method", ""),
        },
    }
    for col, name in input_headers:
        value = _normalize_scalar(ws.cell(row=row, column=col).value)
        if value is not None:
            sequence["inputs"][name] = value
    for col, name in output_headers:
        value = _normalize_scalar(ws.cell(row=row, column=col).value)
        if value is not None:
            sequence["expected"][name] = value
            if isinstance(value, dict) and value.get("verification_required"):
                warnings.append(
                    {
                        "code": "verification_required_expected",
                        "message": f"{unit_meta['unit_name']} seq {sequence_no}: expected '{name}' needs manual verification.",
                    }
                )
    if not sequence["expected"]:
        warnings.append(
            {
                "code": "empty_expected",
                "message": f"{unit_meta['unit_name']} seq {sequence_no}: no expected outputs mapped.",
            }
        )
    return sequence, warnings


def _parse_tc_block(ws: Any, start_row: int, end_row: int, cols: Dict[str, Any]) -> Dict[str, Any]:
    def _cell(key: str, fallback: int) -> str:
        col = cols.get(key)
        col = col if col is not None else fallback
        return _clean_text(ws.cell(row=start_row, column=col).value)

    def _cell_opt(key: str) -> str:
        # header_driven 템플릿에서 부재 컬럼은 None → "" (입력컬럼 오독 방지).
        col = cols.get(key)
        return _clean_text(ws.cell(row=start_row, column=col).value) if col is not None else ""

    component = _cell("component", _COMPONENT_COL)
    tc_id = _cell("tc_id", _TC_ID_COL)
    unit_name = _cell("name", _NAME_COL)
    description = _cell_opt("description")
    precondition = _cell_opt("precondition")
    test_method = _cell_opt("test_method")
    gen_method = _cell_opt("gen_method") or _cell_opt("tc_gen_method")
    related_token = _cell_opt("related")
    # related_token(Related ID/SUDS 컬럼)도 채굴 대상에 포함(deep-review W4) — KJPDS02_PV는
    # Description/Precondition 컬럼이 없어 산문만 쓰면 복구된 1013 유닛이 요구 추적 0이 된다.
    # Related ID 컬럼은 존재하므로 여기서 요구/설계 ID를 뽑아 추적성을 살린다(정규식 게이트=무해).
    related_ids = _extract_related_ids(description, precondition, related_token)
    input_headers = _header_names(ws, start_row, cols["input_start"] or _INPUT_COL_START, cols["input_end"] or _INPUT_COL_END)
    output_headers = _header_names(ws, start_row, cols["output_start"] or _OUTPUT_COL_START, cols["output_end"] or _OUTPUT_COL_END)
    unit = {
        "unit_name": unit_name,
        "prototype": "",
        "component": component,
        "fid": related_token,
        "metadata": {
            "gen_method": gen_method,
            "test_method": test_method,
            "related_ids": related_ids,
        },
        "test_cases": [],
        "warnings": [],
    }
    unit_meta = {
        "unit_name": unit_name,
        "base_tc_id": tc_id,
        "precondition": precondition,
        "gen_method": gen_method,
        "test_method": test_method,
    }
    _seq_gate = cols["seq_no"] or _SEQ_NO_COL
    for row in range(start_row + 1, end_row + 1):
        if ws.cell(row=row, column=_seq_gate).value in (None, ""):
            continue
        test_case, warnings = _parse_sequence_row(ws, row, input_headers, output_headers, unit_meta, cols)
        test_case["metadata"] = {
            "related_ids": related_ids,
            "fid": related_token,
            "component": component,
        }
        test_case["source"] = {
            "sheet": _TC_SHEET,
            "tc_row": start_row,
            "sequence_row": row,
        }
        unit["test_cases"].append(test_case)
        unit["warnings"].extend(warnings)
    if not unit["test_cases"]:
        unit["warnings"].append(
            {"code": "empty_test_case_block", "message": f"{unit_name or tc_id}: no sequence rows found."}
        )
    return unit


def build_vectorcast_model(
    suts_path: str,
    *,
    project_id: str = "HDPDM01",
    target_functions: Optional[Iterable[str]] = None,
    source_bytes: Optional[bytes] = None,
) -> Dict[str, Any]:
    target_set = {str(x or "").strip().lower() for x in (target_functions or []) if str(x or "").strip()}
    # cloudium(U:\)에서는 backend가 파일을 직접 열 수 없어 workbook을 worker가 읽은 bytes로 받는다.
    # source_bytes가 주어지면 그 bytes로, 아니면 로컬 경로에서 연다(로컬 모드 하위호환).
    if source_bytes is not None:
        import io as _io
        workbook = load_workbook(_io.BytesIO(source_bytes), keep_vba=True, data_only=False)
    else:
        workbook = load_workbook(suts_path, keep_vba=True, data_only=False)
    if _TC_SHEET not in workbook.sheetnames:
        raise ValueError(f"missing worksheet: {_TC_SHEET}")
    ws = workbook[_TC_SHEET]
    cols = _detect_columns(ws)
    units: List[Dict[str, Any]] = []
    export_warnings: List[Dict[str, str]] = []
    _detect_warn = cols.pop("_detect_warning", None)
    if _detect_warn:
        export_warnings.append({"code": "header_detect_fallback", "message": str(_detect_warn)})
    for start_row, end_row in _iter_tc_blocks(ws, cols):
        unit = _parse_tc_block(ws, start_row, end_row, cols)
        if target_set and unit["unit_name"].strip().lower() not in target_set:
            continue
        if not unit["unit_name"]:
            export_warnings.append(
                {"code": "missing_unit_name", "message": f"TC row {start_row}: unit name is empty."}
            )
        if unit["warnings"]:
            export_warnings.extend(unit["warnings"])
        units.append(unit)
    workbook.close()
    return {
        "schema_version": "1.0",
        "project_id": project_id,
        "source": {
            "suts_path": str(suts_path) if source_bytes is not None else str(Path(suts_path).resolve()),
            "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        },
        "units": units,
        "export_warnings": export_warnings,
    }


def write_warnings_md(model: Dict[str, Any], out_path: str) -> None:
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    warnings = list(model.get("export_warnings") or [])
    lines = [
        "# SUTS -> VectorCAST Export Warnings",
        "",
        f"- Units: `{len(model.get('units') or [])}`",
        f"- Warnings: `{len(warnings)}`",
        "",
    ]
    if not warnings:
        lines.append("- none")
    else:
        for item in warnings:
            lines.append(f"- `{item.get('code')}`: {item.get('message')}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def export_suts_to_vectorcast_model(
    suts_path: str,
    output_json: str,
    *,
    warnings_md: str = "",
    target_functions: Optional[Iterable[str]] = None,
    project_id: str = "HDPDM01",
) -> Dict[str, Any]:
    model = build_vectorcast_model(
        suts_path,
        project_id=project_id,
        target_functions=target_functions,
    )
    out_path = Path(output_json)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(model, ensure_ascii=False, indent=2), encoding="utf-8")
    if warnings_md:
        write_warnings_md(model, warnings_md)
    return model


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export SUTS XLSM to VectorCAST intermediate JSON.")
    parser.add_argument("--suts-path", required=True)
    parser.add_argument("--output-json", required=True)
    parser.add_argument("--warnings-md", default="")
    parser.add_argument("--target-functions", default="")
    parser.add_argument("--project-id", default="HDPDM01")
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    targets = [x.strip() for x in str(args.target_functions or "").split(",") if x.strip()]
    model = export_suts_to_vectorcast_model(
        args.suts_path,
        args.output_json,
        warnings_md=args.warnings_md,
        target_functions=targets,
        project_id=args.project_id,
    )
    print(f"VECTORCAST_JSON={Path(args.output_json).resolve()}")
    print(f"VECTORCAST_UNITS={len(model.get('units') or [])}")
    print(f"VECTORCAST_WARNINGS={len(model.get('export_warnings') or [])}")
    if args.warnings_md:
        print(f"VECTORCAST_WARNINGS_MD={Path(args.warnings_md).resolve()}")


if __name__ == "__main__":
    main()
